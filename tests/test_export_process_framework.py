from __future__ import annotations

import hashlib
import json
import subprocess
import sys

import pytest
from openpyxl import load_workbook

from netconsole.core.i18n import I18n
from netconsole.services.export.export_handlers import run_generic_export_handler
from netconsole.services.export.export_task_builders import (
    INLINE_ROW_LIMIT,
    app_logs_csv_spec,
    command_reference_markdown_spec,
    markdown_text_file_spec,
    result_file_rows_source,
    table_csv_source_spec,
    table_csv_spec,
)
from netconsole.services.export_task_models import ExportJob
from netconsole.services.trackside_ap_business import (
    TRACKSIDE_AP_BUSINESS_COLUMNS,
    TracksideApExportCancelled,
    export_trackside_ap_business_xlsx,
)


def test_export_job_keeps_legacy_export_type_compatibility() -> None:
    job = ExportJob(job_id="job-1", export_type="mesh_link_detail", output_path="detail.xlsx", tmp_path="detail.xlsx.tmp")

    assert job.job_type == "mesh_link_detail"
    assert job.export_type == "mesh_link_detail"
    assert ExportJob.from_dict(job.to_dict()).job_type == "mesh_link_detail"


def test_export_worker_import_does_not_load_pyside6() -> None:
    result = subprocess.run(
        [
            sys.executable,
            "-c",
            "import netconsole.export_worker, sys; print(any(name.startswith('PySide6') for name in sys.modules))",
        ],
        check=True,
        capture_output=True,
        text=True,
    )

    assert result.stdout.strip() == "False"


def test_mesh_link_detail_export_opens_derived_database_read_only(tmp_path) -> None:
    from netconsole.export_worker import _run_mesh_link_detail_export
    from netconsole.repositories.mesh_mr_repository import MeshMrRepository

    database = tmp_path / "mesh.sqlite"
    repository = MeshMrRepository(database)
    with repository._connect() as conn:
        conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
    wal = database.with_name(f"{database.name}-wal")
    before = (hashlib.sha256(database.read_bytes()).hexdigest(), database.stat().st_mtime_ns, wal.exists())

    job = ExportJob(
        job_id="mesh-read-only",
        job_type="mesh_link_detail_export",
        output_path=str(tmp_path / "links.xlsx"),
        tmp_path=str(tmp_path / "links.xlsx.tmp"),
        db_path=str(database),
        filters={"source_file_id": 1},
    )

    with pytest.raises(RuntimeError, match="暂无可导出的链路明细数据"):
        _run_mesh_link_detail_export(job)

    after = (hashlib.sha256(database.read_bytes()).hexdigest(), database.stat().st_mtime_ns, wal.exists())
    assert after == before


def test_generic_table_csv_handler_writes_utf8_sig_and_replaces_tmp(tmp_path) -> None:
    output = tmp_path / "table.csv"
    tmp = tmp_path / "table.csv.tmp"
    job = ExportJob(
        job_id="csv-job",
        job_type="table_csv",
        output_path=str(output),
        tmp_path=str(tmp),
        params={
            "payload": {
                "columns": [{"key": "name", "title": "名称"}, {"key": "mac", "title": "MAC"}],
                "rows": [{"name": "宁波站", "mac": "00aa-bbcc-ddee"}],
            }
        },
    )

    result = run_generic_export_handler(job)

    assert result["row_count"] == 1
    assert output.read_bytes().startswith(b"\xef\xbb\xbf")
    assert "宁波站" in output.read_text(encoding="utf-8-sig")
    assert not tmp.exists()


def test_table_csv_spec_uses_inline_rows_source(tmp_path) -> None:
    output = tmp_path / "source_table.csv"
    spec = table_csv_spec(
        output,
        columns=[("名称", "name"), ("MAC", "mac")],
        rows=[{"name": "宁波站", "mac": "00aa-bbcc-ddee"}],
        allow_inline_rows=True,
        inline_reason="单元测试小型内联数据",
    )
    job = spec.to_job("source-csv-job")
    job = ExportJob.from_dict({**job.to_dict(), "tmp_path": str(tmp_path / "source_table.csv.tmp")})

    result = run_generic_export_handler(job)

    assert result["row_count"] == 1
    payload = spec.payload
    assert payload["source"]["type"] == "inline_rows"
    assert "rows" not in payload
    assert "宁波站" in output.read_text(encoding="utf-8-sig")


def test_table_csv_source_spec_reads_jsonl_result_file_in_export_process(tmp_path) -> None:
    result_file = tmp_path / "toolbox_result.jsonl"
    result_file.write_text(
        '{"name":"宁波站","status":"正常"}\n{"name":"鼓楼站","status":"告警"}\n',
        encoding="utf-8",
    )
    output = tmp_path / "source_table.csv"
    spec = table_csv_source_spec(
        output,
        columns=[("名称", "name"), ("状态", "status")],
        source=result_file_rows_source(result_file),
    )
    job = spec.to_job("jsonl-source-csv-job")
    job = ExportJob.from_dict({**job.to_dict(), "tmp_path": str(tmp_path / "source_table.csv.tmp")})

    result = run_generic_export_handler(job)

    assert result["row_count"] == 2
    text = output.read_text(encoding="utf-8-sig")
    assert "宁波站" in text
    assert "鼓楼站" in text


def test_inline_rows_builder_rejects_large_payload(tmp_path) -> None:
    rows = ({"name": str(index)} for index in range(INLINE_ROW_LIMIT + 1))

    with pytest.raises(ValueError, match="inline_rows"):
        table_csv_spec(
            tmp_path / "too_many.csv",
            columns=[("名称", "name")],
            rows=rows,
            allow_inline_rows=True,
            inline_reason="单元测试超限内联数据",
        )


def test_generic_table_xlsx_handler_writes_header_and_rows(tmp_path) -> None:
    output = tmp_path / "table.xlsx"
    job = ExportJob(
        job_id="xlsx-job",
        job_type="table_xlsx",
        output_path=str(output),
        tmp_path=str(tmp_path / "table.xlsx.tmp"),
        params={
            "payload": {
                "sheet_name": "测试",
                "columns": [{"key": "name", "title": "名称"}, {"key": "status", "title": "状态"}],
                "rows": [{"name": "AP-1", "status": "正常"}],
            }
        },
    )

    result = run_generic_export_handler(job)

    assert result["row_count"] == 1
    sheet = load_workbook(output)["测试"]
    assert [sheet.cell(1, column).value for column in (1, 2)] == ["名称", "状态"]
    assert [sheet.cell(2, column).value for column in (1, 2)] == ["AP-1", "正常"]


def test_generic_markdown_text_handler_writes_utf8(tmp_path) -> None:
    output = tmp_path / "commands.md"
    job = ExportJob(
        job_id="md-job",
        job_type="markdown_text",
        output_path=str(output),
        tmp_path=str(tmp_path / "commands.md.tmp"),
        params={"payload": {"text": "# 命令清单\n\n- display version"}},
    )

    result = run_generic_export_handler(job)

    assert result["row_count"] > 0
    assert output.read_text(encoding="utf-8").startswith("# 命令清单")


def test_generic_markdown_text_handler_reads_text_file_in_export_process(tmp_path) -> None:
    source = tmp_path / "collect.log"
    source.write_text("display current-configuration\n中文日志", encoding="utf-8")
    output = tmp_path / "collect_export.txt"
    spec = markdown_text_file_spec(output, text_file=source, title="采集日志")
    job = spec.to_job("md-file-job")
    job = ExportJob.from_dict({**job.to_dict(), "tmp_path": str(tmp_path / "collect_export.txt.tmp")})

    result = run_generic_export_handler(job)

    assert result["row_count"] > 0
    assert output.read_text(encoding="utf-8") == source.read_text(encoding="utf-8")


def test_command_reference_export_reads_resource_in_export_process(tmp_path) -> None:
    resource = tmp_path / "command_reference.json"
    resource.write_text(
        json.dumps(
            {
                "items": [
                    {
                        "id": "display-version",
                        "module": "设备管理",
                        "category": "巡检",
                        "command_template": "display version",
                        "purpose": "查看版本",
                    }
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    output = tmp_path / "commands.md"
    spec = command_reference_markdown_spec(output, resource_path=resource, selected_ids=["display-version"])
    job = spec.to_job("command-reference-job")
    job = ExportJob.from_dict({**job.to_dict(), "tmp_path": str(tmp_path / "commands.md.tmp")})

    result = run_generic_export_handler(job)

    assert result["row_count"] == 1
    assert "display version" in output.read_text(encoding="utf-8")


def test_command_reference_export_keeps_empty_selection_empty(tmp_path) -> None:
    resource = tmp_path / "command_reference.json"
    resource.write_text(
        json.dumps({"items": [{"id": "display-version", "command_template": "display version"}]}),
        encoding="utf-8",
    )
    output = tmp_path / "commands.md"
    spec = command_reference_markdown_spec(output, resource_path=resource, selected_ids=[])
    job = spec.to_job("command-reference-empty")
    job = ExportJob.from_dict({**job.to_dict(), "tmp_path": str(tmp_path / "commands.md.tmp")})

    run_generic_export_handler(job)

    assert "display version" not in output.read_text(encoding="utf-8")


def test_app_log_export_applies_page_offset_and_limit_in_export_process(tmp_path) -> None:
    log_path = tmp_path / "app.log"
    log_path.write_text(
        "2026-01-01 00:00:01 | INFO | FIRST | first\n"
        "2026-01-01 00:00:02 | INFO | SECOND | second\n"
        "2026-01-01 00:00:03 | INFO | THIRD | third\n",
        encoding="utf-8",
    )
    output = tmp_path / "current_page.csv"
    spec = app_logs_csv_spec(output, log_path=log_path, offset=1, limit=1)
    job = spec.to_job("app-log-page-job")
    job = ExportJob.from_dict({**job.to_dict(), "tmp_path": str(tmp_path / "current_page.csv.tmp")})

    result = run_generic_export_handler(job)

    assert result["row_count"] == 1
    lines = output.read_text(encoding="utf-8-sig").splitlines()
    assert len(lines) == 3
    assert lines[0].startswith("#NETCONSOLE_META,")


def test_trackside_xlsx_export_reports_progress(tmp_path) -> None:
    i18n = I18n("zh_CN")
    rows = [
        {"site": "Station A", "device_name": "SW-1", "interface_name": "GigabitEthernet1/0/1", "link_status": "UP"},
        {"site": "Station A", "device_name": "SW-1", "interface_name": "GigabitEthernet1/0/2", "link_status": "DOWN"},
    ]
    events: list[tuple[str, int, int, str]] = []

    export_trackside_ap_business_xlsx(
        tmp_path / "trackside_progress.xlsx",
        rows,
        TRACKSIDE_AP_BUSINESS_COLUMNS,
        [i18n.t(key) for key, _field in TRACKSIDE_AP_BUSINESS_COLUMNS],
        progress_callback=lambda stage, current, total, message: events.append((stage, current, total, message)),
    )

    assert events[0] == ("write_trackside_rows", 0, 2, "正在写入轨旁AP业务明细 0/2")
    assert ("write_trackside_rows", 2, 2, "正在写入轨旁AP业务明细 2/2") in events
    assert any(stage == "save_workbook" for stage, _current, _total, _message in events)


def test_trackside_xlsx_export_honors_cancel_callback(tmp_path) -> None:
    i18n = I18n("zh_CN")
    export_path = tmp_path / "trackside_cancelled.xlsx"

    with pytest.raises(TracksideApExportCancelled):
        export_trackside_ap_business_xlsx(
            export_path,
            [{"site": "Station A", "device_name": "SW-1", "interface_name": "GigabitEthernet1/0/1"}],
            TRACKSIDE_AP_BUSINESS_COLUMNS,
            [i18n.t(key) for key, _field in TRACKSIDE_AP_BUSINESS_COLUMNS],
            should_cancel=lambda: True,
        )

    assert not export_path.exists()
