from __future__ import annotations

import io
import json
from dataclasses import replace
from pathlib import Path
import re
import sqlite3
import urllib.error

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
import pytest

from netconsole.core.paths import PathResolver
from netconsole.core.sites import SiteManager
from netconsole.models.wps_sync import (
    TRACKSIDE_AP_WPS_BUSINESS_KEY,
    WpsFreezeMode,
    WpsSyncTarget,
    WpsTargetType,
    build_wps_binding_id,
)
from netconsole.backend.api.wps_sync_router import _site_id
from netconsole.repositories.wps_sync_repository import WpsSyncRepository
from netconsole.services.wps_trackside_ap_sync import (
    STANDARD_TARGET_CODE,
    TracksideApWpsSyncService,
    WpsAirScriptClient,
    WpsHttpResponse,
    WpsRemoteTask,
    WpsStandardSpreadsheetAdapter,
    WpsSyncError,
    WPS_DEPLOYMENT_IDS,
    WPS_SCRIPT_VERSIONS,
    WPS_STANDARD_FORMAT_MIRROR_ENABLED,
    WPS_SYNC_TASK_TYPE,
    _append_column_width_report_warning,
    _cell_format_payload,
    _source_workbook_format_manifest,
    _column_width_probe_verified,
    _column_width_verification_report,
    _openpyxl_color,
    _assert_standard_sync_readiness,
    _sheet_tab_color_probe_verified,
    parse_wps_webhook,
    workbook_dto_from_xlsx,
)

TEST_WEBHOOK_DOCUMENT_ID = "549847228994"
STALE_REMOTE_DOCUMENT_ID = TEST_WEBHOOK_DOCUMENT_ID


def _protect(data: bytes, entropy: bytes) -> bytes:
    return bytes(value ^ entropy[index % len(entropy)] for index, value in enumerate(data))


def _wps_target() -> WpsSyncTarget:
    return WpsSyncTarget(
        target_id="target-1",
        site_id="demo",
        business_key=TRACKSIDE_AP_WPS_BUSINESS_KEY,
        target_code=STANDARD_TARGET_CODE,
        target_type=WpsTargetType.STANDARD_SPREADSHEET,
        credential_id="credential-1",
        target_name="普通表格",
        document_open_url="https://www.kdocs.cn/l/document",
        webhook_url="https://www.kdocs.cn/api/v3/ide/file/document/script/test/sync_task",
        expected_document_id="document",
    )


def _insert_retired_smart_target(
    repository: WpsSyncRepository,
    *,
    credential_id: str,
    create_credential: bool,
) -> None:
    with sqlite3.connect(repository.path) as connection:
        if create_credential:
            connection.execute(
                "INSERT INTO wps_credentials "
                "(credential_id, name, encrypted_token, token_suffix, created_at, updated_at) "
                "VALUES (?, 'retired', ?, 'oken', 'old', 'old')",
                (credential_id, _protect(b"smart-token", repository._entropy(credential_id))),
            )
        connection.execute(
            "INSERT INTO wps_sync_targets "
            "(target_id, binding_id, site_id, business_key, target_code, target_type, "
            "credential_id, target_name, document_open_url, webhook_url, "
            "expected_document_id, created_at, updated_at) "
            "VALUES ('retired-smart-target', 'retired-binding', ?, ?, "
            "'wps_smart_sheet', 'WPS_SMART_SHEET', ?, 'retired', "
            "'https://example.test/retired', 'https://example.test/retired-hook', "
            "'retired', 'old', 'old')",
            (repository.site_id, TRACKSIDE_AP_WPS_BUSINESS_KEY, credential_id),
        )
        connection.commit()


def test_wps_removed_smart_target_migration_preserves_standard_configuration_and_history(
    tmp_path: Path,
) -> None:
    paths = PathResolver(tmp_path)
    repository = WpsSyncRepository(paths, "hangzhou10", protect=_protect, unprotect=_protect)
    standard = repository.upsert_target(
        business_key=TRACKSIDE_AP_WPS_BUSINESS_KEY,
        target_code=STANDARD_TARGET_CODE,
        target_type=WpsTargetType.STANDARD_SPREADSHEET,
        target_name="普通表格",
        document_open_url="https://example.test/standard",
        webhook_url="https://example.test/standard-hook",
        expected_document_id="standard",
        token="secret-token",
        credential_id="standard-credential",
    )
    _insert_retired_smart_target(
        repository,
        credential_id="retired-credential",
        create_credential=True,
    )
    with sqlite3.connect(repository.path) as connection:
        connection.execute(
            "ALTER TABLE wps_sync_targets ADD COLUMN "
            "sheet_order_probe_diagnostic TEXT NOT NULL DEFAULT ''"
        )
        for batch_id in ("mixed-batch", "retired-only-batch"):
            connection.execute(
                "INSERT INTO wps_sync_batches "
                "(batch_id, site_id, business_key, snapshot_revision, snapshot_sha256, "
                "snapshot_generated_at, requested_at, completed_at, status, target_count, "
                "success_target_count, failed_target_count, result_summary) "
                "VALUES (?, ?, ?, 'revision', 'sha', 'old', 'old', 'old', "
                "'PARTIAL_SUCCESS', 2, 1, 1, ?)",
                (
                    batch_id,
                    repository.site_id,
                    TRACKSIDE_AP_WPS_BUSINESS_KEY,
                    json.dumps(
                        {
                            "status": "PARTIAL_SUCCESS",
                            "target_count": 2,
                            "targets": [
                                {"target_code": STANDARD_TARGET_CODE, "status": "SUCCESS"},
                                {"target_code": "wps_smart_sheet", "status": "FAILED"},
                            ],
                        }
                    ),
                ),
            )
        connection.execute(
            "INSERT INTO wps_sync_target_runs "
            "(target_batch_id, batch_id, target_id, target_code, target_type, started_at, "
            "completed_at, status, result_summary) VALUES "
            "('standard-run', 'mixed-batch', ?, ?, 'WPS_STANDARD_SPREADSHEET', "
            "'old', 'old', 'SUCCESS', ?)",
            (
                standard.target_id,
                STANDARD_TARGET_CODE,
                json.dumps({"target_code": STANDARD_TARGET_CODE, "status": "SUCCESS"}),
            ),
        )
        for run_id, batch_id in (
            ("retired-mixed-run", "mixed-batch"),
            ("retired-only-run", "retired-only-batch"),
        ):
            connection.execute(
                "INSERT INTO wps_sync_target_runs "
                "(target_batch_id, batch_id, target_id, target_code, target_type, "
                "started_at, completed_at, status, result_summary) VALUES "
                "(?, ?, 'retired-smart-target', 'wps_smart_sheet', 'WPS_SMART_SHEET', "
                "'old', 'old', 'FAILED', ?)",
                (
                    run_id,
                    batch_id,
                    json.dumps({"target_code": "wps_smart_sheet", "status": "FAILED"}),
                ),
            )
        connection.commit()

    repository.initialize()
    repository.initialize()

    preserved = repository.get_target(
        TRACKSIDE_AP_WPS_BUSINESS_KEY, STANDARD_TARGET_CODE
    )
    assert preserved.target_id == standard.target_id
    assert preserved.document_open_url == "https://example.test/standard"
    assert preserved.webhook_url == "https://example.test/standard-hook"
    assert preserved.binding_id == standard.binding_id
    assert repository.resolve_token(preserved) == "secret-token"
    with sqlite3.connect(repository.path) as connection:
        assert connection.execute(
            "SELECT COUNT(*) FROM wps_sync_targets WHERE target_code = 'wps_smart_sheet'"
        ).fetchone()[0] == 0
        assert connection.execute(
            "SELECT COUNT(*) FROM wps_sync_target_runs WHERE target_type = 'WPS_SMART_SHEET'"
        ).fetchone()[0] == 0
        assert connection.execute(
            "SELECT COUNT(*) FROM wps_credentials WHERE credential_id = 'retired-credential'"
        ).fetchone()[0] == 0
        assert connection.execute(
            "SELECT COUNT(*) FROM wps_sync_batches WHERE batch_id = 'retired-only-batch'"
        ).fetchone()[0] == 0
        batch = connection.execute(
            "SELECT status, target_count, success_target_count, failed_target_count, "
            "result_summary FROM wps_sync_batches WHERE batch_id = 'mixed-batch'"
        ).fetchone()
        columns = {
            str(row[1]) for row in connection.execute("PRAGMA table_info(wps_sync_targets)")
        }
    assert batch[:4] == ("SUCCESS", 1, 1, 0)
    assert "wps_smart_sheet" not in str(json.loads(batch[4]))
    assert "sheet_order_probe_diagnostic" not in columns


def test_wps_removed_smart_target_migration_keeps_shared_standard_credential(
    tmp_path: Path,
) -> None:
    repository = WpsSyncRepository(
        PathResolver(tmp_path), "hangzhou10", protect=_protect, unprotect=_protect
    )
    standard = repository.upsert_target(
        business_key=TRACKSIDE_AP_WPS_BUSINESS_KEY,
        target_code=STANDARD_TARGET_CODE,
        target_type=WpsTargetType.STANDARD_SPREADSHEET,
        target_name="WPS 云文档",
        document_open_url="https://example.test/standard",
        webhook_url="https://example.test/standard-hook",
        expected_document_id="standard",
        token="shared-token",
        credential_id="shared-credential",
    )
    _insert_retired_smart_target(
        repository,
        credential_id=standard.credential_id,
        create_credential=False,
    )

    repository.initialize()

    preserved = repository.get_target(
        TRACKSIDE_AP_WPS_BUSINESS_KEY, STANDARD_TARGET_CODE
    )
    assert repository.resolve_token(preserved) == "shared-token"
    with sqlite3.connect(repository.path) as connection:
        assert connection.execute(
            "SELECT COUNT(*) FROM wps_credentials WHERE credential_id = 'shared-credential'"
        ).fetchone()[0] == 1


def test_retired_wps_product_surface_has_no_implementation_references() -> None:
    root = Path(__file__).parents[1]
    migration = root / "src" / "netconsole" / "repositories" / "wps_sync_repository.py"
    product_files = [
        path
        for base in (root / "src", root / "apps" / "web" / "src", root / "tools")
        for path in base.rglob("*")
        if path.is_file() and path.suffix in {".py", ".ts", ".vue", ".js"}
    ]
    retired_tokens = {
        "WPS_SMART_SHEET",
        "wps_smart_sheet",
        "SmartSheetDTO",
        "SmartFieldDTO",
        "SmartRecordDTO",
        "trackside_ap_smart_sheet_sync.js",
        "trackside_ap_smart_sheet_connection_probe.js",
    }
    for path in product_files:
        if path == migration:
            continue
        source = path.read_text(encoding="utf-8")
        assert not retired_tokens.intersection(source.split()), path
        for token in retired_tokens:
            assert token not in source, f"{path}: {token}"
    for base in (root / "src", root / "apps" / "web" / "src"):
        for path in base.rglob("*"):
            if not path.is_file() or path.suffix not in {".py", ".ts", ".vue"}:
                continue
            if path == migration:
                continue
            source = path.read_text(encoding="utf-8")
            assert "sheet-order-probe" not in source, path
            assert "sheet_order_probe_diagnostic" not in source, path


def test_wps_binding_id_is_stable_across_independent_data_roots(tmp_path: Path) -> None:
    targets = []
    for root in (tmp_path / "first", tmp_path / "restored"):
        repository = WpsSyncRepository(
            PathResolver(data_root=root), "hzl10", protect=_protect, unprotect=_protect
        )
        targets.append(
            repository.upsert_target(
                business_key=TRACKSIDE_AP_WPS_BUSINESS_KEY,
                target_code=STANDARD_TARGET_CODE,
                target_type=WpsTargetType.STANDARD_SPREADSHEET,
                target_name="普通表格",
                document_open_url="https://example.test/standard",
                webhook_url="https://example.test/standard-hook",
                expected_document_id="standard",
            )
        )

    assert targets[0].target_id != targets[1].target_id
    assert targets[0].binding_id == targets[1].binding_id
    assert targets[0].binding_id == build_wps_binding_id(
        "hzl10", TRACKSIDE_AP_WPS_BUSINESS_KEY, STANDARD_TARGET_CODE
    )
    assert build_wps_binding_id(
        "other-site", TRACKSIDE_AP_WPS_BUSINESS_KEY, STANDARD_TARGET_CODE
    ) != targets[0].binding_id


def test_wps_binding_id_migration_is_additive_repeatable_and_preserves_rows(
    tmp_path: Path,
) -> None:
    repository = WpsSyncRepository(
        PathResolver(tmp_path), "hzl10", protect=_protect, unprotect=_protect
    )
    original = repository.upsert_target(
        business_key=TRACKSIDE_AP_WPS_BUSINESS_KEY,
        target_code=STANDARD_TARGET_CODE,
        target_type=WpsTargetType.STANDARD_SPREADSHEET,
        target_name="保留的普通表格",
        document_open_url="https://example.test/standard",
        webhook_url="https://example.test/standard-hook",
        expected_document_id="standard",
        token="preserved-token",
    )
    with sqlite3.connect(repository.path) as connection:
        connection.execute("ALTER TABLE wps_sync_targets DROP COLUMN binding_id")
        connection.commit()

    repository.initialize()
    repository.initialize()

    migrated = repository.get_target(TRACKSIDE_AP_WPS_BUSINESS_KEY, STANDARD_TARGET_CODE)
    with sqlite3.connect(repository.path) as connection:
        columns = {
            str(row[1])
            for row in connection.execute("PRAGMA table_info(wps_sync_targets)")
        }
        row_count = connection.execute(
            "SELECT COUNT(*) FROM wps_sync_targets WHERE target_id = ?",
            (original.target_id,),
        ).fetchone()[0]
    assert "binding_id" in columns
    assert "column_width_probe_diagnostic" in columns
    assert row_count == 1
    assert migrated.target_id == original.target_id
    assert migrated.target_name == "保留的普通表格"
    assert migrated.binding_id == build_wps_binding_id(
        "hzl10", TRACKSIDE_AP_WPS_BUSINESS_KEY, STANDARD_TARGET_CODE
    )
    assert repository.resolve_token(migrated) == "preserved-token"


def test_wps_target_run_migration_adds_async_recovery_columns_repeatably(
    tmp_path: Path,
) -> None:
    repository = WpsSyncRepository(
        PathResolver(tmp_path), "hzl10", protect=_protect, unprotect=_protect
    )
    repository.initialize()
    with sqlite3.connect(repository.path) as connection:
        for column in (
            "remote_task_id",
            "remote_task_type",
            "remote_task_status",
            "remote_task_submitted_at",
            "remote_task_last_polled_at",
            "remote_task_finished_at",
            "request_payload_json",
            "source_format_manifest_json",
        ):
            connection.execute(f"ALTER TABLE wps_sync_target_runs DROP COLUMN {column}")
        connection.commit()

    repository.initialize()
    repository.initialize()

    with sqlite3.connect(repository.path) as connection:
        columns = {
            str(row[1])
            for row in connection.execute("PRAGMA table_info(wps_sync_target_runs)")
        }
    assert {
        "remote_task_id",
        "remote_task_type",
        "remote_task_status",
        "remote_task_submitted_at",
        "remote_task_last_polled_at",
        "remote_task_finished_at",
        "request_payload_json",
        "source_format_manifest_json",
    } <= columns


def test_wps_webhook_parser_derives_sync_submit_and_poll_endpoints() -> None:
    endpoints = parse_wps_webhook(
        f"https://www.kdocs.cn/api/v3/ide/file/{TEST_WEBHOOK_DOCUMENT_ID}/"
        "script/V2-2o35ebQ25Bb3Uyrnii2U3o/sync_task"
    )

    assert endpoints.host == "www.kdocs.cn"
    assert endpoints.file_id == TEST_WEBHOOK_DOCUMENT_ID
    assert endpoints.script_id == "V2-2o35ebQ25Bb3Uyrnii2U3o"
    assert endpoints.sync_task_url.endswith("/sync_task")
    assert endpoints.async_task_url.endswith("/task")
    assert endpoints.task_status_url == "https://www.kdocs.cn/api/v3/script/task"


def test_wps_target_configuration_rejects_non_kdocs_webhook(tmp_path: Path) -> None:
    service = TracksideApWpsSyncService(PathResolver(tmp_path))
    with pytest.raises(WpsSyncError, match="kdocs.cn"):
        service.configure_target(
            "hangzhou10",
            STANDARD_TARGET_CODE,
            webhook_url="https://localhost/api/sync_task",
        )


def test_wps_public_targets_expose_only_cloud_document_deployment_identity(
    tmp_path: Path,
) -> None:
    targets = TracksideApWpsSyncService(PathResolver(tmp_path)).list_targets("hangzhou10")
    by_code = {target["target_code"]: target for target in targets}

    assert list(by_code) == [STANDARD_TARGET_CODE]
    assert by_code[STANDARD_TARGET_CODE]["expected_script_version"] == WPS_SCRIPT_VERSIONS[STANDARD_TARGET_CODE]
    assert by_code[STANDARD_TARGET_CODE]["expected_deployment_id"] == WPS_DEPLOYMENT_IDS[STANDARD_TARGET_CODE]


def test_wps_configuration_isolated_per_site_and_rejects_shared_document_script(
    tmp_path: Path,
) -> None:
    paths = PathResolver(tmp_path)
    service = TracksideApWpsSyncService(paths)
    hangzhou = "hangzhou10"
    ningbo10 = "ningbo10"
    ningbo = "ningbo12"
    hangzhou_webhook = (
        "https://www.kdocs.cn/api/v3/ide/file/hangzhou-document/"
        "script/hangzhou-script/sync_task"
    )
    ningbo_webhook = (
        "https://www.kdocs.cn/api/v3/ide/file/ningbo-document/"
        "script/ningbo-script/sync_task"
    )

    service.configure_target(
        hangzhou,
        STANDARD_TARGET_CODE,
        token="hangzhou-token",
        document_open_url="https://www.kdocs.cn/l/hangzhou-document",
        webhook_url=hangzhou_webhook,
    )
    unconfigured_ningbo = service.list_targets(ningbo)[0]
    unconfigured_ningbo10 = service.list_targets(ningbo10)[0]

    assert unconfigured_ningbo["document_open_url"] == ""
    assert unconfigured_ningbo["webhook_url"] == ""
    assert unconfigured_ningbo["expected_document_id"] == ""
    assert unconfigured_ningbo["token_configured"] is False
    assert unconfigured_ningbo10["document_open_url"] == ""
    assert unconfigured_ningbo10["webhook_url"] == ""
    assert unconfigured_ningbo10["token_configured"] is False

    with pytest.raises(WpsSyncError) as conflict:
        service.configure_target(
            ningbo,
            STANDARD_TARGET_CODE,
            document_open_url="https://www.kdocs.cn/l/hangzhou-document",
            webhook_url=hangzhou_webhook,
        )
    assert conflict.value.code == "WPS_DOCUMENT_SITE_CONFLICT"

    service.configure_target(
        ningbo,
        STANDARD_TARGET_CODE,
        token="ningbo-token",
        document_open_url="https://www.kdocs.cn/l/ningbo-document",
        webhook_url=ningbo_webhook,
    )
    hangzhou_target = service.list_targets(hangzhou)[0]
    ningbo_target = service.list_targets(ningbo)[0]
    hangzhou_credential = service._repository(hangzhou).get_target(
        TRACKSIDE_AP_WPS_BUSINESS_KEY, STANDARD_TARGET_CODE
    ).credential_id
    ningbo_credential = service._repository(ningbo).get_target(
        TRACKSIDE_AP_WPS_BUSINESS_KEY, STANDARD_TARGET_CODE
    ).credential_id

    assert hangzhou_target["document_open_url"] == "https://www.kdocs.cn/l/hangzhou-document"
    assert ningbo_target["document_open_url"] == "https://www.kdocs.cn/l/ningbo-document"
    assert hangzhou_credential != ningbo_credential
    assert len({
        paths.site_sync_dir(site) / "wps_sync.sqlite"
        for site in (hangzhou, ningbo10, ningbo)
    }) == 3


def test_wps_targets_reject_a_stale_site_context(tmp_path: Path) -> None:
    from types import SimpleNamespace

    from fastapi import HTTPException

    request = SimpleNamespace(
        app=SimpleNamespace(
            state=SimpleNamespace(
                paths=PathResolver(tmp_path),
                trackside_ap_business_query_service=SimpleNamespace(
                    current_site_id=lambda: "hangzhou10"
                ),
            )
        )
    )

    assert _site_id(request, "hangzhou10") == "hangzhou10"
    with pytest.raises(HTTPException) as mismatch:
        _site_id(request, "ningbo12")
    assert mismatch.value.status_code == 409
    assert mismatch.value.detail["code"] == "WPS_SITE_CONTEXT_MISMATCH"


def test_wps_service_rejects_every_nonstandard_target_code(tmp_path: Path) -> None:
    service = TracksideApWpsSyncService(PathResolver(tmp_path))

    with pytest.raises(WpsSyncError) as configured:
        service.configure_target("hangzhou10", "unsupported-target")
    with pytest.raises(WpsSyncError) as synced:
        service.sync("hangzhou10", target_codes=("unsupported-target",))

    assert configured.value.code == "WPS_TARGET_UNSUPPORTED"
    assert synced.value.code == "WPS_TARGET_UNSUPPORTED"


def test_wps_public_target_does_not_trust_stale_bound_state_after_stable_id_migration(
    tmp_path: Path,
) -> None:
    service = TracksideApWpsSyncService(PathResolver(tmp_path))
    service.list_targets("hzl10")
    repository = service._repository("hzl10")
    target = repository.get_target(TRACKSIDE_AP_WPS_BUSINESS_KEY, STANDARD_TARGET_CODE)
    repository.update_target_remote_state(
        target.target_id,
        binding_status="BOUND",
        result={
            "remote_binding_id": f"wst_{'a' * 32}",
            "remote_site_id": target.site_id,
            "remote_business_key": target.business_key,
        },
        persist_runtime_identity=False,
    )

    refreshed = {item["target_code"]: item for item in service.list_targets("hzl10")}

    assert refreshed[STANDARD_TARGET_CODE]["binding_status"] == "UNKNOWN"
    assert refreshed[STANDARD_TARGET_CODE]["remote_binding_id"].startswith("wst_")


def test_wps_default_target_name_uses_site_display_name_and_preserves_custom_name(tmp_path: Path) -> None:
    paths = PathResolver(tmp_path)
    manager = SiteManager(paths)
    manager.create_site("hzl10", display_name="杭州地铁10号线")
    manager.create_site("nbl12", display_name="宁波地铁12号线")
    service = TracksideApWpsSyncService(paths)

    hzl10 = {item["target_code"]: item for item in service.list_targets("hzl10")}
    nbl12 = {item["target_code"]: item for item in service.list_targets("nbl12")}
    assert hzl10[STANDARD_TARGET_CODE]["target_name"] == "杭州地铁10号线轨旁AP业务-WPS云文档"
    assert nbl12[STANDARD_TARGET_CODE]["target_name"] == "宁波地铁12号线轨旁AP业务-WPS云文档"

    service.configure_target("hzl10", STANDARD_TARGET_CODE, document_open_url="https://www.kdocs.cn/l/custom")
    repository = service._repository("hzl10")
    target = repository.get_target(TRACKSIDE_AP_WPS_BUSINESS_KEY, STANDARD_TARGET_CODE)
    repository.upsert_target(
        business_key=target.business_key,
        target_code=target.target_code,
        target_type=target.target_type,
        target_name="用户自定义 WPS 文档",
        document_open_url=target.document_open_url,
        webhook_url=target.webhook_url,
        expected_document_id=target.expected_document_id,
        credential_id=target.credential_id,
    )
    refreshed = {item["target_code"]: item for item in service.list_targets("hzl10")}
    assert refreshed[STANDARD_TARGET_CODE]["target_name"] == "用户自定义 WPS 文档"


@pytest.mark.parametrize(
    ("override", "expected_code"),
    [
        ({"script_version": "old-standard"}, "WPS_SCRIPT_VERSION_MISMATCH"),
        ({"deployment_id": "stale-deployment"}, "WPS_DEPLOYMENT_ID_MISMATCH"),
        ({"target_code": "another_target"}, "WPS_TARGET_CODE_MISMATCH"),
    ],
)
def test_wps_connection_test_rejects_stale_or_cross_target_script_identity(
    override: dict[str, str],
    expected_code: str,
) -> None:
    body = {
        "success": True,
        "protocol_version": 2,
        "script_version": WPS_SCRIPT_VERSIONS[STANDARD_TARGET_CODE],
        "deployment_id": WPS_DEPLOYMENT_IDS[STANDARD_TARGET_CODE],
        "target_type": "WPS_STANDARD_SPREADSHEET",
        "target_code": STANDARD_TARGET_CODE,
        "document_id": "document",
        "runtime_capability": "DEPLOYMENT_PENDING",
        **override,
    }

    class FakeClient:
        def post(self, target, *, token, argv):
            return WpsHttpResponse(status_code=200, body=body)

    with pytest.raises(WpsSyncError) as captured:
        WpsStandardSpreadsheetAdapter(FakeClient()).connection_test(
            _wps_target(), "test-only-token"
        )

    assert captured.value.code == expected_code
    assert captured.value.details["phase"] == "PROTOCOL_HANDSHAKE"


def test_wps_connection_test_reports_remote_document_identity_mismatch() -> None:
    target = replace(_wps_target(), expected_document_id="536585421042")

    class FakeClient:
        def post(self, target, *, token, argv):
            return WpsHttpResponse(
                status_code=200,
                body={
                    "success": True,
                    "protocol_version": 2,
                    "script_version": WPS_SCRIPT_VERSIONS[STANDARD_TARGET_CODE],
                    "deployment_id": WPS_DEPLOYMENT_IDS[STANDARD_TARGET_CODE],
                    "target_type": "WPS_STANDARD_SPREADSHEET",
                    "target_code": STANDARD_TARGET_CODE,
                    "document_id": STALE_REMOTE_DOCUMENT_ID,
                    "runtime_capability": "DEPLOYMENT_PENDING",
                },
            )

    with pytest.raises(WpsSyncError) as captured:
        WpsStandardSpreadsheetAdapter(FakeClient()).connection_test(
            target, "test-only-token"
        )

    assert captured.value.code == "WPS_DOCUMENT_IDENTITY_MISMATCH"
    assert captured.value.details["expected_document_id"] == "536585421042"
    assert captured.value.details["remote_document_id"] == STALE_REMOTE_DOCUMENT_ID
    assert "WPS_DOCUMENT_IDENTITY_MISMATCH" in str(captured.value)
    assert "预期文档 ID：536585421042" in str(captured.value)
    assert f"远端脚本声明：{STALE_REMOTE_DOCUMENT_ID}" in str(captured.value)
    assert "重新复制脚本并全量替换" in captured.value.details["suggestion"]


@pytest.mark.parametrize(
    "body",
    [
        {
            "status": "finished",
            "error": "",
            "data": {
                "result": json.dumps({
                    "success": True,
                    "protocol_version": 2,
                    "target_type": "WPS_STANDARD_SPREADSHEET",
                    "document_id": "document",
                }),
            },
        },
        {
            "status": "finished",
            "error": "",
            "data": {
                "result": {
                    "success": True,
                    "protocol_version": 2,
                    "target_type": "WPS_STANDARD_SPREADSHEET",
                    "document_id": "document",
                },
            },
        },
    ],
)
def test_wps_connection_test_unwraps_official_sync_task_envelope(body: dict[str, object]) -> None:
    class FakeClient:
        def post(self, target, *, token, argv):
            assert token == "test-only-token"
            return WpsHttpResponse(status_code=200, body=body)

    result = WpsStandardSpreadsheetAdapter(FakeClient()).connection_test(_wps_target(), "test-only-token")
    assert result["phase"] == "SUCCESS"
    assert result["document_id"] == "document"


def test_wps_connection_test_classifies_only_old_random_binding_as_legacy() -> None:
    target = _wps_target()
    legacy_binding_id = f"wst_{'a' * 32}"

    class FakeClient:
        def post(self, target, *, token, argv):
            return WpsHttpResponse(
                status_code=200,
                body={
                    "success": True,
                    "protocol_version": 2,
                    "target_type": target.target_type.value,
                    "target_code": target.target_code,
                    "document_id": target.expected_document_id,
                    "binding_id": legacy_binding_id,
                    "site_id": target.site_id,
                    "business_key": target.business_key,
                },
            )

    result = WpsStandardSpreadsheetAdapter(FakeClient()).connection_test(
        target, "test-only-token"
    )

    assert result["binding_status"] == "LEGACY_BINDING_ID_MISMATCH"
    assert result["local_binding_id"] == target.binding_id
    assert result["remote_binding_id"] == legacy_binding_id
    assert result["binding_id_match"] is False
    assert result["document_match"] is True
    assert result["document_identity_match"] is True
    assert result["site_match"] is True
    assert result["site_identity_match"] is True
    assert result["business_match"] is True
    assert result["business_identity_match"] is True
    assert result["target_match"] is True


def test_wps_connection_test_rejects_nonlegacy_binding_or_business_identity_as_mismatch() -> None:
    target = _wps_target()

    class FakeClient:
        def post(self, target, *, token, argv):
            return WpsHttpResponse(
                status_code=200,
                body={
                    "success": True,
                    "protocol_version": 2,
                    "target_type": target.target_type.value,
                    "target_code": target.target_code,
                    "document_id": target.expected_document_id,
                    "binding_id": "not-a-legacy-random-id",
                    "site_id": "other-site",
                    "business_key": target.business_key,
                },
            )

    result = WpsStandardSpreadsheetAdapter(FakeClient()).connection_test(
        target, "test-only-token"
    )

    assert result["binding_status"] == "MISMATCH"
    assert result["site_match"] is False


def test_wps_legacy_binding_migration_accepts_already_migrated_result() -> None:
    target = _wps_target()
    legacy_binding_id = f"wst_{'c' * 32}"

    class FakeClient:
        def post(self, actual_target, *, token, argv):
            assert actual_target is target
            assert argv["operation"] == "migrate_legacy_binding"
            assert argv["document_id"] == target.expected_document_id
            assert argv["expected_old_binding_id"] == legacy_binding_id
            assert argv["new_binding_id"] == target.binding_id
            return WpsHttpResponse(
                status_code=200,
                body={
                    "success": True,
                    "protocol_version": 2,
                    "target_type": target.target_type.value,
                    "target_code": target.target_code,
                    "document_id": target.expected_document_id,
                    "binding_status": "BOUND",
                    "binding_id": target.binding_id,
                    "remote_binding_id": target.binding_id,
                    "remote_document_id": target.expected_document_id,
                    "remote_site_id": target.site_id,
                    "remote_business_key": target.business_key,
                    "remote_target_code": target.target_code,
                    "remote_target_type": target.target_type.value,
                    "migrated": False,
                    "already_migrated": True,
                },
            )

    result = WpsStandardSpreadsheetAdapter(FakeClient()).migrate_legacy_binding(
        target,
        "test-only-token",
        expected_old_binding_id=legacy_binding_id,
    )

    assert result["binding_status"] == "BOUND"
    assert result["binding_id_match"] is True
    assert result["already_migrated"] is True


def test_wps_formal_sync_preserves_remote_business_error_before_success_identity_checks() -> None:
    target = _wps_target()

    class FakeClient:
        def post(self, target, *, token, argv):
            return WpsHttpResponse(
                status_code=200,
                body={
                    "success": False,
                    "protocol_version": 2,
                    "target_type": target.target_type.value,
                    "document_id": target.expected_document_id,
                    "error_code": "WPS_DOCUMENT_BINDING_MISMATCH",
                    "message": "远端文档绑定与当前请求不一致",
                    "binding_status": "MISMATCH",
                    "failed_sheet": "_NetConsoleSyncMeta",
                    "failed_operation": "ASSERT_BINDING",
                },
            )

    with pytest.raises(WpsSyncError) as captured:
        WpsStandardSpreadsheetAdapter(FakeClient()).sync(
            target,
            "test-only-token",
            {
                "target_batch_id": "batch-expected",
                "site_id": target.site_id,
                "business_key": target.business_key,
                "snapshot_revision": "revision-1",
                "snapshot_sha256": "sha-1",
            },
        )

    assert captured.value.code == "WPS_DOCUMENT_BINDING_MISMATCH"
    assert captured.value.code != "WPS_RESPONSE_IDENTITY_MISMATCH"
    assert captured.value.details["binding_status"] == "MISMATCH"
    assert captured.value.details["failed_operation"] == "ASSERT_BINDING"


def test_wps_success_identity_mismatch_reports_expected_and_remote_values() -> None:
    target = _wps_target()

    class FakeClient:
        def post(self, target, *, token, argv):
            return WpsHttpResponse(
                status_code=200,
                body={
                    "success": True,
                    "protocol_version": 2,
                    "target_type": target.target_type.value,
                    "document_id": target.expected_document_id,
                    "target_batch_id": "batch-remote",
                    "site_id": target.site_id,
                    "business_key": target.business_key,
                    "snapshot_revision": "revision-remote",
                    "snapshot_sha256": "sha-remote",
                },
            )

    with pytest.raises(WpsSyncError) as captured:
        WpsStandardSpreadsheetAdapter(FakeClient()).sync(
            target,
            "test-only-token",
            {
                "target_batch_id": "batch-expected",
                "site_id": target.site_id,
                "business_key": target.business_key,
                "snapshot_revision": "revision-expected",
                "snapshot_sha256": "sha-expected",
            },
        )

    assert captured.value.code == "WPS_RESPONSE_IDENTITY_MISMATCH"
    assert captured.value.details["expected_target_batch_id"] == "batch-expected"
    assert captured.value.details["remote_target_batch_id"] == "batch-remote"
    assert captured.value.details["expected_revision"] == "revision-expected"
    assert captured.value.details["remote_revision"] == "revision-remote"


@pytest.mark.parametrize(
    ("body", "code"),
    [
        ({"status": "finished", "error": "", "data": {"result": None}}, "WPS_SCRIPT_RESULT_EMPTY"),
        ({"status": "finished", "error": "", "data": {"result": "not-json"}}, "WPS_SCRIPT_RESULT_INVALID"),
        ({"status": "running", "error": "", "data": {}}, "WPS_SCRIPT_STATUS_INVALID"),
        ({"status": "finished", "error": "permission denied", "error_details": {"name": "Forbidden"}}, "WPS_SCRIPT_EXECUTION_FAILED"),
    ],
)
def test_wps_connection_test_reports_envelope_failures(body: dict[str, object], code: str) -> None:
    class FakeClient:
        def post(self, target, *, token, argv):
            return WpsHttpResponse(status_code=200, body=body)

    with pytest.raises(WpsSyncError) as captured:
        WpsStandardSpreadsheetAdapter(FakeClient()).connection_test(_wps_target(), "test-only-token")
    assert captured.value.code == code
    assert captured.value.details["phase"] == "SCRIPT_EXECUTION"


@pytest.mark.parametrize(
    ("status", "body", "expected_code"),
    [
        (403, b'{"code":"DOCUMENT_PERMISSION_DENIED","message":"no access"}', "WPS_DOCUMENT_PERMISSION_DENIED"),
        (403, b"token expired", "WPS_TOKEN_INVALID"),
        (403, b"script is nil", "WPS_SCRIPT_NOT_AVAILABLE"),
        (403, b"<html><title>Forbidden</title></html>", "WPS_REMOTE_FORBIDDEN"),
        (401, b'{"error_code":"TOKEN_INVALID","msg":"invalid token"}', "WPS_TOKEN_INVALID"),
        (404, b"not found", "WPS_WEBHOOK_NOT_FOUND"),
        (429, b"rate limited", "WPS_REMOTE_RATE_LIMITED"),
    ],
)
def test_wps_http_errors_keep_safe_remote_diagnostics(
    monkeypatch: pytest.MonkeyPatch,
    status: int,
    body: bytes,
    expected_code: str,
) -> None:
    token = "test-only-token"

    def raise_http_error(request, *, timeout):
        raise urllib.error.HTTPError(
            request.full_url,
            status,
            "remote failure",
            {"Content-Type": "application/json"},
            io.BytesIO(body),
        )

    monkeypatch.setattr("urllib.request.urlopen", raise_http_error)
    with pytest.raises(WpsSyncError) as captured:
        WpsAirScriptClient().post(_wps_target(), token=token, argv={"operation": "connection_test"})

    error = captured.value
    assert error.code == expected_code
    assert error.details["phase"] == "HTTP_AUTH"
    assert error.details["http_status"] == status
    assert token not in str(error)
    assert token not in str(error.details)
    if status == 403:
        assert "HTTP 403" in str(error)
        assert "建议" in str(error)


def test_wps_http_response_and_request_headers_are_bounded(monkeypatch: pytest.MonkeyPatch) -> None:
    target = _wps_target()
    observed: dict[str, str] = {}

    class FakeResponse:
        status = 200

        def __enter__(self):
            return self

        def __exit__(self, *args):
            return False

        def read(self, size):
            observed["read_size"] = str(size)
            return json.dumps({"success": True}).encode("utf-8")

    def open_request(request, *, timeout):
        observed.update({key.lower(): value for key, value in request.headers.items()})
        observed["timeout"] = str(timeout)
        return FakeResponse()

    monkeypatch.setattr("urllib.request.urlopen", open_request)
    response = WpsAirScriptClient().post(target, token="test-only-token", argv={})
    assert response.status_code == 200
    assert observed["content-type"] == "application/json"
    assert observed["accept"] == "application/json"
    assert observed["user-agent"].startswith("NetConsole/")
    assert int(observed["read_size"]) == 20 * 1024 * 1024 + 1


def test_wps_async_client_submits_and_url_encodes_task_id(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    observed: list[tuple[str, str, bytes | None, str]] = []
    responses = [
        {
            "data": {"task_id": "GN/KU3+task=="},
            "task_id": "GN/KU3+task==",
            "task_type": "open_air_script",
        },
        {"status": "running", "error": "", "data": {"result": None, "logs": []}},
    ]

    class FakeResponse:
        status = 200

        def __init__(self, body: dict[str, object]) -> None:
            self.body = body

        def __enter__(self):
            return self

        def __exit__(self, *args):
            return False

        def read(self, size):
            return json.dumps(self.body).encode("utf-8")

    def open_request(request, *, timeout):
        observed.append(
            (
                request.method,
                request.full_url,
                request.data,
                request.headers.get("Airscript-token", ""),
            )
        )
        return FakeResponse(responses.pop(0))

    monkeypatch.setattr("urllib.request.urlopen", open_request)
    client = WpsAirScriptClient()
    remote = client.submit_async(
        _wps_target(),
        token="test-only-token",
        argv={"operation": "sync_trackside_ap_business"},
    )
    polled = client.poll_async_task(
        _wps_target(),
        token="test-only-token",
        task_id=remote.task_id,
    )

    assert remote == WpsRemoteTask(
        task_id="GN/KU3+task==",
        task_type="open_air_script",
    )
    assert observed[0][0] == "POST"
    assert observed[0][1].endswith("/script/test/task")
    assert json.loads(observed[0][2].decode("utf-8"))["Context"]["argv"]["operation"] == "sync_trackside_ap_business"
    assert observed[1][0] == "GET"
    assert observed[1][1].endswith(
        "/api/v3/script/task?task_id=GN%2FKU3%2Btask%3D%3D"
    )
    assert observed[0][3] == observed[1][3] == "test-only-token"
    assert polled.body == {
        "status": "running",
        "error": "",
        "data": {"result": None, "logs": []},
    }


def test_workbook_dto_uses_prepend_mode_for_overview(tmp_path: Path) -> None:
    path = tmp_path / "sample.xlsx"
    workbook = Workbook()
    workbook.active.title = "轨旁AP业务"
    workbook.active.append(["站点", "AP"])
    workbook.active.append(["A", "AP-1"])
    overview = workbook.create_sheet("AP上线情况概览")
    overview.append(["同步时间", "上线率"])
    overview.append(["2026-08-07 10:00:00", 1])
    workbook.save(path)
    workbook.close()

    dto = workbook_dto_from_xlsx(path)
    assert [sheet.sheet_name for sheet in dto.sheets] == ["轨旁AP业务", "AP上线情况概览"]
    assert dto.sheets[0].sync_mode.value == "FULL_REPLACE"
    assert dto.sheets[1].sync_mode.value == "PREPEND_SNAPSHOT"
    assert dto.sheets[1].logical_sheet_key == "ap_online_history_overview"
    assert dto.sheets[1].tab_color == ""
    assert dto.sheets[1].row_count == 2
    assert [sheet.sheet_order for sheet in dto.sheets] == [0, 1]


def test_workbook_dto_preserves_sheet_order_and_compresses_format_runs(
    tmp_path: Path,
) -> None:
    path = tmp_path / "styled.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "轨旁AP业务"
    sheet.append(["站点", "上线率", "备注"])
    sheet.append(["A", 0.5, "第一行"])
    sheet.append(["B", 0.75, "第二行"])
    header_fill = PatternFill(fill_type="solid", fgColor="FF4472C4")
    thin = Side(style="thin", color="FFD1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in sheet[1]:
        cell.font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in sheet.iter_rows(min_row=2, max_row=3, min_col=1, max_col=3):
        for cell in row:
            cell.font = Font(name="Microsoft YaHei", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
    for row in (2, 3):
        sheet.cell(row=row, column=2).number_format = "0.00%"
    sheet.merge_cells("A4:C4")
    sheet["A4"] = "合并说明"
    sheet.column_dimensions["A"].width = 18
    sheet.row_dimensions[1].height = 24
    sheet.freeze_panes = "B2"
    sheet.sheet_properties.tabColor = "FF70AD47"
    hidden = workbook.create_sheet("隐藏业务页")
    hidden.sheet_state = "hidden"
    workbook.save(path)
    workbook.close()

    dto = workbook_dto_from_xlsx(path, include_format_mirror=True)
    first, second = dto.sheets
    assert first.sheet_order == 0
    assert second.sheet_order == 1
    assert first.sheet_visible is True
    assert second.sheet_visible is False
    assert first.tab_color == "#70AD47"
    assert first.merges == ["A4:C4"]
    assert first.column_widths["A"] == 18
    assert first.row_heights["1"] == 24
    assert first.auto_fit_columns == ("A", "B", "C")
    assert first.column_layouts["A"]["layout_type"] == "normal"
    assert first.auto_fit_rows is True
    assert first.freeze_mode is WpsFreezeMode.FIRST_ROW_ONLY
    assert first.auto_filter == ""
    assert {sample["label"] for sample in first.verification_samples} >= {
        "header",
        "first_data",
        "last_data",
    }
    first_data_sample = next(
        sample for sample in first.verification_samples if "first_data" in sample["label"]
    )
    percentage_sample = next(
        item for item in first_data_sample["format_cells"] if item["range"] == "B2"
    )
    assert percentage_sample["expected_display_text"] == "50.00%"
    header_font_run = next(
        run for run in first.format_runs if run.range == "A1:C1" and run.font
    )
    header_fill_run = next(
        run for run in first.format_runs if run.range == "A1:C1" and run.fill
    )
    header_alignment_run = next(
        run for run in first.format_runs if run.range == "A1:C1" and run.alignment
    )
    percentage_run = next(
        run
        for run in first.format_runs
        if run.range == "B2:B3" and run.number_format == "0.00%"
    )
    border_run = next(run for run in first.format_runs if run.border)
    assert header_font_run.font == {
        "name": "Microsoft YaHei",
        "size": 11.0,
        "bold": True,
        "italic": False,
        "underline": "",
        "strike": False,
        "color": "#FFFFFF",
    }
    assert header_fill_run.fill["fg_color"] == "#4472C4"
    assert header_alignment_run.alignment["wrap_text"] is True
    assert percentage_run.number_format == "0.00%"
    assert border_run.border["left"] == {
        "style": "thin",
        "color": "#D1D5DB",
    }
    assert len(first.format_runs) < first.row_count * first.column_count * 5
    serialized = first.to_dict()
    assert "format_runs" in serialized
    assert "fonts" not in serialized
    assert "borders" not in serialized


def test_workbook_dto_emits_stable_layout_types_for_real_business_columns(
    tmp_path: Path,
) -> None:
    path = tmp_path / "column-layouts.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "轨旁AP业务"
    sheet.append([f"字段{index}" for index in range(1, 17)])
    sheet.append([f"值{index}" for index in range(1, 17)])
    for column in range(1, 17):
        sheet.column_dimensions[chr(64 + column)].width = 18
    workbook.save(path)
    workbook.close()

    dto = workbook_dto_from_xlsx(path, include_format_mirror=True).sheets[0]

    assert dto.auto_fit_columns == tuple(chr(64 + column) for column in range(1, 17))
    assert dto.column_layouts["G"]["layout_type"] == "identifier"
    assert dto.column_layouts["L"]["layout_type"] == "datetime"
    assert dto.column_layouts["P"] == {
        "layout_type": "long_text",
        "min_width": 16.0,
        "max_width": 48.0,
        "wrap_text": True,
    }


def test_workbook_dto_enforces_final_freeze_contract(tmp_path: Path) -> None:
    path = tmp_path / "freeze-contract.xlsx"
    workbook = Workbook()
    business = workbook.active
    business.title = "轨旁AP业务"
    business.append(["站点", "AP"])
    business.append(["A", "AP-1"])
    business.freeze_panes = "B2"
    overview = workbook.create_sheet("AP上线情况概览")
    overview.append(["日期", "上线率"])
    overview.freeze_panes = "B2"
    workbook.save(path)
    workbook.close()

    dto = workbook_dto_from_xlsx(path, include_format_mirror=True)
    assert [sheet.freeze_mode for sheet in dto.sheets] == [
        WpsFreezeMode.FIRST_ROW_ONLY,
        WpsFreezeMode.NONE,
    ]
    serialized_sheets = dto.to_dict()["sheets"]
    assert [sheet["freeze_mode"] for sheet in serialized_sheets] == [
        "FIRST_ROW_ONLY",
        "NONE",
    ]
    assert all("freeze_panes" not in sheet for sheet in serialized_sheets)


def test_workbook_dto_can_include_only_explicit_column_widths(tmp_path: Path) -> None:
    path = tmp_path / "dimensions.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "轨旁AP业务"
    sheet.append(["站点", "AP名称"])
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 32.5
    sheet.row_dimensions[1].height = 26
    sheet["A1"].fill = PatternFill(fill_type="solid", fgColor="FFC6EFCE")
    workbook.save(path)
    workbook.close()

    dto = workbook_dto_from_xlsx(path, include_column_widths=True)

    assert dto.sheets[0].column_widths == {"A": 18.0, "B": 32.5}
    assert dto.sheets[0].row_heights == {}
    assert dto.sheets[0].format_runs == ()
    assert dto.sheets[0].merges == []


def test_column_width_manifest_preserves_source_dto_widths_and_headers(
    tmp_path: Path,
) -> None:
    path = tmp_path / "column-width-manifest.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "轨旁AP业务"
    sheet.append(["归属站点", "室内交换机"])
    sheet.column_dimensions["A"].width = 22.55
    sheet.column_dimensions["B"].width = 40
    workbook.save(path)
    workbook.close()
    dto = workbook_dto_from_xlsx(path, include_column_widths=True)

    manifest = _source_workbook_format_manifest(path, dto)

    assert manifest["column_widths"] == [
        {
            "sheet_name": "轨旁AP业务",
            "column": "A",
            "range": "A:A",
            "column_label": "归属站点",
            "source_mode": "EXPLICIT",
            "local_workbook_width": 22.55,
            "sheet_dto_width": 22.55,
            "sheet_dto_mode": "EXPLICIT",
            "auto_fit_min_width": 8.0,
            "auto_fit_max_width": 60.0,
        },
        {
            "sheet_name": "轨旁AP业务",
            "column": "B",
            "range": "B:B",
            "column_label": "室内交换机",
            "source_mode": "EXPLICIT",
            "local_workbook_width": 40.0,
            "sheet_dto_width": 40.0,
            "sheet_dto_mode": "EXPLICIT",
            "auto_fit_min_width": 8.0,
            "auto_fit_max_width": 60.0,
        },
    ]
    assert manifest["totals"]["column_count"] == 2
    assert manifest["totals"]["explicit_width_count"] == 2


def test_column_width_verification_report_classifies_all_pipeline_stages() -> None:
    columns = ("A", "B", "C", "D")
    local_widths = {"A": 18.0, "B": 22.0, "C": 30.0, "D": 40.0}
    dto_widths = {"A": 17.0, "B": 22.0, "C": 30.0, "D": 40.0}
    payload_widths = {"A": 17.0, "B": 21.0, "C": 30.0, "D": 40.0}
    remote_widths = {"A": 17.0, "B": 21.0, "C": 28.0, "D": 40.0}
    manifest = [
        {
            "sheet_name": "轨旁AP业务",
            "column": column,
            "range": f"{column}:{column}",
            "column_label": column,
            "local_workbook_width": local_widths[column],
            "sheet_dto_width": dto_widths[column],
        }
        for column in columns
    ]
    request_payload = {
        "workbook": {
            "sheets": [
                {
                    "sheet_name": "轨旁AP业务",
                    "column_widths": payload_widths,
                }
            ]
        }
    }
    remote_result = {
        "column_width_result": {
            "attempted_count": 4,
            "items": [
                {
                    "sheet_name": "轨旁AP业务",
                    "column": column,
                    "requested_width": payload_widths[column],
                    "before_column_width": 8.43,
                    "remote_column_width": remote_widths[column],
                    "before_width_points": 59.01,
                    "remote_width_points": (
                        59.01 if column == "D" else remote_widths[column] * 7
                    ),
                    "physical_width_change_points": (
                        0.0 if column == "D" else remote_widths[column] * 7 - 59.01
                    ),
                    "read_back": True,
                }
                for column in columns
            ],
        }
    }

    report = _column_width_verification_report(
        manifest=manifest,
        request_payload=request_payload,
        remote_result=remote_result,
        enabled=True,
    )

    assert report["status"] == "FAILED"
    assert report["total_columns"] == 4
    assert report["dto_match_count"] == 3
    assert report["payload_match_count"] == 3
    assert report["attempted_count"] == 4
    assert report["read_back_count"] == 4
    assert report["physical_read_back_count"] == 4
    assert report["verified_count"] == 1
    assert report["warning_count"] == 1
    assert report["failed_count"] == 3
    assert report["stage_counts"] == {
        "WORKBOOK_DTO_WIDTH_MISMATCH": 1,
        "WPS_PAYLOAD_WIDTH_MISMATCH": 1,
        "WPS_COLUMN_WIDTH_APPLY_MISMATCH": 1,
        "WPS_COLUMN_WIDTH_VALUE_VERIFIED": 1,
    }
    assert report["largest_differences"][0]["column"] == "C"
    assert report["largest_differences"][0]["difference"] == 2.0
    assert [item["column"] for item in report["representative_columns"]] == ["A", "B", "C"]
    assert report["items"][-1]["physical_width_status"] == "APPLY_MISMATCH"


def test_column_width_report_failure_becomes_noncritical_format_warning() -> None:
    remote_result: dict[str, object] = {
        "format_results": {"column_width": {"status": "SUCCESS"}},
        "format_warnings": [],
    }
    report = {
        "status": "FAILED",
        "total_columns": 143,
        "attempted_count": 143,
        "read_back_count": 143,
        "physical_read_back_count": 143,
        "verified_count": 141,
        "warning_count": 0,
        "failed_count": 2,
        "verified_ratio": 0.986,
        "stage_counts": {
            "WPS_COLUMN_WIDTH_VALUE_VERIFIED": 141,
            "WPS_COLUMN_WIDTH_APPLY_MISMATCH": 2,
        },
        "largest_differences": [],
        "representative_columns": [],
    }

    _append_column_width_report_warning(remote_result, report)

    assert remote_result["format_warning_count"] == 1
    assert remote_result["format_warnings"] == [
        {
            "sheet_name": "_NetConsoleColumnWidths",
            "feature": "column_width_verification",
            "reason": "生产列宽自动验收：验证通过 141/143，告警 0，失败 2",
        }
    ]
    assert remote_result["format_results"]["column_width"]["status"] == "FAILED"


@pytest.mark.parametrize(
    ("color", "expected"),
    [
        (Color(rgb="00000000"), ""),
        (Color(rgb="00DBEAFE"), "#DBEAFE"),
        (Color(rgb="00DCFCE7"), "#DCFCE7"),
        (Color(rgb="00D1D5DB"), "#D1D5DB"),
        (Color(rgb="FFFFFFFF"), "#FFFFFF"),
        (Color(rgb="FFC6EFCE"), "#C6EFCE"),
        (Color(rgb="FFDDEBF7"), "#DDEBF7"),
        (Color(theme=1), ""),
        (Color(indexed=0), ""),
        (None, ""),
    ],
)
def test_openpyxl_color_only_returns_explicit_opaque_rgb(
    color: Color | None,
    expected: str,
) -> None:
    assert _openpyxl_color(color) == expected


def test_format_payload_does_not_turn_missing_or_transparent_fill_black() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "no fill"
    sheet["A2"] = "transparent"
    sheet["A2"].fill = PatternFill(fill_type="solid", fgColor="00000000")
    sheet["A3"] = "opaque"
    sheet["A3"].fill = PatternFill(fill_type="solid", fgColor="FFC6EFCE")

    assert "fill" not in _cell_format_payload(sheet["A1"])
    assert "fill" not in _cell_format_payload(sheet["A2"])
    assert _cell_format_payload(sheet["A3"])["fill"]["fg_color"] == "#C6EFCE"
    workbook.close()


def test_column_width_report_accepts_autofit_readback_with_clamp() -> None:
    report = _column_width_verification_report(
        manifest=[
            {
                "sheet_name": "轨旁AP业务",
                "column": "P",
                "range": "P:P",
                "column_label": "AP业务判定原因",
                "source_mode": "AUTO_FIT",
                "local_workbook_width": None,
                "sheet_dto_width": None,
                "sheet_dto_mode": "AUTO_FIT",
                "auto_fit_min_width": 8.0,
                "auto_fit_max_width": 40.0,
            }
        ],
        request_payload={
            "workbook": {
                "sheets": [
                    {
                        "sheet_name": "轨旁AP业务",
                        "column_widths": {},
                        "auto_fit_columns": ["P"],
                    }
                ]
            }
        },
        remote_result={
            "column_width_result": {
                "attempted_count": 1,
                "items": [
                    {
                        "sheet_name": "轨旁AP业务",
                        "column": "P",
                        "mode": "AUTO_FIT",
                        "applied": True,
                        "clamped": True,
                        "remote_column_width": 40.0,
                        "remote_width_points": 280.0,
                        "verified": True,
                    }
                ],
            }
        },
        enabled=True,
    )

    assert report["status"] == "SUCCESS"
    assert report["auto_fit_requested_count"] == 1
    assert report["auto_fit_applied_count"] == 1
    assert report["clamped_count"] == 1
    assert report["stage_counts"] == {"WPS_COLUMN_WIDTH_AUTOFIT_VERIFIED": 1}


def test_column_width_report_accepts_autofit_with_local_minimum() -> None:
    report = _column_width_verification_report(
        manifest=[
            {
                "sheet_name": "轨旁AP业务",
                "column": "P",
                "range": "P:P",
                "column_label": "AP业务判定原因",
                "source_mode": "AUTO_FIT_WITH_LOCAL_MIN",
                "local_workbook_width": 42.0,
                "sheet_dto_width": 42.0,
                "sheet_dto_mode": "AUTO_FIT_WITH_LOCAL_MIN",
                "layout_type": "long_text",
                "layout_min_width": 16.0,
                "layout_max_width": 48.0,
            }
        ],
        request_payload={
            "workbook": {
                "sheets": [
                    {
                        "sheet_name": "轨旁AP业务",
                        "column_widths": {"P": 42.0},
                        "auto_fit_columns": ["P"],
                    }
                ]
            }
        },
        remote_result={
            "column_width_result": {
                "attempted_count": 1,
                "items": [
                    {
                        "sheet_name": "轨旁AP业务",
                        "column": "P",
                        "mode": "AUTO_FIT_WITH_LOCAL_MIN",
                        "local_workbook_width": 42.0,
                        "auto_fit_width": 35.0,
                        "requested_width": 42.0,
                        "remote_column_width": 42.0,
                        "remote_width_points": 294.0,
                        "applied": True,
                        "verified": True,
                    }
                ],
            }
        },
        enabled=True,
    )

    assert report["status"] == "SUCCESS"
    assert report["auto_fit_requested_count"] == 1
    assert report["auto_fit_applied_count"] == 1
    assert report["verified_count"] == 1
    assert report["items"][0]["wps_auto_fit_width"] == 35.0


def test_sheet_tab_color_formal_sync_gate_requires_successful_probe() -> None:
    target = _wps_target()
    assert _sheet_tab_color_probe_verified(target) is False

    target.sheet_tab_color_probe_diagnostic.update(
        {
            "status": "SUCCESS",
            "sheet_tab_color_verified": True,
        }
    )
    assert _sheet_tab_color_probe_verified(target) is True

    target.sheet_tab_color_probe_diagnostic["sheet_tab_color_verified"] = False
    assert _sheet_tab_color_probe_verified(target) is False


def test_column_width_probe_diagnostic_recognizes_successful_readback() -> None:
    target = _wps_target()
    assert _column_width_probe_verified(target) is False

    target.column_width_probe_diagnostic.update(
        {
            "status": "SUCCESS",
            "column_width_verified": True,
        }
    )
    assert _column_width_probe_verified(target) is True

    target.column_width_probe_diagnostic["column_width_verified"] = False
    assert _column_width_probe_verified(target) is False


def test_standard_workbook_format_mirror_is_enabled() -> None:
    assert WPS_STANDARD_FORMAT_MIRROR_ENABLED is True



def test_standard_airscript_applies_formats_after_the_stable_writer() -> None:
    script_path = (
        Path(__file__).parents[1]
        / "tools"
        / "wps_airscript"
        / "trackside_ap_standard_spreadsheet_sync.js"
    )
    script = script_path.read_text(encoding="utf-8")

    assert f'const SCRIPT_VERSION = "{WPS_SCRIPT_VERSIONS[STANDARD_TARGET_CODE]}";' in script
    assert f'const DEPLOYMENT_ID = "{WPS_DEPLOYMENT_IDS[STANDARD_TARGET_CODE]}";' in script
    assert "const FORMAT_MIRROR_ENABLED = true;" in script
    assert "function writeStableSheet(sheetDto, previousManaged)" in script
    assert "function clearFullReplaceTarget" in script
    assert "target.ClearContents();" in script
    assert "target.ClearFormats();" in script
    assert "target.UnMerge();" in script
    assert "managed_ranges_json" in script
    assert "FORMAT_MIRROR_ENABLED && args.format_mirror_enabled === true" in script
    assert "const result = skipRepeatedPrepend" in script
    assert ": writeStableSheet(runtimeSheet, managedRanges" in script
    assert "applyBusinessFormatting(formatSheets, formatWarnings, () =>" in script
    assert "sheetOrderVerification = reorderBusinessSheets(sheets);" in script
    assert "const sheet = sheets.Add();" in script
    assert "sheets.Add(null" not in script
    assert "range.ClearFormats()" in script
    assert "targetRange.Rows.AutoFit()" in script
    assert "window.FreezePanes = false" in script
    assert "function applyWorkbookFreezeLayout" in script
    assert "function expectedFreezeState(sheetDto)" in script
    assert "function resetWindowPaneState(sheet)" in script
    assert "function selectFirstRowFreezeAnchor(sheet)" in script
    assert "Application.ActiveCell" in script
    assert "WPS_FREEZE_SELECTION_FAILED" in script
    assert "WPS_FREEZE_REACTIVATION_READBACK_FAILED" in script
    assert "applyWorkbookFreezeLayout(sheets" in script
    assert 'sheet.Name, "freeze_panes", expected.mode' in script
    assert "sheetDto.freeze_panes" not in script
    assert "function columnNumber(" not in script
    assert "function selectFreezeAnchor(" not in script
    sheet_formatting = script.split("function applySheetFormatting", 1)[1].split(
        "function applyBusinessFormatting", 1
    )[0]
    assert "applyFreezePanes" not in sheet_formatting
    sync_body = script.split("function sync(payload)", 1)[1].split(
        "function sheetIsHidden", 1
    )[0]
    assert sync_body.index("applyBusinessFormatting(formatSheets") < sync_body.index(
        "sheetOrderVerification = reorderBusinessSheets(sheets)"
    )
    assert sync_body.index("applyBusinessSheetTabColors(sheets") < sync_body.index(
        "applyWorkbookFreezeLayout(sheets"
    )
    assert "row.RowHeight = entry.height" in script
    assert "Columns.AutoFit API unavailable" in script
    assert "AUTO_FIT_WITH_LOCAL_MIN" in script
    assert "Math.max(" in script
    assert "layout.max_width" in script
    assert 'sheet.AutoFilterMode = false' in script
    assert "sheetDto.format_runs" in script
    assert "verifiedFormatOperation" in script
    assert "range.Interior.Color" in script
    assert "range.Font.Bold" in script
    assert "range.NumberFormat" in script
    assert "range.HorizontalAlignment" in script
    assert "range.VerticalAlignment" in script
    assert "range.WrapText" in script
    assert "range.Borders.Item" in script
    assert "applyAllBorders" in script
    assert "xlInsideHorizontal" in script
    assert "xlInsideVertical" in script
    assert "all_borders: true" in script
    assert "window.SplitColumn = 0" in script
    assert re.findall(r"window\.SplitRow\s*=\s*([^;]+);", script) == ["0"]
    assert "expected_frozen_columns" in script
    assert "sheet.Range(merge).Merge()" in script
    assert "range.MergeArea && range.MergeArea.Address" in script
    assert "expected_values: cells[row - 1].slice" in script
    assert "sheet.Move(first)" in script
    assert "sheet.Move(null, last)" in script
    assert "sheet.Move({" not in script
    assert 'error_code: "WPS_SHEET_ORDER_VERIFY_FAILED"' in script
    assert 'if (args.operation === "sheet_order_probe") return sheetOrderProbe(args);' in script
    assert 'if (args.operation === "sheet_tab_color_probe") return sheetTabColorProbe(args);' in script
    assert 'if (args.operation === "column_width_probe") return columnWidthProbe(args);' in script
    assert "const columnWidthEnabled = args.column_width_enabled === true;" in script
    assert "applyBusinessColumnWidths(sheets, formatWarnings)" in script
    assert 'sheet.Name, "column_width"' in script
    assert "requested_width" in script
    assert "remote_column_width" in script
    assert "remote_width_points" in script
    assert "physical_width_change_points" in script
    assert "WPS_COLUMN_WIDTH_VALUE_VERIFIED" in script
    assert "function rowHeightProbe" not in script
    assert "applyRowHeights" in script
    assert "format_mirror_enabled: formatMirrorEnabled" in script
    assert "function toWpsColor(value)" in script
    assert 'sheet.Tab.Color = expected;' in script
    assert '"sheet_tab_color"' in script
    assert "sheet_order:" in script
    assert '"sheet_tab_color",\n      "Tab.Color"' in script
    assert "applyBusinessSheetTabColors(sheets, formatWarnings, mirroredFormatResults)" in script
    assert '.startsWith("_NetConsole")' in script
    assert 'status: publicWarnings.length ? "SUCCESS_WITH_WARNINGS" : "SUCCESS"' in script
    freeze_finalize_tail = sync_body.split(
        "if (formatMirrorEnabled) applyWorkbookFreezeLayout(sheets, formatWarnings, mirroredFormatResults);",
        1,
    )[1]
    for forbidden in (".Activate(", ".Select(", ".AutoFit(", ".Move(", ".AutoFilter("):
        assert forbidden not in freeze_finalize_tail
    assert 'if (args.operation === "migrate_legacy_binding") return migrateLegacyBinding(args);' in script
    assert "function updateBindingIdOnly(newBindingId)" in script
    assert 'sheet.Range(`B${index + 1}`).Value2 = String(newBindingId || "");' in script
    assert "LEGACY_BINDING_ID_MISMATCH" in script
    migration_block = script.split("function migrateLegacyBinding(args)", 1)[1].split(
        "function addFormatWarning", 1
    )[0]
    assert "updateBindingIdOnly(newBindingId)" in migration_block
    assert "expected_old_binding_id" in migration_block
    assert "new_binding_id" in migration_block
    assert "ensureSheet(" not in migration_block
    assert "writeStableSheet(" not in migration_block
    assert "args.workbook" not in migration_block
    assert script.rstrip().endswith("return main();")


def test_standard_probe_and_sync_scripts_share_deployment_identity() -> None:
    root = Path(__file__).parents[1] / "tools" / "wps_airscript"
    sync_script = (root / "trackside_ap_standard_spreadsheet_sync.js").read_text(
        encoding="utf-8"
    )
    probe_script = (
        root / "trackside_ap_standard_spreadsheet_connection_probe.js"
    ).read_text(encoding="utf-8")
    for expected in (
        WPS_SCRIPT_VERSIONS[STANDARD_TARGET_CODE],
        WPS_DEPLOYMENT_IDS[STANDARD_TARGET_CODE],
    ):
        assert expected in sync_script
        assert expected in probe_script
    assert sync_script.rstrip().endswith("return main();")
    assert probe_script.rstrip().endswith("return main();")
    assert "LEGACY_BINDING_ID_MISMATCH" in sync_script
    assert "LEGACY_BINDING_ID_MISMATCH" in probe_script
    assert "WPS_BINDING_MIGRATION_REQUIRES_SYNC_SCRIPT" in probe_script
    assert ".Worksheets.Add" not in probe_script
    assert "ensureSheet(" not in probe_script


def test_cloud_document_sync_uses_single_standard_workbook_payload(
    monkeypatch, tmp_path: Path
) -> None:
    class FakeClient:
        def __init__(self) -> None:
            self.payloads: list[dict[str, object]] = []

        def post(self, target, *, token, argv):
            assert token == "test-token"
            self.payloads.append(dict(argv))
            body = {
                "success": True,
                "protocol_version": 2,
                "script_version": "test",
                "target_type": target.target_type.value,
                "document_id": target.expected_document_id,
                "target_batch_id": argv.get("target_batch_id"),
                "site_id": argv.get("site_id"),
                "business_key": argv.get("business_key"),
                "snapshot_revision": argv.get("snapshot_revision"),
                "snapshot_sha256": argv.get("snapshot_sha256"),
            }
            body.update(
                {
                        "column_width_result": {
                            "attempted_count": 1,
                            "items": [
                                {
                                    "sheet_name": "轨旁AP业务",
                                    "column": "A",
                                    "mode": "EXPLICIT",
                                    "requested_width": 22.55,
                                    "before_column_width": 8.43,
                                    "remote_column_width": 22.55,
                                    "before_width_points": 59.01,
                                    "remote_width_points": 157.85,
                                    "physical_width_change_points": 98.84,
                                    "read_back": True,
                                    "applied": True,
                                    "verified": True,
                                }
                            ],
                        },
                        "format_results": {
                            "column_width": {"status": "SUCCESS"},
                            "row_height": {"status": "SUCCESS"},
                        },
                        "format_warnings": [],
                }
            )
            return type(
                "Response",
                (),
                {
                    "status_code": 200,
                    "body": body,
                },
            )

    fake = FakeClient()
    service = TracksideApWpsSyncService(PathResolver(tmp_path), client=fake)
    service.configure_target(
        "hzl10",
        STANDARD_TARGET_CODE,
        document_open_url="https://www.kdocs.cn/l/standard",
        webhook_url="https://www.kdocs.cn/api/v3/ide/file/standard/script/test/sync_task",
    )
    monkeypatch.setenv("NETCONSOLE_WPS_STANDARD_AIRSCRIPT_TOKEN", "test-token")
    repository = service._repository("hzl10")
    for target in repository.list_targets(TRACKSIDE_AP_WPS_BUSINESS_KEY):
        repository.set_runtime_capability(target.target_id, "VERIFIED")
        repository.update_target_diagnostic(
            target.target_id,
            operation="sheet_tab_color_probe",
            diagnostic={
                "status": "SUCCESS",
                "sheet_tab_color_verified": True,
            },
        )
    monkeypatch.setattr(
        service,
        "_build_snapshot",
        lambda site_id: {
            "business_revision": "revision-1",
            "created_at": "2026-08-07T10:00:00+08:00",
        },
    )
    from netconsole.models.wps_sync import WorkbookDTO, WorkbookSheetDTO, WpsSyncMode

    workbook = WorkbookDTO(
        sheets=(
            WorkbookSheetDTO(
                logical_sheet_key="trackside_ap_business",
                sheet_name="轨旁AP业务",
                sync_mode=WpsSyncMode.FULL_REPLACE,
                cells=[["归属站点"]],
                row_count=1,
                column_count=1,
                column_widths={"A": 22.55},
            ),
        )
    )
    manifest = [
        {
            "sheet_name": "轨旁AP业务",
            "column": "A",
            "range": "A:A",
            "column_label": "归属站点",
            "source_mode": "EXPLICIT",
            "local_workbook_width": 22.55,
            "sheet_dto_width": 22.55,
            "sheet_dto_mode": "EXPLICIT",
            "auto_fit_min_width": 8.0,
            "auto_fit_max_width": 60.0,
        }
    ]
    monkeypatch.setattr(
        service,
        "_build_workbook_dto",
        lambda site_id, batch_id, snapshot: (
            workbook,
            "sha-1",
            10,
            {"column_widths": manifest, "sheets": [], "totals": {}},
        ),
    )
    result = service.sync("hzl10")
    assert result["status"] == "SUCCESS"
    assert len(fake.payloads) == 1
    payload = fake.payloads[0]
    assert payload["snapshot_revision"] == "revision-1"
    assert payload["snapshot_sha256"] == "sha-1"
    assert payload["target_code"] == STANDARD_TARGET_CODE
    assert payload["sheet_tab_color_enabled"] is True
    assert payload["column_width_enabled"] is True
    assert payload["format_mirror_enabled"] is True
    assert "workbook" in payload
    assert "row_height_enabled" not in payload
    standard_result = result["targets"][0]
    report = standard_result["column_width_verification_report"]
    assert report["status"] == "SUCCESS"
    assert report["local_explicit_width_count"] == 1
    assert report["dto_match_count"] == 1
    assert report["payload_match_count"] == 1
    assert report["attempted_count"] == 1
    assert report["read_back_count"] == 1
    assert report["physical_read_back_count"] == 1
    assert report["verified_count"] == 1
    assert report["failed_count"] == 0
    assert report["representative_columns"][0]["column_label"] == "归属站点"


def test_wps_sync_aggregates_noncritical_format_warnings(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    class FakeClient:
        def post(self, target, *, token, argv):
            return WpsHttpResponse(
                status_code=200,
                body={
                    "success": True,
                    "protocol_version": 2,
                    "target_type": target.target_type.value,
                    "document_id": target.expected_document_id,
                    "target_batch_id": argv.get("target_batch_id"),
                    "site_id": argv.get("site_id"),
                    "business_key": argv.get("business_key"),
                    "snapshot_revision": argv.get("snapshot_revision"),
                    "snapshot_sha256": argv.get("snapshot_sha256"),
                    "format_warning_count": 1,
                    "format_warnings": [
                        {
                            "sheet_name": "轨旁AP业务",
                            "feature": "freeze_panes",
                            "reason": "runtime unsupported",
                        }
                    ],
                },
            )

    service = TracksideApWpsSyncService(PathResolver(tmp_path), client=FakeClient())
    service.configure_target(
        "hzl10",
        STANDARD_TARGET_CODE,
        document_open_url="https://www.kdocs.cn/l/standard",
        webhook_url="https://www.kdocs.cn/api/v3/ide/file/standard/script/test/sync_task",
    )
    monkeypatch.setenv("NETCONSOLE_WPS_STANDARD_AIRSCRIPT_TOKEN", "test-token")
    repository = service._repository("hzl10")
    target = repository.get_target(TRACKSIDE_AP_WPS_BUSINESS_KEY, STANDARD_TARGET_CODE)
    repository.set_runtime_capability(target.target_id, "VERIFIED")
    monkeypatch.setattr(
        service,
        "_build_snapshot",
        lambda site_id: {
            "business_revision": "revision-1",
            "created_at": "2026-08-07T10:00:00+08:00",
        },
    )
    from netconsole.models.wps_sync import WorkbookDTO

    monkeypatch.setattr(
        service,
        "_build_workbook_dto",
        lambda site_id, batch_id, snapshot: (
            WorkbookDTO(sheets=()),
            "sha-1",
            10,
            {"column_widths": [], "sheets": [], "totals": {}},
        ),
    )

    result = service.sync("hzl10", target_codes=[STANDARD_TARGET_CODE])

    assert result["status"] == "SUCCESS_WITH_WARNINGS"
    assert result["success_count"] == 1
    assert result["failed_count"] == 0
    assert result["warning_count"] == 1
    assert result["targets"][0]["status"] == "SUCCESS_WITH_WARNINGS"
    assert result["targets"][0]["format_warnings"][0]["feature"] == "freeze_panes"


def test_wps_sync_is_registered_as_a_job_center_handler() -> None:
    from netconsole.services.job_center.job_registry import registered_task_types

    assert WPS_SYNC_TASK_TYPE in registered_task_types()


def test_wps_runtime_probe_identity_is_persisted_and_invalidated_by_webhook_change(
    tmp_path: Path,
) -> None:
    paths = PathResolver(tmp_path)
    service = TracksideApWpsSyncService(paths)
    service.configure_target(
        "hzl10",
        STANDARD_TARGET_CODE,
        document_open_url="https://www.kdocs.cn/l/standard",
        webhook_url="https://www.kdocs.cn/api/v3/ide/file/standard/script/script-old/sync_task",
    )
    repository = service._repository("hzl10")
    target = repository.get_target(TRACKSIDE_AP_WPS_BUSINESS_KEY, STANDARD_TARGET_CODE)
    repository.update_target_remote_state(
        target.target_id,
        binding_status="BOUND",
        runtime_capability="VERIFIED",
        result={
            "document_id": "standard",
            "script_id": "script-old",
            "script_version": WPS_SCRIPT_VERSIONS[STANDARD_TARGET_CODE],
            "deployment_id": WPS_DEPLOYMENT_IDS[STANDARD_TARGET_CODE],
            "binding_id": target.binding_id,
            "site_id": "hzl10",
            "business_key": TRACKSIDE_AP_WPS_BUSINESS_KEY,
        },
    )

    verified = repository.get_target(TRACKSIDE_AP_WPS_BUSINESS_KEY, STANDARD_TARGET_CODE)
    assert verified.runtime_probe_document_id == "standard"
    assert verified.runtime_probe_script_id == "script-old"
    assert verified.runtime_capability == "VERIFIED"
    assert verified.binding_status == "BOUND"

    updated = service.configure_target(
        "hzl10",
        STANDARD_TARGET_CODE,
        webhook_url="https://www.kdocs.cn/api/v3/ide/file/standard/script/script-new/sync_task",
    )
    assert updated["runtime_capability"] == "DEPLOYMENT_PENDING"
    assert updated["binding_status"] == "BOUND"
    assert updated["runtime_probe_script_id"] == ""
    assert updated["expected_script_id"] == "script-new"


def _seed_verified_wps_target(
    service: TracksideApWpsSyncService,
    *,
    script_id: str = "script-one",
) -> tuple[WpsSyncRepository, WpsSyncTarget]:
    service.configure_target(
        "hzl10",
        STANDARD_TARGET_CODE,
        document_open_url="https://www.kdocs.cn/l/standard",
        webhook_url=(
            "https://www.kdocs.cn/api/v3/ide/file/standard/"
            f"script/{script_id}/sync_task"
        ),
        token="seed-token",
        enabled=True,
    )
    repository = service._repository("hzl10")
    target = repository.get_target(TRACKSIDE_AP_WPS_BUSINESS_KEY, STANDARD_TARGET_CODE)
    identity = {
        "document_id": "standard",
        "script_id": script_id,
        "script_version": WPS_SCRIPT_VERSIONS[STANDARD_TARGET_CODE],
        "deployment_id": WPS_DEPLOYMENT_IDS[STANDARD_TARGET_CODE],
        "binding_id": target.binding_id,
        "site_id": "hzl10",
        "business_key": TRACKSIDE_AP_WPS_BUSINESS_KEY,
    }
    repository.update_target_remote_identity(target.target_id, result=identity)
    repository.update_target_remote_state(
        target.target_id,
        binding_status="BOUND",
        result=identity,
        runtime_capability="VERIFIED",
    )
    for operation in (
        "connection_test",
        "runtime_write_probe",
        "sync_test_sheet",
        "sheet_tab_color_probe",
        "column_width_probe",
    ):
        repository.update_target_diagnostic(
            target.target_id,
            operation=operation,
            diagnostic={
                "executed_at": "2026-08-07T10:00:00+08:00",
                "status": "SUCCESS",
                "script_version": identity["script_version"],
                "deployment_id": identity["deployment_id"],
                "script_id": identity["script_id"],
                "document_id": identity["document_id"],
                "operation": operation,
                "message": "ok",
            },
        )
    return repository, repository.get_target(TRACKSIDE_AP_WPS_BUSINESS_KEY, STANDARD_TARGET_CODE)


def _stub_wps_sync_snapshot(
    monkeypatch: pytest.MonkeyPatch,
    service: TracksideApWpsSyncService,
) -> None:
    from netconsole.models.wps_sync import WorkbookDTO

    monkeypatch.setattr(
        service,
        "_build_snapshot",
        lambda site_id: {
            "business_revision": "revision-async",
            "created_at": "2026-08-09T10:00:00+08:00",
        },
    )
    monkeypatch.setattr(
        service,
        "_build_workbook_dto",
        lambda site_id, batch_id, snapshot: (
            WorkbookDTO(sheets=()),
            "sha-async",
            128,
            {"column_widths": [], "sheets": [], "totals": {"sheet_count": 0}},
        ),
    )


class _AsyncSyncClient(WpsAirScriptClient):
    def __init__(self, effects: list[object]) -> None:
        self.effects = effects
        self.submit_count = 0
        self.poll_count = 0
        self.submitted_batch_ids: list[str] = []
        self.last_payload: dict[str, object] = {}

    def submit_async(self, target, *, token, argv):
        self.submit_count += 1
        self.last_payload = dict(argv)
        self.submitted_batch_ids.append(str(argv.get("target_batch_id") or ""))
        return WpsRemoteTask(
            task_id="GN/KU3B3+remote-task==",
            task_type="open_air_script",
        )

    def poll_async_task(self, target, *, token, task_id):
        self.poll_count += 1
        effect = self.effects.pop(0)
        if isinstance(effect, BaseException):
            raise effect
        if effect == "running":
            return WpsHttpResponse(
                status_code=200,
                body={"status": "running", "error": "", "data": {"result": None}},
            )
        if effect == "remote_error":
            return WpsHttpResponse(
                status_code=200,
                body={
                    "status": "finished",
                    "error": "script failed",
                    "error_details": {"name": "Error", "msg": "write failed"},
                    "data": {"result": None},
                },
            )
        payload = self.last_payload
        result = {
            "success": True,
            "protocol_version": 2,
            "script_version": WPS_SCRIPT_VERSIONS[STANDARD_TARGET_CODE],
            "deployment_id": WPS_DEPLOYMENT_IDS[STANDARD_TARGET_CODE],
            "script_id": "script-one",
            "runtime_capability": "VERIFIED",
            "target_type": target.target_type.value,
            "target_code": target.target_code,
            "document_id": target.expected_document_id,
            "binding_id": target.binding_id,
            "binding_status": "BOUND",
            "target_batch_id": payload.get("target_batch_id"),
            "site_id": payload.get("site_id"),
            "business_key": payload.get("business_key"),
            "snapshot_revision": payload.get("snapshot_revision"),
            "snapshot_sha256": payload.get("snapshot_sha256"),
            "format_warnings": [],
        }
        return WpsHttpResponse(
            status_code=200,
            body={
                "status": "finished",
                "error": "",
                "data": {"result": json.dumps(result, ensure_ascii=False)},
            },
        )


def test_wps_full_sync_uses_async_submit_poll_and_persists_masked_remote_task(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    client = _AsyncSyncClient(["running", "finished"])
    service = TracksideApWpsSyncService(
        PathResolver(tmp_path),
        client=client,
        remote_task_poll_interval_seconds=0,
    )
    repository, _ = _seed_verified_wps_target(service)
    _stub_wps_sync_snapshot(monkeypatch, service)
    progress_messages: list[object] = []

    result = service.sync(
        "hzl10",
        target_codes=[STANDARD_TARGET_CODE],
        progress=lambda stage, current, total, message: progress_messages.append(message),
    )

    assert result["status"] == "SUCCESS"
    assert client.submit_count == 1
    assert client.poll_count == 2
    assert result["targets"][0]["remote_task_id_masked"] == "GN/KU3B3...sk=="
    assert all("GN/KU3B3+remote-task==" not in str(item) for item in progress_messages)
    with sqlite3.connect(repository.path) as connection:
        row = connection.execute(
            "SELECT remote_task_id, remote_task_type, remote_task_status, "
            "remote_task_submitted_at, remote_task_last_polled_at, remote_task_finished_at "
            "FROM wps_sync_target_runs"
        ).fetchone()
    assert row[0] == "GN/KU3B3+remote-task=="
    assert row[1] == "open_air_script"
    assert row[2] == "finished"
    assert all(row[index] for index in (3, 4, 5))
    recent = repository.recent_batches(TRACKSIDE_AP_WPS_BUSINESS_KEY)
    assert recent[0]["targets"][0]["remote_task_id_masked"] == "GN/KU3B3...sk=="
    assert "GN/KU3B3+remote-task==" not in str(recent)
    assert "request_payload_json" not in str(recent)


def test_wps_poll_connection_reset_keeps_task_id_and_restart_resumes_without_submit(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    clock = [0.0]
    client = _AsyncSyncClient(
        [WpsSyncError("REMOTE_POLL_TEMPORARY_FAILED", "[WinError 10054]")]
    )
    paths = PathResolver(tmp_path)
    first = TracksideApWpsSyncService(
        paths,
        client=client,
        remote_task_max_wait_seconds=1,
        remote_task_poll_interval_seconds=1,
        sleep=lambda seconds: clock.__setitem__(0, clock[0] + seconds),
        monotonic=lambda: clock[0],
    )
    repository, _ = _seed_verified_wps_target(first)
    _stub_wps_sync_snapshot(monkeypatch, first)

    unknown = first.sync("hzl10", target_codes=[STANDARD_TARGET_CODE])

    assert unknown["status"] == "REMOTE_RESULT_UNKNOWN"
    assert unknown["targets"][0]["status"] == "REMOTE_RESULT_UNKNOWN"
    with sqlite3.connect(repository.path) as connection:
        pending = connection.execute(
            "SELECT status, remote_task_id, completed_at FROM wps_sync_target_runs"
        ).fetchone()
    assert pending == (
        "REMOTE_RESULT_UNKNOWN",
        "GN/KU3B3+remote-task==",
        "",
    )

    client.effects.append("finished")
    resumed = TracksideApWpsSyncService(
        paths,
        client=client,
        remote_task_poll_interval_seconds=0,
    ).sync("hzl10", target_codes=[STANDARD_TARGET_CODE])

    assert resumed["status"] == "SUCCESS"
    assert client.submit_count == 1
    assert client.submitted_batch_ids == [resumed["targets"][0]["target_batch_id"]]


def test_wps_submit_timeout_retries_same_target_batch_id(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    class SubmitTimeoutClient(_AsyncSyncClient):
        def submit_async(self, target, *, token, argv):
            self.submit_count += 1
            self.last_payload = dict(argv)
            self.submitted_batch_ids.append(str(argv.get("target_batch_id") or ""))
            if self.submit_count == 1:
                raise WpsSyncError(
                    "ASYNC_SUBMIT_FAILED",
                    "The read operation timed out",
                    details={"submission_outcome": "UNKNOWN"},
                )
            return WpsRemoteTask(
                task_id="GN/KU3B3+remote-task==",
                task_type="open_air_script",
            )

    client = SubmitTimeoutClient(["finished"])
    paths = PathResolver(tmp_path)
    first = TracksideApWpsSyncService(paths, client=client)
    _seed_verified_wps_target(first)
    _stub_wps_sync_snapshot(monkeypatch, first)

    unknown = first.sync("hzl10", target_codes=[STANDARD_TARGET_CODE])
    resumed = TracksideApWpsSyncService(
        paths,
        client=client,
        remote_task_poll_interval_seconds=0,
    ).sync("hzl10", target_codes=[STANDARD_TARGET_CODE])

    assert unknown["status"] == "REMOTE_RESULT_UNKNOWN"
    assert unknown["targets"][0]["error_code"] == "ASYNC_SUBMIT_FAILED"
    assert resumed["status"] == "SUCCESS"
    assert client.submit_count == 2
    assert len(set(client.submitted_batch_ids)) == 1


def test_wps_remote_execution_error_is_terminal_and_keeps_remote_identity(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    client = _AsyncSyncClient(["remote_error"])
    service = TracksideApWpsSyncService(
        PathResolver(tmp_path),
        client=client,
        remote_task_poll_interval_seconds=0,
    )
    repository, _ = _seed_verified_wps_target(service)
    _stub_wps_sync_snapshot(monkeypatch, service)

    result = service.sync("hzl10", target_codes=[STANDARD_TARGET_CODE])

    assert result["status"] == "FAILED"
    assert result["targets"][0]["error_code"] == "WPS_REMOTE_EXECUTION_FAILED"
    with sqlite3.connect(repository.path) as connection:
        row = connection.execute(
            "SELECT status, remote_task_status, remote_task_id FROM wps_sync_target_runs"
        ).fetchone()
    assert row == (
        "FAILED",
        "failed",
        "GN/KU3B3+remote-task==",
    )


@pytest.mark.parametrize(
    "update",
    [
        {"document_open_url": "https://www.kdocs.cn/l/standard"},
        {"webhook_url": "https://www.kdocs.cn/api/v3/ide/file/standard/script/script-one/sync_task"},
        {"timeout_seconds": 60},
        {"enabled": False},
        {"token": "rotated-token"},
    ],
)
def test_wps_target_configuration_noop_and_non_identity_changes_keep_deployment(
    tmp_path: Path,
    update: dict[str, object],
) -> None:
    service = TracksideApWpsSyncService(PathResolver(tmp_path))
    repository, target = _seed_verified_wps_target(service)

    service.configure_target("hzl10", STANDARD_TARGET_CODE, **update)

    refreshed = repository.get_target(TRACKSIDE_AP_WPS_BUSINESS_KEY, STANDARD_TARGET_CODE)
    assert refreshed.runtime_capability == "VERIFIED"
    assert refreshed.binding_status == "BOUND"
    assert refreshed.runtime_probe_script_id == "script-one"
    assert refreshed.remote_script_id == "script-one"
    assert refreshed.remote_deployment_id == WPS_DEPLOYMENT_IDS[STANDARD_TARGET_CODE]
    assert refreshed.column_width_probe_diagnostic["status"] == "SUCCESS"
    if "token" in update:
        assert repository.resolve_token(target) == "rotated-token"


def test_wps_target_configuration_document_identity_change_clears_all_runtime_state(
    tmp_path: Path,
) -> None:
    service = TracksideApWpsSyncService(PathResolver(tmp_path))
    repository, target = _seed_verified_wps_target(service)

    service.configure_target(
        "hzl10",
        STANDARD_TARGET_CODE,
        webhook_url="https://www.kdocs.cn/api/v3/ide/file/other-document/script/script-one/sync_task",
    )

    refreshed = repository.get_target(TRACKSIDE_AP_WPS_BUSINESS_KEY, STANDARD_TARGET_CODE)
    assert refreshed.expected_document_id == "other-document"
    assert refreshed.runtime_capability == "DEPLOYMENT_PENDING"
    assert refreshed.binding_status == "BOUND"
    assert refreshed.remote_script_id == ""
    assert refreshed.connection_diagnostic == {}
    assert refreshed.runtime_probe_diagnostic == {}
    assert refreshed.sync_test_diagnostic == {}
    assert refreshed.sheet_tab_color_probe_diagnostic == {}
    assert refreshed.column_width_probe_diagnostic == {}
    assert refreshed.remote_identity_verified_at == ""
    assert target.remote_script_id == "script-one"


def test_wps_expected_deployment_change_downgrades_verification_but_keeps_binding(
    tmp_path: Path,
) -> None:
    service = TracksideApWpsSyncService(PathResolver(tmp_path))
    repository, target = _seed_verified_wps_target(service)
    repository.update_target_remote_state(
        target.target_id,
        binding_status="BOUND",
        runtime_capability="VERIFIED",
        result={
            "document_id": "standard",
            "script_id": "script-one",
            "script_version": "9.9.9-standard",
            "deployment_id": "stale-deployment",
        },
    )

    listed = {
        item["target_code"]: item
        for item in service.list_targets("hzl10")
    }[STANDARD_TARGET_CODE]
    assert listed["runtime_capability"] == "DEPLOYMENT_PENDING"
    assert listed["binding_status"] == "BOUND"
    assert listed["runtime_probe_deployment_id"] == ""
    assert listed["connection_diagnostic"] == {}


def test_wps_operation_diagnostics_keep_old_connection_failure_separate_from_new_probe(
    tmp_path: Path,
) -> None:
    old_result = {
        "success": True,
        "protocol_version": 2,
        "script_version": "2.1.0-standard",
        "deployment_id": "trackside-ap-standard-2.1.0",
        "script_id": "script-one",
        "target_type": "WPS_STANDARD_SPREADSHEET",
        "target_code": STANDARD_TARGET_CODE,
        "document_id": "standard",
        "runtime_capability": "DEPLOYMENT_PENDING",
    }
    current_result = {
        **old_result,
        "script_version": WPS_SCRIPT_VERSIONS[STANDARD_TARGET_CODE],
        "deployment_id": WPS_DEPLOYMENT_IDS[STANDARD_TARGET_CODE],
        "runtime_capability": "VERIFIED",
        "binding_status": "BOUND",
    }

    class FakeClient:
        def post(self, target, *, token, argv):
            if argv["operation"] == "connection_test":
                return WpsHttpResponse(status_code=200, body=old_result)
            return WpsHttpResponse(status_code=200, body=current_result)

    service = TracksideApWpsSyncService(PathResolver(tmp_path), client=FakeClient())
    service.configure_target(
        "hzl10",
        STANDARD_TARGET_CODE,
        document_open_url="https://www.kdocs.cn/l/standard",
        webhook_url="https://www.kdocs.cn/api/v3/ide/file/standard/script/script-one/sync_task",
        token="test-token",
    )
    with pytest.raises(WpsSyncError) as failure:
        service.connection_test("hzl10", STANDARD_TARGET_CODE)
    assert failure.value.code == "WPS_SCRIPT_VERSION_MISMATCH"
    service.runtime_write_probe("hzl10", STANDARD_TARGET_CODE)
    repository = service._repository("hzl10")
    target = repository.get_target(TRACKSIDE_AP_WPS_BUSINESS_KEY, STANDARD_TARGET_CODE)
    assert target.connection_diagnostic["status"] == "FAILED"
    assert target.connection_diagnostic["script_version"] == "2.1.0-standard"
    assert target.runtime_probe_diagnostic["status"] == "SUCCESS"
    assert target.runtime_probe_diagnostic["script_version"] == WPS_SCRIPT_VERSIONS[STANDARD_TARGET_CODE]
    assert target.remote_script_version == ""


def test_wps_runtime_probe_visibility_warning_keeps_core_verified_and_diagnostic(
    tmp_path: Path,
) -> None:
    capabilities = {
        "worksheet_enum": True,
        "worksheet_item": True,
        "worksheet_create": True,
        "scalar_value2": True,
        "matrix_value2": True,
        "used_range": True,
        "clear_contents": True,
        "entire_row_insert": True,
        "sheet_visibility": False,
    }
    body = {
        "success": True,
        "status": "SUCCESS_WITH_WARNINGS",
        "protocol_version": 2,
        "script_version": WPS_SCRIPT_VERSIONS[STANDARD_TARGET_CODE],
        "deployment_id": WPS_DEPLOYMENT_IDS[STANDARD_TARGET_CODE],
        "script_id": "script-one",
        "target_type": "WPS_STANDARD_SPREADSHEET",
        "target_code": STANDARD_TARGET_CODE,
        "document_id": "standard",
        "binding_status": "BOUND",
        "runtime_capability": "VERIFIED",
        "core_verified": True,
        "full_replace_ready": True,
        "prepend_snapshot_ready": True,
        "capabilities": capabilities,
        "core_capabilities": {
            key: value for key, value in capabilities.items() if key != "sheet_visibility"
        },
        "optional_capabilities": {"sheet_visibility": False},
        "warnings": [
            {
                "capability": "sheet_visibility",
                "message": "WPS 当前运行时无法确认系统 Sheet 隐藏状态",
            }
        ],
    }

    class FakeClient:
        def post(self, target, *, token, argv):
            return WpsHttpResponse(status_code=200, body=body)

    service = TracksideApWpsSyncService(PathResolver(tmp_path), client=FakeClient())
    service.configure_target(
        "hzl10",
        STANDARD_TARGET_CODE,
        document_open_url="https://www.kdocs.cn/l/standard",
        webhook_url="https://www.kdocs.cn/api/v3/ide/file/standard/script/script-one/sync_task",
        token="test-token",
    )

    result = service.runtime_write_probe("hzl10", STANDARD_TARGET_CODE)
    target = service._repository("hzl10").get_target(
        TRACKSIDE_AP_WPS_BUSINESS_KEY, STANDARD_TARGET_CODE
    )

    assert result["runtime_capability"] == "VERIFIED"
    assert target.runtime_capability == "VERIFIED"
    assert target.runtime_probe_diagnostic["status"] == "SUCCESS_WITH_WARNINGS"
    assert target.runtime_probe_diagnostic["core_verified"] is True
    assert target.runtime_probe_diagnostic["core_capabilities"]["matrix_value2"] is True
    assert target.runtime_probe_diagnostic["optional_capabilities"]["sheet_visibility"] is False
    assert target.runtime_probe_diagnostic["warnings"][0]["capability"] == "sheet_visibility"


def test_wps_sheet_tab_color_probe_is_independent_and_persists_verification(
    tmp_path: Path,
) -> None:
    operations: list[str] = []
    body = {
        "success": True,
        "status": "SUCCESS",
        "protocol_version": 2,
        "script_version": WPS_SCRIPT_VERSIONS[STANDARD_TARGET_CODE],
        "deployment_id": WPS_DEPLOYMENT_IDS[STANDARD_TARGET_CODE],
        "script_id": "script-one",
        "target_type": "WPS_STANDARD_SPREADSHEET",
        "target_code": STANDARD_TARGET_CODE,
        "document_id": "standard",
        "binding_status": "BOUND",
        "runtime_capability": "VERIFIED",
        "sheet_tab_color_verified": True,
        "expected_tab_color": "#C6EFCE",
        "actual_tab_color": 13561798,
        "probe_sheet": "_NetConsoleSyncTest",
        "message": "Sheet 标签颜色探针通过",
    }

    class FakeClient:
        def post(self, target, *, token, argv):
            operations.append(str(argv["operation"]))
            return WpsHttpResponse(status_code=200, body=body)

    service = TracksideApWpsSyncService(PathResolver(tmp_path), client=FakeClient())
    service.configure_target(
        "hzl10",
        STANDARD_TARGET_CODE,
        document_open_url="https://www.kdocs.cn/l/standard",
        webhook_url="https://www.kdocs.cn/api/v3/ide/file/standard/script/script-one/sync_task",
        token="test-token",
    )

    result = service.sheet_tab_color_probe("hzl10", STANDARD_TARGET_CODE)
    target = service._repository("hzl10").get_target(
        TRACKSIDE_AP_WPS_BUSINESS_KEY,
        STANDARD_TARGET_CODE,
    )

    assert operations == ["sheet_tab_color_probe"]
    assert result["sheet_tab_color_verified"] is True
    assert target.sheet_tab_color_probe_diagnostic["status"] == "SUCCESS"
    assert target.sheet_tab_color_probe_diagnostic["expected_tab_color"] == "#C6EFCE"
    assert target.runtime_capability == "DEPLOYMENT_PENDING"
    assert target.runtime_probe_diagnostic == {}
    assert target.sync_test_diagnostic == {}
    assert target.binding_status == "UNKNOWN"


def test_wps_column_width_probe_is_independent_and_persists_verification(
    tmp_path: Path,
) -> None:
    operations: list[str] = []
    body = {
        "success": True,
        "status": "SUCCESS",
        "protocol_version": 2,
        "script_version": WPS_SCRIPT_VERSIONS[STANDARD_TARGET_CODE],
        "deployment_id": WPS_DEPLOYMENT_IDS[STANDARD_TARGET_CODE],
        "script_id": "script-one",
        "target_type": "WPS_STANDARD_SPREADSHEET",
        "target_code": STANDARD_TARGET_CODE,
        "document_id": "standard",
        "binding_status": "BOUND",
        "runtime_capability": "VERIFIED",
        "column_width_verified": True,
        "expected_column_widths": {"A": 8, "B": 15, "C": 25, "D": 40},
        "actual_column_widths": {"A": 8, "B": 15, "C": 25, "D": 40},
        "probe_sheet": "_NetConsoleSyncTest",
        "message": "列宽探针通过",
    }

    class FakeClient:
        def post(self, target, *, token, argv):
            operations.append(str(argv["operation"]))
            return WpsHttpResponse(status_code=200, body=body)

    service = TracksideApWpsSyncService(PathResolver(tmp_path), client=FakeClient())
    service.configure_target(
        "hzl10",
        STANDARD_TARGET_CODE,
        document_open_url="https://www.kdocs.cn/l/standard",
        webhook_url="https://www.kdocs.cn/api/v3/ide/file/standard/script/script-one/sync_task",
        token="test-token",
    )

    result = service.column_width_probe("hzl10", STANDARD_TARGET_CODE)
    target = service._repository("hzl10").get_target(
        TRACKSIDE_AP_WPS_BUSINESS_KEY,
        STANDARD_TARGET_CODE,
    )

    assert operations == ["column_width_probe"]
    assert result["column_width_verified"] is True
    assert target.column_width_probe_diagnostic["status"] == "SUCCESS"
    assert target.column_width_probe_diagnostic["expected_column_widths"] == body["expected_column_widths"]
    assert target.runtime_capability == "DEPLOYMENT_PENDING"
    assert target.runtime_probe_diagnostic == {}
    assert target.sync_test_diagnostic == {}
    assert target.sheet_tab_color_probe_diagnostic == {}
    assert target.binding_status == "UNKNOWN"


def test_wps_runtime_probe_core_failure_persists_each_capability(
    tmp_path: Path,
) -> None:
    body = {
        "success": False,
        "status": "FAILED",
        "protocol_version": 2,
        "script_version": WPS_SCRIPT_VERSIONS[STANDARD_TARGET_CODE],
        "deployment_id": WPS_DEPLOYMENT_IDS[STANDARD_TARGET_CODE],
        "script_id": "script-one",
        "target_type": "WPS_STANDARD_SPREADSHEET",
        "target_code": STANDARD_TARGET_CODE,
        "document_id": "standard",
        "binding_status": "BOUND",
        "runtime_capability": "DEPLOYMENT_PENDING",
        "core_verified": False,
        "full_replace_ready": True,
        "prepend_snapshot_ready": False,
        "capabilities": {"matrix_value2": True, "entire_row_insert": False},
        "core_capabilities": {"matrix_value2": True, "entire_row_insert": False},
        "optional_capabilities": {"sheet_visibility": True},
        "capability_failures": [
            {"capability": "entire_row_insert", "message": "能力验证结果为未通过"}
        ],
        "error_code": "WPS_RUNTIME_PROBE_VERIFY_FAILED",
        "message": "运行时核心能力探针未通过：entire_row_insert",
    }

    class FakeClient:
        def post(self, target, *, token, argv):
            return WpsHttpResponse(status_code=200, body=body)

    service = TracksideApWpsSyncService(PathResolver(tmp_path), client=FakeClient())
    service.configure_target(
        "hzl10",
        STANDARD_TARGET_CODE,
        document_open_url="https://www.kdocs.cn/l/standard",
        webhook_url="https://www.kdocs.cn/api/v3/ide/file/standard/script/script-one/sync_task",
        token="test-token",
    )

    with pytest.raises(WpsSyncError) as failure:
        service.runtime_write_probe("hzl10", STANDARD_TARGET_CODE)

    assert failure.value.code == "WPS_RUNTIME_PROBE_VERIFY_FAILED"
    target = service._repository("hzl10").get_target(
        TRACKSIDE_AP_WPS_BUSINESS_KEY, STANDARD_TARGET_CODE
    )
    assert target.runtime_probe_diagnostic["status"] == "FAILED"
    assert target.runtime_probe_diagnostic["full_replace_ready"] is True
    assert target.runtime_probe_diagnostic["prepend_snapshot_ready"] is False
    assert target.runtime_probe_diagnostic["core_capabilities"]["entire_row_insert"] is False


def test_wps_successful_connection_test_persists_identity_and_diagnostic_atomically(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    body = {
        "success": True,
        "protocol_version": 2,
        "script_version": WPS_SCRIPT_VERSIONS[STANDARD_TARGET_CODE],
        "deployment_id": WPS_DEPLOYMENT_IDS[STANDARD_TARGET_CODE],
        "script_id": "script-one",
        "target_type": "WPS_STANDARD_SPREADSHEET",
        "target_code": STANDARD_TARGET_CODE,
        "document_id": "standard",
        "binding_status": "BOUND",
        "binding_id": "binding-one",
        "site_id": "hzl10",
        "site_name": "杭州地铁10号线",
        "business_key": TRACKSIDE_AP_WPS_BUSINESS_KEY,
        "runtime_capability": "DEPLOYMENT_PENDING",
    }

    class FakeClient:
        def post(self, target, *, token, argv):
            return WpsHttpResponse(status_code=200, body=body)

    service = TracksideApWpsSyncService(PathResolver(tmp_path), client=FakeClient())
    service.configure_target(
        "hzl10",
        STANDARD_TARGET_CODE,
        document_open_url="https://www.kdocs.cn/l/standard",
        webhook_url="https://www.kdocs.cn/api/v3/ide/file/standard/script/script-one/sync_task",
        token="test-token",
    )
    calls: list[str] = []
    original = WpsSyncRepository._update_target

    def observe(self, target_id, assignment, values):
        calls.append(str(assignment))
        return original(self, target_id, assignment, values)

    monkeypatch.setattr(WpsSyncRepository, "_update_target", observe)
    service.connection_test("hzl10", STANDARD_TARGET_CODE)

    connection_updates = [item for item in calls if "connection_diagnostic" in item]
    assert len(connection_updates) == 1
    assert "remote_script_version" in connection_updates[0]
    assert "remote_deployment_id" in connection_updates[0]
    assert "remote_script_id" in connection_updates[0]
    assert "remote_identity_verified_at" in connection_updates[0]
    target = service._repository("hzl10").get_target(
        TRACKSIDE_AP_WPS_BUSINESS_KEY, STANDARD_TARGET_CODE
    )
    assert target.last_test_status == "SUCCESS"
    assert target.remote_script_id == "script-one"
    assert target.connection_diagnostic["status"] == "SUCCESS"


def test_wps_revalidate_deployment_runs_all_probes_and_restores_verified(
    tmp_path: Path,
) -> None:
    operations: list[str] = []
    identity = {
        "protocol_version": 2,
        "script_version": WPS_SCRIPT_VERSIONS[STANDARD_TARGET_CODE],
        "deployment_id": WPS_DEPLOYMENT_IDS[STANDARD_TARGET_CODE],
        "script_id": "script-one",
        "target_type": "WPS_STANDARD_SPREADSHEET",
        "target_code": STANDARD_TARGET_CODE,
        "document_id": "standard",
        "binding_status": "BOUND",
        "binding_id": "binding-one",
        "site_id": "hzl10",
        "site_name": "杭州地铁10号线",
        "business_key": TRACKSIDE_AP_WPS_BUSINESS_KEY,
        "runtime_capability": "VERIFIED",
    }

    class FakeClient:
        def post(self, target, *, token, argv):
            operations.append(str(argv["operation"]))
            return WpsHttpResponse(
                status_code=200,
                body={
                    **identity,
                    "success": True,
                    "message": f"{argv['operation']} ok",
                },
            )

    service = TracksideApWpsSyncService(PathResolver(tmp_path), client=FakeClient())
    service.configure_target(
        "hzl10",
        STANDARD_TARGET_CODE,
        document_open_url="https://www.kdocs.cn/l/standard",
        webhook_url="https://www.kdocs.cn/api/v3/ide/file/standard/script/script-one/sync_task",
        token="test-token",
    )
    result = service.revalidate_deployment("hzl10", STANDARD_TARGET_CODE)
    assert operations == ["connection_test", "runtime_write_probe", "sync_test_sheet"]
    assert result["runtime_capability"] == "VERIFIED"
    target = service._repository("hzl10").get_target(
        TRACKSIDE_AP_WPS_BUSINESS_KEY, STANDARD_TARGET_CODE
    )
    assert target.runtime_capability == "VERIFIED"
    assert target.binding_status == "BOUND"
    assert target.remote_script_id == "script-one"
    assert target.connection_diagnostic["status"] == "SUCCESS"
    assert target.runtime_probe_diagnostic["status"] == "SUCCESS"
    assert target.sync_test_diagnostic["status"] == "SUCCESS"


def test_wps_formal_sync_gate_treats_operation_diagnostics_as_evidence_only(
    tmp_path: Path,
) -> None:
    service = TracksideApWpsSyncService(PathResolver(tmp_path))
    repository, target = _seed_verified_wps_target(service)
    _assert_standard_sync_readiness(target)

    stale = dict(target.sync_test_diagnostic)
    stale["deployment_id"] = "trackside-ap-standard-2.2"
    repository.update_target_diagnostic(
        target.target_id,
        operation="sync_test_sheet",
        diagnostic=stale,
    )
    _assert_standard_sync_readiness(
        repository.get_target(TRACKSIDE_AP_WPS_BUSINESS_KEY, STANDARD_TARGET_CODE)
    )


def test_wps_sync_rejects_unknown_and_unconfirmed_unbound_binding(tmp_path: Path) -> None:
    service = TracksideApWpsSyncService(PathResolver(tmp_path))
    service.configure_target(
        "hzl10",
        STANDARD_TARGET_CODE,
        document_open_url="https://www.kdocs.cn/l/standard",
        webhook_url="https://www.kdocs.cn/api/v3/ide/file/standard/script/script-one/sync_task",
        enabled=True,
    )
    repository = service._repository("hzl10")
    target = repository.get_target(TRACKSIDE_AP_WPS_BUSINESS_KEY, STANDARD_TARGET_CODE)
    repository.set_runtime_capability(target.target_id, "VERIFIED")

    with pytest.raises(WpsSyncError) as unknown:
        service.sync("hzl10", target_codes=[STANDARD_TARGET_CODE])
    assert unknown.value.code == "WPS_BINDING_STATUS_UNKNOWN"

    repository.update_target_remote_state(
        target.target_id,
        binding_status="UNBOUND",
        result={
            "document_id": "standard",
            "script_id": "script-one",
            "script_version": WPS_SCRIPT_VERSIONS[STANDARD_TARGET_CODE],
            "deployment_id": WPS_DEPLOYMENT_IDS[STANDARD_TARGET_CODE],
        },
    )
    with pytest.raises(WpsSyncError) as unbound:
        service.sync("hzl10", target_codes=[STANDARD_TARGET_CODE])
    assert unbound.value.code == "WPS_DOCUMENT_UNBOUND"

    repository.update_target_remote_state(
        target.target_id,
        binding_status="LEGACY_BINDING_ID_MISMATCH",
        result={"remote_binding_id": f"wst_{'a' * 32}"},
        persist_runtime_identity=False,
    )
    with pytest.raises(WpsSyncError) as legacy:
        service.sync("hzl10", target_codes=[STANDARD_TARGET_CODE])
    assert legacy.value.code == "WPS_LEGACY_BINDING_ID_MISMATCH"


def test_wps_legacy_binding_migration_rechecks_identity_and_runs_verification_chain(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    operations: list[dict[str, object]] = []
    remote_binding_id = [f"wst_{'b' * 32}"]

    class FakeClient:
        def post(self, target, *, token, argv):
            operations.append(dict(argv))
            assert token == "test-token"
            operation = str(argv["operation"])
            base = {
                "success": True,
                "protocol_version": 2,
                "script_version": WPS_SCRIPT_VERSIONS[STANDARD_TARGET_CODE],
                "deployment_id": WPS_DEPLOYMENT_IDS[STANDARD_TARGET_CODE],
                "script_id": "test",
                "runtime_capability": "VERIFIED",
                "target_type": target.target_type.value,
                "target_code": target.target_code,
                "document_id": target.expected_document_id,
                "binding_status": (
                    "BOUND"
                    if remote_binding_id[0] == target.binding_id
                    else "LEGACY_BINDING_ID_MISMATCH"
                ),
                "local_binding_id": target.binding_id,
                "remote_binding_id": remote_binding_id[0],
                "remote_document_id": target.expected_document_id,
                "remote_site_id": target.site_id,
                "remote_business_key": target.business_key,
                "remote_target_code": target.target_code,
                "remote_target_type": target.target_type.value,
            }
            if operation == "migrate_legacy_binding":
                assert argv["document_id"] == target.expected_document_id
                assert argv["expected_old_binding_id"] == remote_binding_id[0]
                assert argv["new_binding_id"] == target.binding_id
                previous_binding_id = remote_binding_id[0]
                remote_binding_id[0] = target.binding_id
                base.update(
                    {
                        "binding_status": "BOUND",
                        "binding_id": target.binding_id,
                        "remote_binding_id": target.binding_id,
                        "binding_id_match": True,
                        "migrated": True,
                        "already_migrated": False,
                        "previous_binding_id": previous_binding_id,
                        "message": "旧版绑定标识已迁移",
                    }
                )
            return WpsHttpResponse(
                status_code=200,
                body=base,
            )

    service = TracksideApWpsSyncService(PathResolver(tmp_path), client=FakeClient())
    service.configure_target(
        "hzl10",
        STANDARD_TARGET_CODE,
        document_open_url="https://www.kdocs.cn/l/standard",
        webhook_url="https://www.kdocs.cn/api/v3/ide/file/standard/script/test/sync_task",
    )
    monkeypatch.setenv("NETCONSOLE_WPS_STANDARD_AIRSCRIPT_TOKEN", "test-token")
    repository = service._repository("hzl10")
    target = repository.get_target(TRACKSIDE_AP_WPS_BUSINESS_KEY, STANDARD_TARGET_CODE)
    legacy_binding_id = remote_binding_id[0]
    repository.update_target_remote_state(
        target.target_id,
        binding_status="LEGACY_BINDING_ID_MISMATCH",
        result={
            "remote_binding_id": legacy_binding_id,
            "remote_site_id": target.site_id,
            "remote_business_key": target.business_key,
        },
        persist_runtime_identity=False,
    )

    result = service.migrate_legacy_binding("hzl10", STANDARD_TARGET_CODE)

    assert [item["operation"] for item in operations] == [
        "connection_test",
        "migrate_legacy_binding",
        "connection_test",
        "runtime_write_probe",
        "sync_test_sheet",
    ]
    assert result["migrated"] is True
    assert result["already_migrated"] is False
    assert result["previous_binding_id"] == legacy_binding_id
    assert result["verification"]["runtime_capability"] == "VERIFIED"
    refreshed = repository.get_target(TRACKSIDE_AP_WPS_BUSINESS_KEY, STANDARD_TARGET_CODE)
    assert refreshed.binding_status == "BOUND"
    assert refreshed.remote_binding_id == target.binding_id
    assert refreshed.connection_diagnostic["binding_id_match"] is True


def test_wps_connection_test_rejects_webhook_script_id_mismatch() -> None:
    body = {
        "success": True,
        "protocol_version": 2,
        "script_version": WPS_SCRIPT_VERSIONS[STANDARD_TARGET_CODE],
        "deployment_id": WPS_DEPLOYMENT_IDS[STANDARD_TARGET_CODE],
        "script_id": "other-script",
        "target_type": "WPS_STANDARD_SPREADSHEET",
        "target_code": STANDARD_TARGET_CODE,
        "document_id": "document",
        "runtime_capability": "DEPLOYMENT_PENDING",
    }

    class FakeClient:
        def post(self, target, *, token, argv):
            return WpsHttpResponse(status_code=200, body=body)

    with pytest.raises(WpsSyncError) as captured:
        WpsStandardSpreadsheetAdapter(FakeClient()).connection_test(
            _wps_target(), "test-only-token"
        )
    assert captured.value.code == "WPS_SCRIPT_ID_MISMATCH"
