from pathlib import Path

from openpyxl import Workbook

from netconsole.core.database import Database
from netconsole.repositories.ac_repository import AcRepository
from netconsole.services.ap_extension_import import (
    AP_SWITCH_PORT_POINT_TABLE,
    AP_NAME_MAC_LIST,
    PIS_LAYOUT_TABLE,
    SIGNAL_AB_NETWORK_TABLE,
    STANDARD_TEMPLATE_TYPE,
    ApExtensionImportService,
    normalize_ap_mac,
    parse_mileage,
    standard_export_row,
    standard_template_headers,
)


def make_database(tmp_path: Path) -> Database:
    database = Database(tmp_path / "devices.db")
    database.initialize()
    return database


def test_normalize_ap_mac_accepts_common_formats():
    assert normalize_ap_mac("4ce9-e4ef-5c20").normalized == "4ce9e4ef5c20"
    assert normalize_ap_mac("4c:e9:e4:ef:5c:20").normalized == "4ce9e4ef5c20"
    assert normalize_ap_mac("4ce9.e4ef.5c20").normalized == "4ce9e4ef5c20"
    assert normalize_ap_mac("4ce9e4ef5c20").display == "4ce9-e4ef-5c20"
    assert normalize_ap_mac("bad-mac").error == "MAC格式无效"


def test_parse_mileage_keeps_raw_and_parses_common_forms():
    assert parse_mileage("K6+491").meters == 6491
    assert parse_mileage("DK6+491").meters == 6491
    assert parse_mileage("AK0+207.267").meters == 207.267
    assert parse_mileage("ZDK0+035").meters == 35
    assert parse_mileage("YDK1+020").meters == 1020
    assert parse_mileage("CDK1+170").meters == 1170
    assert parse_mileage("RDK12+345").meters == 12345
    assert parse_mileage("6491").meters == 6491
    assert parse_mileage("bad").raw == "bad"
    assert parse_mileage("bad").meters is None
    assert parse_mileage("bad").error == "里程无法解析"


def test_preview_pis_left_right_layout_without_mac(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "PIS点位表"
    sheet.append(["车站", "轨旁AP归属区间", "左线里程", "右线里程", "距上一个AP", "曲线半径", "曲线开始", "曲线终点", "AP编号", "备注"])
    sheet.append(["金家渡", "金家渡-墩祥街", "K6+491", "", "90", "450", "K6+300", "K6+700", "Z01-01", "左线点位"])
    sheet.append(["金家渡", "金家渡-墩祥街", "", "K6+580", "95", "", "", "", "Y01-01", "右线点位"])
    path = tmp_path / "pis.xlsx"
    workbook.save(path)

    preview = ApExtensionImportService().preview_file(path, "smart_design")

    assert preview.template_type == PIS_LAYOUT_TABLE
    assert preview.confidence_score >= 60
    assert preview.summary["missing_mac_rows"] == 2
    assert preview.standard_rows[0]["line_side"] == "左线"
    assert preview.standard_rows[0]["direction"] == "下行"
    assert preview.standard_rows[0]["mileage_m"] == 6491
    assert preview.standard_rows[0]["curve_radius_m"] == 450
    assert preview.standard_rows[1]["line_side"] == "右线"
    assert preview.standard_rows[1]["direction"] == "上行"


def test_standard_export_row_formats_mileage_with_line_prefix():
    row = {
        "station_name": "金家渡",
        "line_side": "左线",
        "direction": "下行",
        "mileage_text": "K0+035",
        "mileage_m": 35,
    }

    values = standard_export_row(row)
    headers = list(standard_template_headers())

    assert values[headers.index("里程原文")] == "ZDK0+035"
    assert values[headers.index("里程米")] == 35


def test_preview_signal_ab_network_sheet_recognizes_section_and_yard(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "A网"
    sheet.append(["3.中医药大学", "", "", "", "3.中医药大学", "", ""])
    sheet.append(["", "", "", "", "", "ap0201_a", "5866-bab3-1111"])
    sheet.append(["4.联庄", "", "", "", "4.联庄", "", ""])
    sheet.append(["", "", "", "", "", "ap0303_a", "5866-bab3-0a40"])
    sheet.append(["88.勾庄车辆段(库内）", "ap8841_a", "4ce9-e4ef-7b80", "", "", "", ""])
    sheet.append(["90.七堡停车场（库内）", "ap8020_a", "4ce9-e4f1-29a0", "", "", "", ""])
    path = tmp_path / "杭4AP布点表 - 副本A.xlsx"
    workbook.save(path)

    preview = ApExtensionImportService().preview_file(path, "smart_design")
    rows = {row["ap_name"]: row for row in preview.standard_rows}

    assert preview.template_type == SIGNAL_AB_NETWORK_TABLE
    assert rows["ap0303_a"]["line_name"] == "杭州地铁4号线"
    assert rows["ap0303_a"]["system_type"] == "信号"
    assert rows["ap0303_a"]["network_domain"] == "A网"
    assert rows["ap0303_a"]["belong_type"] == "section"
    assert rows["ap0303_a"]["station_name"] == ""
    assert rows["ap0303_a"]["section_name"] == "联庄-中医药大学"
    assert rows["ap0303_a"]["section_start_station"] == "中医药大学"
    assert rows["ap0303_a"]["section_end_station"] == "联庄"
    assert rows["ap8841_a"]["belong_type"] == "yard"
    assert rows["ap8841_a"]["station_name"] == "勾庄车辆段"
    assert rows["ap8841_a"]["yard_name"] == "勾庄车辆段"
    assert rows["ap8841_a"]["area_name"] == "库内"
    assert rows["ap8020_a"]["station_name"] == "七堡停车场"


def test_preview_signal_b_network_sheet_keeps_network_domain(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "B网"
    sheet.append(["3.中医药大学", "", "", "", "3.中医药大学", "", ""])
    sheet.append(["4.联庄", "", "", "", "4.联庄", "", ""])
    sheet.append(["", "", "", "", "", "ap0303_b", "5866-bab2-77c0"])
    path = tmp_path / "杭4AP布点表 - 副本b.xlsx"
    workbook.save(path)

    preview = ApExtensionImportService().preview_file(path, "smart_design")

    assert preview.template_type == SIGNAL_AB_NETWORK_TABLE
    assert preview.standard_rows[0]["network_domain"] == "B网"
    assert preview.standard_rows[0]["ap_name"] == "ap0303_b"
    assert preview.standard_rows[0]["section_name"] == "联庄-中医药大学"


def test_preview_ap_name_mac_list_uses_segment_title(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "AP清单"
    sheet.append(["32.金家渡"])
    sheet.append(["AP名称", "MAC"])
    sheet.append(["JJD-AP-01", "4c:e9:e4:ef:5c:20"])
    path = tmp_path / "mac-list.xlsx"
    workbook.save(path)

    preview = ApExtensionImportService().preview_file(path, "smart_design")

    assert preview.template_type == AP_NAME_MAC_LIST
    row = preview.standard_rows[0]
    assert row["station_name"] == "金家渡"
    assert row["ap_name"] == "JJD-AP-01"
    assert row["ap_mac_norm"] == "4ce9e4ef5c20"


def test_preview_scans_multiple_sheets(tmp_path):
    workbook = Workbook()
    workbook.active.title = "说明"
    workbook.active.append(["无数据"])
    sheet = workbook.create_sheet("AP清单")
    sheet.append(["AP名称", "MAC"])
    sheet.append(["AP-01", "4ce9-e4ef-5c20"])
    path = tmp_path / "multi.xlsx"
    workbook.save(path)

    preview = ApExtensionImportService().preview_file(path, "smart_design")

    assert [sheet.sheet_name for sheet in preview.sheets] == ["AP清单"]
    assert preview.summary["total_rows"] == 1


def test_preview_ap_switch_port_point_table_maps_access_fields_without_mileage(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "轨旁AP业务"
    sheet.append(["归属站点", "室内交换机", "接口名称", "AP_MAC", "AP名称", "区间", "AP编号"])
    sheet.append(["11-高桥西", "11-高桥西1", "GE1/0/1", "1c94-6876-8ee0", "1c94-6876-8ee0", "高桥西-高桥-上行", "AP0127"])
    sheet.append(["50-高桥西停车场", "50-高桥西停车场1", "GE1/0/2", "1c94-6876-8ee1", "1c94-6876-8ee1", "", "AP0683"])
    sheet.append(["11-高桥西", "11-高桥西1", "GE1/0/3", "-", "-", "", ""])
    path = tmp_path / "ap-switch-port.xlsx"
    workbook.save(path)

    preview = ApExtensionImportService().preview_file(path)

    assert preview.template_type == AP_SWITCH_PORT_POINT_TABLE
    assert preview.confidence_score == 95
    assert preview.summary["total_rows"] == 3
    assert preview.summary["missing_mac_rows"] == 1
    assert preview.summary["error_rows"] == 0
    section, yard, placeholder = preview.standard_rows
    assert section["station_name"] == "高桥西"
    assert section["section_name"] == "高桥西-高桥-上行"
    assert section["section_start_station"] == "高桥西"
    assert section["section_end_station"] == "高桥"
    assert section["direction"] == "上行"
    assert section["belong_type"] == "section"
    assert section["uplink_switch"] == "11-高桥西1"
    assert section["uplink_port"] == "GE1/0/1"
    assert section["mileage_text"] == ""
    assert yard["station_name"] == "高桥西停车场"
    assert yard["belong_type"] == "yard"
    assert placeholder["ap_mac_display"] == "-"


def test_standard_template_import_keeps_direction_mileage_and_infers_section(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "AP扩展信息模板"
    sheet.append(["AP名称", "AP_MAC", "归属类型", "归属站点", "归属区间", "区间起点站", "区间终点站", "场段", "区域", "里程", "点位说明", "上下行"])
    sheet.append(["AP-CLD_01", "10b6-5e92-d3e0", "场段", "高桥南车辆段", "", "", "", "高桥南车辆段", "", "", "场段", ""])
    sheet.append(["AP-S_1214", "10b6-5e92-c780", "站点", "古林站", "", "", "", "", "", 5276, "正线站台两侧", "上行"])
    sheet.append(["AP-S_1215", "10b6-5e92-f340", "站点", "古林站", "", "", "", "", "", 5426, "正线", "上行"])
    sheet.append(["AP-S_1218", "083b-e9ec-b980", "站点", "古林站", "", "", "", "", "", 5876, "正线", "上行"])
    sheet.append(["AP-S_1301", "94a7-482c-2360", "站点", "云林西路站", "", "", "", "", "", 6026, "正线", "上行"])
    sheet.append(["AP-CRD-X_1603", "94a7-482c-2440", "区间", "卖面桥站", "", "", "", "", "", 450, "出入段线", "下行"])
    path = tmp_path / "ningbo_like_template.xlsx"
    workbook.save(path)

    preview = ApExtensionImportService().preview_file(path, "standard_template")
    rows = {row["ap_name"]: row for row in preview.standard_rows}

    assert rows["AP-S_1214"]["direction"] == "上行"
    assert rows["AP-S_1214"]["line_side"] == "右线"
    assert rows["AP-S_1214"]["mileage_text"] == "5276"
    assert rows["AP-S_1214"]["mileage_m"] == 5276
    assert rows["AP-S_1214"]["location_desc"] == "正线站台两侧"
    assert rows["AP-S_1214"]["belong_type"] == "section"
    assert rows["AP-S_1214"]["section_name"] == "古林站-云林西路站"
    assert rows["AP-S_1214"]["section_start_station"] == "古林站"
    assert rows["AP-S_1214"]["section_end_station"] == "云林西路站"
    assert rows["AP-CRD-X_1603"]["direction"] == "下行"
    assert rows["AP-CRD-X_1603"]["belong_type"] == "section"
    assert rows["AP-CLD_01"]["belong_type"] == "yard"
    assert rows["AP-CLD_01"]["section_name"] == ""


def test_repository_imports_unbound_points_and_matches_resources_by_mac(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    repository.replace_fit_ap_resources("ac-1", [{"ap_name": "AP-01", "ap_mac": "4ce9-e4ef-5c20"}])
    stats = repository.import_ap_extension_points(
        [
            {
                "station_name": "金家渡",
                "line_side": "左线",
                "direction": "下行",
                "mileage_text": "K6+491",
                "mileage_m": 6491,
                "ap_point_code": "Z01-01",
            },
            {
                "station_name": "金家渡",
                "belong_type": "station",
                "ap_name": "AP-01",
                "ap_mac_display": "4c:e9:e4:ef:5c:20",
                "power_station": "金家渡变电所",
            },
            {
                "belong_type": "section",
                "section_name": "联庄-中医药大学",
                "section_start_station": "中医药大学",
                "section_end_station": "联庄",
                "ap_name": "AP-02",
                "ap_mac_display": "5866-bab3-0a40",
            },
        ],
        source_file="design.xlsx",
        template_type=PIS_LAYOUT_TABLE,
    )

    assert stats["success_rows"] == 3
    rows = repository.list_ap_extension_points()
    assert {row["match_status"] for row in rows} == {"unbound_no_mac", "matched_by_mac", "extension_not_online"}
    section_row = next(row for row in rows if row["ap_name"] == "AP-02")
    assert section_row["belong_type"] == "section"
    assert section_row["section_name"] == "联庄-中医药大学"
    resource = repository.list_fit_ap_resources_with_metadata("ac-1")[0]
    assert resource["extension_station_name"] == "金家渡"
    assert resource["extension_belong_type"] == "station"
    assert resource["extension_power_station"] == "金家渡变电所"
    assert resource["site"] in (None, "")


def test_repository_extension_search_matches_prefixed_mileage(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    repository.import_ap_extension_points(
        [
            {
                "station_name": "金家渡",
                "line_side": "左线",
                "direction": "下行",
                "mileage_text": "K0+035",
                "mileage_m": 35,
                "ap_point_code": "Z01-01",
            }
        ],
        source_file="design.xlsx",
        template_type=PIS_LAYOUT_TABLE,
    )

    rows = repository.list_ap_extension_points(search="ZDK0+035")

    assert len(rows) == 1
    assert rows[0]["ap_point_code"] == "Z01-01"


def test_repository_extension_search_matches_direction_section_and_location_desc(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    repository.import_ap_extension_points(
        [
            {
                "station_name": "古林站",
                "section_name": "古林站-云林西路站",
                "section_start_station": "古林站",
                "section_end_station": "云林西路站",
                "line_side": "右线",
                "direction": "上行",
                "mileage_text": "5276",
                "mileage_m": 5276,
                "location_desc": "正线站台两侧",
                "ap_name": "AP-S_1214",
                "ap_mac_display": "10b6-5e92-c780",
            }
        ],
        source_file="template.xlsx",
        template_type=STANDARD_TEMPLATE_TYPE,
    )

    assert repository.list_ap_extension_points(search="上行")[0]["ap_name"] == "AP-S_1214"
    assert repository.list_ap_extension_points(search="5276")[0]["ap_name"] == "AP-S_1214"
    assert repository.list_ap_extension_points(search="正线站台")[0]["ap_name"] == "AP-S_1214"
    assert repository.list_ap_extension_points(search="古林站-云林西路站")[0]["ap_name"] == "AP-S_1214"
