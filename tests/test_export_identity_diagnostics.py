from __future__ import annotations

import sqlite3
from copy import deepcopy
from pathlib import Path

from openpyxl import Workbook, load_workbook

from netconsole.services.export_identity_diagnostics import (
    DEFAULT_SAMPLE_LIMIT,
    DEFAULT_WARNING_LIMIT,
    ExportIdentityDiagnostics,
    unavailable_export_identity_diagnostics,
)
from netconsole.services.mesh_link_detail_export import export_mesh_link_details_xlsx
from netconsole.services.online_mr_analysis_report_exporter import OnlineMrAnalysisReportExporter
from netconsole.services.export.export_job import ExportJob


def test_export_identity_diagnostics_counts_risks_without_mutating_rows() -> None:
    rows = [
        {
            "record_seq": 1,
            "peer_mac": "30f5-277a-5a2f",
            "peer_ap_mac": "30:f5:27:7a:5a:2f",
            "peer_radio_mac": "30f5.277a.5a2f",
            "peer_ap_name": "30f5-277a-5a2f",
            "peer_resolve_source": "legacy_mapping",
        },
        {
            "record_seq": 2,
            "peer_mac": "",
            "peer_ap_mac": "",
            "bssid": "30f5-277a-5a3f",
            "min_mr_rssi": 31,
            "standby_peer_mac": "30f5-277a-5a4f",
        },
        {
            "record_seq": 3,
            "peer_mac": "30f5-277a-5a4f",
            "peer_ap_mac": "30f5-277a-5a5f",
            "peer_radio_mac": "30f5-277a-5a5f",
            "peer_ap_name": "AP-03",
        },
    ]
    original = deepcopy(rows)

    report = ExportIdentityDiagnostics("unit").inspect_mesh_link_detail_rows(rows)

    assert rows == original
    assert report.available is True
    assert report.total_rows == 3
    assert report.duplicate_peer_radio_mac_rows == 2
    assert report.peer_mac_equals_ap_mac_rows == 1
    assert report.peer_mac_equals_peer_radio_mac_rows == 1
    assert report.ap_name_mac_like_rows == 1
    assert report.radio_or_bssid_only_rows == 1
    assert report.missing_ap_mac_rows == 1
    assert report.missing_peer_mac_rows == 1
    assert report.missing_min_rssi_rows == 2
    assert report.missing_backup_link_rows == 2
    assert report.has_mapping_source_field is True
    assert report.has_peer_radio_mac_field is True


def test_export_identity_diagnostics_limits_samples_and_warnings() -> None:
    rows = [
        {
            "record_seq": index,
            "peer_mac": "30f5-277a-5a2f",
            "peer_ap_mac": "30f5-277a-5a2f",
            "peer_radio_mac": "30f5-277a-5a2f",
            "peer_ap_name": "30f5-277a-5a2f",
        }
        for index in range(100)
    ]

    payload = ExportIdentityDiagnostics("limits", sample_limit=999, warning_limit=999).inspect_mesh_link_detail_rows(rows).to_dict()

    assert len(payload["samples"]) == DEFAULT_SAMPLE_LIMIT
    assert len(payload["warnings"]) <= DEFAULT_WARNING_LIMIT


def test_unavailable_export_identity_diagnostics_is_structured() -> None:
    payload = unavailable_export_identity_diagnostics("mesh_link_detail", RuntimeError("diagnostics failed"))

    assert payload["available"] is False
    assert payload["export_type"] == "mesh_link_detail"
    assert payload["total_rows"] == 0
    assert "RuntimeError" in str(payload["error"])


def test_export_identity_diagnostics_marks_unknown_row_shape_unavailable() -> None:
    report = ExportIdentityDiagnostics("unknown").inspect_mesh_link_detail_rows([{"value": 1}])

    assert report.available is False
    assert report.total_rows == 1


def test_mesh_export_keeps_workbook_contract_and_returns_diagnostics(tmp_path: Path) -> None:
    target = tmp_path / "mesh.xlsx"
    row = _mesh_row()

    result = export_mesh_link_details_xlsx(target, [row], total_rows=1)

    diagnostics = result["export_identity_diagnostics"]
    assert diagnostics["available"] is True
    assert diagnostics["total_rows"] == 1
    assert diagnostics["peer_mac_equals_ap_mac_rows"] == 1
    assert diagnostics["peer_mac_equals_peer_radio_mac_rows"] == 1
    workbook = load_workbook(target)
    sheet = workbook["链路明细"]
    headers = [cell.value for cell in sheet[1]]
    assert headers[:8] == ["序号", "采样时间", "Radio", "链路状态", "Peer MAC", "对端AP MAC", "对端AP名称", "归属站点"]
    assert "归属来源" not in headers
    assert "Peer Radio MAC" not in headers
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref


def test_mesh_export_diagnostics_failure_does_not_fail_export(tmp_path: Path, monkeypatch) -> None:
    target = tmp_path / "mesh-diagnostics-failed.xlsx"

    def fail_diagnostics(*_args, **_kwargs):
        raise RuntimeError("diagnostics failed")

    monkeypatch.setattr(ExportIdentityDiagnostics, "inspect_mesh_link_detail_row", fail_diagnostics)

    result = export_mesh_link_details_xlsx(target, [_mesh_row()], total_rows=1)

    assert target.exists()
    assert result["export_identity_diagnostics"]["available"] is False
    assert load_workbook(target)["链路明细"]["E2"].value == "30f5-277a-5a2f"


def test_mesh_export_worker_finished_result_contains_diagnostics(tmp_path: Path, monkeypatch) -> None:
    from netconsole import export_worker

    class FakeRepository:
        def __init__(self, _path: Path) -> None:
            pass

        def count_link_details(self, _filters: dict[str, object]) -> int:
            return 1

        def query_active_link_build_order(self, *_args) -> list[dict[str, object]]:
            return []

        def query_source_files(self, _limit: int, _offset: int) -> tuple[int, list[dict[str, object]]]:
            return 0, []

        def export_rows(self, _name: str) -> list[dict[str, object]]:
            return []

        def iter_link_details(self, _filters: dict[str, object], *, batch_size: int):
            assert batch_size == 2000
            yield _mesh_row()

    events: list[dict[str, object]] = []
    monkeypatch.setattr(export_worker, "MeshMrRepository", FakeRepository)
    monkeypatch.setattr(export_worker, "_emit", events.append)
    output_path = tmp_path / "worker-mesh.xlsx"
    tmp_output = tmp_path / "worker-mesh.xlsx.tmp"
    job = ExportJob(
        job_id="mesh-diagnostics",
        job_type="mesh_link_detail",
        db_path=str(tmp_path / "mesh.sqlite"),
        output_path=str(output_path),
        tmp_path=str(tmp_output),
    )

    export_worker._run_mesh_link_detail(job)

    assert output_path.exists()
    assert events[-1]["type"] == "finished"
    assert events[-1]["result"]["export_identity_diagnostics"]["total_rows"] == 1
    assert not list(tmp_path.glob("*.diagnostics.json"))


def test_online_mr_compat_export_keeps_detail_columns_and_exposes_diagnostics(tmp_path: Path) -> None:
    session_dir, db_path = _online_mr_database(tmp_path)
    workbook = Workbook()
    exporter = OnlineMrAnalysisReportExporter()

    exporter._append_offline_report_sheets(workbook, session_dir, db_path)

    sheet = workbook["链路明细"]
    headers = [cell.value for cell in sheet[1]]
    assert headers == [
        "采样时间",
        "设备时间",
        "Radio",
        "状态",
        "PeerMac",
        "当前PEER AP名称",
        "AP MAC",
        "归属站点",
        "归属区间",
        "Peer Radio MAC",
        "MR RSSI",
        "BSSID",
        "Mesh接口",
        "Online Time",
    ]
    assert sheet["E2"].value == sheet["G2"].value == sheet["J2"].value == "30f5-277a-5a2f"
    diagnostics = exporter.result_metadata["export_identity_diagnostics"]
    assert diagnostics["available"] is True
    assert diagnostics["total_rows"] == 1
    assert diagnostics["peer_mac_equals_ap_mac_rows"] == 1
    assert diagnostics["peer_mac_equals_peer_radio_mac_rows"] == 1
    assert diagnostics["has_mapping_source_field"] is False


def test_online_mr_compat_diagnostics_failure_does_not_fail_workbook(tmp_path: Path, monkeypatch) -> None:
    session_dir, db_path = _online_mr_database(tmp_path)
    workbook = Workbook()
    exporter = OnlineMrAnalysisReportExporter()

    def fail_diagnostics(*_args, **_kwargs):
        raise RuntimeError("diagnostics failed")

    monkeypatch.setattr(ExportIdentityDiagnostics, "inspect_online_mr_detail_rows", fail_diagnostics)

    exporter._append_offline_report_sheets(workbook, session_dir, db_path)

    assert workbook["链路明细"]["E2"].value == "30f5-277a-5a2f"
    assert exporter.result_metadata["export_identity_diagnostics"]["available"] is False


def test_export_identity_diagnostics_static_boundaries() -> None:
    project_root = Path(__file__).resolve().parents[1]
    diagnostics_source = (project_root / "netconsole" / "services" / "export_identity_diagnostics.py").read_text(encoding="utf-8")
    online_source = (project_root / "netconsole" / "services" / "online_mr_analysis_report_exporter.py").read_text(encoding="utf-8")

    for forbidden in ("PySide6", "netconsole.ui", "repositories", "sqlite3", "Workbook", "save(", "write_text", "open("):
        assert forbidden not in diagnostics_source
    assert "peer_mac, belong_station, belong_section, peer_mac, mr_rssi, bssid" in online_source
    assert '"PeerMac", "当前PEER AP名称", "AP MAC"' in online_source
    assert '"Peer Radio MAC"' in online_source


def _mesh_row() -> dict[str, object]:
    return {
        "record_seq": 1,
        "sample_time": "2026-07-11 10:00:00.000",
        "radio": 1,
        "link_state": "ACTIVE",
        "peer_mac_raw": "30f5-277a-5a2f",
        "peer_mac_normalized": "30:f5:27:7a:5a:2f",
        "peer_ap_mac": "30f5-277a-5a2f",
        "peer_radio_mac": "30f5-277a-5a2f",
        "peer_ap_name": "30f5-277a-5a2f",
        "peer_site": "站点A",
        "belong_section": "区间A",
        "peer_resolve_source": "legacy_mapping",
        "peer_radio": "Radio 1",
        "duration_text": "0d 00h 00m 03s",
        "link_count": 1,
        "metrics_json": "{}",
    }


def _online_mr_database(tmp_path: Path) -> tuple[Path, Path]:
    session_dir = tmp_path / "session"
    db_path = session_dir / "parsed" / "online_diagnosis.sqlite"
    db_path.parent.mkdir(parents=True)
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            """
            CREATE TABLE main_link_samples (
                id INTEGER PRIMARY KEY,
                collector_time TEXT,
                device_time TEXT,
                device_clock TEXT,
                radio INTEGER,
                link_state TEXT,
                peer_mac TEXT,
                resolved_peer_name TEXT,
                peer_name TEXT,
                belong_station TEXT,
                belong_section TEXT,
                mr_rssi REAL,
                bssid TEXT,
                mesh_interface TEXT,
                online_time TEXT
            )
            """
        )
        conn.execute(
            """
            INSERT INTO main_link_samples (
                collector_time, device_time, device_clock, radio, link_state, peer_mac,
                resolved_peer_name, peer_name, belong_station, belong_section, mr_rssi,
                bssid, mesh_interface, online_time
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "2026-07-11 10:00:00.000",
                "2026-07-11 10:00:00.100",
                "",
                1,
                "ACTIVE",
                "30f5-277a-5a2f",
                "30f5-277a-5a2f",
                "",
                "站点A",
                "区间A",
                32,
                "30f5-277a-5a3f",
                "WLAN-Mesh1/0/1",
                "00h 01m 00s",
            ),
        )
    return session_dir, db_path
