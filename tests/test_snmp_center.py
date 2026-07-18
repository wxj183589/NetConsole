from __future__ import annotations

import os
from pathlib import Path
from zipfile import ZipFile

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

import pytest
from openpyxl import Workbook

from netconsole.core.paths import PathResolver
from netconsole.models.device import Device
from netconsole.models.snmp_models import DeviceSnmpProfileResult, SnmpProfile, SnmpQueryRequest, SnmpQueryResult, SnmpSetRequest, SnmpVarBind
from netconsole.repositories.global_mib_repository import GlobalMibRepository
from netconsole.repositories.site_snmp_repository import SiteSnmpRepository
from netconsole.services.comware_version_service import parse_comware_version
from netconsole.services.mib_product_reference_compare_service import MibProductReferenceCompareService
from netconsole.services.mib_resource_service import MibResourceService
from netconsole.services.file_contract import ImportValidationError
from netconsole.services.snmp_recommend_service import SnmpRecommendService
from netconsole.services import snmp_client as snmp_client_module
from netconsole.services.snmp_client import SnmpClient, _WireResponse, _WireVarBind, _encode_snmp_value, normalize_oid
from netconsole.services.snmp_query_service import SnmpQueryService








def test_snmp_center_initializes_global_and_site_databases(tmp_path: Path):
    paths = PathResolver(tmp_path)
    paths.ensure_global_mib_dirs()
    paths.ensure_site_snmp_dirs("demo")
    global_repo = GlobalMibRepository(paths.global_mib_db_path())
    site_repo = SiteSnmpRepository(paths.site_snmp_db_path("demo"))

    global_repo.initialize()
    site_repo.initialize()

    dictionaries = global_repo.list_dictionary_sets()
    templates = global_repo.list_templates()
    objects = global_repo.list_objects("sysName")

    assert paths.global_mib_db_path().exists()
    assert paths.site_snmp_db_path("demo").exists()
    assert any(item["name"] == "内置通用字典" for item in dictionaries)
    assert any(item["template_name"] == "系统信息" for item in templates)
    assert objects[0]["oid"] == "1.3.6.1.2.1.1.5"


def test_mib_import_indexes_objects_and_deduplicates_by_hash(tmp_path: Path):
    paths = PathResolver(tmp_path)
    service = MibResourceService(paths)
    mib_file = tmp_path / "TEST-MIB.mib"
    mib_file.write_text(
        """
TEST-MIB DEFINITIONS ::= BEGIN
IMPORTS
    enterprises FROM SNMPv2-SMI;

testRoot OBJECT IDENTIFIER ::= { enterprises 99999 }

testScalar OBJECT-TYPE
    SYNTAX      INTEGER { up(1), down(2) }
    MAX-ACCESS  read-only
    STATUS      current
    DESCRIPTION "Test scalar"
    ::= { testRoot 1 }
END
""".strip(),
        encoding="utf-8",
    )

    first = service.import_paths([mib_file], vendor="TestVendor")
    second = service.import_paths([mib_file], vendor="TestVendor")
    repo = GlobalMibRepository(paths.global_mib_db_path())
    objects = repo.list_objects("testScalar")
    with repo.connect() as conn:
        conn.execute("UPDATE mib_objects SET parent_oid = '' WHERE name = 'testScalar'")
        conn.commit()
    prefix_children = repo.list_oid_children("1.3.6.1.4.1.99999")

    assert first.imported == 1
    assert second.duplicated == 1
    assert objects[0]["module_name"] == "TEST-MIB"
    assert objects[0]["oid"] == "1.3.6.1.4.1.99999.1"
    assert objects[0]["enum_map_json"] == '{"1": "up", "2": "down"}'
    assert any(row["name"] == "testScalar" for row in prefix_children)


def test_batch_import_resolves_dependencies_by_module_definition_and_suffix(tmp_path: Path):
    paths = PathResolver(tmp_path)
    service = MibResourceService(paths)
    archive = tmp_path / "H3C-V9-V7-Comware_MIB-20260610.zip"
    with ZipFile(archive, "w") as package:
        package.writestr(
            "rfc2578.sm2",
            """
SNMPv2-SMI DEFINITIONS ::= BEGIN
enterprises OBJECT IDENTIFIER ::= { private 1 }
END
""".strip(),
        )
        package.writestr(
            "hh3c-oid.mib",
            """
HH3C-OID-MIB DEFINITIONS ::= BEGIN
IMPORTS enterprises FROM SNMPv2-SMI;
hh3c OBJECT IDENTIFIER ::= { enterprises 25506 }
END
""".strip(),
        )
        package.writestr(
            "mesh.txt",
            """
HH3C-DOT11S-MESH-MIB DEFINITIONS ::= BEGIN
IMPORTS hh3c FROM HH3C-OID-MIB;
hh3cDot11 OBJECT IDENTIFIER ::= { hh3c 2 }
hh3cDot11sMesh OBJECT IDENTIFIER ::= { hh3cDot11 75 }
hh3cDot11sMeshLinkStatusTable OBJECT-TYPE
    SYNTAX SEQUENCE OF MeshEntry
    MAX-ACCESS not-accessible
    STATUS current
    DESCRIPTION "Mesh table"
    ::= { hh3cDot11sMesh 11 }
END
""".strip(),
        )
        package.writestr("readme.txt", "not a mib")

    report = service.import_paths([archive])
    repo = GlobalMibRepository(paths.global_mib_db_path())
    modules = repo.list_modules()
    packages = repo.list_source_packages()
    templates = repo.list_templates()

    assert report.failed == 0
    assert {row["module_name"] for row in modules} >= {"SNMPv2-SMI", "HH3C-OID-MIB", "HH3C-DOT11S-MESH-MIB"}
    assert any(row["package_name"] == "H3C-Comware-V7V9-20260610" for row in packages)
    assert not repo.list_missing_dependency_summary()
    assert any(row["template_name"] == "Mesh 链路状态模板" for row in templates)


def test_builtin_h3c_initialization_registers_packages_once(tmp_path: Path):
    builtin_dir = tmp_path / "resources" / "builtin_mibs" / "h3c" / "comware_v5_20210918"
    builtin_dir.mkdir(parents=True)
    archive = builtin_dir / "H3C-V5-Comware_MIB-20210918.zip"
    with ZipFile(archive, "w") as package:
        package.writestr(
            "rfc2578.sm2",
            "SNMPv2-SMI DEFINITIONS ::= BEGIN\nenterprises OBJECT IDENTIFIER ::= { private 1 }\nEND",
        )
        package.writestr(
            "hh3c-oid.mib",
            "HH3C-OID-MIB DEFINITIONS ::= BEGIN\nIMPORTS enterprises FROM SNMPv2-SMI;\nhh3c OBJECT IDENTIFIER ::= { enterprises 25506 }\nEND",
        )

    paths = PathResolver(tmp_path)
    service = MibResourceService(paths)
    first = service.initialize_builtin_resources()
    second = service.initialize_builtin_resources()
    repo = GlobalMibRepository(paths.global_mib_db_path())
    packages = repo.list_source_packages()

    assert first.imported == 2
    assert second.imported == 0
    assert len(packages) == 1
    assert packages[0]["source_type"] == "builtin_h3c_comware_package"
    assert packages[0]["version_line"] == "V5"


def test_comware_release_and_product_reference_recommendation(tmp_path: Path):
    paths = PathResolver(tmp_path)
    repo = GlobalMibRepository(paths.global_mib_db_path())
    repo.initialize()
    xlsx = tmp_path / "20260522-H3C 无线控制器产品 MIB参考-R16xx-6W100.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "对象"
    sheet.append(["MIB模块", "对象名称", "OID", "中文含义", "实现规格", "支持情况"])
    sheet.append(["HH3C-DOT11S-MESH-MIB", "hh3cDot11sMeshNbrRSSI", "1.3.6.1.4.1.25506.2.75.11.3.3.1.11", "邻居 RSSI", "R16xx 支持", "支持读取"])
    workbook.save(xlsx)

    report = MibResourceService(paths, repo).import_paths([xlsx])
    profile = DeviceSnmpProfileResult(
        device_name="NBDT12HX-WX3540X-AC1",
        vendor="H3C",
        device_type="AC",
        model="WX3540X",
        system="Comware",
        system_version="9.1.081",
        os_family="Comware",
        os_major="V9",
        release="1608P01",
        release_number="1608",
        release_patch="P01",
        release_series="R16xx",
        sys_descr="H3C Comware Software, Version 9.1.081, Release 1608P01",
    )
    recommendations = SnmpRecommendService(repo).recommend_product_references(Device(name="NBDT12HX-WX3540X-AC1", device_vendor="H3C", device_type="AC"), profile)
    override = repo.find_product_object_override(module_name="HH3C-DOT11S-MESH-MIB", object_name="hh3cDot11sMeshNbrRSSI", numeric_oid="1.3.6.1.4.1.25506.2.75.11.3.3.1.11")
    parsed = parse_comware_version(profile.sys_descr)

    assert report.imported == 1
    assert parsed.software_version == "9.1.081"
    assert parsed.release == "1608P01"
    assert parsed.release_series == "R16xx"
    assert recommendations[0].reference_name == "20260522-H3C 无线控制器产品 MIB参考-R16xx-6W100"
    assert "R16xx" in "；".join(recommendations[0].reasons)
    assert override is not None
    assert override["implementation_spec"] == "R16xx 支持"


def test_product_reference_compare_matches_by_oid_and_normalizes_category_numbers(tmp_path: Path):
    paths = PathResolver(tmp_path)
    repo = GlobalMibRepository(paths.global_mib_db_path())
    repo.initialize()

    left_xlsx = tmp_path / "20251226-H3C 无线控制器产品 MIB参考-R12xx_E12xx-6W101.xlsx"
    right_xlsx = tmp_path / "20260522-H3C 无线控制器产品 MIB参考-R16xx-6W100.xlsx"

    left_book = Workbook()
    left_sheet = left_book.active
    left_sheet.title = "表节点"
    left_sheet.append(["分册名", "模块名", "MIB文件名", "根节点", "父节点名称", "功能描述", "操作支持情况", "子节点名称及OID", "最大访问权限", "数据类型", "有效范围", "含义", "实现规格"])
    left_sheet.append(["11-WLAN", "01-HH3C-DOT11S-MESH-MIB", "hh3c-dot11s-mesh.mib", "hh3cDot11sMesh", "hh3cDot11sMeshNbrStatusTable", "邻居 RSSI", "读取约束：支持", "hh3cDot11sMeshNbrRSSI（1.3.6.1.4.1.25506.2.75.11.3.3.1.11）", "read-only", "Integer32", "-100..0", "旧 RSSI", "R12xx 支持"])
    left_book.save(left_xlsx)

    right_book = Workbook()
    right_sheet = right_book.active
    right_sheet.title = "表节点"
    right_sheet.append(["分册名", "模块名", "MIB文件名", "根节点", "父节点名称", "功能描述", "操作支持情况", "子节点名称及OID", "最大访问权限", "数据类型", "有效范围", "含义", "实现规格"])
    right_sheet.append(["13-WLAN", "13-HH3C-DOT11S-MESH-MIB", "hh3c-dot11s-mesh.mib", "hh3cDot11sMesh", "hh3cDot11sMeshNbrStatusTable", "邻居 RSSI", "读取约束：支持", "hh3cDot11sMeshNbrRSSI（1.3.6.1.4.1.25506.2.75.11.3.3.1.11）", "read-only", "Integer32", "-120..0", "新 RSSI", "R16xx 支持"])
    right_sheet.append(["13-WLAN", "13-HH3C-DOT11S-MESH-MIB", "hh3c-dot11s-mesh.mib", "hh3cDot11sMesh", "hh3cDot11sMeshLinkStatusTable", "链路状态", "读取约束：支持", "hh3cDot11sMeshLinkStatus（1.3.6.1.4.1.25506.2.75.11.3.1.1.5）", "read-only", "Integer32", "1..2", "链路状态", "R16xx 新增"])
    right_book.save(right_xlsx)

    report = MibResourceService(paths, repo).import_paths([left_xlsx, right_xlsx])
    references = repo.list_product_references()
    left_id = next(int(row["id"]) for row in references if row["release_series"] == "R12xx,E12xx")
    right_id = next(int(row["id"]) for row in references if row["release_series"] == "R16xx")

    result = MibProductReferenceCompareService(repo).compare(left_id, right_id)
    stored_rows = repo.list_product_reference_compare_results(left_id, right_id, limit=1000)
    right_objects = repo.list_product_object_overrides(right_id)

    assert report.imported == 2
    assert any(row["module_name"] == "HH3C-DOT11S-MESH-MIB" for row in right_objects)
    assert result.summary["objects_added"] == 1
    assert any(row["diff_type"] == "category_changed" and row["field_name"] == "category_name" for row in stored_rows)
    assert any(row["diff_type"] == "changed" and row["field_name"] == "取值范围" for row in stored_rows)
    assert any(row["diff_type"] == "added" and row["numeric_oid"] == "1.3.6.1.4.1.25506.2.75.11.3.1.1.5" for row in stored_rows)


def test_product_reference_import_parses_chinese_headers_and_rebuilds_tree(tmp_path: Path):
    paths = PathResolver(tmp_path)
    repo = GlobalMibRepository(paths.global_mib_db_path())
    repo.initialize()
    xlsx = tmp_path / "20260522-H3C 无线控制器产品 MIB参考-R16xx-6W100.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "全局节点"
    sheet.append(["分册名", "模块名", "MIB文件名", "根节点", "全局节点名称及OID", "最大访问权限", "数据类型", "有效范围", "含义", "实现规格"])
    sheet.append(["13-WLAN", "13-HH3C-DOT11S-MESH-MIB", "hh3c-dot11s-mesh.mib", "hh3cDot11sMesh", "hh3cDot11sMeshPflMeshID (1.3.6.1.4.1.25506.2.75.11.1.2.1.2)", "read-only", "OCTET STRING", "SIZE(0..32)", "Mesh ID", "R16xx 支持"])
    table_sheet = workbook.create_sheet("表节点")
    table_sheet.append(["分册名", "模块名", "MIB文件名", "根节点", "父节点名称", "功能描述", "操作支持情况", "子节点名称及OID", "最大访问权限", "数据类型", "有效范围", "含义", "实现规格", "表节点信息"])
    table_sheet.append(["13-WLAN", "13-HH3C-DOT11-APMT-MIB", "hh3c-dot11-apmt.mib", "hh3cDot11APMT", "hh3cDot11APObjectStatusTable", "AP 状态", "读取约束：支持", "hh3cDot11APID (1.3.6.1.4.1.25506.2.75.2.1.1.1.1)", "not-accessible", "Integer32", "1..2147483647", "AP ID", "R16xx 支持", "索引节点是 hh3cDot11APID"])
    table_sheet.append(["13-WLAN", "13-HH3C-FLASH-MAN-MIB", "h3c-flash-man.mib", "hh3cFlashMan", "hh3cFlashTable", "Flash 芯片描述", "读取约束：支持", "hh3cFlhChipDescr (1.3.6.1.4.1.25506.2.5.1.1.3.1.1.3)", "read-only", "OCTET STRING", "", "Flash 描述", "R16xx 支持", ""])
    table_sheet.append(["13-WLAN", "13-HH3C-IPSEC-MONITOR-V2-MIB", "h3c-ipsec-monitor-v2.mib", "hh3cIPsecMonitorV2", "hh3cIPsecTrap", "IPsec 认证失败", "读取约束：支持", "hh3cIpsecAuthFailureTrapCntV2 (1.3.6.1.4.1.25506.2.126.1.8.5)", "read-only", "Counter32", "", "IPsec Trap", "R16xx 支持", ""])
    workbook.save(xlsx)

    report = MibResourceService(paths, repo).import_paths([xlsx])
    references = repo.list_product_references()
    reference_id = int(references[0]["id"])
    objects = repo.list_product_reference_objects(reference_id)
    top_nodes = repo.list_product_reference_tree_nodes(reference_id, None)
    root_id = int(top_nodes[0]["id"])
    categories = repo.list_product_reference_tree_nodes(reference_id, root_id)

    assert report.failed == 0
    assert any(row["object_name"] == "hh3cDot11sMeshPflMeshID" for row in objects)
    assert any(row["numeric_oid"] == "1.3.6.1.4.1.25506.2.75.11.1.2.1.2" for row in objects)
    assert any(row["node_type"] == "category" and row["display_name"] == "13-WLAN" for row in categories)
    category_id = int(next(row["id"] for row in categories if row["display_name"] == "13-WLAN"))
    modules = repo.list_product_reference_tree_nodes(reference_id, category_id)
    module_names = [str(row["display_name"]) for row in modules]
    assert all(not str(row.get("numeric_oid") or "") for row in categories)
    assert all(not str(row.get("numeric_oid") or "") for row in modules)
    assert module_names.index("HH3C-FLASH-MAN-MIB") < module_names.index("HH3C-DOT11-APMT-MIB")
    assert module_names.index("HH3C-DOT11-APMT-MIB") < module_names.index("HH3C-IPSEC-MONITOR-V2-MIB")

    with repo.connect() as conn:
        conn.execute("DELETE FROM mib_product_reference_tree_nodes WHERE reference_id = ?", (reference_id,))
        conn.execute("DELETE FROM mib_product_reference_objects WHERE reference_id = ?", (reference_id,))
        conn.commit()
    rebuilt = repo.rebuild_product_reference_tree(reference_id)
    rebuilt_top = repo.list_product_reference_tree_nodes(reference_id, None)
    rebuilt_root_id = int(rebuilt_top[0]["id"])
    rebuilt_categories = repo.list_product_reference_tree_nodes(reference_id, rebuilt_root_id)

    assert rebuilt["object_count"] == 4
    assert rebuilt["node_count"] > 2
    assert any(row["display_name"] == "13-WLAN" for row in rebuilt_categories)
    rebuilt_category_id = int(next(row["id"] for row in rebuilt_categories if row["display_name"] == "13-WLAN"))
    rebuilt_modules = repo.list_product_reference_tree_nodes(reference_id, rebuilt_category_id)
    rebuilt_module_names = [str(row["display_name"]) for row in rebuilt_modules]
    assert all(not str(row.get("numeric_oid") or "") for row in rebuilt_categories)
    assert all(not str(row.get("numeric_oid") or "") for row in rebuilt_modules)
    assert rebuilt_module_names.index("HH3C-FLASH-MAN-MIB") < rebuilt_module_names.index("HH3C-DOT11-APMT-MIB")
    assert rebuilt_module_names.index("HH3C-DOT11-APMT-MIB") < rebuilt_module_names.index("HH3C-IPSEC-MONITOR-V2-MIB")
















def test_snmp_query_service_dispatches_browser_operations(tmp_path: Path):
    class FakeSnmpClient:
        def __init__(self) -> None:
            self.calls: list[tuple[str, int | None]] = []

        def _result(self, method: str, oid: str) -> SnmpQueryResult:
            request = SnmpQueryRequest(profile=SnmpProfile(host="127.0.0.1"), method=method, oid=oid, save_history=False)
            return SnmpQueryResult(request=request, rows=[SnmpVarBind(oid=oid, value="ok")], status="success")

        def get(self, profile: SnmpProfile, oid: str) -> SnmpQueryResult:
            self.calls.append(("get", None))
            return self._result("Get", oid)

        def get_next(self, profile: SnmpProfile, oid: str) -> SnmpQueryResult:
            self.calls.append(("get_next", None))
            return self._result("GetNext", oid)

        def get_bulk(self, profile: SnmpProfile, oid: str, *, max_repetitions: int = 10) -> SnmpQueryResult:
            self.calls.append(("get_bulk", max_repetitions))
            return self._result("GetBulk", oid)

        def get_subtree(self, profile: SnmpProfile, oid: str, *, max_rows: int = 200, cancel_checker=None) -> SnmpQueryResult:
            self.calls.append(("get_subtree", max_rows))
            return self._result("GetSubtree", oid)

        def walk(self, profile: SnmpProfile, oid: str, *, max_rows: int = 200, cancel_checker=None) -> SnmpQueryResult:
            self.calls.append(("walk", max_rows))
            return self._result("Walk", oid)

        def bulk_walk(self, profile: SnmpProfile, oid: str, *, max_repetitions: int = 10, max_rows: int = 200, cancel_checker=None) -> SnmpQueryResult:
            self.calls.append(("bulk_walk", max_repetitions))
            return self._result("BulkWalk", oid)

        def table_walk(self, profile: SnmpProfile, oid: str, *, max_repetitions: int = 10, max_rows: int = 200, cancel_checker=None) -> SnmpQueryResult:
            self.calls.append(("table_walk", max_rows))
            return self._result("TableWalk", oid)

    repo = SiteSnmpRepository(tmp_path / "snmp.db")
    repo.initialize()
    fake_client = FakeSnmpClient()
    service = SnmpQueryService(repo, client=fake_client)  # type: ignore[arg-type]
    profile = SnmpProfile(host="127.0.0.1")

    for method in ["Get", "GetNext", "GetBulk", "GetSubtree", "Walk", "BulkWalk", "TableWalk"]:
        request = SnmpQueryRequest(profile=profile, method=method, oid="1.3.6.1", max_repetitions=7, max_rows=9, save_history=False)
        assert service.run(request).status == "success"

    assert fake_client.calls == [
        ("get", None),
        ("get_next", None),
        ("get_bulk", 7),
        ("get_subtree", 9),
        ("walk", 9),
        ("bulk_walk", 7),
        ("table_walk", 9),
    ]


def test_snmp_client_getbulk_keeps_good_rows_before_end_of_mib(monkeypatch):
    class FakeWireClient:
        responses: list[_WireResponse] = []

        def __init__(self, profile: SnmpProfile) -> None:
            self.profile = profile

        def request(self, oids: list[str], *, pdu_type: int, max_repetitions: int = 10) -> _WireResponse:
            assert pdu_type == 0xA5
            assert max_repetitions == 10
            return self.responses.pop(0)

    FakeWireClient.responses = [
        _WireResponse(
            status="end_of_mib_view",
            error_message="end",
            varbinds=[
                _WireVarBind("1.3.6.1.4.1.25506.2.75.2.1.1.1.1.1", "ap1", "OCTET STRING"),
                _WireVarBind("1.3.6.1.4.1.25506.2.75.2.1.1.1.1.2", "ap2", "OCTET STRING"),
                _WireVarBind("1.3.6.1.4.1.25506.2.75.2.1.1.1.1", "endOfMibView", "endOfMibView", status="end_of_mib_view", error_message="end"),
            ],
        )
    ]
    monkeypatch.setattr(snmp_client_module, "_SnmpWireClient", FakeWireClient)

    result = SnmpClient().get_bulk(SnmpProfile(host="127.0.0.1"), "1.3.6.1.4.1.25506.2.75.2.1.1.1.1", max_repetitions=10)

    assert result.status == "success"
    assert [row.value for row in result.rows] == ["ap1", "ap2"]


def test_snmp_client_walk_reports_empty_subtree(monkeypatch):
    class FakeWireClient:
        def __init__(self, profile: SnmpProfile) -> None:
            self.profile = profile

        def request(self, oids: list[str], *, pdu_type: int, max_repetitions: int = 10) -> _WireResponse:
            assert pdu_type == 0xA1
            return _WireResponse(
                status="success",
                error_message="",
                varbinds=[_WireVarBind("1.3.6.1.4.1.25506.999", "outside", "OCTET STRING")],
            )

    monkeypatch.setattr(snmp_client_module, "_SnmpWireClient", FakeWireClient)

    result = SnmpClient().walk(SnmpProfile(host="127.0.0.1"), "1.3.6.1.4.1.25506.2.75", max_rows=10)

    assert result.status == "empty_table"
    assert result.rows == []
    assert "返回 0 条" in result.error_message


def test_mib_objects_are_sorted_by_numeric_oid(tmp_path: Path):
    repo = GlobalMibRepository(tmp_path / "global_mib.db")
    repo.initialize()
    now = "2026-07-06T00:00:00"
    with repo.connect() as conn:
        module_id = conn.execute(
            """
            INSERT INTO mib_modules (module_name, status, created_at, updated_at)
            VALUES (?, ?, ?, ?)
            """,
            ("ORDER-MIB", "compiled", now, now),
        ).lastrowid
        conn.executemany(
            """
            INSERT INTO mib_objects (module_id, name, oid, parent_oid, syntax, access, status, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (module_id, "item75", "1.3.6.1.4.1.25506.2.75.2.3.0.40", "1.3.6.1.4.1.25506.2", "Integer32", "read-only", "current", now, now),
                (module_id, "item5", "1.3.6.1.4.1.25506.2.5.1.1.3.1.1.3", "1.3.6.1.4.1.25506.2", "Integer32", "read-only", "current", now, now),
                (module_id, "item126", "1.3.6.1.4.1.25506.2.126.1.8.5", "1.3.6.1.4.1.25506.2", "Integer32", "read-only", "current", now, now),
            ],
        )
        conn.commit()

    rows = repo.list_objects(module_id=int(module_id), limit=20)
    children = repo.list_oid_children("1.3.6.1.4.1.25506.2", module_ids=[int(module_id)], limit=20)

    expected = [
        "1.3.6.1.4.1.25506.2.5.1.1.3.1.1.3",
        "1.3.6.1.4.1.25506.2.75.2.3.0.40",
        "1.3.6.1.4.1.25506.2.126.1.8.5",
    ]
    assert [row["oid"] for row in rows] == expected
    assert [row["oid"] for row in children] == expected


def test_snmp_client_requires_numeric_oid_for_final_query():
    assert normalize_oid(".1.3.6.1.2.1.1.5.0") == "1.3.6.1.2.1.1.5.0"
    assert normalize_oid("sysName.0") == "1.3.6.1.2.1.1.5.0"
    with pytest.raises(ValueError, match="数字 OID"):
        normalize_oid("sysName")


def test_snmp_set_is_disabled_by_default_and_logged(tmp_path: Path):
    paths = PathResolver(tmp_path)
    repo = SiteSnmpRepository(paths.site_snmp_db_path("demo"))
    repo.initialize()
    request = SnmpSetRequest(profile=SnmpProfile(host="127.0.0.1", community_rw="private"), oid="1.3.6.1.2.1.1.5.0", data_type="DisplayString", value="demo")

    result = SnmpQueryService(repo).set_value(request)

    assert result.status == "cancelled"
    assert repo.snmp_set_enabled() is False
    history = repo.list_set_history()
    assert history[0]["oid"] == "1.3.6.1.2.1.1.5.0"
    assert history[0]["status"] == "cancelled"


def test_snmp_set_requires_rw_community_when_enabled(tmp_path: Path):
    paths = PathResolver(tmp_path)
    repo = SiteSnmpRepository(paths.site_snmp_db_path("demo"))
    repo.initialize()
    repo.set_snmp_set_enabled(True)
    request = SnmpSetRequest(profile=SnmpProfile(host="127.0.0.1", community_ro="public", community_rw=""), oid="1.3.6.1.2.1.1.5.0", data_type="DisplayString", value="demo")

    result = SnmpQueryService(repo).set_value(request)

    assert result.status == "auth_failed"
    assert "community_rw" in result.error_message


def test_snmp_set_value_encoder_validates_basic_types():
    assert _encode_snmp_value("Integer", "12") == b"\x02\x01\x0c"
    assert _encode_snmp_value("HexString", "0x01020AFF") == b"\x04\x04\x01\x02\n\xff"
    assert _encode_snmp_value("IpAddress", "192.168.1.1") == b"\x40\x04\xc0\xa8\x01\x01"
    with pytest.raises(ValueError):
        _encode_snmp_value("Gauge32", "-1")


def test_mib_import_rejects_unrelated_text_before_creating_database(tmp_path: Path) -> None:
    paths = PathResolver(tmp_path)
    unrelated = tmp_path / "notes.txt"
    unrelated.write_text("这只是普通说明文件，不是 MIB。", encoding="utf-8")

    with pytest.raises(ImportValidationError, match="不是 NetConsole 支持的导入文件"):
        MibResourceService(paths).import_paths([unrelated])

    assert not paths.global_mib_db_path().exists()
