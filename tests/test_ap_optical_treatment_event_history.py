from __future__ import annotations

from pathlib import Path
import json

from openpyxl import load_workbook

from netconsole.core.database import Database
from netconsole.repositories.ac_repository import AcRepository
from netconsole.services.optical_retention import (
    update_ap_optical_treatment,
    upsert_optical_current_and_history,
)
from netconsole.services.trackside_ap_business import (
    AP_OPTICAL_TREATMENT_RECORD_COLUMNS,
    TRACKSIDE_AP_BUSINESS_EXPORT_COLUMNS,
    _ap_optical_treatment_row_fill_status,
    export_trackside_ap_business_xlsx,
    format_optical_event_status,
)
from netconsole.services.trackside_ap_export_service import (
    _export_persisted_optical_treatments,
)
from netconsole.core.i18n import I18n
from scripts.backfill_ap_optical_treatment_events import _apply_plan, _plan


def _record(database: Database, payload: dict[str, object], collected_at: str) -> None:
    site_id = database.path.parent.parent.name or database.path.parent.name
    with database.connect() as connection:
        for side in ("AP", "SWITCH"):
            projection = upsert_optical_current_and_history(
                connection,
                {**payload, "collected_at": collected_at},
                site_id=site_id,
                side=side,
                now=collected_at,
            )
            if projection is not None:
                update_ap_optical_treatment(
                    connection,
                    site_id=site_id,
                    ap_identity=str(projection["ap_identity"]),
                    source_row={**payload, "collected_at": collected_at},
                    now=collected_at,
                )
        connection.commit()


def _payload(**values: object) -> dict[str, object]:
    return {
        "ap_uuid": "ap-event-1",
        "ap_name": "AP-EVENT-1",
        "ap_mac": "0011-2233-4455",
        "status": "success",
        "ap_optical_status": "normal",
        "switch_optical_status": "normal",
        "rx_power": "-10",
        "switch_rx_power": "-10",
        **values,
    }


def test_event_lifecycle_keeps_history_and_summary_projection(tmp_path: Path) -> None:
    database = Database(tmp_path / "devices.db")
    database.initialize()

    _record(database, _payload(source_revision="normal-1"), "2026-09-01T00:00:00")
    alarm = _payload(
        source_revision="alarm-1",
        ap_optical_status="alarm",
        rx_power="-20",
    )
    _record(database, alarm, "2026-09-01T00:01:00")
    _record(
        database,
        _payload(source_revision="warning-1", ap_optical_status="warning", rx_power="-15"),
        "2026-09-01T00:02:00",
    )
    _record(
        database,
        _payload(
            source_revision="both-1",
            ap_optical_status="alarm",
            switch_optical_status="alarm",
            rx_power="-21",
            switch_rx_power="-22",
        ),
        "2026-09-01T00:03:00",
    )

    repository = AcRepository(database)
    events = repository.list_ap_optical_treatment_events()
    assert len(events) == 1
    assert events[0]["event_status"] == "OPEN"
    assert events[0]["worst_abnormal_side"] == "BOTH"
    assert events[0]["worst_severity"] == "alarm"
    assert events[0]["worst_ap_rx_dbm"] == "-21"
    assert events[0]["worst_switch_rx_dbm"] == "-22"

    _record(
        database,
        _payload(
            source_revision="failed-1",
            status="collection_failed",
            ap_optical_status="collection_failed",
            switch_optical_status="collection_failed",
            rx_power="-5",
            switch_rx_power="-5",
        ),
        "2026-09-01T00:04:00",
    )
    failed_events = repository.list_ap_optical_treatment_events()
    assert len(failed_events) == 1
    assert failed_events[0]["event_status"] == "OPEN"
    assert failed_events[0]["worst_ap_rx_dbm"] == "-21"

    _record(
        database,
        _payload(
            source_revision="stale-1",
            status="stale",
            ap_optical_status="stale",
            switch_optical_status="stale",
            rx_power="-4",
            switch_rx_power="-4",
        ),
        "2026-09-01T00:04:30",
    )
    _record(
        database,
        _payload(
            source_revision="not-collected-1",
            status="not_collected",
            ap_optical_status="not_collected",
            switch_optical_status="not_collected",
            rx_power="-4",
            switch_rx_power="-4",
        ),
        "2026-09-01T00:04:45",
    )
    assert len(repository.list_ap_optical_treatment_events()) == 1

    _record(
        database,
        _payload(
            source_revision="normal-2",
            ap_optical_status="normal",
            switch_optical_status="normal",
            rx_power="-9",
            switch_rx_power="-8",
        ),
        "2026-09-01T00:05:00",
    )
    _record(database, alarm, "2026-09-01T00:01:00")
    _record(database, alarm, "2026-09-01T00:01:00")
    assert len(repository.list_ap_optical_treatment_events()) == 2
    _record(
        database,
        _payload(source_revision="normal-3", rx_power="-9", switch_rx_power="-8"),
        "2026-09-01T00:06:00",
    )
    _record(
        database,
        _payload(source_revision="alarm-3", ap_optical_status="alarm", rx_power="-23"),
        "2026-09-01T00:07:00",
    )
    _record(
        database,
        _payload(source_revision="normal-4", rx_power="-9", switch_rx_power="-8"),
        "2026-09-01T00:08:00",
    )

    events = repository.list_ap_optical_treatment_events()
    assert len(events) == 3
    assert all(event["event_status"] == "RESOLVED" for event in events)
    summary = repository.list_ap_optical_treatments()
    assert len(summary) == 1
    assert summary[0]["recurrence_count"] == 2
    assert summary[0]["current_status"] in {"NORMAL", "RECOVERED"}


def test_event_migration_is_additive_and_idempotent(tmp_path: Path) -> None:
    database = Database(tmp_path / "devices.db")
    database.initialize()
    database.initialize()
    with database.connect_readonly() as connection:
        columns = {
            str(row[1])
            for row in connection.execute(
                "PRAGMA table_info(ap_optical_treatment_events)"
            )
        }
        assert {"event_uuid", "first_detected_at", "worst_rx_dbm", "evidence_json"} <= columns
        assert connection.execute(
            "SELECT COUNT(*) FROM ap_optical_treatment"
        ).fetchone()[0] == 0
        assert connection.execute("PRAGMA quick_check").fetchone()[0] == "ok"


def test_event_history_export_keeps_three_events_for_one_ap(tmp_path: Path) -> None:
    events = [
        {
            "event_uuid": f"event-{index}",
            "ap_uuid": "ap-export-1",
            "ap_name": "AP-EXPORT-1",
            "ap_mac": "0011-2233-4455",
            "serial_number": "SN-EXPORT-1",
            "station_name": "站点A",
            "worst_abnormal_side": "AP",
            "switch_name": "SW-A",
            "switch_interface": "XGE1/0/1",
            "issue_type": "alarm",
            "first_detected_at": f"2026-09-01T00:0{index}:00",
            "first_rx_dbm": f"-{20 + index}",
            "worst_rx_dbm": f"-{21 + index}",
            "recovered_rx_dbm": f"-{10 + index}",
            "event_status": "RESOLVED",
            "treatment_status": "PENDING",
            "resolved_at": f"2026-09-01T00:1{index}:00",
        }
        for index in range(1, 4)
    ]
    exported = _export_persisted_optical_treatments(events)
    assert len(exported) == 3
    assert [row["worst_rx_power"] for row in exported] == ["-22", "-23", "-24"]
    assert all(row["event_status"] == "RESOLVED" for row in exported)

    output = tmp_path / "event-history.xlsx"
    i18n = I18n("zh_CN")
    export_trackside_ap_business_xlsx(
        output,
        [],
        TRACKSIDE_AP_BUSINESS_EXPORT_COLUMNS,
        [i18n.t(key) for key, _field in TRACKSIDE_AP_BUSINESS_EXPORT_COLUMNS],
        ap_optical_treatment_rows=exported,
        ap_optical_treatment_columns=AP_OPTICAL_TREATMENT_RECORD_COLUMNS,
        ap_optical_treatment_headers=[
            i18n.t(key) for key, _field in AP_OPTICAL_TREATMENT_RECORD_COLUMNS
        ],
    )
    workbook = load_workbook(output)
    sheet = workbook["AP光衰处理记录"]
    assert sheet.max_row == 4
    assert [sheet.cell(row, 2).value for row in range(2, 5)] == [
        "AP-EXPORT-1",
        "AP-EXPORT-1",
        "AP-EXPORT-1",
    ]
    assert _ap_optical_treatment_row_fill_status(
        {"event_status": "OPEN", "treatment_status": "已处理"}
    ) == "alarm"
    assert _ap_optical_treatment_row_fill_status(
        {"event_status": "RESOLVED", "treatment_status": "未处理"}
    ) == "normal"


def test_optical_treatment_export_uses_current_side_specific_runtime_value() -> None:
    common = {
        "station_name": "站点A",
        "serial_number": "SN-1",
        "event_status": "OPEN",
        "treatment_status": "PENDING",
        "first_detected_at": "2026-09-01T00:00:00",
    }
    ap_open = {
        **common,
        "event_uuid": "event-ap-open",
        "ap_uuid": "ap-open",
        "ap_name": "AP-OPEN",
        "worst_abnormal_side": "AP",
        "first_ap_rx_dbm": "-20.0",
        "worst_ap_rx_dbm": "-22.0",
        "recovered_ap_rx_dbm": "",
    }
    switch_open = {
        **common,
        "event_uuid": "event-switch-open",
        "ap_uuid": "switch-open",
        "ap_name": "AP-SWITCH",
        "worst_abnormal_side": "SWITCH",
        "first_switch_rx_dbm": "-18.0",
        "worst_switch_rx_dbm": "-19.0",
        "recovered_switch_rx_dbm": "",
    }
    resolved = {
        **common,
        "event_uuid": "event-resolved",
        "ap_uuid": "ap-resolved",
        "ap_name": "AP-RESOLVED",
        "worst_abnormal_side": "AP",
        "first_ap_rx_dbm": "-17",
        "worst_ap_rx_dbm": "-20",
        "recovered_ap_rx_dbm": "-10",
        "event_status": "RESOLVED",
        "resolved_at": "2026-09-01T00:05:00",
    }

    exported = _export_persisted_optical_treatments(
        [ap_open, switch_open, resolved],
        business_rows=[
            {
                "ap_uuid": "ap-open",
                "ap_rx_power": "-18.0",
                "ap_optical_status": "alarm",
            },
            {
                "ap_uuid": "switch-open",
                "ap_rx_power": "-8.0",
                "switch_rx_power": "-16.0",
                "switch_optical_status": "warning",
            },
            {
                "ap_uuid": "ap-resolved",
                "ap_rx_power": "-8",
                "ap_optical_status": "normal",
            },
        ],
        current_optical_rows=[
            {"ap_identity": "ap-open", "side": "AP", "rx_dbm": "-19.0", "status": "abnormal"},
            {"ap_identity": "switch-open", "side": "SWITCH", "rx_dbm": "-15.0", "status": "abnormal"},
            {"ap_identity": "ap-resolved", "side": "AP", "rx_dbm": "-7", "status": "normal"},
        ],
    )

    assert exported[0]["first_rx_power"] == "-20.0"
    assert exported[0]["worst_rx_power"] == "-22.0"
    assert exported[0]["fixed_rx_power"] is None
    assert exported[0]["current_rx_power"] == "-18.0"
    assert exported[0]["event_status"] == "OPEN"
    assert exported[1]["current_rx_power"] == "-16.0"
    assert exported[1]["current_rx_power"] != "-8.0"
    assert exported[2]["first_rx_power"] == "-17"
    assert exported[2]["worst_rx_power"] == "-20"
    assert exported[2]["fixed_rx_power"] == "-10"
    assert exported[2]["current_rx_power"] == "-8"
    assert exported[2]["event_status"] == "RESOLVED"


def test_optical_event_status_is_chinese_only_in_user_visible_workbook(tmp_path: Path) -> None:
    rows = _export_persisted_optical_treatments(
        [
            {"event_uuid": "open", "ap_uuid": "ap-open", "event_status": "OPEN", "treatment_status": "PENDING"},
            {"event_uuid": "resolved", "ap_uuid": "ap-resolved", "event_status": "RESOLVED", "treatment_status": "PENDING"},
            {"event_uuid": "unknown", "ap_uuid": "ap-unknown", "event_status": "INTERNAL_UNKNOWN", "treatment_status": "PENDING"},
        ]
    )
    assert format_optical_event_status("OPEN") == "未恢复"
    assert format_optical_event_status("RESOLVED") == "已恢复"
    assert format_optical_event_status("INTERNAL_UNKNOWN") == "未知"

    output = tmp_path / "localized-event-status.xlsx"
    i18n = I18n("zh_CN")
    export_trackside_ap_business_xlsx(
        output,
        [],
        TRACKSIDE_AP_BUSINESS_EXPORT_COLUMNS,
        [i18n.t(key) for key, _field in TRACKSIDE_AP_BUSINESS_EXPORT_COLUMNS],
        ap_optical_treatment_rows=rows,
        ap_optical_treatment_columns=AP_OPTICAL_TREATMENT_RECORD_COLUMNS,
        ap_optical_treatment_headers=[
            i18n.t(key) for key, _field in AP_OPTICAL_TREATMENT_RECORD_COLUMNS
        ],
    )
    sheet = load_workbook(output, data_only=True)["AP光衰处理记录"]
    event_status_column = 14
    visible_statuses = [sheet.cell(row, event_status_column).value for row in range(2, 5)]
    assert visible_statuses == ["未恢复", "已恢复", "未知"]
    visible_values = [cell.value for row in sheet.iter_rows() for cell in row]
    assert "OPEN" not in visible_values
    assert "RESOLVED" not in visible_values


def test_optical_treatment_export_falls_back_to_matching_switch_side() -> None:
    exported = _export_persisted_optical_treatments(
        [
            {
                "event_uuid": "event-switch-fallback",
                "ap_uuid": "ap-switch-fallback",
                "worst_abnormal_side": "SWITCH",
                "first_switch_rx_dbm": "-18",
                "worst_switch_rx_dbm": "-19",
                "event_status": "OPEN",
                "treatment_status": "PENDING",
            }
        ],
        current_optical_rows=[
            {"ap_identity": "ap-switch-fallback", "side": "AP", "rx_dbm": "-8", "status": "normal"},
            {"ap_identity": "ap-switch-fallback", "side": "SWITCH", "rx_dbm": "-16", "status": "normal"},
        ],
        persisted_summary_rows=[
            {"ap_identity": "ap-switch-fallback", "current_switch_rx_dbm": "-15", "current_switch_status": "normal"},
        ],
    )

    assert exported[0]["current_rx_power"] == "-16"


def test_optical_treatment_export_uses_dash_for_invalid_current_value() -> None:
    exported = _export_persisted_optical_treatments(
        [
            {
                "event_uuid": "event-offline",
                "ap_uuid": "ap-offline-current",
                "worst_abnormal_side": "AP",
                "first_ap_rx_dbm": "-24.2",
                "worst_ap_rx_dbm": "-24.2",
                "event_status": "OPEN",
                "treatment_status": "PENDING",
            }
        ],
        business_rows=[
            {
                "ap_uuid": "ap-offline-current",
                "ap_rx_power": "0",
                "ap_optical_status": "AP Offline",
            }
        ],
    )

    assert exported[0]["current_rx_power"] == "-"
    assert exported[0]["current_rx_power"] not in {0, "0", "0.00", "-0.00"}

    normal_placeholder = _export_persisted_optical_treatments(
        [
            {
                "event_uuid": "event-zero-placeholder",
                "ap_uuid": "ap-zero-placeholder",
                "worst_abnormal_side": "AP",
                "first_ap_rx_dbm": "-20",
                "worst_ap_rx_dbm": "-20",
                "event_status": "OPEN",
                "treatment_status": "PENDING",
            }
        ],
        business_rows=[
            {
                "ap_uuid": "ap-zero-placeholder",
                "ap_rx_power": "0.00 dBm",
                "ap_optical_status": "normal",
            }
        ],
    )
    assert normal_placeholder[0]["current_rx_power"] == "-"


def test_optical_treatment_record_columns_have_only_business_fields() -> None:
    fields = [field for _key, field in AP_OPTICAL_TREATMENT_RECORD_COLUMNS]
    assert fields == [
        "site",
        "ap_name",
        "ap_mac",
        "serial_number",
        "side",
        "device_name",
        "interface_name",
        "issue_type",
        "first_found_at",
        "first_rx_power",
        "worst_rx_power",
        "fixed_rx_power",
        "current_rx_power",
        "event_status",
        "treatment_status",
        "remark",
        "completed_at",
    ]
    assert "current_status" not in fields


def test_backfill_plan_is_dry_run_by_default_and_idempotent(tmp_path: Path) -> None:
    database = Database(tmp_path / "devices.db")
    database.initialize()
    with database.connect() as connection:
        connection.execute(
            """
            INSERT INTO ap_optical_treatment (
                site_id, ap_identity, ap_uuid, ap_name, ap_mac,
                current_status, current_abnormal_side, first_abnormal_side,
                first_detected_at, last_abnormal_at, current_ap_status,
                first_ap_rx_dbm, current_ap_rx_dbm, source_revision
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "site-1",
                "ap-summary-1",
                "ap-summary-1",
                "AP-SUMMARY-1",
                "0011-2233-4455",
                "ABNORMAL",
                "AP",
                "AP",
                "2026-09-01T00:00:00",
                "2026-09-01T00:01:00",
                "alarm",
                "-20",
                "-20",
                "summary-revision",
            ),
        )
        connection.commit()

    evidence_dir = tmp_path / "evidence"
    evidence_dir.mkdir()
    full = {
        "audit_id": "CANONICAL-1",
        "classification": "RECOVERABLE_FROM_EXISTING_PERSISTED_EVIDENCE",
        "identity_resolution_status": "RESOLVED_PERSISTED_IDENTITY_NO_TREATMENT",
        "resolved_ap_uuid": "ap-history-1",
        "ap_name": "AP-HISTORY-1",
        "ap_mac": "0011-2233-4466",
        "recoverable_first_detected_at": "2026-08-01T00:00:00",
        "recoverable_resolved_at": "2026-08-01T00:05:00",
        "recoverable_first_rx": "-25",
        "recoverable_recovered_rx": "-8",
        "recoverable_side": "ap",
        "recoverable_issue_type": "alarm",
        "recoverable_occurrence": "FULL",
        "source_revision_evidence": "history-1",
        "event_issue_rows": {"time_max": "2026-08-01T00:04:00"},
        "raw_log_evidence": "PRESENT_DEV",
    }
    legacy = {
        "audit_id": "RECURRENCE-LEGACY",
        "classification": "LEGACY_ONLY_EVIDENCE",
        "identity_resolution_status": "RESOLVED_PERSISTED_IDENTITY_NO_TREATMENT",
        "resolved_ap_uuid": "ap-legacy-only",
    }
    (evidence_dir / "canonical_missing_42_trace.json").write_text(
        json.dumps({"rows": [full]}, ensure_ascii=False), encoding="utf-8"
    )
    (evidence_dir / "recurrence_26_trace.json").write_text(
        json.dumps({"rows": [legacy]}, ensure_ascii=False), encoding="utf-8"
    )

    with database.connect_readonly() as connection:
        plan = _plan(
            connection,
            site="site-1",
            evidence_dir=evidence_dir,
            now="2026-09-01T01:00:00",
        )
    assert plan["would_create"] == 2
    assert plan["legacy_only_skipped"] == 1
    assert plan["conflicts"] == []
    assert plan["unresolved"] == []

    with database.connect() as connection:
        _apply_plan(connection, plan, "2026-09-01T01:00:00")
    with database.connect_readonly() as connection:
        second_plan = _plan(
            connection,
            site="site-1",
            evidence_dir=evidence_dir,
            now="2026-09-01T01:01:00",
        )
        assert second_plan["would_create"] == 0
        assert connection.execute(
            "SELECT COUNT(*) FROM ap_optical_treatment_events WHERE site_id='site-1'"
        ).fetchone()[0] == 2
