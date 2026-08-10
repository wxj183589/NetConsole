from __future__ import annotations

import sqlite3
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook
from netconsole.models.ap_identity_index import ApIdentityBatchResult, ApIdentityMatch
from netconsole.services.ap_identity.normalizers import normalize_mac_key
from netconsole.services.mesh_ap_coverage_export import export_mesh_ap_coverage_audit_xlsx
from netconsole.services.rail_transit.mesh_ap_coverage_audit_service import (
    MeshApCoverageAuditService,
)


class _FakeQueryService:
    def __init__(self, contexts: dict[str, SimpleNamespace]) -> None:
        self._contexts = contexts

    def _context(self, _site_id: str, session_id: str) -> SimpleNamespace:
        return self._contexts[session_id]

    @staticmethod
    def _connect_readonly(path: Path) -> sqlite3.Connection:
        connection = sqlite3.connect(path)
        connection.row_factory = sqlite3.Row
        return connection

    @staticmethod
    def _table_columns(connection: sqlite3.Connection, table: str) -> set[str]:
        return {str(row[1]) for row in connection.execute(f"PRAGMA table_info({table})")}


class _FakeIdentityQuery:
    def __init__(self, peer_to_ap: dict[str, str], ap_metadata: dict[str, dict[str, str]]) -> None:
        self._peer_to_ap = peer_to_ap
        self._ap_metadata = ap_metadata

    def resolve_peer_macs(self, macs: list[str]) -> ApIdentityBatchResult:
        matches: dict[str, ApIdentityMatch] = {}
        for mac in macs:
            key = normalize_mac_key(mac) or ""
            physical = self._peer_to_ap.get(key, "")
            metadata = self._ap_metadata.get(normalize_mac_key(physical) or "", {})
            matches[key] = ApIdentityMatch(
                status="matched" if physical else "unresolved",
                query_mac=key,
                effective_ap_mac=physical,
                effective_ap_name=metadata.get("name", ""),
                station=metadata.get("station", ""),
                section=metadata.get("section", ""),
                direction=metadata.get("direction", ""),
                base_record_id=metadata.get("base_id", ""),
                unresolved_reason="exact_alias_not_found" if not physical else "",
            )
        return ApIdentityBatchResult(
            revision=1,
            index_status="ready",
            requested_count=len(macs),
            normalized_count=len(macs),
            distinct_count=len(matches),
            matched_count=sum(item.matched for item in matches.values()),
            unresolved_count=sum(not item.matched for item in matches.values()),
            ambiguous_count=0,
            invalid_count=0,
            matches=matches,
        )

    def resolve_current_ap_mac(self, mac: str) -> ApIdentityMatch:
        metadata = self._ap_metadata.get(normalize_mac_key(mac) or "", {})
        return ApIdentityMatch(
            status="matched",
            effective_ap_mac=mac,
            effective_ap_name=metadata.get("name", ""),
            station=metadata.get("station", ""),
            section=metadata.get("section", ""),
            direction=metadata.get("direction", ""),
            base_record_id=metadata.get("base_id", ""),
        )


def _detail_database(path: Path, rows: list[tuple[str, str, int, str, str, str]]) -> None:
    with sqlite3.connect(path) as connection:
        connection.execute(
            """
            CREATE TABLE mesh_links (
                sample_time TEXT,
                link_state TEXT,
                link_count INTEGER,
                peer_radio_mac TEXT,
                peer_mac_normalized TEXT,
                peer_mac_raw TEXT,
                peer_ap_name TEXT,
                peer_site TEXT,
                peer_section TEXT
            )
            """
        )
        connection.executemany(
            """
            INSERT INTO mesh_links VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (sample_time, state, link_count, radio_mac, radio_mac, radio_mac, name, station, section)
                for sample_time, state, link_count, radio_mac, name, station, section in rows
            ],
        )


def _base_item(
    identifier: str,
    mac: str,
    *,
    station: str,
    section: str,
    location_class: str = "MAINLINE",
    participates_in_mainline: bool = True,
) -> SimpleNamespace:
    return SimpleNamespace(
        id=f"ap:{identifier}",
        mac=mac,
        station=station,
        section=section,
        direction="上行",
        location_class=location_class,
        participates_in_mainline=participates_in_mainline,
    )


def test_ap_coverage_uses_valid_links_identity_and_observed_route_scope(tmp_path: Path) -> None:
    source_a = tmp_path / "source-a.sqlite"
    source_b = tmp_path / "source-b.sqlite"
    _detail_database(
        source_a,
        [
            ("2026-08-07 10:00:00.000", "ACTIVE", 1, "0000-0000-0011", "AP-01", "A站", "A-B"),
            ("2026-08-07 10:00:01.000", "STANDBY", 2, "0000-0000-0099", "未知 AP", "A站", "A-B"),
            ("2026-08-07 10:00:02.000", "ACTIVE", 0, "0000-0000-0022", "AP-02", "A站", "A-B"),
        ],
    )
    _detail_database(
        source_b,
        [
            ("2026-08-07 10:05:00.000", "STANDBY", 1, "0000-0000-0022", "AP-02", "B站", "B-C"),
        ],
    )
    contexts = {
        "a": SimpleNamespace(
            site_id="demo",
            session_id="a",
            mr_name="列车24-MR-CT",
            detail_db=source_a,
            source={"original_filename": "ct.log", "first_sample_time": "2026-08-07 10:00:00.000", "last_sample_time": "2026-08-07 10:00:02.000"},
        ),
        "b": SimpleNamespace(
            site_id="demo",
            session_id="b",
            mr_name="列车24-MR-CW",
            detail_db=source_b,
            source={"original_filename": "cw.log", "first_sample_time": "2026-08-07 10:05:00.000", "last_sample_time": "2026-08-07 10:05:00.000"},
        ),
    }
    physical_macs = {index: f"0000-0000-01{index:02d}" for index in range(1, 9)}
    metadata = {
        normalize_mac_key(physical_macs[index]) or "": {
            "name": f"AP-{index:02d}",
            "station": "A站" if index in {1, 3} else "B站" if index == 2 else "C站",
            "section": "A-B" if index in {1, 3} else "B-C" if index == 2 else "C-D",
            "direction": "上行",
            "base_id": f"{index}",
        }
        for index in range(1, 9)
    }
    identity = _FakeIdentityQuery(
        {
            normalize_mac_key("0000-0000-0011") or "": physical_macs[1],
            normalize_mac_key("0000-0000-0022") or "": physical_macs[2],
        },
        metadata,
    )
    base_items = [
        _base_item(str(index), physical_macs[index], station=metadata[normalize_mac_key(physical_macs[index]) or ""]["station"], section=metadata[normalize_mac_key(physical_macs[index]) or ""]["section"])
        for index in range(1, 6)
    ]
    base_items.extend(
        [
            _base_item("6", physical_macs[6], station="车辆段", section="车辆段", location_class="DEPOT", participates_in_mainline=False),
            _base_item("7", physical_macs[7], station="停车场", section="停车场", location_class="PARKING_YARD", participates_in_mainline=False),
            _base_item("8", physical_macs[8], station="出入段线", section="出入段线", location_class="DEPOT_CONNECTION", participates_in_mainline=False),
        ]
    )
    ac_details = [
        SimpleNamespace(ap=SimpleNamespace(mac=item.mac, name=f"AP-{index:02d}", status="Online", updated_at="2026-08-11"))
        for index, item in enumerate(base_items, start=1)
    ]
    base_query = SimpleNamespace(
        list_ap_status_items=lambda _site_id: base_items,
        ac_query=SimpleNamespace(list_all_ap_details=lambda _site_id: ac_details),
    )

    result = MeshApCoverageAuditService(
        _FakeQueryService(contexts), base_query, identity_query=identity  # type: ignore[arg-type]
    ).audit("demo", ["a", "b"])

    assert result.summary.expected_mainline_count == 5
    assert result.summary.expected_route_scope_count == 3
    assert result.summary.connected_count == 2
    assert result.summary.unconnected_count == 1
    assert result.summary.full_mainline_unconnected_count == 3
    assert result.summary.unmatched_observed_count == 1
    assert result.summary.excluded_count == 3
    assert result.summary.coverage_percent == 66.67
    assert {item.ap_name for item in result.connected} == {"AP-01", "AP-02"}
    assert {item.ap_name for item in result.unconnected if item.in_observed_route_scope} == {"AP-03"}
    assert next(item for item in result.connected if item.ap_name == "AP-02").seen_in_source_a is False
    assert next(item for item in result.connected if item.ap_name == "AP-02").seen_in_source_b is True
    assert result.unmatched[0].triangle_link_count == 1
    assert {item.exclude_reason for item in result.excluded} == {"车辆段", "停车场", "出入段线"}

    output = tmp_path / "coverage.xlsx"
    assert export_mesh_ap_coverage_audit_xlsx(output, result) == 9
    assert output.is_file()
    workbook = load_workbook(output, read_only=True)
    assert workbook.sheetnames == ["核查摘要", "未连接 AP", "已连接 AP", "资料未匹配", "排除 AP"]
