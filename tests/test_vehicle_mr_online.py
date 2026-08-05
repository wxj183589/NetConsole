from __future__ import annotations

import json
import sqlite3
from pathlib import Path


from netconsole.core.database import Database
from netconsole.core.paths import PathResolver
from netconsole.models.ap_identity_index import ApIdentityBatchResult, ApIdentityMatch
from netconsole.models.device import Device
from netconsole.models.online_mr_models import OnlineMrConnectionConfig
from netconsole.repositories.device_group_repository import DeviceGroupRepository
from netconsole.repositories.device_repository import DeviceRepository
from netconsole.services.ap_identity import ApIdentityQueryService
from netconsole.services.ap_identity.normalizers import normalize_mac_key
from netconsole.services import vehicle_mr_online as vehicle_mr_online_service
from netconsole.services.export.export_handlers import run_generic_export_handler
from netconsole.services.export.export_task_builders import vehicle_mr_history_xlsx_spec
from netconsole.services.export_task_models import ExportJob
from netconsole.services.vehicle_mr_online import (
    H3CComwareV9VehicleMrMeshLinkParser,
    MatchedAp,
    ONLINE_POLICY_DUAL_ACTIVE,
    ONLINE_POLICY_SINGLE_TAIL,
    ONLINE_POLICY_SINGLE_TC1,
    ONLINE_POLICY_SINGLE_TC2,
    UNKNOWN_STATION,
    VehicleMrIdentitySnapshot,
    VehicleMrMeshLink,
    VehicleMrMeshParseResult,
    VehicleMrEndState,
    VehicleMrOnlineStore,
    VehicleMrTrainState,
    backfill_fit_ap_resource_station_from_optical,
    build_mapping_lookup,
    build_registered_trains,
    build_canonical_train_key,
    build_train_states,
    choose_best_links,
    load_trackside_ap_lookup,
    load_vehicle_mr_mapping_trains,
    match_ap,
    normalize_mac,
    normalize_train_no,
    parse_ac_clock_line,
    parse_train_identity,
    resolve_vehicle_mr_identity_snapshot,
    resolve_ap_station,
)
from netconsole.services.job_center.handlers.legacy_tasks import (
    _jsonable_vehicle_ap_lookup,
)






SAMPLE = """<NBDT12HX-WX3540X-AC1>display clock
22:01:45 Beijing Thu 06/25/2026
Time Zone : Beijing add 08:00:00
<NBDT12HX-WX3540X-AC1>display  wlan  mesh-link ap
AP name: 30f5-277a-8440
 Peer Name              Peer Mac       Local Mac      Status     RSSI Packets(Rx/Tx)
 NBL12-LC06-MR-CW       74ad-cb9d-317f 30f5-277a-844f Forwarding 41   7995/5412

AP name: bc5a-3457-cde0
 Peer Name              Peer Mac       Local Mac      Status     RSSI Packets(Rx/Tx)
 NBL12-LC06-MR-CT       74ad-cb9d-3321 bc5a-3457-cdef Forwarding 58   3202/5050
<NBDT12HX-WX3540X-AC1>
"""


def test_h3c_v9_vehicle_mr_mesh_link_parser_parses_sample() -> None:
    result = H3CComwareV9VehicleMrMeshLinkParser().parse(SAMPLE)

    assert result.parse_status == "OK"
    assert result.ac_time == "2026-06-25 22:01:45"
    assert len(result.links) == 2
    assert {link.local_ap_name for link in result.links} == {"30f5-277a-8440", "bc5a-3457-cde0"}
    by_peer = {link.peer_name: link for link in result.links}
    assert by_peer["NBL12-LC06-MR-CT"].rssi == 58
    assert by_peer["NBL12-LC06-MR-CW"].rx_packets == 7995
    assert by_peer["NBL12-LC06-MR-CW"].tx_packets == 5412


def test_h3c_v9_parser_keeps_peer_name_with_spaces_and_canonical_identity() -> None:
    sample = """AP name: AP-TCC-16
 Peer Name              Peer Mac       Local Mac      Status     RSSI Packets(Rx/Tx)
 Nbl06-LC12-AP- CT      78a1-3e52-4d4f 30f5-277a-478f Forwarding 39   53157/270011
 Nbl06-LC20-AP- CT      78a1-3e52-554f 30f5-277a-479f Forwarding 41   57/71

AP name: AP-TCC-15
 Peer Name              Peer Mac       Local Mac      Status     RSSI Packets(Rx/Tx)
 Nbl06-LC14-AP-CW       eccd-4c04-b30f 30f5-277a-4e1f Forwarding 43   448470/1852345
"""

    result = H3CComwareV9VehicleMrMeshLinkParser().parse(sample)
    identities = {link.peer_name: parse_train_identity(link.peer_name) for link in result.links}

    assert result.parse_status == "OK"
    assert [link.peer_name for link in result.links] == ["Nbl06-LC12-AP- CT", "Nbl06-LC20-AP- CT", "Nbl06-LC14-AP-CW"]
    assert result.links[0].peer_name_canonical == "Nbl06-LC12-AP-CT"
    assert identities["Nbl06-LC12-AP- CT"].train_no == "12"
    assert identities["Nbl06-LC12-AP- CT"].car_end_label == "TC1"
    assert identities["Nbl06-LC14-AP-CW"].car_end_label == "TC2"


def test_vehicle_mr_online_uses_ap_names_and_status_filter() -> None:
    result = VehicleMrMeshParseResult(
        "2026-06-28 10:00:00",
        [
            VehicleMrMeshLink("AP-A", "Nbl06-LC14-AP-CW", rssi=37, status="Down"),
            VehicleMrMeshLink("AP-B", "Nbl06-LC14-AP-CW", rssi=43, status="Forwarding"),
            VehicleMrMeshLink("AP-C", "Nbl06-LC14-AP-CW", rssi=39, status="Active"),
        ],
    )

    trains = build_train_states({"LC14": VehicleMrTrainState("LC14", "14", True)}, result, {})

    assert trains[0].tc2.seen is True
    assert trains[0].tc2.ap_name == "AP-B"
    assert trains[0].status == "在线"


def test_parse_ac_clock_line_full_datetime_and_fallback() -> None:
    fallback = __import__("datetime").datetime(2026, 6, 26, 1, 2, 3)

    assert parse_ac_clock_line("22:01:45 Beijing Thu 06/25/2026", fallback) == "2026-06-25 22:01:45"
    assert parse_ac_clock_line("22:01:45", fallback) == "2026-06-26 22:01:45"
    assert parse_ac_clock_line("bad clock", fallback) == "2026-06-26 01:02:03"


def test_train_no_canonical_names_merge() -> None:
    values = ["06车", "6车", "列车06", "列车6", "LC06", "NBL12-LC06", "列车06-MR-CT", "NBL12-LC06-MR-CT"]

    assert {normalize_train_no(value) for value in values} == {"06"}
    assert {build_canonical_train_key(value) for value in values} == {"train:06"}


def test_parse_train_identity_merges_ct_and_cw_to_same_train() -> None:
    ct = parse_train_identity("NBL12-LC06-MR-CT")
    cw = parse_train_identity("NBL12-LC06-MR-CW")

    assert ct is not None
    assert cw is not None
    assert ct.train_id == "NBL12-LC06"
    assert cw.train_id == "NBL12-LC06"
    assert ct.train_no == "06"
    assert ct.car_end_label == "TC1"
    assert cw.car_end_label == "TC2"


def test_train_status_online_partial_offline_and_unregistered() -> None:
    registered = {"NBL12-LC06": VehicleMrTrainState("NBL12-LC06", "06", True)}
    ap_lookup = {"ap-a": MatchedAp("AP-A", "鼓楼站")}

    both = build_train_states(
        registered,
        VehicleMrMeshParseResult(
            "22:01:45",
            [
                VehicleMrMeshLink("AP-A", "NBL12-LC06-MR-CT", rssi=50),
                VehicleMrMeshLink("AP-A", "NBL12-LC06-MR-CW", rssi=40),
            ],
        ),
        ap_lookup,
    )
    assert both[0].status == "在线"

    only_tc1 = build_train_states(registered, VehicleMrMeshParseResult("22:01:46", [VehicleMrMeshLink("AP-A", "NBL12-LC06-MR-CT", rssi=50)]), ap_lookup)
    assert only_tc1[0].status == "在线"

    only_tc2 = build_train_states(registered, VehicleMrMeshParseResult("22:01:47", [VehicleMrMeshLink("AP-A", "NBL12-LC06-MR-CW", rssi=50)]), ap_lookup)
    assert only_tc2[0].status == "在线"

    offline = build_train_states(registered, VehicleMrMeshParseResult("22:01:48", []), ap_lookup)
    assert offline[0].status == "离线"

    unregistered = build_train_states(registered, VehicleMrMeshParseResult("22:01:49", [VehicleMrMeshLink("AP-A", "NBL12-LC19-MR-CT", rssi=60)]), ap_lookup)
    by_id = {train.train_id: train for train in unregistered}
    assert by_id["NBL12-LC19"].is_registered is False
    assert by_id["NBL12-LC19"].status == "在线"


def test_registered_train_no_merges_ac_peer_with_different_prefix() -> None:
    registered = {"列车06": VehicleMrTrainState("列车06", "06", True)}
    result = VehicleMrMeshParseResult(
        "00:22:05",
        [
            VehicleMrMeshLink("AP-A", "NBL12-LC06-MR-CT", rssi=45),
            VehicleMrMeshLink("AP-B", "NBL12-LC06-MR-CW", rssi=52),
        ],
    )

    trains = build_train_states(registered, result, {})

    assert len(trains) == 1
    assert trains[0].train_id == "列车06"
    assert trains[0].display_name == "06车"
    assert trains[0].is_registered is True
    assert trains[0].status == "在线"


def test_best_ap_chooses_highest_rssi_and_ignores_missing_rssi() -> None:
    links = [
        VehicleMrMeshLink("AP-A", "NBL12-LC06-MR-CT", rssi=37),
        VehicleMrMeshLink("AP-B", "NBL12-LC06-MR-CT", rssi=None),
        VehicleMrMeshLink("AP-C", "NBL12-LC06-MR-CT", rssi=58),
    ]

    best = choose_best_links(links)

    assert best["NBL12-LC06-MR-CT"].local_ap_name == "AP-C"
    states = build_train_states({}, VehicleMrMeshParseResult("22:01:45", links), {})
    assert states[0].tc1.station == "未知车站"
    assert states[0].tc1.ap_name == "AP-C"


def test_registered_vehicle_mr_devices_prebuild_train_rows(tmp_path: Path) -> None:
    paths = PathResolver(tmp_path)
    database = Database(paths.site_db_path("demo"))
    database.initialize()
    repository = DeviceRepository(database)
    group = DeviceGroupRepository(database, "demo").create("车载")
    for index in range(1, 19):
        for offset, suffix in enumerate(("CT", "CW")):
            repository.create(Device(name=f"NBL12-LC{index:02d}-MR-{suffix}", group_id=group.id, device_type="FAT-AP", primary_address=f"192.0.2.{index * 2 - 1 + offset}"))

    trains = build_registered_trains(repository.list(), {group.id: group.name})

    assert len(trains) == 18
    assert "NBL12-LC01" in trains
    assert "NBL12-LC18" in trains


def test_vehicle_group_chinese_names_and_station_fallback_generate_ordered_trains(tmp_path: Path) -> None:
    paths = PathResolver(tmp_path)
    database = Database(paths.site_db_path("demo"))
    database.initialize()
    repository = DeviceRepository(database)
    group = DeviceGroupRepository(database, "demo").create("车载")
    for index in range(1, 19):
        if index in {1, 18}:
            repository.create(Device(name=f"列车{index:02d}-MR-CT", group_id=group.id, device_type="FAT-AP", primary_address=f"192.0.2.{index}", station=f"{index:02d}车车头"))
            repository.create(Device(name=f"列车{index:02d}-MR-CW", group_id=group.id, device_type="FAT-AP", primary_address=f"192.0.3.{index}", station=f"{index:02d}车车尾"))
        else:
            repository.create(Device(name=f"MR-{index:02d}-A", group_id=group.id, device_type="FAT-AP", primary_address=f"192.0.2.{index}", station=f"{index:02d}车车头"))
            repository.create(Device(name=f"MR-{index:02d}-B", group_id=group.id, device_type="FAT-AP", primary_address=f"192.0.3.{index}", station=f"{index:02d}车车尾"))

    trains = build_registered_trains(repository.list(), {group.id: group.name})
    ordered = sorted(trains.values(), key=lambda train: int(train.train_no))

    assert len(ordered) == 18
    assert [train.display_name for train in ordered[:2]] == ["01车", "02车"]
    assert ordered[-1].display_name == "18车"


def test_optional_vehicle_mr_mapping_table_prebuilds_trains(tmp_path: Path) -> None:
    paths = PathResolver(tmp_path)
    database = Database(paths.site_db_path("demo"))
    database.initialize()
    repository = DeviceRepository(database)
    with database.connect() as conn:
        conn.execute("CREATE TABLE vehicle_mr_mapping (peer_name TEXT, train_id TEXT, train_no TEXT, car_end TEXT)")
        conn.execute("INSERT INTO vehicle_mr_mapping VALUES ('custom', 'MAP-LC06', '06', 'TC1')")
        conn.commit()

    trains = load_vehicle_mr_mapping_trains(repository)

    assert trains["MAP-LC06"].display_name == "06车"


def test_mapping_import_and_peer_lookup_priority(tmp_path: Path) -> None:
    store = VehicleMrOnlineStore(PathResolver(tmp_path), "demo")
    count = store.import_mapping_rows(
        [
            {"车次": "1车", "TC1": "0101", "TC2": "0106", "备注": ""},
            {"车次": "2车", "TC1": "0201", "TC2": "0206", "备注": ""},
        ]
    )
    lookup = build_mapping_lookup(store.list_mappings())
    result = VehicleMrMeshParseResult("2026-06-26 01:15:38", [VehicleMrMeshLink("AP-A", "0101", rssi=45), VehicleMrMeshLink("AP-B", "0106", rssi=46)])

    trains = build_train_states({}, result, {}, {}, lookup)

    assert count == 2
    assert lookup["0101"].train_no == "01"
    assert trains[0].display_name == "01车"
    assert trains[0].status == "在线"


def test_mapping_edit_replaces_old_peer(tmp_path: Path) -> None:
    store = VehicleMrOnlineStore(PathResolver(tmp_path), "demo")
    store.import_mapping_rows([{"车次": "1车", "TC1": "0101", "TC2": "0106", "备注": ""}])
    mapping = store.list_mappings()[0]
    mapping.tc1_peer_name = "01A"
    store.save_mappings([mapping])
    lookup = build_mapping_lookup(store.list_mappings())

    assert "01A" in lookup
    assert "0101" not in lookup




def test_online_policy_dual_active_marks_missing_end_abnormal() -> None:
    registered = {
        "列车06": VehicleMrTrainState("列车06", "06", True, online_policy=ONLINE_POLICY_DUAL_ACTIVE),
    }

    trains = build_train_states(registered, VehicleMrMeshParseResult("00:01:00", [VehicleMrMeshLink("AP-A", "列车06-MR-CT", rssi=45)]), {})

    assert trains[0].status == "异常单端"
    assert trains[0].status_reason == "tc2_missing"


def test_online_policy_single_tc1_and_tc2_expected_end_rules() -> None:
    tc1_registered = {"列车06": VehicleMrTrainState("列车06", "06", True, online_policy=ONLINE_POLICY_SINGLE_TC1)}
    tc1_expected = build_train_states(tc1_registered, VehicleMrMeshParseResult("00:01:00", [VehicleMrMeshLink("AP-A", "列车06-MR-CT", rssi=45)]), {})
    tc1_unexpected = build_train_states(tc1_registered, VehicleMrMeshParseResult("00:02:00", [VehicleMrMeshLink("AP-B", "列车06-MR-CW", rssi=45)]), {})

    assert tc1_expected[0].status == "在线"
    assert tc1_expected[0].expected_end == "TC1"
    assert tc1_unexpected[0].status == "非预期端在线"

    tc2_registered = {"列车07": VehicleMrTrainState("列车07", "07", True, online_policy=ONLINE_POLICY_SINGLE_TC2)}
    tc2_expected = build_train_states(tc2_registered, VehicleMrMeshParseResult("00:03:00", [VehicleMrMeshLink("AP-C", "列车07-MR-CW", rssi=45)]), {})

    assert tc2_expected[0].status == "在线"
    assert tc2_expected[0].expected_end == "TC2"


def test_online_policy_single_tail_unknown_direction_treats_any_end_online() -> None:
    registered = {"列车06": VehicleMrTrainState("列车06", "06", True, online_policy=ONLINE_POLICY_SINGLE_TAIL)}

    trains = build_train_states(registered, VehicleMrMeshParseResult("00:01:00", [VehicleMrMeshLink("AP-A", "列车06-MR-CT", rssi=45)]), {})

    assert trains[0].status == "在线"
    assert trains[0].direction == "未知"
    assert trains[0].status_reason == "direction_unknown_any_end_online"


def test_mapping_import_old_template_defaults_online_policy_auto(tmp_path: Path) -> None:
    store = VehicleMrOnlineStore(PathResolver(tmp_path), "demo")

    store.import_mapping_rows([{"车次": "1车", "TC1": "0101", "TC2": "0106", "备注": "old"}])

    mapping = store.list_mappings()[0]
    assert mapping.online_policy == "auto"
    assert build_mapping_lookup([mapping])["0101"].online_policy == "auto"


def test_mapping_import_and_save_preserve_online_policy(tmp_path: Path) -> None:
    store = VehicleMrOnlineStore(PathResolver(tmp_path), "demo")

    store.import_mapping_rows([{"车次": "1车", "TC1": "0101", "TC2": "0106", "在线策略": "双端在线", "备注": ""}])
    mapping = store.list_mappings()[0]

    assert mapping.online_policy == ONLINE_POLICY_DUAL_ACTIVE
    trains = build_train_states({}, VehicleMrMeshParseResult("00:01:00", [VehicleMrMeshLink("AP-A", "0101", rssi=45)]), {}, {}, build_mapping_lookup([mapping]))
    assert trains[0].status == "异常单端"


def test_pass_events_store_policy_reason_fields(tmp_path: Path) -> None:
    store = VehicleMrOnlineStore(PathResolver(tmp_path), "demo")
    train = VehicleMrTrainState(
        "列车06",
        "06",
        True,
        status="异常单端",
        current_station="鼓楼站",
        last_ac_time="2026-06-26 09:00:00",
        tc1=VehicleMrEndState(True, "鼓楼站", "AP-A", 46, "2026-06-26 09:00:00"),
        online_policy=ONLINE_POLICY_DUAL_ACTIVE,
        status_reason="tc2_missing",
    )

    store.persist_snapshot("s1", 1, VehicleMrMeshParseResult(train.last_ac_time, []), [train], {}, 10)

    event = store.list_events("列车06", 1)[0]
    assert event["online_policy"] == ONLINE_POLICY_DUAL_ACTIVE
    assert event["status_reason"] == "tc2_missing"
    assert event["status"] == "异常单端"


def test_vehicle_mr_store_persists_unregistered_current_state(tmp_path: Path) -> None:
    paths = PathResolver(tmp_path)
    store = VehicleMrOnlineStore(paths, "demo")
    train = VehicleMrTrainState("NBL12-LC19", "19", False, status="单端在线", current_station="未知车站", last_ac_time="22:01:45")
    result = VehicleMrMeshParseResult("22:01:45", [VehicleMrMeshLink("AP-X", "NBL12-LC19-MR-CT", rssi=61)])
    store.persist_snapshot("s1", 1, result, [train], {}, 12)

    reloaded = store.list_current_states()

    assert reloaded[0].train_id == "NBL12-LC19"
    assert reloaded[0].is_registered is False
    with sqlite3.connect(store.db_path) as conn:
        assert conn.execute("SELECT COUNT(*) FROM vehicle_mr_online_links").fetchone()[0] == 1


def test_cleanup_history_keeps_recent_mapping_current_state_and_external_tables(tmp_path: Path) -> None:
    paths = PathResolver(tmp_path)
    database = Database(paths.site_db_path("demo"))
    database.initialize()
    repository = DeviceRepository(database)
    repository.create(Device(name="AC1", device_type="AC", primary_address="192.0.2.1"))
    store = VehicleMrOnlineStore(paths, "demo")
    now = __import__("datetime").datetime(2026, 6, 26, 0, 0, 0)
    old = "2026-05-26 00:00:00"
    recent = "2026-05-28 00:00:00"
    store.import_mapping_rows([{"车次": "1车", "TC1": "0101", "TC2": "0106", "备注": ""}])
    with store.connect() as conn:
        conn.execute("INSERT INTO vehicle_mr_online_sessions (session_id, created_at) VALUES ('old', ?)", (old,))
        conn.execute("INSERT INTO vehicle_mr_online_sessions (session_id, created_at) VALUES ('recent', ?)", (recent,))
        conn.execute("INSERT INTO vehicle_mr_online_snapshots (session_id, created_at) VALUES ('old', ?)", (old,))
        conn.execute("INSERT INTO vehicle_mr_online_links (session_id, created_at) VALUES ('old', ?)", (old,))
        conn.execute("INSERT INTO vehicle_mr_train_pass_events (train_id, created_at) VALUES ('old', ?)", (old,))
        conn.execute("INSERT INTO vehicle_mr_train_current_state (train_id, train_no, updated_at) VALUES ('列车01', '01', ?)", (old,))

    deleted = store.cleanup_history(30, now=now)

    with store.connect() as conn:
        assert conn.execute("SELECT COUNT(*) FROM vehicle_mr_online_sessions WHERE session_id='old'").fetchone()[0] == 0
        assert conn.execute("SELECT COUNT(*) FROM vehicle_mr_online_sessions WHERE session_id='recent'").fetchone()[0] == 1
        assert conn.execute("SELECT COUNT(*) FROM vehicle_mr_train_mapping").fetchone()[0] == 1
        assert conn.execute("SELECT COUNT(*) FROM vehicle_mr_train_current_state").fetchone()[0] == 1
    assert repository.list()
    assert deleted >= 4


def test_vehicle_mr_does_not_match_same_name_or_ap_base_mac(tmp_path: Path) -> None:
    paths = PathResolver(tmp_path)
    database = Database(paths.site_db_path("demo"))
    database.initialize()
    repository = DeviceRepository(database)
    with database.connect() as conn:
        conn.execute(
            """
            INSERT INTO ac_fit_ap_resources (
                ac_device_uuid, ap_uuid, ap_name, ap_mac, site, collected_at, updated_at
            ) VALUES ('ac1', 'ap1', 'bc5a-3457-a740', 'bc5a-3457-a740', '鼓楼站', 'now', 'now')
            """
        )
        conn.commit()
    ApIdentityQueryService(database).rebuild_index("test_ac_refresh_succeeded")
    query_service = load_trackside_ap_lookup(repository)
    result = VehicleMrMeshParseResult("00:22:05", [VehicleMrMeshLink("bc5a-3457-a740", "列车06-MR-CT", local_mac="bc5a-3457-a740", rssi=46)])
    identity_snapshot = resolve_vehicle_mr_identity_snapshot(query_service, result.links)
    trains = build_train_states(
        {"列车06": VehicleMrTrainState("列车06", "06", True)},
        result,
        identity_snapshot,
    )

    assert trains[0].current_station == UNKNOWN_STATION
    assert trains[0].tc1.display() == "未知车站 / bc5a-3457-a740 / 46"
    store = VehicleMrOnlineStore(paths, "demo")
    store.persist_snapshot("s1", 1, result, trains, identity_snapshot, 10)
    with sqlite3.connect(store.db_path) as conn:
        row = conn.execute("SELECT matched_station, matched_ap_name, match_method, match_score FROM vehicle_mr_online_links").fetchone()
    assert row == (
        UNKNOWN_STATION,
        "bc5a-3457-a740",
        "unmatched",
        0,
    )


def test_resolve_ap_station_prefers_optical_site() -> None:
    station, source = resolve_ap_station(
        {
            "resource.ap_name": "30f5-2787-a560",
            "resource.site": "-",
            "metadata.site_name": "-",
            "optical.site": "11云龙车辆段",
        }
    )

    assert station == "11云龙车辆段"
    assert source == "optical.site"


def test_optical_site_backfills_empty_fit_ap_resource_site(tmp_path: Path) -> None:
    paths = PathResolver(tmp_path)
    database = Database(paths.site_db_path("demo"))
    database.initialize()
    repository = DeviceRepository(database)
    with database.connect() as conn:
        conn.execute(
            """
            INSERT INTO ac_fit_ap_resources (
                ac_device_uuid, ap_uuid, ap_name, ap_mac, site, collected_at, updated_at
            ) VALUES ('ac1', 'ap1', '30f5-2787-a560', '30f5-2787-a560', '-', 'now', 'now')
            """
        )
        conn.execute(
            """
            INSERT INTO ac_fit_ap_optical (
                ac_device_uuid, ap_uuid, ap_name, ap_mac, site
            ) VALUES ('ac1', 'ap1', '30f5-2787-a560', '30f5-2787-a560', '11云龙车辆段')
            """
        )
        conn.commit()

    changed = backfill_fit_ap_resource_station_from_optical(repository)

    assert changed == 1
    with database.connect() as conn:
        assert conn.execute("SELECT site FROM ac_fit_ap_resources WHERE ap_name='30f5-2787-a560'").fetchone()[0] == "11云龙车辆段"


def test_optical_site_backfill_does_not_match_same_name_with_different_mac(tmp_path: Path) -> None:
    paths = PathResolver(tmp_path)
    database = Database(paths.site_db_path("demo"))
    database.initialize()
    repository = DeviceRepository(database)
    with database.connect() as conn:
        conn.execute(
            """
            INSERT INTO ac_fit_ap_resources (
                ac_device_uuid, ap_uuid, ap_name, ap_mac, site, collected_at, updated_at
            ) VALUES ('ac1', 'resource-ap', 'same-name', '0011-2233-4455', '-', 'now', 'now')
            """
        )
        conn.execute(
            """
            INSERT INTO ac_fit_ap_optical (
                ac_device_uuid, ap_uuid, ap_name, ap_mac, site
            ) VALUES ('ac1', 'optical-ap', 'same-name', '0011-2233-5566', 'Station A')
            """
        )
        conn.commit()

    assert backfill_fit_ap_resource_station_from_optical(repository) == 0
    with database.connect() as conn:
        assert conn.execute("SELECT site FROM ac_fit_ap_resources WHERE ap_uuid='resource-ap'").fetchone()[0] == "-"


def test_optical_site_backfill_skips_ambiguous_mac_sites(tmp_path: Path) -> None:
    paths = PathResolver(tmp_path)
    database = Database(paths.site_db_path("demo"))
    database.initialize()
    repository = DeviceRepository(database)
    with database.connect() as conn:
        conn.execute(
            """
            INSERT INTO ac_fit_ap_resources (
                ac_device_uuid, ap_uuid, ap_name, ap_mac, site, collected_at, updated_at
            ) VALUES ('ac1', 'resource-ap', 'ap-a', '0011-2233-4455', '-', 'now', 'now')
            """
        )
        conn.executemany(
            """
            INSERT INTO ac_fit_ap_optical (
                ac_device_uuid, ap_uuid, ap_name, ap_mac, site
            ) VALUES ('ac1', ?, ?, ?, ?)
            """,
            [
                ("optical-a", "renamed-a", "0011-2233-4455", "Station A"),
                ("optical-b", "renamed-b", "0011-2233-4455", "Station B"),
            ],
        )
        conn.commit()

    assert backfill_fit_ap_resource_station_from_optical(repository) == 0
    with database.connect() as conn:
        assert conn.execute("SELECT site FROM ac_fit_ap_resources WHERE ap_uuid='resource-ap'").fetchone()[0] == "-"


def test_online_vehicle_mr_uses_optical_site_for_station_display(tmp_path: Path) -> None:
    paths = PathResolver(tmp_path)
    database = Database(paths.site_db_path("demo"))
    database.initialize()
    repository = DeviceRepository(database)
    with database.connect() as conn:
        conn.execute(
            """
            INSERT INTO ac_fit_ap_resources (
                ac_device_uuid, ap_uuid, ap_name, ap_mac, site, rid1_bbssid,
                collected_at, updated_at
            ) VALUES (
                'ac1', 'ap1', '30f5-2787-a560', '30f5-2787-a560', '-',
                '30f5-2787-a56f', 'now', 'now'
            )
            """
        )
        conn.execute("INSERT INTO ac_fit_ap_optical (ac_device_uuid, ap_uuid, ap_name, ap_mac, site) VALUES ('ac1', 'ap1', '30f5-2787-a560', '30f5-2787-a560', '11云龙车辆段')")
        conn.commit()
    ApIdentityQueryService(database).rebuild_index("test_legacy_optical_loaded")
    query_service = load_trackside_ap_lookup(repository)
    result = VehicleMrMeshParseResult("00:22:05", [VehicleMrMeshLink("30f5-2787-a560", "NBL12-LC06-MR-CT", local_mac="30f5-2787-a56f", rssi=45)])

    identity_snapshot = resolve_vehicle_mr_identity_snapshot(query_service, result.links)
    trains = build_train_states(
        {"列车06": VehicleMrTrainState("列车06", "06", True)},
        result,
        identity_snapshot,
    )

    assert trains[0].current_station == "11云龙车辆段"
    assert trains[0].tc1.display() == "11云龙车辆段 / 30f5-2787-a560 / 45"


def test_vehicle_mr_lookup_job_payload_is_json_safe(tmp_path: Path) -> None:
    paths = PathResolver(tmp_path)
    database = Database(paths.site_db_path("demo"))
    database.initialize()
    with database.connect() as conn:
        conn.execute(
            """
            INSERT INTO ac_fit_ap_optical (
                ac_device_uuid, ap_uuid, ap_name, ap_mac, site
            ) VALUES ('ac1', 'ap1', 'AP-1', '30f5-2787-a560', '鼓楼站')
            """
        )
        conn.commit()
    ApIdentityQueryService(database).rebuild_index("test_legacy_optical_loaded")

    payload = _jsonable_vehicle_ap_lookup(
        load_trackside_ap_lookup(DeviceRepository(database))
    )

    assert payload["identity_entities"][0]["station"] == "鼓楼站"
    json.dumps(payload, ensure_ascii=False)


def test_radio_mac_different_from_canonical_mac_stays_unresolved() -> None:
    snapshot = VehicleMrIdentitySnapshot()

    matched = match_ap("unknown-ap", snapshot, "bc5a-3457-a750")

    assert normalize_mac("BC5A-3457-A740") == "bc:5a:34:57:a7:40"
    assert matched is None




def test_vehicle_mr_identity_snapshot_batches_50000_facts_once() -> None:
    macs = [f"02000000{index:04x}" for index in range(200)]
    links = [
        VehicleMrMeshLink(
            local_ap_name=f"AP-{index % 200}",
            peer_name="NBL12-LC06-MR-CT",
            local_mac=macs[index % 200],
            rssi=40,
        )
        for index in range(50_000)
    ]

    class QueryService:
        def __init__(self) -> None:
            self.calls: list[tuple[list[object], str | None]] = []

        def resolve_peer_macs(self, values, *, ap_role=None):
            self.calls.append((list(values), ap_role))
            keys = tuple(
                dict.fromkeys(
                    key for value in values if (key := normalize_mac_key(value))
                )
            )
            matches = {
                key: ApIdentityMatch(
                    status="matched",
                    identity_revision=41,
                    query_mac=key,
                    matched_entity_id=f"entity-{key}",
                    effective_ap_name=f"AP-{key[-2:]}",
                    effective_ap_mac=key,
                    station="鼓楼站",
                    section="鼓楼-东门口",
                    matched_alias_type="ac_radio_mac",
                    matched_source="ac_runtime",
                    match_rule="ac_radio_mac",
                    match_confidence=100,
                    radio_id=1,
                )
                for key in keys
            }
            return ApIdentityBatchResult(
                revision=41,
                index_status="ready",
                requested_count=len(values),
                normalized_count=len(values),
                distinct_count=len(keys),
                matched_count=len(keys),
                unresolved_count=0,
                ambiguous_count=0,
                invalid_count=0,
                matches=matches,
            )

    query = QueryService()
    snapshot = resolve_vehicle_mr_identity_snapshot(query, links)  # type: ignore[arg-type]

    assert len(query.calls) == 1
    assert query.calls[0][1] == "trackside"
    assert len(query.calls[0][0]) == 50_000
    assert snapshot.revision == 41
    assert snapshot.requested_count == 50_000
    assert snapshot.distinct_count == 200


def test_vehicle_mr_identity_snapshot_preserves_ambiguous_invalid_and_revision(tmp_path: Path) -> None:
    ambiguous_mac = "020000000001"
    unresolved_mac = "020000000002"
    links = [
        VehicleMrMeshLink("AP-A", "NBL12-LC06-MR-CT", local_mac=ambiguous_mac, rssi=40),
        VehicleMrMeshLink("AP-B", "NBL12-LC06-MR-CW", local_mac=unresolved_mac, rssi=41),
        VehicleMrMeshLink("AP-C", "NBL12-LC07-MR-CT", local_mac="invalid", rssi=42),
    ]

    class QueryService:
        def resolve_peer_macs(self, values, *, ap_role=None):
            assert ap_role == "trackside"
            return ApIdentityBatchResult(
                revision=52,
                index_status="ready",
                requested_count=len(values),
                normalized_count=2,
                distinct_count=2,
                matched_count=0,
                unresolved_count=1,
                ambiguous_count=1,
                invalid_count=1,
                matches={
                    ambiguous_mac: ApIdentityMatch(
                        status="ambiguous",
                        identity_revision=52,
                        query_mac=ambiguous_mac,
                        unresolved_reason="duplicate_exact_alias",
                    ),
                    unresolved_mac: ApIdentityMatch(
                        status="unresolved",
                        identity_revision=52,
                        query_mac=unresolved_mac,
                        unresolved_reason="exact_alias_not_found",
                    ),
                },
            )

    snapshot = resolve_vehicle_mr_identity_snapshot(QueryService(), links)  # type: ignore[arg-type]
    assert snapshot.match_for("invalid").status == "invalid"

    store = VehicleMrOnlineStore(PathResolver(tmp_path), "demo")
    store.persist_snapshot(
        "s1",
        1,
        VehicleMrMeshParseResult("2026-06-28 10:00:00", links),
        [],
        snapshot,
        20,
    )
    with sqlite3.connect(store.db_path) as conn:
        rows = conn.execute(
            "SELECT local_ap_name, identity_status, identity_reason, identity_revision FROM vehicle_mr_online_links ORDER BY id"
        ).fetchall()
        metadata = conn.execute(
            "SELECT identity_ambiguous_count, identity_unresolved_count, identity_invalid_count FROM vehicle_mr_online_snapshots"
        ).fetchone()
    assert rows == [
        ("AP-A", "ambiguous", "duplicate_exact_alias", 52),
        ("AP-B", "unresolved", "exact_alias_not_found", 52),
        ("AP-C", "invalid", "invalid_peer_mac", 52),
    ]
    assert metadata == (1, 1, 1)


def test_vehicle_mr_next_sample_observes_new_identity_revision() -> None:
    link = VehicleMrMeshLink(
        "AP-A",
        "NBL12-LC06-MR-CT",
        local_mac="020000000001",
        rssi=40,
    )

    class QueryService:
        def __init__(self) -> None:
            self.calls = 0

        def resolve_peer_macs(self, values, *, ap_role=None):
            self.calls += 1
            revision = 60 + self.calls
            key = normalize_mac_key(values[0]) or ""
            match = ApIdentityMatch(
                status="matched",
                identity_revision=revision,
                query_mac=key,
                matched_entity_id="entity-1",
                effective_ap_name="AP-A",
                effective_ap_mac="020000000000",
                station="鼓楼站",
                matched_alias_type="ac_radio_mac",
                matched_source="ac_runtime",
                match_rule="ac_radio_mac",
                match_confidence=100,
            )
            return ApIdentityBatchResult(
                revision=revision,
                index_status="ready",
                requested_count=1,
                normalized_count=1,
                distinct_count=1,
                matched_count=1,
                unresolved_count=0,
                ambiguous_count=0,
                invalid_count=0,
                matches={key: match},
            )

    query = QueryService()
    first = resolve_vehicle_mr_identity_snapshot(query, [link])  # type: ignore[arg-type]
    second = resolve_vehicle_mr_identity_snapshot(query, [link])  # type: ignore[arg-type]

    assert query.calls == 2
    assert first.revision == 61
    assert first.match_for(link.local_mac).identity_revision == 61
    assert second.revision == 62
    assert second.match_for(link.local_mac).identity_revision == 62


def test_vehicle_mr_store_upgrades_legacy_identity_columns_idempotently(tmp_path: Path) -> None:
    paths = PathResolver(tmp_path)
    db_path = paths.online_mr_root("demo") / "parsed" / "vehicle_mr_online.sqlite"
    db_path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(db_path) as conn:
        conn.executescript(
            """
            CREATE TABLE vehicle_mr_online_snapshots (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                session_id TEXT NOT NULL,
                sample_index INTEGER,
                ac_time TEXT,
                local_time TEXT,
                command_duration_ms INTEGER,
                link_count INTEGER,
                parse_status TEXT,
                error_message TEXT,
                created_at TEXT
            );
            CREATE TABLE vehicle_mr_online_links (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                session_id TEXT,
                snapshot_id INTEGER,
                ac_time TEXT,
                peer_name TEXT,
                peer_mac TEXT,
                local_ap_name TEXT,
                local_mac TEXT,
                status TEXT,
                rssi INTEGER,
                rx_packets INTEGER,
                tx_packets INTEGER,
                train_id TEXT,
                train_display_name TEXT,
                train_no TEXT,
                car_end TEXT,
                car_end_label TEXT,
                matched_station TEXT,
                matched_ap_name TEXT,
                match_method TEXT,
                match_score INTEGER,
                station_source TEXT,
                created_at TEXT
            );
            """
        )

    VehicleMrOnlineStore(paths, "demo")
    VehicleMrOnlineStore(paths, "demo")

    with sqlite3.connect(db_path) as conn:
        snapshot_columns = {
            row[1] for row in conn.execute("PRAGMA table_info(vehicle_mr_online_snapshots)")
        }
        link_columns = {
            row[1] for row in conn.execute("PRAGMA table_info(vehicle_mr_online_links)")
        }
    assert {
        "identity_revision",
        "identity_index_status",
        "identity_matched_count",
        "identity_invalid_count",
        "identity_mapped_at",
    } <= snapshot_columns
    assert {
        "identity_entity_id",
        "identity_revision",
        "identity_status",
        "identity_source",
        "identity_reason",
        "matched_alias_type",
        "matched_ap_mac",
        "matched_radio_id",
        "matched_section",
    } <= link_columns


def test_vehicle_mr_collector_reuses_one_identity_snapshot(monkeypatch) -> None:
    link = VehicleMrMeshLink(
        "AP-A",
        "NBL12-LC06-MR-CT",
        local_mac="020000000001",
        rssi=40,
    )
    parse_result = VehicleMrMeshParseResult("2026-06-28 10:00:00", [link])

    class QueryService:
        def resolve_peer_macs(self, values, *, ap_role=None):
            key = normalize_mac_key(values[0]) or ""
            return ApIdentityBatchResult(
                revision=71,
                index_status="ready",
                requested_count=1,
                normalized_count=1,
                distinct_count=1,
                matched_count=1,
                unresolved_count=0,
                ambiguous_count=0,
                invalid_count=0,
                matches={
                    key: ApIdentityMatch(
                        status="matched",
                        identity_revision=71,
                        query_mac=key,
                        matched_entity_id="entity-1",
                        effective_ap_name="AP-A",
                        effective_ap_mac="020000000000",
                        station="鼓楼站",
                        matched_alias_type="ac_radio_mac",
                        matched_source="ac_runtime",
                        match_rule="ac_radio_mac",
                        match_confidence=100,
                    )
                },
            )

    class Store:
        def __init__(self) -> None:
            self.persisted_snapshot = None

        def load_current_states(self):
            return {}

        def persist_snapshot(self, _session_id, _sample_index, _result, _trains, snapshot, _duration):
            self.persisted_snapshot = snapshot
            return 1

        def update_session(self, *_args, **_kwargs):
            return None

    built_snapshots = []
    original_build_train_states = vehicle_mr_online_service.build_train_states

    def tracked_build_train_states(*args, **kwargs):
        built_snapshots.append(args[2])
        return original_build_train_states(*args, **kwargs)

    monkeypatch.setattr(
        vehicle_mr_online_service,
        "build_train_states",
        tracked_build_train_states,
    )
    store = Store()
    collector = vehicle_mr_online_service.VehicleMrOnlineCollector(
        ac=Device(name="AC-1", device_type="AC", primary_address="192.0.2.1"),
        site_name="demo",
        interval_seconds=10,
        store=store,  # type: ignore[arg-type]
        registered_trains={"列车06": VehicleMrTrainState("列车06", "06", True)},
        identity_query_service=QueryService(),  # type: ignore[arg-type]
        mapping_lookup={},
        connection_config=OnlineMrConnectionConfig(
            site="demo",
            mr_id="ac-1",
            mr_name="AC-1",
            safe_mr_name="ac-1",
            device_id=1,
            device_name="AC-1",
            host="192.0.2.1",
        ),
        parser=type("Parser", (), {"parse": lambda _self, _raw: parse_result})(),
    )
    collector.connection = type(
        "Connection",
        (),
        {"send_command": lambda _self, _command, _timeout: "ok"},
    )()
    collector.session_id = "s1"

    collector.run_once()

    assert len(built_snapshots) == 1
    assert built_snapshots[0] is store.persisted_snapshot
    assert store.persisted_snapshot.revision == 71


def test_pass_events_are_persisted_for_online_ap_station_and_offline_changes(tmp_path: Path) -> None:
    paths = PathResolver(tmp_path)
    store = VehicleMrOnlineStore(paths, "demo")
    train_id = "列车06"
    states = [
        VehicleMrTrainState(train_id, "06", True, status="单端在线", current_station="鼓楼站", last_ac_time="00:01:00", tc1=VehicleMrEndState(True, "鼓楼站", "AP-A", 46, "00:01:00")),
        VehicleMrTrainState(train_id, "06", True, status="单端在线", current_station="鼓楼站", last_ac_time="00:02:00", tc1=VehicleMrEndState(True, "鼓楼站", "AP-B", 48, "00:02:00")),
        VehicleMrTrainState(train_id, "06", True, status="单端在线", current_station="西门口站", last_ac_time="00:03:00", tc1=VehicleMrEndState(True, "西门口站", "AP-C", 50, "00:03:00")),
        VehicleMrTrainState(train_id, "06", True, status="离线", current_station="-", last_ac_time="00:04:00", tc1=VehicleMrEndState(False, last_seen_at="00:03:00")),
    ]
    for index, state in enumerate(states, start=1):
        store.persist_snapshot("s1", index, VehicleMrMeshParseResult(state.last_ac_time, []), [state], {}, 5)

    reloaded = VehicleMrOnlineStore(paths, "demo").list_events(train_id, 200)
    event_types = [row["event_type"] for row in reversed(reloaded)]

    assert event_types == ["online", "ap_changed", "station_changed", "offline"]












def test_vehicle_mr_store_query_events_filters_by_time_end_status_station_and_ap(tmp_path: Path) -> None:
    store = VehicleMrOnlineStore(PathResolver(tmp_path), "demo")
    rows = [
        ("列车06", "06车", "06", "TC1", "2026-06-26 01:00:00", "在线", "鼓楼站", "AP-A", 46, "online"),
        ("列车06", "06车", "06", "TC2", "2026-06-26 01:20:00", "在线", "西门口站", "AP-B", 48, "station_changed"),
        ("列车06", "06车", "06", "TC1", "2026-06-26 02:00:00", "离线", "-", "AP-C", None, "offline"),
    ]
    with store.connect() as conn:
        conn.executemany(
            """
            INSERT INTO vehicle_mr_train_pass_events (
                train_id, train_display_name, train_no, car_end_label, event_time,
                status, station, ap_name, rssi, event_type, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [(*row, row[4]) for row in rows],
        )

    assert len(store.query_events("列车06", "2026-06-26 00:00:00", "2026-06-26 01:30:00")) == 2
    assert len(store.query_events("06车", "2026-06-26 00:00:00", "2026-06-26 01:30:00")) == 2
    assert len(store.query_events("NBL12-LC06", "2026-06-26 00:00:00", "2026-06-26 01:30:00")) == 2
    assert [row["car_end_label"] for row in store.query_events("06车", "2026-06-26 00:00:00", "2026-06-26 01:30:00", car_end_label="TC1")] == ["TC1"]
    assert [row["station"] for row in store.query_events("06车", "2026-06-26 00:00:00", "2026-06-26 01:30:00", station="西门")] == ["西门口站"]
    assert [row["ap_name"] for row in store.query_events("06车", "2026-06-26 00:00:00", "2026-06-26 01:30:00", ap_name="AP-B")] == ["AP-B"]
    assert [row["event_type"] for row in store.query_events("06车", "2026-06-26 00:00:00", "2026-06-26 03:00:00", status="离线")] == ["offline"]


def test_vehicle_mr_history_export_writes_rows(tmp_path: Path) -> None:
    path = tmp_path / "history.xlsx"
    paths = PathResolver(tmp_path)
    store = VehicleMrOnlineStore(paths, "demo")
    with store.connect() as conn:
        conn.execute(
            """
            INSERT INTO vehicle_mr_train_pass_events (
                train_id, train_display_name, train_no, car_end_label, event_time,
                status, station, ap_name, rssi, event_type, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            ("列车06", "06车", "06", "TC1", "2026-06-26 01:00:00", "在线", "鼓楼站", "AP-A", 46, "online", "2026-06-26 01:00:00"),
        )

    spec = vehicle_mr_history_xlsx_spec(
        path,
        app_root=paths.app_root,
        data_root=paths.data_root,
        site_name="demo",
        train_id="列车06",
        filters={"start_time": "2026-06-26 00:00:00", "end_time": "2026-06-26 02:00:00"},
        title="导出历史记录",
    )
    job = spec.to_job("vehicle-mr-history-test")
    job = ExportJob.from_dict({**job.to_dict(), "tmp_path": str(tmp_path / "history.xlsx.tmp")})
    run_generic_export_handler(job)

    from openpyxl import load_workbook

    sheet = load_workbook(path).active
    assert sheet.cell(1, 1).value == "时间"
    assert sheet.cell(2, 4).value == "鼓楼站"
