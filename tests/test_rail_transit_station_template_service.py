from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook

from netconsole.models.api.rail_transit_base_data import SectionDTO, StationDTO
from netconsole.services.rail_transit.base_data_query_service import (
    RailTransitBaseDataQueryService,
)
from netconsole.services.rail_transit.station_template_service import (
    LINE_PARAM_HEADERS,
    SECTION_HEADERS,
    STATION_HEADERS,
    StationTemplateService,
)
from tests.support.rail_transit_base_data_fixture import build_rail_transit_base_data_fixture


def _service(tmp_path: Path) -> StationTemplateService:
    paths, _database = build_rail_transit_base_data_fixture(tmp_path)
    return StationTemplateService(paths, RailTransitBaseDataQueryService(paths))


def test_new_template_has_four_sheets_and_round_trips_station_and_section_fields(
    tmp_path: Path,
) -> None:
    service = _service(tmp_path)
    stations = [
        StationDTO(
            id="station:low",
            node_uid="node-low",
            name="高桥西",
            code="11",
            sort_order=11,
            path_code="MAIN",
            structure_type="underground",
            platform_layout="island",
            center_mileage_text="K12+345.5",
            center_mileage_m=12345.5,
            is_line_terminal=True,
            is_service_terminal=True,
            turnback_capable=True,
            track_facilities=["turnback_track", "storage_track"],
            turnback_direction="both",
            terminal_extension_enabled=True,
            terminal_endpoint_label="端点",
            terminal_extension_distance_m=180,
            terminal_endpoint_mileage_text="K12+165.5",
            source_kind="manual",
        ),
        StationDTO(
            id="station:high",
            node_uid="node-high",
            name="高桥",
            code="12",
            sort_order=12,
            path_code="MAIN",
            structure_type="underground",
            platform_layout="island",
            source_kind="manual",
        ),
    ]
    sections = [
        SectionDTO(
            id="section:auto",
            name="高桥西-高桥-上行",
            section_code="AUTO-1",
            section_kind="between_stations",
            path_code="MAIN",
            direction_role="increasing",
            line_direction="上行",
            start_node_type="station",
            start_node_uid="node-low",
            start_station="高桥西",
            end_node_type="station",
            end_node_uid="node-high",
            end_station="高桥",
            line_side="上行",
            auto_generated=True,
            generation_key="MAIN|between|node-low|node-high|increasing",
            manual_override_fields=["section_mileage_start_m", "section_mileage_source"],
            section_mileage_start_m=12345.5,
            section_mileage_end_m=13000,
            section_mileage_source="manual",
            source_kind="generated",
            ap_count=4,
            mileage_min=12360,
            mileage_max=12980,
            remark="保留备注",
        )
    ]
    content = service._workbook_bytes(
        {
            "line_name": "测试线",
            "system_type": "PIS",
            "network_domain": "default",
            "main_path_code": "MAIN",
            "increasing_direction_name": "上行",
            "decreasing_direction_name": "下行",
            "increasing_direction_line_side": "左线",
            "decreasing_direction_line_side": "右线",
            "station_source_group_name": "车站",
        },
        stations,
        sections,
    )
    workbook = load_workbook(BytesIO(content))
    assert workbook.sheetnames == ["01_线路参数", "02_线路节点", "03_区间配置", "字段说明"]
    assert tuple(cell.value for cell in workbook["02_线路节点"][1]) == STATION_HEADERS
    assert tuple(cell.value for cell in workbook["03_区间配置"][1]) == SECTION_HEADERS
    assert workbook["02_线路节点"]["H2"].value == "K12+345.5"
    assert workbook["02_线路节点"]["N2"].value == "折返线、存车线"

    # AP 数量和 AP 里程统计是只读导出字段，导入修改必须被忽略。
    workbook["03_区间配置"]["S2"] = 9
    workbook["03_区间配置"]["T2"] = "0–99999 m"
    stream = BytesIO()
    workbook.save(stream)
    preview = service.preview("demo", stream.getvalue(), "基础资料.xlsx")

    station = next(row.proposed_station for row in preview.rows if row.name == "高桥西")
    section = preview.section_rows[0].proposed_section
    assert station is not None
    assert station.track_facilities == ["turnback_track", "storage_track"]
    assert station.center_mileage_text == "K12+345.5"
    assert station.center_mileage_m == 12345.5
    assert preview.line_metadata["increasing_direction_line_side"] == "左线"
    assert preview.line_metadata["decreasing_direction_line_side"] == "右线"
    assert section is not None
    assert section.line_side == "左线"
    assert section.generation_key == "MAIN|between|node-low|node-high|increasing"
    assert (section.section_mileage_start_m, section.section_mileage_end_m) == (12345.5, 13000)
    assert section.section_mileage_source == "manual"
    assert section.manual_override_fields == ["section_mileage_source", "section_mileage_start_m"]
    assert section.ap_count == 0
    assert section.mileage_min is None
    assert section.mileage_max is None


def test_old_three_sheet_template_imports_without_deleting_sections(tmp_path: Path) -> None:
    service = _service(tmp_path)
    workbook = Workbook()
    line_sheet = workbook.active
    line_sheet.title = "01_线路参数"
    line_sheet.append(list(LINE_PARAM_HEADERS))
    line_sheet.append(["测试线", "PIS", "default", "MAIN", "上行", "下行", "右线", "左线", "车站", "station", ""])
    station_sheet = workbook.create_sheet("02_线路节点")
    station_sheet.append([
        "来源站点值",
        "节点编码",
        "节点名称",
        "节点类型",
        "所属路径",
        "主线顺序",
        "参与方向判断",
        "车站结构",
        "站台形式",
        "线路端点",
        "运营终点",
        "可折返",
        "折返类型",
        "折返方向",
        "启用",
        "备注",
    ])
    station_sheet.append([
        "11-高桥西",
        "11",
        "高桥西",
        "普通车站",
        "MAIN",
        11,
        "是",
        "",
        "",
        "是",
        "是",
        "是",
        "中间折返线/存车线",
        "双向",
        "是",
        "",
    ])
    workbook.create_sheet("字段说明")
    stream = BytesIO()
    workbook.save(stream)

    preview = service.preview("demo", stream.getvalue(), "旧模板.xlsx")

    assert preview.valid is True
    assert preview.section_sheet_present is False
    assert preview.section_rows == []
    assert any(issue.code == "station_template_sections_missing" for issue in preview.issues)
    station = preview.rows[0].proposed_station
    assert station is not None
    assert station.structure_type == "underground"
    assert station.platform_layout == "island"
    assert station.track_facilities == ["turnback_track", "storage_track"]


def test_old_section_columns_preserve_existing_physical_mileage_and_overrides(
    tmp_path: Path,
) -> None:
    service = _service(tmp_path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "区间编码",
        "区间名称",
        "区间类型",
        "所属路径",
        "方向角色",
        "线路方向",
        "起始节点类型",
        "起始节点",
        "终到节点类型",
        "终到节点",
        "自动生成",
        "生成标识",
        "启用",
        "AP数量",
        "里程范围",
        "备注",
    ])
    sheet.append([
        "AUTO-OLD",
        "高桥西-高桥-上行",
        "站间区间",
        "MAIN",
        "站序递增",
        "上行",
        "车站",
        "高桥西",
        "车站",
        "高桥",
        "是",
        "MAIN|between|node-low|node-high|increasing",
        "是",
        0,
        "--",
        "旧模板",
    ])
    stations = {
        "高桥西": StationDTO(id="station:low", node_uid="node-low", name="高桥西"),
        "高桥": StationDTO(id="station:high", node_uid="node-high", name="高桥"),
    }
    existing = SectionDTO(
        id="section:existing",
        name="高桥西-高桥-上行",
        section_code="AUTO-OLD",
        section_kind="between_stations",
        path_code="MAIN",
        direction_role="increasing",
        line_direction="上行",
        start_node_type="station",
        start_node_uid="node-low",
        start_station="高桥西",
        end_node_type="station",
        end_node_uid="node-high",
        end_station="高桥",
        line_side="上行",
        auto_generated=True,
        generation_key="MAIN|between|node-low|node-high|increasing",
        manual_override_fields=["section_mileage_start_m", "section_mileage_source"],
        section_mileage_start_m=160,
        section_mileage_end_m=1801,
        section_mileage_source="manual",
        source_kind="generated",
        remark="旧模板",
    )
    issues = []

    rows = service._preview_sections(
        sheet,
        {
            "main_path_code": "MAIN",
            "increasing_direction_name": "上行",
            "decreasing_direction_name": "下行",
        },
        stations,
        [existing],
        issues,
    )

    proposed = rows[0].proposed_section
    assert proposed is not None
    assert rows[0].action == "update"
    assert proposed.line_side == "右线"
    assert (proposed.section_mileage_start_m, proposed.section_mileage_end_m) == (160, 1801)
    assert proposed.section_mileage_source == "manual"
    assert proposed.manual_override_fields == ["section_mileage_start_m", "section_mileage_source"]
