from __future__ import annotations

import hashlib
import json
from dataclasses import replace
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import Workbook, load_workbook

from tests.support.job_process_test_support import FakeExportProcessAdapter, FakeLocalProcessAdapter
from netconsole.application.rail_transit.web_application_service import RailTransitWebApplicationService, RailTransitWebError
from netconsole.application.web_artifacts import WebArtifactError
from netconsole.core.database import Database
from netconsole.core.paths import PathResolver
from netconsole.models.device import Device
from netconsole.services.job_center.job_context import JobContext
from netconsole.services.job_center.job_registry import registered_task_types
from netconsole.services.job_center.task_application_service import TaskApplicationService
from netconsole.services.export.export_handlers import run_generic_export_handler
from netconsole.services.rail_transit import trackside_ap_business_query_service, trackside_ap_update_job, trackside_optical_collection
from netconsole.services.trackside_ap_export_service import (
    TracksideApBusinessLoadResult,
    load_trackside_ap_business_snapshot,
)
from netconsole.services.trackside_ap_plan_io import (
    bind_trackside_plan_station,
    normalize_trackside_plan_row,
)
from netconsole.repositories.ac_repository import AcRepository, TRACKSIDE_AP_PLAN_MODE
from netconsole.repositories.device_fact_repository import DeviceFactRepository
from netconsole.repositories.device_group_repository import DeviceGroupRepository
from netconsole.repositories.device_repository import DeviceRepository


def _stable_station_id(node_uid: str) -> str:
    return f"station:{hashlib.sha1(node_uid.encode('utf-8')).hexdigest()[:12]}"


def _seed_base_stations(
    repository: AcRepository,
    station_names: list[str],
    *,
    site_id: str = "demo",
) -> dict[str, str]:
    result = repository.import_ap_extension_points(
        [
            {
                "site_id": site_id,
                "belong_type": "__base_station__",
                "station_id": _stable_station_id(f"station-node-{index}"),
                "station_name": station_name,
                "raw_payload_json": json.dumps(
                    {
                        "node_uid": f"station-node-{index}",
                        "canonical_station_name": station_name,
                        "sort_order": index,
                    },
                    ensure_ascii=False,
                ),
            }
            for index, station_name in enumerate(station_names, start=1)
        ],
        source_file="station-fixture.xlsx",
        template_type="station_fixture",
    )
    assert result["error_rows"] == 0
    return {
        str(row["station_name"]): str(row["station_id"])
        for row in repository.list_ap_extension_points()
        if str(row.get("belong_type") or "") == "__base_station__"
    }


def _seed_effective_trackside_scope(
    repository: AcRepository,
    samples: list[tuple[str, int, int, int, str]],
    *,
    site_id: str = "demo",
    extra_references: list[dict[str, object]] | None = None,
    extra_resources: list[dict[str, object]] | None = None,
    numbered_display: bool = True,
) -> None:
    extension_rows: list[dict[str, object]] = []
    plan_rows: list[dict[str, object]] = []
    resource_rows: list[dict[str, object]] = []
    for station_index, (
        station_name,
        planned,
        reference_count,
        online_count,
        remark,
    ) in enumerate(samples, start=1):
        display_name = (
            f"{station_index:02d}{station_name}"
            if numbered_display
            else station_name
        )
        node_uid = f"station-node-{station_index}"
        station_id = _stable_station_id(node_uid)
        extension_rows.append(
            {
                "site_id": site_id,
                "belong_type": "__base_station__",
                "station_id": station_id,
                "station_name": display_name,
                "raw_payload_json": json.dumps(
                    {
                        "node_uid": node_uid,
                        "canonical_station_name": station_name,
                        "sort_order": station_index,
                    },
                    ensure_ascii=False,
                ),
            }
        )
        plan_rows.append(
            {
                "station_id": station_id,
                "sequence_no": station_index,
                "station_name": station_name,
                "ap_count": planned,
                "management_vlan": 920 + station_index,
                "ap_management_vlans": str(920 + station_index),
                "remark": remark,
                "sort_order": station_index - 1,
            }
        )
        for ap_index in range(reference_count):
            ap_name = f"AP-{station_index:02d}-{ap_index:03d}"
            ap_uuid = f"ap-{station_index}-{ap_index}"
            ap_mac = f"{station_index:02x}{ap_index:010x}"
            extension_rows.append(
                {
                    "site_id": site_id,
                    "belong_type": "station",
                    "station_id": station_id,
                    "station_name": station_name,
                    "ap_name": ap_name,
                    "ap_mac_norm": ap_mac,
                    "raw_payload_json": json.dumps(
                        {
                            "station_node_uid": node_uid,
                            "operation_status": "in_service",
                            "project_id": site_id,
                            "ap_uuid": ap_uuid,
                        },
                        ensure_ascii=False,
                    ),
                }
            )
            resource_rows.append(
                {
                    "ap_uuid": ap_uuid,
                    "ap_name": ap_name,
                    "ap_mac": ap_mac,
                    "state": "R/M" if ap_index < online_count else "I",
                    "updated_at": "2026-07-30T11:30:25+08:00",
                }
            )

    extension_rows.extend(extra_references or [])
    resource_rows.extend(extra_resources or [])
    result = repository.import_ap_extension_points(
        extension_rows,
        source_file="scope-fixture.xlsx",
        template_type="trackside_ap_scope_fixture",
    )
    assert result["error_rows"] == 0
    repository.replace_trackside_ap_plan_rows(TRACKSIDE_AP_PLAN_MODE, plan_rows)
    repository.replace_fit_ap_resources("ac-fixture", resource_rows)
    with repository.database.connect() as connection:
        for station_index, sample in enumerate(samples, start=1):
            connection.execute(
                "UPDATE devices SET station_id = ? WHERE station = ?",
                (_stable_station_id(f"station-node-{station_index}"), sample[0]),
            )
        connection.commit()


def _snapshot() -> TracksideApBusinessLoadResult:
    return TracksideApBusinessLoadResult(
        generation=0,
        site_name="demo",
        rows=[
            {
                "site": "站点A",
                "device_name": "SW-A",
                "interface_name": "XGE1/0/1",
                "link_status": "UP",
                "switch_rx_power": -10.5,
                "switch_optical_status": "normal",
                "ap_uuid": "ap-uuid-a",
                "ap_mac": "0011-2233-4455",
                "ap_name": "AP-A",
                "ap_rx_power": -11.2,
                "ap_optical_status": "normal",
                "ap_side_has_data": True,
            },
            {
                "site": "站点B",
                "device_name": "SW-B",
                "interface_name": "XGE1/0/2",
                "switch_optical_status": "warning",
                "ap_uuid": "ap-uuid-b",
                "ap_mac": "0011-2233-4456",
                "ap_name": "AP-B",
                "ap_side_has_data": False,
            },
            {
                "site": "站点C",
                "device_name": "SW-C",
                "interface_name": "XGE1/0/3",
                "switch_optical_status": "no_module",
                "ap_side_has_data": False,
            },
        ],
        device_count=3,
        query_ms=3,
        build_ms=4,
        candidate_ap_interface_count=3,
        row_count=3,
        fit_ap_resource_count=1,
        identity_shadow={"status": "matched"},
    )


def test_trackside_terminal_targets_use_exact_switch_and_fit_ap_ids() -> None:
    switch = Device(
        device_uuid="switch-device-1",
        name="同名设备",
        device_type="SW",
        primary_address="192.0.2.10",
        ssh_enabled=True,
        ssh_username="admin",
        ssh_password="secret",
    )
    same_name = Device(
        device_uuid="switch-device-2",
        name="同名设备",
        device_type="SW",
        primary_address="192.0.2.11",
        ssh_enabled=True,
        ssh_username="admin",
        ssh_password="secret",
    )
    terminal_devices = trackside_ap_business_query_service.TracksideApBusinessQueryService._terminal_devices_by_uuid(
        [switch, same_name]
    )

    linked_row = trackside_ap_business_query_service.TracksideApBusinessQueryService._row(
        {
            "device_uuid": "switch-device-1",
            "device_name": "同名设备",
            "ac_device_uuid": "ac-device-1",
            "ap_uuid": "fit-ap-1",
            "ap_name": "AP-DISPLAY-NAME",
            "ap_mac": "0011-2233-4455",
            "ap_ip": "192.0.2.20",
            "ap_state": "R/M",
        },
        "normal",
        terminal_devices,
    )
    assert linked_row.switch_device_uuid == "switch-device-1"
    assert linked_row.switch_terminal_available is True
    assert linked_row.ap_terminal_ac_id == "ac-device-1"
    assert linked_row.ap_terminal_ap_id == "fit-ap-1"
    assert linked_row.ap_terminal_available is True

    missing_row = trackside_ap_business_query_service.TracksideApBusinessQueryService._row(
        {
            "device_name": "同名设备",
            "ap_uuid": "fit-resource-uuid",
            "ap_name": "AP-DISPLAY-NAME",
        },
        "normal",
        terminal_devices,
    )
    assert missing_row.switch_device_uuid == ""
    assert missing_row.switch_terminal_available is False
    assert missing_row.ap_terminal_ac_id == ""
    assert missing_row.ap_terminal_ap_id == ""
    assert missing_row.ap_terminal_available is False
    assert missing_row.ap_terminal_unavailable_reason == "未关联到 FIT-AP 资源"


def test_trackside_fit_ap_terminal_rejects_missing_ip_or_offline_state() -> None:
    missing_ip = trackside_ap_business_query_service.TracksideApBusinessQueryService._fit_ap_terminal_status(
        {"ac_device_uuid": "ac-1", "ap_uuid": "ap-1", "ap_state": "R/M"}
    )
    offline = trackside_ap_business_query_service.TracksideApBusinessQueryService._fit_ap_terminal_status(
        {"ac_device_uuid": "ac-1", "ap_uuid": "ap-1", "ap_ip": "192.0.2.20", "ap_state": "Idle"}
    )

    assert missing_ip == (
        "ac-1",
        "ap-1",
        False,
        "当前 AP 没有 IP，无法打开外部终端",
    )
    assert offline == (
        "ac-1",
        "ap-1",
        False,
        "当前 AP 离线或状态异常，无法打开外部终端",
    )


def test_business_snapshot_and_online_overview_share_effective_scope(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    database = Database(tmp_path / "scope.sqlite")
    database.initialize()
    repository = AcRepository(database)
    _seed_effective_trackside_scope(
        repository,
        [("双陈站", 1, 1, 1, "")],
        extra_references=[
            {
                "site_id": "demo",
                "belong_type": "station",
                "station_name": "双陈站",
                "ap_name": "暂停 AP",
                "ap_mac_norm": "aabbccdde001",
                "raw_payload_json": json.dumps(
                    {
                        "station_node_uid": "station-node-1",
                        "operation_status": "suspended",
                        "project_id": "demo",
                        "ap_uuid": "ap-suspended",
                    },
                    ensure_ascii=False,
                ),
            }
        ],
        extra_resources=[
            {
                "ap_uuid": "ap-suspended",
                "ap_name": "暂停 AP",
                "ap_mac": "aabbccdde001",
                "state": "R/M",
            }
        ],
    )
    monkeypatch.setattr(
        "netconsole.services.trackside_ap_export_service.build_trackside_ap_business_rows",
        lambda *_args, **_kwargs: [
            {
                "site": "双陈站",
                "device_uuid": "switch-1",
                "interface_name": "XGE1/0/1",
                "ap_uuid": "ap-1-0",
                "ap_name": "AP-01-000",
                "ap_mac": "010000000000",
            },
            {
                "site": "双陈站",
                "device_uuid": "switch-1",
                "interface_name": "XGE1/0/2",
                "ap_uuid": "ap-suspended",
                "ap_name": "暂停 AP",
                "ap_mac": "aabbccdde001",
            },
        ],
    )

    snapshot = load_trackside_ap_business_snapshot(
        DeviceRepository(database),
        "demo",
        generation=1,
    )

    assert len(snapshot.rows) == 1
    assert snapshot.rows[0]["site"] == "01-双陈站"
    assert snapshot.scope is not None
    overview = snapshot.scope.overview_export_rows()
    assert [row["site"] for row in overview] == ["01-双陈站", "合计"]
    assert overview[0]["online"] == 1
    assert snapshot.scope.excluded_device_count == 1


def _station_option_snapshot() -> TracksideApBusinessLoadResult:
    def row(site: str, ap_name: str, ap_mac: str, switch_status: str = "normal") -> dict[str, object | None]:
        return {
            "site": site,
            "device_name": f"SW-{ap_name}",
            "interface_name": "XGE1/0/1",
            "link_status": "UP",
            "switch_optical_status": switch_status,
            "ap_name": ap_name,
            "ap_mac": ap_mac,
            "ap_side_has_data": True,
        }

    rows = [
        row("02-云龙火车站", "AP-YL-01", "00:11:22:33:44:01", "warning"),
        row(" 02-云龙火车站 ", "AP-YL-02", "00:11:22:33:44:02"),
        row("01-小洋江站", "AP-XYJ-01", "00:11:22:33:44:03"),
        row("10-站点", "AP-10-01", "00:11:22:33:44:10"),
        row("", "AP-BLANK-01", "00:11:22:33:44:fe"),
        row("   ", "AP-BLANK-02", "00:11:22:33:44:ff"),
    ]
    return TracksideApBusinessLoadResult(
        generation=0,
        site_name="demo",
        rows=rows,
        device_count=4,
        query_ms=3,
        build_ms=4,
        candidate_ap_interface_count=6,
        row_count=len(rows),
        fit_ap_resource_count=4,
        identity_shadow={"status": "matched"},
    )


def test_trackside_query_reuses_snapshot_filter_and_optical_status(monkeypatch, tmp_path: Path) -> None:
    monkeypatch.setattr(trackside_ap_business_query_service, "Database", lambda _path: object())
    monkeypatch.setattr(trackside_ap_business_query_service, "DeviceRepository", lambda _database: SimpleNamespace(list=lambda: []))
    monkeypatch.setattr(trackside_ap_business_query_service, "load_trackside_ap_business_snapshot", lambda *_args, **_kwargs: _snapshot())
    service = trackside_ap_business_query_service.TracksideApBusinessQueryService(
        PathResolver(app_root=tmp_path, data_root=tmp_path)
    )

    page = service.list_rows("demo", optical_anomaly_only=True)

    assert page.total == 1
    assert page.items[0].device_name == "SW-B"
    assert page.items[0].ap_uuid == "ap-uuid-b"
    assert page.items[0].optical_severity == "warning"
    assert page.optical_abnormal_count == 1
    assert page.identity_shadow == {"status": "matched"}


def test_trackside_query_maps_switch_snapshot_times_and_statuses(
    monkeypatch,
    tmp_path: Path,
) -> None:
    snapshot = _snapshot()
    snapshot.rows[0].update(
        {
            "switch_interface_updated_at": "2026-08-03T18:04:29",
            "switch_optical_updated_at": "2026-08-03T18:04:30",
            "switch_interface_data_status": "current",
            "switch_optical_data_status": "stale",
        }
    )
    monkeypatch.setattr(
        trackside_ap_business_query_service,
        "Database",
        lambda _path: object(),
    )
    monkeypatch.setattr(
        trackside_ap_business_query_service,
        "DeviceRepository",
        lambda _database: SimpleNamespace(list=lambda: []),
    )
    monkeypatch.setattr(
        trackside_ap_business_query_service,
        "load_trackside_ap_business_snapshot",
        lambda *_args, **_kwargs: snapshot,
    )

    page = trackside_ap_business_query_service.TracksideApBusinessQueryService(
        PathResolver(app_root=tmp_path, data_root=tmp_path)
    ).list_rows("demo", station="站点A")

    assert page.total == 1
    assert page.items[0].switch_interface_updated_at == "2026-08-03T18:04:29"
    assert page.items[0].switch_optical_updated_at == "2026-08-03T18:04:30"
    assert page.items[0].switch_interface_data_status == "current"
    assert page.items[0].switch_optical_data_status == "stale"


def _seed_trackside_switch(
    repository: DeviceRepository,
    *,
    name: str,
    station: str,
    address: str,
) -> Device:
    group = next(
        (
            item
            for item in DeviceGroupRepository(repository.database, "demo").list()
            if item.name == "车站"
        ),
        None,
    )
    if group is None:
        group = DeviceGroupRepository(repository.database, "demo").create("车站")
    device = repository.create(
        Device(
            name=name,
            station=station,
            group_id=int(group.id or 0),
            device_type="SW",
            device_vendor="H3C",
            project_phase="phase_1",
            work_scope_status="included",
            primary_address=address,
        )
    )
    DeviceFactRepository(repository.database).replace_device_interfaces(
        str(device.device_uuid),
        [{
            "interface_name": "XGE1/0/1",
            "description": "Trackside AP",
            "link_status": "UP",
            "port_status": "access",
            "pvid": "921",
        }],
    )
    return device


def test_trackside_snapshot_keeps_switch_rows_when_fit_ap_resources_fail(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    database = Database(tmp_path / "fit-resource-failure.sqlite")
    database.initialize()
    repository = DeviceRepository(database)
    _seed_trackside_switch(
        repository,
        name="SW-A",
        station="站点A",
        address="192.0.2.11",
    )
    _seed_effective_trackside_scope(
        AcRepository(database),
        [("站点A", 1, 1, 1, "")],
        numbered_display=False,
    )
    monkeypatch.setattr(
        AcRepository,
        "list_all_fit_ap_resources_with_metadata",
        lambda _self: (_ for _ in ()).throw(RuntimeError("resource reset")),
    )

    snapshot = load_trackside_ap_business_snapshot(repository, "demo", 1)

    assert snapshot.partial_data is True
    assert snapshot.source_statuses["fit_ap_resources"] == "failed"
    assert snapshot.candidate_ap_interface_count == 1
    assert snapshot.row_count == 1
    assert snapshot.rows[0]["device_name"] == "SW-A"
    assert snapshot.unavailable_sources[0]["code"] == "FIT_AP_RESOURCES_UNAVAILABLE"


def test_trackside_snapshot_reads_history_store_after_lldp_legacy_tables_are_retired(
    tmp_path: Path,
) -> None:
    database = Database(tmp_path / "retired-lldp-history.sqlite")
    database.initialize()
    repository = DeviceRepository(database)
    _seed_trackside_switch(
        repository,
        name="SW-A",
        station="站点A",
        address="192.0.2.11",
    )
    ac_repository = AcRepository(database)
    _seed_effective_trackside_scope(
        ac_repository,
        [("站点A", 1, 1, 1, "")],
        numbered_display=False,
    )
    with database.connect() as conn:
        conn.execute("DROP TABLE ap_lldp_history")
        conn.execute("DROP TABLE ac_fit_ap_lldp_history")
        ac_repository.history_store.record_event(
            conn,
            kind="fit_ap_lldp",
            entity_key="ap-1-0",
            payload={
                "ac_device_uuid": "ac-fixture",
                "ap_uuid": "ap-1-0",
                "ap_name": "AP-01-000",
                "local_interface": "GigabitEthernet1/0/1",
                "local_interface_normalized": "ge1/0/1",
                "lldp_neighbor": "SW-A",
                "neighbor_name": "SW-A",
                "neighbor_device_name": "SW-A",
                "neighbor_interface": "XGE1/0/1",
                "neighbor_mac": "00:11:22:33:44:55",
                "neighbor_mac_normalized": "001122334455",
            },
            collected_at="2026-08-01T10:00:00",
            meaningful_fields=(
                "ap_uuid",
                "local_interface_normalized",
                "neighbor_interface",
                "neighbor_mac_normalized",
            ),
        )
        conn.commit()

    snapshot = load_trackside_ap_business_snapshot(repository, "demo", 1)

    assert snapshot.source_statuses["ap_lldp_history"] == "loaded"
    assert not any(
        item["code"] == "AP_LLDP_HISTORY_UNAVAILABLE"
        for item in snapshot.unavailable_sources
    )
    assert snapshot.partial_data is False
    assert any(row.get("ap_uuid") == "ap-1-0" for row in snapshot.rows)


def test_trackside_snapshot_uses_device_and_ac_without_base_data(
    tmp_path: Path,
) -> None:
    database = Database(tmp_path / "device-ac-primary.sqlite")
    database.initialize()
    repository = DeviceRepository(database)
    switch = _seed_trackside_switch(
        repository,
        name="SW-A",
        station="设备管理站点A",
        address="192.0.2.11",
    )
    AcRepository(database).replace_fit_ap_resources(
        "ac-fixture",
        [
            {
                "ap_uuid": "ap-runtime-1",
                "ap_name": "AP-RUNTIME-1",
                "ap_mac": "0011-2233-4455",
                "state": "R/M",
            }
        ],
    )
    DeviceFactRepository(database).replace_lldp_neighbors(
        str(switch.device_uuid),
        [
            {
                "local_interface": "XGE1/0/1",
                "neighbor_mac": "00:11:22:33:44:55",
            }
        ],
        preserve_existing=False,
    )

    snapshot = load_trackside_ap_business_snapshot(repository, "demo", 1)

    assert snapshot.scope is not None
    assert snapshot.scope.scope_ap_reference_count == 0
    assert snapshot.fit_ap_resource_count == 1
    assert snapshot.fit_ap_unmatched_online_count == 1
    assert snapshot.row_count == 1
    assert snapshot.rows[0]["site"] == "设备管理站点A"
    assert snapshot.rows[0]["station_id"] == ""
    assert snapshot.rows[0]["ap_uuid"] == "ap-runtime-1"
    assert snapshot.rows[0]["ap_mac"] == "0011-2233-4455"


def test_trackside_snapshot_projects_unique_legacy_station_and_matches_fit_by_mac(
    tmp_path: Path,
) -> None:
    database = Database(tmp_path / "legacy-station-projection.sqlite")
    database.initialize()
    repository = AcRepository(database)
    _seed_base_stations(repository, ["16-双陈站"])
    imported = repository.import_ap_extension_points(
        [
            {
                "site_id": "demo",
                "belong_type": "station",
                "station_name": "双陈站",
                "ap_name": f"AP-SC-{index}",
                "ap_mac_norm": f"0011223344{index:02d}",
                "raw_payload_json": json.dumps(
                    {
                        "operation_status": "in_service",
                        "project_id": "demo",
                    },
                    ensure_ascii=False,
                ),
            }
            for index in (1, 2)
        ],
        source_file="legacy-trackside.xlsx",
        template_type="legacy_trackside_fixture",
    )
    assert imported["error_rows"] == 0
    repository.replace_fit_ap_resources(
        "ac-fixture",
        [
            {
                "ap_uuid": f"fit-{index}",
                "ap_name": f"AP-SC-{index}",
                "ap_mac": f"0011223344{index:02d}",
                "state": "R/M",
            }
            for index in (1, 2, 3)
        ],
    )

    with database.connect() as observer:
        before_data_version = int(observer.execute("PRAGMA data_version").fetchone()[0])
        snapshot = load_trackside_ap_business_snapshot(
            DeviceRepository(database),
            "demo",
            generation=1,
        )
        after_data_version = int(observer.execute("PRAGMA data_version").fetchone()[0])

    assert snapshot.scope is not None
    assert snapshot.scope.scope_ap_reference_count == 2
    assert snapshot.fit_ap_resource_count == 3
    assert snapshot.fit_ap_matched_count == 2
    assert snapshot.fit_ap_unmatched_online_count == 1
    assert before_data_version == after_data_version


def test_trackside_snapshot_keeps_other_devices_when_one_interface_source_fails(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    database = Database(tmp_path / "device-fact-failure.sqlite")
    database.initialize()
    repository = DeviceRepository(database)
    failed_device = _seed_trackside_switch(
        repository,
        name="SW-A",
        station="站点A",
        address="192.0.2.11",
    )
    healthy_device = _seed_trackside_switch(
        repository,
        name="SW-B",
        station="站点B",
        address="192.0.2.12",
    )
    _seed_effective_trackside_scope(
        AcRepository(database),
        [
            ("站点A", 1, 1, 1, ""),
            ("站点B", 1, 1, 1, ""),
        ],
        numbered_display=False,
    )
    original = DeviceFactRepository.list_device_interfaces

    def list_interfaces(
        fact_repository: DeviceFactRepository,
        device_uuid: str,
    ) -> list[dict[str, object | None]]:
        if device_uuid == str(failed_device.device_uuid):
            raise RuntimeError("device reset")
        return original(fact_repository, device_uuid)

    def bulk_interfaces(*_args, **_kwargs):
        raise RuntimeError("bulk reader unavailable")

    monkeypatch.setattr(DeviceFactRepository, "list_device_interfaces", list_interfaces)
    monkeypatch.setattr(DeviceFactRepository, "list_device_interfaces_for_uuids", bulk_interfaces)

    snapshot = load_trackside_ap_business_snapshot(repository, "demo", 1)

    assert snapshot.partial_data is True
    assert snapshot.source_statuses["interfaces"] == "partial"
    assert snapshot.candidate_ap_interface_count == 1
    assert {row["device_name"] for row in snapshot.rows} == {healthy_device.name}
    assert any(
        item["source"] == "interfaces"
        and item["device_id"] == str(failed_device.device_uuid)
        for item in snapshot.unavailable_sources
    )


def test_trackside_query_maps_partial_source_status(monkeypatch, tmp_path: Path) -> None:
    snapshot = replace(
        _snapshot(),
        partial_data=True,
        source_statuses={"fit_ap_resources": "failed"},
        unavailable_sources=[{
            "source": "fit_ap_resources",
            "label": "FIT-AP 资源",
            "code": "FIT_AP_RESOURCES_UNAVAILABLE",
            "message": "FIT-AP 资源暂时不可用。",
            "device_id": "",
        }],
    )
    monkeypatch.setattr(trackside_ap_business_query_service, "Database", lambda _path: object())
    monkeypatch.setattr(trackside_ap_business_query_service, "DeviceRepository", lambda _database: SimpleNamespace(list=lambda: []))
    monkeypatch.setattr(
        trackside_ap_business_query_service,
        "load_trackside_ap_business_snapshot",
        lambda *_args, **_kwargs: snapshot,
    )

    page = trackside_ap_business_query_service.TracksideApBusinessQueryService(
        PathResolver(app_root=tmp_path, data_root=tmp_path)
    ).list_rows("demo")

    assert page.partial_data is True
    assert page.source_statuses == {"fit_ap_resources": "failed"}
    assert page.unavailable_sources[0].code == "FIT_AP_RESOURCES_UNAVAILABLE"


def test_trackside_query_applies_business_threshold_to_both_rx_sides(
    monkeypatch,
    tmp_path: Path,
) -> None:
    original = _snapshot()
    snapshot = replace(
        original,
        rows=[
            {
                **original.rows[0],
                "pvid": 71,
                "vlan": "Native/PVID 71; Tagged 201",
                "switch_rx_power": -19.10,
                "switch_optical_status": "normal",
                "ap_rx_power": -7.72,
                "ap_optical_status": "normal",
            }
        ],
    )
    monkeypatch.setattr(trackside_ap_business_query_service, "Database", lambda _path: object())
    monkeypatch.setattr(trackside_ap_business_query_service, "DeviceRepository", lambda _database: SimpleNamespace(list=lambda: []))
    monkeypatch.setattr(
        trackside_ap_business_query_service,
        "load_trackside_ap_business_snapshot",
        lambda *_args, **_kwargs: snapshot,
    )

    page = trackside_ap_business_query_service.TracksideApBusinessQueryService(
        PathResolver(app_root=tmp_path, data_root=tmp_path)
    ).list_rows("demo", optical_anomaly_only=True)

    assert page.total == 1
    assert page.optical_abnormal_count == 1
    assert page.items[0].pvid == 71
    assert page.items[0].vlan == "Tagged 201"
    assert page.items[0].switch_device_optical_status == "normal"
    assert page.items[0].switch_optical_status == "abnormal"
    assert page.items[0].ap_device_optical_status == "normal"
    assert page.items[0].ap_business_optical_status == "abnormal"
    assert page.items[0].ap_optical_status == "normal"
    assert page.items[0].ap_business_threshold_dbm == -13.90
    assert "交换机侧收光 -19.10 dBm 低于业务门限 -13.90 dBm" in page.items[0].ap_business_reason
    assert page.items[0].optical_severity == "abnormal"


def test_trackside_query_counts_multiple_abnormal_interfaces_once_per_ap(monkeypatch, tmp_path: Path) -> None:
    snapshot = _snapshot()
    snapshot.rows.append(
        {
            "site": "站点B",
            "device_name": "SW-B",
            "interface_name": "XGE1/0/3",
            "switch_optical_status": "critical",
            "ap_uuid": "ap-uuid-b",
            "ap_mac": "0011-2233-4456",
            "ap_name": "AP-B",
            "ap_side_has_data": False,
        }
    )
    monkeypatch.setattr(trackside_ap_business_query_service, "Database", lambda _path: object())
    monkeypatch.setattr(trackside_ap_business_query_service, "DeviceRepository", lambda _database: SimpleNamespace(list=lambda: []))
    monkeypatch.setattr(trackside_ap_business_query_service, "load_trackside_ap_business_snapshot", lambda *_args, **_kwargs: snapshot)

    page = trackside_ap_business_query_service.TracksideApBusinessQueryService(
        PathResolver(app_root=tmp_path, data_root=tmp_path)
    ).list_rows("demo", optical_anomaly_only=True)

    assert page.total == 2
    assert page.optical_abnormal_count == 1


def test_trackside_query_station_options_use_full_snapshot_before_filters(monkeypatch, tmp_path: Path) -> None:
    snapshot = _station_option_snapshot()
    monkeypatch.setattr(trackside_ap_business_query_service, "Database", lambda _path: object())
    monkeypatch.setattr(trackside_ap_business_query_service, "DeviceRepository", lambda _database: SimpleNamespace(list=lambda: []))
    monkeypatch.setattr(trackside_ap_business_query_service, "load_trackside_ap_business_snapshot", lambda *_args, **_kwargs: snapshot)
    service = trackside_ap_business_query_service.TracksideApBusinessQueryService(
        PathResolver(app_root=tmp_path, data_root=tmp_path)
    )
    expected = ["01-小洋江站", "02-云龙火车站", "10-站点"]

    paged = service.list_rows("demo", page_size=1)
    assert paged.station_options == expected
    assert len(paged.items) == 1

    station_filtered = service.list_rows("demo", station="02-云龙火车站")
    assert station_filtered.station_options == expected
    assert station_filtered.total == 2
    assert {item.site.strip() for item in station_filtered.items} == {"02-云龙火车站"}

    anomaly_filtered = service.list_rows("demo", optical_anomaly_only=True)
    assert anomaly_filtered.station_options == expected

    query_filtered = service.list_rows("demo", query="AP-10")
    assert query_filtered.station_options == expected

    snapshot.identity_query_entities["001122334a01"] = "entity-yl-01"
    next(
        row for row in snapshot.rows if row.get("ap_name") == "AP-YL-01"
    )["ap_identity_entity_id"] = "entity-yl-01"
    radio_filtered = service.list_rows(
        "demo",
        query="0011-2233-4a01",
    )
    assert radio_filtered.total == 1
    assert radio_filtered.items[0].ap_name == "AP-YL-01"


def test_trackside_update_job_calls_existing_collection_service(monkeypatch, tmp_path: Path) -> None:
    captured: dict[str, object] = {}
    progress: list[tuple[str, int, int, str]] = []
    identity_rebuilds: list[str] = []
    monkeypatch.setattr(trackside_ap_update_job, "Database", lambda _path: object())
    monkeypatch.setattr(trackside_ap_update_job, "DeviceRepository", lambda _database: object())
    monkeypatch.setattr(trackside_ap_update_job, "load_trackside_ap_business_snapshot", lambda *_args, **_kwargs: _snapshot())
    monkeypatch.setattr(
        trackside_ap_update_job,
        "ApIdentityQueryService",
        lambda _database: SimpleNamespace(
            rebuild_index=lambda reason: identity_rebuilds.append(reason)
        ),
    )

    class Result:
        session_id = "session-1"
        status = "PARTIAL_SUCCESS"
        scope = "station"
        target_label = "站点A"
        success_count = 746
        failed_count = 0
        skipped_count = 1
        actionable_skipped_count = 0
        ignored_skipped_count = 1
        skipped_reason_counts = {"no_station_switches": 1}
        skipped = [trackside_optical_collection.TracksideSkippedTarget("车站", "SWITCH", "no_station_switches")]
        target_count = 746
        concurrency = 64
        requested_concurrency = 1000
        effective_concurrency = 2
        platform_concurrency_limit = 64
        fit_ap_effective_concurrency = 2
        fit_ap_round_summaries = [{"ac_device_uuid": "ac-1", "rounds": []}]
        fit_ap_resource_count = 746
        fit_ap_optical_success_count = 746
        fit_ap_optical_failed_count = 0
        candidate_ap_interface_count = 2
        current_lldp_port_count = 2
        preserved_lldp_port_count = 0
        warning_count = 1
        warning_reason_counts = {"switch_interface_snapshot_invalid": 1}
        warnings = ["接口摘要无效"]
        port_errors = []

    def collect(repository, site_id, paths, rows, **kwargs):
        captured.update(repository=repository, site_id=site_id, paths=paths, rows=rows, **kwargs)
        kwargs["progress_callback"](1, 2)
        return Result()

    monkeypatch.setattr(trackside_ap_update_job, "collect_trackside_optical", collect)
    context = JobContext(
        "task-1",
        "trackside_ap_optical_update",
        {"site_name": "demo", "db_path": str(tmp_path / "site.sqlite"), "station": "站点A"},
        lambda stage, current, total, message: progress.append((stage, current, total, message)),
        lambda: False,
        PathResolver(app_root=tmp_path, data_root=tmp_path),
    )

    result = trackside_ap_update_job.run_trackside_ap_optical_update(context)

    assert captured["site_id"] == "demo"
    assert captured["target_station"] == "站点A"
    assert captured["rows"] == _snapshot().rows
    assert result["session_id"] == "session-1"
    assert result["status"] == "PARTIAL_SUCCESS"
    assert result["terminal_state"] == "COMPLETED"
    assert result["success_count"] == 746
    assert result["failed_count"] == 0
    assert result["skipped_count"] == 1
    assert result["actionable_skipped_count"] == 0
    assert result["ignored_skipped_count"] == 1
    assert result["skipped_reason_counts"] == {"no_station_switches": 1}
    assert result["skipped"][0]["reason"] == "no_station_switches"
    assert result["requested_concurrency"] == 1000
    assert result["effective_concurrency"] == 2
    assert result["warning_count"] == 1
    assert result["warning_reason_counts"] == {
        "switch_interface_snapshot_invalid": 1
    }
    assert result["has_warning"] is True
    assert progress[-1] == ("trackside_ap_optical_update", 1, 2, "正在更新轨旁 AP 光衰")
    assert identity_rebuilds == ["trackside_ap_optical_refresh_succeeded"]


def test_trackside_update_job_rejects_partial_source_snapshot(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    partial_snapshot = replace(
        _snapshot(),
        partial_data=True,
        source_statuses={"interfaces": "partial"},
        unavailable_sources=[{
            "source": "interfaces",
            "label": "交换机接口事实",
            "code": "SWITCH_INTERFACES_UNAVAILABLE",
            "message": "交换机接口事实暂时不可用。",
            "device_id": "switch-1",
        }],
    )
    collection_called = False

    def unexpected_collection(*_args, **_kwargs):
        nonlocal collection_called
        collection_called = True
        pytest.fail("部分数据快照不得进入采集")

    monkeypatch.setattr(
        trackside_ap_update_job,
        "collect_trackside_optical",
        unexpected_collection,
    )
    monkeypatch.setattr(trackside_ap_update_job, "Database", lambda _path: object())
    monkeypatch.setattr(trackside_ap_update_job, "DeviceRepository", lambda _database: object())
    monkeypatch.setattr(
        trackside_ap_update_job,
        "load_trackside_ap_business_snapshot",
        lambda *_args, **_kwargs: partial_snapshot,
    )
    context = JobContext(
        "task-partial",
        "trackside_ap_optical_update",
        {"site_name": "demo", "db_path": str(tmp_path / "site.sqlite")},
        lambda *_args: None,
        None,
        PathResolver(app_root=tmp_path, data_root=tmp_path),
    )

    with pytest.raises(
        RuntimeError,
        match="轨旁 AP 光衰更新所需数据不完整：交换机接口事实暂时不可用",
    ):
        trackside_ap_update_job.run_trackside_ap_optical_update(context)

    assert collection_called is False


@pytest.mark.parametrize(
    ("status", "expected"),
    [
        ("SUCCESS", "COMPLETED"),
        ("PARTIAL_SUCCESS", "COMPLETED"),
        ("NO_TARGET", "COMPLETED"),
        ("FAILED", "FAILED"),
        ("CANCELLED", "CANCELLED"),
    ],
)
def test_trackside_update_job_maps_result_status_to_task_terminal_state(
    monkeypatch,
    tmp_path: Path,
    status: str,
    expected: str,
) -> None:
    monkeypatch.setattr(trackside_ap_update_job, "Database", lambda _path: object())
    monkeypatch.setattr(trackside_ap_update_job, "DeviceRepository", lambda _database: object())
    monkeypatch.setattr(trackside_ap_update_job, "load_trackside_ap_business_snapshot", lambda *_args, **_kwargs: _snapshot())
    monkeypatch.setattr(
        trackside_ap_update_job,
        "collect_trackside_optical",
        lambda *_args, **_kwargs: SimpleNamespace(
            session_id="session-1",
            status=status,
            scope="all",
            target_label="",
            success_count=0 if status == "FAILED" else 1,
            failed_count=2 if status == "FAILED" else 0,
            skipped_count=0,
            target_count=2,
            concurrency=64,
            requested_concurrency=64,
            effective_concurrency=2,
            platform_concurrency_limit=64,
            fit_ap_effective_concurrency=2,
            fit_ap_round_summaries=[],
            fit_ap_resource_count=1,
            fit_ap_optical_success_count=0,
            fit_ap_optical_failed_count=0,
            candidate_ap_interface_count=2,
            current_lldp_port_count=2,
            preserved_lldp_port_count=0,
        ),
    )
    context = JobContext(
        "task-1",
        "trackside_ap_optical_update",
        {"site_name": "demo", "db_path": str(tmp_path / "site.sqlite")},
        lambda *_args: None,
        lambda: False,
        PathResolver(app_root=tmp_path, data_root=tmp_path),
    )

    result = trackside_ap_update_job.run_trackside_ap_optical_update(context)

    assert result["status"] == status
    assert result["terminal_state"] == expected


def test_trackside_application_starts_scoped_update(tmp_path: Path) -> None:
    paths = PathResolver(app_root=tmp_path, data_root=tmp_path)
    paths.ensure_site_dirs("demo")
    tasks = TaskApplicationService(paths=paths, site_name="demo")
    process = FakeLocalProcessAdapter(tasks)
    export = FakeExportProcessAdapter(tasks)
    service = RailTransitWebApplicationService(
        paths,
        tasks,
        process_adapter=process,  # type: ignore[arg-type]
        export_adapter=export,  # type: ignore[arg-type]
    )

    update = service.start_trackside_ap_update("demo", station="站点A")
    duplicate = service.start_trackside_ap_update("demo", station="站点A")
    update_job = process.jobs[update.task_id]
    assert duplicate.task_id == update.task_id
    assert len(process.jobs) == 1
    assert update_job.task_type == "trackside_ap_optical_update"
    assert update_job.params["station"] == "站点A"


def test_trackside_application_lists_zte_adapter_and_finalizes_sample_artifact(
    tmp_path: Path,
) -> None:
    paths = PathResolver(app_root=tmp_path, data_root=tmp_path)
    paths.ensure_site_dirs("demo")
    database = Database(paths.site_db_path("demo"))
    database.initialize()
    device = DeviceRepository(database).create(
        Device(
            device_uuid="11111111-1111-4111-8111-111111111111",
            name="ZTE-SW-01",
            station="站点A",
            device_vendor="ZTE",
            device_type="SW",
            primary_address="192.0.2.10",
        )
    )
    catalog = (
        trackside_ap_business_query_service.TracksideApBusinessQueryService(
            paths
        ).list_switch_adapters("demo")
    )
    assert catalog.total == 1
    assert catalog.items[0].device_uuid == device.device_uuid
    assert (
        catalog.items[0].adapter.adaptation_status
        == "C89E-4 Release 已验证；其他 ZXR10/5960X 型号需逐型号复核"
    )
    assert catalog.items[0].adapter.profile.profile_id == (
        "zte_zxr10_5960x_es_v2"
    )
    assert {
        item.key: item.status for item in catalog.items[0].adapter.capabilities
    }["lldp"] == "VERIFIED"

    tasks = TaskApplicationService(paths=paths, site_name="demo")
    process = FakeLocalProcessAdapter(tasks)
    service = RailTransitWebApplicationService(
        paths,
        tasks,
        process_adapter=process,  # type: ignore[arg-type]
        export_adapter=FakeExportProcessAdapter(tasks),  # type: ignore[arg-type]
    )
    started = service.start_switch_vendor_sample(
        "demo",
        device_uuid=str(device.device_uuid),
        vendor="ZTE",
        command_profile="zte_zxr10_5960x_es_v2",
        selected_interface="xgei-0/1/1/2",
        requested_commands=["device_version", "lldp_global"],
    )
    job = process.jobs[started.task_id]
    assert job.task_type == "switch_vendor_sample_collect"
    assert job.params["requested_commands"] == [
        "device_version",
        "lldp_global",
    ]
    assert job.params["selected_interface"] == "xgei-0/1/1/2"
    output_path = Path(str(job.params["artifact_output_path"]))
    assert output_path.name.startswith("zte-adapter-sample-ZTE-SW-01-")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_bytes(b"fixture-zip")

    process.complete(started.task_id, {"status": "PARTIAL_SUCCESS"})

    completed = service.get_task("demo", started.task_id)
    assert completed.action == "switch_vendor_sample_collect"
    assert completed.available is True
    assert completed.artifact_name == output_path.name
    opened_path, opened_name = service.open_switch_vendor_sample(
        "demo",
        completed.artifact_id,
    )
    assert opened_path == output_path
    assert opened_name == output_path.name
    assert "switch_vendor_sample_collect" in registered_task_types()


def test_trackside_application_rejects_mismatched_sample_profile(
    tmp_path: Path,
) -> None:
    paths = PathResolver(app_root=tmp_path, data_root=tmp_path)
    paths.ensure_site_dirs("demo")
    database = Database(paths.site_db_path("demo"))
    database.initialize()
    device = DeviceRepository(database).create(
        Device(
            device_uuid="22222222-2222-4222-8222-222222222222",
            name="ZTE-SW-02",
            device_vendor="ZTE",
            device_type="SW",
        )
    )
    tasks = TaskApplicationService(paths=paths, site_name="demo")
    process = FakeLocalProcessAdapter(tasks)
    service = RailTransitWebApplicationService(
        paths,
        tasks,
        process_adapter=process,  # type: ignore[arg-type]
        export_adapter=FakeExportProcessAdapter(tasks),  # type: ignore[arg-type]
    )

    with pytest.raises(RailTransitWebError) as exc_info:
        service.start_switch_vendor_sample(
            "demo",
            device_uuid=str(device.device_uuid),
            vendor="ZTE",
            command_profile="h3c_comware_trackside_v1",
        )

    assert exc_info.value.code == "SWITCH_PROFILE_MISMATCH"
    assert process.jobs == {}


def test_trackside_application_rejects_h3c_vendor_sample(
    tmp_path: Path,
) -> None:
    paths = PathResolver(app_root=tmp_path, data_root=tmp_path)
    paths.ensure_site_dirs("demo")
    database = Database(paths.site_db_path("demo"))
    database.initialize()
    device = DeviceRepository(database).create(
        Device(
            device_uuid="33333333-3333-4333-8333-333333333333",
            name="H3C-SW-01",
            device_vendor="H3C",
            device_type="SW",
        )
    )
    tasks = TaskApplicationService(paths=paths, site_name="demo")
    process = FakeLocalProcessAdapter(tasks)
    service = RailTransitWebApplicationService(
        paths,
        tasks,
        process_adapter=process,  # type: ignore[arg-type]
        export_adapter=FakeExportProcessAdapter(tasks),  # type: ignore[arg-type]
    )

    with pytest.raises(RailTransitWebError) as exc_info:
        service.start_switch_vendor_sample(
            "demo",
            device_uuid=str(device.device_uuid),
            vendor="H3C",
            command_profile="h3c_comware_trackside_v1",
        )

    assert exc_info.value.code == "SWITCH_SAMPLE_VENDOR_UNSUPPORTED"
    assert process.jobs == {}


def test_trackside_application_validates_update_scope_and_ap_identity(tmp_path: Path) -> None:
    paths = PathResolver(app_root=tmp_path, data_root=tmp_path)
    paths.ensure_site_dirs("demo")
    database = Database(paths.site_db_path("demo"))
    database.initialize()
    ac_uuid_1 = "00000000-0000-4000-8000-000000000001"
    ac_uuid_2 = "00000000-0000-4000-8000-000000000002"
    device_repository = DeviceRepository(database)
    device_repository.create(
        Device(
            device_uuid=ac_uuid_1,
            name="AC-1",
            device_vendor="H3C",
            device_type="AC",
            primary_address="10.0.0.1",
        )
    )
    device_repository.create(
        Device(
            device_uuid=ac_uuid_2,
            name="AC-2",
            device_vendor="H3C",
            device_type="AC",
            primary_address="10.0.0.2",
        )
    )
    repository = AcRepository(database)
    repository.replace_fit_ap_resources(
        ac_uuid_1,
        [
            {
                "ac_device_uuid": ac_uuid_1,
                "ap_uuid": "ap-1",
                "ap_name": "AP-A",
                "ap_mac": "bc5a-3457-8cc0",
                "ap_ip": "10.0.0.1",
                "site": "站点A",
            }
        ],
    )
    repository.replace_fit_ap_resources(
        ac_uuid_2,
        [
            {
                "ac_device_uuid": ac_uuid_2,
                "ap_uuid": "ap-2",
                "ap_name": "AP-B",
                "ap_mac": "305f-277a-1880",
                "ap_ip": "10.0.0.2",
                "site": "站点B",
            }
        ],
    )
    tasks = TaskApplicationService(paths=paths, site_name="demo")
    process = FakeLocalProcessAdapter(tasks)
    service = RailTransitWebApplicationService(
        paths,
        tasks,
        process_adapter=process,  # type: ignore[arg-type]
        export_adapter=FakeExportProcessAdapter(tasks),  # type: ignore[arg-type]
    )

    assert "trackside_ap_optical_update" in registered_task_types()

    all_update = service.start_trackside_ap_update("demo")
    all_job = process.jobs[all_update.task_id]
    process.complete(all_update.task_id, {"success_count": 1})
    station_update = service.start_trackside_ap_update("demo", station="站点A")
    station_job = process.jobs[station_update.task_id]
    process.complete(station_update.task_id, {"success_count": 1})
    uuid_update = service.start_trackside_ap_update("demo", ap_uuid="ap-1")
    uuid_job = process.jobs[uuid_update.task_id]
    process.complete(uuid_update.task_id, {"success_count": 1})
    mac_update = service.start_trackside_ap_update("demo", ap_mac="BC5A-3457-8CC0")
    mac_job = process.jobs[mac_update.task_id]
    process.complete(mac_update.task_id, {"success_count": 1})
    ap_update = service.start_trackside_ap_update("demo", ap_uuid="ap-1", ap_mac="bc5a-3457-8cc0", ap_name="AP-A")
    ap_job = process.jobs[ap_update.task_id]
    process.complete(ap_update.task_id, {"success_count": 1})

    assert all_job.params["station"] == ""
    assert all_job.params["ap_uuid"] == ""
    assert all_job.params["resource_keys"] == [
        "site:demo|trackside_ap_business_data",
        f"site:demo|ac:{ac_uuid_1}|fit_ap_optical",
        f"site:demo|ac:{ac_uuid_2}|fit_ap_optical",
        "site:demo|trackside_ap_optical|scope:all",
    ]
    assert station_job.params["station"] == "站点A"
    assert station_job.params["ap_uuid"] == ""
    assert station_job.params["resource_keys"] == [
        "site:demo|trackside_ap_business_data",
        f"site:demo|ac:{ac_uuid_1}|fit_ap_optical",
        "site:demo|trackside_ap_optical|scope:station:站点a",
    ]
    for job in (uuid_job, mac_job, ap_job):
        assert job.params["station"] == ""
        assert job.params["ap_uuid"] == "ap-1"
        assert job.params["ap_mac"] == "bc:5a:34:57:8c:c0"
        assert job.params["ap_name"] == "AP-A"
        assert job.params["ac_uuid"] == ac_uuid_1
        assert job.params["device_uuid"] == ac_uuid_1
        assert job.params["resource_keys"] == [
            "site:demo|trackside_ap_business_data",
            f"site:demo|ac:{ac_uuid_1}|fit_ap_optical",
            (
                "site:demo|trackside_ap_optical|scope:"
                "ap_uuid:ap-1|ap_mac:bc:5a:34:57:8c:c0|ap_name:ap-a"
            ),
        ]

    with pytest.raises(RailTransitWebError) as name_deprecated:
        service.start_trackside_ap_update("demo", ap_name="AP-A")
    assert name_deprecated.value.code == "TRACKSIDE_UPDATE_AP_NAME_DEPRECATED"

    with pytest.raises(RailTransitWebError) as scope_conflict:
        service.start_trackside_ap_update("demo", station="站点A", ap_uuid="ap-1")
    assert scope_conflict.value.code == "TRACKSIDE_UPDATE_SCOPE_CONFLICT"

    with pytest.raises(RailTransitWebError) as ap_conflict:
        service.start_trackside_ap_update("demo", ap_uuid="ap-1", ap_mac="305f-277a-1880", ap_name="AP-A")
    assert ap_conflict.value.code == "TRACKSIDE_UPDATE_AP_CONFLICT"

    with pytest.raises(RailTransitWebError) as ap_not_found:
        service.start_trackside_ap_update("demo", ap_mac="ffff-ffff-ffff")
    assert ap_not_found.value.code == "TRACKSIDE_UPDATE_AP_NOT_FOUND"
    assert process.jobs == {}

    with pytest.raises(RailTransitWebError) as invalid_mac:
        service.start_trackside_ap_update("demo", ap_mac="0011-2233-G455")
    assert invalid_mac.value.code == "AP_MAC_INVALID"


def test_trackside_application_rejects_unbound_ap_target(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    paths = PathResolver(app_root=tmp_path, data_root=tmp_path)
    paths.ensure_site_dirs("demo")
    database = Database(paths.site_db_path("demo"))
    database.initialize()
    tasks = TaskApplicationService(paths=paths, site_name="demo")
    process = FakeLocalProcessAdapter(tasks)
    service = RailTransitWebApplicationService(
        paths,
        tasks,
        process_adapter=process,  # type: ignore[arg-type]
        export_adapter=FakeExportProcessAdapter(tasks),  # type: ignore[arg-type]
    )

    monkeypatch.setattr(
        service,
        "_resolve_trackside_ap_update_target",
        lambda *_args, **_kwargs: {
            "ap_uuid": "ap-1",
            "ap_mac": "bc5a-3457-8cc0",
            "ap_name": "AP-A",
            "ac_device_uuid": "",
        },
    )

    with pytest.raises(RailTransitWebError) as exc_info:
        service.start_trackside_ap_update("demo", ap_uuid="ap-1")

    assert exc_info.value.code == "TRACKSIDE_UPDATE_AP_AC_MISSING"
    assert process.jobs == {}


def test_trackside_collection_no_target_does_not_fake_success(tmp_path: Path) -> None:
    paths = PathResolver(app_root=tmp_path, data_root=tmp_path)
    paths.ensure_site_dirs("demo")
    database = Database(paths.site_db_path("demo"))
    database.initialize()
    repository = DeviceRepository(database)

    result = trackside_ap_update_job.collect_trackside_optical(
        repository,
        "demo",
        paths,
        [],
        target_ap_uuid="ap-1",
        target_ap_mac="00:11:22:33:44:55",
        target_ap_name="AP-A",
    )

    assert result.status == "NO_TARGET"
    assert result.target_count == 0
    assert result.success_count == 0
    assert result.failed_count == 0


@pytest.mark.parametrize(
    ("success_count", "failed_count", "actionable_skipped_count", "cancelled", "expected"),
    [
        (746, 0, 0, False, "SUCCESS"),
        (745, 1, 0, False, "PARTIAL_SUCCESS"),
        (745, 0, 1, False, "PARTIAL_SUCCESS"),
        (0, 38, 0, False, "FAILED"),
        (0, 0, 3, False, "FAILED"),
        (0, 0, 0, False, "NO_TARGET"),
        (1, 1, 0, True, "CANCELLED"),
    ],
)
def test_trackside_collection_status_classification(
    success_count: int,
    failed_count: int,
    actionable_skipped_count: int,
    cancelled: bool,
    expected: str,
) -> None:
    assert (
        trackside_optical_collection._trackside_update_status(
            success_count=success_count,
            failed_count=failed_count,
            actionable_skipped_count=actionable_skipped_count,
            cancelled=cancelled,
        )
        == expected
    )


def test_trackside_skipped_classification_ignores_optional_station_switch_branch() -> None:
    skipped = [
        trackside_optical_collection.TracksideSkippedTarget("车站", "SWITCH", "no_station_switches"),
        trackside_optical_collection.TracksideSkippedTarget("AP-A", "FIT_AP", "connection_incomplete"),
    ]

    actionable, ignored, reason_counts = trackside_optical_collection.classify_trackside_skipped(skipped)

    assert actionable == 1
    assert ignored == 1
    assert reason_counts == {"no_station_switches": 1, "connection_incomplete": 1}


def test_trackside_plan_preview_save_export_and_artifact_download(tmp_path: Path) -> None:
    paths = PathResolver(app_root=tmp_path, data_root=tmp_path)
    paths.ensure_site_dirs("demo")
    database = Database(paths.site_db_path("demo"))
    database.initialize()
    repository = AcRepository(database)
    station_ids = _seed_base_stations(
        repository,
        ["站点A", "站点B", "小洋江站"],
    )
    repository.replace_trackside_ap_plan_rows(
        TRACKSIDE_AP_PLAN_MODE,
        [{
            "station_id": station_ids["站点A"],
            "sequence_no": 1,
            "station_name": "站点A",
            "ap_count": 20,
            "management_vlan": 921,
            "remark": "原值",
        }],
    )
    tasks = TaskApplicationService(paths=paths, site_name="demo")
    process = FakeLocalProcessAdapter(tasks)
    export = FakeExportProcessAdapter(tasks)
    service = RailTransitWebApplicationService(
        paths,
        tasks,
        process_adapter=process,  # type: ignore[arg-type]
        export_adapter=export,  # type: ignore[arg-type]
    )
    content = (
        "序号,车站ID,车站名称,AP数量,AP管理VLAN,备注\r\n"
        f"1,{station_ids['站点A']},站点A,30,921,覆盖值\r\n"
        f"2,{station_ids['站点B']},站点B,10,922,新增值\r\n"
    ).encode("utf-8-sig")

    preview = service.preview_trackside_ap_plan(
        "demo", file_name="trackside.csv", content=content, duplicate_strategy="replace"
    )
    assert preview.can_apply is True
    assert preview.duplicate_count == 1
    assert preview.valid_count == 1
    assert {row.station_name for row in preview.result_rows} == {"站点A", "站点B"}
    assert (
        next(
            row for row in preview.result_rows if row.station_name == "站点A"
        ).planned_ap_count
        == 30
    )

    started = service.start_trackside_ap_plan_save(
        "demo",
        rows=[row.model_dump() for row in preview.result_rows],
        explicit_confirmation=True,
        audit={"source": "test"},
    )
    assert process.jobs[started.task_id].task_type == "trackside_ap_plan_save"
    assert process.jobs[started.task_id].params["audit"] == {"source": "test"}
    current = service.start_trackside_ap_plan_export("demo", template=False)
    current_job = export.jobs[current.task_id]
    run_generic_export_handler(current_job)
    current_content = Path(current_job.output_path).read_bytes()
    export.complete(current.task_id, current_content)
    completed = service.get_task("demo", current.task_id)
    assert completed.available is True
    current_path, current_name = service.open_trackside_ap_plan_export(
        "demo",
        completed.artifact_id,
    )
    assert current_name.startswith("demo_轨旁AP规划及上线概览_")
    current_workbook = load_workbook(current_path)
    assert current_workbook.sheetnames == [
        "AP规划",
        "AP上线情况概览",
        "_netconsole_meta",
    ]
    assert [
        current_workbook["AP规划"].cell(1, column).value
        for column in range(1, 7)
    ] == [
        "序号",
        "车站ID",
        "车站名称",
        "AP数量",
        "AP管理VLAN",
        "备注",
    ]
    assert current_workbook["AP规划"].freeze_panes == "A2"
    assert current_workbook["AP规划"].auto_filter.ref == "A1:F2"
    assert current_workbook["AP规划"]["A2"].value == 1
    assert current_workbook["AP规划"]["B2"].value == station_ids["站点A"]
    assert current_workbook["AP规划"]["C2"].value == "站点A"
    assert current_workbook["AP规划"]["D2"].value == 20
    assert current_workbook["AP规划"]["D2"].data_type == "n"
    assert current_workbook["AP规划"]["D2"].number_format == "0"
    assert current_workbook["AP规划"]["E2"].value == 921
    assert current_workbook["AP规划"]["E2"].data_type == "n"
    assert current_workbook["AP规划"]["E2"].number_format == "0"
    assert current_workbook["AP规划"]["F2"].alignment.wrap_text is True
    overview = current_workbook["AP上线情况概览"]
    assert [overview.cell(1, column).value for column in range(1, 7)] == [
        "归属站点",
        "规划AP总数量",
        "实际上线",
        "未上线",
        "上线率",
        "备注",
    ]
    assert [
        [overview.cell(row, column).value for column in range(1, 6)]
            for row in range(2, 4)
        ] == [
            ["站点A", 20, 0, 20, 0],
            ["合计", 20, 0, 20, 0.0],
        ]
    assert overview.max_row == 3
    assert overview["E3"].number_format == "0.0%"
    assert overview["E3"].font.bold is True
    metadata = json.loads(current_workbook["_netconsole_meta"]["B1"].value)
    assert current_workbook["_netconsole_meta"].sheet_state == "hidden"
    assert metadata["template_type"] == "trackside_ap_station_plan"
    assert metadata["schema_version"] == 4
    assert metadata["project_id"] == "demo"
    assert metadata["line_id"] == "current"
    current_workbook.close()
    current_roundtrip = service.preview_trackside_ap_plan(
        "demo",
        file_name=current_name,
        content=current_path.read_bytes(),
        duplicate_strategy="replace",
    )
    assert current_roundtrip.can_apply is True
    assert [
        row.station_name
        for row in current_roundtrip.result_rows
    ] == ["站点A"]
    assert current_roundtrip.result_rows[0].planned_ap_count == 20
    snapshot = tasks.repository("demo").get(current.task_id)
    assert snapshot is not None
    assert snapshot.result["sha256"] == hashlib.sha256(current_content).hexdigest()
    assert snapshot.result["size_bytes"] == len(current_content)

    template = service.start_trackside_ap_plan_export("demo", template=True)
    template_job = export.jobs[template.task_id]
    run_generic_export_handler(template_job)
    template_content = Path(template_job.output_path).read_bytes()
    export.complete(template.task_id, template_content)
    template_task = service.get_task("demo", template.task_id)
    template_path, template_name = service.open_trackside_ap_plan_export(
        "demo",
        template_task.artifact_id,
    )
    assert template_name == "轨旁AP逐站规划模板.xlsx"
    template_workbook = load_workbook(template_path)
    assert template_workbook.sheetnames == [
        "AP规划",
        "字段说明",
        "_netconsole_meta",
    ]
    assert template_workbook["AP规划"].max_row == 1
    template_headers = [
        template_workbook["AP规划"].cell(1, column).value
        for column in range(1, 7)
    ]
    assert template_headers == [
        "序号",
        "车站ID",
        "车站名称",
        "AP数量",
        "AP管理VLAN",
        "备注",
    ]
    assert not {
        "VLAN组编号",
        "VLAN组名称",
        "实际上线",
        "未上线",
        "上线率",
        "revision",
    }.intersection(template_headers)
    template_workbook["AP规划"].append(
        [
            1,
            station_ids["小洋江站"],
            "小洋江站",
            28,
            921,
            "左线01、02无法铺设，核减2个AP，原30个AP",
        ]
    )
    template_workbook.save(template_path)
    template_workbook.close()
    repository.replace_trackside_ap_plan_rows(TRACKSIDE_AP_PLAN_MODE, [])
    roundtrip = service.preview_trackside_ap_plan(
        "demo",
        file_name="轨旁AP逐站规划模板.xlsx",
        content=template_path.read_bytes(),
        duplicate_strategy="replace",
    )
    assert roundtrip.can_apply is True, roundtrip.model_dump()
    roundtrip_row = next(
        row
        for row in roundtrip.result_rows
        if row.station_name == "小洋江站"
    )
    assert roundtrip_row.planned_ap_count == 28
    assert roundtrip_row.management_vlan == 921


@pytest.mark.parametrize(
    ("ap_count", "management_vlan", "expected_vlan"),
    [
        (0, None, None),
        (0, 71, 71),
    ],
)
def test_trackside_plan_zero_count_accepts_empty_or_valid_vlan(
    ap_count: int,
    management_vlan: object,
    expected_vlan: int | None,
) -> None:
    row = normalize_trackside_plan_row(
        {
            "station_id": "station-1",
            "sequence_no": 1,
            "station_name": "站点A",
            "planned_ap_count": ap_count,
            "management_vlan": management_vlan,
        }
    )

    assert row["ap_count"] == 0
    assert row["management_vlan"] == expected_vlan
    assert row["ap_management_vlans"] == (
        "" if expected_vlan is None else str(expected_vlan)
    )


@pytest.mark.parametrize(
    ("ap_count", "management_vlan", "message"),
    [
        (1, None, "AP数量大于 0 时必填"),
        (-1, None, "AP数量：必须是非负整数"),
        (0, 0, "必须在 1～4094 范围内"),
        (0, 4095, "必须在 1～4094 范围内"),
        (0, 71.5, "AP管理VLAN：必须是整数"),
    ],
)
def test_trackside_plan_rejects_invalid_count_or_vlan(
    ap_count: int,
    management_vlan: object,
    message: str,
) -> None:
    with pytest.raises(ValueError, match=message):
        normalize_trackside_plan_row(
            {
                "station_id": "station-1",
                "sequence_no": 1,
                "station_name": "站点A",
                "planned_ap_count": ap_count,
                "management_vlan": management_vlan,
            }
        )


def test_trackside_plan_import_and_save_keep_empty_vlan(tmp_path: Path) -> None:
    paths = PathResolver(app_root=tmp_path, data_root=tmp_path)
    paths.ensure_site_dirs("demo")
    database = Database(paths.site_db_path("demo"))
    database.initialize()
    repository = AcRepository(database)
    station_ids = _seed_base_stations(repository, ["站点A"])
    tasks = TaskApplicationService(paths=paths, site_name="demo")
    process = FakeLocalProcessAdapter(tasks)
    export = FakeExportProcessAdapter(tasks)
    service = RailTransitWebApplicationService(
        paths,
        tasks,
        process_adapter=process,  # type: ignore[arg-type]
        export_adapter=export,  # type: ignore[arg-type]
    )

    preview = service.preview_trackside_ap_plan(
        "demo",
        file_name="zero.csv",
        content=(
            "序号,车站ID,车站名称,AP数量,AP管理VLAN,备注\r\n"
            f"1,{station_ids['站点A']},站点A,0,,尚未规划\r\n"
        ).encode("utf-8-sig"),
        duplicate_strategy="replace",
    )
    assert preview.can_apply is True, [
        (row.status, row.message, row.row) for row in preview.rows
    ]
    assert preview.error_count == 0
    assert len(preview.result_rows) == 1, preview.model_dump()
    assert preview.result_rows[0].management_vlan is None

    started = service.start_trackside_ap_plan_save(
        "demo",
        rows=[row.model_dump() for row in preview.result_rows],
        explicit_confirmation=True,
    )
    saved_row = process.jobs[started.task_id].params["rows"][0]

    assert saved_row["management_vlan"] is None
    assert saved_row["ap_management_vlans"] == ""

    repository.replace_trackside_ap_plan_rows(
        TRACKSIDE_AP_PLAN_MODE,
        process.jobs[started.task_id].params["rows"],
    )
    current = service.start_trackside_ap_plan_export("demo", template=False)
    current_job = export.jobs[current.task_id]
    run_generic_export_handler(current_job)
    workbook = load_workbook(current_job.output_path)
    try:
        assert workbook["AP规划"]["D2"].value == 0
        assert workbook["AP规划"]["E2"].value is None
    finally:
        workbook.close()


def test_legacy_grouped_trackside_plan_import_uses_station_values(
    tmp_path: Path,
) -> None:
    paths = PathResolver(app_root=tmp_path, data_root=tmp_path)
    paths.ensure_site_dirs("demo")
    database = Database(paths.site_db_path("demo"))
    database.initialize()
    repository = AcRepository(database)
    station_ids = _seed_base_stations(repository, ["站点A", "站点B"])
    repository.replace_trackside_ap_plan_rows(
        TRACKSIDE_AP_PLAN_MODE,
        [
            {
                "station_id": station_ids["站点A"],
                "station_name": "站点A",
                "ap_count": 1,
                "ap_start_address": "10.1.0.10",
                "mask_length": 24,
                "ap_gateway": "10.1.0.1",
                "ap_management_vlans": "71",
                "remark": "",
            },
            {
                "station_id": station_ids["站点B"],
                "station_name": "站点B",
                "ap_count": 1,
                "ap_start_address": "10.1.0.11",
                "mask_length": 24,
                "ap_gateway": "10.1.0.1",
                "ap_management_vlans": "71",
                "remark": "",
            },
        ],
    )
    tasks = TaskApplicationService(paths=paths, site_name="demo")
    service = RailTransitWebApplicationService(
        paths,
        tasks,
        process_adapter=FakeLocalProcessAdapter(tasks),  # type: ignore[arg-type]
        export_adapter=FakeExportProcessAdapter(tasks),  # type: ignore[arg-type]
    )
    content = (
        "AP管理VLAN规划方式,VLAN组编号,VLAN组名称,管理VLAN,网络地址,"
        "子网掩码,默认网关,组AP起始地址,组成员站点ID,组成员站点,"
        "车站名称,AP数量,AP起始地址,"
        "掩码,AP网关,AP管理VLAN,备注\r\n"
        "station_grouped,G001,一号组,71,10.1.0.0,255.255.255.0,"
        f"10.1.0.1,10.1.0.10,\"{station_ids['站点A']},{station_ids['站点B']}\",站点A、站点B,"
        "站点A,1,10.1.0.10,24,10.1.0.1,71,\r\n"
        "station_grouped,G001,一号组,72,10.2.0.0,255.255.255.0,"
        f"10.2.0.1,10.2.0.10,\"{station_ids['站点A']},{station_ids['站点B']}\",站点A、站点B,"
        "站点B,1,10.2.0.10,24,10.2.0.1,72,\r\n"
    ).encode("utf-8-sig")

    preview = service.preview_trackside_ap_plan(
        "demo",
        file_name="grouped.csv",
        content=content,
        duplicate_strategy="replace",
    )

    assert preview.can_apply is True
    assert preview.error_count == 0
    assert preview.legacy_schema is True
    assert preview.message == "已识别旧版 VLAN 分组模板，将转换为逐站 AP 规划。"
    assert [row.management_vlan for row in preview.result_rows] == [71, 72]
    assert preview.result_plan is not None
    assert [row.station_name for row in preview.result_plan.items] == ["站点A", "站点B"]


def test_legacy_xlsx_plan_import_maps_count_and_group_vlan_fallback(
    tmp_path: Path,
) -> None:
    paths = PathResolver(app_root=tmp_path, data_root=tmp_path)
    paths.ensure_site_dirs("demo")
    database = Database(paths.site_db_path("demo"))
    database.initialize()
    station_ids = _seed_base_stations(AcRepository(database), ["小洋江站"])
    tasks = TaskApplicationService(paths=paths, site_name="demo")
    service = RailTransitWebApplicationService(
        paths,
        tasks,
        process_adapter=FakeLocalProcessAdapter(tasks),  # type: ignore[arg-type]
        export_adapter=FakeExportProcessAdapter(tasks),  # type: ignore[arg-type]
    )
    legacy_path = tmp_path / "legacy-trackside-plan.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "轨旁AP规划"
    sheet.append(
        [
            "车站名称",
            "组成员站点ID",
            "AP数量",
            "AP起始地址",
            "掩码",
            "AP网关",
            "AP管理VLAN",
            "备注",
            "VLAN组编号",
            "管理VLAN",
        ]
    )
    sheet.append(
        [
            "小洋江站",
            station_ids["小洋江站"],
            28,
            "10.122.221.X",
            "/24",
            "10.122.221.254",
            "",
            "核减2个AP，原30个AP",
            "G001",
            921,
        ]
    )
    workbook.create_sheet("字段说明")
    workbook.save(legacy_path)
    workbook.close()

    preview = service.preview_trackside_ap_plan(
        "demo",
        file_name=legacy_path.name,
        content=legacy_path.read_bytes(),
        duplicate_strategy="replace",
    )

    assert preview.can_apply is True
    assert preview.legacy_schema is True
    assert preview.message == "已识别旧版 VLAN 分组模板，将转换为逐站 AP 规划。"
    assert len(preview.result_rows) == 1
    assert preview.result_rows[0].planned_ap_count == 28
    assert preview.result_rows[0].management_vlan == 921


def test_trackside_plan_preview_finds_second_row_header_and_all_fifteen_rows(
    tmp_path: Path,
) -> None:
    paths = PathResolver(app_root=tmp_path, data_root=tmp_path)
    paths.ensure_site_dirs("demo")
    database = Database(paths.site_db_path("demo"))
    database.initialize()
    repository = AcRepository(database)
    station_names = [f"站点{index:02d}" for index in range(1, 16)]
    station_ids = _seed_base_stations(repository, station_names)
    tasks = TaskApplicationService(paths=paths, site_name="demo")
    service = RailTransitWebApplicationService(
        paths,
        tasks,
        process_adapter=FakeLocalProcessAdapter(tasks),  # type: ignore[arg-type]
        export_adapter=FakeExportProcessAdapter(tasks),  # type: ignore[arg-type]
    )
    source_path = tmp_path / "fifteen-stations.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "AP规划"
    sheet.append(["轨旁 AP 逐站规划"])
    sheet.append(["序号", "车站ID", "车站名称", "AP数量", "AP管理VLAN", "备注"])
    for index, station_name in enumerate(station_names, start=1):
        sheet.append([index, station_ids[station_name], station_name, index * 2, 921, ""])
    workbook.save(source_path)
    workbook.close()

    preview = service.preview_trackside_ap_plan(
        "demo",
        file_name=source_path.name,
        content=source_path.read_bytes(),
        duplicate_strategy="replace",
    )

    assert preview.total_count == 15
    assert preview.valid_count == 15
    assert preview.error_count == 0
    assert preview.can_apply is True
    assert [row.row_number for row in preview.rows] == list(range(3, 18))
    assert len(preview.result_rows) == 15
    assert {row.management_vlan for row in preview.result_rows} == {921}


def test_trackside_plan_preview_keeps_errors_and_applies_valid_rows(
    tmp_path: Path,
) -> None:
    paths = PathResolver(app_root=tmp_path, data_root=tmp_path)
    paths.ensure_site_dirs("demo")
    database = Database(paths.site_db_path("demo"))
    database.initialize()
    station_ids = _seed_base_stations(AcRepository(database), ["小洋江站", "横溪站"])
    tasks = TaskApplicationService(paths=paths, site_name="demo")
    service = RailTransitWebApplicationService(
        paths,
        tasks,
        process_adapter=FakeLocalProcessAdapter(tasks),  # type: ignore[arg-type]
        export_adapter=FakeExportProcessAdapter(tasks),  # type: ignore[arg-type]
    )
    content = (
        "序号,车站ID,车站名称,AP数量,AP管理VLAN,备注\r\n"
        f"1,{station_ids['小洋江站']},15-小洋江站,28,921,有效\r\n"
        "2,station:missing,不存在站,invalid,922,错误\r\n"
        f"3,{station_ids['横溪站']},横溪站,126,921,同VLAN合法\r\n"
    ).encode("utf-8-sig")

    preview = service.preview_trackside_ap_plan(
        "demo",
        file_name="partial.csv",
        content=content,
        duplicate_strategy="replace",
    )

    assert preview.can_apply is True
    assert preview.valid_count == 2
    assert preview.error_count == 1
    assert [row.station_name for row in preview.result_rows] == [
        "小洋江站",
        "横溪站",
    ]
    error_row = next(row for row in preview.rows if row.status == "error")
    assert error_row.row_number == 3
    assert error_row.row is not None
    assert error_row.row["station_name"] == "不存在站"
    assert "AP数量" in error_row.message


def test_trackside_plan_station_binding_requires_stable_id() -> None:
    stations = [
        {"id": "station-1", "name": "01小洋江站"},
        {"id": "station-2", "name": "02小洋江站"},
    ]

    exact = bind_trackside_plan_station(
        {"station_id": "station-1", "station_name": "旧名称"},
        stations,
        row_number=2,
    )

    assert exact["station_id"] == "station-1"
    assert exact["station_name"] == "01小洋江站"
    with pytest.raises(ValueError, match="车站ID：必填"):
        bind_trackside_plan_station(
            {"station_name": "15-小洋江站"},
            stations,
            row_number=3,
        )


def test_legacy_grouped_trackside_plan_import_ignores_retired_network_fields(
    tmp_path: Path,
) -> None:
    paths = PathResolver(app_root=tmp_path, data_root=tmp_path)
    paths.ensure_site_dirs("demo")
    database = Database(paths.site_db_path("demo"))
    database.initialize()
    repository = AcRepository(database)
    station_ids = _seed_base_stations(repository, ["站点A", "站点B"])
    repository.replace_trackside_ap_plan_rows(
        TRACKSIDE_AP_PLAN_MODE,
        [
            {
                "station_id": station_ids[station_name],
                "station_name": station_name,
                "ap_count": 1,
                "ap_start_address": "",
                "mask_length": None,
                "ap_gateway": "",
                "ap_management_vlans": "71",
                "remark": "",
            }
            for station_name in ("站点A", "站点B")
        ],
    )
    tasks = TaskApplicationService(paths=paths, site_name="demo")
    service = RailTransitWebApplicationService(
        paths,
        tasks,
        process_adapter=FakeLocalProcessAdapter(tasks),  # type: ignore[arg-type]
        export_adapter=FakeExportProcessAdapter(tasks),  # type: ignore[arg-type]
    )
    content = (
        "AP管理VLAN规划方式,VLAN组编号,VLAN组名称,管理VLAN,网络地址,"
        "子网掩码,默认网关,组AP起始地址,组成员站点ID,组成员站点,"
        "车站名称,AP数量,AP起始地址,掩码,AP网关,AP管理VLAN,备注\r\n"
        "station_grouped,G001,一号组,71,invalid-network,invalid-mask,"
        f"invalid-gateway,invalid-start,\"{station_ids['站点A']},{station_ids['站点B']}\",站点A、站点B,"
        "站点A,1,invalid-a,invalid-mask,invalid-gateway,71,\r\n"
        "station_grouped,G001,一号组,71,192.0.2.0,255.255.255.0,"
        f"203.0.113.254,198.51.100.10,\"{station_ids['站点A']},{station_ids['站点B']}\",站点A、站点B,"
        "站点B,1,invalid-b,another-mask,another-gateway,71,\r\n"
    ).encode("utf-8-sig")

    preview = service.preview_trackside_ap_plan(
        "demo",
        file_name="grouped.csv",
        content=content,
        duplicate_strategy="replace",
    )

    assert preview.can_apply is True
    assert preview.error_count == 0
    assert all(row.status == "duplicate" for row in preview.rows)
    assert [row.management_vlan for row in preview.result_rows] == [71, 71]


def test_trackside_ap_online_status_uses_planned_targets_and_weighted_total(
    tmp_path: Path,
) -> None:
    paths = PathResolver(app_root=tmp_path, data_root=tmp_path)
    paths.ensure_site_dirs("demo")
    database = Database(paths.site_db_path("demo"))
    database.initialize()
    repository = AcRepository(database)
    tasks = TaskApplicationService(paths=paths, site_name="demo")
    export = FakeExportProcessAdapter(tasks)
    service = RailTransitWebApplicationService(
        paths,
        tasks,
        process_adapter=FakeLocalProcessAdapter(tasks),  # type: ignore[arg-type]
        export_adapter=export,  # type: ignore[arg-type]
    )
    samples = [
        ("小洋江站", 28, 28, 28, "左线01、02无法铺设，核减2个AP，原30个AP"),
        ("云龙火车站站", 56, 56, 56, ""),
        ("横溪站", 126, 126, 77, ""),
        ("塘溪站", 138, 138, 133, ""),
        ("鄞州咸祥", 134, 134, 134, ""),
        ("象山贤庠", 134, 134, 131, ""),
        ("大徐站", 94, 94, 94, ""),
        ("丹城站", 78, 78, 0, ""),
        ("滨海大道站", 56, 56, 0, ""),
        ("大目湾站", 34, 34, 0, "路段未铺设，核减8个AP，原42个AP"),
        ("云龙车辆段", 67, 67, 66, ""),
    ]
    _seed_effective_trackside_scope(
        repository,
        samples,
        extra_references=[
            {
                "site_id": "demo",
                "belong_type": "station",
                "station_name": "小洋江站",
                "ap_name": "暂停 AP",
                "ap_mac_norm": "aabbccdde001",
                "raw_payload_json": json.dumps(
                    {
                        "station_node_uid": "station-node-1",
                        "operation_status": "suspended",
                        "project_id": "demo",
                        "ap_uuid": "ap-suspended",
                    },
                    ensure_ascii=False,
                ),
            },
            {
                "site_id": "demo",
                "belong_type": "station",
                "station_name": "小洋江站",
                "ap_name": "退役 AP",
                "ap_mac_norm": "aabbccdde002",
                "raw_payload_json": json.dumps(
                    {
                        "station_node_uid": "station-node-1",
                        "operation_status": "retired",
                        "project_id": "demo",
                        "ap_uuid": "ap-retired",
                    },
                    ensure_ascii=False,
                ),
            },
        ],
        extra_resources=[
            {
                "ap_uuid": "ap-suspended",
                "ap_name": "暂停 AP",
                "ap_mac": "aabbccdde001",
                    "state": "R/M",
                "updated_at": "2026-07-30T11:30:25+08:00",
            },
            {
                "ap_uuid": "ap-retired",
                "ap_name": "退役 AP",
                "ap_mac": "aabbccdde002",
                    "state": "R/M",
                "updated_at": "2026-07-30T11:30:25+08:00",
            },
            {
                "ap_uuid": "ap-unassigned",
                "ap_name": "未分配 AP",
                "ap_mac": "aabbccdde003",
                    "state": "R/M",
                "updated_at": "2026-07-30T11:30:25+08:00",
            },
        ],
    )

    status = service.get_trackside_ap_online_status("demo")

    assert status.planned_ap_count == 945
    assert status.actual_online_count == 719
    assert status.offline_count == 226
    assert status.online_rate == 76.1
    assert status.unassigned_count == 1
    assert "1 个 AC 在线 AP" in status.warning
    assert status.excluded_device_count == 2
    assert status.fit_ap_resource_total_count == 948
    assert status.fit_ap_matched_count == 945
    assert status.fit_ap_unmatched_online_count == 1
    assert status.cache_hit is False
    assert status.revision
    assert status.source_revision["fit_ap_count"] == 948
    assert status.excluded_items == []
    assert status.unmatched_online_items == []
    assert status.unassigned_items == []
    cached_status = service.get_trackside_ap_online_status("demo")
    assert cached_status.cache_hit is True
    assert cached_status.revision == status.revision
    excluded_page = service.list_trackside_ap_online_excluded(
        "demo",
        page=1,
        page_size=1,
    )
    unmatched_page = service.list_trackside_ap_online_unmatched(
        "demo",
        page=1,
        page_size=1,
    )
    assert excluded_page.total == 4
    assert len(excluded_page.items) == 1
    assert unmatched_page.total == 1
    assert unmatched_page.items[0].item_id == "ap-unassigned"
    assert "actual_ap_count" not in status.model_dump()
    assert "online_count" not in status.model_dump()
    by_name = {row.station_name: row for row in status.items}
    assert by_name["01-小洋江站"].planned_ap_count == 28
    assert by_name["01-小洋江站"].actual_online_count == 28
    assert by_name["01-小洋江站"].offline_count == 0
    assert by_name["01-小洋江站"].online_rate == 100.0
    assert by_name["03-横溪站"].online_rate == 61.1
    assert by_name["04-塘溪站"].online_rate == 96.4
    assert by_name["06-象山贤庠"].online_rate == 97.8
    assert by_name["10-大目湾站"].planned_ap_count == 34
    assert by_name["10-大目湾站"].actual_online_count == 0
    assert by_name["10-大目湾站"].offline_count == 34
    assert by_name["10-大目湾站"].online_rate == 0.0
    assert by_name["11-云龙车辆段"].online_rate == 98.5
    assert all(row.count_anomaly is False for row in status.items)

    current = service.start_trackside_ap_plan_export("demo", template=False)
    current_job = export.jobs[current.task_id]
    run_generic_export_handler(current_job)
    workbook = load_workbook(current_job.output_path)
    overview = workbook["AP上线情况概览"]
    assert [overview.cell(2, column).value for column in range(1, 6)] == [
        "01-小洋江站",
        28,
        28,
        0,
        1,
    ]
    assert [overview.cell(4, column).value for column in range(1, 6)] == [
        "03-横溪站",
        126,
        77,
        49,
        0.611,
    ]
    assert [overview.cell(11, column).value for column in range(1, 6)] == [
        "10-大目湾站",
        34,
        0,
        34,
        0,
    ]
    assert [overview.cell(13, column).value for column in range(1, 5)] == [
        "合计",
        945,
        719,
        226,
    ]
    assert "未计入业务统计" in str(overview.cell(13, 6).value)
    assert overview["E13"].value == pytest.approx(0.761)
    assert overview["E13"].number_format == "0.0%"
    assert overview["E13"].font.bold is True
    workbook.close()


def test_trackside_online_status_refreshes_from_exact_switch_lldp_station_evidence(
    tmp_path: Path,
) -> None:
    paths = PathResolver(app_root=tmp_path, data_root=tmp_path)
    paths.ensure_site_dirs("demo")
    database = Database(paths.site_db_path("demo"))
    database.initialize()
    repository = AcRepository(database)
    station_ids = _seed_base_stations(repository, ["站点A"])
    repository.replace_trackside_ap_plan_rows(
        TRACKSIDE_AP_PLAN_MODE,
        [
            {
                "station_id": station_ids["站点A"],
                "station_name": "站点A",
                "ap_count": 1,
                "ap_management_vlans": "921",
            }
        ],
    )
    repository.replace_fit_ap_resources(
        "ac-fixture",
        [
            {
                "ap_uuid": "ap-runtime-1",
                "ap_name": "AP-RUNTIME-1",
                "ap_mac": "0011-2233-4455",
                "state": "R/M",
            }
        ],
    )
    group = DeviceGroupRepository(database, "demo").create("车站")
    switch = DeviceRepository(database).create(
        Device(
            name="SW-A",
            primary_address="192.0.2.10",
            device_type="SW",
            group_id=group.id,
            station="站点A",
            station_id="",
            project_phase="phase_1",
            work_scope_status="included",
        )
    )
    tasks = TaskApplicationService(paths=paths, site_name="demo")
    service = RailTransitWebApplicationService(
        paths,
        tasks,
        process_adapter=FakeLocalProcessAdapter(tasks),  # type: ignore[arg-type]
        export_adapter=FakeExportProcessAdapter(tasks),  # type: ignore[arg-type]
    )

    before = service.get_trackside_ap_online_status("demo")
    assert before.actual_online_count == 0
    assert before.fit_ap_unmatched_online_count == 1

    DeviceFactRepository(database).replace_lldp_neighbors(
        str(switch.device_uuid),
        [
            {
                "local_interface": "XGE1/0/1",
                "neighbor_mac": "00:11:22:33:44:55",
                "collected_at": "2026-08-02T12:00:00+08:00",
            }
        ],
        preserve_existing=False,
    )

    after = service.get_trackside_ap_online_status("demo")
    assert after.cache_hit is False
    assert after.revision != before.revision
    assert after.actual_online_count == 1
    assert after.fit_ap_matched_count == 1
    assert after.fit_ap_unmatched_online_count == 0
    assert after.scope_ap_reference_count == 0
    assert after.source_revision["station_switch_lldp_count"] == 1

    snapshot = load_trackside_ap_business_snapshot(
        DeviceRepository(database),
        "demo",
        generation=1,
    )
    assert snapshot.scope is not None
    assert snapshot.scope.resources[0]["station_id"] == station_ids["站点A"]
    assert snapshot.scope.resources[0]["site"] == "站点A"


def test_trackside_online_status_recovers_from_ac_lldp_after_switch_identity_arrives(
    tmp_path: Path,
) -> None:
    paths = PathResolver(app_root=tmp_path, data_root=tmp_path)
    paths.ensure_site_dirs("demo")
    database = Database(paths.site_db_path("demo"))
    database.initialize()
    repository = AcRepository(database)
    station_ids = _seed_base_stations(repository, ["站点A"])
    repository.replace_trackside_ap_plan_rows(
        TRACKSIDE_AP_PLAN_MODE,
        [{
            "station_id": station_ids["站点A"],
            "station_name": "站点A",
            "ap_count": 1,
            "ap_management_vlans": "921",
        }],
    )
    repository.replace_fit_ap_resources(
        "ac-fixture",
        [{
            "ap_uuid": "ap-runtime-1",
            "ap_name": "AP-RUNTIME-1",
            "ap_mac": "0011-2233-4455",
            "state": "R/M",
            "lldp_neighbor_name": "HZDT-SC",
            "lldp_neighbor_interface": "gei-0/3/0/1",
            "lldp_match_status": "matched",
        }],
    )
    group = DeviceGroupRepository(database, "demo").create("车站")
    switch = DeviceRepository(database).create(
        Device(
            name="SW-ZTE-A",
            primary_address="192.0.2.10",
            device_type="SW",
            group_id=group.id,
            station="站点A",
            station_id="",
            project_phase="phase_1",
            work_scope_status="included",
        )
    )
    tasks = TaskApplicationService(paths=paths, site_name="demo")
    service = RailTransitWebApplicationService(
        paths,
        tasks,
        process_adapter=FakeLocalProcessAdapter(tasks),  # type: ignore[arg-type]
        export_adapter=FakeExportProcessAdapter(tasks),  # type: ignore[arg-type]
    )

    before = service.get_trackside_ap_online_status("demo")
    assert before.actual_online_count == 0
    assert before.fit_ap_switch_not_found_count == 1
    before_detail = service.list_trackside_ap_online_unmatched("demo")
    assert before_detail.items[0].reason_code == "SWITCH_NOT_FOUND"

    DeviceFactRepository(database).upsert_device_fact({
        "device_uuid": str(switch.device_uuid),
        "sysname": "hzdt_sc.example.com",
        "collected_at": "2026-08-06T12:00:00+08:00",
        "updated_at": "2026-08-06T12:00:00+08:00",
    })

    after = service.get_trackside_ap_online_status("demo")
    assert after.cache_hit is False
    assert after.revision != before.revision
    assert after.actual_online_count == 1
    assert after.fit_ap_matched_online_count == 1
    assert after.fit_ap_unmatched_online_count == 0
    assert after.source_revision["station_switch_fact_count"] == 1

    snapshot = load_trackside_ap_business_snapshot(
        DeviceRepository(database),
        "demo",
        generation=1,
    )
    assert snapshot.scope is not None
    assert snapshot.scope.resources[0]["_scope_binding_source"] == "ac_lldp_switch_identity"


def test_trackside_online_status_retries_when_source_revision_changes(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    paths = PathResolver(app_root=tmp_path, data_root=tmp_path)
    paths.ensure_site_dirs("demo")
    database = Database(paths.site_db_path("demo"))
    database.initialize()
    repository = AcRepository(database)
    _seed_effective_trackside_scope(
        repository,
        [("站点A", 1, 1, 1, "")],
        numbered_display=False,
    )
    tasks = TaskApplicationService(paths=paths, site_name="demo")
    service = RailTransitWebApplicationService(
        paths,
        tasks,
        process_adapter=FakeLocalProcessAdapter(tasks),  # type: ignore[arg-type]
        export_adapter=FakeExportProcessAdapter(tasks),  # type: ignore[arg-type]
    )
    original_revision = service._trackside_online_revision
    calls = 0

    def changing_revision(site_id: str, current_repository: AcRepository):
        nonlocal calls
        calls += 1
        revision, source_revision = original_revision(site_id, current_repository)
        return (
            "stale-first-read" if calls == 1 else revision,
            source_revision,
        )

    monkeypatch.setattr(service, "_trackside_online_revision", changing_revision)

    status = service.get_trackside_ap_online_status("demo")

    assert calls == 4
    assert status.actual_online_count == 1
    assert status.cache_hit is False


def test_trackside_ap_online_status_ignores_reference_count_and_flags_excess(
    tmp_path: Path,
) -> None:
    paths = PathResolver(app_root=tmp_path, data_root=tmp_path)
    paths.ensure_site_dirs("demo")
    database = Database(paths.site_db_path("demo"))
    database.initialize()
    repository = AcRepository(database)
    tasks = TaskApplicationService(paths=paths, site_name="demo")
    service = RailTransitWebApplicationService(
        paths,
        tasks,
        process_adapter=FakeLocalProcessAdapter(tasks),  # type: ignore[arg-type]
        export_adapter=FakeExportProcessAdapter(tasks),  # type: ignore[arg-type]
    )
    _seed_effective_trackside_scope(
        repository,
        [
            ("站点A", 5, 2, 1, ""),
            ("站点B", 1, 2, 2, ""),
            ("站点C", 0, 1, 1, ""),
        ],
        numbered_display=False,
    )

    status = service.get_trackside_ap_online_status("demo")
    by_name = {row.station_name: row for row in status.items}

    assert by_name["站点A"].planned_ap_count == 5
    assert by_name["站点A"].actual_online_count == 1
    assert by_name["站点A"].offline_count == 4
    assert by_name["站点A"].online_rate == 20.0
    assert by_name["站点B"].actual_online_count == 1
    assert by_name["站点B"].offline_count == 0
    assert by_name["站点B"].online_rate is None
    assert by_name["站点B"].count_anomaly is True
    assert by_name["站点B"].status == "over_planned"
    assert (
        by_name["站点B"].warning
        == "实际上线 AP 数量超过当前规划数量，请检查规划资料或 AP 归属关系。"
    )
    assert by_name["站点C"].planned_ap_count == 0
    assert by_name["站点C"].actual_online_count == 0
    assert by_name["站点C"].offline_count == 0
    assert by_name["站点C"].online_rate is None
    assert by_name["站点C"].count_anomaly is True
    assert status.planned_ap_count == 6
    assert status.actual_online_count == 2
    assert status.offline_count == 4
    assert status.online_rate is None
    assert status.count_anomaly is True
    assert status.scope_device_count == 5


def test_trackside_ap_base_template_and_draft_export_use_controlled_artifacts(tmp_path: Path) -> None:
    paths = PathResolver(app_root=tmp_path, data_root=tmp_path)
    paths.ensure_site_dirs("demo")
    Database(paths.site_db_path("demo")).initialize()
    tasks = TaskApplicationService(paths=paths, site_name="demo")
    export = FakeExportProcessAdapter(tasks)
    service = RailTransitWebApplicationService(
        paths,
        tasks,
        process_adapter=FakeLocalProcessAdapter(tasks),  # type: ignore[arg-type]
        export_adapter=export,  # type: ignore[arg-type]
    )

    with pytest.raises(WebArtifactError, match="路径不在受控目录"):
        service.artifact_store.reserve(
            site_id="demo",
            owner=service._OWNER,
            source="trackside_ap_plan",
            artifact_type="xlsx",
            task_id="outside-task",
            task_type=service._ARTIFACT_TASK_TYPES["trackside_ap_plan"],
            output_root=tmp_path / "outside",
            preferred_name="outside.xlsx",
        )

    template = service.start_trackside_ap_base_export("demo", template=True)
    template_job = export.jobs[template.task_id]
    run_generic_export_handler(template_job)
    template_content = Path(template_job.output_path).read_bytes()
    export.complete(template.task_id, template_content)
    template_task = service.get_task("demo", template.task_id)
    template_path, _name = service.open_trackside_ap_base_export("demo", template_task.artifact_id)
    workbook = load_workbook(template_path)
    assert workbook.sheetnames[:2] == ["轨旁AP", "字段说明"]
    assert [cell.value for cell in workbook["轨旁AP"][1]][:6] == ["AP名称", "点位编号", "AP厂商", "AP MAC", "管理 IP", "型号"]
    assert workbook["字段说明"]["A2"].value == "AP名称"
    workbook.close()

    draft = service.start_trackside_ap_base_export(
        "demo",
        template=False,
        rows=[
            {
                "id": "new:1",
                "site_id": "demo",
                "line_name": "宁波地铁1号线",
                "name": "",
                "point_code": "AP0127",
                "vendor": "H3C",
                "mac": "1c94-6876-8ee0",
                "station": "高桥西",
                "section": "高桥西-高桥-上行",
                "section_start_station": "高桥西",
                "section_end_station": "高桥",
                "mileage": {"raw": "", "normalized": "", "meters": None, "valid": False},
                "direction": "上行",
                "runtime": {"fit_ap_status": "unknown", "optical_status": "no_data"},
                "record_kind": "section",
                "base_metadata": {"uplink_switch": "11-高桥西1", "uplink_port": "GE1/0/1"},
            }
        ],
    )
    draft_job = export.jobs[draft.task_id]
    run_generic_export_handler(draft_job)
    draft_content = Path(draft_job.output_path).read_bytes()
    export.complete(draft.task_id, draft_content)
    draft_task = service.get_task("demo", draft.task_id)
    draft_path, draft_name = service.open_trackside_ap_base_export("demo", draft_task.artifact_id)
    assert draft_name.startswith("demo_轨旁AP基础资料_")
    workbook = load_workbook(draft_path)
    header = [cell.value for cell in workbook["轨旁AP"][1]]
    values = [cell.value for cell in workbook["轨旁AP"][2]]
    row = dict(zip(header, values, strict=True))
    assert row["点位编号"] == "AP0127"
    assert row["AP厂商"] == "H3C"
    assert row["上联交换机"] == "11-高桥西1"
    assert row["上联端口"] == "GE1/0/1"
    workbook.close()
