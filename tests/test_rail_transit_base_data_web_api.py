from __future__ import annotations

import hashlib
import json
from io import BytesIO
from pathlib import Path
from urllib.parse import unquote

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from rail_transit_base_data_fixture import build_rail_transit_base_data_fixture
from netconsole.backend.api.main import create_app
from netconsole.core.database import Database
from netconsole.core.runtime_mode import RuntimeMode
from netconsole.application.rail_transit.base_data_application_service import (
    RailTransitBaseDataApplicationService,
)
from netconsole.repositories.rail_transit_base_data_repository import (
    RailTransitBaseDataRepository,
)


class _NoopAsyncService:
    async def start(self) -> None:
        return None

    async def stop(self) -> None:
        return None


def _fingerprint(path: Path) -> tuple[str, int]:
    return hashlib.sha256(path.read_bytes()).hexdigest(), path.stat().st_mtime_ns


def _app(paths, tmp_path: Path):
    return create_app(
        RuntimeMode.SERVER,
        paths=paths,
        task_service=object(),  # type: ignore[arg-type]
        agent_service=_NoopAsyncService(),  # type: ignore[arg-type]
        traffic_service=_NoopAsyncService(),  # type: ignore[arg-type]
        frontend_dist=tmp_path / "missing",
    )


def _desktop_app(paths, tmp_path: Path):
    return create_app(
        RuntimeMode.DESKTOP,
        paths=paths,
        task_service=object(),  # type: ignore[arg-type]
        agent_service=_NoopAsyncService(),  # type: ignore[arg-type]
        traffic_service=_NoopAsyncService(),  # type: ignore[arg-type]
        frontend_dist=tmp_path / "missing",
        desktop_session_token="d" * 40,
        rail_base_data_write_feature_enabled=True,
        rail_base_data_desktop_write_enabled=True,
    )


def _insert_station_source_devices(db_path: Path) -> None:
    now = "2026-07-22T08:00:00"
    database = Database(db_path)
    with database.connect() as conn:
        station_group = conn.execute(
            "INSERT INTO device_groups (site_id, name, sort_order, created_at, updated_at) VALUES ('demo', '车站', 2, ?, ?)",
            (now, now),
        ).lastrowid
        other_group = conn.execute(
            "INSERT INTO device_groups (site_id, name, sort_order, created_at, updated_at) VALUES ('demo', '其他', 3, ?, ?)",
            (now, now),
        ).lastrowid
        rows = [
            ("station-wuxiang-1", "32-五乡1", "SYS-WX1", "32-五乡", station_group, "10.20.0.1"),
            ("station-wuxiang-2", "32-五乡2", "SYS-WX2", "32-五乡", station_group, "10.20.0.2"),
            ("station-baozhuang", "33-宝幢1", "33-错误系统名", "33-宝幢", station_group, "10.20.0.3"),
            ("station-yard", "50-高桥西停车场", "YARD-SYS", "50-高桥西停车场", station_group, "10.20.0.4"),
            ("station-depot", "52-天童庄车辆段", "DEPOT-SYS", "52-天童庄车辆段", station_group, "10.20.0.5"),
            ("station-empty", "错误设备名称99", "SYS-EMPTY", "", station_group, "10.20.0.6"),
            ("other-station", "99-不应读取", "OTHER-SYS", "99-不应读取", other_group, "10.20.0.7"),
        ]
        rows.extend(
            (f"station-bulk-{index}", f"批量设备{index}", f"BULK-{index}", "34-批量站", station_group, f"10.21.0.{index}")
            for index in range(1, 206)
        )
        conn.executemany(
            """
            INSERT INTO devices (
                device_uuid, name, system_name, station, group_id, primary_address,
                device_vendor, device_type, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, 'H3C', 'SWITCH', ?, ?)
            """,
            [(uuid, name, system_name, station, group_id, address, now, now) for uuid, name, system_name, station, group_id, address in rows],
        )
        conn.commit()
    with database.connect() as conn:
        conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")


def test_router_delegates_import_policy_without_guard_or_source_policy_access() -> None:
    router_source = (
        Path(__file__).parents[1] / "src/netconsole/backend/api/rail_transit_base_data_router.py"
    ).read_text(encoding="utf-8")

    assert "get_import_policy" in router_source
    assert ".guard" not in router_source
    assert "source_policy" not in router_source
    assert "import_policy_rows" not in router_source


def test_base_data_api_defaults_to_locked_and_redacts_credentials(tmp_path: Path) -> None:
    paths, db_path = build_rail_transit_base_data_fixture(tmp_path)
    app = create_app(
        RuntimeMode.SERVER,
        paths=paths,
        task_service=object(),  # type: ignore[arg-type]
        agent_service=_NoopAsyncService(),  # type: ignore[arg-type]
        traffic_service=_NoopAsyncService(),  # type: ignore[arg-type]
        frontend_dist=tmp_path / "missing",
    )
    with TestClient(app) as client:
        assert client.get("/api/rail-transit/base-data/summary").status_code == 200
        before = _fingerprint(db_path)
        responses = [
            client.get("/api/rail-transit/base-data/summary"),
            client.get("/api/rail-transit/base-data/stations"),
            client.get("/api/rail-transit/base-data/sections"),
            client.get("/api/rail-transit/base-data/aps?page=1&page_size=2"),
            client.get("/api/rail-transit/base-data/trains"),
            client.get("/api/rail-transit/base-data/mrs"),
            client.get("/api/rail-transit/base-data/issues"),
            client.get("/api/rail-transit/base-data/issues/groups"),
            client.get("/api/rail-transit/base-data/import-policies"),
            client.get("/api/rail-transit/base-data/relations"),
        ]
        ap_id = responses[3].json()["items"][0]["id"]
        train_id = responses[4].json()["items"][0]["id"]
        mr_id = responses[5].json()["items"][0]["id"]
        responses.extend(
            [
                client.get(f"/api/rail-transit/base-data/aps/{ap_id}"),
                client.get(f"/api/rail-transit/base-data/trains/{train_id}"),
                client.get(f"/api/rail-transit/base-data/mrs/{mr_id}"),
            ]
        )
        preview = client.post(
            "/api/rail-transit/base-data/import-preview",
            files={
                "file": (
                    "preview.json",
                    json.dumps(
                        [{"ap_name": "AP-Preview", "ap_mac_display": "0011-2233-4455"}],
                        ensure_ascii=False,
                    ).encode("utf-8"),
                    "application/json",
                )
            },
        )
        responses.append(preview)
    assert all(response.status_code == 200 for response in responses)
    text = "".join(response.text for response in responses).casefold()
    assert "private-user" not in text
    assert "private-pass" not in text
    assert "password" not in text
    assert _fingerprint(db_path) == before

    routes = {
        (path, method.upper())
        for path, operations in app.openapi()["paths"].items()
        if path.startswith("/api/rail-transit/base-data")
        for method in operations
    }
    assert {path for path, method in routes if method == "POST"} == {
        "/api/rail-transit/base-data/import-preview",
        "/api/rail-transit/base-data/station-template-preview",
        "/api/rail-transit/base-data/section-generation-preview",
        "/api/rail-transit/base-data/stations/delete-preflight",
        "/api/rail-transit/base-data/clear-all",
        "/api/rail-transit/base-data/import-apply",
        "/api/rail-transit/base-data/import-operations/{operation_id}/rollback",
        "/api/rail-transit/base-data/validate",
        "/api/rail-transit/base-data/changes",
    }
    assert not any(method in {"PUT", "PATCH", "DELETE"} for _path, method in routes)
    assert responses[8].json()["write_enabled"] is False


def test_station_source_preview_uses_station_field_only_and_is_read_only(tmp_path: Path) -> None:
    paths, db_path = build_rail_transit_base_data_fixture(tmp_path)
    _insert_station_source_devices(db_path)
    with TestClient(_app(paths, tmp_path)) as client:
        before = _fingerprint(db_path)
        response = client.get("/api/rail-transit/base-data/station-source-preview")
    assert response.status_code == 200
    assert _fingerprint(db_path) == before
    payload = response.json()
    assert payload["group_found"] is True
    assert payload["source_field"] == "station"
    assert payload["scanned_device_count"] == 211
    assert payload["empty_station_device_count"] == 1
    assert payload["unique_station_value_count"] == 5

    by_name = {item["name"]: item for item in payload["candidates"]}
    assert by_name["五乡"]["code"] == "32"
    assert by_name["五乡"]["sort_order"] == 32
    assert by_name["五乡"]["source_device_count"] == 2
    assert by_name["五乡"]["proposed_station"]["structure_type"] == "underground"
    assert by_name["五乡"]["proposed_station"]["platform_layout"] == "island"
    assert by_name["宝幢"]["code"] == "33"
    assert by_name["批量站"]["source_device_count"] == 205
    assert by_name["高桥西停车场"]["node_type"] == "parking_lot"
    assert by_name["高桥西停车场"]["path_code"] == "UNASSIGNED"
    assert by_name["高桥西停车场"]["sort_order"] is None
    assert by_name["高桥西停车场"]["participates_in_direction"] is False
    assert by_name["高桥西停车场"]["proposed_station"]["structure_type"] == "unknown"
    assert by_name["高桥西停车场"]["proposed_station"]["platform_layout"] == "unknown"
    assert by_name["天童庄车辆段"]["node_type"] == "depot"
    assert by_name["天童庄车辆段"]["sort_order"] is None
    text = response.text
    assert "五乡1" not in text
    assert "五乡2" not in text
    assert "33-错误系统名" not in text
    assert "错误设备名称99" not in text
    assert "99-不应读取" not in text
    assert any(issue["code"] == "station_source_value_empty" for issue in payload["issues"])


def test_station_source_preview_normalizes_prefix_variants_and_matches_existing_name(
    tmp_path: Path,
) -> None:
    paths, db_path = build_rail_transit_base_data_fixture(tmp_path)
    now = "2026-07-28T08:00:00"
    database = Database(db_path)
    with database.connect() as conn:
        conn.execute(
            "UPDATE ap_extension_points SET section_end_station = '3.车站C' WHERE section_end_station = '车站C'"
        )
        group_id = conn.execute(
            "INSERT INTO device_groups (site_id, name, sort_order, created_at, updated_at) VALUES ('demo', '车站', 2, ?, ?)",
            (now, now),
        ).lastrowid
        conn.executemany(
            """
            INSERT INTO devices (
                device_uuid, name, system_name, station, group_id, primary_address,
                device_vendor, device_type, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, 'H3C', 'SWITCH', ?, ?)
            """,
            [
                ("station-prefix-a", "设备A", "SYS-A", "01车站A", group_id, "10.23.0.1", now, now),
                ("station-prefix-b", "设备B", "SYS-B", "1.车站A", group_id, "10.23.0.2", now, now),
                ("station-no-prefix", "设备C", "SYS-C", "车站B", group_id, "10.23.0.3", now, now),
                ("station-depot", "设备D", "SYS-D", "11云龙车辆段", group_id, "10.23.0.4", now, now),
                ("station-existing-numbered", "设备E", "SYS-E", "车站C", group_id, "10.23.0.5", now, now),
            ],
        )
        conn.commit()

    with TestClient(_app(paths, tmp_path)) as client:
        payload = client.get("/api/rail-transit/base-data/station-source-preview").json()

    by_name = {item["name"]: item for item in payload["candidates"]}
    assert len(payload["candidates"]) == 4
    assert by_name["车站A"]["source_device_count"] == 2
    assert by_name["车站A"]["source_order"] == 1
    assert by_name["车站A"]["match_status"] == "exact_source_key"
    assert by_name["车站A"]["matched_station_name"] == "车站A"
    assert by_name["车站A"]["suggested_action"] == "覆盖现有"
    assert by_name["车站B"]["sort_order"] is None
    assert by_name["车站B"]["match_status"] in {
        "exact_source_key",
        "canonical_name",
        "canonical_name_and_type",
    }
    assert by_name["车站C"]["match_status"] in {
        "exact_source_key",
        "canonical_name",
        "canonical_name_and_type",
    }
    assert by_name["车站C"]["matched_station_name"] == "3.车站C"
    assert by_name["云龙车辆段"]["source_order"] == 11
    assert by_name["云龙车辆段"]["sort_order"] is None
    assert by_name["云龙车辆段"]["participates_in_direction"] is False
    assert not any(
        issue["code"] == "station_source_parse_failed"
        for candidate in payload["candidates"]
        for issue in candidate["issues"]
    )


def test_station_source_preview_reports_missing_group_without_writes(tmp_path: Path) -> None:
    paths, db_path = build_rail_transit_base_data_fixture(tmp_path)
    with TestClient(_app(paths, tmp_path)) as client:
        before = _fingerprint(db_path)
        payload = client.get("/api/rail-transit/base-data/station-source-preview").json()
    assert _fingerprint(db_path) == before
    assert payload["group_found"] is False
    assert payload["candidates"] == []
    assert payload["issues"][0]["code"] == "station_source_group_missing"


def test_station_source_preview_blocks_same_name_with_different_node_type(
    tmp_path: Path,
) -> None:
    paths, db_path = build_rail_transit_base_data_fixture(tmp_path)
    repository = RailTransitBaseDataRepository(paths)
    repository.apply_base_data_changes(
        "demo",
        repository.base_data_revision("demo"),
        [{
            "entity_type": "station",
            "action": "create",
            "values": {
                "name": "云龙车辆段",
                "node_uid": "existing-yunlong",
                "node_type": "station",
                "path_code": "MAIN",
                "sort_order": 11,
                "participates_in_direction": True,
            },
        }],
    )
    now = "2026-07-28T08:00:00"
    with Database(db_path).connect() as conn:
        group_id = conn.execute(
            "INSERT INTO device_groups (site_id, name, sort_order, created_at, updated_at) VALUES ('demo', '车站', 2, ?, ?)",
            (now, now),
        ).lastrowid
        conn.execute(
            """
            INSERT INTO devices (
                device_uuid, name, system_name, station, group_id, primary_address,
                device_vendor, device_type, created_at, updated_at
            ) VALUES ('station-node-type', '设备A', 'SYS-A', '11云龙车辆段', ?, '10.24.0.1', 'H3C', 'SWITCH', ?, ?)
            """,
            (group_id, now, now),
        )
        conn.commit()

    with TestClient(_app(paths, tmp_path)) as client:
        payload = client.get("/api/rail-transit/base-data/station-source-preview").json()

    candidate = payload["candidates"][0]
    assert candidate["match_status"] == "conflict"
    assert candidate["suggested_action"] == "处理来源冲突"
    assert any(
        issue["code"] == "station_source_node_type_conflict"
        for issue in candidate["issues"]
    )


def test_station_source_preview_flags_code_and_name_conflicts(tmp_path: Path) -> None:
    paths, db_path = build_rail_transit_base_data_fixture(tmp_path)
    now = "2026-07-22T08:00:00"
    database = Database(db_path)
    with database.connect() as conn:
        group_id = conn.execute(
            "INSERT INTO device_groups (site_id, name, sort_order, created_at, updated_at) VALUES ('demo', '车站', 2, ?, ?)",
            (now, now),
        ).lastrowid
        conn.executemany(
            """
            INSERT INTO devices (
                device_uuid, name, system_name, station, group_id, primary_address,
                device_vendor, device_type, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, 'H3C', 'SWITCH', ?, ?)
            """,
            [
                ("station-code-a", "设备A", "SYS-A", "10-甲站", group_id, "10.22.0.1", now, now),
                ("station-code-b", "设备B", "SYS-B", "10-乙站", group_id, "10.22.0.2", now, now),
                ("station-name-a", "设备C", "SYS-C", "11-丙站", group_id, "10.22.0.3", now, now),
                ("station-name-b", "设备D", "SYS-D", "12-丙站", group_id, "10.22.0.4", now, now),
            ],
        )
        conn.commit()
    with TestClient(_app(paths, tmp_path)) as client:
        payload = client.get("/api/rail-transit/base-data/station-source-preview").json()
    issue_codes = {
        issue["code"]
        for candidate in payload["candidates"]
        for issue in candidate["issues"]
    }
    assert payload["conflict_count"] == 4
    assert "station_source_code_conflict" in issue_codes
    assert "station_source_name_conflict" in issue_codes


def test_station_source_preview_matches_real_style_numbered_batch_without_duplicates(
    tmp_path: Path,
) -> None:
    paths, db_path = build_rail_transit_base_data_fixture(tmp_path)
    repository = RailTransitBaseDataRepository(paths)
    existing = RailTransitBaseDataApplicationService._station_values(
        {
            "name": "1.小洋江站",
            "code": "01",
            "sort_order": 1,
            "node_type": "station",
            "path_code": "MAIN",
            "structure_type": "elevated",
            "platform_layout": "side",
            "center_mileage_text": "K1+234",
            "remark": "人工维护字段必须保留",
        },
        "create",
    )
    repository.apply_base_data_changes(
        "demo",
        repository.base_data_revision("demo"),
        [
            {
                "entity_type": "station",
                "action": "create",
                "entity_id": "new:station-numbered",
                "values": existing,
            }
        ],
    )
    station_values = [
        "01小洋江站",
        "02云龙火车站",
        "03甲站",
        "04乙站",
        "05丙站",
        "06丁站",
        "07戊站",
        "08己站",
        "09庚站",
        "10辛站",
        "11云龙车辆段",
    ]
    now = "2026-07-27T08:00:00"
    with Database(db_path).connect() as conn:
        group_id = conn.execute(
            "INSERT INTO device_groups (site_id, name, sort_order, created_at, updated_at) "
            "VALUES ('demo', '车站', 2, ?, ?)",
            (now, now),
        ).lastrowid
        rows = []
        sequence = 0
        for index, station in enumerate(station_values):
            count = 2 if index < 7 else 3
            for _ in range(count):
                sequence += 1
                rows.append(
                    (
                        f"station-real-{sequence}",
                        f"现场设备{sequence}",
                        f"REAL-{sequence}",
                        station,
                        group_id,
                        f"10.27.{sequence // 250}.{sequence % 250 + 1}",
                        now,
                        now,
                    )
                )
        conn.executemany(
            """
            INSERT INTO devices (
                device_uuid, name, system_name, station, group_id, primary_address,
                device_vendor, device_type, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, 'H3C', 'SWITCH', ?, ?)
            """,
            rows,
        )
        conn.commit()

    with TestClient(_app(paths, tmp_path)) as client:
        before = _fingerprint(db_path)
        payload = client.get(
            "/api/rail-transit/base-data/station-source-preview"
        ).json()

    assert _fingerprint(db_path) == before
    assert payload["scanned_device_count"] == 26
    assert payload["unique_station_value_count"] == 11
    assert payload["normal_station_count"] == 10
    assert payload["special_node_count"] == 1
    assert payload["manual_review_count"] == 0
    by_name = {item["name"]: item for item in payload["candidates"]}
    assert by_name["小洋江站"]["matched_station_name"] == "1.小洋江站"
    assert by_name["小洋江站"]["match_status"] in {
        "exact_source_key",
        "canonical_name",
        "canonical_name_and_type",
    }
    assert by_name["小洋江站"]["suggested_action"] == "覆盖现有"
    assert by_name["小洋江站"]["processing_strategy"] == "overwrite_existing"
    assert by_name["小洋江站"]["matched_station_ids"]
    assert by_name["云龙车辆段"]["node_type"] == "depot"
    assert by_name["云龙车辆段"]["sort_order"] is None
    assert {item["name"] for item in payload["candidates"]} == set(
        station.removeprefix(f"{index:02d}")
        for index, station in enumerate(station_values, start=1)
    )


def test_station_template_download_preview_and_export_are_structured_xlsx(tmp_path: Path) -> None:
    paths, db_path = build_rail_transit_base_data_fixture(tmp_path)
    _insert_station_source_devices(db_path)
    with TestClient(_app(paths, tmp_path)) as client:
        template = client.get("/api/rail-transit/base-data/station-template")
        exported = client.get("/api/rail-transit/base-data/station-template-export")
    assert template.status_code == 200
    assert exported.status_code == 200
    assert template.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    assert "线路站点与区间基础资料模板.xlsx" in unquote(template.headers["content-disposition"])
    assert "线路站点与区间基础资料.xlsx" in unquote(exported.headers["content-disposition"])
    workbook = load_workbook(BytesIO(template.content))
    assert workbook.sheetnames == ["01_线路参数", "02_线路节点", "03_区间配置", "字段说明"]
    assert workbook["01_线路参数"]["J2"].value == "station"
    exported_wb = load_workbook(BytesIO(exported.content))
    assert exported_wb["02_线路节点"].max_row >= 2
    assert "03_区间配置" in exported_wb.sheetnames

    upload = Workbook()
    upload.active.title = "01_线路参数"
    upload.active.append(["线路名称", "项目类型", "网络类型", "主线路径编码", "站序递增方向名称", "站序递减方向名称", "设备来源分组", "设备来源字段", "备注"])
    upload.active.append(["测试线", "PIS", "default", "MAIN", "上行", "下行", "车站", "station", ""])
    nodes = upload.create_sheet("02_线路节点")
    nodes.append(["来源站点值", "节点编码", "节点名称", "节点类型", "所属路径", "主线顺序", "参与方向判断", "车站结构", "站台形式", "线路端点", "运营终点", "可折返", "折返类型", "折返方向", "启用", "备注"])
    nodes.append(["32-五乡", "32", "五乡", "普通车站", "MAIN", 32, "是", "地下", "岛式", "否", "否", "否", "无", "无", "是", "模板导入"])
    nodes.append(["50-高桥西停车场", "50", "高桥西停车场", "停车场", "UNASSIGNED", "", "否", "未填写", "未填写", "否", "否", "否", "无", "无", "是", ""])
    upload.create_sheet("字段说明")
    stream = BytesIO()
    upload.save(stream)
    with TestClient(_app(paths, tmp_path)) as client:
        before = _fingerprint(db_path)
        preview = client.post(
            "/api/rail-transit/base-data/station-template-preview",
            files={"file": ("stations.xlsx", stream.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert preview.status_code == 200
    assert _fingerprint(db_path) == before
    payload = preview.json()
    assert payload["valid"] is True
    assert payload["line_metadata"]["station_source_field"] == "station"
    assert payload["rows"][0]["proposed_station"]["structure_type"] == "underground"
    assert payload["rows"][1]["proposed_station"]["node_type"] == "parking_lot"
    assert payload["rows"][1]["proposed_station"]["sort_order"] is None


def test_section_generation_preview_uses_request_draft_and_is_read_only(
    tmp_path: Path,
) -> None:
    paths, db_path = build_rail_transit_base_data_fixture(tmp_path)
    with TestClient(_app(paths, tmp_path)) as client:
        revision = client.get(
            "/api/rail-transit/base-data/revision"
        ).json()["base_revision"]
        before = _fingerprint(db_path)
        request = {
            "site_id": "demo",
            "base_revision": revision,
            "line_metadata": {
                "main_path_code": "MAIN",
                "increasing_direction_name": "上行",
                "decreasing_direction_name": "下行",
            },
            "stations": [
                {
                    "id": "new:low",
                    "node_uid": "node-low",
                    "name": "草稿甲站",
                    "sort_order": 11,
                    "path_code": "MAIN",
                    "participates_in_direction": True,
                    "structure_type": "underground",
                    "platform_layout": "island",
                    "center_mileage_text": "100",
                    "center_mileage_m": 100,
                },
                {
                    "id": "new:high",
                    "node_uid": "node-high",
                    "name": "草稿乙站",
                    "sort_order": 20,
                    "path_code": "MAIN",
                    "participates_in_direction": True,
                    "structure_type": "underground",
                    "platform_layout": "island",
                    "center_mileage_text": "200",
                    "center_mileage_m": 200,
                },
            ],
            "current_sections": [],
        }
        preview = client.post(
            "/api/rail-transit/base-data/section-generation-preview",
            json=request,
        )
        stale = client.post(
            "/api/rail-transit/base-data/section-generation-preview",
            json={**request, "base_revision": "0" * 64},
        )

    assert preview.status_code == 200
    assert stale.status_code == 409
    assert stale.json()["detail"]["code"] == "BASE_DATA_REVISION_CONFLICT"
    assert _fingerprint(db_path) == before
    sections = {
        item["proposed_section"]["name"]: item["proposed_section"]
        for item in preview.json()["generated_sections"]
    }
    assert sections["草稿甲站-草稿乙站-上行"]["start_station"] == "草稿甲站"
    assert sections["草稿甲站-草稿乙站-下行"]["start_station"] == "草稿乙站"
    assert sections["草稿甲站-草稿乙站-上行"]["section_mileage_start_m"] == 100
    assert sections["草稿甲站-草稿乙站-上行"]["section_mileage_end_m"] == 200
    assert sections["草稿甲站-草稿乙站-下行"]["section_mileage_start_m"] == 100
    assert sections["草稿甲站-草稿乙站-下行"]["section_mileage_end_m"] == 200


def test_station_template_real_file_preview_save_export_and_reimport_round_trip(
    tmp_path: Path,
) -> None:
    paths, db_path = build_rail_transit_base_data_fixture(tmp_path)
    template_path = tmp_path / "线路站点与区间基础资料模板.xlsx"
    edited_path = tmp_path / "线路站点与区间基础资料导入.xlsx"
    exported_path = tmp_path / "线路站点与区间基础资料.xlsx"

    with TestClient(_desktop_app(paths, tmp_path)) as client:
        session_response = client.post(
            "/__desktop_session",
            data={"token": "d" * 40},
            follow_redirects=False,
        )
        assert session_response.status_code == 303

        template = client.get("/api/rail-transit/base-data/station-template")
        assert template.status_code == 200
        template_path.write_bytes(template.content)
        workbook = load_workbook(template_path)
        assert workbook.sheetnames == [
            "01_线路参数",
            "02_线路节点",
            "03_区间配置",
            "字段说明",
        ]
        line_sheet = workbook["01_线路参数"]
        line_sheet.append([])
        line_sheet["A2"] = "XLSX闭环验收线"
        line_sheet["B2"] = "PIS"
        node_sheet = workbook["02_线路节点"]
        node_sheet.append([
            "81-验收甲站",
            "81",
            "验收甲站",
            "普通车站",
            "MAIN",
            81,
            "是",
            "K81+000",
            "地下",
            "岛式",
            "是",
            "是",
            "是",
            "折返线、存车线",
            "双向",
            "是",
            "端点",
            120,
            "K80+880",
            "是",
            "多设施验收",
        ])
        node_sheet.append([
            "82-验收乙站",
            "82",
            "验收乙站",
            "普通车站",
            "MAIN",
            82,
            "是",
            "82+250.5",
            "",
            "",
            "否",
            "否",
            "否",
            "",
            "无",
            "否",
            "端点",
            "",
            "",
            "是",
            "",
        ])
        section_sheet = workbook["03_区间配置"]
        section_sheet.append([
            "AUTO-ROUNDTRIP",
            "验收甲站-验收乙站-上行",
            "站间区间",
            "MAIN",
            "站序递增",
            "上行",
            "车站",
            "验收甲站",
            "车站",
            "验收乙站",
            "是",
            "MAIN|between|acceptance-low|acceptance-high|increasing",
            81000,
            82250.5,
            "否",
            "自动生成",
            "",
            "是",
            9,
            "0-99999 m",
            "区间往返验收",
        ])
        workbook.save(edited_path)
        load_workbook(edited_path).close()

        before_preview_revision = client.get(
            "/api/rail-transit/base-data/revision"
        ).json()["base_revision"]
        preview = client.post(
            "/api/rail-transit/base-data/station-template-preview",
            files={
                "file": (
                    edited_path.name,
                    edited_path.read_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert preview.status_code == 200
        payload = preview.json()
        assert payload["valid"] is True
        assert payload["create_count"] == 3
        assert client.get(
            "/api/rail-transit/base-data/revision"
        ).json()["base_revision"] == before_preview_revision

        station_value_fields = {
            "node_uid",
            "name",
            "code",
            "line_name",
            "sort_order",
            "remark",
            "source_station_value",
            "source_station_key",
            "node_type",
            "path_code",
            "participates_in_direction",
            "structure_type",
            "platform_layout",
            "center_mileage_text",
            "center_mileage_m",
            "is_line_terminal",
            "is_service_terminal",
            "turnback_capable",
            "turnback_type",
            "track_facilities",
            "turnback_direction",
            "terminal_extension_enabled",
            "terminal_endpoint_label",
            "terminal_extension_distance_m",
            "terminal_endpoint_mileage_text",
            "enabled",
            "source_kind",
        }
        section_value_fields = {
            "name",
            "section_code",
            "section_kind",
            "path_code",
            "direction_role",
            "line_direction",
            "start_node_type",
            "start_node_uid",
            "start_station",
            "end_node_type",
            "end_node_uid",
            "end_station",
            "line_side",
            "auto_generated",
            "generation_key",
            "manual_override_fields",
            "section_mileage_start_m",
            "section_mileage_end_m",
            "section_mileage_open_end",
            "section_mileage_source",
            "enabled",
            "source_kind",
            "remark",
        }
        changes = [
            {
                "entity_type": "site_metadata",
                "action": "update",
                "entity_id": "current",
                "values": payload["line_metadata"],
            },
            *[
                {
                    "entity_type": "station",
                    "action": "create",
                    "entity_id": f"new:roundtrip-station:{index}",
                    "values": {
                        key: value
                        for key, value in row["proposed_station"].items()
                        if key in station_value_fields
                    },
                }
                for index, row in enumerate(payload["rows"], start=1)
            ],
            *[
                {
                    "entity_type": "section",
                    "action": "create",
                    "entity_id": f"new:roundtrip-section:{index}",
                    "values": {
                        key: value
                        for key, value in row["proposed_section"].items()
                        if key in section_value_fields
                    },
                }
                for index, row in enumerate(payload["section_rows"], start=1)
            ],
        ]
        validation = client.post(
            "/api/rail-transit/base-data/validate",
            json={
                "site_id": "demo",
                "base_revision": before_preview_revision,
                "changes": changes,
            },
        )
        assert validation.status_code == 200
        assert validation.json()["valid"] is True
        saved = client.post(
            "/api/rail-transit/base-data/changes",
            json={
                "site_id": "demo",
                "base_revision": before_preview_revision,
                "changes": changes,
                "explicit_confirmation": True,
            },
        )
        assert saved.status_code == 200

        before_export = _fingerprint(db_path)
        exported = client.get("/api/rail-transit/base-data/station-template-export")
        assert exported.status_code == 200
        exported_path.write_bytes(exported.content)
        assert _fingerprint(db_path) == before_export
        exported_workbook = load_workbook(exported_path)
        exported_nodes = {
            row[2].value: row
            for row in exported_workbook["02_线路节点"].iter_rows(min_row=2)
            if row[2].value
        }
        exported_sections = {
            row[1].value: row
            for row in exported_workbook["03_区间配置"].iter_rows(min_row=2)
            if row[1].value
        }
        assert exported_nodes["验收甲站"][7].value == "K81+000"
        assert exported_nodes["验收甲站"][13].value == "折返线、存车线"
        assert exported_nodes["验收乙站"][7].value == "82+250.5"
        section_row = exported_sections["验收甲站-验收乙站-上行"]
        assert section_row[11].value == "MAIN|between|acceptance-low|acceptance-high|increasing"
        assert section_row[12].value == 81000
        assert section_row[13].value == 82250.5
        assert section_row[15].value == "自动生成"
        assert section_row[18].value == 0
        assert section_row[19].value == "--"
        exported_workbook.close()

        reimported = client.post(
            "/api/rail-transit/base-data/station-template-preview",
            files={
                "file": (
                    exported_path.name,
                    exported_path.read_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert reimported.status_code == 200
        reimported_payload = reimported.json()
        first_station = next(
            row["proposed_station"]
            for row in reimported_payload["rows"]
            if row["name"] == "验收甲站"
        )
        first_section = next(
            row["proposed_section"]
            for row in reimported_payload["section_rows"]
            if row["name"] == "验收甲站-验收乙站-上行"
        )
        assert first_station["track_facilities"] == [
            "turnback_track",
            "storage_track",
        ]
        assert first_station["center_mileage_m"] == 81000
        assert first_section["ap_count"] == 0
        assert first_section["mileage_min"] is None
        assert first_section["mileage_max"] is None
        assert first_section["section_mileage_start_m"] == 81000
        assert first_section["section_mileage_end_m"] == 82250.5
        assert first_section["section_mileage_open_end"] is False
        assert first_section["section_mileage_source"] == "generated"
