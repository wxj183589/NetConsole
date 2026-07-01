import warnings

from openpyxl import Workbook

from netconsole.utils.excel_workbook import load_workbook_without_unsupported_image_warning


def test_load_workbook_helper_suppresses_openpyxl_wmf_warning(tmp_path, monkeypatch):
    path = tmp_path / "sample.xlsx"
    Workbook().save(path)

    def fake_load_workbook(*_args, **_kwargs):
        warnings.warn("wmf image format is not supported so the image is being dropped", UserWarning)
        return "workbook"

    monkeypatch.setattr("netconsole.utils.excel_workbook.load_workbook", fake_load_workbook)

    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        workbook = load_workbook_without_unsupported_image_warning(path, data_only=True)

    assert workbook == "workbook"
    assert caught == []
