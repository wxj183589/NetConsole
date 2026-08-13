from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from urllib.parse import unquote

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from tests.support.web_parity_test_support import FakeExportProcessAdapter, FakeLocalProcessAdapter
from netconsole.application.rail_transit import web_application_service as rail_transit_web_application_service
from netconsole.application.rail_transit.web_application_service import (
    RailTransitWebApplicationService,
)
from netconsole.backend.api.main import create_app
from netconsole.core.database import Database
from netconsole.core.i18n import I18n
from netconsole.core.paths import PathResolver
from netconsole.core.sites import SiteManager
from netconsole.core.runtime_mode import RuntimeMode
from netconsole.models.device import Device
from netconsole.repositories.device_repository import DeviceRepository
from netconsole.services.ap_online_overview import AP_ONLINE_OVERVIEW_COLUMNS
from netconsole.services.export.export_job import ExportJob
from netconsole.services.file_contract import attach_export_metadata
from netconsole.services.job_center.task_application_service import TaskApplicationService
from netconsole.services.job_center.query_service import JobCenterQueryService
from netconsole.services.offline_ap_ledger import (
    OFFLINE_AP_LEDGER_COLUMNS,
    OFFLINE_AP_STATS_COLUMNS,
    offline_ap_headers,
)
from netconsole.services.trackside_ap_business import (
    AP_OPTICAL_TREATMENT_RECORD_COLUMNS,
    CURRENT_OPTICAL_ABNORMAL_COLUMNS,
    NEW_ONLINE_AP_OVERVIEW_COLUMNS,
    TRACKSIDE_AP_BUSINESS_SHEET_DEFINITIONS,
    TRACKSIDE_AP_BUSINESS_EXPORT_COLUMNS,
    TracksideApExportCancelled,
    build_ap_online_history_block,
    export_trackside_ap_business_xlsx,
    trackside_export_fill_status,
)
from netconsole.services.rail_transit.trackside_ap_business_snapshot import (
    TracksideApBusinessSnapshotError,
)
from netconsole.services.trackside_ap_export_service import build_trackside_ap_business_export_name
from netconsole.utils.interface_normalize import display_interface_name


class _NoopAsyncService:
    async def start(self) -> None:
        return None

    async def stop(self) -> None:
        return None


def _enable_feature(app, feature_id: str) -> None:
    app.state.feature_gate.features[feature_id] = {
        "visible": True,
        "enabled": True,
        "client_package": True,
        "internal_only": False,
    }


def test_trackside_business_export_name_uses_site_display_name_and_sanitizes_windows_chars() -> None:
    created_at = datetime(2026, 7, 21, 23, 45, 1)
    assert (
        build_trackside_ap_business_export_name("宁波地铁12号线", created_at)
        == "宁波地铁12号线_轨旁AP业务_20260721_234501.xlsx"
    )
    special = build_trackside_ap_business_export_name("测试/线路:A网", created_at)
    assert special == "测试_线路_A网_轨旁AP业务_20260721_234501.xlsx"
    assert all(char not in special for char in '<>:"/\\|?*')
    assert ".xlsx.xlsx" not in build_trackside_ap_business_export_name("测试.xlsx", created_at)
    with pytest.raises(ValueError, match="缺少局点名称"):
        build_trackside_ap_business_export_name("", created_at)


def test_trackside_export_contract_and_fill_matrix(tmp_path: Path) -> None:
    output = tmp_path / "trackside-fill-matrix.xlsx"
    i18n = I18n("zh_CN")
    rows = [
        {
            "site": "16-双陈站",
            "device_name": "SW-16",
            "interface_name": "gei-0/3/0/41",
            "link_status": "UP",
            "switch_rx_power": "-8.00",
            "switch_optical_status": "normal",
            "ap_mac": "0011-2233-4455",
            "ap_name": "AP-Normal",
            "ap_rx_power": "-8.00",
            "ap_optical_status": "normal",
            "ap_side_has_data": True,
        },
        {
            "site": "16-双陈站",
            "device_name": "SW-16",
            "interface_name": "gei-0/3/0/42",
            "link_status": "DOWN",
            "switch_optical_status": "no_module",
            "ap_mac": "-",
            "ap_name": "-",
            "ap_optical_status": "-",
        },
        {
            "site": "25-千秋广场站",
            "device_name": "SW-25",
            "interface_name": "gei-0/4/0/42",
            "link_status": "DOWN",
            "switch_optical_status": "no_light",
            "ap_mac": "-",
            "ap_name": "-",
            "ap_optical_status": "-",
        },
        {
            "site": "25-千秋广场站",
            "device_name": "SW-25",
            "interface_name": "gei-0/4/0/43",
            "link_status": "DOWN",
            "switch_rx_power": "-36.96",
            "switch_optical_status": "no_light",
            "ap_mac": "0011-2233-4466",
            "ap_name": "AP-Historical",
            "has_historical_lldp": True,
            "lldp_history_status": "stale_snapshot",
            "ap_rx_power": "-7.99",
            "ap_optical_status": "offline",
            "is_ap_offline": True,
        },
        {
            "site": "16-双陈站",
            "device_name": "SW-16",
            "interface_name": "gei-0/3/0/44",
            "link_status": "UP",
            "switch_rx_power": "-19.10",
            "switch_optical_status": "normal",
            "ap_mac": "0011-2233-4477",
            "ap_name": "AP-Abnormal",
            "ap_rx_power": "-8.00",
            "ap_optical_status": "normal",
            "ap_side_has_data": True,
        },
    ]
    export_trackside_ap_business_xlsx(
        output,
        rows,
        TRACKSIDE_AP_BUSINESS_EXPORT_COLUMNS,
        [i18n.t(key) for key, _field in TRACKSIDE_AP_BUSINESS_EXPORT_COLUMNS],
        current_optical_abnormal_headers=[
            i18n.t(key) for key, _field in CURRENT_OPTICAL_ABNORMAL_COLUMNS
        ],
    )

    workbook = load_workbook(output)
    main = workbook["轨旁AP业务"]
    abnormal = workbook["当前异常光衰"]
    assert [cell.value for cell in main[1]] == [
        i18n.t(key) for key, _field in TRACKSIDE_AP_BUSINESS_EXPORT_COLUMNS
    ]
    abnormal_headers = [cell.value for cell in abnormal[1]]
    assert abnormal_headers == [
        i18n.t(key) for key, _field in CURRENT_OPTICAL_ABNORMAL_COLUMNS
    ]
    assert len(abnormal_headers) == 19
    assert not {"AP业务光衰", "AP业务判定原因", "光衰判定"}.intersection(abnormal_headers)
    interface_column = abnormal_headers.index("接口名称") + 1
    assert [
        abnormal.cell(row=index, column=interface_column).value
        for index in range(2, abnormal.max_row + 1)
    ] == ["gei-0/3/0/44", "gei-0/4/0/43"]

    main_headers = [cell.value for cell in main[1]]
    main_interface_column = main_headers.index("接口名称") + 1
    fills = {
        main.cell(row=index, column=main_interface_column).value: main.cell(row=index, column=1).fill
        for index in range(2, main.max_row + 1)
    }
    assert fills["gei-0/3/0/41"].fill_type == "solid"
    assert fills["gei-0/3/0/41"].fgColor.rgb == "FFDCFCE7"
    assert fills["gei-0/3/0/42"].fill_type is None
    assert fills["gei-0/4/0/42"].fill_type is None
    assert fills["gei-0/4/0/43"].fill_type == "solid"
    assert fills["gei-0/3/0/44"].fill_type == "solid"
    assert trackside_export_fill_status(rows[1]) is None
    assert trackside_export_fill_status(rows[2]) is None


def test_trackside_business_export_uses_frozen_snapshot_without_optical_update_lock(
    tmp_path: Path,
) -> None:
    paths = PathResolver(app_root=tmp_path, data_root=tmp_path)
    paths.ensure_site_dirs("demo")
    Database(paths.site_db_path("demo")).initialize()
    SiteManager(paths).save_site_metadata("demo", {"display_name": "测试局点"})
    tasks = TaskApplicationService(paths=paths, site_name="demo")
    process = FakeLocalProcessAdapter(tasks)
    export = FakeExportProcessAdapter(tasks)
    service = RailTransitWebApplicationService(
        paths,
        tasks,
        process_adapter=process,  # type: ignore[arg-type]
        export_adapter=export,  # type: ignore[arg-type]
    )

    service.start_trackside_ap_update("demo")

    started = service.start_trackside_ap_business_export("demo")

    snapshot = tasks.repository("demo").get(started.task_id)
    assert snapshot is not None
    assert snapshot.task_type == "web_export_trackside_ap_business"
    assert snapshot.resource_keys == []


def test_trackside_business_export_api_uses_owned_artifact_and_supports_cancel(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    paths = PathResolver(app_root=tmp_path, data_root=tmp_path)
    database = Database(paths.site_db_path("demo"))
    database.initialize()
    zte_device = DeviceRepository(database).create(
        Device(
            device_uuid="11111111-1111-4111-8111-111111111111",
            name="ZTE-SW-01",
            station="站点A",
            device_vendor="ZTE",
            device_type="SW",
            primary_address="192.0.2.10",
        )
    )
    SiteManager(paths).save_site_metadata("demo", {"display_name": "宁波地铁12号线"})
    paths.config_dir.mkdir(parents=True, exist_ok=True)
    paths.app_config_path.write_text('{"current_site":"demo"}', encoding="utf-8")

    class FrozenDateTime(datetime):
        @classmethod
        def now(cls, tz=None):  # type: ignore[override]
            return datetime(2026, 7, 21, 23, 45, 1, tzinfo=tz)

    monkeypatch.setattr(rail_transit_web_application_service, "datetime", FrozenDateTime)
    app = create_app(
        RuntimeMode.SERVER,
        paths=paths,
        agent_service=_NoopAsyncService(),  # type: ignore[arg-type]
        traffic_service=_NoopAsyncService(),  # type: ignore[arg-type]
        frontend_dist=tmp_path / "missing",
    )
    process = FakeLocalProcessAdapter(app.state.task_service)
    export = FakeExportProcessAdapter(app.state.task_service)
    app.state.rail_transit_web_application_service = RailTransitWebApplicationService(
        paths,
        app.state.task_service,
        process_adapter=process,  # type: ignore[arg-type]
        export_adapter=export,  # type: ignore[arg-type]
    )
    _enable_feature(app, "capability.trackside_ap.export")
    _enable_feature(app, "capability.trackside_ap.base_io")
    _enable_feature(app, "capability.trackside_ap.plan_export")
    _enable_feature(app, "capability.rail_transit.task_control")
    _enable_feature(app, "rail.zte_trackside_switch_adapter")

    with TestClient(app) as client:
        started = client.post(
            "/api/rail-transit/trackside-ap-business/export",
            json={"optical_anomaly_only": True},
        )
        assert started.status_code == 202
        payload = started.json()
        expected_name = "宁波地铁12号线_轨旁AP业务_20260721_234501.xlsx"
        assert payload["action"] == "trackside_ap_business_export"
        assert payload["artifact_id"]
        assert payload["artifact_name"] == expected_name
        assert str(tmp_path) not in json.dumps(payload, ensure_ascii=False)

        task_id = payload["task_id"]
        job = export.jobs[task_id]
        assert job.job_type == "trackside_ap_business"
        assert job.site_name == "demo"
        assert Path(job.db_path) == paths.site_db_path("demo")
        assert job.params["language"] == "zh_CN"
        assert Path(str(job.params["snapshot_staging_root"])) == paths.staging_dir
        assert job.params["expected_revision"] == ""
        assert job.params["scope_context"]["site_id"] == "demo"
        assert job.params["scope_context"]["display_name"] == "宁波地铁12号线"
        assert job.params["selected_row_ids"] == []
        assert "optical_anomaly_only" not in job.params
        assert "snapshot_path" not in job.params
        assert "snapshot_sha256" not in job.params
        assert Path(job.output_path).name == expected_name

        export.complete(
            task_id,
            b"xlsx-fixture",
            result={
                "snapshot_id": "snapshot-1",
                "business_revision": "b" * 64,
                "export_revision": "e" * 64,
                "source_revisions": {"base_data_revision": "1"},
                "content_sha256": "c" * 64,
                "export_content_sha256": "d" * 64,
                "snapshot_created_at": "2026-08-06T01:00:00+08:00",
                "export_kind": "trackside_ap_business",
                "row_count": 2,
                "abnormal_count": 1,
                "unresolved_count": 1,
                "ambiguous_count": 0,
                "identity_distinct_count": 2,
                "identity_revision": 7,
                "snapshot_build_ms": 12,
                "snapshot_retry_count": 1,
                "export_render_ms": 8,
            },
        )
        completed = client.get(
            f"/api/rail-transit/trackside-ap-business/tasks/{task_id}"
        )
        assert completed.status_code == 200
        completed_payload = completed.json()
        assert completed_payload["available"] is True
        assert completed_payload["artifact_name"] == expected_name
        assert completed_payload["result_summary"] == {
            "snapshot_id": "snapshot-1",
            "business_revision": "b" * 64,
            "export_revision": "e" * 64,
            "source_revisions": {"base_data_revision": "1"},
            "content_sha256": "c" * 64,
            "export_content_sha256": "d" * 64,
            "snapshot_created_at": "2026-08-06T01:00:00+08:00",
            "export_kind": "trackside_ap_business",
            "row_count": 2,
            "abnormal_count": 1,
            "unresolved_count": 1,
            "ambiguous_count": 0,
            "identity_distinct_count": 2,
            "identity_revision": 7,
            "snapshot_build_ms": 12,
            "snapshot_retry_count": 1,
            "export_render_ms": 8,
        }
        job_detail = JobCenterQueryService(paths).get_task("demo", task_id)
        assert job_detail is not None
        assert job_detail.details["source_revisions"] == {
            "base_data_revision": "1"
        }
        assert job_detail.details["export_kind"] == "trackside_ap_business"
        assert job_detail.details["abnormal_count"] == 1
        assert str(tmp_path) not in json.dumps(completed_payload, ensure_ascii=False)

        download = client.get(
            "/api/rail-transit/trackside-ap-business/artifacts/"
            f"{payload['artifact_id']}/download"
        )
        assert download.status_code == 200
        assert download.content == b"xlsx-fixture"
        assert expected_name in unquote(download.headers["content-disposition"])

        recovered = client.post("/api/rail-transit/trackside-ap-business/export")
        recovered_payload = recovered.json()
        recovered_name = "宁波地铁12号线_轨旁AP业务_20260721_234501.xlsx"
        export.callbacks.pop(recovered_payload["task_id"])
        export.complete(recovered_payload["task_id"], b"recovered-fixture")
        recovered_tasks = client.post("/api/rail-transit/trackside-ap-business/tasks/recover")
        assert recovered_tasks.status_code == 200
        recovered_items = recovered_tasks.json()
        recovered_item = next(item for item in recovered_items if item["task_id"] == recovered_payload["task_id"])
        assert recovered_item["artifact_name"] == recovered_name
        recovered_detail = client.get(
            f"/api/rail-transit/trackside-ap-business/tasks/{recovered_payload['task_id']}"
        )
        assert recovered_detail.json()["artifact_name"] == recovered_name

        cancelling = client.post("/api/rail-transit/trackside-ap-business/export")
        cancelling_payload = cancelling.json()
        cancelled = client.post(
            "/api/rail-transit/trackside-ap-business/tasks/"
            f"{cancelling_payload['task_id']}/cancel"
        )
        assert cancelled.status_code == 200
        assert cancelled.json()["status"] == "CANCELLED"
        missing = client.get(
            "/api/rail-transit/trackside-ap-business/artifacts/"
            f"{cancelling_payload['artifact_id']}/download"
        )
        assert missing.status_code == 404

        plan_template = client.post(
            "/api/rail-transit/trackside-ap-business/plan/export",
            json={"template": True},
        )
        assert plan_template.status_code == 202
        assert export.jobs[plan_template.json()["task_id"]].job_type == "multi_sheet_xlsx"
        export.complete(plan_template.json()["task_id"], b"plan-template-fixture")

        base_template = client.post(
            "/api/rail-transit/trackside-ap-business/base/export",
            json={"template": True},
        )
        assert base_template.status_code == 202
        assert export.jobs[base_template.json()["task_id"]].job_type == "trackside_ap_base_xlsx"
        export.complete(base_template.json()["task_id"], b"base-template-fixture")

        rename_export = client.post(
            "/api/rail-transit/trackside-ap-business/base/rename-commands/export",
            json={},
        )
        assert rename_export.status_code == 202
        rename_payload = rename_export.json()
        expected_rename_name = "轨旁AP重命名命令_宁波地铁12号线_20260721_234501.txt"
        assert rename_payload["action"] == "trackside_ap_rename_command_export"
        assert rename_payload["artifact_name"] == expected_rename_name
        assert export.jobs[rename_payload["task_id"]].job_type == "trackside_ap_rename_commands"
        export.complete(rename_payload["task_id"], b"rename-fixture")
        rename_download = client.get(
            "/api/rail-transit/trackside-ap-business/base/rename-commands/artifacts/"
            f"{rename_payload['artifact_id']}/download"
        )
        assert rename_download.status_code == 200
        assert rename_download.content == b"rename-fixture"
        assert expected_rename_name in unquote(rename_download.headers["content-disposition"])

        adapters = client.get(
            "/api/rail-transit/trackside-ap-business/switch-adapters"
        )
        assert adapters.status_code == 200
        assert adapters.json()["items"][0]["device_uuid"] == zte_device.device_uuid
        assert (
            adapters.json()["items"][0]["adapter"]["adaptation_status"]
            == "C89E-4 Release 已验证；其他 ZXR10/5960X 型号需逐型号复核"
        )

        sample = client.post(
            "/api/rail-transit/trackside-ap-business/switch-adapters/sample",
            json={
                "device_uuid": zte_device.device_uuid,
                "vendor": "ZTE",
                "command_profile": "zte_zxr10_5960x_es_v2",
                "selected_interface": "xgei-0/1/1/2",
                "requested_commands": ["device_version", "lldp_global"],
            },
        )
        assert sample.status_code == 202
        sample_payload = sample.json()
        assert sample_payload["action"] == "switch_vendor_sample_collect"
        sample_job = process.jobs[sample_payload["task_id"]]
        sample_output = Path(str(sample_job.params["artifact_output_path"]))
        sample_output.parent.mkdir(parents=True, exist_ok=True)
        sample_output.write_bytes(b"zip-fixture")
        process.complete(sample_payload["task_id"], {"status": "PARTIAL_SUCCESS"})

        sample_detail = client.get(
            f"/api/rail-transit/trackside-ap-business/tasks/{sample_payload['task_id']}"
        )
        assert sample_detail.status_code == 200
        assert sample_detail.json()["available"] is True
        assert sample_detail.json()["artifact_name"].startswith(
            "zte-adapter-sample-ZTE-SW-01-20260721_"
        )
        sample_download = client.get(
            "/api/rail-transit/trackside-ap-business/switch-adapters/artifacts/"
            f"{sample_payload['artifact_id']}/download"
        )
        assert sample_download.status_code == 200
        assert sample_download.content == b"zip-fixture"

    schema = app.openapi()
    assert "/api/rail-transit/trackside-ap-business/export" in schema["paths"]
    assert (
        "/api/rail-transit/trackside-ap-business/artifacts/{artifact_id}/download"
        in schema["paths"]
    )
    assert "/api/rail-transit/trackside-ap-business/plan/export" in schema["paths"]
    assert "/api/rail-transit/trackside-ap-business/base/export" in schema["paths"]
    assert "/api/rail-transit/trackside-ap-business/base/rename-commands/export" in schema["paths"]
    assert "/api/rail-transit/trackside-ap-business/base/rename-commands/artifacts/{artifact_id}/download" in schema["paths"]
    assert "/api/rail-transit/trackside-ap-business/switch-adapters" in schema["paths"]
    assert (
        "/api/rail-transit/trackside-ap-business/switch-adapters/sample"
        in schema["paths"]
    )


def test_trackside_business_workbook_preserves_sheets_and_export_style(
    tmp_path: Path,
) -> None:
    i18n = I18n("zh_CN")
    output = tmp_path / "trackside.xlsx"
    rows = [
        {
            "site": "站点A",
            "device_name": "SW-A",
            "interface_name": "GigabitEthernet2/0/1",
            "link_status": "UP",
            "switch_rx_power": -10.5,
            "switch_optical_status": "normal",
            "ap_name": "AP-A",
            "ap_mac": "0011-2233-4455",
            "ap_rx_power": -11.2,
            "ap_optical_status": "normal",
            "ap_side_has_data": True,
        }
    ]
    export_trackside_ap_business_xlsx(
        output,
        rows,
        TRACKSIDE_AP_BUSINESS_EXPORT_COLUMNS,
        [i18n.t(key) for key, _field in TRACKSIDE_AP_BUSINESS_EXPORT_COLUMNS],
        [
            {
                "site": "站点A",
                "total": 2,
                "online": 1,
                "offline": 1,
                "online_rate": "50.0%",
                "remark": "",
            },
            {
                "site": "合计",
                "total": 2,
                "online": 1,
                "offline": 1,
                "online_rate": "50.0%",
                "remark": "",
            },
        ],
        AP_ONLINE_OVERVIEW_COLUMNS,
        [i18n.t(key) for key, _field in AP_ONLINE_OVERVIEW_COLUMNS],
        [],
        NEW_ONLINE_AP_OVERVIEW_COLUMNS,
        [i18n.t(key) for key, _field in NEW_ONLINE_AP_OVERVIEW_COLUMNS],
        "新增上线AP概览",
        [
            {
                "site": "站点A",
                "device_name": "SW-A",
                "interface_name": "Ten-GigabitEthernet1/0/1",
            }
        ],
        AP_OPTICAL_TREATMENT_RECORD_COLUMNS,
        [i18n.t(key) for key, _field in AP_OPTICAL_TREATMENT_RECORD_COLUMNS],
        "AP光衰处理记录",
        {},
        [
            {
                "ap_name": "AP-B",
                "historical_switch_interface": "Twenty-FiveGigE1/0/2",
            }
        ],
        offline_ap_headers(OFFLINE_AP_STATS_COLUMNS),
        offline_ap_headers(OFFLINE_AP_LEDGER_COLUMNS),
        snapshot_generated_at="2026-08-07T01:30:00Z",
        export_updated_at="2026-08-07T12:15:38+00:00",
    )
    attach_export_metadata(
        output,
        effective_suffix=".xlsx",
        export_type="trackside_ap_business",
        payload={"source_module": "rail.trackside_ap_business"},
    )

    workbook = load_workbook(output)
    business_sheets = [
        "AP上线情况概览",
        "轨旁AP业务",
        "当前异常光衰",
        "AP光衰处理记录",
        "AP离线情况",
        "AP离线台账",
        "新增上线AP概览",
        "待关联在线AP",
        "交换机光模块统计",
    ]
    assert workbook.sheetnames == [*business_sheets, "_netconsole_meta"]
    for name in business_sheets:
        sheet = workbook[name]
        assert sheet.freeze_panes == (None if name == "AP上线情况概览" else "A2")
        assert sheet.auto_filter.ref == (
            None if name == "AP上线情况概览" else sheet.dimensions
        )
        header_row = 3 if name == "AP上线情况概览" else 1
        assert all(
            cell.alignment.horizontal == "center"
            and cell.alignment.vertical == "center"
            and cell.alignment.wrap_text is True
            for cell in sheet[header_row]
        )
        assert all(
            dimension.width is not None and dimension.width > 0
            for dimension in sheet.column_dimensions.values()
        )

    main = workbook["轨旁AP业务"]
    main_headers = [cell.value for cell in main[1]]
    interface_column = main_headers.index("接口名称") + 1
    assert main.cell(2, interface_column).value == "GE2/0/1"
    reason_column = main_headers.index("AP业务判定原因") + 1
    assert main.cell(2, reason_column).alignment.horizontal == "left"
    assert main.cell(2, reason_column).alignment.vertical == "center"
    assert main.cell(2, reason_column).alignment.wrap_text is True
    assert main.cell(2, 1).alignment.horizontal == "center"
    assert main.row_dimensions[1].height == 24
    assert main.row_dimensions[2].height == 24
    assert main["A1"].fill.fgColor.rgb == "FFDBEAFE"
    assert main["A2"].fill.fgColor.rgb == "FFDCFCE7"
    assert main["A1"].border.left.color.rgb == "FFD1D5DB"

    current_abnormal = workbook["当前异常光衰"]
    assert current_abnormal.max_row == 2
    assert str(current_abnormal["A2"].value).startswith("当前无异常光衰")
    assert current_abnormal["A2"].fill.fgColor.rgb != "FFFEE2E2"

    treatment = workbook["AP光衰处理记录"]
    treatment_headers = [cell.value for cell in treatment[1]]
    treatment_interface_column = treatment_headers.index("接口名称") + 1
    assert treatment.cell(2, treatment_interface_column).value == "XGE1/0/1"

    overview = workbook["AP上线情况概览"]
    assert overview["A1"].value == "日期：2026-08-07"
    assert overview["A2"].value == "更新时间：2026-08-07 20:15:38"
    assert [cell.value for cell in overview[3]] == [
        i18n.t(key) for key, _field in AP_ONLINE_OVERVIEW_COLUMNS
    ]
    assert overview["E4"].value == 0.5
    assert overview["E4"].number_format == "0.0%"
    assert overview["E5"].value == 0.5
    assert overview["E5"].number_format == "0.0%"
    assert all(cell.value is None for cell in overview[overview.max_row])
    assert overview.row_dimensions[overview.max_row].height == 16
    assert all(
        all(
            side is None or side.style is None
            for side in (
                cell.border.left,
                cell.border.right,
                cell.border.top,
                cell.border.bottom,
            )
        )
        for cell in overview[overview.max_row]
    )

    colors = {
        definition.sheet_name: definition.tab_color
        for definition in TRACKSIDE_AP_BUSINESS_SHEET_DEFINITIONS
        if definition.tab_color
    }
    assert colors == {
        "AP上线情况概览": "#C6EFCE",
        "轨旁AP业务": "#C6EFCE",
        "当前异常光衰": "#FFEB9C",
        "AP光衰处理记录": "#DDEBF7",
        "AP离线情况": "#D9D9D9",
        "AP离线台账": "#D9D9D9",
    }
    for name, color in colors.items():
        assert workbook[name].sheet_properties.tabColor.rgb == f"FF{color[1:]}"
    for name in ("新增上线AP概览", "待关联在线AP", "交换机光模块统计"):
        assert workbook[name].sheet_properties.tabColor is None

    ledger = workbook["AP离线台账"]
    ledger_headers = [cell.value for cell in ledger[1]]
    historical_interface_column = ledger_headers.index("历史邻居接口") + 1
    assert ledger.cell(2, historical_interface_column).value == "25GE1/0/2"


def test_trackside_business_sheet_registry_keeps_overview_identities_distinct() -> None:
    definitions = {
        item.sheet_name: item for item in TRACKSIDE_AP_BUSINESS_SHEET_DEFINITIONS
    }

    history = definitions["AP上线情况概览"]
    new_online = definitions["新增上线AP概览"]
    assert history.stable_key == "ap_online_history_overview"
    assert history.sync_mode == "PREPEND_SNAPSHOT"
    assert history.freeze_mode == "NONE"
    assert new_online.stable_key == "newly_online_ap_overview"
    assert new_online.sync_mode == "FULL_REPLACE"
    assert new_online.freeze_mode == "FIRST_ROW_ONLY"
    assert all(
        item.freeze_mode == "FIRST_ROW_ONLY"
        for item in TRACKSIDE_AP_BUSINESS_SHEET_DEFINITIONS
        if item.stable_key != "ap_online_history_overview"
    )
    assert history.stable_key != new_online.stable_key


def test_ap_online_history_block_has_dynamic_blank_separator() -> None:
    rows = [
        {
            "site": "站点A",
            "total": 2,
            "online": 1,
            "offline": 1,
            "online_rate": "50.0%",
            "remark": "",
        },
        {
            "site": "合计",
            "total": 2,
            "online": 1,
            "offline": 1,
            "online_rate": "50.0%",
            "remark": "",
        },
    ]
    headers = ["归属站点", "规划AP总数量", "上线", "未上线", "上线率", "备注"]

    block = build_ap_online_history_block(
        rows,
        AP_ONLINE_OVERVIEW_COLUMNS,
        headers,
        snapshot_generated_at="2026-08-07T01:30:00Z",
        updated_at="2026-08-07T12:15:38Z",
    )

    assert block.cells() == [
        ["日期：2026-08-07", None, None, None, None, None],
        ["更新时间：2026-08-07 20:15:38", None, None, None, None, None],
        headers,
        ["站点A", "2", "1", "1", 0.5, "-"],
        ["合计", "2", "1", "1", 0.5, "-"],
        [None, None, None, None, None, None],
    ]


@pytest.mark.parametrize(
    ("source", "expected"),
    [
        ("GigabitEthernet2/0/1", "GE2/0/1"),
        ("Ten-GigabitEthernet1/0/1", "XGE1/0/1"),
        ("Twenty-FiveGigE1/0/2", "25GE1/0/2"),
        ("FortyGigE1/0/3", "40GE1/0/3"),
        ("HundredGigE1/0/4", "100GE1/0/4"),
        ("Bridge-Aggregation1", "Bridge-Aggregation1"),
    ],
)
def test_display_interface_name(source: str, expected: str) -> None:
    assert display_interface_name(source) == expected


def test_trackside_export_worker_cleans_temporary_file_on_cancel(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    from netconsole import export_worker

    output = tmp_path / "output.xlsx"
    temporary = tmp_path / "output.xlsx.task.tmp"
    temporary.write_bytes(b"partial")
    job = ExportJob(
        job_id="trackside-cancel",
        job_type="trackside_ap_business",
        site_name="demo",
        output_path=str(output),
        tmp_path=str(temporary),
        cancel_path=str(tmp_path / "cancel"),
        db_path=str(tmp_path / "site.sqlite"),
        params={
            "snapshot_path": str(tmp_path / "snapshot.json"),
            "snapshot_sha256": "a" * 64,
        },
    )

    def cancelled(**_kwargs):
        raise TracksideApExportCancelled("导出已取消")

    monkeypatch.setattr(
        export_worker,
        "export_trackside_ap_business_from_snapshot",
        cancelled,
    )

    assert export_worker.run_job(job) == 2
    assert not temporary.exists()
    assert not output.exists()


def test_trackside_export_worker_reports_missing_snapshot_code(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    from netconsole import export_worker

    events: list[dict[str, object]] = []
    monkeypatch.setattr(export_worker, "_emit", events.append)
    job = ExportJob(
        job_id="trackside-missing",
        job_type="trackside_ap_business",
        site_name="demo",
        output_path=str(tmp_path / "output.xlsx"),
        tmp_path=str(tmp_path / "output.xlsx.task.tmp"),
        params={
            "snapshot_path": str(tmp_path / "missing-snapshot.json"),
            "snapshot_sha256": "a" * 64,
            "_web_public_result": {
                "artifact_id": "artifact-1",
                "artifact_name": "轨旁AP业务.xlsx",
            },
        },
    )

    assert export_worker.run_job(job) == 1
    terminal = events[-1]
    assert terminal["type"] == "error"
    assert terminal["error_code"] == "TRACKSIDE_AP_SNAPSHOT_NOT_FOUND"
    assert terminal["result"] == {
        "error_code": "TRACKSIDE_AP_SNAPSHOT_NOT_FOUND",
        "error_message": "轨旁 AP 导出快照不存在，请重新导出。",
    }


def test_trackside_export_worker_reports_prepare_failure_after_task_creation(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    from netconsole import export_worker

    events: list[dict[str, object]] = []
    monkeypatch.setattr(export_worker, "_emit", events.append)

    def fail_prepare(**_kwargs):
        raise TracksideApBusinessSnapshotError(
            "TRACKSIDE_AP_SNAPSHOT_STALE",
            "轨旁 AP 数据已更新，请刷新后重新导出。",
        )

    monkeypatch.setattr(
        export_worker,
        "export_trackside_ap_business_prepare_and_render",
        fail_prepare,
    )
    job = ExportJob(
        job_id="trackside-prepare-failed",
        job_type="trackside_ap_business",
        site_name="宁波地铁12号线",
        output_path=str(tmp_path / "output.xlsx"),
        tmp_path=str(tmp_path / "output.xlsx.task.tmp"),
        db_path=str(tmp_path / "site.sqlite"),
        params={
            "language": "zh_CN",
            "snapshot_staging_root": str(tmp_path / "staging"),
            "_web_public_result": {
                "artifact_id": "artifact-prepare-failed",
                "artifact_name": "轨旁AP业务.xlsx",
            },
        },
    )

    assert export_worker.run_job(job) == 1
    terminal = events[-1]
    assert terminal["type"] == "error"
    assert terminal["error_code"] == "TRACKSIDE_AP_SNAPSHOT_STALE"
    assert terminal["result"] == {
        "error_code": "TRACKSIDE_AP_SNAPSHOT_STALE",
        "error_message": "轨旁 AP 数据已更新，请刷新后重新导出。",
    }
