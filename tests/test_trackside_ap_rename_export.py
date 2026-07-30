from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from netconsole.services.ap_extension_import import _field_for_header
from netconsole.services.trackside_ap_base_export import (
    _export_row,
    export_trackside_ap_base_xlsx_task,
)
from netconsole.services.trackside_ap_rename_export import (
    TracksideApRenameConflictError,
    build_trackside_ap_rename_commands,
    build_trackside_ap_rename_export_name,
    export_trackside_ap_rename_commands_task,
    h3c_ap_mac,
)


def _row(*, mac: str, point_code: str, name: str = "AC-真实名称", **extra: object) -> dict[str, object]:
    return {"mac": mac, "point_code": point_code, "name": name, **extra}


def test_normalizes_common_mac_formats_to_h3c_command_format() -> None:
    assert h3c_ap_mac("28:c9:7a:3e:5d:a0") == "28c9-7a3e-5da0"
    assert h3c_ap_mac("28-c9-7a-3e-5d-a0") == "28c9-7a3e-5da0"
    assert h3c_ap_mac("28c97a3e5da0") == "28c9-7a3e-5da0"
    assert h3c_ap_mac("not-a-mac") == ""


def test_uses_point_code_and_records_skipped_rows() -> None:
    result = build_trackside_ap_rename_commands(
        [
            _row(mac="28:c9:7a:3e:5d:a0", point_code="ap1", name="AP0127"),
            _row(mac="", point_code="ap2"),
            _row(mac="bad", point_code="ap3"),
            _row(mac="40:f3:4d:c6:50:60", point_code=""),
            _row(mac="40:f3:4d:c6:50:61", point_code="AP0129", name="AP0129"),
            _row(mac="40:f3:4d:c6:50:62", point_code="ap4; delete"),
        ]
    )

    assert result["commands"] == ["wlan rename-ap 28c9-7a3e-5da0 ap1"]
    assert result["valid_command_count"] == 1
    assert result["skipped_count"] == 5
    assert any("名称已一致" in warning for warning in result["warnings"])
    assert any("不允许" in warning for warning in result["warnings"])


def test_conflicting_mac_or_target_blocks_export() -> None:
    with pytest.raises(TracksideApRenameConflictError, match="MAC"):
        build_trackside_ap_rename_commands(
            [
                _row(mac="28c97a3e5da0", point_code="ap1"),
                _row(mac="28c97a3e5da0", point_code="ap2"),
            ]
        )

    with pytest.raises(TracksideApRenameConflictError, match="点位编号"):
        build_trackside_ap_rename_commands(
            [
                _row(mac="28c97a3e5da0", point_code="ap1"),
                _row(mac="40f34dc65060", point_code="AP1"),
            ]
        )


def test_export_writes_utf8_bom_crlf_and_stable_name(tmp_path: Path) -> None:
    output = tmp_path / "rename.txt"
    result = export_trackside_ap_rename_commands_task(
        output,
        {
            "draft_rows": [
                _row(mac="40:f3:4d:c6:50:60", point_code="ap2", station="乙站"),
                _row(mac="28:c9:7a:3e:5d:a0", point_code="ap1", station="甲站"),
            ],
            "site_display_name": "宁波地铁1号线",
            "generated_at": "2026-07-24 23:30:00",
        },
    )

    content = output.read_bytes()
    assert content.startswith(b"\xef\xbb\xbf")
    assert b"\r\n" in content
    assert b"wlan rename-ap 28c9-7a3e-5da0 ap1" in content
    assert result["row_count"] == 2
    assert build_trackside_ap_rename_export_name("宁波地铁1号线", datetime(2026, 7, 24, 23, 30)) == (
        "轨旁AP重命名命令_宁波地铁1号线_20260724_233000.txt"
    )


def test_base_export_separates_runtime_name_from_point_code() -> None:
    row = _export_row(
        {
            "name": "旧点表名称",
            "point_code": "AP1",
            "line_side": "右线",
            "runtime": {"fit_ap_name": "AP0127"},
        }
    )

    assert row["ap_name"] == "AP0127"
    assert row["ap_point_code"] == "AP1"
    assert row["line_side"] == "右线"


def test_base_export_uses_point_code_as_empty_ap_name_fallback() -> None:
    row = _export_row({"name": "", "point_code": "AP0127"})

    assert row["ap_name"] == "AP0127"
    assert row["ap_point_code"] == "AP0127"


def test_import_issue_export_writes_structured_problem_details(tmp_path: Path) -> None:
    output = tmp_path / "issues.xlsx"
    count = export_trackside_ap_base_xlsx_task(
        output,
        {
            "issue_rows": [
                {
                    "row_number": 12,
                    "result": "CONFLICT",
                    "severity": "error",
                    "code": "identity_conflict",
                    "field_name": "ap_mac_display",
                    "original_value": "0011-2233-4455",
                    "message": "同一 MAC 对应不同点位编号",
                    "suggested_action": "核对点位编号",
                    "ap_name": "",
                    "point_code": "AP0127",
                    "ap_mac": "0011-2233-4455",
                }
            ]
        },
    )

    workbook = load_workbook(output, read_only=True)
    sheet = workbook["问题明细"]
    assert count == 1
    assert [cell.value for cell in next(sheet.iter_rows())][:4] == [
        "源行号",
        "分类",
        "级别",
        "问题代码",
    ]
    assert [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))][:4] == [
        12,
        "CONFLICT",
        "error",
        "identity_conflict",
    ]


def test_ap_name_number_header_maps_to_point_code_only() -> None:
    assert _field_for_header("AP名称编号") == "ap_point_code"
    assert _field_for_header("AP名称") == "ap_name"
