import json
from pathlib import Path
from types import SimpleNamespace

import pytest

from netconsole.core.bootstrap import create_demo_context
from netconsole.core.database import Database
from netconsole.core.feature_flags import FeatureGate
from netconsole.core.i18n import I18n
from netconsole.core.paths import PathResolver
from netconsole.core.sources.ap_source import compute_ap_status
from netconsole.core.sources.switch_source import build_switch_data_lookup
from netconsole.core.state_engine import compute_state, STATUS_COLORS
from netconsole.models.device import Device
from netconsole.parsers.h3c.ac.fit_ap_lldp_neighbor_parser import (
    parse_fit_ap_lldp_neighbor,
)
from netconsole.parsers.h3c.ac.fit_ap_optical_parser import (
    parse_fit_ap_lldp,
    parse_fit_ap_optical,
    parse_fit_ap_transceiver,
    parse_fit_ap_transceiver_diagnosis_snapshots,
)
from netconsole.parsers.h3c.ac.state_mapper import (
    classify_fit_ap_state,
    map_fit_ap_state,
)
from netconsole.parsers.h3c.ac.system_usage_parser import parse_cpu_usage, parse_memory
from netconsole.parsers.h3c.ac.wlan_ap_address_parser import parse_wlan_ap_addresses
from netconsole.parsers.h3c.ac.wlan_ap_connection_record_parser import (
    parse_wlan_ap_connection_record_rows,
    parse_wlan_ap_connection_records,
)
from netconsole.parsers.h3c.ac.wlan_ap_parser import (
    parse_wlan_ap_list,
    parse_wlan_ap_summary,
)
from netconsole.parsers.h3c.ac.wlan_ap_radio_parser import parse_wlan_ap_radios
from netconsole.parsers.h3c.ac.wlan_ap_radio_type_parser import (
    parse_wlan_ap_radio_types,
)
from netconsole.repositories.ac_repository import AcRepository, TRACKSIDE_AP_PLAN_MODE
from netconsole.repositories.device_fact_repository import DeviceFactRepository
from netconsole.repositories.device_group_repository import DeviceGroupRepository
from netconsole.repositories.device_repository import DeviceRepository
from netconsole.services.fit_ap_import_export import (
    AP_EXTENSION_TEMPLATE_FIELDS,
    FitApImportExportService,
)
from netconsole.services.fit_ap_link_info import (
    format_h3c_mac,
    merge_lldp_payload,
    normalize_interface_key,
    normalize_mac as normalize_link_mac,
    normalize_lldp_payload,
    resolve_fit_ap_link_info,
    resolve_optical_match_status,
)
from netconsole.services import h3c_ac_collect_service
from netconsole.services import command_guard
from netconsole.services.device_web_service import build_https_url, parse_https_port
from netconsole.services.ap_online_overview import (
    build_ap_online_overview_rows,
    export_ap_online_overview_xlsx,
)
from netconsole.services.ac.ac_optical_service import enrich_fit_ap_optical_rows
from netconsole.parsers.h3c.ac.wlan_ap_lldp_parser import parse_wlan_ap_lldp
from netconsole.parsers.h3c.ac.wlan_ap_radio_verbose_parser import (
    parse_wlan_ap_radio_verbose_bbssid,
)
from netconsole.services.h3c_ac_collect_service import (
    FIT_AP_RESOURCE_COMMANDS,
    FIT_AP_RESOURCE_OPTIONAL_COMMANDS,
    RESOURCE_COMMANDS,
    collect_h3c_ac_info,
    collect_h3c_fit_ap_resources,
)
from netconsole.services.h3c_ac_collect_service import FitApOpticalCollectResult
from netconsole.services.h3c_ac_collect_service import _select_fit_ap_neighbor_match
from netconsole.services.rail_transit import trackside_optical_collection
from netconsole.services.rail_transit.trackside_optical_collection import (
    DEFAULT_TRACKSIDE_OPTICAL_CONCURRENCY,
    TRACKSIDE_OPTICAL_COMMANDS,
    OpticalCommandAdapter,
    UnsupportedVendor,
    build_station_switch_targets,
    build_trackside_ap_targets,
    collect_trackside_optical,
    dedupe_targets,
)
from netconsole.services.netmiko_connection import normalize_command_output
from netconsole.services.neighbor_matcher import (
    NeighborMatchResult,
    find_neighbor_optical_module,
    find_neighbor_rx_power,
    match_ap_from_device_lldp,
    match_neighbor_device,
    normalize_interface_name,
)
from netconsole.services.trackside_ap_business import (
    AP_OPTICAL_TREATMENT_RECORD_COLUMNS,
    CURRENT_OPTICAL_ABNORMAL_COLUMNS,
    NEW_ONLINE_AP_OVERVIEW_COLUMNS,
    TRACKSIDE_AP_BUSINESS_COLUMNS,
    TRACKSIDE_AP_BUSINESS_EXPORT_COLUMNS,
    TREATMENT_CLOSED_LABEL,
    TREATMENT_OPEN_LABEL,
    build_ap_optical_treatment_records,
    build_new_online_ap_overview_rows,
    build_trackside_ap_business_rows,
    count_current_optical_abnormal_aps,
    count_current_optical_abnormal_by_site,
    description_contains_ap,
    enrich_trackside_export_rows,
    export_trackside_ap_business_xlsx,
    filter_trackside_ap_business_rows,
    format_ap_side_alarm,
    format_trackside_display_value,
    has_ap_side_optical_data,
    is_current_optical_abnormal_row,
    is_trackside_ap_interface,
    normalize_link_state,
    normalize_trackside_ap_business_row,
    normalize_vlan_text,
    parse_vlan_set,
    sort_trackside_ap_business_rows,
    _optical_status_from_history,
    normalize_interface_name as normalize_trackside_interface_name,
    normalize_mac,
    pvid_matches_trackside_plan,
    trackside_row_status,
)
from netconsole.parsers.h3c.ac.wlan_ap_unauthenticated_parser import (
    WLAN_AP_UNAUTHENTICATED_SOURCE,
)
from netconsole.core.optical_severity_engine import display_optical_status


FIXTURES = Path(__file__).parent / "fixtures" / "h3c"
AC_FIXTURES = FIXTURES / "ac"


def fixture(name: str) -> str:
    return (FIXTURES / name).read_text(encoding="utf-8")


def ac_fixture(name: str) -> str:
    return (AC_FIXTURES / name).read_text(encoding="utf-8")


class FakeConnection:
    def __init__(self, outputs=None):
        self.commands = []
        self.calls = []
        self.disconnected = False
        self.outputs = outputs or {}
        self.output_history = {}

    def send_command(self, command, read_timeout=None):
        self.commands.append(command)
        self.calls.append({"command": command, "read_timeout": read_timeout})
        if command in self.outputs:
            value = self.outputs[command]
            if isinstance(value, Exception):
                raise value
            self.output_history[command] = value
            return value
        value = {
            "screen-length disable": "",
            "display wlan ap all": fixture("display_wlan_ap_all.txt"),
            "display wlan ap all address": fixture("display_wlan_ap_all_address.txt"),
            "display wlan ap all radio": fixture("display_wlan_ap_all_radio.txt"),
            "display wlan ap unauthenticated": "Total number of connected auto APs: 0\n\nAP information:\nAP name APID State Model Serial ID Dev-Type Work-mode\n",
            "display wlan ap all radio verbose filter bbssid": "AP name              RID bbssid\nAP1                  1   0011-2233-4455\n",
            "display wlan ap all lldp": "AP name                        Local Interface          Neighbor Name                  Neighbor MAC    Neighbor Interface\nAP1                            GE1/0/2                  N/A                            903f-8645-6e00  GigabitEthernet2/0/19\n",
            "display cpu-usage": fixture("display_cpu_usage.txt"),
            "display memory": fixture("display_memory.txt"),
            "display ip https | include port": "HTTPS port: 443\n",
            "display ip https": "HTTPS port: 443\n",
            "display version": fixture("display_version.txt"),
            "display device": fixture("display_device_ac.txt"),
            "display device manuinfo": fixture("display_device_manuinfo.txt"),
        }[command]
        self.output_history[command] = value
        return value

    def disconnect(self):
        self.disconnected = True


class FakeTimingConnection:
    def __init__(self, outputs=None):
        self.commands = []
        self.calls = []
        self.disconnected = False
        self.outputs = outputs or {}

    def send_command_timing(self, command, **kwargs):
        self.commands.append(command)
        self.calls.append(
            {
                "command": command,
                "read_timeout": kwargs.get("read_timeout"),
                "strip_prompt": kwargs.get("strip_prompt"),
                "strip_command": kwargs.get("strip_command"),
            }
        )
        if command in self.outputs:
            value = self.outputs[command]
            if isinstance(value, Exception):
                raise value
            return value
        return f"<AC>{command}\nCommand {command} Result=Success\n<AC>"

    def disconnect(self):
        self.disconnected = True


class FakeOpticalConnection:
    instances = []

    def __init__(self, **kwargs):
        self.host = kwargs.get("host")
        self.commands = []
        self.disconnected = False
        FakeOpticalConnection.instances.append(self)

    def send_command(self, command, **_kwargs):
        self.commands.append(command)
        if self.host == "10.0.0.99":
            raise RuntimeError("boom")
        if command == "screen-length disable":
            return ""
        if command == "display interface brief":
            return """
Brief information on interfaces in bridge mode:
Link: ADM - administratively down; Stby - standby
Interface            Link Speed   Duplex Type PVID Description
GE1/0/1              UP   1G      F(a)   A    921  To AP
"""
        if command == "display interface":
            return """
GigabitEthernet1/0/1 current state: UP
Line protocol current state: UP
Description: To AP
PVID: 921
Port link-type: access
Untagged VLANs: 921
"""
        if command == "display lldp neighbor-information list":
            return """
Local Interface    Chassis ID          Port ID             System Name
GE1/0/1            bc5a-3457-cbe0      GigabitEthernet1/0/2 AP-1
"""
        if command == "display transceiver diagnosis interface":
            return """
GigabitEthernet1/0/1 transceiver diagnostic information:
  Current diagnostic parameters:
    Temp.(C) Voltage(V) Bias(mA) RX power(dBm) TX power(dBm)
    32.1 3.31 6.4 -6.10 -3.20
  Alarm thresholds:
    Low  -10.00 2.90 1.00 -20.00 -8.00
    High 80.00 3.70 12.00 0.00 0.00
  Warning thresholds:
    Low  -5.00 3.00 2.00 -15.00 -6.00
    High 70.00 3.60 10.00 -1.00 -1.00
"""
        return ""

    def disconnect(self):
        self.disconnected = True


def make_database(tmp_path):
    database = Database(tmp_path / "devices.db")
    database.initialize()
    return database


def make_feature_gate(root: Path, *, hidden: tuple[str, ...]) -> FeatureGate:
    runtime = root / "runtime"
    runtime.mkdir(parents=True)
    (runtime / "build_info.json").write_text(
        json.dumps({"edition": "customer", "feature_profile": "customer"}),
        encoding="utf-8",
    )
    features = {
        feature_id: {"visible": False, "enabled": False} for feature_id in hidden
    }
    (runtime / "feature_flags.json").write_text(
        json.dumps(
            {"schema_version": 1, "profile": "customer", "features": features},
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    return FeatureGate(root)


def create_station_switch(
    repository: DeviceRepository, site_name: str, **kwargs
) -> Device:
    group_repository = DeviceGroupRepository(repository.database, site_name)
    station_group = group_repository.find_by_name(
        "\u8f66\u7ad9"
    ) or group_repository.create("\u8f66\u7ad9")
    payload = {
        "group_id": station_group.id,
        "device_type": "SW",
        "device_vendor": "H3C",
        **kwargs,
    }
    return repository.create(Device(**payload))


def make_ac_device():
    return Device(
        device_uuid="22222222-2222-4222-8222-222222222222",
        name="AC",
        device_type="AC",
        ip_address="10.0.0.51",
        ssh_enabled=1,
        ssh_username="admin",
        ssh_password="Admin@123",
    )


def test_ac_tables_are_created(tmp_path):
    database = make_database(tmp_path)

    with database.connect() as conn:
        table_names = {
            row["name"]
            for row in conn.execute(
                "SELECT name FROM sqlite_master WHERE type = 'table'"
            ).fetchall()
        }
        optical_history_columns = [
            row["name"]
            for row in conn.execute(
                "PRAGMA table_info(ac_fit_ap_optical_history)"
            ).fetchall()
        ]
        resource_columns = [
            row["name"]
            for row in conn.execute("PRAGMA table_info(ac_fit_ap_resources)").fetchall()
        ]
        radio_history_columns = [
            row["name"]
            for row in conn.execute(
                "PRAGMA table_info(ac_fit_ap_radio_history)"
            ).fetchall()
        ]

    assert "ac_ap_summary" in table_names
    assert "ac_fit_ap_resources" in table_names
    assert "ac_fit_ap_optical" in table_names
    assert "ac_fit_ap_metadata" in table_names
    assert "ac_fit_ap_resource_history" in table_names
    assert "ac_station_ap_capacity" in table_names
    assert "ac_trackside_ap_plan" in table_names
    assert "ac_trackside_ap_plan_settings" in table_names
    assert "ac_fit_ap_optical_history" in table_names
    assert "ac_fit_ap_radio_history" in table_names
    assert "ac_fit_ap_lldp_history" in table_names
    for column in (
        "voltage",
        "bias_current",
        "rx_low_alarm",
        "tx_high_warning",
        "module_vendor",
        "wavelength",
        "transmission_distance",
        "connector_type",
    ):
        assert column in optical_history_columns
    for column in (
        "connection_state",
        "connection_time",
        "rid1_status",
        "rid1_mode",
        "rid1_band",
        "rid1_usage",
        "rid1_clients",
    ):
        assert column in resource_columns
    for column in ("status", "mode", "band", "usage", "clients"):
        assert column in radio_history_columns


def test_ac_repository_summary_upsert_and_replace_lists(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    repository.upsert_ac_ap_summary({"ac_device_uuid": "ac-1", "total_aps": 1})
    repository.upsert_ac_ap_summary({"ac_device_uuid": "ac-1", "total_aps": 2})

    assert repository.get_ac_ap_summary("ac-1")["total_aps"] == 2

    repository.replace_fit_ap_resources(
        "ac-1", [{"ap_name": "ap-a"}, {"ap_name": "ap-b"}]
    )
    assert [row["ap_name"] for row in repository.list_fit_ap_resources("ac-1")] == [
        "ap-a",
        "ap-b",
    ]
    repository.replace_fit_ap_resources("ac-1", [{"ap_name": "ap-c"}])
    assert [row["ap_name"] for row in repository.list_fit_ap_resources("ac-1")] == [
        "ap-c"
    ]

    repository.replace_fit_ap_optical(
        "ac-1",
        [
            {
                "ap_name": "ap-c",
                "status": "success",
                "neighbor_interface": "GigabitEthernet1/0/1",
                "rx_power": "-7.55",
            }
        ],
    )
    assert (
        repository.list_fit_ap_optical("ac-1")[0]["neighbor_interface"]
        == "GigabitEthernet1/0/1"
    )
    # 名称只用于展示/兼容查询，缺少 MAC/UUID 时不得把光衰行关联到资源。
    assert repository.get_fit_ap_optical_by_ap("ac-1", "ap-c") is None
    assert repository.get_fit_ap_resource("ac-1", "ap-c")["ap_name"] == "ap-c"


def test_fit_ap_resource_refresh_persists_radio_and_connection_fields_without_erasing_bssid(
    tmp_path,
):
    repository = AcRepository(make_database(tmp_path))
    repository.replace_fit_ap_resources(
        "ac-1",
        [
            {
                "ap_name": "ap-a",
                "serial_number": "SN-001",
                "connection_ip": "2001::3",
                "connection_state": "Run",
                "connection_time": "05-06 09:47:44",
                "rid1_status": "Up",
                "rid1_mode": "802.11n",
                "rid1_band": "5GHz",
                "rid1_channel": "149",
                "rid1_bandwidth": "40",
                "rid1_usage": "27",
                "rid1_tx_power": "24",
                "rid1_clients": 3,
                "rid1_bbssid": "0011-2233-4455",
            }
        ],
    )
    repository.replace_fit_ap_resources(
        "ac-1", [{"ap_name": "ap-a", "serial_number": "SN-001", "rid1_channel": "153"}]
    )

    row = repository.get_fit_ap_resource("ac-1", "ap-a")
    assert row["connection_state"] == "Run"
    assert row["rid1_status"] == "Up"
    assert row["rid1_usage"] == "27"
    assert row["rid1_clients"] == 3
    assert row["rid1_channel"] == "153"
    assert row["rid1_bbssid"] == "0011-2233-4455"
    history = repository.list_fit_ap_radio_history_by_ap(str(row["ap_uuid"]))
    assert history[0]["status"] == "Up"
    assert history[0]["clients"] == 3


def test_single_fit_ap_upsert_keeps_other_resources(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    repository.replace_fit_ap_resources(
        "ac-1",
        [
            {"ap_name": "AP1", "rid1_channel": "149", "rid1_bbssid": "old-bssid"},
            {"ap_name": "AP2", "rid1_channel": "6"},
        ],
    )
    ap1_uuid = repository.get_fit_ap_resource("ac-1", "AP1")["ap_uuid"]

    repository.upsert_fit_ap_resource(
        "ac-1",
        {"ap_uuid": ap1_uuid, "ap_name": "AP1", "rid1_channel": "153", "rid1_bbssid": "new-bssid"},
    )

    rows = repository.list_fit_ap_resources("ac-1")
    assert [row["ap_name"] for row in rows] == ["AP1", "AP2"]
    assert repository.get_fit_ap_resource("ac-1", "AP1")["rid1_bbssid"] == "new-bssid"
    assert repository.get_fit_ap_resource("ac-1", "AP2")["rid1_channel"] == "6"


def test_fit_ap_resources_match_by_serial_number_and_keep_ap_uuid(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    repository.replace_fit_ap_resources(
        "ac-1",
        [
            {
                "ap_name": "old-business-name",
                "ap_ip": "10.0.0.1",
                "ap_mac": "0011-2233-4455",
                "serial_number": "SN-001",
                "state": "R/M",
            }
        ],
    )
    first = repository.list_fit_ap_resources("ac-1")[0]

    repository.replace_fit_ap_resources(
        "ac-1",
        [
            {
                "ap_name": "new-business-name",
                "ap_ip": "10.0.0.99",
                "ap_mac": "0011-2233-9999",
                "serial_number": "SN-001",
                "state": "R/B",
            }
        ],
    )
    second = repository.list_fit_ap_resources("ac-1")[0]

    assert second["ap_uuid"] == first["ap_uuid"]
    assert second["ap_name"] == "new-business-name"
    assert second["ap_ip"] == "10.0.0.99"
    assert second["ap_mac"] == "0011-2233-9999"
    assert second["state"] == "R/B"
    assert len(repository.list_fit_ap_resources("ac-1")) == 1


def test_fit_ap_resources_reuse_ap_uuid_when_apid_changes(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    repository.replace_fit_ap_resources(
        "ac-1",
        [
            {
                "ap_name": "ap-a",
                "apid": "1346",
                "ap_mac": "0011-2233-4455",
                "serial_number": "SN-001",
            }
        ],
    )
    first = repository.list_fit_ap_resources("ac-1")[0]

    repository.replace_fit_ap_resources(
        "ac-1",
        [
            {
                "ap_name": "ap-a",
                "apid": "2001",
                "ap_mac": "0011-2233-4455",
                "serial_number": "SN-001",
            }
        ],
    )
    second = repository.list_fit_ap_resources("ac-1")[0]
    entity = repository.list_ap_entities("ac-1")[0]

    assert second["ap_uuid"] == first["ap_uuid"]
    assert second["apid"] == "2001"
    assert entity["ap_uuid"] == first["ap_uuid"]
    assert entity["ap_id"] == "2001"


def test_fit_ap_resource_offline_refresh_preserves_identity_and_clears_runtime_state(
    tmp_path,
):
    repository = AcRepository(make_database(tmp_path))
    repository.replace_fit_ap_resources(
        "ac-1",
        [
            {
                "ap_name": "bc5a-3457-b5e0",
                "apid": "1346",
                "ap_mac": "bc5a-3457-b5e0",
                "serial_number": "SN-OFFLINE-1",
                "ap_ip": "10.122.0.10",
                "state": "R/M",
                "rid1_status": "Up",
                "rid1_channel": "149",
                "rid2_status": "Up",
                "rid2_channel": "153",
            }
        ],
    )
    first = repository.list_fit_ap_resources("ac-1")[0]

    repository.replace_fit_ap_resources(
        "ac-1",
        [
            {
                "ap_name": "bc5a-3457-b5e0",
                "apid": "1346",
                "ap_mac": None,
                "serial_number": None,
                "ap_ip": None,
                "state": "I",
            }
        ],
    )

    current = repository.list_fit_ap_resources("ac-1")[0]
    entities = repository.list_ap_entities("ac-1")
    assert current["ap_uuid"] == first["ap_uuid"]
    assert current["ap_mac"] == "bc5a-3457-b5e0"
    assert current["serial_number"] == "SN-OFFLINE-1"
    assert current["ap_ip"] is None
    assert current["state"] == "I"
    assert current["rid1_status"] == "Down"
    assert current["rid2_status"] == "Down"
    assert current["rid1_channel"] is None
    assert len(entities) == 1
    assert entities[0]["ap_uuid"] == first["ap_uuid"]
    assert entities[0]["ap_mac"] == "bc5a-3457-b5e0"
    assert entities[0]["ap_ip"] is None
    assert entities[0]["is_offline"] == 1


def test_fit_ap_resource_refresh_does_not_restore_identity_from_legacy_history(
    tmp_path,
):
    database = make_database(tmp_path)
    repository = AcRepository(database)
    repository.replace_fit_ap_resources(
        "ac-1",
        [
            {
                "ap_name": "AP-HISTORY",
                "apid": "88",
                "ap_mac": "0011-2233-4455",
                "serial_number": "SN-HISTORY",
            }
        ],
    )
    first = repository.list_fit_ap_resources("ac-1")[0]
    with database.connect() as connection:
        connection.execute(
            "UPDATE ac_fit_ap_resources SET ap_mac = NULL, serial_number = NULL WHERE ap_uuid = ?",
            (first["ap_uuid"],),
        )
        connection.execute(
            "UPDATE ap_entities SET ap_mac = NULL, serial_number = NULL WHERE ap_uuid = ?",
            (first["ap_uuid"],),
        )
        connection.execute(
            "UPDATE ac_fit_ap_resource_history SET ap_mac = NULL, serial_number = NULL WHERE ap_uuid = ?",
            (first["ap_uuid"],),
        )
        connection.commit()

    repository.replace_fit_ap_resources(
        "ac-1",
        [{"ap_name": "AP-HISTORY", "apid": "88", "state": "I"}],
    )
    recovered = repository.list_fit_ap_resources("ac-1")[0]

    assert recovered["ap_uuid"] == first["ap_uuid"]
    assert recovered["ap_mac"] is None
    assert recovered["serial_number"] is None


def test_fit_ap_resource_name_continuity_does_not_guess_when_existing_name_is_ambiguous(
    tmp_path,
):
    repository = AcRepository(make_database(tmp_path))
    repository.replace_fit_ap_resources(
        "ac-1",
        [
            {"ap_name": "DUPLICATE", "ap_mac": "0011-2233-4455", "serial_number": "SN-1"},
            {"ap_name": "DUPLICATE", "ap_mac": "0011-2233-5566", "serial_number": "SN-2"},
        ],
    )
    previous_uuids = {row["ap_uuid"] for row in repository.list_fit_ap_resources("ac-1")}

    repository.replace_fit_ap_resources(
        "ac-1",
        [{"ap_name": "DUPLICATE", "ap_mac": None, "serial_number": None, "state": "I"}],
    )
    current = repository.list_fit_ap_resources("ac-1")[0]

    assert current["ap_uuid"] not in previous_uuids
    assert current["ap_mac"] is None
    assert current["serial_number"] is None


def test_fit_ap_resource_does_not_derive_mac_from_mac_like_name(tmp_path):
    repository = AcRepository(make_database(tmp_path))

    repository.replace_fit_ap_resources(
        "ac-1",
        [{"ap_name": "bc5a-3457-b5e0", "apid": "99", "state": "I"}],
    )

    current = repository.list_fit_ap_resources("ac-1")[0]
    assert current["ap_mac"] is None


def test_fit_ap_resources_do_not_merge_different_identity_with_same_apid(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    repository.replace_fit_ap_resources(
        "ac-1",
        [
            {
                "ap_name": "ap-a",
                "apid": "1346",
                "ap_mac": "0011-2233-4455",
                "serial_number": "SN-001",
            },
            {
                "ap_name": "ap-b",
                "apid": "1346",
                "ap_mac": "0011-2233-5566",
                "serial_number": "SN-002",
            },
        ],
    )

    rows = repository.list_fit_ap_resources("ac-1")
    assert len(rows) == 2
    assert len({row["ap_uuid"] for row in rows}) == 2
    assert len(repository.list_ap_entities("ac-1")) == 2


def test_fit_ap_resources_same_name_different_macs_are_distinct(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    repository.replace_fit_ap_resources(
        "ac-1",
        [
            {"ap_name": "same-name", "ap_mac": "0011-2233-4455"},
            {"ap_name": "same-name", "ap_mac": "0011-2233-5566"},
        ],
    )

    rows = repository.list_fit_ap_resources("ac-1")
    assert len(rows) == 2
    assert {row["ap_mac"] for row in rows} == {"0011-2233-4455", "0011-2233-5566"}


def test_fit_ap_resources_same_name_without_identity_are_not_deduplicated(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    repository.replace_fit_ap_resources(
        "ac-1",
        [{"ap_name": "same-name"}, {"ap_name": "same-name"}],
    )

    rows = repository.list_fit_ap_resources("ac-1")
    assert len(rows) == 2
    assert len({row["ap_uuid"] for row in rows}) == 2


def test_fit_ap_resources_same_name_hardware_replacement_creates_new_identity(
    tmp_path,
):
    database = make_database(tmp_path)
    repository = AcRepository(database)
    repository.replace_fit_ap_resources(
        "ac-1",
        [
            {
                "ap_name": "ap-a",
                "apid": "1346",
                "ap_ip": "10.0.0.1",
                "ap_mac": "0011-2233-4455",
                "serial_number": "SN-OLD",
                "model": "old-model",
                "state": "I",
            }
        ],
    )
    first = repository.list_fit_ap_resources("ac-1")[0]
    with database.connect() as conn:
        conn.execute(
            """
            UPDATE ap_entities
            SET station = ?, milestone = ?, direction = ?, location_note = ?
            WHERE ap_uuid = ?
            """,
            ("Station A", "K1+100", "up", "old note", first["ap_uuid"]),
        )
        conn.commit()

    repository.replace_fit_ap_resources(
        "ac-1",
        [
            {
                "ap_name": "ap-a",
                "apid": "2001",
                "ap_ip": "10.0.0.99",
                "ap_mac": "0011-2233-9999",
                "serial_number": "SN-NEW",
                "model": "new-model",
                "state": "R/M",
            }
        ],
    )
    second = repository.list_fit_ap_resources("ac-1")[0]
    with database.connect() as conn:
        entities = {
            row["ap_uuid"]: dict(row)
            for row in conn.execute("SELECT * FROM ap_entities ORDER BY ap_uuid")
        }
    history = repository.list_fit_ap_resource_history("ac-1")

    assert second["ap_uuid"] != first["ap_uuid"]
    assert len(entities) == 2
    assert entities[second["ap_uuid"]]["ap_mac"] == "0011-2233-9999"
    assert entities[second["ap_uuid"]]["serial_number"] == "SN-NEW"
    assert entities[second["ap_uuid"]]["model"] == "new-model"
    assert entities[second["ap_uuid"]]["ap_ip"] == "10.0.0.99"
    assert entities[second["ap_uuid"]]["ap_id"] == "2001"
    assert entities[second["ap_uuid"]]["state"] == "R/M"
    assert entities[second["ap_uuid"]]["station"] in (None, "")
    assert entities[second["ap_uuid"]]["milestone"] in (None, "")
    assert entities[second["ap_uuid"]]["direction"] in (None, "")
    assert entities[second["ap_uuid"]]["location_note"] in (None, "")
    assert {row["serial_number"] for row in history} == {"SN-OLD", "SN-NEW"}


def test_fit_ap_resources_allow_empty_or_repeated_apid_without_unique_failure(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    repository.replace_fit_ap_resources(
        "ac-1",
        [
            {
                "ap_name": "ap-a",
                "apid": "",
                "ap_mac": "0011-2233-4455",
                "serial_number": "SN-001",
            },
            {
                "ap_name": "ap-b",
                "apid": "",
                "ap_mac": "0011-2233-5566",
                "serial_number": "SN-002",
            },
            {
                "ap_name": "ap-c",
                "apid": "1346",
                "ap_mac": "0011-2233-6677",
                "serial_number": "SN-003",
            },
            {
                "ap_name": "ap-d",
                "apid": "1346",
                "ap_mac": "0011-2233-7788",
                "serial_number": "SN-004",
            },
        ],
    )

    rows = repository.list_fit_ap_resources("ac-1")
    assert len(rows) == 4
    assert len({row["ap_uuid"] for row in rows}) == 4


def test_fit_ap_resources_repeated_update_is_idempotent(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    rows = [
        {
            "ap_name": "ap-a",
            "apid": "1346",
            "ap_mac": "0011-2233-4455",
            "serial_number": "SN-001",
        },
        {
            "ap_name": "ap-b",
            "apid": "1346",
            "ap_mac": "0011-2233-5566",
            "serial_number": "SN-002",
        },
    ]

    repository.replace_fit_ap_resources("ac-1", rows)
    first_uuids = [row["ap_uuid"] for row in repository.list_fit_ap_resources("ac-1")]
    repository.replace_fit_ap_resources("ac-1", rows)
    second_rows = repository.list_fit_ap_resources("ac-1")

    assert [row["ap_uuid"] for row in second_rows] == first_uuids
    assert len(second_rows) == 2
    assert len(repository.list_ap_entities("ac-1")) == 2


def test_fit_ap_optical_and_metadata_use_ap_uuid_association(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    repository.replace_fit_ap_resources(
        "ac-1", [{"ap_name": "ap-a", "serial_number": "SN-001"}]
    )
    ap_uuid = repository.list_fit_ap_resources("ac-1")[0]["ap_uuid"]

    repository.replace_fit_ap_optical(
        "ac-1", [{"ap_uuid": ap_uuid, "ap_name": "renamed-ap", "rx_power": "-7.55"}]
    )
    repository.upsert_fit_ap_metadata(
        {"ap_uuid": ap_uuid, "ap_name": "renamed-ap", "site_name": "Station A"}
    )

    assert repository.get_fit_ap_optical_by_uuid("ac-1", ap_uuid)["rx_power"] == "-7.55"
    assert repository.get_fit_ap_metadata_by_uuid(ap_uuid)["site_name"] == "Station A"


def test_fit_ap_optical_history_is_appended_and_sorted(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    repository.replace_fit_ap_resources(
        "ac-1",
        [
            {
                "ap_name": "ap-a",
                "serial_number": "SN-001",
                "ap_mac": "0011-2233-4455",
            }
        ],
    )
    ap_uuid = repository.list_fit_ap_resources("ac-1")[0]["ap_uuid"]

    repository.replace_fit_ap_optical(
        "ac-1",
        [
            {
                "ap_uuid": ap_uuid,
                "ap_name": "ap-a",
                "rx_power": "-8",
                "collected_at": "2026-01-01T00:00:00",
            }
        ],
    )
    repository.replace_fit_ap_optical(
        "ac-1",
        [
            {
                "ap_uuid": ap_uuid,
                "ap_name": "ap-a",
                "rx_power": "-7",
                "temperature": "31.2",
                "voltage": "3.31",
                "bias_current": "5.10",
                "tx_power": "-4.5",
                "rx_low_alarm": "-19.00",
                "rx_high_alarm": "-3.00",
                "tx_low_alarm": "-11.00",
                "tx_high_alarm": "-1.00",
                "rx_low_warning": "-16.99",
                "rx_high_warning": "-5.00",
                "tx_low_warning": "-9.00",
                "tx_high_warning": "-3.00",
                "module_model": "SFP-GE-LX-SM1310",
                "module_serial_number": "OPT-001",
                "module_vendor": "H3C",
                "wavelength": "1310 nm",
                "transmission_distance": "10 km",
                "connector_type": "LC",
                "collected_at": "2026-01-02T00:00:00",
            }
        ],
    )

    history = repository.list_fit_ap_optical_history_by_ap(ap_uuid)
    assert [row["rx_power"] for row in history] == ["-7", "-7"]
    assert {row["side"] for row in history} == {"ap", "switch"}
    assert history[0]["ap_mac"] == "0011-2233-4455"
    assert history[0]["voltage"] == "3.31"
    assert history[0]["bias_current"] == "5.10"
    assert history[0]["rx_low_alarm"] == "-19.00"
    assert history[0]["tx_high_warning"] == "-3.00"
    assert history[0]["module_vendor"] == "H3C"
    assert history[0]["wavelength"] == "1310 nm"
    assert history[0]["transmission_distance"] == "10 km"
    assert history[0]["connector_type"] == "LC"
    assert repository.list_fit_ap_optical("ac-1")[0]["rx_power"] == "-7"
    with repository.database.connect() as conn:
        legacy_count = conn.execute(
            "SELECT COUNT(*) FROM ac_fit_ap_optical_history WHERE ap_uuid = ?",
            (ap_uuid,),
        ).fetchone()[0]
        projection_count = conn.execute(
            "SELECT COUNT(*) FROM ap_optical_history WHERE ap_uuid = ?",
            (ap_uuid,),
        ).fetchone()[0]
    assert legacy_count == 0
    assert projection_count == 0


def test_fit_ap_optical_telemetry_jitter_does_not_create_change_event(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    repository.replace_fit_ap_resources(
        "ac-1", [{"ap_name": "ap-a", "serial_number": "SN-001"}]
    )
    ap_uuid = repository.list_fit_ap_resources("ac-1")[0]["ap_uuid"]
    repository.replace_fit_ap_optical(
        "ac-1", [{"ap_uuid": ap_uuid, "rx_power": "-10.5", "temperature": "30.0", "collected_at": "2026-01-01T00:00:00"}]
    )
    repository.replace_fit_ap_optical(
        "ac-1", [{"ap_uuid": ap_uuid, "rx_power": "-10.6", "temperature": "30.1", "collected_at": "2026-01-01T00:01:00"}]
    )
    with repository.database.connect() as conn:
        pending = conn.execute(
            "SELECT COUNT(*) FROM optical_history WHERE ap_identity = ? AND side = 'AP'",
            (ap_uuid,),
        ).fetchone()[0]
    assert pending == 1

    repository.replace_fit_ap_optical(
        "ac-1", [{"ap_uuid": ap_uuid, "rx_power": "-10.6", "optical_alarm_status": "no_light", "collected_at": "2026-01-01T00:02:00"}]
    )
    with repository.database.connect() as conn:
        pending = conn.execute(
            "SELECT COUNT(*) FROM optical_history WHERE ap_identity = ? AND side = 'AP'",
            (ap_uuid,),
        ).fetchone()[0]
    assert pending == 2


def test_fit_ap_optical_identical_refresh_is_incremental_and_change_only(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    resources = [
        {"ap_uuid": f"ap-{index}", "ap_name": f"AP-{index}", "ap_mac": f"0011-2233-{index:04x}"}
        for index in range(981)
    ]
    repository.replace_fit_ap_resources("ac-1", resources)
    rows = [
        {"ap_uuid": item["ap_uuid"], "ap_name": item["ap_name"], "rx_power": "-10.00", "tx_power": "-5.00", "status": "success"}
        for item in resources
    ]
    repository.replace_fit_ap_optical("ac-1", rows)
    with repository.database.connect() as conn:
        first_events = conn.execute("SELECT COUNT(*) FROM optical_history WHERE side = 'AP'").fetchone()[0]
    repository.replace_fit_ap_optical("ac-1", rows)
    assert len(repository.list_fit_ap_optical("ac-1")) == 981
    with repository.database.connect() as conn:
        assert conn.execute("SELECT COUNT(*) FROM ac_fit_ap_optical WHERE ac_device_uuid = 'ac-1'").fetchone()[0] == 981
        assert conn.execute("SELECT COUNT(*) FROM optical_history WHERE side = 'AP'").fetchone()[0] == first_events


def test_fit_ap_optical_history_power_tolerance_and_status_change(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    repository.replace_fit_ap_resources("ac-1", [{"ap_uuid": "ap-1", "ap_name": "AP-1"}])
    base = {"ap_uuid": "ap-1", "rx_power": "-10.00", "tx_power": "-5.00", "status": "success"}
    repository.replace_fit_ap_optical("ac-1", [base])
    repository.replace_fit_ap_optical("ac-1", [{**base, "rx_power": "-10.05"}])
    with repository.database.connect() as conn:
        assert conn.execute(
            "SELECT COUNT(*) FROM optical_history WHERE ap_identity = 'ap-1' AND side = 'AP'"
        ).fetchone()[0] == 1
    repository.replace_fit_ap_optical("ac-1", [{**base, "rx_power": "-10.25"}])
    repository.replace_fit_ap_optical("ac-1", [{**base, "rx_power": "-10.26", "optical_alarm_status": "alarm"}])
    with repository.database.connect() as conn:
        assert conn.execute(
            "SELECT COUNT(*) FROM optical_history WHERE ap_identity = 'ap-1' AND side = 'AP'"
        ).fetchone()[0] == 3


def test_fit_ap_lldp_history_is_appended_and_sorted(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    repository.replace_fit_ap_resources(
        "ac-1", [{"ap_name": "ap-a", "serial_number": "SN-001", "ap_mac": "0011"}]
    )
    ap_uuid = repository.list_fit_ap_resources("ac-1")[0]["ap_uuid"]

    repository.replace_fit_ap_optical(
        "ac-1",
        [
            {
                "ap_uuid": ap_uuid,
                "ap_name": "ap-a",
                "ap_mac": "0011",
                "interface_name": "GigabitEthernet1/0/1",
                "lldp_neighbor": "SW01",
                "neighbor_interface": "GE1/0/1",
                "neighbor_mac": "aaaa-bbbb-cccc",
                "neighbor_device_name": "HX_1",
                "collected_at": "2026-01-01T00:00:00",
            }
        ],
    )
    repository.replace_fit_ap_optical(
        "ac-1",
        [
            {
                "ap_uuid": ap_uuid,
                "ap_name": "ap-a",
                "ap_mac": "0011",
                "interface_name": "GigabitEthernet1/0/2",
                "lldp_neighbor": "SW02",
                "neighbor_interface": "GE1/0/2",
                "neighbor_mac": "dddd-eeee-ffff",
                "neighbor_device_name": "HX_2",
                "collected_at": "2026-01-02T00:00:00",
            }
        ],
    )

    history = repository.list_fit_ap_lldp_history_by_ap(ap_uuid)

    assert [row["lldp_neighbor"] for row in history] == ["SW02"]
    assert history[0]["local_interface"] == "GigabitEthernet1/0/2"
    assert history[0]["neighbor_device_name"] == "HX_2"
    current = repository.list_current_ap_lldp_states([ap_uuid])
    assert len(current) == 1
    assert current[0]["lldp_neighbor"] == "SW02"


def test_current_fit_ap_lldp_converges_repeated_relation_to_one(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    rows = [
        {
            "ap_uuid": "ap-1",
            "ap_mac": "0011-2233-4455",
            "local_interface": "GigabitEthernet1/0/1",
            "neighbor_mac": "aa11-bbcc-ddee",
            "neighbor_interface": "GE1/0/48",
            "lldp_neighbor": "SW-1",
            "collected_at": f"2026-08-01T00:{index:02d}:00",
            "id": index,
        }
        for index in range(100)
    ]
    repository.list_fit_ap_lldp_history_by_ap = lambda ap_uuid, limit=100: rows  # type: ignore[method-assign]

    current = repository.list_current_fit_ap_lldp_by_ap("ap-1")

    assert len(current) == 1
    assert current[0]["collected_at"] == "2026-08-01T00:99:00"


def test_current_fit_ap_lldp_reads_bounded_current_with_ac_scope(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    repository.replace_fit_ap_resources(
        "ac-a",
        [
            {
                "ap_uuid": "ap-shared",
                "ap_name": "AP-shared",
                "ap_mac": "0011-2233-4455",
                "lldp_source": "ac_bulk_lldp",
                "lldp_local_interface": "GigabitEthernet1/0/1",
                "lldp_neighbor_name": "SW-A",
                "lldp_neighbor_mac": "aa11-bbcc-ddee",
                "lldp_neighbor_interface": "GE1/0/1",
                "collected_at": "2026-08-01T00:00:00",
            }
        ],
    )
    repository.replace_fit_ap_resources(
        "ac-b",
        [
            {
                "ap_uuid": "ap-shared",
                "ap_name": "AP-shared",
                "ap_mac": "0011-2233-4455",
                "lldp_source": "ac_bulk_lldp",
                "lldp_local_interface": "GigabitEthernet1/0/2",
                "lldp_neighbor_name": "SW-B",
                "lldp_neighbor_mac": "ff11-2233-4455",
                "lldp_neighbor_interface": "GE1/0/2",
                "collected_at": "2026-08-01T00:01:00",
            }
        ],
    )

    current = repository.list_current_fit_ap_lldp_by_ap(
        "ap-shared", ac_device_uuid="ac-a"
    )

    assert len(current) == 1
    assert current[0]["lldp_neighbor"] == "SW-A"
    assert current[0]["local_interface"] == "GigabitEthernet1/0/1"


def test_current_fit_ap_lldp_keeps_latest_relations_after_switch_change(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    rows = [
        {
            "ap_uuid": "ap-1",
            "ap_mac": "0011-2233-4455",
            "local_interface": "GE1/0/1",
            "neighbor_mac": neighbor_mac,
            "neighbor_interface": "GE1/0/48",
            "lldp_neighbor": neighbor_name,
            "collected_at": collected_at,
            "id": index,
        }
        for index, (neighbor_mac, neighbor_name, collected_at) in enumerate(
            (
                ("aa11-bbcc-ddee", "SW-1", "2026-08-01T00:00:00"),
                ("ff11-2233-4455", "SW-2", "2026-08-01T00:01:00"),
            ),
            start=1,
        )
    ]
    repository.list_fit_ap_lldp_history_by_ap = lambda ap_uuid, limit=100: rows  # type: ignore[method-assign]

    current = repository.list_current_fit_ap_lldp_by_ap("ap-1")

    assert len(current) == 2
    assert {row["lldp_neighbor"] for row in current} == {"SW-1", "SW-2"}


def test_fit_ap_resource_lldp_merges_ap_direct_and_marks_history_changes(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    repository.replace_fit_ap_resources(
        "ac-1",
        [
            {
                "ap_name": "ap-a",
                "serial_number": "SN-001",
                "ap_mac": "0011-2233-4455",
                "lldp_source": "ac_bulk_lldp",
                "lldp_local_interface": "GE1/0/2",
                "lldp_neighbor_name": "N/A",
                "lldp_neighbor_mac": "903f-8645-6e00",
                "lldp_neighbor_interface": "GE2/0/19",
                "collected_at": "2026-01-01T00:00:00",
            }
        ],
    )
    ap_uuid = repository.list_fit_ap_resources("ac-1")[0]["ap_uuid"]

    direct_payload = {
        "ap_uuid": ap_uuid,
        "ap_name": "ap-a",
        "ap_mac": "0011-2233-4455",
        "lldp_source": "ap_direct_lldp",
        "lldp_local_interface": "GigabitEthernet1/0/2",
        "lldp_neighbor_name": "HX_1",
        "lldp_neighbor_mac": "90:3f:86:45:6e:00",
        "lldp_neighbor_interface": "GigabitEthernet2/0/19",
        "interface_name": "GigabitEthernet1/0/2",
        "rx_power": "-7.55",
        "tx_power": "-6.09",
        "collected_at": "2026-01-02T00:00:00",
    }
    repository.replace_fit_ap_optical("ac-1", [direct_payload])
    repository.replace_fit_ap_optical(
        "ac-1", [{**direct_payload, "collected_at": "2026-01-03T00:00:00"}]
    )

    resource = repository.get_fit_ap_resource_by_uuid("ac-1", ap_uuid)
    history = repository.list_fit_ap_lldp_history_by_ap(ap_uuid)

    assert resource["lldp_neighbor_name"] == "HX_1"
    assert resource["lldp_source"] == "merged"
    assert resource["lldp_match_status"] == "matched"
    assert resource["optical_interface"] == "GigabitEthernet1/0/2"
    assert resource["optical_rx_power"] == -7.55
    assert resource["optical_match_status"] == "matched"
    assert [row["source"] for row in history[:1]] == ["merged"]
    assert history[0]["change_kind"] == "change"
    with repository.database.connect() as conn:
        legacy_count = conn.execute(
            "SELECT COUNT(*) FROM ac_fit_ap_lldp_history WHERE ap_uuid = ?",
            (ap_uuid,),
        ).fetchone()[0]
        projection_count = conn.execute(
            "SELECT COUNT(*) FROM ap_lldp_history WHERE ap_uuid = ?",
            (ap_uuid,),
        ).fetchone()[0]
    assert legacy_count == 0
    assert projection_count == 0


def test_fit_ap_resource_history_uses_bounded_recent_without_legacy_writes(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    sample = {
        "ap_name": "ap-change-aware",
        "ap_mac": "0011-2233-4455",
        "serial_number": "SN-001",
        "state": "R/M",
        "collected_at": "2026-01-01T00:00:00",
    }

    for _ in range(100):
        repository.replace_fit_ap_resources("ac-1", [sample])

    ap_uuid = repository.list_fit_ap_resources("ac-1")[0]["ap_uuid"]
    history = repository.list_fit_ap_resource_history("ac-1")
    assert [row["ap_uuid"] for row in history] == [ap_uuid]
    with repository.database.connect() as conn:
        legacy_count = conn.execute(
            "SELECT COUNT(*) FROM ac_fit_ap_resource_history WHERE ap_uuid = ?",
            (ap_uuid,),
        ).fetchone()[0]
        outbox_count = conn.execute(
            "SELECT COUNT(*) FROM sqlite_master WHERE type = 'table' AND name = 'history_outbox'"
        ).fetchone()[0]
        recent_count = conn.execute(
            "SELECT COUNT(*) FROM fit_ap_resource_recent WHERE ac_device_uuid='ac-1' AND ap_uuid = ?",
            (ap_uuid,),
        ).fetchone()[0]
    assert legacy_count == 0
    assert outbox_count == 0
    assert recent_count == 1


def test_fit_ap_optical_failed_row_does_not_overwrite_valid_rx(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    repository.replace_fit_ap_resources(
        "ac-1", [{"ap_name": "ap-a", "serial_number": "SN-001"}]
    )
    ap_uuid = repository.list_fit_ap_resources("ac-1")[0]["ap_uuid"]

    repository.replace_fit_ap_optical(
        "ac-1",
        [
            {
                "ap_uuid": ap_uuid,
                "ap_name": "ap-a",
                "rx_power": "-7.34",
                "status": "success",
                "collected_at": "2026-01-01T00:00:00",
            }
        ],
    )
    repository.replace_fit_ap_optical(
        "ac-1",
        [
            {
                "ap_uuid": ap_uuid,
                "ap_name": "ap-a",
                "rx_power": "",
                "status": "timeout",
                "collected_at": "2026-01-02T00:00:00",
            }
        ],
    )

    row = repository.get_fit_ap_optical_by_uuid("ac-1", ap_uuid)
    assert row["rx_power"] == "-7.34"
    assert row["ap_name"] == "ap-a"


def test_fit_ap_optical_lldp_only_success_is_not_treated_as_failure(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    repository.replace_fit_ap_resources(
        "ac-1", [{"ap_name": "ap-a", "serial_number": "SN-001"}]
    )
    ap_uuid = repository.list_fit_ap_resources("ac-1")[0]["ap_uuid"]

    repository.replace_fit_ap_optical(
        "ac-1",
        [
            {
                "ap_uuid": ap_uuid,
                "ap_name": "ap-a",
                "status": "success",
                "neighbor_device_name": "SW01",
                "neighbor_interface": "GE1/0/1",
            }
        ],
    )

    row = repository.get_fit_ap_optical_by_uuid("ac-1", ap_uuid)
    assert row["neighbor_device_name"] == "SW01"
    assert h3c_ac_collect_service._is_fit_ap_optical_success_row(dict(row))


def test_fit_ap_optical_retry_targets_and_concurrency():
    resources = [
        {"ap_uuid": "ap-ok", "ap_name": "AP-OK"},
        {"ap_uuid": "ap-empty", "ap_name": "AP-EMPTY"},
        {"ap_uuid": "ap-missing", "ap_name": "AP-MISSING"},
    ]
    round_rows = [
        {"ap_uuid": "ap-ok", "status": "success", "rx_power": "-7.34"},
        {"ap_uuid": "ap-empty", "status": "success", "rx_power": ""},
    ]

    retry = h3c_ac_collect_service._retry_fit_ap_optical_targets(resources, round_rows)

    assert [row["ap_uuid"] for row in retry] == ["ap-empty", "ap-missing"]
    assert (
        h3c_ac_collect_service.retry_fit_ap_optical_concurrency(
            1000, floor=100, ratio=0.5
        )
        == 500
    )
    assert (
        h3c_ac_collect_service.retry_fit_ap_optical_concurrency(
            120, floor=100, ratio=0.5
        )
        == 100
    )
    assert (
        h3c_ac_collect_service.retry_fit_ap_optical_concurrency(
            64, floor=100, ratio=0.5
        )
        == 32
    )


def test_fit_ap_optical_final_rows_count_each_ap_once_after_retry():
    rows = [
        {
            "ap_uuid": "ap-retried",
            "status": "failed",
            "error_message": "connect_timeout: first round",
            "collected_at": "2026-01-01T00:00:00",
        },
        {
            "ap_uuid": "ap-stable",
            "status": "success",
            "rx_power": "-8.10",
            "collected_at": "2026-01-01T00:00:00",
        },
        {
            "ap_uuid": "ap-retried",
            "status": "success",
            "rx_power": "-7.20",
            "collected_at": "2026-01-01T00:00:01",
        },
    ]

    final_rows = h3c_ac_collect_service._final_fit_ap_optical_rows(rows)

    assert len(final_rows) == 2
    by_uuid = {str(row["ap_uuid"]): row for row in final_rows}
    assert by_uuid["ap-retried"]["status"] == "success"
    assert by_uuid["ap-retried"]["rx_power"] == "-7.20"
    assert by_uuid["ap-stable"]["status"] == "success"


def test_fit_ap_optical_round_isolates_single_future_failure(monkeypatch, tmp_path):
    ac_device = make_ac_device()
    ac_device.device_uuid = "ac-1"
    resources = [
        {"ap_uuid": "ap-ok", "ap_name": "AP-OK", "ap_ip": "10.0.0.10"},
        {"ap_uuid": "ap-fail", "ap_name": "AP-FAIL", "ap_ip": "10.0.0.11"},
    ]

    def collect_one(_ac_device, ap_row, *_args, **_kwargs):
        if ap_row["ap_uuid"] == "ap-fail":
            raise OSError(36, "Resource deadlock avoided")
        return {
            "ac_device_uuid": "ac-1",
            "ap_uuid": ap_row["ap_uuid"],
            "ap_name": ap_row["ap_name"],
            "status": "success",
            "rx_power": "-7.10",
        }

    monkeypatch.setattr(h3c_ac_collect_service, "_collect_single_fit_ap_optical", collect_one)
    progress: list[tuple[int, int]] = []

    rows = h3c_ac_collect_service._collect_fit_ap_optical_round(
        ac_device,
        resources,
        "demo",
        "run-1",
        tmp_path,
        PathResolver(tmp_path),
        64,
        lambda: False,
        lambda current, total: progress.append((current, total)),
    )

    by_uuid = {row["ap_uuid"]: row for row in rows}
    assert by_uuid["ap-ok"]["status"] == "success"
    assert by_uuid["ap-fail"]["status"] == "failed"
    assert str(by_uuid["ap-fail"]["error_message"]).startswith("log_write_failed:")
    assert progress[-1] == (2, 2)


def test_fit_ap_optical_round_emits_structured_item_progress(monkeypatch, tmp_path):
    ac_device = make_ac_device()
    ac_device.device_uuid = "ac-1"
    resources = [
        {"ap_uuid": f"ap-{index}", "ap_name": f"AP-{index}", "ap_ip": f"10.0.0.{index}", "site": "A站"}
        for index in range(1, 4)
    ]
    events: list[dict[str, object]] = []

    def collect_one(_ac_device, ap_row, *_args, **_kwargs):
        return {
            **dict(ap_row),
            "ac_device_uuid": _ac_device.device_uuid,
            "status": "success",
            "rx_power": "-7.88",
            "tx_power": "-2.10",
        }

    monkeypatch.setattr(h3c_ac_collect_service, "_collect_single_fit_ap_optical", collect_one)

    rows = h3c_ac_collect_service._collect_fit_ap_optical_round(
        ac_device,
        resources,
        "demo",
        "run-1",
        tmp_path,
        PathResolver(tmp_path),
        3,
        lambda: False,
        lambda _current, _total: None,
        round_index=1,
        item_progress=events.append,
    )

    completed = [event for event in events if event["event"] == "ap_completed"]
    assert len(rows) == 3
    assert len(completed) == 3
    assert sorted(event["index"] for event in completed) == [1, 2, 3]
    assert {event["total"] for event in completed} == {3}
    assert {event["ap_name"] for event in completed} == {"AP-1", "AP-2", "AP-3"}
    assert completed[-1]["success_count"] == 3
    assert completed[-1]["failed_count"] == 0
    assert completed[-1]["rx_power"] == "-7.88"
    assert isinstance(completed[-1]["elapsed_ms"], int)


def test_fit_ap_optical_round_reports_failed_ap_reason_and_retry_event(monkeypatch, tmp_path):
    ac_device = make_ac_device()
    ac_device.device_uuid = "ac-1"
    resources = [
        {"ap_uuid": "ap-ok", "ap_name": "AP-OK", "ap_ip": "10.0.0.21"},
        {"ap_uuid": "ap-fail", "ap_name": "AP-FAIL", "ap_ip": "10.0.0.22"},
    ]
    events: list[dict[str, object]] = []

    def collect_one(_ac_device, ap_row, *_args, **_kwargs):
        if ap_row["ap_uuid"] == "ap-fail":
            raise TimeoutError("connect timeout")
        return {**dict(ap_row), "ac_device_uuid": _ac_device.device_uuid, "status": "success"}

    monkeypatch.setattr(h3c_ac_collect_service, "_collect_single_fit_ap_optical", collect_one)

    rows = h3c_ac_collect_service._collect_fit_ap_optical_round(
        ac_device,
        resources,
        "demo",
        "run-1",
        tmp_path,
        PathResolver(tmp_path),
        2,
        lambda: False,
        lambda _current, _total: None,
        round_index=2,
        item_progress=events.append,
        retry=True,
    )

    retry_started = [event for event in events if event["event"] == "ap_retry_started"]
    failed = [event for event in events if event.get("status") == "failed"]
    assert len(rows) == 2
    assert len(retry_started) == 2
    assert failed[0]["ap_name"] == "AP-FAIL"
    assert failed[0]["round"] == 2
    assert failed[0]["reason_code"] == "connect_timeout"
    assert "connect_timeout" in str(failed[0]["error_message"])


def test_fit_ap_optical_round_skips_executor_when_no_targets(monkeypatch, tmp_path):
    ac_device = make_ac_device()

    def forbidden_executor(*_args, **_kwargs):
        raise AssertionError("empty FIT-AP target list must not create a thread pool")

    monkeypatch.setattr(h3c_ac_collect_service, "ThreadPoolExecutor", forbidden_executor)

    rows = h3c_ac_collect_service._collect_fit_ap_optical_round(
        ac_device,
        [],
        "demo",
        "run-empty",
        tmp_path,
        PathResolver(tmp_path),
        64,
        lambda: False,
        lambda _current, _total: None,
    )

    assert rows == []


def test_fit_ap_radio_history_is_appended_from_resource_rows(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    repository.replace_fit_ap_resources(
        "ac-1",
        [
            {
                "ap_name": "ap-a",
                "serial_number": "SN-001",
                "rid1_channel": "149",
                "rid1_bandwidth": "40",
                "rid1_tx_power": "24",
                "collected_at": "2026-01-01T00:00:00",
                "collect_run_uuid": "run-1",
            }
        ],
    )
    ap_uuid = repository.list_fit_ap_resources("ac-1")[0]["ap_uuid"]
    repository.replace_fit_ap_resources(
        "ac-1",
        [
            {
                "ap_uuid": ap_uuid,
                "ap_name": "ap-a",
                "serial_number": "SN-001",
                "rid1_channel": "153",
                "rid1_bandwidth": "80",
                "rid1_tx_power": "25",
                "collected_at": "2026-01-02T00:00:00",
                "collect_run_uuid": "run-2",
            }
        ],
    )

    history = repository.list_fit_ap_radio_history_by_ap(ap_uuid)

    assert history[0]["rid"] == 1
    assert [row["channel"] for row in history] == ["153"]
    assert history[0]["bandwidth"] == "80"
    assert history[0]["tx_power"] == "25"


def test_fit_ap_radio_state_change_creates_one_bounded_event(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    repository.replace_fit_ap_resources(
        "ac-1", [{"ap_name": "ap-a", "serial_number": "SN-001", "rid1_channel": "149", "rid1_usage": "10", "rid1_clients": "2"}]
    )
    ap_uuid = repository.list_fit_ap_resources("ac-1")[0]["ap_uuid"]
    repository.replace_fit_ap_resources(
        "ac-1", [{"ap_uuid": ap_uuid, "ap_name": "ap-a", "serial_number": "SN-001", "rid1_channel": "149", "rid1_usage": "20", "rid1_clients": "3", "collected_at": "2026-01-01T00:01:00"}]
    )
    history = repository.list_fit_ap_radio_history_by_ap(ap_uuid)
    assert len(history) == 1
    assert history[0]["usage"] == "20"
    assert history[0]["clients"] == 3


def test_fit_ap_resource_history_is_appended_from_resource_rows(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    repository.replace_fit_ap_resources(
        "ac-1",
        [
            {
                "ap_name": "ap-a",
                "ap_mac": "0011-2233-4455",
                "ap_ip": "10.0.0.1",
                "serial_number": "SN-001",
                "state_raw": "R/M",
                "state_display": "运行(主)",
                "site": "Station A",
                "collected_at": "2026-01-01T00:00:00",
                "collect_run_uuid": "run-1",
                "raw_log_path": "raw.log",
            }
        ],
    )
    row = repository.list_fit_ap_resource_history("ac-1")[0]

    assert row["ap_name"] == "ap-a"
    assert row["ap_mac"] == "0011-2233-4455"
    assert row["ap_ip"] == "10.0.0.1"
    assert row["state_raw"] == "R/M"
    assert row["site_name"] == "Station A"


def test_station_ap_capacity_can_be_saved_and_loaded(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    repository.upsert_station_ap_capacity("Station A", 56)

    assert repository.list_station_ap_capacities()["Station A"] == 56


def test_station_ap_capacity_remark_can_be_saved_and_loaded(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    repository.upsert_station_ap_capacity("Station A", 56)
    repository.upsert_station_ap_remark("Station A", "Need field check")

    details = repository.list_station_ap_capacity_details()
    assert details["Station A"]["ap_total"] == 56
    assert details["Station A"]["remark"] == "Need field check"


def test_trackside_ap_plan_repository_unified_capacity_and_pvid_plan(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    repository.upsert_station_ap_capacity("Station A", 12)
    repository.upsert_station_ap_remark("Station A", "Keep this remark")
    repository.replace_trackside_ap_plan_rows(
        TRACKSIDE_AP_PLAN_MODE,
        [
            {"station_name": "Station A", "ap_count": 10, "ap_management_vlans": "921"},
            {
                "station_name": "Station B",
                "ap_count": 56,
                "ap_management_vlans": "922,923",
            },
        ],
    )

    assert repository.get_trackside_ap_plan_mode() == TRACKSIDE_AP_PLAN_MODE
    details = repository.list_active_trackside_plan_capacity_details()
    assert details["Station A"]["ap_total"] == 10
    assert details["Station A"]["remark"] == "Keep this remark"
    assert details["Station A"]["source"] == "trackside_plan"
    active_plan = repository.get_active_trackside_pvid_plan()
    assert active_plan["station_vlans"]["Station A"] == {921}
    assert active_plan["all_vlans"] == {921, 922, 923}
    assert active_plan["station_totals"]["Station B"] == 56

    repository.upsert_trackside_ap_plan_row(
        TRACKSIDE_AP_PLAN_MODE,
        {"station_name": "Station A", "ap_count": 30, "ap_management_vlans": "921"},
    )
    updated_details = repository.list_active_trackside_plan_capacity_details()
    assert updated_details["Station A"]["ap_total"] == 30
    assert updated_details["Station A"]["remark"] == "Keep this remark"


def test_ap_entity_station_normalizes_aliases_and_preserves_existing_station(tmp_path):
    database = make_database(tmp_path)
    repository = AcRepository(database)
    ac = make_ac_device()

    repository.replace_fit_ap_resources(
        ac.device_uuid,
        [
            {
                "ap_uuid": "ap-1",
                "ap_name": "AP-1",
                "serial_number": "SN-AP-1",
                "site_name": "FIT Station",
                "state": "R/M",
            }
        ],
    )
    with database.connect() as conn:
        row = conn.execute(
            "SELECT station FROM ap_entities WHERE ap_uuid = 'ap-1'"
        ).fetchone()
    assert row["station"] == "FIT Station"

    repository.replace_fit_ap_resources(
        ac.device_uuid,
        [
            {
                "ap_uuid": "ap-1",
                "ap_name": "AP-1",
                "serial_number": "SN-AP-1",
                "ap_station": "LLDP Station",
                "site_name": "FIT Station 2",
                "state": "I",
            }
        ],
    )
    with database.connect() as conn:
        row = conn.execute(
            "SELECT station FROM ap_entities WHERE ap_uuid = 'ap-1'"
        ).fetchone()
    assert row["station"] == "FIT Station"


def test_ap_and_trackside_station_headers_display_ownership_station():
    i18n = I18n("zh_CN")

    assert i18n.t("ac.station") == "归属站点"
    assert i18n.t("field.station") == "归属站点"
    assert [i18n.t(key) for key, _field in TRACKSIDE_AP_BUSINESS_COLUMNS][
        0
    ] == "归属站点"


def test_trackside_ap_plan_unified_listing_ignores_legacy_modes(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    repository.replace_trackside_ap_plan_rows(
        "single_vlan",
        [{"station_name": "Station A", "ap_count": 10, "ap_management_vlans": "921"}],
    )
    repository.replace_trackside_ap_plan_rows(
        "multi_vlan",
        [
            {
                "station_name": "Station B",
                "ap_count": 56,
                "ap_management_vlans": "922,923",
            }
        ],
    )

    unified_rows = repository.list_trackside_ap_plan(TRACKSIDE_AP_PLAN_MODE)
    assert unified_rows == []

    active_plan = repository.get_active_trackside_pvid_plan()
    assert active_plan["planning_mode"] == "station_rows"
    assert active_plan["station_vlans"] == {}
    assert active_plan["all_vlans"] == set()

    repository.upsert_trackside_ap_plan_row(
        TRACKSIDE_AP_PLAN_MODE,
        {"station_name": "Station B", "ap_count": 34, "ap_management_vlans": "922"},
    )
    refreshed_rows = repository.list_trackside_ap_plan(TRACKSIDE_AP_PLAN_MODE)
    assert refreshed_rows[0]["mode"] == TRACKSIDE_AP_PLAN_MODE
    assert refreshed_rows[0]["ap_count"] == 34


def test_station_online_summary_history_table_created(tmp_path):
    database = make_database(tmp_path)
    with database.connect() as conn:
        row = conn.execute(
            "SELECT name FROM sqlite_master WHERE type = 'table' AND name = 'ac_station_online_summary_history'"
        ).fetchone()

    assert row is not None


def test_station_online_summary_history_save_and_list_desc(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    rows = [
        {
            "site": "Station A",
            "total": 5,
            "online": 4,
            "offline": 1,
            "online_rate": "80.0%",
            "remark": "First",
        },
        {
            "site": "合计",
            "total": 5,
            "online": 4,
            "offline": 1,
            "online_rate": "80.0%",
            "remark": "",
        },
    ]

    assert (
        repository.save_station_online_summary_history(
            rows, collected_at="2026-01-01T00:00:00"
        )
        == 1
    )
    repository.save_station_online_summary_history(
        [{**rows[0], "remark": "Second"}], collected_at="2026-01-02T00:00:00"
    )
    history = repository.list_station_online_summary_history("Station A")

    assert [row["remark"] for row in history] == ["Second", "First"]
    assert history[0]["site_name"] == "Station A"
    assert history[0]["online_count"] == 4


def test_station_ap_capacity_overrides_incomplete_planned_total(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    repository.upsert_station_ap_capacity("Station A", 56)

    rows = [{"ap_name": "AP-A", "ap_mac": "0011-2233-4455", "site_name": "Station A", "state": "R/M"}]
    overview = build_ap_online_overview_rows(
        planned_aps=rows,
        fit_ap_resources=rows,
        capacities=repository.list_station_ap_capacities(),
    )

    assert overview[0]["total"] == 56
    assert overview[0]["online"] == 1
    assert overview[0]["offline"] == 55


def test_ac_repository_metadata_crud_batch_edit_and_delete(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    repository.replace_fit_ap_resources(
        "ac-1", [{"ap_name": "ap-a"}, {"ap_name": "ap-b"}]
    )
    repository.replace_fit_ap_optical(
        "ac-1", [{"ap_name": "ap-a"}, {"ap_name": "ap-b"}]
    )

    repository.upsert_fit_ap_metadata(
        {
            "ap_name": "ap-a",
            "site_name": "S1",
            "belong_type": "section",
            "belong_section": "联庄-中医药大学",
            "section_start_station": "中医药大学",
            "section_end_station": "联庄",
        }
    )
    metadata = repository.get_fit_ap_metadata("ap-a")
    assert metadata["site_name"] == "S1"
    assert metadata["belong_type"] == "section"
    assert metadata["belong_section"] == "联庄-中医药大学"
    assert repository.update_fit_ap_site(["ap-a", "ap-b"], "S2") == 2
    assert repository.get_fit_ap_metadata("ap-b")["site_name"] == "S2"
    assert repository.delete_fit_aps("ac-1", ["ap-a"]) == 1
    assert [row["ap_name"] for row in repository.list_fit_ap_resources("ac-1")] == [
        "ap-b"
    ]
    assert repository.get_fit_ap_metadata("ap-a") is None


def test_wlan_ap_parser_extracts_summary_and_ap_list():
    summary = parse_wlan_ap_summary(fixture("display_wlan_ap_all.txt"))
    rows = parse_wlan_ap_list(fixture("display_wlan_ap_all.txt"))

    assert summary["total_aps"] == 2
    assert summary["online_aps"] == 2
    assert summary["remaining_local_ap_licenses"] == 59998
    assert rows[0]["ap_name"] == "4c6f-d608-0400"
    assert rows[0]["model"] == "WA6320-HCL"


def test_wlan_ap_parser_skips_state_legend_and_accepts_master_state():
    rows = parse_wlan_ap_list(
        """
 State : I = Idle,      J  = Join,       JA = JoinAck,    IL = ImageLoad
         C = Config,    DC = DataCheck,  R  = Run,   M = Master,  B = Backup

AP name                        APID  State Model           Serial ID
4c6f-d608-0400                 1     R/M   WA6320-HCL      H3C_4C-6F-D6-08-04-00
"""
    )

    assert len(rows) == 1
    assert rows[0]["state"] == "R/M"
    assert rows[0]["serial_number"] == "H3C_4C-6F-D6-08-04-00"


def test_wlan_ap_summary_counts_only_running_master_and_backup_as_online():
    summary = parse_wlan_ap_summary(
        """
Total number of APs: 3
AP name APID State Model Serial ID Group name Online time
AP-MASTER 1 R/M WA6624X SN-1 default-group 1:00:00
AP-BACKUP 2 R/B WA6624X SN-2 default-group 1:00:00
AP-JOIN 3 JA WA6624X SN-3 default-group 0:00:01
"""
    )

    assert summary["online_aps"] == 2
    assert summary["offline_aps"] == 1


def test_wlan_ap_parser_handles_v9_states_and_chinese_gb2312_output():
    output = """
 State : I = Idle,      J  = Join,       JA = JoinAck,    IL = ImageLoad
         C = Config,    DC = DataCheck,  R  = Run,   M = Master,  B = Backup

AP name                        APID State Model           Serial ID            Group name             Online time   Clients Mode  IP address
站厅_AP-01                     10   JA   WA6624X         SN-CN-0001           一层无线               0:00:01:02    0       Fit   10.1.1.10
4c6f-d608-0400                 11   IL   WA6320-HCL      SN-MAC-0002          default-group          12 days       0       Fit   10.1.1.11
ap_under-score                 12   DC   WA6630          SN-DC-0003           grp-01                 1:02:03:04    0       Fit   10.1.1.12
"""
    decoded = normalize_command_output(output.encode("gb2312"), "gb18030")
    rows = parse_wlan_ap_list(decoded)

    assert [row["ap_name"] for row in rows] == [
        "站厅_AP-01",
        "4c6f-d608-0400",
        "ap_under-score",
    ]
    assert rows[0]["state_display"] == "JoinAck"
    assert rows[1]["state_display"] == "ImageLoad"
    assert rows[2]["state_display"] == "DataCheck"
    assert rows[0]["group_name"] == "一层无线"


def test_wlan_ap_address_and_radio_parsers():
    addresses = parse_wlan_ap_addresses(fixture("display_wlan_ap_all_address.txt"))
    radios = parse_wlan_ap_radios(fixture("display_wlan_ap_all_radio.txt"))

    assert addresses["4c6f-d608-0400"]["ap_ip"] == "10.0.0.61"
    assert addresses["4c6f-d608-0400"]["ap_mac"] == "4c6f-d608-0400"
    assert radios["4c6f-d608-0400"]["rid1_channel"] == "1"
    assert radios["4c6f-d608-0400"]["rid2_tx_power"] == "17"


def test_address_radio_merge_by_ap_name_with_chinese_names():
    _summary, resources = h3c_ac_collect_service.parse_ac_resource_outputs(
        {
            "display wlan ap all": """
AP name                        APID State Model           Serial ID            Group name             Online time   Clients Mode  IP address
站厅_AP-01                     10   R/M  WA6624X         SN-CN-0001           一层无线               1:02:03:04    0       Fit   10.1.1.10
""",
            "display wlan ap all address": """
AP name                          IP address                       MAC address
站厅_AP-01                       10.1.1.10                        10b6-5e92-d3e0
""",
            "display wlan ap all radio": """
AP name                  RID State Channel          BW    Usage TxPower Clients
站厅_AP-01               1   Up    149              40    27    24      0
站厅_AP-01               2   Up    6                20    1     17      0
""",
        },
        "ac-1",
        "run-1",
        "files/rail_transit/trackside_ap/raw/ac/run-1/ac.log",
    )

    assert resources[0]["ap_name"] == "站厅_AP-01"
    assert resources[0]["apid"] == "10"
    assert resources[0]["ap_mac"] == "10b6-5e92-d3e0"
    assert resources[0]["rid1_channel"] == "149"
    assert resources[0]["rid2_tx_power"] == "17"


def test_address_output_missing_offline_ap_keeps_persisted_mac(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    repository.replace_fit_ap_resources(
        "ac-1",
        [
            {
                "ap_name": "AP-OFFLINE",
                "apid": "11",
                "ap_mac": "bc5a-3457-b5e0",
                "serial_number": "SN-OFFLINE",
                "ap_ip": "10.1.1.11",
                "state": "R/M",
            }
        ],
    )
    previous_uuid = repository.list_fit_ap_resources("ac-1")[0]["ap_uuid"]
    _summary, resources = h3c_ac_collect_service.parse_ac_resource_outputs(
        {
            "display wlan ap all": """
AP name             APID State Model      Serial ID     Group name       Online time
AP-ONLINE           10   R/M   WA6624X    SN-ONLINE    default-group    1:02:03:04
AP-OFFLINE          11   I     WA6624X    N/A          default-group    0:00:00:00
""",
            "display wlan ap all address": """
AP name             IP address       MAC address
AP-ONLINE           10.1.1.10       bc5a-3457-b520
""",
        },
        "ac-1",
        "run-offline",
        "files/rail_transit/trackside_ap/raw/ac/run-offline/ac.log",
    )

    repository.replace_fit_ap_resources("ac-1", resources)
    by_name = {row["ap_name"]: row for row in repository.list_fit_ap_resources("ac-1")}
    assert by_name["AP-ONLINE"]["ap_mac"] == "bc5a-3457-b520"
    assert by_name["AP-OFFLINE"]["ap_uuid"] == previous_uuid
    assert by_name["AP-OFFLINE"]["ap_mac"] == "bc5a-3457-b5e0"
    assert by_name["AP-OFFLINE"]["ap_ip"] is None
    assert by_name["AP-OFFLINE"]["state"] == "I"


def test_state_cpu_and_memory_parsers():
    assert map_fit_ap_state("R/M") == "\u8fd0\u884c(\u4e3b)"
    assert map_fit_ap_state("R/B") == "\u8fd0\u884c(\u5907)"
    assert map_fit_ap_state("JA") == "JoinAck"
    assert classify_fit_ap_state("R/M") == "online"
    assert classify_fit_ap_state("R/B") == "online"
    assert classify_fit_ap_state("JA") == "offline"
    assert classify_fit_ap_state("") == "unknown"
    assert parse_cpu_usage(fixture("display_cpu_usage.txt")) == {
        "cpu_5s": 16,
        "cpu_1m": 18,
        "cpu_5m": 18,
        "cpu_usage": "16%",
    }
    memory = parse_memory(fixture("display_memory.txt"))
    assert memory["memory_total"] == 770180
    assert memory["memory_free_ratio"] == 53.0
    assert memory["memory_usage"] == "47%"
    memory_table = parse_memory(
        "Mem:        770180    366008    404172         0      3848    156656       52.9%"
    )
    assert memory_table["memory_total"] == 770180
    assert memory_table["memory_used"] == 366008
    assert memory_table["memory_usage"] == "47%"


def test_wlan_ap_radio_parser_handles_state_column():
    radios = parse_wlan_ap_radios(
        """
AP name                  RID State Channel          BW    Usage TxPower Clients
                                                    (MHz) (%)   (dBm)
4c6f-d608-0400           1   Down  52(auto)         80    0     20      0
"""
    )

    assert radios["4c6f-d608-0400"]["rid1_channel"] == "52(auto)"
    assert radios["4c6f-d608-0400"]["rid1_bandwidth"] == "80"
    assert radios["4c6f-d608-0400"]["rid1_status"] == "Down"
    assert radios["4c6f-d608-0400"]["rid1_usage"] == "0"
    assert radios["4c6f-d608-0400"]["rid1_tx_power"] == "20"
    assert radios["4c6f-d608-0400"]["rid1_clients"] == "0"


def test_wlan_ap_connection_record_and_radio_type_parsers_follow_h3c_output():
    records = parse_wlan_ap_connection_records(
        """
AP name                          IP address    State      Time
ap1                              2001::3       Run        05-06 09:47:44
ap2                              N/A           Offline    05-06 09:50:38
"""
    )
    radio_types = parse_wlan_ap_radio_types(
        """
AP name                  RID  AP state  Radio state  Radio type
ap1                      1    Up        Up           802.11n(5GHz)
ap1                      2    Up        Down         802.11n(2.4GHz)
"""
    )

    assert records["ap1"] == {
        "ap_name": "ap1",
        "connection_ip": "2001::3",
        "connection_state": "Run",
        "connection_time": "05-06 09:47:44",
    }
    assert records["ap2"]["connection_ip"] is None
    assert radio_types["ap1"]["rid1_status"] == "Up"
    assert radio_types["ap1"]["rid1_mode"] == "802.11n"
    assert radio_types["ap1"]["rid1_band"] == "5GHz"
    assert radio_types["ap1"]["rid2_status"] == "Down"
    assert radio_types["ap1"]["rid2_band"] == "2.4GHz"


def test_wlan_ap_connection_record_rows_keep_raw_time_and_resolve_cross_year():
    rows = parse_wlan_ap_connection_record_rows(
        """
AP name                          IP address    State      Time
0011-2233-4455                   10.0.0.1      Run        12-31 23:59:58
AP-2                             N/A           Offline    01-01 00:00:03
""",
        collected_at="2026-01-01 00:00:10",
        ac_id="ac-1",
        site_key="NBO12",
    )

    assert rows == [
        {
            "ap_name": "0011-2233-4455",
            "ap_mac": "001122334455",
            "ip_address": "10.0.0.1",
            "state": "Run",
            "connection_ip": "10.0.0.1",
            "connection_state": "Run",
            "connection_time": "12-31 23:59:58",
            "raw_time": "12-31 23:59:58",
            "resolved_time": "2025-12-31 23:59:58",
            "ac_id": "ac-1",
            "site_key": "NBO12",
            "collected_at": "2026-01-01 00:00:10",
            "raw_line": "0011-2233-4455                   10.0.0.1      Run        12-31 23:59:58",
        },
        {
            "ap_name": "AP-2",
            "ap_mac": None,
            "ip_address": None,
            "state": "Offline",
            "connection_ip": None,
            "connection_state": "Offline",
            "connection_time": "01-01 00:00:03",
            "raw_time": "01-01 00:00:03",
            "resolved_time": "2026-01-01 00:00:03",
            "ac_id": "ac-1",
            "site_key": "NBO12",
            "collected_at": "2026-01-01 00:00:10",
            "raw_line": "AP-2                             N/A           Offline    01-01 00:00:03",
        },
    ]


def test_connection_record_authority_persists_stable_transitions(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    repository.replace_fit_ap_resources(
        "ac-1",
        [
            {
                "ap_uuid": "ap-1",
                "ap_name": "AP-1",
                "ap_mac": "0011-2233-4455",
                "serial_number": "SN-1",
                "site_key": "NBO12",
                "state": "R/M",
                "collected_at": "2026-01-01 10:00:00",
            }
        ],
    )

    def apply(state: str, collected_at: str, resolved_time: str) -> None:
        repository.apply_fit_ap_connection_records(
            "ac-1",
            [
                {
                    "ap_name": "AP-1",
                    "ap_mac": "0011-2233-4455",
                    "state": state,
                    "connection_state": state,
                    "raw_time": resolved_time[5:],
                    "resolved_time": resolved_time,
                    "collected_at": collected_at,
                    "site_key": "NBO12",
                }
            ],
        )

    apply("Run", "2026-01-01 10:05:00", "2026-01-01 10:00:00")
    apply("Offline", "2026-01-01 11:05:00", "2026-01-01 11:00:00")
    offline = repository.list_offline_ap_entities("ac-1")
    assert len(offline) == 1
    assert offline[0]["connection_state"] == "Offline"
    assert offline[0]["last_online_time"] == "2026-01-01 10:00:00"
    assert offline[0]["offline_time"] == "2026-01-01 11:05:00"

    apply("Run", "2026-01-01 12:05:00", "2026-01-01 12:00:00")
    entity = next(row for row in repository.list_ap_entities("ac-1") if row["ap_uuid"] == "ap-1")
    assert entity["connection_state"] == "Run"
    assert entity["last_online_time"] == "2026-01-01 12:00:00"
    assert entity["offline_time"] == ""
    assert entity["last_state_change_at"] == "2026-01-01 12:05:00"
    assert entity["last_connection_record_seen_at"] == "2026-01-01 12:05:00"
    assert entity["connection_reonline_count"] == 1
    assert repository.list_offline_ap_entities("ac-1") == []


def test_fit_ap_optical_parser_extracts_lldp_and_power_summary():
    parsed = parse_fit_ap_optical(
        fixture("display_fit_ap_lldp.txt"),
        fixture("display_fit_ap_transceiver_diagnosis.txt"),
    )

    assert "SW01-DEMO" in parsed["lldp_neighbor"]
    assert parsed["rx_power"] == "-3.21"
    assert parsed["tx_power"] == "-2.85"
    assert "optical_alarm_status" not in parsed


def test_fit_ap_lldp_parser_handles_fit_ap_table_format():
    parsed = parse_fit_ap_lldp(
        """
System Name          Local Interface Chassis ID      Port ID
HX_1                 GE1/0/2         903f-8645-6e00  GigabitEthernet2/0/19
"""
    )

    assert parsed["lldp_neighbor"] == "HX_1"
    assert parsed["interface_name"] == "GigabitEthernet1/0/2"
    assert parsed["neighbor_mac"] == "903f-8645-6e00"
    assert parsed["neighbor_interface"] == "GigabitEthernet2/0/19"


def test_fit_ap_lldp_parser_does_not_fallback_after_invalid_supported_table_row():
    parsed = parse_fit_ap_lldp(
        """
Local Interface Chassis ID Port ID System Name
GE1/0/1 H3C H3C H3C
"""
    )

    assert parsed["lldp_neighbor"] is None
    assert parsed["interface_name"] is None
    assert parsed["neighbor_interface"] is None
    assert parsed["neighbor_mac"] is None


def test_fit_ap_transceiver_parser_extracts_diagnosis_interface_and_manuinfo():
    parsed = parse_fit_ap_transceiver(
        """
GigabitEthernet1/0/2 transceiver diagnostic information:
Current diagnostic parameters:
Temp.(C) Voltage(V) Bias(mA) RX power(dBm) TX power(dBm)
43       3.31       6.10     -7.55         -6.09
Alarm thresholds:
High     90         3.63      0.00          0.00
Low      -10        2.97      -20.00        -20.00
Warning thresholds:
High     85         3.50      -1.00         -1.00
Low      -5         3.00      -17.00        -17.00
""",
        """
GigabitEthernet1/0/2 transceiver information:
Transceiver Type           : 1000_BASE_LX_SFP
Wavelength(nm)             : 1310
Transfer Distance(km)      : 10
Connector Type             : LC
Vendor Name                : H3C
""",
        """
GigabitEthernet1/0/2 transceiver manufacture information:
Manu. Serial Number        : SN123456
Vendor Name                : H3C
""",
    )

    assert parsed["interface_name"] == "GigabitEthernet1/0/2"
    assert parsed["temperature"] == "43"
    assert parsed["rx_power"] == "-7.55"
    assert parsed["tx_power"] == "-6.09"
    assert parsed["module_model"] == "1000_BASE_LX_SFP"
    assert parsed["wavelength"] == "1310 nm"
    assert parsed["transmission_distance"] == "10 km"
    assert parsed["module_serial_number"] == "SN123456"


def test_fit_ap_optical_parser_handles_real_machine_lldp_and_transceiver():
    lldp = parse_fit_ap_lldp(ac_fixture("real_fit_ap_lldp_neighbor.txt"))
    optical = parse_fit_ap_transceiver(
        ac_fixture("real_fit_ap_transceiver_diagnosis.txt")
    )

    assert lldp["lldp_neighbor"] == "HX_1"
    assert lldp["neighbor_interface"] == "GigabitEthernet2/0/19"
    assert lldp["neighbor_mac"] == "903f-8645-6e00"
    assert optical["interface_name"] == "GigabitEthernet1/0/2"
    assert optical["temperature"] == "43"
    assert optical["rx_power"] == "-7.55"
    assert optical["tx_power"] == "-6.09"


def test_real_machine_wlan_parsers_read_large_fixture():
    rows = parse_wlan_ap_list(ac_fixture("real_display_wlan_ap_all.txt"))
    addresses = parse_wlan_ap_addresses(
        ac_fixture("real_display_wlan_ap_all_address.txt")
    )
    radios = parse_wlan_ap_radios(ac_fixture("real_display_wlan_ap_all_radio.txt"))

    assert len(rows) >= 500
    assert rows[0]["ap_name"] == "AP-CLD_01"
    assert rows[0]["serial_number"] == "219801A4588249E00063"
    assert rows[0]["group_name"] == "cld-tcc-dcc"
    assert rows[0]["online_time"] == "27:02:49:39"
    assert addresses["AP-CLD_01"]["ap_ip"] == "10.62.113.177"
    assert radios["AP-CLD_01"]["rid1_channel"] == "149"
    assert radios["AP-CLD_01"]["rid1_status"] == "Up"
    assert radios["AP-CLD_01"]["rid1_usage"] == "27"
    assert radios["AP-CLD_01"]["rid1_clients"] == "0"


def test_wlan_ap_radio_verbose_bbssid_parser_groups_by_ap_name():
    parsed = parse_wlan_ap_radio_verbose_bbssid(
        """
Total number of APs: 932

                             Radio Filtered Information
  bbssid = Base BSSID

AP name              RID bbssid
30f5-277a-0ea0       1   30f5-277a-0ea0
30f5-277a-0ea0       2   30f5-277a-0eb0
30f5-277a-0ee0       1   30f5-277a-0ee0
30f5-277a-0ee0       2   30f5-277a-0ef0
30f5-277a-0f00       1   30f5-277a-0f00
"""
    )

    assert parsed["30f5-277a-0ea0"]["rid1_bbssid"] == "30f5-277a-0ea0"
    assert parsed["30f5-277a-0ea0"]["rid2_bbssid"] == "30f5-277a-0eb0"
    assert parsed["30f5-277a-0ee0"]["rid2_bbssid"] == "30f5-277a-0ef0"


def test_wlan_ap_lldp_parser_keeps_neighbor_interface():
    parsed = parse_wlan_ap_lldp(
        """
AP name                        Local Interface          Neighbor Name                  Neighbor MAC    Neighbor Interface
30f5-277a-0ea0                 GE1/0/2                  N/A                            903f-8645-6e00  GigabitEthernet2/0/19
30f5-277a-0ee0                 GE1/0/2                  N/A                            903f-8645-a600  GigabitEthernet2/0/8
30f5-277a-0f00                 GE1/0/2                  N/A                            903f-8645-fa00  GigabitEthernet2/0/38
"""
    )

    assert parsed["30f5-277a-0ea0"]["lldp_local_interface"] == "GigabitEthernet1/0/2"
    assert parsed["30f5-277a-0ea0"]["lldp_neighbor_name"] == "N/A"
    assert parsed["30f5-277a-0ea0"]["lldp_neighbor_mac"] == "903f-8645-6e00"
    assert (
        parsed["30f5-277a-0ea0"]["lldp_neighbor_interface"] == "GigabitEthernet2/0/19"
    )
    assert parsed["30f5-277a-0ea0"]["lldp_source"] == "ac_bulk_lldp"
    assert parsed["30f5-277a-0ea0"]["lldp_local_interface_normalized"] == "ge1/0/2"
    assert parsed["30f5-277a-0ea0"]["lldp_neighbor_mac_normalized"] == "903f86456e00"


def test_fit_ap_direct_lldp_parser_normalizes_neighbor_row():
    parsed = parse_fit_ap_lldp_neighbor(
        """
System Name          Local Interface  Chassis ID       Port ID
HX_1                 GE1/0/2          903f-8645-6e00   GigabitEthernet2/0/19
"""
    )

    assert parsed["lldp_source"] == "ap_direct_lldp"
    assert parsed["lldp_neighbor_name"] == "HX_1"
    assert parsed["lldp_local_interface"] == "GigabitEthernet1/0/2"
    assert parsed["lldp_local_interface_normalized"] == "ge1/0/2"
    assert parsed["lldp_neighbor_mac"] == "903f-8645-6e00"
    assert parsed["lldp_neighbor_mac_normalized"] == "903f86456e00"
    assert parsed["lldp_neighbor_interface"] == "GigabitEthernet2/0/19"


def test_fit_ap_direct_lldp_parser_supports_wa6522_and_keeps_generic_name_as_raw_evidence():
    parsed = parse_fit_ap_lldp_neighbor(
        """
Local Interface Chassis ID      Port ID                         System Name
GE1/0/1         2c4c-7d66-c492  GigabitEthernet1/0/4            H3C
"""
    )

    assert parsed["lldp_local_interface_normalized"] == "ge1/0/1"
    assert parsed["lldp_neighbor_mac_normalized"] == "2c4c7d66c492"
    assert parsed["lldp_neighbor_interface"] == "GigabitEthernet1/0/4"
    assert parsed["lldp_neighbor_name"] == "H3C"


def test_lldp_normalization_preserves_explicit_partial_and_unknown_statuses():
    partial = normalize_lldp_payload(
        {
            "lldp_local_interface": "GE1/0/1",
            "lldp_neighbor_mac": "2c4c-7d66-c492",
            "lldp_neighbor_interface": "GigabitEthernet1/0/4",
            "lldp_match_status": "partial",
        },
        "ap_direct_lldp",
    )
    unknown = normalize_lldp_payload(
        {
            "lldp_neighbor_mac": "2c4c-7d66-c492",
            "lldp_match_status": "unknown",
        },
        "ap_direct_lldp",
    )

    assert partial["lldp_match_status"] == "partial"
    assert unknown["lldp_match_status"] == "unknown"


def test_fit_ap_transceiver_snapshot_parser_marks_optical_interface():
    snapshots = parse_fit_ap_transceiver_diagnosis_snapshots(
        """
GigabitEthernet1/0/2 transceiver diagnostic information:
Current diagnostic parameters:
Temp.(C) Voltage(V) Bias(mA) RX power(dBm) TX power(dBm)
43       3.31       6.10     -7.55         -6.09
"""
    )

    assert snapshots[0]["interface_name"] == "GigabitEthernet1/0/2"
    assert snapshots[0]["optical_interface"] == "GigabitEthernet1/0/2"
    assert snapshots[0]["optical_interface_normalized"] == "ge1/0/2"
    assert snapshots[0]["rx_power"] == "-7.55"


def test_fit_ap_link_normalization_and_merge_prioritize_ap_direct_lldp():
    assert (
        normalize_interface_key("GE1/0/2")
        == normalize_interface_key("GigabitEthernet1/0/2")
        == "ge1/0/2"
    )
    assert normalize_link_mac("90:3f:86:45:6e:00") == "903f86456e00"
    assert format_h3c_mac("903f86456e00") == "903f-8645-6e00"

    merged = merge_lldp_payload(
        {
            "lldp_source": "ac_bulk_lldp",
            "lldp_local_interface": "GE1/0/2",
            "lldp_neighbor_name": "N/A",
            "lldp_neighbor_mac": "903f-8645-6e00",
            "lldp_neighbor_interface": "GE2/0/19",
        },
        {
            "lldp_source": "ap_direct_lldp",
            "lldp_local_interface": "GigabitEthernet1/0/2",
            "lldp_neighbor_name": "HX_1",
            "lldp_neighbor_mac": "90:3f:86:45:6e:00",
            "lldp_neighbor_interface": "GigabitEthernet2/0/19",
        },
    )

    assert merged["lldp_neighbor_name"] == "HX_1"
    assert merged["lldp_source"] == "merged"
    assert merged["lldp_confidence"] == 90
    assert merged["lldp_match_status"] == "matched"
    assert (
        resolve_optical_match_status(
            merged, {"optical_interface": "GE1/0/2", "rx_power": "-7.55"}
        )
        == "matched"
    )
    assert (
        resolve_optical_match_status(
            merged, {"optical_interface": "GE1/0/3", "rx_power": "-7.55"}
        )
        == "conflict"
    )


def test_fit_ap_link_view_model_maps_legacy_fields_to_current_fields():
    resolved = resolve_fit_ap_link_info(
        {
            "lldp_neighbor": "HX_1",
            "neighbor_interface": "GigabitEthernet2/0/19",
            "neighbor_mac": "903f-8645-6e00",
            "neighbor_device_name": "04-横溪站",
            "neighbor_rx_power": "-7.55",
            "interface_name": "GE1/0/2",
        }
    )

    assert resolved["lldp_neighbor_name"] == "HX_1"
    assert resolved["lldp_neighbor_interface"] == "GigabitEthernet2/0/19"
    assert resolved["lldp_neighbor_mac"] == "903f-8645-6e00"
    assert resolved["lldp_neighbor_mac_normalized"] == "903f86456e00"
    assert resolved["neighbor_device_name"] == "04-横溪站"
    assert resolved["lldp_source"] == "legacy_compat"
    assert resolved["lldp_match_status"] == "matched"
    assert resolved["optical_rx_power"] == "-7.55"
    assert resolved["optical_match_status"] == "matched"


def test_h3c_ac_collect_service_uses_mock_netmiko(monkeypatch, tmp_path):
    connection = FakeConnection()
    monkeypatch.setattr(
        h3c_ac_collect_service.netmiko_connection,
        "ConnectHandler",
        lambda **_kwargs: connection,
    )
    database = make_database(tmp_path)
    repository = AcRepository(database)

    result = collect_h3c_fit_ap_resources(
        make_ac_device(), "demo", repository=repository, paths=PathResolver(tmp_path)
    )

    assert result.success is True
    assert result.summary_updated is True
    assert result.fit_ap_resources_updated == 2
    assert result.fit_ap_snapshot_status == "SUCCESS_WITH_ROWS"
    assert connection.commands == ["screen-length disable", *RESOURCE_COMMANDS]
    assert connection.disconnected is True
    assert Path(result.raw_log_path).is_file()
    assert "display wlan ap all radio verbose filter bbssid" in Path(
        result.raw_log_path
    ).read_text(encoding="utf-8")
    assert next(
        call["read_timeout"]
        for call in connection.calls
        if call["command"] == "display wlan ap all radio verbose filter bbssid"
    ) == 120
    assert result.bbssid_collect_status == "success"
    assert result.bbssid_error is None
    assert (
        repository.get_ac_ap_summary("22222222-2222-4222-8222-222222222222")[
            "total_aps"
        ]
        == 2
    )
    assert (
        repository.list_fit_ap_resources("22222222-2222-4222-8222-222222222222")[0][
            "ap_ip"
        ]
        == "10.0.0.61"
    )


def test_enable_ap_remote_login_uses_per_command_timeouts(monkeypatch, tmp_path):
    connection = FakeTimingConnection()
    monkeypatch.setattr(
        h3c_ac_collect_service.netmiko_connection,
        "ConnectHandler",
        lambda **_kwargs: connection,
    )
    device = make_ac_device()
    device.device_vendor = "H3C"

    result = h3c_ac_collect_service.run_h3c_ac_action(
        device,
        "demo",
        "enable_ap_remote_login",
        repository=AcRepository(make_database(tmp_path)),
        paths=PathResolver(tmp_path),
    )

    assert result.success is True
    assert connection.commands == [
        "screen-length disable",
        "system-view",
        "probe",
        "wlan ap-execute all exec-console enable",
        "return",
        "quit",
    ]
    timeouts = {call["command"]: call["read_timeout"] for call in connection.calls}
    assert timeouts["screen-length disable"] == 15
    assert timeouts["system-view"] == 15
    assert timeouts["probe"] == 30
    assert timeouts["wlan ap-execute all exec-console enable"] == 120
    assert timeouts["return"] == 30
    assert timeouts["quit"] == 30
    assert all(
        call["strip_prompt"] is False and call["strip_command"] is False
        for call in connection.calls
    )


def test_enable_ap_remote_login_treats_tail_read_timeout_as_success(
    monkeypatch, tmp_path
):
    timeout = RuntimeError(
        "return: read_channel_timing's absolute timer expired. "
        "The network device was continually outputting data for longer than 10 seconds."
    )
    connection = FakeTimingConnection({"return": timeout})
    monkeypatch.setattr(
        h3c_ac_collect_service.netmiko_connection,
        "ConnectHandler",
        lambda **_kwargs: connection,
    )
    device = make_ac_device()
    device.device_vendor = "H3C"

    result = h3c_ac_collect_service.run_h3c_ac_action(
        device,
        "demo",
        "enable_ap_remote_login",
        repository=AcRepository(make_database(tmp_path)),
        paths=PathResolver(tmp_path),
    )

    assert result.success is True
    return_result = next(
        item for item in result.command_results if item.command == "return"
    )
    assert return_result.success is True
    assert str(return_result.error_message).startswith("warning: read timeout")
    assert "treated as success" in return_result.output
    assert connection.commands[-1] == "quit"


def test_enable_ap_remote_login_keeps_real_command_error_failed(monkeypatch, tmp_path):
    connection = FakeTimingConnection(
        {
            "wlan ap-execute all exec-console enable": "% Unrecognized command found at '^' position."
        }
    )
    monkeypatch.setattr(
        h3c_ac_collect_service.netmiko_connection,
        "ConnectHandler",
        lambda **_kwargs: connection,
    )
    device = make_ac_device()
    device.device_vendor = "H3C"

    result = h3c_ac_collect_service.run_h3c_ac_action(
        device,
        "demo",
        "enable_ap_remote_login",
        repository=AcRepository(make_database(tmp_path)),
        paths=PathResolver(tmp_path),
    )

    assert result.success is False
    assert "% Unrecognized command" in str(result.error_message)


def test_h3c_ac_resource_only_collect_skips_overview_commands(monkeypatch, tmp_path):
    connection = FakeConnection()
    monkeypatch.setattr(
        h3c_ac_collect_service.netmiko_connection,
        "ConnectHandler",
        lambda **_kwargs: connection,
    )
    database = make_database(tmp_path)
    repository = AcRepository(database)

    result = collect_h3c_fit_ap_resources(
        make_ac_device(),
        "demo",
        repository=repository,
        paths=PathResolver(tmp_path),
    )

    assert result.success is True
    assert result.summary_updated is True
    assert result.https_port_collected is False
    assert connection.commands == [
        "screen-length disable",
        *FIT_AP_RESOURCE_COMMANDS,
        *FIT_AP_RESOURCE_OPTIONAL_COMMANDS,
    ]
    assert "display cpu-usage" not in connection.commands
    assert "display memory" not in connection.commands
    assert "display version" not in connection.commands
    assert "display device" not in connection.commands
    summary = repository.get_ac_ap_summary("22222222-2222-4222-8222-222222222222")
    assert summary["total_aps"] == 2
    assert summary["online_aps"] == 2
    assert summary["offline_aps"] == 0
    assert summary["cpu_usage"] is None
    assert summary["model"] is None
    assert (
        repository.list_fit_ap_resources("22222222-2222-4222-8222-222222222222")[0][
            "ap_ip"
        ]
        == "10.0.0.61"
    )


def test_h3c_ac_resource_collect_reports_optional_bbssid_failure_without_losing_resources(
    monkeypatch, tmp_path
):
    command = "display wlan ap all radio verbose filter bbssid"
    connection = FakeConnection({command: TimeoutError("verbose output timed out")})
    monkeypatch.setattr(
        h3c_ac_collect_service.netmiko_connection,
        "ConnectHandler",
        lambda **_kwargs: connection,
    )
    repository = AcRepository(make_database(tmp_path))

    result = collect_h3c_fit_ap_resources(
        make_ac_device(),
        "demo",
        repository=repository,
        paths=PathResolver(tmp_path),
    )

    assert result.success is True
    assert result.fit_ap_resources_updated == 2
    assert result.bbssid_rows_parsed == 0
    assert result.bbssid_collect_status == "failed"
    assert "verbose output timed out" in str(result.bbssid_error)
    assert command in [
        item.command for item in result.command_results if not item.success
    ]
    raw_text = Path(result.raw_log_path).read_text(encoding="utf-8")
    assert command in raw_text
    assert "verbose output timed out" in raw_text


def test_h3c_fit_ap_deep_refresh_uses_verified_verbose_command_and_only_upserts_target(
    monkeypatch, tmp_path
):
    connection = FakeConnection(
        {
            "display wlan ap all radio verbose filter bbssid": (
                "AP name              RID bbssid\n"
                "4c6f-d608-0400       1   0011-2233-4455\n"
            ),
        }
    )
    monkeypatch.setattr(
        h3c_ac_collect_service.netmiko_connection,
        "ConnectHandler",
        lambda **_kwargs: connection,
    )
    repository = AcRepository(make_database(tmp_path))
    ac_uuid = "22222222-2222-4222-8222-222222222222"
    repository.replace_fit_ap_resources(
        ac_uuid,
        [
            {"ap_name": "4c6f-d608-0400", "rid1_bbssid": "old-bssid"},
            {"ap_name": "AP-KEEP", "rid1_channel": "6"},
        ],
    )
    target = repository.get_fit_ap_resource(ac_uuid, "4c6f-d608-0400")

    result = collect_h3c_fit_ap_resources(
        make_ac_device(),
        "demo",
        repository=repository,
        paths=PathResolver(tmp_path),
        target_ap_uuid=str(target["ap_uuid"]),
    )

    assert result.success is True
    assert result.fit_ap_snapshot_status == "NOT_COLLECTED"
    assert "display wlan ap all radio verbose filter bbssid" in connection.commands
    assert "display wlan ap unauthenticated" not in connection.commands
    assert (
        repository.get_fit_ap_resource(ac_uuid, "4c6f-d608-0400")["rid1_bbssid"]
        == "0011-2233-4455"
    )
    assert repository.get_fit_ap_resource(ac_uuid, "AP-KEEP")["rid1_channel"] == "6"


def test_h3c_fit_ap_single_detail_uses_name_verbose_command(monkeypatch, tmp_path):
    output = ac_fixture("real_display_wlan_ap_verbose.txt").split("AP name : AP-NB12-02", 1)[0]
    connection = FakeConnection(
        {"display wlan ap name AP-NB12-01 verbose": output}
    )
    monkeypatch.setattr(
        h3c_ac_collect_service.netmiko_connection,
        "ConnectHandler",
        lambda **_kwargs: connection,
    )
    repository = AcRepository(make_database(tmp_path))
    ac_uuid = "22222222-2222-4222-8222-222222222222"
    repository.replace_fit_ap_resources(
        ac_uuid,
        [
            {
                "ap_name": "AP-NB12-01",
                "ap_mac": "28c9-7a3e-5da0",
                "serial_number": "219801A3L68257P005M3",
            },
            {"ap_name": "AP-KEEP", "rid1_channel": "6"},
        ],
    )
    target = repository.get_fit_ap_resource(ac_uuid, "AP-NB12-01")

    result = collect_h3c_fit_ap_resources(
        make_ac_device(),
        "demo",
        repository=repository,
        paths=PathResolver(tmp_path),
        target_ap_uuid=str(target["ap_uuid"]),
    )

    assert result.success is True
    assert connection.commands == [
        "screen-length disable",
        "display wlan ap name AP-NB12-01 verbose",
    ]
    detail = repository.get_fit_ap_detail(str(target["ap_uuid"]))
    assert detail is not None
    assert detail["software_version"] == "Version 7.1.064, Release 2619P08"
    assert repository.get_fit_ap_resource(ac_uuid, "AP-KEEP")["rid1_channel"] == "6"


def test_h3c_ac_resource_only_collect_preserves_static_summary_fields(
    monkeypatch, tmp_path
):
    connection = FakeConnection()
    monkeypatch.setattr(
        h3c_ac_collect_service.netmiko_connection,
        "ConnectHandler",
        lambda **_kwargs: connection,
    )
    database = make_database(tmp_path)
    repository = AcRepository(database)
    ac_uuid = "22222222-2222-4222-8222-222222222222"
    repository.upsert_ac_ap_summary(
        {
            "ac_device_uuid": ac_uuid,
            "total_aps": 82,
            "online_aps": 58,
            "offline_aps": 24,
            "cpu_usage": "16%",
            "memory_usage": "47%",
            "model": "WX-AC",
            "serial_number": "SN-AC",
            "software_version": "Version 7.1",
        }
    )

    collect_h3c_fit_ap_resources(
        make_ac_device(),
        "demo",
        repository=repository,
        paths=PathResolver(tmp_path),
    )

    summary = repository.get_ac_ap_summary(ac_uuid)
    assert summary["total_aps"] == 2
    assert summary["online_aps"] == 2
    assert summary["offline_aps"] == 0
    assert summary["cpu_usage"] == "16%"
    assert summary["memory_usage"] == "47%"
    assert summary["model"] == "WX-AC"
    assert summary["serial_number"] == "SN-AC"
    assert summary["software_version"] == "Version 7.1"


def test_h3c_ac_resource_only_collect_does_not_overwrite_summary_when_ap_all_fails(
    monkeypatch, tmp_path
):
    connection = FakeConnection({"display wlan ap all": RuntimeError("command failed")})
    monkeypatch.setattr(
        h3c_ac_collect_service.netmiko_connection,
        "ConnectHandler",
        lambda **_kwargs: connection,
    )
    database = make_database(tmp_path)
    repository = AcRepository(database)
    ac_uuid = "22222222-2222-4222-8222-222222222222"
    repository.upsert_ac_ap_summary(
        {
            "ac_device_uuid": ac_uuid,
            "total_aps": 82,
            "online_aps": 58,
            "offline_aps": 24,
        }
    )

    result = collect_h3c_fit_ap_resources(
        make_ac_device(),
        "demo",
        repository=repository,
        paths=PathResolver(tmp_path),
    )

    summary = repository.get_ac_ap_summary(ac_uuid)
    assert result.success is False
    assert summary["total_aps"] == 82
    assert summary["online_aps"] == 58
    assert summary["offline_aps"] == 24


def test_h3c_ac_resource_collect_failure_preserves_current_resources(
    monkeypatch, tmp_path
):
    connection = FakeConnection({"display wlan ap all": RuntimeError("command failed")})
    monkeypatch.setattr(
        h3c_ac_collect_service.netmiko_connection,
        "ConnectHandler",
        lambda **_kwargs: connection,
    )
    repository = AcRepository(make_database(tmp_path))
    ac_uuid = "22222222-2222-4222-8222-222222222222"
    repository.replace_fit_ap_resources(
        ac_uuid,
        [{"ap_uuid": "ap-keep", "ap_name": "AP-KEEP", "serial_number": "SN-KEEP"}],
    )

    result = collect_h3c_fit_ap_resources(
        make_ac_device(),
        "demo",
        repository=repository,
        paths=PathResolver(tmp_path),
    )

    assert result.success is False
    assert result.fit_ap_snapshot_status == "FAILED"
    assert [row["ap_uuid"] for row in repository.list_fit_ap_resources(ac_uuid)] == ["ap-keep"]


def test_h3c_ac_resource_partial_rows_with_required_command_failure_preserves_current(
    monkeypatch, tmp_path
):
    connection = FakeConnection(
        {"display wlan ap all address": RuntimeError("address command failed")}
    )
    monkeypatch.setattr(
        h3c_ac_collect_service.netmiko_connection,
        "ConnectHandler",
        lambda **_kwargs: connection,
    )
    repository = AcRepository(make_database(tmp_path))
    ac_uuid = "22222222-2222-4222-8222-222222222222"
    repository.replace_fit_ap_resources(
        ac_uuid,
        [
            {"ap_uuid": "ap-sn001", "ap_name": "AP-SN001", "serial_number": "SN001"},
            {"ap_uuid": "ap-sn002", "ap_name": "AP-SN002", "serial_number": "SN002"},
        ],
    )

    result = collect_h3c_fit_ap_resources(
        make_ac_device(),
        "demo",
        repository=repository,
        paths=PathResolver(tmp_path),
    )

    current = repository.list_fit_ap_resources(ac_uuid)
    assert parse_wlan_ap_list(connection.output_history["display wlan ap all"])
    assert result.success is False
    assert result.fit_ap_snapshot_status == "FAILED"
    assert result.fit_ap_resources_updated == 0
    assert "display wlan ap all address" in [
        item.command for item in result.command_results if not item.success
    ]
    assert sorted(row["serial_number"] for row in current) == ["SN001", "SN002"]


def test_h3c_ac_resource_invalid_snapshot_is_failed_without_empty_replace(
    monkeypatch, tmp_path
):
    connection = FakeConnection(
        {
            "display wlan ap all": "AP output cannot be parsed\n",
            "display wlan ap all address": "",
            "display wlan ap all radio": "",
        }
    )
    monkeypatch.setattr(
        h3c_ac_collect_service.netmiko_connection,
        "ConnectHandler",
        lambda **_kwargs: connection,
    )
    repository = AcRepository(make_database(tmp_path))
    ac_uuid = "22222222-2222-4222-8222-222222222222"
    repository.replace_fit_ap_resources(
        ac_uuid,
        [{"ap_uuid": "ap-old", "ap_name": "AP-OLD", "serial_number": "SN-OLD"}],
    )

    result = collect_h3c_fit_ap_resources(
        make_ac_device(),
        "demo",
        repository=repository,
        paths=PathResolver(tmp_path),
    )

    assert result.success is False
    assert result.fit_ap_snapshot_status == "FAILED"
    assert [row["serial_number"] for row in repository.list_fit_ap_resources(ac_uuid)] == [
        "SN-OLD"
    ]


def test_h3c_ac_resource_success_empty_replaces_only_that_ac_current(
    monkeypatch, tmp_path
):
    empty_outputs = {
        "display wlan ap all": "Total number of APs: 0\n",
        "display wlan ap all address": "",
        "display wlan ap all radio": "",
    }
    connection = FakeConnection(empty_outputs)
    monkeypatch.setattr(
        h3c_ac_collect_service.netmiko_connection,
        "ConnectHandler",
        lambda **_kwargs: connection,
    )
    repository = AcRepository(make_database(tmp_path))
    ac_uuid = "22222222-2222-4222-8222-222222222222"
    repository.replace_fit_ap_resources(
        ac_uuid,
        [{"ap_uuid": "ap-old", "ap_name": "AP-OLD", "serial_number": "SN-OLD"}],
    )

    result = collect_h3c_fit_ap_resources(
        make_ac_device(),
        "demo",
        repository=repository,
        paths=PathResolver(tmp_path),
    )

    assert result.success is True
    assert result.fit_ap_snapshot_status == "SUCCESS_EMPTY"
    assert repository.list_fit_ap_resources(ac_uuid) == []


def test_wlan_ap_summary_derives_offline_from_total_minus_online():
    summary = parse_wlan_ap_summary(
        """
Total number of APs: 882
Total number of connected APs: 756
"""
    )

    assert summary["total_aps"] == 882
    assert summary["online_aps"] == 756
    assert summary["offline_aps"] == 126


def test_h3c_ac_collect_service_emits_progress_stages(monkeypatch, tmp_path):
    connection = FakeConnection()
    monkeypatch.setattr(
        h3c_ac_collect_service.netmiko_connection,
        "ConnectHandler",
        lambda **_kwargs: connection,
    )
    messages: list[str] = []

    result = collect_h3c_fit_ap_resources(
        make_ac_device(),
        "demo",
        repository=AcRepository(make_database(tmp_path)),
        paths=PathResolver(tmp_path),
        progress=messages.append,
    )

    assert result.success is True
    assert any("正在连接AC" in message for message in messages)
    assert any("display wlan ap all" in message for message in messages)
    assert any("正在解析FIT-AP资源" in message for message in messages)
    assert any("正在写入数据库" in message for message in messages)
    assert any("正在保存 FIT-AP 主资源" in message for message in messages)
    assert any("FIT-AP 主资源保存完成" in message for message in messages)
    assert any("AP Identity 更新完成" in message for message in messages)
    assert any("更新完成" in message for message in messages)


def test_h3c_ac_collect_service_saves_https_port(monkeypatch, tmp_path):
    connection = FakeConnection()
    monkeypatch.setattr(
        h3c_ac_collect_service.netmiko_connection,
        "ConnectHandler",
        lambda **_kwargs: connection,
    )
    database = make_database(tmp_path)
    device_repository = DeviceRepository(database)
    ac_device = device_repository.create(make_ac_device())

    result = collect_h3c_ac_info(
        ac_device,
        "demo",
        repository=AcRepository(database),
        paths=PathResolver(tmp_path),
    )

    assert result.success is True
    assert result.https_port == 443
    assert result.https_port_collected is True
    assert result.https_port_persisted is True
    assert result.https_port_error is None
    assert device_repository.get(int(ac_device.id)).https_port == 443


def test_h3c_ac_collect_service_saves_non_default_https_port(monkeypatch, tmp_path):
    connection = FakeConnection(
        {
            "display ip https": "<AC>display ip https\nHTTPS port: 10443\nOperation status : Enabled\n<AC>"
        }
    )
    monkeypatch.setattr(
        h3c_ac_collect_service.netmiko_connection,
        "ConnectHandler",
        lambda **_kwargs: connection,
    )
    database = make_database(tmp_path)
    device_repository = DeviceRepository(database)
    ac_device = device_repository.create(make_ac_device())

    result = collect_h3c_ac_info(
        ac_device,
        "demo",
        repository=AcRepository(database),
        paths=PathResolver(tmp_path),
    )

    assert result.success is True
    assert result.https_port == 10443
    assert result.https_port_collected is True
    assert result.https_port_persisted is True
    assert device_repository.get(int(ac_device.id)).https_port == 10443


def test_h3c_ac_collect_service_reports_https_port_save_failure(monkeypatch, tmp_path):
    connection = FakeConnection({"display ip https": "HTTPS port: 10443\n"})
    monkeypatch.setattr(
        h3c_ac_collect_service.netmiko_connection,
        "ConnectHandler",
        lambda **_kwargs: connection,
    )
    monkeypatch.setattr(
        h3c_ac_collect_service.DeviceRepository,
        "update_https_port",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            RuntimeError("no such column: https_port")
        ),
    )
    database = make_database(tmp_path)
    device_repository = DeviceRepository(database)
    ac_device = device_repository.create(make_ac_device())

    result = collect_h3c_ac_info(
        ac_device,
        "demo",
        repository=AcRepository(database),
        paths=PathResolver(tmp_path),
    )

    assert result.success is True
    assert result.https_port == 10443
    assert result.https_port_collected is True
    assert result.https_port_persisted is False
    assert "no such column: https_port" in str(result.https_port_error)
    assert device_repository.get(int(ac_device.id)).https_port is None


def test_https_port_parser_and_url_builder_are_strict():
    assert parse_https_port("HTTPS port: 443") == 443
    assert parse_https_port("HTTPS port: 10443") == 10443
    assert (
        parse_https_port(
            "<AC>display ip https | include port\r\nHTTPS port : 8443\r\n<AC>"
        )
        == 8443
    )
    assert parse_https_port("\x1b[24D HTTPS port：443") == 443
    assert parse_https_port("HTTP port: 80\nSSH server port: 22") is None
    assert parse_https_port("HTTPS port: 70000") is None
    assert build_https_url("10.122.100.10", 443) == "https://10.122.100.10:443"
    assert build_https_url("2001:db8::10", 443) == "https://[2001:db8::10]:443"
    assert build_https_url("", 443) is None


def test_h3c_ac_collect_service_falls_back_to_full_https_command(monkeypatch, tmp_path):
    connection = FakeConnection(
        {
            "display ip https": "",
            "display ip https | include port": "<AC>display ip https | include port\nHTTPS port: 8443\n<AC>",
        }
    )
    monkeypatch.setattr(
        h3c_ac_collect_service.netmiko_connection,
        "ConnectHandler",
        lambda **_kwargs: connection,
    )
    database = make_database(tmp_path)
    device_repository = DeviceRepository(database)
    ac_device = device_repository.create(make_ac_device())

    result = collect_h3c_ac_info(
        ac_device,
        "demo",
        repository=AcRepository(database),
        paths=PathResolver(tmp_path),
    )

    assert result.success is True
    assert result.https_port == 8443
    assert connection.commands[-2:] == [
        "display ip https",
        "display ip https | include port",
    ]
    assert device_repository.get(int(ac_device.id)).https_port == 8443


def test_h3c_ac_collect_service_keeps_existing_https_port_on_collect_failure(
    monkeypatch, tmp_path
):
    connection = FakeConnection(
        {
            "display ip https | include port": RuntimeError("unsupported"),
            "display ip https": "",
        }
    )
    monkeypatch.setattr(
        h3c_ac_collect_service.netmiko_connection,
        "ConnectHandler",
        lambda **_kwargs: connection,
    )
    database = make_database(tmp_path)
    device_repository = DeviceRepository(database)
    ac_device = device_repository.create(make_ac_device())
    device_repository.update_https_port(int(ac_device.id), 443)

    result = collect_h3c_ac_info(
        device_repository.get(int(ac_device.id)),
        "demo",
        repository=AcRepository(database),
        paths=PathResolver(tmp_path),
    )

    assert result.success is True
    assert result.https_port is None
    assert device_repository.get(int(ac_device.id)).https_port == 443


def test_h3c_ac_collect_service_validates_commands_before_execution(
    monkeypatch, tmp_path
):
    calls = []
    connection = FakeConnection()
    monkeypatch.setattr(
        h3c_ac_collect_service.netmiko_connection,
        "ConnectHandler",
        lambda **_kwargs: connection,
    )
    monkeypatch.setattr(
        h3c_ac_collect_service.command_guard,
        "validate_command_list",
        lambda commands, context: calls.append((list(commands), context)),
    )
    database = make_database(tmp_path)
    repository = AcRepository(database)

    collect_h3c_fit_ap_resources(
        make_ac_device(), "demo", repository=repository, paths=PathResolver(tmp_path)
    )

    assert calls == [
        (["screen-length disable", *RESOURCE_COMMANDS], "ac_fit_ap_resource_collect")
    ]


def test_build_new_online_ap_overview_rows_uses_current_online_without_prior_resource_history():
    current_resources = [
        {
            "collect_run_uuid": "run-2",
            "collected_at": "2026-06-30 10:00:00",
            "site": "Station A",
            "ap_name": "AP-New",
            "ap_mac": "0011-2233-4455",
            "serial_number": "SN-NEW",
            "state": "R/M",
        },
        {
            "collect_run_uuid": "run-2",
            "collected_at": "2026-06-30 10:00:00",
            "site": "Station A",
            "ap_name": "AP-Idle",
            "ap_mac": "0011-2233-4466",
            "serial_number": "SN-IDLE",
            "state": "Idle",
        },
        {
            "collect_run_uuid": "run-2",
            "collected_at": "2026-06-30 10:00:00",
            "site": "Station A",
            "ap_name": "AP-Old-Apid-Changed",
            "ap_mac": "0011-2233-4477",
            "serial_number": "SN-OLD",
            "state": "R/M",
            "apid": "999",
        },
    ]
    history_rows = [
        {
            "collect_run_uuid": "run-1",
            "collected_at": "2026-06-29 10:00:00",
            "ap_name": "AP-Old",
            "ap_mac": "0011-2233-4477",
            "serial_number": "SN-OLD",
            "apid": "1",
        }
    ]
    trackside_rows = [
        {
            "site": "Station A",
            "ap_name": "AP-New",
            "ap_mac": "0011-2233-4455",
            "serial_number": "SN-NEW",
            "device_name": "SW-1",
            "interface_name": "GigabitEthernet1/0/1",
        }
    ]

    rows = build_new_online_ap_overview_rows(
        current_resources,
        history_rows,
        trackside_rows,
        unauthenticated_rows=[
            {
                "site_key": "Station A",
                "ap_name": "AP-New",
                "ap_mac": "0011-2233-4455",
                "serial_number": "SN-NEW",
                "source": WLAN_AP_UNAUTHENTICATED_SOURCE,
                "collected_at": "2026-06-30 10:00:00",
            }
        ],
    )

    assert [row["ap_name"] for row in rows] == ["AP-New"]
    assert rows[0]["device_name"] == "SW-1"
    assert rows[0]["interface_name"] == "GigabitEthernet1/0/1"
    assert [field for _key, field in NEW_ONLINE_AP_OVERVIEW_COLUMNS][:18] == [
        "site",
        "device_name",
        "interface_name",
        "link_status",
        "port_type",
        "description",
        "pvid",
        "vlan",
        "switch_rx_power",
        "switch_optical_status",
        "ap_mac",
        "ap_name",
        "ap_rx_power",
        "ap_device_optical_status",
        "ap_optical_status",
        "ap_business_threshold_dbm",
        "ap_business_reason",
        "updated_at",
    ]
    assert [field for _key, field in NEW_ONLINE_AP_OVERVIEW_COLUMNS][-1] == "suggestion"


def test_build_ap_optical_treatment_records_closes_and_opens_by_history():
    trackside_rows = [
        {
            "site": "Station A",
            "ap_uuid": "ap-1",
            "ap_name": "AP-1",
            "ap_mac": "0011-2233-4455",
            "serial_number": "SN-1",
            "device_uuid": "sw-1",
            "device_name": "SW-1",
            "interface_name": "GigabitEthernet1/0/1",
            "ap_rx_power": "-8.00",
            "ap_optical_status": "normal",
            "switch_rx_power": "-21.00",
            "switch_optical_status": "warning",
            "updated_at": "2026-06-30 10:00:00",
        }
    ]
    ap_history = [
        {
            "id": 1,
            "ap_uuid": "ap-1",
            "rx_power": "-22.00",
            "optical_alarm_status": "warning",
            "collected_at": "2026-06-30 09:00:00",
        },
        {
            "id": 2,
            "ap_uuid": "ap-1",
            "rx_power": "-8.00",
            "optical_alarm_status": "normal",
            "collected_at": "2026-06-30 09:30:00",
        },
        {
            "id": 3,
            "ap_uuid": "ap-2",
            "rx_power": "-24.00",
            "optical_alarm_status": "warning",
            "collected_at": "2026-06-30 09:00:00",
        },
    ]
    switch_history = [
        {
            "id": 1,
            "device_uuid": "sw-1",
            "interface_name": "GE1/0/1",
            "rx_power": "-20.00",
            "optical_alarm_status": "failed",
            "collected_at": "2026-06-30 08:00:00",
        }
    ]

    records = build_ap_optical_treatment_records(
        trackside_rows, ap_history, switch_history
    )

    assert [field for _key, field in AP_OPTICAL_TREATMENT_RECORD_COLUMNS][0] == "site"
    assert len(records) == 2
    closed = next(row for row in records if row["side"] == "AP侧")
    open_record = next(row for row in records if row["side"] == "交换机侧")
    assert closed["treatment_status"] == TREATMENT_CLOSED_LABEL
    assert closed["completed_at"] == "2026-06-30 09:30:00"
    assert open_record["treatment_status"] == TREATMENT_OPEN_LABEL
    assert open_record["first_found_at"] == "2026-06-30 10:00:00"


def test_trackside_ap_business_treatment_records_complete_ap_identity_and_normalize_interfaces():
    trackside_rows = [
        {
            "site": "Station A",
            "ap_name": "AP-GE",
            "ap_mac": "0011-2233-4455",
            "device_uuid": "sw-1",
            "device_name": "SW-1",
            "interface_name": "GigabitEthernet2/0/1",
            "switch_rx_power": "-8.00",
            "switch_optical_status": "warning",
            "updated_at": "2026-06-30 10:00:00",
        },
        {
            "site": "Station B",
            "ap_name": "AP-BAGG",
            "ap_mac": "00aa-bbcc-ddee",
            "device_uuid": "sw-2",
            "device_name": "SW-2",
            "interface_name": "Bridge-Aggregation121",
            "switch_rx_power": "-8.00",
            "switch_optical_status": "warning",
            "updated_at": "2026-06-30 10:00:00",
        },
    ]
    ap_history = [
        {
            "id": 1,
            "ap_name": "AP-GE",
            "ap_mac": "0011.2233.4455",
            "site": "Station A",
            "rx_power": "-24.00",
            "optical_alarm_status": "warning",
            "collected_at": "2026-06-30 09:00:00",
        }
    ]
    switch_history = [
        {
            "id": 1,
            "device_uuid": "sw-1",
            "interface_name": "GE2/0/1",
            "rx_power": "-24.00",
            "optical_alarm_status": "normal",
            "collected_at": "2026-06-30 09:00:00",
        },
        {
            "id": 2,
            "device_uuid": "sw-2",
            "interface_name": "BAGG121",
            "rx_power": "-24.00",
            "optical_alarm_status": "normal",
            "collected_at": "2026-06-30 09:00:00",
        },
    ]
    resources = [
            {"ap_name": "AP-GE", "ap_mac": "0011-2233-4455", "serial_number": "SN-GE", "site": "Station A"},
            {"ap_name": "AP-BAGG", "ap_mac": "00aa-bbcc-ddee", "serial_number": "SN-BAGG", "site": "Station B"},
    ]

    records = build_ap_optical_treatment_records(
        trackside_rows, ap_history, switch_history, resources
    )

    switch_records = [row for row in records if row["side"] == "交换机侧"]
    assert {row["interface_name"] for row in switch_records} == {
        "GigabitEthernet2/0/1",
        "Bridge-Aggregation121",
    }
    assert {row["serial_number"] for row in switch_records} == {"SN-GE", "SN-BAGG"}
    ap_record = next(row for row in records if row["side"] == "AP侧")
    assert ap_record["ap_name"] == "AP-GE"
    assert ap_record["ap_mac"] == "0011-2233-4455"
    assert ap_record["serial_number"] == "SN-GE"


def test_trackside_ap_business_treatment_records_complete_identity_by_serial_only():
    trackside_rows = [
        {
            "site": "Station A",
            "ap_name": "-",
            "ap_mac": "-",
            "serial_number": "SN-ONLY",
            "device_uuid": "sw-1",
            "device_name": "SW-1",
            "interface_name": "GigabitEthernet2/0/1",
            "switch_rx_power": "-24.00",
            "switch_optical_status": "warning",
            "updated_at": "2026-06-30 10:00:00",
        }
    ]
    resources = [
        {
            "serial_number": "SN-ONLY",
            "ap_name": "AP-SERIAL",
            "ap_mac": "083b.e9ec.da40",
            "site": "Station A",
        }
    ]

    records = build_ap_optical_treatment_records(trackside_rows, [], [], resources)

    assert len(records) == 1
    assert records[0]["ap_name"] == "AP-SERIAL"
    assert records[0]["ap_mac"] == "083b-e9ec-da40"


def test_trackside_ap_business_treatment_records_do_not_guess_mac_or_serial_from_name():
    trackside_rows = [
        {
            "site": "Station A",
            "ap_name": "30f5-277a-0ea0",
            "ap_mac": "",
            "serial_number": "",
            "device_uuid": "sw-1",
            "device_name": "SW-1",
            "interface_name": "GigabitEthernet2/0/1",
            "switch_rx_power": "-24.00",
            "switch_optical_status": "warning",
            "updated_at": "2026-06-30 10:00:00",
        }
    ]

    records = build_ap_optical_treatment_records(trackside_rows, [], [], [])

    assert records[0]["ap_name"] == "30f5-277a-0ea0"
    assert records[0]["ap_mac"] == ""
    assert records[0]["serial_number"] == ""


def test_trackside_ap_business_treatment_records_use_offline_ledger_by_serial():
    trackside_rows = [
        {
            "site": "Station A",
            "ap_name": "-",
            "ap_mac": "-",
            "serial_number": "SN-OFFLINE",
            "device_uuid": "sw-1",
            "device_name": "SW-1",
            "interface_name": "GigabitEthernet2/0/1",
            "switch_rx_power": "-24.00",
            "switch_optical_status": "warning",
            "updated_at": "2026-06-30 10:00:00",
        }
    ]
    offline_ledger_rows = [
        {
            "site": "Station A",
            "ap_name": "AP-OFFLINE",
            "ap_mac": "083b.e9ec.da40",
            "serial_number": "SN-OFFLINE",
        }
    ]

    records = build_ap_optical_treatment_records(
        trackside_rows, [], [], [], [], offline_ledger_rows=offline_ledger_rows
    )

    assert records[0]["ap_name"] == "AP-OFFLINE"
    assert records[0]["ap_mac"] == "083b-e9ec-da40"


def test_trackside_ap_business_treatment_records_use_offline_ledger_by_switch_interface():
    trackside_rows = [
        {
            "site": "Station A",
            "ap_name": "-",
            "ap_mac": "-",
            "serial_number": "",
            "device_uuid": "sw-1",
            "device_name": "SW-1",
            "interface_name": "GigabitEthernet2/0/1",
            "switch_rx_power": "-24.00",
            "switch_optical_status": "warning",
            "updated_at": "2026-06-30 10:00:00",
        }
    ]
    offline_ledger_rows = [
        {
            "site": "Station A",
            "ap_name": "AP-PORT",
            "ap_mac": "083be9ecda40",
            "serial_number": "SN-PORT",
            "historical_switch_name": "SW-1",
            "historical_switch_interface": "GE2/0/1",
        }
    ]

    records = build_ap_optical_treatment_records(
        trackside_rows, [], [], [], [], offline_ledger_rows=offline_ledger_rows
    )

    assert records[0]["ap_name"] == "AP-PORT"
    assert records[0]["ap_mac"] == "083b-e9ec-da40"
    # Offline LLDP/interface history may recover display identity, but it is
    # not an AP Identity/FIT-AP current source for the canonical serial.
    assert records[0]["serial_number"] == ""


def test_trackside_ap_business_treatment_records_leave_serial_blank_when_identity_is_ambiguous_or_unmatched():
    trackside_rows = [
        {
            "site": "Station A",
            "ap_name": "AP-1",
            "ap_mac": "0011-2233-4455",
            "serial_number": "",
            "device_uuid": "sw-1",
            "device_name": "SW-1",
            "interface_name": "GigabitEthernet2/0/1",
            "switch_rx_power": "-24.00",
            "switch_optical_status": "warning",
            "updated_at": "2026-06-30 10:00:00",
        },
        {
            "site": "Station B",
            "ap_name": "AP-2",
            "ap_mac": "00aa-bbcc-ddee",
            "serial_number": "",
            "device_uuid": "sw-2",
            "device_name": "SW-2",
            "interface_name": "GigabitEthernet2/0/2",
            "switch_rx_power": "-24.00",
            "switch_optical_status": "warning",
            "updated_at": "2026-06-30 10:00:00",
        },
    ]
    resources = [
        {
            "site": "Station A",
            "ap_mac": "0011-2233-4455",
            "serial_number": "SN-A",
        },
        {
            "site": "Station B",
            "ap_mac": "00aa-bbcc-ddee",
            "serial_number": "SN-B1",
        },
        {
            "site": "Station B",
            "ap_mac": "00aa-bbcc-ddee",
            "serial_number": "SN-B2",
        },
    ]

    records = build_ap_optical_treatment_records(trackside_rows, [], [], resources)

    by_site = {row["site"]: row for row in records}
    assert by_site["Station A"]["serial_number"] == "SN-A"
    assert by_site["Station B"]["serial_number"] == ""


def test_trackside_ap_business_treatment_records_ignore_unmatched_switch_history():
    records = build_ap_optical_treatment_records(
        [],
        [],
        [
            {
                "device_uuid": "sw-ordinary",
                "interface_name": "GE1/0/1",
                "rx_power": "-24",
                "optical_alarm_status": "warning",
            }
        ],
        [],
        [],
        offline_ledger_rows=[],
    )

    assert records == []


def test_trackside_export_omits_ap_port_change_columns_and_sheet(tmp_path):
    from openpyxl import load_workbook

    export_path = tmp_path / "trackside_without_ap_port_change.xlsx"
    i18n = I18n("zh_CN")
    rows = [
        {
            "site": "Station A",
            "device_name": "SW-1",
            "interface_name": "GigabitEthernet2/0/6",
            "link_status": "UP",
            "switch_rx_power": "-8",
            "switch_optical_status": "normal",
            "ap_mac": "0011-2233-4455",
            "ap_name": "AP-1",
            "ap_rx_power": "-8",
            "ap_optical_status": "normal",
            "ap_port_change": "N/A GigabitEthernet2/0/6 -> SW-1 GigabitEthernet2/0/6",
            "ap_port_change_reason": "交换机变化",
            "previous_switch": "N/A",
            "previous_interface": "GigabitEthernet2/0/6",
            "current_switch": "SW-1",
            "current_interface": "GigabitEthernet2/0/6",
            "history_compared_at": "2026-07-04T15:59:07",
        }
    ]

    export_trackside_ap_business_xlsx(
        export_path,
        rows,
        TRACKSIDE_AP_BUSINESS_EXPORT_COLUMNS,
        [i18n.t(key) for key, _field in TRACKSIDE_AP_BUSINESS_EXPORT_COLUMNS],
    )

    workbook = load_workbook(export_path)
    assert "AP端口变化" not in workbook.sheetnames
    headers = [cell.value for cell in workbook["轨旁AP业务"][1]]
    for forbidden in (
        "AP端口变化",
        "AP端口变化原因",
        "上次交换机",
        "上次端口",
        "本次交换机",
        "本次端口",
        "历史对比时间",
    ):
        assert forbidden not in headers


def test_trackside_export_sorts_main_sheet_by_switch_then_interface(tmp_path):
    from openpyxl import load_workbook

    export_path = tmp_path / "trackside_station_order.xlsx"
    i18n = I18n("zh_CN")
    rows = [
        {"site": "10站", "device_name": "SW-10", "interface_name": "GE1/0/2"},
        {"site": "2站", "device_name": "SW-2", "interface_name": "GE1/0/10"},
        {"site": "2站", "device_name": "SW-2", "interface_name": "GE1/0/2"},
        {"site": "01站", "device_name": "SW-1", "interface_name": "GE1/0/1"},
    ]

    export_trackside_ap_business_xlsx(
        export_path,
        rows,
        TRACKSIDE_AP_BUSINESS_EXPORT_COLUMNS,
        [i18n.t(key) for key, _field in TRACKSIDE_AP_BUSINESS_EXPORT_COLUMNS],
    )

    workbook = load_workbook(export_path, read_only=True)
    sheet = workbook["轨旁AP业务"]
    assert [
        (row[1], row[2])
        for row in sheet.iter_rows(min_row=2, values_only=True)
    ] == [
        ("SW-1", "GE1/0/1"),
        ("SW-2", "GE1/0/2"),
        ("SW-2", "GE1/0/10"),
        ("SW-10", "GE1/0/2"),
    ]


def test_trackside_ap_business_export_adds_current_optical_abnormal_sheet(tmp_path):
    from openpyxl import load_workbook

    export_path = tmp_path / "trackside_current_abnormal.xlsx"
    i18n = I18n("zh_CN")
    rows = [
        {
            "site": "Station A",
            "device_name": "SW-1",
            "interface_name": "GigabitEthernet1/0/1",
            "link_status": "UP",
            "switch_rx_power": "-8",
            "switch_optical_status": "normal",
            "ap_name": "AP-Normal",
            "ap_rx_power": "-8",
            "ap_optical_status": "normal",
        },
        {
            "site": "Station A",
            "device_name": "SW-1",
            "interface_name": "GigabitEthernet1/0/2",
            "link_status": "UP",
            "switch_rx_power": "-19.10",
            "switch_optical_status": "normal",
            "ap_name": "AP-Warning",
            "ap_state": "R/M",
            "ap_rx_power": "-8",
            "ap_optical_status": "normal",
        },
        {
            "site": "Station B",
            "device_name": "SW-2",
            "interface_name": "GigabitEthernet1/0/3",
            "link_status": "UP",
            "switch_rx_power": "-8",
            "switch_optical_status": "normal",
            "ap_name": "AP-NoLight",
            "ap_state": "R/B",
            "ap_rx_power": "-40",
            "ap_optical_status": "no_light",
            "ap_side_has_data": True,
        },
        {
            "site": "Station C",
            "device_name": "SW-3",
            "interface_name": "GigabitEthernet1/0/4",
            "link_status": "DOWN",
            "switch_rx_power": "-36.96",
            "switch_optical_status": "no_light",
            "ap_name": "AP-Offline",
            "ap_mac": "30f5-2787-91c0",
            "ap_rx_power": "-7.99",
            "ap_optical_status": "offline",
            "is_ap_offline": True,
        },
        {
            "site": "Station D",
            "device_name": "SW-4",
            "interface_name": "GigabitEthernet1/0/5",
            "link_status": "DOWN",
            "switch_rx_power": "-36.96",
            "switch_optical_status": "no_light",
            "ap_mac": "-",
            "ap_name": "-",
            "ap_rx_power": "-",
            "ap_optical_status": "-",
        },
    ]

    export_trackside_ap_business_xlsx(
        export_path,
        rows,
        TRACKSIDE_AP_BUSINESS_COLUMNS,
        [i18n.t(key) for key, _field in TRACKSIDE_AP_BUSINESS_COLUMNS],
    )

    workbook = load_workbook(export_path)
    assert workbook.sheetnames[:3] == ["AP上线情况概览", "轨旁AP业务", "当前异常光衰"]
    source_sheet = workbook["轨旁AP业务"]
    abnormal_sheet = workbook["当前异常光衰"]
    abnormal_headers = [cell.value for cell in abnormal_sheet[1]]
    assert abnormal_headers == [
        i18n.t(key) for key, _field in CURRENT_OPTICAL_ABNORMAL_COLUMNS
    ]
    assert [
        abnormal_sheet.cell(row=row, column=3).value
        for row in range(2, abnormal_sheet.max_row + 1)
    ] == ["GE1/0/2", "GE1/0/3", "GE1/0/4"]
    reason_column = abnormal_headers.index("异常原因") + 1
    assert abnormal_sheet.cell(row=3, column=reason_column).value == "AP侧业务光衰异常"
    detail_column = abnormal_headers.index("异常说明") + 1
    assert "AP 侧收光无光：-40.00 dBm" in abnormal_sheet.cell(
        row=3,
        column=detail_column,
    ).value
    switch_status_column = next(
        index
        for index, (_key, field) in enumerate(TRACKSIDE_AP_BUSINESS_COLUMNS, start=1)
        if field == "switch_optical_status"
    )
    assert source_sheet.cell(row=3, column=switch_status_column).value == "光衰大"
    assert abnormal_sheet.cell(row=4, column=reason_column).value == "交换机侧业务光衰异常"
    online_status_column = abnormal_headers.index("AP 在线状态") + 1
    assert [
        abnormal_sheet.cell(row=row, column=online_status_column).value
        for row in range(2, abnormal_sheet.max_row + 1)
    ] == ["在线", "在线", "离线"]
    assert (
        abnormal_sheet["A2"].fill.fgColor.rgb
        == source_sheet["A3"].fill.fgColor.rgb
        == "FFFEE2E2"
    )
    assert abnormal_sheet["A1"].font.bold
    assert abnormal_sheet.freeze_panes == "A2"
    assert abnormal_sheet.auto_filter.ref == abnormal_sheet.dimensions
    assert (
        abnormal_sheet.column_dimensions["A"].width
        == source_sheet.column_dimensions["A"].width
    )


def test_trackside_ap_business_export_empty_current_optical_abnormal_sheet(tmp_path):
    from openpyxl import load_workbook

    export_path = tmp_path / "trackside_no_current_abnormal.xlsx"
    i18n = I18n("zh_CN")

    export_trackside_ap_business_xlsx(
        export_path,
        [
            {
                "site": "Station A",
                "device_name": "SW-1",
                "interface_name": "GigabitEthernet1/0/1",
                "link_status": "UP",
                "switch_rx_power": "-8",
                "switch_optical_status": "normal",
                "ap_name": "AP-Normal",
                "ap_rx_power": "-8",
                "ap_optical_status": "normal",
            },
            {
                "site": "Station B",
                "device_name": "SW-2",
                "interface_name": "GigabitEthernet1/0/2",
                "link_status": "DOWN",
                "switch_rx_power": "-36.96",
                "switch_optical_status": "no_light",
                "ap_name": "-",
                "ap_rx_power": "-",
                "ap_optical_status": "-",
            },
        ],
        TRACKSIDE_AP_BUSINESS_COLUMNS,
        [i18n.t(key) for key, _field in TRACKSIDE_AP_BUSINESS_COLUMNS],
    )

    sheet = load_workbook(export_path)["当前异常光衰"]
    assert sheet.max_row == 2
    assert sheet["A2"].value == "当前无异常光衰（已排除无 AP 绑定、无光模块和非告警光功率）"


def test_current_optical_abnormal_is_independent_from_ap_online_state():
    assert is_current_optical_abnormal_row(
        {
            "link_status": "DOWN",
            "switch_rx_power": "-36.96",
            "switch_optical_status": "no_light",
            "ap_mac": "30f5-2787-91c0",
            "ap_name": "30f5-2787-91c0",
            "ap_rx_power": "-7.99",
            "ap_optical_status": "offline",
        }
    )


def test_current_optical_abnormal_keeps_stale_latest_valid_observation():
    stale_abnormal = {
        "ap_mac": "30f5-2787-ab01",
        "ap_name": "AP-stale-abnormal",
        "ap_rx_power": "-26.99",
        "ap_optical_status": "normal",
        "data_freshness": "stale",
        "updated_at": "2026-08-06T17:45:20",
    }
    stale_normal = {
        **stale_abnormal,
        "ap_mac": "30f5-2787-ab02",
        "ap_rx_power": "-7.10",
    }

    assert is_current_optical_abnormal_row(stale_abnormal)
    assert not is_current_optical_abnormal_row(stale_normal)


def test_current_optical_abnormal_does_not_consume_retired_history():
    base = {
        "ap_uuid": "ap-cache",
        "ap_mac": "30f5-2787-ab03",
        "ap_name": "AP-cache",
        "ap_rx_power": "-",
        "ap_optical_status": "not_collected",
        "collection_status": "timeout",
        "updated_at": "2026-08-10T09:00:00",
    }
    old_abnormal = {
        "ap_uuid": "ap-cache",
        "rx_power": "-26.99",
        "collected_at": "2026-08-06T17:45:20",
        "status": "success",
    }
    newer_normal = {
        **old_abnormal,
        "rx_power": "-7.10",
        "collected_at": "2026-08-09T17:45:20",
    }

    cached = enrich_trackside_export_rows(
        [base], ap_optical_history_rows=[old_abnormal, newer_normal]
    )[0]
    assert cached == base
    assert not is_current_optical_abnormal_row(cached)


def test_current_optical_abnormal_does_not_reuse_cache_after_explicit_no_module():
    row = {
        "ap_uuid": "ap-no-module",
        "ap_mac": "30f5-2787-ab04",
        "ap_name": "AP-no-module",
        "ap_rx_power": "-",
        "switch_rx_power": "-",
        "switch_optical_status": "no_module",
        "raw_status": "no module",
    }
    history = [{"ap_uuid": "ap-no-module", "rx_power": "-26.99", "collected_at": "2026-08-06T17:45:20"}]

    enriched = enrich_trackside_export_rows([row], ap_optical_history_rows=history)[0]
    assert enriched == row
    assert not is_current_optical_abnormal_row(enriched)
    assert not is_current_optical_abnormal_row(
        {
            "ap_mac": "30f5-2787-afc1",
            "ap_name": "30f5-2787-afc1",
            "ap_optical_status": "alarm",
            "updated_at": "2020-01-01T00:00:00+00:00",
        }
    )
    assert not is_current_optical_abnormal_row(
        {
            "ap_mac": "30f5-2787-91c1",
            "ap_name": "30f5-2787-91c1",
            "ap_optical_status": "offline",
            "switch_optical_status": "normal",
        }
    )
    assert not is_current_optical_abnormal_row(
        {
            "link_status": "DOWN",
            "switch_rx_power": "-36.96",
            "switch_optical_status": "no_light",
            "ap_mac": "-",
            "ap_name": "-",
            "ap_rx_power": "-",
            "ap_optical_status": "-",
        }
    )
    assert is_current_optical_abnormal_row(
        {
            "link_status": "DOWN",
            "switch_rx_power": "-36.96",
            "switch_optical_status": "no_light",
            "ap_mac": "30f5-xxxx-yyyy",
            "ap_name": "30f5-xxxx-yyyy",
            "ap_optical_status": "normal",
        }
    )
    assert is_current_optical_abnormal_row(
        {
            "ap_mac": "30f5-2787-afc0",
            "ap_name": "30f5-2787-afc0",
            "ap_rx_power": "-22.01",
            "ap_optical_status": "alarm",
            "ap_side_has_data": True,
        }
    )


def test_enrich_fit_ap_optical_rows_adds_ap_mac_and_station_from_resources():
    rows = [{"ap_name": "AP-A", "site": None, "optical_alarm_status": "normal"}]
    resources = [{"ap_name": "AP-A", "ap_mac": "0011-2233-4455", "site": "Station A"}]

    enriched = enrich_fit_ap_optical_rows(rows, resources)

    assert enriched[0]["ap_mac"] == "0011-2233-4455"
    assert enriched[0]["site"] == "Station A"


def test_enrich_fit_ap_optical_rows_uses_unassigned_and_filters_invalid_neighbor():
    rows = [
        {"ap_name": "AP-A", "neighbor_device_name": "* -- -- Nearest customer bridge"}
    ]

    enriched = enrich_fit_ap_optical_rows(rows, [])

    assert enriched[0]["site"] == "未归属"
    assert enriched[0]["neighbor_device_name"] is None


def test_display_optical_status_uses_chinese_labels():
    assert display_optical_status("normal") == "正常"
    assert display_optical_status("warning") == "提示告警"
    assert display_optical_status("alarm") == "一般告警"
    assert display_optical_status("link_abnormal") == "链路异常"
    assert display_optical_status("link_down") == "链路断开"
    assert display_optical_status("no_light") == "无光"
    assert display_optical_status("skipped") == "未检查"
    assert display_optical_status("not_collected") == "未采集"
    assert display_optical_status("unknown") == "未知"


def test_optical_color_legend_does_not_expose_threshold_rules():
    assert "RX警告下限" not in I18n("zh_CN").t("details.optical_color_legend")
    assert "RX low warning" not in I18n("en_US").t("details.optical_color_legend")


def test_ap_online_overview_rows_count_states_and_total_bottom():
    rows = [
        {
            "ap_name": "AP-2-1",
            "site": "02云龙火车站站",
            "state": "R/M",
            "connection_reonline_count": 10,
        },
        {"ap_name": "AP-2-2", "site": "02云龙火车站站", "state": "R/B"},
        {"ap_name": "AP-1-1", "site": "01小洋江站", "state": "I"},
        {"ap_name": "AP-1-2", "site": "01小洋江站", "state": "JA"},
        {"ap_name": "AP-1-3", "site": "01小洋江站", "state": "R/M"},
    ]

    overview = build_ap_online_overview_rows(planned_aps=rows, fit_ap_resources=rows)

    assert overview[0] == {
        "site": "01小洋江站",
        "total": 3,
        "online": 1,
        "offline": 2,
        "optical_problem_count": 0,
        "remark": "",
        "online_rate": "33.3%",
    }
    assert overview[1] == {
        "site": "02云龙火车站站",
        "total": 2,
        "online": 2,
        "offline": 0,
        "optical_problem_count": 0,
        "remark": "",
        "online_rate": "100.0%",
    }
    assert overview[-1] == {
        "site": "\u5408\u8ba1",
        "total": 5,
        "online": 3,
        "offline": 2,
        "optical_problem_count": 0,
        "remark": "",
        "online_rate": "60.0%",
    }


def test_ap_online_overview_uses_fit_ap_resource_site_capacity_and_unassigned():
    resources = [
        {"ap_uuid": "ap-1", "site_name": "Metadata Station", "state": "R/M"},
        {"ap_uuid": "ap-2", "site": "", "state": "JA"},
    ]
    overview = build_ap_online_overview_rows(
        planned_aps=[],
        fit_ap_resources=resources,
        capacity_details={
            "Metadata Station": {"ap_total": 5, "remark": "Keep watching"}
        },
    )

    assert [row["site"] for row in overview] == ["Metadata Station", "\u5408\u8ba1"]
    assert overview[0]["total"] == 5
    assert overview[0]["online"] == 1
    assert overview[0]["offline"] == 4
    assert overview[0]["remark"] == "Keep watching"
    assert overview[1]["online"] == 1


def test_ap_online_overview_matches_dirty_resource_site_back_to_plan():
    resources = [
        {"ap_mac": "0011-2233-4455", "ap_name": "AP-A", "site": "Demo", "state": "R/M"}
    ]
    plans = [{"AP_MAC": "0011.2233.4455", "ap_name": "AP-A", "station": "01小洋江站"}]

    overview = build_ap_online_overview_rows(
        planned_aps=plans, fit_ap_resources=resources
    )

    assert [row["site"] for row in overview] == ["01小洋江站", "\u5408\u8ba1"]
    assert overview[0]["total"] == 1
    assert overview[0]["online"] == 1
    assert overview[0]["offline"] == 0


def test_ap_online_overview_unmatched_online_does_not_enter_unassigned():
    resources = [
        {"ap_mac": "00aa-bbcc-ddee", "ap_name": "AP-Z", "site": "Demo", "state": "R/M"}
    ]
    plans = [{"ap_mac": "0011-2233-4455", "ap_name": "AP-A", "station": "01小洋江站"}]

    overview = build_ap_online_overview_rows(
        planned_aps=plans, fit_ap_resources=resources
    )

    assert [row["site"] for row in overview] == ["01小洋江站", "\u5408\u8ba1"]
    assert overview[0]["online"] == 0
    assert overview[-1]["online"] == 0
    assert all(row["site"] != "Demo" for row in overview)


def test_ap_online_overview_excludes_bulk_unmatched_when_plan_coverage_is_missing():
    resources = [
        {"ap_mac": f"00aa-bbcc-{index:04x}", "site": "Demo", "state": "R/M"}
        for index in range(20)
    ]
    plans = [{"ap_mac": "0011-2233-4455", "station": "Station A"}]
    capacities = {
        "Station A": {"ap_total": 30, "remark": ""},
        "Station B": {"ap_total": 56, "remark": ""},
    }

    overview = build_ap_online_overview_rows(
        planned_aps=plans, fit_ap_resources=resources, capacity_details=capacities
    )

    assert [row["site"] for row in overview] == [
        "Station A",
        "Station B",
        "\u5408\u8ba1",
    ]
    assert all(row["site"] != "Demo" for row in overview)
    assert overview[-1]["total"] == 86
    assert overview[-1]["online"] == 0


def test_ap_online_overview_uses_ap_metadata_as_total_baseline():
    plan_rows = [
        *(
            {
                "ap_uuid": f"s1-plan-{index}",
                "ap_name": f"S1-AP-{index}",
                "site_name": "01小洋江站",
            }
            for index in range(30)
        ),
        *(
            {
                "ap_uuid": f"s2-plan-{index}",
                "ap_name": f"S2-AP-{index}",
                "site_name": "02云龙火车站站",
            }
            for index in range(56)
        ),
        {
            "ap_uuid": "unknown-plan-0",
            "ap_name": "UNKNOWN-AP-0",
            "site_name": "\u672a\u5f52\u5c5e",
        },
    ]
    resource_rows = [
        *(
            {
                "ap_uuid": f"s1-plan-{index}",
                "ap_name": f"S1-AP-{index}",
                "site": "\u672a\u5f52\u5c5e",
                "state": "R/M",
            }
            for index in range(26)
        ),
        *(
            {
                "ap_uuid": f"s2-plan-{index}",
                "ap_name": f"S2-AP-{index}",
                "site": "\u672a\u5f52\u5c5e",
                "state": "R/M",
            }
            for index in range(48)
        ),
        {
            "ap_uuid": "unknown-plan-0",
            "ap_name": "UNKNOWN-AP-0",
            "site": "\u672a\u5f52\u5c5e",
            "state": "R/M",
        },
    ]

    overview = build_ap_online_overview_rows(
        planned_aps=plan_rows, fit_ap_resources=resource_rows
    )

    assert overview[0] == {
        "site": "01小洋江站",
        "total": 30,
        "online": 26,
        "offline": 4,
        "optical_problem_count": 0,
        "remark": "",
        "online_rate": "86.7%",
    }
    assert overview[1] == {
        "site": "02云龙火车站站",
        "total": 56,
        "online": 48,
        "offline": 8,
        "optical_problem_count": 0,
        "remark": "",
        "online_rate": "85.7%",
    }
    assert overview[2] == {
        "site": "\u672a\u5f52\u5c5e",
        "total": 1,
        "online": 1,
        "offline": 0,
        "optical_problem_count": 0,
        "remark": "",
        "online_rate": "100.0%",
    }
    assert overview[-1] == {
        "site": "\u5408\u8ba1",
        "total": 87,
        "online": 75,
        "offline": 12,
        "optical_problem_count": 0,
        "remark": "",
        "online_rate": "86.2%",
    }


def test_ap_online_overview_does_not_use_fit_ap_resource_count_as_total_when_plan_exists():
    plan_rows = [
        {"ap_uuid": f"plan-{index}", "site_name": "Station A"} for index in range(948)
    ]
    resource_rows = [
        {"ap_uuid": f"plan-{index}", "site": "\u672a\u5f52\u5c5e", "state": "R/M"}
        for index in range(773)
    ]

    overview = build_ap_online_overview_rows(
        planned_aps=plan_rows, fit_ap_resources=resource_rows
    )

    assert overview[0]["site"] == "Station A"
    assert overview[0]["total"] == 948
    assert overview[0]["online"] == 773
    assert overview[0]["offline"] == 175
    assert overview[0]["online_rate"] == "81.5%"
    assert overview[-1]["total"] == 948
    assert overview[-1]["online"] == 773
    assert overview[-1]["offline"] == 175
    assert overview[-1]["online_rate"] == "81.5%"


def test_ap_online_overview_capacity_total_takes_priority_over_incomplete_plan():
    planned_aps = [
        {"ap_name": "AP-A-1", "site_name": "01小洋江站"},
        {"ap_name": "AP-B-1", "site_name": "02云龙火车站"},
    ]
    rows = build_ap_online_overview_rows(
        planned_aps=planned_aps,
        fit_ap_resources=[],
        capacity_details={
            "01小洋江站": {"ap_total": 30, "remark": ""},
            "02云龙火车站": {"ap_total": 56, "remark": ""},
            "08丹城站": {"ap_total": 78, "remark": ""},
            "10大目湾站": {"ap_total": 34, "remark": "大目湾校减-8"},
        },
    )
    by_site = {row["site"]: row for row in rows}

    assert by_site["01小洋江站"]["total"] == 30
    assert by_site["02云龙火车站"]["total"] == 56
    assert by_site["08丹城站"]["total"] == 78
    assert by_site["10大目湾站"]["total"] == 34
    assert by_site["10大目湾站"]["remark"] == "大目湾校减-8"


def test_ap_online_overview_matches_by_known_resource_station_when_metadata_key_missing():
    resources = [
        *(
            {"ap_name": f"UNKNOWN-A-{index}", "site_name": "01小洋江站", "state": "R/M"}
            for index in range(26)
        ),
        *(
            {"ap_name": f"UNKNOWN-B-{index}", "site_name": "02云龙火车站", "state": "R/M"}
            for index in range(48)
        ),
    ]
    rows = build_ap_online_overview_rows(
        planned_aps=[],
        fit_ap_resources=resources,
        capacity_details={
            "01小洋江站": {"ap_total": 30, "remark": ""},
            "02云龙火车站": {"ap_total": 56, "remark": ""},
        },
    )
    by_site = {row["site"]: row for row in rows}

    assert by_site["01小洋江站"]["online"] == 26
    assert by_site["02云龙火车站"]["online"] == 48
    assert "未归属" not in by_site
    assert rows[-1]["total"] == 86
    assert rows[-1]["online"] == 74


def test_ap_online_overview_matches_online_by_metadata_name_even_when_resource_site_is_dirty():
    metadata_rows = [
        *(
            {"ap_name": f"STA-A-{index}", "site_name": "01小洋江站"}
            for index in range(26)
        ),
        *(
            {"ap_name": f"STA-B-{index}", "site_name": "02云龙火车站"}
            for index in range(48)
        ),
    ]
    resources = [
        *(
            {"ap_name": f" STA-A-{index} ", "site_name": "Demo", "state": "R/M"}
            for index in range(26)
        ),
        *(
            {"ap_name": f"STA-B-{index}", "site": "体育中心站", "state": "R/M"}
            for index in range(48)
        ),
    ]
    rows = build_ap_online_overview_rows(
        metadata_rows=metadata_rows,
        fit_ap_resources=resources,
        capacity_details={
            "01小洋江站": {"ap_total": 30, "remark": ""},
            "02云龙火车站": {"ap_total": 56, "remark": ""},
        },
    )
    by_site = {row["site"]: row for row in rows}

    assert by_site["01小洋江站"]["online"] == 26
    assert by_site["02云龙火车站"]["online"] == 48
    assert "Demo" not in by_site
    assert "体育中心站" not in by_site
    assert "未归属" not in by_site


def test_ap_online_overview_metadata_empty_uses_optical_site_by_uuid():
    rows = build_ap_online_overview_rows(
        metadata_rows=[{"ap_uuid": "", "ap_name": "", "site_name": ""}],
        fit_ap_resources=[
            {
                "ap_uuid": "A",
                "ap_name": "AP001",
                "ap_mac": "30f5-277a-82c0",
                "state": "R/M",
            }
        ],
        optical_rows=[
            {
                "ap_uuid": "A",
                "ap_name": "AP001",
                "ap_mac": "30f5-277a-82c0",
                "site": "01小洋江站",
            }
        ],
        capacity_details={"01小洋江站": {"ap_total": 30, "remark": ""}},
    )
    by_site = {row["site"]: row for row in rows}

    assert by_site["01小洋江站"]["online"] == 1
    assert "未归属" not in by_site


def test_ap_online_overview_matches_optical_site_by_mac_without_uuid():
    rows = build_ap_online_overview_rows(
        metadata_rows=[],
        fit_ap_resources=[
            {"ap_name": "AP001", "ap_mac": "30:f5:27:7a:82:c0", "state": "R/M"}
        ],
        optical_rows=[
            {"ap_name": "OTHER", "ap_mac": "30f5-277a-82c0", "site": "01小洋江站"}
        ],
        capacity_details={"01小洋江站": {"ap_total": 30, "remark": ""}},
    )

    assert {row["site"]: row for row in rows}["01小洋江站"]["online"] == 1


def test_ap_online_overview_matches_optical_site_by_name_without_uuid_or_mac():
    rows = build_ap_online_overview_rows(
        metadata_rows=[],
        fit_ap_resources=[{"ap_name": " AP - 001 ", "state": "R/B"}],
        optical_rows=[{"ap_name": "AP-001", "site": "02云龙火车站"}],
        capacity_details={"02云龙火车站": {"ap_total": 56, "remark": ""}},
    )

    assert {row["site"]: row for row in rows}["02云龙火车站"]["online"] == 1


def test_ap_online_overview_resource_dirty_site_does_not_override_optical_site():
    rows = build_ap_online_overview_rows(
        metadata_rows=[],
        fit_ap_resources=[{"ap_mac": "30f5-277a-82c0", "site": "Demo", "state": "R/M"}],
        optical_rows=[{"ap_mac": "30f5-277a-82c0", "site": "01小洋江站"}],
        capacity_details={"01小洋江站": {"ap_total": 30, "remark": ""}},
    )
    by_site = {row["site"]: row for row in rows}

    assert by_site["01小洋江站"]["online"] == 1
    assert "Demo" not in by_site


def test_ap_online_overview_falls_back_to_known_resource_site_when_no_optical_match():
    rows = build_ap_online_overview_rows(
        metadata_rows=[],
        fit_ap_resources=[{"ap_name": "UNKNOWN", "site": "01小洋江站", "state": "R/M"}],
        optical_rows=[],
        capacity_details={"01小洋江站": {"ap_total": 30, "remark": ""}},
    )
    by_site = {row["site"]: row for row in rows}

    assert by_site["01小洋江站"]["online"] == 1
    assert "未归属" not in by_site


def test_ap_online_overview_unknown_dirty_resource_site_is_excluded_from_unassigned():
    rows = build_ap_online_overview_rows(
        metadata_rows=[],
        fit_ap_resources=[{"ap_name": "UNKNOWN", "site": "体育中心站", "state": "R/M"}],
        optical_rows=[],
        capacity_details={"01小洋江站": {"ap_total": 30, "remark": ""}},
    )
    by_site = {row["site"]: row for row in rows}

    assert "未归属" not in by_site
    assert "体育中心站" not in by_site


def test_ap_online_overview_dirty_unknown_station_does_not_create_station_row():
    rows = build_ap_online_overview_rows(
        planned_aps=[],
        fit_ap_resources=[{"ap_name": "UNKNOWN", "site": "Demo", "state": "R/M"}],
        capacity_details={"01小洋江站": {"ap_total": 30, "remark": ""}},
    )
    by_site = {row["site"]: row for row in rows}

    assert "Demo" not in by_site
    assert "未归属" not in by_site


def _make_standard_ap_online_overview_fixture():
    station_totals = {
        "01小洋江站": 30,
        "02云龙火车站": 56,
        "03横溪站": 126,
        "04横溪站": 138,
        "05鄞州咸祥": 134,
        "06象山贤庠": 134,
        "07大徐站": 94,
        "08丹城站": 78,
        "09滨海大道站": 56,
        "10大目湾站": 34,
        "11云龙车辆段": 67,
        "未归属": 1,
    }
    station_online = {
        "01小洋江站": 26,
        "02云龙火车站": 48,
        "03横溪站": 111,
        "04横溪站": 135,
        "05鄞州咸祥": 105,
        "06象山贤庠": 41,
        "07大徐站": 94,
        "08丹城站": 78,
        "09滨海大道站": 48,
        "10大目湾站": 34,
        "11云龙车辆段": 52,
        "未归属": 1,
    }
    planned_aps = []
    resources = []
    optical_rows = []
    for station, total in station_totals.items():
        for index in range(total):
            ap_name = f"{station}-AP-{index:03d}"
            mac = f"30f527{len(planned_aps):06x}"[-12:]
            planned_aps.append(
                {"ap_name": ap_name, "ap_mac": mac, "site_name": station}
            )
            if index < station_online[station]:
                dirty_site = (
                    "Demo"
                    if station == "01小洋江站" and index == 0
                    else (
                        "体育中心站"
                        if station == "02云龙火车站" and index == 0
                        else station
                    )
                )
                resources.append(
                    {
                        "ap_name": f" {ap_name} ",
                        "ap_mac": mac.upper(),
                        "site": dirty_site,
                        "state": "R/M",
                    }
                )
                optical_rows.append(
                    {"ap_name": ap_name, "ap_mac": mac, "site": station}
                )
    return planned_aps, resources, optical_rows


def _standard_ap_capacity_details():
    return {
        "01小洋江站": {"ap_total": 30, "remark": ""},
        "02云龙火车站": {"ap_total": 56, "remark": ""},
        "03横溪站": {"ap_total": 126, "remark": ""},
        "04横溪站": {"ap_total": 138, "remark": ""},
        "05鄞州咸祥": {"ap_total": 134, "remark": ""},
        "06象山贤庠": {"ap_total": 134, "remark": ""},
        "07大徐站": {"ap_total": 94, "remark": ""},
        "08丹城站": {"ap_total": 78, "remark": ""},
        "09滨海大道站": {"ap_total": 56, "remark": ""},
        "10大目湾站": {"ap_total": 34, "remark": "大目湾校减-8"},
        "11云龙车辆段": {"ap_total": 67, "remark": ""},
        "未归属": {"ap_total": 1, "remark": ""},
    }


def test_ap_online_overview_standard_large_sample_matches_expected_totals():
    planned_aps, resources, optical_rows = _make_standard_ap_online_overview_fixture()
    rows = build_ap_online_overview_rows(
        metadata_rows=planned_aps,
        fit_ap_resources=resources,
        optical_rows=optical_rows,
        capacity_details=_standard_ap_capacity_details(),
    )
    by_site = {row["site"]: row for row in rows}

    assert by_site["01小洋江站"]["online"] == 26
    assert by_site["08丹城站"]["total"] == 78
    assert by_site["08丹城站"]["online"] == 78
    assert by_site["10大目湾站"]["total"] == 34
    assert by_site["10大目湾站"]["remark"] == "大目湾校减-8"
    assert "Demo" not in by_site
    assert "体育中心站" not in by_site
    assert rows[-1] == {
        "site": "\u5408\u8ba1",
        "total": 948,
        "online": 773,
        "offline": 175,
        "optical_problem_count": 0,
        "remark": "",
        "online_rate": "81.5%",
    }


def test_ap_online_overview_name_match_without_mac_and_unmatched_is_excluded():
    planned_aps = [{"ap_name": " AP - 001 ", "site_name": "01小洋江站"}]
    resources = [
        {"ap_name": "AP-001", "site": "Demo", "state": "R/M"},
        {"ap_name": "AP-Z", "site": "体育中心站", "state": "R/B"},
    ]
    rows = build_ap_online_overview_rows(
        planned_aps=planned_aps, fit_ap_resources=resources
    )
    by_site = {row["site"]: row for row in rows}

    assert by_site["01小洋江站"]["online"] == 1
    assert "未归属" not in by_site
    assert "Demo" not in by_site
    assert "体育中心站" not in by_site


def test_ap_online_overview_unassigned_only_comes_from_metadata_rows():
    planned_aps = [
        {"ap_mac": "30f5-277a-11a0", "site_name": "未归属", "state": "I"},
        {"ap_mac": "30f5-277a-1e00", "site_name": "未归属", "state": "I"},
    ]
    resources = [
        {"ap_mac": f"30f5-277a-{index:04x}", "site": "", "state": "R/M"}
        for index in range(13)
    ]

    rows = build_ap_online_overview_rows(
        planned_aps=planned_aps, fit_ap_resources=resources
    )
    by_site = {row["site"]: row for row in rows}

    assert by_site["未归属"]["total"] == 2
    assert by_site["未归属"]["online"] == 0
    assert by_site["未归属"]["offline"] == 2
    assert rows[-1]["online"] == 0


def test_ap_online_overview_deduplicates_by_uuid_serial_and_mac():
    rows = [
        {
            "ap_uuid": "ap-1",
            "serial_number": "SN-1",
            "ap_mac": "mac-1",
            "site": "S1",
            "state": "R/M",
        },
        {
            "ap_uuid": "ap-1",
            "serial_number": "SN-1",
            "ap_mac": "mac-1",
            "site": "S1",
            "state": "R/M",
        },
        {"serial_number": "SN-2", "ap_mac": "mac-2", "site": "S1", "state": "R/B"},
        {"serial_number": "SN-2", "ap_mac": "mac-2", "site": "S1", "state": "R/B"},
        {"ap_mac": "mac-3", "site": "S1", "state": "I"},
        {"ap_mac": "mac-3", "site": "S1", "state": "I"},
    ]

    overview = build_ap_online_overview_rows(
        planned_aps=rows, fit_ap_resources=rows, capacities={"S1": 5}
    )

    assert overview[0]["total"] == 5
    assert overview[0]["online"] == 2
    assert overview[0]["offline"] == 3


def test_export_ap_online_overview_xlsx_contains_colors_and_alignment(tmp_path):
    from openpyxl import load_workbook

    export_path = tmp_path / "overview.xlsx"
    plan_rows = [
        {"ap_name": "AP-1", "site_name": "01小洋江站"},
        {"ap_name": "AP-2", "site_name": "02云龙火车站站"},
    ]
    resource_rows = [
        {"ap_name": "AP-1", "site": "01小洋江站", "state": "R/M"},
        {"ap_name": "AP-2", "site": "02云龙火车站站", "state": "I"},
    ]
    rows = build_ap_online_overview_rows(
        planned_aps=plan_rows,
        fit_ap_resources=resource_rows,
        capacities={"02云龙火车站站": 1},
    )
    rows[0]["optical_problem_count"] = 1
    rows[-1]["optical_problem_count"] = 1
    rows[1]["remark"] = "Need check"
    headers = [
        "Station", "AP Total", "Online", "Offline", "Online Rate",
        "Optical Problems", "Remark",
    ]

    export_ap_online_overview_xlsx(export_path, rows, headers)

    workbook = load_workbook(export_path)
    sheet = workbook["AP Online Overview"]
    assert export_path.exists()
    assert [cell.value for cell in sheet[1]] == headers
    assert sheet.freeze_panes == "A2"
    assert sheet["A1"].alignment.horizontal == "center"
    assert sheet["A2"].alignment.horizontal == "center"
    assert sheet.column_dimensions["A"].width > len("车站")
    assert sheet["A2"].fill.fgColor.rgb == "FFDCFCE7"
    assert sheet["D3"].fill.fgColor.rgb == "FFFEE2E2"
    assert sheet["F2"].fill.fgColor.rgb == "FFFEE2E2"
    assert sheet["F3"].fill.fgColor.rgb != "FFFEE2E2"
    assert sheet["F4"].fill.fgColor.rgb == "FFFEE2E2"
    assert sheet["G3"].value == "Need check"


def test_trackside_ap_description_filter_is_case_insensitive():
    assert description_contains_ap("To_AP01")
    assert description_contains_ap("ap-001")
    assert description_contains_ap("Ap uplink")
    assert description_contains_ap("aP access")
    assert not description_contains_ap("camera")


def test_trackside_ap_normalizes_mac_and_interface_names():
    assert normalize_mac("bc5a-3457-cbe0") == "bc5a-3457-cbe0"
    assert normalize_mac("bc:5a:34:57:cb:e0") == "bc5a-3457-cbe0"
    assert normalize_mac("bc5a.3457.cbe0") == "bc5a-3457-cbe0"
    assert normalize_mac("BC-5A-34-57-CB-E0") == "bc5a-3457-cbe0"
    assert normalize_trackside_interface_name("GE2/0/22") == "GigabitEthernet2/0/22"
    assert (
        normalize_trackside_interface_name("GigabitEthernet2/0/22")
        == "GigabitEthernet2/0/22"
    )
    assert (
        normalize_trackside_interface_name("XGE1/0/49") == "Ten-GigabitEthernet1/0/49"
    )
    assert normalize_trackside_interface_name("BAGG1") == "Bridge-Aggregation1"


def test_trackside_vlan_parser_normalizes_ranges_and_rejects_invalid_values():
    assert parse_vlan_set("21，925-927; 21 922") == {21, 922, 925, 926, 927}
    assert normalize_vlan_text("922,21,921-922") == "21,921,922"
    for value in ("0", "4095", "abc", "930-925"):
        with pytest.raises(ValueError):
            parse_vlan_set(value)


def test_trackside_ap_interface_matches_description_or_pvid_plan():
    plan = {
        "mode": TRACKSIDE_AP_PLAN_MODE,
        "station_vlans": {"Station A": {921}},
        "all_vlans": {921},
        "station_totals": {"Station A": 30},
    }
    switch = Device(name="SW", station="Station A", device_uuid="sw-1")

    base = {"interface_name": "GigabitEthernet1/0/1", "port_status": "access"}
    assert is_trackside_ap_interface(
        switch, {**base, "description": "", "pvid": "921"}, plan
    ) == (True, "pvid")
    assert is_trackside_ap_interface(
        switch, {**base, "description": "to AP", "pvid": "1"}, plan
    ) == (True, "description")
    assert is_trackside_ap_interface(
        switch, {**base, "description": "to AP", "pvid": "921"}, plan
    ) == (True, "description+pvid")
    assert is_trackside_ap_interface(
        switch, {**base, "description": "", "pvid": "922"}, plan
    ) == (False, "none")


def test_trackside_ap_interface_uses_line_vlan_candidates_without_station_assumption():
    plan = {
        "mode": TRACKSIDE_AP_PLAN_MODE,
        "station_vlans": {"Station A": {921}, "Station B": {922}},
        "all_vlans": {921, 922},
        "station_totals": {"Station A": 30, "Station B": 56},
    }

    assert pvid_matches_trackside_plan("Station A", "921", plan) is True
    assert pvid_matches_trackside_plan("Station A", "922", plan) is True
    assert pvid_matches_trackside_plan("", "922", plan) is True


def test_trackside_ap_business_rows_include_pvid_match_source():
    switch = Device(name="HX_1", station="Station A", device_uuid="sw-1")
    rows = build_trackside_ap_business_rows(
        [switch],
        {
            "sw-1": [
                {
                    "interface_name": "GigabitEthernet2/0/10",
                    "description": "",
                    "pvid": "921",
                }
            ]
        },
        {"sw-1": []},
        [],
        trackside_ap_plan={
            "mode": TRACKSIDE_AP_PLAN_MODE,
            "station_vlans": {"Station A": {921}},
            "all_vlans": {921},
        },
    )

    assert len(rows) == 1
    assert rows[0]["match_source"] == "pvid"
    assert format_trackside_display_value("match_source", rows[0]) == "PVID匹配"


def test_trackside_ap_business_rows_use_base_station_display_name():
    switch = Device(
        name="SW-1",
        station="设备管理旧站名",
        station_id="station:01",
        device_uuid="sw-1",
    )
    rows = build_trackside_ap_business_rows(
        [switch],
        {
            "sw-1": [
                {
                    "interface_name": "GigabitEthernet1/0/1",
                    "description": "to AP",
                    "pvid": "921",
                }
            ]
        },
        {"sw-1": []},
        [],
        station_names={"station:01": "01-基础资料站"},
    )

    assert len(rows) == 1
    assert rows[0]["site"] == "01-基础资料站"


def test_trackside_ap_business_rows_join_interface_optical_and_fit_ap_data():
    switch = Device(
        name="HX_1", sysname="HX_SYS", station="Station A", device_uuid="sw-1"
    )
    rows = build_trackside_ap_business_rows(
        [switch],
        {
            "sw-1": [
                {
                    "interface_name": "GigabitEthernet2/0/10",
                    "link_status": "UP",
                    "protocol_status": "UP",
                    "description": "To_AP10",
                    "port_status": "trunk",
                    "pvid": "1",
                    "vlan": "10",
                    "updated_at": "2026-01-01T00:00:00",
                },
                {"interface_name": "GigabitEthernet2/0/2", "description": "camera"},
                {"interface_name": "GigabitEthernet2/0/1", "description": "ap-001"},
            ]
        },
        {
            "sw-1": [
                {
                    "interface_name": "GigabitEthernet2/0/10",
                    "rx_power": "-6.10",
                    "status": "normal",
                },
                {
                    "interface_name": "GigabitEthernet2/0/1",
                    "rx_power": "-7.20",
                    "status": "warning",
                },
            ]
        },
        [
            {
                "ac_device_uuid": "ac-1",
                "ap_uuid": "ap-10",
                "ap_mac": "bc5a-3457-cbe0",
                "neighbor_device_name": "HX_1",
                "neighbor_interface": "GigabitEthernet2/0/10",
                "ap_name": "AP10",
                "rx_power": "-14.35",
                "rx_low_alarm": "-19.00",
                "rx_low_warning": "-16.99",
                "updated_at": "2026-01-02T00:00:00",
            }
        ],
    )

    assert [row["interface_name"] for row in rows] == [
        "GigabitEthernet2/0/1",
        "GigabitEthernet2/0/10",
    ]
    assert rows[1]["switch_rx_power"] == "-6.10"
    assert rows[1]["link_status"] == "UP"
    assert format_trackside_display_value("link_status", rows[1]) == "UP"
    assert format_trackside_display_value("port_type", rows[1]) == "trunk"
    assert rows[1]["ap_rx_power"] == "-14.35"
    assert rows[1]["ap_optical_status"] == "abnormal"
    assert rows[1]["ap_name"] == "AP10"


def test_trackside_ap_business_keeps_same_ap_on_different_interfaces():
    switch = Device(
        name="04-横溪站1", sysname="HX_1", station="03横溪站", device_uuid="sw-hx-1"
    )
    rows = build_trackside_ap_business_rows(
        [switch],
        {
            "sw-hx-1": [
                {
                    "interface_name": "GigabitEthernet2/0/16",
                    "link_status": "UP",
                    "description": "AP",
                    "port_status": "trunk",
                    "pvid": "923",
                },
                {
                    "interface_name": "GigabitEthernet2/0/32",
                    "link_status": "DOWN",
                    "description": "AP",
                    "port_status": "trunk",
                    "pvid": "923",
                },
            ]
        },
        {"sw-hx-1": []},
        [],
        {
            "sw-hx-1": [
                {
                    "local_interface": "GE2/0/16",
                    "neighbor_mac": "bc5a-3457-9c60",
                    "neighbor_sysname": "bc5a-3457-9c60",
                },
                {
                    "local_interface": "GE2/0/32",
                    "neighbor_mac": "bc5a-3457-9c60",
                    "neighbor_sysname": "bc5a-3457-9c60",
                },
            ]
        },
        [],
    )

    assert [row["interface_name"] for row in rows] == [
        "GigabitEthernet2/0/16",
        "GigabitEthernet2/0/32",
    ]
    assert {row["ap_mac"] for row in rows} == {"bc5a-3457-9c60"}


def test_trackside_ap_business_link_and_port_type_are_separate():
    assert normalize_link_state("up") == "UP"
    assert normalize_link_state("Administratively DOWN") == "DOWN"
    assert normalize_link_state("") == "-"
    assert (
        format_trackside_display_value(
            "link_status", {"link_status": "DOWN", "port_type": "access"}
        )
        == "DOWN"
    )
    assert (
        format_trackside_display_value(
            "port_type", {"link_status": "DOWN", "port_type": "access"}
        )
        == "access"
    )
    assert (
        format_trackside_display_value(
            "port_type", {"port_type": "DOWN", "port_status": "UP"}
        )
        == "unknown"
    )


def test_trackside_ap_business_offline_ap_keeps_link_state():
    rows = build_trackside_ap_business_rows(
        [],
        {
            "sw-1": [
                {
                    "interface_name": "GigabitEthernet2/0/10",
                    "link_status": "DOWN",
                    "port_status": "access",
                    "pvid": "921",
                }
            ]
        },
        {"sw-1": [{"interface_name": "GigabitEthernet2/0/10", "rx_power": "-6.10"}]},
        [],
        offline_ap_ledger_rows=[
            {
                "site": "Station A",
                "device_uuid": "sw-1",
                "historical_switch_name": "HX_1",
                "historical_switch_interface": "GigabitEthernet2/0/10",
                "ap_mac": "30f5-277a-15e0",
                "ap_name": "AP-IDLE",
                "ap_status": "Idle",
            }
        ],
    )

    assert rows[0]["is_ap_offline"] is True
    assert format_trackside_display_value("link_status", rows[0]) == "DOWN"
    assert format_trackside_display_value("port_type", rows[0]) == "access"
    assert (
        format_trackside_display_value("ap_optical_status", rows[0])
        == "未知"
    )


def test_trackside_ap_business_matches_fit_ap_by_lldp_neighbor_mac():
    switch = Device(name="HX_1", station="Station A", device_uuid="sw-1")
    rows = build_trackside_ap_business_rows(
        [switch],
        {
            "sw-1": [
                {"interface_name": "GigabitEthernet2/0/22", "description": "To_AP22"}
            ]
        },
        {
            "sw-1": [
                {
                    "interface_name": "GigabitEthernet2/0/22",
                    "rx_power": "-6.10",
                    "status": "normal",
                }
            ]
        },
        [
            {
                "ac_device_uuid": "ac-1",
                "ap_uuid": "ap-22",
                "ap_mac": "bc5a-3457-cbe0",
                "ap_name": "Business-AP-22",
                "rx_power": "-14.35",
                "rx_low_alarm": "-19.00",
                "rx_low_warning": "-16.99",
            }
        ],
        {
            "sw-1": [
                {
                    "local_interface": "GE2/0/22",
                    "neighbor_mac": "BC:5A:34:57:CB:E0",
                    "neighbor_interface": "GigabitEthernet1/0/2",
                }
            ]
        },
    )

    assert rows[0]["ap_mac"] == "bc5a-3457-cbe0"
    assert rows[0]["ap_name"] == "Business-AP-22"
    assert rows[0]["ap_rx_power"] == "-14.35"
    assert rows[0]["ap_optical_status"] == "abnormal"


def test_trackside_ap_business_matches_fit_ap_resource_by_lldp_neighbor_mac():
    switch = Device(name="HX_1", station="Station A", device_uuid="sw-1")
    rows = build_trackside_ap_business_rows(
        [switch],
        {"sw-1": [{"interface_name": "GigabitEthernet2/0/23", "description": "AP23"}]},
        {
            "sw-1": [
                {
                    "interface_name": "GigabitEthernet2/0/23",
                    "rx_power": "-6.20",
                    "status": "normal",
                }
            ]
        },
        [],
        {"sw-1": [{"local_interface": "GE2/0/23", "neighbor_mac": "bc5a.3457.cbe1"}]},
        [
            {
                "ac_device_uuid": "ac-1",
                "ap_uuid": "ap-23",
                "ap_mac": "bc5a-3457-cbe1",
                "ap_name": "Renamed-AP-23",
            }
        ],
    )

    assert rows[0]["ap_mac"] == "bc5a-3457-cbe1"
    assert rows[0]["ap_name"] == "Renamed-AP-23"
    assert rows[0]["ap_rx_power"] is None
    assert rows[0]["switch_optical_status"] == "normal"
    assert rows[0]["ap_optical_status"] == "unknown"
    assert has_ap_side_optical_data(rows[0]) is False
    assert format_ap_side_alarm(rows[0]) == "未知"


def test_trackside_ap_business_keeps_neighbor_mac_when_fit_ap_not_found():
    switch = Device(name="HX_1", station="Station A", device_uuid="sw-1")
    rows = build_trackside_ap_business_rows(
        [switch],
        {"sw-1": [{"interface_name": "GigabitEthernet2/0/24", "description": "AP24"}]},
        {
            "sw-1": [
                {
                    "interface_name": "GigabitEthernet2/0/24",
                    "rx_power": "-6.20",
                    "rx_low_alarm": "-20.00",
                    "rx_low_warning": "-8.00",
                }
            ]
        },
        [],
        {"sw-1": [{"local_interface": "GE2/0/24", "neighbor_mac": "bc5a-3457-cbe2"}]},
    )

    assert rows[0]["ap_mac"] == "bc5a-3457-cbe2"
    assert rows[0]["ap_name"] is None
    assert rows[0]["ap_rx_power"] is None
    assert rows[0]["switch_optical_status"] == "notice"
    assert rows[0]["ap_optical_status"] == "unknown"
    assert has_ap_side_optical_data(rows[0]) is False
    assert format_ap_side_alarm(rows[0]) == "未知"


def test_trackside_ap_business_keeps_switch_and_ap_status_separate():
    switch = Device(name="HX_1", station="Station A", device_uuid="sw-1")
    rows = build_trackside_ap_business_rows(
        [switch],
        {"sw-1": [{"interface_name": "GigabitEthernet2/0/25", "description": "AP25"}]},
        {
            "sw-1": [
                {
                    "interface_name": "GigabitEthernet2/0/25",
                    "rx_power": "-6.20",
                    "status": "normal",
                }
            ]
        },
        [
            {
                "ap_uuid": "ap-25",
                "ap_mac": "bc5a-3457-cbe3",
                "ap_name": "AP25",
                "rx_power": "-20.00",
                "rx_low_alarm": "-19.00",
                "rx_low_warning": "-16.99",
            }
        ],
        {"sw-1": [{"local_interface": "GE2/0/25", "neighbor_mac": "bc5a-3457-cbe3"}]},
    )

    assert rows[0]["switch_optical_status"] == "normal"
    assert rows[0]["ap_optical_status"] == "abnormal"
    assert rows[0]["ap_device_optical_status"] == "alarm"
    assert trackside_row_status(rows[0]) == "abnormal"


def test_trackside_ap_business_row_status_uses_more_severe_side():
    assert (
        trackside_row_status(
            {"switch_optical_status": "warning", "ap_optical_status": "normal"}
        )
        == "warning"
    )
    assert (
        trackside_row_status(
            {
                "switch_optical_status": "normal",
                "ap_rx_power": "-17.80",
                "ap_optical_status": "alarm",
                "ap_side_has_data": True,
            }
        )
        == "abnormal"
    )


def test_trackside_ap_side_missing_data_formats_as_dash():
    row = {
        "ap_mac": "-",
        "ap_name": "-",
        "ap_rx_power": "-",
        "ap_tx_power": "",
        "ap_optical_status": "no_module",
        "ap_side_has_data": False,
    }

    assert has_ap_side_optical_data(row) is False
    assert format_ap_side_alarm(row) == "未知"
    assert format_trackside_display_value("ap_mac", row) == "-"
    assert format_trackside_display_value("ap_name", row) == "-"
    assert format_trackside_display_value("ap_rx_power", row) == "-"
    assert format_trackside_display_value("ap_tx_power", row) == "-"


def test_trackside_ap_side_unmatched_optical_record_formats_as_dash():
    row = {
        "ap_mac": "bc5a-3457-cbe1",
        "ap_name": "Renamed-AP-23",
        "ap_rx_power": None,
        "ap_optical_status": "no_module",
        "ap_side_has_data": False,
    }

    assert has_ap_side_optical_data(row) is False
    assert format_ap_side_alarm(row) == "未知"


def test_trackside_ap_side_explicit_no_module_keeps_no_module_label():
    row = {
        "ap_mac": "bc5a-3457-cbe1",
        "ap_name": "AP23",
        "ap_rx_power": None,
        "ap_tx_power": None,
        "ap_optical_status": "no_module",
        "raw_status": "no module",
        "ap_side_has_data": True,
    }

    assert has_ap_side_optical_data(row) is True
    assert format_ap_side_alarm(row) == "未知"


def test_trackside_ap_side_normal_and_notice_format_from_computed_status():
    normal = {
        "ap_mac": "bc5a-3457-cbe1",
        "ap_name": "AP23",
        "ap_rx_power": "-7.55",
        "ap_tx_power": "1.20",
        "ap_optical_status": "normal",
        "ap_side_has_data": True,
    }
    notice = {**normal, "ap_rx_power": "-14.35", "ap_optical_status": "notice"}

    assert format_ap_side_alarm(normal) == "正常"
    assert format_ap_side_alarm(notice) == "光衰大"


def test_trackside_ap_side_unknown_with_rx_power_recomputes_for_display():
    row = {
        "ap_mac": "bc5a-3457-cbe1",
        "ap_name": "AP23",
        "ap_rx_power": "-14.41",
        "ap_optical_status": "unknown",
        "ap_side_has_data": True,
    }

    assert format_ap_side_alarm(row) == "光衰大"
    assert format_trackside_display_value("ap_optical_status", row) == "光衰大"


def test_trackside_history_unknown_with_rx_power_recomputes_ap_status():
    row = {
        "rx_power": "-19.07",
        "rx_low_alarm": None,
        "rx_low_warning": None,
        "optical_alarm_status": "unknown",
    }

    assert _optical_status_from_history(row, "ap") == "abnormal"


def test_trackside_ap_optical_status_uses_default_profile_without_thresholds():
    switch = Device(name="HX_1", station="Station A", device_uuid="sw-1")
    rows = build_trackside_ap_business_rows(
        [switch],
        {
            "sw-1": [
                {"interface_name": "GigabitEthernet2/0/10", "description": "To_AP10"}
            ]
        },
        {"sw-1": [{"interface_name": "GigabitEthernet2/0/10", "rx_power": "-6.10"}]},
        [
            {
                "ap_uuid": "ap-10",
                "ap_mac": "bc5a-3457-cbe0",
                "ap_name": "AP10",
                "neighbor_device_name": "HX_1",
                "neighbor_interface": "GigabitEthernet2/0/10",
                "rx_power": "-14.41",
            }
        ],
    )

    assert rows[0]["ap_rx_power"] == "-14.41"
    assert rows[0]["ap_optical_status"] == "abnormal"
    assert format_trackside_display_value("ap_optical_status", rows[0]) == "光衰大"


def test_trackside_wa6522_is_not_applicable_and_excluded_from_optical_anomalies():
    row = normalize_trackside_ap_business_row(
        {
            "model": " wa6522 ",
            "ap_mac": "0011-2233-4455",
            "ap_name": "AP-WA6522",
            "ap_rx_power": "-30",
            "ap_optical_status": "no_light",
            "switch_rx_power": "-30",
            "switch_optical_status": "alarm",
            "ap_side_has_data": True,
        }
    )

    assert row["ap_optical_applicable"] is False
    assert row["ap_optical_status"] == "not_applicable"
    assert row["switch_optical_status"] == "not_applicable"
    assert row["optical_severity"] == "not_applicable"
    assert row["ap_business_reason"] == "该型号使用网口接入，不适用 AP 光模块光衰检测。"
    assert is_current_optical_abnormal_row(row) is False
    assert count_current_optical_abnormal_aps([row]) == 0


def test_trackside_switch_rx_fixed_threshold_drives_business_status_and_count():
    row = normalize_trackside_ap_business_row(
        {
            "model": "WA6528X-E",
            "ap_mac": "0011-2233-4455",
            "ap_name": "AP-WA6528X-E",
            "ap_rx_power": "-7.72",
            "ap_optical_status": "normal",
            "switch_rx_power": "-19.10",
            "switch_optical_status": "normal",
            "ap_side_has_data": True,
        }
    )

    assert row["ap_optical_status"] == "normal"
    assert row["switch_device_optical_status"] == "normal"
    assert row["switch_optical_status"] == "abnormal"
    assert row["ap_business_optical_status"] == "abnormal"
    assert row["optical_severity"] == "abnormal"
    assert "交换机侧收光 -19.10 dBm 低于业务门限 -13.90 dBm" in str(
        row["ap_business_reason"]
    )
    assert is_current_optical_abnormal_row(row) is True
    assert count_current_optical_abnormal_aps([row]) == 1


def test_current_optical_problem_count_groups_by_station_and_deduplicates_ap():
    first = normalize_trackside_ap_business_row(
        {
            "site_key": "NBO12",
            "station_name": "宁波地铁12号线01",
            "site": "宁波地铁12号线01",
            "model": "WA6528X-E",
            "ap_mac": "0011-2233-4455",
            "ap_name": "AP-1",
            "ap_rx_power": "-7.72",
            "ap_optical_status": "normal",
            "switch_rx_power": "-19.10",
            "switch_optical_status": "normal",
            "ap_side_has_data": True,
        }
    )
    duplicate_side = {**first, "interface_name": "Bridge-Aggregation1"}
    second = {
        **first,
        "station_name": "宁波地铁12号线02",
        "site": "宁波地铁12号线02",
        "ap_mac": "00aa-bbcc-ddee",
        "ap_name": "AP-2",
    }

    assert count_current_optical_abnormal_by_site([first, duplicate_side, second]) == {
        "宁波地铁12号线01": 1,
        "宁波地铁12号线02": 1,
    }


def test_current_optical_problem_uses_one_business_predicate_and_excludes_non_problems():
    def row(**overrides: object) -> dict[str, object | None]:
        result: dict[str, object | None] = {
            "site": "站点A",
            "station_name": "站点A",
            "model": "WA6528X-E",
            "ap_uuid": "ap-1",
            "ap_mac": "0011-2233-4455",
            "ap_name": "AP-1",
            "ap_identity_entity_id": "entity-1",
            "identity_match_status": "matched",
            "ap_side_has_data": True,
            "ap_rx_power": "-7.72",
            "ap_business_optical_status": "normal",
            "ap_device_optical_status": "normal",
            "ap_optical_status": "normal",
            "switch_rx_power": "-19.10",
            "switch_optical_status": "abnormal",
            "switch_device_optical_status": "abnormal",
        }
        result.update(overrides)
        return result

    valid = row()
    duplicate_side = row(interface_name="Bridge-Aggregation1")
    ap_side_valid = row(
        ap_uuid="ap-2",
        ap_mac="0011-2233-4456",
        ap_name="AP-2",
        ap_identity_entity_id="entity-2",
        ap_rx_power="-19.10",
        ap_business_optical_status="abnormal",
        ap_device_optical_status="abnormal",
        ap_optical_status="abnormal",
        switch_rx_power="-7.72",
        switch_optical_status="normal",
        switch_device_optical_status="normal",
    )

    assert count_current_optical_abnormal_aps([valid, duplicate_side, ap_side_valid]) == 2
    assert count_current_optical_abnormal_by_site(
        [valid, duplicate_side, ap_side_valid]
    ) == {"站点A": 2}

    excluded_rows = [
        row(
            switch_optical_status="collection_failed",
            switch_device_optical_status="collection_failed",
        ),
        row(
            switch_optical_status="not_collected",
            switch_device_optical_status="not_collected",
        ),
        row(
            switch_optical_status="unknown",
            switch_device_optical_status="unknown",
        ),
        row(primary_reason_code="EMPTY_CONFIGURED_PORT"),
        row(identity_match_status="unresolved"),
        row(
            switch_rx_power="-7.72",
            switch_optical_status="normal",
            switch_device_optical_status="normal",
            lldp_match_status="LLDP_SNAPSHOT_STALE",
        ),
    ]
    assert count_current_optical_abnormal_aps(excluded_rows) == 0
    assert count_current_optical_abnormal_by_site(excluded_rows) == {}

    unassigned = row(
        ap_uuid="ap-unassigned",
        ap_mac="0011-2233-4466",
        ap_name="AP-未归属",
        ap_identity_entity_id="entity-unassigned",
        site="",
        station_name="",
        station="",
    )
    assert count_current_optical_abnormal_by_site([unassigned]) == {"未归属": 1}


def test_trackside_wa6522_display_is_not_applicable_before_row_normalization():
    row = {
        "model": "WA6522",
        "ap_name": "AP-WA6522",
        "ap_mac": "0011-2233-4455",
        "ap_rx_power": "-30",
        "ap_optical_status": "alarm",
    }

    assert format_ap_side_alarm(row) == "不适用"
    assert format_trackside_display_value("ap_optical_status", row) == "不适用"


def test_trackside_row_status_marks_missing_ap_side_data_incomplete():
    row = {
        "switch_optical_status": "normal",
        "ap_optical_status": "alarm",
        "ap_side_has_data": False,
    }

    assert trackside_row_status(row) == "unknown"


def test_trackside_ap_business_filter_by_site_and_search():
    rows = [
        {
            "site": "Station A",
            "ap_name": "AP-A",
            "device_name": "HX_1",
            "interface_name": "GigabitEthernet1/0/1",
        },
        {
            "site": "Station B",
            "ap_name": "AP-B",
            "device_name": "HX_2",
            "interface_name": "GigabitEthernet1/0/2",
        },
        {
            "site": "Station C",
            "ap_name": "AP-C",
            "device_name": "HX_3",
            "interface_name": "GigabitEthernet1/0/3",
        },
    ]

    assert len(filter_trackside_ap_business_rows(rows, "", "")) == 3
    assert len(filter_trackside_ap_business_rows(rows, None, "")) == 3
    assert [
        row["ap_name"]
        for row in filter_trackside_ap_business_rows(rows, "Station A", "")
    ] == ["AP-A"]
    assert [
        row["ap_name"] for row in filter_trackside_ap_business_rows(rows, "", "hx_2")
    ] == ["AP-B"]
    assert [
        row["ap_name"]
        for row in filter_trackside_ap_business_rows(rows, "Station B", "1/0/2")
    ] == ["AP-B"]


def test_trackside_ap_business_sort_uses_natural_switch_and_interface_order():
    rows = [
        {"site": "10站", "device_name": "SW-10", "interface_name": "GE1/0/2"},
        {"site": "2站", "device_name": "SW-2", "interface_name": "GE1/0/10"},
        {"site": "2站", "device_name": "SW-2", "interface_name": "GE1/0/2"},
        {"site": "01站", "device_name": "SW-1", "interface_name": "GE1/0/1"},
    ]

    assert [
        (row["device_name"], row["interface_name"])
        for row in sort_trackside_ap_business_rows(rows)
    ] == [
        ("SW-1", "GE1/0/1"),
        ("SW-2", "GE1/0/2"),
        ("SW-2", "GE1/0/10"),
        ("SW-10", "GE1/0/2"),
    ]


def test_trackside_ap_i18n_zh_cn_keys_are_translated():
    zh = I18n("zh_CN")
    assert zh.t("rail_transit.trackside_ap_service") == "\u8f68\u65c1AP\u4e1a\u52a1"
    assert zh.t("trackside_ap.update") == "\u66f4\u65b0"
    assert zh.t("trackside_ap.cancel_update") == "\u53d6\u6d88\u66f4\u65b0"
    assert zh.t("trackside_ap.not_collected") == "\u672a\u91c7\u96c6"
    assert (
        zh.t("trackside_ap.vendor_not_supported")
        == "\u5f53\u524d\u5382\u5546\u6682\u672a\u9002\u914d\u5149\u8870\u91c7\u96c6\u547d\u4ee4"
    )


def test_trackside_ap_progress_text_has_no_mojibake_or_i18n_key():
    zh = I18n("zh_CN")
    en = I18n("en_US")
    zh_text = zh.t("trackside_ap.collecting_progress", done=0, total=1200)
    en_text = en.t("trackside_ap.collection_summary", success=26, failed=0, skipped=640)

    assert zh_text == "\u91c7\u96c6\u4e2d\uff1a0 / 1200"
    assert en_text == "Completed: 26 succeeded, 0 failed, 640 skipped"
    assert "?" not in zh_text
    assert "trackside_ap." not in zh_text
    assert "trackside_ap." not in en_text


def test_trackside_optical_command_adapter_supports_h3c_and_zte_aliases_and_rejects_huawei():
    assert (
        OpticalCommandAdapter.get_optical_diagnosis_commands("H3C", "SW")
        == TRACKSIDE_OPTICAL_COMMANDS
    )
    assert (
        OpticalCommandAdapter.get_optical_diagnosis_commands(
            "\u65b0\u534e\u4e09", "\u4ea4\u6362\u673a"
        )
        == TRACKSIDE_OPTICAL_COMMANDS
    )
    for vendor in ("ZTE", "\u4e2d\u5174"):
        assert OpticalCommandAdapter.get_optical_diagnosis_commands(
            vendor, "SW"
        ) == ("show version", "show interface brief", "show opticalinfo brief")
    for vendor in ("Huawei", "\u534e\u4e3a"):
        with pytest.raises(UnsupportedVendor):
            OpticalCommandAdapter.get_optical_diagnosis_commands(vendor, "SW")


def test_trackside_station_switch_target_filter_uses_station_group_and_switch_types(
    tmp_path,
):
    database = make_database(tmp_path)
    repository = DeviceRepository(database)
    groups = DeviceGroupRepository(database, "demo")
    station = groups.create("车站")
    onboard = groups.create("车载")
    switch_a = repository.create(
        Device(
            name="A",
            group_id=station.id,
            device_type="SW",
            device_vendor="H3C",
            ip_address="10.0.0.1",
            ssh_username="u",
            ssh_password="p",
        )
    )
    repository.create(
        Device(
            name="B",
            group_id=station.id,
            device_type="FAT-AP",
            ip_address="10.0.0.2",
            ssh_username="u",
            ssh_password="p",
        )
    )
    repository.create(
        Device(
            name="C",
            group_id=onboard.id,
            device_type="FAT-AP",
            ip_address="10.0.0.3",
            ssh_username="u",
            ssh_password="p",
        )
    )
    switch_d = repository.create(
        Device(
            name="D",
            group_id=station.id,
            device_type="交换机",
            device_vendor="\u65b0\u534e\u4e09",
            ip_address="10.0.0.4",
            ssh_username="u",
            ssh_password="p",
        )
    )

    targets, skipped = build_station_switch_targets(repository, "demo")

    assert [target.device_id for target in targets] == [switch_a.id, switch_d.id]
    assert skipped == []


def test_trackside_station_switch_target_filter_can_scope_station(tmp_path):
    database = make_database(tmp_path)
    repository = DeviceRepository(database)
    station_group = DeviceGroupRepository(database, "demo").create("车站")
    station_a = repository.create(
        Device(
            name="A",
            station="Station A",
            group_id=station_group.id,
            device_type="SW",
            device_vendor="H3C",
            ip_address="10.0.0.1",
            ssh_username="u",
            ssh_password="p",
        )
    )
    repository.create(
        Device(
            name="B",
            station="Station B",
            group_id=station_group.id,
            device_type="SW",
            device_vendor="H3C",
            ip_address="10.0.0.2",
            ssh_username="u",
            ssh_password="p",
        )
    )

    targets, skipped = build_station_switch_targets(
        repository, "demo", station="Station A"
    )

    assert [target.device_id for target in targets] == [station_a.id]
    assert skipped == []


def test_trackside_station_switch_target_filter_matches_numbered_display_name(tmp_path):
    database = make_database(tmp_path)
    repository = DeviceRepository(database)
    station_group = DeviceGroupRepository(database, "demo").create("车站")
    station_a = repository.create(
        Device(
            name="A",
            station="01Station A",
            group_id=station_group.id,
            device_type="SW",
            device_vendor="H3C",
            ip_address="10.0.0.1",
            ssh_username="u",
            ssh_password="p",
        )
    )
    repository.create(
        Device(
            name="B",
            station="02Station B",
            group_id=station_group.id,
            device_type="SW",
            device_vendor="H3C",
            ip_address="10.0.0.2",
            ssh_username="u",
            ssh_password="p",
        )
    )

    targets, skipped = build_station_switch_targets(
        repository, "demo", station="01-Station A"
    )

    assert [target.device_id for target in targets] == [station_a.id]
    assert skipped == []


def test_trackside_station_switch_target_skips_unsupported_vendor(tmp_path):
    database = make_database(tmp_path)
    repository = DeviceRepository(database)
    station = DeviceGroupRepository(database, "demo").create("车站")
    repository.create(
        Device(
            name="HW",
            group_id=station.id,
            device_type="SW",
            device_vendor="Huawei",
            ip_address="10.0.0.5",
            ssh_username="u",
            ssh_password="p",
        )
    )

    targets, skipped = build_station_switch_targets(repository, "demo")

    assert targets == []
    assert skipped[0].reason == "vendor_not_supported"


def test_trackside_ap_targets_skip_missing_connection_info(tmp_path):
    database = make_database(tmp_path)
    device_repository = DeviceRepository(database)
    connectable = device_repository.create(
        Device(
            name="AP-OK",
            device_type="FAT-AP",
            ip_address="10.0.0.10",
            ssh_username="u",
            ssh_password="p",
        )
    )
    device_repository.create(
        Device(
            name="AP-NO-PASSWORD",
            device_type="FAT-AP",
            ip_address="10.0.0.11",
            ssh_username="u",
            ssh_password="",
        )
    )
    ac_repository = AcRepository(database)
    ac_repository.replace_fit_ap_resources(
        "ac-1",
        [
            {"ap_uuid": "ap-ok", "ap_name": "AP-OK", "ap_ip": "10.0.0.10"},
            {"ap_uuid": "ap-no-ip", "ap_name": "AP-NO-IP", "ap_ip": ""},
            {
                "ap_uuid": "ap-no-password",
                "ap_name": "AP-NO-PASSWORD",
                "ap_ip": "10.0.0.11",
            },
        ],
    )

    targets, skipped = build_trackside_ap_targets(
        ac_repository,
        device_repository,
        [{"ap_uuid": "ap-ok"}, {"ap_uuid": "ap-no-ip"}, {"ap_uuid": "ap-no-password"}],
    )

    assert [target.device_id for target in targets] == [connectable.id]
    assert {item.name for item in skipped} == {"AP-NO-IP", "AP-NO-PASSWORD"}


def test_trackside_fit_ap_branch_skips_without_resources_instead_of_failing(
    tmp_path,
    monkeypatch,
):
    database = make_database(tmp_path)
    repository = DeviceRepository(database)
    repository.create(
        Device(
            name="AC-EMPTY",
            device_type="AC",
            device_vendor="H3C",
            ip_address="10.0.0.20",
            ssh_username="u",
            ssh_password="p",
        )
    )
    monkeypatch.setattr(
        trackside_optical_collection,
        "collect_h3c_ac_resources",
        lambda *_args, **_kwargs: SimpleNamespace(success=True),
    )
    monkeypatch.setattr(
        trackside_optical_collection,
        "collect_h3c_fit_ap_optical",
        lambda *_args, **_kwargs: pytest.fail("无 FIT-AP 资源时不应启动 AP 光衰采集"),
    )

    results, total, skipped, failures = trackside_optical_collection._collect_fit_ap_optical_subtasks(
        repository,
        "demo",
        PathResolver(tmp_path),
        concurrency=1,
        cancel_event=trackside_optical_collection.Event(),
    )

    assert results == []
    assert total == 0
    assert failures == []
    assert [(item.target_type, item.reason) for item in skipped] == [
        ("FIT_AP", "no_fit_ap_resource")
    ]


def test_trackside_fit_ap_default_collection_enumerates_all_h3c_ac_roles(
    tmp_path,
    monkeypatch,
):
    database = make_database(tmp_path)
    repository = DeviceRepository(database)
    ac_a = repository.create(
        Device(
            name="AC-A",
            device_type="AC",
            device_vendor="H3C",
            ip_address="10.0.0.30",
            ssh_username="u",
            ssh_password="p",
        )
    )
    ac_b = repository.create(
        Device(
            name="AC-B",
            device_type="wireless_controller",
            device_vendor="H3C",
            ip_address="10.0.0.31",
            ssh_username="u",
            ssh_password="p",
        )
    )
    repository.create(
        Device(
            name="SW-IGNORED",
            device_type="SW",
            device_vendor="H3C",
            ip_address="10.0.0.32",
            ssh_username="u",
            ssh_password="p",
        )
    )
    resource_calls: list[str] = []
    optical_calls: list[str] = []
    progress_events: list[dict[str, object]] = []

    def fake_resource_collect(device, *_args, **_kwargs):
        device_uuid = str(device.device_uuid)
        resource_calls.append(device_uuid)
        AcRepository(database).replace_fit_ap_resources(
            device_uuid,
            [{"ap_uuid": f"{device_uuid}-ap", "ap_name": device.name, "ap_ip": "10.0.1.1"}],
        )
        return SimpleNamespace(success=True, error_message="")

    def fake_optical_collect(**kwargs):
        device_uuid = str(kwargs["ac_device"].device_uuid)
        optical_calls.append(device_uuid)
        return SimpleNamespace(success=True, ac_device_uuid=device_uuid, optical_rows=[])

    monkeypatch.setattr(
        trackside_optical_collection,
        "collect_h3c_ac_resources",
        fake_resource_collect,
    )
    monkeypatch.setattr(
        trackside_optical_collection,
        "collect_h3c_fit_ap_optical",
        fake_optical_collect,
    )

    results, total, skipped, failures = trackside_optical_collection._collect_fit_ap_optical_subtasks(
        repository,
        "demo",
        PathResolver(tmp_path),
        concurrency=1,
        cancel_event=trackside_optical_collection.Event(),
        ac_progress_callback=progress_events.append,
    )

    expected = {str(ac_a.device_uuid), str(ac_b.device_uuid)}
    assert set(resource_calls) == expected
    assert set(optical_calls) == expected
    assert len(resource_calls) == len(optical_calls) == 2
    assert total == 2
    assert len(results) == 2
    assert skipped == []
    assert failures == []
    assert {event["ac_device_uuid"] for event in progress_events} == expected
    assert {event["ac_total"] for event in progress_events} == {2}


def test_trackside_fit_ap_branch_reports_ac_resource_failure_as_failure(
    tmp_path,
    monkeypatch,
):
    database = make_database(tmp_path)
    repository = DeviceRepository(database)
    ac = repository.create(
        Device(
            name="AC-FAIL",
            device_type="AC",
            device_vendor="H3C",
            ip_address="10.0.0.21",
            ssh_username="u",
            ssh_password="p",
        )
    )
    monkeypatch.setattr(
        trackside_optical_collection,
        "collect_h3c_ac_resources",
        lambda *_args, **_kwargs: SimpleNamespace(
            success=False,
            error_message="SSH 连接失败",
        ),
    )

    results, total, skipped, failures = trackside_optical_collection._collect_fit_ap_optical_subtasks(
        repository,
        "demo",
        PathResolver(tmp_path),
        concurrency=1,
        cancel_event=trackside_optical_collection.Event(),
    )

    assert results == []
    assert total == 0
    assert skipped == []
    assert len(failures) == 1
    assert failures[0]["device_uuid"] == str(ac.device_uuid)
    assert failures[0]["host"] == "10.0.0.21"
    assert failures[0]["stage"] == "trackside_ap.ac_resource_refresh"
    assert failures[0]["reason_code"] == "fit_ap_resource_failed"
    assert failures[0]["message"] == "SSH 连接失败"


def test_trackside_collection_aggregates_ac_resource_failure_as_failed_unit(
    tmp_path,
    monkeypatch,
):
    database = make_database(tmp_path)
    repository = DeviceRepository(database)
    ac = repository.create(
        Device(
            name="AC-FAIL",
            device_type="AC",
            device_vendor="H3C",
            ip_address="10.0.0.22",
            ssh_username="u",
            ssh_password="p",
        )
    )
    monkeypatch.setattr(
        trackside_optical_collection,
        "collect_h3c_ac_resources",
        lambda *_args, **_kwargs: SimpleNamespace(
            success=False,
            error_message="AC 资源刷新失败",
        ),
    )

    result = collect_trackside_optical(
        repository,
        "demo",
        PathResolver(tmp_path),
        [],
        concurrency=1,
    )

    assert result.status == "FAILED"
    assert result.target_count == 1
    assert result.success_count == 0
    assert result.failed_count == 1
    assert result.fit_ap_resource_failed_count == 1
    assert result.failure_reason_counts == {"fit_ap_resource_failed": 1}
    assert result.failures[0]["device_uuid"] == str(ac.device_uuid)
    assert result.failures[0]["host"] == "10.0.0.22"
    with (result.session_dir / "session_meta.json").open(encoding="utf-8") as handle:
        session_meta = json.load(handle)
    assert session_meta["fit_ap_resource_status"] == "FAILED"
    assert session_meta["fit_ap_optical_status"] == "FAILED"


def test_trackside_collection_dedupes_by_device_id_and_uses_default_concurrency(
    tmp_path,
):
    database = make_database(tmp_path)
    repository = DeviceRepository(database)
    groups = DeviceGroupRepository(database, "demo")
    station = groups.create("车站")
    shared = repository.create(
        Device(
            name="Shared",
            group_id=station.id,
            device_type="SW",
            ip_address="10.0.0.10",
            ssh_username="u",
            ssh_password="p",
        )
    )
    ac_repository = AcRepository(database)
    ac_repository.replace_fit_ap_resources(
        "ac-1", [{"ap_uuid": "ap-shared", "ap_name": "Shared", "ap_ip": "10.0.0.10"}]
    )
    switch_targets, _ = build_station_switch_targets(repository, "demo")
    ap_targets, _ = build_trackside_ap_targets(
        ac_repository, repository, [{"ap_uuid": "ap-shared"}]
    )

    assert DEFAULT_TRACKSIDE_OPTICAL_CONCURRENCY == 64
    assert len(dedupe_targets([*switch_targets, *ap_targets])) == 1
    assert switch_targets[0].device_id == shared.id


def test_trackside_optical_collection_runs_commands_writes_database_and_skips_raw_files(
    tmp_path, monkeypatch
):
    database = make_database(tmp_path)
    repository = DeviceRepository(database)
    groups = DeviceGroupRepository(database, "demo")
    station = groups.create("车站")
    repository.create(
        Device(
            name="OK",
            group_id=station.id,
            device_type="SW",
            ip_address="10.0.0.10",
            ssh_username="u",
            ssh_password="p",
        )
    )
    repository.create(
        Device(
            name="FAIL",
            group_id=station.id,
            device_type="SW",
            ip_address="10.0.0.99",
            ssh_username="u",
            ssh_password="p",
        )
    )
    FakeOpticalConnection.instances = []
    monkeypatch.setattr(
        trackside_optical_collection.netmiko_connection,
        "ConnectHandler",
        FakeOpticalConnection,
    )

    result = collect_trackside_optical(
        repository,
        "demo",
        PathResolver(tmp_path),
        [],
        concurrency=DEFAULT_TRACKSIDE_OPTICAL_CONCURRENCY,
    )

    assert result.concurrency == 64
    assert result.requested_concurrency == 64
    assert result.success_count == 1
    assert result.failed_count == 1
    assert result.failure_reason_counts == {"device_collection_failed": 1}
    assert len(result.failures) == 1
    assert result.failures[0]["device_name"] == "FAIL"
    assert result.failures[0]["host"] == "10.0.0.99"
    assert result.failures[0]["stage"] == "trackside_ap.switch.collect"
    expected_commands = [
        "screen-length disable",
        "display interface brief",
        "display transceiver diagnosis interface",
        "display lldp neighbor-information list",
    ]
    assert any(
        connection.commands == expected_commands
        for connection in FakeOpticalConnection.instances
    )
    assert not (result.session_dir / "raw").exists()
    assert (result.session_dir / "session_meta.json").exists()
    parsed_dir = PathResolver(tmp_path).trackside_ap_update_parsed_session_dir(
        "demo", result.session_id
    )
    assert not (parsed_dir / "trackside_update_results.sqlite").exists()
    ok_device = next(device for device in repository.list() if device.name == "OK")
    interfaces = DeviceFactRepository(database).list_device_interfaces(
        ok_device.device_uuid
    )
    assert interfaces[0]["pvid"] == "921"
    assert interfaces[0]["description"] == "To AP"


def test_trackside_update_preserves_current_on_failed_or_invalid_snapshot(
    tmp_path,
):
    database = make_database(tmp_path)
    repository = DeviceRepository(database)
    switch = create_station_switch(
        repository,
        "demo",
        name="SW-PRESERVE",
        ip_address="10.0.0.30",
    )
    facts = DeviceFactRepository(database)
    facts.replace_device_interfaces(
        str(switch.device_uuid),
        [{"interface_name": "GE1/0/1", "link_status": "UP", "pvid": "100"}],
    )
    facts.replace_optical_modules(
        str(switch.device_uuid),
        [{"interface_name": "GE1/0/1", "rx_power": "-7.00", "module_serial_number": "OLD"}],
    )
    target = trackside_optical_collection.TracksideOpticalTarget(
        key="device:1",
        name=switch.name,
        host="10.0.0.30",
        port=22,
        protocol="ssh",
        target_type="SWITCH",
        group_name="车站",
        device=switch,
        device_uuid=str(switch.device_uuid),
    )

    trackside_optical_collection._persist_result(
        repository,
        AcRepository(database),
        trackside_optical_collection.TracksideDeviceCollectionResult(
            target,
            False,
            error_message="SSH 连接失败",
        ),
    )
    assert facts.list_device_interfaces(str(switch.device_uuid))[0]["pvid"] == "100"
    assert facts.list_optical_modules(str(switch.device_uuid))[0]["rx_power"] == "-7.00"

    trackside_optical_collection._persist_result(
        repository,
        AcRepository(database),
        trackside_optical_collection.TracksideDeviceCollectionResult(
            target,
            True,
            interfaces=[{"interface_name": "GE1/0/1", "link_status": "DOWN", "pvid": "200"}],
            rows=[{"interface_name": "GE1/0/1", "rx_power": "-20.00"}],
            interface_snapshot_status="PARTIAL",
            optical_snapshot_status="EMPTY",
        ),
    )
    assert facts.list_device_interfaces(str(switch.device_uuid))[0]["pvid"] == "100"
    assert facts.list_optical_modules(str(switch.device_uuid))[0]["rx_power"] == "-7.00"


def test_trackside_update_combines_fit_ap_service_and_station_switch_collection(
    tmp_path, monkeypatch
):
    database = make_database(tmp_path)
    repository = DeviceRepository(database)
    groups = DeviceGroupRepository(database, "demo")
    station = groups.create("车站")
    switch = repository.create(
        Device(
            name="SW",
            group_id=station.id,
            device_type="SW",
            ip_address="10.0.0.10",
            ssh_username="u",
            ssh_password="p",
        )
    )
    ac = repository.create(make_ac_device())
    ac_repo = AcRepository(database)
    ac_repo.replace_fit_ap_resources(
        ac.device_uuid,
        [
            {
                "ap_uuid": "ap-1",
                "serial_number": "SN-1",
                "ap_name": "AP1",
                "ap_ip": "10.0.0.21",
            },
            {
                "ap_uuid": "ap-2",
                "serial_number": "SN-2",
                "ap_name": "AP2",
                "ap_ip": "10.0.0.22",
            },
            {
                "ap_uuid": "ap-skip",
                "serial_number": "SN-SKIP",
                "ap_name": "AP-SKIP",
                "ap_ip": "",
            },
        ],
    )
    paths = PathResolver(tmp_path)
    fit_calls = []
    resource_calls = []

    def fake_resource_collect(
        ac_device, site_name, repository=None, paths=None, progress=None, should_cancel=None, refresh_ac_overview=True
    ):
        resource_calls.append((ac_device.device_uuid, site_name, refresh_ac_overview))
        run_dir = paths.trackside_ap_raw_dir(site_name) / "ac" / "resource-run"
        run_dir.mkdir(parents=True, exist_ok=True)
        (run_dir / "resource.log").write_text("resource raw", encoding="utf-8")
        repository.replace_fit_ap_resources(
            ac_device.device_uuid,
            [
                {
                    "ap_uuid": "ap-1",
                    "serial_number": "SN-1",
                    "ap_name": "AP1",
                    "ap_ip": "10.0.0.21",
                },
                {
                    "ap_uuid": "ap-2",
                    "serial_number": "SN-2",
                    "ap_name": "AP2",
                    "ap_ip": "10.0.0.22",
                },
                {
                    "ap_uuid": "ap-skip",
                    "serial_number": "SN-SKIP",
                    "ap_name": "AP-SKIP",
                    "ap_ip": "",
                },
            ],
        )
        return SimpleNamespace(
            success=True, collect_run_uuid="resource-run", error_message=None
        )

    def fake_fit_collect(
        ac_device,
        site_name,
        repository=None,
        paths=None,
        max_workers=None,
        progress=None,
        item_progress=None,
        target_ap_uuids=None,
        target_ap_macs=None,
        target_ap_names=None,
        target_stations=None,
        should_cancel=None,
    ):
        fit_calls.append(
            (
                ac_device.device_uuid,
                site_name,
                max_workers,
                target_ap_uuids,
                target_ap_macs,
                target_ap_names,
                target_stations,
            )
        )
        run_dir = paths.trackside_ap_raw_dir(site_name) / "ac" / "fit-run" / "fit_ap"
        run_dir.mkdir(parents=True, exist_ok=True)
        (run_dir / "AP1.log").write_text("fit raw", encoding="utf-8")
        return FitApOpticalCollectResult(
            True, False, str(ac_device.device_uuid), "fit-run", 2, 0, None
        )

    FakeOpticalConnection.instances = []
    monkeypatch.setattr(
        trackside_optical_collection, "collect_h3c_ac_resources", fake_resource_collect
    )
    monkeypatch.setattr(
        trackside_optical_collection, "collect_h3c_fit_ap_optical", fake_fit_collect
    )
    monkeypatch.setattr(
        trackside_optical_collection.netmiko_connection,
        "ConnectHandler",
        FakeOpticalConnection,
    )

    result = collect_trackside_optical(
        repository, "demo", paths, [], concurrency=DEFAULT_TRACKSIDE_OPTICAL_CONCURRENCY
    )

    assert resource_calls == [(ac.device_uuid, "demo", False)]
    assert fit_calls == [(ac.device_uuid, "demo", 64, None, None, None, None)]
    assert result.fit_ap_total == 3
    assert result.station_switch_total == 1
    assert result.success_count == 3
    assert result.failed_count == 0
    assert result.skipped_count == 1
    assert result.target_count == 4
    assert not (result.session_dir / "raw").exists()
    assert switch.id is not None


def test_trackside_ap_update_scopes_switch_to_target_ap_and_reports_offline(
    tmp_path, monkeypatch
):
    database = make_database(tmp_path)
    repository = DeviceRepository(database)
    target_switch = create_station_switch(
        repository,
        "demo",
        name="SW-A",
        station="Station A",
        ip_address="10.0.0.10",
        ssh_username="u",
        ssh_password="p",
    )
    create_station_switch(
        repository,
        "demo",
        name="SW-B",
        station="Station A",
        ip_address="10.0.0.11",
        ssh_username="u",
        ssh_password="p",
    )
    ac = repository.create(make_ac_device())
    ac_repo = AcRepository(database)
    ac_repo.replace_fit_ap_resources(
        ac.device_uuid,
        [
            {
                "ap_uuid": "ap-1",
                "ap_name": "AP1",
                "ap_mac": "bc5a-3457-cbe0",
                "ap_ip": "10.0.0.21",
                "site": "Station A",
                "state": "R/M",
            },
            {
                "ap_uuid": "ap-2",
                "ap_name": "AP2",
                "ap_mac": "bc5a-3457-cbe1",
                "ap_ip": "10.0.0.22",
                "site": "Station A",
                "state": "R/M",
            },
        ],
    )
    paths = PathResolver(tmp_path)
    fit_calls = []

    def fake_resource_collect(
        ac_device, site_name, repository=None, paths=None, progress=None, should_cancel=None, refresh_ac_overview=True
    ):
        repository.replace_fit_ap_resources(
            ac_device.device_uuid,
            [
                {
                    "ap_uuid": "ap-1",
                    "ap_name": "AP1",
                    "ap_mac": "bc5a-3457-cbe0",
                    "ap_ip": "10.0.0.21",
                    "site": "Station A",
                    "state": "I",
                    "state_display": "Idle",
                },
                {
                    "ap_uuid": "ap-2",
                    "ap_name": "AP2",
                    "ap_mac": "bc5a-3457-cbe1",
                    "ap_ip": "10.0.0.22",
                    "site": "Station A",
                    "state": "R/M",
                    "state_display": "Online",
                },
            ],
        )
        return SimpleNamespace(
            success=True, collect_run_uuid="resource-run", error_message=None
        )

    def fake_fit_collect(
        ac_device,
        site_name,
        repository=None,
        paths=None,
        max_workers=None,
        progress=None,
        item_progress=None,
        target_ap_uuids=None,
        target_ap_macs=None,
        target_ap_names=None,
        target_stations=None,
        should_cancel=None,
    ):
        fit_calls.append(
            (target_ap_uuids, target_ap_macs, target_ap_names, target_stations)
        )
        return FitApOpticalCollectResult(
            True, False, str(ac_device.device_uuid), "fit-run", 1, 0, None
        )

    FakeOpticalConnection.instances = []
    monkeypatch.setattr(
        trackside_optical_collection, "collect_h3c_ac_resources", fake_resource_collect
    )
    monkeypatch.setattr(
        trackside_optical_collection, "collect_h3c_fit_ap_optical", fake_fit_collect
    )
    monkeypatch.setattr(
        trackside_optical_collection.netmiko_connection,
        "ConnectHandler",
        FakeOpticalConnection,
    )

    result = collect_trackside_optical(
        repository,
        "demo",
        paths,
        [
            {
                "site": "Station A",
                "ap_uuid": "ap-1",
                "ap_name": "AP1",
                "device_uuid": target_switch.device_uuid,
                "device_name": target_switch.name,
            }
        ],
        concurrency=DEFAULT_TRACKSIDE_OPTICAL_CONCURRENCY,
        target_ap_uuid="ap-1",
    )

    assert [connection.host for connection in FakeOpticalConnection.instances] == [
        "10.0.0.10"
    ]
    assert fit_calls == [(["ap-1"], None, None, ["Station A"])]
    assert result.scope == "ap"
    assert result.target_label == "AP1"
    assert result.target_ap_offline is True
    assert result.switch_scope == "ap_switch"
    assert result.switch_scope_reason == "current_trackside_row"
    assert result.station_switch_total == 1
    with (result.session_dir / "session_meta.json").open(encoding="utf-8") as handle:
        meta = json.load(handle)
    assert meta["target_ap_offline"] is True
    assert meta["switch_scope"] == "ap_switch"


def test_trackside_progress_tracker_prevents_switch_branch_from_reaching_full_progress():
    events: list[tuple[int, int, dict[str, object]]] = []
    tracker = trackside_optical_collection.TracksideOpticalProgressTracker(
        switch_total=26,
        progress_callback=lambda current, total, details: events.append((current, total, details)),
    )
    target = trackside_optical_collection.TracksideOpticalTarget(
        key="device:1",
        name="SW",
        host="10.0.0.10",
        port=22,
        protocol="ssh",
        target_type="SWITCH",
        group_name="车站",
        device=Device(name="SW"),
    )

    tracker.handle_fit_ap_event({"event": "plan_ready", "phase": "fit_ap_optical", "total": 974, "ac_device_uuid": "ac-1"})
    for _index in range(26):
        tracker.mark_switch_completed(
            trackside_optical_collection.TracksideDeviceCollectionResult(target, True),
            persist_elapsed_ms=125,
        )

    current, total, details = events[-1]
    assert details["logical_total"] == 1000
    assert current == 26
    assert total == 1001
    assert current < total
    assert details["prevent_running_100"] is True
    assert details["elapsed_ms"] == 125


def test_trackside_progress_tracker_counts_retried_ap_only_once():
    events: list[tuple[int, int, dict[str, object]]] = []
    tracker = trackside_optical_collection.TracksideOpticalProgressTracker(
        switch_total=0,
        progress_callback=lambda current, total, details: events.append((current, total, details)),
    )

    tracker.handle_fit_ap_event({"event": "plan_ready", "phase": "fit_ap_optical", "total": 1, "ac_device_uuid": "ac-1"})
    tracker.handle_fit_ap_event({"event": "ap_completed", "phase": "fit_ap_optical", "ap_identity": "ap:1", "ap_name": "AP-1", "status": "failed"})
    tracker.handle_fit_ap_event({"event": "ap_retry_started", "phase": "fit_ap_optical", "ap_identity": "ap:1", "ap_name": "AP-1", "status": "retrying"})
    tracker.handle_fit_ap_event({"event": "ap_completed", "phase": "fit_ap_optical", "ap_identity": "ap:1", "ap_name": "AP-1", "status": "success"})

    current, _total, details = events[-1]
    assert current == 1
    assert details["fit_ap_completed"] == 1
    assert details["success_count"] == 1
    assert details["failed_count"] == 0


def test_trackside_ap_business_export_excludes_internal_fields(tmp_path):
    from openpyxl import load_workbook

    rows = [
        {
            "site": "Station A",
            "device_name": "HX_1",
            "interface_name": "GigabitEthernet1/0/1",
            "ap_ip": "10.0.0.10",
            "host": "10.0.0.10",
            "source_device": "AC-1",
            "collection_status": "success",
        }
    ]
    i18n = I18n("en_US")
    export_path = tmp_path / "trackside_hidden.xlsx"

    export_trackside_ap_business_xlsx(
        export_path,
        rows,
        TRACKSIDE_AP_BUSINESS_COLUMNS,
        [i18n.t(key) for key, _field in TRACKSIDE_AP_BUSINESS_COLUMNS],
    )

    sheet = load_workbook(export_path)["轨旁AP业务"]
    headers = [cell.value for cell in sheet[1]]
    assert "TX功率" not in headers
    assert "IP Address" not in headers
    assert "Host" not in headers
    assert "Host Address" not in headers
    assert "Management IP" not in headers
    assert "Source Device" not in headers
    assert "Collection Status" not in headers
    assert "Match Source" not in headers
    assert "trackside.collection_status" not in headers
    assert "collection_status" not in headers
    assert "source_device" not in headers
    assert "host" not in headers


def test_trackside_ap_business_export_formats_missing_ap_side_as_dash(tmp_path):
    from openpyxl import load_workbook

    rows = [
        {
            "site": "Station A",
            "device_name": "HX_1",
            "interface_name": "GigabitEthernet1/0/1",
            "switch_optical_status": "normal",
            "ap_mac": None,
            "ap_name": None,
            "ap_rx_power": None,
            "ap_tx_power": None,
            "ap_optical_status": "no_module",
            "ap_side_has_data": False,
        },
        {
            "site": "Station A",
            "device_name": "HX_1",
            "interface_name": "GigabitEthernet1/0/2",
            "switch_optical_status": "normal",
            "ap_mac": "bc5a-3457-cbe1",
            "ap_name": "AP23",
            "ap_rx_power": None,
            "ap_tx_power": None,
            "ap_optical_status": "no_module",
            "raw_status": "no module",
            "ap_side_has_data": True,
        },
    ]
    i18n = I18n("zh_CN")
    export_path = tmp_path / "trackside_ap_missing.xlsx"

    export_trackside_ap_business_xlsx(
        export_path,
        rows,
        TRACKSIDE_AP_BUSINESS_COLUMNS,
        [i18n.t(key) for key, _field in TRACKSIDE_AP_BUSINESS_COLUMNS],
    )

    sheet = load_workbook(export_path)["轨旁AP业务"]
    headers = [cell.value for cell in sheet[1]]
    alarm_column = headers.index("AP业务光衰") + 1
    ap_mac_column = headers.index("AP_MAC") + 1
    ap_name_column = headers.index("AP名称") + 1
    ap_rx_column = headers.index("AP侧收光(dBm)") + 1
    assert "TX功率" not in headers
    assert sheet.cell(2, alarm_column).value == "未知"
    assert sheet.cell(2, ap_mac_column).value == "-"
    assert sheet.cell(2, ap_name_column).value == "-"
    assert sheet.cell(2, ap_rx_column).value == "-"
    assert sheet.cell(2, alarm_column).value != "无光模块"
    assert sheet.cell(3, alarm_column).value == "未知"


def test_demo_data_contains_ac_management_rows(tmp_path):
    context = create_demo_context(PathResolver(tmp_path))
    ac = next(
        device
        for device in context.repository.list(device_type="AC")
        if device.ip_address == "10.0.0.51"
    )
    repository = AcRepository(context.database)

    assert (
        repository.get_ac_ap_summary(ac.device_uuid)["remaining_local_ap_licenses"]
        == 59998
    )
    assert repository.get_ac_ap_summary(ac.device_uuid)["cpu_usage"] == "16%"
    assert [
        row["ap_name"] for row in repository.list_fit_ap_resources(ac.device_uuid)
    ] == ["4c6f-d608-0400", "4c6f-de4b-0500"]
    assert len(repository.list_fit_ap_optical(ac.device_uuid)) == 2
    assert repository.get_fit_ap_metadata("4c6f-d608-0400")["site_name"] == "体育中心站"


def test_import_and_export_fit_ap_metadata(tmp_path):
    from openpyxl import Workbook

    repository = AcRepository(make_database(tmp_path))
    repository.replace_fit_ap_resources(
        "ac-1",
        [{"ap_name": "ap-a", "ap_mac": "30f5-277a-1b00", "serial_number": "SN-1"}],
    )
    service = FitApImportExportService(repository)
    import_path = tmp_path / "metadata.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(AP_EXTENSION_TEMPLATE_FIELDS)
    sheet.append(
        [
            "renamed-ap",
            "30F5:277A:1B00",
            "站点",
            "Station X",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "K12+450",
            "Platform",
            "上下行",
            "",
        ]
    )
    workbook.save(import_path)

    result = service.import_metadata_file(import_path)

    assert result.updated == 1
    assert result.skipped == 0
    entity = repository.list_ap_entities("ac-1")[0]
    assert entity["station"] == "Station X"
    assert entity["milestone"] == "12450"
    assert entity["location_note"] == "Platform"
    assert entity["direction"] == "上下行"

    export_path = tmp_path / "export.csv"
    service.export_ap_csv(
        export_path,
        [
            {
                "ap_name": "ap-a",
                "ap_ip": "10.0.0.1",
                "state_display": "运行(主)",
                "site": "体育中心站",
                "mileage": "1020",
                "direction": "上行",
            }
        ],
    )
    text = export_path.read_text(encoding="utf-8-sig")
    assert "AP名称" in text
    assert "ap-a" in text
    assert "YDK1+020" in text
    assert "LLDP" not in text
    assert "BSSID" not in text
    assert "RID1信道" in text
    assert "RID2信道" in text
    assert "RID3信道" not in text
    headers = text.splitlines()[0].split(",")
    assert (
        headers.index("RID2功率")
        < headers.index("归属站点")
        < headers.index("更新时间")
    )
    assert "归属区间" in headers
    assert "归属类型" in headers


def test_import_ap_extension_metadata_skips_empty_or_unmatched_mac(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    repository.replace_fit_ap_resources(
        "ac-1", [{"ap_name": "ap-a", "ap_mac": "30f5-277a-1b00"}]
    )
    service = FitApImportExportService(repository)

    result = service.import_metadata_rows(
        AP_EXTENSION_TEMPLATE_FIELDS,
        [
            ["ap-a", "", "Empty Station", "", "", ""],
            ["ap-a", "ffff-ffff-ffff", "Unknown Station", "", "", ""],
        ],
    )

    assert result.updated == 0
    assert result.skipped == 2
    assert any("AP_MAC is empty or invalid" in error for error in result.errors)
    assert any("not matched" in error for error in result.errors)
    assert repository.list_ap_entities("ac-1")[0]["station"] == ""


def test_import_ap_extension_metadata_rejects_legacy_matching_headers(tmp_path):
    repository = AcRepository(make_database(tmp_path))
    repository.replace_fit_ap_resources(
        "ac-1", [{"ap_name": "ap-a", "ap_mac": "30f5-277a-1b00"}]
    )
    service = FitApImportExportService(repository)

    with pytest.raises(ValueError, match="Unsupported AP metadata template header"):
        service.import_metadata_rows(
            ["AP名称", "归属站点", "里程", "点位说明", "上下行"],
            [["ap-a", "Legacy Station", "", "", ""]],
        )

    assert repository.list_ap_entities("ac-1")[0]["station"] == ""


def test_export_ap_extension_template_xlsx_contains_editable_headers_and_entity_station(
    tmp_path,
):
    from openpyxl import load_workbook

    service = FitApImportExportService(AcRepository(make_database(tmp_path)))
    export_path = tmp_path / "ap_extension_template.xlsx"

    service.export_ap_extension_template_xlsx(
        export_path,
        [
            {
                "ap_uuid": "ap-1",
                "ap_name": "AP-1",
                "apid": "101",
                "ap_ip": "10.0.0.10",
                "ap_mac": "0011-2233-4455",
                "model": "WA6338",
                "serial_number": "SN-001",
                "state_display": "Idle",
                "site": "Resource Station",
            }
        ],
        [
            {
                "ap_uuid": "ap-1",
                "ap_mac": "0011-2233-4455",
                "station": "Entity Station",
                "direction": "uplink",
                "milestone": "K12+450",
                "location_note": "platform",
            }
        ],
    )

    sheet = load_workbook(export_path).active
    headers = [cell.value for cell in sheet[1]]

    assert headers == AP_EXTENSION_TEMPLATE_FIELDS
    assert "归属站点" in headers
    for forbidden in (
        "AP_IP",
        "APID",
        "SN",
        "型号",
        "状态",
        "AP状态",
        "AP组",
        "在线时长",
        "更新时间",
    ):
        assert forbidden not in headers
    assert "站点/位置" not in headers
    assert "site" not in headers
    assert "site_name" not in headers
    assert "station" not in headers
    assert sheet.max_row == 2
    assert sheet["A2"].value == "AP-1"
    assert sheet["B2"].value == "0011-2233-4455"
    assert sheet["D2"].value == "Entity Station"
    assert sheet["L2"].value == "K12+450"
    assert sheet["M2"].value == "platform"
    assert sheet["N2"].value == "uplink"
    assert sheet["A1"].font.bold
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:O2"


def test_export_ap_extension_template_xlsx_allows_empty_template(tmp_path):
    from openpyxl import load_workbook

    service = FitApImportExportService(AcRepository(make_database(tmp_path)))
    export_path = tmp_path / "empty_ap_extension_template.xlsx"

    service.export_ap_extension_template_xlsx(export_path, [], [])

    sheet = load_workbook(export_path).active
    assert [cell.value for cell in sheet[1]] == AP_EXTENSION_TEMPLATE_FIELDS
    assert sheet.max_row == 1


def test_neighbor_matcher_matches_sysname_mac_and_rx_power(tmp_path):
    paths = PathResolver(tmp_path)
    database = make_database(paths.site_db_path("demo").parent)
    device_repository = DeviceRepository(database)
    device = device_repository.create(
        Device(name="HX Device", sysname="HX_1", ip_address="10.0.0.2")
    )
    fact_repository = DeviceFactRepository(database)
    fact_repository.replace_device_interfaces(
        device.device_uuid,
        [
            {
                "interface_name": "GigabitEthernet2/0/19",
                "mac_address": "903f-8645-6e00",
                "collected_at": "2026-01-01T00:00:00",
                "updated_at": "2026-01-01T00:00:00",
            }
        ],
    )
    fact_repository.replace_optical_modules(
        device.device_uuid,
        [
            {
                "interface_name": "GigabitEthernet2/0/19",
                "rx_power": "-6.66 dBm",
                "collected_at": "2026-01-01T00:00:00",
                "updated_at": "2026-01-01T00:00:00",
            }
        ],
    )

    by_sysname = match_neighbor_device("demo", neighbor_sysname="HX_1", paths=paths)
    by_mac = match_neighbor_device("demo", neighbor_mac="903f-8645-6e00", paths=paths)

    assert by_sysname.device_uuid == device.device_uuid
    assert by_sysname.matched_by == "sysname"
    assert by_sysname.station is None
    assert by_mac.device_uuid == device.device_uuid
    assert by_mac.matched_by == "mac"
    assert (
        find_neighbor_rx_power(
            "demo", device.device_uuid, "GigabitEthernet2/0/19", paths=paths
        )
        == "-6.66 dBm"
    )


def test_neighbor_matcher_ignores_generic_h3c_neighbor_name(tmp_path):
    paths = PathResolver(tmp_path)
    database = make_database(paths.site_db_path("demo").parent)
    DeviceRepository(database).create(
        Device(name="H3C", sysname="H3C", station="Station A", ip_address="10.0.0.2")
    )

    result = match_neighbor_device("demo", neighbor_sysname="H3C", paths=paths)

    assert result.match_status == "unresolved"
    assert result.device_uuid is None


def test_neighbor_matcher_reports_ambiguous_chassis_mac(tmp_path):
    paths = PathResolver(tmp_path)
    database = make_database(paths.site_db_path("demo").parent)
    device_repository = DeviceRepository(database)
    first = device_repository.create(Device(name="SW-1", ip_address="10.0.0.2"))
    second = device_repository.create(Device(name="SW-2", ip_address="10.0.0.3"))
    facts = DeviceFactRepository(database)
    for device in (first, second):
        facts.replace_device_interfaces(
            device.device_uuid,
            [{"interface_name": "GE1/0/1", "mac_address": "903f-8645-6e00"}],
        )

    result = match_neighbor_device("demo", neighbor_mac="903f-8645-6e00", paths=paths)

    assert result.match_status == "ambiguous"
    assert result.candidate_count == 2
    assert result.device_uuid is None


def test_fit_ap_neighbor_evidence_prefers_direct_mac_and_marks_port_conflict():
    direct = NeighborMatchResult(
        device_uuid="switch-direct",
        device_name="SW-direct",
        matched_by="mac",
        match_status="matched",
    )
    reverse = NeighborMatchResult(
        device_uuid="switch-direct",
        device_name="SW-direct",
        local_interface="GigabitEthernet1/0/9",
        matched_by="device_lldp",
        match_status="matched",
    )

    selected, conflict = _select_fit_ap_neighbor_match(
        direct,
        reverse,
        "GigabitEthernet1/0/4",
    )

    assert selected is direct
    assert conflict is True


def test_fit_ap_association_repair_is_dry_run_by_default_and_keeps_history(tmp_path):
    paths = PathResolver(tmp_path)
    database = make_database(paths.site_db_path("demo").parent)
    repository = AcRepository(database)
    repository.replace_fit_ap_resources(
        "ac-1",
        [{"ap_uuid": "ap-1", "ap_name": "AP-1", "ap_mac": "28c9-7a3e-5da0"}],
    )
    repository.replace_fit_ap_optical(
        "ac-1",
        [
            {
                "ap_uuid": "ap-1",
                "ap_name": "AP-1",
                "ap_mac": "28c9-7a3e-5da0",
                "status": "success",
                "rx_power": "-7.1",
                "neighbor_device_name": "GE1/0/1",
                "neighbor_interface": "H3C",
                "lldp_neighbor_name": "H3C",
                "lldp_neighbor_interface": "GigabitEthernet1/0/4",
            }
        ],
    )
    before_history = len(repository.list_fit_ap_optical_history_by_ap("ap-1"))

    preview = repository.repair_invalid_fit_ap_association_projection("ac-1")
    assert preview["applied"] is False
    assert preview["candidate_count"] == 1
    assert repository.list_fit_ap_optical("ac-1")[0]["neighbor_interface"] == "H3C"

    result = repository.repair_invalid_fit_ap_association_projection("ac-1", apply=True)
    assert result["applied"] is True
    assert result["cleared_optical_rows"] == 1
    repaired = repository.list_fit_ap_optical("ac-1")
    assert len(repaired) == 1
    assert repaired[0]["rx_power"] == "-7.1"
    assert repaired[0]["neighbor_interface"] == ""
    assert repository.get_fit_ap_resource_by_uuid("ac-1", "ap-1")["lldp_neighbor_interface"] == ""
    assert len(repository.list_fit_ap_optical_history_by_ap("ap-1")) == before_history


def test_neighbor_matcher_reverse_matches_ap_mac_from_device_lldp(tmp_path):
    paths = PathResolver(tmp_path)
    database = make_database(paths.site_db_path("demo").parent)
    device_repository = DeviceRepository(database)
    device = device_repository.create(
        Device(name="HX Switch", station="Station A", ip_address="10.0.0.2")
    )
    fact_repository = DeviceFactRepository(database)
    fact_repository.replace_lldp_neighbors(
        device.device_uuid,
        [
            {
                "local_interface": "GE2/0/22",
                "neighbor_sysname": "bc5a-3457-cbe0",
                "neighbor_mac": "bc5a-3457-cbe0",
                "neighbor_interface": "GigabitEthernet1/0/2",
                "collected_at": "2026-01-01T00:00:00",
                "updated_at": "2026-01-01T00:00:00",
            }
        ],
    )

    match = match_ap_from_device_lldp(
        "demo", ap_mac="bc5a-3457-cbe0", paths=paths
    )

    assert match.device_uuid == device.device_uuid
    assert match.device_name == "HX Switch"
    assert match.station == "Station A"
    assert match.local_interface == "GE2/0/22"
    assert match.ap_interface == "GigabitEthernet1/0/2"


def test_neighbor_matcher_reverse_matches_ap_sysname_from_device_lldp(tmp_path):
    paths = PathResolver(tmp_path)
    database = make_database(paths.site_db_path("demo").parent)
    device_repository = DeviceRepository(database)
    device = device_repository.create(
        Device(name="HX Switch", station="Station B", ip_address="10.0.0.2")
    )
    DeviceFactRepository(database).replace_lldp_neighbors(
        device.device_uuid,
        [
            {
                "local_interface": "GE2/0/23",
                "neighbor_sysname": "bc5a-3457-cbe1",
                "neighbor_mac": "",
                "neighbor_interface": "GigabitEthernet1/0/2",
                "collected_at": "2026-01-01T00:00:00",
                "updated_at": "2026-01-01T00:00:00",
            }
        ],
    )

    match = match_ap_from_device_lldp(
        "demo", ap_name="bc5a-3457-cbe1", paths=paths
    )

    assert match.device_uuid is None
    assert match.local_interface is None


def test_neighbor_optical_module_matches_interface_alias(tmp_path):
    paths = PathResolver(tmp_path)
    database = make_database(paths.site_db_path("demo").parent)
    device_repository = DeviceRepository(database)
    device = device_repository.create(
        Device(name="HX Switch", station="Station A", ip_address="10.0.0.2")
    )
    DeviceFactRepository(database).replace_optical_modules(
        device.device_uuid,
        [
            {
                "interface_name": "GigabitEthernet2/0/22",
                "rx_power": "-4.44 dBm",
                "collected_at": "2026-01-01T00:00:00",
                "updated_at": "2026-01-01T00:00:00",
            }
        ],
    )

    module = find_neighbor_optical_module(
        "demo", device.device_uuid, "GE2/0/22", paths=paths
    )

    assert normalize_interface_name("GE2/0/22") == "GigabitEthernet2/0/22"
    assert normalize_interface_name("XGE1/0/49") == "Ten-GigabitEthernet1/0/49"
    assert module["rx_power"] == "-4.44 dBm"


def test_neighbor_matcher_prefers_fact_sysname_and_returns_station(tmp_path):
    paths = PathResolver(tmp_path)
    database = make_database(paths.site_db_path("demo").parent)
    device_repository = DeviceRepository(database)
    device = device_repository.create(
        Device(
            name="HX Device",
            sysname="HX_DEVICE",
            station="Station A",
            ip_address="10.0.0.2",
        )
    )
    fact_repository = DeviceFactRepository(database)
    fact_repository.upsert_device_fact(
        {
            "device_uuid": device.device_uuid,
            "sysname": "HX_1",
            "collected_at": "2026-01-01T00:00:00",
            "updated_at": "2026-01-01T00:00:00",
        }
    )

    match = match_neighbor_device(
        "demo", neighbor_sysname="HX_1", paths=paths
    )

    assert match.device_uuid == device.device_uuid
    assert match.device_name == "HX Device"
    assert match.station == "Station A"


def test_fit_ap_optical_command_guard_allows_only_expected_commands():
    allowed = {
        "screen-length disable",
        "display lldp neighbor-information list",
        "display transceiver diagnosis interface",
        "display transceiver interface",
        "display transceiver manuinfo interface",
    }

    assert all(
        command_guard.is_command_allowed(command, "fit_ap_collect")
        for command in allowed
    )
    assert (
        command_guard.is_command_allowed("display interface", "fit_ap_collect") is False
    )
    assert command_guard.is_command_allowed("reboot", "fit_ap_collect") is False


def test_ac_log_event_names_do_not_contain_password():
    event_names = (
        "AC_COLLECT_STARTED",
        "AC_COLLECT_SUCCESS",
        "AC_COLLECT_FAILED",
        "FIT_AP_RESOURCE_UPDATED",
        "FIT_AP_OPTICAL_STARTED",
        "FIT_AP_OPTICAL_SUCCESS",
        "FIT_AP_OPTICAL_PARTIAL_SUCCESS",
        "FIT_AP_OPTICAL_FAILED",
        "FIT_AP_TELNET_FAILED",
    )

    assert all("PASSWORD" not in event for event in event_names)


def test_database_runtime_has_no_legacy_migration_chain():
    text = (
        Path(__file__).parents[1] / "src" / "netconsole" / "core" / "database.py"
    ).read_text(encoding="utf-8")

    assert "ALTER TABLE ac_trackside_ap_plan ADD COLUMN remark TEXT" in text
    assert (
        "ALTER TABLE rail_ap_vlan_allocations_reference_migration\n"
        "            RENAME TO rail_ap_vlan_allocations"
    ) in text
    assert text.count("DROP TABLE") == 4
    assert 'conn.execute("DROP TABLE rail_ap_vlan_allocations")' in text
    assert 'conn.execute("DROP TABLE ac_fit_ap_resources")' in text
    assert 'SELECT COUNT(*) AS count FROM ac_fit_ap_resources' in text
    assert "旧 ac_fit_ap_resources 表缺少身份字段，且包含数据，拒绝无损迁移" in text
    assert "migrate_old_" not in text
    assert "legacy_table_adapter" not in text


def test_build_device_optical_status_lookup_indexes_by_name_and_sysname(tmp_path):
    """The lookup must resolve both device.name and device.sysname."""
    database = make_database(tmp_path / "data" / "sites" / "demo" / "db")
    device_repository = DeviceRepository(database)
    device = device_repository.create(
        Device(name="HX Switch", sysname="HX_1", ip_address="10.0.0.2")
    )
    fact_repository = DeviceFactRepository(database)
    fact_repository.replace_optical_modules(
        device.device_uuid,
        [
            {
                "interface_name": "GigabitEthernet2/0/10",
                "status": "warning",
                "collected_at": "2026-01-01T00:00:00",
            }
        ],
    )

    devices = device_repository.list()
    optical_by_device = {
        str(device.device_uuid): fact_repository.list_optical_modules(
            device.device_uuid
        )
    }
    lookup = build_switch_data_lookup(devices, optical_by_device)

    # Both name and sysname should resolve to the raw optical module row
    name_result = lookup.get(("hx switch", "gigabitethernet2/0/10"))
    assert name_result is not None
    assert name_result.get("collected_at") == "2026-01-01T00:00:00"
    sysname_result = lookup.get(("hx_1", "gigabitethernet2/0/10"))
    assert sysname_result is not None
    # Non-existent name should not resolve
    assert lookup.get(("nonexistent", "gigabitethernet2/0/10")) is None


def test_trackside_ap_optical_status_computes_from_raw_data():
    """Trackside ap_optical_status must be computed real-time from raw FIT-AP rx_power."""
    switch = Device(name="HX_1", station="Station A", device_uuid="sw-1")
    rows = build_trackside_ap_business_rows(
        [switch],
        {
            "sw-1": [
                {"interface_name": "GigabitEthernet2/0/10", "description": "To_AP10"}
            ]
        },
        {"sw-1": [{"interface_name": "GigabitEthernet2/0/10", "rx_power": "-6.10"}]},
        [
            {
                "ap_uuid": "ap-10",
                "ap_mac": "bc5a-3457-cbe0",
                "ap_name": "AP10",
                "neighbor_device_name": "HX_1",
                "neighbor_interface": "GigabitEthernet2/0/10",
                "rx_power": "-20.32",
                "rx_low_alarm": "-20.00",
                "rx_low_warning": "-17.00",
            }
        ],
    )
    # AP business status is independent of the FIT-AP resource module threshold.
    assert rows[0]["ap_device_optical_status"] == "alarm"
    assert rows[0]["ap_optical_status"] == "abnormal"
    # switch_optical_status is computed real-time: -6.10 → normal
    assert rows[0]["switch_optical_status"] == "normal"


def test_device_management_projection_keeps_fit_ap_module_threshold():
    switch = Device(name="HX_1", station="Station A", device_uuid="sw-1")

    rows = build_trackside_ap_business_rows(
        [switch],
        {
            "sw-1": [
                {
                    "interface_name": "GigabitEthernet2/0/10",
                    "description": "To_AP10",
                    "pvid": 71,
                    "vlan": "Native/PVID 71; Tagged 201",
                }
            ]
        },
        {},
        [
            {
                "ap_uuid": "ap-10",
                "ap_mac": "bc5a-3457-cbe0",
                "ap_name": "AP10",
                "neighbor_device_name": "HX_1",
                "neighbor_interface": "GigabitEthernet2/0/10",
                "rx_power": "-17.80",
                "rx_low_alarm": "-28.20",
                "rx_low_warning": "-25.00",
                "data_freshness": "fresh",
            }
        ],
        business_projection=False,
    )

    assert rows[0]["vlan"] == "Tagged 201"
    assert rows[0]["ap_optical_status"] == "normal"
    assert "ap_business_optical_status" not in rows[0]
    assert "ap_business_threshold_dbm" not in rows[0]


# ── Unified State Architecture tests ──────────────────────────────────────────


def test_state_engine_compute_state_returns_unified_result():
    """compute_state must return a StateResult with all status fields populated."""
    from netconsole.core.state_engine import StateResult

    result = compute_state(
        {
            "switch_rx_power": "-14.35",
            "switch_alarm_low": "-19.00",
            "switch_warning_low": "-16.99",
            "fit_ap_row": {
                "rx_power": "-20.32",
                "rx_low_alarm": "-20.00",
                "rx_low_warning": "-17.00",
            },
        }
    )
    assert isinstance(result, StateResult)
    assert result.switch_status == "abnormal"
    assert result.ap_status == "abnormal"
    assert result.optical_status == "abnormal"  # fixed AP business threshold
    assert result.severity > 0
    assert result.color == STATUS_COLORS["alarm"]


def test_state_engine_minus_36_96_is_below_ap_business_threshold():
    """A valid AP RX value below -13.90 dBm is treated as attenuation."""
    fit_ap_row = {
        "rx_power": "-36.96",
        "rx_low_alarm": "-20.00",
        "rx_low_warning": "-17.00",
        "optical_alarm_status": "no_light",
    }
    result = compute_state({"fit_ap_row": fit_ap_row})
    assert result.ap_status == "abnormal"
    assert result.color == STATUS_COLORS["abnormal"]
    assert result.color == "FEE2E2"


def test_state_engine_minus_20_32_uses_fixed_ap_business_threshold():
    """rx_power = -20.32 is below the fixed AP business threshold."""
    fit_ap_row = {
        "rx_power": "-20.32",
        "rx_low_alarm": "-20.00",
        "rx_low_warning": "-17.00",
    }
    result = compute_state({"fit_ap_row": fit_ap_row})
    assert result.ap_status == "abnormal"
    assert result.color == STATUS_COLORS["abnormal"]
    assert result.color == "FEE2E2"


@pytest.mark.parametrize(
    ("reported_status", "expected"),
    [("critical", "abnormal"), ("no_light", "no_light"), ("链路断开", "link_down")],
)
def test_ap_status_preserves_explicit_backend_abnormal_statuses(
    reported_status: str,
    expected: str,
) -> None:
    assert compute_ap_status(
        {"rx_power": "-10.00", "optical_alarm_status": reported_status}
    ) == expected


def test_state_engine_link_down_unified_pink():
    """link_down / link_abnormal must produce unified pink colour."""
    for status in ("link_down", "link_abnormal"):
        result = compute_state(
            {
                "switch_rx_power": "-10.00",
                "switch_port_status": "DOWN",
                "fit_ap_row": {"rx_power": "-10.00", "ap_port_status": "DOWN"},
            }
        )
        assert result.optical_status in ("link_abnormal",)
        assert result.color == STATUS_COLORS["link_abnormal"]


def test_state_engine_three_pages_same_input_same_output():
    """Given the same input, FIT-AP / Trackside / DeviceDetail must produce identical statuses."""
    from netconsole.core.view_models import (
        FITAPViewModel,
        TracksideViewModel,
        DeviceDetailViewModel,
    )

    # Raw optical module data for switch side — computes to "alarm"
    switch_data_lookup = {
        ("hx_1", "gigabitethernet2/0/10"): {
            "rx_power": "-20.00",
            "rx_low_alarm": "-19.00",
            "rx_low_warning": "-16.99",
            "port_status": "access",
        }
    }

    fit_ap_vm = FITAPViewModel(switch_data_lookup=switch_data_lookup)
    trackside_vm = TracksideViewModel(switch_data_lookup=switch_data_lookup)
    device_vm = DeviceDetailViewModel(switch_data_lookup=switch_data_lookup)

    fit_ap_row = {
        "device_name": "HX_1",
        "local_interface": "GigabitEthernet2/0/10",
        "neighbor_device_name": "HX_1",
        "neighbor_interface": "GigabitEthernet2/0/10",
        "rx_power": "-14.35",
        "rx_low_alarm": "-19.00",
        "rx_low_warning": "-16.99",
    }

    # FIT-AP ViewModel
    fit_ap_result = fit_ap_vm.populate_row(fit_ap_row)
    assert fit_ap_result.switch_status == "alarm"
    assert fit_ap_result.ap_status == "abnormal"

    # Trackside ViewModel (using same FIT-AP row)
    trackside_row = {
        "device_name": "HX_1",
        "interface_name": "GigabitEthernet2/0/10",
    }
    trackside_result = trackside_vm.populate_row(trackside_row, fit_ap_row)
    assert trackside_result.switch_status == fit_ap_result.switch_status
    assert trackside_result.ap_status == fit_ap_result.ap_status
    assert trackside_result.color == fit_ap_result.color

    # DeviceDetail ViewModel (switch only)
    device_color = device_vm.get_color(
        device_name="HX_1",
        interface_name="GigabitEthernet2/0/10",
    )
    assert device_color == STATUS_COLORS["alarm"]
