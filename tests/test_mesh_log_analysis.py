from __future__ import annotations

import gzip
import json
import os
import sqlite3
import time
from datetime import datetime
from pathlib import Path

import numpy as np
import pytest
from matplotlib.dates import date2num
from openpyxl import load_workbook

from netconsole.core.database import Database
from netconsole.models.device import Device
from netconsole.models.mesh_log_models import EVENT_ACTIVE_SWITCH, EVENT_MULTI_ACTIVE, EVENT_NO_ACTIVE
from netconsole.parsers import mesh_log_parser
from netconsole.parsers.mesh_log_parser import MeshLogParser, calculate_signal
from netconsole.core.paths import PathResolver
from netconsole.repositories.device_group_repository import DeviceGroupRepository
from netconsole.repositories.device_repository import DeviceRepository
from netconsole.repositories.mesh_catalog_repository import MeshCatalogRepository
from netconsole.repositories.mesh_mr_repository import MeshMrRepository
from netconsole.models.mesh_analysis_params import mesh_analysis_params_to_json
from netconsole.services.mesh_log_analysis_service import MeshLogAnalysisService
from netconsole.services.mesh_import_service import MeshImportService
from netconsole.services.mesh_peer_mapping_service import MeshPeerMappingService
from netconsole.services.mesh_storage_service import MeshStorageService
from netconsole.services.network_tools.trackside_bssid_resolver import TracksideApBssidResolver
from netconsole.services.rail_transit.constants import VEHICLE_MR_GROUP_NAME


LINE_A = "[1] Active 30f5-277a-5a2f 2025/12/03 10:12:30 0d 00h 00m 03s 1 36/43 2%/4% 45%/47% 3/1 15/27 60/72060 88/105 0/5000 2/297 314/0 0/93 0/0 0/0 0/0"
LINE_B = "[1] Active 30f5-277a-5a3f 2025/12/03 10:12:30 0d 00h 00m 03s 1 37/44 2%/4% 45%/47% 3/1 15/27 60/72060 88/105 0/5000 2/297 314/0 0/93 0/0 0/0 0/0"
LINE_STANDBY = "[1] Standy 30f5-277a-5a4f 2025/12/03 10:12:30 0d 00h 00m 03s 1 30/31 2%/4% 45%/47% 3/1 15/27 60/72060 88/105 0/5000 2/297 314/0 0/93 0/0 0/0 0/0"


def test_parse_timestamp_and_standy(tmp_path):
    path = tmp_path / "14CW-01-2026_02_01_1meshlog.log"
    path.write_text("[1] 2025/12/03 10:12:33.579 (3)\n" + LINE_STANDBY + "\n", encoding="utf-8")
    _, records, issues = MeshLogParser().parse_file(path)
    assert not [issue for issue in issues if issue.issue_type in {"缺少采样时间", "无法识别链路状态"}]
    assert records[0].radio == 1
    assert records[0].sample_time == datetime(2025, 12, 3, 10, 12, 33, 579000)
    assert records[0].timestamp_tag == "3"
    assert records[0].link_state_raw == "Standy"
    assert records[0].link_state == "STANDBY"


def test_signal_dbm_calculation():
    metrics = {"local_rssi_db": 36, "peer_rssi_db": 43, "local_noise_raw": 88, "peer_noise_raw": 105}
    assert calculate_signal(metrics) == (-88, -105, -52, -62)


def test_active_switch_observed_window(tmp_path):
    path = tmp_path / "14CW-01-2026_06_19-meshlog.log"
    path.write_text(
        "\n".join(
            [
                "[1] 2025/12/03 10:12:35.964",
                LINE_A,
                "[1] 2025/12/03 10:12:36.681",
                LINE_B,
            ]
        ),
        encoding="utf-8",
    )
    result = MeshLogAnalysisService("demo", tmp_path).analyze([path])
    switches = [event for event in result.switch_events if event.event_type == EVENT_ACTIVE_SWITCH]
    assert len(switches) == 1
    assert switches[0].from_peer_mac == "30f5-277a-5a2f"
    assert switches[0].to_peer_mac == "30f5-277a-5a3f"
    assert switches[0].observed_window_ms == 717


def test_no_active_and_multi_active(tmp_path):
    path = tmp_path / "meshlog.log"
    path.write_text(
        "\n".join(
            [
                "[1] 2025/12/03 10:12:33.000",
                LINE_STANDBY,
                "[1] 2025/12/03 10:12:34.000",
                LINE_A,
                LINE_B,
            ]
        ),
        encoding="utf-8",
    )
    result = MeshLogAnalysisService("demo", tmp_path).analyze([path])
    assert any(event.event_type == EVENT_NO_ACTIVE for event in result.switch_events)
    assert any(event.event_type == EVENT_MULTI_ACTIVE for event in result.switch_events)


def test_gzip_matches_plain_text(tmp_path):
    text = "[1] 2025/12/03 10:12:33.579 (3)\n" + LINE_A + "\n"
    plain = tmp_path / "a-meshlog.log"
    gz = tmp_path / "a-meshlog.log.gz"
    plain.write_text(text, encoding="utf-8")
    with gzip.open(gz, "wt", encoding="utf-8") as file:
        file.write(text)
    _, plain_records, _ = MeshLogParser().parse_file(plain)
    _, gz_records, _ = MeshLogParser().parse_file(gz)
    assert plain_records[0].metrics == gz_records[0].metrics
    assert plain_records[0].local_signal_dbm == gz_records[0].local_signal_dbm


def test_duplicate_records_are_removed(tmp_path):
    first = tmp_path / "14CW-01-2026_02_01_1meshlog.log"
    second = tmp_path / "14CW-01-2026_02_01_2meshlog.log"
    content = "[1] 2025/12/03 10:12:33.579 (3)\n" + LINE_A + "\n"
    first.write_text(content, encoding="utf-8")
    second.write_text(content, encoding="utf-8")
    result = MeshLogAnalysisService("demo", tmp_path).analyze([first, second])
    assert result.summary.raw_record_count == 2
    assert result.summary.record_count == 1
    assert result.summary.duplicate_record_count == 1


def test_dynamic_same_filename_archives_without_overwrite(tmp_path):
    paths = PathResolver(tmp_path)
    storage = MeshStorageService("demo", paths)
    profile = storage.create_mr_profile("14CW-01")
    source = tmp_path / "meshlog.log"
    source.write_text("[1] 2025/12/03 10:12:33.579\n" + LINE_A + "\n", encoding="utf-8")
    MeshImportService("demo", paths).import_files(profile, [source])
    source.write_text("[1] 2025/12/03 10:12:34.579\n" + LINE_B + "\n", encoding="utf-8")
    MeshImportService("demo", paths).import_files(profile, [source])
    archived = list(paths.mesh_mr_raw_dir("demo", profile.safe_folder_name).rglob("meshlog*.log"))
    assert len(archived) == 2
    assert len({item.name for item in archived}) == 2


def test_mesh_page_syncs_vehicle_mr_group_profiles_in_natural_order(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.ui.pages.mesh_log_analysis_page import MeshLogAnalysisPage

    database = Database(tmp_path / "devices.db")
    database.initialize()
    repository = DeviceRepository(database)
    groups = DeviceGroupRepository(database, "demo")
    onboard = groups.create(VEHICLE_MR_GROUP_NAME)
    station = groups.create("车站")
    repository.create(Device(name="MR10", primary_address="192.0.2.10", group_id=onboard.id))
    mr2 = repository.create(Device(name="MR2", primary_address="192.0.2.2", group_id=onboard.id))
    repository.create(Device(name="SW1", primary_address="192.0.2.20", group_id=station.id))

    page = MeshLogAnalysisPage(repository, I18n("zh_CN"), "demo", PathResolver(tmp_path))

    assert page.has_loaded is False
    assert page.mr_table.rowCount() == 0
    page.refresh_all()
    assert [page.mr_table.item(row, 0).text() for row in range(page.mr_table.rowCount())] == ["MR2", "MR10"]
    assert page.create_mr_button.parent() is None
    assert Path(tmp_path / "data" / "sites" / "demo" / "files" / "rail_transit" / "mr_raw_mesh" / "MR2" / "raw").exists()
    assert Path(tmp_path / "data" / "sites" / "demo" / "files" / "rail_transit" / "mr_raw_mesh" / "MR2" / "parsed").exists()
    assert Path(tmp_path / "data" / "sites" / "demo" / "files" / "rail_transit" / "mr_raw_mesh" / "MR2" / "outputs").exists()

    repository.update_group(int(mr2.id), station.id)
    page.refresh_all()

    assert [page.mr_table.item(row, 0).text() for row in range(page.mr_table.rowCount())] == ["MR10"]
    assert Path(tmp_path / "data" / "sites" / "demo" / "files" / "rail_transit" / "mr_raw_mesh" / "MR2").exists()


def test_mesh_page_first_show_empty_state_exits_loading(tmp_path):
    qt_app = _app()
    from netconsole.core.i18n import I18n
    from netconsole.ui.pages.mesh_log_analysis_page import MeshLogAnalysisPage

    database = Database(tmp_path / "devices.db")
    database.initialize()
    page = MeshLogAnalysisPage(DeviceRepository(database), I18n("zh_CN"), "demo", PathResolver(tmp_path))

    page.first_show_refresh(force=True)
    deadline = time.time() + 1.0
    while page.is_loading and time.time() < deadline:
        qt_app.processEvents()
        time.sleep(0.01)

    assert page.is_loading is False
    assert page.page_state == "empty"
    assert "暂无 MR 原始 MESH 日志" in page.progress_label.text()


def test_mesh_link_detail_export_writes_xlsx_with_centered_content(tmp_path, monkeypatch):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.ui.pages import mesh_log_analysis_page as page_module
    from netconsole.ui.pages.mesh_log_analysis_page import MeshLogAnalysisPage

    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("MR2")
    source = tmp_path / "meshlog.log"
    source.write_text("[1] 2025/12/03 10:12:33.000\n" + LINE_A + "\n" + LINE_STANDBY + "\n", encoding="utf-8")
    MeshImportService("demo", paths).import_files(profile, [source])
    page = MeshLogAnalysisPage(I18n("zh_CN"), "demo", paths)
    page.current_profile = profile
    target = paths.mesh_mr_export_dir("demo", profile.safe_folder_name) / "MR2_链路明细.xlsx"
    messages: list[str] = []
    monkeypatch.setattr(page_module.QFileDialog, "getSaveFileName", lambda *_args, **_kwargs: (str(target), "Excel Files (*.xlsx)"))
    monkeypatch.setattr(page_module.QMessageBox, "information", lambda *_args: messages.append(str(_args[-1])) or None)

    page.export_link_details()
    deadline = time.time() + 5
    while page.export_worker is not None and time.time() < deadline:
        _app().processEvents()
        time.sleep(0.01)
    _drain_qt_events()

    assert target.exists()
    workbook = load_workbook(target)
    assert set(workbook.sheetnames) >= {"导出说明", "统计汇总", "链路明细", "主链路建链顺序", "事件明细", "分析参数"}
    sheet = workbook["链路明细"]
    headers = [cell.value for cell in sheet[1]]
    assert "归属来源" not in headers
    assert "Peer Radio MAC" not in headers
    assert headers[:8] == ["序号", "采样时间", "Radio", "链路状态", "Peer MAC", "对端AP MAC", "对端AP名称", "归属站点"]
    assert sheet["A1"].font.bold
    assert sheet["A2"].alignment.horizontal == "center"
    assert sheet["A2"].alignment.vertical == "center"
    assert sheet["D2"].font.bold
    assert sheet["D2"].font.color.rgb.endswith("15803D")
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref
    assert sheet.column_dimensions["B"].width >= 24
    assert sheet.column_dimensions["E"].width >= 20
    build_sheet = workbook["主链路建链顺序"]
    assert build_sheet["A1"].value == "序号"
    assert workbook["分析参数"]["A1"].value == "统计项"
    assert build_sheet["Q1"].value == "配置切换时间(ms)"
    assert build_sheet["T1"].value == "建链结果"
    assert build_sheet["T2"].value == "短时建链"
    headers = [cell.value for cell in sheet[1]]
    for header in ("采样时间", "Radio", "Peer MAC", "链路状态", "对端射频口", "建链时间", "链路时长", "链路数量", "MR侧RSSI", "对端RSSI", "MR侧CPU", "对端CPU", "MR侧噪声", "对端噪声", "发送繁忙度", "接收繁忙度", "总发送繁忙度", "总接收繁忙度", "备注"):
        assert header in headers
    assert "归属来源" not in headers
    assert "Peer Radio MAC" not in headers
    state_index = headers.index("链路状态") + 1
    assert sheet.cell(2, state_index).value == "主链路"
    assert any("链路明细已导出" in message for message in messages)


def test_mesh_page_state_filter_defaults_to_raw_and_filters_active_standby(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.ui.pages.mesh_log_analysis_page import MeshLogAnalysisPage

    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-01")
    source = tmp_path / "meshlog.log"
    source.write_text("[1] 2025/12/03 10:12:33.000\n" + LINE_A + "\n" + LINE_STANDBY + "\n", encoding="utf-8")
    MeshImportService("demo", paths).import_files(profile, [source])
    page = MeshLogAnalysisPage(I18n("en_US"), "demo", paths)
    page.current_profile = profile

    page.tabs.setCurrentIndex(1)
    page.refresh_current_mr_data()
    _wait_for_mesh_tab_load(page)
    assert page._current_link_filters()["state"] is None
    assert page.link_table.rowCount() == 2

    page.state_filter.setCurrentIndex(1)
    page.refresh_link_table()
    _wait_for_mesh_tab_load(page)
    assert page._current_link_filters()["state"] == "ACTIVE"
    assert {page.link_table.item(row, 3).text() for row in range(page.link_table.rowCount())} == {"ACTIVE"}

    page.state_filter.setCurrentIndex(2)
    page.refresh_link_table()
    _wait_for_mesh_tab_load(page)
    assert page._current_link_filters()["state"] == "STANDBY"
    assert {page.link_table.item(row, 3).text() for row in range(page.link_table.rowCount())} == {"STANDBY"}

    page._clear_link_filters()
    assert page.state_filter.currentIndex() == 0
    assert page._current_link_filters()["state"] is None


def test_duplicate_sha_is_skipped(tmp_path):
    paths = PathResolver(tmp_path)
    storage = MeshStorageService("demo", paths)
    profile = storage.create_mr_profile("14CW-01")
    source = tmp_path / "meshlog.log"
    source.write_text("[1] 2025/12/03 10:12:33.579\n" + LINE_A + "\n", encoding="utf-8")
    service = MeshImportService("demo", paths)
    service.import_files(profile, [source])
    duplicate = service.import_files(profile, [source])
    repo = MeshMrRepository(paths.mesh_mr_db_path("demo", profile.safe_folder_name))
    assert duplicate.duplicate_count == 1
    assert len(repo.list_source_files()) == 1


def test_mesh_import_creates_compact_schema_without_raw_payload_columns(tmp_path):
    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-01")
    source = tmp_path / "meshlog.log"
    source.write_text("[1] 2025/12/03 10:12:33.579\n" + LINE_A + "\n", encoding="utf-8")

    MeshImportService("demo", paths).import_files(profile, [source])
    repo = MeshMrRepository(paths.mesh_mr_db_path("demo", profile.safe_folder_name))

    with sqlite3.connect(repo.path) as conn:
        schema_version = conn.execute("SELECT value FROM meta WHERE key = 'schema_version'").fetchone()[0]
        assert schema_version == "meshlog_compact_v2_single_log"
        parsed_db_path = Path(conn.execute("SELECT parsed_db_path FROM source_files").fetchone()[0])
        assert parsed_db_path.name.endswith(".mesh.sqlite")
        assert conn.execute("SELECT COUNT(*) FROM mesh_links").fetchone()[0] == 0
    with sqlite3.connect(parsed_db_path) as conn:
        schema_version = conn.execute("SELECT value FROM meta WHERE key = 'schema_version'").fetchone()[0]
        assert schema_version == "meshlog_compact_v2_single_log"
        table_names = {row[0] for row in conn.execute("SELECT name FROM sqlite_master WHERE type IN ('table', 'view')").fetchall()}
        assert {"active_points", "active_segments", "switch_events", "rssi_stats", "diagnosis_events"} <= table_names
        forbidden = {"raw_line", "raw_text", "raw_block", "raw_payload", "full_command_output", "debug_text", "metrics_json", "deltas_json", "raw_file"}
        for table in ("mesh_links", "parse_issues", "samples", "active_points"):
            columns = {row[1] for row in conn.execute(f"PRAGMA table_info({table})").fetchall()}
            assert not (columns & forbidden)
        assert conn.execute("SELECT COUNT(*) FROM active_points").fetchone()[0] == 1
        assert conn.execute("SELECT COUNT(*) FROM active_segments").fetchone()[0] == 1
        assert conn.execute("SELECT COUNT(*) FROM rssi_stats").fetchone()[0] >= 1


def test_mesh_repository_recovers_when_schema_meta_table_is_missing(tmp_path):
    db_path = tmp_path / "mesh.sqlite"
    conn = sqlite3.connect(db_path)
    conn.execute("CREATE TABLE mesh_links(raw_line TEXT)")
    conn.commit()
    conn.close()

    repo = MeshMrRepository(db_path)
    conn = sqlite3.connect(db_path)
    conn.execute("DROP TABLE schema_meta")
    conn.commit()
    conn.close()

    assert not repo.needs_derived_analysis_rebuild()
    with sqlite3.connect(db_path) as conn:
        version = conn.execute("SELECT value FROM meta WHERE key = 'schema_version'").fetchone()[0]
        columns = {row[1] for row in conn.execute("PRAGMA table_info(mesh_links)").fetchall()}
    assert version == "meshlog_compact_v2_single_log"
    assert "raw_line" not in columns
    assert list(tmp_path.glob("mesh.sqlite.legacy_*"))


def test_multiple_mrs_use_isolated_databases(tmp_path):
    paths = PathResolver(tmp_path)
    storage = MeshStorageService("demo", paths)
    mr1 = storage.create_mr_profile("14CW-01")
    mr2 = storage.create_mr_profile("14CW-02")
    source1 = tmp_path / "mr1.log"
    source2 = tmp_path / "mr2.log"
    source1.write_text("[1] 2025/12/03 10:12:33.579\n" + LINE_A + "\n", encoding="utf-8")
    source2.write_text("[1] 2025/12/03 10:12:33.579\n" + LINE_B + "\n", encoding="utf-8")
    service = MeshImportService("demo", paths)
    service.import_files(mr1, [source1])
    service.import_files(mr2, [source2])
    assert paths.mesh_mr_db_path("demo", mr1.safe_folder_name) != paths.mesh_mr_db_path("demo", mr2.safe_folder_name)
    total1, rows1 = MeshMrRepository(paths.mesh_mr_db_path("demo", mr1.safe_folder_name)).query_links(10, 0)
    total2, rows2 = MeshMrRepository(paths.mesh_mr_db_path("demo", mr2.safe_folder_name)).query_links(10, 0)
    assert total1 == total2 == 1
    assert rows1[0]["peer_mac_raw"] == "30f5-277a-5a2f"
    assert rows2[0]["peer_mac_raw"] == "30f5-277a-5a3f"


def test_mesh_links_keep_import_record_sequence_order(tmp_path):
    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-01")
    first = tmp_path / "first.log"
    second = tmp_path / "second.log"
    first.write_text("[1] 2025/12/03 10:00:03.000\n" + LINE_A + "\n", encoding="utf-8")
    second.write_text("[1] 2025/12/03 10:00:01.000\n" + LINE_B + "\n", encoding="utf-8")

    MeshImportService("demo", paths).import_files(profile, [first, second])
    repo = MeshMrRepository(paths.mesh_mr_db_path("demo", profile.safe_folder_name))
    _, links = repo.query_links(10, 0)

    assert [link["peer_mac_raw"] for link in links] == ["30f5-277a-5a2f", "30f5-277a-5a3f"]
    assert [link["record_seq"] for link in links] == [1, 2]
    assert [link["source_file_order"] for link in links] == [1, 2]


def test_mesh_query_sorts_record_seq_as_integer(tmp_path):
    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-01")
    source = tmp_path / "meshlog.log"
    lines = ["[1] 2025/12/03 10:00:00.000"]
    for suffix in ("1f", "2f", "3f", "4f", "5f"):
        lines.append(LINE_A.replace("30f5-277a-5a2f", f"30f5-277a-5a{suffix}"))
    source.write_text("\n".join(lines), encoding="utf-8")

    MeshImportService("demo", paths).import_files(profile, [source])
    repo = MeshMrRepository(paths.mesh_mr_db_path("demo", profile.safe_folder_name))
    detail_path = Path(repo.list_source_files()[0]["parsed_db_path"])
    with sqlite3.connect(detail_path) as conn:
        ids = [row[0] for row in conn.execute("SELECT id FROM mesh_links ORDER BY id ASC").fetchall()]
        for link_id, seq in zip(ids, [1000, 10, 1, 100, 2], strict=True):
            conn.execute("UPDATE mesh_links SET record_seq = ? WHERE id = ?", (seq, link_id))

    _, rows = repo.query_links(10, 0)

    assert [row["record_seq"] for row in rows] == [1, 2, 10, 100, 1000]


def test_mesh_link_table_sorts_record_seq_as_integer():
    from PySide6.QtCore import Qt
    from PySide6.QtWidgets import QTableWidget

    from netconsole.ui.pages.mesh_log_analysis_page import _set_row

    _app()
    table = QTableWidget(0, 1)
    table.setSortingEnabled(True)
    values = [1, 10, 100, 1000, 2]
    table.setRowCount(len(values))
    for row, value in enumerate(values):
        _set_row(table, row, [value])

    table.sortItems(0, Qt.AscendingOrder)

    assert [int(table.item(row, 0).text()) for row in range(table.rowCount())] == [1, 2, 10, 100, 1000]


def test_mesh_peer_mapping_keeps_ap_mac_and_radio_mac_separate(tmp_path):
    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-01")
    source = tmp_path / "meshlog.log"
    source.write_text("[1] 2025/12/03 10:12:33.000\n" + LINE_A + "\n", encoding="utf-8")
    MeshImportService("demo", paths).import_files(profile, [source])
    repo = MeshMrRepository(paths.mesh_mr_db_path("demo", profile.safe_folder_name))

    repo.upsert_peer_mappings(
        [
            {
                "peer_mac_normalized": "30f5277a5a2f",
                "peer_ap_name": "AP-01",
                "peer_ap_mac": "30f5277a5a3f",
                "peer_radio_mac": "30f5277a5a2f",
                "peer_radio_label": "radio2",
                "peer_site": "站点",
                "match_rule": "resolver",
                "match_confidence": 90,
            }
        ]
    )
    repo.refresh_peer_mapping_on_links()
    _, links = repo.query_links(10, 0)

    assert links[0]["peer_mac_raw"] == "30f5-277a-5a2f"
    assert links[0]["peer_ap_mac"] == "30f5277a5a3f"
    assert links[0]["peer_radio_mac"] == "30f5277a5a2f"


def test_mesh_peer_mapping_service_uses_h3c_radio_rule_fields(tmp_path):
    paths = PathResolver(tmp_path)
    service = MeshPeerMappingService("demo", paths)
    service._resolver = TracksideApBssidResolver([{"ap_name": "AP-01", "ap_mac": "083b-e9ec-da40", "site_name": "S1"}])

    resolved = service.resolve("083b-e9ec-da5f")

    assert resolved is not None
    assert resolved["peer_mac_normalized"] == "083be9ecda5f"
    assert resolved["peer_radio_mac"] == "083be9ecda5f"
    assert resolved["peer_ap_mac"] == "083be9ecda40"
    assert resolved["peer_ap_name"] == "AP-01"
    assert resolved["peer_site"] == "S1"
    assert resolved["peer_radio_label"] == "radio2"


def test_mesh_peer_mapping_service_returns_serial_number_for_h3c_peer_mac(tmp_path):
    paths = PathResolver(tmp_path)
    service = MeshPeerMappingService("demo", paths)
    service._resolver = TracksideApBssidResolver(
        [{"ap_name": "bc5a-3457-cbe0", "ap_mac": "bc5a-3457-cbe0", "station": "03镇驼站", "serial_number": "TEST-SN-001"}]
    )

    resolved = service.resolve("bc5a-3457-cbef")

    assert resolved is not None
    assert resolved["peer_ap_name"] == "bc5a-3457-cbe0"
    assert resolved["peer_ap_mac"] == "bc5a3457cbe0"
    assert resolved["peer_radio_mac"] == "bc5a3457cbef"
    assert resolved["peer_site"] == "03镇驼站"
    assert resolved["peer_serial_number"] == "TEST-SN-001"
    assert resolved["serial_number"] == "TEST-SN-001"
    assert resolved["peer_radio_label"] == "radio1"


def test_multiple_source_files_can_filter_links_and_charts(tmp_path):
    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-01")
    first = tmp_path / "first-meshlog.log"
    second = tmp_path / "second-meshlog.log"
    first.write_text("[1] 2025/12/03 10:00:00.000\n" + LINE_A + "\n", encoding="utf-8")
    second.write_text("[1] 2025/12/03 10:00:01.000\n" + LINE_A + "\n", encoding="utf-8")

    MeshImportService("demo", paths).import_files(profile, [first, second])
    repo = MeshMrRepository(paths.mesh_mr_db_path("demo", profile.safe_folder_name))
    sources = repo.list_source_files()
    assert len(sources) == 2

    first_id = int(sources[0]["id"])
    second_id = int(sources[1]["id"])
    total_all, all_links = repo.query_links(100, 0)
    total_first, first_links = repo.query_links(100, 0, {"source_file_id": first_id})
    total_second, second_links = repo.query_links(100, 0, {"source_file_id": second_id})

    assert total_all == 2
    assert total_first == len(first_links) == 1
    assert total_second == len(second_links) == 1
    assert int(first_links[0]["source_file_id"]) == first_id
    assert int(second_links[0]["source_file_id"]) == second_id

    chart_first = repo.query_peer_chart_segments(int(first_links[0]["id"]), source_file_id=first_id)
    chart_all = repo.query_peer_chart_segments(int(first_links[0]["id"]))
    assert [row["sample_time"] for row in chart_first["run_segment"]["rows"]] == ["2025-12-03 10:00:00.000"]
    assert chart_all["run_segment"]["rows"] == []
    assert "source_file_id" in chart_all["message"]

    first_db = Path(sources[0]["parsed_db_path"])
    second_db = Path(sources[1]["parsed_db_path"])
    conn = sqlite3.connect(first_db)
    try:
        conn.execute(
            """
            INSERT INTO switch_events (
                event_type, event_time, radio, previous_sample_time, current_sample_time,
                observed_window_ms, from_peer_mac, to_peer_mac, details_json, source_file_id, source_line_number
            ) VALUES
                ('ACTIVE_SWITCH', '2025-12-03 10:00:00.000', 1, NULL, '2025-12-03 10:00:00.000', NULL, NULL, NULL, '{}', 1, 2)
            """,
        )
        conn.commit()
    finally:
        conn.close()
    conn = sqlite3.connect(second_db)
    try:
        conn.execute(
            """
            INSERT INTO parse_issues (source_file_id, source_file, line_number, severity, issue_type, field_name, message, raw_line_start, raw_line_end)
            VALUES (1, 'second.log', 9, 'ERROR', 'x', 'x', 'second', 9, 9)
            """,
        )
        conn.commit()
    finally:
        conn.close()

    event_total, events = repo.query_events(100, 0, first_id)
    issue_total, issues = repo.query_issues(100, 0, second_id)
    assert event_total == len(events) == 1
    assert int(events[0]["source_file_id"]) == first_id
    assert issue_total == len(issues) == 1
    assert int(issues[0]["source_file_id"]) == second_id

    result = repo.delete_parsed_data_by_source_file(first_id)
    assert result.ok
    assert result.deleted_links == 1
    assert repo.query_links(100, 0, {"source_file_id": first_id})[0] == 0
    assert repo.query_links(100, 0, {"source_file_id": second_id})[0] == 1
    assert repo.query_events(100, 0, first_id)[0] == 0
    assert repo.query_issues(100, 0, first_id)[0] == 0
    first_status = next(source for source in repo.list_source_files() if int(source["id"]) == first_id)
    assert first_status["file_status"] == "parsed_deleted"

    second_archive = Path(str(next(source for source in repo.list_source_files() if int(source["id"]) == second_id)["archived_path"]))
    second_archive.unlink()
    repo.mark_source_file_deleted(second_id)
    assert repo.query_links(100, 0, {"source_file_id": second_id})[0] == 1
    second_status = next(source for source in repo.list_source_files() if int(source["id"]) == second_id)
    assert second_status["file_status"] == "deleted"


def test_list_source_files_uses_cached_file_status_without_path_exists(tmp_path, monkeypatch):
    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-01")
    source = tmp_path / "meshlog.log"
    source.write_text("[1] 2025/12/03 10:12:33.579\n" + LINE_A + "\n", encoding="utf-8")
    MeshImportService("demo", paths).import_files(profile, [source])
    repo = MeshMrRepository(paths.mesh_mr_db_path("demo", profile.safe_folder_name))

    monkeypatch.setattr(Path, "exists", lambda _path: (_ for _ in ()).throw(AssertionError("Path.exists should not be called")))

    rows = repo.list_source_files()

    assert len(rows) == 1
    assert rows[0]["file_status"] == "ok"
    assert rows[0]["file_exists"] == 1


def test_counter_reset_generates_event_without_negative_delta(tmp_path):
    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-01")
    line_high = LINE_A.replace("2/297", "100/200")
    line_low = LINE_A.replace("2/297", "1/2")
    source = tmp_path / "meshlog.log"
    source.write_text(
        "\n".join(["[1] 2025/12/03 10:12:33.000", line_high, "[1] 2025/12/03 10:12:34.000", line_low]),
        encoding="utf-8",
    )
    MeshImportService("demo", paths).import_files(profile, [source])
    repo = MeshMrRepository(paths.mesh_mr_db_path("demo", profile.safe_folder_name))
    _, events = repo.query_events(100, 0)
    _, links = repo.query_links(100, 0)
    assert any(event["event_type"] == "COUNTER_RESET" for event in events)
    assert all("-" not in str(link["deltas_json"]) for link in links)


def test_same_peer_different_establish_time_creates_sessions(tmp_path):
    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-01")
    second_session = LINE_A.replace("2025/12/03 10:12:30", "2025/12/03 10:13:00")
    source = tmp_path / "meshlog.log"
    source.write_text(
        "\n".join(["[1] 2025/12/03 10:12:33.000", LINE_A, "[1] 2025/12/03 10:13:33.000", second_session]),
        encoding="utf-8",
    )
    MeshImportService("demo", paths).import_files(profile, [source])
    sessions = MeshMrRepository(paths.mesh_mr_db_path("demo", profile.safe_folder_name)).export_rows("mesh_sessions")
    assert len(sessions) == 2


def test_import_hash_once_and_duplicate_does_not_parse(tmp_path, monkeypatch):
    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-01")
    source = tmp_path / "meshlog.log"
    source.write_text("[1] 2025/12/03 10:12:33.579\n" + LINE_A + "\n", encoding="utf-8")
    hash_calls = 0
    original_hash = mesh_log_parser.sha256_file
    original_parse = MeshLogParser.parse_file

    def counting_hash(path):
        nonlocal hash_calls
        hash_calls += 1
        return original_hash(path)

    parse_calls = 0

    def counting_parse(self, *args, **kwargs):
        nonlocal parse_calls
        parse_calls += 1
        return original_parse(self, *args, **kwargs)

    monkeypatch.setattr("netconsole.services.mesh_import_service.sha256_file", counting_hash)
    monkeypatch.setattr(MeshLogParser, "parse_file", counting_parse)
    service = MeshImportService("demo", paths)
    service.import_files(profile, [source])
    service.import_files(profile, [source])
    assert hash_calls == 2
    assert parse_calls == 1


def test_query_links_returns_stable_sample_group_index(tmp_path):
    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-01")
    source = tmp_path / "meshlog.log"
    source.write_text(
        "\n".join(
            [
                "[1] 2025/12/03 10:12:33.000",
                LINE_A,
                LINE_STANDBY,
                "[1] 2025/12/03 10:12:34.000",
                LINE_B,
                LINE_STANDBY.replace("30f5-277a-5a4f", "30f5-277a-5a5f"),
            ]
        ),
        encoding="utf-8",
    )
    MeshImportService("demo", paths).import_files(profile, [source])
    _, rows = MeshMrRepository(paths.mesh_mr_db_path("demo", profile.safe_folder_name)).query_links(10, 0)
    groups_by_time = {}
    for row in rows:
        groups_by_time.setdefault(row["sample_time"], set()).add(row["sample_group_index"])
    assert all(len(groups) == 1 for groups in groups_by_time.values())
    assert len({next(iter(groups)) for groups in groups_by_time.values()}) == 2


def test_mesh_import_resolves_peer_ap_name_site_and_radio(tmp_path):
    paths = PathResolver(tmp_path)
    site_db = Database(paths.site_db_path("demo"))
    site_db.initialize()
    with site_db.connect() as conn:
        now = datetime.now().isoformat(timespec="seconds")
        conn.execute(
            """
            INSERT INTO ac_fit_ap_resources (
                ac_device_uuid, ap_uuid, ap_name, ap_mac, site, collected_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            ("ac-1", "ap-1", "AP-01", "30f5-277a-5a10", "Ningbo Station", now, now),
        )
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-01")
    source = tmp_path / "meshlog.log"
    source.write_text("[1] 2025/12/03 10:12:33.000\n" + LINE_A + "\n", encoding="utf-8")
    MeshImportService("demo", paths).import_files(profile, [source])
    repo = MeshMrRepository(paths.mesh_mr_db_path("demo", profile.safe_folder_name))
    _, rows = repo.query_links(10, 0, {"peer": "AP-01"})
    assert len(rows) == 1
    assert rows[0]["peer_ap_name"] == "AP-01"
    assert rows[0]["peer_site"] == "Ningbo Station"
    assert rows[0]["peer_radio"] == "radio2"
    assert rows[0]["peer_radio_label"] == "radio2"
    assert rows[0]["peer_radio_mac"] == "30f5277a5a2f"
    assert rows[0]["peer_ap_mac"] == "30f5277a5a10"
    assert rows[0]["peer_resolve_source"] == "h3c_radio_2_ap_mac_nibble_plus_1"
    cache_rows = repo.export_rows("mesh_peer_resolve_cache")
    assert cache_rows[0]["peer_mac"] == "30f5277a5a2f"
    assert cache_rows[0]["peer_ap_name"] == "AP-01"


def test_mesh_repository_builds_downsampled_link_aggregates(tmp_path):
    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-01")
    source = tmp_path / "meshlog.log"
    source.write_text(
        "\n".join(
            [
                "[1] 2025/12/03 10:12:33.000",
                LINE_A,
                "[1] 2025/12/03 10:12:34.000",
                LINE_A.replace("36/43", "38/45"),
            ]
        ),
        encoding="utf-8",
    )
    MeshImportService("demo", paths).import_files(profile, [source])
    repo = MeshMrRepository(paths.mesh_mr_db_path("demo", profile.safe_folder_name))
    rows = repo.query_link_aggregates(bucket_seconds=10)
    assert rows == []
    detail_path = Path(repo.list_source_files()[0]["parsed_db_path"])
    with sqlite3.connect(detail_path) as conn:
        assert conn.execute("SELECT COUNT(*) FROM sqlite_master WHERE name = 'mesh_link_aggregates'").fetchone()[0] == 0


def _app():
    os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
    from PySide6.QtWidgets import QApplication

    return QApplication.instance() or QApplication([])


def _drain_qt_events(iterations: int = 10) -> None:
    app = _app()
    for _ in range(iterations):
        app.processEvents()
        time.sleep(0.01)


def _wait_for_mesh_tab_load(page, timeout: float = 5.0) -> None:
    app = _app()
    deadline = time.time() + timeout
    while time.time() < deadline:
        app.processEvents()
        current_tab = page._current_tab_name()
        overlay = getattr(page, "tab_overlays", {}).get(current_tab)
        if getattr(page, "tab_load_worker", None) is None and (overlay is None or not overlay.isVisible()):
            return
        time.sleep(0.01)
    raise AssertionError("Timed out waiting for mesh tab load")


def test_mesh_page_column_width_persists_and_active_style(tmp_path):
    _app()
    from PySide6.QtCore import Qt

    from netconsole.core.i18n import I18n
    from netconsole.ui.pages.mesh_log_analysis_page import MeshLogAnalysisPage

    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-01")
    source = tmp_path / "meshlog.log"
    source.write_text("[1] 2025/12/03 10:12:33.000\n" + LINE_A + "\n" + LINE_STANDBY + "\n", encoding="utf-8")
    MeshImportService("demo", paths).import_files(profile, [source])
    page = MeshLogAnalysisPage(I18n("en_US"), "demo", paths)
    page.link_table.setColumnWidth(3, 260)
    page.refresh_all()
    page.tabs.setCurrentIndex(1)
    page.refresh_current_mr_data()
    _wait_for_mesh_tab_load(page)
    assert page.link_table.columnWidth(3) == 260
    assert page.link_table.rowCount() == 2
    data = page.link_table.item(0, 0).data(Qt.UserRole)
    assert data["link_state"] == "ACTIVE"
    assert page.link_table.item(0, 0).font().bold()


def test_mesh_link_table_auto_width_keeps_metric_headers_visible(tmp_path):
    _app()
    from PySide6.QtCore import Qt
    from PySide6.QtWidgets import QHeaderView

    from netconsole.core.i18n import I18n
    from netconsole.ui.pages.mesh_log_analysis_page import MeshLogAnalysisPage

    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-01")
    source = tmp_path / "meshlog.log"
    source.write_text("[1] 2025/12/03 10:12:33.000\n" + LINE_A + "\n" + LINE_STANDBY + "\n", encoding="utf-8")
    MeshImportService("demo", paths).import_files(profile, [source])
    page = MeshLogAnalysisPage(I18n("zh_CN"), "demo", paths)
    page.refresh_all()
    page.tabs.setCurrentIndex(1)
    page.refresh_current_mr_data()
    _wait_for_mesh_tab_load(page)

    header = page.link_table.horizontalHeader()
    assert header.sectionResizeMode(13) == QHeaderView.Interactive
    assert page.link_table.horizontalScrollBarPolicy() == Qt.ScrollBarAsNeeded
    assert page.link_table.wordWrap() is False
    assert page.link_table.textElideMode() == Qt.TextElideMode.ElideRight
    for column in range(13, 25):
        header_text = page.link_table.horizontalHeaderItem(column).text()
        expected = header.fontMetrics().horizontalAdvance(header_text) + 32
        assert page.link_table.columnWidth(column) >= expected


def test_active_build_order_headers_are_chinese_and_autosized(tmp_path):
    _app()
    from PySide6.QtCore import Qt
    from PySide6.QtWidgets import QHeaderView

    from netconsole.core.i18n import I18n
    from netconsole.ui.pages.mesh_log_analysis_page import MeshLogAnalysisPage

    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-01")
    source = tmp_path / "meshlog.log"
    source.write_text("[1] 2025/12/03 10:12:33.000\n" + LINE_A + "\n" + LINE_STANDBY + "\n", encoding="utf-8")
    MeshImportService("demo", paths).import_files(profile, [source])
    page = MeshLogAnalysisPage(I18n("zh_CN"), "demo", paths)
    page.current_profile = profile
    repo = MeshMrRepository(paths.mesh_mr_db_path("demo", profile.safe_folder_name))
    page._render_active_build_order(repo)

    headers = [page.active_build_order_table.horizontalHeaderItem(column).text() for column in range(page.active_build_order_table.columnCount())]
    assert "mesh_analysis.min_rssi" not in headers
    assert headers[2] == "主链路 PeerMac"
    assert headers[12] == "MR侧最低RSSI"
    assert headers[14] == "发送繁忙度"
    assert headers[16] == "配置切换时间(ms)"
    assert headers[18] == "是否同AP射频切换"
    assert headers[20] == "判定原因"
    assert headers[21] == "是否AP回切"
    assert headers[23] == "乒乓类型"
    assert headers[31] == "乒乓判定原因"
    header = page.active_build_order_table.horizontalHeader()
    assert header.sectionResizeMode(6) == QHeaderView.Interactive
    assert page.active_build_order_table.horizontalScrollBarPolicy() == Qt.ScrollBarAsNeeded
    for column in (6, 7, 31, 32):
        header_text = page.active_build_order_table.horizontalHeaderItem(column).text()
        expected = header.fontMetrics().horizontalAdvance(header_text) + 32
        assert page.active_build_order_table.columnWidth(column) >= expected


def test_active_build_order_uses_source_snapshot_before_site_fallback_and_temp_override(tmp_path):
    repo = MeshMrRepository(tmp_path / "sample.mesh.sqlite")
    _insert_mesh_samples(repo.path, first_count=2, second_count=0)
    snapshot = mesh_analysis_params_to_json(
        {
            "main_link_switch_time_ms": 2500,
            "short_link_tolerance_ms": 500,
            "merge_same_physical_ap_dual_radio": True,
        }
    )
    with sqlite3.connect(repo.path) as conn:
        conn.execute("UPDATE source_files SET analysis_params_json = ?", (snapshot,))

    rows = repo.query_active_link_build_order(
        fallback_analysis_params={"main_link_switch_time_ms": 5000, "short_link_tolerance_ms": 500}
    )
    assert rows[0]["main_link_switch_time_ms"] == 2500
    assert rows[0]["short_link_tolerance_ms"] == 500
    assert rows[0]["build_result"] == "normal"
    assert "容差范围" in rows[0]["judge_reason"]

    override_rows = repo.query_active_link_build_order(
        analysis_params={"main_link_switch_time_ms": 3000, "short_link_tolerance_ms": 500}
    )
    assert override_rows[0]["main_link_switch_time_ms"] == 3000
    assert override_rows[0]["build_result"] == "short"


def test_active_build_order_marks_ap_pingpong_by_physical_ap_sequence():
    from netconsole.repositories.mesh_mr_repository import _active_build_order_rows_from_points

    rows = [
        _active_point("2025-12-03 10:00:00.000", "30f5277a5a2f", "aaaa-0000-0001", "AP-A", "radio1"),
        _active_point("2025-12-03 10:00:01.000", "30f5277a5a3f", "bbbb-0000-0001", "AP-B", "radio1"),
        _active_point("2025-12-03 10:00:02.000", "30f5277a5a2f", "aaaa-0000-0001", "AP-A", "radio1"),
    ]
    result = _active_build_order_rows_from_points(rows, {"main_link_switch_time_ms": 2500, "pingpong_tolerance_ms": 500})

    middle = result[1]
    assert middle["is_ap_return_event"] is True
    assert middle["is_pingpong_abnormal"] is True
    assert middle["pingpong_type"] == "AP乒乓切换异常"
    assert middle["middle_ap_dwell_ms"] == 1000
    assert "明显小于配置切换时间 2500ms" in middle["pingpong_judgment_reason"]


def test_active_build_order_rssi_stats_use_same_valid_samples():
    from netconsole.repositories.mesh_mr_repository import _active_build_order_rows_from_points

    rows = [
        {**_active_point("2025-12-03 10:00:00.000", "30f5277a5a2f", "aaaa-0000-0001", "AP-A", "radio1"), "local_rssi_db": 0},
        {**_active_point("2025-12-03 10:00:01.000", "30f5277a5a2f", "aaaa-0000-0001", "AP-A", "radio1"), "local_rssi_db": 40},
        {**_active_point("2025-12-03 10:00:02.000", "30f5277a5a2f", "aaaa-0000-0001", "AP-A", "radio1"), "local_rssi_db": 50},
    ]
    result = _active_build_order_rows_from_points(rows, {"main_link_switch_time_ms": 2500, "pingpong_tolerance_ms": 500})

    segment = result[0]
    assert segment["avg_mr_rssi"] == 30
    assert segment["min_mr_rssi"] == 0
    assert segment["max_mr_rssi"] == 50
    assert segment["min_mr_rssi"] <= segment["avg_mr_rssi"] <= segment["max_mr_rssi"]


def test_active_build_order_separates_critical_and_normal_return_events():
    from netconsole.repositories.mesh_mr_repository import _active_build_order_rows_from_points

    critical_rows = [
        _active_point("2025-12-03 10:00:00.000", "30f5277a5a2f", "aaaa-0000-0001", "AP-A", "radio1"),
        _active_point("2025-12-03 10:00:01.000", "30f5277a5a3f", "bbbb-0000-0001", "AP-B", "radio1"),
        _active_point("2025-12-03 10:00:02.000", "30f5277a5a3f", "bbbb-0000-0001", "AP-B", "radio1"),
        _active_point("2025-12-03 10:00:03.000", "30f5277a5a2f", "aaaa-0000-0001", "AP-A", "radio1"),
    ]
    normal_rows = [
        _active_point("2025-12-03 10:10:00.000", "30f5277a5a2f", "aaaa-0000-0001", "AP-A", "radio1"),
        _active_point("2025-12-03 10:10:01.000", "30f5277a5a3f", "bbbb-0000-0001", "AP-B", "radio1"),
        _active_point("2025-12-03 10:10:02.000", "30f5277a5a3f", "bbbb-0000-0001", "AP-B", "radio1"),
        _active_point("2025-12-03 10:10:03.000", "30f5277a5a3f", "bbbb-0000-0001", "AP-B", "radio1"),
        _active_point("2025-12-03 10:10:04.000", "30f5277a5a3f", "bbbb-0000-0001", "AP-B", "radio1"),
        _active_point("2025-12-03 10:10:05.000", "30f5277a5a2f", "aaaa-0000-0001", "AP-A", "radio1"),
    ]

    critical = _active_build_order_rows_from_points(critical_rows, {"main_link_switch_time_ms": 2500, "pingpong_tolerance_ms": 500})[1]
    normal = _active_build_order_rows_from_points(normal_rows, {"main_link_switch_time_ms": 2500, "pingpong_tolerance_ms": 500})[1]

    assert critical["pingpong_type"] == "临界回切"
    assert critical["is_pingpong_abnormal"] is False
    assert normal["pingpong_type"] == "普通回切事件"
    assert normal["is_pingpong_abnormal"] is False


def test_active_build_order_marks_same_physical_ap_radio_roundtrip_not_ap_pingpong():
    from netconsole.repositories.mesh_mr_repository import _active_build_order_rows_from_points

    rows = [
        _active_point("2025-12-03 10:00:00.000", "30f5277a5a2f", "aaaa-0000-0001", "AP-A", "radio1"),
        _active_point("2025-12-03 10:00:01.000", "30f5277a5a3f", "aaaa-0000-0001", "AP-A", "radio2"),
        _active_point("2025-12-03 10:00:02.000", "30f5277a5a2f", "aaaa-0000-0001", "AP-A", "radio1"),
    ]
    result = _active_build_order_rows_from_points(rows, {"main_link_switch_time_ms": 2500, "pingpong_tolerance_ms": 500})

    middle = result[1]
    assert middle["pingpong_type"] == "同AP射频往返"
    assert middle["is_ap_return_event"] is False
    assert middle["is_pingpong_abnormal"] is False


def test_peer_chart_run_context_includes_same_time_standby_from_other_session(tmp_path):
    from netconsole.ui.mesh_chart_payload import build_chart_payload

    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-01")
    source = tmp_path / "meshlog.log"
    source.write_text("[1] 2025/12/03 10:12:33.000\n" + LINE_A + "\n" + LINE_STANDBY + "\n", encoding="utf-8")
    MeshImportService("demo", paths).import_files(profile, [source])
    repo = MeshMrRepository(paths.mesh_mr_db_path("demo", profile.safe_folder_name))
    total, rows = repo.query_links(10, 0, {"state": "ACTIVE"})
    assert total == 1
    anchor = rows[0]

    segments = repo.query_peer_chart_initial_segments(int(anchor["id"]), source_file_id=anchor["source_file_id"])
    run_rows = segments["run_segment"]["rows"]
    assert {row["link_state"] for row in run_rows} == {"ACTIVE", "STANDBY"}

    payload = build_chart_payload(segments["peer_segment"], segments["run_segment"])
    backups = [items for items in payload["backup_links_by_index"] if items]
    assert backups
    assert backups[0][0]["peer_mac"] == "30f5277a5a4f"
    assert payload["main_links_by_index"][0]["peer_mac"] == "30f5277a5a2f"


def test_mesh_page_double_click_source_opens_filtered_link_details(tmp_path):
    app = _app()
    from PySide6.QtCore import Qt

    from netconsole.core.i18n import I18n
    from netconsole.ui.pages.mesh_log_analysis_page import MeshLogAnalysisPage

    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-01")
    first = tmp_path / "first-meshlog.log"
    second = tmp_path / "second-meshlog.log"
    first.write_text("[1] 2025/12/03 10:12:33.000\n" + LINE_A + "\n" + LINE_STANDBY + "\n", encoding="utf-8")
    second.write_text("[1] 2025/12/03 10:12:34.000\n" + LINE_B + "\n", encoding="utf-8")
    MeshImportService("demo", paths).import_files(profile, [first, second])
    repo = MeshMrRepository(paths.mesh_mr_db_path("demo", profile.safe_folder_name))
    sources = repo.list_source_files()
    target_source_id = int(next(row["id"] for row in sources if row["original_filename"] == first.name))
    page = MeshLogAnalysisPage(I18n("en_US"), "demo", paths)
    page.current_profile = profile
    page._render_sources(repo)
    row = _source_table_row(page, target_source_id)

    page.source_table.cellDoubleClicked.emit(row, 0)
    app.processEvents()
    app.processEvents()
    _wait_for_mesh_tab_load(page)

    assert page.tabs.currentIndex() == 1
    assert page.current_source_file_id == target_source_id
    assert page.link_table.rowCount() == 2
    record_seq = [int(page.link_table.item(index, 0).text()) for index in range(page.link_table.rowCount())]
    assert record_seq == sorted(record_seq)
    source_ids = {int(page.link_table.item(index, 0).data(Qt.UserRole)["source_file_id"]) for index in range(page.link_table.rowCount())}
    assert source_ids == {target_source_id}

    for _ in range(10):
        page.source_table.cellDoubleClicked.emit(row, 0)
        app.processEvents()
    assert page.tabs.currentIndex() == 1
    assert page.current_source_file_id == target_source_id


def test_mesh_page_peer_dialog_uses_row_source_file_id_in_all_files_mode(tmp_path, monkeypatch):
    _app()
    from PySide6.QtCore import Qt

    from netconsole.core.i18n import I18n
    from netconsole.ui.pages.mesh_log_analysis_page import MeshLogAnalysisPage

    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-01")
    first = tmp_path / "first-meshlog.log"
    second = tmp_path / "second-meshlog.log"
    first.write_text("[1] 2025/12/03 10:00:00.000\n" + LINE_A + "\n", encoding="utf-8")
    second.write_text("[1] 2025/12/03 10:00:01.000\n" + LINE_B + "\n", encoding="utf-8")
    MeshImportService("demo", paths).import_files(profile, [first, second])

    page = MeshLogAnalysisPage(I18n("zh_CN"), "demo", paths)
    page.current_profile = profile
    page.current_source_file_id = None
    page.tabs.setCurrentIndex(1)
    page.refresh_link_table()
    _wait_for_mesh_tab_load(page)
    captured: list[tuple[str, int | None, str, int | None, int | None]] = []
    monkeypatch.setattr(page, "_open_peer_dialog", lambda peer, radio, session, link_id=None, source_file_id=None: captured.append((peer, radio, session, link_id, source_file_id)))

    row_data = page.link_table.item(0, 0).data(Qt.UserRole)
    for column in (4, 5, 6, 8, 9):
        page._open_peer_from_link_cell(0, column)

    assert len(captured) == 5
    assert {item[4] for item in captured} == {int(row_data["source_file_id"])}
    assert {item[3] for item in captured} == {int(row_data["id"])}


def test_mesh_page_double_click_source_without_links_shows_empty_detail(tmp_path):
    app = _app()
    from netconsole.core.i18n import I18n
    from netconsole.ui.pages.mesh_log_analysis_page import MeshLogAnalysisPage

    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-01")
    repo = MeshMrRepository(paths.mesh_mr_db_path("demo", profile.safe_folder_name))
    empty = tmp_path / "empty-meshlog.log"
    empty.write_text("", encoding="utf-8")
    source_id = repo.insert_file_result(
        profile.mr_id,
        empty,
        empty,
        "empty-sha",
        0,
        None,
        "test",
        "ok",
        None,
        None,
        0,
        0,
        0,
        0,
        0,
        "",
        [],
        [],
        [],
    )
    page = MeshLogAnalysisPage(I18n("en_US"), "demo", paths)
    page.current_profile = profile
    page._render_sources(repo)

    page.source_table.cellDoubleClicked.emit(_source_table_row(page, source_id), 0)
    app.processEvents()
    app.processEvents()

    assert page.tabs.currentIndex() == 1
    assert page.current_source_file_id == source_id
    assert page.link_table.rowCount() == 0


def _source_table_row(page, source_file_id: int) -> int:
    from PySide6.QtCore import Qt

    for row in range(page.source_table.rowCount()):
        item = page.source_table.item(row, 0)
        data = item.data(Qt.UserRole) if item is not None else None
        if isinstance(data, dict) and int(data.get("id") or 0) == int(source_file_id):
            return row
    raise AssertionError(f"source row not found: {source_file_id}")


def test_peer_dialog_draws_signal_lines(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.ui.dialogs.mesh_peer_detail_dialog import MeshPeerDetailDialog

    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-01")
    second_session = LINE_A.replace("2025/12/03 10:12:30", "2025/12/03 10:13:00")
    source = tmp_path / "meshlog.log"
    source.write_text(
        "\n".join(["[1] 2025/12/03 10:12:33.000", LINE_A, "[1] 2025/12/03 10:13:33.000", second_session]),
        encoding="utf-8",
    )
    MeshImportService("demo", paths).import_files(profile, [source])
    rows = MeshMrRepository(paths.mesh_mr_db_path("demo", profile.safe_folder_name)).query_peer_series("30f5277a5a2f")
    assert rows == sorted(rows, key=lambda row: row["sample_time"])
    dialog = MeshPeerDetailDialog(I18n("en_US"), profile, paths.mesh_mr_db_path("demo", profile.safe_folder_name), "30f5277a5a2f", auto_load=False)
    dialog._on_loaded([{"metrics": json.loads(str(row["metrics_json"])), "deltas": json.loads(str(row["deltas_json"])), **row} for row in rows])
    signal_axis = dialog.figures["signal"].axes[0]
    assert len(signal_axis.lines) >= 2
    assert dialog.canvases["signal"].figure is dialog.figures["signal"]


def test_pagination_widget_can_jump_to_numeric_page():
    _app()
    from PySide6.QtTest import QSignalSpy

    from netconsole.core.i18n import I18n
    from netconsole.ui.pagination import PaginationState
    from netconsole.ui.widgets.pagination_widget import PaginationWidget

    widget = PaginationWidget(I18n("en_US"))
    spy = QSignalSpy(widget.pageChanged)
    widget.set_state(PaginationState(page_size=200, current_page=1, total_items=1000, total_pages=5))
    widget.page_jump_spin.setValue(4)
    widget.page_jump_button.click()
    assert spy.count() == 1
    assert spy.at(0)[0] == 4


def test_peer_context_segment_splits_same_peer_after_gap(tmp_path):
    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-01")
    source = tmp_path / "meshlog.log"
    source.write_text(
        "\n".join(
            [
                "[1] 2025/12/03 10:00:00.000",
                LINE_A,
                "[1] 2025/12/03 10:00:01.000",
                LINE_A.replace("60/72060", "61/72061"),
                "[1] 2025/12/03 10:30:00.000",
                LINE_A.replace("60/72060", "62/72062"),
            ]
        ),
        encoding="utf-8",
    )
    MeshImportService("demo", paths).import_files(profile, [source])
    repo = MeshMrRepository(paths.mesh_mr_db_path("demo", profile.safe_folder_name))
    _, rows = repo.query_links(100, 0)
    first_anchor = next(row for row in rows if str(row["sample_time"]).endswith("10:00:00.000"))
    late_anchor = next(row for row in rows if str(row["sample_time"]).endswith("10:30:00.000"))
    first_segment = repo.query_peer_context_segment(int(first_anchor["id"]))
    late_segment = repo.query_peer_context_segment(int(late_anchor["id"]))
    assert [row["sample_time"] for row in first_segment["rows"]] == ["2025-12-03 10:00:00.000", "2025-12-03 10:00:01.000"]
    assert [row["sample_time"] for row in late_segment["rows"]] == ["2025-12-03 10:30:00.000"]


def test_anchor_dialog_signal_chart_uses_raw_positive_rssi(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.ui.dialogs.mesh_peer_detail_dialog import MeshPeerDetailDialog

    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-01")
    source = tmp_path / "meshlog.log"
    source.write_text("[1] 2025/12/03 10:12:33.000\n" + LINE_A + "\n", encoding="utf-8")
    MeshImportService("demo", paths).import_files(profile, [source])
    repo = MeshMrRepository(paths.mesh_mr_db_path("demo", profile.safe_folder_name))
    _, links = repo.query_links(10, 0)
    anchor_id = int(links[0]["id"])
    payload = {
        "peer_segment": repo.query_peer_context_segment(anchor_id),
        "run_segment": repo.query_run_context_segment(anchor_id),
        "active_timeline": repo.query_active_timeline(anchor_id),
    }
    for segment_key in ("peer_segment", "run_segment"):
        payload[segment_key]["rows"] = [{"metrics": json.loads(str(row["metrics_json"])), "deltas": json.loads(str(row["deltas_json"])), **row} for row in payload[segment_key]["rows"]]
    for row in payload["active_timeline"]["rows"]:
        for key in ("active", "next_active"):
            if isinstance(row.get(key), dict):
                row[key] = {"metrics": json.loads(str(row[key]["metrics_json"])), "deltas": json.loads(str(row[key]["deltas_json"])), **row[key]}
    dialog = MeshPeerDetailDialog(I18n("en_US"), profile, paths.mesh_mr_db_path("demo", profile.safe_folder_name), "30f5277a5a2f", radio=1, auto_load=False, anchor_link_id=anchor_id)
    dialog._on_loaded(payload)
    signal_axis = dialog.figures["signal"].axes[0]
    y_values = [value for line in signal_axis.lines for value in line.get_ydata()]
    assert 36 in y_values
    assert 43 in y_values
    assert -52 not in y_values
    assert "active_next_rssi" not in dialog.tab_keys
    assert "active_channel_load" not in dialog.tab_keys


def test_run_segment_query_uses_anchor_boundaries_not_second_run(tmp_path):
    repo = MeshMrRepository(tmp_path / "sample.mesh.sqlite")
    _insert_mesh_samples(repo.path, first_count=10000, second_count=10000)
    segment = repo.query_run_context_segment(5000)
    sample_times = {row["sample_time"] for row in segment["rows"]}
    assert len(sample_times) == 10000
    assert "2025-12-03 10:00:00.000" in sample_times
    assert "2025-12-03 13:16:40.000" not in sample_times


def test_full_active_chart_query_loads_beyond_link_detail_page_size(tmp_path):
    from netconsole.ui.mesh_chart_payload import build_chart_payload

    repo = MeshMrRepository(tmp_path / "sample.mesh.sqlite")
    _insert_mesh_samples(repo.path, first_count=1005, second_count=0)

    segment = repo.query_active_link_chart_segments()
    payload = build_chart_payload(segment["peer_segment"], segment["run_segment"])

    assert len(segment["run_segment"]["rows"]) == 1005
    assert segment["run_segment"]["query_active_count"] == 1005
    assert payload["metadata"]["sample_count"] == 1005
    assert payload["metadata"]["query_active_count"] == 1005
    assert payload["metadata"]["full_active_payload"] is True


def test_worker_uses_single_chart_segment_query(monkeypatch, tmp_path):
    from netconsole.ui import mesh_peer_series_worker
    from netconsole.ui.mesh_peer_series_worker import MeshPeerSeriesWorker

    calls = {"chart": 0, "run": 0}

    class FakeRepo:
        def __init__(self, path):
            self.path = path

        def query_peer_chart_segments(self, anchor_link_id):
            calls["chart"] += 1
            row = _chart_row(1, "2025-12-03 10:00:00.000")
            return {"peer_segment": {"anchor": row, "rows": [row]}, "run_segment": {"anchor": row, "rows": [row], "events": []}}

        def query_run_context_segment(self, anchor_link_id):
            calls["run"] += 1
            return {}

    monkeypatch.setattr(mesh_peer_series_worker, "MeshMrRepository", FakeRepo)
    worker = MeshPeerSeriesWorker(tmp_path / "mesh.sqlite", "30f5277a5a2f", 1, anchor_link_id=1, source_file_id=1)
    received = []
    worker.loaded.connect(lambda payload: received.append(("loaded", payload)))
    worker.loaded_full.connect(lambda payload: received.append(("loaded_full", payload)))
    worker.run()
    assert calls == {"chart": 1, "run": 0}
    assert [kind for kind, _payload in received] == ["loaded_full"]
    assert "chart_payload" in received[0][1]


def test_anchor_centering_for_middle_start_and_end(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.models.mesh_log_models import MeshMrProfile
    from netconsole.ui.dialogs.mesh_peer_detail_dialog import MeshPeerDetailDialog

    profile = MeshMrProfile("mr", "MR", "MR", datetime.now(), datetime.now())
    dialog = MeshPeerDetailDialog(I18n("en_US"), profile, tmp_path / "mesh.sqlite", "30f5277a5a2f", auto_load=False)
    dialog.visible_sample_count = 60
    dialog._on_loaded({"chart_payload": _chart_payload(300, 150), "peer_segment": {}})
    assert dialog.window_start_index == 120
    dialog._on_loaded({"chart_payload": _chart_payload(300, 10), "peer_segment": {}})
    assert dialog.window_start_index == 0
    dialog._on_loaded({"chart_payload": _chart_payload(300, 295), "peer_segment": {}})
    assert dialog.window_start_index == 240


def test_chart_draws_only_visible_window_and_reuses_figure(tmp_path, monkeypatch):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.models.mesh_log_models import MeshMrProfile
    from netconsole.ui.dialogs.mesh_peer_detail_dialog import MeshPeerDetailDialog

    profile = MeshMrProfile("mr", "MR", "MR", datetime.now(), datetime.now())
    dialog = MeshPeerDetailDialog(I18n("en_US"), profile, tmp_path / "mesh.sqlite", "30f5277a5a2f", auto_load=False)
    dialog._on_loaded({"chart_payload": _chart_payload(100000, 50000), "peer_segment": {}})
    signal_axis = dialog.figures["signal"].axes[0]
    line_lengths = [len(line.get_xdata()) for line in signal_axis.lines if len(line.get_xdata()) > 2]
    assert line_lengths and max(line_lengths) <= 124
    clear_calls = 0
    original_clear = dialog.figures["signal"].clear

    def counting_clear(*args, **kwargs):
        nonlocal clear_calls
        clear_calls += 1
        return original_clear(*args, **kwargs)

    monkeypatch.setattr(dialog.figures["signal"], "clear", counting_clear)
    dialog.time_scrollbar.setValue(dialog.time_scrollbar.value() + 10)
    dialog._render_current_tab()
    assert clear_calls == 0
    line_lengths = [len(line.get_xdata()) for line in signal_axis.lines if len(line.get_xdata()) > 2]
    assert max(line_lengths) <= 124


def test_all_view_downsamples_and_preserves_important_points(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.models.mesh_log_models import MeshMrProfile
    from netconsole.ui.dialogs.mesh_peer_detail_dialog import MeshPeerDetailDialog

    profile = MeshMrProfile("mr", "MR", "MR", datetime.now(), datetime.now())
    dialog = MeshPeerDetailDialog(I18n("en_US"), profile, tmp_path / "mesh.sqlite", "30f5277a5a2f", auto_load=False)
    payload = _chart_payload(100000, 50000)
    payload["peer_series"]["local_rssi"][12345] = 999
    payload["important_indices"] = np.asarray([12345, 50000], dtype=np.int32)
    dialog._on_loaded({"chart_payload": payload, "peer_segment": {}})
    dialog.visible_samples_combo.setCurrentIndex(4)
    dialog._visible_samples_changed()
    signal_axis = dialog.figures["signal"].axes[0]
    rendered_y = [value for line in signal_axis.lines for value in line.get_ydata()]
    max_points = dialog._max_render_points("signal")
    line_lengths = [len(line.get_xdata()) for line in signal_axis.lines if len(line.get_xdata()) > 2]
    assert max(line_lengths) <= max_points + len(payload["important_indices"]) + 2
    assert 999 in rendered_y


def test_center_selected_sample_button_restores_anchor_without_requery(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.models.mesh_log_models import MeshMrProfile
    from netconsole.ui.dialogs.mesh_peer_detail_dialog import MeshPeerDetailDialog

    profile = MeshMrProfile("mr", "MR", "MR", datetime.now(), datetime.now())
    dialog = MeshPeerDetailDialog(I18n("en_US"), profile, tmp_path / "mesh.sqlite", "30f5277a5a2f", auto_load=False)
    dialog.visible_sample_count = 60
    dialog._on_loaded({"chart_payload": _chart_payload(300, 150), "peer_segment": {}})
    dialog.time_scrollbar.setValue(20)
    assert dialog.user_moved_window
    dialog.center_selected_sample()
    assert dialog.window_start_index == 120
    assert not dialog.user_moved_window


def test_mesh_time_axis_uses_sample_time_not_scientific_offset(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.models.mesh_log_models import MeshMrProfile
    from netconsole.ui.dialogs.mesh_peer_detail_dialog import MeshPeerDetailDialog

    labels = ["2026-06-19 23:12:20.000", "2026-06-19 23:12:25.000", "2026-06-19 23:12:30.000"]
    profile = MeshMrProfile("mr", "MR", "MR", datetime.now(), datetime.now())
    dialog = MeshPeerDetailDialog(I18n("en_US"), profile, tmp_path / "mesh.sqlite", "30f5277a5a2f", auto_load=False)
    dialog._on_loaded({"chart_payload": _chart_payload_from_labels(labels, 1), "peer_segment": {}})
    axis = dialog.figures["signal"].axes[0]
    dialog.canvases["signal"].draw()
    tick_text = " ".join(label.get_text() for label in axis.get_xticklabels())
    assert "23:12:" in tick_text
    assert ".0002" not in tick_text
    assert "2.040" not in axis.xaxis.get_offset_text().get_text()
    assert not axis.xaxis.get_offset_text().get_visible()


def test_mesh_time_axis_preserves_millisecond_distances_and_hover_time(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.models.mesh_log_models import MeshMrProfile
    from netconsole.ui.dialogs.mesh_peer_detail_dialog import MeshPeerDetailDialog

    labels = ["2025-12-03 10:12:33.579", "2025-12-03 10:12:33.964", "2025-12-03 10:12:34.964"]
    payload = _chart_payload_from_labels(labels, 1)
    profile = MeshMrProfile("mr", "MR", "MR", datetime.now(), datetime.now())
    dialog = MeshPeerDetailDialog(I18n("en_US"), profile, tmp_path / "mesh.sqlite", "30f5277a5a2f", auto_load=False)
    dialog._on_loaded({"chart_payload": payload, "peer_segment": {}})
    numeric = payload["timestamp_numeric"]
    seconds_01 = (numeric[1] - numeric[0]) * 86400
    seconds_12 = (numeric[2] - numeric[1]) * 86400
    assert seconds_01 == pytest.approx(0.385, abs=0.001)
    assert seconds_12 == pytest.approx(1.0, abs=0.001)
    hover = dialog.hover_controllers["signal"]
    assert "2025-12-03 10:12:33.964" in hover.tooltip_text(1)


def test_mesh_time_axis_cross_day_shows_date(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.models.mesh_log_models import MeshMrProfile
    from netconsole.ui.dialogs.mesh_peer_detail_dialog import MeshPeerDetailDialog

    labels = ["2025-12-03 23:59:59.000", "2025-12-04 00:00:01.000"]
    profile = MeshMrProfile("mr", "MR", "MR", datetime.now(), datetime.now())
    dialog = MeshPeerDetailDialog(I18n("en_US"), profile, tmp_path / "mesh.sqlite", "30f5277a5a2f", auto_load=False)
    dialog._on_loaded({"chart_payload": _chart_payload_from_labels(labels, 0), "peer_segment": {}})
    axis = dialog.figures["signal"].axes[0]
    dialog.canvases["signal"].draw()
    tick_text = " ".join(label.get_text() for label in axis.get_xticklabels())
    assert "12-03" in tick_text or "12-04" in tick_text
    assert dialog.chart_payload["timestamp_labels"] == labels


def test_hover_snaps_to_nearest_master_sample_and_shows_signal_metrics(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.models.mesh_log_models import MeshMrProfile
    from netconsole.ui.dialogs.mesh_peer_detail_dialog import MeshPeerDetailDialog

    labels = [f"2025-12-03 10:00:{index:02d}.000" for index in range(20)]
    payload = _chart_payload_from_labels(labels, 10)
    payload["peer_series"]["local_rssi"][10] = 24
    payload["peer_series"]["peer_rssi"][10] = 34
    profile = MeshMrProfile("mr", "MR", "MR", datetime.now(), datetime.now())
    dialog = MeshPeerDetailDialog(I18n("en_US"), profile, tmp_path / "mesh.sqlite", "30f5277a5a2f", auto_load=False)
    dialog._on_loaded({"chart_payload": payload, "peer_segment": {}})
    hover = dialog.hover_controllers["signal"]
    between = (payload["timestamp_numeric"][10] + payload["timestamp_numeric"][11]) / 2 - 0.000001
    assert hover.nearest_index(between) == 10
    text = hover.tooltip_text(10)
    assert "2025-12-03 10:00:10.000" in text
    assert "30f5-277a-5a2f" in text
    assert "Active Link" in text
    assert "MR-AP_RSSI: 24/34" in text
    assert "MR RSSI: 24" not in text
    assert "Peer RSSI: 34" not in text
    assert "Raw RSSI margin" not in text


def test_current_active_rssi_hover_shows_only_current_peer(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.models.mesh_log_models import MeshMrProfile
    from netconsole.ui.dialogs.mesh_peer_detail_dialog import MeshActiveLinkChartDialog

    labels = ["2025-12-03 10:00:00.000", "2025-12-03 10:00:01.000"]
    payload = _chart_payload_from_labels(labels, 0)
    payload["active_peer_macs"][0] = "30f5277a5a2f"
    payload["active_series"]["active_local_rssi"][0] = 24
    payload["active_peer_rssi"] = np.asarray([34, np.nan], dtype=np.float32)
    profile = MeshMrProfile("mr", "MR", "MR", datetime.now(), datetime.now())
    dialog = MeshActiveLinkChartDialog(I18n("en_US"), profile, tmp_path / "mesh.sqlite", auto_load=False)
    dialog._on_loaded({"chart_payload": payload, "peer_segment": {}})
    dialog.tabs.setCurrentIndex(dialog.tab_keys.index("active_next_rssi"))
    text = dialog.hover_controllers["active_next_rssi"].tooltip_text(0)
    assert "30f5-277a-5a2f" in text
    assert "MR-AP_RSSI: 24/34" in text
    assert "30f5-277a-5a3f" not in text
    assert "Next Active" not in text
    assert "Next Active MR RSSI" not in text
    assert "Next Active Link" not in text
    assert "Current Active Peer RSSI" not in text
    assert "Next Active Peer RSSI" not in text


def test_hover_reuses_artists_on_repeated_motion(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.models.mesh_log_models import MeshMrProfile
    from netconsole.ui.dialogs.mesh_peer_detail_dialog import MeshPeerDetailDialog

    labels = [f"2025-12-03 10:00:{index:02d}.000" for index in range(20)]
    profile = MeshMrProfile("mr", "MR", "MR", datetime.now(), datetime.now())
    dialog = MeshPeerDetailDialog(I18n("en_US"), profile, tmp_path / "mesh.sqlite", "30f5277a5a2f", auto_load=False)
    dialog._on_loaded({"chart_payload": _chart_payload_from_labels(labels, 10), "peer_segment": {}})
    axis = dialog.figures["signal"].axes[0]
    hover = dialog.hover_controllers["signal"]
    before_lines = len(axis.lines)
    before_annotations = len(axis.texts)
    before_collections = len(axis.collections)

    class Event:
        inaxes = axis
        xdata = dialog.chart_payload["timestamp_numeric"][10]
        x = axis.transData.transform((xdata, 0))[0]
        y = 100
        ydata = 24

    for _index in range(100):
        hover.latest_event = Event()
        hover._process_latest_event()
    assert len(axis.lines) == before_lines
    assert len(axis.texts) == before_annotations
    assert len(axis.collections) == before_collections
    assert hover.popup is dialog.hover_controllers["signal"].popup
    hover.hide_hover()
    assert not hover.popup.isVisible()


def test_hover_popup_positions_inside_screen_and_uses_wrapping():
    _app()
    from PySide6.QtCore import QPoint
    from PySide6.QtGui import QGuiApplication

    from netconsole.ui.mesh_chart_hover_popup import MeshChartHoverPopup

    popup = MeshChartHoverPopup()
    popup.set_tooltip_text("Sample Time:\n2025-12-03 10:12:33.579\nPeerMac: <30f5-277a-5a2f>\n\nMR RSSI: 24\n" + "long text " * 30)
    assert popup.label.wordWrap()
    assert popup.label.maximumWidth() == 520
    assert "&lt;30f5-277a-5a2f&gt;" in popup.label.text()
    screen = QGuiApplication.primaryScreen()
    available = screen.availableGeometry()
    popup.show_at(available.bottomRight() - QPoint(2, 2))
    assert available.contains(popup.frameGeometry())
    assert popup.frameGeometry().right() < available.right()
    assert popup.frameGeometry().bottom() < available.bottom()
    popup.hide()
    popup.deleteLater()


def test_hover_uses_qt_popup_not_matplotlib_annotation(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.models.mesh_log_models import MeshMrProfile
    from netconsole.ui.dialogs.mesh_peer_detail_dialog import MeshPeerDetailDialog

    profile = MeshMrProfile("mr", "MR", "MR", datetime.now(), datetime.now())
    dialog = MeshPeerDetailDialog(I18n("en_US"), profile, tmp_path / "mesh.sqlite", "30f5277a5a2f", auto_load=False)
    dialog._on_loaded({"chart_payload": _chart_payload(20, 10), "peer_segment": {}})
    axis = dialog.figures["signal"].axes[0]
    hover = dialog.hover_controllers["signal"]
    before_texts = len(axis.texts)
    before_lines = len(axis.lines)

    class Event:
        inaxes = axis
        xdata = dialog.chart_payload["timestamp_numeric"][10]
        x = axis.transData.transform((xdata, 0))[0]
        y = 100
        ydata = 24

    hover.latest_event = Event()
    hover._process_latest_event()
    assert hover.popup.isVisible()
    assert "2025-12-03" in hover.popup.label.text()
    assert len(axis.texts) == before_texts
    assert len(axis.lines) == before_lines


def test_peer_dialog_is_maximizable_and_canvas_expands(tmp_path):
    _app()
    from PySide6.QtCore import Qt
    from PySide6.QtWidgets import QSizePolicy

    from netconsole.core.i18n import I18n
    from netconsole.models.mesh_log_models import MeshMrProfile
    from netconsole.ui.dialogs.mesh_peer_detail_dialog import MeshPeerDetailDialog

    profile = MeshMrProfile("mr", "MR", "MR", datetime.now(), datetime.now())
    dialog = MeshPeerDetailDialog(I18n("en_US"), profile, tmp_path / "mesh.sqlite", "30f5277a5a2f", auto_load=False)
    dialog._on_loaded({"chart_payload": _chart_payload(300, 150), "peer_segment": {}})
    dialog.time_scrollbar.setValue(25)
    assert dialog.windowFlags() & Qt.WindowMaximizeButtonHint
    assert not dialog.maximumSize().isValid() or dialog.maximumSize().width() > 1180
    assert dialog.canvases["signal"].sizePolicy().horizontalPolicy() == QSizePolicy.Expanding
    current_tab = dialog.tabs.currentIndex()
    current_scroll = dialog.time_scrollbar.value()
    dialog.showMaximized()
    dialog.showNormal()
    assert dialog.tabs.currentIndex() == current_tab
    assert dialog.time_scrollbar.value() == current_scroll


def test_peer_dialog_keeps_only_single_peer_tabs_and_active_dialog_has_full_active_tabs(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.models.mesh_log_models import MeshMrProfile
    from netconsole.ui.dialogs.mesh_peer_detail_dialog import MeshActiveLinkChartDialog, MeshPeerDetailDialog

    profile = MeshMrProfile("mr", "MR", "MR", datetime.now(), datetime.now())
    dialog = MeshPeerDetailDialog(I18n("en_US"), profile, tmp_path / "mesh.sqlite", "30f5277a5a2f", auto_load=False)
    dialog._on_loaded({"chart_payload": _chart_payload(100, 50), "peer_segment": {}})
    assert dialog.tab_keys == ["signal", "rssi_noise", "load"]
    tab_titles = [dialog.tabs.tabText(index) for index in range(dialog.tabs.count())]
    assert "All Active Link RSSI" not in tab_titles
    active_dialog = MeshActiveLinkChartDialog(I18n("en_US"), profile, tmp_path / "mesh.sqlite", auto_load=False)
    active_dialog._on_loaded({"chart_payload": _chart_payload(100, 50), "peer_segment": {}})
    assert active_dialog.tab_keys == ["active_next_rssi", "active_channel_load"]
    assert "Active / Next Active RSSI" not in tab_titles
    assert all("Next Active" not in title for title in tab_titles)
    assert "Link Duration" not in tab_titles
    assert "Transmission Reliability" not in tab_titles
    assert "Link State" not in tab_titles
    assert "Negotiated Rate (Raw Device Value)" not in tab_titles
    assert "duration" not in dialog.figures
    assert "reliability" not in dialog.canvases
    assert "rate" not in dialog.canvases
    assert "state" not in dialog.figures
    assert "duration" not in dialog.chart_artists
    assert "reliability" not in dialog.hover_controllers
    assert "rate" not in dialog.hover_controllers


def test_peer_dialog_uses_callback_without_qt_parent(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.models.mesh_log_models import MeshMrProfile
    from netconsole.ui.dialogs.mesh_peer_detail_dialog import MeshPeerDetailDialog, MeshSelectedPoint

    jumps: list[dict[str, object]] = []
    profile = MeshMrProfile("mr", "MR", "MR", datetime.now(), datetime.now())
    dialog = MeshPeerDetailDialog(
        I18n("en_US"),
        profile,
        tmp_path / "mesh.sqlite",
        "30f5277a5a2f",
        parent=None,
        auto_load=False,
        detail_jump_handler=jumps.append,
    )

    assert dialog.parent() is None
    dialog.jump_to_detail_row(MeshSelectedPoint(0, "s1", "2026-01-01 00:00:00.000", "30f5277a5a2f", "AP", "Site", "1", "radio1", "ACTIVE"))
    assert jumps and jumps[0]["session_id"] == "s1"
    assert "state" not in dialog.interaction_controllers


def test_current_active_rssi_chart_only_draws_current_mr_side_line(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.models.mesh_log_models import MeshMrProfile
    from netconsole.ui.dialogs.mesh_peer_detail_dialog import MeshActiveLinkChartDialog

    profile = MeshMrProfile("mr", "MR", "MR", datetime.now(), datetime.now())
    dialog = MeshActiveLinkChartDialog(I18n("en_US"), profile, tmp_path / "mesh.sqlite", auto_load=False)
    dialog._on_loaded({"chart_payload": _chart_payload(100, 50), "peer_segment": {}})
    dialog.tabs.setCurrentIndex(dialog.tab_keys.index("active_next_rssi"))
    fields = set(dialog.chart_artists["active_next_rssi"]["lines"])
    assert fields == {"active.active_local_rssi"}
    text = dialog.hover_controllers["active_next_rssi"].tooltip_text(0)
    assert "MR-AP_RSSI:" in text
    assert "Current Active MR RSSI" not in text
    assert "Next Active MR RSSI" not in text
    assert "Next Active Link" not in text
    assert "Current Active Peer RSSI" not in text
    assert "Next Active Peer RSSI" not in text


def test_active_channel_load_chart_only_draws_mr_side_busy(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.models.mesh_log_models import MeshMrProfile
    from netconsole.ui.dialogs.mesh_peer_detail_dialog import MeshActiveLinkChartDialog

    profile = MeshMrProfile("mr", "MR", "MR", datetime.now(), datetime.now())
    dialog = MeshActiveLinkChartDialog(I18n("en_US"), profile, tmp_path / "mesh.sqlite", auto_load=False)
    dialog._on_loaded({"chart_payload": _chart_payload(100, 50), "peer_segment": {}})
    dialog.tabs.setCurrentIndex(dialog.tab_keys.index("active_channel_load"))
    fields = set(dialog.chart_artists["active_channel_load"]["lines"])
    assert fields == {"active.active_local_tx_busy", "active.active_local_rx_busy"}
    axis = dialog.chart_artists["active_channel_load"]["axis"]
    assert axis.get_ylim() == pytest.approx((0, 100))
    text = dialog.hover_controllers["active_channel_load"].tooltip_text(0)
    assert "MR Tx Busy" in text
    assert "MR Rx Busy" in text
    assert "Next Active MR Tx Busy" not in text
    assert "Next Active MR Rx Busy" not in text
    assert "Peer Tx Busy" not in text
    assert "Peer Rx Busy" not in text


def test_active_role_rssi_does_not_break_on_peer_switch(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.models.mesh_log_models import MeshMrProfile
    from netconsole.ui.dialogs.mesh_peer_detail_dialog import MeshPeerDetailDialog

    payload = _chart_payload(4, 1)
    payload["active_peer_macs"] = ["a", "a", "b", "b"]
    payload["switch_indices"] = np.asarray([2], dtype=np.int32)
    profile = MeshMrProfile("mr", "MR", "MR", datetime.now(), datetime.now())
    dialog = MeshPeerDetailDialog(I18n("en_US"), profile, tmp_path / "mesh.sqlite", "30f5277a5a2f", auto_load=False)
    dialog.chart_payload = payload
    rendered = dialog._render_series_values("active.active_local_rssi", payload["active_series"]["active_local_rssi"])
    assert np.isfinite(rendered[2])


def test_active_channel_load_does_not_break_on_peer_switch(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.models.mesh_log_models import MeshMrProfile
    from netconsole.ui.dialogs.mesh_peer_detail_dialog import MeshPeerDetailDialog

    payload = _chart_payload(4, 1)
    payload["active_peer_macs"] = ["a", "a", "b", "b"]
    profile = MeshMrProfile("mr", "MR", "MR", datetime.now(), datetime.now())
    dialog = MeshPeerDetailDialog(I18n("en_US"), profile, tmp_path / "mesh.sqlite", "30f5277a5a2f", auto_load=False)
    dialog.chart_payload = payload
    rendered = dialog._render_series_values("active.active_local_tx_busy", payload["active_series"]["active_local_tx_busy"])
    assert np.isfinite(rendered[2])


def test_active_role_series_breaks_on_long_gap_only(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.models.mesh_log_models import MeshMrProfile
    from netconsole.ui.dialogs.mesh_peer_detail_dialog import MeshPeerDetailDialog

    labels = ["2025-12-03 10:00:00.000", "2025-12-03 10:00:01.000", "2025-12-03 10:10:00.000"]
    payload = _chart_payload_from_labels(labels, 0)
    payload["metadata"]["continuity_gap_seconds"] = 5
    profile = MeshMrProfile("mr", "MR", "MR", datetime.now(), datetime.now())
    dialog = MeshPeerDetailDialog(I18n("en_US"), profile, tmp_path / "mesh.sqlite", "30f5277a5a2f", auto_load=False)
    dialog.chart_payload = payload
    rendered = dialog._render_series_values("active.active_local_rssi", payload["active_series"]["active_local_rssi"])
    assert np.isfinite(rendered[1])
    assert np.isnan(rendered[2])


def test_active_role_short_missing_point_draws_dashed_bridge(tmp_path):
    _app()
    from matplotlib.collections import LineCollection
    from netconsole.core.i18n import I18n
    from netconsole.models.mesh_log_models import MeshMrProfile
    from netconsole.ui.dialogs.mesh_peer_detail_dialog import MeshActiveLinkChartDialog

    profile = MeshMrProfile("mr", "MR", "MR", datetime.now(), datetime.now())
    dialog = MeshActiveLinkChartDialog(I18n("en_US"), profile, tmp_path / "mesh.sqlite", auto_load=False)
    payload = _chart_payload(5, 0)
    payload["active_series"]["active_local_rssi"][2] = np.nan
    dialog._on_loaded({"chart_payload": payload, "peer_segment": {}})
    assert np.isnan(payload["active_series"]["active_local_rssi"][2])
    collections = dialog.chart_artists["active_next_rssi"]["collections"]
    assert any(isinstance(collection, LineCollection) for collection in collections)


def test_chart_payload_does_not_load_removed_chart_series(tmp_path):
    from netconsole.ui.mesh_chart_payload import build_chart_payload

    row = _chart_row(1, "2025-12-03 10:00:00.000")
    row["duration_seconds"] = 3
    row["expected_duration_seconds"] = 4
    row["duration_deviation_seconds"] = -1
    row["deltas"] = {"delta_local_retry": 1, "delta_peer_retry": 2, "delta_local_err": 3, "delta_peer_err": 4}
    payload = build_chart_payload({"anchor": row, "rows": [row]}, {"anchor": row, "rows": [row], "events": []})
    peer_series = payload["peer_series"]
    for key in ("duration", "expected_duration", "delta_local_retry", "delta_peer_retry", "delta_local_err", "delta_peer_err", "local_rate", "peer_rate"):
        assert key not in peer_series
    for key in ("active_peer_rssi", "next_peer_rssi", "active_peer_tx_busy", "active_peer_rx_busy", "next_peer_tx_busy", "next_peer_rx_busy"):
        assert key not in payload["active_series"]
    assert "peer_duration_deviation" not in payload
    assert row["duration_seconds"] == 3
    assert row["deltas"]["delta_local_err"] == 3


def test_active_channel_load_payload_does_not_include_next_or_peer_side_metrics():
    from netconsole.ui.mesh_chart_payload import build_chart_payload

    rows = [
        _payload_row(1, "2025-12-03 10:00:01.000", "30f5-277a-5a2f", "ACTIVE", 30, 50, 1, 2),
        _payload_row(2, "2025-12-03 10:00:01.000", "30f5-277a-5a3f", "STANDBY", 40, 60, 10, 20),
        _payload_row(3, "2025-12-03 10:00:02.000", "30f5-277a-5a2f", "ACTIVE", 31, 51, 3, 4),
        _payload_row(4, "2025-12-03 10:00:02.000", "30f5-277a-5a3f", "STANDBY", 42, 62, 12, 22),
        _payload_row(5, "2025-12-03 10:00:03.000", "30f5-277a-5a2f", "ACTIVE", 32, 52, 5, 6),
        _payload_row(6, "2025-12-03 10:00:03.000", "30f5-277a-5a3f", "STANDBY", 45, 65, 15, 24),
        _payload_row(7, "2025-12-03 10:00:04.000", "30f5-277a-5a3f", "ACTIVE", 46, 66, 16, 26),
    ]
    payload = build_chart_payload({"anchor": rows[0], "rows": rows}, {"anchor": rows[0], "rows": rows, "events": []})
    assert payload["active_series"]["active_local_tx_busy"][:3].tolist() == [1, 3, 5]
    assert payload["active_series"]["active_local_rx_busy"][:3].tolist() == [2, 4, 6]
    for key in ("next_local_tx_busy", "next_local_rx_busy", "active_peer_tx_busy", "active_peer_rx_busy", "next_peer_tx_busy", "next_peer_rx_busy"):
        assert key not in payload["active_series"]


def test_chart_payload_uses_compact_v2_scalar_metrics_without_json_payload():
    from netconsole.ui.mesh_chart_payload import build_chart_payload

    rows = [
        {
            "id": 1,
            "sample_time": "2025-12-03 10:00:01.000",
            "link_state": "ACTIVE",
            "peer_mac_normalized": "30f5277a5a2f",
            "peer_ap_name": "AP-01",
            "peer_site": "Station-01",
            "peer_radio": "radio1",
            "local_rssi_db": 36,
            "peer_rssi_db": 43,
            "local_tx_busy": 3,
            "local_rx_busy": 5,
        },
        {
            "id": 2,
            "sample_time": "2025-12-03 10:00:02.000",
            "link_state": "ACTIVE",
            "peer_mac_normalized": "30f5277a5a2f",
            "peer_ap_name": "AP-01",
            "peer_site": "Station-01",
            "peer_radio": "radio1",
            "local_rssi_db": 37,
            "peer_rssi_db": 44,
            "local_tx_busy": 4,
            "local_rx_busy": 6,
        },
    ]
    payload = build_chart_payload({"anchor": rows[0], "rows": rows}, {"anchor": rows[0], "rows": rows, "events": []})
    assert payload["peer_series"]["local_rssi"].tolist() == [36, 37]
    assert payload["active_series"]["active_local_rssi"].tolist() == [36, 37]
    assert payload["active_series"]["active_local_tx_busy"].tolist() == [3, 4]
    assert payload["active_series"]["active_local_rx_busy"].tolist() == [5, 6]


def test_current_active_payload_excludes_next_active_fields():
    from netconsole.ui.mesh_chart_payload import build_chart_payload

    rows = [
        _payload_row(1, "2025-12-03 10:00:01.000", "30f5-277a-5a2f", "ACTIVE", 24, 34),
        _payload_row(2, "2025-12-03 10:00:02.000", "30f5-277a-5a2f", "ACTIVE", 25, 35),
        _payload_row(3, "2025-12-03 10:00:02.000", "30f5-277a-5a3f", "STANDBY", 45, 47),
        _payload_row(4, "2025-12-03 10:00:03.000", "30f5-277a-5a3f", "ACTIVE", 44, 46),
    ]
    payload = build_chart_payload({"anchor": rows[0], "rows": rows[:2]}, {"anchor": rows[0], "rows": rows, "events": []})
    assert "next_peer_macs" not in payload
    assert "next_peer_change_indices" not in payload
    assert "next_local_rssi" not in payload["active_series"]


def test_current_active_payload_uses_canonical_mac_formats_for_runs():
    from netconsole.ui.mesh_chart_payload import build_chart_payload

    rows = [
        _payload_row(1, "2025-12-03 10:00:01.000", "30f5-277a-5a2f", "ACTIVE", 24, 34),
        _payload_row(2, "2025-12-03 10:00:02.000", "30F5277A5A3F", "ACTIVE", 44, 46),
    ]
    rows[1]["peer_mac_normalized"] = ""
    payload = build_chart_payload({"anchor": rows[0], "rows": [rows[0]]}, {"anchor": rows[0], "rows": rows, "events": []})
    assert [run["peer_mac"] for run in payload["active_runs"]] == ["30f5277a5a2f", "30f5277a5a3f"]


def test_hover_standby_links_use_compact_rssi_format_and_payload_order(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.models.mesh_log_models import MeshMrProfile
    from netconsole.ui.dialogs.mesh_peer_detail_dialog import MeshActiveLinkChartDialog
    from netconsole.ui.mesh_chart_payload import build_chart_payload

    rows = [
        _payload_row(1, "2025-12-03 10:00:00.000", "30f5-277a-5a2f", "ACTIVE", 45, 53),
        _payload_row(2, "2025-12-03 10:00:00.000", "30f5-277a-5a3f", "STANDBY", 27, 37),
        _payload_row(3, "2025-12-03 10:00:00.000", "30f5-277a-5a4f", "STANDBY", 31, None),
        _payload_row(4, "2025-12-03 10:00:00.000", "30f5-277a-5a5f", "STANDBY", None, 39),
        _payload_row(5, "2025-12-03 10:00:00.000", "30f5-277a-5a6f", "STANDBY", 33, 40),
        _payload_row(6, "2025-12-03 10:00:00.000", "30f5-277a-5a7f", "STANDBY", 34, 41),
        _payload_row(7, "2025-12-03 10:00:00.000", "30f5-277a-5a8f", "STANDBY", 35, 42),
    ]
    rows[0]["peer_ap_name"] = "AP-X_3111"
    rows[0]["peer_site"] = "31 Site"
    rows[1]["peer_ap_name"] = "AP-X_3110"
    rows[1]["peer_site"] = "31 Site"
    rows[1]["peer_radio"] = "radio2"
    rows[2]["peer_ap_name"] = ""
    rows[2]["peer_site"] = ""
    rows[2]["peer_radio"] = "radio2"
    rows[3]["peer_ap_name"] = "AP-X_3109"
    rows[3]["peer_site"] = "31 Site"
    rows[3]["peer_radio"] = "radio2"
    payload = build_chart_payload({"anchor": rows[0], "rows": [rows[0]]}, {"anchor": rows[0], "rows": rows, "events": []})

    profile = MeshMrProfile("mr", "MR", "MR", datetime.now(), datetime.now())
    dialog = MeshActiveLinkChartDialog(I18n("en_US"), profile, tmp_path / "mesh.sqlite", auto_load=False)
    dialog._on_loaded({"chart_payload": payload, "peer_segment": {}})
    text = dialog.hover_controllers["active_next_rssi"].tooltip_text(0)

    assert "AP-X_3111 / 31 Site" in text
    assert "MR-AP_RSSI: 45/53" in text
    assert "1. AP-X_3110 / 31 Site" in text
    assert "PeerMac: 30f5-277a-5a3f" in text
    assert "Peer Radio: radio2" in text
    assert "MR-AP_RSSI: 27/37" in text
    assert "State: STANDBY" in text
    assert "2. 30f5-277a-5a4f" in text
    assert "MR-AP_RSSI: 31/-" in text
    assert "5. 30f5-277a-5a7f" in text
    assert "6." not in text
    assert "……另有 1 条备份链路" in text


def test_active_payload_backup_links_are_isolated_by_source_and_allow_nearby_time():
    from netconsole.ui.mesh_chart_payload import build_chart_payload

    active = _payload_row(1, "2025-12-03 10:00:00.000", "30f5-277a-5a2f", "ACTIVE", 45, 53, source_file_id=1)
    same_source = _payload_row(2, "2025-12-03 10:00:00.600", "30f5-277a-5a3f", "STANDBY", 27, 37, source_file_id=1)
    other_source = _payload_row(3, "2025-12-03 10:00:00.000", "30f5-277a-5a4f", "STANDBY", 31, 39, source_file_id=2)
    payload = build_chart_payload({"anchor": active, "rows": [active]}, {"anchor": active, "rows": [active, same_source, other_source], "events": []})

    backups = payload["standby_links_by_index"][0]
    assert [item["peer_mac"] for item in backups] == ["30f5277a5a3f"]
    assert backups[0]["source_file_id"] == 1


def test_aba_active_switch_preserves_three_runs_and_rapid_flap():
    from netconsole.ui.mesh_chart_payload import build_chart_payload

    rows = [
        _payload_row(1, "2025-12-03 10:00:01.000", "30f5-277a-5a2f", "ACTIVE", 24, 34),
        _payload_row(2, "2025-12-03 10:00:01.000", "30f5-277a-5a3f", "STANDBY", 40, 42),
        _payload_row(3, "2025-12-03 10:00:02.000", "30f5-277a-5a2f", "ACTIVE", 25, 35),
        _payload_row(4, "2025-12-03 10:00:02.000", "30f5-277a-5a3f", "STANDBY", 41, 43),
        _payload_row(5, "2025-12-03 10:00:03.000", "30f5-277a-5a2f", "STANDBY", 26, 36),
        _payload_row(6, "2025-12-03 10:00:03.000", "30f5-277a-5a3f", "ACTIVE", 42, 44),
        _payload_row(7, "2025-12-03 10:00:04.000", "30f5-277a-5a2f", "ACTIVE", 27, 37),
        _payload_row(8, "2025-12-03 10:00:04.000", "30f5-277a-5a3f", "STANDBY", 43, 45),
    ]
    events = [
        {"event_type": EVENT_ACTIVE_SWITCH, "event_time": "2025-12-03 10:00:03.000"},
        {"event_type": EVENT_ACTIVE_SWITCH, "event_time": "2025-12-03 10:00:04.000", "is_rapid_flap": True},
    ]
    payload = build_chart_payload({"anchor": rows[0], "rows": rows}, {"anchor": rows[0], "rows": rows, "events": events, "estimated_interval_seconds": 1.0})
    assert [run["peer_mac"] for run in payload["active_runs"]] == ["30f5277a5a2f", "30f5277a5a3f", "30f5277a5a2f"]
    assert payload["switch_indices"].tolist() == [2, 3]
    assert "next_peer_macs" not in payload
    assert payload["rapid_flap_indices"].tolist() == [3]
    assert payload["rapid_flaps"][0]["is_rapid_flap"] is True


def test_active_switch_hover_includes_from_and_to_ap_names():
    from types import SimpleNamespace

    from netconsole.core.i18n import I18n
    from netconsole.ui.mesh_chart_hover import MeshChartHoverController
    from netconsole.ui.mesh_chart_payload import build_chart_payload

    rows = [
        _payload_row(1, "2025-12-03 10:00:01.000", "083b-e9ec-de2f", "ACTIVE", 24, 34),
        _payload_row(2, "2025-12-03 10:00:01.000", "94a7-482c-1def", "STANDBY", 40, 42),
        _payload_row(3, "2025-12-03 10:00:02.000", "083b-e9ec-de2f", "STANDBY", 25, 35),
        _payload_row(4, "2025-12-03 10:00:02.000", "94a7-482c-1def", "ACTIVE", 41, 43),
    ]
    rows[0]["peer_ap_name"] = "AP-X_3109"
    rows[0]["peer_site"] = "31 Site"
    rows[1]["peer_ap_name"] = "AP-X_3110"
    rows[1]["peer_site"] = "31 Site"
    rows[2]["peer_ap_name"] = "AP-X_3109"
    rows[2]["peer_site"] = "31 Site"
    rows[3]["peer_ap_name"] = "AP-X_3110"
    rows[3]["peer_site"] = "31 Site"
    events = [
        {
            "event_type": EVENT_ACTIVE_SWITCH,
            "event_time": "2025-12-03 10:00:02.000",
            "from_peer_mac": "083b-e9ec-de2f",
            "to_peer_mac": "94a7-482c-1def",
        }
    ]

    payload = build_chart_payload({"anchor": rows[0], "rows": rows}, {"anchor": rows[0], "rows": rows, "events": events})
    event = payload["events_by_index"][1][0]
    assert event["from_peer_ap_name"] == "AP-X_3109"
    assert event["to_peer_ap_name"] == "AP-X_3110"

    controller = SimpleNamespace(payload=payload, i18n=I18n("zh_CN"))
    lines = MeshChartHoverController._event_lines(controller, 1)
    assert any("AP-X_3109 / 31 Site / 083b-e9ec-de2f" in line for line in lines)
    assert any("AP-X_3110 / 31 Site / 94a7-482c-1def" in line for line in lines)


def test_no_active_and_multi_active_do_not_generate_active_series():
    from netconsole.ui.mesh_chart_payload import build_chart_payload

    rows = [
        _payload_row(1, "2025-12-03 10:00:01.000", "30f5-277a-5a2f", "STANDBY", 24, 34),
        _payload_row(2, "2025-12-03 10:00:02.000", "30f5-277a-5a2f", "ACTIVE", 25, 35),
        _payload_row(3, "2025-12-03 10:00:02.000", "30f5-277a-5a3f", "ACTIVE", 45, 47),
    ]
    payload = build_chart_payload({"anchor": rows[0], "rows": rows}, {"anchor": rows[0], "rows": rows, "events": []})
    assert payload["no_active_indices"].tolist() == [0]
    assert payload["multi_active_indices"].tolist() == [1]
    assert np.isnan(payload["active_series"]["active_local_rssi"][0])
    assert np.isnan(payload["active_series"]["active_local_rssi"][1])


def test_hover_same_index_updates_content_once_and_only_moves_popup(tmp_path, monkeypatch):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.models.mesh_log_models import MeshMrProfile
    from netconsole.ui.dialogs.mesh_peer_detail_dialog import MeshPeerDetailDialog

    profile = MeshMrProfile("mr", "MR", "MR", datetime.now(), datetime.now())
    dialog = MeshPeerDetailDialog(I18n("en_US"), profile, tmp_path / "mesh.sqlite", "30f5277a5a2f", auto_load=False)
    dialog._on_loaded({"chart_payload": _chart_payload(100, 50), "peer_segment": {}})
    axis = dialog.figures["signal"].axes[0]
    hover = dialog.hover_controllers["signal"]
    counts = {"text": 0, "draw": 0, "move": 0}
    original_text = hover.tooltip_text
    original_draw = dialog.canvases["signal"].draw_idle
    original_show_at = hover.popup.show_at

    def tooltip_text(index):
        counts["text"] += 1
        return original_text(index)

    def draw_idle():
        counts["draw"] += 1
        return original_draw()

    def show_at(*args, **kwargs):
        counts["move"] += 1
        return original_show_at(*args, **kwargs)

    monkeypatch.setattr(hover, "tooltip_text", tooltip_text)
    monkeypatch.setattr(dialog.canvases["signal"], "draw_idle", draw_idle)
    monkeypatch.setattr(hover.popup, "show_at", show_at)

    class Event:
        inaxes = axis
        xdata = dialog.chart_payload["timestamp_numeric"][50]
        x = axis.transData.transform((xdata, 0))[0]
        y = 100
        ydata = 24

    for offset in range(100):
        event = Event()
        event.x += min(offset % 3, 1)
        hover.latest_event = event
        hover._process_latest_event()
    assert counts["text"] == 1
    assert counts["draw"] == 1
    assert counts["move"] == 100
    assert len(hover.content_cache) == 1


def test_hover_content_cache_hits_when_returning_to_index(tmp_path, monkeypatch):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.models.mesh_log_models import MeshMrProfile
    from netconsole.ui.dialogs.mesh_peer_detail_dialog import MeshPeerDetailDialog

    profile = MeshMrProfile("mr", "MR", "MR", datetime.now(), datetime.now())
    dialog = MeshPeerDetailDialog(I18n("en_US"), profile, tmp_path / "mesh.sqlite", "30f5277a5a2f", auto_load=False)
    dialog._on_loaded({"chart_payload": _chart_payload(100, 20), "peer_segment": {}})
    hover = dialog.hover_controllers["signal"]
    calls = {"text": 0}
    original = hover.tooltip_text

    def tooltip_text(index):
        calls["text"] += 1
        return original(index)

    monkeypatch.setattr(hover, "tooltip_text", tooltip_text)
    assert hover._cached_tooltip_text(10)
    assert hover._cached_tooltip_text(11)
    assert hover._cached_tooltip_text(10)
    assert calls["text"] == 2
    assert len(hover.content_cache) == 2


def test_wheel_zoom_keeps_cursor_sample_position_and_shared_window(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.models.mesh_log_models import MeshMrProfile
    from netconsole.ui.dialogs.mesh_peer_detail_dialog import MeshPeerDetailDialog

    profile = MeshMrProfile("mr", "MR", "MR", datetime.now(), datetime.now())
    dialog = MeshPeerDetailDialog(I18n("en_US"), profile, tmp_path / "mesh.sqlite", "30f5277a5a2f", auto_load=False)
    dialog._on_loaded({"chart_payload": _chart_payload(1000, 450), "peer_segment": {}})
    dialog.set_time_window(400, 120, "preset")
    cursor_index = 460
    dialog.zoom_time_window_at("signal", dialog.chart_payload["timestamp_numeric"][cursor_index], 1)
    assert dialog.visible_sample_count < 120
    ratio = (cursor_index - dialog.window_start_index) / max(dialog.visible_sample_count - 1, 1)
    assert ratio == pytest.approx(60 / 119, abs=0.05)
    assert dialog.time_scrollbar.value() == dialog.window_start_index
    dialog.tabs.setCurrentIndex(dialog.tab_keys.index("load"))
    assert dialog.window_start_index == dialog.time_window_controller.window_start_index
    assert dialog.visible_sample_count == dialog.time_window_controller.visible_sample_count


def test_wheel_zoom_out_clamps_to_all_without_blank_area(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.models.mesh_log_models import MeshMrProfile
    from netconsole.ui.dialogs.mesh_peer_detail_dialog import MeshPeerDetailDialog

    profile = MeshMrProfile("mr", "MR", "MR", datetime.now(), datetime.now())
    dialog = MeshPeerDetailDialog(I18n("en_US"), profile, tmp_path / "mesh.sqlite", "30f5277a5a2f", auto_load=False)
    dialog._on_loaded({"chart_payload": _chart_payload(1000, 100), "peer_segment": {}})
    dialog.set_time_window(300, 900, "preset")
    for _index in range(4):
        dialog.zoom_time_window_at("signal", dialog.chart_payload["timestamp_numeric"][500], -1)
    assert dialog.effective_visible_sample_count() == 1000
    assert dialog.window_start_index == 0
    assert dialog.time_scrollbar.maximum() == 0


def test_custom_visible_samples_combo_replaces_single_custom_item(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.models.mesh_log_models import MeshMrProfile
    from netconsole.ui.dialogs.mesh_peer_detail_dialog import MeshPeerDetailDialog

    profile = MeshMrProfile("mr", "MR", "MR", datetime.now(), datetime.now())
    dialog = MeshPeerDetailDialog(I18n("zh_CN"), profile, tmp_path / "mesh.sqlite", "30f5277a5a2f", auto_load=False)
    dialog._on_loaded({"chart_payload": _chart_payload(1000, 100), "peer_segment": {}})
    dialog.set_time_window(0, 96, "wheel")
    assert dialog.visible_samples_combo.currentText() == "自定义（96）"
    dialog.set_time_window(0, 48, "wheel")
    assert dialog.visible_samples_combo.currentText() == "自定义（48）"
    assert sum(1 for index in range(dialog.visible_samples_combo.count()) if "自定义" in dialog.visible_samples_combo.itemText(index)) == 1
    dialog.visible_samples_combo.setCurrentIndex(dialog.visible_samples_combo.findData(120))
    assert dialog.visible_sample_count == 120


def test_left_drag_pans_timeline_and_pauses_hover(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.models.mesh_log_models import MeshMrProfile
    from netconsole.ui.dialogs.mesh_peer_detail_dialog import MeshPeerDetailDialog

    profile = MeshMrProfile("mr", "MR", "MR", datetime.now(), datetime.now())
    dialog = MeshPeerDetailDialog(I18n("en_US"), profile, tmp_path / "mesh.sqlite", "30f5277a5a2f", auto_load=False)
    dialog._on_loaded({"chart_payload": _chart_payload(1000, 200), "peer_segment": {}})
    dialog.set_time_window(200, 100, "preset")
    interaction = dialog.interaction_controllers["signal"]
    interaction.axis_pixel_width = 1000
    interaction.visible_sample_count = 100
    interaction.drag_start_mouse_x = 500
    interaction.drag_start_window_index = 200
    interaction.dragging = True

    class Event:
        x = 400

    interaction.on_motion(Event())
    interaction._apply_pending_drag()
    assert dialog.window_start_index == 210
    assert dialog.time_scrollbar.value() == 210
    assert dialog.hover_controllers["signal"].paused
    interaction.on_release(Event())
    dialog._resume_hover_after_interaction()
    assert not dialog.hover_controllers["signal"].paused


def test_drag_threshold_does_not_pan(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.models.mesh_log_models import MeshMrProfile
    from netconsole.ui.dialogs.mesh_peer_detail_dialog import MeshPeerDetailDialog

    profile = MeshMrProfile("mr", "MR", "MR", datetime.now(), datetime.now())
    dialog = MeshPeerDetailDialog(I18n("en_US"), profile, tmp_path / "mesh.sqlite", "30f5277a5a2f", auto_load=False)
    dialog._on_loaded({"chart_payload": _chart_payload(1000, 200), "peer_segment": {}})
    dialog.set_time_window(200, 100, "preset")
    interaction = dialog.interaction_controllers["signal"]
    interaction.axis_pixel_width = 1000
    interaction.visible_sample_count = 100
    interaction.drag_start_mouse_x = 500
    interaction.drag_start_window_index = 200
    interaction.dragging = True

    class Event:
        x = 498

    interaction.on_motion(Event())
    interaction.on_release(Event())
    assert dialog.window_start_index == 200
    assert dialog.interaction_state == "IDLE"


def test_active_switch_event_uses_raw_positive_rssi(tmp_path):
    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-01")
    first = LINE_A.replace("36/43", "24/34")
    second = LINE_B.replace("37/44", "44/46")
    source = tmp_path / "meshlog.log"
    source.write_text("\n".join(["[1] 2025/12/03 10:00:01.000", first, "[1] 2025/12/03 10:00:02.000", second]), encoding="utf-8")
    MeshImportService("demo", paths).import_files(profile, [source])
    repo = MeshMrRepository(paths.mesh_mr_db_path("demo", profile.safe_folder_name))
    _, events = repo.query_events(100, 0)
    switch = next(event for event in events if event["event_type"] == EVENT_ACTIVE_SWITCH)
    details = json.loads(str(switch["details_json"]))
    assert details["from_local_rssi"] == 24
    assert details["to_local_rssi"] == 44
    assert details["from_peer_rssi"] == 34
    assert details["to_peer_rssi"] == 46
    assert details["from_local_signal_dbm"] < 0
    assert details["to_local_signal_dbm"] < 0


def test_aba_active_switch_events_are_not_merged_and_return_is_rapid_flap(tmp_path):
    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-01")
    source = tmp_path / "meshlog.log"
    source.write_text(
        "\n".join(
            [
                "[1] 2025/12/03 10:00:01.000",
                LINE_A,
                "[1] 2025/12/03 10:00:02.000",
                LINE_B,
                "[1] 2025/12/03 10:00:03.000",
                LINE_A,
            ]
        ),
        encoding="utf-8",
    )
    MeshImportService("demo", paths).import_files(profile, [source])
    repo = MeshMrRepository(paths.mesh_mr_db_path("demo", profile.safe_folder_name))
    _, events = repo.query_events(100, 0)
    switches = [event for event in events if event["event_type"] == EVENT_ACTIVE_SWITCH]
    assert [event["event_time"] for event in switches] == ["2025-12-03 10:00:02.000", "2025-12-03 10:00:03.000"]
    assert switches[0]["from_peer_mac"] == "30f5-277a-5a2f"
    assert switches[0]["to_peer_mac"] == "30f5-277a-5a3f"
    assert switches[1]["from_peer_mac"] == "30f5-277a-5a3f"
    assert switches[1]["to_peer_mac"] == "30f5-277a-5a2f"
    details = json.loads(str(switches[1]["details_json"]))
    assert details["is_rapid_flap"] is True
    assert details["rapid_flap_middle_peer"] == "30f5277a5a3f"


def test_mesh_repository_lists_pages_in_record_sequence_order(tmp_path):
    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-01")
    first = tmp_path / "first.log"
    second = tmp_path / "second.log"
    first.write_text("[1] 2025/12/03 10:00:01.000\n" + LINE_A + "\n", encoding="utf-8")
    second.write_text("[1] 2025/12/03 10:00:03.000\n" + LINE_B + "\n", encoding="utf-8")
    MeshImportService("demo", paths).import_files(profile, [second, first])
    repo = MeshMrRepository(paths.mesh_mr_db_path("demo", profile.safe_folder_name))
    _, links = repo.query_links(1, 0)
    assert links[0]["sample_time"] == "2025-12-03 10:00:03.000"
    assert links[0]["source_file_order"] == 1
    _, events = repo.query_events(10, 0)
    assert [event["event_time"] for event in events if event["event_type"] == EVENT_ACTIVE_SWITCH] == []
    sources = repo.list_source_files()
    assert [source["first_sample_time"] for source in sources] == ["2025-12-03 10:00:01.000", "2025-12-03 10:00:03.000"]
    first_detail = Path(sources[0]["parsed_db_path"])
    with sqlite3.connect(first_detail) as conn:
        conn.execute(
            """
            INSERT INTO parse_issues (source_file_id, source_file, line_number, severity, issue_type, field_name, message, raw_line_start, raw_line_end)
            VALUES (1, 'z.log', 9, 'ERROR', 'x', 'x', 'x', 9, 9), (1, 'a.log', 2, 'ERROR', 'x', 'x', 'x', 2, 2)
            """
        )
    _, issues = repo.query_issues(10, 0)
    assert [(issue["source_file"], issue["line_number"]) for issue in issues] == [("a.log", 2), ("z.log", 9)]


def test_mesh_page_defaults_to_1000_and_removes_export_buttons(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.ui.pages.mesh_log_analysis_page import MESH_DEFAULT_PAGE_SIZE, MeshLogAnalysisPage
    from netconsole.ui.pagination import DEFAULT_PAGE_SIZE

    page = MeshLogAnalysisPage(I18n("en_US"), "demo", PathResolver(tmp_path))
    assert DEFAULT_PAGE_SIZE == 200
    assert MESH_DEFAULT_PAGE_SIZE == 1000
    assert page.page_size == 1000
    assert page.source_pagination.page_size_combo.currentData() == 1000
    assert page.link_pagination.page_size_combo.currentData() == 1000
    assert page.event_pagination.page_size_combo.currentData() == 1000
    assert page.issue_pagination.page_size_combo.currentData() == 1000
    assert not hasattr(page, "export_detail_button")
    assert not hasattr(page, "export_event_button")
    assert not hasattr(page, "export_session_button")
    assert not hasattr(page, "export_table")


def test_mesh_page_mr_selection_is_debounced(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.ui.pages.mesh_log_analysis_page import MeshLogAnalysisPage

    page = MeshLogAnalysisPage(I18n("en_US"), "demo", PathResolver(tmp_path))
    assert page.mr_selection_timer.isSingleShot()
    assert 100 <= page.mr_selection_timer.interval() <= 150


def test_mesh_page_refresh_all_does_not_double_load_selected_mr(tmp_path, monkeypatch):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.ui.pages.mesh_log_analysis_page import MeshLogAnalysisPage

    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-01")
    page = MeshLogAnalysisPage(I18n("en_US"), "demo", paths)
    calls = []

    def load_profile(mr_id):
        calls.append(mr_id)
        page.current_profile = page.profile_by_id.get(mr_id)

    monkeypatch.setattr(page, "_load_profile_by_id", load_profile)
    page.refresh_all(select_mr_id=profile.mr_id)
    assert calls == [profile.mr_id]


def test_mesh_page_tab_lazy_loads_only_current_tab(tmp_path, monkeypatch):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.ui.pages.mesh_log_analysis_page import MeshLogAnalysisPage

    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-01")
    page = MeshLogAnalysisPage(I18n("en_US"), "demo", paths)
    page.current_profile = profile
    page.dirty_tabs = {"source", "link", "event", "issue"}
    calls = []
    monkeypatch.setattr(page, "_ensure_current_derived_analysis", lambda repo: None)
    monkeypatch.setattr(page, "_start_tab_load", lambda tab: calls.append(tab))
    page.tabs.setCurrentIndex(1)
    assert calls == ["link"]
    assert "source" in page.dirty_tabs
    assert "event" in page.dirty_tabs


def test_mesh_page_link_tab_shows_loading_overlay_until_async_load_finishes(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.ui.pages.mesh_log_analysis_page import MeshLogAnalysisPage

    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-01")
    source = tmp_path / "meshlog.log"
    source.write_text("[1] 2025/12/03 10:12:33.000\n" + LINE_A + "\n", encoding="utf-8")
    MeshImportService("demo", paths).import_files(profile, [source])
    page = MeshLogAnalysisPage(I18n("zh_CN"), "demo", paths)
    page.current_profile = profile
    page.dirty_tabs.add("link")

    page.tabs.setCurrentIndex(1)

    assert not page.tab_overlays["link"].isHidden()
    assert page.tab_overlays["link"].spinner.timer.isActive()
    assert "正在加载链路明细" in page.progress_label.text()
    _wait_for_mesh_tab_load(page)
    assert page.tab_overlays["link"].isHidden()
    assert not page.tab_overlays["link"].spinner.timer.isActive()
    assert page.link_table.rowCount() == 1


def test_mesh_page_repository_cache_reuses_current_mr_repo(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.ui.pages.mesh_log_analysis_page import MeshLogAnalysisPage

    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-01")
    page = MeshLogAnalysisPage(I18n("en_US"), "demo", paths)
    page.current_profile = profile
    assert page._repo() is page._repo()


def test_parse_issues_empty_state_and_count(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.ui.pages.mesh_log_analysis_page import MeshLogAnalysisPage

    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-01")
    repo = MeshMrRepository(paths.mesh_mr_db_path("demo", profile.safe_folder_name))
    page = MeshLogAnalysisPage(I18n("en_US"), "demo", paths)
    page.current_profile = profile
    page.refresh_parse_issues(repo)
    assert page.tabs.tabText(4) == "Parse Issues (0)"
    assert not page.issue_empty_widget.isHidden()
    assert page.issue_table.isHidden()
    assert page.issue_pagination.isHidden()
    assert page.issue_empty_title.text() == "No parsing issues found"


def test_parse_issues_table_returns_when_issues_exist(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.ui.pages.mesh_log_analysis_page import MeshLogAnalysisPage

    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-01")
    repo = MeshMrRepository(paths.mesh_mr_db_path("demo", profile.safe_folder_name))
    with sqlite3.connect(repo.path) as conn:
        conn.execute(
            """
            INSERT INTO parse_issues (
                source_file_id, source_file, line_number, severity, issue_type,
                field_name, message, raw_line_start, raw_line_end
            ) VALUES (NULL, 'bad.log', 7, 'ERROR', 'field_count', 'metrics', 'too few fields', 7, 7)
            """
        )
    page = MeshLogAnalysisPage(I18n("en_US"), "demo", paths)
    page.current_profile = profile
    page.refresh_parse_issues(repo)
    assert page.tabs.tabText(4) == "Parse Issues (1)"
    assert page.issue_empty_widget.isHidden()
    assert not page.issue_table.isHidden()
    assert not page.issue_pagination.isHidden()
    assert page.issue_table.rowCount() == 1
    assert page.issue_table.item(0, 0).text() == "bad.log"


def test_old_derived_events_upgrade_to_raw_positive_rssi(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.ui.mesh_log_workers import MeshDerivedAnalysisRebuildWorker
    from netconsole.ui.pages.mesh_log_analysis_page import MeshLogAnalysisPage

    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-01")
    first = LINE_A.replace("36/43", "24/34")
    second = LINE_B.replace("37/44", "44/46")
    source = tmp_path / "meshlog.log"
    source.write_text("\n".join(["[1] 2025/12/03 10:00:01.000", first, "[1] 2025/12/03 10:00:02.000", second]), encoding="utf-8")
    MeshImportService("demo", paths).import_files(profile, [source])
    db_path = paths.mesh_mr_db_path("demo", profile.safe_folder_name)
    detail_path = Path(MeshMrRepository(db_path).list_source_files()[0]["parsed_db_path"])
    with sqlite3.connect(detail_path) as conn:
        conn.execute("DELETE FROM schema_meta WHERE key = 'derived_analysis_version'")
        conn.execute(
            "UPDATE switch_events SET details_json = ? WHERE event_type = ?",
            (json.dumps({"from_local_signal_dbm": -65, "to_local_signal_dbm": -50, "from_peer_signal_dbm": -71, "to_peer_signal_dbm": -60}), EVENT_ACTIVE_SWITCH),
        )
    repo = MeshMrRepository(db_path)
    assert repo.needs_derived_analysis_rebuild()
    link_count_before = repo.summary()["link_record_count"]
    worker = MeshDerivedAnalysisRebuildWorker(db_path)
    completed = []
    worker.completed.connect(lambda: completed.append(True))
    worker.run()
    assert completed
    assert not repo.needs_derived_analysis_rebuild()
    assert repo.summary()["link_record_count"] == link_count_before
    _, events = repo.query_events(100, 0)
    details = json.loads(str(next(event for event in events if event["event_type"] == EVENT_ACTIVE_SWITCH)["details_json"]))
    assert details["from_local_rssi"] == 24
    assert details["to_local_rssi"] == 44

    page = MeshLogAnalysisPage(I18n("en_US"), "demo", paths)
    page.current_profile = MeshCatalogRepository(paths.mesh_catalog_path("demo")).get_profile(profile.mr_id)
    page._render_events(repo)
    assert page.event_table.item(0, 6).text() == "24"
    assert page.event_table.item(0, 7).text() == "44"


def test_session_filter_visibility_and_order(tmp_path):
    _app()
    from netconsole.core.i18n import I18n
    from netconsole.models.mesh_log_models import MeshMrProfile
    from netconsole.ui.dialogs.mesh_peer_detail_dialog import MeshPeerDetailDialog

    profile = MeshMrProfile("mr", "MR", "MR", datetime.now(), datetime.now())
    dialog = MeshPeerDetailDialog(I18n("en_US"), profile, tmp_path / "mesh.sqlite", "30f5277a5a2f", auto_load=False)
    no_session = _chart_payload(10, 1)
    no_session["session_options"] = []
    no_session["peer_session_ids"] = [""] * 10
    dialog._on_loaded({"chart_payload": no_session, "peer_segment": {}})
    assert dialog.session_filter_container.isHidden()

    one_session = _chart_payload(10, 1)
    one_session["session_options"] = [{"session_id": "s1", "first_sample_time": "2025-12-03 10:00:00.000", "last_sample_time": "2025-12-03 10:00:09.000"}]
    one_session["peer_session_ids"] = ["s1"] * 10
    dialog._on_loaded({"chart_payload": one_session, "peer_segment": {}})
    assert dialog.session_filter_container.isHidden()
    assert dialog.current_session_id == "s1"

    multi = _chart_payload(30, 1)
    multi["session_options"] = [
        {"session_id": "s2", "first_sample_time": "2025-12-03 10:01:00.000", "last_sample_time": "2025-12-03 10:01:09.000"},
        {"session_id": "s1", "first_sample_time": "2025-12-03 10:00:00.000", "last_sample_time": "2025-12-03 10:00:09.000"},
        {"session_id": "s3", "first_sample_time": "2025-12-03 10:02:00.000", "last_sample_time": "2025-12-03 10:02:09.000"},
    ]
    multi["session_options"] = sorted(multi["session_options"], key=lambda item: item["first_sample_time"])
    multi["peer_session_ids"] = ["s1"] * 10 + ["s2"] * 10 + ["s3"] * 10
    dialog.initial_session_id = "s2"
    dialog._on_loaded({"chart_payload": multi, "peer_segment": {}})
    assert not dialog.session_filter_container.isHidden()
    assert dialog.session_filter.count() == 4
    assert dialog.session_filter.itemData(0) == ""
    assert [dialog.session_filter.itemData(index) for index in range(1, 4)] == ["s1", "s2", "s3"]
    assert dialog.session_filter.currentData() == "s2"


def _insert_mesh_samples(db_path, first_count: int, second_count: int) -> None:
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            """
            INSERT INTO source_files (
                mr_id, original_path, archived_path, original_filename, archived_filename, sha256,
                file_size, file_mtime, imported_at, parser_version, parse_status
            ) VALUES ('mr', 'a.log', 'a.log', 'a.log', 'a.log', 'sha', 1, NULL, '2025-12-03 00:00:00.000', 'test', 'ok')
            """
        )
        rows = []
        links = []
        link_id = 1
        first_start = datetime(2025, 12, 3, 10, 0, 0)
        second_start = datetime.fromtimestamp(first_start.timestamp() + first_count + 1800)
        for run_start, count in ((first_start, first_count), (second_start, second_count)):
            for index in range(count):
                sample_time = run_start.timestamp() + index
                dt = datetime.fromtimestamp(sample_time).strftime("%Y-%m-%d %H:%M:%S.000")
                rows.append((link_id, 1, 1, dt, int(sample_time * 1000), ""))
                links.append(
                        (
                            link_id,
                            link_id,
                            1,
                            1,
                            link_id,
                            1,
                        1,
                        dt,
                        "Active",
                        "ACTIVE",
                        "30f5-277a-5a2f",
                        "30f5277a5a2f",
                        "2025-12-03 09:59:00.000",
                        "0d 00h 00m 01s",
                        1,
                        "s1",
                        36,
                        43,
                        f"fp-{link_id}",
                    )
                )
                link_id += 1
        conn.executemany("INSERT INTO samples(id, source_file_id, radio, sample_time, sample_time_epoch_ms, timestamp_tag) VALUES (?, ?, ?, ?, ?, ?)", rows)
        conn.executemany(
            """
                INSERT INTO mesh_links (
                    id, sample_id, source_file_id, source_file_order, record_seq, source_line_number, radio, sample_time,
                    link_state_raw, link_state, peer_mac_raw, peer_mac_normalized, establish_time,
                    duration_text, duration_seconds, session_id, local_rssi_db, peer_rssi_db, record_fingerprint
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
            links,
        )
    MeshMrRepository(db_path).rebuild_derived_analysis()


def _active_point(sample_time: str, peer: str, ap_mac: str, ap_name: str, peer_radio: str) -> dict[str, object]:
    normalized = "".join(character for character in peer.lower() if character in "0123456789abcdef")
    return {
        "id": int(datetime.fromisoformat(sample_time).timestamp() * 1000),
        "source_file_id": 1,
        "radio": 1,
        "sample_time": sample_time,
        "peer_mac_raw": peer,
        "peer_mac_normalized": normalized,
        "peer_mac": normalized,
        "peer_ap_name": ap_name,
        "peer_site": "03横溪站",
        "peer_radio": peer_radio,
        "peer_radio_label": peer_radio,
        "peer_ap_mac": ap_mac,
        "peer_radio_mac": peer,
        "duration_seconds": 1,
        "local_rssi_db": 35,
        "peer_rssi_db": 38,
        "local_tx_busy": 1,
        "local_rx_busy": 3,
        "peer_tx_busy": 1,
        "peer_rx_busy": 3,
        "source_file": "mesh.log",
    }


def _chart_row(link_id: int, sample_time: str) -> dict[str, object]:
    return {
        "id": link_id,
        "sample_time": sample_time,
        "peer_mac_normalized": "30f5277a5a2f",
        "peer_mac_raw": "30f5-277a-5a2f",
        "link_state": "ACTIVE",
        "metrics_json": '{"local_rssi_db": 36, "peer_rssi_db": 43, "local_tx_busy": 3, "peer_tx_busy": 4, "local_rx_busy": 5, "peer_rx_busy": 6}',
        "deltas_json": "{}",
        "metrics": {"local_rssi_db": 36, "peer_rssi_db": 43, "local_tx_busy": 3, "peer_tx_busy": 4, "local_rx_busy": 5, "peer_rx_busy": 6},
        "deltas": {},
    }


def _payload_row(
    link_id: int,
    sample_time: str,
    peer_mac: str,
    state: str,
    local_rssi: int | None,
    peer_rssi: int | None,
    local_tx_busy: int = 3,
    local_rx_busy: int = 5,
    source_file_id: int = 1,
    radio: int = 1,
) -> dict[str, object]:
    normalized = "".join(character for character in peer_mac.lower() if character in "0123456789abcdef")
    return {
        "id": link_id,
        "source_file_id": source_file_id,
        "radio": radio,
        "sample_time": sample_time,
        "peer_mac_normalized": normalized,
        "peer_mac_raw": peer_mac,
        "link_state": state,
        "metrics_json": json.dumps({"local_rssi_db": local_rssi, "peer_rssi_db": peer_rssi, "local_tx_busy": local_tx_busy, "local_rx_busy": local_rx_busy}),
        "deltas_json": "{}",
        "metrics": {
            "local_rssi_db": local_rssi,
            "peer_rssi_db": peer_rssi,
            "local_tx_busy": local_tx_busy,
            "peer_tx_busy": 4,
            "local_rx_busy": local_rx_busy,
            "peer_rx_busy": 6,
        },
        "deltas": {},
    }


def _chart_payload(count: int, anchor_index: int) -> dict[str, object]:
    timestamps = [datetime.fromtimestamp(datetime(2025, 12, 3, 10, 0, 0).timestamp() + index) for index in range(count)]
    labels = [value.strftime("%Y-%m-%d %H:%M:%S.000") for value in timestamps]
    numeric = np.asarray([date2num(value) for value in timestamps], dtype=np.float64)
    peer = {
        "local_rssi": np.linspace(30, 60, count, dtype=np.float32),
        "peer_rssi": np.linspace(40, 70, count, dtype=np.float32),
        "local_noise": np.full(count, 88, dtype=np.float32),
        "peer_noise": np.full(count, 105, dtype=np.float32),
        "local_tx_busy": np.full(count, 3, dtype=np.float32),
        "peer_tx_busy": np.full(count, 4, dtype=np.float32),
        "local_rx_busy": np.full(count, 5, dtype=np.float32),
        "peer_rx_busy": np.full(count, 6, dtype=np.float32),
        "state": np.ones(count, dtype=np.int8),
    }
    active = {
        "active_local_rssi": peer["local_rssi"].copy(),
        "active_local_tx_busy": peer["local_tx_busy"].copy(),
        "active_local_rx_busy": peer["local_rx_busy"].copy(),
    }
    return {
        "metadata": {
            "anchor": {"peer_mac_normalized": "30f5277a5a2f"},
            "anchor_index": anchor_index,
            "anchor_sample_time": labels[anchor_index],
            "segment_start": labels[0],
            "segment_end": labels[-1],
            "estimated_interval_seconds": 1.0,
            "sample_count": count,
            "peer_sample_count": count,
            "backend": "matplotlib-cpu",
        },
        "timestamps": timestamps,
        "timestamp_labels": labels,
        "timestamp_numeric": numeric,
        "peer_series": peer,
        "active_series": active,
        "active_peer_macs": ["30f5277a5a2f"] * count,
        "peer_macs": ["30f5277a5a2f"] * count,
        "peer_link_states": ["ACTIVE"] * count,
        "peer_establish_times": ["2025-12-03 09:59:00.000"] * count,
        "peer_session_ids": ["s1"] * count,
        "session_options": [{"session_id": "s1", "first_sample_time": labels[0], "last_sample_time": labels[-1]}],
        "events_by_index": {},
        "switch_indices": np.asarray([anchor_index], dtype=np.int32),
        "no_active_indices": np.asarray([], dtype=np.int32),
        "multi_active_indices": np.asarray([], dtype=np.int32),
        "peer_change_indices": np.asarray([], dtype=np.int32),
        "important_indices": np.asarray([anchor_index], dtype=np.int32),
    }


def _chart_payload_from_labels(labels: list[str], anchor_index: int) -> dict[str, object]:
    payload = _chart_payload(len(labels), anchor_index)
    timestamps = [datetime.fromisoformat(label) for label in labels]
    payload["timestamps"] = timestamps
    payload["timestamp_labels"] = labels
    payload["timestamp_numeric"] = np.asarray([date2num(value) for value in timestamps], dtype=np.float64)
    payload["metadata"]["anchor_sample_time"] = labels[anchor_index]
    payload["metadata"]["segment_start"] = labels[0]
    payload["metadata"]["segment_end"] = labels[-1]
    return payload
