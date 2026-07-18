from __future__ import annotations

import json
import sqlite3
import sys
from datetime import datetime, timedelta
from pathlib import Path
from threading import Event
from types import SimpleNamespace

import pytest

from netconsole.core.paths import PathResolver
from netconsole.models.device import Device
from netconsole.models.online_mr_models import (
    CONFIG_COLLECT_COMMANDS,
    INIT_COMMANDS,
    TERMINAL_MONITOR_INIT_COMMANDS,
    STATE_ABORTED,
    STATE_COLLECTING,
    STATE_CONNECTING,
    STATE_STOPPING,
    STATE_STOPPED,
    TASK_CONFIG_COLLECT,
    TASK_CHANNEL_BUSY,
    TASK_AP_RADIO_STATISTICS,
    TASK_INTERFACE_RATE,
    TASK_MESH_LINK,
    TASK_SWITCH_HISTORY,
    TASK_WIRELESS_STATUS,
    OnlineMrConnectionConfig,
    OnlineMrIntervals,
    OnlineMrSnapshot,
    repeat_command_group,
)
from netconsole.core.ping.fping_v5_runner import build_fping_v5_args
from netconsole.services.fping_legacy_parser import (
    aggregate_ping_for_active_segment,
    parse_fping_lines,
    parse_fping_summary,
)
from netconsole.services.fping_v5 import detect_fping_version, find_fping_tool
from netconsole.services.background_job import BackgroundJob
from netconsole.services.background_tasks import run_background_task
from netconsole.services.online_mr_collector import NetmikoShellConnection, RepeatSshSession
from netconsole.repositories.device_repository import DeviceRepository
from netconsole.services.network_tools.iperf_parser import parse_iperf_lines, read_iperf_text
from netconsole.services.online_mr_parser import parse_ap_radio_statistics_text, parse_channel_busy_text, parse_mesh_link_text, parse_switch_history_text
from netconsole.services.online_mr_analysis_report_exporter import OnlineMrAnalysisReportExporter
from netconsole.services.online_mr_chart_builder import OnlineMrChartBuilder
from netconsole.services.online_mr_terminal_log_parser import parse_active_link_endpoint, parse_active_link_switch_logs, switch_reason_text
from netconsole.services.online_mr.core.event_model import EVENT_BUSY_SAMPLE, EVENT_FPING_V5_SAMPLE, EVENT_MESH_SAMPLE, OnlineMrEvent
from netconsole.services.online_mr.core.realtime_model import RealtimeAggregator, build_realtime_state
from netconsole.services.online_mr.core.realtime_cache import OnlineMrRawEvent, OnlineMrRealtimeCache
from netconsole.services.online_mr.core.realtime_parser import OnlineMrRealtimeParser
from netconsole.services.online_mr.parser.event_parser_engine import EventParserEngine
from netconsole.services.online_mr.realtime.sliding_window_buffer import SlidingWindowBuffer
from netconsole.services.mesh_storage_service import MeshStorageService
from netconsole.services.online_mr_collector import (
    NORMAL_DISPLAY_PREPARE_COMMANDS,
    PROBE_STREAM_PREPARE_COMMANDS,
    STREAM_PREPARE_COMMANDS,
    OnlineMrCollectionManager,
    OnlineMrCollector,
    stream_prepare_commands,
)
from netconsole.services.online_mr_session_store import COLLECTOR_OUTPUT_RAW_FILE, DEVICE_TERMINAL_MONITOR_RAW_FILE, OnlineMrSessionStore
from netconsole.services.rail_transit.online_mr_diagnosis_parser import PARSER_VERSION, OnlineMrDiagnosisParser, TimeSyncSample, estimate_device_time_from_local


LINE_A = "[1] Active 30f5-277a-5a2f 2025/12/03 10:12:30 0d 00h 00m 03s 1 36/43 2%/4% 45%/47% 3/1 15/27 60/72060 88/105 0/5000 2/297 314/0 0/93 0/0 0/0 0/0"






class FakeConnection:
    def __init__(self, outputs: dict[str, str] | None = None, fail_on: set[str] | None = None) -> None:
        self.outputs = outputs or {}
        self.fail_on = fail_on or set()
        self.commands: list[str] = []
        self.closed = False

    def send_command(self, command: str, timeout: int) -> str:
        self.commands.append(command)
        if command in self.fail_on:
            self.closed = True
            raise RuntimeError("connection closed")
        return self.outputs.get(command, f"{command}\nOK")

    def close(self) -> None:
        self.closed = True




class Factory:
    def __init__(self, connections: list[FakeConnection]) -> None:
        self.connections = connections
        self.created: list[FakeConnection] = []

    def __call__(self, config: OnlineMrConnectionConfig) -> FakeConnection:
        connection = self.connections.pop(0)
        self.created.append(connection)
        return connection


def _config(tmp_path: Path) -> tuple[PathResolver, OnlineMrConnectionConfig]:
    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("MR-01")
    config = OnlineMrConnectionConfig(
        site="demo",
        mr_id=profile.mr_id,
        mr_name=profile.display_name,
        safe_mr_name=profile.safe_folder_name,
        device_id=1,
        device_name="FAT-AP-01",
        host="192.0.2.10",
        username="admin",
        password="secret",
        reconnect_interval=0,
    )
    return paths, config


def _collector(tmp_path: Path, connection: FakeConnection | None = None) -> tuple[OnlineMrCollector, FakeConnection]:
    paths, config = _config(tmp_path)
    connection = connection or FakeConnection({"display wlan mesh-link": LINE_A})
    collector = OnlineMrCollector(config, OnlineMrSessionStore(paths), connection_factory=lambda _: connection, sleeper=lambda _: None)
    return collector, connection








def _create_onboard_device(repository: DeviceRepository, group_id: int, name: str, device_type: str = "FAT-AP") -> Device:
    return repository.create(
        Device(
            name=name,
            group_id=group_id,
            device_type=device_type,
            ip_address=f"192.0.2.{len(name) + 10}",
            ssh_enabled=1,
            ssh_port=22,
            ssh_username="admin",
            ssh_password="secret",
        )
    )


def test_concurrency_limit_rejects_third_collector() -> None:
    manager = OnlineMrCollectionManager(max_concurrent=2)
    manager.register("s1", object())
    manager.register("s2", object())
    assert manager.running_count() == 2
    with pytest.raises(RuntimeError, match="online_mr.max_two_running"):
        manager.register("s3", object())


def test_manager_allows_session_then_device_registration_for_second_collector() -> None:
    manager = OnlineMrCollectionManager(max_concurrent=2)
    first = object()
    second = object()
    manager.register("s1", first)
    manager.register_device(1, first)
    manager.register("s2", second)
    manager.register_device(2, second)
    assert manager.running_count() == 2
    with pytest.raises(RuntimeError, match="online_mr.max_two_running"):
        manager.register_device(3, object())


def test_init_command_order_is_exact(tmp_path: Path) -> None:
    collector, connection = _collector(tmp_path)
    collector.start()
    assert connection.commands[: len(INIT_COMMANDS)] == list(INIT_COMMANDS)


def test_init_status_written_without_init_raw_log(tmp_path: Path) -> None:
    collector, _connection = _collector(tmp_path)
    collector.start()

    session_dir = collector.session.session_dir
    meta = json.loads((session_dir / "session_meta.json").read_text(encoding="utf-8"))
    assert not (session_dir / "raw" / "init_raw.log").exists()
    assert meta["init"]["status"] == "success"
    assert meta["init"]["commands"] == list(INIT_COMMANDS)
    assert meta["init"]["started_at"]
    assert meta["init"]["ended_at"]
    raw = (session_dir / "raw" / COLLECTOR_OUTPUT_RAW_FILE).read_text(encoding="utf-8")
    assert "===== INIT START" not in raw
    assert "screen-length disable -> OK" not in raw


def test_init_failure_records_meta_without_raw_echo(tmp_path: Path) -> None:
    collector, _connection = _collector(tmp_path, FakeConnection(fail_on={"terminal monitor"}))
    collector.start()

    session_dir = collector.session.session_dir
    meta = json.loads((session_dir / "session_meta.json").read_text(encoding="utf-8"))
    assert not (session_dir / "raw" / "init_raw.log").exists()
    assert meta["init"]["status"] == "failed"
    assert "terminal monitor" in meta["init"]["error_message"]


def test_stop_during_init_does_not_reenter_collecting_or_write_init_success(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)

    class StopDuringInitConnection(FakeConnection):
        def __init__(self) -> None:
            super().__init__()
            self.collector: OnlineMrCollector | None = None

        def send_command(self, command: str, timeout: int) -> str:
            result = super().send_command(command, timeout)
            if command == INIT_COMMANDS[0] and self.collector is not None:
                self.collector.request_stop()
            return result

    connection = StopDuringInitConnection()
    collector = OnlineMrCollector(config, OnlineMrSessionStore(paths), connection_factory=lambda _: connection)
    connection.collector = collector

    meta = collector.start()

    log_text = (collector.session.session_dir / "logs" / "collector.log").read_text(encoding="utf-8")
    assert meta.status == STATE_STOPPING
    assert collector.status == STATE_STOPPING
    assert "state=COLLECTING" not in log_text
    assert "init_status=success" not in log_text


def test_scheduler_intervals_with_fake_clock(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    config.intervals = OnlineMrIntervals(mesh_link=2, channel_busy=2, ap_radio_statistics=5, switch_history=300)
    connection = FakeConnection({"display wlan mesh-link": LINE_A})
    collector = OnlineMrCollector(config, OnlineMrSessionStore(paths), connection_factory=lambda _: connection, sleeper=lambda _: None)
    collector.start()
    assert set(collector.run_due_tasks(0.0)) == {"mesh_link", "channel_busy", "ap_radio_statistics", "switch_history", "interface_rate"}
    assert collector.run_due_tasks(1.0) == []
    assert set(collector.run_due_tasks(2.0)) == {"mesh_link", "channel_busy", "interface_rate"}
    assert set(collector.run_due_tasks(4.0)) == {"mesh_link", "channel_busy", "interface_rate"}
    assert collector.run_due_tasks(5.0) == ["ap_radio_statistics"]
    assert "switch_history" in collector.run_due_tasks(300.0)


def test_auto_reconnect_reruns_init_and_continues(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    first = FakeConnection(fail_on={"display wlan mesh-link"})
    second = FakeConnection({"display wlan mesh-link": LINE_A})
    factory = Factory([first, second])
    collector = OnlineMrCollector(config, OnlineMrSessionStore(paths), connection_factory=factory, sleeper=lambda _: None)
    collector.start()
    sample_id = collector.run_once(TASK_MESH_LINK)
    assert collector.status == STATE_COLLECTING
    assert collector.stats.reconnect_count == 1
    assert sample_id == -1
    assert second.commands[: len(INIT_COMMANDS)] == list(INIT_COMMANDS)
    assert any("reconnect_count=1" in line for line in (collector.session.session_dir / "raw" / "reconnect.log").read_text(encoding="utf-8").splitlines())


def test_raw_persistence_after_mesh_link(tmp_path: Path) -> None:
    collector, _ = _collector(tmp_path)
    collector.start()
    collector.run_once(TASK_MESH_LINK)
    raw_path = collector.session.session_dir / "raw" / "mesh_link_raw.log"
    raw = raw_path.read_text(encoding="utf-8")
    assert "display wlan mesh-link" in raw
    assert LINE_A in raw
    meta = json.loads((collector.session.session_dir / "session_meta.json").read_text(encoding="utf-8"))
    assert meta["stats"]["mesh_link_success"] == 1


def test_realtime_session_collects_config_inside_same_session(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    config.collect_config_on_start = True
    config_connection = FakeConnection({command: f"{command}\nOK" for command in CONFIG_COLLECT_COMMANDS})
    realtime_connection = FakeConnection()
    collector = OnlineMrCollector(config, OnlineMrSessionStore(paths), connection_factory=Factory([config_connection, realtime_connection]))

    meta = collector.start()

    config_path = collector.session.session_dir / "outputs" / "current_configuration.txt"
    root_config_dir = collector.session.session_dir.parent.parent / "config"
    saved_meta = json.loads((collector.session.session_dir / "session_meta.json").read_text(encoding="utf-8"))
    assert meta.session_id == collector.session.meta.session_id
    assert config_path.exists()
    assert root_config_dir.exists() is False
    assert config_connection.commands == list(CONFIG_COLLECT_COMMANDS)
    assert saved_meta["session_type"] == "realtime"
    assert saved_meta["config_collect_enabled"] is True
    assert saved_meta["config_collect_status"] == "success"
    assert saved_meta["config_file_path"] == str(config_path)
    assert saved_meta["raw_log_path"] == str(collector.session.session_dir / "raw" / COLLECTOR_OUTPUT_RAW_FILE)
    assert "display current-configuration" in config_path.read_text(encoding="utf-8")


def test_config_only_session_writes_config_and_meta(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    connection = FakeConnection({command: f"{command}\nOK" for command in CONFIG_COLLECT_COMMANDS})
    collector = OnlineMrCollector(config, OnlineMrSessionStore(paths), connection_factory=lambda _: connection)

    meta = collector.collect_config_only()

    config_path = collector.session.session_dir / "outputs" / "current_configuration.txt"
    saved_meta = json.loads((collector.session.session_dir / "session_meta.json").read_text(encoding="utf-8"))
    assert meta.session_type == "config_only"
    assert saved_meta["session_type"] == "config_only"
    assert saved_meta["config_collect_status"] == "success"
    assert saved_meta["config_file_path"] == str(config_path)
    assert saved_meta["raw_log_path"] == str(collector.session.session_dir / "raw" / COLLECTOR_OUTPUT_RAW_FILE)
    assert saved_meta["ended_at"]
    assert connection.commands == list(CONFIG_COLLECT_COMMANDS)


def test_config_collect_failure_keeps_session_meta(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    config.collect_config_on_start = True
    collector = OnlineMrCollector(
        config,
        OnlineMrSessionStore(paths),
        connection_factory=Factory([FakeConnection(fail_on={"display current-configuration"}), FakeConnection()]),
    )

    collector.start()

    saved_meta = json.loads((collector.session.session_dir / "session_meta.json").read_text(encoding="utf-8"))
    assert saved_meta["config_collect_status"] == "failed"
    assert saved_meta["config_error"]
    assert collector.status == STATE_COLLECTING


def _prepare_parsed_channel_busy_session(tmp_path: Path, count: int = 3):
    paths, config = _config(tmp_path)
    session = OnlineMrSessionStore(paths).create_session(config)
    parser = OnlineMrDiagnosisParser(session.session_dir)
    parser._ensure_tables()
    with sqlite3.connect(session.db_path) as conn:
        for index in range(count):
            conn.execute(
                """
                INSERT INTO channel_busy_records (
                    session_id, device_time, device_clock, time_source, radio, ctl_channel, bandwidth,
                    record_interval, row_index, ctl_busy, tx_busy, rx_busy,
                    raw_file, raw_line_start, raw_line_end
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    session.meta.session_id,
                    f"2026-07-03 19:00:0{index}.000",
                    None,
                    "device_record",
                    1,
                    None,
                    None,
                    None,
                    1,
                    7 + index,
                    4 + index,
                    3 + index,
                    "raw/channel_busy_raw.log",
                    index,
                    index + 1,
                ),
            )
        conn.execute(
            """
            INSERT INTO active_segments (
                session_id, radio, active_peer_mac, start_time, end_time, sample_count,
                avg_mr_rssi, min_mr_rssi, max_mr_rssi, event_type, details_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                session.meta.session_id,
                1,
                "1111-2222-3333",
                "2026-07-03 19:00:00.000",
                "2026-07-03 19:01:00.000",
                count,
                35,
                30,
                40,
                "NORMAL",
                "{}",
            ),
        )
        conn.execute(
            """
            INSERT INTO online_parse_metadata (
                session_id, parsed_at, parser_version, raw_fingerprint, row_counts, status, error_summary
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                session.meta.session_id,
                "2026-07-03 19:02:00.000",
                PARSER_VERSION,
                parser.raw_fingerprint(),
                json.dumps({"channel_samples": count, "active_segments": 1}),
                "OK",
                "",
            ),
        )
    return session, config


def _insert_main_link_sample(
    conn: sqlite3.Connection,
    session_id: str,
    collected_at: str,
    *,
    radio: int = 1,
    link_state: str = "ACTIVE",
    peer_mac: str = "bc5a-3457-cbef",
    peer_name: str | None = None,
    resolved_peer_name: str | None = None,
    rssi: int = -36,
    station: str = "",
    section: str = "",
    online_time: str = "00h 00m 01s",
) -> None:
    _ensure_test_new_parsed_tables(conn)
    conn.execute(
        """
        INSERT INTO main_link_samples (
            session_id, collector_time, device_time, device_clock, time_source, radio, link_state,
            peer_name, peer_mac, peer_mac_normalized, resolved_peer_name, mr_rssi,
            bssid, mesh_interface, belong_station, belong_section, belong_type,
            belonging_source, online_time, raw_file, raw_line_start, raw_line_end
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            session_id,
            collected_at,
            collected_at,
            None,
            "collector",
            radio,
            link_state,
            peer_name if peer_name is not None else peer_mac,
            peer_mac,
            peer_mac.replace("-", ""),
            resolved_peer_name if resolved_peer_name is not None else peer_name if peer_name is not None else peer_mac,
            rssi,
            "",
            "",
            station,
            section,
            "unknown",
            "",
            online_time,
            "raw/mesh_link_raw.log",
            0,
            1,
        ),
    )


def _insert_channel_busy_record(
    conn: sqlite3.Connection,
    session_id: str,
    collected_at: str,
    *,
    radio: int = 1,
    ctl_busy: int = 7,
    tx_busy: int = 4,
    rx_busy: int = 3,
) -> None:
    _ensure_test_new_parsed_tables(conn)
    conn.execute(
        """
        INSERT INTO channel_busy_records (
            session_id, device_time, time_source, radio,
            row_index, ctl_busy, tx_busy, rx_busy, raw_file, raw_line_start, raw_line_end
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (session_id, collected_at, "device_record_time", radio, 1, ctl_busy, tx_busy, rx_busy, "raw/channel_busy_raw.log", 0, 1),
    )


def _insert_fping_sample(
    conn: sqlite3.Connection,
    session_id: str,
    collected_at: str,
    *,
    target_ip: str = "127.0.0.1",
    success: int = 1,
    latency_ms: float | None = 2.5,
) -> None:
    _ensure_test_new_parsed_tables(conn)
    conn.execute(
        """
        INSERT INTO fping_samples (
            session_id, collector_time, local_time, device_aligned_time, clock_offset_ms,
            offset_source, time_source, target_ip, target_name, seq,
            success, latency_ms, loss_percent, status
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (session_id, collected_at, collected_at, collected_at, 0.0, "nearest_sample", "local_tool", target_ip, "", 1, success, latency_ms, 0 if success else 100, "OK" if success else "TIMEOUT"),
    )
    conn.execute(
        """
        INSERT INTO fping_1s_summary (
            session_id, bucket_time, local_bucket_time, device_bucket_time, clock_offset_ms,
            target_ip, target_name, sent, received, lost, loss_percent,
            avg_latency_ms, min_latency_ms, max_latency_ms, jitter_ms, status
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            session_id,
            collected_at[:19],
            collected_at[:19],
            collected_at[:19],
            0.0,
            target_ip,
            "",
            1,
            success,
            1 - success,
            0 if success else 100,
            latency_ms if success else None,
            latency_ms if success else None,
            latency_ms if success else None,
            0 if success else None,
            "OK" if success else "LOSS",
        ),
    )


def _insert_interface_rate_sample(
    conn: sqlite3.Connection,
    session_id: str,
    collected_at: str,
    *,
    direction: str = "inbound",
    interface_name: str = "GE1/0/1",
    total_pps: int = 100,
) -> None:
    _ensure_test_new_parsed_tables(conn)
    conn.execute(
        """
        INSERT INTO interface_rate_samples (
            session_id, device_time, device_clock, time_source, interface_name, interface_normalized,
            direction, total_pps, broadcast_pps, multicast_pps, usage_percent,
            raw_file, raw_line_start, raw_line_end
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (session_id, collected_at, None, "device_clock", interface_name, interface_name, direction, total_pps, 0, 0, None, "raw/interface_rate_raw.log", 0, 1),
    )


def _insert_switch_realtime_event(
    conn: sqlite3.Connection,
    session_id: str,
    collected_at: str,
    *,
    device_name: str = "MR-01",
    old_peer_name: str = "AP-A",
    old_peer_mac: str = "1111-2222-3333",
    old_rssi: int = 30,
    old_station: str = "站点A",
    old_section: str = "",
    new_peer_name: str = "AP-B",
    new_peer_mac: str = "1111-2222-4444",
    new_rssi: int = 40,
    new_station: str = "站点B",
    new_section: str = "",
    peer_quantity: int = 2,
    link_quantity: int = 1,
    reason_code: int | None = 2,
    reason_text: str = "主动切换（未开启移动链路优化）",
) -> None:
    _ensure_test_new_parsed_tables(conn)
    conn.execute(
        """
        INSERT INTO switch_realtime_events (
            session_id, device_time, time_source, device_name,
            old_peer_name, old_peer_mac, old_rssi, old_belong_station, old_belong_section,
            new_peer_name, new_peer_mac, new_rssi, new_belong_station, new_belong_section,
            peer_quantity, link_quantity, switch_reason_code, switch_reason_text,
            raw_file, raw_line_start, raw_line_end
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            session_id,
            collected_at,
            "device_event_time",
            device_name,
            old_peer_name,
            old_peer_mac,
            old_rssi,
            old_station,
            old_section,
            new_peer_name,
            new_peer_mac,
            new_rssi,
            new_station,
            new_section,
            peer_quantity,
            link_quantity,
            reason_code,
            reason_text,
            "raw/terminal_monitor_raw.log",
            0,
            1,
        ),
    )


def _ensure_test_new_parsed_tables(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS main_link_samples (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id TEXT, collector_time TEXT, device_time TEXT, device_clock TEXT, time_source TEXT,
            radio INTEGER, link_state TEXT, peer_name TEXT, peer_mac TEXT,
            peer_mac_normalized TEXT, resolved_peer_name TEXT, mr_rssi INTEGER,
            bssid TEXT, mesh_interface TEXT, belong_station TEXT, belong_section TEXT,
            belong_type TEXT, belonging_source TEXT, online_time TEXT,
            raw_file TEXT, raw_line_start INTEGER, raw_line_end INTEGER
        );
        CREATE TABLE IF NOT EXISTS channel_busy_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id TEXT, device_time TEXT, device_clock TEXT, time_source TEXT,
            radio INTEGER, ctl_channel INTEGER, bandwidth INTEGER, record_interval INTEGER,
            row_index INTEGER, ctl_busy INTEGER, tx_busy INTEGER, rx_busy INTEGER,
            raw_file TEXT, raw_line_start INTEGER, raw_line_end INTEGER
        );
        CREATE TABLE IF NOT EXISTS fping_samples (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id TEXT, collector_time TEXT, local_time TEXT, device_aligned_time TEXT,
            clock_offset_ms REAL, offset_source TEXT, time_source TEXT, target_ip TEXT,
            target_name TEXT, seq INTEGER, success INTEGER, latency_ms REAL,
            loss_percent REAL, status TEXT
        );
        CREATE TABLE IF NOT EXISTS fping_1s_summary (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id TEXT, bucket_time TEXT, local_bucket_time TEXT, device_bucket_time TEXT,
            clock_offset_ms REAL, target_ip TEXT, target_name TEXT,
            sent INTEGER, received INTEGER, lost INTEGER, loss_percent REAL,
            avg_latency_ms REAL, min_latency_ms REAL, max_latency_ms REAL,
            jitter_ms REAL, status TEXT
        );
        CREATE TABLE IF NOT EXISTS time_sync_samples (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id TEXT NOT NULL, collector_time TEXT NOT NULL, device_time TEXT NOT NULL,
            offset_ms REAL NOT NULL, source TEXT, raw_file TEXT,
            raw_line_start INTEGER, raw_line_end INTEGER
        );
        CREATE TABLE IF NOT EXISTS interface_rate_samples (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id TEXT, device_time TEXT, device_clock TEXT, time_source TEXT,
            interface_name TEXT, interface_normalized TEXT, direction TEXT, total_pps REAL,
            broadcast_pps REAL, multicast_pps REAL, usage_percent REAL, raw_file TEXT,
            raw_line_start INTEGER, raw_line_end INTEGER
        );
        CREATE TABLE IF NOT EXISTS switch_realtime_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id TEXT, device_time TEXT, time_source TEXT,
            device_name TEXT, old_peer_name TEXT, old_peer_mac TEXT, old_rssi INTEGER,
            old_belong_station TEXT, old_belong_section TEXT, new_peer_name TEXT,
            new_peer_mac TEXT, new_rssi INTEGER, new_belong_station TEXT,
            new_belong_section TEXT, peer_quantity INTEGER, link_quantity INTEGER,
            switch_reason_code INTEGER, switch_reason_text TEXT, raw_file TEXT,
            raw_line_start INTEGER, raw_line_end INTEGER
        );
        """
    )


class _ShutdownWorker:
    def __init__(self) -> None:
        self.cancelled = False
        self.stopped = False

    def cancel(self) -> None:
        self.cancelled = True

    def stop(self) -> None:
        self.stopped = True


def test_raw_log_files_split_collector_output_from_terminal_monitor(tmp_path: Path) -> None:
    collector, _ = _collector(tmp_path)
    collector.start()
    collector.run_once(TASK_MESH_LINK)

    collector_raw = collector.session.session_dir / "raw" / COLLECTOR_OUTPUT_RAW_FILE
    mesh_raw = collector.session.session_dir / "raw" / "mesh_link_raw.log"
    terminal_raw = collector.session.session_dir / "raw" / DEVICE_TERMINAL_MONITOR_RAW_FILE
    collector_text = collector_raw.read_text(encoding="utf-8")
    mesh_text = mesh_raw.read_text(encoding="utf-8")
    terminal_text = terminal_raw.read_text(encoding="utf-8")
    assert "display wlan mesh-link" not in collector_text
    assert LINE_A not in collector_text
    assert "display wlan mesh-link" in mesh_text
    assert LINE_A in mesh_text
    assert "[collector=repeat]" not in terminal_text
    assert "display current-configuration" not in terminal_text
    assert "display wlan mesh-link" not in terminal_text


def test_append_raw_does_not_mirror_to_collector_output_by_default(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    session = OnlineMrSessionStore(paths).create_session(config)

    session.append_raw(TASK_CONFIG_COLLECT, "\n".join(CONFIG_COLLECT_COMMANDS), "display current-configuration\nversion 7.1.064\nlocal-user admin")

    collector_text = (session.session_dir / "raw" / COLLECTOR_OUTPUT_RAW_FILE).read_text(encoding="utf-8")
    config_text = (session.session_dir / "raw" / "config_collect_raw.log").read_text(encoding="utf-8")
    assert "display current-configuration" in config_text
    assert "version 7.1.064" in config_text
    assert "local-user admin" in config_text
    assert "display current-configuration" not in collector_text
    assert "version 7.1.064" not in collector_text
    assert "local-user admin" not in collector_text


def test_raw_append_methods_write_separate_files(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    session = OnlineMrSessionStore(paths).create_session(config)

    session.append_collector_output_raw("[collector=repeat] display wlan mesh-link\n")
    session.append_device_terminal_monitor_raw("%Jul  3 18:23:32:224 2026 MR SHELL/5/SHELL_LOGIN: admin logged in\n")

    collector_text = (session.session_dir / "raw" / COLLECTOR_OUTPUT_RAW_FILE).read_text(encoding="utf-8")
    terminal_text = (session.session_dir / "raw" / DEVICE_TERMINAL_MONITOR_RAW_FILE).read_text(encoding="utf-8")
    assert "[collector=repeat]" in collector_text
    assert "SHELL_LOGIN" not in collector_text
    assert "SHELL_LOGIN" in terminal_text
    assert "[collector=repeat]" not in terminal_text


def test_realtime_cache_tracks_latest_snapshot_without_file_polling(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    cache = OnlineMrRealtimeCache()
    collector = OnlineMrCollector(
        config,
        OnlineMrSessionStore(paths),
        connection_factory=lambda _: FakeConnection({"display wlan mesh-link": LINE_A}),
        realtime_cache=cache,
    )
    collector.start()
    collector.run_once(TASK_MESH_LINK)

    snapshot = cache.get_latest_snapshot(1)
    assert snapshot is not None
    assert snapshot.active_peer == "30f5-277a-5a2f"
    assert cache.get_session_realtime_table(snapshot.session_id) is snapshot

    cache.close_session(snapshot.session_id)

    assert cache.get_latest_snapshot(1) is None
    assert cache.get_latest_snapshot(1, site_id="demo") is None


def test_realtime_cache_clear_device_latest_removes_old_snapshot() -> None:
    cache = OnlineMrRealtimeCache()
    snapshot = OnlineMrSnapshot(session_id="old-session", status=STATE_STOPPED, device_id=7)
    cache.register_session(site_id="demo", session_id=snapshot.session_id, device_id=7, snapshot=snapshot)

    cache.clear_device_latest(site_id="demo", device_id=7)

    assert cache.get_latest_snapshot(7, site_id="demo") is None
    assert cache.get_latest_snapshot(7) is None


def test_realtime_parser_parses_raw_cache_event_without_file_polling() -> None:
    parser = OnlineMrRealtimeParser()
    event = OnlineMrRawEvent(
        timestamp=datetime.now(),
        session_id="session-1",
        device_id=1,
        source="ssh-repeat",
        task_type=TASK_MESH_LINK,
        raw=LINE_A,
    )

    parsed = parser.parse_raw_event(event)

    assert parsed is not None
    assert parsed.module == "mesh"
    assert parsed.payload["peer_mac"] == "30f5-277a-5a2f"
    assert parsed.payload["link_state"] == "ACTIVE"




def test_repeat_stream_invokes_callback_before_archival(tmp_path: Path) -> None:
    stop_event = Event()
    raw_path = tmp_path / "raw" / "mesh_link_raw.log"
    seen: list[str] = []

    class FakeInteractiveConnection:
        def __init__(self) -> None:
            self.read_count = 0
            self.writes: list[str] = []

        def write_channel(self, text: str) -> None:
            self.writes.append(text)

        def read_channel(self) -> str:
            self.read_count += 1
            if self.read_count == 1:
                return f"{LINE_A}\n"
            stop_event.set()
            return ""

        def disconnect(self) -> None:
            pass

    connection = object.__new__(NetmikoShellConnection)
    connection.connection = FakeInteractiveConnection()
    connection._tunnel_session = None

    def callback(_stamp: datetime, line: str) -> None:
        seen.append(line)
        stop_event.set()

    connection.run_repeat_stream(
        ("display wlan mesh-link",),
        raw_path,
        stop_event,
        timeout=1,
        line_callback=callback,
    )

    assert seen == [LINE_A]
    text = raw_path.read_text(encoding="utf-8")
    assert LINE_A in text
    assert "display wlan mesh-link" in text
    assert "\x03" in "".join(connection.connection.writes)


def test_terminal_monitor_stream_uses_dedicated_init_commands() -> None:
    stop_event = Event()
    seen: list[str] = []

    class FakeInteractiveConnection:
        def __init__(self) -> None:
            self.read_count = 0
            self.writes: list[str] = []

        def write_channel(self, text: str) -> None:
            self.writes.append(text)

        def read_channel(self) -> str:
            self.read_count += 1
            if self.read_count == 1:
                return "%Jul  3 18:23:32:224 2026 MR SHELL/5/SHELL_LOGIN: admin logged in\n"
            stop_event.set()
            return ""

        def disconnect(self) -> None:
            pass

    connection = object.__new__(NetmikoShellConnection)
    connection.connection = FakeInteractiveConnection()
    connection._tunnel_session = None

    connection.run_terminal_monitor_stream(
        TERMINAL_MONITOR_INIT_COMMANDS,
        stop_event,
        timeout=1,
        line_callback=seen.append,
    )

    assert connection.connection.writes == [f"{command}\n" for command in TERMINAL_MONITOR_INIT_COMMANDS]
    assert "SHELL_LOGIN" in "".join(seen)
    assert "display wlan mesh-link" not in "".join(connection.connection.writes)


def test_repeat_stream_uses_minimal_prepare_and_does_not_write_init_status(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)

    class FakeRepeatConnection(FakeConnection):
        def __init__(self) -> None:
            super().__init__()
            self.repeat_commands: tuple[str, ...] = ()

        def run_repeat_stream(self, commands, raw_path, stop_event, timeout: int, line_callback=None) -> None:
            self.repeat_commands = tuple(commands)
            stamp = datetime.now()
            raw_path.parent.mkdir(parents=True, exist_ok=True)
            raw_path.write_text(f"{stamp.isoformat(sep=' ', timespec='milliseconds')} [collector=repeat] START commands:\n" + "\n".join(commands) + f"\n{LINE_A}\n", encoding="utf-8")
            if line_callback is not None:
                line_callback(stamp, LINE_A)
            stop_event.set()

    main_connection = FakeConnection()
    repeat_connection = FakeRepeatConnection()
    collector = OnlineMrCollector(config, OnlineMrSessionStore(paths), connection_factory=Factory([main_connection, repeat_connection]))

    collector.start()
    collector._start_repeat_thread(TASK_MESH_LINK)
    for thread in list(collector._stream_threads):
        thread.join(timeout=1)

    log_text = (collector.session.session_dir / "logs" / "collector.log").read_text(encoding="utf-8")
    raw_text = (collector.session.session_dir / "raw" / "mesh_link_raw.log").read_text(encoding="utf-8")
    assert log_text.count("init_status=success") == 1
    assert repeat_connection.commands == list(NORMAL_DISPLAY_PREPARE_COMMANDS)
    assert stream_prepare_commands(TASK_MESH_LINK) == STREAM_PREPARE_COMMANDS
    assert "terminal logging level 7" not in repeat_connection.commands
    assert "system-view" not in repeat_connection.commands
    assert "probe" not in repeat_connection.commands
    assert repeat_connection.repeat_commands == repeat_command_group(TASK_MESH_LINK, interval=config.intervals.mesh_link)
    assert LINE_A in raw_text


@pytest.mark.parametrize(
    ("task_type", "expected_command", "raw_name"),
    [
        (TASK_CHANNEL_BUSY, "display ar5drv 1 channelbusy", "channel_busy_raw.log"),
        (TASK_AP_RADIO_STATISTICS, "display ar5drv 1 statistics", "ap_radio_statistics_raw.log"),
        (TASK_WIRELESS_STATUS, "display ar5drv 1 client all rssi", "wireless_status_raw.log"),
    ],
)
def test_ar5drv_repeat_stream_enters_probe_without_rewriting_init_status(
    tmp_path: Path,
    task_type: str,
    expected_command: str,
    raw_name: str,
) -> None:
    paths, config = _config(tmp_path)

    class FakeRepeatConnection(FakeConnection):
        def __init__(self) -> None:
            super().__init__()
            self.repeat_commands: tuple[str, ...] = ()

        def run_repeat_stream(self, commands, raw_path, stop_event, timeout: int, line_callback=None) -> None:
            self.repeat_commands = tuple(commands)
            stamp = datetime.now()
            raw_path.parent.mkdir(parents=True, exist_ok=True)
            raw_path.write_text(
                f"{stamp.isoformat(sep=' ', timespec='milliseconds')} [collector=repeat] START commands:\n"
                + "\n".join(commands)
                + f"\n{expected_command}\nOK\n",
                encoding="utf-8",
            )
            if line_callback is not None:
                line_callback(stamp, expected_command)
            stop_event.set()

    main_connection = FakeConnection()
    repeat_connection = FakeRepeatConnection()
    collector = OnlineMrCollector(config, OnlineMrSessionStore(paths), connection_factory=Factory([main_connection, repeat_connection]))

    collector.start()
    collector._start_repeat_thread(task_type)
    for thread in list(collector._stream_threads):
        thread.join(timeout=1)

    log_text = (collector.session.session_dir / "logs" / "collector.log").read_text(encoding="utf-8")
    raw_text = (collector.session.session_dir / "raw" / raw_name).read_text(encoding="utf-8")
    assert log_text.count("init_status=success") == 1
    assert f"collector={task_type} prepare_status=success" in log_text
    assert repeat_connection.commands == list(PROBE_STREAM_PREPARE_COMMANDS)
    assert "terminal monitor" not in repeat_connection.commands
    assert "terminal logging level 7" not in repeat_connection.commands
    assert expected_command in repeat_connection.repeat_commands
    if task_type == TASK_WIRELESS_STATUS:
        assert repeat_connection.repeat_commands == (
            "display clock",
            "display ar5drv 1 client all rssi",
            "display ar5drv 1 client all status",
            "repeat 3 delay 3",
        )
        assert "display ar5drv 1 client all status" in raw_text
    assert expected_command in raw_text


def test_ar5drv_repeat_stream_stops_when_probe_prepare_fails(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)

    class FakeRepeatConnection(FakeConnection):
        def __init__(self) -> None:
            super().__init__(outputs={"probe": "probe\n          ^\n% Unrecognized command found at '^' position."})
            self.repeat_started = False

        def run_repeat_stream(self, commands, raw_path, stop_event, timeout: int, line_callback=None) -> None:
            self.repeat_started = True
            raise AssertionError("repeat stream must not start after probe prepare failure")

    main_connection = FakeConnection()
    repeat_connection = FakeRepeatConnection()
    collector = OnlineMrCollector(config, OnlineMrSessionStore(paths), connection_factory=Factory([main_connection, repeat_connection]))

    collector.start()
    collector._start_repeat_thread(TASK_CHANNEL_BUSY)
    for thread in list(collector._stream_threads):
        thread.join(timeout=1)

    log_text = (collector.session.session_dir / "logs" / "collector.log").read_text(encoding="utf-8")
    assert repeat_connection.commands == list(PROBE_STREAM_PREPARE_COMMANDS)
    assert repeat_connection.repeat_started is False
    assert "collector=channel_busy prepare_status=failed reason=probe_failed" in log_text
    assert log_text.count("init_status=success") == 1
    raw_text = (collector.session.session_dir / "raw" / "channel_busy_raw.log").read_text(encoding="utf-8")
    assert "display ar5drv" not in raw_text
    assert "repeat 2 delay" not in raw_text


def test_streaming_switch_history_uses_normal_display_connection_after_main_init(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    main_connection = FakeConnection()
    switch_connection = FakeConnection(
        {
            "display clock": "2026-07-06 20:27:42",
            "display wlan mesh-link switch-history": "display wlan mesh-link switch-history\nTotal 0",
        }
    )
    collector = OnlineMrCollector(config, OnlineMrSessionStore(paths), connection_factory=Factory([main_connection, switch_connection]))

    collector.start()
    collector._streaming_mode = True
    assert collector._replace_main_connection_for_stream_task(TASK_SWITCH_HISTORY) is True
    collector.run_once(TASK_SWITCH_HISTORY)

    latest_text = (collector.session.session_dir / "raw" / "switch_history_latest.log").read_text(encoding="utf-8")
    assert main_connection.closed is True
    assert switch_connection.commands[: len(NORMAL_DISPLAY_PREPARE_COMMANDS)] == list(NORMAL_DISPLAY_PREPARE_COMMANDS)
    assert "terminal monitor" not in switch_connection.commands
    assert "terminal logging level 7" not in switch_connection.commands
    assert "system-view" not in switch_connection.commands
    assert "probe" not in switch_connection.commands
    assert "display wlan mesh-link switch-history" in latest_text
    assert "SHELL/" not in latest_text


def test_stop_state_prevents_new_repeat_connection_and_start_log(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    main_connection = FakeConnection()
    unused_repeat_connection = FakeConnection()
    factory = Factory([main_connection, unused_repeat_connection])
    collector = OnlineMrCollector(config, OnlineMrSessionStore(paths), connection_factory=factory)

    collector.start()
    collector.request_stop()
    collector._start_repeat_thread(TASK_MESH_LINK)

    collector_text = (collector.session.session_dir / "raw" / COLLECTOR_OUTPUT_RAW_FILE).read_text(encoding="utf-8")
    assert factory.created == [main_connection]
    assert "[collector=repeat] START" not in collector_text


def test_sqlite_writes_live_samples_and_active_peer(tmp_path: Path) -> None:
    collector, _ = _collector(tmp_path)
    collector.start()
    collector.run_once(TASK_MESH_LINK)
    with sqlite3.connect(collector.session.db_path) as conn:
        assert conn.execute("SELECT COUNT(*) FROM live_samples").fetchone()[0] == 1
        row = conn.execute("SELECT link_state, peer_mac_raw FROM live_mesh_links").fetchone()
    assert row == ("ACTIVE", "30f5-277a-5a2f")


def test_stop_updates_meta_and_closes_connection(tmp_path: Path) -> None:
    collector, connection = _collector(tmp_path)
    collector.start()
    collector.stop()
    meta = json.loads((collector.session.session_dir / "session_meta.json").read_text(encoding="utf-8"))
    assert connection.closed is True
    assert meta["ended_at"]
    assert meta["status"] == STATE_STOPPED


def test_recovery_marks_stale_collecting_meta_aborted(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    store = OnlineMrSessionStore(paths)
    session = store.create_session(config)
    session.update_status(STATE_COLLECTING)
    changed = store.mark_stale_sessions_aborted("demo")
    assert changed
    meta = json.loads((session.session_dir / "session_meta.json").read_text(encoding="utf-8"))
    assert meta["status"] == STATE_ABORTED


def test_recovery_marks_stale_sessions_through_background_task(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    session = OnlineMrSessionStore(paths).create_session(config)
    session.update_status(STATE_COLLECTING)

    result = run_background_task(
        BackgroundJob(
            task_type="online_mr_mark_stale_sessions",
            params={"site_name": "demo", "app_root": str(paths.app_root), "data_root": str(paths.data_root)},
        )
    )

    assert result == {"changed_count": 1}
    assert json.loads((session.session_dir / "session_meta.json").read_text(encoding="utf-8"))["status"] == STATE_ABORTED






def test_collector_snapshot_before_session_is_pending_with_device_identity(tmp_path: Path) -> None:
    collector, _connection = _collector(tmp_path)
    collector.status = STATE_CONNECTING

    snapshot = collector.snapshot()

    assert snapshot.status == STATE_CONNECTING
    assert snapshot.session_id == "pending:1"
    assert snapshot.device_id == 1
    assert snapshot.device_name == "FAT-AP-01"
    assert snapshot.host == "192.0.2.10"


def test_collector_snapshot_overrides_stale_latest_status(tmp_path: Path) -> None:
    collector, _connection = _collector(tmp_path)
    collector.start()
    collector.latest_snapshot = OnlineMrSnapshot(
        collector.session.meta.session_id,
        STATE_COLLECTING,
        device_id=collector.config.device_id,
        device_name=collector.config.device_name,
        host=collector.config.host,
        active_peer="30f5-277a-5a2f",
        local_rssi=36,
    )
    collector.status = STATE_STOPPED

    snapshot = collector.snapshot()

    assert snapshot.status == STATE_STOPPED
    assert collector.latest_snapshot.status == STATE_COLLECTING
    assert snapshot.active_peer == "30f5-277a-5a2f"




























def test_parse_failure_saves_raw_marks_failed_and_loop_continues(tmp_path: Path) -> None:
    connection = FakeConnection({"display wlan mesh-link": "not a mesh table", "display ar5drv 1 channelbusy": "TxBusy: 11 RxBusy: 22"})
    collector, _ = _collector(tmp_path, connection)
    collector.start()
    collector.run_once(TASK_MESH_LINK)
    collector.run_once(TASK_CHANNEL_BUSY)
    raw = (collector.session.session_dir / "raw" / "mesh_link_raw.log").read_text(encoding="utf-8")
    assert "not a mesh table" in raw
    with sqlite3.connect(collector.session.db_path) as conn:
        statuses = [row[0] for row in conn.execute("SELECT parse_status FROM live_samples ORDER BY id")]
        busy_count = conn.execute("SELECT COUNT(*) FROM live_channel_busy").fetchone()[0]
    assert statuses == ["FAILED", "OK"]
    assert busy_count == 1


def test_run_forever_does_not_create_second_session_after_explicit_start(tmp_path: Path) -> None:
    collector, _ = _collector(tmp_path)
    meta = collector.start()
    collector.cancelled = True
    collector.run_forever()
    sessions = list((collector.session.session_dir.parent).iterdir())
    assert [path.name for path in sessions] == [meta.session_id]


def test_fping_tool_discovery_from_project_resources(tmp_path: Path) -> None:
    exe = tmp_path / "resources" / "tools" / "windows-x64" / "fping" / "fping.exe"
    exe.parent.mkdir(parents=True)
    exe.write_text("fake", encoding="utf-8")
    (exe.parent / "cygwin1.dll").write_text("fake", encoding="utf-8")
    assert find_fping_tool(PathResolver(tmp_path)) == exe.resolve()


def test_fping_v5_version_detects_json_support(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    exe = tmp_path / "fping.exe"
    exe.write_text("fake", encoding="utf-8")
    (tmp_path / "cygwin1.dll").write_text("fake", encoding="utf-8")

    def runner(*args, **kwargs):
        class Result:
            stdout = "/fping_v5/fping: Version 5.5" if "-v" in args[0] else "-J, --json output in JSON format"
            stderr = ""
            returncode = 1

        return Result()

    monkeypatch.setattr("netconsole.core.ping.fping_v5_runner.subprocess.run", runner)
    status = detect_fping_version(exe)
    assert status.found is True
    assert status.version == "5.5"
    assert status.json_supported is True


def test_fping_command_args_are_list_with_expected_parameters(tmp_path: Path) -> None:
    args = build_fping_v5_args(tmp_path / "fping.exe", "127.0.0.1", 10, 100, 1256)
    assert args == [
        str(tmp_path / "fping.exe"),
        "-J",
        "-b",
        "1256",
        "-l",
        "-p",
        "10",
        "-t",
        "100",
        "127.0.0.1",
    ]


def test_fping_success_and_failure_lines_parse_with_midnight_rollover() -> None:
    rows = parse_fping_lines(
        [
            "23:59:59.990 : Reply[6] from 10.62.90.252: bytes=64 time=4.9 ms TTL=255",
            "00:00:00.010 : Request timed out",
        ],
        datetime(2025, 12, 20, 12, 0, 0),
        default_target="10.62.90.252",
    )
    assert rows[0]["seq"] == 6
    assert rows[0]["success"] is True
    assert rows[0]["latency_ms"] == 4.9
    assert rows[0]["ttl"] == 255
    assert rows[0]["bytes"] == 64
    assert rows[1]["success"] is False
    assert rows[1]["latency_ms"] is None
    assert str(rows[1]["collected_at"]).startswith("2025-12-21")


def test_fping_summary_parse() -> None:
    summary = parse_fping_summary(
        "Packets: Sent = 97358, Received = 96573, Lost = 785 (0.806% loss)\n"
        "Minimum = 1.5 ms, Maximum = 534.4 ms, Average = 5.6 ms",
        "10.62.90.252",
    )
    assert summary["sent"] == 97358
    assert summary["received"] == 96573
    assert summary["lost"] == 785
    assert summary["loss_percent"] == 0.806
    assert summary["max_latency_ms"] == 534.4


def test_mesh_parser_normalizes_active_variants() -> None:
    line = LINE_A.replace("Active", "Active(ax)")

    records, status, error = parse_mesh_link_text(line, datetime(2025, 12, 3, 10, 12, 33))

    assert status == "OK"
    assert error == ""
    assert records[0].link_state == "ACTIVE"
    assert records[0].metrics["local_rssi_db"] == 36


def test_online_mesh_parser_accepts_peer_name_table_format() -> None:
    records, status, error = parse_mesh_link_text(
        " Peer Name              Peer MAC       RSSI BSSID          Interface         Link state       Online time\n"
        " bc5a-3457-7540         bc5a-3457-755f 51   74ad-cb9d-3321 WLAN-MeshLink694  Active(ax)       00h 36m 52s\n",
        datetime(2026, 6, 27, 3, 23, 54),
    )

    assert status == "OK"
    assert error == ""
    assert len(records) == 1
    assert records[0].link_state == "ACTIVE"
    assert records[0].peer_mac_raw == "bc5a-3457-755f"
    assert records[0].metrics["local_rssi_db"] == 51
    assert records[0].metrics["online_time"] == "00h 36m 52s"


def test_online_mesh_parser_accepts_empty_peer_name_table_format() -> None:
    records, status, error = parse_mesh_link_text(
        " Peer Name              Peer MAC       RSSI BSSID          Interface         Link state       Online time\n"
        "                        4ce9-e4f1-b880 53   5cf7-9605-960f WLAN-MeshLink25   Active(a)        00h 43m 10s\n",
        datetime(2026, 7, 7, 1, 29, 36),
    )

    assert status == "OK"
    assert error == ""
    assert len(records) == 1
    assert records[0].metrics["peer_name"] == ""
    assert records[0].peer_mac_raw == "4ce9-e4f1-b880"
    assert records[0].peer_mac_normalized == "4ce9e4f1b880"
    assert records[0].metrics["local_rssi_db"] == 53
    assert records[0].metrics["bssid"] == "5cf7-9605-960f"
    assert records[0].metrics["interface"] == "WLAN-MeshLink25"
    assert records[0].metrics["radio_mode"] == "a"
    assert records[0].metrics["online_time"] == "00h 43m 10s"
    assert records[0].link_state == "ACTIVE"


def test_online_mesh_parser_accepts_empty_peer_name_standby_online_time() -> None:
    records, status, error = parse_mesh_link_text(
        " Peer Name              Peer MAC       RSSI BSSID          Interface         Link state       Online time\n"
        "                        4ce9-e4ef-aae0 30   5cf7-9605-960f WLAN-MeshLink24   Standby(a)       00h 43m 07s\n",
        datetime(2026, 7, 7, 1, 29, 36),
    )

    assert status == "OK"
    assert error == ""
    assert len(records) == 1
    assert records[0].metrics["peer_name"] == ""
    assert records[0].peer_mac_raw == "4ce9-e4ef-aae0"
    assert records[0].link_state_raw == "Standby(a)"
    assert records[0].link_state == "STANDBY"
    assert records[0].metrics["online_time"] == "00h 43m 07s"


@pytest.mark.parametrize("peer_name", ["AP-X_3111", "AP-S_3406", "30f5-277a-0ea0", "083b-e9ec-da40"])
def test_online_mesh_parser_accepts_common_ap_names(peer_name: str) -> None:
    records, status, error = parse_mesh_link_text(
        " Peer Name              Peer MAC       RSSI BSSID          Interface         Link state       Online time\n"
        f" {peer_name:<22} 083b-e9ec-da40 39   74ad-cb9d-3321 WLAN-MeshLink694  Active(ax)       00h 36m 52s\n",
        datetime(2026, 6, 27, 3, 23, 54),
    )

    assert status == "OK"
    assert error == ""
    assert len(records) == 1
    assert records[0].metrics["peer_name"] == peer_name
    assert records[0].peer_mac_raw == "083b-e9ec-da40"


def test_channel_busy_parser_keeps_table_rows_with_structured_fields() -> None:
    rows = parse_channel_busy_text(
        "Date/Month/Year: 26/06/2026\n"
        "Ctl Channel: 165\n"
        "BandWidth: 1\n"
        "Record Interval(s):  9\n"
        "      Time(h/m/s):   CtlBusy(%) TxBusy(%)  RxBusy(%)  ExtBusy(%)\n"
        "01     22:08:24          4          1          3          -\n"
        "02     22:08:15          7          5          6          -\n"
    )

    assert len(rows) == 2
    assert rows[0]["row_index"] == 1
    assert rows[0]["sample_time"] == "2026-06-26 22:08:24"
    assert rows[0]["channel_busy_sample_time"] == "2026-06-26 22:08:24"
    assert rows[0]["channel_busy_total"] == 4
    assert rows[0]["ctl_channel"] == 165
    assert rows[0]["bandwidth"] == 1
    assert rows[0]["record_interval"] == 9
    assert rows[0]["ctl_busy"] == 4
    assert rows[0]["tx_busy"] == 1
    assert rows[0]["rx_busy"] == 3


def test_channel_busy_parser_strips_collector_prefix_and_does_not_map_bandwidth_to_busy() -> None:
    rows = parse_channel_busy_text(
        "2026-07-07 03:05:11.465 [collector=repeat] RX  Ctl Channel: 165\n"
        "2026-07-07 03:05:11.465 [collector=repeat] RX  BandWidth: 1\n"
        "2026-07-07 03:05:11.465 [collector=repeat] RX  Record Interval(s):  9\n"
        "2026-07-07 03:05:11.465 [collector=repeat] RX  CurrentTime: 03:05:13\n"
        "2026-07-07 03:05:11.465 [collector=repeat] RX        Time(h/m/s):   CtlBusy(%) TxBusy(%)  RxBusy(%)\n"
        "2026-07-07 03:05:11.465 [collector=repeat] RX  01     03:05:07         81          2         77\n"
    )

    assert rows[0]["ctl_channel"] == 165
    assert rows[0]["bandwidth"] == 1
    assert rows[0]["ctl_busy"] == 81
    assert rows[0]["tx_busy"] == 2
    assert rows[0]["rx_busy"] == 77


def test_switch_history_parser_accepts_h3c_rows() -> None:
    rows = parse_switch_history_text(
        " Peer Name              Peer MAC          Reason            In/Out RSSI Switched At    ActiveTime\n"
        " bc5a-3457-cde0         bc5a-3457-cdef(A) N/A               54/0        06-27 20:32:35 01h 07m 41s\n"
        "                        0000-0000-0000(A) Link establish    0 /0        06-27 20:32:27 00h 00m 07s\n"
        " bc5a-3457-cc60         bc5a-3457-cc7f(A) Active link fault 27/27       06-27 20:32:27 00h 00m 00s\n",
        datetime(2026, 6, 27, 21, 40, 12),
    )

    assert len(rows) == 3
    assert rows[0]["switch_time"] == "2026-06-27 20:32:27"
    assert rows[0]["to_peer_name"] == ""
    assert rows[0]["reason"] == "Link establish"
    assert rows[1]["from_peer_mac"] == "0000-0000-0000"
    assert rows[1]["to_peer_name"] == "bc5a-3457-cc60"
    assert rows[1]["reason"] == "Active link fault"
    assert rows[2]["switch_time"] == "2026-06-27 20:32:35"
    assert rows[2]["from_peer_name"] == "bc5a-3457-cc60"
    assert rows[2]["to_peer_name"] == "bc5a-3457-cde0"
    assert rows[2]["to_peer_mac"] == "bc5a-3457-cdef"
    assert rows[2]["to_peer_mac_normalized"] == "bc5a3457cdef"
    assert rows[2]["reason"] == "N/A"
    assert rows[2]["in_rssi"] == 54


def test_switch_history_parser_fills_from_peer_in_time_order() -> None:
    rows = parse_switch_history_text(
        " Peer Name              Peer MAC          Reason            In/Out RSSI Switched At    ActiveTime\n"
        " bc5a-3457-cc60         bc5a-3457-cc6f(A) N/A               41/0        07-03 19:19:27 00h 00m 15s\n"
        " bc5a-3457-cbe0         bc5a-3457-cbef(A) Better RSSI       36/33       07-03 19:12:27 00h 07m 00s\n"
        " bc5a-3457-6ba0         bc5a-3457-6baf(A) Better RSSI       35/28       07-03 19:12:12 00h 00m 14s\n"
        " bc5a-3457-cbe0         bc5a-3457-cbef(A) Better RSSI       35/27       07-03 19:11:55 00h 00m 17s\n"
        " bc5a-3457-cc60         bc5a-3457-cc6f(A) Better RSSI       41/26       07-03 19:11:48 00h 00m 07s\n",
        datetime(2026, 7, 3, 20, 0, 0),
    )

    assert rows[1]["switch_time"] == "2026-07-03 19:11:55"
    assert rows[1]["from_peer_name"] == "bc5a-3457-cc60"
    assert rows[1]["from_peer_mac"] == "bc5a-3457-cc6f"
    assert rows[1]["to_peer_name"] == "bc5a-3457-cbe0"
    assert rows[1]["to_peer_mac"] == "bc5a-3457-cbef"
    assert rows[2]["from_peer_name"] == "bc5a-3457-cbe0"
    assert rows[2]["to_peer_name"] == "bc5a-3457-6ba0"
    assert rows[-1]["from_peer_name"] == "bc5a-3457-cbe0"
    assert rows[-1]["to_peer_name"] == "bc5a-3457-cc60"


def test_active_link_switch_parser_parses_terminal_monitor_log() -> None:
    rows = parse_active_link_switch_logs(
        "%Jul  3 19:19:27:496 2026 NBL12-LC05-MR-CT WMESH/5/MESH_ACTIVELINK_SWITCH: "
        "Switch an active link from bc5a-3457-cbe0_bc5a-3457-cbef(33) to bc5a-3457-cc60_bc5a-3457-cc6f(41): "
        "peer quantity = 14, link quantity = 3, switch reason = 2.\n",
    )

    assert len(rows) == 1
    row = rows[0]
    assert row.log_time == datetime(2026, 7, 3, 19, 19, 27, 496000)
    assert row.device_name == "NBL12-LC05-MR-CT"
    assert row.from_peer_name == "bc5a-3457-cbe0"
    assert row.from_peer_mac == "bc5a-3457-cbef"
    assert row.from_peer_rssi == 33
    assert row.to_peer_name == "bc5a-3457-cc60"
    assert row.to_peer_mac == "bc5a-3457-cc6f"
    assert row.to_peer_rssi == 41
    assert row.peer_quantity == 14
    assert row.link_quantity == 3
    assert row.switch_reason_code == 2
    assert row.switch_reason_text == "主动切换（未开启移动链路优化）"


def test_active_link_switch_parser_accepts_mac_only_endpoint() -> None:
    rows = parse_active_link_switch_logs(
        "%Jul  7 01:52:36:077 2026 HZ4DCS-MR6628E-T-040661-B WMESH/5/MESH_ACTIVELINK_SWITCH: "
        "Switch an active link from _4ce9-e4f1-b880(26) to _4ce9-e4ef-aae0(36): "
        "peer quantity = 5, link quantity = 2, switch reason = 2.\n",
    )

    assert len(rows) == 1
    row = rows[0]
    assert row.log_time == datetime(2026, 7, 7, 1, 52, 36, 77000)
    assert row.from_peer_name == ""
    assert row.from_peer_mac == "4ce9-e4f1-b880"
    assert row.from_peer_rssi == 26
    assert row.to_peer_name == ""
    assert row.to_peer_mac == "4ce9-e4ef-aae0"
    assert row.to_peer_rssi == 36
    assert row.peer_quantity == 5
    assert row.link_quantity == 2
    assert row.switch_reason_code == 2


def test_active_link_switch_parser_marks_empty_link() -> None:
    endpoint = parse_active_link_endpoint("NA_0000-0000-0000(0)")

    assert endpoint.is_empty_link is True
    assert endpoint.display_peer_name == "空链路"
    assert endpoint.radio_mac_display == "-"
    assert endpoint.rssi_display == "-"

    rows = parse_active_link_switch_logs(
        "%Jul  3 19:19:27:496 2026 NBL12-LC05-MR-CT WMESH/5/MESH_ACTIVELINK_SWITCH: "
        "Switch an active link from NA_0000-0000-0000(0) to bc5a-3457-cc60_bc5a-3457-cc6f(41): "
        "peer quantity = 14, link quantity = 3, switch reason = 1.\n",
    )

    assert rows[0].from_is_empty_link is True
    assert rows[0].from_station == "-"
    assert rows[0].from_serial_number == "-"
    assert rows[0].from_resolve_rule == "empty_link"
    assert rows[0].from_radio_mac_display == "-"
    assert rows[0].switch_reason_text == "首个 Mesh 链路建立"


def test_active_link_switch_unknown_reason_text() -> None:
    assert switch_reason_text(99) == "未知原因(99)"


def test_ap_radio_statistics_parser_extracts_required_counters() -> None:
    parsed = parse_ap_radio_statistics_text(
        "[Radio Statistics]\n"
        " TxFrameAllCnt       : 3949759\n"
        " TxFrameAllBytes     : 809134274\n"
        " RxFrameAllCnt       : 4902715\n"
        " RxFrameAllBytes     : 648170255\n"
        " TxRetryFrmCnt       : 198448               0                    0                    157395\n"
        " TxErrFrmCnt         : 162931               0                    12                   64835\n"
        " TxDiscardFrmCnt     : 162099               0                    0                    64532\n"
    )

    counters = parsed["counters"]
    assert counters["TxFrameAllCnt"] == 3949759
    assert counters["RxFrameAllCnt"] == 4902715
    assert counters["TxRetryFrmCnt"] == 355843
    assert parsed["retry_count"] == 355843
    assert parsed["error_count"] == 227778
    assert parsed["discard_count"] == 226631


def test_ap_radio_statistics_parser_extracts_latest_busy_values_from_prefixed_log() -> None:
    parsed = parse_ap_radio_statistics_text(
        "2026-07-07 03:05:11.465 [collector=repeat] RX [Radio Statistics]\n"
        "2026-07-07 03:05:11.465 [collector=repeat] RX ChannelBusy: 65\n"
        "2026-07-07 03:05:11.465 [collector=repeat] RX TxBusy: 12\n"
        "2026-07-07 03:05:11.465 [collector=repeat] RX RxBusy: 34\n"
    )

    assert parsed["channel_busy_total"] == 65
    assert parsed["ctl_busy"] == 65
    assert parsed["tx_busy"] == 12
    assert parsed["rx_busy"] == 34
    assert parsed["channel_busy_sample_time"] == "2026-07-07 03:05:11"


def test_realtime_state_unifies_mesh_busy_and_ping_fields() -> None:
    now = datetime(2026, 6, 27, 10, 0, 0)
    events = [
        OnlineMrEvent(now, "s1", 7, "ssh", "mesh", EVENT_MESH_SAMPLE, {"peer_mac": "30f5-277a-5a2f", "mr_rssi": 36, "link_state": "ACTIVE"}),
        OnlineMrEvent(now, "s1", 7, "ssh", "busy", EVENT_BUSY_SAMPLE, {"ctl_busy": 4, "tx_busy": 1, "rx_busy": 3}),
        OnlineMrEvent(now, "s1", 7, "fping_v5", "fping", EVENT_FPING_V5_SAMPLE, {"loss_rate_percent": 0.5, "avg_rtt_ms": 2.5}),
    ]

    state = build_realtime_state(
        device_id=7,
        device_name="MR",
        status="COLLECTING",
        events=events,
        sample_count=3,
        resolve_peer=lambda _mac: {"peer_ap_name": "AP-01", "peer_site": "宁波站"},
    )

    assert state.peer_name == "AP-01"
    assert state.peer_site == "宁波站"
    assert state.peer_station == "宁波站"
    assert state.mr_rssi == 36
    assert state.channel_busy_total == 4
    assert state.ctl_busy == 4
    assert state.loss == 0.5
    assert state.rtt == 2.5


def test_realtime_state_preserves_latest_valid_busy_when_later_stats_are_blank() -> None:
    now = datetime(2026, 6, 27, 10, 0, 0)
    state = build_realtime_state(
        device_id=7,
        device_name="MR",
        status="COLLECTING",
        events=[
            OnlineMrEvent(now, "s1", 7, "ssh", "busy", EVENT_BUSY_SAMPLE, {"ctl_busy": 55, "tx_busy": 12, "rx_busy": 34, "sample_time": "2026-06-27 10:00:00"}),
            OnlineMrEvent(now + timedelta(seconds=1), "s1", 7, "ssh", "stats", "STATS_SAMPLE", {"counters": {}}),
        ],
    )

    assert state.channel_busy_total == 55
    assert state.tx_busy == 12
    assert state.rx_busy == 34
    assert state.channel_busy_sample_time == "2026-06-27 10:00:00"


def test_realtime_state_continues_after_unresolved_peer_mac() -> None:
    now = datetime(2026, 7, 3, 18, 0, 0)
    calls: list[str] = []

    def resolve_peer(value: str) -> dict[str, object]:
        calls.append(value)
        if value == "bc5a-3457-7540":
            return {"peer_ap_name": "bc5a-3457-7540", "peer_site": "某站", "match_rule": "fit_ap_ap_mac_exact"}
        return {"peer_ap_name": "", "peer_site": "", "match_rule": "unresolved"}

    state = build_realtime_state(
        device_id=7,
        device_name="MR",
        status="COLLECTING",
        events=[
            OnlineMrEvent(
                now,
                "s1",
                7,
                "ssh",
                "mesh",
                EVENT_MESH_SAMPLE,
                {
                    "peer_name": "bc5a-3457-7540",
                    "peer_mac": "bc5a-3457-755f",
                    "peer_mac_normalized": "bc5a3457755f",
                    "bssid": "74ad-cb9d-345f",
                    "link_state": "ACTIVE",
                },
            )
        ],
        resolve_peer=resolve_peer,
    )

    assert calls[0] == "bc5a-3457-7540"
    assert state.peer_site == "某站"
    assert state.peer_station == "某站"


def test_realtime_state_does_not_let_standby_override_active_peer() -> None:
    now = datetime(2026, 6, 27, 10, 0, 0)
    events = [
        OnlineMrEvent(now, "s1", 7, "ssh", "mesh", EVENT_MESH_SAMPLE, {"peer_name": "active-peer", "peer_mac": "30f5-277a-5a2f", "mr_rssi": 36, "link_state": "ACTIVE"}),
        OnlineMrEvent(now + timedelta(seconds=1), "s1", 7, "ssh", "mesh", EVENT_MESH_SAMPLE, {"peer_name": "standby-peer", "peer_mac": "30f5-277a-5a3f", "mr_rssi": 10, "link_state": "STANDBY"}),
    ]

    state = build_realtime_state(device_id=7, device_name="MR", status="COLLECTING", events=events)

    assert state.peer_name == "active-peer"
    assert state.peer_mac == "30f5-277a-5a2f"
    assert state.mr_rssi == 36


def test_event_parser_extracts_simple_mesh_peer_fields() -> None:
    parser = EventParserEngine()
    event = OnlineMrEvent(
        datetime(2026, 6, 27, 10, 0, 0),
        "s1",
        7,
        "ssh",
        "mesh",
        EVENT_MESH_SAMPLE,
        {},
        raw="bc5a-3457-c8a0         bc5a-3457-c8bf 35   74ad-cb9d-317f WLAN-MeshLink774  Active(ax)       00h 05m 28s",
    )

    parser.on_event(event)

    latest = parser.latest("mesh")
    assert latest is not None
    assert latest["peer_name"] == "bc5a-3457-c8a0"
    assert latest["peer_mac"] == "bc5a-3457-c8bf"
    assert latest["mr_rssi"] == 35
    assert latest["bssid"] == "74ad-cb9d-317f"
    assert latest["interface"] == "WLAN-MeshLink774"
    assert latest["link_state"] == "ACTIVE"


def test_realtime_parser_extracts_latest_channel_busy_sample_from_stream_table() -> None:
    parser = OnlineMrRealtimeParser()
    parsed = parser.parse_raw_event(
        OnlineMrEvent(
            datetime(2026, 7, 7, 3, 5, 11, 465000),
            "s1",
            7,
            "ssh_stream",
            "busy",
            EVENT_BUSY_SAMPLE,
            {},
            raw=(
                "Date/Month/Year: 07/07/2026\n"
                "CurrentTime: 03:05:13\n"
                "      Time(h/m/s):   CtlBusy(%) TxBusy(%)  RxBusy(%)  ExtBusy(%)\n"
                "01     03:05:07          81          2          77          -\n"
                "02     03:04:58          82          3          78          -\n"
            ),
        )
    )

    assert parsed is not None
    assert parsed.payload["channel_busy_total"] == 81
    assert parsed.payload["tx_busy"] == 2
    assert parsed.payload["rx_busy"] == 77
    assert parsed.payload["channel_busy_sample_time"] == "2026-07-07 03:05:07"


def test_realtime_aggregator_updates_stats_and_iperf_fields() -> None:
    now = datetime(2026, 6, 27, 10, 0, 0)
    aggregator = RealtimeAggregator(device_id=7, device_name="MR", status="COLLECTING")

    aggregator.update(OnlineMrEvent(now, "s1", 7, "ssh", "stats", "STATS_SAMPLE", {"retry_count": 12}))
    state = aggregator.update(OnlineMrEvent(now, "s1", 7, "iperf3", "iperf", "IPERF3_SAMPLE", {"throughput_mbps": 88.0, "retransmits": 2}))

    assert state.retry_count == 12
    assert state.retry == 12
    assert state.iperf_mbps == 88.0
    assert state.retrans == 2


def test_sliding_window_keeps_latest_module_after_window_trim() -> None:
    buffer = SlidingWindowBuffer(window_seconds=60)
    old_mesh = OnlineMrEvent(
        datetime(2026, 6, 27, 10, 0, 0),
        "s1",
        1,
        "ssh_stream",
        "mesh",
        EVENT_MESH_SAMPLE,
        {"peer_name": "AP-X_3111", "link_state": "ACTIVE"},
        raw="AP-X_3111 083b-e9ec-da40 39 74ad-cb9d-3321 WLAN-MeshLink694 Active(ax)",
    )
    busy = OnlineMrEvent(
        datetime(2026, 6, 27, 10, 2, 0),
        "s1",
        1,
        "ssh_stream",
        "busy",
        EVENT_BUSY_SAMPLE,
        {"tx_busy": 1, "rx_busy": 3},
    )

    buffer.add(old_mesh)
    buffer.add(busy)
    events = buffer.get_window()

    assert old_mesh in events
    assert busy in events
    state = build_realtime_state(device_id=1, device_name="MR", status=STATE_COLLECTING, events=events)
    assert state.peer_name == "AP-X_3111"
    assert state.tx_busy == 1


def test_iperf_json_parser_extracts_udp_1m_result() -> None:
    rows = parse_iperf_lines(
        [
            json.dumps(
                {
                    "intervals": [
                        {
                            "sum": {
                                "start": 0,
                                "end": 1,
                                "seconds": 1,
                                "bytes": 125000,
                                "bits_per_second": 1000000,
                                "jitter_ms": 0.12,
                                "lost_packets": 0,
                                "packets": 86,
                                "lost_percent": 0,
                            }
                        }
                    ],
                    "end": {
                        "sum": {
                            "start": 0,
                            "end": 10,
                            "seconds": 10,
                            "bytes": 1250000,
                            "bits_per_second": 1000000,
                            "jitter_ms": 0.2,
                            "lost_packets": 0,
                            "packets": 860,
                            "lost_percent": 0,
                        }
                    },
                }
            )
        ]
    )

    assert rows
    assert rows[0]["bitrate_mbps"] == 1.0
    assert rows[-1]["loss_percent"] == 0.0


def test_iperf_text_reader_accepts_powershell_utf16_json(tmp_path: Path) -> None:
    path = tmp_path / "iperf_client_raw.json"
    path.write_text('{"end":{"sum":{"bits_per_second":1000000}}}', encoding="utf-16")

    rows = parse_iperf_lines(read_iperf_text(path).splitlines())

    assert rows[-1]["bitrate_mbps"] == 1.0


def test_active_segment_ping_aggregation() -> None:
    start = datetime(2025, 12, 20, 10, 0, 0)
    samples = []
    for index in range(1000):
        samples.append(
            {
                "collected_at": (start + timedelta(milliseconds=index * 10)).isoformat(sep=" ", timespec="milliseconds"),
                "success": index >= 10,
                "latency_ms": None if index < 10 else 5.0,
            }
        )
    result = aggregate_ping_for_active_segment(samples, start, start + timedelta(seconds=10))
    assert result["ping_sent"] == 1000
    assert result["ping_lost"] == 10
    assert result["ping_loss_percent"] == 1.0
    assert result["max_consecutive_loss"] == 10


def test_repeat_command_groups_match_required_sequences() -> None:
    assert repeat_command_group(TASK_MESH_LINK, interval=1) == (
        "display clock",
        "display wlan mesh-link",
        "repeat 2 delay 1",
    )
    assert repeat_command_group(TASK_CHANNEL_BUSY, interval=9, radio_id=1) == (
        "display clock",
        "display ar5drv 1 channelbusy",
        "repeat 2 delay 9",
    )
    assert "display ar5drv 3 channelbusy" in repeat_command_group(TASK_CHANNEL_BUSY, interval=9, radio_id=3)
    assert repeat_command_group(TASK_AP_RADIO_STATISTICS, interval=10, radio_id=1)[1] == "display ar5drv 1 statistics"
    assert repeat_command_group(TASK_WIRELESS_STATUS, interval=3, radio_id=1) == (
        "display clock",
        "display ar5drv 1 client all rssi",
        "display ar5drv 1 client all status",
        "repeat 3 delay 3",
    )
    assert repeat_command_group(TASK_WIRELESS_STATUS, interval=3, radio_id=3) == (
        "display clock",
        "display ar5drv 3 client all rssi",
        "display ar5drv 3 client all status",
        "repeat 3 delay 3",
    )
    assert repeat_command_group(TASK_SWITCH_HISTORY, interval=300)[-1] == "repeat 2 delay 300"
    assert repeat_command_group(TASK_INTERFACE_RATE, interval=2) == (
        "display clock",
        "dis counters rate inbound interface",
        "dis counters rate outbound interface",
        "repeat 3 delay 2",
    )


def test_repeat_session_stop_sends_ctrl_c_and_closes() -> None:
    connection = FakeConnection()
    session = RepeatSshSession(connection, TASK_MESH_LINK, 1)
    session.stop()
    assert "\x03" in connection.commands
    assert connection.closed is True


def test_online_diagnosis_database_contains_required_tables(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    session = OnlineMrSessionStore(paths).create_session(config)
    with sqlite3.connect(session.db_path) as conn:
        tables = {row[0] for row in conn.execute("SELECT name FROM sqlite_master WHERE type='table'")}
    assert {
        "ping_samples",
        "ping_summary",
        "live_samples",
        "live_mesh_links",
        "live_channel_busy",
        "live_radio_statistics_raw_index",
        "live_switch_history_latest",
        "live_interface_rates",
        "live_terminal_events",
        "live_events",
        "collector_logs",
    }.issubset(tables)


def test_session_raw_directory_precreates_required_files(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    session = OnlineMrSessionStore(paths).create_session(config)
    raw_names = {path.name for path in (session.session_dir / "raw").iterdir()}
    assert {
        "mesh_link_raw.log",
        "channel_busy_raw.log",
        "ap_radio_statistics_raw.log",
        "switch_history_latest.log",
        "interface_rate_raw.log",
        "fping_v5_raw.log",
        "fping_v5_samples.jsonl",
    }.issubset(raw_names)
    assert "init_raw.log" not in raw_names
    assert "iperf_client_raw.log" not in raw_names






def test_default_online_mr_intervals_and_radio() -> None:
    config = OnlineMrIntervals()
    assert config.mesh_link == 1
    assert config.channel_busy == 9
    assert config.ap_radio_statistics == 10
    assert config.switch_history == 300
    assert config.interface_rate == 2
    assert config.wireless_status == 3




































def test_online_mr_active_rssi_interactive_points_fill_nearby_metrics(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    session = OnlineMrSessionStore(paths).create_session(config)
    parser = OnlineMrDiagnosisParser(session.session_dir)
    parser._ensure_tables()
    with sqlite3.connect(session.db_path) as conn:
        for index, (collected_at, state, rssi) in enumerate(
            [
                ("2026-07-03 19:00:00.000", "ACTIVE", -36),
                ("2026-07-03 19:00:01.000", "STANDBY", -80),
            ]
        ):
            _insert_main_link_sample(conn, session.meta.session_id, collected_at, link_state=state, peer_mac=f"peer-{index}", rssi=rssi)
        _insert_channel_busy_record(conn, session.meta.session_id, "2026-07-03 19:00:05.000", ctl_busy=7, tx_busy=4, rx_busy=3)
        _insert_fping_sample(conn, session.meta.session_id, "2026-07-03 19:00:01.000", latency_ms=2.5)
        _insert_interface_rate_sample(conn, session.meta.session_id, "2026-07-03 19:00:03.000", direction="inbound", interface_name="WLAN-MESH1", total_pps=100)
        _insert_interface_rate_sample(conn, session.meta.session_id, "2026-07-03 19:00:03.000", direction="outbound", interface_name="WLAN-MESH1", total_pps=80)

    points = OnlineMrChartBuilder(session.db_path).build_active_rssi_interactive_points()

    assert len(points) == 1
    point = points[0]
    assert point.peer_mac == "peer-0"
    assert point.rssi == 36.0
    assert point.ctl_busy == "7" or point.ctl_busy == 7
    assert point.tx_busy == 4
    assert point.rx_busy == 3
    assert point.ping_loss == 0
    assert point.ping_avg_latency == 2.5
    assert point.inbound_pps == 100.0
    assert point.outbound_pps == 80.0


















def test_online_mr_parse_metadata_cache_valid_and_invalidates_on_raw_change(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    session = OnlineMrSessionStore(paths).create_session(config)
    raw_path = session.session_dir / "raw" / "mesh_link_raw.log"
    raw_path.write_text(f"2026-07-03 19:00:00 >>> display clock ; display wlan mesh-link\n{LINE_A}\n", encoding="utf-8")
    parser = OnlineMrDiagnosisParser(session.session_dir)

    summary = parser.parse()
    cached = OnlineMrDiagnosisParser(session.session_dir).cached_summary_if_valid()

    assert summary.mesh_samples == 1
    assert cached is not None
    assert cached.cache_used is True
    with sqlite3.connect(session.db_path) as conn:
        row = conn.execute("SELECT parser_version, raw_fingerprint, status FROM online_parse_metadata WHERE session_id = ?", (session.meta.session_id,)).fetchone()
    assert row[0]
    assert row[1]
    assert row[2] == "OK"

    raw_path.write_text(raw_path.read_text(encoding="utf-8") + "\n# changed", encoding="utf-8")

    assert OnlineMrDiagnosisParser(session.session_dir).cached_summary_if_valid() is None


def test_online_mr_chart_builder_active_rssi_switch_empty_link_and_export(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    session = OnlineMrSessionStore(paths).create_session(config)
    parser = OnlineMrDiagnosisParser(session.session_dir)
    parser._ensure_tables()
    with sqlite3.connect(session.db_path) as conn:
        _insert_main_link_sample(conn, session.meta.session_id, "2026-07-03 19:00:00.000", peer_mac="active", rssi=-36)
        _insert_main_link_sample(conn, session.meta.session_id, "2026-07-03 19:00:00.000", link_state="STANDBY", peer_mac="standby", rssi=-80)
        _insert_switch_realtime_event(
            conn,
            session.meta.session_id,
            "2026-07-03 19:01:00.000",
            device_name=config.device_name,
            old_peer_name="AP-A",
            old_peer_mac="1111-2222-3333",
            old_rssi=32,
            old_station="站点A",
            new_peer_name="NA",
            new_peer_mac="0000-0000-0000",
            new_rssi=0,
            new_station="-",
            peer_quantity=2,
            link_quantity=1,
            reason_code=4,
            reason_text="被动切换或强制断开后切换",
        )

    builder = OnlineMrChartBuilder(session.db_path)
    rssi = builder.build_active_rssi_series()
    switch = builder.build_switch_rssi_series()

    assert rssi.series[0].points == [("2026-07-03 19:00:00.000", 36.0)]
    assert switch.series[0].points == [("2026-07-03 19:01:00.000", 32)]
    assert switch.series[1].points == []
    assert switch.tooltip_rows[0]["to_peer_name"] == "空链路"

    export_path = tmp_path / "report.xlsx"
    OnlineMrAnalysisReportExporter().export(session.session_dir, export_path)
    from openpyxl import load_workbook

    workbook = load_workbook(export_path)
    assert {
        "综合结论",
        "会话信息",
        "质量总览",
        "时间轴质量分析",
        "fping业务质量",
        "Mesh主链路质量",
        "Peer稳定性分析",
        "切换影响分析",
        "丢包关联分析",
        "异常事件清单",
        "空口繁忙度分析",
        "射频统计分析",
        "接口速率分析",
        "链路重建与连接异常",
        "原始证据片段",
        "参数配置",
        "主链路信号趋势表",
        "主链路信号趋势图",
        "主链路切换前后信号趋势表",
        "主链路切换前后信号趋势图",
        "信道繁忙度趋势表",
        "信道繁忙度趋势图",
        "Ping丢包率趋势表",
        "Ping丢包率趋势图",
        "主链路切换原因统计表",
    }.issubset(set(workbook.sheetnames))


def test_vehicle_mr_offline_report_exports_diagnostic_sheets_without_default_details(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    session = OnlineMrSessionStore(paths).create_session(config)
    parser = OnlineMrDiagnosisParser(session.session_dir)
    parser._ensure_tables()
    with sqlite3.connect(session.db_path) as conn:
        _insert_main_link_sample(conn, session.meta.session_id, "2026-07-03 19:00:00.000", peer_mac="active", rssi=36, station="站点A", section="区间A")
        _insert_main_link_sample(conn, session.meta.session_id, "2026-07-03 19:00:00.000", link_state="STANDBY", peer_mac="standby", rssi=30, station="站点A", section="区间A")
        _insert_channel_busy_record(conn, session.meta.session_id, "2026-07-03 19:00:00.000", tx_busy=10, rx_busy=12)
        _insert_fping_sample(conn, session.meta.session_id, "2026-07-03 19:00:00.000", latency_ms=2.5)
        _insert_interface_rate_sample(conn, session.meta.session_id, "2026-07-03 19:00:00.000", total_pps=128)
        _insert_switch_realtime_event(conn, session.meta.session_id, "2026-07-03 19:01:00.000", old_rssi=35, new_rssi=36)

    from netconsole.services.vehicle_mr_offline_excel_report import VehicleMrOfflineExcelReportExporter
    from openpyxl import load_workbook

    export_path = tmp_path / "vehicle_report.xlsx"
    VehicleMrOfflineExcelReportExporter().export(session.session_dir, export_path)
    workbook = load_workbook(export_path)

    expected_order = [
        "报告总览",
        "会话信息",
        "数据完整性",
        "质量评分",
        "时间轴质量概览",
        "fping业务质量",
        "Mesh主链路区段",
        "Peer质量排名",
        "切换影响分析",
        "丢包关联分析",
        "异常事件清单",
        "空口繁忙度分析",
        "射频统计分析",
        "接口速率分析",
        "链路重建与连接异常",
        "原始证据片段",
        "参数配置",
    ]
    assert workbook.sheetnames == expected_order
    assert "fping原始样本" not in workbook.sheetnames
    assert "趋势图表" not in workbook.sheetnames
    assert workbook["报告总览"]["A2"].value == "报告类型"
    assert workbook["fping业务质量"].max_row >= 2
    assert workbook["Mesh主链路区段"]["A2"].value == "未生成主链路区段数据，请先执行离线解析。"
    assert workbook["异常事件清单"]["A2"].value == "未发现异常事件。"


def test_online_mr_switch_rssi_chart_keeps_active_context_and_skips_empty_zero(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    session = OnlineMrSessionStore(paths).create_session(config)
    parser = OnlineMrDiagnosisParser(session.session_dir)
    parser._ensure_tables()
    with sqlite3.connect(session.db_path) as conn:
        for index, (collected_at, state, rssi) in enumerate(
            [
                ("2026-07-03 19:00:50.000", "ACTIVE", -35),
                ("2026-07-03 19:00:55.000", "ACTIVE", -36),
                ("2026-07-03 19:01:05.000", "ACTIVE", -42),
                ("2026-07-03 19:01:10.000", "STANDBY", -80),
            ]
        ):
            _insert_main_link_sample(conn, session.meta.session_id, collected_at, link_state=state, peer_mac=f"peer-{index}", rssi=rssi)
        _insert_switch_realtime_event(
            conn,
            session.meta.session_id,
            "2026-07-03 19:01:00.000",
            device_name=config.device_name,
            old_peer_name="AP-A",
            old_peer_mac="1111-2222-3333",
            old_rssi=34,
            old_station="站点A",
            new_peer_name="NA",
            new_peer_mac="0000-0000-0000",
            new_rssi=0,
            new_station="-",
            reason_code=4,
            reason_text="被动切换或强制断开后切换",
        )

    chart = OnlineMrChartBuilder(session.db_path).build_switch_rssi_series()

    before_points = chart.series[0].points
    after_points = chart.series[1].points
    assert before_points == [
        ("2026-07-03 19:00:50.000", 35.0),
        ("2026-07-03 19:00:55.000", 36.0),
        ("2026-07-03 19:01:00.000", 34.0),
    ]
    assert after_points == [("2026-07-03 19:01:05.000", 42.0)]
    assert all(value != 0 for _time, value in before_points + after_points)
    assert chart.tooltip_rows[0]["to_peer_name"] == "空链路"


def test_online_mr_switch_log_rssi_chart_uses_switch_event_rssi_only(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    session = OnlineMrSessionStore(paths).create_session(config)
    parser = OnlineMrDiagnosisParser(session.session_dir)
    parser._ensure_tables()
    with sqlite3.connect(session.db_path) as conn:
        _insert_main_link_sample(conn, session.meta.session_id, "2026-07-03 19:00:55.000", link_state="ACTIVE", peer_mac="active-before", rssi=-36)
        _insert_main_link_sample(conn, session.meta.session_id, "2026-07-03 19:01:05.000", link_state="ACTIVE", peer_mac="active-after", rssi=-42)
        _insert_switch_realtime_event(
            conn,
            session.meta.session_id,
            "2026-07-03 19:01:00.000",
            device_name=config.device_name,
            old_peer_name="AP-A",
            old_peer_mac="1111-2222-3333",
            old_rssi=34,
            old_station="站点A",
            new_peer_name="NA",
            new_peer_mac="0000-0000-0000",
            new_rssi=0,
            new_station="-",
            reason_code=4,
            reason_text="被动切换或强制断开后切换",
        )

    chart = OnlineMrChartBuilder(session.db_path).build_switch_log_rssi_series()

    assert chart.series[0].points == [("2026-07-03 19:01:00.000", 34.0)]
    assert chart.series[1].points == []
    assert chart.tooltip_rows[0]["to_peer_name"] == "空链路"
    assert chart.events[0].severity == "warning"


def test_online_mr_chart_builder_interface_pps_keeps_interface_series(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    session = OnlineMrSessionStore(paths).create_session(config)
    parser = OnlineMrDiagnosisParser(session.session_dir)
    parser._ensure_tables()
    with sqlite3.connect(session.db_path) as conn:
        _insert_interface_rate_sample(conn, session.meta.session_id, "2026-07-03 19:00:00.000", direction="inbound", interface_name="WLAN-MESH1", total_pps=300)
        _insert_interface_rate_sample(conn, session.meta.session_id, "2026-07-03 19:00:00.000", direction="inbound", interface_name="WLAN-MESH1", total_pps=310)
        _insert_interface_rate_sample(conn, session.meta.session_id, "2026-07-03 19:00:00.000", direction="outbound", interface_name="WLAN-MESH1", total_pps=250)
        _insert_interface_rate_sample(conn, session.meta.session_id, "2026-07-03 19:00:00.000", direction="inbound", interface_name="XGE1/0/1", total_pps=90)

    chart = OnlineMrChartBuilder(session.db_path).build_interface_rate_series()

    assert [series.name for series in chart.series] == ["WLAN-MESH1 入方向PPS", "WLAN-MESH1 出方向PPS"]
    assert chart.series[0].points == [("2026-07-03 19:00:00.000", 310.0)]
    assert chart.series[1].points == [("2026-07-03 19:00:00.000", 250.0)]
    assert all("广播PPS" not in series.name and "组播PPS" not in series.name for series in chart.series)































def test_netmiko_shell_connection_falls_back_to_tunnel_and_releases_session(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    _paths, config = _config(tmp_path)
    device = Device(
        primary_address="10.0.0.1",
        backup_address="10.0.1.1",
        ssh_enabled=1,
        ssh_username="admin",
        ssh_password="secret",
        tunnel_enabled=1,
        tunnel1_enabled=1,
        tunnel1_host="jump1",
        tunnel1_username="jump",
    )
    from netconsole.services.netmiko_connection import connection_targets

    config.connection_targets = tuple(connection_targets(device))
    calls: list[str] = []
    closed: list[bool] = []

    class FakeNetmiko:
        def send_command_timing(self, command, **_kwargs):
            return f"{command}\nOK"

        def disconnect(self):
            closed.append(True)

    def fake_connect(**params):
        calls.append(str(params["host"]))
        if params["host"] != "127.0.0.1":
            raise RuntimeError("direct failed")
        return FakeNetmiko()

    class FakeSession:
        local_host = "127.0.0.1"
        local_port = 10022

        def close(self):
            closed.append(True)

    monkeypatch.setitem(sys.modules, "netmiko", SimpleNamespace(ConnectHandler=fake_connect))
    monkeypatch.setattr("netconsole.services.online_mr_collector.TunnelManager.open_tunnel", lambda *_args: FakeSession())

    connection = NetmikoShellConnection(config)
    output = connection.send_command("display clock", 10)
    connection.close()

    assert output.endswith("OK")
    assert config.connection_method == "tunnel1"
    assert calls == ["10.0.0.1", "10.0.1.1", "127.0.0.1"]
    assert closed == [True, True]














































def test_online_mr_diagnosis_parser_rebuilds_raw_session_tables(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    session = OnlineMrSessionStore(paths).create_session(config)
    (session.session_dir / "raw" / "mesh_link_raw.log").write_text(
        f"2025-12-03 10:12:30 >>> display clock ; display wlan mesh-link\n{LINE_A}\n",
        encoding="utf-8",
    )
    (session.session_dir / "raw" / "channel_busy_raw.log").write_text(
        "2025-12-03 10:12:31 >>> display clock ; display ar5drv 1 channelbusy\nTxBusy: 11 RxBusy: 22\n"
        "2025-12-03 10:12:32 >>> display clock ; display ar5drv 1 channelbusy\nTxBusy: 11 RxBusy: 22\n",
        encoding="utf-8",
    )
    (session.session_dir / "raw" / "interface_rate_raw.log").write_text(
        "2025-12-03 10:12:33 >>> display clock ; dis counters rate inbound interface\ninterface raw\n",
        encoding="utf-8",
    )
    (session.session_dir / "raw" / "ap_radio_statistics_raw.log").write_text(
        "2025-12-03 10:12:34 >>> display clock ; display ar5drv 1 statistics\n"
        "[Radio Statistics]\n"
        " TxFrameAllCnt       : 100\n"
        " TxFrameAllBytes     : 200\n"
        " RxFrameAllCnt       : 300\n"
        " RxFrameAllBytes     : 400\n"
        " TxRetryFrmCnt       : 5 6 7 8\n"
        " TxErrFrmCnt         : 1 2 3 4\n"
        " TxDiscardFrmCnt     : 9 10 11 12\n",
        encoding="utf-8",
    )
    (session.session_dir / "raw" / "Fping.txt").write_text(
        "10:12:30.500 : Reply[6] from 10.62.90.252: bytes=64 time=4.9 ms TTL=255\n"
        "10:12:31.500 : Request timed out\n",
        encoding="utf-8",
    )
    (session.session_dir / "raw" / "iperf_client_raw.log").write_text(
        "[  5]   0.00-1.00   sec  10.5 MBytes  88.1 Mbits/sec  0   256 KBytes\n",
        encoding="utf-8",
    )

    summary = OnlineMrDiagnosisParser(session.session_dir).parse()

    assert summary.mesh_samples == 1
    assert summary.radio_stats_samples == 1
    assert summary.ping_samples == 2
    assert summary.iperf_samples == 1
    assert summary.active_segments >= 1
    with sqlite3.connect(session.db_path) as conn:
        assert conn.execute("SELECT COUNT(*) FROM main_link_samples").fetchone()[0] >= 1
        assert conn.execute("SELECT COUNT(*) FROM fping_samples").fetchone()[0] == 2
        assert conn.execute("SELECT COUNT(*) FROM iperf_intervals").fetchone()[0] == 1
        assert conn.execute("SELECT COUNT(*) FROM radio_statistics_samples").fetchone()[0] >= 1
        assert conn.execute("SELECT COUNT(*) FROM active_segments").fetchone()[0] >= 1
        assert conn.execute("SELECT COUNT(*) FROM active_segment_metrics").fetchone()[0] >= 1


def test_online_mr_diagnosis_parser_accepts_stream_rx_raw(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    session = OnlineMrSessionStore(paths).create_session(config)
    (session.session_dir / "raw" / "mesh_link_raw.log").write_text(
        "2025-12-03 10:12:30 [collector=repeat] START commands:\n"
        "display clock\n"
        "display wlan mesh-link\n"
        "repeat 2 delay 1\n"
        "2025-12-03 10:12:31.001 [collector=repeat] RX display clock\n"
        "2025-12-03 10:12:31.001 [collector=repeat] RX [MR-probe]display wlan mesh-link\n"
        f"2025-12-03 10:12:31.001 [collector=repeat] RX {LINE_A}\n"
        "2025-12-03 10:12:32.002 [collector=repeat] RX display clock\n"
        "2025-12-03 10:12:32.002 [collector=repeat] RX [MR-probe]display wlan mesh-link\n"
        f"2025-12-03 10:12:32.002 [collector=repeat] RX {LINE_A}\n",
        encoding="utf-8",
    )

    summary = OnlineMrDiagnosisParser(session.session_dir).parse()

    assert summary.mesh_samples == 2
    with sqlite3.connect(session.db_path) as conn:
        assert conn.execute("SELECT COUNT(*) FROM main_link_samples").fetchone()[0] >= 2


def _prompted_mesh_stream_block(clock_stamp: str, mesh_stamp: str, rssi: int, active_peer: str = "30f5-277a-169f") -> str:
    return (
        f"{clock_stamp} [collector=repeat] RX <NBL12-LC07-MR-CT>display clock\n"
        f"{clock_stamp} [collector=repeat] RX 20:50:18 BeiJing Mon 07/06/2026\n"
        f"{clock_stamp} [collector=repeat] RX Time Zone : BeiJing add 08:00:00\n"
        f"{mesh_stamp} [collector=repeat] RX <NBL12-LC07-MR-CT>display wlan mesh-link\n"
        f"{mesh_stamp} [collector=repeat] RX  Peer Name              Peer MAC       RSSI BSSID          Interface         Link state       Online time\n"
        f"{mesh_stamp} [collector=repeat] RX  30f5-277a-1680         {active_peer} {rssi}   74ad-cb9d-318f WLAN-MeshLink881  Active(ax)       00h 21m 36s\n"
        f"{mesh_stamp} [collector=repeat] RX  bc5a-3457-a740         bc5a-3457-a74f 22   74ad-cb9d-318f WLAN-MeshLink434  Standby(ax)      00h 00m 01s\n"
        f"{mesh_stamp} [collector=repeat] RX  bc5a-3457-9f60         bc5a-3457-9f6f 21   74ad-cb9d-318f WLAN-MeshLink429  Standby(ax)      00h 00m 09s\n"
    )


def test_online_mr_diagnosis_parser_splits_prompted_mesh_stream_into_samples(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    session = OnlineMrSessionStore(paths).create_session(config)
    (session.session_dir / "raw" / "mesh_link_raw.log").write_text(
        "2026-07-06 20:59:58 [collector=repeat] START commands:\n"
        "display clock\n"
        "display wlan mesh-link\n"
        "repeat 2 delay 1\n"
        + _prompted_mesh_stream_block("2026-07-06 20:59:59.100", "2026-07-06 20:59:59.200", 31)
        + _prompted_mesh_stream_block("2026-07-06 21:00:00.100", "2026-07-06 21:00:00.200", 32)
        + _prompted_mesh_stream_block("2026-07-06 21:00:01.100", "2026-07-06 21:00:01.200", 33),
        encoding="utf-8",
    )

    summary = OnlineMrDiagnosisParser(session.session_dir).parse()

    assert summary.mesh_samples == 3
    with sqlite3.connect(session.db_path) as conn:
        sample_rows = conn.execute(
            """
            SELECT collector_time, device_clock, 'display clock' || char(10) || 'display wlan mesh-link'
            FROM main_link_samples
            WHERE UPPER(link_state) LIKE 'ACTIVE%'
            ORDER BY collector_time
            """
        ).fetchall()
        active_count = conn.execute("SELECT COUNT(*) FROM main_link_samples WHERE UPPER(link_state) LIKE 'ACTIVE%'").fetchone()[0]
        total_count = conn.execute("SELECT COUNT(*) FROM main_link_samples").fetchone()[0]
        segment = conn.execute("SELECT active_peer_mac, start_time, end_time, sample_count, event_type, avg_mr_rssi FROM active_segments").fetchone()
        metadata = conn.execute("SELECT parser_version, row_counts FROM online_parse_metadata WHERE session_id = ?", (session.meta.session_id,)).fetchone()
    assert [row[0] for row in sample_rows] == [
        "2026-07-06 20:59:59.200",
        "2026-07-06 21:00:00.200",
        "2026-07-06 21:00:01.200",
    ]
    assert all(row[1] == "20:50:18 BeiJing Mon 07/06/2026" for row in sample_rows)
    assert all(row[2] == "display clock\ndisplay wlan mesh-link" for row in sample_rows)
    assert active_count == 3
    assert total_count == 9
    assert segment[0] == "30f5277a169f"
    assert segment[1] == "2026-07-06 20:59:59.200"
    assert segment[2] == "2026-07-06 21:00:01.200"
    assert segment[3] == 3
    assert segment[4] == "ACTIVE"
    assert segment[5] == pytest.approx(32.0)
    row_counts = json.loads(metadata[1])
    assert metadata[0] == PARSER_VERSION
    assert row_counts["main_link_sample_count"] == 9
    assert row_counts["active_link_count"] == 3
    assert row_counts["analysis_start"] == "2026-07-06 20:59:59.200"
    assert row_counts["analysis_end"] == "2026-07-06 21:00:01.200"


def test_online_mr_cached_summary_rejects_collapsed_mesh_sample_parse(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    session = OnlineMrSessionStore(paths).create_session(config)
    (session.session_dir / "raw" / "mesh_link_raw.log").write_text(
        _prompted_mesh_stream_block("2026-07-06 20:59:59.100", "2026-07-06 20:59:59.200", 31)
        + _prompted_mesh_stream_block("2026-07-06 21:00:00.100", "2026-07-06 21:00:00.200", 32),
        encoding="utf-8",
    )
    parser = OnlineMrDiagnosisParser(session.session_dir)
    parser._ensure_tables()
    with sqlite3.connect(session.db_path) as conn:
        conn.execute(
            "INSERT INTO live_samples (session_id, task_type, collected_at, command_group, raw_file, raw_offset_start, raw_offset_end, parse_status) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (session.meta.session_id, TASK_MESH_LINK, "2026-07-06 20:59:59.200", "display clock\ndisplay wlan mesh-link", "raw/mesh_link_raw.log", 0, 9999, "OK"),
        )
        sample_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        conn.execute("INSERT INTO live_mesh_links (sample_id, link_state, peer_mac_raw, peer_mac_normalized, local_rssi_db) VALUES (?, ?, ?, ?, ?)", (sample_id, "ACTIVE", "30f5-277a-169f", "30f5277a169f", 31))
        conn.execute(
            "INSERT INTO active_segments (session_id, radio, active_peer_mac, start_time, end_time, sample_count, event_type) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (session.meta.session_id, 1, "30f5277a169f,30f5277a169f,30f5277a169f,30f5277a169f,30f5277a169f", "2026-07-06 20:59:59.200", "2026-07-06 20:59:59.200", 1, "MULTI_ACTIVE"),
        )
        conn.execute(
            "INSERT INTO online_parse_metadata (session_id, parsed_at, parser_version, raw_fingerprint, row_counts, status, error_summary) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (
                session.meta.session_id,
                "2026-07-06 21:03:00.000",
                PARSER_VERSION,
                parser.raw_fingerprint(),
                json.dumps({"mesh_samples": 1, "main_link_sample_count": 1}, ensure_ascii=False),
                "OK",
                "",
            ),
        )

    assert OnlineMrDiagnosisParser(session.session_dir).cached_summary_if_valid() is None
    assert OnlineMrDiagnosisParser(session.session_dir).cache_status() == "stale"






def test_online_mr_diagnosis_parser_writes_mesh_online_time(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    session = OnlineMrSessionStore(paths).create_session(config)
    (session.session_dir / "raw" / "mesh_link_raw.log").write_text(
        "2026-07-07 01:29:30 >>> display wlan mesh-link\n"
        " Peer Name              Peer MAC       RSSI BSSID          Interface         Link state       Online time\n"
        "                        4ce9-e4ef-aae0 30   5cf7-9605-960f WLAN-MeshLink24   Standby(a)       00h 43m 07s\n"
        "                        4ce9-e4f1-b880 53   5cf7-9605-960f WLAN-MeshLink25   Active(a)        00h 43m 04s\n",
        encoding="utf-8",
    )

    summary = OnlineMrDiagnosisParser(session.session_dir).parse()

    assert summary.mesh_samples == 1
    with sqlite3.connect(session.db_path) as conn:
        rows = conn.execute(
            "SELECT link_state, peer_mac, online_time FROM main_link_samples ORDER BY link_state ASC"
        ).fetchall()
        parser_version = conn.execute("SELECT parser_version FROM online_parse_metadata WHERE session_id = ?", (session.meta.session_id,)).fetchone()[0]
    assert rows == [
        ("ACTIVE", "4ce9-e4f1-b880", "00h 43m 04s"),
        ("STANDBY", "4ce9-e4ef-aae0", "00h 43m 07s"),
    ]
    assert parser_version == PARSER_VERSION




def test_online_mr_diagnosis_parser_accepts_stream_channel_busy_table(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    session = OnlineMrSessionStore(paths).create_session(config)
    (session.session_dir / "raw" / "channel_busy_raw.log").write_text(
        "2025-12-03 10:12:30 [collector=repeat] START commands:\n"
        "display clock\n"
        "display ar5drv 1 channelbusy\n"
        "repeat 2 delay 9\n"
        "2025-12-03 10:12:31.001 [collector=repeat] RX display clock\n"
        "2025-12-03 10:12:31.001 [collector=repeat] RX 03:05:13 BeiJing Tue 07/07/2026\n"
        "2025-12-03 10:12:31.001 [collector=repeat] RX [MR-probe]display ar5drv 1 channelbusy\n"
        "2025-12-03 10:12:31.001 [collector=repeat] RX ChannelBusy information\n"
        "2025-12-03 10:12:31.001 [collector=repeat] RX        Time(h/m/s):   CtlBusy(%) TxBusy(%)  RxBusy(%)  ExtBusy(%)\n"
        "2025-12-03 10:12:31.001 [collector=repeat] RX  01     03:05:07         81          2         77          -\n"
        "2025-12-03 10:12:31.001 [collector=repeat] RX  02     03:04:58         82          3         78          -\n",
        encoding="utf-8",
    )

    summary = OnlineMrDiagnosisParser(session.session_dir).parse()

    assert summary.channel_samples == 1
    with sqlite3.connect(session.db_path) as conn:
        row = conn.execute("SELECT device_time, ctl_busy, tx_busy, rx_busy FROM channel_busy_records").fetchone()
        count = conn.execute("SELECT COUNT(*) FROM channel_busy_records").fetchone()[0]
    assert row == ("2026-07-07 03:05:07", 81, 2, 77)
    assert count == 1


def test_online_mr_diagnosis_parser_groups_fping_v5_by_target(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    session = OnlineMrSessionStore(paths).create_session(config)
    (session.session_dir / "raw" / "fping_v5_samples.jsonl").write_text(
        "\n".join(
            [
                json.dumps({"ts": "2026-06-27T04:58:18.793", "target": "10.122.6.249", "seq": 0, "ok": True, "rtt_ms": 65.6, "timeout_ms": 100, "size": 64, "error": "", "backend": "fping_v5_json", "raw_type": "resp"}),
                json.dumps({"ts": "2026-06-27T04:58:18.803", "target": "10.122.6.250", "seq": 0, "ok": False, "rtt_ms": None, "timeout_ms": 100, "size": 64, "error": "timeout", "backend": "fping_v5_json", "raw_type": "timeout"}),
                json.dumps({"ts": "2026-06-27T04:58:19.793", "target": "10.122.6.249", "seq": 1, "ok": True, "rtt_ms": 70.0, "timeout_ms": 100, "size": 64, "error": "", "backend": "fping_v5_json", "raw_type": "resp"}),
            ]
        ),
        encoding="utf-8",
    )

    summary = OnlineMrDiagnosisParser(session.session_dir).parse()

    assert summary.ping_samples == 3
    with sqlite3.connect(session.db_path) as conn:
        rows = conn.execute(
            """
            SELECT target_ip, SUM(sent), SUM(received), SUM(sent - received), AVG(loss_percent), MAX(max_latency_ms)
            FROM fping_1s_summary
            GROUP BY target_ip
            ORDER BY target_ip
            """
        ).fetchall()
    assert rows == [
        ("10.122.6.249", 2, 2, 0, 0.0, 70.0),
        ("10.122.6.250", 1, 0, 1, 100.0, None),
    ]


def test_online_mr_diagnosis_parser_falls_back_to_fping_v5_raw_log(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    session = OnlineMrSessionStore(paths).create_session(config)
    (session.session_dir / "raw" / "fping_v5_raw.log").write_text(
        "\n".join(
            [
                '2026-07-07T01:29:19.341 {"resp": {"host": "172.28.29.45", "seq": 0, "size": 64, "rtt": 1.10}}',
                '2026-07-07T01:29:19.382 {"timeout": {"host": "172.28.29.45", "seq": 1}}',
            ]
        ),
        encoding="utf-8",
    )

    summary = OnlineMrDiagnosisParser(session.session_dir).parse()

    assert summary.ping_samples == 2
    with sqlite3.connect(session.db_path) as conn:
        rows = conn.execute("SELECT collector_time, target_ip, seq, success, latency_ms FROM fping_samples ORDER BY seq").fetchall()
        summary_row = conn.execute(
            """
            SELECT target_ip, SUM(sent), SUM(received), SUM(sent - received), AVG(loss_percent), MAX(max_latency_ms)
            FROM fping_1s_summary
            GROUP BY target_ip
            """
        ).fetchone()
    assert rows == [
        ("2026-07-07 01:29:19.341", "172.28.29.45", 0, 1, 1.1),
        ("2026-07-07 01:29:19.382", "172.28.29.45", 1, 0, None),
    ]
    assert summary_row == ("172.28.29.45", 2, 1, 1, 50.0, 1.1)


def test_online_mr_estimates_device_time_from_local_offset() -> None:
    sample = TimeSyncSample(
        collector_time=datetime(2026, 7, 7, 2, 32, 58, 532000),
        device_time=datetime(2026, 7, 7, 2, 33, 0),
        offset_ms=1468.0,
        source="mesh_link_display_clock",
    )

    aligned, offset_ms, source = estimate_device_time_from_local(datetime(2026, 7, 7, 2, 32, 58, 532000), [sample])

    assert aligned == datetime(2026, 7, 7, 2, 33, 0)
    assert offset_ms == 1468.0
    assert source == "first_sample"


def test_online_mr_diagnosis_parser_aligns_fping_raw_with_mesh_clock(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    session = OnlineMrSessionStore(paths).create_session(config)
    (session.session_dir / "raw" / "mesh_link_raw.log").write_text(
        "2026-07-07 01:29:17.532 [collector=repeat] RX <MR>display clock\n"
        "2026-07-07 01:29:17.532 [collector=repeat] RX 01:29:19 BeiJing Tue 07/07/2026\n",
        encoding="utf-8",
    )
    (session.session_dir / "raw" / "fping_v5_raw.log").write_text(
        '2026-07-07T01:29:19.341 {"resp": {"host": "172.28.29.45", "seq": 0, "size": 64, "rtt": 1.10}}\n',
        encoding="utf-8",
    )

    summary = OnlineMrDiagnosisParser(session.session_dir).parse()

    assert summary.ping_samples == 1
    with sqlite3.connect(session.db_path) as conn:
        sync_row = conn.execute("SELECT collector_time, device_time, offset_ms, source FROM time_sync_samples").fetchone()
        fping_row = conn.execute(
            "SELECT local_time, device_aligned_time, clock_offset_ms, offset_source FROM fping_samples"
        ).fetchone()
    assert sync_row == ("2026-07-07 01:29:17.532", "2026-07-07 01:29:19.000", 1468.0, "mesh_link_display_clock")
    assert fping_row == ("2026-07-07 01:29:19.341", "2026-07-07 01:29:20.809", 1468.0, "last_sample")


def test_online_mr_diagnosis_parser_aligns_iperf_with_mesh_clock(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    session = OnlineMrSessionStore(paths).create_session(config)
    session.meta.started_at = datetime(2026, 7, 7, 1, 29, 19)
    session.write_meta()
    (session.session_dir / "raw" / "mesh_link_raw.log").write_text(
        "2026-07-07 01:29:17.532 [collector=repeat] RX <MR>display clock\n"
        "2026-07-07 01:29:17.532 [collector=repeat] RX 01:29:19 BeiJing Tue 07/07/2026\n",
        encoding="utf-8",
    )
    (session.session_dir / "raw" / "iperf_client_raw.log").write_text(
        "[  5]   0.00-1.00   sec  10.5 MBytes  88.1 Mbits/sec  0   256 KBytes\n",
        encoding="utf-8",
    )

    summary = OnlineMrDiagnosisParser(session.session_dir).parse()

    assert summary.iperf_samples == 1
    with sqlite3.connect(session.db_path) as conn:
        row = conn.execute(
            """
            SELECT collector_time, interval_center_time, device_aligned_time,
                   device_interval_center_time, clock_offset_ms, offset_source, time_source
            FROM iperf_intervals
            """
        ).fetchone()
    assert row is not None
    assert row[0]
    assert row[1:] == (
        "2026-07-07 01:29:19.500",
        "2026-07-07 01:29:20.968",
        "2026-07-07 01:29:20.968",
        1468.0,
        "last_sample",
        "mr_device_clock_aligned",
    )


def test_online_mr_diagnosis_parser_keeps_fping_local_time_without_offset(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    session = OnlineMrSessionStore(paths).create_session(config)
    (session.session_dir / "raw" / "fping_v5_raw.log").write_text(
        '2026-07-07T01:29:19.341 {"resp": {"host": "172.28.29.45", "seq": 0, "size": 64, "rtt": 1.10}}\n',
        encoding="utf-8",
    )

    OnlineMrDiagnosisParser(session.session_dir).parse()

    with sqlite3.connect(session.db_path) as conn:
        fping_row = conn.execute(
            "SELECT local_time, device_aligned_time, clock_offset_ms, offset_source FROM fping_samples"
        ).fetchone()
        summary_row = conn.execute("SELECT bucket_time, local_bucket_time, device_bucket_time FROM fping_1s_summary").fetchone()
    assert fping_row == ("2026-07-07 01:29:19.341", None, None, "none")
    assert summary_row == ("2026-07-07 01:29:19", "2026-07-07 01:29:19", None)


def test_online_mr_diagnosis_parser_groups_fping_by_device_bucket(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    session = OnlineMrSessionStore(paths).create_session(config)
    (session.session_dir / "raw" / "mesh_link_raw.log").write_text(
        "2026-07-07 01:29:17.532 [collector=repeat] RX <MR>display clock\n"
        "2026-07-07 01:29:17.532 [collector=repeat] RX 01:29:19 BeiJing Tue 07/07/2026\n",
        encoding="utf-8",
    )
    (session.session_dir / "raw" / "fping_v5_raw.log").write_text(
        "\n".join(
            [
                '2026-07-07T01:29:19.341 {"resp": {"host": "172.28.29.45", "seq": 0, "size": 64, "rtt": 1.10}}',
                '2026-07-07T01:29:19.382 {"resp": {"host": "172.28.29.45", "seq": 1, "size": 64, "rtt": 1.20}}',
                '2026-07-07T01:29:19.414 {"resp": {"host": "172.28.29.45", "seq": 2, "size": 64, "rtt": 1.30}}',
            ]
        ),
        encoding="utf-8",
    )

    OnlineMrDiagnosisParser(session.session_dir).parse()

    with sqlite3.connect(session.db_path) as conn:
        row = conn.execute(
            "SELECT local_bucket_time, device_bucket_time, sent, received, avg_latency_ms FROM fping_1s_summary"
        ).fetchone()
    assert row == ("2026-07-07 01:29:19", "2026-07-07 01:29:20", 3, 3, 1.2)


def test_online_mr_diagnosis_parser_finds_alternate_switch_history_filename(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    session = OnlineMrSessionStore(paths).create_session(config)
    (session.session_dir / "raw" / "mesh-link switch-history.txt").write_text(
        " Peer Name              Peer MAC          Reason            In/Out RSSI Switched At    ActiveTime\n"
        " bc5a-3457-cde0         bc5a-3457-cdef(A) N/A               54/0        06-27 20:32:35 01h 07m 41s\n",
        encoding="utf-8",
    )

    summary = OnlineMrDiagnosisParser(session.session_dir).parse()

    assert summary.switch_history_samples == 1
    with sqlite3.connect(session.db_path) as conn:
        row = conn.execute("SELECT switch_reason_text, new_peer_mac FROM switch_history_events").fetchone()
    assert row == ("N/A", "bc5a-3457-cdef")


def test_online_mr_diagnosis_parser_keeps_active_logs_terminal_only(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    session = OnlineMrSessionStore(paths).create_session(config)
    (session.session_dir / "raw" / "terminal_monitor_raw.log").write_text(
        "%Jul  3 19:19:27:496 2026 NBL12-LC05-MR-CT WMESH/5/MESH_ACTIVELINK_SWITCH: "
        "Switch an active link from bc5a-3457-cbe0_bc5a-3457-cbef(33) to bc5a-3457-cc60_bc5a-3457-cc6f(41): "
        "peer quantity = 14, link quantity = 3, switch reason = 2.\n",
        encoding="utf-8",
    )
    (session.session_dir / "raw" / "switch_history_latest.log").write_text(
        " Peer Name              Peer MAC          Reason            In/Out RSSI Switched At    ActiveTime\n"
        " bc5a-3457-cbe0         bc5a-3457-cbef(A) Better RSSI       33/27       07-03 19:19:26 00h 00m 01s\n"
        " bc5a-3457-cc60         bc5a-3457-cc6f(A) Better RSSI       41/33       07-03 19:19:28 00h 00m 01s\n",
        encoding="utf-8",
    )

    summary = OnlineMrDiagnosisParser(session.session_dir).parse()

    assert summary.active_link_switch_logs == 1
    assert summary.switch_history_samples == 2
    with sqlite3.connect(session.db_path) as conn:
        terminal_row = conn.execute(
            """
            SELECT time_source, device_time, device_name, old_peer_name, old_peer_mac, old_rssi,
                   new_peer_name, new_peer_mac, new_rssi, peer_quantity, link_quantity,
                   switch_reason_code, switch_reason_text
            FROM switch_realtime_events
            WHERE time_source = 'device_event_time'
            """
        ).fetchone()
        switch_history_event_count = conn.execute(
            """
            SELECT COUNT(*)
            FROM switch_history_events
            """
        ).fetchone()[0]
        active_sources = conn.execute("SELECT DISTINCT time_source FROM switch_realtime_events").fetchall()
    assert terminal_row == (
        "device_event_time",
        "2026-07-03 19:19:27.496",
        config.device_name,
        "bc5a-3457-cbe0",
        "bc5a-3457-cbef",
        33,
        "bc5a-3457-cc60",
        "bc5a-3457-cc6f",
        41,
        14,
        3,
        2,
        "主动切换（未开启移动链路优化）",
    )
    assert switch_history_event_count == 2
    assert active_sources == [("device_event_time",)]


def test_online_mr_diagnosis_parser_writes_empty_link_without_resolver(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    session = OnlineMrSessionStore(paths).create_session(config)
    (session.session_dir / "raw" / "terminal_monitor_raw.log").write_text(
        "%Jul  3 19:19:27:496 2026 NBL12-LC05-MR-CT WMESH/5/MESH_ACTIVELINK_SWITCH: "
        "Switch an active link from NA_0000-0000-0000(0) to bc5a-3457-cc60_bc5a-3457-cc6f(41): "
        "peer quantity = 14, link quantity = 3, switch reason = 1.\n",
        encoding="utf-8",
    )

    OnlineMrDiagnosisParser(session.session_dir).parse()

    with sqlite3.connect(session.db_path) as conn:
        row = conn.execute(
            """
            SELECT old_peer_name, old_peer_mac, old_rssi, old_belong_station
            FROM switch_realtime_events
            WHERE time_source = 'device_event_time'
            """
        ).fetchone()
    assert row == ("空链路", None, None, "-")


def test_online_mr_switch_history_never_populates_active_link_switch_logs(tmp_path: Path) -> None:
    paths, config = _config(tmp_path)
    session = OnlineMrSessionStore(paths).create_session(config)
    (session.session_dir / "raw" / "switch_history_latest.log").write_text(
        " Peer Name              Peer MAC          Reason            In/Out RSSI Switched At    ActiveTime\n"
        "                        0000-0000-0000(A) Link establish    0 /0        07-03 19:19:26 00h 00m 01s\n",
        encoding="utf-8",
    )

    OnlineMrDiagnosisParser(session.session_dir).parse()

    with sqlite3.connect(session.db_path) as conn:
        active_count = conn.execute("SELECT COUNT(*) FROM switch_realtime_events").fetchone()[0]
        switch_history_event_count = conn.execute("SELECT COUNT(*) FROM switch_history_events").fetchone()[0]
    builder = OnlineMrChartBuilder(session.db_path)
    chart = builder.build_switch_rssi_series()
    log_chart = builder.build_switch_log_rssi_series()

    assert active_count == 0
    assert switch_history_event_count == 1
    assert chart.series[1].points == []
    assert log_chart.series[0].points == []
    assert log_chart.series[1].points == []


def test_online_mr_diagnosis_parser_skips_locked_fping_file(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    paths, config = _config(tmp_path)
    session = OnlineMrSessionStore(paths).create_session(config)
    (session.session_dir / "raw" / "mesh_link_raw.log").write_text(
        f"2025-12-03 10:12:30 >>> display clock ; display wlan mesh-link\n{LINE_A}\n",
        encoding="utf-8",
    )
    (session.session_dir / "raw" / "Fping.txt").write_text("locked", encoding="utf-8")

    import netconsole.services.rail_transit.online_mr_diagnosis_parser as parser_module

    original_reader = parser_module.read_text_with_retry

    def fake_reader(path: Path, *args, **kwargs):
        if path.name == "Fping.txt":
            raise PermissionError("locked")
        return original_reader(path, *args, **kwargs)

    monkeypatch.setattr(parser_module, "read_text_with_retry", fake_reader)

    summary = OnlineMrDiagnosisParser(session.session_dir).parse()

    assert summary.mesh_samples == 1
    assert summary.ping_samples == 0
    assert summary.issues >= 1
