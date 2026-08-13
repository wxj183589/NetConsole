from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from tests.support.ac_management_web_fixture import build_ac_management_fixture
from netconsole.core.database import Database
from netconsole.repositories.ac_repository import AcRepository
from netconsole.services.ac.fit_ap_resource_export import (
    export_fit_ap_resource_xlsx,
    make_fit_ap_resource_filename,
)
from netconsole.services.ac.fit_ap_resource_identity import coalesce_fit_ap_resource_rows
from netconsole.services.export.export_handlers import run_generic_export_handler
from netconsole.services.export.export_task_builders import fit_ap_resource_xlsx_spec


def _export(
    tmp_path: Path,
    *,
    scope: str,
    selected_ap_ids: list[str] | None = None,
    filters: dict[str, str] | None = None,
) -> tuple[Path, dict[str, int]]:
    paths, _db_path, _files = build_ac_management_fixture(tmp_path)
    output = tmp_path / f"{scope}.xlsx"
    result = export_fit_ap_resource_xlsx(
        output,
        {
            "app_root": str(paths.app_root),
            "data_root": str(paths.data_root),
            "site_name": "demo",
            "ac_uuid": "ac-1",
            "scope": scope,
            "selected_ap_ids": selected_ap_ids or [],
            "filters": filters or {},
            "requested_at": "2026-07-29T00:42:15+08:00",
        },
    )
    return output, result


def test_fit_ap_resource_export_keeps_one_ap_per_row_and_radio_details(tmp_path: Path) -> None:
    output, result = _export(tmp_path, scope="all")
    workbook = load_workbook(output, data_only=True)

    assert result == {"row_count": 3, "ap_count": 3, "radio_count": 6, "warning_count": 3}
    assert workbook.sheetnames == ["AP资源清单", "Radio明细", "导出说明"]
    ap_sheet = workbook["AP资源清单"]
    radio_sheet = workbook["Radio明细"]
    assert ap_sheet.max_row == 4
    assert radio_sheet.max_row == 7
    assert ap_sheet.freeze_panes == "A2"
    assert radio_sheet.freeze_panes == "A2"
    assert ap_sheet.auto_filter.ref
    assert not ap_sheet.merged_cells.ranges

    headers = {cell.value: cell.column for cell in ap_sheet[1]}
    header_values = [cell.value for cell in ap_sheet[1]]
    assert "项目名称" not in header_values
    assert "线路名称" not in header_values
    assert header_values.index("AP序列号") == header_values.index("AP型号") + 1
    assert header_values.index("AP序列号") + 1 == header_values.index("AP状态")
    assert header_values.index("AP硬件版本") > header_values.index("AP序列号")
    assert header_values.index("AP软件版本") > header_values.index("AP硬件版本")
    assert header_values.index("AP Boot版本") > header_values.index("AP软件版本")
    assert header_values.index("详细信息更新时间") > header_values.index("AP Boot版本")
    ap_names = [ap_sheet.cell(row, headers["AP名称"]).value for row in range(2, ap_sheet.max_row + 1)]
    assert len(ap_names) == len(set(ap_names)) == 3
    topology_rows = [
        tuple(str(ap_sheet.cell(row, headers[name]).value or "") for name in ("连接交换机", "连接端口", "AP名称"))
        for row in range(2, ap_sheet.max_row + 1)
    ]
    assert [row[1] for row in topology_rows] == [
        "GigabitEthernet1/0/1",
        "GigabitEthernet1/0/2",
        "GigabitEthernet1/0/3",
    ]
    assert ap_sheet.cell(2, headers["AP IP"]).number_format == "@"
    assert ap_sheet.cell(2, headers["AP MAC"]).number_format == "@"
    assert ap_sheet.cell(2, headers["AC软件版本"]).number_format == "@"
    assert ap_sheet.cell(2, headers["AP序列号"]).number_format == "@"
    assert 14 <= ap_sheet.column_dimensions[get_column_letter(headers["AP序列号"])].width <= 32
    assert 18 <= ap_sheet.column_dimensions[get_column_letter(headers["Radio更新时间"])].width <= 26
    radio_headers = {cell.value: cell.column for cell in radio_sheet[1]}
    assert 18 <= radio_sheet.column_dimensions[get_column_letter(radio_headers["Radio MAC"])].width <= 22
    assert ap_sheet.cell(2, headers["数据完整性"]).alignment.wrap_text is True
    assert ap_sheet.cell(2, headers["备注"]).alignment.wrap_text is True
    instruction_sheet = workbook["导出说明"]
    instruction_fields = [instruction_sheet.cell(row, 1).value for row in range(2, instruction_sheet.max_row + 1)]
    assert "项目名称" not in instruction_fields
    assert "线路名称" not in instruction_fields
    assert instruction_sheet.cell(2, 2).alignment.wrap_text is True
    unauth_row = next(row for row in range(2, ap_sheet.max_row + 1) if ap_sheet.cell(row, headers["AP名称"]).value == "AP-Unauth")
    assert "缺少光衰" in str(ap_sheet.cell(unauth_row, headers["数据完整性"]).value)
    assert "未采集AP详细信息" in str(ap_sheet.cell(unauth_row, headers["数据完整性"]).value)


def test_fit_ap_resource_export_scopes_are_not_limited_by_page(tmp_path: Path) -> None:
    filtered, filtered_result = _export(tmp_path / "filtered", scope="filtered", filters={"status": "offline"})
    selected, selected_result = _export(
        tmp_path / "selected",
        scope="selected",
        selected_ap_ids=["ap-online", "ap-offline"],
    )

    filtered_book = load_workbook(filtered, read_only=True, data_only=True)
    selected_book = load_workbook(selected, read_only=True, data_only=True)
    assert filtered_result["ap_count"] == 1
    assert filtered_book["AP资源清单"].max_row == 2
    assert selected_result["ap_count"] == 2
    assert selected_book["AP资源清单"].max_row == 3


def test_fit_ap_resource_export_rejects_foreign_selection_and_empty_scope(tmp_path: Path) -> None:
    paths, _db_path, _files = build_ac_management_fixture(tmp_path)
    payload = {
        "app_root": str(paths.app_root),
        "data_root": str(paths.data_root),
        "site_name": "demo",
        "ac_uuid": "ac-1",
        "scope": "selected",
        "filters": {},
    }
    with pytest.raises(ValueError, match="不属于当前 AC"):
        export_fit_ap_resource_xlsx(
            tmp_path / "foreign.xlsx",
            {**payload, "selected_ap_ids": ["ap-online", "foreign-ap"]},
        )
    with pytest.raises(ValueError, match="没有可导出"):
        export_fit_ap_resource_xlsx(
            tmp_path / "empty.xlsx",
            {**payload, "selected_ap_ids": [], "filters": {"query": "不存在的AP"}},
        )


def test_fit_ap_resource_export_deduplicates_by_normalized_mac(tmp_path: Path) -> None:
    paths, db_path, _files = build_ac_management_fixture(tmp_path)
    with Database(db_path).connect() as conn:
        conn.execute(
            """
            INSERT INTO ac_fit_ap_resources (
                ac_device_uuid, ap_uuid, ap_name, ap_ip, ap_mac, model,
                state, collected_at, updated_at
            ) VALUES (
                'ac-1', 'ap-duplicate', 'AP-Duplicate', '10.0.1.99',
                '000000000001', 'WA-Test', 'R/M',
                '2026-07-29T00:42:15+08:00', '2026-07-29T00:42:15+08:00'
            )
            """
        )
        conn.commit()
    output = tmp_path / "deduplicated.xlsx"
    result = export_fit_ap_resource_xlsx(
        output,
        {
            "app_root": str(paths.app_root),
            "data_root": str(paths.data_root),
            "site_name": "demo",
            "ac_uuid": "ac-1",
            "scope": "all",
            "filters": {},
            "selected_ap_ids": [],
        },
    )
    workbook = load_workbook(output, read_only=True, data_only=True)
    assert result["ap_count"] == 3
    assert workbook["AP资源清单"].max_row == 4


def test_fit_ap_resource_export_coalesces_duplicate_unauthenticated_row_by_serial(
    tmp_path: Path,
) -> None:
    paths, db_path, _files = build_ac_management_fixture(tmp_path)
    with Database(db_path).connect() as conn:
        conn.execute(
            """
            INSERT INTO ac_fit_ap_unauthenticated (
                ac_device_uuid, ap_name, apid, state, state_display, model,
                serial_number, collected_at, updated_at
            ) VALUES (
                'ac-1', 'AP-Online', '1', 'R/M', '运行(主)', 'WA-Test',
                'SECRET-SN-1', '2026-07-29T00:42:15+08:00',
                '2026-07-29T00:42:15+08:00'
            )
            """
        )
        conn.commit()

    output = tmp_path / "serial-duplicate.xlsx"
    result = export_fit_ap_resource_xlsx(
        output,
        {
            "app_root": str(paths.app_root),
            "data_root": str(paths.data_root),
            "site_name": "demo",
            "ac_uuid": "ac-1",
            "scope": "all",
            "filters": {},
        },
    )
    workbook = load_workbook(output, data_only=True)
    sheet = workbook["AP资源清单"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    ap_rows = [
        {
            "name": sheet.cell(row, headers["AP名称"]).value,
            "mac": sheet.cell(row, headers["AP MAC"]).value,
            "status": sheet.cell(row, headers["在线状态"]).value,
        }
        for row in range(2, sheet.max_row + 1)
    ]

    assert result["ap_count"] == 3
    assert [row for row in ap_rows if row["name"] == "AP-Online"] == [
        {"name": "AP-Online", "mac": "0000-0000-0001", "status": "未认证"}
    ]


def test_fit_ap_resource_coalescing_preserves_distinct_or_ambiguous_same_names() -> None:
    distinct = coalesce_fit_ap_resource_rows(
        [
            {"ac_device_uuid": "ac-1", "ap_name": "same", "ap_mac": "0011-2233-4455"},
            {"ac_device_uuid": "ac-1", "ap_name": "same", "ap_mac": "0011-2233-5566"},
        ]
    )
    ambiguous = coalesce_fit_ap_resource_rows(
        [
            {"ac_device_uuid": "ac-1", "ap_name": "same"},
            {"ac_device_uuid": "ac-1", "ap_name": "same"},
        ]
    )

    assert len(distinct) == 2
    assert len(ambiguous) == 2


def test_fit_ap_resource_export_keeps_serial_text_and_reports_missing(tmp_path: Path) -> None:
    paths, db_path, _files = build_ac_management_fixture(tmp_path)
    with Database(db_path).connect() as conn:
        conn.execute("UPDATE ac_fit_ap_resources SET serial_number = ? WHERE ap_uuid = 'ap-online'", ("SN-A1B2",))
        conn.execute("UPDATE ac_fit_ap_resources SET serial_number = ? WHERE ap_uuid = 'ap-offline'", ("001234567890",))
        conn.execute("DELETE FROM ac_fit_ap_resources WHERE ap_uuid = 'ap-unauth'")
        conn.execute("UPDATE ac_fit_ap_unauthenticated SET serial_number = '' WHERE ap_name = 'AP-Unauth'")
        conn.commit()
    output = tmp_path / "serials.xlsx"
    result = export_fit_ap_resource_xlsx(
        output,
        {
            "app_root": str(paths.app_root),
            "data_root": str(paths.data_root),
            "site_name": "demo",
            "ac_uuid": "ac-1",
            "scope": "all",
            "filters": {},
        },
    )
    workbook = load_workbook(output, data_only=True)
    sheet = workbook["AP资源清单"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    values = {
        sheet.cell(row, headers["AP名称"]).value: sheet.cell(row, headers["AP序列号"])
        for row in range(2, sheet.max_row + 1)
    }
    assert result["ap_count"] == 3
    assert values["AP-Online"].value == "SN-A1B2"
    assert values["AP-Offline"].value == "001234567890"
    assert values["AP-Offline"].data_type == "s"
    assert values["AP-Offline"].number_format == "@"
    unauth_row = next(row for row in range(2, sheet.max_row + 1) if sheet.cell(row, headers["AP名称"]).value == "AP-Unauth")
    assert sheet.cell(unauth_row, headers["AP序列号"]).value in (None, "")
    assert "缺少AP序列号" in str(sheet.cell(unauth_row, headers["数据完整性"]).value)


def test_fit_ap_resource_export_uses_topology_natural_sort_and_resequences(tmp_path: Path) -> None:
    paths, db_path, _files = build_ac_management_fixture(tmp_path)
    repository = AcRepository(Database(db_path))
    for port in (20, 6, 18, 1):
        repository.upsert_fit_ap_resource(
            "ac-1",
            {
                "ap_uuid": f"sort-{port}",
                "ap_name": f"SORT-{port}",
                "ap_ip": f"10.0.2.{port}",
                "ap_mac": f"0000-0002-{port:04d}",
                "model": "WA-Sort",
                "serial_number": f"SORT-SN-{port}",
                "state": "R/M",
                "state_display": "运行(主)",
                "rid1_status": "Up",
                "rid1_mode": "802.11n",
                "rid1_band": "5GHz",
                "rid1_channel": "1",
                "rid1_clients": 1,
                "lldp_neighbor_name": "SW-A",
                "lldp_neighbor_interface": f"GigabitEthernet2/0/{port}",
                "lldp_match_status": "matched",
            },
        )
    repository.upsert_fit_ap_resource(
        "ac-1",
        {
            "ap_uuid": "sort-missing-port",
            "ap_name": "SORT-MISSING-PORT",
            "ap_mac": "0000-0002-0090",
            "model": "WA-Sort",
            "serial_number": "SORT-SN-PORT",
            "state": "R/M",
            "state_display": "运行(主)",
            "lldp_neighbor_name": "SW-A",
            "lldp_match_status": "matched",
        },
    )
    repository.upsert_fit_ap_resource(
        "ac-1",
        {
            "ap_uuid": "sort-missing-switch",
            "ap_name": "SORT-MISSING-SWITCH",
            "ap_mac": "0000-0002-0091",
            "model": "WA-Sort",
            "serial_number": "SORT-SN-SWITCH",
            "state": "R/M",
            "state_display": "运行(主)",
        },
    )
    output = tmp_path / "topology.xlsx"
    export_fit_ap_resource_xlsx(
        output,
        {
            "app_root": str(paths.app_root),
            "data_root": str(paths.data_root),
            "site_name": "demo",
            "ac_uuid": "ac-1",
            "scope": "filtered",
            "filters": {"query": "SORT-"},
        },
    )
    workbook = load_workbook(output, data_only=True)
    sheet = workbook["AP资源清单"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    rows = [
        (
            sheet.cell(row, headers["连接交换机"]).value or "",
            sheet.cell(row, headers["连接端口"]).value or "",
            sheet.cell(row, headers["AP名称"]).value,
        )
        for row in range(2, sheet.max_row + 1)
    ]
    assert [row[1] for row in rows[:4]] == [
        "GigabitEthernet2/0/1",
        "GigabitEthernet2/0/6",
        "GigabitEthernet2/0/18",
        "GigabitEthernet2/0/20",
    ]
    assert rows[4][2] == "SORT-MISSING-PORT"
    assert rows[5][2] == "SORT-MISSING-SWITCH"
    assert [sheet.cell(row, headers["序号"]).value for row in range(2, sheet.max_row + 1)] == list(range(1, 7))
    radio_sheet = workbook["Radio明细"]
    radio_seq = [radio_sheet.cell(row, 1).value for row in range(2, radio_sheet.max_row + 1)]
    assert radio_seq == list(range(1, len(radio_seq) + 1))


def test_fit_ap_resource_export_metadata_uses_schema_v2_and_new_columns(tmp_path: Path) -> None:
    paths, db_path, _files = build_ac_management_fixture(tmp_path)
    output = tmp_path / "metadata.xlsx"
    spec = fit_ap_resource_xlsx_spec(
        output,
        db_path=db_path,
        site_name="demo",
        ac_uuid="ac-1",
        scope="all",
        app_root=paths.app_root,
        data_root=paths.data_root,
    )
    job = spec.to_job("fit-ap-metadata").with_runtime_paths(tmp_path=str(tmp_path / "metadata.tmp"), cancel_path=str(tmp_path / "cancel"))
    run_generic_export_handler(job)
    workbook = load_workbook(output, data_only=True)
    assert workbook["_netconsole_meta"].sheet_state == "hidden"
    metadata = json.loads(workbook["_netconsole_meta"]["B1"].value)
    assert metadata["schema_version"] == 2
    ap_columns = metadata["required_columns"]["AP资源清单"]
    instruction_columns = metadata["required_columns"]["导出说明"]
    assert "项目名称" not in ap_columns
    assert "线路名称" not in ap_columns
    assert "AP序列号" in ap_columns
    assert "项目名称" not in instruction_columns
    assert "线路名称" not in instruction_columns


def test_fit_ap_resource_filename_is_windows_safe_and_keeps_chinese() -> None:
    name = make_fit_ap_resource_filename("宁波:10号线", "251/无线*控制器")
    assert name.startswith("FIT-AP资源_宁波_10号线_251_无线_控制器_")
    assert name.endswith(".xlsx")
    assert not any(value in name for value in '<>:"/\\|?*')
