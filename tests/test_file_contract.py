from __future__ import annotations

import csv
import json
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from netconsole.services.file_contract import (
    CSV_META_MARKER,
    META_SHEET,
    ZIP_MANIFEST,
    ImportValidationError,
    attach_export_metadata,
    validate_csv_import,
    validate_excel_import,
    validate_json_import,
    validate_zip_import,
)


def test_excel_contract_accepts_matching_module_and_rejects_other_module(tmp_path: Path) -> None:
    path = tmp_path / "plan.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "轨旁AP规划"
    sheet.append(["车站名称", "AP数量"])
    sheet.append(["车站A", 2])
    workbook.save(path)

    attach_export_metadata(path, effective_suffix=".xlsx", export_type="table_xlsx", payload={"source_module": "ac.trackside_ap_plan"})

    result = validate_excel_import(
        path,
        expected_module="ac.trackside_ap_plan",
        required_sheets=["轨旁AP规划"],
        required_headers={"轨旁AP规划": ["车站名称", "AP数量"]},
        allow_legacy=False,
    )
    saved = load_workbook(path, read_only=False)
    try:
        assert result.legacy is False
        assert saved[META_SHEET].sheet_state == "hidden"
        metadata = json.loads(saved[META_SHEET].cell(1, 2).value)
        assert metadata["required_sheets"] == ["轨旁AP规划"]
        assert metadata["required_columns"]["轨旁AP规划"][:2] == ["车站名称", "AP数量"]
    finally:
        saved.close()

    with pytest.raises(ImportValidationError, match="当前模块不能导入"):
        validate_excel_import(
            path,
            expected_module="rail.car_network_point_table",
            required_headers={"轨旁AP规划": ["车站名称", "AP数量"]},
            allow_legacy=False,
        )


def test_unrecognized_excel_and_empty_file_are_rejected(tmp_path: Path) -> None:
    unrelated = tmp_path / "unrelated.xlsx"
    workbook = Workbook()
    workbook.active.append(["姓名", "电话"])
    workbook.active.append(["张三", "1"])
    workbook.save(unrelated)

    with pytest.raises(ImportValidationError, match="缺少必要字段"):
        validate_excel_import(
            unrelated,
            expected_module="devices",
            required_headers={"设备": ["设备名称", "主用地址"]},
            allow_legacy=True,
        )

    empty = tmp_path / "empty.xlsx"
    empty.write_bytes(b"")
    with pytest.raises(ImportValidationError, match="文件为空"):
        validate_excel_import(empty, expected_module="devices")


def test_csv_contract_checks_metadata_headers_columns_and_data(tmp_path: Path) -> None:
    path = tmp_path / "devices.csv"
    path.write_text("设备名称,主用地址\nSW1,192.0.2.1\n", encoding="utf-8-sig")
    attach_export_metadata(path, effective_suffix=".csv", export_type="device_csv", payload={"source_module": "devices"})

    result = validate_csv_import(path, expected_module="devices", required_headers=["设备名称", "主用地址"], allow_legacy=False)
    rows = list(csv.reader(path.read_text(encoding="utf-8-sig").splitlines()))
    metadata = json.loads(rows[0][1])
    assert rows[0][0] == CSV_META_MARKER
    assert metadata["required_columns"]["data"] == ["设备名称", "主用地址"]
    assert result.row_count == 1

    with pytest.raises(ImportValidationError, match="当前模块不能导入"):
        validate_csv_import(path, expected_module="ac.fit_ap", required_headers=["设备名称"], allow_legacy=False)

    bad_columns = tmp_path / "bad_columns.csv"
    bad_columns.write_text("设备名称,主用地址\nSW1\n", encoding="utf-8")
    with pytest.raises(ImportValidationError, match="列数量"):
        validate_csv_import(bad_columns, expected_module="devices", required_headers=["设备名称", "主用地址"])

    no_data = tmp_path / "no_data.csv"
    no_data.write_text("设备名称,主用地址\n", encoding="utf-8")
    with pytest.raises(ImportValidationError, match="文件为空"):
        validate_csv_import(no_data, expected_module="devices", required_headers=["设备名称", "主用地址"])

    unrelated = tmp_path / "unrelated.csv"
    unrelated.write_text("姓名,电话\n张三,1\n", encoding="utf-8")
    with pytest.raises(ImportValidationError, match="缺少必要字段"):
        validate_csv_import(unrelated, expected_module="devices", required_headers=["设备名称", "主用地址"])


def test_json_contract_checks_structure_module_and_schema(tmp_path: Path) -> None:
    path = tmp_path / "data.json"
    path.write_text(json.dumps([{"name": "SW1"}], ensure_ascii=False), encoding="utf-8")
    attach_export_metadata(path, effective_suffix=".json", export_type="table_json", payload={"source_module": "devices"})

    result = validate_json_import(path, expected_module="devices", required_keys=["name"])
    assert result.row_count == 1

    payload = json.loads(path.read_text(encoding="utf-8"))
    payload["_netconsole_meta"]["schema_version"] = 999
    path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    with pytest.raises(ImportValidationError, match="文件版本不兼容"):
        validate_json_import(path, expected_module="devices")

    broken = tmp_path / "broken.json"
    broken.write_text("{", encoding="utf-8")
    with pytest.raises(ImportValidationError, match="文件已损坏或无法读取"):
        validate_json_import(broken, expected_module="devices")

    unrelated = tmp_path / "unrelated.json"
    unrelated.write_text(json.dumps({"foo": "bar"}), encoding="utf-8")
    with pytest.raises(ImportValidationError, match="不是 NetConsole 导出的文件"):
        validate_json_import(unrelated, expected_module="devices")


def test_zip_contract_checks_manifest_required_files_and_path_traversal(tmp_path: Path) -> None:
    path = tmp_path / "bundle.zip"
    with zipfile.ZipFile(path, "w") as archive:
        archive.writestr("data/items.json", "[]")
    attach_export_metadata(path, effective_suffix=".zip", export_type="zip_files", payload={"source_module": "history"})

    result = validate_zip_import(path, expected_module="history", required_files=["data/items.json"])
    with zipfile.ZipFile(path) as archive:
        manifest = json.loads(archive.read(ZIP_MANIFEST).decode("utf-8"))
    assert result.row_count == 2
    assert manifest["files"] == ["data/items.json"]

    with pytest.raises(ImportValidationError, match="当前模块不能导入"):
        validate_zip_import(path, expected_module="settings")
    with pytest.raises(ImportValidationError, match="缺少必要文件"):
        validate_zip_import(path, expected_module="history", required_files=["data/missing.json"])

    no_manifest = tmp_path / "no_manifest.zip"
    with zipfile.ZipFile(no_manifest, "w") as archive:
        archive.writestr("data.txt", "plain text")
    with pytest.raises(ImportValidationError, match="缺少 manifest"):
        validate_zip_import(no_manifest, expected_module="history")

    unsafe = tmp_path / "unsafe.zip"
    with zipfile.ZipFile(unsafe, "w") as archive:
        archive.writestr("../outside.txt", "bad")
    with pytest.raises(ImportValidationError, match="不安全路径"):
        validate_zip_import(unsafe, expected_module="history", allow_external=True, allowed_external_suffixes=[".txt"])
