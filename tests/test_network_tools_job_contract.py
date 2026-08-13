from __future__ import annotations

import asyncio

import csv

import hashlib

import json

import os

from pathlib import Path

import threading

import time

from types import SimpleNamespace

from fastapi import FastAPI

from fastapi.testclient import TestClient

from openpyxl import load_workbook

from netconsole.backend.api.network_tools_router import router

from netconsole.core.feature_flags import FeatureGate

from netconsole.core.paths import PathResolver

from netconsole.models.task_snapshot import TaskSnapshot, utc_now_iso

from netconsole.models.task_state import TaskState

from netconsole.services.job_center.job_events import progress_event

from netconsole.services.job_center.job_models import JobSpec

from netconsole.services.job_center.job_runner import run_job

from netconsole.services.job_center.task_application_service import TaskApplicationService

from netconsole.services.job_center.worker_protocol import encode_event

from netconsole.services.network_tools import job_handlers

from netconsole.services.network_tools.application_service import NetworkToolsApplicationService

from netconsole.services.network_tools.job_handlers import (
    NETWORK_BATCH_PING_TASK,
    NETWORK_CONTINUOUS_PING_TASK,
    NETWORK_SINGLE_PING_TASK,
    NETWORK_TASK_SOURCE,
    NETWORK_TOOLBOX_EXPORT_TASK,
    NETWORK_TOOL_OWNER,
    NETWORK_WIRELESS_SCAN_TASK,
)

from netconsole.services.network_tools.toolbox.ping_tools import PingResult

from netconsole.services.network_tools.toolbox.fping_runner import FpingAvailability

from netconsole.services.windows_network_manager import NetworkAdapterInfo

class FakeTrafficService:
    paths = None
    site_name = "demo"

    async def start_tcp_port_test(self, config, execution_target):
        raise AssertionError("TCP Traffic adapter is not part of this test")

class DispatchingProcessAdapter:
    """测试用正式 Registry 分发器；保留 TaskRuntime 事件和取消协议。"""

    def __init__(self, task_service: TaskApplicationService, *, force_release: threading.Event | None = None) -> None:
        self.task_service = task_service
        self.force_release = force_release
        self.jobs: list[JobSpec] = []
        self._threads: dict[str, threading.Thread] = {}
        self._cancel: dict[str, threading.Event] = {}
        self.forced_jobs: set[str] = set()
        self.shutdown_calls = 0

    def start_job(self, job: JobSpec, *, on_complete=None) -> str:
        launch = self.task_service.prepare(job)
        self.task_service.mark_running(launch.job.job_id)
        self.jobs.append(launch.job)
        cancel = threading.Event()
        self._cancel[launch.job.job_id] = cancel
        thread = threading.Thread(
            target=self._run,
            args=(launch.job, cancel, on_complete),
            daemon=True,
        )
        self._threads[launch.job.job_id] = thread
        thread.start()
        return launch.job.job_id

    def _run(self, job: JobSpec, cancel: threading.Event, on_complete) -> None:
        def progress(stage: str, current: int, total: int, message: str) -> None:
            self.task_service.feed_stdout(
                job.job_id,
                encode_event(progress_event(job.job_id, stage, current, total, message)).encode("utf-8"),
            )

        result = run_job(job, progress_callback=progress, should_cancel=cancel.is_set)
        self.task_service.feed_stdout(job.job_id, encode_event(result.to_event()).encode("utf-8"))
        exit_code = 0 if result.ok else 2 if result.cancelled else 1
        payload = self.task_service.complete(job.job_id, exit_code)
        if on_complete is not None:
            on_complete(SimpleNamespace(job_id=job.job_id, task_type=job.task_type, exit_code=exit_code, payload=payload, cancelled=result.cancelled, forced=job.job_id in self.forced_jobs))

    def is_running(self, job_id: str) -> bool:
        thread = self._threads.get(job_id)
        return bool(thread and thread.is_alive())

    def cancel_job(self, job_id: str) -> bool:
        thread = self._threads.get(job_id)
        if thread is None or not thread.is_alive():
            return False
        self.task_service.request_cancel(job_id)
        self._cancel[job_id].set()
        if self.force_release is not None:
            def force_release() -> None:
                time.sleep(0.05)
                if thread.is_alive():
                    self.forced_jobs.add(job_id)
                    self.force_release.set()

            threading.Thread(target=force_release, daemon=True).start()
        return True

    def wait(self, job_id: str, timeout: float = 5) -> bool:
        thread = self._threads.get(job_id)
        if thread is None:
            return True
        thread.join(timeout)
        return not thread.is_alive()

    def shutdown(self, timeout_seconds: float = 5) -> None:
        self.shutdown_calls += 1
        for job_id in tuple(self._threads):
            self.cancel_job(job_id)
        for thread in tuple(self._threads.values()):
            thread.join(timeout_seconds)

def _service(
    tmp_path: Path,
    *,
    force_release: threading.Event | None = None,
    wireless_service: object | None = None,
    network_manager: object | None = None,
) -> tuple[NetworkToolsApplicationService, DispatchingProcessAdapter]:
    paths = PathResolver(tmp_path)
    task_service = TaskApplicationService(paths=paths, site_name="demo")
    adapter = DispatchingProcessAdapter(task_service, force_release=force_release)
    service = NetworkToolsApplicationService(
        FakeTrafficService(),
        task_service=task_service,
        paths=paths,
        site_name="demo",
        wireless_scan_service=wireless_service,
        network_manager=network_manager,
        process_adapter=adapter,
    )
    return service, adapter

def _app(tmp_path: Path, service: NetworkToolsApplicationService) -> FastAPI:
    app = FastAPI()
    app.include_router(router, prefix="/api")
    app.state.network_tools_service = service
    app.state.feature_gate = FeatureGate(tmp_path)
    return app

def _wait_task(service: NetworkToolsApplicationService, adapter: DispatchingProcessAdapter, task_id: str, *, wireless: bool = False) -> TaskSnapshot:
    assert adapter.wait(task_id, 10)
    task = service.get_wireless_task(task_id) if wireless else service.get_network_task(task_id)
    assert task is not None
    return task


def test_registered_fake_probe_uses_worker_task_minimal_snapshot_paging_and_export(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr(
        job_handlers,
        "run_single_ping",
        lambda *_args, **_kwargs: PingResult(
            target="fake-host",
            status="online",
            latency_ms=1.2,
            received=1,
            sent=1,
            raw_output="secret raw output",
        ),
    )
    service, adapter = _service(tmp_path)

    task = asyncio.run(service.start_network_task(kind="single_ping", target="fake-host"))
    completed = _wait_task(service, adapter, task.task_id)

    assert completed.task_type == NETWORK_SINGLE_PING_TASK
    assert completed.owner == NETWORK_TOOL_OWNER
    assert completed.source == NETWORK_TASK_SOURCE
    assert completed.result == {"result_id": task.task_id, "row_count": 1}
    page = service.list_network_task_results(task.task_id, offset=0, limit=1)
    assert page["total"] == 1
    assert page["items"][0]["target"] == "fake-host"
    assert "raw_output" not in page["items"][0]
    assert NETWORK_SINGLE_PING_TASK in job_handlers.HANDLERS

    with TestClient(_app(tmp_path, service)) as client:
        submitted = client.post(
            f"/api/network-tools/runs/{task.task_id}/export",
            json={"format": "csv", "filename": "probe.csv"},
        )
    assert submitted.status_code == 202
    export_id = submitted.json()["task"]["id"]
    exported = _wait_task(service, adapter, export_id)
    assert exported.task_type == NETWORK_TOOLBOX_EXPORT_TASK
    assert set(exported.result) == {
        "artifact_id",
        "filename",
        "format",
        "owner",
        "parent_id",
        "physical_name",
        "result_id",
        "row_count",
        "sha256",
        "site_name",
        "size",
        "source",
        "task_id",
        "task_type",
    }
    artifact = service.get_network_export_artifact(export_id)
    path, filename, opened = service.open_network_artifact(str(artifact["artifact_id"]))
    assert filename == "probe.csv"
    assert opened == artifact
    assert hashlib.sha256(path.read_bytes()).hexdigest() == artifact["sha256"]
    manifest = json.loads(path.with_suffix(".json").read_text(encoding="utf-8"))
    assert manifest["task_id"] == export_id
    assert manifest["parent_id"] == task.task_id
    assert manifest["site_name"] == "demo"
    assert manifest["owner"] == NETWORK_TOOL_OWNER
    assert manifest["source"] == NETWORK_TASK_SOURCE
    assert manifest["task_type"] == NETWORK_TOOLBOX_EXPORT_TASK

    restored_adapter = DispatchingProcessAdapter(service.task_service)
    restored = NetworkToolsApplicationService(
        FakeTrafficService(),
        task_service=service.task_service,
        paths=service.paths,
        process_adapter=restored_adapter,
    )
    assert restored.get_network_export_artifact(export_id)["sha256"] == artifact["sha256"]


def test_batch_probe_dispatches_registered_handler_and_never_embeds_rows(tmp_path: Path, monkeypatch) -> None:
    def fake_batch(targets, *, progress=None, **_kwargs):
        rows = [PingResult(target=target, status="online", received=1, sent=1, raw_output="raw") for target in targets]
        for row in rows:
            if progress:
                progress(row)
        return rows

    monkeypatch.setattr(job_handlers, "run_batch_ping", fake_batch)
    service, adapter = _service(tmp_path)
    targets = [f"fake-{index}" for index in range(30)]
    task = asyncio.run(service.start_network_task(kind="batch_ping", targets=targets))
    completed = _wait_task(service, adapter, task.task_id)

    assert completed.task_type == NETWORK_BATCH_PING_TASK
    assert completed.result == {"result_id": task.task_id, "row_count": len(targets)}
    first = service.list_network_task_results(task.task_id, offset=0, limit=7)
    second = service.list_network_task_results(task.task_id, offset=7, limit=7)
    assert len(first["items"]) == len(second["items"]) == 7
    assert first["items"][0]["target"] == "fake-0"
    assert second["items"][0]["target"] == "fake-7"
    assert all("raw_output" not in row for row in first["items"])


def test_long_probe_results_support_bounded_offset_and_incremental_cursor_paging(tmp_path: Path) -> None:
    service, _adapter = _service(tmp_path)
    task_id = "a" * 32
    now = utc_now_iso()
    service.task_service.repository("demo").save(
        TaskSnapshot(
            task_id=task_id,
            task_type=NETWORK_BATCH_PING_TASK,
            task_name="长结果分页",
            status=TaskState.COMPLETED,
            created_time=now,
            updated_time=now,
            finished_time=now,
            current=10_000,
            owner=NETWORK_TOOL_OWNER,
            source=NETWORK_TASK_SOURCE,
            site_name="demo",
            result={"result_id": task_id, "row_count": 10_000},
        )
    )
    result_path = service.paths.toolbox_outputs_dir("demo") / f"{task_id}.jsonl"
    result_path.parent.mkdir(parents=True, exist_ok=True)
    lines = [json.dumps({"sequence": index, "target": f"host-{index}"}) for index in range(10_000)]
    lines.insert(9_005, "{broken-json")
    result_path.write_text("\n".join(lines) + "\n", encoding="utf-8")

    deep_page = service.list_network_task_results(task_id, offset=9_000, limit=5)
    cursor_page = service.list_network_task_results(
        task_id,
        offset=deep_page["next_offset"],
        limit=5,
        cursor=deep_page["next_cursor"],
    )

    assert deep_page["total"] == 10_000
    assert [row["sequence"] for row in deep_page["items"]] == list(range(9_000, 9_005))
    assert deep_page["has_more"] is True
    assert [row["sequence"] for row in cursor_page["items"]] == list(range(9_005, 9_010))
    assert cursor_page["next_offset"] == 9_010
    assert cursor_page["next_cursor"] > deep_page["next_cursor"]

    with TestClient(_app(tmp_path, service)) as client:
        invalid_cursor = client.get(
            f"/api/network-tools/runs/{task_id}/results",
            params={"offset": 9_000, "limit": 5, "cursor": deep_page["next_cursor"] + 1},
        )
    assert invalid_cursor.status_code == 422


def test_active_probe_incomplete_tail_does_not_request_immediate_cursor_page(
    tmp_path: Path,
) -> None:
    service, _adapter = _service(tmp_path)
    task_id = "b" * 32
    now = utc_now_iso()
    service.task_service.repository("demo").save(
        TaskSnapshot(
            task_id=task_id,
            task_type=NETWORK_CONTINUOUS_PING_TASK,
            task_name="持续 Ping",
            status=TaskState.RUNNING,
            created_time=now,
            updated_time=now,
            current=2,
            owner=NETWORK_TOOL_OWNER,
            source=NETWORK_TASK_SOURCE,
            site_name="demo",
            result={"result_id": task_id},
        )
    )
    result_path = service.paths.toolbox_outputs_dir("demo") / f"{task_id}.jsonl"
    result_path.parent.mkdir(parents=True, exist_ok=True)
    result_path.write_bytes(b'{"sequence": 1}\n{"sequence": 2')

    page = service.list_network_task_results(task_id, limit=500, cursor=0)

    assert page["items"] == [{"sequence": 1}]
    assert page["total"] == 2
    assert page["next_offset"] == 1
    assert page["has_more"] is False


def test_subnet_ping_uses_existing_adapter_source_full_range_and_reports_engine(tmp_path: Path, monkeypatch) -> None:
    captured: dict[str, object] = {}

    class FakeNetworkManager:
        def list_adapters(self):
            return [NetworkAdapterInfo(name="Ethernet 2", interface_index=12, status="Up", ipv4_addresses=["192.168.50.10/30"])]

    def fake_batch(targets, *, source_ip="", progress=None, **_kwargs):
        captured["targets"] = list(targets)
        captured["source_ip"] = source_ip
        rows = [PingResult(target=target, status="online", received=1, sent=1) for target in targets]
        for row in rows:
            if progress:
                progress(row)
        return rows

    unavailable = FpingAvailability(False, error="测试环境未提供 fping")
    monkeypatch.setattr(job_handlers, "discover_fping", lambda *_args, **_kwargs: unavailable)
    monkeypatch.setattr(job_handlers, "run_batch_ping", fake_batch)
    monkeypatch.setattr("netconsole.services.network_tools.application_service.discover_fping", lambda *_args, **_kwargs: unavailable)
    service, adapter = _service(tmp_path, network_manager=FakeNetworkManager())

    with TestClient(_app(tmp_path, service)) as client:
        environment = client.get("/api/network-tools/toolbox/probe-environment")
        submitted = client.post(
            "/api/network-tools/tasks",
            json={
                "kind": "subnet_ping",
                "target": "192.168.50.10/30",
                "source_ip": "192.168.50.10",
                "usable_only": False,
            },
        )

    assert environment.status_code == 200
    assert environment.json()["adapters"][0]["interface_index"] == 12
    assert environment.json()["scan_engine"] == "系统 ping"
    assert submitted.status_code == 202
    task_id = submitted.json()["task"]["id"]
    completed = _wait_task(service, adapter, task_id)
    assert completed.result["engine"] == "系统 ping（测试环境未提供 fping）"
    assert adapter.jobs[-1].params["source_ip"] == "192.168.50.10"
    assert adapter.jobs[-1].params["usable_only"] is False
    assert captured == {
        "targets": ["192.168.50.8", "192.168.50.9", "192.168.50.10", "192.168.50.11"],
        "source_ip": "192.168.50.10",
    }


def test_continuous_ping_cancel_preserves_samples_for_paging_and_csv_xlsx_export(tmp_path: Path, monkeypatch) -> None:
    sample = 0

    def fake_ping(target: str, **_kwargs) -> PingResult:
        nonlocal sample
        sample += 1
        return PingResult(target=target, status="online", latency_ms=float(sample), received=1, sent=1)

    monkeypatch.setattr(job_handlers, "run_single_ping", fake_ping)
    service, adapter = _service(tmp_path)
    task = asyncio.run(service.start_network_task(kind="continuous_ping", target="fake-host", interval_ms=1))
    result_path = service.paths.toolbox_outputs_dir("demo") / f"{task.task_id}.jsonl"
    deadline = time.monotonic() + 30
    while time.monotonic() < deadline:
        if result_path.is_file() and len(result_path.read_text(encoding="utf-8").splitlines()) >= 1001:
            break
        time.sleep(0.01)
    else:
        raise AssertionError("持续 Ping 在 1000 个样本前后未保持增量写入")
    assert adapter.is_running(task.task_id)
    running = service.get_network_task(task.task_id)
    assert running is not None and running.status is TaskState.RUNNING

    with TestClient(_app(tmp_path, service)) as client:
        stopping = client.post(f"/api/network-tools/runs/{task.task_id}/cancel")
        assert stopping.status_code == 200
        assert stopping.json()["status"] == "STOPPING"
        terminal = _wait_task(service, adapter, task.task_id)
        assert terminal.task_type == NETWORK_CONTINUOUS_PING_TASK
        assert terminal.status is TaskState.CANCELLED

        first_page = client.get(f"/api/network-tools/runs/{task.task_id}/results", params={"offset": 0, "limit": 2})
        second_page = client.get(f"/api/network-tools/runs/{task.task_id}/results", params={"offset": 2, "limit": 2})
        assert first_page.status_code == second_page.status_code == 200
        total = first_page.json()["total"]
        assert total >= 1001
        assert len(first_page.json()["items"]) == 2
        assert second_page.json()["items"]

        exported_paths: dict[str, Path] = {}
        for file_format in ("csv", "xlsx"):
            submitted = client.post(f"/api/network-tools/runs/{task.task_id}/export", json={"format": file_format})
            assert submitted.status_code == 202
            export_id = submitted.json()["task"]["id"]
            exported = _wait_task(service, adapter, export_id)
            assert exported.status is TaskState.COMPLETED
            artifact = service.get_network_export_artifact(export_id)
            exported_paths[file_format] = service.open_network_artifact(str(artifact["artifact_id"]))[0]

    with exported_paths["csv"].open("r", encoding="utf-8-sig", newline="") as handle:
        csv_rows = list(csv.reader(handle))
        assert csv_rows[0][0] == "#NETCONSOLE_META"
        assert len(csv_rows) == total + 2
    workbook = load_workbook(exported_paths["xlsx"], read_only=True)
    try:
        assert sum(1 for row in workbook.active.iter_rows(values_only=True) if row[0] == "fake-host") == total
    finally:
        workbook.close()


def test_toolbox_and_wireless_task_scope_rejects_site_owner_source_and_exact_type_mismatches(tmp_path: Path) -> None:
    service, adapter = _service(tmp_path)
    repository = service.task_service.repository("demo")
    now = utc_now_iso()
    cases = (
        ("wrong-site", "other", NETWORK_TOOL_OWNER, "local", NETWORK_SINGLE_PING_TASK),
        ("wrong-owner", "demo", "other", "local", NETWORK_SINGLE_PING_TASK),
        ("wrong-source", "demo", NETWORK_TOOL_OWNER, "agent", NETWORK_SINGLE_PING_TASK),
        ("wrong-type", "demo", NETWORK_TOOL_OWNER, "local", "network_tools.single_ping_extra"),
    )
    for task_id, site, owner, source, task_type in cases:
        repository.save(TaskSnapshot(
            task_id=task_id,
            task_type=task_type,
            task_name="foreign",
            status=TaskState.RUNNING,
            created_time=now,
            updated_time=now,
            owner=owner,
            source=source,
            site_name=site,
            owner_pid=os.getpid(),
        ))
    repository.save(TaskSnapshot(
        task_id="wireless-only",
        task_type=NETWORK_WIRELESS_SCAN_TASK,
        task_name="wireless",
        status=TaskState.COMPLETED,
        created_time=now,
        updated_time=now,
        owner=NETWORK_TOOL_OWNER,
        source="local",
        site_name="demo",
    ))

    app = _app(tmp_path, service)
    app.state.feature_gate.features["capability.network_tools.wireless_scan"] = {"visible": True, "enabled": True}
    with TestClient(app) as client:
        for task_id, *_rest in cases:
            assert client.get(f"/api/network-tools/runs/{task_id}").status_code == 404
            assert client.post(f"/api/network-tools/runs/{task_id}/cancel").status_code == 404
        assert client.get("/api/network-tools/runs/wireless-only").status_code == 404
        assert client.get("/api/network-tools/wireless-scan/tasks/wireless-only").status_code == 200
        assert client.get("/api/network-tools/wireless-scan/tasks/wireless-only/events").status_code == 200
        assert client.post("/api/network-tools/wireless-scan/tasks/wrong-owner/cancel").status_code == 404
    assert not adapter.jobs


def test_active_network_task_is_reconciled_when_its_worker_is_missing(tmp_path: Path) -> None:
    paths = PathResolver(tmp_path)
    task_service = TaskApplicationService(paths=paths, site_name="demo")
    now = utc_now_iso()
    task_service.repository("demo").save(TaskSnapshot(
        task_id="orphan-network-task",
        task_type=NETWORK_SINGLE_PING_TASK,
        task_name="orphan",
        status=TaskState.RUNNING,
        created_time=now,
        updated_time=now,
        owner=NETWORK_TOOL_OWNER,
        source=NETWORK_TASK_SOURCE,
        site_name="demo",
        owner_pid=os.getpid(),
    ))
    adapter = DispatchingProcessAdapter(task_service)
    service = NetworkToolsApplicationService(
        FakeTrafficService(),
        task_service=task_service,
        paths=paths,
        process_adapter=adapter,
    )

    recovered = task_service.repository("demo").get("orphan-network-task")
    assert recovered is not None and recovered.status is TaskState.FAILED
    assert "Worker" in recovered.error_message
    assert service.get_network_task("orphan-network-task") is not None


def test_network_task_target_items_have_length_limit(tmp_path: Path) -> None:
    service, _adapter = _service(tmp_path)
    with TestClient(_app(tmp_path, service)) as client:
        response = client.post(
            "/api/network-tools/tasks",
            json={"kind": "batch_ping", "targets": ["x" * 256]},
        )
    assert response.status_code == 422
