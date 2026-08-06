from __future__ import annotations

from datetime import datetime
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

from netconsole.services.mesh_analysis_excel_report import EMPTY_PARSE_ISSUES_TEXT, MeshAnalysisExcelReportExporter, REPORT_FIELD_LABELS, SHEET_DEFINITIONS, translate_report_value
from netconsole.services.mesh_analysis_excel_report import MAX_EMBEDDED_CHART_POINTS, _downsample_chart_rows
from netconsole.services.mesh_analysis_report import (
    MeshAnalysisReportModel,
    MeshAnalysisReportService,
    MeshReportOptions,
    build_active_anomalies,
    build_analysis_parameter_rows,
    build_active_segments,
    build_channel_busy_statistics,
    build_link_establishment_order,
    build_peer_lifecycle,
    build_rssi_statistics,
    build_switch_sequence,
    detect_flap_switches,
    _active_path_report_rows,
)
from netconsole.services.mesh_chart_payload import build_chart_payload
from netconsole.services.mesh_link_detail_export import active_build_order_row_values
from netconsole.services.mesh_report_process import _analysis_params_for_report
from netconsole.services.rail_transit.mesh_analysis_query_service import MeshAnalysisQueryService
from netconsole.core.paths import PathResolver
from netconsole.repositories.mesh_mr_repository import MeshMrRepository
from netconsole.services.mesh_import_service import MeshImportService
from netconsole.services.mesh_storage_service import MeshStorageService
from netconsole.services.mesh_quality_analysis import (
    MeshQualityRules,
    analyze_anomaly_events,
    analyze_link_rebuilds,
    analyze_switch_events,
    build_active_segments as build_quality_active_segments,
    build_busy_analysis,
    build_overview,
    build_sample_quality,
    get_threshold_template,
    load_threshold_templates,
    normalize_samples,
    percentile,
    template_overview_fields,
)


PEER_A = "30f5277a5a2f"
PEER_B = "30f5277a5a30"
PEER_C = "30f5277a5a31"


def _row(
    sample_time: str,
    peer: str,
    state: str = "ACTIVE",
    radio: int = 1,
    mr_rssi: int = 41,
    peer_rssi: int = 39,
    peer_ap_name: str | None = None,
    peer_ap_mac: str | None = None,
    peer_radio: str | None = None,
) -> dict[str, object]:
    return {
        "sample_time": sample_time,
        "radio": radio,
        "link_state": state,
        "peer_mac_normalized": peer,
        "peer_ap_name": peer_ap_name or f"AP-{peer[-4:]}",
        "peer_ap_mac": peer_ap_mac or peer,
        "peer_site": "03横溪站",
        "peer_radio": peer_radio or "radio1",
        "peer_radio_label": peer_radio or "radio1",
        "peer_radio_mac": peer,
        "establish_time": "2025-12-03 09:59:00.000",
        "duration_seconds": 60,
        "mr_rssi": mr_rssi,
        "peer_rssi": peer_rssi,
        "local_tx_busy": 12,
        "local_rx_busy": 8,
        "peer_tx_busy": 6,
        "peer_rx_busy": 4,
        "local_rate_raw": 866,
        "archived_filename": "mesh.log",
        "source_line_number": 1,
        "raw_line": f"{sample_time} {state} {peer}",
    }


def _source_row(source_file_id: int, sample_time: str, peer: str, state: str, radio: int = 1, mr_rssi: int | None = 41, peer_rssi: int | None = 39) -> dict[str, object]:
    row = _row(sample_time, peer, state, radio=radio, mr_rssi=mr_rssi if mr_rssi is not None else 0, peer_rssi=peer_rssi if peer_rssi is not None else 0)
    row["source_file_id"] = source_file_id
    row["archived_filename"] = f"source-{source_file_id}.log"
    row["source_line_number"] = source_file_id * 100
    if mr_rssi is None:
        row["mr_rssi"] = None
    if peer_rssi is None:
        row["peer_rssi"] = None
    return row


def test_active_segments_preserve_aba_runs_and_switch_sequence():
    rows = [
        _row("2025-12-03 10:00:00.000", PEER_A, mr_rssi=40),
        _row("2025-12-03 10:00:01.000", PEER_A, mr_rssi=42),
        _row("2025-12-03 10:00:02.000", PEER_B, mr_rssi=55),
        _row("2025-12-03 10:00:03.000", PEER_A, mr_rssi=41),
    ]
    segments = build_active_segments(rows)
    assert [segment["active_peer_mac"] for segment in segments] == [PEER_A, PEER_B, PEER_A]
    assert segments[0]["sample_count"] == 2
    switches = build_switch_sequence(segments)
    assert [(switch["from_peer_mac"], switch["to_peer_mac"]) for switch in switches] == [(PEER_A, PEER_B), (PEER_B, PEER_A)]
    assert switches[0]["from_mr_rssi"] == 41
    assert switches[0]["to_mr_rssi"] == 55


def test_report_active_segments_rssi_stats_are_complete_and_consistent():
    rows = [
        _row("2025-12-03 10:00:00.000", PEER_A, mr_rssi=0),
        _row("2025-12-03 10:00:01.000", PEER_A, mr_rssi=40),
        _row("2025-12-03 10:00:02.000", PEER_A, mr_rssi=50),
    ]
    segment = build_active_segments(rows)[0]

    assert segment["avg_mr_rssi"] == 45
    assert segment["min_mr_rssi"] == 40
    assert segment["max_mr_rssi"] == 50
    assert segment["p10_mr_rssi"] == 41
    assert segment["min_mr_rssi"] <= segment["avg_mr_rssi"] <= segment["max_mr_rssi"]


def test_mesh_threshold_templates_follow_business_scenario_not_wifi_generation():
    templates = load_threshold_templates()

    assert list(templates)[:2] == ["pis_wifi6_40_80_standard", "pis_wifi6_80_high_quality"]
    assert templates["pis_wifi6_40_80_standard"].label == "PIS - Wi-Fi6 - 40/80M - 标准间隔"
    assert templates["pis_wifi6_40_80_standard"].rules.rssi_good_threshold == 32
    assert templates["pis_wifi6_40_80_standard"].rules.backup_available_threshold == 32
    assert templates["dcs_wifi6_dot11a_20_far"].business_type == "DCS/信号"
    assert templates["dcs_wifi6_dot11a_20_far"].working_mode == "强制 dot11a"
    assert templates["dcs_wifi6_dot11a_20_far"].rules.rssi_good_threshold == 28
    assert templates["dcs_wifi6_dot11a_20_far"].rules.backup_available_threshold == 28
    assert templates["dcs_wifi6_dot11a_20_far"].rules.no_backup_min_seconds == 15
    assert "Wi-Fi6" in templates["dcs_wifi6_dot11a_20_far"].label
    assert "dot11a/20M" in templates["dcs_wifi6_dot11a_20_far"].label


def test_template_overview_fields_explain_selected_evaluation_context():
    template = get_threshold_template("pis_wifi6_40_80_standard")
    fields = template_overview_fields(template.key, template.rules)

    assert fields["评估模板"] == "PIS - Wi-Fi6 - 40/80M - 标准间隔"
    assert fields["业务类型"] == "PIS"
    assert fields["实际工作模式"] == "Wi-Fi6 / 11ax"
    assert fields["RSSI 良好线"] == 32
    assert fields["可用备份线"] == 32
    assert fields["无备份风险窗口"] == "5秒"

    overview = build_overview(
        "MR-01",
        "",
        "MR_RAW_MESH_LOG",
        [],
        [],
        [],
        [],
        [],
        [],
        [],
        100,
        "优秀",
        [],
        template.key,
        template.rules,
    )
    assert overview["报告名称"] == "MR-01"
    assert overview["评估模板"] == template.label
    assert overview["RSSI 良好线"] == template.rules.rssi_good_threshold






def test_flap_detects_aba_inside_window_and_ignores_non_flap():
    flap_segments = build_active_segments(
        [
            _row("2025-12-03 10:00:00.000", PEER_A),
            _row("2025-12-03 10:00:01.000", PEER_B),
            _row("2025-12-03 10:00:02.000", PEER_A),
        ]
    )
    flap = detect_flap_switches(flap_segments, flap_window_seconds=5)[0]
    assert flap["flap_type"] == "AP乒乓切换异常"
    assert flap["is_pingpong_abnormal"] is True
    non_flap_segments = build_active_segments(
        [
            _row("2025-12-03 10:00:00.000", PEER_A),
            _row("2025-12-03 10:00:01.000", PEER_B),
            _row("2025-12-03 10:00:02.000", PEER_C),
        ]
    )
    assert detect_flap_switches(non_flap_segments, flap_window_seconds=5) == []


def test_flap_detects_critical_normal_and_same_ap_radio_return_without_abnormal():
    critical_segments = build_active_segments(
        [
            _row("2025-12-03 10:00:00.000", PEER_A),
            _row("2025-12-03 10:00:01.000", PEER_B),
            _row("2025-12-03 10:00:02.000", PEER_B),
            _row("2025-12-03 10:00:03.000", PEER_B),
            _row("2025-12-03 10:00:04.000", PEER_A),
        ]
    )
    normal_segments = build_active_segments(
        [
            _row("2025-12-03 10:10:00.000", PEER_A),
            _row("2025-12-03 10:10:01.000", PEER_B),
            _row("2025-12-03 10:10:02.000", PEER_B),
            _row("2025-12-03 10:10:03.000", PEER_B),
            _row("2025-12-03 10:10:04.000", PEER_B),
            _row("2025-12-03 10:10:05.000", PEER_B),
            _row("2025-12-03 10:10:06.000", PEER_A),
        ]
    )
    same_ap_segments = build_active_segments(
        [
            _row("2025-12-03 10:20:00.000", PEER_A, peer_ap_name="AP-A", peer_ap_mac="aaaa", peer_radio="radio1"),
            _row("2025-12-03 10:20:01.000", PEER_B, peer_ap_name="AP-A", peer_ap_mac="aaaa", peer_radio="radio2"),
            _row("2025-12-03 10:20:02.000", PEER_A, peer_ap_name="AP-A", peer_ap_mac="aaaa", peer_radio="radio1"),
        ]
    )

    critical = detect_flap_switches(critical_segments, flap_window_seconds=5)[0]
    normal = detect_flap_switches(normal_segments, flap_window_seconds=10)[0]
    same_ap = detect_flap_switches(same_ap_segments, flap_window_seconds=5)[0]

    assert critical["flap_type"] == "临界回切"
    assert critical["is_pingpong_abnormal"] is False
    assert normal["flap_type"] == "普通回切事件"
    assert normal["is_pingpong_abnormal"] is False
    assert same_ap["flap_type"] == "同AP射频往返"
    assert same_ap["is_pingpong_abnormal"] is False


def test_link_establishment_order_lifecycle_and_anomalies():
    rows = [
        _row("2025-12-03 10:00:00.000", PEER_A, "STANDBY"),
        _row("2025-12-03 10:00:01.000", PEER_A, "ACTIVE"),
        _row("2025-12-03 10:00:02.000", PEER_B, "ACTIVE"),
        _row("2025-12-03 10:00:03.000", PEER_A, "ACTIVE"),
        _row("2025-12-03 10:00:04.000", PEER_B, "STANDBY"),
    ]
    segments = build_active_segments(rows)
    switches = build_switch_sequence(segments)
    order = build_link_establishment_order(rows)
    assert [item["peer_mac"] for item in order] == [PEER_A, PEER_B]
    lifecycle = build_peer_lifecycle(rows, segments, switches)
    peer_a = next(item for item in lifecycle if item["peer_mac"] == PEER_A)
    assert peer_a["standby_sample_count"] == 1
    assert peer_a["active_segment_count"] == 2

    anomalies = build_active_anomalies(
        [
            _row("2025-12-03 10:01:00.000", PEER_A, "STANDBY"),
            _row("2025-12-03 10:01:01.000", PEER_A, "ACTIVE"),
            _row("2025-12-03 10:01:01.000", PEER_B, "ACTIVE"),
        ]
    )
    assert [item["anomaly_type"] for item in anomalies] == ["NO_ACTIVE", "MULTI_ACTIVE"]


def test_rssi_and_channel_busy_statistics_keep_raw_positive_values():
    rows = [_row("2025-12-03 10:00:00.000", PEER_A, mr_rssi=43, peer_rssi=37), _row("2025-12-03 10:00:01.000", PEER_A, mr_rssi=45, peer_rssi=39)]
    rssi = build_rssi_statistics(rows)[0]
    busy = build_channel_busy_statistics(rows)[0]
    assert rssi["mr_rssi_avg"] == 44
    assert rssi["mr_rssi_min"] == 43
    assert rssi["peer_rssi_max"] == 39
    assert busy["local_tx_busy_avg"] == 12


def test_excel_report_contains_required_sheets_headers_and_empty_parse_issue_text(tmp_path):
    model = MeshAnalysisReportModel(
        mr_name="14CW-01",
        report_name="14CW-01",
        generated_at=datetime.now(),
        options=MeshReportOptions(),
        overview={"mr_name": "14CW-01"},
        active_segments=build_active_segments([_row("2025-12-03 10:00:00.000", PEER_A)]),
    )
    path = MeshAnalysisExcelReportExporter().export(model, tmp_path / "report.xlsx")
    workbook = load_workbook(path)
    assert workbook.sheetnames == [definition[0] for definition in SHEET_DEFINITIONS]
    assert [cell.value for cell in workbook["主链路切换分析"][1]][:4] == [
        REPORT_FIELD_LABELS["sequence"],
        REPORT_FIELD_LABELS["radio"],
        REPORT_FIELD_LABELS["switch_time"],
        REPORT_FIELD_LABELS["from_peer"],
    ]
    assert workbook["解析问题"]["F2"].value == EMPTY_PARSE_ISSUES_TEXT


def test_formal_report_reuses_repository_results_and_exports_fixed_active_series(tmp_path: Path):
    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-01")
    source = tmp_path / "meshlog.log"
    active_a = "[1] Active 30f5-277a-5a2f 2025/12/03 10:12:30 0d 00h 00m 03s 1 36/43 2%/4% 45%/47% 3/1 15/27 60/72060 88/105 0/5000 2/297 314/0 0/93 0/0 0/0 0/0"
    active_b = active_a.replace("Active 30f5-277a-5a2f", "Active 30f5-277a-5a3f").replace("36/43", "37/44")
    standby = active_a.replace("Active 30f5-277a-5a2f", "Standy 30f5-277a-5a4f").replace("36/43", "30/31")
    source.write_text(
        "\n".join(
            [
                "[1] 2025/12/03 10:12:33.000 (2)",
                active_a,
                standby,
                "[1] 2025/12/03 10:12:34.000 (4)",
                active_b,
                "[1] 2025/12/03 10:12:35.000 (5)",
                active_a,
            ]
        ),
        encoding="utf-8",
    )
    MeshImportService("demo", paths).import_files(profile, [source])
    index_path = paths.mesh_mr_db_path("demo", profile.safe_folder_name)
    index_repository = MeshMrRepository(index_path, read_only=True)
    source_file = index_repository.list_source_files()[0]
    source_file_id = int(source_file["id"])
    expected_order = index_repository.query_active_link_build_order(source_file_id)

    model = MeshAnalysisReportService(index_path, profile.display_name).build_report(
        MeshReportOptions(source_file_id=source_file_id)
    )

    assert [row["active_peer_mac"] for row in model.active_build_order] == [
        row["active_peer_mac"] for row in expected_order
    ]
    assert len(model.link_details) == 4
    assert {row["timestamp_tag"] for row in model.link_details} == {"2", "4", "5"}
    assert len(model.active_path_rssi) == 3
    assert len(model.active_path_busy) == 3
    assert [row["visit_sequence"] for row in model.peer_visit_statistics] == [1, 1, 2]

    report_path = MeshAnalysisExcelReportExporter().export(model, tmp_path / "formal-report.xlsx")
    workbook = load_workbook(report_path)
    expected_business_sheets = [definition[0] for definition in SHEET_DEFINITIONS]
    assert workbook.sheetnames[: len(expected_business_sheets)] == expected_business_sheets
    assert {"链路明细", "全量链路明细", "Peer质量排名", "AP统计"}.isdisjoint(workbook.sheetnames)
    assert all(workbook[name].sheet_state == "hidden" for name in workbook.sheetnames[len(expected_business_sheets) :])
    build_sheet = workbook["主链路建链顺序"]
    assert [cell.value for cell in build_sheet[2]] == [
        None if value == "" else value for value in active_build_order_row_values(model.active_build_order[0])
    ]
    busy_sheet = workbook["空口负载分析"]
    busy_headers = [cell.value for cell in busy_sheet[1]]
    busy_values = dict(zip(busy_headers, (cell.value for cell in busy_sheet[2])))
    assert "CtlBusy" not in " ".join(str(value) for value in busy_headers)
    assert busy_values["MR侧 TxBusy"] == model.active_path_busy[0]["local_tx_busy"]
    assert len(workbook["全部 ACTIVE RSSI 分析"]._charts[0].series) == 2
    assert len(workbook["空口负载分析"]._charts[0].series) == 2

    parameter_sheet = workbook["分析参数与阈值"]
    assert str(parameter_sheet.merged_cells) == "A1:L1"
    assert parameter_sheet.freeze_panes == "A3"
    assert parameter_sheet.auto_filter.ref.startswith("A2:L")
    parameter_headers = [cell.value for cell in parameter_sheet[2]]
    parameter_rows = {
        row[1].value: dict(zip(parameter_headers, (cell.value for cell in row)))
        for row in parameter_sheet.iter_rows(min_row=3)
    }
    assert parameter_rows["主链路切换基准时间"]["参数来源"] == "source_snapshot"
    assert parameter_rows["实际短时建链阈值"]["计算后有效值"] == 3500
    assert parameter_rows["同物理 AP 双射频合并"]["当前值"] == "是"
    assert parameter_rows["乒乓返回窗口"]["Source 快照值"] == "未配置"


def test_report_peer_busy_keeps_same_millisecond_timestamp_tags_isolated():
    rows = [
        {
            "id": 1,
            "source_file_id": 7,
            "sample_time": "2025-12-03 10:00:00.000",
            "timestamp_tag": "2",
            "radio": 1,
            "link_state": "ACTIVE",
            "peer_mac_normalized": PEER_A,
            "local_tx_busy": 1,
            "local_rx_busy": 2,
            "peer_tx_busy": 11,
            "peer_rx_busy": 12,
        },
        {
            "id": 2,
            "source_file_id": 7,
            "sample_time": "2025-12-03 10:00:00.000",
            "timestamp_tag": "4",
            "radio": 1,
            "link_state": "ACTIVE",
            "peer_mac_normalized": PEER_A,
            "local_tx_busy": 3,
            "local_rx_busy": 4,
            "peer_tx_busy": 21,
            "peer_rx_busy": 22,
        },
    ]
    chart = build_chart_payload({}, {"rows": rows, "events": []})

    _rssi, busy = _active_path_report_rows(chart, [{"id": 7, "archived_filename": "mesh.log"}])

    assert [row["timestamp_tag"] for row in busy] == ["2", "4"]
    assert [row["peer_tx_busy"] for row in busy] == [11, 21]
    assert [row["peer_rx_busy"] for row in busy] == [12, 22]


def test_report_time_window_filters_complete_segments_links_charts_and_visits(tmp_path: Path):
    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-02")
    source = tmp_path / "mesh-window.log"
    base = "[1] Active {peer} 2025/12/03 10:12:30 0d 00h 00m 03s 1 36/43 2%/4% 45%/47% 3/1 15/27 60/72060 88/105 0/5000 2/297 314/0 0/93 0/0 0/0 0/0"
    source.write_text(
        "\n".join(
            [
                "[1] 2025/12/03 10:12:33.000 (2)",
                base.format(peer="30f5-277a-5a2f"),
                "[1] 2025/12/03 10:12:34.000 (3)",
                base.format(peer="30f5-277a-5a3f"),
                "[1] 2025/12/03 10:12:35.000 (4)",
                base.format(peer="30f5-277a-5a2f"),
            ]
        ),
        encoding="utf-8",
    )
    MeshImportService("demo", paths).import_files(profile, [source])
    index_path = paths.mesh_mr_db_path("demo", profile.safe_folder_name)
    source_id = int(MeshMrRepository(index_path, read_only=True).list_source_files()[0]["id"])

    model = MeshAnalysisReportService(index_path, profile.display_name).build_report(
        MeshReportOptions(
            source_file_id=source_id,
            start_time="2025-12-03 10:12:34.000",
            end_time="2025-12-03 10:12:34.000",
        )
    )

    assert [row["sample_time"] for row in model.link_details] == ["2025-12-03 10:12:34.000"]
    assert [row["build_start_time"] for row in model.active_build_order] == ["2025-12-03 10:12:34.000"]
    assert [row["sample_time"] for row in model.active_path_rssi] == ["2025-12-03 10:12:34.000"]
    assert [row["sample_time"] for row in model.active_path_busy] == ["2025-12-03 10:12:34.000"]
    assert len(model.peer_visit_statistics) == 1


def test_embedded_chart_downsampling_preserves_key_points_and_has_fixed_limit():
    rows = [
        {
            "sample_time": f"2025-12-03 10:{index // 60:02d}:{index % 60:02d}.000",
            "mr_rssi": index % 80,
            "peer_rssi": 80 - (index % 80),
            "chart_key_point": index in {123, 30_000, 58_143},
        }
        for index in range(58_144)
    ]

    sampled = _downsample_chart_rows(rows, ("mr_rssi", "peer_rssi"))

    assert len(sampled) <= MAX_EMBEDDED_CHART_POINTS
    assert {123, 30_000, 58_143} <= {rows.index(row) for row in sampled if row.get("chart_key_point")}
    assert sampled[0] is rows[0]
    assert sampled[-1] is rows[-1]


def test_embedded_chart_downsampling_caps_all_key_points():
    rows = [
        {
            "sample_time": f"2025-12-03 10:{index // 60:02d}:{index % 60:02d}.000",
            "mr_rssi": index,
            "chart_key_point": True,
        }
        for index in range(MAX_EMBEDDED_CHART_POINTS + 1_000)
    ]

    sampled = _downsample_chart_rows(rows, ("mr_rssi",))

    assert len(sampled) == MAX_EMBEDDED_CHART_POINTS
    assert sampled[0] is rows[0]
    assert sampled[-1] is rows[-1]


def test_empty_active_series_does_not_create_embedded_chart(tmp_path: Path):
    model = MeshAnalysisReportModel(
        mr_name="14CW-03",
        report_name="14CW-03",
        generated_at=datetime.now(),
        options=MeshReportOptions(),
        active_path_rssi=[{"sample_time": "2025-12-03 10:00:00.000", "mr_rssi": "", "peer_rssi": ""}],
        active_path_busy=[{"sample_time": "2025-12-03 10:00:00.000", "local_tx_busy": "", "local_rx_busy": ""}],
    )

    workbook = load_workbook(MeshAnalysisExcelReportExporter().export(model, tmp_path / "empty-chart.xlsx"))

    assert workbook["全部 ACTIVE RSSI 分析"]._charts == []
    assert workbook["空口负载分析"]._charts == []
    assert "_ACTIVE_RSSI图表数据" not in workbook.sheetnames
    assert "_ACTIVE_BUSY图表数据" not in workbook.sheetnames


def test_report_analysis_params_keep_override_source_site_default_priority():
    source = {"analysis_params_json": '{"main_link_switch_time_ms":2222,"short_link_tolerance_ms":444}'}
    site = {"main_link_switch_time_ms": 3333, "pingpong_tolerance_ms": 777}

    assert _analysis_params_for_report(MeshReportOptions(site_analysis_params=site), source).main_link_switch_time_ms == 2222
    assert (
        _analysis_params_for_report(
            MeshReportOptions(
                analysis_params_override={"main_link_switch_time_ms": 1111},
                site_analysis_params=site,
            ),
            source,
        ).main_link_switch_time_ms
        == 1111
    )
    resolved = _analysis_params_for_report(
        MeshReportOptions(
            analysis_params_override={"main_link_switch_time_ms": 1111},
            site_analysis_params=site,
        ),
        source,
    )
    assert resolved.short_link_tolerance_ms == 444
    assert resolved.pingpong_tolerance_ms == 777
    assert _analysis_params_for_report(MeshReportOptions(site_analysis_params=site), {}).main_link_switch_time_ms == 3333
    assert _analysis_params_for_report(MeshReportOptions(), {}).main_link_switch_time_ms == 4000


def test_analysis_parameter_rows_show_each_candidate_and_final_source():
    options = MeshReportOptions(
        analysis_params_override={"main_link_switch_time_ms": 1111},
        site_analysis_params={"pingpong_tolerance_ms": 777},
        rssi_excellent_threshold=43,
    )
    source_files = [
        {
            "analysis_params_json": (
                '{"main_link_switch_time_ms":2222,"short_link_tolerance_ms":444,'
                '"merge_same_physical_ap_dual_radio":false}'
            )
        }
    ]
    rows = build_analysis_parameter_rows(options, source_files, MeshQualityRules(rssi_excellent_threshold=43), [])
    by_name = {str(row["parameter_name"]): row for row in rows}

    switch = by_name["主链路切换基准时间"]
    assert switch["effective_value"] == 1111
    assert switch["parameter_source"] == "report_override"
    assert switch["report_override"] == 1111
    assert switch["source_snapshot"] == 2222
    assert switch["site_config"] is None
    assert switch["global_default"] == 4000
    assert by_name["短时判定容差"]["parameter_source"] == "source_snapshot"
    assert by_name["乒乓判定容差"]["parameter_source"] == "site_config"
    assert by_name["RSSI 优"]["parameter_source"] == "report_override"
    assert by_name["RSSI 良"]["parameter_source"] == "global_default"
    assert by_name["同物理 AP 双射频合并"]["effective_value"] is False


def test_report_and_page_do_not_enrich_unresolved_peer_from_location_snapshot(
    tmp_path: Path,
):
    class BaseQuery:
        @staticmethod
        def list_aps(*_args, **_kwargs):
            return SimpleNamespace(
                items=[
                    SimpleNamespace(
                        name="轨旁AP-01",
                        mac=PEER_A,
                        station="车站A",
                        section="区间A-B",
                        mileage=SimpleNamespace(raw="K12+300"),
                        line_side="上行",
                    )
                ],
                total=1,
            )

    paths = PathResolver(tmp_path)
    profile = MeshStorageService("demo", paths).create_mr_profile("14CW-04")
    source = tmp_path / "mesh-location.log"
    source.write_text(
        "[1] 2025/12/03 10:12:33.000 (2)\n"
        "[1] Active 30f5-277a-5a2f 2025/12/03 10:12:30 0d 00h 00m 03s 1 "
        "36/43 2%/4% 45%/47% 3/1 15/27 60/72060 88/105 0/5000 2/297 314/0 0/93 0/0 0/0 0/0\n",
        encoding="utf-8",
    )
    MeshImportService("demo", paths).import_files(profile, [source])
    index_path = paths.mesh_mr_db_path("demo", profile.safe_folder_name)
    source_id = int(MeshMrRepository(index_path, read_only=True).list_source_files()[0]["id"])
    session_id = f"{profile.mr_id}:{source_id}"
    query = MeshAnalysisQueryService(paths, base_query=BaseQuery())
    snapshot_rows = query.ap_location_snapshot("demo").to_serializable()

    page_row = query.list_active_build_order("demo", session_id).items[0]
    report = MeshAnalysisReportService(index_path, profile.display_name).build_report(
        MeshReportOptions(source_file_id=source_id, ap_location_snapshot=tuple(snapshot_rows))
    )
    report_row = report.active_build_order[0]

    assert report_row["peer_ap_name"] == ""
    assert page_row.peer_ap_name is None
    assert report_row["peer_ap_mac"] == ""
    assert page_row.peer_ap_mac is None
    assert report_row["station"] == ""
    assert page_row.station is None
    assert report_row["section"] == ""
    assert page_row.section is None
    assert report_row["mileage"] == ""
    assert page_row.mileage is None
    assert report_row["line_side"] == ""
    assert page_row.line_side is None
    assert report_row["identity_status"] == page_row.identity_status == "unresolved"
    for rows in (
        report.link_details,
        report.active_path_rssi,
        report.active_path_busy,
        report.peer_visit_statistics,
    ):
        assert rows[0]["station"] == ""
        assert rows[0]["section"] == ""
        assert rows[0]["mileage"] == ""
        assert rows[0]["line_side"] == ""


def test_excel_report_emits_row_progress_for_large_sheets(tmp_path):
    rows = [{"sample_time": f"2025-12-03 10:00:{index % 60:02d}.000"} for index in range(1001)]
    model = MeshAnalysisReportModel(
        mr_name="14CW-01",
        report_name="14CW-01",
        generated_at=datetime.now(),
        options=MeshReportOptions(),
        raw_evidence=rows,
    )
    stages: list[str] = []

    MeshAnalysisExcelReportExporter().export(model, tmp_path / "report.xlsx", progress=lambda _value, stage: stages.append(stage))

    assert any(stage.startswith("excel_sheet_rows:") and ":1000:1001" in stage for stage in stages)






def test_report_export_translates_internal_enum_values():
    assert translate_report_value("quality_level", "EXCELLENT") == "优秀"
    assert translate_report_value("switch_type", "LATE_SWITCH") == "切换滞后"
    assert translate_report_value("event_type", "NO_BACKUP") == "无可用备份"
    assert translate_report_value("rebuild_type", "DURATION_RESET") == "持续时间回退"
    assert translate_report_value("link_state", "ACTIVE") == "ACTIVE 主链路"
    assert translate_report_value("data_source_type", "MR_RAW_MESH_LOG") == "MR原始MESH日志"
    assert translate_report_value("fping_loss_rate", None) == "N/A"
    assert translate_report_value("related_event_type", "SHORT_SEGMENT_SWITCH") == "短时切换"


def test_quality_sample_point_and_fping_na():
    rules = MeshQualityRules()
    rows = normalize_samples(
        [
            _row("2025-12-03 10:00:00", PEER_A, "ACTIVE", mr_rssi=42),
            _row("2025-12-03 10:00:00", PEER_B, "STANDBY", mr_rssi=35),
        ]
    )
    sample = build_sample_quality(rows, rules)[0]
    assert sample["quality_level"] == "EXCELLENT"
    assert sample["available_backup_count"] == 1
    assert rows[0]["fping_loss_rate"] is None


def test_quality_sample_point_uses_full_link_group_for_backup_counts():
    rules = MeshQualityRules(backup_available_threshold=32, backup_strong_threshold=42, no_backup_min_seconds=0)
    rows = normalize_samples(
        [
            _source_row(1, "2025-12-03 10:00:00.100", PEER_A, "ACTIVE 主链路", mr_rssi=45),
            _source_row(1, "2025-12-03 10:00:00.250", PEER_B, "STANDBY 备链", mr_rssi=35),
            _source_row(1, "2025-12-03 10:00:00.300", PEER_C, "Standy", mr_rssi=None, peer_rssi=43),
            _source_row(2, "2025-12-03 10:00:00.100", PEER_A, "ACTIVE", mr_rssi=45),
        ]
    )

    samples = build_sample_quality(rows, rules)
    source1 = next(sample for sample in samples if sample["source_file_id"] == 1)
    source2 = next(sample for sample in samples if sample["source_file_id"] == 2)
    events = analyze_anomaly_events(samples, rows, rules)

    assert source1["standby_peer_count"] == 2
    assert source1["available_backup_count"] == 2
    assert source1["strong_backup_count"] == 1
    assert source1["best_backup_peer_key"] == PEER_C
    assert source1["best_backup_rssi"] == 43
    assert source1["backup_judgment_reason"] == "Active 存在，已识别 2 条可用备链。"
    assert source2["standby_peer_count"] == 0
    assert source2["available_backup_count"] == 0
    assert not any(event["event_type"] == "NO_BACKUP" and event["source_file"] == "source-1.log" for event in events)


def test_no_backup_diagnosis_distinguishes_missing_and_weak_standby():
    rules = MeshQualityRules(backup_available_threshold=32, no_backup_min_seconds=0)
    missing_rows = normalize_samples([_source_row(1, "2025-12-03 10:00:00", PEER_A, "ACTIVE", mr_rssi=45)])
    weak_rows = normalize_samples(
        [
            _source_row(2, "2025-12-03 10:00:00", PEER_A, "ACTIVE", mr_rssi=45),
            _source_row(2, "2025-12-03 10:00:00", PEER_B, "备链", mr_rssi=25),
        ]
    )

    missing_event = analyze_anomaly_events(build_sample_quality(missing_rows, rules), missing_rows, rules)[0]
    weak_event = analyze_anomaly_events(build_sample_quality(weak_rows, rules), weak_rows, rules)[0]

    assert "没有任何 STANDBY/备链记录" in missing_event["diagnosis"]
    assert "最佳备链 RSSI 25" in weak_event["diagnosis"]
    assert weak_event["standby_peer_count_min"] == 1
    assert weak_event["best_backup_rssi"] == 25


def test_quality_anomaly_event_merging_for_no_backup_weak_no_active_multi_active_and_busy():
    rules = MeshQualityRules(no_backup_min_seconds=1, weak_active_min_seconds=1, busy_warning_threshold=60, busy_bad_threshold=75)
    rows = normalize_samples(
        [
            _row("2025-12-03 10:00:00", PEER_A, "ACTIVE", mr_rssi=28),
            _row("2025-12-03 10:00:01", PEER_A, "ACTIVE", mr_rssi=27),
            _row("2025-12-03 10:00:02", PEER_A, "STANDBY"),
            _row("2025-12-03 10:00:03", PEER_A, "ACTIVE"),
            _row("2025-12-03 10:00:03", PEER_B, "ACTIVE"),
            {**_row("2025-12-03 10:00:04", PEER_A, "ACTIVE"), "local_tx_busy": 80, "local_rx_busy": 10},
        ]
    )
    samples = build_sample_quality(rows, rules)
    events = analyze_anomaly_events(samples, rows, rules)
    event_types = {event["event_type"] for event in events}
    assert {"NO_BACKUP", "WEAK_ACTIVE", "NO_ACTIVE", "MULTI_ACTIVE", "HIGH_BUSY"} <= event_types
    assert len([event for event in events if event["event_type"] == "NO_BACKUP"]) == 1


def test_quality_switch_late_weak_target_and_flap_detection():
    rules = MeshQualityRules(switch_late_window_seconds=5, switch_target_window_seconds=5, flap_window_seconds=30, short_active_segment_seconds=0)
    rows = normalize_samples(
        [
            _row("2025-12-03 10:00:00", PEER_A, "ACTIVE", mr_rssi=24),
            _row("2025-12-03 10:00:00", PEER_B, "STANDBY", mr_rssi=40),
            _row("2025-12-03 10:00:01", PEER_A, "ACTIVE", mr_rssi=23),
            _row("2025-12-03 10:00:01", PEER_B, "STANDBY", mr_rssi=41),
            _row("2025-12-03 10:00:02", PEER_B, "ACTIVE", mr_rssi=24),
            _row("2025-12-03 10:00:03", PEER_A, "ACTIVE", mr_rssi=41),
        ]
    )
    samples = build_sample_quality(rows, rules)
    segments = build_quality_active_segments(samples, rows, rules)
    switches = analyze_switch_events(segments, samples, rules)
    assert any(switch["switch_type"] == "FLAP_SWITCH" for switch in switches)
    late_rows = normalize_samples(
        [
            _row("2025-12-03 10:01:00", PEER_A, "ACTIVE", mr_rssi=24),
            _row("2025-12-03 10:01:00", PEER_B, "STANDBY", mr_rssi=40),
            _row("2025-12-03 10:01:01", PEER_A, "ACTIVE", mr_rssi=23),
            _row("2025-12-03 10:01:01", PEER_B, "STANDBY", mr_rssi=41),
            _row("2025-12-03 10:01:02", PEER_B, "ACTIVE", mr_rssi=24),
        ]
    )
    late_samples = build_sample_quality(late_rows, rules)
    late_segments = build_quality_active_segments(late_samples, late_rows, rules)
    late_switches = analyze_switch_events(late_segments, late_samples, rules)
    assert late_switches[0]["switch_type"] in {"LATE_SWITCH", "WEAK_TARGET_SWITCH"}


def test_quality_switch_with_missing_target_rssi_does_not_raise():
    rules = MeshQualityRules(
        switch_late_window_seconds=5,
        switch_target_window_seconds=5,
        short_active_segment_seconds=0,
    )
    rows = normalize_samples(
        [
            _row("2025-12-03 10:02:00", PEER_A, "ACTIVE", mr_rssi=40),
            _row("2025-12-03 10:02:01", PEER_B, "ACTIVE", mr_rssi=None),
        ]
    )
    samples = build_sample_quality(rows, rules)
    segments = build_quality_active_segments(samples, rows, rules)

    events = analyze_switch_events(segments, samples, rules)

    assert events


def test_quality_switch_long_return_is_not_pingpong_abnormal():
    rules = MeshQualityRules(
        switch_late_window_seconds=5,
        switch_target_window_seconds=5,
        flap_window_seconds=10,
        short_active_segment_seconds=0,
        main_link_switch_time_ms=2500,
        pingpong_tolerance_ms=500,
        pingpong_return_window_ms=10000,
    )
    rows = normalize_samples(
        [
            _row("2025-12-03 10:00:00", PEER_A, "ACTIVE", mr_rssi=40),
            _row("2025-12-03 10:00:01", PEER_B, "ACTIVE", mr_rssi=40),
            _row("2025-12-03 10:00:02", PEER_B, "ACTIVE", mr_rssi=40),
            _row("2025-12-03 10:00:03", PEER_B, "ACTIVE", mr_rssi=40),
            _row("2025-12-03 10:00:04", PEER_B, "ACTIVE", mr_rssi=40),
            _row("2025-12-03 10:00:05", PEER_B, "ACTIVE", mr_rssi=40),
            _row("2025-12-03 10:00:06", PEER_A, "ACTIVE", mr_rssi=40),
        ]
    )
    samples = build_sample_quality(rows, rules)
    segments = build_quality_active_segments(samples, rows, rules)
    switches = analyze_switch_events(segments, samples, rules)

    assert not any(switch["switch_type"] == "FLAP_SWITCH" for switch in switches)
    assert any(switch["pingpong_type"] == "普通回切事件" and not switch["is_pingpong_abnormal"] for switch in switches)


def test_quality_rebuild_busy_and_percentiles():
    assert percentile([10, 20, 30, 40], 0.1) == 13
    assert percentile([10, 20, 30, 40], 0.9) == 37
    rows = normalize_samples(
        [
            {**_row("2025-12-03 10:00:00", PEER_A, "ACTIVE"), "link_count": 1, "duration_seconds": 20, "establish_time": "2025-12-03 09:59:00", "local_tx_busy": 10, "local_rx_busy": 10},
            {**_row("2025-12-03 10:00:01", PEER_A, "ACTIVE"), "link_count": 2, "duration_seconds": 21, "establish_time": "2025-12-03 09:59:00", "local_tx_busy": 80, "local_rx_busy": 20},
            {**_row("2025-12-03 10:00:02", PEER_A, "ACTIVE"), "link_count": 2, "duration_seconds": 1, "establish_time": "2025-12-03 10:00:02", "local_tx_busy": 30, "local_rx_busy": 82},
        ]
    )
    rebuilds = analyze_link_rebuilds(rows)
    assert {event["rebuild_type"] for event in rebuilds} >= {"LINKCNT_INCREASE", "DURATION_RESET"}
    busy = build_busy_analysis(rows, MeshQualityRules(busy_warning_threshold=60, busy_bad_threshold=75))[0]
    assert busy["busy_level"] == "BAD"
