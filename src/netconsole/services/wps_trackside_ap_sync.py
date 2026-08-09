from __future__ import annotations

import ipaddress
import json
import os
import re
import shutil
import time
import urllib.error
import urllib.request
from urllib.parse import urlencode, urlsplit, urlunsplit
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Mapping, Sequence
from uuid import uuid4

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from netconsole.core.database import Database
from netconsole.core.paths import PathResolver
from netconsole.core.sites import SiteManager
from netconsole.core.version import APP_VERSION
from netconsole.models.wps_sync import (
    TRACKSIDE_AP_WPS_BUSINESS_KEY,
    WPS_SYNC_PROTOCOL_VERSION,
    WorkbookFormatRunDTO,
    WorkbookDTO,
    WorkbookSheetDTO,
    WpsFreezeMode,
    WpsSyncMode,
    WpsSyncTarget,
    WpsTargetType,
)
from netconsole.repositories.device_repository import DeviceRepository
from netconsole.repositories.wps_sync_repository import WpsSyncRepository
from netconsole.services.rail_transit.trackside_ap_business_snapshot import (
    canonical_json_bytes,
    content_sha256,
)
from netconsole.services.trackside_ap_export_service import (
    _render_trackside_ap_business_export,
    build_trackside_ap_business_export_snapshot,
)
from netconsole.services.trackside_ap_business import (
    TRACKSIDE_COLUMN_LAYOUT_LIMITS,
    trackside_ap_business_column_layout_types,
    trackside_ap_business_sheet_definition,
)


STANDARD_TARGET_CODE = "wps_standard_spreadsheet"
WPS_SYNC_TASK_TYPE = "trackside_ap_wps_sync"
WPS_SYNC_OWNER = "web_rail_transit"
WPS_STANDARD_FORMAT_MIRROR_ENABLED = True
WPS_SCRIPT_VERSIONS = {
    STANDARD_TARGET_CODE: "2.8.4-standard",
}
WPS_DEPLOYMENT_IDS = {
    STANDARD_TARGET_CODE: "trackside-ap-standard-2.8.4",
}
WPS_RUNTIME_CAPABILITIES = {
    STANDARD_TARGET_CODE: "DEPLOYMENT_PENDING",
}
_VALID_TARGET_CODES = {STANDARD_TARGET_CODE}
_TARGET_TOKEN_ENV = {
    STANDARD_TARGET_CODE: "NETCONSOLE_WPS_STANDARD_AIRSCRIPT_TOKEN",
}
_META_SHEET_NAMES = {"_netconsole_meta", "_NetConsoleSyncMeta", "_NetConsoleSyncRuns"}
_SAFE_ERROR_RE = re.compile(r"(?i)(airscript-token|authorization|token)\s*[:=]\s*\S+")
_MAX_PAYLOAD_BYTES = 20 * 1024 * 1024
_MAX_ERROR_BODY_BYTES = 64 * 1024
_MAX_ERROR_TEXT = 500
_LEGACY_BINDING_ID_RE = re.compile(r"^wst_[0-9a-f]{32}$", re.IGNORECASE)
_WPS_WEBHOOK_PATH_RE = re.compile(
    r"^/api/v3/ide/file/(?P<file_id>[A-Za-z0-9_-]{3,160})/"
    r"script/(?P<script_id>[A-Za-z0-9_-]{3,160})/sync_task$"
)
WPS_REMOTE_TASK_MAX_WAIT_SECONDS = 600.0
WPS_REMOTE_TASK_POLL_INTERVAL_SECONDS = 1.5
_REMOTE_TASK_ID_MAX_LENGTH = 4096
DEFAULT_TARGETS = (
    {
        "target_code": STANDARD_TARGET_CODE,
        "target_type": WpsTargetType.STANDARD_SPREADSHEET,
        "document_open_url": "",
        "webhook_url": "",
        "expected_document_id": "",
    },
)


class WpsSyncError(RuntimeError):
    def __init__(
        self,
        code: str,
        message: str,
        *,
        details: Mapping[str, object] | None = None,
    ) -> None:
        super().__init__(_sanitize_error(message))
        self.code = code
        self.details = _sanitize_details(details or {})


@dataclass(frozen=True)
class WpsHttpResponse:
    status_code: int
    body: object


@dataclass(frozen=True)
class WpsWebhookEndpoints:
    host: str
    file_id: str
    script_id: str
    sync_task_url: str
    async_task_url: str
    task_status_url: str


@dataclass(frozen=True)
class WpsRemoteTask:
    task_id: str
    task_type: str
    status: str = "submitted"


class WpsAirScriptClient:
    def post(
        self,
        target: WpsSyncTarget,
        *,
        token: str,
        argv: Mapping[str, object],
    ) -> WpsHttpResponse:
        if not token:
            raise WpsSyncError(
                "WPS_TOKEN_MISSING",
                "WPS 脚本凭据未配置",
                details=_target_error_details(target, phase="LOCAL_CONFIGURATION"),
            )
        body = json.dumps(
            {"Context": {"argv": dict(argv)}},
            ensure_ascii=False,
            separators=(",", ":"),
        ).encode("utf-8")
        if len(body) > _MAX_PAYLOAD_BYTES:
            raise WpsSyncError(
                "WPS_PAYLOAD_TOO_LARGE",
                f"WPS 同步请求超过 {_MAX_PAYLOAD_BYTES} 字节限制",
            )
        return self._request_json(
            target,
            token=token,
            url=target.webhook_url,
            method="POST",
            body=body,
            connection_error_code="WPS_CONNECTION_FAILED",
            connection_error_prefix="WPS 连接失败",
        )

    def submit_async(
        self,
        target: WpsSyncTarget,
        *,
        token: str,
        argv: Mapping[str, object],
    ) -> WpsRemoteTask:
        if not token:
            raise WpsSyncError(
                "WPS_TOKEN_MISSING",
                "WPS 脚本凭据未配置",
                details=_target_error_details(target, phase="LOCAL_CONFIGURATION"),
            )
        endpoints = parse_wps_webhook(target.webhook_url)
        body = json.dumps(
            {"Context": {"argv": dict(argv)}},
            ensure_ascii=False,
            separators=(",", ":"),
        ).encode("utf-8")
        if len(body) > _MAX_PAYLOAD_BYTES:
            raise WpsSyncError(
                "WPS_PAYLOAD_TOO_LARGE",
                f"WPS 同步请求超过 {_MAX_PAYLOAD_BYTES} 字节限制",
            )
        response = self._request_json(
            target,
            token=token,
            url=endpoints.async_task_url,
            method="POST",
            body=body,
            connection_error_code="ASYNC_SUBMIT_FAILED",
            connection_error_prefix="WPS 异步任务提交失败",
        )
        payload = response.body
        if not isinstance(payload, Mapping):
            raise WpsSyncError(
                "WPS_ASYNC_SUBMIT_RESPONSE_INVALID",
                "WPS 异步任务提交返回结构无效",
                details=_target_error_details(target, phase="ASYNC_SUBMIT"),
            )
        nested = payload.get("data")
        nested_data = nested if isinstance(nested, Mapping) else {}
        task_id = str(payload.get("task_id") or nested_data.get("task_id") or "").strip()
        task_type = str(payload.get("task_type") or nested_data.get("task_type") or "").strip()
        if not task_id or len(task_id) > _REMOTE_TASK_ID_MAX_LENGTH:
            raise WpsSyncError(
                "WPS_ASYNC_TASK_ID_MISSING",
                "WPS 异步任务提交未返回有效 task_id",
                details=_target_error_details(target, phase="ASYNC_SUBMIT"),
            )
        return WpsRemoteTask(task_id=task_id, task_type=task_type or "open_air_script")

    def poll_async_task(
        self,
        target: WpsSyncTarget,
        *,
        token: str,
        task_id: str,
    ) -> WpsHttpResponse:
        if not token:
            raise WpsSyncError(
                "WPS_TOKEN_MISSING",
                "WPS 脚本凭据未配置",
                details=_target_error_details(target, phase="LOCAL_CONFIGURATION"),
            )
        normalized_task_id = str(task_id or "").strip()
        if not normalized_task_id or len(normalized_task_id) > _REMOTE_TASK_ID_MAX_LENGTH:
            raise WpsSyncError("WPS_ASYNC_TASK_ID_INVALID", "WPS 远端 task_id 无效")
        endpoints = parse_wps_webhook(target.webhook_url)
        query = urlencode({"task_id": normalized_task_id})
        return self._request_json(
            target,
            token=token,
            url=f"{endpoints.task_status_url}?{query}",
            method="GET",
            body=None,
            connection_error_code="REMOTE_POLL_TEMPORARY_FAILED",
            connection_error_prefix="WPS 远端任务查询暂时失败",
        )

    @staticmethod
    def _request_json(
        target: WpsSyncTarget,
        *,
        token: str,
        url: str,
        method: str,
        body: bytes | None,
        connection_error_code: str,
        connection_error_prefix: str,
    ) -> WpsHttpResponse:
        request = urllib.request.Request(
            url,
            data=body,
            method=method,
            headers={
                "Content-Type": "application/json",
                "Accept": "application/json",
                "AirScript-Token": token,
                "User-Agent": f"NetConsole/{APP_VERSION}",
            },
        )
        try:
            with urllib.request.urlopen(
                request,
                timeout=target.timeout_seconds,
            ) as response:
                raw = response.read(_MAX_PAYLOAD_BYTES + 1)
                status_code = int(response.status)
        except urllib.error.HTTPError as exc:
            raise _http_error_from_response(exc, target, token) from None
        except (urllib.error.URLError, TimeoutError, OSError) as exc:
            raise WpsSyncError(
                connection_error_code,
                f"{connection_error_prefix}：{_sanitize_error(str(exc))}",
                details={
                    **_target_error_details(
                        target,
                        phase=(
                            "ASYNC_SUBMIT"
                            if connection_error_code == "ASYNC_SUBMIT_FAILED"
                            else "REMOTE_POLL"
                            if connection_error_code == "REMOTE_POLL_TEMPORARY_FAILED"
                            else "HTTP_REQUEST"
                        ),
                    ),
                    "submission_outcome": (
                        "UNKNOWN" if connection_error_code == "ASYNC_SUBMIT_FAILED" else ""
                    ),
                },
            ) from None
        if len(raw) > _MAX_PAYLOAD_BYTES:
            raise WpsSyncError("WPS_RESPONSE_TOO_LARGE", "WPS 返回内容超过安全限制")
        try:
            decoded = json.loads(raw.decode("utf-8"))
        except (UnicodeDecodeError, json.JSONDecodeError):
            raise WpsSyncError(
                "WPS_RESPONSE_INVALID",
                "WPS 返回了非 JSON 内容",
                details=_target_error_details(target, phase="SCRIPT_EXECUTION"),
            ) from None
        if not isinstance(decoded, dict):
            raise WpsSyncError(
                "WPS_RESPONSE_INVALID",
                "WPS 返回结构无效",
                details=_target_error_details(target, phase="SCRIPT_EXECUTION"),
            )
        return WpsHttpResponse(status_code=status_code, body=decoded)


class BaseWpsAdapter:
    def __init__(self, client: WpsAirScriptClient) -> None:
        self.client = client

    def connection_test(self, target: WpsSyncTarget, token: str) -> dict[str, Any]:
        response = self.client.post(
            target,
            token=token,
            argv={
                "protocol_version": WPS_SYNC_PROTOCOL_VERSION,
                "operation": "connection_test",
                "target_code": target.target_code,
                "target_type": target.target_type.value,
                "site_id": target.site_id,
                "business_key": target.business_key,
                "binding_id": target.binding_id,
                "document_id": target.expected_document_id,
                "script_id": _script_id_from_webhook(target.webhook_url),
            },
        )
        result = _unwrap_wps_sync_task_response(response, target, token=token)
        self._validate_common(target, result)
        if not bool(result.get("success")):
            raise _remote_result_error(target, result, token)
        return {
            "http_status": response.status_code,
            "phase": "SUCCESS",
            **_with_binding_diagnostics(target, result),
        }

    def runtime_write_probe(self, target: WpsSyncTarget, token: str) -> dict[str, Any]:
        response = self.client.post(
            target,
            token=token,
            argv={
                "protocol_version": WPS_SYNC_PROTOCOL_VERSION,
                "operation": "runtime_write_probe",
                "target_code": target.target_code,
                "target_type": target.target_type.value,
                "site_id": target.site_id,
                "business_key": target.business_key,
                "binding_id": target.binding_id,
                "probe_id": f"wps_probe_{uuid4().hex}",
                "script_id": _script_id_from_webhook(target.webhook_url),
            },
        )
        result = _unwrap_wps_sync_task_response(response, target, token=token)
        self._validate_common(target, result)
        if not bool(result.get("success")):
            raise _remote_result_error(target, result, token)
        if str(result.get("runtime_capability") or "") != "VERIFIED":
            raise WpsSyncError(
                "WPS_RUNTIME_PROBE_UNVERIFIED",
                "WPS 运行时写入探针未完成验证",
                details=_target_error_details(target, phase="RUNTIME_WRITE_PROBE"),
            )
        return {"http_status": response.status_code, "phase": "SUCCESS", **result}

    def sync_test_sheet(self, target: WpsSyncTarget, token: str) -> dict[str, Any]:
        response = self.client.post(
            target,
            token=token,
            argv={
                "protocol_version": WPS_SYNC_PROTOCOL_VERSION,
                "operation": "sync_test_sheet",
                "target_code": target.target_code,
                "target_type": target.target_type.value,
                "site_id": target.site_id,
                "business_key": target.business_key,
                "binding_id": target.binding_id,
                "probe_id": f"wps_sync_test_{uuid4().hex}",
                "script_id": _script_id_from_webhook(target.webhook_url),
            },
        )
        result = _unwrap_wps_sync_task_response(response, target, token=token)
        self._validate_common(target, result)
        if not bool(result.get("success")):
            raise _remote_result_error(target, result, token)
        return {"http_status": response.status_code, "phase": "SUCCESS", **result}

    def sheet_tab_color_probe(
        self,
        target: WpsSyncTarget,
        token: str,
    ) -> dict[str, Any]:
        response = self.client.post(
            target,
            token=token,
            argv={
                "protocol_version": WPS_SYNC_PROTOCOL_VERSION,
                "operation": "sheet_tab_color_probe",
                "target_code": target.target_code,
                "target_type": target.target_type.value,
                "site_id": target.site_id,
                "business_key": target.business_key,
                "binding_id": target.binding_id,
                "probe_id": f"wps_sheet_tab_color_{uuid4().hex}",
                "script_id": _script_id_from_webhook(target.webhook_url),
            },
        )
        result = _unwrap_wps_sync_task_response(response, target, token=token)
        self._validate_common(target, result)
        if not bool(result.get("success")):
            raise _remote_result_error(target, result, token)
        if not bool(result.get("sheet_tab_color_verified")):
            raise WpsSyncError(
                "WPS_SHEET_TAB_COLOR_VERIFY_FAILED",
                "WPS Sheet 标签颜色探针未通过读回校验",
                details={
                    **_target_error_details(target, phase="SHEET_TAB_COLOR_PROBE"),
                    "expected_tab_color": result.get("expected_tab_color") or "",
                    "actual_tab_color": result.get("actual_tab_color") or "",
                },
            )
        return {"http_status": response.status_code, "phase": "SUCCESS", **result}

    def column_width_probe(
        self,
        target: WpsSyncTarget,
        token: str,
    ) -> dict[str, Any]:
        response = self.client.post(
            target,
            token=token,
            argv={
                "protocol_version": WPS_SYNC_PROTOCOL_VERSION,
                "operation": "column_width_probe",
                "target_code": target.target_code,
                "target_type": target.target_type.value,
                "site_id": target.site_id,
                "business_key": target.business_key,
                "binding_id": target.binding_id,
                "probe_id": f"wps_column_width_{uuid4().hex}",
                "script_id": _script_id_from_webhook(target.webhook_url),
            },
        )
        result = _unwrap_wps_sync_task_response(response, target, token=token)
        self._validate_common(target, result)
        if not bool(result.get("success")):
            raise _remote_result_error(target, result, token)
        if not bool(result.get("column_width_verified")):
            raise WpsSyncError(
                "WPS_COLUMN_WIDTH_VERIFY_FAILED",
                "WPS 列宽探针未通过写后读回校验",
                details={
                    **_target_error_details(target, phase="COLUMN_WIDTH_PROBE"),
                    "expected_column_widths": result.get("expected_column_widths") or {},
                    "actual_column_widths": result.get("actual_column_widths") or {},
                },
            )
        return {"http_status": response.status_code, "phase": "SUCCESS", **result}

    def migrate_legacy_binding(
        self,
        target: WpsSyncTarget,
        token: str,
        *,
        expected_old_binding_id: str,
    ) -> dict[str, Any]:
        response = self.client.post(
            target,
            token=token,
            argv={
                "protocol_version": WPS_SYNC_PROTOCOL_VERSION,
                "operation": "migrate_legacy_binding",
                "document_id": target.expected_document_id,
                "target_code": target.target_code,
                "target_type": target.target_type.value,
                "site_id": target.site_id,
                "business_key": target.business_key,
                "expected_old_binding_id": expected_old_binding_id,
                "new_binding_id": target.binding_id,
                "script_id": _script_id_from_webhook(target.webhook_url),
            },
        )
        result = _unwrap_wps_sync_task_response(response, target, token=token)
        self._validate_common(target, result)
        if not bool(result.get("success")):
            raise _remote_result_error(target, result, token)
        diagnosed = _with_binding_diagnostics(target, result)
        migrated = bool(diagnosed.get("migrated"))
        already_migrated = bool(diagnosed.get("already_migrated"))
        previous_binding_id = str(diagnosed.get("previous_binding_id") or "")
        if (
            str(diagnosed.get("binding_status") or "") != "BOUND"
            or not bool(diagnosed.get("binding_id_match"))
            or (migrated and previous_binding_id != expected_old_binding_id)
            or not (migrated or already_migrated)
        ):
            raise WpsSyncError(
                "WPS_BINDING_MIGRATION_VERIFY_FAILED",
                "WPS 旧版绑定标识迁移结果未通过身份校验",
                details={
                    **_target_error_details(target, phase="DOCUMENT_IDENTITY"),
                    **_binding_error_details(diagnosed),
                    "expected_old_binding_id": expected_old_binding_id,
                    "previous_binding_id": previous_binding_id,
                    "migrated": migrated,
                    "already_migrated": already_migrated,
                },
            )
        return {"http_status": response.status_code, "phase": "SUCCESS", **diagnosed}

    def sync(
        self,
        target: WpsSyncTarget,
        token: str,
        payload: Mapping[str, object],
    ) -> dict[str, Any]:
        request_payload = dict(payload)
        request_payload.setdefault("script_id", _script_id_from_webhook(target.webhook_url))
        response = self.client.post(target, token=token, argv=request_payload)
        return self.validate_sync_response(target, token, payload, response)

    def validate_sync_response(
        self,
        target: WpsSyncTarget,
        token: str,
        payload: Mapping[str, object],
        response: WpsHttpResponse,
    ) -> dict[str, Any]:
        result = _unwrap_wps_sync_task_response(response, target, token=token)
        self._validate_common(target, result)
        if not bool(result.get("success")):
            raise _remote_result_error(target, result, token)
        mismatched = []
        for key in (
            "target_batch_id",
            "site_id",
            "business_key",
            "snapshot_revision",
            "snapshot_sha256",
        ):
            if str(result.get(key) or "") != str(payload.get(key) or ""):
                mismatched.append(key)
        if mismatched:
            details = {
                **_target_error_details(target, phase="DOCUMENT_IDENTITY"),
                "mismatched_fields": mismatched,
                "expected_target_batch_id": str(payload.get("target_batch_id") or ""),
                "remote_target_batch_id": str(result.get("target_batch_id") or ""),
                "expected_site_id": str(payload.get("site_id") or ""),
                "remote_site_id": str(result.get("site_id") or ""),
                "expected_business_key": str(payload.get("business_key") or ""),
                "remote_business_key": str(result.get("business_key") or ""),
                "expected_revision": str(payload.get("snapshot_revision") or ""),
                "remote_revision": str(result.get("snapshot_revision") or ""),
                "expected_snapshot_sha256": str(payload.get("snapshot_sha256") or ""),
                "remote_snapshot_sha256": str(result.get("snapshot_sha256") or ""),
            }
            raise WpsSyncError(
                "WPS_RESPONSE_IDENTITY_MISMATCH",
                f"WPS 返回的 {mismatched[0]} 与本次同步不一致",
                details=details,
            )
        return {"http_status": response.status_code, "phase": "SUCCESS", **result}

    @staticmethod
    def _validate_common(target: WpsSyncTarget, result: Mapping[str, object]) -> None:
        if int(result.get("protocol_version") or 0) != WPS_SYNC_PROTOCOL_VERSION:
            raise WpsSyncError(
                "WPS_PROTOCOL_MISMATCH",
                "WPS 脚本协议版本不兼容",
                details=_target_error_details(target, phase="PROTOCOL_HANDSHAKE"),
            )
        if str(result.get("target_type") or "") != target.target_type.value:
            raise WpsSyncError(
                "WPS_TARGET_TYPE_MISMATCH",
                "WPS 目标类型不匹配",
                details=_target_error_details(target, phase="PROTOCOL_HANDSHAKE"),
            )
        if str(result.get("document_id") or "") != target.expected_document_id:
            raise WpsSyncError(
                "WPS_DOCUMENT_ID_MISMATCH",
                "WPS 文档身份校验失败",
                details=_target_error_details(target, phase="DOCUMENT_IDENTITY"),
            )
        # Older local test doubles did not expose the deployment identity. Keep
        # them readable, but validate every identity field returned by a real
        # AirScript deployment so stale or cross-target webhooks fail closed.
        expected_script_version = WPS_SCRIPT_VERSIONS.get(target.target_code, "")
        returned_script_version = result.get("script_version")
        identity_required = "runtime_capability" in result
        if identity_required and not str(returned_script_version or ""):
            raise WpsSyncError(
                "WPS_SCRIPT_VERSION_MISSING",
                "WPS 脚本未返回脚本版本",
                details=_target_error_details(target, phase="PROTOCOL_HANDSHAKE"),
            )
        if (
            returned_script_version is not None
            and str(returned_script_version) not in {expected_script_version, "test"}
        ):
            raise WpsSyncError(
                "WPS_SCRIPT_VERSION_MISMATCH",
                "WPS 脚本版本与本地期望不一致",
                details={
                    **_target_error_details(target, phase="PROTOCOL_HANDSHAKE"),
                    "expected_script_version": expected_script_version,
                    "remote_script_version": str(returned_script_version),
                },
            )
        expected_deployment_id = WPS_DEPLOYMENT_IDS.get(target.target_code, "")
        returned_deployment_id = result.get("deployment_id")
        if identity_required and not str(returned_deployment_id or ""):
            raise WpsSyncError(
                "WPS_DEPLOYMENT_ID_MISSING",
                "WPS 脚本未返回部署 ID",
                details=_target_error_details(target, phase="PROTOCOL_HANDSHAKE"),
            )
        if returned_deployment_id is not None and str(returned_deployment_id) != expected_deployment_id:
            raise WpsSyncError(
                "WPS_DEPLOYMENT_ID_MISMATCH",
                "WPS 脚本部署身份与本地期望不一致",
                details={
                    **_target_error_details(target, phase="PROTOCOL_HANDSHAKE"),
                    "expected_deployment_id": expected_deployment_id,
                    "remote_deployment_id": str(returned_deployment_id),
                },
            )
        expected_script_id = _script_id_from_webhook(target.webhook_url)
        returned_script_id = result.get("script_id")
        if (
            identity_required
            and not str(returned_script_id or "")
            and str(result.get("target_code") or "") == target.target_code
        ):
            raise WpsSyncError(
                "WPS_SCRIPT_ID_MISSING",
                "WPS 脚本未返回脚本 ID",
                details=_target_error_details(target, phase="PROTOCOL_HANDSHAKE"),
            )
        if returned_script_id is not None and str(returned_script_id) not in {expected_script_id, "test"}:
            raise WpsSyncError(
                "WPS_SCRIPT_ID_MISMATCH",
                "WPS 脚本 ID 与 webhook 配置不一致",
                details={
                    **_target_error_details(target, phase="PROTOCOL_HANDSHAKE"),
                    "expected_script_id": expected_script_id,
                    "remote_script_id": str(returned_script_id),
                },
            )
        returned_target_code = result.get("target_code")
        if identity_required and not str(returned_target_code or ""):
            raise WpsSyncError(
                "WPS_TARGET_CODE_MISSING",
                "WPS 脚本未返回目标代码",
                details=_target_error_details(target, phase="PROTOCOL_HANDSHAKE"),
            )
        if returned_target_code is not None and str(returned_target_code) != target.target_code:
            raise WpsSyncError(
                "WPS_TARGET_CODE_MISMATCH",
                "WPS 脚本目标代码与本地目标不一致",
                details={
                    **_target_error_details(target, phase="PROTOCOL_HANDSHAKE"),
                    "expected_target_code": target.target_code,
                    "remote_target_code": str(returned_target_code),
                },
            )


class WpsStandardSpreadsheetAdapter(BaseWpsAdapter):
    pass


class TracksideApWpsSyncService:
    def __init__(
        self,
        paths: PathResolver,
        *,
        client: WpsAirScriptClient | None = None,
        remote_task_max_wait_seconds: float = WPS_REMOTE_TASK_MAX_WAIT_SECONDS,
        remote_task_poll_interval_seconds: float = WPS_REMOTE_TASK_POLL_INTERVAL_SECONDS,
        sleep: Callable[[float], None] = time.sleep,
        monotonic: Callable[[], float] = time.monotonic,
    ) -> None:
        self.paths = paths
        self.client = client or WpsAirScriptClient()
        self.remote_task_max_wait_seconds = max(1.0, float(remote_task_max_wait_seconds))
        self.remote_task_poll_interval_seconds = max(
            0.0, float(remote_task_poll_interval_seconds)
        )
        self._sleep = sleep
        self._monotonic = monotonic
        self.adapters = {
            WpsTargetType.STANDARD_SPREADSHEET: WpsStandardSpreadsheetAdapter(
                self.client
            ),
        }

    def list_targets(self, site_id: str) -> list[dict[str, Any]]:
        repository = self._repository(site_id)
        self._ensure_default_targets(repository)
        return [
            self._public_target(
                target,
                env_token_configured=self._env_token(target.target_code) != "",
            )
            for target in repository.list_targets(TRACKSIDE_AP_WPS_BUSINESS_KEY)
            if target.target_code in _VALID_TARGET_CODES
        ]

    def configure_target(
        self,
        site_id: str,
        target_code: str,
        *,
        token: str | None = None,
        document_open_url: str | None = None,
        webhook_url: str | None = None,
        enabled: bool | None = None,
        timeout_seconds: int | None = None,
    ) -> dict[str, Any]:
        _require_standard_target_code(target_code)
        repository = self._repository(site_id)
        self._ensure_default_targets(repository)
        target = repository.get_target(TRACKSIDE_AP_WPS_BUSINESS_KEY, target_code)
        selected_document_url = (
            target.document_open_url
            if document_open_url is None
            else _validate_wps_url(document_open_url, kind="document")
        )
        selected_webhook_url = (
            target.webhook_url
            if webhook_url is None
            else _validate_wps_url(webhook_url, kind="webhook")
        )
        selected_document_id = (
            target.expected_document_id
            if webhook_url is None
            else _document_id_from_webhook(selected_webhook_url)
        )
        selected_enabled = target.enabled if enabled is None else enabled
        configured = repository.upsert_target(
            business_key=target.business_key,
            target_code=target.target_code,
            target_type=target.target_type,
            target_name=target.target_name,
            document_open_url=selected_document_url,
            webhook_url=selected_webhook_url,
            expected_document_id=selected_document_id,
            enabled=selected_enabled,
            timeout_seconds=(
                target.timeout_seconds
                if timeout_seconds is None
                else timeout_seconds
            ),
            token=token,
            credential_id=target.credential_id,
        )
        connection_identity_changed = any(
            (
                target.document_open_url != configured.document_open_url,
                target.webhook_url != configured.webhook_url,
                target.expected_document_id != configured.expected_document_id,
            )
        )
        if connection_identity_changed:
            repository.clear_runtime_probe_identity(configured.target_id)
            configured = repository.get_target(
                configured.business_key, configured.target_code
            )
        return self._public_target(
            configured,
            env_token_configured=self._env_token(configured.target_code) != "",
        )

    def connection_test(self, site_id: str, target_code: str) -> dict[str, Any]:
        _require_standard_target_code(target_code)
        repository = self._repository(site_id)
        self._ensure_default_targets(repository)
        target = repository.get_target(TRACKSIDE_AP_WPS_BUSINESS_KEY, target_code)
        try:
            _validate_target_configuration(target)
            result = self.adapters[target.target_type].connection_test(
                target,
                self._token(repository, target),
            )
        except WpsSyncError as exc:
            if not exc.details:
                exc.details.update(_target_error_details(target, phase="LOCAL_CONFIGURATION"))
            repository.record_connection_test(
                target.target_id,
                result=None,
                diagnostic=_operation_diagnostic(
                    "connection_test",
                    status="FAILED",
                    message=str(exc),
                    values=exc.details,
                ),
            )
            raise
        repository.record_connection_test(
            target.target_id,
            result=result,
            diagnostic=_operation_diagnostic(
                "connection_test",
                status="SUCCESS",
                message=str(result.get("message") or "连接测试通过"),
                values=result,
            ),
        )
        return _sanitize_result(result)

    def migrate_legacy_binding(self, site_id: str, target_code: str) -> dict[str, Any]:
        _require_standard_target_code(target_code)
        repository = self._repository(site_id)
        self._ensure_default_targets(repository)
        target = repository.get_target(TRACKSIDE_AP_WPS_BUSINESS_KEY, target_code)
        if target.target_type is not WpsTargetType.STANDARD_SPREADSHEET:
            raise WpsSyncError(
                "WPS_BINDING_MIGRATION_UNSUPPORTED",
                "当前仅支持普通在线表格迁移旧版绑定标识",
            )
        _validate_target_configuration(target)
        preflight = self.connection_test(site_id, target_code)
        target = repository.get_target(TRACKSIDE_AP_WPS_BUSINESS_KEY, target_code)
        binding_status = str(preflight.get("binding_status") or "UNKNOWN").upper()
        remote_binding_id = str(preflight.get("remote_binding_id") or "")
        already_migrated = (
            binding_status == "BOUND" and remote_binding_id == target.binding_id
        )
        if binding_status != "LEGACY_BINDING_ID_MISMATCH" and not already_migrated:
            raise WpsSyncError(
                "WPS_BINDING_MIGRATION_NOT_ALLOWED",
                "只有业务身份完全一致的旧版绑定标识可以迁移",
                details={
                    **_target_error_details(target, phase="BINDING_GATE"),
                    "binding_status": binding_status,
                    "local_binding_id": target.binding_id,
                    "remote_binding_id": remote_binding_id,
                    **_binding_error_details(preflight),
                },
            )
        result = self.adapters[target.target_type].migrate_legacy_binding(
            target,
            self._token(repository, target),
            expected_old_binding_id=remote_binding_id,
        )
        repository.update_target_remote_state(
            target.target_id,
            binding_status="BOUND",
            result=result,
            persist_runtime_identity=False,
        )
        try:
            verification = self.revalidate_deployment(site_id, target_code)
        except WpsSyncError as exc:
            raise WpsSyncError(
                "WPS_BINDING_MIGRATED_VERIFICATION_FAILED",
                "远端旧版绑定标识已迁移，但后续部署验证未全部通过",
                details={
                    **_target_error_details(target, phase="POST_MIGRATION_VERIFICATION"),
                    "migrated": bool(result.get("migrated")),
                    "already_migrated": bool(result.get("already_migrated")),
                    "previous_binding_id": str(
                        result.get("previous_binding_id") or remote_binding_id
                    ),
                    "verification_error_code": exc.code,
                    "verification_error_message": _sanitize_error(str(exc)),
                    "verification_error_details": _sanitize_result(exc.details),
                },
            ) from exc
        connection = verification["connection_test"]
        return {
            **connection,
            "message": str(
                result.get("message") or "旧版绑定标识已迁移并完成部署验证"
            ),
            "migrated": bool(result.get("migrated")),
            "already_migrated": bool(result.get("already_migrated")),
            "previous_binding_id": str(
                result.get("previous_binding_id") or remote_binding_id
            ),
            "verification": verification,
        }

    def revalidate_deployment(self, site_id: str, target_code: str) -> dict[str, Any]:
        """Re-run the complete ordinary-spreadsheet deployment verification chain."""
        _require_standard_target_code(target_code)
        repository = self._repository(site_id)
        self._ensure_default_targets(repository)
        target = repository.get_target(TRACKSIDE_AP_WPS_BUSINESS_KEY, target_code)
        if target.target_type is not WpsTargetType.STANDARD_SPREADSHEET:
            raise WpsSyncError(
                "WPS_REVALIDATE_UNSUPPORTED",
                "当前仅支持普通在线表格重新验证部署",
            )
        repository.clear_runtime_probe_identity(target.target_id)
        try:
            connection = self.connection_test(site_id, target_code)
            probe = self.runtime_write_probe(site_id, target_code)
            sync_test = self.sync_test_sheet(site_id, target_code)
        except WpsSyncError:
            repository.set_runtime_capability(target.target_id, "DEPLOYMENT_PENDING")
            raise
        return {
            "target_code": target_code,
            "connection_test": connection,
            "runtime_write_probe": probe,
            "sync_test_sheet": sync_test,
            "runtime_capability": "VERIFIED",
        }

    def runtime_write_probe(self, site_id: str, target_code: str) -> dict[str, Any]:
        _require_standard_target_code(target_code)
        repository = self._repository(site_id)
        self._ensure_default_targets(repository)
        target = repository.get_target(TRACKSIDE_AP_WPS_BUSINESS_KEY, target_code)
        try:
            _validate_target_configuration(target)
            result = self.adapters[target.target_type].runtime_write_probe(
                target, self._token(repository, target)
            )
        except WpsSyncError as exc:
            details = dict(exc.details) or _target_error_details(
                target, phase="LOCAL_CONFIGURATION"
            )
            repository.update_target_diagnostic(
                target.target_id,
                operation="runtime_write_probe",
                diagnostic=_operation_diagnostic(
                    "runtime_write_probe",
                    status="FAILED",
                    message=str(exc),
                    values=details,
                ),
            )
            raise
        repository.update_target_diagnostic(
            target.target_id,
            operation="runtime_write_probe",
            diagnostic=_operation_diagnostic(
                "runtime_write_probe",
                status=str(result.get("status") or "SUCCESS"),
                message=str(result.get("message") or "运行时写入探针通过"),
                values=result,
            ),
        )
        repository.update_target_remote_state(
            target.target_id,
            binding_status=str(result.get("binding_status") or "UNKNOWN"),
            result=result,
            runtime_capability="VERIFIED",
        )
        return _sanitize_result(result)

    def sync_test_sheet(self, site_id: str, target_code: str) -> dict[str, Any]:
        _require_standard_target_code(target_code)
        repository = self._repository(site_id)
        self._ensure_default_targets(repository)
        target = repository.get_target(TRACKSIDE_AP_WPS_BUSINESS_KEY, target_code)
        if target.target_type is not WpsTargetType.STANDARD_SPREADSHEET:
            raise WpsSyncError("WPS_SYNC_TEST_UNSUPPORTED", "同步测试 Sheet 仅支持普通在线表格")
        try:
            _validate_target_configuration(target)
            result = self.adapters[target.target_type].sync_test_sheet(
                target, self._token(repository, target)
            )
        except WpsSyncError as exc:
            details = dict(exc.details) or _target_error_details(
                target, phase="LOCAL_CONFIGURATION"
            )
            repository.update_target_diagnostic(
                target.target_id,
                operation="sync_test_sheet",
                diagnostic=_operation_diagnostic(
                    "sync_test_sheet",
                    status="FAILED",
                    message=str(exc),
                    values=details,
                ),
            )
            raise
        repository.update_target_diagnostic(
            target.target_id,
            operation="sync_test_sheet",
            diagnostic=_operation_diagnostic(
                "sync_test_sheet",
                status="SUCCESS",
                message=str(result.get("message") or "同步测试 Sheet 通过"),
                values=result,
            ),
        )
        repository.update_target_remote_state(
            target.target_id,
            binding_status=str(result.get("binding_status") or "UNKNOWN"),
            result=result,
            persist_runtime_identity=False,
        )
        return _sanitize_result(result)

    def sheet_tab_color_probe(self, site_id: str, target_code: str) -> dict[str, Any]:
        _require_standard_target_code(target_code)
        repository = self._repository(site_id)
        self._ensure_default_targets(repository)
        target = repository.get_target(TRACKSIDE_AP_WPS_BUSINESS_KEY, target_code)
        if target.target_type is not WpsTargetType.STANDARD_SPREADSHEET:
            raise WpsSyncError(
                "WPS_SHEET_TAB_COLOR_PROBE_UNSUPPORTED",
                "Sheet 标签颜色探针仅支持普通在线表格",
            )
        try:
            _validate_target_configuration(target)
            result = self.adapters[target.target_type].sheet_tab_color_probe(
                target,
                self._token(repository, target),
            )
        except WpsSyncError as exc:
            details = dict(exc.details) or _target_error_details(
                target,
                phase="LOCAL_CONFIGURATION",
            )
            repository.update_target_diagnostic(
                target.target_id,
                operation="sheet_tab_color_probe",
                diagnostic=_operation_diagnostic(
                    "sheet_tab_color_probe",
                    status="FAILED",
                    message=str(exc),
                    values=details,
                ),
            )
            raise
        repository.update_target_diagnostic(
            target.target_id,
            operation="sheet_tab_color_probe",
            diagnostic=_operation_diagnostic(
                "sheet_tab_color_probe",
                status=str(result.get("status") or "SUCCESS"),
                message=str(result.get("message") or "Sheet 标签颜色探针通过"),
                values=result,
            ),
        )
        return _sanitize_result(result)

    def column_width_probe(self, site_id: str, target_code: str) -> dict[str, Any]:
        _require_standard_target_code(target_code)
        repository = self._repository(site_id)
        self._ensure_default_targets(repository)
        target = repository.get_target(TRACKSIDE_AP_WPS_BUSINESS_KEY, target_code)
        if target.target_type is not WpsTargetType.STANDARD_SPREADSHEET:
            raise WpsSyncError(
                "WPS_COLUMN_WIDTH_PROBE_UNSUPPORTED",
                "列宽探针仅支持普通在线表格",
            )
        try:
            _validate_target_configuration(target)
            result = self.adapters[target.target_type].column_width_probe(
                target,
                self._token(repository, target),
            )
        except WpsSyncError as exc:
            details = dict(exc.details) or _target_error_details(
                target,
                phase="LOCAL_CONFIGURATION",
            )
            repository.update_target_diagnostic(
                target.target_id,
                operation="column_width_probe",
                diagnostic=_operation_diagnostic(
                    "column_width_probe",
                    status="FAILED",
                    message=str(exc),
                    values=details,
                ),
            )
            raise
        repository.update_target_diagnostic(
            target.target_id,
            operation="column_width_probe",
            diagnostic=_operation_diagnostic(
                "column_width_probe",
                status=str(result.get("status") or "SUCCESS"),
                message=str(result.get("message") or "列宽探针通过"),
                values=result,
            ),
        )
        return _sanitize_result(result)

    def sync(
        self,
        site_id: str,
        *,
        target_codes: Sequence[str] = (),
        expected_revision: str = "",
        initialize_binding: bool = False,
        progress: Callable[[str, int, int, str], None] | None = None,
        should_cancel: Callable[[], None] | None = None,
    ) -> dict[str, Any]:
        repository = self._repository(site_id)
        self._ensure_default_targets(repository)
        requested_codes = tuple(
            dict.fromkeys(target_codes or (STANDARD_TARGET_CODE,))
        )
        if not requested_codes or any(code not in _VALID_TARGET_CODES for code in requested_codes):
            raise WpsSyncError(
                "WPS_TARGET_UNSUPPORTED",
                "当前版本仅支持 WPS 普通在线表格同步",
            )
        targets_by_code = {
            target.target_code: target
            for target in repository.list_targets(TRACKSIDE_AP_WPS_BUSINESS_KEY)
        }
        targets = [targets_by_code[code] for code in requested_codes]
        disabled = [target.target_name for target in targets if not target.enabled]
        if disabled:
            raise WpsSyncError(
                "WPS_TARGET_DISABLED",
                f"WPS 目标未启用：{', '.join(disabled)}",
            )
        for target in targets:
            if not isinstance(self.client, WpsAirScriptClient):
                continue
            binding_status = _effective_binding_status(target)
            if binding_status == "MISMATCH":
                raise WpsSyncError(
                    "WPS_DOCUMENT_BINDING_MISMATCH",
                    f"WPS 文档绑定与当前局点不一致：{target.target_name}",
                    details={
                        **_target_error_details(target, phase="BINDING_GATE"),
                        "binding_status": binding_status,
                        "remote_site_id": target.remote_site_id,
                        "remote_business_key": target.remote_business_key,
                    },
                )
            if binding_status == "LEGACY_BINDING_ID_MISMATCH":
                raise WpsSyncError(
                    "WPS_LEGACY_BINDING_ID_MISMATCH",
                    f"WPS 文档仍使用旧版绑定标识，请先升级：{target.target_name}",
                    details={
                        **_target_error_details(target, phase="BINDING_GATE"),
                        "binding_status": binding_status,
                        "local_binding_id": target.binding_id,
                        "remote_binding_id": target.remote_binding_id,
                    },
                )
            if binding_status == "UNKNOWN":
                raise WpsSyncError(
                    "WPS_BINDING_STATUS_UNKNOWN",
                    f"WPS 文档绑定状态未知，请先执行连接测试：{target.target_name}",
                    details=_target_error_details(target, phase="BINDING_GATE"),
                )
            if binding_status == "UNBOUND" and not initialize_binding:
                raise WpsSyncError(
                    "WPS_DOCUMENT_UNBOUND",
                    f"WPS 文档尚未绑定当前局点，必须显式确认后才能写入：{target.target_name}",
                    details={
                        **_target_error_details(target, phase="BINDING_GATE"),
                        "binding_status": binding_status,
                    },
                )
            if binding_status not in {"BOUND", "UNBOUND"}:
                raise WpsSyncError(
                    "WPS_BINDING_STATUS_UNKNOWN",
                    f"WPS 文档绑定状态无法识别：{target.target_name}",
                    details=_target_error_details(target, phase="BINDING_GATE"),
                )
        unverified = [target.target_name for target in targets if target.runtime_capability != "VERIFIED"]
        if unverified:
            raise WpsSyncError(
                "WPS_RUNTIME_WRITE_PROBE_REQUIRED",
                f"WPS 写入探针尚未验证：{', '.join(unverified)}",
            )
        if isinstance(self.client, WpsAirScriptClient):
            for target in targets:
                _assert_runtime_identity(target)

        resumable = repository.find_resumable_batch(
            TRACKSIDE_AP_WPS_BUSINESS_KEY,
            requested_codes,
        )
        if resumable is not None:
            if progress is not None:
                progress(
                    "wps_remote_resume",
                    30,
                    100,
                    "发现未完成的 WPS 远端任务，正在恢复查询",
                )
            return self._execute_persisted_batch(
                repository,
                targets,
                resumable,
                progress=progress,
                should_cancel=should_cancel,
            )

        if should_cancel is not None:
            should_cancel()
        if progress is not None:
            progress("wps_snapshot", 5, 100, "正在冻结轨旁 AP 业务快照")
        snapshot = self._build_snapshot(site_id)
        revision = str(snapshot.get("business_revision") or "")
        if expected_revision and revision != expected_revision:
            raise WpsSyncError("TRACKSIDE_AP_SNAPSHOT_STALE", "轨旁 AP 数据已更新，请刷新后重试")
        batch_id = f"wps_{uuid4().hex}"
        workbook, snapshot_sha256, payload_size, source_format_manifest = self._build_workbook_dto(
            site_id,
            batch_id,
            snapshot,
        )
        if progress is not None:
            progress("wps_workbook", 30, 100, f"已生成统一工作簿数据集（{len(workbook.sheets)} 个 Sheet）")
            for sheet in workbook.sheets:
                progress(
                    "wps_workbook_dimensions",
                    30,
                    100,
                    (
                        f"{sheet.sheet_name}: columns={sheet.column_count}, "
                        f"column_widths={len(sheet.column_widths)}"
                    ),
                )
        generated_at = str(snapshot.get("created_at") or _now())
        repository.create_batch(
            batch_id=batch_id,
            business_key=TRACKSIDE_AP_WPS_BUSINESS_KEY,
            revision=revision,
            snapshot_sha256=snapshot_sha256,
            snapshot_generated_at=generated_at,
            target_count=len(targets),
        )
        initial_summary = {
            "batch_id": batch_id,
            "site_id": site_id,
            "business_key": TRACKSIDE_AP_WPS_BUSINESS_KEY,
            "snapshot_revision": revision,
            "snapshot_sha256": snapshot_sha256,
            "snapshot_generated_at": generated_at,
            "payload_bytes": payload_size,
            "sheet_count": len(workbook.sheets),
            "status": "RUNNING",
            "target_count": len(targets),
            "success_count": 0,
            "failed_count": 0,
            "warning_count": 0,
            "partial_success": False,
            "targets": [],
        }
        repository.update_batch_state(
            batch_id,
            status="RUNNING",
            summary=initial_summary,
        )
        persisted_runs: list[dict[str, Any]] = []
        for target in targets:
            target_batch_id = f"{batch_id}_{target.target_code}"
            request_payload = {
                "protocol_version": WPS_SYNC_PROTOCOL_VERSION,
                "operation": "sync_trackside_ap_business",
                "parent_batch_id": batch_id,
                "target_batch_id": target_batch_id,
                "target_type": target.target_type.value,
                "target_code": target.target_code,
                "binding_id": target.binding_id,
                "initialize_binding": bool(initialize_binding),
                "site_id": site_id,
                "site_name": self._site_display_name(site_id),
                "business_key": TRACKSIDE_AP_WPS_BUSINESS_KEY,
                "snapshot_revision": revision,
                "snapshot_sha256": snapshot_sha256,
                "snapshot_generated_at": generated_at,
                "requested_at": _now(),
                "format_mirror_enabled": WPS_STANDARD_FORMAT_MIRROR_ENABLED,
                "sheet_tab_color_enabled": _sheet_tab_color_probe_verified(target),
                "column_width_enabled": True,
                "runtime_capability": target.runtime_capability,
                "runtime_probe_verified": target.runtime_capability == "VERIFIED",
                "workbook": workbook.to_dict(),
                "script_id": _script_id_from_webhook(target.webhook_url),
            }
            repository.create_target_run(
                target_batch_id=target_batch_id,
                batch_id=batch_id,
                target=target,
                request_payload=request_payload,
                source_format_manifest=source_format_manifest,
            )
            persisted_runs.append(
                {
                    "target_batch_id": target_batch_id,
                    "batch_id": batch_id,
                    "target_id": target.target_id,
                    "target_code": target.target_code,
                    "target_type": target.target_type.value,
                    "status": "REMOTE_SUBMITTING",
                    "remote_task_id": "",
                    "remote_task_id_masked": "",
                    "remote_task_type": "",
                    "remote_task_status": "",
                    "remote_task_submitted_at": "",
                    "remote_task_last_polled_at": "",
                    "remote_task_finished_at": "",
                    "request_payload": request_payload,
                    "source_format_manifest": source_format_manifest,
                    "result_summary": {},
                }
            )
        batch = {**initial_summary, "targets": persisted_runs}
        return self._execute_persisted_batch(
            repository,
            targets,
            batch,
            progress=progress,
            should_cancel=should_cancel,
        )

    def _execute_persisted_batch(
        self,
        repository: WpsSyncRepository,
        targets: Sequence[WpsSyncTarget],
        batch: Mapping[str, object],
        *,
        progress: Callable[[str, int, int, object], None] | None,
        should_cancel: Callable[[], None] | None,
    ) -> dict[str, Any]:
        target_by_code = {target.target_code: target for target in targets}
        raw_runs = batch.get("targets")
        runs = [dict(run) for run in raw_runs if isinstance(run, Mapping)] if isinstance(raw_runs, list) else []
        batch_id = str(batch.get("batch_id") or "")
        revision = str(batch.get("snapshot_revision") or "")
        results: list[dict[str, Any]] = []
        for target_index, run in enumerate(runs, start=1):
            if should_cancel is not None:
                should_cancel()
            target = target_by_code.get(str(run.get("target_code") or ""))
            if target is None:
                raise WpsSyncError(
                    "WPS_TARGET_INVALID",
                    "未完成批次包含当前请求之外的 WPS 目标",
                )
            stored_result = run.get("result_summary")
            if (
                str(run.get("status") or "")
                in {"SUCCESS", "SUCCESS_WITH_WARNINGS", "FAILED"}
                and isinstance(stored_result, Mapping)
                and stored_result
            ):
                results.append(dict(stored_result))
                continue
            request_payload = run.get("request_payload")
            source_format_manifest = run.get("source_format_manifest")
            if not isinstance(request_payload, Mapping):
                raise WpsSyncError(
                    "WPS_REMOTE_RESUME_CONTEXT_MISSING",
                    "WPS 未完成任务缺少可恢复请求上下文",
                )
            try:
                _validate_target_configuration(target)
                if progress is not None:
                    progress(
                        "wps_target_sync",
                        30 + int((target_index - 1) * 65 / max(len(runs), 1)),
                        100,
                        f"正在处理 {target.target_name}",
                    )
                if isinstance(self.client, WpsAirScriptClient):
                    response = self._execute_async_target(
                        repository,
                        target,
                        run,
                        dict(request_payload),
                        progress=progress,
                        should_cancel=should_cancel,
                    )
                else:
                    response = self.adapters[target.target_type].sync(
                        target,
                        self._token(repository, target),
                        request_payload,
                    )
                manifest = (
                    dict(source_format_manifest)
                    if isinstance(source_format_manifest, Mapping)
                    else {}
                )
                column_width_report = _column_width_verification_report(
                    manifest=manifest.get("column_widths") or [],
                    request_payload=request_payload,
                    remote_result=response,
                    enabled=True,
                )
                response["column_width_verification_report"] = column_width_report
                response["source_workbook_format_manifest"] = manifest
                _append_column_width_report_warning(response, column_width_report)
                format_warnings = response.get("format_warnings")
                target_status = (
                    "SUCCESS_WITH_WARNINGS"
                    if isinstance(format_warnings, list) and format_warnings
                    else "SUCCESS"
                )
                public = {
                    "target_code": target.target_code,
                    "target_name": target.target_name,
                    "target_type": target.target_type.value,
                    "target_batch_id": str(run.get("target_batch_id") or ""),
                    **_sanitize_result(response),
                    "status": target_status,
                }
                repository.update_target_remote_state(
                    target.target_id,
                    binding_status=str(response.get("binding_status") or "UNKNOWN"),
                    result=response,
                )
                repository.complete_target_run(
                    str(run.get("target_batch_id") or ""),
                    status=target_status,
                    result=public,
                )
                repository.update_target_sync(
                    target.target_id,
                    status=target_status,
                    revision=revision,
                )
            except WpsSyncError as exc:
                recoverable = exc.code in {
                    "ASYNC_SUBMIT_FAILED",
                    "WPS_ASYNC_TASK_ID_MISSING",
                    "REMOTE_POLL_TEMPORARY_FAILED",
                    "REMOTE_RESULT_UNKNOWN",
                }
                public = {
                    "target_code": target.target_code,
                    "target_name": target.target_name,
                    "target_type": target.target_type.value,
                    "target_batch_id": str(run.get("target_batch_id") or ""),
                    "status": "REMOTE_RESULT_UNKNOWN" if recoverable else "FAILED",
                    "error_code": exc.code,
                    "message": str(exc),
                    **_sanitize_result(dict(exc.details)),
                }
                if recoverable:
                    repository.update_remote_task_poll(
                        str(run.get("target_batch_id") or ""),
                        status="REMOTE_RESULT_UNKNOWN",
                        remote_status="unknown",
                        error_code=exc.code,
                        error_message=str(exc),
                    )
                else:
                    repository.complete_target_run(
                        str(run.get("target_batch_id") or ""),
                        status="FAILED",
                        result=public,
                        error_code=exc.code,
                        error_message=str(exc),
                        remote_task_status=(
                            "failed"
                            if exc.code == "WPS_REMOTE_EXECUTION_FAILED"
                            else "finished"
                        ),
                    )
                repository.update_target_sync(
                    target.target_id,
                    status=str(public["status"]),
                    revision=revision,
                )
            results.append(public)
            if progress is not None:
                progress(
                    "wps_target_sync",
                    30 + int(target_index * 65 / max(len(runs), 1)),
                    100,
                    f"{target.target_name}同步状态：{public['status']}",
                )

        success_count = sum(
            result["status"] in {"SUCCESS", "SUCCESS_WITH_WARNINGS"}
            for result in results
        )
        unknown_count = sum(
            result["status"] == "REMOTE_RESULT_UNKNOWN" for result in results
        )
        failed_count = sum(result["status"] == "FAILED" for result in results)
        warning_count = sum(_target_format_warning_count(result) for result in results)
        status = (
            "REMOTE_RESULT_UNKNOWN"
            if unknown_count > 0
            else "SUCCESS_WITH_WARNINGS"
            if failed_count == 0 and warning_count > 0
            else "SUCCESS"
            if failed_count == 0
            else "FAILED"
            if success_count == 0
            else "PARTIAL_SUCCESS"
        )
        initial_summary = batch.get("result_summary")
        persisted_summary = (
            dict(initial_summary) if isinstance(initial_summary, Mapping) else {}
        )
        summary = {
            **persisted_summary,
            "batch_id": batch_id,
            "site_id": str(batch.get("site_id") or repository.site_id),
            "business_key": str(
                batch.get("business_key") or TRACKSIDE_AP_WPS_BUSINESS_KEY
            ),
            "snapshot_revision": revision,
            "snapshot_sha256": str(batch.get("snapshot_sha256") or ""),
            "snapshot_generated_at": str(batch.get("snapshot_generated_at") or ""),
            "payload_bytes": int(persisted_summary.get("payload_bytes") or batch.get("payload_bytes") or 0),
            "sheet_count": int(persisted_summary.get("sheet_count") or batch.get("sheet_count") or 0),
            "status": status,
            "target_count": len(results),
            "success_count": success_count,
            "failed_count": failed_count,
            "unknown_count": unknown_count,
            "warning_count": warning_count,
            "partial_success": status == "PARTIAL_SUCCESS",
            "targets": results,
        }
        if unknown_count:
            repository.update_batch_state(
                batch_id,
                status="REMOTE_RESULT_UNKNOWN",
                summary=summary,
            )
        else:
            repository.complete_batch(
                batch_id,
                status=status,
                success_count=success_count,
                failed_count=failed_count,
                summary=summary,
            )
        if progress is not None:
            progress("wps_complete", 100, 100, f"WPS 云文档同步完成：{status}")
        return summary

    def _execute_async_target(
        self,
        repository: WpsSyncRepository,
        target: WpsSyncTarget,
        run: Mapping[str, object],
        request_payload: dict[str, object],
        *,
        progress: Callable[[str, int, int, object], None] | None,
        should_cancel: Callable[[], None] | None,
    ) -> dict[str, Any]:
        target_batch_id = str(run.get("target_batch_id") or "")
        task_id = str(run.get("remote_task_id") or "")
        task_type = str(run.get("remote_task_type") or "")
        submitted_at = str(run.get("remote_task_submitted_at") or "")
        token = self._token(repository, target)
        if not task_id:
            if progress is not None:
                progress("wps_remote_submit", 35, 100, "正在提交 WPS 远端任务")
            remote = self.client.submit_async(
                target,
                token=token,
                argv=request_payload,
            )
            task_id = remote.task_id
            task_type = remote.task_type
            submitted_at = _now()
            repository.update_remote_task_submitted(
                target_batch_id,
                task_id=task_id,
                task_type=task_type,
                remote_status=remote.status,
            )
            if progress is not None:
                progress(
                    "wps_remote_submitted",
                    40,
                    100,
                    {
                        "message": "WPS 远端任务已提交",
                        "remote_task_id_masked": _mask_remote_task_id(task_id),
                        "remote_task_type": task_type,
                        "remote_task_status": remote.status,
                        "remote_task_submitted_at": submitted_at,
                    },
                )
        elif progress is not None:
            progress(
                "wps_remote_resume",
                40,
                100,
                {
                    "message": "正在恢复 WPS 远端任务查询",
                    "remote_task_id_masked": _mask_remote_task_id(task_id),
                    "remote_task_type": task_type,
                    "remote_task_status": str(run.get("remote_task_status") or "unknown"),
                    "remote_task_submitted_at": submitted_at,
                    "remote_task_last_polled_at": str(
                        run.get("remote_task_last_polled_at") or ""
                    ),
                },
            )

        deadline = self._monotonic() + self.remote_task_max_wait_seconds
        while True:
            if should_cancel is not None:
                should_cancel()
            if self._monotonic() >= deadline:
                raise WpsSyncError(
                    "REMOTE_RESULT_UNKNOWN",
                    "WPS 任务已经提交，但在总等待时间内未能确认远端执行结果",
                    details={
                        **_target_error_details(target, phase="REMOTE_POLL"),
                        "remote_task_id_masked": _mask_remote_task_id(task_id),
                        "remote_task_status": "unknown",
                        "remote_task_submitted_at": submitted_at,
                    },
                )
            try:
                polled = self.client.poll_async_task(
                    target,
                    token=token,
                    task_id=task_id,
                )
            except WpsSyncError as exc:
                repository.update_remote_task_poll(
                    target_batch_id,
                    status="REMOTE_RESULT_UNKNOWN",
                    remote_status="unknown",
                    error_code=exc.code,
                    error_message=str(exc),
                )
                if exc.code in {
                    "REMOTE_POLL_TEMPORARY_FAILED",
                    "WPS_REMOTE_UNAVAILABLE",
                    "WPS_REMOTE_RATE_LIMITED",
                }:
                    if progress is not None:
                        progress(
                            "wps_remote_poll_retry",
                            45,
                            100,
                            {
                                "message": "WPS 远端任务查询暂时失败，正在继续查询",
                                "remote_task_id_masked": _mask_remote_task_id(task_id),
                                "remote_task_status": "unknown",
                                "remote_task_last_polled_at": _now(),
                                "remote_error_code": exc.code,
                            },
                        )
                    self._wait_for_next_poll(deadline)
                    continue
                raise WpsSyncError(
                    "REMOTE_RESULT_UNKNOWN",
                    "WPS 任务已经提交，当前无法确认远端执行结果",
                    details={
                        **_target_error_details(target, phase="REMOTE_POLL"),
                        "remote_task_id_masked": _mask_remote_task_id(task_id),
                        "remote_task_status": "unknown",
                        "poll_error_code": exc.code,
                        "poll_error_message": str(exc),
                    },
                ) from exc
            body = polled.body
            if not isinstance(body, Mapping):
                raise WpsSyncError(
                    "REMOTE_RESULT_UNKNOWN",
                    "WPS 远端任务查询返回结构无效",
                    details={
                        **_target_error_details(target, phase="REMOTE_POLL"),
                        "remote_task_id_masked": _mask_remote_task_id(task_id),
                    },
                )
            remote_status = str(body.get("status") or "").strip().casefold()
            polled_at = _now()
            if remote_status in {"pending", "queued", "starting", "running"}:
                repository.update_remote_task_poll(
                    target_batch_id,
                    status="REMOTE_RUNNING",
                    remote_status=remote_status,
                )
                if progress is not None:
                    progress(
                        "wps_remote_running",
                        50,
                        100,
                        {
                            "message": "WPS 远端任务执行中",
                            "remote_task_id_masked": _mask_remote_task_id(task_id),
                            "remote_task_type": task_type,
                            "remote_task_status": remote_status,
                            "remote_task_submitted_at": submitted_at,
                            "remote_task_last_polled_at": polled_at,
                        },
                    )
                self._wait_for_next_poll(deadline)
                continue
            if remote_status not in {"finished", "success", "completed"}:
                repository.update_remote_task_poll(
                    target_batch_id,
                    status="REMOTE_RESULT_UNKNOWN",
                    remote_status=remote_status or "unknown",
                    error_code="WPS_REMOTE_TASK_STATUS_INVALID",
                    error_message="WPS 远端任务状态无法识别",
                )
                raise WpsSyncError(
                    "REMOTE_RESULT_UNKNOWN",
                    f"WPS 远端任务状态无法识别：{remote_status or '未知'}",
                    details={
                        **_target_error_details(target, phase="REMOTE_POLL"),
                        "remote_task_id_masked": _mask_remote_task_id(task_id),
                        "remote_task_status": remote_status or "unknown",
                    },
                )
            repository.update_remote_task_poll(
                target_batch_id,
                status="REMOTE_FINISHED",
                remote_status=remote_status,
            )
            if progress is not None:
                progress(
                    "wps_remote_finished",
                    90,
                    100,
                    {
                        "message": "WPS 远端执行完成，正在解析结果",
                        "remote_task_id_masked": _mask_remote_task_id(task_id),
                        "remote_task_type": task_type,
                        "remote_task_status": remote_status,
                        "remote_task_submitted_at": submitted_at,
                        "remote_task_last_polled_at": polled_at,
                        "remote_task_finished_at": polled_at,
                    },
                )
            try:
                result = self.adapters[target.target_type].validate_sync_response(
                    target,
                    token,
                    request_payload,
                    polled,
                )
            except WpsSyncError as exc:
                if exc.code != "WPS_SCRIPT_EXECUTION_FAILED":
                    raise
                raise WpsSyncError(
                    "WPS_REMOTE_EXECUTION_FAILED",
                    str(exc),
                    details={
                        **dict(exc.details),
                        "remote_task_id_masked": _mask_remote_task_id(task_id),
                        "remote_task_status": remote_status,
                    },
                ) from exc
            result.update(
                {
                    "remote_task_id_masked": _mask_remote_task_id(task_id),
                    "remote_task_type": task_type,
                    "remote_task_status": remote_status,
                    "remote_task_submitted_at": submitted_at,
                    "remote_task_last_polled_at": polled_at,
                    "remote_task_finished_at": polled_at,
                }
            )
            return result

    def _wait_for_next_poll(self, deadline: float) -> None:
        remaining = max(0.0, deadline - self._monotonic())
        delay = min(self.remote_task_poll_interval_seconds, remaining)
        if delay > 0:
            self._sleep(delay)

    def recent_batches(self, site_id: str, limit: int = 10) -> list[dict[str, object]]:
        repository = self._repository(site_id)
        self._ensure_default_targets(repository)
        return repository.recent_batches(TRACKSIDE_AP_WPS_BUSINESS_KEY, limit)

    def _build_snapshot(self, site_id: str) -> dict[str, object]:
        metadata = SiteManager(self.paths).load_site_metadata(site_id)
        scope_context = {
            **metadata,
            "site_id": site_id,
            "site_display_name": self._site_display_name(site_id),
            "generated_at": _now(),
        }
        return build_trackside_ap_business_export_snapshot(
            DeviceRepository(Database(self.paths.site_db_path(site_id))),
            site_id,
            scope_context=scope_context,
        )

    def _build_workbook_dto(
        self,
        site_id: str,
        batch_id: str,
        snapshot: Mapping[str, object],
    ) -> tuple[WorkbookDTO, str, int, dict[str, Any]]:
        root = (self.paths.temp_dir / "wps_trackside_ap" / site_id / batch_id).resolve()
        root.mkdir(parents=True, exist_ok=False)
        output = root / "trackside-ap-business.xlsx"
        tmp = root / "trackside-ap-business.xlsx.tmp"
        try:
            _render_trackside_ap_business_export(
                snapshot,
                output_path=output,
                tmp_path=tmp,
                language="zh_CN",
                progress_callback=None,
                should_cancel=None,
            )
            workbook = workbook_dto_from_xlsx(
                output,
                include_format_mirror=WPS_STANDARD_FORMAT_MIRROR_ENABLED,
                include_column_widths=True,
            )
            source_format_manifest = _source_workbook_format_manifest(output, workbook)
            digest_payload = {
                "site_id": site_id,
                "business_key": TRACKSIDE_AP_WPS_BUSINESS_KEY,
                "snapshot_revision": str(snapshot.get("business_revision") or ""),
                "workbook": workbook.to_dict(),
            }
            serialized = canonical_json_bytes(digest_payload)
            if len(serialized) > _MAX_PAYLOAD_BYTES:
                raise WpsSyncError(
                    "WPS_PAYLOAD_TOO_LARGE",
                    "轨旁 AP 工作簿超过 WPS 单次同步安全大小",
                )
            return (
                workbook,
                content_sha256(digest_payload),
                len(serialized),
                source_format_manifest,
            )
        finally:
            shutil.rmtree(root, ignore_errors=True)

    def _ensure_default_targets(self, repository: WpsSyncRepository) -> None:
        existing_targets = repository.list_targets(TRACKSIDE_AP_WPS_BUSINESS_KEY)
        existing = {target.target_code: target for target in existing_targets}
        credential_owners: dict[str, str] = {}
        for target in existing_targets:
            owner = credential_owners.setdefault(target.credential_id, target.target_code)
            if owner == target.target_code:
                continue
            token = repository.resolve_token(target)
            repository.upsert_target(
                business_key=target.business_key,
                target_code=target.target_code,
                target_type=target.target_type,
                target_name=target.target_name,
                document_open_url=target.document_open_url,
                webhook_url=target.webhook_url,
                expected_document_id=target.expected_document_id,
                enabled=target.enabled,
                timeout_seconds=target.timeout_seconds,
                token=token or None,
                credential_id=f"wsc_{uuid4().hex}",
            )
        for definition in DEFAULT_TARGETS:
            code = str(definition["target_code"])
            target = existing.get(code)
            default_name = _default_target_name(
                self._site_display_name(repository.site_id), code
            )
            if target is not None:
                if _is_legacy_default_target_name(target.target_name, code):
                    repository.upsert_target(
                        business_key=target.business_key,
                        target_code=target.target_code,
                        target_type=target.target_type,
                        target_name=default_name,
                        document_open_url=target.document_open_url,
                        webhook_url=target.webhook_url,
                        expected_document_id=target.expected_document_id,
                        enabled=target.enabled,
                        timeout_seconds=target.timeout_seconds,
                        credential_id=target.credential_id,
                    )
                continue
            repository.upsert_target(
                business_key=TRACKSIDE_AP_WPS_BUSINESS_KEY,
                **definition,
                target_name=default_name,
                credential_id=f"wsc_{uuid4().hex}",
            )
        for target in repository.list_targets(TRACKSIDE_AP_WPS_BUSINESS_KEY):
            if (
                target.target_type is WpsTargetType.STANDARD_SPREADSHEET
                and target.runtime_capability == "VERIFIED"
            ):
                expected_identity = {
                    "document_id": target.expected_document_id,
                    "script_id": _script_id_from_webhook(target.webhook_url),
                    "script_version": WPS_SCRIPT_VERSIONS.get(target.target_code, ""),
                    "deployment_id": WPS_DEPLOYMENT_IDS.get(target.target_code, ""),
                }
                runtime_identity = {
                    "document_id": target.runtime_probe_document_id,
                    "script_id": target.runtime_probe_script_id,
                    "script_version": target.runtime_probe_script_version,
                    "deployment_id": target.runtime_probe_deployment_id,
                }
                if any(
                    actual and actual != expected_identity[key]
                    for key, actual in runtime_identity.items()
                ):
                    repository.clear_runtime_probe_identity(target.target_id)

    def _repository(self, site_id: str) -> WpsSyncRepository:
        selected = SiteManager(self.paths).validate_site_name(site_id)
        return WpsSyncRepository(self.paths, selected)

    def _site_display_name(self, site_id: str) -> str:
        metadata = SiteManager(self.paths).load_site_metadata(site_id)
        return str(metadata.get("display_name") or site_id)

    @staticmethod
    def _env_token(target_code: str) -> str:
        name = _TARGET_TOKEN_ENV.get(target_code, "")
        return str(os.environ.get(name) or "").strip() if name else ""

    def _token(self, repository: WpsSyncRepository, target: WpsSyncTarget) -> str:
        return repository.resolve_token(target) or self._env_token(target.target_code)

    @staticmethod
    def _public_target(
        target: WpsSyncTarget,
        *,
        env_token_configured: bool,
    ) -> dict[str, Any]:
        result = target.public_dict()
        result["binding_status"] = _effective_binding_status(target)
        result["expected_script_version"] = WPS_SCRIPT_VERSIONS.get(target.target_code, "")
        result["expected_deployment_id"] = WPS_DEPLOYMENT_IDS.get(target.target_code, "")
        result["expected_script_id"] = (
            _script_id_from_webhook(target.webhook_url)
            if target.webhook_url
            else ""
        )
        result["runtime_capability"] = target.runtime_capability or WPS_RUNTIME_CAPABILITIES.get(target.target_code, "RUNTIME_UNVERIFIED")
        result["token_configured"] = bool(
            target.token_configured or env_token_configured
        )
        if env_token_configured and not target.token_configured:
            result["token_suffix"] = ""
        return result


def workbook_dto_from_xlsx(
    path: str | Path,
    *,
    include_format_mirror: bool = False,
    include_column_widths: bool = False,
) -> WorkbookDTO:
    workbook = load_workbook(Path(path), data_only=False, read_only=False)
    try:
        sheets: list[WorkbookSheetDTO] = []
        for worksheet in workbook.worksheets:
            if worksheet.title in _META_SHEET_NAMES or worksheet.title.startswith("_NetConsole"):
                continue
            definition = trackside_ap_business_sheet_definition(worksheet.title)
            max_row = int(worksheet.max_row or 0)
            max_column = int(worksheet.max_column or 0)
            cells = [
                [worksheet.cell(row=row, column=column).value for column in range(1, max_column + 1)]
                for row in range(1, max_row + 1)
            ]
            sync_mode = (
                WpsSyncMode(definition.sync_mode)
                if definition is not None
                else WpsSyncMode.FULL_REPLACE
            )
            sheets.append(
                WorkbookSheetDTO(
                    logical_sheet_key=(
                        definition.stable_key
                        if definition is not None
                        else _logical_sheet_key(worksheet.title)
                    ),
                    sheet_name=worksheet.title,
                    sync_mode=sync_mode,
                    cells=cells,
                    row_count=max_row,
                    column_count=max_column,
                    sheet_order=len(sheets),
                    sheet_visible=(worksheet.sheet_state == "visible") if include_format_mirror else True,
                    tab_color=(
                        _openpyxl_color(worksheet.sheet_properties.tabColor)
                        if include_format_mirror
                        else ""
                    ),
                    merges=(
                        [str(value) for value in worksheet.merged_cells.ranges]
                        if include_format_mirror
                        else []
                    ),
                    row_heights={
                        str(index): float(dimension.height)
                        for index, dimension in worksheet.row_dimensions.items()
                        if dimension.height is not None
                    } if include_format_mirror else {},
                    column_widths={
                        str(index): float(dimension.width)
                        for index, dimension in worksheet.column_dimensions.items()
                        if dimension.width is not None
                    } if include_format_mirror or include_column_widths else {},
                    auto_fit_columns=(
                        tuple(get_column_letter(column) for column in range(1, max_column + 1))
                        if include_format_mirror
                        else ()
                    ),
                    column_layouts=(
                        {
                            get_column_letter(column): {
                                "layout_type": layout_type,
                                "min_width": TRACKSIDE_COLUMN_LAYOUT_LIMITS[layout_type][0],
                                "max_width": TRACKSIDE_COLUMN_LAYOUT_LIMITS[layout_type][1],
                                "wrap_text": layout_type == "long_text",
                            }
                            for column, layout_type in enumerate(
                                trackside_ap_business_column_layout_types(
                                    worksheet.title,
                                    max_column,
                                ),
                                start=1,
                            )
                        }
                        if include_format_mirror
                        else {}
                    ),
                    auto_fit_rows=bool(include_format_mirror and max_row),
                    format_runs=(
                        _format_runs_from_worksheet(
                            worksheet,
                            max_row=max_row,
                            max_column=max_column,
                        )
                        if include_format_mirror
                        else ()
                    ),
                    freeze_mode=(
                        WpsFreezeMode(definition.freeze_mode)
                        if include_format_mirror and definition is not None
                        else WpsFreezeMode.FIRST_ROW_ONLY
                        if include_format_mirror
                        else WpsFreezeMode.NONE
                    ),
                    auto_filter=(
                        str(worksheet.auto_filter.ref or "")
                        if include_format_mirror
                        else ""
                    ),
                    verification_samples=(
                        _verification_samples_from_worksheet(
                            worksheet,
                            max_row=max_row,
                            max_column=max_column,
                            header_row=(
                                3
                                if definition is not None
                                and definition.stable_key
                                == "ap_online_history_overview"
                                else 1
                            ),
                        )
                        if include_format_mirror
                        else ()
                    ),
                )
            )
        return WorkbookDTO(sheets=tuple(sheets))
    finally:
        workbook.close()


def _source_workbook_format_manifest(
    path: str | Path,
    workbook_dto: WorkbookDTO,
) -> dict[str, Any]:
    dto_sheets = {sheet.sheet_name: sheet for sheet in workbook_dto.sheets}
    workbook = load_workbook(Path(path), data_only=False, read_only=False)
    try:
        column_manifest: list[dict[str, Any]] = []
        sheet_manifest: list[dict[str, Any]] = []
        for worksheet in workbook.worksheets:
            dto_sheet = dto_sheets.get(worksheet.title)
            if dto_sheet is None:
                continue
            definition = trackside_ap_business_sheet_definition(worksheet.title)
            header_row = (
                3
                if definition is not None
                and definition.stable_key == "ap_online_history_overview"
                else 1
            )
            font_styles: set[str] = set()
            fill_styles: set[str] = set()
            alignment_styles: set[str] = set()
            number_formats: set[str] = set()
            border_styles: set[str] = set()
            for row in worksheet.iter_rows():
                for cell in row:
                    payload = _cell_format_payload(cell)
                    font_styles.add(_format_signature({"font": payload.get("font") or {}}))
                    if payload.get("fill"):
                        fill_styles.add(_format_signature({"fill": payload["fill"]}))
                    if payload.get("alignment"):
                        alignment_styles.add(
                            _format_signature({"alignment": payload["alignment"]})
                        )
                    if payload.get("number_format"):
                        number_formats.add(str(payload["number_format"]))
                    if payload.get("border"):
                        border_styles.add(
                            _format_signature({"border": payload["border"]})
                        )
            for column_index in range(1, int(worksheet.max_column or 0) + 1):
                column_name = get_column_letter(column_index)
                dimension = worksheet.column_dimensions.get(column_name)
                explicit_width = (
                    float(dimension.width)
                    if dimension is not None and dimension.width is not None
                    else None
                )
                auto_fit_enabled = column_name in dto_sheet.auto_fit_columns
                source_mode = (
                    "AUTO_FIT_WITH_LOCAL_MIN"
                    if auto_fit_enabled and explicit_width is not None
                    else "AUTO_FIT"
                    if auto_fit_enabled
                    else "EXPLICIT"
                    if explicit_width is not None
                    else "AUTO_FIT"
                )
                layout = dict(dto_sheet.column_layouts.get(column_name) or {})
                header_value = worksheet.cell(row=header_row, column=column_index).value
                column_manifest.append(
                    {
                        "sheet_name": worksheet.title,
                        "column": column_name,
                        "range": f"{column_name}:{column_name}",
                        "column_label": (
                            str(header_value).strip()
                            if header_value is not None
                            else ""
                        ),
                        "source_mode": source_mode,
                        "local_workbook_width": explicit_width,
                        "sheet_dto_width": _safe_float(
                            dto_sheet.column_widths.get(column_name)
                        ),
                        "sheet_dto_mode": (
                            "AUTO_FIT_WITH_LOCAL_MIN"
                            if auto_fit_enabled and column_name in dto_sheet.column_widths
                            else "AUTO_FIT"
                            if auto_fit_enabled
                            else "EXPLICIT"
                            if column_name in dto_sheet.column_widths
                            else "AUTO_FIT"
                            if column_name in dto_sheet.auto_fit_columns
                            else "MISSING"
                        ),
                        "auto_fit_min_width": dto_sheet.auto_fit_min_width,
                        "auto_fit_max_width": dto_sheet.auto_fit_max_width,
                        **(
                            {
                                "layout_type": str(
                                    layout.get("layout_type") or "normal"
                                ),
                                "layout_min_width": _safe_float(
                                    layout.get("min_width")
                                ),
                                "layout_max_width": _safe_float(
                                    layout.get("max_width")
                                ),
                            }
                            if layout
                            else {}
                        ),
                    }
                )
            sheet_manifest.append(
                {
                    "sheet_name": worksheet.title,
                    "row_count": int(worksheet.max_row or 0),
                    "column_count": int(worksheet.max_column or 0),
                    "explicit_width_count": len(dto_sheet.column_widths),
                    "auto_fit_column_count": len(dto_sheet.auto_fit_columns),
                    "explicit_row_height_count": len(dto_sheet.row_heights),
                    "row_auto_fit": dto_sheet.auto_fit_rows,
                    "format_run_count": len(dto_sheet.format_runs),
                    "font_style_count": len(font_styles),
                    "fill_style_count": len(fill_styles),
                    "alignment_style_count": len(alignment_styles),
                    "number_format_count": len(number_formats),
                    "merge_range_count": len(dto_sheet.merges),
                    "border_style_count": len(border_styles),
                    "freeze_mode": dto_sheet.freeze_mode.value,
                    "auto_filter": dto_sheet.auto_filter,
                    "verification_sample_count": len(dto_sheet.verification_samples),
                }
            )
        return {
            "sheets": sheet_manifest,
            "column_widths": column_manifest,
            "totals": {
                "sheet_count": len(sheet_manifest),
                "column_count": len(column_manifest),
                "explicit_width_count": sum(
                    int(item["explicit_width_count"]) for item in sheet_manifest
                ),
                "auto_fit_column_count": sum(
                    int(item["auto_fit_column_count"]) for item in sheet_manifest
                ),
                "explicit_row_height_count": sum(
                    int(item["explicit_row_height_count"]) for item in sheet_manifest
                ),
                "format_run_count": sum(
                    int(item["format_run_count"]) for item in sheet_manifest
                ),
            },
        }
    finally:
        workbook.close()


def _verification_samples_from_worksheet(
    worksheet: Any,
    *,
    max_row: int,
    max_column: int,
    header_row: int,
) -> tuple[dict[str, Any], ...]:
    if max_row <= 0 or max_column <= 0:
        return ()
    data_start = min(header_row + 1, max_row)
    data_end = max_row
    while data_end > header_row and all(
        worksheet.cell(row=data_end, column=column).value in (None, "")
        for column in range(1, max_column + 1)
    ):
        data_end -= 1
    candidates: dict[int, set[str]] = {header_row: {"header"}}
    if data_end >= data_start:
        candidates.setdefault(data_start, set()).add("first_data")
        candidates.setdefault((data_start + data_end) // 2, set()).add("middle_data")
        candidates.setdefault(data_end, set()).add("last_data")

    fill_rows: dict[str, int] = {}
    for row_index in range(data_start, data_end + 1):
        colors = {
            str((_cell_format_payload(worksheet.cell(row=row_index, column=column)).get("fill") or {}).get("fg_color") or "")
            for column in range(1, max_column + 1)
        }
        for color in sorted(value for value in colors if value):
            fill_rows.setdefault(color, row_index)
    for color, row_index in list(fill_rows.items())[:4]:
        candidates.setdefault(row_index, set()).add(f"fill_{color}")

    last_column = get_column_letter(max_column)
    samples: list[dict[str, Any]] = []
    for row_index in sorted(candidates):
        format_cells: list[dict[str, Any]] = []
        previous_signature = ""
        for column in range(1, max_column + 1):
            cell = worksheet.cell(row=row_index, column=column)
            payload = _cell_format_payload(cell)
            signature = _format_signature(payload)
            if signature == previous_signature:
                continue
            previous_signature = signature
            format_cell = {
                "range": cell.coordinate,
                "expected": payload,
            }
            expected_display_text = _expected_display_text(cell)
            if expected_display_text is not None:
                format_cell["expected_display_text"] = expected_display_text
            format_cells.append(format_cell)
            if len(format_cells) >= 8:
                break
        samples.append(
            {
                "label": ",".join(sorted(candidates[row_index])),
                "row": row_index,
                "range": f"A{row_index}:{last_column}{row_index}",
                "expected_values": [
                    worksheet.cell(row=row_index, column=column).value
                    for column in range(1, max_column + 1)
                ],
                "format_cells": format_cells,
            }
        )
    return tuple(samples)


def _expected_display_text(cell: Any) -> str | None:
    value = cell.value
    if not isinstance(value, (int, float)) or isinstance(value, bool):
        return None
    match = re.fullmatch(r"0(?:\.(0+))?%", str(cell.number_format or "").strip())
    if not match:
        return None
    decimal_places = len(match.group(1) or "")
    return f"{float(value) * 100:.{decimal_places}f}%"


def _format_runs_from_worksheet(
    worksheet: Any,
    *,
    max_row: int,
    max_column: int,
) -> tuple[WorkbookFormatRunDTO, ...]:
    completed: list[dict[str, Any]] = []
    for feature in ("font", "fill", "number_format", "alignment", "border"):
        completed.extend(
            _format_runs_for_feature(
                worksheet,
                feature=feature,
                max_row=max_row,
                max_column=max_column,
            )
        )
    completed.sort(
        key=lambda item: (
            int(item["start_row"]),
            int(item["start_column"]),
            int(item["end_row"]),
            int(item["end_column"]),
        )
    )
    return tuple(_format_run_dto(item) for item in completed)


def _format_runs_for_feature(
    worksheet: Any,
    *,
    feature: str,
    max_row: int,
    max_column: int,
) -> list[dict[str, Any]]:
    active: dict[tuple[int, int, str], dict[str, Any]] = {}
    completed: list[dict[str, Any]] = []

    for row_index in range(1, max_row + 1):
        segments: list[tuple[int, int, str, dict[str, Any]]] = []
        start_column = 1
        blank_border_row = feature == "border" and all(
            worksheet.cell(row=row_index, column=column).value in (None, "")
            for column in range(1, max_column + 1)
        )
        first_payload = (
            {}
            if blank_border_row
            else _feature_format_payload(
                worksheet.cell(row=row_index, column=1),
                feature,
            )
        )
        signature = _format_signature(first_payload)
        for column_index in range(2, max_column + 2):
            payload = (
                {}
                if blank_border_row
                else _feature_format_payload(
                    worksheet.cell(row=row_index, column=column_index),
                    feature,
                )
                if column_index <= max_column
                else None
            )
            next_signature = _format_signature(payload) if payload is not None else ""
            if next_signature == signature:
                continue
            if first_payload:
                segments.append(
                    (start_column, column_index - 1, signature, first_payload)
                )
            start_column = column_index
            first_payload = payload or {}
            signature = next_signature

        current_keys: set[tuple[int, int, str]] = set()
        for start_column, end_column, signature, payload in segments:
            key = (start_column, end_column, signature)
            current_keys.add(key)
            run = active.get(key)
            if run is None:
                active[key] = {
                    "start_row": row_index,
                    "end_row": row_index,
                    "start_column": start_column,
                    "end_column": end_column,
                    "payload": payload,
                }
            else:
                run["end_row"] = row_index
        for key in tuple(active):
            if key not in current_keys:
                completed.append(active.pop(key))

    completed.extend(active.values())
    return completed


def _feature_format_payload(cell: Any, feature: str) -> dict[str, Any]:
    value = _cell_format_payload(cell).get(feature)
    return {feature: value} if value not in (None, "", {}, []) else {}


def _format_run_dto(value: Mapping[str, Any]) -> WorkbookFormatRunDTO:
    start = f"{get_column_letter(int(value['start_column']))}{int(value['start_row'])}"
    end = f"{get_column_letter(int(value['end_column']))}{int(value['end_row'])}"
    payload = dict(value["payload"])
    return WorkbookFormatRunDTO(
        range=start if start == end else f"{start}:{end}",
        font=dict(payload.get("font") or {}),
        fill=dict(payload.get("fill") or {}),
        number_format=str(payload.get("number_format") or ""),
        alignment=dict(payload.get("alignment") or {}),
        border=dict(payload.get("border") or {}),
    )


def _cell_format_payload(cell: Any) -> dict[str, Any]:
    font = {
        "name": str(cell.font.name or ""),
        "size": float(cell.font.sz) if cell.font.sz is not None else None,
        "bold": bool(cell.font.bold),
        "italic": bool(cell.font.italic),
        "underline": str(cell.font.underline or ""),
        "strike": bool(cell.font.strike),
        "color": _openpyxl_color(cell.font.color),
    }
    fill = {
        "fill_type": str(cell.fill.fill_type or ""),
        "fg_color": _openpyxl_color(cell.fill.fgColor),
        "bg_color": _openpyxl_color(cell.fill.bgColor),
    }
    alignment = {
        "horizontal": str(cell.alignment.horizontal or ""),
        "vertical": str(cell.alignment.vertical or ""),
        "wrap_text": bool(cell.alignment.wrap_text),
        "text_rotation": int(cell.alignment.text_rotation or 0),
        "shrink_to_fit": bool(cell.alignment.shrink_to_fit),
    }
    border = {
        side_name: _border_side_payload(getattr(cell.border, side_name, None))
        for side_name in ("left", "right", "top", "bottom", "diagonal")
    }
    border = {key: value for key, value in border.items() if value}
    if border.get("diagonal"):
        border["diagonal"]["up"] = bool(cell.border.diagonalUp)
        border["diagonal"]["down"] = bool(cell.border.diagonalDown)
    payload: dict[str, Any] = {"font": font}
    if fill["fill_type"] and fill["fg_color"]:
        payload["fill"] = fill
    payload["number_format"] = str(cell.number_format or "General")
    if any(
        value not in {"", False, 0}
        for value in alignment.values()
    ):
        payload["alignment"] = alignment
    if border:
        payload["border"] = border
    return payload


def _border_side_payload(side: Any) -> dict[str, Any]:
    if side is None or not side.style:
        return {}
    return {
        "style": str(side.style),
        "color": _openpyxl_color(side.color),
    }


def _openpyxl_color(color: Any) -> str:
    if color is None or str(getattr(color, "type", "")) != "rgb":
        return ""
    value = str(getattr(color, "rgb", "") or "").strip().lstrip("#")
    if len(value) == 8:
        alpha, value = value[:2].upper(), value[-6:].upper()
        # Excel/openpyxl commonly serializes explicit fills and borders as
        # 00RRGGBB even though the color is visibly opaque. Treat that legacy
        # alpha as opaque for non-black RGB values, while preserving a true
        # transparent black (00000000) as no color.
        if alpha not in {"00", "FF"} or (alpha == "00" and value == "000000"):
            return ""
    if len(value) != 6 or not re.fullmatch(r"[0-9A-Fa-f]{6}", value):
        return ""
    return f"#{value.upper()}"


def _format_signature(payload: Mapping[str, Any]) -> str:
    return json.dumps(payload, ensure_ascii=True, sort_keys=True, separators=(",", ":"))


def _target_format_warning_count(result: Mapping[str, Any]) -> int:
    warnings = result.get("format_warnings")
    if isinstance(warnings, list):
        return len(warnings)
    try:
        return max(0, int(result.get("format_warning_count") or 0))
    except (TypeError, ValueError):
        return 0


def _column_width_verification_report(
    *,
    manifest: Sequence[Mapping[str, Any]],
    request_payload: Mapping[str, Any],
    remote_result: Mapping[str, Any],
    enabled: bool,
) -> dict[str, Any]:
    if not enabled:
        return {
            "status": "NOT_ENABLED",
            "total_columns": 0,
            "local_explicit_width_count": 0,
            "auto_fit_requested_count": 0,
            "explicit_applied_count": 0,
            "auto_fit_applied_count": 0,
            "clamped_count": 0,
            "dto_match_count": 0,
            "payload_match_count": 0,
            "attempted_count": 0,
            "read_back_count": 0,
            "physical_read_back_count": 0,
            "verified_count": 0,
            "warning_count": 0,
            "failed_count": 0,
            "verified_ratio": 0.0,
            "stage_counts": {},
            "largest_differences": [],
            "representative_columns": [],
            "items": [],
        }

    payload_widths: dict[tuple[str, str], float | None] = {}
    payload_modes: dict[tuple[str, str], str] = {}
    workbook_payload = request_payload.get("workbook")
    if isinstance(workbook_payload, Mapping):
        payload_sheets = workbook_payload.get("sheets")
        if isinstance(payload_sheets, list):
            for sheet in payload_sheets:
                if not isinstance(sheet, Mapping):
                    continue
                sheet_name = str(sheet.get("sheet_name") or "")
                widths = sheet.get("column_widths")
                if not isinstance(widths, Mapping):
                    continue
                for column, width in widths.items():
                    key = (sheet_name, str(column).upper())
                    payload_widths[key] = _safe_float(width)
                    payload_modes[key] = "EXPLICIT"
                auto_fit_columns = sheet.get("auto_fit_columns")
                if isinstance(auto_fit_columns, list):
                    for column in auto_fit_columns:
                        key = (sheet_name, str(column).upper())
                        payload_modes[key] = (
                            "AUTO_FIT_WITH_LOCAL_MIN"
                            if key in payload_widths
                            else "AUTO_FIT"
                        )

    remote_items: dict[tuple[str, str], Mapping[str, Any]] = {}
    raw_column_result = remote_result.get("column_width_result")
    if isinstance(raw_column_result, Mapping):
        raw_items = raw_column_result.get("items")
        if isinstance(raw_items, list):
            for item in raw_items:
                if not isinstance(item, Mapping):
                    continue
                key = (
                    str(item.get("sheet_name") or ""),
                    str(item.get("column") or "").upper(),
                )
                if all(key):
                    remote_items[key] = item

    items: list[dict[str, Any]] = []
    stage_counts: dict[str, int] = {}
    dto_match_count = 0
    payload_match_count = 0
    read_back_count = 0
    physical_read_back_count = 0
    verified_count = 0
    warning_count = 0
    failed_count = 0
    explicit_applied_count = 0
    auto_fit_applied_count = 0
    clamped_count = 0
    for source in manifest:
        sheet_name = str(source.get("sheet_name") or "")
        column = str(source.get("column") or "").upper()
        key = (sheet_name, column)
        source_mode = str(source.get("source_mode") or "EXPLICIT").upper()
        dto_mode = str(source.get("sheet_dto_mode") or "").upper()
        payload_mode = payload_modes.get(key, "MISSING")
        local_width = _safe_float(source.get("local_workbook_width"))
        dto_width = _safe_float(source.get("sheet_dto_width"))
        if not dto_mode:
            dto_mode = "EXPLICIT" if dto_width is not None else "MISSING"
        payload_width = payload_widths.get(key)
        remote = remote_items.get(key, {})
        remote_present = bool(remote)
        remote_mode = str(remote.get("mode") or "EXPLICIT").upper()
        wps_requested_width = _safe_float(remote.get("requested_width"))
        before_column_width = _safe_float(remote.get("before_column_width"))
        remote_column_width = _safe_float(remote.get("remote_column_width"))
        before_width_points = _safe_float(remote.get("before_width_points"))
        remote_width_points = _safe_float(remote.get("remote_width_points"))
        difference = (
            abs(remote_column_width - wps_requested_width)
            if remote_column_width is not None and wps_requested_width is not None
            else None
        )
        if source_mode.startswith("AUTO_FIT"):
            dto_matches = dto_mode == source_mode and (
                local_width is None
                or _width_matches(local_width, dto_width, tolerance=0.01)
            )
            payload_matches = payload_mode == source_mode and (
                dto_width is None
                or _width_matches(dto_width, payload_width, tolerance=0.01)
            )
            remote_request_matches = remote_mode == source_mode and (
                payload_width is None
                or _width_matches(
                    payload_width,
                    _safe_float(remote.get("local_workbook_width")),
                    tolerance=0.01,
                )
            )
            remote_matches = bool(remote.get("verified")) and remote_column_width is not None
        else:
            dto_matches = (
                dto_mode == "EXPLICIT"
                and _width_matches(local_width, dto_width, tolerance=0.01)
            )
            payload_matches = (
                payload_mode == "EXPLICIT"
                and _width_matches(dto_width, payload_width, tolerance=0.01)
            )
            remote_request_matches = (
                remote_mode == "EXPLICIT"
                and _width_matches(
                    payload_width,
                    wps_requested_width,
                    tolerance=0.01,
                )
            )
            remote_matches = _width_matches(
                payload_width,
                remote_column_width,
                tolerance=0.5,
            )
        if dto_matches:
            dto_match_count += 1
        if payload_matches:
            payload_match_count += 1
        if remote_column_width is not None:
            read_back_count += 1
        if remote_width_points is not None:
            physical_read_back_count += 1

        if not dto_matches:
            classification = "WORKBOOK_DTO_WIDTH_MISMATCH"
            reason = "本地 XLSX 列宽模式或数值与 SheetDTO 不一致"
        elif not payload_matches or (remote_present and not remote_request_matches):
            classification = "WPS_PAYLOAD_WIDTH_MISMATCH"
            reason = "SheetDTO、序列化 payload 或 WPS 接收值不一致"
        elif not remote_present or not remote_matches:
            classification = "WPS_COLUMN_WIDTH_APPLY_MISMATCH"
            reason = str(remote.get("reason") or "WPS ColumnWidth 写后读回不一致")
        elif source_mode.startswith("AUTO_FIT"):
            classification = "WPS_COLUMN_WIDTH_AUTOFIT_VERIFIED"
            reason = "WPS AutoFit、本地宽度下限和布局边界计算后读回一致"
        else:
            classification = "WPS_COLUMN_WIDTH_VALUE_VERIFIED"
            reason = "本地 XLSX、SheetDTO、payload 与 WPS ColumnWidth 读回一致"

        verified = classification in {
            "WPS_COLUMN_WIDTH_VALUE_VERIFIED",
            "WPS_COLUMN_WIDTH_AUTOFIT_VERIFIED",
        }
        if verified:
            verified_count += 1
        else:
            failed_count += 1
        stage_counts[classification] = stage_counts.get(classification, 0) + 1
        if bool(remote.get("applied")):
            if remote_mode.startswith("AUTO_FIT"):
                auto_fit_applied_count += 1
            else:
                explicit_applied_count += 1
        if bool(remote.get("clamped")):
            clamped_count += 1

        physical_width_status = "READ_BACK"
        if remote_width_points is None:
            physical_width_status = "READBACK_MISSING"
            warning_count += 1
        elif (
            before_column_width is not None
            and before_width_points is not None
            and wps_requested_width is not None
            and abs(before_column_width - wps_requested_width) > 0.5
            and abs(remote_width_points - before_width_points) <= 0.01
        ):
            physical_width_status = "APPLY_MISMATCH"
            warning_count += 1

        items.append(
            {
                **dict(source),
                "payload_mode": payload_mode,
                "remote_mode": remote_mode,
                "payload_requested_width": payload_width,
                "wps_requested_width": wps_requested_width,
                "wps_auto_fit_width": _safe_float(remote.get("auto_fit_width")),
                "before_column_width": before_column_width,
                "remote_column_width": remote_column_width,
                "before_width_points": before_width_points,
                "remote_width_points": remote_width_points,
                "difference": round(difference, 4) if difference is not None else None,
                "physical_width_change_points": _safe_float(
                    remote.get("physical_width_change_points")
                ),
                "read_back": remote_column_width is not None,
                "physical_width_status": physical_width_status,
                "clamped": bool(remote.get("clamped")),
                "verified": verified,
                "classification": classification,
                "reason": reason,
            }
        )

    total = len(items)
    verified_ratio = round(verified_count / total, 4) if total else 0.0
    largest_differences = sorted(
        items,
        key=lambda item: (
            item.get("difference") is not None,
            float(item.get("difference") or 0.0),
        ),
        reverse=True,
    )[:10]
    representative_columns = [
        item
        for item in items
        if item["sheet_name"] == "轨旁AP业务"
        and item["column"] in {"A", "B", "C", "G", "H", "P"}
    ]
    status = (
        "FAILED"
        if failed_count
        else "SUCCESS_WITH_WARNINGS"
        if warning_count
        else "SUCCESS"
    )
    return {
        "status": status,
        "tolerance": 0.5,
        "total_columns": total,
        "local_explicit_width_count": sum(
            1 for item in items if item.get("local_workbook_width") is not None
        ),
        "auto_fit_requested_count": sum(
            1 for item in items if str(item.get("source_mode") or "").startswith("AUTO_FIT")
        ),
        "explicit_applied_count": explicit_applied_count,
        "auto_fit_applied_count": auto_fit_applied_count,
        "clamped_count": clamped_count,
        "dto_match_count": dto_match_count,
        "payload_match_count": payload_match_count,
        "attempted_count": _safe_int(
            raw_column_result.get("attempted_count")
            if isinstance(raw_column_result, Mapping)
            else 0
        ),
        "read_back_count": read_back_count,
        "physical_read_back_count": physical_read_back_count,
        "verified_count": verified_count,
        "warning_count": warning_count,
        "failed_count": failed_count,
        "verified_ratio": verified_ratio,
        "stage_counts": stage_counts,
        "largest_differences": largest_differences,
        "representative_columns": representative_columns,
        "items": items,
    }


def _append_column_width_report_warning(
    remote_result: dict[str, Any],
    report: Mapping[str, Any],
) -> None:
    format_results = remote_result.get("format_results")
    if not isinstance(format_results, dict):
        format_results = {}
        remote_result["format_results"] = format_results
    column_result = format_results.get("column_width")
    if not isinstance(column_result, dict):
        column_result = {}
        format_results["column_width"] = column_result
    for key in (
        "status",
        "local_explicit_width_count",
        "auto_fit_requested_count",
        "explicit_applied_count",
        "auto_fit_applied_count",
        "clamped_count",
        "attempted_count",
        "read_back_count",
        "physical_read_back_count",
        "verified_count",
        "warning_count",
        "failed_count",
        "verified_ratio",
        "stage_counts",
        "largest_differences",
        "representative_columns",
    ):
        column_result[key] = report.get(key)

    if str(report.get("status") or "") in {"SUCCESS", "NOT_ENABLED"}:
        return
    warnings = remote_result.get("format_warnings")
    if not isinstance(warnings, list):
        warnings = []
        remote_result["format_warnings"] = warnings
    warnings.append(
        {
            "sheet_name": "_NetConsoleColumnWidths",
            "feature": "column_width_verification",
            "reason": (
                f"生产列宽自动验收：验证通过 {int(report.get('verified_count') or 0)}/"
                f"{int(report.get('total_columns') or 0)}，告警 {int(report.get('warning_count') or 0)}，"
                f"失败 {int(report.get('failed_count') or 0)}"
            ),
        }
    )
    remote_result["format_warning_count"] = len(warnings)


def _safe_float(value: object) -> float | None:
    try:
        number = float(value)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return None
    return number if number == number and abs(number) != float("inf") else None


def _safe_int(value: object) -> int:
    try:
        return max(0, int(value))
    except (TypeError, ValueError):
        return 0


def _width_matches(
    left: float | None,
    right: float | None,
    *,
    tolerance: float,
) -> bool:
    return left is not None and right is not None and abs(left - right) <= tolerance


def _sheet_tab_color_probe_verified(target: WpsSyncTarget) -> bool:
    diagnostic = target.sheet_tab_color_probe_diagnostic
    return (
        str(diagnostic.get("status") or "").upper()
        in {"SUCCESS", "SUCCESS_WITH_WARNINGS"}
        and bool(diagnostic.get("sheet_tab_color_verified"))
    )


def _column_width_probe_verified(target: WpsSyncTarget) -> bool:
    diagnostic = target.column_width_probe_diagnostic
    return (
        str(diagnostic.get("status") or "").upper()
        in {"SUCCESS", "SUCCESS_WITH_WARNINGS"}
        and bool(diagnostic.get("column_width_verified"))
    )


def _logical_sheet_key(name: str) -> str:
    definition = trackside_ap_business_sheet_definition(name)
    if definition is not None:
        return definition.stable_key
    digest = content_sha256({"sheet_name": name})[:12]
    return f"sheet_{digest}"


def _sanitize_error(value: str) -> str:
    sanitized = _SAFE_ERROR_RE.sub("credential=<redacted>", str(value or ""))
    sanitized = re.sub(r"(?i)(cookie|set-cookie|x-api-key)\s*[:=]\s*\S+", "header=<redacted>", sanitized)
    sanitized = re.sub(r"<[^>]{0,200}>", " ", sanitized)
    return re.sub(r"\s+", " ", sanitized).strip()[:_MAX_ERROR_TEXT]


def _sanitize_details(value: Mapping[str, object]) -> dict[str, object]:
    result: dict[str, object] = {}
    for key, item in value.items():
        key_text = str(key)
        if re.search(r"(?i)(token|authorization|cookie|header|webhook)", key_text):
            continue
        if isinstance(item, Mapping):
            result[key_text] = _sanitize_details(item)
        elif isinstance(item, list):
            result[key_text] = [
                _sanitize_error(str(entry)) if isinstance(entry, str) else entry
                for entry in item[:20]
            ]
        elif isinstance(item, str):
            result[key_text] = _sanitize_error(item)
        elif item is None or isinstance(item, (bool, int, float)):
            result[key_text] = item
    return result


def _target_error_details(target: WpsSyncTarget, *, phase: str) -> dict[str, object]:
    return {
        "phase": phase,
        "target_code": target.target_code,
        "document_id": target.expected_document_id,
        "local_binding_id": target.binding_id,
        "expected_script_id": (
            _script_id_from_webhook(target.webhook_url)
            if target.webhook_url
            else ""
        ),
        "expected_script_version": WPS_SCRIPT_VERSIONS.get(target.target_code, ""),
        "expected_deployment_id": WPS_DEPLOYMENT_IDS.get(target.target_code, ""),
    }


def _with_binding_diagnostics(
    target: WpsSyncTarget,
    result: Mapping[str, object],
) -> dict[str, Any]:
    diagnosed = dict(result)
    remote_binding_id = str(
        result.get("remote_binding_id") or result.get("binding_id") or ""
    )
    remote_document_id = str(
        result.get("remote_document_id") or result.get("document_id") or ""
    )
    remote_site_id = str(result.get("remote_site_id") or result.get("site_id") or "")
    remote_site_name = str(
        result.get("remote_site_name") or result.get("site_name") or ""
    )
    remote_business_key = str(
        result.get("remote_business_key") or result.get("business_key") or ""
    )
    remote_target_code = str(
        result.get("remote_target_code") or result.get("target_code") or ""
    )
    remote_target_type = str(
        result.get("remote_target_type") or result.get("target_type") or ""
    )
    returned_status = str(result.get("binding_status") or "UNKNOWN").upper()

    if not remote_binding_id:
        binding_status = (
            returned_status
            if returned_status
            in {"BOUND", "UNBOUND", "LEGACY_BINDING_ID_MISMATCH", "MISMATCH"}
            else "UNKNOWN"
        )
        if target.target_type is WpsTargetType.STANDARD_SPREADSHEET and returned_status == "UNKNOWN":
            binding_status = "UNBOUND"
        binding_match = False
    else:
        document_match = remote_document_id == target.expected_document_id
        site_match = remote_site_id == target.site_id
        business_match = remote_business_key == target.business_key
        target_code_match = remote_target_code == target.target_code
        target_type_match = remote_target_type == target.target_type.value
        business_identity_matches = all(
            (
                document_match,
                site_match,
                business_match,
                target_code_match,
                target_type_match,
            )
        )
        binding_match = remote_binding_id == target.binding_id
        if business_identity_matches and binding_match:
            binding_status = "BOUND"
        elif business_identity_matches and _LEGACY_BINDING_ID_RE.fullmatch(remote_binding_id):
            binding_status = "LEGACY_BINDING_ID_MISMATCH"
        else:
            binding_status = "MISMATCH"

    document_match = bool(remote_document_id) and remote_document_id == target.expected_document_id
    site_match = bool(remote_site_id) and remote_site_id == target.site_id
    business_match = bool(remote_business_key) and remote_business_key == target.business_key
    target_code_match = bool(remote_target_code) and remote_target_code == target.target_code
    target_type_match = bool(remote_target_type) and remote_target_type == target.target_type.value
    diagnosed.update(
        {
            "binding_status": binding_status,
            "local_binding_id": target.binding_id,
            "remote_binding_id": remote_binding_id,
            "binding_id_match": binding_match,
            "remote_document_id": remote_document_id,
            "remote_site_id": remote_site_id,
            "remote_site_name": remote_site_name,
            "remote_business_key": remote_business_key,
            "remote_target_code": remote_target_code,
            "remote_target_type": remote_target_type,
            "document_match": document_match,
            "document_identity_match": document_match,
            "site_match": site_match,
            "site_identity_match": site_match,
            "business_match": business_match,
            "business_identity_match": business_match,
            "target_code_match": target_code_match,
            "target_type_match": target_type_match,
            "target_match": target_code_match and target_type_match,
        }
    )
    return diagnosed


def _binding_error_details(result: Mapping[str, object]) -> dict[str, object]:
    keys = (
        "binding_status",
        "local_binding_id",
        "remote_binding_id",
        "binding_id_match",
        "remote_document_id",
        "remote_site_id",
        "remote_business_key",
        "remote_target_code",
        "remote_target_type",
        "document_match",
        "document_identity_match",
        "site_match",
        "site_identity_match",
        "business_match",
        "business_identity_match",
        "target_code_match",
        "target_type_match",
        "target_match",
    )
    return {key: result.get(key) for key in keys if key in result}


def _operation_diagnostic(
    operation: str,
    *,
    status: str,
    message: str,
    values: Mapping[str, object],
) -> dict[str, object]:
    sanitized = _sanitize_result(values)
    diagnostic: dict[str, object] = {
        "executed_at": _now(),
        "status": str(status or "FAILED").upper(),
        "script_version": str(
            sanitized.get("remote_script_version")
            or sanitized.get("script_version")
            or ""
        ),
        "deployment_id": str(
            sanitized.get("remote_deployment_id")
            or sanitized.get("deployment_id")
            or ""
        ),
        "script_id": str(
            sanitized.get("remote_script_id")
            or sanitized.get("script_id")
            or ""
        ),
        "document_id": str(sanitized.get("document_id") or ""),
        "operation": str(operation),
        "message": _sanitize_error(message),
    }
    for key in (
        "phase",
        "http_status",
        "remote_error_code",
        "remote_message",
        "suggestion",
        "target_code",
        "runtime_capability",
        "core_verified",
        "full_replace_ready",
        "prepend_snapshot_ready",
        "append_history_ready",
        "verified_record_batch_size",
        "capabilities",
        "core_capabilities",
        "optional_capabilities",
        "capability_failures",
        "warnings",
        "sheet_order_verified",
        "sheet_move_before_verified",
        "sheet_move_after_verified",
        "system_sheet_order_verified",
        "expected_sheet_order",
        "actual_sheet_order",
        "actual_sheet_order_all",
        "sheet_tab_color_verified",
        "expected_tab_color",
        "actual_tab_color",
        "column_width_verified",
        "expected_column_widths",
        "actual_column_widths",
        "probe_sheet_visible",
        "probe_sheet",
        "probe_cleanup",
        "binding_status",
        "local_binding_id",
        "remote_binding_id",
        "binding_id_match",
        "remote_document_id",
        "remote_site_id",
        "remote_site_name",
        "remote_business_key",
        "remote_target_code",
        "remote_target_type",
        "document_match",
        "document_identity_match",
        "site_match",
        "site_identity_match",
        "business_match",
        "business_identity_match",
        "target_code_match",
        "target_type_match",
        "target_match",
    ):
        if sanitized.get(key) not in (None, ""):
            diagnostic[key] = sanitized[key]
    return diagnostic


def _effective_binding_status(target: WpsSyncTarget) -> str:
    status = str(target.binding_status or "UNKNOWN").upper()
    remote_binding_id = str(target.remote_binding_id or "")
    if (
        status != "BOUND"
        or not remote_binding_id
        or remote_binding_id == target.binding_id
    ):
        return status

    diagnostic = target.connection_diagnostic or {}
    identity_matches = (
        diagnostic.get("document_identity_match", diagnostic.get("document_match")),
        diagnostic.get("site_identity_match", diagnostic.get("site_match")),
        diagnostic.get("business_identity_match", diagnostic.get("business_match")),
        diagnostic.get("target_code_match"),
        diagnostic.get("target_type_match"),
    )
    if all(value is True for value in identity_matches):
        if _LEGACY_BINDING_ID_RE.fullmatch(remote_binding_id):
            return "LEGACY_BINDING_ID_MISMATCH"
        return "MISMATCH"
    if any(value is False for value in identity_matches):
        return "MISMATCH"
    # Historical connection rows may say BOUND because target_id used to be
    # the binding identity. Require a fresh connection test before trusting
    # that state after the stable binding ID migration.
    return "UNKNOWN"


def _remote_result_error(
    target: WpsSyncTarget,
    result: Mapping[str, object],
    token: str,
) -> WpsSyncError:
    diagnosed = _with_binding_diagnostics(target, result)
    details = {
        **_target_error_details(target, phase="SCRIPT_EXECUTION"),
        **_binding_error_details(diagnosed),
        **_sanitize_result(
            {
                key: diagnosed.get(key)
                for key in (
                    "failed_sheet",
                    "failed_operation",
                    "written_sheet_count",
                    "written_row_count",
                    "runtime_error_name",
                    "runtime_error_stack",
                    "binding_status",
                    "runtime_capability",
                    "core_verified",
                    "full_replace_ready",
                    "prepend_snapshot_ready",
                    "capabilities",
                    "core_capabilities",
                    "optional_capabilities",
                    "capability_failures",
                    "warnings",
                )
            }
        ),
    }
    return WpsSyncError(
        str(result.get("error_code") or "WPS_REMOTE_FAILED"),
        _sanitize_error(
            str(result.get("message") or "WPS 远端同步失败").replace(
                token, "<redacted>"
            )
        ),
        details=details,
    )


def _default_target_name(site_display_name: str, target_code: str) -> str:
    del target_code
    return f"{site_display_name}轨旁AP业务-WPS云文档"


def _is_legacy_default_target_name(value: str, target_code: str) -> bool:
    del target_code
    text = str(value or "").replace(" ", "").strip()
    return text.endswith("轨旁AP业务-普通在线表格")


def _unwrap_wps_sync_task_response(
    response: WpsHttpResponse,
    target: WpsSyncTarget,
    *,
    token: str = "",
) -> dict[str, Any]:
    body = response.body
    if not isinstance(body, Mapping):
        raise WpsSyncError(
            "WPS_RESPONSE_INVALID",
            "WPS 返回结构无效",
            details=_target_error_details(target, phase="SCRIPT_EXECUTION"),
        )

    # Keep compatibility with existing direct protocol-object test doubles while
    # preferring the official sync_task execution envelope in production.
    if "protocol_version" in body and "data" not in body:
        return dict(body)

    status = str(body.get("status") or "").strip().casefold()
    details = _target_error_details(target, phase="SCRIPT_EXECUTION")
    if status not in {"finished", "success", "completed"}:
        details["execution_status"] = status or "missing"
        raise WpsSyncError(
            "WPS_SCRIPT_STATUS_INVALID",
            f"WPS 脚本执行状态未完成：{status or '未知'}",
            details=details,
        )

    outer_error = body.get("error")
    if outer_error:
        error_details = body.get("error_details")
        remote_message = _remote_error_message(outer_error, error_details)
        if token:
            remote_message = remote_message.replace(token, "<redacted>")
        raise WpsSyncError(
            "WPS_SCRIPT_EXECUTION_FAILED",
            f"WPS 脚本执行失败：{remote_message}",
            details={**details, "remote_message": remote_message},
        )

    data = body.get("data")
    if not isinstance(data, Mapping):
        raise WpsSyncError(
            "WPS_SCRIPT_RESULT_EMPTY",
            "WPS 脚本响应缺少 data.result",
            details=details,
        )
    raw_result = data.get("result")
    if raw_result is None:
        raise WpsSyncError(
            "WPS_SCRIPT_RESULT_EMPTY",
            "WPS 脚本返回结果为空",
            details=details,
        )
    if isinstance(raw_result, Mapping):
        return dict(raw_result)
    if not isinstance(raw_result, str) or not raw_result.strip():
        raise WpsSyncError(
            "WPS_SCRIPT_RESULT_INVALID",
            "WPS 脚本返回结果不是 JSON 对象",
            details=details,
        )
    try:
        decoded = json.loads(raw_result)
    except json.JSONDecodeError:
        raise WpsSyncError(
            "WPS_SCRIPT_RESULT_INVALID",
            "WPS 脚本返回结果不是有效 JSON",
            details=details,
        ) from None
    if not isinstance(decoded, Mapping):
        raise WpsSyncError(
            "WPS_SCRIPT_RESULT_INVALID",
            "WPS 脚本返回结果不是 JSON 对象",
            details=details,
        )
    return dict(decoded)


def _remote_error_message(value: object, details: object = None) -> str:
    values: list[str] = []
    if isinstance(value, Mapping):
        for key in ("message", "msg", "error_description", "error"):
            if value.get(key):
                values.append(str(value[key]))
    elif value:
        values.append(str(value))
    if isinstance(details, Mapping):
        for key in ("name", "message", "msg", "error_description"):
            if details.get(key):
                values.append(str(details[key]))
    return _sanitize_error("；".join(values) or "远端未提供具体原因")


def _http_error_from_response(
    exc: urllib.error.HTTPError,
    target: WpsSyncTarget,
    token: str,
) -> WpsSyncError:
    raw = b""
    try:
        raw = exc.read(_MAX_ERROR_BODY_BYTES + 1)
    except (OSError, ValueError):
        pass
    too_large = len(raw) > _MAX_ERROR_BODY_BYTES
    raw = raw[:_MAX_ERROR_BODY_BYTES]
    remote_code, remote_message, response_format, request_id = _extract_http_error(raw)
    if token:
        remote_code = remote_code.replace(token, "<redacted>")
        remote_message = remote_message.replace(token, "<redacted>")
        request_id = request_id.replace(token, "<redacted>")
    if too_large:
        remote_message = "WPS 错误响应正文超过安全限制"
        response_format = "oversized"
    code = _classify_http_error(int(exc.code), remote_code, remote_message)
    status_text = int(exc.code)
    fallback = {
        401: "WPS 令牌无效或已过期",
        403: "请检查脚本令牌、文档权限和文档共享脚本状态",
        404: "请检查 webhook 地址和脚本发布状态",
        429: "WPS 请求过于频繁，请稍后重试",
    }.get(status_text, "请检查 WPS 服务状态和网络连接")
    if code == "WPS_SCRIPT_NOT_AVAILABLE":
        fallback = "请在对应文档中发布文档共享脚本，并确认 webhook 与该脚本匹配"
    elif code == "WPS_DOCUMENT_PERMISSION_DENIED":
        fallback = "请确认当前令牌所属账号拥有该文档的访问和编辑权限"
    elif code == "WPS_SCRIPT_PERMISSION_DENIED":
        fallback = "请确认 webhook 对应脚本已发布且允许通过脚本令牌执行"
    elif code == "WPS_TOKEN_INVALID":
        fallback = "请在 WPS 重新生成令牌，并在对应目标配置中重新保存"
    message = (
        f"WPS 请求失败（HTTP {status_text}），错误码：{remote_code or code}，"
        f"原因：{remote_message or '远端未提供具体原因'}。建议：{fallback}"
    )
    details = {
        **_target_error_details(target, phase="HTTP_AUTH"),
        "http_status": status_text,
        "remote_error_code": remote_code,
        "remote_message": remote_message,
        "response_format": response_format,
        "request_id": request_id,
        "suggestion": fallback,
    }
    # Explicitly include the token in sanitization input without retaining it.
    safe_message = _sanitize_error(message.replace(token, "<redacted>")) if token else message
    return WpsSyncError(code, safe_message, details=details)


def _extract_http_error(raw: bytes) -> tuple[str, str, str, str]:
    if not raw:
        return "", "", "empty", ""
    text = _sanitize_error(raw.decode("utf-8", errors="replace"))
    try:
        decoded = json.loads(raw.decode("utf-8-sig"))
    except (UnicodeDecodeError, json.JSONDecodeError):
        return "", text, "text", ""
    if not isinstance(decoded, Mapping):
        return "", text, "json", ""
    nested = decoded.get("error_details")
    code = ""
    for key in ("code", "error_code"):
        if decoded.get(key):
            code = _sanitize_error(str(decoded[key]))
            break
    message = _remote_error_message(decoded, nested)
    request_id = ""
    for key in ("request_id", "trace_id"):
        if decoded.get(key):
            request_id = _sanitize_error(str(decoded[key]))
            break
    return code, message, "json", request_id


def _classify_http_error(status: int, remote_code: str, remote_message: str) -> str:
    haystack = f"{remote_code} {remote_message}".casefold()
    if status == 401 or re.search(r"token|api[_ -]?key|expired|过期|令牌", haystack):
        return "WPS_TOKEN_INVALID"
    if status == 403:
        if re.search(r"account|账号|user.*document|账户.*文档|不一致", haystack):
            return "WPS_ACCOUNT_DOCUMENT_MISMATCH"
        if re.search(r"script.*(permission|access)|脚本.*(权限|访问)", haystack):
            return "WPS_SCRIPT_PERMISSION_DENIED"
        if re.search(r"script.*(not found|unavailable|nil|empty)|脚本.*(不存在|不可用|为空)", haystack):
            return "WPS_SCRIPT_NOT_AVAILABLE"
        if re.search(r"document|file|sheet|文档|文件|表格", haystack):
            return "WPS_DOCUMENT_PERMISSION_DENIED"
        return "WPS_REMOTE_FORBIDDEN"
    if status == 404:
        return "WPS_WEBHOOK_NOT_FOUND"
    if status == 429:
        return "WPS_REMOTE_RATE_LIMITED"
    if status >= 500:
        return "WPS_REMOTE_UNAVAILABLE"
    return f"WPS_HTTP_{status}"


def _validate_wps_url(value: str, *, kind: str) -> str:
    normalized = str(value or "").strip()
    try:
        parsed = urlsplit(normalized)
    except ValueError:
        raise WpsSyncError("WPS_URL_INVALID", "WPS 地址格式无效") from None
    hostname = str(parsed.hostname or "").casefold().rstrip(".")
    if parsed.scheme.casefold() != "https" or not hostname:
        raise WpsSyncError("WPS_URL_INVALID", "WPS 地址必须使用 HTTPS")
    if parsed.username or parsed.password or parsed.query or parsed.fragment:
        raise WpsSyncError("WPS_URL_INVALID", "WPS 地址不得包含凭据、查询参数或片段")
    try:
        ipaddress.ip_address(hostname)
    except ValueError:
        pass
    else:
        raise WpsSyncError("WPS_URL_INVALID", "WPS 地址不得使用 IP 地址")
    if hostname != "kdocs.cn" and not hostname.endswith(".kdocs.cn"):
        raise WpsSyncError("WPS_URL_INVALID", "WPS 地址必须属于 kdocs.cn")
    try:
        port = parsed.port
    except ValueError:
        raise WpsSyncError("WPS_URL_INVALID", "WPS 地址端口无效") from None
    if port not in (None, 443):
        raise WpsSyncError("WPS_URL_INVALID", "WPS 地址仅允许标准 HTTPS 端口")
    if kind == "webhook" and not _WPS_WEBHOOK_PATH_RE.fullmatch(parsed.path):
        raise WpsSyncError(
            "WPS_WEBHOOK_INVALID",
            "AirScript webhook 路径必须包含有效的文档 ID、脚本 ID 和 /sync_task",
        )
    if kind == "document" and not parsed.path.startswith("/l/"):
        raise WpsSyncError("WPS_DOCUMENT_URL_INVALID", "在线文档地址必须使用 kdocs.cn/l/ 链接")
    return normalized


def parse_wps_webhook(value: str) -> WpsWebhookEndpoints:
    normalized = _validate_wps_url(value, kind="webhook")
    parsed = urlsplit(normalized)
    matched = _WPS_WEBHOOK_PATH_RE.fullmatch(parsed.path)
    if matched is None:
        raise WpsSyncError("WPS_WEBHOOK_INVALID", "AirScript webhook 路径无效")
    file_id = matched.group("file_id")
    script_id = matched.group("script_id")
    base = (parsed.scheme, parsed.netloc)
    return WpsWebhookEndpoints(
        host=str(parsed.hostname or ""),
        file_id=file_id,
        script_id=script_id,
        sync_task_url=normalized,
        async_task_url=urlunsplit(
            (*base, f"/api/v3/ide/file/{file_id}/script/{script_id}/task", "", "")
        ),
        task_status_url=urlunsplit((*base, "/api/v3/script/task", "", "")),
    )


def _document_id_from_webhook(value: str) -> str:
    return parse_wps_webhook(value).file_id


def _script_id_from_webhook(value: str) -> str:
    return parse_wps_webhook(value).script_id


def _require_standard_target_code(target_code: str) -> None:
    if str(target_code or "").strip() != STANDARD_TARGET_CODE:
        raise WpsSyncError(
            "WPS_TARGET_UNSUPPORTED",
            "当前版本仅支持 WPS 普通在线表格同步",
        )


def _validate_target_configuration(target: WpsSyncTarget) -> None:
    _validate_wps_url(target.document_open_url, kind="document")
    webhook = _validate_wps_url(target.webhook_url, kind="webhook")
    if _document_id_from_webhook(webhook) != target.expected_document_id:
        raise WpsSyncError("WPS_DOCUMENT_ID_MISMATCH", "webhook 文档身份与已保存配置不一致")
    _script_id_from_webhook(webhook)


def _assert_runtime_identity(target: WpsSyncTarget) -> None:
    """Reject a VERIFIED probe if its persisted deployment identity is stale."""
    if not target.last_runtime_probe_at:
        return
    expected_script_id = _script_id_from_webhook(target.webhook_url)
    expected = {
        "document_id": target.expected_document_id,
        "script_id": expected_script_id,
        "script_version": WPS_SCRIPT_VERSIONS.get(target.target_code, ""),
        "deployment_id": WPS_DEPLOYMENT_IDS.get(target.target_code, ""),
    }
    actual = {
        "document_id": target.runtime_probe_document_id,
        "script_id": target.runtime_probe_script_id,
        "script_version": target.runtime_probe_script_version,
        "deployment_id": target.runtime_probe_deployment_id,
    }
    mismatches = {
        key: {"expected": expected[key], "actual": actual[key]}
        for key in expected
        if actual[key] and actual[key] != expected[key]
    }
    missing = [key for key in expected if not actual[key]]
    if mismatches or missing:
        raise WpsSyncError(
            "WPS_DEPLOYMENT_IDENTITY_MISMATCH",
            f"WPS 运行时部署身份已过期，请重新执行连接测试和写入探针：{target.target_name}",
            details={
                **_target_error_details(target, phase="DEPLOYMENT_IDENTITY"),
                "expected_script_id": expected_script_id,
                "mismatches": mismatches,
                "missing": missing,
            },
        )


def _assert_standard_sync_readiness(target: WpsSyncTarget) -> None:
    # Diagnostics and the remote identity snapshot are evidence for the UI and
    # task history. The actual write gate remains the current runtime probe plus
    # the document binding checks above; stale historical diagnostics must not
    # lock a target after a successful probe.
    if target.target_type is WpsTargetType.STANDARD_SPREADSHEET:
        _assert_runtime_identity(target)


def _sanitize_result(value: Mapping[str, object]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, item in value.items():
        if "token" in str(key).casefold() or "header" in str(key).casefold():
            continue
        if isinstance(item, str):
            result[str(key)] = _sanitize_error(item)
        elif isinstance(item, Mapping):
            result[str(key)] = _sanitize_result(item)
        elif isinstance(item, list):
            result[str(key)] = [
                _sanitize_result(entry) if isinstance(entry, Mapping) else entry
                for entry in item
            ]
        else:
            result[str(key)] = item
    return result


def _mask_remote_task_id(value: str) -> str:
    normalized = str(value or "").strip()
    if not normalized:
        return ""
    if len(normalized) <= 12:
        return f"{normalized[:4]}...{normalized[-2:]}"
    return f"{normalized[:8]}...{normalized[-4:]}"


def _now() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


__all__ = [
    "DEFAULT_TARGETS",
    "STANDARD_TARGET_CODE",
    "WPS_SYNC_OWNER",
    "WPS_SYNC_TASK_TYPE",
    "WPS_DEPLOYMENT_IDS",
    "WPS_RUNTIME_CAPABILITIES",
    "WPS_SCRIPT_VERSIONS",
    "WPS_STANDARD_FORMAT_MIRROR_ENABLED",
    "WPS_REMOTE_TASK_MAX_WAIT_SECONDS",
    "TracksideApWpsSyncService",
    "WpsAirScriptClient",
    "WpsRemoteTask",
    "WpsStandardSpreadsheetAdapter",
    "WpsSyncError",
    "WpsWebhookEndpoints",
    "parse_wps_webhook",
    "workbook_dto_from_xlsx",
]
