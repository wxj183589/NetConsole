from __future__ import annotations

from collections.abc import Iterable
from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBTLE_FILL = PatternFill("solid", fgColor="F3F4F6")
HEADER_FONT = Font(bold=True, color="FFFFFF")
NORMAL_FONT = Font(color="111827")
ACTIVE_FONT = Font(bold=True, color="15803D")
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def display_width(value: object, *, limit: int = 160) -> int:
    text = str(value or "")
    width = 0
    for char in text[:limit]:
        width += 2 if ord(char) > 127 else 1
    return width


def excel_width(value: object, *, minimum: float = 10.0, maximum: float = 42.0) -> float:
    return min(max(display_width(value) + 2, minimum), maximum)


def format_link_state(value: object) -> str:
    text = str(value or "").strip()
    upper = text.upper()
    if upper.startswith("ACTIVE"):
        return "主链路"
    if "STANDBY" in upper or "BACKUP" in upper:
        return "备链"
    if upper in {"UNKNOWN", ""}:
        return "未知" if upper else ""
    return text


def format_empty(value: object, *, placeholder: str = "-") -> object:
    if value is None or value == "":
        return placeholder
    return value


def append_rows_sheet(
    workbook,
    title: str,
    headers: list[str] | tuple[str, ...],
    rows: Iterable[Iterable[Any]],
    *,
    empty_message: str = "本 Sheet 暂无可用数据。",
    max_width: float = 48.0,
):
    sheet = workbook.create_sheet(title)
    sheet.append(list(headers))
    written = 0
    for row in rows:
        sheet.append([format_empty(value, placeholder="") for value in row])
        written += 1
    if written == 0:
        sheet.append([empty_message, *["" for _ in headers[1:]]])
    apply_standard_sheet_style(sheet, max_width=max_width)
    return sheet


def apply_standard_sheet_style(sheet, *, max_scan_rows: int = 2000, max_width: float = 48.0) -> None:
    if sheet.max_row >= 1:
        for cell in sheet[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            cell.border = THIN_BORDER
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.font = NORMAL_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = THIN_BORDER
    sheet.freeze_panes = "A2"
    if sheet.max_column and sheet.max_row:
        sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
    apply_column_widths(sheet, max_scan_rows=max_scan_rows, max_width=max_width)


def apply_column_widths(sheet, *, max_scan_rows: int = 2000, min_width: float = 10.0, max_width: float = 48.0) -> None:
    max_row = min(sheet.max_row or 1, max_scan_rows)
    for column_index in range(1, (sheet.max_column or 0) + 1):
        width = min_width
        for row_index in range(1, max_row + 1):
            width = max(width, excel_width(sheet.cell(row_index, column_index).value, minimum=min_width, maximum=max_width))
        sheet.column_dimensions[get_column_letter(column_index)].width = width
