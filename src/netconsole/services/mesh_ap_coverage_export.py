from __future__ import annotations

from pathlib import Path
from typing import Callable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from netconsole.models.api.mesh_analysis import MeshApCoverageAuditDTO, MeshApCoverageRowDTO


_HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
_HEADERS = (
    ("AP 名称", "ap_name"),
    ("物理 AP MAC", "physical_ap_mac"),
    ("Peer Radio MAC", "radio_mac"),
    ("所属站点", "station"),
    ("所属区间", "section"),
    ("线路方向", "direction"),
    ("FIT-AP 状态", "fit_ap_status"),
    ("来源 A 出现", "seen_in_source_a"),
    ("来源 B 出现", "seen_in_source_b"),
    ("ACTIVE 次数", "active_count"),
    ("STANDBY 次数", "standby_count"),
    ("LinkCnt=2 次数", "triangle_link_count"),
    ("首次出现", "first_seen"),
    ("最后出现", "last_seen"),
    ("Identity 状态", "identity_status"),
    ("Identity 说明", "identity_reason"),
    ("经过范围", "in_observed_route_scope"),
    ("排除原因", "exclude_reason"),
    ("结果", "result"),
    ("说明", "description"),
)


def export_mesh_ap_coverage_audit_xlsx(
    output_path: Path,
    audit: MeshApCoverageAuditDTO,
    *,
    should_cancel: Callable[[], bool] | None = None,
) -> int:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "核查摘要"
    source_a, source_b = audit.sources
    values = (
        ("局点", audit.site_id),
        ("来源 A", f"{source_a.mr_name} · {source_a.original_filename}"),
        ("来源 A 时间范围", f"{source_a.first_sample_time or '—'} — {source_a.last_sample_time or '—'}"),
        ("来源 A Peer Radio / 物理 AP", f"{source_a.distinct_peer_radio_count} / {source_a.distinct_canonical_ap_count}"),
        ("来源 B", f"{source_b.mr_name} · {source_b.original_filename}"),
        ("来源 B 时间范围", f"{source_b.first_sample_time or '—'} — {source_b.last_sample_time or '—'}"),
        ("来源 B Peer Radio / 物理 AP", f"{source_b.distinct_peer_radio_count} / {source_b.distinct_canonical_ap_count}"),
        ("AP Identity scope / revision", f"{audit.identity_summary.identity_scope} / {audit.identity_summary.identity_revision}"),
        ("AP Identity 索引状态", audit.identity_summary.index_status),
        ("MESH Peer Radio / 物理 AP", f"{audit.identity_summary.mesh_distinct_peer_radio_count} / {audit.identity_summary.mesh_distinct_canonical_ap_count}"),
        ("已持久化 Identity 命中", audit.identity_summary.persisted_matched_count),
        ("Identity fallback 命中 / 请求 / 未匹配", f"{audit.identity_summary.fallback_matched_count} / {audit.identity_summary.fallback_requested_count} / {audit.identity_summary.fallback_unmatched_count}"),
        ("默认范围口径", "本次经过范围" if audit.summary.route_scope_mode == "observed_route" else "全正线（缺少可用经过范围）"),
        ("正线 FIT-AP", audit.summary.expected_mainline_count),
        ("本次范围 FIT-AP", audit.summary.expected_route_scope_count),
        ("已连接", audit.summary.connected_count),
        ("未连接", audit.summary.unconnected_count),
        ("资料未匹配", audit.summary.unmatched_observed_count),
        ("已排除非正线", audit.summary.excluded_count),
        ("覆盖率", f"{audit.summary.coverage_percent:.2f}%"),
    )
    for row_index, (label, value) in enumerate(values, start=1):
        summary_sheet.cell(row_index, 1, label)
        summary_sheet.cell(row_index, 2, value)
        summary_sheet.cell(row_index, 1).font = Font(bold=True)
    summary_sheet.column_dimensions["A"].width = 22
    summary_sheet.column_dimensions["B"].width = 80

    for title, rows in (
        ("未连接 AP", audit.unconnected),
        ("已连接 AP", audit.connected),
        ("资料未匹配", audit.unmatched),
        ("排除 AP", audit.excluded),
    ):
        if should_cancel and should_cancel():
            raise MeshApCoverageExportCancelled("导出已取消")
        sheet = workbook.create_sheet(title)
        _write_rows(sheet, rows)
    workbook.save(output_path)
    return len(audit.connected) + len(audit.unconnected) + len(audit.unmatched) + len(audit.excluded)


class MeshApCoverageExportCancelled(Exception):
    pass


def _write_rows(sheet, rows: list[MeshApCoverageRowDTO]) -> None:
    sheet.append([label for label, _field in _HEADERS])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
    for row in rows:
        values = []
        for _label, field in _HEADERS:
            value = getattr(row, field)
            if field in {"seen_in_source_a", "seen_in_source_b", "in_observed_route_scope"}:
                value = "是" if value else "否"
            values.append(value or "")
        sheet.append(values)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, (label, _field) in enumerate(_HEADERS, start=1):
        width = min(42, max(12, len(label) + 4))
        sheet.column_dimensions[get_column_letter(index)].width = width
