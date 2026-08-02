from __future__ import annotations

from collections.abc import Iterable, Mapping, Sequence


def weighted_text_length(value: object) -> int:
    length = 0
    for char in str(value or ""):
        length += 2 if ord(char) > 127 else 1
    return length


def excel_column_width(value: object, minimum: float = 8.0, maximum: float = 60.0, field: str | None = None, header: str | None = None) -> float:
    field_min, field_max = excel_column_width_bounds(field, header)
    effective_min = max(minimum, field_min)
    effective_max = min(maximum, field_max) if maximum else field_max
    width = weighted_text_length(value) * 1.15 + 3
    return max(effective_min, min(width, effective_max))


def excel_column_width_bounds(field: str | None = None, header: str | None = None) -> tuple[float, float]:
    field_text = str(field or "").casefold()
    header_text = str(header or "").casefold()
    text = f"{field_text} {header_text}"
    if any(token in text for token in ("序号", "source_line_number", "line_number", "record_seq")):
        return 8.0, 12.0
    if any(token in text for token in ("mac", "peermac")):
        return 18.0, 22.0
    if any(token in text for token in ("time", "时间", "duration", "时长", "updated_at", "created_at", "collected_at", "timestamp", "datetime")):
        return 18.0, 26.0
    if any(token in text for token in ("source_file", "archived_filename", "file", "文件", "路径", "path")):
        return 30.0, 80.0
    if any(token in text for token in ("状态", "state", "result", "结果")):
        return 10.0, 18.0
    if any(token in text for token in ("序列号", "serial_number", "serial no", "serialno")):
        return 14.0, 32.0
    if any(token in text for token in ("连接端口", "邻居端口", "interface", "端口", "port")):
        return 18.0, 48.0
    if any(token in text for token in ("软件版本", "software_version", "version")):
        return 18.0, 48.0
    if any(token in text for token in ("数据完整性", "备注", "说明", "description", "remark", "note")):
        return 18.0, 40.0
    if any(token in text for token in ("rssi", "busy", "采样点数", "sample_count", "链路数", "link_count")):
        return 10.0, 14.0
    if field_text in {"radio_id", "radio_count"} or header_text in {"radio id", "radio数量"}:
        return 8.0, 12.0
    if any(token in text for token in ("ap名称", "ap name", "peer_ap_name", "站点", "station", "peer_site")):
        return 18.0, 32.0
    return 8.0, 40.0


def calculate_excel_column_widths(
    headers: Sequence[object],
    rows: Iterable[Mapping[str, object | None] | Sequence[object]],
    fields: Sequence[str] | None = None,
) -> list[float]:
    widths = [
        excel_column_width(header, field=fields[index] if fields and index < len(fields) else None, header=str(header))
        for index, header in enumerate(headers)
    ]
    for row in rows:
        for index in range(len(widths)):
            field = fields[index] if fields and index < len(fields) else str(headers[index])
            if isinstance(row, Mapping):
                value = row.get(field)
            else:
                value = row[index] if index < len(row) else None
            widths[index] = max(widths[index], excel_column_width(value, field=field, header=str(headers[index])))
    return widths


def apply_worksheet_column_widths(
    sheet,
    headers: Sequence[object],
    rows: Iterable[Mapping[str, object | None] | Sequence[object]],
    fields: Sequence[str] | None = None,
    maximum: float = 60.0,
) -> None:
    from openpyxl.utils import get_column_letter

    for index, width in enumerate(calculate_excel_column_widths(headers, rows, fields), start=1):
        sheet.column_dimensions[get_column_letter(index)].width = min(width, maximum)


def apply_worksheet_autofit(sheet, maximum: float = 60.0) -> None:
    from openpyxl.utils import get_column_letter

    for column_index in range(1, sheet.max_column + 1):
        header_value = sheet.cell(row=1, column=column_index).value
        width = excel_column_width(header_value, header=str(header_value or ""), maximum=maximum)
        for cell in sheet[get_column_letter(column_index)]:
            width = max(width, excel_column_width(cell.value, header=str(header_value or ""), maximum=maximum))
        sheet.column_dimensions[get_column_letter(column_index)].width = min(width, maximum)
