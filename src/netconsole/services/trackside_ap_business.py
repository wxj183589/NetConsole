from __future__ import annotations

from collections.abc import Mapping
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from ipaddress import ip_address
from pathlib import Path
import re
from time import perf_counter

from netconsole.adapters.trackside_switch import resolve_trackside_switch_adapter
from netconsole.core import app_logger
from netconsole.core.ap_optical_capability import (
    OPTICAL_NOT_APPLICABLE_REASON,
    OPTICAL_NOT_APPLICABLE_STATUS,
    is_ap_optical_applicable,
)
from netconsole.core.optical_severity_engine import (
    classify_optical_health,
    compute_optical_severity,
    compute_zte_optical_severity,
    display_optical_status,
    is_optical_health_abnormal,
    normalize_zte_optical_record,
    worse_optical_severity,
)
from netconsole.core.sources.switch_source import build_switch_data_lookup
from netconsole.models.device import Device, is_device_eligible_for_automatic_collection
from netconsole.models.trackside_switch import CommandCapabilityState
from netconsole.parsers.h3c.ac.state_mapper import (
    classify_fit_ap_state,
    normalize_fit_ap_state_token,
)
from netconsole.parsers.h3c.ac.wlan_ap_unauthenticated_parser import (
    WLAN_AP_UNAUTHENTICATED_SOURCE,
)
from netconsole.repositories.device_group_repository import DeviceGroupRepository
from netconsole.services.ap_online_overview import (
    AP_ONLINE_OVERVIEW_COLUMNS,
    overview_row_fill,
)
from netconsole.services.offline_ap_ledger import (
    OFFLINE_AP_LEDGER_COLUMNS,
    OFFLINE_AP_STATS_COLUMNS,
    write_offline_ap_ledger_sheet,
    write_offline_ap_stats_sheet,
)
from netconsole.services.ac.fit_ap_resource_identity import coalesce_fit_ap_resource_rows
from netconsole.services.ap_business_optical import (
    AP_BUSINESS_RX_MIN_DBM,
    evaluate_ap_business_rx,
    evaluate_dual_rx_business_detail,
)
from netconsole.services.rail_transit.trackside_ap_runtime_snapshot import (
    TracksideApRuntimeSnapshot,
    classify_lldp_history_status,
    deduplicate_lldp_snapshot_rows,
    select_latest_lldp_snapshot_rows,
)
from netconsole.services.ap_identity.normalizers import format_mac, normalize_mac_key
from netconsole.utils.interface_normalize import display_interface_name, normalize_interface_name
from netconsole.utils.interface_sort import interface_sort_key
from netconsole.utils.natural_sort import natural_text_key
from netconsole.utils.station_normalize import normalize_station_value


TRACKSIDE_ATTENUATION_SAMPLE_WINDOW = timedelta(minutes=30)
TRACKSIDE_RX_NORMAL_MIN_DBM = AP_BUSINESS_RX_MIN_DBM
_AP_OFFLINE_REASONS = frozenset({"switch_offline", "ac_idle", "ac_offline"})
_PRESERVED_SWITCH_MODULE_STATUSES = {
    "abnormal",
    "unverified",
    "dom_unavailable",
    "offline",
    "no_module",
}
_TRACKSIDE_NATIVE_PVID_SEGMENT_RE = re.compile(
    r"^native\s*/\s*pvid\s*[:=]?\s*\d{1,4}$",
    re.IGNORECASE,
)
_AP_IDENTITY_AMBIGUOUS = "_ambiguous"
_TREATMENT_SERIAL_IDENTITY_SOURCES = frozenset(
    {"trackside_row", "fit_ap_resource"}
)


def normalize_mac(value: object) -> str | None:
    """Compatibility display helper for trackside AP user-facing rows."""

    return format_mac(value) or None


TRACKSIDE_AP_BUSINESS_INTERNAL_FIELDS = {
    "host",
    "host_address",
    "management_ip",
    "ap_ip",
    "source_device",
    "collection_status",
    "offline_reason",
    "status_reason",
    "data_source",
    "switch_collection_status",
    "ap_rx_low_alarm",
    "ap_rx_low_warning",
    "switch_system_name",
    "switch_primary_address",
    "switch_backup_address",
    "switch_identity",
}

TRACKSIDE_AP_BUSINESS_VISIBLE_COLUMNS = (
    ("ac.station", "site"),
    ("ac.indoor_switch", "device_name"),
    ("details.interface_name", "interface_name"),
    ("details.link", "link_status"),
    ("details.port_type", "port_type"),
    ("details.port_description", "description"),
    ("details.pvid", "pvid"),
    ("trackside.planned_management_vlan", "planned_management_vlan"),
    ("trackside.pvid_plan_status", "pvid_plan_status"),
    ("trackside.vlan_group", "vlan_group_name"),
    ("details.vlan", "vlan"),
    ("ac.indoor_switch_rx_power", "switch_rx_power"),
    ("trackside.switch_optical_status", "switch_optical_status"),
    ("ac.ap_mac", "ap_mac"),
    ("ac.ap_name", "ap_name"),
    ("ac.ap_side_rx_power", "ap_rx_power"),
    ("trackside.ap_device_optical_status", "ap_device_optical_status"),
    ("trackside.ap_optical_status", "ap_optical_status"),
    ("trackside.ap_business_threshold", "ap_business_threshold_dbm"),
    ("trackside.ap_business_reason", "ap_business_reason"),
    ("trackside_ap.last_collected_at", "updated_at"),
)

TRACKSIDE_AP_BUSINESS_COLUMNS = TRACKSIDE_AP_BUSINESS_VISIBLE_COLUMNS

TRACKSIDE_AP_BUSINESS_EXPORT_COLUMNS = (
    ("ac.station", "site"),
    ("ac.indoor_switch", "device_name"),
    ("details.interface_name", "interface_name"),
    ("details.link", "link_status"),
    ("ac.indoor_switch_rx_power", "switch_rx_power"),
    ("trackside.switch_optical_status", "switch_optical_status"),
    ("ac.ap_mac", "ap_mac"),
    ("ac.ap_name", "ap_name"),
    ("ac.ap_side_rx_power", "ap_rx_power"),
    ("trackside.ap_device_optical_status", "ap_device_optical_status"),
    ("trackside.ap_optical_status", "ap_optical_status"),
    ("trackside_ap.last_collected_at", "updated_at"),
    ("trackside.ap_business_threshold", "ap_business_threshold_dbm"),
    ("trackside.ap_business_reason", "ap_business_reason"),
)

CURRENT_OPTICAL_ABNORMAL_COLUMNS = (
    ("ac.station", "site"),
    ("ac.indoor_switch", "device_name"),
    ("details.interface_name", "interface_name"),
    ("details.link", "link_status"),
    ("ac.indoor_switch_rx_power", "switch_rx_power"),
    ("trackside.switch_optical_status", "switch_optical_status"),
    ("ac.ap_mac", "ap_mac"),
    ("ac.ap_name", "ap_name"),
    ("ac.ap_side_rx_power", "ap_rx_power"),
    ("trackside.ap_device_optical_status", "ap_device_optical_status"),
    ("trackside.export.ap_online_status", "ap_online_status"),
    ("trackside.export.abnormal_side", "side"),
    ("trackside.export.abnormal_level", "level"),
    ("trackside.export.abnormal_reason", "reason"),
    ("trackside_ap.last_collected_at", "updated_at"),
    ("trackside.ap_business_threshold", "ap_business_threshold_dbm"),
    ("trackside.export.abnormal_detail", "detail"),
)

NEW_ONLINE_AP_OVERVIEW_COLUMNS = (
    ("trackside.export.station", "site"),
    ("trackside.export.switch", "device_name"),
    ("details.interface_name", "interface_name"),
    ("details.link", "link_status"),
    ("details.port_type", "port_type"),
    ("details.port_description", "description"),
    ("details.pvid", "pvid"),
    ("details.vlan", "vlan"),
    ("ac.indoor_switch_rx_power", "switch_rx_power"),
    ("trackside.switch_optical_status", "switch_optical_status"),
    ("ac.ap_mac", "ap_mac"),
    ("ac.ap_name", "ap_name"),
    ("ac.ap_side_rx_power", "ap_rx_power"),
    ("trackside.ap_device_optical_status", "ap_device_optical_status"),
    ("trackside.ap_optical_status", "ap_optical_status"),
    ("trackside.ap_business_threshold", "ap_business_threshold_dbm"),
    ("trackside.ap_business_reason", "ap_business_reason"),
    ("trackside_ap.last_collected_at", "updated_at"),
    ("ac.register_status", "register_status"),
    ("ac.new_online_status", "new_online_status"),
    ("trackside.export.identity_source", "identity_source"),
    ("ac.new_online_source", "source"),
    ("ac.current_unauthenticated", "current_unauthenticated"),
    ("ac.current_resource_exists", "current_resource_exists"),
    ("ac.last_unauthenticated_at", "last_unauthenticated_at"),
    ("trackside.export.first_seen_at", "first_seen_at"),
    ("trackside.export.ac_device", "ac_device_name"),
    ("APID", "apid"),
    ("ap.ip_address", "ap_ip"),
    ("ap.model", "model"),
    ("ap.serial_number", "serial_number"),
    ("ac.group_name", "group_name"),
    ("ac.state", "state_display"),
    ("trackside.export.identity_entity_id", "identity_entity_id"),
    ("trackside.export.baseline_collected_at", "baseline_collected_at"),
    ("trackside.export.current_collected_at", "current_collected_at"),
    ("trackside.export.suggestion", "suggestion"),
)

TRACKSIDE_AP_UNMATCHED_ONLINE_COLUMNS = (
    ("trackside.export.unmatched_ap_name", "ap_name"),
    ("trackside.export.unmatched_ap_mac", "mac"),
    ("trackside.export.unmatched_ac_status", "ac_status"),
    ("trackside.export.unmatched_runtime_station", "runtime_station_text"),
    ("trackside.export.unmatched_association_status", "association_status"),
    ("trackside.export.unmatched_reason_code", "reason_code"),
    ("trackside.export.unmatched_fit_ap_collected_at", "fit_ap_collected_at"),
    ("trackside.export.unmatched_lldp_collected_at", "lldp_collected_at"),
    ("trackside.export.unmatched_lldp_candidate_count", "lldp_candidate_count"),
    ("trackside.export.unmatched_ap_mac_raw", "ap_mac_raw"),
    ("trackside.export.unmatched_ap_mac_normalized", "ap_mac_normalized"),
    ("trackside.export.unmatched_planning_record_id", "planning_record_id"),
    ("trackside.export.unmatched_planning_station", "planning_station_name"),
    ("trackside.export.unmatched_plan_station_id", "plan_station_id"),
    ("trackside.export.unmatched_planning_match_method", "planning_match_method"),
    ("trackside.export.unmatched_lldp_exists", "lldp_exists"),
    ("trackside.export.unmatched_lldp_local_interface", "lldp_local_interface"),
    ("trackside.export.unmatched_lldp_remote_device", "lldp_remote_device_name"),
    ("trackside.export.unmatched_lldp_system_name", "lldp_system_name"),
    ("trackside.export.unmatched_lldp_management_ip", "lldp_management_ip"),
    ("trackside.export.unmatched_lldp_chassis_id", "lldp_chassis_id"),
    ("trackside.export.unmatched_switch_candidate_count", "switch_candidate_count"),
    ("trackside.export.unmatched_switch_device_id", "matched_switch_device_id"),
    ("trackside.export.unmatched_switch_match_method", "switch_match_method"),
    ("trackside.export.unmatched_failure_stage", "failure_stage"),
    ("trackside.export.unmatched_source_revisions", "source_revisions"),
    ("trackside.export.unmatched_snapshot_revision", "snapshot_revision"),
    ("trackside.export.unmatched_snapshot_created_at", "snapshot_created_at"),
    ("trackside.export.unmatched_reason", "reason"),
    ("trackside.export.unmatched_suggestion", "suggested_action"),
)

AP_OPTICAL_TREATMENT_RECORD_COLUMNS = (
    ("ac.station", "site"),
    ("ac.ap_name", "ap_name"),
    ("ac.ap_mac", "ap_mac"),
    ("ap.serial_number", "serial_number"),
    ("ac.ap_id", "ap_id"),
    ("trackside.section_name", "section_name"),
    ("trackside.direction", "direction"),
    ("trackside.export.side", "side"),
    ("ac.indoor_switch", "device_name"),
    ("details.interface_name", "interface_name"),
    ("trackside.export.issue_type", "issue_type"),
    ("trackside.export.first_found_at", "first_found_at"),
    ("trackside.export.first_rx_power", "first_rx_power"),
    ("trackside.export.fixed_rx_power", "fixed_rx_power"),
    ("trackside.export.current_rx_power", "current_rx_power"),
    ("trackside.export.current_status", "current_status"),
    ("trackside.export.treatment_status", "treatment_status"),
    ("trackside.export.remark", "remark"),
    ("trackside.export.completed_at", "completed_at"),
)

OPTICAL_TREATMENT_ISSUE_STATUSES = {
    "abnormal",
    "alarm",
    "link_abnormal",
    "link_down",
    "no_light",
    "notice",
    "warning",
}
OPTICAL_TREATMENT_IGNORED_STATUSES = {"normal", "unknown", "not_collected", "skipped", "failed", "timeout", "offline", "no_module", ""}
AP_SIDE_LABEL = "AP\u4fa7"
SWITCH_SIDE_LABEL = "\u4ea4\u6362\u673a\u4fa7"
TREATMENT_OPEN_LABEL = "\u672a\u5904\u7406"
TREATMENT_CLOSED_LABEL = "\u5df2\u5904\u7406"
ISSUE_TYPE_NOTICE_LABEL = "\u5149\u8870\u9884\u8b66"
ISSUE_TYPE_ALARM_LABEL = "\u5149\u8870\u544a\u8b66"
ISSUE_TYPE_LINK_ABNORMAL_LABEL = "\u94fe\u8def\u5f02\u5e38"
ISSUE_TYPE_NO_LIGHT_LABEL = "\u65e0\u5149"
ISSUE_TYPE_OPTICAL_ABNORMAL_LABEL = "\u5149\u8870\u5f02\u5e38"

TRACKSIDE_AP_BUSINESS_HEADER_TOOLTIPS = {
    "site": "trackside.tooltip.station",
    "link_status": "trackside.tooltip.link",
    "port_type": "trackside.tooltip.port_type",
    "switch_rx_power": "trackside.tooltip.switch_rx_power",
    "ap_optical_status": "trackside.tooltip.ap_optical_status",
}

TRACKSIDE_AP_DEVICE_COLUMNS = (
    ("details.interface_name", "interface_name"),
    ("details.link", "link_status"),
    ("details.port_type", "port_type"),
    ("details.port_description", "description"),
    ("details.pvid", "pvid"),
    ("details.vlan", "vlan"),
    ("ac.indoor_switch_rx_power", "switch_rx_power"),
    ("trackside.switch_optical_status", "switch_optical_status"),
    ("ac.ap_mac", "ap_mac"),
    ("ac.ap_name", "ap_name"),
    ("ac.ap_side_rx_power", "ap_rx_power"),
    ("trackside.ap_device_optical_status", "ap_device_optical_status"),
    ("trackside.ap_optical_status", "ap_optical_status"),
    ("trackside.ap_business_threshold", "ap_business_threshold_dbm"),
    ("trackside.ap_business_reason", "ap_business_reason"),
    ("field.updated_at", "updated_at"),
)

TRACKSIDE_OPTICAL_COLOR_RGB = {
    "normal": "DCFCE7",
    "notice": "FEF9C3",
    "warning": "FEF9C3",
    "alarm": "FEE2E2",
    "abnormal": "FEE2E2",
    "link_abnormal": "FFE4E6",
    "link_down": "FFE4E6",
    "no_light": "E5E7EB",
    "no_module": "F3F4F6",
    "skipped": "F3F4F6",
    "offline": "E5E7EB",
    "not_applicable": "F3F4F6",
}
TRACKSIDE_EXPORT_HEADER_FILL = "DBEAFE"
TRACKSIDE_EXPORT_NORMAL_FILL = TRACKSIDE_OPTICAL_COLOR_RGB["normal"]
TRACKSIDE_EXPORT_WARNING_FILL = TRACKSIDE_OPTICAL_COLOR_RGB["warning"]
TRACKSIDE_EXPORT_ALARM_FILL = TRACKSIDE_OPTICAL_COLOR_RGB["alarm"]
_TRACKSIDE_LONG_TEXT_HEADER_TOKENS = (
    "原因",
    "备注",
    "说明",
    "建议",
    "详情",
    "未插光模块端口",
    "reason",
    "remark",
    "description",
    "note",
)
CURRENT_OPTICAL_ABNORMAL_SHEET_TITLE = "当前异常光衰"
CURRENT_OPTICAL_ABNORMAL_EMPTY_TEXT = "当前无异常光衰（已排除无 AP 绑定、无光模块和非告警光功率）"
TRACKSIDE_EXPORT_ROW_HEIGHT = 24.0
TRACKSIDE_OVERVIEW_SEPARATOR_ROW_HEIGHT = 16.0


@dataclass(frozen=True)
class TracksideApBusinessSheetDefinition:
    stable_key: str
    sheet_name: str
    order: int
    sync_mode: str = "FULL_REPLACE"
    tab_color: str = ""
    freeze_mode: str = "FIRST_ROW_ONLY"


TRACKSIDE_AP_BUSINESS_SHEET_DEFINITIONS = (
    TracksideApBusinessSheetDefinition(
        "ap_online_history_overview",
        "AP上线情况概览",
        10,
        sync_mode="PREPEND_SNAPSHOT",
        tab_color="#C6EFCE",
        freeze_mode="NONE",
    ),
    TracksideApBusinessSheetDefinition(
        "trackside_ap_business",
        "轨旁AP业务",
        20,
        tab_color="#C6EFCE",
    ),
    TracksideApBusinessSheetDefinition(
        "current_optical_abnormal",
        CURRENT_OPTICAL_ABNORMAL_SHEET_TITLE,
        30,
        tab_color="#FFEB9C",
    ),
    TracksideApBusinessSheetDefinition(
        "ap_optical_treatment_records",
        "AP光衰处理记录",
        40,
        tab_color="#DDEBF7",
    ),
    TracksideApBusinessSheetDefinition(
        "ap_offline_status",
        "AP离线情况",
        50,
        tab_color="#D9D9D9",
    ),
    TracksideApBusinessSheetDefinition(
        "ap_offline_ledger",
        "AP离线台账",
        60,
        tab_color="#D9D9D9",
    ),
    TracksideApBusinessSheetDefinition(
        "newly_online_ap_overview",
        "新增上线AP概览",
        70,
    ),
    TracksideApBusinessSheetDefinition(
        "unmatched_online_ap",
        "待关联在线AP",
        80,
    ),
    TracksideApBusinessSheetDefinition(
        "switch_optical_module_summary",
        "交换机光模块统计",
        90,
    ),
)

TRACKSIDE_COLUMN_LAYOUT_LIMITS = {
    "compact": (8.0, 16.0),
    "normal": (8.0, 24.0),
    "identifier": (10.0, 28.0),
    "datetime": (12.0, 24.0),
    "long_text": (16.0, 48.0),
}
_TRACKSIDE_LONG_TEXT_FIELDS = {
    "ap_business_reason",
    "description",
    "detail",
    "reason",
    "remark",
    "source_revisions",
    "suggestion",
    "suggested_action",
    "missing_ports",
}
_TRACKSIDE_IDENTIFIER_FIELDS = {
    "ap_ip",
    "ap_mac",
    "apid",
    "identity_entity_id",
    "matched_switch_device_id",
    "plan_station_id",
    "planning_record_id",
    "serial_number",
}
_TRACKSIDE_COMPACT_FIELDS = {
    "ap_rx_power",
    "fixed_rx_power",
    "first_rx_power",
    "current_rx_power",
    "lldp_candidate_count",
    "module_count",
    "offline_aps",
    "offline_locatable",
    "offline_rate",
    "offline_unlocatable",
    "offline_with_lldp",
    "offline_without_lldp",
    "online",
    "online_aps",
    "online_rate",
    "pvid",
    "switch_rx_power",
    "total",
    "total_aps",
    "vlan",
}
_TRACKSIDE_SHEET_COLUMNS = {
    "ap_online_history_overview": AP_ONLINE_OVERVIEW_COLUMNS,
    "trackside_ap_business": TRACKSIDE_AP_BUSINESS_EXPORT_COLUMNS,
    "current_optical_abnormal": CURRENT_OPTICAL_ABNORMAL_COLUMNS,
    "ap_optical_treatment_records": AP_OPTICAL_TREATMENT_RECORD_COLUMNS,
    "ap_offline_status": OFFLINE_AP_STATS_COLUMNS,
    "ap_offline_ledger": OFFLINE_AP_LEDGER_COLUMNS,
    "newly_online_ap_overview": NEW_ONLINE_AP_OVERVIEW_COLUMNS,
    "unmatched_online_ap": TRACKSIDE_AP_UNMATCHED_ONLINE_COLUMNS,
    "switch_optical_module_summary": (
        ("switch", "device_name"),
        ("module_count", "module_count"),
        ("missing_port_count", "missing_port_count"),
        ("missing_ports", "missing_ports"),
    ),
}
_TRACKSIDE_AP_BUSINESS_SHEETS_BY_NAME = {
    definition.sheet_name: definition
    for definition in TRACKSIDE_AP_BUSINESS_SHEET_DEFINITIONS
}


def trackside_ap_business_sheet_definition(
    sheet_name: str,
) -> TracksideApBusinessSheetDefinition | None:
    return _TRACKSIDE_AP_BUSINESS_SHEETS_BY_NAME.get(str(sheet_name or ""))


def trackside_ap_business_column_layout_types(
    sheet_name: str,
    column_count: int,
) -> tuple[str, ...]:
    definition = trackside_ap_business_sheet_definition(sheet_name)
    columns = _TRACKSIDE_SHEET_COLUMNS.get(
        definition.stable_key if definition is not None else "",
        (),
    )
    layouts = [
        _trackside_column_layout_type(field)
        for _key, field in columns[: max(0, column_count)]
    ]
    layouts.extend(["normal"] * max(0, column_count - len(layouts)))
    return tuple(layouts)


def _trackside_column_layout_type(field: str) -> str:
    if field in _TRACKSIDE_LONG_TEXT_FIELDS:
        return "long_text"
    if field in _TRACKSIDE_IDENTIFIER_FIELDS:
        return "identifier"
    if field.endswith("_at") or field.endswith("_time"):
        return "datetime"
    if field in _TRACKSIDE_COMPACT_FIELDS:
        return "compact"
    return "normal"


@dataclass(frozen=True)
class ApOnlineHistoryBlockDTO:
    snapshot_date: str
    updated_at: str
    headers: tuple[str, ...]
    rows: tuple[tuple[object, ...], ...]

    def cells(self) -> list[list[object | None]]:
        width = len(self.headers)
        return [
            [f"日期：{self.snapshot_date}", *([None] * max(width - 1, 0))],
            [f"更新时间：{self.updated_at}", *([None] * max(width - 1, 0))],
            list(self.headers),
            *[list(row) for row in self.rows],
            [None] * width,
        ]


class TracksideApExportCancelled(RuntimeError):
    """Raised when a trackside AP export is cancelled."""


def normalize_trackside_vlan_display(value: object) -> str:
    """Remove the Native/PVID fragment duplicated by the dedicated PVID column."""

    text = str(value or "").strip()
    if not text:
        return "—"
    parts = []
    for raw_part in re.split(r"[;；]+", text):
        part = re.sub(r"\s+", " ", raw_part).strip()
        if not part or _TRACKSIDE_NATIVE_PVID_SEGMENT_RE.fullmatch(part):
            continue
        parts.append(part)
    return "; ".join(parts) or "—"


def normalize_trackside_ap_business_row(
    row: dict[str, object | None],
    *,
    business_projection: bool = True,
) -> dict[str, object | None]:
    """Return one row with normalized device facts and optional business semantics."""

    normalized = dict(row)
    normalized["vlan"] = normalize_trackside_vlan_display(normalized.get("vlan"))
    if normalized.get("switch_interface_data_status") in {"stale", "missing"}:
        normalized["link_status"] = "-"
        normalized["protocol_status"] = None
    if normalized.get("switch_optical_data_status") in {"stale", "missing"}:
        for field in (
            "switch_rx_power",
            "switch_tx_power",
            "switch_rx_low_alarm",
            "switch_rx_high_alarm",
            "switch_tx_low_alarm",
            "switch_tx_high_alarm",
        ):
            normalized[field] = None
        normalized["switch_optical_status"] = "not_collected"
    switch_device_status = _normalized_optical_status(
        normalized.get("switch_optical_status")
    )
    normalized["switch_device_optical_status"] = (
        _normalized_optical_status(normalized.get("switch_device_optical_status"))
        or switch_device_status
        or "unknown"
    )
    device_status = _normalized_optical_status(
        normalized.get("ap_device_optical_status")
        or normalized.get("ap_optical_status")
    )
    optical_applicable = is_ap_optical_applicable(
        normalized.get("model") or normalized.get("ap_model")
    )
    normalized["ap_optical_applicable"] = optical_applicable
    if not optical_applicable:
        normalized["ap_rx_power"] = None
        normalized["ap_tx_power"] = None
        normalized["ap_device_optical_status"] = OPTICAL_NOT_APPLICABLE_STATUS
        normalized["switch_device_optical_status"] = OPTICAL_NOT_APPLICABLE_STATUS
        normalized["switch_optical_status"] = OPTICAL_NOT_APPLICABLE_STATUS
        normalized["ap_business_optical_status"] = OPTICAL_NOT_APPLICABLE_STATUS
        normalized["ap_business_threshold_dbm"] = None
        normalized["ap_business_reason"] = OPTICAL_NOT_APPLICABLE_REASON
        normalized["ap_optical_status"] = OPTICAL_NOT_APPLICABLE_STATUS
        normalized["optical_severity"] = OPTICAL_NOT_APPLICABLE_STATUS
        return normalized
    if not business_projection:
        normalized["switch_optical_status"] = switch_device_status
        normalized["ap_optical_status"] = device_status
        normalized["optical_severity"] = _trackside_row_status_with_ap_status(
            normalized,
            device_status if has_ap_side_optical_data(normalized) else "",
        )
        return normalized

    ap_side_has_data = has_ap_side_optical_data(normalized)
    evaluation = evaluate_dual_rx_business_detail(
        normalized.get("ap_rx_power") if ap_side_has_data else None,
        normalized.get("switch_rx_power"),
        ap_reported_status=device_status if ap_side_has_data else "",
        switch_reported_status=switch_device_status,
        ap_data_freshness=(
            normalized.get("ap_optical_data_freshness")
            or normalized.get("data_freshness")
        ),
        switch_data_freshness=normalized.get("switch_optical_data_status"),
    )
    normalized["ap_device_optical_status"] = device_status or "unknown"
    normalized["ap_business_optical_status"] = evaluation.status
    normalized["ap_business_threshold_dbm"] = evaluation.threshold_dbm
    normalized["ap_business_reason"] = evaluation.reason
    normalized["ap_optical_status"] = evaluation.ap_status
    normalized["switch_optical_status"] = evaluation.switch_status
    normalized["optical_severity"] = trackside_row_status(normalized)
    return normalized

AP_SIDE_DISPLAY_FIELDS = {"ap_rx_power", "ap_tx_power"}
AP_SIDE_MISSING_DISPLAY = "-"
MATCH_SOURCE_LABELS = {
    "description": "\u63cf\u8ff0\u5339\u914d",
    "pvid": "PVID\u5339\u914d",
    "description+pvid": "\u63cf\u8ff0+PVID",
    "none": "-",
}


def description_contains_ap(description: object) -> bool:
    return "ap" in str(description or "").casefold()


def parse_vlan_set(value: object) -> set[int]:
    if value in (None, ""):
        return set()
    tokens = [item for item in re.split(r"[,，;；\s]+", str(value).strip()) if item]
    vlans: set[int] = set()
    for token in tokens:
        if "-" in token:
            parts = token.split("-", 1)
            if len(parts) != 2 or not parts[0].isdigit() or not parts[1].isdigit():
                raise ValueError(f"invalid VLAN range: {token}")
            start = int(parts[0])
            end = int(parts[1])
            if start > end:
                raise ValueError(f"invalid VLAN range: {token}")
            for vlan in range(start, end + 1):
                _validate_vlan(vlan)
                vlans.add(vlan)
            continue
        if not token.isdigit():
            raise ValueError(f"invalid VLAN: {token}")
        vlan = int(token)
        _validate_vlan(vlan)
        vlans.add(vlan)
    return vlans


def normalize_vlan_text(value: object) -> str:
    return ",".join(str(vlan) for vlan in sorted(parse_vlan_set(value)))


def pvid_matches_trackside_plan(device_station: object, pvid: object, active_plan: dict | None) -> bool:
    if not active_plan:
        return False
    try:
        vlan = int(str(pvid or "").strip())
    except ValueError:
        return False
    if vlan <= 0:
        return False
    # 这里只用于发现可能的轨旁 AP 端口，不能用交换机所属站点把
    # 候选 VLAN 收窄；最终核验必须在 AP 身份解析后执行。
    return (
        vlan in set(active_plan.get("all_vlans") or set())
        if isinstance(active_plan, dict)
        else False
    )


def effective_pvid_plan(
    *,
    ap_mac: object,
    ap_name: object,
    station_id: object = "",
    station_name: object = "",
    pvid: object,
    active_plan: dict | None,
) -> dict[str, object]:
    del ap_name, station_name
    if not isinstance(active_plan, dict):
        return {"pvid_plan_status": "unresolved"}
    mac = normalize_mac_key(ap_mac) or ""
    network = (active_plan.get("ap_networks_by_mac") or {}).get(mac) if mac else None
    if network is None:
        station_key = str(station_id or "").strip()
        station_vlans = (
            (active_plan.get("station_vlans_by_id") or {}).get(station_key)
            if station_key
            else None
        )
        vlan_values = sorted(
            int(value) for value in (station_vlans or set())
        )
        if len(vlan_values) == 1:
            network = {
                "management_vlan": vlan_values[0],
                "planning_station_id": station_key,
            }
    if not isinstance(network, dict):
        return {"pvid_plan_status": "unresolved"}
    try:
        actual = int(str(pvid or "").strip())
        planned = int(network.get("management_vlan") or 0)
    except (TypeError, ValueError):
        return {**network, "pvid_plan_status": "unresolved"}
    return {
        **network,
        "planning_station_id": str(
            network.get("planning_station_id")
            or network.get("station_id")
            or station_id
            or ""
        ).strip(),
        "planned_management_vlan": planned,
        "pvid_plan_status": "matched" if actual == planned else "mismatched",
    }


def _station_consistency_projection(
    switch_station_id: object,
    ap_station_id: object,
    planning_station_id: object,
) -> dict[str, str]:
    switch_id = str(switch_station_id or "").strip()
    ap_id = str(ap_station_id or "").strip()
    planning_id = str(planning_station_id or "").strip()
    authoritative = [value for value in (switch_id, ap_id) if value]
    if len(set(authoritative)) > 1:
        return {
            "switch_station_id": switch_id,
            "ap_station_id": ap_id,
            "planning_station_id": planning_id,
            "effective_station_id": "",
            "station_consistency_status": "conflict",
            "station_consistency_reason": "SWITCH_AP_STATION_CONFLICT",
        }
    effective_id = ap_id or switch_id
    if planning_id and effective_id and planning_id != effective_id:
        return {
            "switch_station_id": switch_id,
            "ap_station_id": ap_id,
            "planning_station_id": planning_id,
            "effective_station_id": "",
            "station_consistency_status": "conflict",
            "station_consistency_reason": "PLANNING_STATION_CONFLICT",
        }
    if effective_id:
        return {
            "switch_station_id": switch_id,
            "ap_station_id": ap_id,
            "planning_station_id": planning_id,
            "effective_station_id": effective_id,
            "station_consistency_status": "consistent" if switch_id and ap_id else "partial",
            "station_consistency_reason": "" if switch_id and ap_id else "ONE_SIDE_STATION_ID_MISSING",
        }
    return {
        "switch_station_id": "",
        "ap_station_id": "",
        "planning_station_id": planning_id,
        "effective_station_id": planning_id,
        "station_consistency_status": "planning_only" if planning_id else "unresolved",
        "station_consistency_reason": "ONLY_PLANNING_STATION_ID" if planning_id else "STATION_ID_MISSING",
    }


def is_switch_device_type(device_type: object) -> bool:
    return str(device_type or "").strip().casefold() in {"sw", "switch", "交换机"}


def filter_station_switch_devices(
    devices: list[Device],
    database,
    site_name: str,
    *,
    project_phase: str = "",
) -> list[Device]:
    groups = {group.id: group.name for group in DeviceGroupRepository(database, site_name).list()}
    return [
        device
        for device in devices
        if groups.get(device.group_id or -1, "") == "车站"
        and is_switch_device_type(device.device_type)
        and is_trackside_device_eligible(device, project_phase=project_phase)
    ]


def is_trackside_device_eligible(
    device: Device,
    *,
    project_phase: str = "",
) -> bool:
    """Trackside automatic scopes use current debugging targets in the configured phase."""

    if not is_device_eligible_for_automatic_collection(device):
        return False
    expected_phase = str(project_phase or "").strip().casefold()
    if not expected_phase:
        return True
    return str(device.project_phase or "").strip().casefold() == expected_phase


def is_trackside_layer2_interface(interface: dict[str, object | None]) -> bool:
    name = str(interface.get("interface_name") or "").strip()
    name_key = name.casefold()
    if not name_key:
        return False
    if name_key.startswith(("vlan-interface", "loopback", "inloopback", "null")):
        return False
    if str(interface.get("interface_type") or "").strip().casefold() in {"三层", "l3", "layer3", "route", "routed"}:
        return False
    if str(interface.get("port_status") or "").strip().casefold() == "route":
        return False
    if str(interface.get("ip_address") or "").strip():
        return False
    layer2_markers = (
        interface.get("pvid"),
        interface.get("vlan"),
        interface.get("port_status"),
        interface.get("link_type"),
        interface.get("port_link_type"),
    )
    if any(str(value or "").strip() for value in layer2_markers):
        return True
    return bool(re.match(r"^(?:GigabitEthernet|Ten-GigabitEthernet|XGE|Bridge-Aggregation)", name, re.IGNORECASE))


def is_trackside_ap_interface(device: Device, interface: dict[str, object | None], active_plan: dict | None = None) -> tuple[bool, str]:
    if not is_trackside_layer2_interface(interface):
        return False, "none"
    by_description = description_contains_ap(interface.get("description"))
    by_pvid = pvid_matches_trackside_plan(getattr(device, "station", ""), interface.get("pvid"), active_plan)
    if by_description and by_pvid:
        return True, "description+pvid"
    if by_description:
        return True, "description"
    if by_pvid:
        return True, "pvid"
    return False, "none"


def _validate_vlan(vlan: int) -> None:
    if vlan < 1 or vlan > 4094:
        raise ValueError(f"VLAN out of range: {vlan}")


def build_device_optical_status_lookup(
    devices: list[Device],
    optical_by_device: dict[str, list[dict[str, object | None]]],
) -> dict:
    """Backward-compatible alias — delegates to ``switch_source.build_switch_data_lookup``."""
    return build_switch_data_lookup(devices, optical_by_device)


def build_trackside_ap_business_rows(
    devices: list[Device],
    interfaces_by_device: dict[str, list[dict[str, object | None]]],
    optical_by_device: dict[str, list[dict[str, object | None]]],
    fit_ap_optical_rows: list[dict[str, object | None]],
    lldp_by_device: dict[str, list[dict[str, object | None]]] | None = None,
    fit_ap_resource_rows: list[dict[str, object | None]] | None = None,
    device_optical_status_lookup: dict[tuple[str, str], str] | None = None,
    trackside_ap_plan: dict | None = None,
    offline_ap_ledger_rows: list[dict[str, object | None]] | None = None,
    historical_lldp_rows: list[dict[str, object | None]] | None = None,
    station_names: Mapping[str, str] | None = None,
    latest_switch_collect_runs: Mapping[str, str] | None = None,
    runtime_snapshot: TracksideApRuntimeSnapshot | None = None,
    business_projection: bool = True,
) -> list[dict[str, object | None]]:
    optical_indexes = {device_uuid: _latest_rows_by_normalized_interface(rows, "interface_name") for device_uuid, rows in optical_by_device.items()}
    lldp_indexes = {
        device_uuid: _latest_rows_by_normalized_interface(
            deduplicate_lldp_snapshot_rows(
                select_latest_lldp_snapshot_rows(rows)
            ),
            "local_interface",
        )
        for device_uuid, rows in (lldp_by_device or {}).items()
    }
    fit_ap_optical_rows = merge_fit_ap_rows_by_identity(fit_ap_optical_rows)
    fit_ap_resource_rows = coalesce_fit_ap_resource_rows(fit_ap_resource_rows or [])
    fit_ap_resource_rows = merge_fit_ap_rows_by_identity(fit_ap_resource_rows)
    fit_ap_optical_by_mac: dict[str, dict[str, object | None]] = {}
    fit_ap_optical_by_identity: dict[tuple[str, str], dict[str, object | None]] = {}
    fit_ap_optical_by_switch_interface: dict[tuple[str, str], dict[str, object | None]] = {}
    fit_ap_optical_by_name_mac: dict[str, dict[str, object | None]] = {}
    fit_ap_resource_by_mac: dict[str, dict[str, object | None]] = {}
    fit_ap_resource_by_identity: dict[tuple[str, str], dict[str, object | None]] = {}
    fit_ap_resources_by_identity: dict[tuple[str, str], list[dict[str, object | None]]] = {}
    fit_ap_resources_by_ip: dict[str, list[dict[str, object | None]]] = {}
    fit_ap_resources_by_mac: dict[str, list[dict[str, object | None]]] = {}
    fit_ap_resources_by_name: dict[str, list[dict[str, object | None]]] = {}
    for row in fit_ap_optical_rows:
        mac = normalize_mac_key(row.get("ap_mac"))
        if mac:
            fit_ap_optical_by_mac[mac] = row
        switch_key = (
            _normalize_name(row.get("neighbor_device_name")),
            normalize_interface_name(row.get("neighbor_interface")).casefold(),
        )
        if all(switch_key):
            fit_ap_optical_by_switch_interface[switch_key] = row
        name_as_mac = normalize_mac_key(row.get("ap_name"))
        if name_as_mac:
            fit_ap_optical_by_name_mac[name_as_mac] = row
        identity = ap_identity_key(row)
        if identity:
            fit_ap_optical_by_identity[identity] = row
    for row in fit_ap_resource_rows or []:
        mac = normalize_mac_key(row.get("ap_mac"))
        if mac:
            current = fit_ap_resource_by_mac.get(mac)
            if current is None or _fit_ap_prefer_score(row) >= _fit_ap_prefer_score(current):
                fit_ap_resource_by_mac[mac] = row
            fit_ap_resources_by_mac.setdefault(mac, []).append(row)
        ap_ip = _normalize_ip(row.get("ap_ip") or row.get("management_ip"))
        if ap_ip:
            fit_ap_resources_by_ip.setdefault(ap_ip, []).append(row)
        ap_name = _normalize_name(row.get("ap_name"))
        if ap_name:
            fit_ap_resources_by_name.setdefault(ap_name, []).append(row)
        identity = ap_identity_key(row)
        if identity:
            current = fit_ap_resource_by_identity.get(identity)
            if current is None or _fit_ap_prefer_score(row) >= _fit_ap_prefer_score(current):
                fit_ap_resource_by_identity[identity] = row
            fit_ap_resources_by_identity.setdefault(identity, []).append(row)
    historical_lldp_index = _build_historical_lldp_index(historical_lldp_rows or [])

    result: list[dict[str, object | None]] = []
    for device in devices:
        device_uuid = str(device.device_uuid or "")
        latest_switch_collect_run = str(
            (latest_switch_collect_runs or {}).get(device_uuid) or ""
        )
        lldp_snapshot_present = device_uuid in (lldp_by_device or {})
        try:
            adapter = resolve_trackside_switch_adapter(device)
            adapter_description = adapter.describe_capabilities()
            capability_statuses = {
                item.key: item.status.value
                for item in adapter_description.capabilities
            }
            bidirectional_attenuation_enabled = (
                adapter.capabilities.bidirectional_attenuation
            )
        except ValueError:
            capability_statuses = {}
            bidirectional_attenuation_enabled = True
        device_names = {_normalize_name(device.name), _normalize_name(device.system_name)}
        device_names.discard("")
        optical_index = optical_indexes.get(device_uuid, {})
        lldp_index = lldp_indexes.get(device_uuid, {})
        for interface in interfaces_by_device.get(device_uuid, []):
            matched, match_source = is_trackside_ap_interface(device, interface, trackside_ap_plan)
            if not matched:
                continue
            interface_name = str(interface.get("interface_name") or "")
            normalized_interface = normalize_interface_name(interface_name).casefold()
            interface_data_status = _snapshot_data_status(
                interface,
                latest_switch_collect_run,
            )
            current_interface = (
                interface if interface_data_status == "current" else {}
            )
            link_state = normalize_link_state(
                current_interface.get("link_status") or current_interface.get("link")
            )
            stored_optical = optical_index.get(normalized_interface, {})
            optical_data_status = _snapshot_data_status(
                stored_optical,
                latest_switch_collect_run,
            )
            optical = stored_optical if optical_data_status == "current" else {}
            lldp = lldp_index.get(normalized_interface, {})
            historical_lldp = _find_historical_lldp_row(historical_lldp_index, device_names, interface_name)
            neighbor_mac = normalize_mac_key(lldp.get("neighbor_mac"))
            (
                resource_from_current_lldp,
                ap_match_source,
                ap_match_confidence,
                lldp_match_status,
            ) = _match_fit_ap_resource_from_lldp(
                lldp,
                fit_ap_resources_by_ip,
                fit_ap_resources_by_mac,
                fit_ap_resources_by_name,
            )
            historical_lldp_used = False
            if not lldp and not neighbor_mac and historical_lldp:
                neighbor_mac = normalize_mac_key(
                    historical_lldp.get("ap_mac")
                    or historical_lldp.get("neighbor_mac")
                )
                historical_lldp_used = bool(
                    neighbor_mac
                    or historical_lldp.get("ap_name")
                    or historical_lldp.get("ap_uuid")
                )
            resource_from_neighbor = (
                None
                if lldp_match_status == "AMBIGUOUS"
                else resource_from_current_lldp or fit_ap_resource_by_mac.get(neighbor_mac)
            )
            if not ap_match_source and resource_from_neighbor and lldp:
                ap_match_source = "LLDP_MAC"
                ap_match_confidence = 92
                lldp_match_status = "MATCHED"
            identity_from_neighbor = ap_identity_key(resource_from_neighbor or historical_lldp or {})
            fit_ap_from_identity = fit_ap_optical_by_identity.get(identity_from_neighbor) if identity_from_neighbor else None
            resource_from_identity = fit_ap_resource_by_identity.get(identity_from_neighbor) if identity_from_neighbor else None
            if lldp_match_status == "AMBIGUOUS":
                fit_ap = {}
                historical_lldp = {}
                historical_lldp_used = False
                neighbor_mac = ""
                fit_ap_from_identity = None
                resource_from_identity = None
            elif resource_from_current_lldp:
                fit_ap = (
                    _merge_resource_with_optical(resource_from_neighbor, fit_ap_optical_by_mac)
                    or resource_from_identity
                    or fit_ap_from_identity
                    or resource_from_neighbor
                    or {}
                )
            else:
                fit_ap = (
                    _merge_resource_with_optical(resource_from_neighbor, fit_ap_optical_by_mac)
                    or resource_from_identity
                    or resource_from_neighbor
                    or fit_ap_from_identity
                    or fit_ap_optical_by_mac.get(neighbor_mac)
                    or fit_ap_optical_by_name_mac.get(neighbor_mac)
                    or {}
                )
                if not fit_ap and not lldp:
                    interface_resource = next(
                        (
                            fit_ap_optical_by_switch_interface.get((name, normalized_interface))
                            for name in device_names
                            if fit_ap_optical_by_switch_interface.get((name, normalized_interface))
                        ),
                        None,
                    )
                    if interface_resource:
                        fit_ap = (
                            _merge_resource_with_optical(
                                fit_ap_resource_by_mac.get(
                                    normalize_mac_key(interface_resource.get("ap_mac"))
                                ),
                                fit_ap_optical_by_mac,
                            )
                            or interface_resource
                        )
                        neighbor_mac = (
                            normalize_mac_key(interface_resource.get("ap_mac")) or ""
                        )
                        ap_match_source = "FIT_AP_INTERFACE_MAC"
                        ap_match_confidence = 80
                        lldp_match_status = "MATCHED"
            switch_collection_status = _switch_collection_status(
                device,
                current_interface,
                optical,
            )
            if device.vendor_key == "zte":
                optical = normalize_zte_optical_record(optical)
                switch_result = compute_zte_optical_severity(optical)
            else:
                switch_result = compute_optical_severity(
                    {
                        "module_present": bool(_has_optical_module_data(optical)),
                        "no_module": _explicit_no_module(optical),
                        "switch_rx_power": optical.get("rx_power"),
                        "switch_port_status": optical.get("port_status"),
                        "alarm_low": optical.get("rx_low_alarm"),
                        "alarm_high": optical.get("rx_high_alarm"),
                        "warning_low": optical.get("rx_low_warning"),
                        "device_type": "switch",
                    }
                )
            collected_module_status = str(optical.get("status") or "").strip().casefold()
            switch_status = (
                collected_module_status
                if collected_module_status in _PRESERVED_SWITCH_MODULE_STATUSES
                else switch_result.severity
            )
            switch_offline = _is_switch_collection_offline(switch_collection_status)
            ap_candidate = {
                "ap_mac": format_mac(fit_ap.get("ap_mac") or neighbor_mac),
                "ap_name": fit_ap.get("ap_name") or (historical_lldp or {}).get("ap_name"),
                "ap_rx_power": fit_ap.get("rx_power"),
                "ap_tx_power": fit_ap.get("tx_power"),
            }
            ap_model = fit_ap.get("model")
            ap_optical_applicable = is_ap_optical_applicable(ap_model)
            if not ap_optical_applicable:
                ap_candidate["ap_rx_power"] = None
                ap_candidate["ap_tx_power"] = None
            ac_idle = _is_ac_idle(fit_ap)
            ac_offline = _is_ac_offline(fit_ap)
            ap_identity_known = any(ap_candidate.get(field) for field in ("ap_mac", "ap_name"))
            ap_side_has_data = ap_optical_applicable and (
                _has_ap_side_optical_data(fit_ap, ap_candidate)
                or ac_offline
                or (switch_offline and ap_identity_known)
            )
            ap_status = ""
            offline_reason = ""
            status_reason = ""
            data_source = "current"
            if historical_lldp_used and not lldp:
                status_reason = "沿用历史LLDP映射，本轮未采到当前LLDP"
                data_source = "historical_lldp"
            if switch_offline:
                switch_status = "offline"
                ap_status = "offline"
                offline_reason = "switch_offline"
                status_reason = "室内交换机离线，轨旁AP跟随离线"
                data_source = "mixed" if ap_identity_known else "stale"
            elif ac_offline:
                ap_status = "offline"
                offline_reason = "ac_idle" if ac_idle else "ac_offline"
                status_reason = _ac_offline_status_reason(fit_ap)
            elif ap_side_has_data:
                ap_result = compute_optical_severity(
                    {
                        "module_present": bool(_has_optical_module_data(fit_ap)) or _explicit_no_module(fit_ap),
                        "no_module": _explicit_no_module(fit_ap),
                        "ap_rx_power": fit_ap.get("rx_power"),
                        "ap_port_status": fit_ap.get("ap_port_status"),
                        "alarm_low": fit_ap.get("rx_low_alarm"),
                        "alarm_high": fit_ap.get("rx_high_alarm"),
                        "warning_low": fit_ap.get("rx_low_warning"),
                        "device_type": "ap",
                    }
                )
                ap_status = ap_result.severity
            local_sample_time = (
                optical.get("updated_at")
                or optical.get("collected_at")
                or current_interface.get("updated_at")
                or current_interface.get("collected_at")
            )
            remote_sample_time = fit_ap.get("updated_at") or fit_ap.get("collected_at")
            attenuation = _build_bidirectional_attenuation(
                local_rx_power=optical.get("rx_power"),
                local_tx_power=optical.get("tx_power"),
                remote_rx_power=ap_candidate["ap_rx_power"] if ap_side_has_data and not switch_offline else None,
                remote_tx_power=ap_candidate["ap_tx_power"] if ap_side_has_data and not switch_offline else None,
                local_status=switch_status,
                remote_status=ap_status,
                association_reliable=ap_match_source in {
                    "LLDP_MAC",
                    "MANUAL",
                    "IMPORTED",
                },
                remote_identity_known=ap_identity_known,
                local_sample_time=local_sample_time,
                remote_sample_time=remote_sample_time,
                calculation_enabled=bidirectional_attenuation_enabled,
            )
            if not lldp_match_status:
                if lldp:
                    lldp_match_status = "UNRESOLVED"
                elif (
                    capability_statuses.get("lldp")
                    == CommandCapabilityState.SAMPLE_REQUIRED.value
                    or (
                        device.vendor_key
                        == "zte"
                        and not lldp_snapshot_present
                    )
                ):
                    lldp_match_status = "SAMPLE_REQUIRED"
                else:
                    lldp_match_status = "NO_NEIGHBOR"
            pvid_projection = effective_pvid_plan(
                ap_mac=ap_candidate["ap_mac"],
                ap_name=ap_candidate["ap_name"],
                station_id=fit_ap.get("station_id"),
                station_name=normalize_station_value(fit_ap),
                pvid=interface.get("pvid"),
                active_plan=trackside_ap_plan,
            )
            station_projection = _station_consistency_projection(
                getattr(device, "station_id", ""),
                fit_ap.get("station_id"),
                pvid_projection.get("planning_station_id"),
            )
            effective_station_id = station_projection["effective_station_id"]
            base_station_name = (
                str((station_names or {}).get(effective_station_id) or "").strip()
                if effective_station_id and station_names is not None
                else str(device.station or normalize_station_value(fit_ap) or "").strip()
            )
            row = {
                    "station_id": effective_station_id,
                    **station_projection,
                    "site": base_station_name,
                    "ac_device_uuid": fit_ap.get("ac_device_uuid"),
                    "ap_uuid": fit_ap.get("ap_uuid") or (historical_lldp or {}).get("ap_uuid"),
                    "serial_number": fit_ap.get("serial_number") or fit_ap_resource_by_identity.get(ap_identity_key(fit_ap) or ("", ""), {}).get("serial_number"),
                    "model": ap_model,
                    "ap_optical_applicable": ap_optical_applicable,
                    "device_uuid": device_uuid,
                    "device_name": device.name,
                    "switch_vendor": device.device_vendor,
                    "switch_system_name": device.system_name,
                    "switch_primary_address": device.primary_address,
                    "switch_backup_address": device.backup_address,
                    "switch_identity": device_uuid,
                    "interface_name": interface_name,
                    "link_status": "DOWN" if switch_offline else link_state,
                    "protocol_status": interface.get("protocol_status") or interface.get("protocol"),
                    "description": interface.get("description"),
                    "port_type": _port_type(interface.get("port_status")),
                    "port_status": interface.get("port_status"),
                    "pvid": interface.get("pvid"),
                    "match_source": match_source,
                    "vlan": interface.get("vlan"),
                    "switch_rx_power": optical.get("rx_power"),
                    "switch_tx_power": optical.get("tx_power"),
                    "switch_rx_low_alarm": optical.get("rx_low_alarm"),
                    "switch_rx_high_alarm": optical.get("rx_high_alarm"),
                    "switch_tx_low_alarm": optical.get("tx_low_alarm"),
                    "switch_tx_high_alarm": optical.get("tx_high_alarm"),
                    "switch_device_optical_status": switch_status,
                    "switch_optical_status": switch_status,
                    "switch_interface_updated_at": interface.get("updated_at") or interface.get("collected_at"),
                    "switch_optical_updated_at": stored_optical.get("updated_at") or stored_optical.get("collected_at"),
                    "switch_interface_data_status": interface_data_status,
                    "switch_optical_data_status": optical_data_status,
                    "ap_mac": ap_candidate["ap_mac"],
                    "ap_name": ap_candidate["ap_name"],
                    "ap_ip": fit_ap.get("ap_ip"),
                    "ap_state": fit_ap.get("state"),
                    "ap_state_display": fit_ap.get("state_display") or fit_ap.get("state_raw"),
                    "ap_rx_power": ap_candidate["ap_rx_power"] if ap_side_has_data and not switch_offline else None,
                    "ap_tx_power": ap_candidate["ap_tx_power"] if ap_side_has_data and not switch_offline else None,
                    "ap_rx_low_alarm": fit_ap.get("rx_low_alarm"),
                    "ap_rx_low_warning": fit_ap.get("rx_low_warning"),
                    "ap_optical_status": ap_status,
                    "ap_optical_data_freshness": fit_ap.get("data_freshness"),
                    "ap_side_has_data": ap_side_has_data,
                    "updated_at": fit_ap.get("updated_at") or optical.get("updated_at") or interface.get("updated_at") or interface.get("collected_at"),
                    "source_device": fit_ap.get("device_name") or fit_ap.get("neighbor_device_name") or device.name,
                    "collection_status": fit_ap.get("status") or ("success" if optical else "not_collected"),
                    "switch_collection_status": switch_collection_status,
                    "offline_reason": offline_reason,
                    "status_reason": status_reason,
                    "data_source": data_source,
                    "ap_match_source": ap_match_source,
                    "ap_match_confidence": ap_match_confidence,
                    "lldp_match_status": lldp_match_status,
                    "lldp_observed_neighbor_mac": format_mac(
                        lldp.get("neighbor_mac") or lldp.get("chassis_id")
                    ),
                    "has_current_lldp": bool(lldp),
                    "has_historical_lldp": bool(historical_lldp),
                    "has_fit_ap_resource": bool(resource_from_neighbor or resource_from_identity),
                    "is_ap_offline": bool(offline_reason),
                    **attenuation,
                    **pvid_projection,
                }
            source_resources = _fit_ap_resource_sources(
                fit_ap,
                fit_ap_resources_by_mac,
                fit_ap_resources_by_identity,
            )
            source_ac_device_uuids = sorted(
                {
                    str(source.get("ac_device_uuid") or "").strip()
                    for source in source_resources
                    if str(source.get("ac_device_uuid") or "").strip()
                }
            )
            if source_ac_device_uuids:
                row["source_ac_device_uuids"] = source_ac_device_uuids
            _ensure_ap_optical_status(row)
            if runtime_snapshot is not None and not lldp:
                if runtime_snapshot.snapshot_status == "lldp_stale":
                    row["lldp_match_status"] = "LLDP_SNAPSHOT_STALE"
                elif runtime_snapshot.has_current_lldp:
                    row["lldp_match_status"] = "LLDP_EXACT_MATCH_PENDING"
                else:
                    row["lldp_match_status"] = "NO_CURRENT_EVIDENCE"
            row["lldp_history_status"] = classify_lldp_history_status(
                [lldp] if lldp else [],
                [historical_lldp] if historical_lldp else [],
            )
            if runtime_snapshot is not None:
                row["runtime_snapshot_status"] = runtime_snapshot.snapshot_status
                row["fit_ap_snapshot_collected_at"] = runtime_snapshot.fit_ap_collected_at
                row["lldp_snapshot_collected_at"] = runtime_snapshot.switch_lldp_collected_at
                row["lldp_snapshot_generation"] = runtime_snapshot.switch_lldp_generation
            elif historical_lldp_used and not lldp:
                row["lldp_history_status"] = "stale_snapshot"
            result.append(row)
    result.extend(
        _offline_ledger_to_trackside_rows(
            offline_ap_ledger_rows or [],
            interfaces_by_device,
            optical_by_device,
            latest_switch_collect_runs,
        )
    )
    result = [
        normalize_trackside_ap_business_row(
            row,
            business_projection=business_projection,
        )
        for row in _merge_duplicate_trackside_rows(result)
    ]
    _log_trackside_identity_coverage(
        devices,
        result,
        interfaces_by_device,
        lldp_by_device or {},
        fit_ap_resource_rows,
        fit_ap_optical_rows,
    )
    return sort_trackside_ap_business_rows(result)


def merge_fit_ap_rows_by_identity(rows: list[dict[str, object | None]]) -> list[dict[str, object | None]]:
    merged: dict[tuple[str, str], dict[str, object | None]] = {}
    passthrough: list[dict[str, object | None]] = []
    for row in rows or []:
        key = _ac_scoped_fit_ap_identity_key(row)
        if not key:
            passthrough.append(dict(row))
            continue
        existing = merged.get(key)
        if existing is None or _fit_ap_prefer_score(row) >= _fit_ap_prefer_score(existing):
            merged[key] = {**existing, **row} if existing else dict(row)
        else:
            for field, value in row.items():
                if _is_missing_display(existing.get(field)) and not _is_missing_display(value):
                    existing[field] = value
    return [*merged.values(), *passthrough]


def _latest_rows_by_normalized_interface(rows: list[dict[str, object | None]], field: str) -> dict[str, dict[str, object | None]]:
    latest: dict[str, dict[str, object | None]] = {}
    for row in rows or []:
        key = normalize_interface_name(row.get(field)).casefold()
        if not key:
            continue
        existing = latest.get(key)
        if existing is None or _fact_prefer_score(row) >= _fact_prefer_score(existing):
            latest[key] = row
    return latest


def _build_historical_lldp_index(rows: list[dict[str, object | None]]) -> dict[tuple[str, str], dict[str, object | None]]:
    latest: dict[tuple[str, str], dict[str, object | None]] = {}
    for row in rows or []:
        interface = normalize_interface_name(row.get("neighbor_interface")).casefold()
        if not interface:
            continue
        for name_field in ("neighbor_device_name", "neighbor_switch_name", "neighbor_switch_sysname", "lldp_neighbor"):
            switch_name = _normalize_name(row.get(name_field))
            if not switch_name:
                continue
            key = (switch_name, interface)
            current = latest.get(key)
            if current is None or _fact_prefer_score(row) >= _fact_prefer_score(current):
                latest[key] = row
    return latest


def _find_historical_lldp_row(
    index: dict[tuple[str, str], dict[str, object | None]],
    device_names: set[str],
    interface_name: object,
) -> dict[str, object | None]:
    interface = normalize_interface_name(interface_name).casefold()
    if not interface:
        return {}
    for name in device_names:
        row = index.get((name, interface))
        if row:
            return row
    return {}


def _log_trackside_identity_coverage(
    devices: list[Device],
    rows: list[dict[str, object | None]],
    interfaces_by_device: dict[str, list[dict[str, object | None]]],
    lldp_by_device: dict[str, list[dict[str, object | None]]],
    fit_ap_resource_rows: list[dict[str, object | None]],
    fit_ap_optical_rows: list[dict[str, object | None]],
) -> None:
    candidate_ap_interface_count = sum(
        1
        for device in devices
        for interface in interfaces_by_device.get(str(device.device_uuid or ""), [])
        if is_trackside_ap_interface(device, interface)[0]
    )
    rows_with_ap_identity = sum(
        1
        for row in rows
        if normalize_mac_key(row.get("ap_mac"))
        or str(row.get("ap_name") or "").strip()
    )
    rows_without_ap_identity = max(len(rows) - rows_with_ap_identity, 0)
    current_lldp_port_count = sum(len(items) for items in lldp_by_device.values())
    preserved_lldp_port_count = sum(1 for row in rows if row.get("data_source") == "historical_lldp")
    fit_ap_optical_success_count = sum(1 for row in fit_ap_optical_rows if str(row.get("status") or "").casefold() == "success")
    fit_ap_optical_failed_count = sum(1 for row in fit_ap_optical_rows if str(row.get("status") or "").casefold() not in {"", "success"})
    app_logger.log_info(
        "TRACKSIDE_AP_IDENTITY_COVERAGE",
        (
            f"station_switch_total={len(devices)}, candidate_ap_interface_count={candidate_ap_interface_count}, "
            f"current_lldp_port_count={current_lldp_port_count}, preserved_lldp_port_count={preserved_lldp_port_count}, "
            f"fit_ap_resource_count={len(fit_ap_resource_rows)}, fit_ap_optical_success_count={fit_ap_optical_success_count}, "
            f"fit_ap_optical_failed_count={fit_ap_optical_failed_count}, trackside_rows_total={len(rows)}, "
            f"rows_with_ap_identity={rows_with_ap_identity}, rows_without_ap_identity={rows_without_ap_identity}"
        ),
    )
    if rows_without_ap_identity:
        app_logger.log_warning(
            "TRACKSIDE_AP_IDENTITY_MISSING",
            (
                f"missing_count={rows_without_ap_identity}, trackside_rows_total={len(rows)}, "
                f"current_lldp_port_count={current_lldp_port_count}, "
                f"fit_ap_resource_count={len(fit_ap_resource_rows)}"
            ),
        )


def ap_identity_key(row: dict[str, object | None] | None) -> tuple[str, str] | None:
    if not row:
        return None
    ap_uuid = str(row.get("ap_uuid") or row.get("ap_identity") or "").strip()
    if ap_uuid and not ap_uuid.casefold().startswith("unauth-"):
        return ("uuid", ap_uuid.casefold())
    mac = normalize_mac_key(row.get("ap_mac") or row.get("mac"))
    if mac:
        return ("mac", mac.casefold())
    serial = str(row.get("serial_number") or row.get("serial") or "").strip()
    if serial and serial.casefold() not in {"-", "n/a", "na", "none"}:
        scope = str(
            row.get("site_key")
            or row.get("site_id")
            or row.get("station_id")
            or row.get("station_name")
            or row.get("station")
            or row.get("site")
            or row.get("ac_device_uuid")
            or ""
        ).strip().casefold()
        return ("serial", f"{scope}:{serial.casefold()}" if scope else serial.casefold())
    return None


def _ac_scoped_fit_ap_identity_key(
    row: dict[str, object | None],
) -> tuple[str, str] | None:
    identity = ap_identity_key(row)
    if identity is None:
        return None
    ac_device_uuid = str(
        row.get("ac_device_uuid") or row.get("device_uuid") or ""
    ).strip().casefold()
    if not ac_device_uuid:
        return identity
    return ("ac-resource", f"{ac_device_uuid}:{identity[0]}:{identity[1]}")


def _fit_ap_resource_sources(
    row: dict[str, object | None],
    by_mac: dict[str, list[dict[str, object | None]]],
    by_identity: dict[tuple[str, str], list[dict[str, object | None]]],
) -> list[dict[str, object | None]]:
    candidates: list[dict[str, object | None]] = []
    identity = ap_identity_key(row)
    if identity:
        candidates.extend(by_identity.get(identity, []))
    mac = normalize_mac_key(row.get("ap_mac") or row.get("mac"))
    if mac:
        candidates.extend(by_mac.get(mac, []))
    unique: dict[tuple[str, str], dict[str, object | None]] = {}
    for candidate in candidates:
        key = (
            str(candidate.get("ac_device_uuid") or ""),
            str(candidate.get("ap_uuid") or candidate.get("ap_name") or ""),
        )
        unique[key] = candidate
    return list(unique.values())


def _fit_ap_prefer_score(row: dict[str, object | None]) -> tuple[int, int, int, str]:
    return (
        1 if _ap_state(row) == "online" else 0,
        1 if str(row.get("ap_ip") or "").strip() else 0,
        1 if _has_optical_module_data(row) or _has_ap_lldp_data(row) else 0,
        f"{row.get('collected_at') or ''}|{row.get('updated_at') or ''}|{_int_value(row.get('id')):020d}",
    )


def _fact_prefer_score(row: dict[str, object | None]) -> tuple[str, str, int]:
    return (
        str(row.get("collected_at") or ""),
        str(row.get("updated_at") or ""),
        _int_value(row.get("id")),
    )


def _int_value(value: object) -> int:
    try:
        return int(str(value or "0"))
    except ValueError:
        return 0


def _has_ap_lldp_data(row: dict[str, object | None]) -> bool:
    return any(str(row.get(field) or "").strip() for field in ("neighbor_device_name", "neighbor_interface", "lldp_neighbor"))


def _merge_duplicate_ap_rows(rows: list[dict[str, object | None]]) -> list[dict[str, object | None]]:
    merged: dict[tuple[str, str], dict[str, object | None]] = {}
    passthrough: list[dict[str, object | None]] = []
    for row in rows:
        key = ap_identity_key(row)
        if not key:
            passthrough.append(row)
            continue
        existing = merged.get(key)
        if existing is None:
            merged[key] = dict(row)
            continue
        if _trackside_ap_row_prefer_score(row) >= _trackside_ap_row_prefer_score(existing):
            merged[key] = _merged_trackside_ap_row(row, existing)
        else:
            merged[key] = _merged_trackside_ap_row(existing, row)
    return [*merged.values(), *passthrough]


def _trackside_ap_row_prefer_score(row: dict[str, object | None]) -> tuple[int, int, int, str]:
    return (
        0 if bool(row.get("is_ap_offline")) or row.get("offline_reason") in _AP_OFFLINE_REASONS else 1,
        1 if _ap_state(row) == "online" else 0,
        1 if str(row.get("ap_ip") or "").strip() else 0,
        str(row.get("updated_at") or ""),
    )


def _merged_trackside_ap_row(primary: dict[str, object | None], secondary: dict[str, object | None]) -> dict[str, object | None]:
    result = dict(primary)
    for field, value in secondary.items():
        if _is_missing_display(result.get(field)) and not _is_missing_display(value):
            result[field] = value
    return result


def build_new_online_ap_overview_rows(
    current_resource_rows: list[dict[str, object | None]],
    resource_history_rows: list[dict[str, object | None]],
    trackside_rows: list[dict[str, object | None]],
    unauthenticated_rows: list[dict[str, object | None]] | None = None,
    unauthenticated_history_rows: list[dict[str, object | None]] | None = None,
) -> list[dict[str, object | None]]:
    del resource_history_rows, unauthenticated_history_rows
    # ``display wlan ap unauthenticated`` Current is the sole membership
    # authority.  FIT-AP resources and their history may enrich a row, but
    # runtime transitions are deliberately not promoted to "new online".
    if unauthenticated_rows is None:
        return []
    return _build_unauthenticated_new_online_rows(
        current_resource_rows,
        trackside_rows,
        unauthenticated_rows or [],
    )


def _build_unauthenticated_new_online_rows(
    current_resource_rows: list[dict[str, object | None]],
    trackside_rows: list[dict[str, object | None]],
    unauthenticated_rows: list[dict[str, object | None]],
) -> list[dict[str, object | None]]:
    resources_by_key = _new_online_identity_index(current_resource_rows)
    trackside_by_key = _new_online_identity_index(trackside_rows)
    rows: list[dict[str, object | None]] = []
    emitted: set[tuple[str, str] | tuple[str, int]] = set()
    for index, source in enumerate(unauthenticated_rows):
        source_name = str(source.get("source") or "").strip()
        if source_name != WLAN_AP_UNAUTHENTICATED_SOURCE:
            continue
        keys = _new_online_identity_keys(source)
        primary_key: tuple[str, str] | tuple[str, int] = keys[0] if keys else ("row", index)
        if primary_key in emitted:
            continue
        emitted.add(primary_key)
        resource = _first_index_match(keys, resources_by_key)
        trackside = _first_index_match(keys, trackside_by_key)
        rows.append(
            {
                "identity_source": "AC未固化Auto AP",
                "source": WLAN_AP_UNAUTHENTICATED_SOURCE,
                "register_status": "未固化",
                "new_online_status": "当前新上线Auto AP",
                "site": (
                    (resource or {}).get("station")
                    or (resource or {}).get("station_name")
                    or trackside.get("station_name")
                    or trackside.get("station")
                    or (resource or {}).get("site")
                    or (resource or {}).get("site_name")
                    or trackside.get("site")
                    or "等待 LLDP 同步"
                ),
                "device_name": trackside.get("device_name") or "-",
                "interface_name": trackside.get("interface_name") or "-",
                "link_status": trackside.get("link_status") or "-",
                "port_type": trackside.get("port_type") or "-",
                "description": trackside.get("description") or "-",
                "pvid": trackside.get("pvid") or "-",
                "vlan": trackside.get("vlan") or "-",
                "switch_rx_power": trackside.get("switch_rx_power") or "-",
                "switch_optical_status": trackside.get("switch_optical_status") or "-",
                "ap_mac": (
                    format_mac((resource or {}).get("ap_mac"))
                    or format_mac(source.get("ap_mac"))
                    or format_mac(source.get("inferred_ap_mac"))
                    or "-"
                ),
                "ap_name": source.get("ap_name") or (resource or {}).get("ap_name") or "-",
                "ap_rx_power": trackside.get("ap_rx_power") or "-",
                "ap_optical_status": trackside.get("ap_optical_status") or "-",
                "updated_at": (resource or {}).get("updated_at") or source.get("updated_at") or source.get("collected_at"),
                "apid": source.get("apid") or (resource or {}).get("apid"),
                "current_unauthenticated": "是",
                "current_resource_exists": "是" if bool(resource) else "否",
                "serial_number": source.get("serial_number") or (resource or {}).get("serial_number"),
                "model": source.get("model") or (resource or {}).get("model"),
                "ap_ip": (resource or {}).get("ap_ip"),
                "group_name": (resource or {}).get("group_name"),
                "state_display": source.get("state_display") or source.get("state_raw") or source.get("state") or (resource or {}).get("state_display"),
                "ac_device_name": (resource or {}).get("ac_device_name") or (resource or {}).get("device_name") or source.get("ac_device_uuid"),
                "last_unauthenticated_at": source.get("collected_at") or source.get("created_at"),
                "suggestion": "新上线Auto AP，确认点位后在AC手动固化AP",
                "first_seen_at": source.get("collected_at") or source.get("created_at"),
                "identity_entity_id": (resource or {}).get("identity_entity_id") or "",
                "baseline_collected_at": "",
                "current_collected_at": (resource or {}).get("collected_at") or source.get("collected_at") or "",
            }
        )
    return sorted(rows, key=lambda row: (str(row.get("site") or ""), str(row.get("ap_name") or "")))


def _new_online_identity_index(rows: list[dict[str, object | None]]) -> dict[tuple[str, str], dict[str, object | None]]:
    index: dict[tuple[str, str], dict[str, object | None]] = {}
    for row in rows or []:
        for key in _new_online_identity_keys(row):
            index.setdefault(key, row)
    return index


def _new_online_identity_keys(row: dict[str, object | None]) -> list[tuple[str, str]]:
    keys: list[tuple[str, str]] = []
    site_key = next(
        (
            str(row.get(field) or "").strip().casefold()
            for field in ("site_key", "site_id", "station_name", "station", "site")
            if str(row.get(field) or "").strip()
        ),
        "",
    )
    entity_id = str(
        row.get("identity_entity_id")
        or row.get("ap_identity_entity_id")
        or row.get("identity_entity_uuid")
        or ""
    ).strip()
    mac = normalize_mac_key(row.get("ap_mac") or row.get("inferred_ap_mac") or row.get("mac"))
    if mac:
        if site_key:
            keys.append(("site_mac", f"{site_key}:{mac.casefold()}"))
    if entity_id:
        keys.append(("entity", entity_id.casefold()))
    serial = str(row.get("serial_number") or row.get("serial") or "").strip()
    if serial and serial not in {"-", "N/A", "n/a"}:
        if site_key:
            keys.append(("site_serial", f"{site_key}:{serial.casefold()}"))
    ac_uuid = str(row.get("ac_device_uuid") or "").strip()
    apid = str(row.get("apid") or row.get("ap_id") or "").strip()
    if site_key and ac_uuid and apid:
        keys.append(("site_apid", f"{site_key}:{ac_uuid.casefold()}:{apid.casefold()}"))
    return keys


def _first_index_match(keys: list[tuple[str, str]], index: dict[tuple[str, str], dict[str, object | None]]) -> dict[str, object | None]:
    for key in keys:
        row = index.get(key)
        if row:
            return row
    return {}


def build_ap_optical_treatment_records(
    trackside_rows: list[dict[str, object | None]],
    ap_optical_history_rows: list[dict[str, object | None]],
    switch_optical_history_rows: list[dict[str, object | None]],
    fit_ap_resource_rows: list[dict[str, object | None]] | None = None,
    resource_history_rows: list[dict[str, object | None]] | None = None,
    offline_ledger_rows: list[dict[str, object | None]] | None = None,
) -> list[dict[str, object | None]]:
    records: list[dict[str, object | None]] = []
    identity_lookup = _ap_identity_lookup_with_sources(
        (
            ("trackside_row", trackside_rows or []),
            ("ap_optical_history", ap_optical_history_rows or []),
            ("fit_ap_resource", fit_ap_resource_rows or []),
            ("resource_history", resource_history_rows or []),
            ("offline_ledger", offline_ledger_rows or []),
        )
    )
    trackside_by_identity = _trackside_rows_by_ap_identity(trackside_rows)
    ap_history_by_identity = _ap_optical_history_by_identity(ap_optical_history_rows)
    switch_history_by_interface = _switch_optical_history_by_interface(switch_optical_history_rows)
    trackside_interface_lookup = _trackside_rows_by_switch_name_interface(trackside_rows)
    offline_ledger_interface_lookup = _offline_ledger_rows_by_switch_interface(offline_ledger_rows or [])
    for _key, current in trackside_by_identity.items():
        history = _ap_history_for_trackside(current, ap_history_by_identity)
        current_item = {
            **current,
            "side": "ap",
            "rx_power": current.get("ap_rx_power"),
            "optical_alarm_status": current.get("ap_optical_status"),
            "collected_at": current.get("updated_at"),
        }
        for record in _build_treatment_records_for_series(AP_SIDE_LABEL, current, history, current_item):
            enrich_treatment_record_ap_identity(
                record,
                identity_lookup=identity_lookup,
                trackside_interface_lookup=trackside_interface_lookup,
                offline_ledger_interface_lookup=offline_ledger_interface_lookup,
            )
            records.append(record)

    switch_rows = _trackside_rows_by_switch_interface(trackside_rows)
    for key, current in switch_rows.items():
        history = switch_history_by_interface.get(key, [])
        current_item = {
            **current,
            "side": "switch",
            "rx_power": current.get("switch_rx_power"),
            "optical_alarm_status": current.get("switch_optical_status"),
            "collected_at": current.get("updated_at"),
        }
        for record in _build_treatment_records_for_series(SWITCH_SIDE_LABEL, current, history, current_item):
            enrich_treatment_record_ap_identity(
                record,
                identity_lookup=identity_lookup,
                trackside_interface_lookup=trackside_interface_lookup,
                offline_ledger_interface_lookup=offline_ledger_interface_lookup,
            )
            records.append(record)
    for record in records:
        enrich_treatment_record_ap_identity(
            record,
            identity_lookup=identity_lookup,
            trackside_interface_lookup=trackside_interface_lookup,
            offline_ledger_interface_lookup=offline_ledger_interface_lookup,
        )
    return sorted(records, key=lambda row: (str(row.get("first_found_at") or ""), str(row.get("site") or ""), str(row.get("ap_name") or "")))


def _build_treatment_records_for_series(
    side_label: str,
    trackside: dict[str, object | None],
    history_rows: list[dict[str, object | None]],
    current_item: dict[str, object | None],
) -> list[dict[str, object | None]]:
    timeline = sorted([dict(row) for row in history_rows or []], key=lambda row: (str(row.get("collected_at") or ""), _int_value(row.get("id"))))
    if current_item:
        timeline.append(dict(current_item))
    records: list[dict[str, object | None]] = []
    open_record: dict[str, object | None] | None = None
    latest_status = ""
    latest_rx_power = None
    for item in timeline:
        status = _optical_treatment_status(item)
        if status in OPTICAL_TREATMENT_IGNORED_STATUSES and status != "normal":
            continue
        collected_at = item.get("collected_at") or item.get("updated_at") or item.get("created_at")
        rx_power = item.get("rx_power")
        latest_status = status or latest_status
        latest_rx_power = rx_power if rx_power not in (None, "") else latest_rx_power
        if status in OPTICAL_TREATMENT_ISSUE_STATUSES:
            if open_record is None:
                open_record = _new_treatment_record(side_label, trackside, item, status, collected_at, rx_power)
            continue
        if status == "normal" and open_record is not None:
            open_record["completed_at"] = collected_at
            open_record["fixed_rx_power"] = rx_power
            open_record["current_rx_power"] = rx_power
            open_record["current_status"] = display_optical_status("normal")
            open_record["treatment_status"] = TREATMENT_CLOSED_LABEL
            records.append(open_record)
            open_record = None
    if open_record is not None:
        open_record["current_rx_power"] = latest_rx_power
        open_record["current_status"] = display_optical_status(latest_status)
        open_record["treatment_status"] = TREATMENT_OPEN_LABEL
        records.append(open_record)
    return records


def _new_treatment_record(
    side_label: str,
    trackside: dict[str, object | None],
    item: dict[str, object | None],
    status: str,
    collected_at: object,
    rx_power: object,
) -> dict[str, object | None]:
    return {
        "site": trackside.get("site"),
        "ap_name": trackside.get("ap_name"),
        "ap_mac": trackside.get("ap_mac"),
        "serial_number": trackside.get("serial_number"),
        "ap_id": trackside.get("ap_id") or trackside.get("apid"),
        "section_name": trackside.get("section_name") or trackside.get("belong_section"),
        "direction": trackside.get("direction"),
        "side": side_label,
        "device_name": trackside.get("device_name"),
        "interface_name": trackside.get("interface_name"),
        "issue_type": _optical_issue_type(status),
        "first_found_at": collected_at,
        "completed_at": "",
        "first_rx_power": rx_power,
        "fixed_rx_power": "",
        "current_rx_power": rx_power,
        "current_status": display_optical_status(status),
        "treatment_status": TREATMENT_OPEN_LABEL,
        "remark": "",
    }


def _optical_treatment_status(row: dict[str, object | None]) -> str:
    status = row.get("optical_alarm_status") or row.get("alarm_status") or row.get("status")
    text = str(status or "").strip().casefold()
    if text in OPTICAL_TREATMENT_ISSUE_STATUSES or text in OPTICAL_TREATMENT_IGNORED_STATUSES:
        return text
    if not row:
        return ""
    return _optical_status_from_history(row, "ap").casefold()


def _optical_issue_type(status: str) -> str:
    return {
        "notice": ISSUE_TYPE_NOTICE_LABEL,
        "warning": ISSUE_TYPE_NOTICE_LABEL,
        "alarm": ISSUE_TYPE_ALARM_LABEL,
        "link_abnormal": ISSUE_TYPE_LINK_ABNORMAL_LABEL,
        "link_down": ISSUE_TYPE_LINK_ABNORMAL_LABEL,
        "no_light": ISSUE_TYPE_NO_LIGHT_LABEL,
    }.get(status, ISSUE_TYPE_OPTICAL_ABNORMAL_LABEL)


def _ap_optical_history_matches_trackside(row: dict[str, object | None], trackside: dict[str, object | None]) -> bool:
    row_serial = str(row.get("serial_number") or "").strip()
    current_serial = str(trackside.get("serial_number") or "").strip()
    if row_serial and current_serial:
        return row_serial == current_serial
    row_mac = normalize_mac_key(row.get("ap_mac"))
    current_mac = normalize_mac_key(trackside.get("ap_mac"))
    if row_mac and current_mac:
        return row_mac == current_mac
    row_name = str(row.get("ap_name") or "").strip().casefold()
    current_name = str(trackside.get("ap_name") or "").strip().casefold()
    if row_name and current_name:
        return row_name == current_name
    row_uuid = str(row.get("ap_uuid") or "").strip()
    current_uuid = str(trackside.get("ap_uuid") or "").strip()
    return bool(row_uuid and current_uuid and row_uuid == current_uuid)


def _ap_identity_lookup(rows: list[dict[str, object | None]]) -> dict[tuple[str, str], dict[str, object | None]]:
    return _ap_identity_lookup_with_sources((("unknown", rows),))


def _ap_identity_lookup_with_sources(
    sources: tuple[tuple[str, list[dict[str, object | None]]], ...],
) -> dict[tuple[str, str], dict[str, object | None]]:
    lookup: dict[tuple[str, str], dict[str, object | None]] = {}
    for source, rows in sources:
        for row in rows or []:
            payload = _ap_identity_payload(row, source)
            for key in _ap_identity_keys(payload):
                existing = lookup.get(key, {})
                merged = _merge_ap_identity_payload(existing, payload)
                if not merged.get("_source"):
                    merged["_source"] = source
                lookup[key] = merged
    return lookup


def _ap_identity_payload(row: dict[str, object | None], source: str = "") -> dict[str, object | None]:
    ap_name = row.get("ap_name")
    ap_mac = format_mac(row.get("ap_mac"))
    serial_number = row.get("serial_number")
    site_key = str(row.get("site_key") or row.get("site_id") or "").strip()
    station_key = next(
        (
            str(row.get(field) or "").strip()
            for field in ("station_name", "station", "ap_station", "ownership_station", "site")
            if str(row.get(field) or "").strip()
            and str(row.get(field) or "").strip().casefold() != site_key.casefold()
        ),
        "",
    )
    return {
        "ap_uuid": row.get("ap_uuid"),
        "ap_name": ap_name,
        "ap_mac": ap_mac,
        "serial_number": serial_number,
        "ap_id": row.get("ap_id") or row.get("apid"),
        "section_name": row.get("section_name") or row.get("belong_section"),
        "direction": row.get("direction"),
        "site_key": site_key,
        "station_key": station_key,
        "site": station_key,
        "device_name": row.get("device_name") or row.get("historical_switch_name"),
        "interface_name": row.get("interface_name") or row.get("historical_switch_interface"),
        "_serial_source": source if not _is_missing_display(serial_number) else "",
        "_source": source,
    }


def _ap_optical_history_by_identity(rows: list[dict[str, object | None]]) -> dict[tuple[str, str], list[dict[str, object | None]]]:
    grouped: dict[tuple[str, str], list[dict[str, object | None]]] = {}
    for row in rows or []:
        for key in _ap_identity_keys(row):
            grouped.setdefault(key, []).append(row)
    return grouped


def _ap_history_for_trackside(
    trackside: dict[str, object | None],
    history_by_identity: dict[tuple[str, str], list[dict[str, object | None]]],
) -> list[dict[str, object | None]]:
    seen: set[int] = set()
    history: list[dict[str, object | None]] = []
    for key in _ap_identity_keys(trackside):
        for row in history_by_identity.get(key, []):
            marker = id(row)
            if marker in seen:
                continue
            seen.add(marker)
            history.append(row)
    return history


def _switch_optical_history_by_interface(rows: list[dict[str, object | None]]) -> dict[tuple[str, str], list[dict[str, object | None]]]:
    grouped: dict[tuple[str, str], list[dict[str, object | None]]] = {}
    for row in rows or []:
        device_uuid = str(row.get("device_uuid") or "")
        interface_key = normalize_interface_name(row.get("interface_name")).casefold()
        if device_uuid and interface_key:
            grouped.setdefault((device_uuid, interface_key), []).append(row)
    return grouped


def _trackside_rows_by_switch_name_interface(rows: list[dict[str, object | None]]) -> dict[tuple[str, str], dict[str, object | None]]:
    lookup: dict[tuple[str, str], dict[str, object | None]] = {}
    for row in rows or []:
        key = (_normalize_name(row.get("device_name")), normalize_interface_name(row.get("interface_name")).casefold())
        if all(key):
            lookup[key] = row
    return lookup


def _offline_ledger_rows_by_switch_interface(rows: list[dict[str, object | None]]) -> dict[tuple[str, str], dict[str, object | None]]:
    lookup: dict[tuple[str, str], dict[str, object | None]] = {}
    for row in rows or []:
        key = (
            _normalize_name(row.get("historical_switch_name") or row.get("device_name")),
            normalize_interface_name(row.get("historical_switch_interface") or row.get("interface_name")).casefold(),
        )
        if all(key):
            lookup[key] = row
    return lookup


def _ap_identity_keys(row: dict[str, object | None]) -> list[tuple[str, str]]:
    keys: list[tuple[str, str]] = []
    site_root = str(row.get("site_key") or row.get("site_id") or "").strip().casefold()
    scope_keys: list[str] = []
    for field in ("site_key", "site_id"):
        value = str(row.get(field) or "").strip().casefold()
        if value and value not in scope_keys:
            scope_keys.append(value)
    for field in ("station_key", "station_name", "station", "site"):
        value = str(row.get(field) or "").strip().casefold()
        if value and value != site_root and value not in scope_keys:
            scope_keys.append(value)
    ap_uuid = str(row.get("ap_uuid") or "").strip()
    if ap_uuid:
        keys.append(("uuid", ap_uuid.casefold()))
    mac = normalize_mac_key(row.get("ap_mac") or row.get("mac"))
    if mac:
        for scope_key in scope_keys:
            keys.append(("site_mac", f"{scope_key}:{mac.casefold()}"))
    serial = str(row.get("serial_number") or row.get("serial") or "").strip()
    if serial and serial not in {"-", "N/A", "n/a"}:
        for scope_key in scope_keys:
            keys.append(("site_serial", f"{scope_key}:{serial.casefold()}"))
    return keys


def _merge_ap_identity_payload(
    primary: dict[str, object | None],
    secondary: dict[str, object | None],
) -> dict[str, object | None]:
    result = dict(primary)
    if result.get(_AP_IDENTITY_AMBIGUOUS):
        return result
    primary_serial = result.get("serial_number")
    secondary_serial = secondary.get("serial_number")
    for field in ("ap_uuid", "ap_mac", "serial_number"):
        first = str(result.get(field) or "").strip().casefold()
        second = str(secondary.get(field) or "").strip().casefold()
        if first and second and first != second:
            return {_AP_IDENTITY_AMBIGUOUS: True}
    for field in (
        "ap_uuid",
        "ap_name",
        "ap_mac",
        "serial_number",
        "ap_id",
        "section_name",
        "direction",
        "site_key",
        "station_key",
        "site",
        "device_name",
        "interface_name",
    ):
        if _is_missing_display(result.get(field)) and not _is_missing_display(secondary.get(field)):
            result[field] = secondary.get(field)
    if (
        _is_missing_display(primary_serial)
        and not _is_missing_display(secondary_serial)
    ):
        result["_serial_source"] = secondary.get("_serial_source") or secondary.get("_source") or ""
    primary_serial_source = str(
        result.get("_serial_source") or result.get("_source") or ""
    )
    secondary_serial_source = str(
        secondary.get("_serial_source") or secondary.get("_source") or ""
    )
    if (
        not _is_missing_display(secondary.get("serial_number"))
        and secondary_serial_source in _TREATMENT_SERIAL_IDENTITY_SOURCES
        and primary_serial_source not in _TREATMENT_SERIAL_IDENTITY_SOURCES
    ):
        result["_serial_source"] = secondary_serial_source
    if _is_missing_display(result.get("_source")) and not _is_missing_display(secondary.get("_source")):
        result["_source"] = secondary.get("_source")
    return result


def enrich_treatment_record_ap_identity(
    record: dict[str, object | None],
    *,
    identity_lookup: dict[tuple[str, str], dict[str, object | None]],
    trackside_interface_lookup: dict[tuple[str, str], dict[str, object | None]],
    offline_ledger_interface_lookup: dict[tuple[str, str], dict[str, object | None]],
) -> tuple[dict[str, object | None], str]:
    before_ap_name = record.get("ap_name")
    before_ap_mac = record.get("ap_mac")
    source = "not_found"
    for key in _ap_identity_keys(record):
        matched = identity_lookup.get(key)
        if matched and not matched.get(_AP_IDENTITY_AMBIGUOUS):
            _fill_treatment_record_identity(record, matched)
            source = f"{matched.get('_source') or 'identity'}_{key[0]}"
            break
    if _record_identity_missing(record):
        interface_key = (_normalize_name(record.get("device_name")), normalize_interface_name(record.get("interface_name")).casefold())
        matched = trackside_interface_lookup.get(interface_key)
        if matched:
            _fill_treatment_record_identity(record, _ap_identity_payload(matched, "trackside_row"))
            source = "trackside_row"
        if _record_identity_missing(record):
            matched = offline_ledger_interface_lookup.get(interface_key)
            if matched:
                _fill_treatment_record_identity(record, _ap_identity_payload(matched, "offline_ledger"))
                source = "offline_ledger_interface"
    changed = before_ap_name != record.get("ap_name") or before_ap_mac != record.get("ap_mac")
    if changed:
        app_logger.log_info(
            "TRACKSIDE_AP_TREATMENT_IDENTITY_ENRICH",
            " ".join(
                [
                    f"side={record.get('side') or ''}",
                    f"station={record.get('site') or ''}",
                    f"switch={record.get('device_name') or ''}",
                    f"interface={record.get('interface_name') or ''}",
                    f"before_ap_name={before_ap_name or '-'}",
                    f"before_ap_mac={before_ap_mac or '-'}",
                    f"serial={record.get('serial_number') or ''}",
                    f"after_ap_name={record.get('ap_name') or '-'}",
                    f"after_ap_mac={record.get('ap_mac') or '-'}",
                    f"source={source}",
                ]
            ),
        )
    elif _record_identity_missing(record):
        app_logger.log_info(
            "TRACKSIDE_AP_TREATMENT_IDENTITY_MISSING",
            " ".join(
                [
                    f"side={record.get('side') or ''}",
                    f"station={record.get('site') or ''}",
                    f"switch={record.get('device_name') or ''}",
                    f"interface={record.get('interface_name') or ''}",
                    f"serial={record.get('serial_number') or ''}",
                    "reason=no_trackside_or_offline_ledger_match",
                ]
            ),
        )
    return record, source


def _fill_treatment_record_identity(record: dict[str, object | None], source: dict[str, object | None]) -> None:
    fields = (
        "ap_name",
        "ap_mac",
        "ap_id",
        "section_name",
        "direction",
        "site",
    )
    for field in fields:
        if _is_missing_display(record.get(field)) and not _is_missing_display(source.get(field)):
            record[field] = source.get(field)
    serial_source = str(source.get("_serial_source") or source.get("_source") or "")
    if (
        serial_source in _TREATMENT_SERIAL_IDENTITY_SOURCES
        and _is_missing_display(record.get("serial_number"))
        and not _is_missing_display(source.get("serial_number"))
    ):
        record["serial_number"] = source.get("serial_number")


def _record_identity_missing(record: dict[str, object | None]) -> bool:
    return (
        _is_missing_display(record.get("ap_name"))
        or _is_missing_display(record.get("ap_mac"))
        or _is_missing_display(record.get("serial_number"))
    )


def _trackside_rows_by_ap_identity(rows: list[dict[str, object | None]]) -> dict[tuple[str, str], dict[str, object | None]]:
    result: dict[tuple[str, str], dict[str, object | None]] = {}
    for row in rows or []:
        key = ap_identity_key(row)
        if key and (key not in result or _trackside_ap_row_prefer_score(row) >= _trackside_ap_row_prefer_score(result[key])):
            result[key] = row
    return result


def _trackside_rows_by_switch_interface(rows: list[dict[str, object | None]]) -> dict[tuple[str, str], dict[str, object | None]]:
    result: dict[tuple[str, str], dict[str, object | None]] = {}
    for row in rows or []:
        device_uuid = str(row.get("device_uuid") or "")
        interface_key = normalize_interface_name(row.get("interface_name")).casefold()
        if device_uuid and interface_key:
            result[(device_uuid, interface_key)] = row
    return result


def _trackside_row_status_with_ap_status(
    row: dict[str, object | None],
    ap_status: str,
) -> str:
    if (
        row.get("offline_reason") == "switch_offline"
        or _is_switch_collection_offline(row.get("switch_collection_status"))
    ):
        return "offline"
    if normalize_link_state(row.get("link_status")) == "DOWN":
        return "link_down"
    switch_status = _normalized_optical_status(row.get("switch_optical_status"))
    if "critical" in {switch_status, ap_status}:
        return "critical"
    return worse_optical_severity(switch_status, ap_status)


def trackside_row_status(row: dict[str, object | None]) -> str:
    if not is_ap_optical_applicable(row.get("model") or row.get("ap_model")):
        return OPTICAL_NOT_APPLICABLE_STATUS
    if (
        row.get("offline_reason") == "switch_offline"
        or _is_switch_collection_offline(row.get("switch_collection_status"))
    ):
        return "offline"
    if normalize_link_state(row.get("link_status")) == "DOWN":
        return "link_down"
    status = _normalized_optical_status(row.get("ap_business_optical_status"))
    if status:
        return status
    return _dual_business_evaluation(row).status


_TRACKSIDE_EXPORT_MISSING_IDENTITY_VALUES = frozenset(
    {"", "-", "—", "n/a", "na", "none", "null"}
)


def _has_trackside_export_identity(value: object) -> bool:
    return str(value or "").strip().casefold() not in _TRACKSIDE_EXPORT_MISSING_IDENTITY_VALUES


def has_trackside_export_ap_evidence(row: dict[str, object | None]) -> bool:
    """Return whether a port has current or historical AP business evidence."""

    for field in (
        "has_current_lldp",
        "has_historical_lldp",
        "has_fit_ap_resource",
    ):
        value = row.get(field)
        if isinstance(value, str):
            if value.strip().casefold() not in {"", "0", "false", "no", "none", "null"}:
                return True
        elif bool(value):
            return True
    if any(_has_trackside_export_identity(row.get(field)) for field in ("ap_uuid", "ap_mac", "ap_name")):
        return True
    history_status = str(row.get("lldp_history_status") or "").strip().casefold()
    return history_status not in {"", "no_current_evidence", "none", "unknown", "-"}


def _trackside_export_switch_statuses(row: dict[str, object | None]) -> set[str]:
    statuses = {
        _normalized_optical_status(row.get(field))
        for field in ("switch_optical_status", "switch_device_optical_status")
    }
    statuses.discard("")
    return statuses


def trackside_export_fill_status(row: dict[str, object | None]) -> str | None:
    """Return the export-only row fill status without changing business semantics."""

    switch_statuses = _trackside_export_switch_statuses(row)
    if "no_module" in switch_statuses or _explicit_no_module(row):
        return None
    if "no_light" in switch_statuses and not has_trackside_export_ap_evidence(row):
        return None
    return trackside_row_status(row)


def is_trackside_optical_abnormal_status(status: object) -> bool:
    """Return whether a trackside optical status belongs in the optical anomaly set."""
    return is_optical_health_abnormal(_normalized_optical_status(status))


def has_ap_side_optical_data(row: dict[str, object | None]) -> bool:
    if not row:
        return False
    if not is_ap_optical_applicable(row.get("model") or row.get("ap_model")):
        return False
    if bool(row.get("is_ap_offline")):
        return True
    if "ap_side_has_data" in row:
        return bool(row.get("ap_side_has_data"))
    if _explicit_no_module(row):
        return True
    if _is_missing_display(row.get("ap_mac")) or _is_missing_display(row.get("ap_name")) or _is_missing_display(row.get("ap_rx_power")):
        return False
    return _has_optical_module_data({"rx_power": row.get("ap_rx_power"), "tx_power": row.get("ap_tx_power")})


def format_ap_side_alarm(row: dict[str, object | None], language: str = "zh") -> str:
    if not is_ap_optical_applicable(row.get("model") or row.get("ap_model")):
        return display_optical_status(OPTICAL_NOT_APPLICABLE_STATUS, language)
    if bool(row.get("is_ap_offline")) or row.get("offline_reason") in _AP_OFFLINE_REASONS:
        return display_optical_status("unknown", language)
    if not has_ap_side_optical_data(row):
        return display_optical_status("unknown", language)
    status = _dual_business_evaluation(row).ap_status
    if classify_optical_health(status) == "no_data":
        status = "unknown"
    return display_optical_status(status, language)


def _ensure_ap_optical_status(row: dict[str, object | None]) -> None:
    if bool(row.get("is_ap_offline")) or row.get("offline_reason") in _AP_OFFLINE_REASONS:
        return
    status = str(row.get("ap_optical_status") or "").strip().casefold()
    if status not in {"", "unknown"} or not _has_valid_rx_power(row.get("ap_rx_power")):
        return
    row["ap_optical_status"] = _compute_ap_optical_status_from_row(row)
    row["ap_side_has_data"] = True


def _ap_optical_status_for_display(row: dict[str, object | None]) -> str:
    if bool(row.get("is_ap_offline")) or row.get("offline_reason") in _AP_OFFLINE_REASONS:
        return "offline"
    status = str(row.get("ap_optical_status") or "").strip().casefold()
    if status in {"", "unknown"} and _has_valid_rx_power(row.get("ap_rx_power")):
        return _compute_ap_optical_status_from_row(row)
    return status


def _compute_ap_optical_status_from_row(row: dict[str, object | None]) -> str:
    return compute_optical_severity(
        {
            "module_present": True,
            "no_module": _explicit_no_module(row),
            "ap_rx_power": row.get("ap_rx_power") or row.get("rx_power"),
            "ap_port_status": row.get("ap_port_status") or row.get("port_status"),
            "alarm_low": row.get("ap_rx_low_alarm") or row.get("rx_low_alarm"),
            "warning_low": row.get("ap_rx_low_warning") or row.get("rx_low_warning"),
            "device_type": "ap",
        }
    ).severity


def _has_valid_rx_power(value: object) -> bool:
    return _float_value(value) is not None


def _float_value(value: object) -> float | None:
    if value is None:
        return None
    match = re.search(r"[-+]?\d+(?:\.\d+)?", str(value))
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def format_trackside_display_value(field: str, row: dict[str, object | None], language: str = "zh") -> str:
    if field == "match_source":
        return _match_source_label(row.get(field), language)
    if field == "port_type":
        return _port_type(row.get("port_type") or row.get("port_status"))
    if field == "link_status":
        return normalize_link_state(row.get("link_status") or row.get("link") or row.get("status"))
    if field == "protocol_status":
        return normalize_link_state(row.get("protocol_status") or row.get("protocol"))
    if field == "vlan":
        return normalize_trackside_vlan_display(row.get("vlan"))
    if field == "ap_optical_status":
        return format_ap_side_alarm(row, language)
    if field == "ap_device_optical_status":
        value = row.get(field)
        return display_optical_status(str(value), language) if value else AP_SIDE_MISSING_DISPLAY
    if field == "ap_business_threshold_dbm":
        return (
            f"AP Rx ≥ {AP_BUSINESS_RX_MIN_DBM:.2f} dBm 且"
            f"交换机 Rx ≥ {AP_BUSINESS_RX_MIN_DBM:.2f} dBm"
        )
    if field in AP_SIDE_DISPLAY_FIELDS and not has_ap_side_optical_data(row):
        return AP_SIDE_MISSING_DISPLAY
    value = row.get(field)
    if field == "switch_optical_status":
        if row.get("offline_reason") == "switch_offline" or _is_switch_collection_offline(row.get("switch_collection_status")):
            return "交换机离线" if not language.startswith("en") else "Switch Offline"
        evaluation = evaluate_dual_rx_business_detail(
            row.get("ap_rx_power"),
            row.get("switch_rx_power"),
            ap_reported_status=(
                row.get("ap_device_optical_status") or row.get("ap_optical_status")
            ),
            switch_reported_status=(
                row.get("switch_device_optical_status") or value
            ),
            ap_data_freshness=(
                row.get("ap_optical_data_freshness") or row.get("data_freshness")
            ),
            switch_data_freshness=row.get("switch_optical_data_status"),
        )
        return display_optical_status(evaluation.switch_status, language)
    if field == "ap_optical_status" and value:
        return display_optical_status(str(value), language)
    return str(value) if value not in (None, "") else AP_SIDE_MISSING_DISPLAY


def _match_source_label(value: object, language: str = "zh") -> str:
    source = str(value or "none")
    if language.startswith("en"):
        return {
            "description": "Description",
            "pvid": "PVID",
            "description+pvid": "Description+PVID",
            "none": "-",
        }.get(source, source or "-")
    return MATCH_SOURCE_LABELS.get(source, source or "-")


def filter_trackside_ap_business_rows(rows: list[dict[str, object | None]], site: object = "", search: object = "") -> list[dict[str, object | None]]:
    site_text = str(site or "").strip()
    search_text = str(search or "").strip().casefold()
    result = rows
    if site_text:
        result = [row for row in result if str(row.get("site") or "").strip() == site_text]
    if search_text:
        fields = ("ap_name", "ap_mac", "device_name", "interface_name", "site", "match_source", "pvid", "vlan")
        result = [row for row in result if any(search_text in str(row.get(field) or "").casefold() for field in fields)]
    return result


def sort_trackside_ap_business_rows(rows: list[dict[str, object | None]]) -> list[dict[str, object | None]]:
    """Keep page and export rows in the same switch/interface order."""

    return sorted(
        (dict(row) for row in rows or []),
        key=_trackside_ap_business_sort_key,
    )


def _trackside_ap_business_sort_key(row: dict[str, object | None]) -> tuple[object, ...]:
    return (
        _trackside_sort_text_key(
            row.get("device_name")
            or row.get("switch_name")
            or row.get("indoor_switch")
            or row.get("室内交换机")
        ),
        interface_sort_key(row.get("interface_name") or row.get("接口名称")),
        _trackside_sort_text_key(row.get("site") or row.get("station") or row.get("归属站点")),
        _trackside_sort_text_key(row.get("ap_name") or row.get("AP名称")),
        _trackside_sort_text_key(row.get("ap_mac") or row.get("AP MAC")),
    )


def _trackside_sort_text_key(value: object) -> tuple[object, ...]:
    text = str(value or "").strip()
    if not text or text == "-":
        return (1, ())
    return (0, natural_text_key(text), text.casefold())


def trackside_station_options(rows: list[dict[str, object | None]]) -> list[str]:
    return sorted(
        {site for site in (str(row.get("site") or "").strip() for row in rows) if site},
        key=natural_text_key,
    )


def build_trackside_site_filter_items(rows: list[dict[str, object | None]], all_label: str) -> list[tuple[str, str]]:
    sites = trackside_station_options(rows)
    return [(all_label, ""), *[(site, site) for site in sites]]


def _emit_trackside_export_progress(progress_callback, stage: str, current: int = 0, total: int = 0, message: str = "") -> None:
    if not progress_callback:
        return
    try:
        progress_callback(stage, int(current or 0), int(total or 0), message or stage)
    except TypeError:
        progress_callback(stage)


def _raise_if_trackside_export_cancelled(should_cancel) -> None:
    if should_cancel and should_cancel():
        raise TracksideApExportCancelled("导出已取消")


_SHANGHAI_TIMEZONE = timezone(timedelta(hours=8), name="Asia/Shanghai")


def _shanghai_datetime(value: str | datetime | None) -> datetime:
    if isinstance(value, datetime):
        parsed = value
    else:
        text = str(value or "").strip().replace("Z", "+00:00")
        try:
            parsed = datetime.fromisoformat(text) if text else datetime.now(_SHANGHAI_TIMEZONE)
        except ValueError:
            parsed = datetime.now(_SHANGHAI_TIMEZONE)
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=_SHANGHAI_TIMEZONE)
    return parsed.astimezone(_SHANGHAI_TIMEZONE)


def build_ap_online_history_block(
    rows: list[dict[str, object | None]],
    columns: tuple[tuple[str, str], ...],
    headers: list[str],
    *,
    snapshot_generated_at: str | datetime | None,
    updated_at: str | datetime | None,
) -> ApOnlineHistoryBlockDTO:
    snapshot_time = _shanghai_datetime(snapshot_generated_at)
    update_time = _shanghai_datetime(updated_at)
    values = tuple(
        tuple(_ap_online_history_value(field, row.get(field)) for _key, field in columns)
        for row in rows
    )
    return ApOnlineHistoryBlockDTO(
        snapshot_date=snapshot_time.strftime("%Y-%m-%d"),
        updated_at=update_time.strftime("%Y-%m-%d %H:%M:%S"),
        headers=tuple(headers),
        rows=values,
    )


def _ap_online_history_value(field: str, value: object | None) -> object:
    if value in (None, ""):
        return "-"
    if field != "online_rate":
        return str(value)
    if isinstance(value, str):
        text = value.strip()
        is_percentage = text.endswith("%")
        if is_percentage:
            text = text[:-1].strip()
        try:
            number = float(text)
        except ValueError:
            return value
        return number / 100.0 if is_percentage or number > 1 else number
    try:
        number = float(value)
    except (TypeError, ValueError):
        return value
    return number / 100.0 if number > 1 else number


def _apply_trackside_workbook_registry(workbook) -> None:
    for target_index, definition in enumerate(TRACKSIDE_AP_BUSINESS_SHEET_DEFINITIONS):
        if definition.sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[definition.sheet_name]
        current_index = workbook.worksheets.index(sheet)
        workbook.move_sheet(sheet, offset=target_index - current_index)
        if definition.tab_color:
            sheet.sheet_properties.tabColor = _opaque_argb(definition.tab_color)


def _opaque_argb(value: object) -> str:
    color = str(value or "").strip().lstrip("#").upper()
    if re.fullmatch(r"[0-9A-F]{6}", color):
        return f"FF{color}"
    if re.fullmatch(r"[0-9A-F]{8}", color) and color.startswith("FF"):
        return color
    raise ValueError(f"invalid opaque RGB color: {value!r}")


def export_trackside_ap_business_xlsx(
    path: Path,
    rows: list[dict[str, object | None]],
    columns: tuple[tuple[str, str], ...],
    headers: list[str],
    ap_online_overview_rows: list[dict[str, object | None]] | None = None,
    ap_online_overview_columns: tuple[tuple[str, str], ...] | None = None,
    ap_online_overview_headers: list[str] | None = None,
    new_online_ap_rows: list[dict[str, object | None]] | None = None,
    new_online_ap_columns: tuple[tuple[str, str], ...] | None = None,
    new_online_ap_headers: list[str] | None = None,
    new_online_ap_sheet_title: str = "\u65b0\u589e\u4e0a\u7ebfAP\u6982\u89c8",
    ap_optical_treatment_rows: list[dict[str, object | None]] | None = None,
    ap_optical_treatment_columns: tuple[tuple[str, str], ...] | None = None,
    ap_optical_treatment_headers: list[str] | None = None,
    ap_optical_treatment_sheet_title: str = "AP\u5149\u8870\u5904\u7406\u8bb0\u5f55",
    offline_ap_stats: dict[str, object | None] | None = None,
    offline_ap_ledger_rows: list[dict[str, object | None]] | None = None,
    offline_ap_stats_headers: list[str] | None = None,
    offline_ap_ledger_headers: list[str] | None = None,
    unmatched_online_rows: list[dict[str, object | None]] | None = None,
    unmatched_online_columns: tuple[tuple[str, str], ...] | None = None,
    unmatched_online_headers: list[str] | None = None,
    unmatched_online_sheet_title: str = "待关联在线AP",
    progress_callback=None,
    should_cancel=None,
    current_optical_abnormal_headers: list[str] | None = None,
    snapshot_generated_at: str | datetime | None = None,
    export_updated_at: str | datetime | None = None,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    from netconsole.services.excel_autosize import apply_worksheet_autofit

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "\u8f68\u65c1AP\u4e1a\u52a1"
    rows = sort_trackside_ap_business_rows(rows)
    def log_write_phase(phase: str, start: float, **values: object) -> None:
        elapsed_ms = int((perf_counter() - start) * 1000)
        details = " ".join(f"{key}={value}" for key, value in values.items())
        app_logger.log_info("TRACKSIDE_EXPORT_PROFILE", f"phase={phase} elapsed_ms={elapsed_ms}" + (f" {details}" if details else ""))

    total_rows = len(rows)
    _emit_trackside_export_progress(progress_callback, "write_trackside_rows", 0, total_rows, f"正在写入轨旁AP业务明细 0/{total_rows}")
    _raise_if_trackside_export_cancelled(should_cancel)
    phase_start = perf_counter()
    sheet.append(headers)
    fills = {
        status: PatternFill(fill_type="solid", fgColor=_opaque_argb(color))
        for status, color in TRACKSIDE_OPTICAL_COLOR_RGB.items()
        if color
    }
    header_fill = PatternFill(
        fill_type="solid",
        fgColor=_opaque_argb(TRACKSIDE_EXPORT_HEADER_FILL),
    )
    for row_index, row in enumerate(rows, start=1):
        sheet.append([_export_value(field, row) for _key, field in columns])
        fill = fills.get(trackside_export_fill_status(row))
        for cell in sheet[sheet.max_row]:
            if fill:
                cell.fill = fill
        if row_index == total_rows or row_index % 100 == 0:
            _emit_trackside_export_progress(
                progress_callback,
                "write_trackside_rows",
                row_index,
                total_rows,
                f"正在写入轨旁AP业务明细 {row_index}/{total_rows}",
            )
            _raise_if_trackside_export_cancelled(should_cancel)
    alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin", color=_opaque_argb("D1D5DB")),
        right=Side(style="thin", color=_opaque_argb("D1D5DB")),
        top=Side(style="thin", color=_opaque_argb("D1D5DB")),
        bottom=Side(style="thin", color=_opaque_argb("D1D5DB")),
    )
    header_font = Font(bold=True)
    _format_export_sheet(sheet, alignment, border, header_font, header_fill)
    sheet.auto_filter.ref = sheet.dimensions
    log_write_phase("write_trackside_sheet", phase_start, rows=len(rows))
    _raise_if_trackside_export_cancelled(should_cancel)
    phase_start = perf_counter()
    _emit_trackside_export_progress(progress_callback, "write_current_optical_abnormal", 0, 0, "正在写入异常光衰")
    build_current_optical_abnormal_sheet(
        workbook,
        sheet,
        rows,
        source_columns=columns,
        headers=current_optical_abnormal_headers,
    )
    log_write_phase("write_current_optical_abnormal_sheet", phase_start, rows=sum(1 for row in rows if is_current_optical_abnormal_export_row(row)))
    _raise_if_trackside_export_cancelled(should_cancel)
    _emit_trackside_export_progress(progress_callback, "write_ap_online_overview", 0, len(ap_online_overview_rows or []), "正在写入AP上线情况")
    phase_start = perf_counter()
    _append_ap_overview_sheet(
        workbook,
        rows,
        alignment,
        border,
        header_font,
        ap_online_overview_rows,
        ap_online_overview_columns,
        ap_online_overview_headers,
        header_fill,
        snapshot_generated_at=snapshot_generated_at,
        updated_at=export_updated_at,
    )
    log_write_phase("write_ap_online_overview_sheet", phase_start, rows=len(ap_online_overview_rows or []))
    _raise_if_trackside_export_cancelled(should_cancel)
    _emit_trackside_export_progress(progress_callback, "write_new_online_ap", 0, len(new_online_ap_rows or []), "正在写入新增上线AP概览")
    phase_start = perf_counter()
    _append_export_rows_sheet(
        workbook,
        new_online_ap_sheet_title,
        new_online_ap_rows or [],
        new_online_ap_columns or NEW_ONLINE_AP_OVERVIEW_COLUMNS,
        new_online_ap_headers or [key for key, _field in NEW_ONLINE_AP_OVERVIEW_COLUMNS],
        alignment,
        border,
        header_font,
        header_fill,
    )
    log_write_phase("write_new_online_ap_sheet", phase_start, rows=len(new_online_ap_rows or []))
    _raise_if_trackside_export_cancelled(should_cancel)
    _emit_trackside_export_progress(progress_callback, "write_unmatched_online_ap", 0, len(unmatched_online_rows or []), "正在写入待关联在线AP")
    phase_start = perf_counter()
    _append_export_rows_sheet(
        workbook,
        unmatched_online_sheet_title,
        unmatched_online_rows or [],
        unmatched_online_columns or TRACKSIDE_AP_UNMATCHED_ONLINE_COLUMNS,
        unmatched_online_headers or [key for key, _field in (unmatched_online_columns or TRACKSIDE_AP_UNMATCHED_ONLINE_COLUMNS)],
        alignment,
        border,
        header_font,
        header_fill,
    )
    log_write_phase("write_unmatched_online_ap_sheet", phase_start, rows=len(unmatched_online_rows or []))
    _raise_if_trackside_export_cancelled(should_cancel)
    _emit_trackside_export_progress(progress_callback, "write_optical_treatment", 0, len(ap_optical_treatment_rows or []), "正在写入AP光衰处理记录")
    phase_start = perf_counter()
    sorted_treatment_rows = _sort_ap_optical_treatment_rows(ap_optical_treatment_rows or [])
    _append_export_rows_sheet(
        workbook,
        ap_optical_treatment_sheet_title,
        sorted_treatment_rows,
        ap_optical_treatment_columns or AP_OPTICAL_TREATMENT_RECORD_COLUMNS,
        ap_optical_treatment_headers or [key for key, _field in AP_OPTICAL_TREATMENT_RECORD_COLUMNS],
        alignment,
        border,
        header_font,
        header_fill,
        fills,
        _ap_optical_treatment_row_fill_status,
        preserve_ap_identity=True,
    )
    log_write_phase("write_optical_treatment_sheet", phase_start, rows=len(sorted_treatment_rows))
    _raise_if_trackside_export_cancelled(should_cancel)
    if offline_ap_stats is not None and offline_ap_ledger_rows is not None:
        _emit_trackside_export_progress(progress_callback, "write_offline_ap_ledger", 0, len(offline_ap_ledger_rows or []), "正在写入离线AP台账")
        phase_start = perf_counter()
        stats_sheet = workbook.create_sheet("AP\u79bb\u7ebf\u60c5\u51b5")
        write_offline_ap_stats_sheet(stats_sheet, offline_ap_stats, offline_ap_stats_headers or [key for key, _field in OFFLINE_AP_STATS_COLUMNS])
        ledger_sheet = workbook.create_sheet("AP\u79bb\u7ebf\u53f0\u8d26")
        display_ledger_rows = [
            {
                **row,
                "historical_switch_interface": display_interface_name(
                    row.get("historical_switch_interface")
                ),
            }
            for row in offline_ap_ledger_rows
        ]
        write_offline_ap_ledger_sheet(
            ledger_sheet,
            display_ledger_rows,
            offline_ap_ledger_headers
            or [key for key, _field in OFFLINE_AP_LEDGER_COLUMNS],
        )
        log_write_phase("write_offline_ap_sheets", phase_start, rows=len(offline_ap_ledger_rows or []))
    _raise_if_trackside_export_cancelled(should_cancel)
    _emit_trackside_export_progress(progress_callback, "write_summary", 0, len(rows), "正在写入汇总统计")
    phase_start = perf_counter()
    _append_switch_optical_summary_sheet(workbook, rows, alignment, border, header_font, header_fill)
    log_write_phase("write_switch_optical_summary_sheet", phase_start, rows=len(rows))
    _raise_if_trackside_export_cancelled(should_cancel)
    _emit_trackside_export_progress(progress_callback, "style_autofit", 0, len(workbook.worksheets), "正在设置样式和列宽")
    phase_start = perf_counter()
    _apply_trackside_workbook_registry(workbook)
    for worksheet in workbook.worksheets:
        definition = trackside_ap_business_sheet_definition(worksheet.title)
        header_row = 3 if definition and definition.stable_key == "ap_online_history_overview" else 1
        _format_export_sheet(
            worksheet,
            alignment,
            border,
            header_font,
            header_fill,
            header_row=header_row,
        )
        if definition and definition.stable_key == "ap_online_history_overview":
            worksheet.auto_filter.ref = None
        elif header_row == 1:
            worksheet.auto_filter.ref = worksheet.dimensions
        apply_worksheet_autofit(worksheet, maximum=60)
    _set_switch_optical_summary_widths(workbook)
    log_write_phase("autofit_sheets", phase_start, sheets=len(workbook.worksheets))
    _raise_if_trackside_export_cancelled(should_cancel)
    _emit_trackside_export_progress(progress_callback, "save_workbook", total_rows, total_rows, "正在保存Excel文件")
    phase_start = perf_counter()
    workbook.save(path)
    log_write_phase("save_workbook", phase_start, path=Path(path).name)


# Legacy aliases removed — status is now computed real-time from raw data.


def _merge_resource_with_optical(resource: dict[str, object | None] | None, optical_by_mac: dict[str, dict[str, object | None]]) -> dict[str, object | None]:
    if not resource:
        return {}
    optical = optical_by_mac.get(normalize_mac_key(resource.get("ap_mac")), {})
    merged = {**resource, **optical}
    # FIT-AP runtime state is authoritative; optical payloads may carry stale
    # or unrelated status fields but must never replace state evidence.
    for field in ("state", "state_raw", "state_display", "is_online"):
        if field in resource and resource.get(field) not in (None, ""):
            merged[field] = resource[field]
    return merged


def _match_fit_ap_resource_from_lldp(
    lldp: dict[str, object | None],
    resources_by_ip: dict[str, list[dict[str, object | None]]],
    resources_by_mac: dict[str, list[dict[str, object | None]]],
    resources_by_name: dict[str, list[dict[str, object | None]]],
) -> tuple[dict[str, object | None] | None, str, int, str]:
    del resources_by_ip, resources_by_name
    if not lldp:
        return None, "", 0, ""
    observed_mac = normalize_mac_key(lldp.get("neighbor_mac") or lldp.get("chassis_id"))
    if observed_mac:
        matches = resources_by_mac.get(observed_mac, [])
        if len(matches) == 1:
            return matches[0], "LLDP_MAC", 100, "MATCHED"
        if len(matches) > 1:
            return None, "LLDP_MAC", 0, "AMBIGUOUS"
        return None, "LLDP_MAC", 0, "UNRESOLVED"
    return None, "", 0, "NO_MAC_EVIDENCE"


def _build_bidirectional_attenuation(
    *,
    local_rx_power: object,
    local_tx_power: object,
    remote_rx_power: object,
    remote_tx_power: object,
    local_status: object,
    remote_status: object,
    association_reliable: bool,
    remote_identity_known: bool,
    local_sample_time: object,
    remote_sample_time: object,
    calculation_enabled: bool = True,
) -> dict[str, object | None]:
    local_rx = _reasonable_power(local_rx_power)
    local_tx = _reasonable_power(local_tx_power)
    remote_rx = _reasonable_power(remote_rx_power)
    remote_tx = _reasonable_power(remote_tx_power)
    result: dict[str, object | None] = {
        "local_rx_power_dbm": local_rx,
        "local_tx_power_dbm": local_tx,
        "remote_rx_power_dbm": remote_rx,
        "remote_tx_power_dbm": remote_tx,
        "forward_loss_db": None,
        "reverse_loss_db": None,
        "calculation_status": "",
        "calculation_reason": "",
        "local_sample_time": str(local_sample_time or ""),
        "remote_sample_time": str(remote_sample_time or ""),
        "sample_time_delta_seconds": None,
    }
    if not calculation_enabled:
        result.update(
            calculation_status="NOT_VERIFIED",
            calculation_reason="REAL_DEVICE_SAMPLE_REQUIRED",
        )
        return result
    unavailable_statuses = {
        "offline",
        "no_module",
        "dom_unavailable",
        "no_light",
        "link_down",
    }
    if str(local_status or "").strip().casefold() in unavailable_statuses or str(
        remote_status or ""
    ).strip().casefold() in unavailable_statuses:
        result.update(
            calculation_status="MODULE_OFFLINE",
            calculation_reason="至少一端光模块离线或无有效 DOM",
        )
        return result
    if remote_rx is None or remote_tx is None:
        status = "REMOTE_DOM_UNAVAILABLE" if remote_identity_known else "SINGLE_ENDED_ONLY"
        reason = (
            "已识别对端，但对端 DOM 数据不完整"
            if remote_identity_known
            else "仅有本端光功率，无法计算双向光衰"
        )
        result.update(calculation_status=status, calculation_reason=reason)
        return result
    if local_rx is None or local_tx is None:
        result.update(
            calculation_status="SINGLE_ENDED_ONLY",
            calculation_reason="本端 DOM 数据不完整，无法计算双向光衰",
        )
        return result
    if not association_reliable:
        result.update(
            calculation_status="NEIGHBOR_UNCERTAIN",
            calculation_reason="对端关系缺少可靠 LLDP 或人工绑定证据",
        )
        return result
    local_time = _parse_sample_time(local_sample_time)
    remote_time = _parse_sample_time(remote_sample_time)
    if local_time is None or remote_time is None:
        result.update(
            calculation_status="STALE_SAMPLE",
            calculation_reason="两端采集时间不完整",
        )
        return result
    delta = int(abs((local_time - remote_time).total_seconds()))
    result["sample_time_delta_seconds"] = delta
    if delta > int(TRACKSIDE_ATTENUATION_SAMPLE_WINDOW.total_seconds()):
        result.update(
            calculation_status="STALE_SAMPLE",
            calculation_reason="两端采集时间超出 30 分钟允许窗口",
        )
        return result
    result.update(
        calculation_status="CALCULATED",
        calculation_reason="两端 DOM、端口映射和采集时间均有效",
        forward_loss_db=round(local_tx - remote_rx, 2),
        reverse_loss_db=round(remote_tx - local_rx, 2),
    )
    return result


def _reasonable_power(value: object) -> float | None:
    parsed = _float_value(value)
    return parsed if parsed is not None and -50.0 <= parsed <= 20.0 else None


def _parse_sample_time(value: object) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        return None
    return parsed if parsed.tzinfo is not None else parsed.replace(tzinfo=timezone.utc)


def _normalize_ip(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    try:
        return str(ip_address(text))
    except ValueError:
        return ""


def _offline_ledger_to_trackside_rows(
    rows: list[dict[str, object | None]],
    interfaces_by_device: dict[str, list[dict[str, object | None]]] | None = None,
    optical_by_device: dict[str, list[dict[str, object | None]]] | None = None,
    latest_switch_collect_runs: Mapping[str, str] | None = None,
) -> list[dict[str, object | None]]:
    result: list[dict[str, object | None]] = []
    interface_indexes = {device_uuid: _latest_rows_by_normalized_interface(items, "interface_name") for device_uuid, items in (interfaces_by_device or {}).items()}
    optical_indexes = {device_uuid: _latest_rows_by_normalized_interface(items, "interface_name") for device_uuid, items in (optical_by_device or {}).items()}
    for row in rows:
        device_uuid = str(row.get("device_uuid") or "")
        interface_key = normalize_interface_name(row.get("historical_switch_interface")).casefold()
        interface = interface_indexes.get(device_uuid, {}).get(interface_key, {})
        latest_switch_collect_run = str(
            (latest_switch_collect_runs or {}).get(device_uuid) or ""
        )
        interface_data_status = _snapshot_data_status(
            interface,
            latest_switch_collect_run,
        )
        current_interface = interface if interface_data_status == "current" else {}
        stored_optical = optical_indexes.get(device_uuid, {}).get(interface_key, {})
        optical_data_status = _snapshot_data_status(
            stored_optical,
            latest_switch_collect_run,
        )
        optical = stored_optical if optical_data_status == "current" else {}
        link_state = normalize_link_state(
            current_interface.get("link_status") or current_interface.get("link")
        )
        switch_collection_status = (
            row.get("switch_collection_status")
            or current_interface.get("switch_collection_status")
            or current_interface.get("collection_status")
        )
        switch_result = compute_optical_severity(
            {
                "module_present": bool(_has_optical_module_data(optical)),
                "no_module": _explicit_no_module(optical),
                "switch_rx_power": optical.get("rx_power"),
                "switch_port_status": optical.get("port_status"),
                "alarm_low": optical.get("rx_low_alarm"),
                "alarm_high": optical.get("rx_high_alarm"),
                "warning_low": optical.get("rx_low_warning"),
                "device_type": "switch",
            }
        )
        switch_status = switch_result.severity
        result.append(
            {
                "site": row.get("site"),
                "ac_device_uuid": row.get("ac_device_uuid"),
                "ap_uuid": row.get("ap_uuid"),
                "device_uuid": device_uuid or row.get("device_uuid"),
                "device_name": row.get("historical_switch_name"),
                "interface_name": row.get("historical_switch_interface"),
                "link_status": link_state,
                "protocol_status": interface.get("protocol_status"),
                "description": interface.get("description") or row.get("offline_remark"),
                "port_type": _port_type(interface.get("port_status")),
                "port_status": interface.get("port_status"),
                "pvid": interface.get("pvid"),
                "match_source": "historical",
                "vlan": interface.get("vlan"),
                "switch_rx_power": optical.get("rx_power"),
                "switch_tx_power": optical.get("tx_power"),
                "switch_optical_status": switch_status,
                "switch_interface_updated_at": interface.get("updated_at") or interface.get("collected_at"),
                "switch_optical_updated_at": stored_optical.get("updated_at") or stored_optical.get("collected_at"),
                "switch_interface_data_status": interface_data_status,
                "switch_optical_data_status": optical_data_status,
                "ap_mac": row.get("ap_mac"),
                "ap_name": row.get("ap_name"),
                "ap_ip": row.get("ap_ip"),
                "ap_state": "Idle",
                "ap_state_display": row.get("ap_status"),
                "ap_rx_power": None,
                "ap_tx_power": None,
                "ap_optical_status": "offline",
                "ap_side_has_data": True,
                "updated_at": row.get("last_lldp_at"),
                "source_device": row.get("historical_switch_name"),
                "collection_status": "historical",
                "is_ap_offline": True,
                "offline_reason": row.get("offline_reason") or "ac_idle",
                "status_reason": row.get("status_reason") or "AC FIT-AP状态为Idle，轨旁AP离线",
                "data_source": row.get("data_source") or "historical",
                "switch_collection_status": switch_collection_status,
                "offline_remark": row.get("offline_remark"),
            }
        )
    return result


def summarize_trackside_ap_online_counts(rows: list[dict[str, object | None]]) -> tuple[int, int]:
    online = 0
    offline = 0
    seen: set[str] = set()
    for row in rows:
        key = str(
            row.get("ap_uuid")
            or normalize_mac_key(row.get("ap_mac"))
            or row.get("serial_number")
            or row.get("ap_name")
            or ""
        ).strip()
        if not key:
            continue
        if key and key in seen:
            continue
        if key:
            seen.add(key)
        state = _ap_state(row)
        if state == "online":
            online += 1
        elif state == "offline":
            offline += 1
    return online, offline


def _normalize_name(value: object) -> str:
    return str(value or "").strip().casefold()


def _port_type(value: object) -> str:
    text = str(value or "").strip().casefold()
    return text if text in {"access", "trunk", "hybrid"} else "unknown"


def normalize_link_state(value: object) -> str:
    text = str(value or "").strip().upper()
    if not text:
        return "-"
    if text in {"UP", "DOWN"}:
        return text
    if "DOWN" in text:
        return "DOWN"
    if "UP" in text:
        return "UP"
    return "-"


def _snapshot_data_status(
    row: Mapping[str, object | None],
    latest_collect_run: str,
) -> str:
    if not row:
        return "missing"
    expected = str(latest_collect_run or "").strip()
    if not expected:
        return "current"
    actual = str(row.get("collect_run_uuid") or "").strip()
    return "current" if actual == expected else "stale"


def _switch_collection_status(device: Device, interface: dict[str, object | None], optical: dict[str, object | None]) -> str:
    for source in (interface, optical):
        for field in ("switch_collection_status", "collection_status", "collect_status"):
            value = str(source.get(field) or "").strip()
            if value:
                return value
    for field in ("switch_collection_status", "collection_status", "collect_status"):
        value = str(getattr(device, field, "") or "").strip()
        if value:
            return value
    return "success" if optical else "not_collected"


def _is_switch_collection_offline(value: object) -> bool:
    text = str(value or "").strip().casefold()
    if not text:
        return False
    return text in {"offline", "switch_offline", "device_offline", "link_down", "离线", "交换机离线"} or "真实离线" in text


def enrich_trackside_export_rows(
    rows: list[dict[str, object | None]],
    *_retired_args: object,
    **_retired_kwargs: object,
) -> list[dict[str, object | None]]:
    """Copy current rows without consulting optical history or deriving changes."""

    return [dict(row) for row in rows]


def _optical_status_from_history(row: dict[str, object | None], device_type: str) -> str:
    if device_type == "ap":
        return evaluate_ap_business_rx(row.get("rx_power"))
    status = row.get("alarm_status") or row.get("optical_alarm_status") or row.get("status")
    normalized_status = str(status or "").strip().casefold()
    if normalized_status in {"normal", "notice", "warning", "alarm", "link_abnormal", "link_down", "no_light", "no_module", "not_collected", "skipped", "offline"}:
        return str(status or "")
    if normalized_status == "unknown" and not _has_valid_rx_power(row.get("rx_power")):
        return str(status or "")
    result = compute_optical_severity(
        {
            "module_present": bool(_has_optical_module_data(row)),
            "rx_power": row.get("rx_power"),
            "alarm_low": row.get("rx_low_alarm"),
            "warning_low": row.get("rx_low_warning"),
            "port_status": row.get("port_status"),
            "device_type": device_type,
        }
    )
    return result.severity


def _previous_lldp_switch(row: dict[str, object | None]) -> object:
    return row.get("neighbor_switch_name") or row.get("neighbor_switch_sysname") or row.get("neighbor_device_name") or row.get("lldp_neighbor")


def _previous_lldp_interface(row: dict[str, object | None]) -> object:
    return row.get("neighbor_interface")


def _is_ac_idle(row: dict[str, object | None]) -> bool:
    for field in ("state", "state_raw", "ap_state", "state_display", "ap_state_display"):
        value = row.get(field)
        if str(value or "").strip():
            return normalize_fit_ap_state_token(value) in {"I", "IDLE"}
    return False


def _is_ac_offline(row: dict[str, object | None]) -> bool:
    return _ap_state(row) == "offline"


def _ac_offline_status_reason(row: dict[str, object | None]) -> str:
    state = next(
        (
            str(row.get(field) or "").strip()
            for field in ("state_display", "state", "state_raw", "ap_state_display", "ap_state")
            if str(row.get(field) or "").strip()
        ),
        "非运行态",
    )
    return f"AC FIT-AP状态为{state}，轨旁AP离线"


def _trackside_merge_key(row: dict[str, object | None]) -> tuple[str, str, str]:
    site = str(row.get("site") or "").strip().casefold()
    switch = str(row.get("device_uuid") or row.get("switch_uuid") or row.get("device_name") or row.get("switch_name") or "").strip().casefold()
    interface = normalize_interface_name(row.get("interface_name")).casefold()
    return site, switch, interface


def _merge_duplicate_trackside_rows(rows: list[dict[str, object | None]]) -> list[dict[str, object | None]]:
    merged: dict[tuple[str, str, str], dict[str, object | None]] = {}
    passthrough: list[dict[str, object | None]] = []
    for row in rows:
        key = _trackside_merge_key(row)
        if not all(key):
            passthrough.append(row)
            continue
        existing = merged.get(key)
        if existing is None:
            merged[key] = dict(row)
            continue
        _merge_trackside_row(existing, row)
    result = [*merged.values(), *passthrough]
    for row in result:
        _apply_trackside_offline_priority(row)
    return result


def _merge_trackside_row(target: dict[str, object | None], source: dict[str, object | None]) -> None:
    prefer_source_fields = (
        "ap_uuid",
        "ap_mac",
        "ap_name",
        "ap_ip",
        "ap_state",
        "ap_state_display",
        "ap_rx_power",
        "ap_tx_power",
        "ap_rx_low_alarm",
        "ap_rx_low_warning",
        "offline_remark",
    )
    for field in prefer_source_fields:
        if _is_missing_display(target.get(field)) and not _is_missing_display(source.get(field)):
            target[field] = source.get(field)
    for field, value in source.items():
        if field not in target or _is_missing_display(target.get(field)):
            target[field] = value
    if bool(source.get("ap_side_has_data")):
        target["ap_side_has_data"] = True
    if bool(source.get("is_ap_offline")):
        target["is_ap_offline"] = True
    if source.get("offline_reason") == "switch_offline" or target.get("offline_reason") != "switch_offline":
        for field in ("offline_reason", "status_reason", "data_source", "switch_collection_status"):
            if source.get(field):
                target[field] = source.get(field)
    if normalize_link_state(source.get("link_status")) == "DOWN":
        target["link_status"] = "DOWN"
    target["port_type"] = _port_type(target.get("port_type") or source.get("port_type") or target.get("port_status") or source.get("port_status"))
    source_ac_device_uuids = sorted(
        {
            *(
                str(value).strip()
                for value in target.get("source_ac_device_uuids", []) or []
                if str(value).strip()
            ),
            *(
                str(value).strip()
                for value in source.get("source_ac_device_uuids", []) or []
                if str(value).strip()
            ),
            *(
                value
                for value in (
                    str(target.get("ac_device_uuid") or "").strip(),
                    str(source.get("ac_device_uuid") or "").strip(),
                )
                if value
            ),
        }
    )
    if source_ac_device_uuids:
        target["source_ac_device_uuids"] = source_ac_device_uuids


def _apply_trackside_offline_priority(row: dict[str, object | None]) -> None:
    switch_offline = row.get("offline_reason") == "switch_offline" or _is_switch_collection_offline(row.get("switch_collection_status"))
    ac_offline = _is_ac_offline(row) or row.get("offline_reason") in {"ac_idle", "ac_offline"}
    if switch_offline:
        row["link_status"] = "DOWN"
        row["switch_optical_status"] = "offline"
        row["ap_optical_status"] = "offline"
        row["ap_side_has_data"] = True
        row["is_ap_offline"] = True
        row["offline_reason"] = "switch_offline"
        row["status_reason"] = "室内交换机离线，轨旁AP跟随离线"
        row["data_source"] = row.get("data_source") or "mixed"
    elif ac_offline:
        row["ap_optical_status"] = "offline"
        row["ap_side_has_data"] = True
        row["is_ap_offline"] = True
        row["offline_reason"] = row.get("offline_reason") or (
            "ac_idle" if _is_ac_idle(row) else "ac_offline"
        )
        row["status_reason"] = row.get("status_reason") or _ac_offline_status_reason(row)
    else:
        _ensure_ap_optical_status(row)
    row["port_type"] = _port_type(row.get("port_type") or row.get("port_status"))


def _export_value(field: str, row: dict[str, object | None], *, preserve_ap_identity: bool = False) -> str:
    if field == "completed_at" and row.get("treatment_status") == TREATMENT_OPEN_LABEL:
        return ""
    if preserve_ap_identity and field in {"ap_name", "ap_mac", "serial_number"}:
        value = row.get(field)
        return str(value) if value not in (None, "") else AP_SIDE_MISSING_DISPLAY
    if field == "interface_name":
        return display_interface_name(row.get(field)) or AP_SIDE_MISSING_DISPLAY
    return format_trackside_display_value(field, row)


def _has_optical_module_data(row: dict[str, object | None]) -> bool:
    if not row:
        return False
    module_fields = (
        "rx_power",
        "tx_power",
        "module_model",
        "module_serial_number",
        "module_vendor",
        "wavelength",
        "transmission_distance",
        "connector_type",
        "rx_low_alarm",
        "rx_low_warning",
    )
    return any(row.get(field) not in (None, "") for field in module_fields)


def _has_ap_side_optical_data(fit_ap: dict[str, object | None], candidate: dict[str, object | None]) -> bool:
    if not fit_ap:
        return False
    if _explicit_no_module(fit_ap):
        return True
    if _is_missing_display(candidate.get("ap_mac")) or _is_missing_display(candidate.get("ap_name")) or _is_missing_display(candidate.get("ap_rx_power")):
        return False
    return _has_optical_module_data(fit_ap)


def _explicit_no_module(row: dict[str, object | None]) -> bool:
    text = " ".join(
        str(row.get(field) or "")
        for field in (
            "optical_alarm_status",
            "status",
            "raw_status",
            "ap_raw_status",
            "error_message",
            "message",
        )
    ).strip().casefold()
    if not text:
        return False
    return any(token in text for token in ("no_module", "no module", "no transceiver", "no-transceiver", "\u65e0\u5149\u6a21\u5757"))


def _is_missing_display(value: object) -> bool:
    return str(value or "").strip() in {"", "-"}


def _format_export_sheet(
    sheet,
    alignment,
    border,
    header_font,
    header_fill=None,
    *,
    header_row: int = 1,
) -> None:
    from openpyxl.styles import Alignment, Border

    # The history sheet is intentionally scrollable; other business sheets keep
    # only the header row visible for large AP datasets.
    is_history_sheet = sheet.title == "AP上线情况概览" or header_row > 1
    sheet.freeze_panes = None if is_history_sheet else "A2"
    long_text_columns = {
        cell.column
        for cell in sheet[header_row]
        if any(
            token in str(cell.value or "").strip().casefold()
            for token in _TRACKSIDE_LONG_TEXT_HEADER_TOKENS
        )
    }
    for row in sheet.iter_rows():
        row_is_blank = all(cell.value in (None, "") for cell in row)
        sheet.row_dimensions[row[0].row].height = (
            TRACKSIDE_OVERVIEW_SEPARATOR_ROW_HEIGHT
            if is_history_sheet and row_is_blank
            else TRACKSIDE_EXPORT_ROW_HEIGHT
        )
        for cell in row:
            if cell.row == header_row:
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
            elif cell.column in long_text_columns:
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                    wrap_text=True,
                )
            else:
                cell.alignment = alignment
            # The trailing blank row separates history blocks and must remain
            # visually empty instead of receiving table borders.
            cell.border = Border() if is_history_sheet and row_is_blank else border
            if cell.row == header_row:
                cell.font = header_font
                if header_fill is not None:
                    cell.fill = header_fill


def is_no_light_optical_row(row: dict[str, object | None]) -> bool:
    status_values = [
        row.get("switch_optical_status"),
        row.get("ap_optical_status"),
        row.get("optical_alarm_status"),
        row.get("alarm_status"),
        row.get("current_status"),
        row.get("status"),
        row.get("raw_status"),
        row.get("ap_raw_status"),
        row.get("module_status"),
        row.get("transceiver_status"),
    ]
    status_text = " ".join(str(value or "") for value in status_values).strip().casefold()
    return any(
        token in status_text
        for token in (
            "no_light",
            "no light",
            "no-light",
            "no_module",
            "no module",
            "no-module",
            "no transceiver",
            "no-transceiver",
            "无光",
            "无光模块",
            "未插光模块",
            "光模块不存在",
        )
    )


def has_valid_ap_binding(row: dict[str, object | None]) -> bool:
    return not _is_missing_display(row.get("ap_mac")) or not _is_missing_display(row.get("ap_name"))


def _normalized_optical_status(value: object) -> str:
    text = str(value or "").strip().casefold()
    return {
        "正常": "normal",
        "偏低关注": "notice",
        "提示告警": "warning",
        "低告警": "warning",
        "高告警": "warning",
        "一般告警": "alarm",
        "严重告警": "alarm",
        "功率异常": "abnormal",
        "链路异常": "link_abnormal",
        "链路断开": "link_down",
        "无光": "no_light",
        "无光模块": "no_module",
        "未插光模块": "no_module",
        "离线": "offline",
        "交换机离线": "offline",
        "-": "",
    }.get(text, text)


def _is_ap_offline_abnormal(row: dict[str, object | None]) -> bool:
    if not has_valid_ap_binding(row):
        return False
    # AP 在线态只能来自 FIT-AP runtime state。光衰的 ``offline``/``no_light``
    # 是链路观测，不得把仍处于 R/M、Run、Up 的 AP 改判为离线。
    runtime_state = _ap_state(row)
    if runtime_state == "online":
        return False
    if runtime_state == "offline":
        return True
    return bool(row.get("is_ap_offline"))


def _is_ap_side_current_abnormal(row: dict[str, object | None]) -> bool:
    if not is_ap_optical_applicable(row.get("model") or row.get("ap_model")):
        return False
    if not has_valid_ap_binding(row):
        return False
    status = _current_optical_export_evaluation(row).ap_status
    return is_optical_health_abnormal(status)


def _dual_business_evaluation(row: dict[str, object | None]):
    ap_side_has_data = has_ap_side_optical_data(row)
    return evaluate_dual_rx_business_detail(
        row.get("ap_rx_power") if ap_side_has_data else None,
        row.get("switch_rx_power"),
        ap_reported_status=(
            row.get("ap_device_optical_status") or row.get("ap_optical_status")
        ) if ap_side_has_data else "",
        switch_reported_status=(
            row.get("switch_device_optical_status")
            or row.get("switch_optical_status")
        ),
        ap_data_freshness=(
            row.get("ap_optical_data_freshness") or row.get("data_freshness")
        ),
        switch_data_freshness=row.get("switch_optical_data_status"),
    )


def _current_optical_export_evaluation(row: dict[str, object | None]):
    """Evaluate the latest valid optical values without turning stale into unknown."""
    return evaluate_dual_rx_business_detail(
        row.get("ap_rx_power") if has_ap_side_optical_data(row) else None,
        row.get("switch_rx_power"),
        ap_reported_status=(
            row.get("ap_device_optical_status") or row.get("ap_optical_status")
        ) if has_ap_side_optical_data(row) else "",
        switch_reported_status=(
            row.get("switch_device_optical_status")
            or row.get("switch_optical_status")
        ),
        ap_data_freshness="",
        switch_data_freshness="",
    )


def _is_switch_side_current_abnormal(row: dict[str, object | None]) -> bool:
    if not is_ap_optical_applicable(row.get("model") or row.get("ap_model")):
        return False
    if not has_valid_ap_binding(row):
        return False
    return is_optical_health_abnormal(_current_optical_export_evaluation(row).switch_status)


def current_optical_abnormal_reason(row: dict[str, object | None]) -> dict[str, str]:
    ap_state = _ap_state(row)
    ap_online_status = "离线" if _is_ap_offline_abnormal(row) or ap_state == "offline" else "在线" if ap_state == "online" else "未知"
    evaluation = _current_optical_export_evaluation(row)
    if _is_ap_side_current_abnormal(row):
        return {
            "ap_online_status": ap_online_status,
            "judgement": "异常",
            "reason": "AP侧业务光衰异常",
            "side": "AP侧",
            "level": display_optical_status(evaluation.ap_status),
            "detail": evaluation.reason,
        }
    if _is_switch_side_current_abnormal(row):
        return {
            "ap_online_status": ap_online_status,
            "judgement": "异常",
            "reason": "交换机侧业务光衰异常",
            "side": "交换机侧",
            "level": display_optical_status(evaluation.switch_status),
            "detail": evaluation.reason,
        }
    return {"ap_online_status": ap_online_status, "judgement": "", "reason": "", "side": "", "level": "", "detail": ""}


def _current_optical_observed_at(row: dict[str, object | None]) -> object:
    """Use the observation time of the side that makes the export abnormal."""
    evaluation = _current_optical_export_evaluation(row)
    if is_optical_health_abnormal(evaluation.ap_status):
        return (
            row.get("ap_last_valid_collected_at")
            or row.get("ap_optical_updated_at")
            or row.get("updated_at")
        )
    if is_optical_health_abnormal(evaluation.switch_status):
        return (
            row.get("switch_last_valid_collected_at")
            or row.get("switch_optical_updated_at")
            or row.get("updated_at")
        )
    return row.get("updated_at")


def is_current_optical_abnormal_export_row(row: dict[str, object | None]) -> bool:
    if not is_ap_optical_applicable(row.get("model") or row.get("ap_model")):
        return False
    switch_statuses = _trackside_export_switch_statuses(row)
    if "no_module" in switch_statuses or _explicit_no_module(row):
        return False
    if "no_light" in switch_statuses and not has_trackside_export_ap_evidence(row):
        return False
    if _is_ap_side_current_abnormal(row):
        return True
    return _is_switch_side_current_abnormal(row)


def count_current_optical_abnormal_aps(rows: list[dict[str, object | None]]) -> int:
    """Count current optical alarms by bound AP identity, not by interface rows."""
    return len({key for row in rows if is_current_optical_abnormal_row(row) if (key := ap_identity_key(row)) is not None})


def count_current_optical_abnormal_by_site(
    rows: list[dict[str, object | None]],
) -> dict[str, int]:
    """Group the current-abnormal sheet population by actual AP station.

    The AP identity is part of the grouping key so one AP with both switch and
    AP side alarms is counted once.  ``site_key`` keeps same-named stations in
    separate site databases from being accidentally merged by callers.
    """

    grouped: dict[tuple[str, str], set[tuple[str, str]]] = {}
    for row in rows or []:
        if not is_current_optical_abnormal_row(row):
            continue
        identity = ap_identity_key(row)
        if identity is None:
            continue
        site_key = str(row.get("site_key") or row.get("site_id") or "").strip().casefold()
        station = str(
            row.get("station_name")
            or row.get("station")
            or row.get("site")
            or "未归属"
        ).strip() or "未归属"
        grouped.setdefault((site_key, station), set()).add(identity)
    result: dict[str, int] = {}
    for (_site_key, station), identities in grouped.items():
        result[station] = result.get(station, 0) + len(identities)
    return result



def is_current_optical_abnormal_row(row: dict[str, object | None]) -> bool:
    return is_current_optical_abnormal_export_row(row)


def build_current_optical_abnormal_sheet(
    workbook,
    source_sheet,
    rows: list[dict[str, object | None]],
    *,
    source_columns: tuple[tuple[str, str], ...] | None = None,
    headers: list[str] | None = None,
) -> None:
    """Build the independent current optical abnormal sheet contract."""

    from copy import copy
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter

    if CURRENT_OPTICAL_ABNORMAL_SHEET_TITLE in workbook.sheetnames:
        del workbook[CURRENT_OPTICAL_ABNORMAL_SHEET_TITLE]
    source_index = workbook.worksheets.index(source_sheet)
    sheet = workbook.create_sheet(CURRENT_OPTICAL_ABNORMAL_SHEET_TITLE, source_index + 1)
    if headers is None:
        from netconsole.core.i18n import TRANSLATIONS

        headers = [
            TRANSLATIONS["zh_CN"].get(key, key)
            for key, _field in CURRENT_OPTICAL_ABNORMAL_COLUMNS
        ]
    sheet.append(headers)
    source_field_columns = {
        field: index
        for index, (_key, field) in enumerate(source_columns or (), start=1)
    }
    for index, (key, field) in enumerate(CURRENT_OPTICAL_ABNORMAL_COLUMNS, start=1):
        source_column = source_field_columns.get(field)
        if source_column:
            width = source_sheet.column_dimensions[get_column_letter(source_column)].width
            if width is not None:
                sheet.column_dimensions[get_column_letter(index)].width = width
        if sheet.column_dimensions[get_column_letter(index)].width is None:
            sheet.column_dimensions[get_column_letter(index)].width = max(12, len(headers[index - 1]) + 4)
        _copy_cell_style(source_sheet.cell(row=1, column=1), sheet.cell(row=1, column=index))

    fills = {
        status: PatternFill(fill_type="solid", fgColor=_opaque_argb(color))
        for status, color in TRACKSIDE_OPTICAL_COLOR_RGB.items()
        if color
    }
    text_fields = {"ap_mac", "ap_name", "ap_online_status"}
    target_row = 2
    for source_row, data in enumerate(rows, start=2):
        if not is_current_optical_abnormal_export_row(data):
            continue
        values = dict(data)
        values.update(current_optical_abnormal_reason(data))
        values["updated_at"] = _current_optical_observed_at(data)
        sheet.append([_export_value(field, values) for _key, field in CURRENT_OPTICAL_ABNORMAL_COLUMNS])
        if source_sheet.row_dimensions[source_row].height is not None:
            sheet.row_dimensions[target_row].height = source_sheet.row_dimensions[source_row].height
        for index, (_key, field) in enumerate(CURRENT_OPTICAL_ABNORMAL_COLUMNS, start=1):
            if field in text_fields:
                sheet.cell(row=target_row, column=index).number_format = "@"
        fill = fills.get(trackside_export_fill_status(data))
        if fill is not None:
            for cell in sheet[target_row]:
                cell.fill = fill
        target_row += 1
    if target_row == 2:
        cell = sheet.cell(row=2, column=1, value=CURRENT_OPTICAL_ABNORMAL_EMPTY_TEXT)
        header_cell = source_sheet.cell(row=1, column=1)
        cell.font = copy(header_cell.font)
        cell.fill = copy(header_cell.fill)
        cell.border = copy(header_cell.border)
        cell.alignment = copy(header_cell.alignment)
        cell.number_format = header_cell.number_format
        cell.protection = copy(header_cell.protection)
        sheet.row_dimensions[2].height = 22
    sheet.freeze_panes = source_sheet.freeze_panes
    sheet.auto_filter.ref = sheet.dimensions


def _copy_cell_style(source_cell, target_cell) -> None:
    from copy import copy

    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.number_format = source_cell.number_format
    target_cell.protection = copy(source_cell.protection)


def _append_export_rows_sheet(
    workbook,
    title: str,
    rows: list[dict[str, object | None]],
    columns: tuple[tuple[str, str], ...],
    headers: list[str],
    alignment,
    border,
    header_font,
    header_fill=None,
    fills: dict[str, object] | None = None,
    row_fill_status_getter=None,
    preserve_ap_identity: bool = False,
) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    text_fields = {"ap_mac", "ap_name", "serial_number", "apid", "ap_ip"}
    for row in rows:
        sheet.append([_export_value(field, row, preserve_ap_identity=preserve_ap_identity) for _key, field in columns])
        for index, (_key, field) in enumerate(columns, start=1):
            if field in text_fields:
                sheet.cell(row=sheet.max_row, column=index).number_format = "@"
        status = row_fill_status_getter(row) if callable(row_fill_status_getter) else None
        fill = (fills or {}).get(status)
        if fill is not None:
            for cell in sheet[sheet.max_row]:
                cell.fill = fill
    _format_export_sheet(sheet, alignment, border, header_font, header_fill)
    sheet.auto_filter.ref = sheet.dimensions


def _sort_ap_optical_treatment_rows(rows: list[dict[str, object | None]]) -> list[dict[str, object | None]]:
    return sorted(
        [dict(row) for row in rows],
        key=_trackside_ap_business_sort_key,
    )


def _ap_optical_treatment_row_fill_status(row: dict[str, object | None]) -> str:
    if row.get("treatment_status") == TREATMENT_CLOSED_LABEL:
        return "normal"
    status_text = " ".join(str(row.get(field) or "") for field in ("current_status", "issue_type")).casefold()
    if "预警" in status_text or "notice" in status_text or "warning" in status_text:
        return "warning"
    if any(token in status_text for token in ("无光", "告警", "异常", "alarm", "abnormal", "no_light")):
        return "alarm"
    return "alarm" if row.get("treatment_status") == TREATMENT_OPEN_LABEL else "normal"


def _append_ap_overview_sheet(
    workbook,
    rows: list[dict[str, object | None]],
    alignment,
    border,
    header_font,
    overview_rows: list[dict[str, object | None]] | None = None,
    overview_columns: tuple[tuple[str, str], ...] | None = None,
    overview_headers: list[str] | None = None,
    header_fill=None,
    *,
    snapshot_generated_at: str | datetime | None = None,
    updated_at: str | datetime | None = None,
) -> None:
    sheet = workbook.create_sheet("AP上线情况概览")
    display_rows = overview_rows or []
    display_columns = overview_columns or AP_ONLINE_OVERVIEW_COLUMNS
    display_headers = overview_headers or [
        key for key, _field in AP_ONLINE_OVERVIEW_COLUMNS
    ]
    block = build_ap_online_history_block(
        display_rows,
        display_columns,
        display_headers,
        snapshot_generated_at=snapshot_generated_at,
        updated_at=updated_at,
    )
    for values in block.cells():
        sheet.append(values)
    online_rate_column = next(
        (
            column_index
            for column_index, (_key, field) in enumerate(display_columns, start=1)
            if field == "online_rate"
        ),
        None,
    )
    if online_rate_column is not None:
        for row_index in range(4, 4 + len(display_rows)):
            sheet.cell(row=row_index, column=online_rate_column).number_format = "0.0%"
    from openpyxl.styles import PatternFill

    for row_index, source_row in enumerate(display_rows, start=4):
        fill = overview_row_fill(source_row)
        if fill:
            for cell in sheet[row_index]:
                cell.fill = fill
        if int(source_row.get("offline") or 0) > 0:
            sheet.cell(row_index, 4).fill = PatternFill(
                fill_type="solid",
                fgColor=_opaque_argb("FEE2E2"),
            )


def _append_switch_optical_summary_sheet(workbook, rows: list[dict[str, object | None]], alignment, border, header_font, header_fill=None) -> None:
    sheet = workbook.create_sheet("\u4ea4\u6362\u673a\u5149\u6a21\u5757\u7edf\u8ba1")
    sheet.append(["\u4ea4\u6362\u673a", "\u5149\u6a21\u5757\u6570\u91cf", "\u672a\u63d2\u5149\u6a21\u5757\u7aef\u53e3\u6570\u91cf", "\u672a\u63d2\u5149\u6a21\u5757\u7aef\u53e3"])
    grouped: dict[str, dict[str, object]] = {}
    for row in rows:
        switch_name = str(row.get("device_name") or "-")
        item = grouped.setdefault(switch_name, {"module_count": 0, "missing_ports": []})
        missing_ports = item["missing_ports"]
        if isinstance(missing_ports, list):
            missing_ports.extend(_normalize_missing_module_ports(row.get("missing_module_ports")))
        if row.get("switch_optical_status") == "no_module" and isinstance(missing_ports, list):
            missing_ports.extend(_normalize_missing_module_ports(row.get("interface_name") or "-"))
        else:
            item["module_count"] = int(item["module_count"]) + 1
    for switch_name in sorted(grouped):
        item = grouped[switch_name]
        missing_ports = _normalize_missing_module_ports(item["missing_ports"])
        sheet.append(
            [
                switch_name,
                item["module_count"],
                len(missing_ports),
                ", ".join(_short_interface_name(port) for port in missing_ports) if missing_ports else "-",
            ]
        )
    _format_export_sheet(sheet, alignment, border, header_font, header_fill)
    sheet.auto_filter.ref = sheet.dimensions


def _normalize_missing_module_ports(value: object) -> list[str]:
    if value in (None, "", "-"):
        return []
    if isinstance(value, (list, tuple, set)):
        return [str(item).strip() for item in value if str(item or "").strip() and str(item or "").strip() != "-"]
    return [part.strip() for part in re.split(r"[,，;；]", str(value)) if part.strip() and part.strip() != "-"]


def _set_switch_optical_summary_widths(workbook) -> None:
    if "\u4ea4\u6362\u673a\u5149\u6a21\u5757\u7edf\u8ba1" not in workbook.sheetnames:
        return
    sheet = workbook["\u4ea4\u6362\u673a\u5149\u6a21\u5757\u7edf\u8ba1"]
    for column, width in {"A": 22, "B": 14, "C": 20, "D": 80}.items():
        sheet.column_dimensions[column].width = width


def _ap_state(row: dict[str, object | None]) -> str:
    status = classify_fit_ap_state(
        row.get("ap_state"),
        row.get("state"),
        row.get("state_raw"),
        row.get("ap_state_display"),
        row.get("state_display"),
    )
    return "" if status == "unknown" else status


def _short_interface_name(value: object) -> str:
    return display_interface_name(value) or "-"
