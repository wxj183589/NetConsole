from __future__ import annotations

import re
from io import BytesIO
from typing import Any, Iterable, Mapping
from uuid import uuid4

from openpyxl import Workbook, load_workbook

from netconsole.core.paths import PathResolver
from netconsole.core.sites import SiteManager
from netconsole.models.api.rail_transit_base_data import (
    SectionDTO,
    StationDTO,
    StationSourceIssueDTO,
    StationTemplatePreviewDTO,
    StationTemplatePreviewRowDTO,
    StationTemplateSectionPreviewRowDTO,
)
from netconsole.services.excel_report_utils import apply_standard_sheet_style
from netconsole.services.rail_transit.base_data_query_service import RailTransitBaseDataQueryService
from netconsole.services.rail_transit.station_source_utils import (
    DEFAULT_MAIN_PATH_CODE,
    DEFAULT_STATION_SOURCE_GROUP,
    NODE_TYPE_LABELS,
    PLATFORM_LAYOUT_LABELS,
    STATION_SOURCE_FIELD,
    STRUCTURE_TYPE_LABELS,
    TRACK_FACILITY_LABELS,
    TURNBACK_DIRECTION_LABELS,
    bool_from_template,
    bool_label,
    label_for,
    legacy_turnback_type_for_facilities,
    normalize_station_source_value,
    normalize_track_facilities,
    parse_station_source_value,
    station_structure_defaults,
    track_facilities_label,
    value_from_label,
)
from netconsole.utils.mileage import parse_track_mileage


LINE_PARAM_HEADERS = (
    "线路名称",
    "项目类型",
    "网络类型",
    "主线路径编码",
    "站序递增方向名称",
    "站序递减方向名称",
    "设备来源分组",
    "设备来源字段",
    "备注",
)
STATION_HEADERS = (
    "来源站点值",
    "节点编码",
    "节点名称",
    "节点类型",
    "所属路径",
    "主线顺序",
    "参与方向判断",
    "中心里程",
    "车站结构",
    "站台形式",
    "线路端点",
    "运营终到/折返",
    "具备折返能力",
    "轨道设施",
    "折返方向",
    "端点延伸区间",
    "端点名称",
    "端点距离（米）",
    "端点里程",
    "启用",
    "备注",
)
SECTION_HEADERS = (
    "区间编码",
    "区间名称",
    "区间类型",
    "所属路径",
    "方向角色",
    "线路方向",
    "起始节点类型",
    "起始节点",
    "终到节点类型",
    "终到节点",
    "自动生成",
    "生成标识",
    "物理起点里程(m)",
    "物理终点里程(m)",
    "开放终点",
    "里程范围来源",
    "人工覆盖字段",
    "启用",
    "AP数量",
    "AP里程统计",
    "备注",
)
SECTION_KIND_LABELS = {
    "between_stations": "站间区间",
    "terminal_extension": "端点延伸",
    "depot_connection": "出入段连接",
    "manual": "人工区间",
    "legacy": "兼容区间",
}
SECTION_DIRECTION_LABELS = {
    "increasing": "站序递增",
    "decreasing": "站序递减",
    "none": "无",
    "unknown": "未知",
}
SECTION_NODE_TYPE_LABELS = {
    "station": "车站",
    "terminal_endpoint": "线路端点",
    "legacy": "兼容节点",
    "unknown": "未知",
}
SECTION_MILEAGE_SOURCE_LABELS = {
    "generated": "自动生成",
    "manual": "人工填写",
    "unavailable": "未生成",
}


class StationTemplateService:
    """线路、站点与区间基础资料 XLSX 模板、导出和草稿预览。"""

    def __init__(self, paths: PathResolver, query_service: RailTransitBaseDataQueryService) -> None:
        self.paths = paths
        self.query_service = query_service

    def build_blank_template(self, site_id: str) -> bytes:
        site_id = SiteManager(self.paths).validate_site_name(site_id)
        return self._workbook_bytes(self._line_metadata(site_id), [], [])

    def export_current(self, site_id: str) -> bytes:
        site_id = SiteManager(self.paths).validate_site_name(site_id)
        return self._workbook_bytes(
            self._line_metadata(site_id),
            self._all_stations(site_id),
            self._all_sections(site_id),
        )

    def preview(self, site_id: str, content: bytes, file_name: str = "") -> StationTemplatePreviewDTO:
        site_id = SiteManager(self.paths).validate_site_name(site_id)
        if len(content) > 10 * 1024 * 1024:
            raise ValueError("基础资料模板文件不能超过 10 MiB")
        if file_name and not file_name.casefold().endswith(".xlsx"):
            raise ValueError("基础资料模板仅支持 XLSX")
        workbook = load_workbook(
            BytesIO(content),
            data_only=True,
            read_only=True,
            keep_links=False,
        )
        try:
            line_sheet = workbook["01_线路参数"]
            node_sheet = workbook["02_线路节点"]
        except KeyError as exc:
            raise ValueError("基础资料模板必须包含 01_线路参数 和 02_线路节点 工作表") from exc
        issues: list[StationSourceIssueDTO] = []
        line_metadata = self._parse_line_metadata(line_sheet)
        if str(line_metadata.get("station_source_field") or "") != STATION_SOURCE_FIELD:
            issues.append(
                self._issue(
                    "error",
                    "station_source_field_invalid",
                    "设备来源字段只能为 station",
                    "设备来源字段",
                    blocking=True,
                )
            )

        existing_stations = self._all_stations(site_id)
        station_rows = self._preview_stations(
            site_id,
            node_sheet,
            line_metadata,
            existing_stations,
            issues,
        )
        station_lookup = {
            station.name: station
            for station in [
                *existing_stations,
                *(row.proposed_station for row in station_rows if row.proposed_station),
            ]
        }
        existing_sections = self._all_sections(site_id)
        section_sheet_present = "03_区间配置" in workbook.sheetnames
        if section_sheet_present:
            section_rows = self._preview_sections(
                workbook["03_区间配置"],
                line_metadata,
                station_lookup,
                existing_sections,
                issues,
            )
        else:
            section_rows = []
            issues.append(
                self._issue(
                    "info",
                    "station_template_sections_missing",
                    "模板未包含区间配置",
                    "03_区间配置",
                )
            )
        all_actions = [row.action for row in station_rows] + [row.action for row in section_rows]
        return StationTemplatePreviewDTO(
            valid=not any(issue.blocking for issue in issues),
            line_metadata=line_metadata,
            rows=station_rows,
            section_rows=section_rows,
            section_sheet_present=section_sheet_present,
            create_count=sum(action == "create" for action in all_actions),
            update_count=sum(action == "update" for action in all_actions),
            unchanged_count=sum(action == "unchanged" for action in all_actions),
            conflict_count=sum(action == "conflict" for action in all_actions),
            blocking_count=sum(issue.blocking for issue in issues),
            issues=issues,
        )

    def _preview_stations(
        self,
        site_id: str,
        sheet: Any,
        line_metadata: Mapping[str, Any],
        existing: list[StationDTO],
        issues: list[StationSourceIssueDTO],
    ) -> list[StationTemplatePreviewRowDTO]:
        existing_by_key = {row.source_station_key: row for row in existing if row.source_station_key}
        existing_by_identity = {
            (row.code.casefold(), row.name.casefold()): row
            for row in existing
            if row.code
        }
        existing_by_name = {
            row.name.casefold(): row
            for row in existing
            if row.name
        }
        rows: list[StationTemplatePreviewRowDTO] = []
        seen_source_keys: set[str] = set()
        for row_number, raw in self._iter_rows(sheet):
            row_issues: list[StationSourceIssueDTO] = []
            try:
                proposed = self._row_to_station(
                    raw,
                    line_metadata=line_metadata,
                    site_id=site_id,
                )
            except ValueError as exc:
                row_issues.append(
                    self._issue(
                        "error",
                        "station_template_invalid_enum",
                        str(exc),
                        blocking=True,
                    )
                )
                proposed = None
            matched: StationDTO | None = None
            if proposed is not None:
                if not proposed.name:
                    row_issues.append(
                        self._issue(
                            "error",
                            "station_name_required",
                            "节点名称不能为空",
                            "节点名称",
                            blocking=True,
                        )
                    )
                if proposed.source_station_key:
                    if proposed.source_station_key in seen_source_keys:
                        row_issues.append(
                            self._issue(
                                "error",
                                "station_source_duplicate",
                                "模板中同一来源站点重复",
                                "来源站点值",
                                blocking=True,
                            )
                        )
                    seen_source_keys.add(proposed.source_station_key)
                matched = (
                    existing_by_key.get(proposed.source_station_key)
                    or existing_by_identity.get((proposed.code.casefold(), proposed.name.casefold()))
                    or existing_by_name.get(proposed.name.casefold())
                )
                if matched:
                    proposed = proposed.model_copy(
                        update={
                            "id": matched.id,
                            "node_uid": matched.node_uid,
                            "ap_count": matched.ap_count,
                            "section_count": matched.section_count,
                            "mileage_min": matched.mileage_min,
                            "mileage_max": matched.mileage_max,
                            "source_device_count": matched.source_device_count,
                            "source_sync_status": matched.source_sync_status,
                        }
                    )
                action = (
                    "update"
                    if matched and self._station_payload(matched) != self._station_payload(proposed)
                    else "unchanged"
                    if matched
                    else "create"
                )
            else:
                action = "conflict"
            valid = not any(issue.blocking for issue in row_issues)
            rows.append(
                StationTemplatePreviewRowDTO(
                    row_number=row_number,
                    source_station_value=proposed.source_station_value if proposed else str(raw.get("来源站点值") or ""),
                    source_station_key=proposed.source_station_key if proposed else "",
                    code=proposed.code if proposed else str(raw.get("节点编码") or ""),
                    name=proposed.name if proposed else str(raw.get("节点名称") or ""),
                    node_type=proposed.node_type if proposed else "unknown",
                    path_code=proposed.path_code if proposed else "",
                    sort_order=proposed.sort_order if proposed else None,
                    participates_in_direction=proposed.participates_in_direction if proposed else False,
                    proposed_station=proposed,
                    action="conflict" if not valid else action,  # type: ignore[arg-type]
                    valid=valid,
                    issues=row_issues,
                )
            )
            issues.extend(row_issues)
        return rows

    def _preview_sections(
        self,
        sheet: Any,
        line_metadata: Mapping[str, Any],
        station_lookup: Mapping[str, StationDTO],
        existing: list[SectionDTO],
        issues: list[StationSourceIssueDTO],
    ) -> list[StationTemplateSectionPreviewRowDTO]:
        existing_by_generation = {
            section.generation_key: section
            for section in existing
            if section.generation_key
        }
        existing_by_code = {
            section.section_code.casefold(): section
            for section in existing
            if section.section_code
        }
        existing_by_name = {
            section.name.casefold(): section
            for section in existing
            if section.name
        }
        rows: list[StationTemplateSectionPreviewRowDTO] = []
        seen_generation_keys: set[str] = set()
        for row_number, raw in self._iter_rows(sheet):
            row_issues: list[StationSourceIssueDTO] = []
            try:
                proposed = self._row_to_section(raw, line_metadata, station_lookup)
            except ValueError as exc:
                row_issues.append(
                    self._issue(
                        "error",
                        "section_generation_conflict",
                        str(exc),
                        blocking=True,
                    )
                )
                proposed = None
            matched: SectionDTO | None = None
            if proposed is not None:
                if proposed.generation_key:
                    if proposed.generation_key in seen_generation_keys:
                        row_issues.append(
                            self._issue(
                                "error",
                                "section_duplicate_generation_key",
                                "模板中自动区间生成标识重复",
                                "生成标识",
                                blocking=True,
                            )
                        )
                    seen_generation_keys.add(proposed.generation_key)
                matched = (
                    existing_by_generation.get(proposed.generation_key)
                    or existing_by_code.get(proposed.section_code.casefold())
                    or existing_by_name.get(proposed.name.casefold())
                )
                if matched:
                    preserved = {
                        "id": matched.id,
                        "ap_count": matched.ap_count,
                        "mileage_min": matched.mileage_min,
                        "mileage_max": matched.mileage_max,
                    }
                    if "物理起点里程(m)" not in raw:
                        preserved.update(
                            {
                                "section_mileage_start_m": matched.section_mileage_start_m,
                                "section_mileage_end_m": matched.section_mileage_end_m,
                                "section_mileage_open_end": matched.section_mileage_open_end,
                                "section_mileage_source": matched.section_mileage_source,
                                "manual_override_fields": matched.manual_override_fields,
                            }
                        )
                    proposed = proposed.model_copy(update=preserved)
                action = (
                    "update"
                    if matched and self._section_payload(matched) != self._section_payload(proposed)
                    else "unchanged"
                    if matched
                    else "create"
                )
            else:
                action = "conflict"
            valid = not any(issue.blocking for issue in row_issues)
            rows.append(
                StationTemplateSectionPreviewRowDTO(
                    row_number=row_number,
                    section_code=proposed.section_code if proposed else str(raw.get("区间编码") or ""),
                    name=proposed.name if proposed else str(raw.get("区间名称") or ""),
                    section_kind=proposed.section_kind if proposed else "manual",
                    path_code=proposed.path_code if proposed else "",
                    direction_role=proposed.direction_role if proposed else "unknown",
                    line_direction=proposed.line_direction if proposed else "",
                    start_node_type=proposed.start_node_type if proposed else "unknown",
                    start_station=proposed.start_station if proposed else str(raw.get("起始节点") or ""),
                    end_node_type=proposed.end_node_type if proposed else "unknown",
                    end_station=proposed.end_station if proposed else str(raw.get("终到节点") or ""),
                    proposed_section=proposed,
                    action="conflict" if not valid else action,  # type: ignore[arg-type]
                    valid=valid,
                    issues=row_issues,
                )
            )
            issues.extend(row_issues)
        return rows

    def _workbook_bytes(
        self,
        metadata: Mapping[str, Any],
        stations: Iterable[StationDTO],
        sections: Iterable[SectionDTO],
    ) -> bytes:
        workbook = Workbook()
        line_sheet = workbook.active
        line_sheet.title = "01_线路参数"
        line_sheet.append(list(LINE_PARAM_HEADERS))
        line_sheet.append(
            [
                metadata.get("line_name") or "",
                metadata.get("system_type") or "",
                metadata.get("network_domain") or "",
                metadata.get("main_path_code") or DEFAULT_MAIN_PATH_CODE,
                metadata.get("increasing_direction_name") or "上行",
                metadata.get("decreasing_direction_name") or "下行",
                metadata.get("station_source_group_name") or DEFAULT_STATION_SOURCE_GROUP,
                STATION_SOURCE_FIELD,
                metadata.get("remark") or "",
            ]
        )
        apply_standard_sheet_style(line_sheet)

        node_sheet = workbook.create_sheet("02_线路节点")
        node_sheet.append(list(STATION_HEADERS))
        for station in stations:
            node_sheet.append(
                [
                    station.source_station_value,
                    station.code,
                    station.name,
                    label_for(NODE_TYPE_LABELS, station.node_type),
                    station.path_code,
                    station.sort_order if station.sort_order is not None else "",
                    bool_label(station.participates_in_direction),
                    station.center_mileage_text,
                    label_for(STRUCTURE_TYPE_LABELS, station.structure_type),
                    label_for(PLATFORM_LAYOUT_LABELS, station.platform_layout),
                    bool_label(station.is_line_terminal),
                    bool_label(station.is_service_terminal),
                    bool_label(station.turnback_capable),
                    track_facilities_label(station.track_facilities),
                    label_for(TURNBACK_DIRECTION_LABELS, station.turnback_direction),
                    bool_label(station.terminal_extension_enabled),
                    station.terminal_endpoint_label,
                    station.terminal_extension_distance_m if station.terminal_extension_distance_m is not None else "",
                    station.terminal_endpoint_mileage_text,
                    bool_label(station.enabled),
                    station.remark,
                ]
            )
        apply_standard_sheet_style(node_sheet)

        section_sheet = workbook.create_sheet("03_区间配置")
        section_sheet.append(list(SECTION_HEADERS))
        for section in sections:
            section_sheet.append(
                [
                    section.section_code,
                    section.name,
                    label_for(SECTION_KIND_LABELS, section.section_kind),
                    section.path_code,
                    label_for(SECTION_DIRECTION_LABELS, section.direction_role),
                    section.line_direction or section.line_side,
                    label_for(SECTION_NODE_TYPE_LABELS, section.start_node_type),
                    section.start_station,
                    label_for(SECTION_NODE_TYPE_LABELS, section.end_node_type),
                    section.end_station,
                    bool_label(section.auto_generated),
                    section.generation_key,
                    section.section_mileage_start_m if section.section_mileage_start_m is not None else "",
                    section.section_mileage_end_m if section.section_mileage_end_m is not None else "",
                    bool_label(section.section_mileage_open_end),
                    label_for(SECTION_MILEAGE_SOURCE_LABELS, section.section_mileage_source),
                    "、".join(section.manual_override_fields),
                    bool_label(section.enabled),
                    section.ap_count,
                    self._mileage_range(section.mileage_min, section.mileage_max),
                    section.remark,
                ]
            )
        apply_standard_sheet_style(section_sheet)
        self._append_field_description(workbook)
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _parse_line_metadata(self, sheet: Any) -> dict[str, Any]:
        row = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
        values = {
            header: row[index] if index < len(row) else ""
            for index, header in enumerate(LINE_PARAM_HEADERS)
        }
        return {
            "line_name": str(values.get("线路名称") or "").strip(),
            "system_type": str(values.get("项目类型") or "").strip(),
            "network_domain": str(values.get("网络类型") or "default").strip() or "default",
            "main_path_code": str(values.get("主线路径编码") or DEFAULT_MAIN_PATH_CODE).strip() or DEFAULT_MAIN_PATH_CODE,
            "increasing_direction_name": str(values.get("站序递增方向名称") or "上行").strip() or "上行",
            "decreasing_direction_name": str(values.get("站序递减方向名称") or "下行").strip() or "下行",
            "station_source_group_name": str(values.get("设备来源分组") or DEFAULT_STATION_SOURCE_GROUP).strip() or DEFAULT_STATION_SOURCE_GROUP,
            "station_source_field": str(values.get("设备来源字段") or STATION_SOURCE_FIELD).strip() or STATION_SOURCE_FIELD,
            "remark": str(values.get("备注") or "").strip(),
        }

    def _iter_rows(self, sheet: Any) -> Iterable[tuple[int, dict[str, Any]]]:
        headers = [
            str(cell.value or "").strip()
            for cell in next(sheet.iter_rows(min_row=1, max_row=1))
        ]
        for row_number, cells in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            values = {
                header: cells[index] if index < len(cells) else ""
                for index, header in enumerate(headers)
            }
            if any(str(value or "").strip() for value in values.values()):
                yield row_number, values

    def _row_to_station(
        self,
        raw: Mapping[str, Any],
        *,
        line_metadata: Mapping[str, Any],
        site_id: str,
    ) -> StationDTO:
        source_value = str(raw.get("来源站点值") or "").strip()
        code = str(raw.get("节点编码") or "").strip()
        name = str(raw.get("节点名称") or "").strip()
        if source_value and (not code or not name):
            parsed = parse_station_source_value(
                source_value,
                main_path_code=str(line_metadata.get("main_path_code") or DEFAULT_MAIN_PATH_CODE),
            )
            code = code or parsed.code
            name = name or parsed.name
        source_display, source_key = normalize_station_source_value(
            source_value or (f"{code}-{name}" if code else name)
        )
        node_type = value_from_label(NODE_TYPE_LABELS, raw.get("节点类型"), default="station")
        if node_type not in NODE_TYPE_LABELS:
            raise ValueError(f"节点类型无效：{raw.get('节点类型')}")
        path_code = str(
            raw.get("所属路径")
            or (DEFAULT_MAIN_PATH_CODE if node_type == "station" else "UNASSIGNED")
        ).strip()
        structure_default, platform_default = station_structure_defaults(
            node_type,
            path_code,
            line_metadata.get("main_path_code"),
        )
        structure_type = value_from_label(
            STRUCTURE_TYPE_LABELS,
            raw.get("车站结构"),
            default=structure_default,
        )
        if structure_type not in STRUCTURE_TYPE_LABELS:
            raise ValueError(f"车站结构无效：{raw.get('车站结构')}")
        platform_layout = value_from_label(
            PLATFORM_LAYOUT_LABELS,
            raw.get("站台形式"),
            default=platform_default,
        )
        if platform_layout not in PLATFORM_LAYOUT_LABELS:
            raise ValueError(f"站台形式无效：{raw.get('站台形式')}")
        legacy_turnback = value_from_label(
            {
                "none": "无",
                "crossover": "渡线",
                "pocket_track": "中间折返线/存车线",
                "tail_track": "站后折返线",
                "loop": "环形折返",
                "depot_connection": "出入段线",
                "other": "其他",
                "unknown": "类型未知",
            },
            raw.get("折返类型"),
            default="none",
        )
        facilities_value = raw.get("轨道设施")
        facilities = normalize_track_facilities(
            facilities_value if facilities_value not in (None, "") else None,
            legacy_turnback_type=legacy_turnback,
        )
        turnback_direction = value_from_label(
            TURNBACK_DIRECTION_LABELS,
            raw.get("折返方向"),
            default="none",
        )
        if turnback_direction not in TURNBACK_DIRECTION_LABELS:
            raise ValueError(f"折返方向无效：{raw.get('折返方向')}")
        sort_order = self._int_or_none(raw.get("主线顺序"))
        participates = bool_from_template(
            raw.get("参与方向判断"),
            default=node_type == "station",
        )
        center_text = str(raw.get("中心里程") or "").strip()
        center_mileage = parse_track_mileage(center_text)
        if center_text and center_mileage.error:
            raise ValueError("中心里程格式无效")
        terminal_distance = self._float_or_none(raw.get("端点距离（米）"))
        if terminal_distance is not None and terminal_distance < 0:
            raise ValueError("端点距离不能为负数")
        terminal_mileage_text = str(raw.get("端点里程") or "").strip()
        if terminal_mileage_text and parse_track_mileage(terminal_mileage_text).error:
            raise ValueError("端点里程格式无效")
        turnback_capable_value = self._first(
            raw,
            "具备折返能力",
            "可折返",
        )
        service_terminal_value = self._first(
            raw,
            "运营终到/折返",
            "运营终点",
        )
        return StationDTO(
            id=f"new:template:{source_key or uuid4().hex}",
            node_uid=str(uuid4()),
            name=name,
            code=code,
            line_name=str(line_metadata.get("line_name") or ""),
            sort_order=sort_order,
            remark=str(raw.get("备注") or "").strip(),
            source_station_value=source_display,
            source_station_key=source_key,
            node_type=node_type,  # type: ignore[arg-type]
            path_code=path_code,
            participates_in_direction=participates,
            structure_type=structure_type,  # type: ignore[arg-type]
            platform_layout=platform_layout,  # type: ignore[arg-type]
            center_mileage_text=center_text,
            center_mileage_m=center_mileage.meters,
            is_line_terminal=bool_from_template(raw.get("线路端点"), default=False),
            is_service_terminal=bool_from_template(service_terminal_value, default=False),
            turnback_capable=bool_from_template(turnback_capable_value, default=False),
            turnback_type=legacy_turnback_type_for_facilities(facilities),  # type: ignore[arg-type]
            track_facilities=facilities,  # type: ignore[arg-type]
            turnback_direction=turnback_direction,  # type: ignore[arg-type]
            terminal_extension_enabled=bool_from_template(raw.get("端点延伸区间"), default=False),
            terminal_endpoint_label=str(raw.get("端点名称") or "端点").strip() or "端点",
            terminal_extension_distance_m=terminal_distance,
            terminal_endpoint_mileage_text=terminal_mileage_text,
            enabled=bool_from_template(raw.get("启用"), default=True),
            source_kind="template",
            source_sync_status="manual",
        )

    def _row_to_section(
        self,
        raw: Mapping[str, Any],
        line_metadata: Mapping[str, Any],
        station_lookup: Mapping[str, StationDTO],
    ) -> SectionDTO:
        name = str(raw.get("区间名称") or "").strip()
        if not name:
            raise ValueError("区间名称不能为空")
        path_code = str(raw.get("所属路径") or line_metadata.get("main_path_code") or DEFAULT_MAIN_PATH_CODE).strip()
        section_kind = value_from_label(
            SECTION_KIND_LABELS,
            raw.get("区间类型"),
            default="manual",
        )
        direction_role = value_from_label(
            SECTION_DIRECTION_LABELS,
            raw.get("方向角色"),
            default="none",
        )
        start_node_type = value_from_label(
            SECTION_NODE_TYPE_LABELS,
            raw.get("起始节点类型"),
            default="legacy",
        )
        end_node_type = value_from_label(
            SECTION_NODE_TYPE_LABELS,
            raw.get("终到节点类型"),
            default="legacy",
        )
        if section_kind not in SECTION_KIND_LABELS:
            raise ValueError("区间类型无效")
        if direction_role not in SECTION_DIRECTION_LABELS:
            raise ValueError("方向角色无效")
        if start_node_type not in SECTION_NODE_TYPE_LABELS or end_node_type not in SECTION_NODE_TYPE_LABELS:
            raise ValueError("区间节点类型无效")
        start_station = str(raw.get("起始节点") or "").strip()
        end_station = str(raw.get("终到节点") or "").strip()
        if not start_station or not end_station or start_station == end_station:
            raise ValueError("区间起始节点和终到节点必须填写且不能相同")
        generation_key = str(raw.get("生成标识") or "").strip()
        auto_generated = bool_from_template(raw.get("自动生成"), default=False)
        start_uid = self._resolve_node_uid(
            start_node_type,
            start_station,
            path_code,
            generation_key,
            station_lookup,
        )
        end_uid = self._resolve_node_uid(
            end_node_type,
            end_station,
            path_code,
            generation_key,
            station_lookup,
        )
        if auto_generated and (not generation_key or not start_uid or not end_uid):
            raise ValueError("自动区间缺少稳定生成标识或正式节点")
        line_direction = str(raw.get("线路方向") or "").strip()
        mileage_columns_present = "物理起点里程(m)" in raw
        mileage_start = self._float_or_none(raw.get("物理起点里程(m)")) if mileage_columns_present else None
        mileage_end = self._float_or_none(raw.get("物理终点里程(m)")) if mileage_columns_present else None
        mileage_open_end = bool_from_template(raw.get("开放终点"), default=False) if mileage_columns_present else False
        mileage_source = value_from_label(
            SECTION_MILEAGE_SOURCE_LABELS,
            raw.get("里程范围来源"),
            default="unavailable",
        ) if mileage_columns_present else "unavailable"
        if mileage_source not in SECTION_MILEAGE_SOURCE_LABELS:
            raise ValueError("区间里程范围来源无效")
        if mileage_start is not None and mileage_start < 0:
            raise ValueError("区间物理起点里程不能小于 0")
        if mileage_end is not None and mileage_end < 0:
            raise ValueError("区间物理终点里程不能小于 0")
        if mileage_open_end:
            if (
                mileage_source == "unavailable"
                or section_kind != "terminal_extension"
                or mileage_start is None
                or mileage_end is not None
            ):
                raise ValueError("开放终点仅适用于端点延伸区间，且必须保留起点、清空终点")
        elif mileage_source != "unavailable":
            if mileage_start is None or mileage_end is None or mileage_end <= mileage_start:
                raise ValueError("非开放区间必须填写有效的物理起点和终点里程")
        elif mileage_start is not None or mileage_end is not None:
            raise ValueError("已填写物理里程时，范围来源不能为未生成")
        override_text = str(raw.get("人工覆盖字段") or "").strip()
        manual_override_fields = sorted(
            {
                field.strip()
                for field in re.split(r"[、,，;；\s]+", override_text)
                if field.strip()
            }
        )
        return SectionDTO(
            id=f"new:template-section:{uuid4().hex}",
            name=name,
            section_code=str(raw.get("区间编码") or "").strip(),
            section_kind=section_kind,  # type: ignore[arg-type]
            path_code=path_code,
            direction_role=direction_role,  # type: ignore[arg-type]
            line_direction=line_direction,
            start_node_type=start_node_type,  # type: ignore[arg-type]
            start_node_uid=start_uid,
            start_station=start_station,
            end_node_type=end_node_type,  # type: ignore[arg-type]
            end_node_uid=end_uid,
            end_station=end_station,
            line_side=line_direction,
            auto_generated=auto_generated,
            generation_key=generation_key,
            manual_override_fields=manual_override_fields,
            section_mileage_start_m=mileage_start,
            section_mileage_end_m=mileage_end,
            section_mileage_open_end=mileage_open_end,
            section_mileage_source=mileage_source,  # type: ignore[arg-type]
            enabled=bool_from_template(raw.get("启用"), default=True),
            source_kind="generated" if auto_generated else "template",
            ap_count=0,
            mileage_min=None,
            mileage_max=None,
            remark=str(raw.get("备注") or "").strip(),
        )

    @staticmethod
    def _resolve_node_uid(
        node_type: str,
        display_name: str,
        path_code: str,
        generation_key: str,
        stations: Mapping[str, StationDTO],
    ) -> str:
        if node_type == "station":
            station = stations.get(display_name)
            return station.node_uid if station else ""
        if node_type != "terminal_endpoint":
            return ""
        match = re.search(r"endpoint:[^|]+\:(?:low|high)", generation_key)
        if match:
            return match.group(0)
        endpoint_station = re.search(r"（(.+?)端）$", display_name)
        if not endpoint_station:
            return ""
        station = stations.get(endpoint_station.group(1))
        if not station:
            return ""
        same_path = sorted(
            (
                item
                for item in stations.values()
                if item.path_code.casefold() == path_code.casefold()
                and item.sort_order is not None
                and item.enabled
                and item.participates_in_direction
            ),
            key=lambda item: item.sort_order or 0,
        )
        if not same_path:
            return ""
        if station.node_uid == same_path[0].node_uid:
            return f"endpoint:{path_code}:low"
        if station.node_uid == same_path[-1].node_uid:
            return f"endpoint:{path_code}:high"
        return ""

    def _append_field_description(self, workbook: Workbook) -> None:
        sheet = workbook.create_sheet("字段说明")
        sheet.append(["字段名称", "是否必填", "含义", "允许值", "默认值", "示例"])
        rows = [
            ("设备来源字段", "是", "固定读取设备管理 devices.station", "station", "station", "station"),
            ("节点类型", "是", "普通车站或特殊节点", "、".join(NODE_TYPE_LABELS.values()), "普通车站", "停车场"),
            ("所属路径", "是", "主线路径或接轨路径", "文本", "MAIN/UNASSIGNED", "MAIN"),
            ("主线顺序", "参与方向判断时必填", "同一路径内方向判断顺序，不要求连续", "整数", "", "32"),
            ("中心里程", "否", "站点中心参考里程，不作为 AP 实际覆盖范围", "ZDK12+345、K12+345、12+345、12345.5", "", "K12+345"),
            ("车站结构", "否", "MAIN 普通新站为空时默认地下", "、".join(STRUCTURE_TYPE_LABELS.values()), "地下/特殊节点未填写", "地下"),
            ("站台形式", "否", "MAIN 普通新站为空时默认岛式", "、".join(PLATFORM_LAYOUT_LABELS.values()), "岛式/特殊节点未填写", "岛式"),
            ("轨道设施", "否", "支持多项，以顿号、逗号或分号分隔", "、".join(TRACK_FACILITY_LABELS.values()), "", "折返线、存车线"),
            ("端点延伸区间", "否", "终点站外侧至线路物理端点仍有轨道时启用", "是/否", "否", "是"),
            ("区间类型", "是", "区间业务类型", "、".join(SECTION_KIND_LABELS.values()), "人工区间", "站间区间"),
            ("方向角色", "是", "站序递增或递减方向", "、".join(SECTION_DIRECTION_LABELS.values()), "无", "站序递增"),
            ("物理起点里程(m)", "有范围时必填", "根据站台中心里程或人工配置得到的区间物理起点", "非负数", "", "152"),
            ("物理终点里程(m)", "非开放范围必填", "区间物理终点；开放终点时必须留空", "大于起点的非负数", "", "1801"),
            ("开放终点", "否", "高里程端未明确终点时使用，不伪造终点数值", "是/否", "否", "是"),
            ("里程范围来源", "是", "区分自动生成、人工填写和未生成的物理范围", "、".join(SECTION_MILEAGE_SOURCE_LABELS.values()), "未生成", "自动生成"),
            ("人工覆盖字段", "否", "保留自动区间已人工调整的字段状态", "字段名，以顿号分隔", "", "section_mileage_start_m"),
            ("AP数量", "只读", "根据轨旁 AP 正式区间归属实时统计，导入时忽略", "非负整数", "0", "9"),
            ("AP里程统计", "只读", "根据关联轨旁 AP 有效里程统计，导入时忽略", "文本", "--", "12000–13000 m"),
            ("生成标识", "自动区间必填", "稳定区间身份，不使用区间名称替代", "文本", "", "MAIN|between|...|increasing"),
        ]
        for row in rows:
            sheet.append(list(row))
        apply_standard_sheet_style(sheet)

    def _line_metadata(self, site_id: str) -> dict[str, Any]:
        metadata = SiteManager(self.paths).load_site_metadata(site_id)
        return {
            "line_name": str(metadata.get("line_name") or ""),
            "system_type": str(metadata.get("system_type") or ""),
            "network_domain": str(metadata.get("network_domain") or "default"),
            "main_path_code": str(metadata.get("main_path_code") or DEFAULT_MAIN_PATH_CODE),
            "increasing_direction_name": str(metadata.get("increasing_direction_name") or "上行"),
            "decreasing_direction_name": str(metadata.get("decreasing_direction_name") or "下行"),
            "station_source_group_name": str(metadata.get("station_source_group_name") or DEFAULT_STATION_SOURCE_GROUP),
            "station_source_field": STATION_SOURCE_FIELD,
            "remark": str(metadata.get("remark") or ""),
        }

    def _all_stations(self, site_id: str) -> list[StationDTO]:
        result: list[StationDTO] = []
        page = 1
        while True:
            data = self.query_service.list_stations(site_id, page=page, page_size=200)
            result.extend(data.items)
            if len(result) >= data.total or not data.items:
                return result
            page += 1

    def _all_sections(self, site_id: str) -> list[SectionDTO]:
        result: list[SectionDTO] = []
        page = 1
        while True:
            data = self.query_service.list_sections(site_id, page=page, page_size=200)
            result.extend(data.items)
            if len(result) >= data.total or not data.items:
                return result
            page += 1

    @staticmethod
    def _station_payload(station: StationDTO) -> dict[str, Any]:
        return {
            key: getattr(station, key)
            for key in (
                "node_uid",
                "name",
                "code",
                "line_name",
                "sort_order",
                "remark",
                "source_station_value",
                "source_station_key",
                "node_type",
                "path_code",
                "participates_in_direction",
                "structure_type",
                "platform_layout",
                "center_mileage_text",
                "center_mileage_m",
                "is_line_terminal",
                "is_service_terminal",
                "turnback_capable",
                "turnback_type",
                "track_facilities",
                "turnback_direction",
                "terminal_extension_enabled",
                "terminal_endpoint_label",
                "terminal_extension_distance_m",
                "terminal_endpoint_mileage_text",
                "enabled",
            )
        }

    @staticmethod
    def _section_payload(section: SectionDTO) -> dict[str, Any]:
        return {
            key: getattr(section, key)
            for key in (
                "name",
                "section_code",
                "section_kind",
                "path_code",
                "direction_role",
                "line_direction",
                "start_node_type",
                "start_node_uid",
                "start_station",
                "end_node_type",
                "end_node_uid",
                "end_station",
                "line_side",
                "auto_generated",
                "generation_key",
                "manual_override_fields",
                "section_mileage_start_m",
                "section_mileage_end_m",
                "section_mileage_open_end",
                "section_mileage_source",
                "enabled",
                "source_kind",
                "remark",
            )
        }

    @staticmethod
    def _first(raw: Mapping[str, Any], *keys: str) -> Any:
        for key in keys:
            if key in raw:
                return raw.get(key)
        return None

    @staticmethod
    def _int_or_none(value: Any) -> int | None:
        text = str(value or "").strip()
        return None if not text else int(float(text))

    @staticmethod
    def _float_or_none(value: Any) -> float | None:
        text = str(value or "").strip()
        return None if not text else float(text)

    @staticmethod
    def _mileage_range(minimum: float | None, maximum: float | None) -> str:
        if minimum is None and maximum is None:
            return "--"
        if maximum is None or minimum == maximum:
            return f"{minimum:g} m" if minimum is not None else "--"
        return f"{minimum:g}–{maximum:g} m"

    @staticmethod
    def _issue(
        severity: str,
        code: str,
        message: str,
        field_name: str = "",
        *,
        blocking: bool = False,
    ) -> StationSourceIssueDTO:
        return StationSourceIssueDTO(
            severity=severity,  # type: ignore[arg-type]
            code=code,
            message=message,
            field_name=field_name,
            blocking=blocking,
        )


__all__ = [
    "LINE_PARAM_HEADERS",
    "SECTION_HEADERS",
    "STATION_HEADERS",
    "StationTemplateService",
]
