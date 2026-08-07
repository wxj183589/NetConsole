from __future__ import annotations

import sqlite3
from copy import deepcopy
from pathlib import Path

from netconsole.services.export_identity_diagnostics import (
    ExportIdentityDiagnostics,
    unavailable_export_identity_diagnostics,
)
from netconsole.services.online_mr_chart_builder import OnlineMrChartBuilder, ChartData
from netconsole.services.online_mr.traffic_analysis import build_iperf_traffic_overview


class OnlineMrAnalysisReportExporter:
    def __init__(self) -> None:
        self._result_metadata: dict[str, object] = {}

    @property
    def result_metadata(self) -> dict[str, object]:
        return deepcopy(self._result_metadata)

    def export(self, session_dir: Path, output_path: Path) -> Path:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, LineChart
        from openpyxl.styles import Alignment, Font, PatternFill

        session_dir = Path(session_dir)
        output_path = Path(output_path)
        self._result_metadata = {}
        db_path = session_dir / "parsed" / "online_diagnosis.sqlite"
        builder = OnlineMrChartBuilder(db_path)
        workbook = Workbook()
        overview = workbook.active
        overview.title = "综合结论"
        self._write_overview(overview, db_path)
        self._append_offline_report_sheets(workbook, session_dir, db_path)

        chart_specs = [
            ("主链路信号趋势表", "主链路信号趋势图", builder.build_active_rssi_series(), LineChart),
            ("主链路切换前后信号趋势表", "主链路切换前后信号趋势图", builder.build_switch_rssi_series(), LineChart),
            ("主链路切换日志RSSI趋势表", "主链路切换日志RSSI趋势图", builder.build_switch_log_rssi_series(), LineChart),
            ("信道繁忙度趋势表", "信道繁忙度趋势图", builder.build_channel_busy_series(), LineChart),
            ("Ping丢包率趋势表", "Ping丢包率趋势图", builder.build_ping_loss_series(), LineChart),
            ("Ping延迟趋势表", "Ping延迟趋势图", builder.build_ping_latency_series(), LineChart),
            ("接口PPS趋势表", "接口PPS趋势图", builder.build_interface_rate_series(), LineChart),
            ("业务打流趋势表", "业务打流趋势图", builder.build_traffic_rate_series(), LineChart),
            ("主链路切换原因统计表", "主链路切换原因统计图", builder.build_switch_reason_summary(), BarChart),
        ]
        for data_sheet_name, chart_sheet_name, chart_data, chart_type in chart_specs:
            data_sheet = workbook.create_sheet(data_sheet_name)
            self._write_chart_data(data_sheet, chart_data)
            chart_sheet = workbook.create_sheet(chart_sheet_name)
            self._write_chart_sheet(chart_sheet, data_sheet, chart_data, chart_type, chart_sheet_name)

        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E79")
                cell.alignment = Alignment(horizontal="center")
            for column in sheet.columns:
                letter = column[0].column_letter
                sheet.column_dimensions[letter].width = min(max(max(len(str(cell.value or "")) for cell in column) + 2, 12), 36)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return output_path

    def _append_offline_report_sheets(self, workbook, session_dir: Path, db_path: Path) -> None:
        mesh_link_detail_rows = self._mesh_link_detail_rows(db_path)
        self._result_metadata = {
            "export_identity_diagnostics": self._inspect_mesh_link_detail_rows(mesh_link_detail_rows)
        }
        sheet_rows = {
            "会话信息": self._session_info_rows(session_dir, db_path),
            "质量总览": [["维度", "结论", "说明"], ["业务连通性", "N/A", "缺少 fping 数据时不按 0 计算"], ["Mesh主链路", "N/A", "基于 parsed 数据计算"], ["空口繁忙度", "N/A", "无数据时显示 N/A"]],
            "时间轴质量分析": [["时间", "Active Peer", "MR侧RSSI", "TxBusy", "RxBusy", "fping丢包率", "平均延迟", "质量等级", "原因"]],
            "fping业务质量": self._fping_quality_rows(db_path),
            "fping 1s聚合": self._fping_1s_rows(db_path),
            "Mesh主链路质量": self._mesh_main_link_rows(db_path),
            "链路明细": mesh_link_detail_rows,
            "Peer稳定性分析": [["Peer", "Active次数", "平均RSSI", "最低RSSI", "切换次数", "结论"]],
            "主链路切换历史": self._switch_history_rows(db_path),
            "主链路切换日志": self._active_switch_log_rows(db_path),
            "切换影响分析": [["切换时间", "原Peer", "新Peer", "切换前RSSI", "切换后RSSI", "切换原因", "是否影响业务", "建议"]],
            "丢包关联分析": [["丢包时间", "丢包率", "Active RSSI", "是否切换", "TxBusy", "RxBusy", "判断"]],
            "异常事件清单": [["事件ID", "事件时间", "事件类型", "级别", "说明", "证据ID"]],
            "空口繁忙度分析": self._channel_busy_rows(db_path),
            "射频统计分析": [["时间", "指标", "当前值", "增量", "结论"]],
            "接口速率分析": self._interface_rate_rows(db_path),
            "业务打流概览": self._traffic_overview_rows(db_path),
            "链路重建与连接异常": [["时间", "Peer", "事件", "RSSI", "原因", "证据ID"]],
            "原始证据片段": [["证据ID", "来源Sheet", "事件类型", "采样时间", "源文件", "源行号"]],
            "参数配置": [["配置项", "值"], ["规则来源", "resources/mesh_quality_rules.json"], ["报告类型", "VEHICLE_MR_REALTIME_OFFLINE"]],
        }
        for name, rows in sheet_rows.items():
            sheet = workbook.create_sheet(name)
            for row in rows:
                sheet.append(["N/A" if value is None else value for value in row])

    def _traffic_overview_rows(self, db_path: Path) -> list[list[object]]:
        rows: list[list[object]] = [[
            "范围", "运行ID", "协议", "方向", "开始时间", "结束时间", "时长(s)",
            "记录数", "平均吞吐(Mbps)", "最小吞吐(Mbps)", "最大吞吐(Mbps)",
            "发送数据(B)", "接收数据(B)", "UDP发送包", "UDP接收包", "UDP丢失包",
            "UDP丢包率(%)", "UDP平均Jitter(ms)", "UDP最小Jitter(ms)", "UDP最大Jitter(ms)",
            "TCP Retransmits", "数据质量说明",
        ]]
        if not db_path.exists():
            return rows
        try:
            with sqlite3.connect(db_path) as conn:
                overview = build_iperf_traffic_overview(conn)
        except sqlite3.Error:
            return rows

        def values(scope: str, item: dict[str, object]) -> list[object]:
            return [
                scope,
                item.get("run_id", ""),
                item.get("protocol", ""),
                item.get("label", item.get("direction", "")),
                item.get("started_at"),
                item.get("ended_at"),
                item.get("duration_seconds"),
                item.get("record_count"),
                item.get("average_mbps"),
                item.get("minimum_mbps"),
                item.get("maximum_mbps"),
                item.get("sent_bytes"),
                item.get("received_bytes"),
                item.get("sent_packets"),
                item.get("received_packets"),
                item.get("lost_packets"),
                item.get("loss_percent"),
                item.get("average_jitter_ms"),
                item.get("minimum_jitter_ms"),
                item.get("maximum_jitter_ms"),
                item.get("retransmits"),
                overview.get("data_quality_note", ""),
            ]

        rows.append(values("整场", {
            **overview.get("overall", {}),
            "protocol": overview.get("protocol", ""),
            "direction": overview.get("direction", ""),
            "started_at": overview.get("started_at"),
            "ended_at": overview.get("ended_at"),
        }))
        for item in overview.get("directions", []):
            rows.append(values("单次运行", item))
        return rows

    @staticmethod
    def _inspect_mesh_link_detail_rows(rows: list[list[object]]) -> dict[str, object]:
        try:
            diagnostics = ExportIdentityDiagnostics("online_mr_compat_detail")
            headers = rows[0] if rows else []
            return diagnostics.inspect_online_mr_detail_rows(rows[1:], headers=headers).to_dict()
        except Exception as exc:
            return unavailable_export_identity_diagnostics("online_mr_compat_detail", exc)

    def _mesh_main_link_rows(self, db_path: Path) -> list[list[object]]:
        rows: list[list[object]] = [["采样时间", "设备时间", "射频ID", "链路状态", "对端名称", "对端MAC", "MR侧RSSI", "归属站点", "归属区间", "归属类型", "归属来源"]]
        if not db_path.exists():
            return rows
        with sqlite3.connect(db_path) as conn:
            try:
                data = conn.execute(
                    """
                    SELECT collector_time, COALESCE(NULLIF(device_time, ''), device_clock, collector_time), radio, link_state,
                           COALESCE(NULLIF(resolved_peer_name, ''), NULLIF(peer_name, ''), peer_mac) AS peer_name,
                           peer_mac, mr_rssi, belong_station, belong_section, belong_type, belonging_source
                    FROM main_link_samples
                    WHERE UPPER(link_state) LIKE 'ACTIVE%'
                    ORDER BY collector_time ASC
                    LIMIT 20000
                    """
                ).fetchall()
            except sqlite3.Error:
                return rows
        rows.extend([list(row) for row in data])
        return rows

    def _mesh_link_detail_rows(self, db_path: Path) -> list[list[object]]:
        rows: list[list[object]] = [["采样时间", "设备时间", "Radio", "状态", "PeerMac", "当前PEER AP名称", "AP MAC", "归属站点", "归属区间", "Peer Radio MAC", "MR RSSI", "BSSID", "Mesh接口", "Online Time"]]
        if not db_path.exists():
            return rows
        try:
            with sqlite3.connect(db_path) as conn:
                data = conn.execute(
                    """
                    SELECT collector_time, COALESCE(NULLIF(device_time, ''), device_clock, collector_time),
                           radio, link_state, peer_mac,
                           COALESCE(NULLIF(resolved_peer_name, ''), NULLIF(peer_name, ''), peer_mac),
                           peer_mac, belong_station, belong_section, peer_mac, mr_rssi, bssid,
                           mesh_interface, online_time
                    FROM main_link_samples
                    ORDER BY collector_time ASC, id ASC
                    LIMIT 20000
                    """
                ).fetchall()
        except sqlite3.Error:
            return rows
        rows.extend([list(row) for row in data])
        return rows

    def _channel_busy_rows(self, db_path: Path) -> list[list[object]]:
        rows: list[list[object]] = [["设备时间", "射频ID", "控制信道", "频宽", "记录间隔", "控制信道繁忙度", "发送繁忙度", "接收繁忙度"]]
        if not db_path.exists():
            return rows
        try:
            with sqlite3.connect(db_path) as conn:
                data = conn.execute(
                    """
                    SELECT device_time, radio, ctl_channel, bandwidth, record_interval, ctl_busy, tx_busy, rx_busy
                    FROM channel_busy_records
                    ORDER BY device_time ASC, id ASC
                    LIMIT 20000
                    """
                ).fetchall()
        except sqlite3.Error:
            return rows
        rows.extend([list(row) for row in data])
        return rows

    def _interface_rate_rows(self, db_path: Path) -> list[list[object]]:
        rows: list[list[object]] = [["设备时间", "方向", "接口", "总PPS", "广播PPS", "组播PPS", "说明"]]
        if not db_path.exists():
            return rows
        try:
            with sqlite3.connect(db_path) as conn:
                data = conn.execute(
                    """
                    SELECT device_time, direction, COALESCE(NULLIF(interface_normalized, ''), interface_name),
                           total_pps, broadcast_pps, multicast_pps, ''
                    FROM interface_rate_samples
                    ORDER BY device_time ASC, id ASC
                    LIMIT 20000
                    """
                ).fetchall()
        except sqlite3.Error:
            return rows
        rows.extend([list(row) for row in data])
        return rows

    def _fping_1s_rows(self, db_path: Path) -> list[list[object]]:
        rows: list[list[object]] = [["时间", "设备对齐时间", "本地时间", "时间偏移(ms)", "目标IP", "目标名称", "发送数", "接收数", "丢失数", "丢包率(%)", "平均延迟(ms)", "最小延迟(ms)", "最大延迟(ms)", "Jitter(ms)", "状态"]]
        if not db_path.exists():
            return rows
        try:
            with sqlite3.connect(db_path) as conn:
                data = conn.execute(
                    """
                    SELECT COALESCE(NULLIF(device_bucket_time, ''), NULLIF(bucket_time, ''), local_bucket_time),
                           COALESCE(NULLIF(device_bucket_time, ''), ''),
                           COALESCE(NULLIF(local_bucket_time, ''), NULLIF(bucket_time, ''), ''),
                           clock_offset_ms,
                           target_ip, COALESCE(target_name, ''), sent, received,
                           COALESCE(lost, sent - received), loss_percent, avg_latency_ms,
                           min_latency_ms, max_latency_ms, jitter_ms, COALESCE(status, '')
                    FROM fping_1s_summary
                    ORDER BY COALESCE(NULLIF(device_bucket_time, ''), NULLIF(bucket_time, ''), local_bucket_time) ASC, target_ip ASC
                    LIMIT 20000
                    """
                ).fetchall()
        except sqlite3.Error:
            return rows
        for row in data:
            values = list(row)
            values[-1] = _fping_status_label(values[-1])
            rows.append(values)
        return rows

    def _fping_quality_rows(self, db_path: Path) -> list[list[object]]:
        rows: list[list[object]] = [["本地时间", "设备对齐时间", "时间偏移(ms)", "目标IP", "序号", "状态", "延迟(ms)", "发送", "接收", "丢包", "丢包率(%)", "平均延迟(ms)", "最大延迟(ms)", "原始记录"]]
        if not db_path.exists():
            return rows
        try:
            with sqlite3.connect(db_path) as conn:
                samples = conn.execute(
                    """
                    SELECT COALESCE(NULLIF(p.local_time, ''), p.collector_time),
                           p.device_aligned_time,
                           p.clock_offset_ms,
                           p.target_ip, p.seq, p.success, p.latency_ms,
                           s.sent, s.received, s.sent - s.received, s.loss_percent,
                           s.avg_latency_ms, s.max_latency_ms,
                           p.status
                    FROM fping_samples p
                    LEFT JOIN fping_1s_summary s ON s.target_ip = p.target_ip
                       AND COALESCE(NULLIF(s.device_bucket_time, ''), s.bucket_time, s.local_bucket_time)
                           = substr(replace(COALESCE(NULLIF(p.device_aligned_time, ''), p.collector_time), 'T', ' '), 1, 19)
                    ORDER BY COALESCE(NULLIF(p.device_aligned_time, ''), p.collector_time) ASC, p.seq ASC
                    LIMIT 20000
                    """
                ).fetchall()
        except sqlite3.Error:
            return rows
        for sample in samples:
            rows.append(
                [
                    sample[0],
                    sample[1],
                    sample[2],
                    sample[3],
                    sample[4],
                    "成功" if int(sample[5] or 0) else "超时",
                    sample[6],
                    sample[7],
                    sample[8],
                    sample[9],
                    sample[10],
                    sample[11],
                    sample[12],
                    sample[13],
                ]
            )
        return rows

    def _switch_history_rows(self, db_path: Path) -> list[list[object]]:
        rows: list[list[object]] = [["切换时间", "射频ID", "原对端名称", "新对端名称", "原AP MAC", "新AP MAC", "原归属站点", "新归属站点", "原归属区间", "新归属区间", "切换原因", "入RSSI", "出RSSI", "Active持续时间"]]
        if not db_path.exists():
            return rows
        try:
            with sqlite3.connect(db_path) as conn:
                events = conn.execute(
                    """
                    SELECT event_time_local, radio, old_peer_name, new_peer_name, old_peer_mac, new_peer_mac,
                           old_belong_station, new_belong_station, old_belong_section, new_belong_section,
                           switch_reason_text, old_rssi, new_rssi, active_duration
                    FROM switch_history_events
                    ORDER BY event_time_local ASC, id ASC
                    LIMIT 20000
                    """
                ).fetchall()
        except sqlite3.Error:
            return rows
        for item in events:
            rows.append(
                [
                    item[0],
                    item[1],
                    item[2] or "-",
                    item[3] or "-",
                    item[4] or "-",
                    item[5] or "-",
                    item[6] or "-",
                    item[7] or "-",
                    item[8] or "-",
                    item[9] or "-",
                    item[10] or "-",
                    item[11] or "-",
                    item[12] or "-",
                    item[13] or "-",
                ]
            )
        return rows

    def _active_switch_log_rows(self, db_path: Path) -> list[list[object]]:
        rows: list[list[object]] = [["设备时间", "设备名称", "原AP名称", "原AP MAC", "原RSSI", "原归属站点", "原归属区间", "新AP名称", "新AP MAC", "新RSSI", "新归属站点", "新归属区间", "Peer数量", "Link数量", "切换原因码", "切换原因"]]
        if not db_path.exists():
            return rows
        try:
            with sqlite3.connect(db_path) as conn:
                data = conn.execute(
                    """
                    SELECT device_time, device_name,
                           old_peer_name, old_peer_mac, old_rssi, old_belong_station, old_belong_section,
                           new_peer_name, new_peer_mac, new_rssi, new_belong_station, new_belong_section,
                           peer_quantity, link_quantity, switch_reason_code, switch_reason_text
                    FROM switch_realtime_events
                    ORDER BY device_time ASC, id ASC
                    LIMIT 20000
                    """
                ).fetchall()
        except sqlite3.Error:
            return rows
        rows.extend([list(row) for row in data])
        return rows

    def _session_info_rows(self, session_dir: Path, db_path: Path) -> list[list[object]]:
        meta_path = session_dir / "session_meta.json"
        rows: list[list[object]] = [["字段", "值"]]
        if meta_path.exists():
            import json

            try:
                meta = json.loads(meta_path.read_text(encoding="utf-8"))
            except json.JSONDecodeError:
                meta = {}
            for key in ("session_id", "mr_name", "device_name", "host", "started_at", "ended_at", "status"):
                rows.append([key, meta.get(key) or "N/A"])
        if db_path.exists():
            with sqlite3.connect(db_path) as conn:
                table_counts = []
                for table in ("main_link_samples", "channel_busy_records", "radio_statistics_samples", "switch_history_events", "switch_realtime_events", "time_sync_samples", "fping_samples", "interface_rate_samples"):
                    try:
                        count = conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
                    except sqlite3.Error:
                        count = "N/A"
                    table_counts.append([table, count])
            rows.append(["parsed_db", str(db_path)])
            rows.extend(table_counts)
        return rows

    def _write_overview(self, sheet, db_path: Path) -> None:
        stats = {
            "设备数量": 0,
            "主链路切换次数": 0,
            "空链路次数": 0,
            "平均RSSI": None,
            "最低RSSI": None,
            "平均信道繁忙度": None,
            "Ping平均丢包率": None,
            "Ping平均延迟": None,
            "时间同步状态": "未建立",
            "平均偏移(ms)": None,
            "最大偏移(ms)": None,
            "最小偏移(ms)": None,
            "fping对齐模式": "local_fallback",
        }
        if db_path.exists():
            with sqlite3.connect(db_path) as conn:
                stats["主链路切换次数"] = conn.execute("SELECT COUNT(*) FROM switch_realtime_events").fetchone()[0]
                stats["空链路次数"] = conn.execute("SELECT COUNT(*) FROM switch_realtime_events WHERE old_peer_mac IS NULL OR new_peer_mac IS NULL").fetchone()[0]
                rssi = conn.execute("SELECT AVG(mr_rssi), MIN(mr_rssi) FROM main_link_samples WHERE UPPER(link_state) LIKE 'ACTIVE%' AND mr_rssi IS NOT NULL").fetchone()
                stats["平均RSSI"], stats["最低RSSI"] = rssi
                busy = conn.execute("SELECT AVG(tx_busy), AVG(rx_busy) FROM channel_busy_records").fetchone()
                if busy and busy[0] is not None and busy[1] is not None:
                    stats["平均信道繁忙度"] = round((float(busy[0]) + float(busy[1])) / 2, 2)
                ping = conn.execute("SELECT AVG(loss_percent), AVG(avg_latency_ms) FROM fping_1s_summary").fetchone()
                stats["Ping平均丢包率"], stats["Ping平均延迟"] = ping
                sync = conn.execute("SELECT COUNT(*), AVG(offset_ms), MAX(offset_ms), MIN(offset_ms) FROM time_sync_samples").fetchone()
                if sync and int(sync[0] or 0) > 0:
                    stats["时间同步状态"] = "已建立"
                    stats["平均偏移(ms)"] = sync[1]
                    stats["最大偏移(ms)"] = sync[2]
                    stats["最小偏移(ms)"] = sync[3]
                    aligned = conn.execute("SELECT COUNT(*) FROM fping_1s_summary WHERE device_bucket_time IS NOT NULL AND device_bucket_time <> ''").fetchone()
                    stats["fping对齐模式"] = "device_aligned" if aligned and int(aligned[0] or 0) > 0 else "local_fallback"
        sheet.append(["指标", "值"])
        for key, value in stats.items():
            sheet.append([key, "" if value is None else value])

    def _write_chart_data(self, sheet, chart_data: ChartData) -> None:
        headers = ["时间/类别"] + [series.name for series in chart_data.series]
        sheet.append(headers)
        x_values: list[object] = []
        for series in chart_data.series:
            for x_value, _y_value in series.points:
                if x_value not in x_values:
                    x_values.append(x_value)
        values_by_series = [{x: y for x, y in series.points} for series in chart_data.series]
        for x_value in x_values:
            sheet.append([x_value] + [values.get(x_value) for values in values_by_series])

    def _write_chart_sheet(self, sheet, data_sheet, chart_data: ChartData, chart_type, title: str) -> None:
        from openpyxl.chart import Reference
        from openpyxl.styles import Font

        sheet["A1"] = title
        sheet["A1"].font = Font(bold=True, size=14)
        if data_sheet.max_row < 2 or data_sheet.max_column < 2:
            sheet["A3"] = chart_data.empty_message
            return
        chart = chart_type()
        chart.title = title
        chart.y_axis.title = chart_data.y_label
        data = Reference(data_sheet, min_col=2, max_col=data_sheet.max_column, min_row=1, max_row=data_sheet.max_row)
        categories = Reference(data_sheet, min_col=1, min_row=2, max_row=data_sheet.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 14
        chart.width = 28
        sheet.add_chart(chart, "A3")


def _fping_status_label(value: object) -> str:
    text = str(value or "").strip().lower()
    if text in {"ok", "success", "normal"}:
        return "正常"
    if text in {"loss", "lost"}:
        return "丢包"
    if text in {"timeout", "time_out"}:
        return "超时"
    if text in {"no_data", "nodata"}:
        return "无数据"
    if text in {"error", "failed", "fail"}:
        return "错误"
    return str(value or "-") or "-"
