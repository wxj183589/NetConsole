from __future__ import annotations

import re
from copy import copy as copy_style
from collections.abc import Callable, Mapping
from datetime import datetime
from pathlib import Path
from typing import Any

from netconsole.core.paths import PathResolver
from netconsole.core.version import APP_VERSION_DISPLAY
from netconsole.services.ac.fit_ap_export_contract import FIT_AP_RESOURCE_EXPORT_SCHEMA_VERSION
from netconsole.services.ac.fit_ap_resource_identity import coalesce_fit_ap_resource_rows
from netconsole.services.ac.query_service import AcManagementQueryService, fit_ap_topology_sort_key
from netconsole.services.ap_extension_import import normalize_ap_mac
from netconsole.services.excel_autosize import apply_worksheet_column_widths
from netconsole.services.export.common_exporters import ExportCancelled
from netconsole.services.export.xlsx_style import apply_basic_sheet_style

ProgressCallback = Callable[[str, int, int, str], None]
CancelCallback = Callable[[], bool]

_INVALID_FILENAME = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
_FILTER_LABELS = {
    "query": "AP名称/IP/MAC",
    "status": "AP状态",
    "optical_status": "光衰状态",
    "station": "归属站点",
    "section": "归属区间",
    "model": "型号",
    "switch": "交换机",
}
_STATUS_LABELS = {
    "online": "在线",
    "offline": "离线",
    "unauthenticated": "未认证",
    "normal": "正常",
    "warning": "一般告警",
    "critical": "严重告警",
    "no_data": "无数据",
}

AP_COLUMNS = (
    ("seq", "序号"),
    ("site_name", "局点名称"),
    ("ac_name", "AC名称"),
    ("ac_ip", "AC管理地址"),
    ("ac_model", "AC型号"),
    ("ac_version", "AC软件版本"),
    ("ap_name", "AP名称"),
    ("ap_ip", "AP IP"),
    ("ap_mac", "AP MAC"),
    ("ap_model", "AP型号"),
    ("serial_number", "AP序列号"),
    ("ap_state", "AP状态"),
    ("ap_hardware_version", "AP硬件版本"),
    ("ap_software_version", "AP软件版本"),
    ("ap_boot_version", "AP Boot版本"),
    ("detail_updated_at", "详细信息更新时间"),
    ("online_status", "在线状态"),
    ("radio_count", "Radio数量"),
    ("radio1_status", "Mesh Radio 1状态"),
    ("radio1_channel", "Mesh Radio 1信道"),
    ("radio1_power", "Mesh Radio 1功率"),
    ("radio2_status", "Mesh Radio 2状态"),
    ("radio2_channel", "Mesh Radio 2信道"),
    ("radio2_power", "Mesh Radio 2功率"),
    ("station", "归属站点"),
    ("section", "归属区间"),
    ("point_code", "点位编号"),
    ("trackside_ap_name", "轨旁AP名称"),
    ("switch_name", "连接交换机"),
    ("switch_ip", "交换机管理地址"),
    ("switch_interface", "连接端口"),
    ("pvid", "端口PVID"),
    ("lldp_status", "LLDP状态"),
    ("lldp_neighbor", "LLDP邻居名称"),
    ("lldp_neighbor_interface", "LLDP邻居端口"),
    ("optical_status", "光衰状态"),
    ("rx_power", "接收光功率（dBm）"),
    ("tx_power", "发送光功率（dBm）"),
    ("optical_updated_at", "光衰采集时间"),
    ("ap_updated_at", "AP资源更新时间"),
    ("lldp_updated_at", "LLDP更新时间"),
    ("radio_updated_at", "Radio更新时间"),
    ("integrity", "数据完整性"),
    ("remark", "备注"),
)

RADIO_COLUMNS = (
    ("seq", "序号"),
    ("ac_name", "AC名称"),
    ("ac_ip", "AC管理地址"),
    ("ap_name", "AP名称"),
    ("ap_ip", "AP IP"),
    ("ap_mac", "AP MAC"),
    ("ap_model", "AP型号"),
    ("radio_id", "Radio ID"),
    ("radio_name", "Radio名称"),
    ("radio_type", "Radio类型"),
    ("band", "射频频段"),
    ("status", "工作状态"),
    ("admin_status", "管理状态"),
    ("channel", "信道"),
    ("bandwidth", "频宽"),
    ("tx_power", "发射功率"),
    ("mesh_role", "Mesh角色"),
    ("bssid", "BSSID"),
    ("radio_mac", "Radio MAC"),
    ("updated_at", "更新时间"),
    ("source", "数据来源"),
    ("remark", "备注"),
)


def make_fit_ap_resource_filename(site_name: str, ac_name: str, now: datetime | None = None) -> str:
    timestamp = (now or datetime.now()).strftime("%Y%m%d_%H%M%S")
    parts = [_safe_filename_part(value) for value in (site_name, ac_name)]
    stem = f"FIT-AP资源_{parts[0]}_{parts[1]}_{timestamp}".strip("_")
    return f"{stem[:170]}.xlsx"


def export_fit_ap_resource_xlsx(
    path: Path,
    payload: Mapping[str, Any],
    progress: ProgressCallback | None = None,
    should_cancel: CancelCallback | None = None,
) -> dict[str, Any]:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, PatternFill

    resolver = PathResolver(
        app_root=_optional_path(payload.get("app_root")),
        data_root=_optional_path(payload.get("data_root")),
    )
    site_name = str(payload.get("site_name") or "")
    ac_id = str(payload.get("ac_uuid") or "")
    filters = dict(payload.get("filters") or {})
    selected_ids = [str(value) for value in payload.get("selected_ap_ids") or [] if str(value)]
    query = AcManagementQueryService(resolver)
    _emit(progress, "query_snapshot", 0, 1, "正在读取 FIT-AP 资源快照")
    _check_cancel(should_cancel)
    ac = query.get_ac_export_identity(site_name, ac_id)
    if ac is None:
        raise ValueError("当前 AC 不存在")
    details = query.list_ap_details_for_export(
        site_name,
        ac_id=ac_id,
        filters=filters,
        selected_ap_ids=selected_ids,
    )
    fixed_ids = {item.ap.id for item in details}
    if selected_ids and fixed_ids != set(selected_ids):
        raise ValueError("已选择 AP 不属于当前 AC")
    if not details:
        raise ValueError("当前范围内没有可导出的 FIT AP")
    details = _deduplicate_details(details, ac_id)
    details.sort(key=lambda detail: fit_ap_topology_sort_key(detail.ap))

    _emit(progress, "build_rows", 0, len(details), "正在整理 AP 与 Radio 数据")
    ap_rows: list[dict[str, object]] = []
    radio_rows: list[dict[str, object]] = []
    warning_count = 0
    for index, detail in enumerate(details, start=1):
        _check_cancel(should_cancel)
        ap = detail.ap
        mac = _display_mac(ap.mac)
        issues = _integrity_issues(detail, mac)
        if issues:
            warning_count += 1
        radio_updated_at = _latest_time(*(radio.updated_at for radio in detail.radios))
        ap_rows.append(
            {
                "seq": index,
                "site_name": site_name,
                "ac_name": ac.name,
                "ac_ip": ac.management_ip,
                "ac_model": ac.model,
                "ac_version": ac.software_version,
                "ap_name": ap.name,
                "ap_ip": ap.ip,
                "ap_mac": mac,
                "ap_model": ap.model,
                "serial_number": ap.serial_number,
                "ap_hardware_version": ap.hardware_version,
                "ap_software_version": ap.software_version,
                "ap_boot_version": ap.boot_version,
                "detail_updated_at": _format_time(ap.detail_updated_at),
                "ap_state": ap.state_display,
                "online_status": _status_label(ap.status),
                "radio_count": len(detail.radios),
                "radio1_status": _radio_value(detail.radios, 1, "status"),
                "radio1_channel": _radio_value(detail.radios, 1, "channel"),
                "radio1_power": _radio_value(detail.radios, 1, "tx_power"),
                "radio2_status": _radio_value(detail.radios, 2, "status"),
                "radio2_channel": _radio_value(detail.radios, 2, "channel"),
                "radio2_power": _radio_value(detail.radios, 2, "tx_power"),
                "station": ap.station,
                "section": ap.section,
                "point_code": ap.point_code,
                "trackside_ap_name": ap.trackside_ap_name,
                "switch_name": detail.lldp.switch_name,
                "switch_ip": detail.lldp.switch_ip,
                "switch_interface": detail.lldp.interface_name,
                "pvid": detail.lldp.vlan,
                "lldp_status": detail.lldp.match_status,
                "lldp_neighbor": detail.lldp.lldp_neighbor,
                "lldp_neighbor_interface": detail.lldp.interface_name,
                "optical_status": _status_label(detail.optical.optical_status),
                "rx_power": _power_text(detail.optical.rx_power or detail.optical.switch_rx_power),
                "tx_power": _power_text(detail.optical.tx_power),
                "optical_updated_at": _format_time(detail.optical.updated_at),
                "ap_updated_at": _format_time(ap.updated_at),
                "lldp_updated_at": _format_time(detail.lldp.updated_at),
                "radio_updated_at": _format_time(radio_updated_at),
                "integrity": "、".join(issues) if issues else "完整",
                "remark": ap.remark,
            }
        )
        for radio in detail.radios:
            radio_rows.append(
                {
                    "_ap_order": index,
                    "seq": len(radio_rows) + 1,
                    "ac_name": ac.name,
                    "ac_ip": ac.management_ip,
                    "ap_name": ap.name,
                    "ap_ip": ap.ip,
                    "ap_mac": mac,
                    "ap_model": ap.model,
                    "radio_id": radio.radio_id,
                    "radio_name": f"Radio {radio.radio_id}",
                    "radio_type": radio.mode,
                    "band": radio.band,
                    "status": radio.status,
                    "admin_status": "",
                    "channel": radio.channel,
                    "bandwidth": radio.bandwidth,
                    "tx_power": radio.tx_power,
                    "mesh_role": "",
                    "bssid": _display_mac(radio.bssid),
                    "radio_mac": "",
                    "updated_at": _format_time(radio.updated_at),
                    "source": "SQLite 已采集数据",
                    "remark": "",
                }
            )
        _emit(progress, "build_rows", index, len(details), f"正在整理 AP {index}/{len(details)}")

    radio_rows.sort(key=lambda row: (int(row.get("_ap_order") or 0), int(row["radio_id"])))
    for index, row in enumerate(radio_rows, start=1):
        row["seq"] = index
        row.pop("_ap_order", None)
    export_time = _format_time(str(payload.get("requested_at") or "")) or datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    scope_text = _scope_text(str(payload.get("scope") or "all"), len(ap_rows))
    data_times = [str(row["ap_updated_at"]) for row in ap_rows if row["ap_updated_at"]]
    instructions = _instruction_rows(
        file_name=path.name,
        export_time=export_time,
        site_name=site_name,
        ac_name=ac.name,
        ac_ip=ac.management_ip,
        scope_text=scope_text,
        filters=filters,
        ap_rows=ap_rows,
        radio_count=len(radio_rows),
        data_times=data_times,
    )

    workbook = Workbook()
    workbook.remove(workbook.active)
    ap_sheet = _write_sheet(workbook, "AP资源清单", AP_COLUMNS, ap_rows)
    radio_sheet = _write_sheet(workbook, "Radio明细", RADIO_COLUMNS, radio_rows)
    instruction_sheet = _write_sheet(workbook, "导出说明", (("field", "字段"), ("value", "值")), instructions)
    text_headers = {
        "AC管理地址", "AC软件版本", "AP名称", "AP IP", "AP MAC", "AP序列号", "AP硬件版本",
        "AP软件版本", "AP Boot版本", "详细信息更新时间", "点位编号",
        "连接端口", "端口PVID", "BSSID", "Radio MAC", "信道",
    }
    for sheet in (ap_sheet, radio_sheet):
        header_index = {cell.value: cell.column for cell in sheet[1]}
        for header in text_headers:
            column = header_index.get(header)
            if column:
                for cell in sheet.iter_cols(min_col=column, max_col=column, min_row=2):
                    for value in cell:
                        value.number_format = "@"
    for sheet, headers in (
        (ap_sheet, {"数据完整性", "备注"}),
        (radio_sheet, {"备注"}),
        (instruction_sheet, {"值"}),
    ):
        header_index = {cell.value: cell.column for cell in sheet[1]}
        for header in headers:
            column = header_index.get(header)
            if not column:
                continue
            for row in sheet.iter_rows(min_row=2, min_col=column, max_col=column):
                cell = row[0]
                alignment = copy_style(cell.alignment)
                alignment.wrap_text = True
                alignment.vertical = alignment.vertical or "top"
                cell.alignment = alignment
    status_fills = {
        "正常": "E2F0D9",
        "一般告警": "FFF2CC",
        "严重告警": "F4CCCC",
        "离线": "E7E6E6",
        "无数据": "D9EAF7",
    }
    for row in ap_sheet.iter_rows(min_row=2):
        for cell in row:
            label = str(cell.value or "")
            if label in status_fills:
                cell.fill = PatternFill(fill_type="solid", fgColor=status_fills[label])
            if cell.column in {AP_COLUMNS.index(("integrity", "数据完整性")) + 1, AP_COLUMNS.index(("remark", "备注")) + 1}:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
    _emit(progress, "save_workbook", len(ap_rows), len(ap_rows), "正在保存 FIT-AP 资源工作簿")
    _check_cancel(should_cancel)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return {
        "row_count": len(ap_rows),
        "ap_count": len(ap_rows),
        "radio_count": len(radio_rows),
        "warning_count": warning_count,
    }


def _write_sheet(workbook, name: str, columns, rows: list[dict[str, object]]):
    sheet = workbook.create_sheet(name)
    sheet.append([title for _key, title in columns])
    for row in rows:
        sheet.append([row.get(key, "") for key, _title in columns])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    apply_basic_sheet_style(sheet, header_row=1, column_count=len(columns))
    apply_worksheet_column_widths(
        sheet,
        [title for _key, title in columns],
        rows,
        [key for key, _title in columns],
        maximum=48,
    )
    return sheet


def _instruction_rows(
    *,
    file_name: str,
    export_time: str,
    site_name: str,
    ac_name: str,
    ac_ip: str,
    scope_text: str,
    filters: dict[str, object],
    ap_rows: list[dict[str, object]],
    radio_count: int,
    data_times: list[str],
) -> list[dict[str, object]]:
    online = sum(row["online_status"] == "在线" for row in ap_rows)
    offline = sum(row["online_status"] == "离线" for row in ap_rows)
    optical_normal = sum(row["optical_status"] == "正常" for row in ap_rows)
    optical_abnormal = sum(row["optical_status"] in {"一般告警", "严重告警"} for row in ap_rows)
    lldp_normal = sum(bool(row["lldp_status"]) and "冲突" not in str(row["lldp_status"]) for row in ap_rows)
    rows = [
        ("文件名称", file_name),
        ("导出时间", export_time),
        ("局点名称", site_name),
        ("AC名称", ac_name),
        ("AC管理地址", ac_ip),
        ("导出范围", scope_text),
        ("页面筛选条件", _filter_text(filters)),
        ("AP总数", len(ap_rows)),
        ("Radio总数", radio_count),
        ("在线AP数量", online),
        ("离线AP数量", offline),
        ("光衰正常数量", optical_normal),
        ("光衰异常数量", optical_abnormal),
        ("LLDP正常数量", lldp_normal),
        ("未关联站点数量", sum(not row["station"] for row in ap_rows)),
        ("数据时间范围", f"{min(data_times)} 至 {max(data_times)}" if data_times else ""),
        ("软件版本", APP_VERSION_DISPLAY),
        ("字段说明", "AP资源清单一台 AP 一行；Radio明细每个持久化 Radio 一行。"),
        ("数据缺失说明", "仅导出已持久化数据；缺失字段保持空白并在“数据完整性”中标识。"),
        ("数据来源", "SQLite 已采集资源快照；导出过程不会连接 AC、AP 或交换机。"),
    ]
    return [{"field": field, "value": value} for field, value in rows]


def _integrity_issues(detail, mac: str) -> list[str]:
    issues: list[str] = []
    if not detail.ap.detail_available:
        issues.append("未采集AP详细信息")
    if detail.optical.optical_status == "no_data":
        issues.append("缺少光衰")
    if not detail.lldp.switch_name and not detail.lldp.interface_name:
        issues.append("缺少LLDP")
    if not detail.ap.station:
        issues.append("未关联站点")
    if not detail.ap.section:
        issues.append("未关联区间")
    if not mac:
        issues.append("缺少AP MAC")
    if not detail.ap.serial_number:
        issues.append("缺少AP序列号")
    if not detail.radios:
        issues.append("缺少Radio信息")
    return issues


def _deduplicate_details(details, ac_id: str):
    projected = []
    for index, detail in enumerate(details):
        projected.append(
            {
                "ac_device_uuid": ac_id,
                "ap_name": detail.ap.name,
                "ap_ip": detail.ap.ip,
                "ap_mac": detail.ap.mac,
                "serial_number": detail.ap.serial_number,
                "model": detail.ap.model,
                "state": detail.ap.status,
                "state_display": detail.ap.state_display,
                "lldp_neighbor_name": detail.lldp.switch_name,
                "lldp_neighbor_interface": detail.lldp.interface_name,
                "optical_updated_at": detail.optical.updated_at,
                **{
                    f"rid{radio.radio_id}_status": radio.status
                    for radio in detail.radios
                },
                "_detail_index": index,
            }
        )
    deduplicated = coalesce_fit_ap_resource_rows(projected)
    return [details[int(item["_detail_index"])] for item in deduplicated]


def _radio_value(radios, radio_id: int, field: str) -> str:
    radio = next((item for item in radios if item.radio_id == radio_id), None)
    return str(getattr(radio, field, "") or "") if radio else ""


def _scope_text(scope: str, count: int) -> str:
    if scope == "filtered":
        return "当前筛选结果"
    if scope == "selected":
        return f"已选择 AP，共 {count} 台"
    return "当前 AC 全部 AP"


def _filter_text(filters: dict[str, object]) -> str:
    values = []
    for key, label in _FILTER_LABELS.items():
        value = str(filters.get(key) or "").strip()
        if value:
            values.append(f"{label}：{_status_label(value)}")
    return "\n".join(values)


def _status_label(value: str) -> str:
    return _STATUS_LABELS.get(str(value or ""), str(value or ""))


def _display_mac(value: object) -> str:
    result = normalize_ap_mac(value)
    return result.display.casefold() if result.valid else ""


def _power_text(value: object) -> str:
    return re.sub(r"\s*dBm\s*$", "", str(value or "").strip(), flags=re.IGNORECASE)


def _format_time(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).astimezone().strftime("%Y-%m-%d %H:%M:%S")
    except ValueError:
        return text[:19].replace("T", " ")


def _latest_time(*values: object) -> str:
    clean = [str(value or "").strip() for value in values if str(value or "").strip()]
    return max(clean) if clean else ""


def _safe_filename_part(value: object) -> str:
    text = _INVALID_FILENAME.sub("_", str(value or "").strip()).rstrip(" .")
    return text[:60] or "未命名"


def _optional_path(value: object) -> Path | None:
    text = str(value or "").strip()
    return Path(text) if text else None


def _emit(progress: ProgressCallback | None, stage: str, current: int, total: int, message: str) -> None:
    if progress:
        progress(stage, current, max(total, 1), message)


def _check_cancel(should_cancel: CancelCallback | None) -> None:
    if should_cancel and should_cancel():
        raise ExportCancelled("导出已取消")


__all__ = [
    "AP_COLUMNS",
    "FIT_AP_RESOURCE_EXPORT_SCHEMA_VERSION",
    "RADIO_COLUMNS",
    "export_fit_ap_resource_xlsx",
    "make_fit_ap_resource_filename",
]
