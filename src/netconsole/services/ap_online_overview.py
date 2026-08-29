from __future__ import annotations

from collections import Counter
from pathlib import Path
import re
import unicodedata

from netconsole.core import app_logger
from netconsole.parsers.h3c.ac.state_mapper import classify_fit_ap_state
from netconsole.services.ap_identity.normalizers import normalize_mac_key
from netconsole.utils.station_normalize import normalize_station_value


AP_ONLINE_OVERVIEW_COLUMNS = (
    ("ac.station", "site"),
    ("trackside.planned_ap_count", "total"),
    ("ac.online", "online"),
    ("ac.offline", "offline"),
    ("ac.online_rate", "online_rate"),
    ("trackside.reonline_count", "reonline_count"),
    ("trackside.reonline_rate", "reonline_rate"),
    ("trackside.optical_problem_count", "optical_problem_count"),
    ("field.remark", "remark"),
)

TOTAL_SITE_LABEL = "合计"
UNASSIGNED_SITE_LABEL = "未归属"
DIRTY_RESOURCE_SITES = {"Demo", "体育中心站"}


class ApOnlineOverviewService:
    @staticmethod
    def build_rows(
        *,
        planned_aps: list[dict[str, object | None]] | None = None,
        metadata_rows: list[dict[str, object | None]] | None = None,
        fit_ap_resources: list[dict[str, object | None]],
        optical_rows: list[dict[str, object | None]] | None = None,
        capacity_details: dict[str, object] | None = None,
    ) -> list[dict[str, object | None]]:
        return build_ap_online_overview_rows(
            metadata_rows=metadata_rows if metadata_rows is not None else planned_aps or [],
            fit_ap_resources=fit_ap_resources,
            optical_rows=optical_rows,
            capacity_details=capacity_details,
        )


def build_ap_online_overview_rows(
    metadata_rows: list[dict[str, object | None]] | None = None,
    fit_ap_resources: list[dict[str, object | None]] | None = None,
    optical_rows: list[dict[str, object | None]] | None = None,
    *,
    planned_aps: list[dict[str, object | None]] | None = None,
    capacity_details: dict[str, object] | None = None,
    capacities: dict[str, object] | None = None,
) -> list[dict[str, object | None]]:
    """Build AP online overview from capacity totals, FIT-AP resources, and optical site data."""
    metadata_source = metadata_rows if metadata_rows is not None else planned_aps or []
    resources = fit_ap_resources or []
    optical_source = optical_rows or []
    capacity_map = capacity_details if capacity_details is not None else capacities
    normalized_resources = [normalize_fit_ap_resource(row) for row in resources]
    raw_metadata_count = len(metadata_source)
    normalized_metadata = [normalize_planned_ap(row) for row in metadata_source if _metadata_row_has_data(row)]
    normalized_optical = [normalize_optical_row(row) for row in optical_source]
    if not normalized_metadata and not capacity_map and normalized_resources and not normalized_optical:
        normalized_metadata = [normalize_planned_ap(row) for row in resources if _resource_site(row) not in DIRTY_RESOURCE_SITES]
    metadata_indexes = build_plan_index(normalized_metadata)
    optical_indexes = build_optical_index(normalized_optical)
    grouped = build_station_totals(normalized_metadata, capacity_map)
    capacity_sites = {str(site or "").strip() for site in (capacity_map or {}) if str(site or "").strip()}
    stats = build_online_counts(normalized_resources, optical_indexes, metadata_indexes, grouped, capacity_sites)
    for site, count in _station_optical_problem_counts(normalized_optical).items():
        item = grouped.setdefault(site, {"site": site, "total": 0, "online": 0, "offline": 0, "remark": ""})
        item["optical_problem_count"] = count
    for site, count in (stats.get("station_reonline_counts") or {}).items():
        item = grouped.setdefault(site, {"site": site, "total": 0, "online": 0, "offline": 0, "remark": ""})
        item["reonline_count"] = count
    rows = build_rows(grouped)
    _log_overview_diagnostics(normalized_metadata, raw_metadata_count, normalized_resources, normalized_optical, capacity_map or {}, rows, stats)
    return rows


def normalize_planned_ap(row: dict[str, object | None]) -> dict[str, object | None]:
    result = dict(row)
    result["_match_mac"] = _row_mac(row)
    result["_match_name"] = _row_name(row)
    result["_match_name_compact"] = _row_name(row, compact=True)
    result["_match_uuid"] = _first_text(row, "ap_uuid", "uuid", "AP_UUID")
    result["_match_serial"] = _first_text(row, "serial_number", "serial", "SN")
    result["_match_apid"] = _first_text(row, "apid", "APID")
    result["_station"] = _plan_site(row)
    return result


def normalize_fit_ap_resource(row: dict[str, object | None]) -> dict[str, object | None]:
    result = dict(row)
    result["_match_mac"] = _row_mac(row)
    result["_match_name"] = _row_name(row)
    result["_match_name_compact"] = _row_name(row, compact=True)
    result["_match_uuid"] = _first_text(row, "ap_uuid", "uuid", "AP_UUID")
    result["_match_serial"] = _first_text(row, "serial_number", "serial", "SN")
    result["_match_apid"] = _first_text(row, "apid", "APID")
    result["_resource_site"] = _resource_site(row)
    result["_online"] = is_fit_ap_online(row)
    return result


def normalize_optical_row(row: dict[str, object | None]) -> dict[str, object | None]:
    result = dict(row)
    result["_match_mac"] = _row_mac(row)
    result["_match_name"] = _row_name(row)
    result["_match_name_compact"] = _row_name(row, compact=True)
    result["_match_uuid"] = _first_text(row, "ap_uuid", "uuid", "AP_UUID")
    result["_station"] = normalize_station_value(row)
    return result


def build_plan_index(planned_aps: list[dict[str, object | None]]) -> dict[str, object]:
    indexes: dict[str, object] = {
        "plans": [],
        "by_mac": {},
        "by_name": {},
        "by_name_compact": {},
        "by_uuid": {},
        "by_serial": {},
        "by_apid": {},
    }
    seen: set[str] = set()
    for plan in planned_aps:
        key = _plan_unique_key(plan)
        if key and key in seen:
            continue
        if key:
            seen.add(key)
        indexes["plans"].append(plan)  # type: ignore[index]
        _index_value(indexes["by_mac"], plan.get("_match_mac"), plan)
        _index_value(indexes["by_name"], plan.get("_match_name"), plan)
        _index_value(indexes["by_name_compact"], plan.get("_match_name_compact"), plan)
        _index_value(indexes["by_uuid"], plan.get("_match_uuid"), plan)
        _index_value(indexes["by_serial"], plan.get("_match_serial"), plan)
        _index_value(indexes["by_apid"], plan.get("_match_apid"), plan)
    return indexes


def build_optical_index(optical_rows: list[dict[str, object | None]]) -> dict[str, object]:
    indexes: dict[str, object] = {
        "by_uuid": {},
        "by_mac": {},
        "by_name": {},
        "by_name_compact": {},
    }
    for row in optical_rows:
        if not str(row.get("_station") or "").strip():
            continue
        _index_value(indexes["by_uuid"], row.get("_match_uuid"), row)
        _index_value(indexes["by_mac"], row.get("_match_mac"), row)
        _index_value(indexes["by_name"], row.get("_match_name"), row)
        _index_value(indexes["by_name_compact"], row.get("_match_name_compact"), row)
    return indexes


def match_resource_to_plan(
    resource: dict[str, object | None],
    indexes: dict[str, object],
) -> tuple[dict[str, object | None] | None, str]:
    checks = (
        ("uuid", resource.get("_match_uuid"), "by_uuid"),
        ("mac", resource.get("_match_mac"), "by_mac"),
        ("serial", resource.get("_match_serial"), "by_serial"),
        ("apid", resource.get("_match_apid"), "by_apid"),
        ("name", resource.get("_match_name"), "by_name"),
        ("name", resource.get("_match_name_compact"), "by_name_compact"),
    )
    for method, value, index_name in checks:
        index = indexes.get(index_name)
        if value and isinstance(index, dict) and value in index:
            return index[value], method
    return None, ""


def match_resource_to_optical(
    resource: dict[str, object | None],
    indexes: dict[str, object],
) -> tuple[dict[str, object | None] | None, str]:
    checks = (
        ("optical_uuid", resource.get("_match_uuid"), "by_uuid"),
        ("optical_mac", resource.get("_match_mac"), "by_mac"),
        ("optical_name", resource.get("_match_name"), "by_name"),
        ("optical_name", resource.get("_match_name_compact"), "by_name_compact"),
    )
    for method, value, index_name in checks:
        index = indexes.get(index_name)
        if value and isinstance(index, dict) and value in index:
            return index[value], method
    return None, ""


def build_station_totals(
    planned_aps: list[dict[str, object | None]],
    capacity_details: dict[str, object] | None = None,
) -> dict[str, dict[str, object | None]]:
    grouped: dict[str, dict[str, object | None]] = {}
    capacity_sites = {str(site or "").strip() for site in (capacity_details or {}) if str(site or "").strip()}
    for site, value in (capacity_details or {}).items():
        site_name = str(site or "").strip()
        if not site_name:
            continue
        capacity_total, remark = _capacity_total_remark(value)
        source = value.get("source") if isinstance(value, dict) else ""
        grouped[site_name] = {"site": site_name, "total": capacity_total, "online": 0, "offline": 0, "remark": remark, "source": source}

    seen: set[str] = set()
    for plan in planned_aps:
        key = _plan_unique_key(plan)
        if key and key in seen:
            continue
        if key:
            seen.add(key)
        site = str(plan.get("_station") or UNASSIGNED_SITE_LABEL)
        if site in capacity_sites:
            continue
        item = grouped.setdefault(site, {"site": site, "total": 0, "online": 0, "offline": 0, "remark": ""})
        item["total"] = int(item["total"] or 0) + 1
    return grouped


def build_online_counts(
    fit_ap_resources: list[dict[str, object | None]],
    optical_indexes: dict[str, object],
    metadata_indexes: dict[str, object],
    grouped: dict[str, dict[str, object | None]],
    capacity_sites: set[str],
) -> dict[str, object]:
    stats: dict[str, object] = {
        "online_resource_count": 0,
        "matched_by_optical_uuid": 0,
        "matched_by_optical_mac": 0,
        "matched_by_optical_name": 0,
        "matched_by_metadata": 0,
        "matched_by_known_resource_site": 0,
        "unmatched_online": 0,
        "station_online_counts": {},
        "station_reonline_counts": {},
        "unmatched_sample": [],
    }
    seen_online: set[str] = set()
    for resource in fit_ap_resources:
        if not bool(resource.get("_online")):
            continue
        stats["online_resource_count"] = int(stats["online_resource_count"] or 0) + 1
        optical, optical_method = match_resource_to_optical(resource, optical_indexes)
        metadata, _metadata_method = match_resource_to_plan(resource, metadata_indexes)
        resource_site = str(resource.get("_resource_site") or "").strip()
        if optical and str(optical.get("_station") or "").strip():
            site = str(optical.get("_station") or "")
            online_key = _resource_unique_key(resource)
            stats[f"matched_by_{optical_method}"] = int(stats.get(f"matched_by_{optical_method}") or 0) + 1
        elif metadata and str(metadata.get("_station") or "").strip():
            site = str(metadata.get("_station") or "")
            online_key = _resource_unique_key(resource)
            stats["matched_by_metadata"] = int(stats["matched_by_metadata"] or 0) + 1
        elif resource_site and resource_site in capacity_sites:
            site = resource_site
            online_key = _resource_unique_key(resource)
            stats["matched_by_known_resource_site"] = int(stats["matched_by_known_resource_site"] or 0) + 1
        else:
            stats["unmatched_online"] = int(stats["unmatched_online"] or 0) + 1
            _append_unmatched_sample(stats, resource)
            continue
        if not online_key or online_key in seen_online:
            continue
        seen_online.add(online_key)
        item = grouped.setdefault(site, {"site": site, "total": 0, "online": 0, "offline": 0, "remark": ""})
        item["online"] = int(item["online"] or 0) + 1
        station_counts = stats["station_online_counts"]
        if isinstance(station_counts, dict):
            station_counts[site] = int(station_counts.get(site) or 0) + 1
        if int(resource.get("connection_reonline_count") or 0) > 0:
            reonline_counts = stats["station_reonline_counts"]
            if isinstance(reonline_counts, dict):
                reonline_counts[site] = int(reonline_counts.get(site) or 0) + 1
    return stats


def build_rows(grouped: dict[str, dict[str, object | None]]) -> list[dict[str, object | None]]:
    result: list[dict[str, object | None]] = []
    for row in sorted(grouped.values(), key=lambda item: _site_sort_key(str(item.get("site") or ""))):
        site = str(row.get("site") or "")
        total = int(row.get("total") or 0)
        online = int(row.get("online") or 0)
        normalized = {
            "site": site,
            "total": total,
            "online": online,
            "offline": max(total - online, 0),
            "reonline_count": int(row.get("reonline_count") or 0),
            "reonline_rate": (
                round(int(row.get("reonline_count") or 0) * 100 / total, 1)
                if total > 0
                else None
            ),
            "optical_problem_count": int(row.get("optical_problem_count") or 0),
            "remark": row.get("remark") or "",
        }
        if row.get("source"):
            normalized["source"] = row.get("source") or ""
        result.append(_with_online_rate(normalized))
    total = sum(int(row.get("total") or 0) for row in result)
    online = sum(int(row.get("online") or 0) for row in result)
    offline = sum(int(row.get("offline") or 0) for row in result)
    reonline = sum(int(row.get("reonline_count") or 0) for row in result)
    optical_problem = sum(int(row.get("optical_problem_count") or 0) for row in result)
    total_row = {
        "site": TOTAL_SITE_LABEL,
        "total": total,
        "online": online,
        "offline": offline,
        "reonline_count": reonline,
        "reonline_rate": round(reonline * 100 / total, 1) if total > 0 else None,
        "optical_problem_count": optical_problem,
        "remark": "",
    }
    return [*result, _with_online_rate(total_row)]


def _station_optical_problem_counts(
    optical_rows: list[dict[str, object | None]],
) -> dict[str, int]:
    abnormal_statuses = {"abnormal", "alarm", "warning", "notice", "link_abnormal", "link_down", "no_light"}
    grouped: dict[str, set[str]] = {}
    for row in optical_rows:
        status = str(
            row.get("current_status")
            or row.get("optical_alarm_status")
            or row.get("status")
            or ""
        ).strip().casefold()
        station = str(row.get("_station") or "").strip()
        identity = str(
            row.get("_match_uuid")
            or row.get("_match_mac")
            or row.get("_match_serial")
            or row.get("_match_name")
            or ""
        ).strip()
        if station and identity and status in abnormal_statuses:
            grouped.setdefault(station, set()).add(identity)
    return {station: len(identities) for station, identities in grouped.items()}


def is_fit_ap_online(row: dict[str, object | None]) -> bool:
    return (
        classify_fit_ap_state(
            row.get("state"),
            row.get("state_raw"),
            row.get("state_display"),
        )
        == "online"
    )


def export_ap_online_overview_xlsx(
    path: Path,
    rows: list[dict[str, object | None]],
    headers: list[str],
    offline_ap_stats: dict[str, object | None] | None = None,
    offline_ap_ledger_rows: list[dict[str, object | None]] | None = None,
    offline_ap_stats_headers: list[str] | None = None,
    offline_ap_ledger_headers: list[str] | None = None,
) -> None:
    from openpyxl import Workbook
    from netconsole.services.offline_ap_ledger import (
        OFFLINE_AP_LEDGER_COLUMNS,
        OFFLINE_AP_STATS_COLUMNS,
        write_offline_ap_ledger_sheet,
        write_offline_ap_stats_sheet,
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "AP Online Overview"
    write_ap_online_overview_sheet(sheet, rows, headers)
    if offline_ap_stats is not None and offline_ap_ledger_rows is not None:
        stats_sheet = workbook.create_sheet("AP离线情况")
        write_offline_ap_stats_sheet(stats_sheet, offline_ap_stats, offline_ap_stats_headers or [key for key, _field in OFFLINE_AP_STATS_COLUMNS])
        ledger_sheet = workbook.create_sheet("离线AP台账")
        write_offline_ap_ledger_sheet(ledger_sheet, offline_ap_ledger_rows, offline_ap_ledger_headers or [key for key, _field in OFFLINE_AP_LEDGER_COLUMNS])
    workbook.save(path)


def write_ap_online_overview_sheet(sheet, rows: list[dict[str, object | None]], headers: list[str]) -> None:
    from openpyxl.styles import PatternFill

    sheet.append(headers)
    for row in rows:
        sheet.append([_display_value(row.get(field)) for _key, field in AP_ONLINE_OVERVIEW_COLUMNS])
        fill = overview_row_fill(row)
        for cell in sheet[sheet.max_row]:
            if fill:
                cell.fill = fill
        if int(row.get("offline") or 0) > 0:
            sheet.cell(sheet.max_row, 4).fill = PatternFill(
                fill_type="solid",
                fgColor="FFFEE2E2",
            )
    _format_export_sheet(sheet)


def _log_overview_diagnostics(
    planned_aps: list[dict[str, object | None]],
    raw_metadata_count: int,
    fit_ap_resources: list[dict[str, object | None]],
    optical_rows: list[dict[str, object | None]],
    capacity_details: dict[str, object],
    rows: list[dict[str, object | None]],
    stats: dict[str, object],
) -> None:
    metadata_counts = Counter(str(row.get("_station") or UNASSIGNED_SITE_LABEL) for row in planned_aps)
    capacity_counts = {
        str(site): _capacity_total_remark(value)[0]
        for site, value in capacity_details.items()
        if str(site or "").strip()
    }
    resource_counts = Counter(str(row.get("_resource_site") or UNASSIGNED_SITE_LABEL) for row in fit_ap_resources)
    state_counts = Counter(str(row.get("state") or row.get("state_raw") or row.get("state_display") or "") for row in fit_ap_resources)
    total_row = rows[-1] if rows else {}
    capacity_total = sum(capacity_counts.values())
    metadata_valid_site_count = sum(1 for row in planned_aps if str(row.get("_station") or "").strip() and str(row.get("_station") or "") != UNASSIGNED_SITE_LABEL)
    source_detail = f"metadata_count={raw_metadata_count}, metadata_valid_count={len(planned_aps)}, metadata_station_counts={dict(metadata_counts)}, capacity_counts={capacity_counts}"
    resource_detail = (
        f"resource_count={len(fit_ap_resources)}, online_count={stats.get('online_resource_count')}, "
        f"resource_station_counts={dict(resource_counts)}, state_counts={dict(state_counts)}"
    )
    match_detail = (
        f"capacity_total={capacity_total}, resource_online={stats.get('online_resource_count')}, "
        f"matched_by_optical_uuid={stats.get('matched_by_optical_uuid')}, matched_by_optical_mac={stats.get('matched_by_optical_mac')}, "
        f"matched_by_optical_name={stats.get('matched_by_optical_name')}, matched_by_metadata={stats.get('matched_by_metadata')}, "
        f"matched_by_known_resource_site={stats.get('matched_by_known_resource_site')}, unmatched_online={stats.get('unmatched_online')}, "
        f"station_online_counts={stats.get('station_online_counts')}, unmatched_sample={stats.get('unmatched_sample')}"
    )
    build_detail = (
        f"capacity_station_count={len(capacity_details)}, capacity_total={capacity_total}, "
        f"fit_ap_resource_count={len(fit_ap_resources)}, online_resource_count={stats.get('online_resource_count')}, "
        f"optical_count={len(optical_rows)}, metadata_count={raw_metadata_count}, "
        f"matched_by_optical_uuid={stats.get('matched_by_optical_uuid')}, matched_by_optical_mac={stats.get('matched_by_optical_mac')}, "
        f"matched_by_optical_name={stats.get('matched_by_optical_name')}, matched_by_metadata={stats.get('matched_by_metadata')}, "
        f"matched_by_known_resource_site={stats.get('matched_by_known_resource_site')}, unmatched_online={stats.get('unmatched_online')}, "
        f"total={total_row.get('total')}, online={total_row.get('online')}, offline={total_row.get('offline')}, "
        f"station_online_counts={stats.get('station_online_counts')}, unmatched_sample={stats.get('unmatched_sample')}"
    )
    app_logger.log_info("AP_PLAN_SOURCE", source_detail)
    app_logger.log_info("FIT_AP_RESOURCE_SOURCE", resource_detail)
    app_logger.log_info("AP_ONLINE_MATCH_DEBUG", match_detail)
    app_logger.log_info("AP_ONLINE_OVERVIEW_BUILD", build_detail)

    online_count = int(stats.get("online_resource_count") or 0)
    unmatched_count = int(stats.get("unmatched_online") or 0)
    if raw_metadata_count > 0 and metadata_valid_site_count == 0:
        app_logger.log_warning("AP_ONLINE_OVERVIEW_METADATA_EMPTY", build_detail)
    if online_count and unmatched_count / online_count > 0.1:
        app_logger.log_warning("AP_ONLINE_OVERVIEW_UNMATCHED_HIGH", f"AP上线概览：未匹配在线AP数量过高，可能 FIT-AP光衰记录缺失或名称/MAC不一致。{build_detail}, {match_detail}")
    planned_sites = {str(row.get("_station") or "") for row in planned_aps}
    visible_dirty_sites = {str(row.get("site") or "") for row in rows if str(row.get("site") or "") in DIRTY_RESOURCE_SITES}
    if visible_dirty_sites - planned_sites:
        app_logger.log_warning("AP_ONLINE_OVERVIEW_DIRTY_SITE_VISIBLE", f"{build_detail}, dirty_sites={visible_dirty_sites}")
    station_counts = stats.get("station_online_counts") if isinstance(stats.get("station_online_counts"), dict) else {}
    zero_online_sites = [
        site
        for site, count in metadata_counts.items()
        if count > 0 and int(station_counts.get(site) or 0) == 0 and online_count > 0
    ]
    if zero_online_sites and len(zero_online_sites) == len(metadata_counts) and online_count:
        app_logger.log_warning("AP_ONLINE_OVERVIEW_ALL_PLANNED_STATIONS_ZERO_ONLINE", f"{build_detail}, sites={zero_online_sites}")


def _index_value(index: object, value: object, plan: dict[str, object | None]) -> None:
    if value and isinstance(index, dict):
        index.setdefault(str(value), plan)


def _first_text(row: dict[str, object | None], *fields: str) -> str:
    for field in fields:
        text = _normalize_match_text(row.get(field))
        if text:
            return text
    return ""


def _metadata_row_has_data(row: dict[str, object | None]) -> bool:
    return bool(
        _row_mac(row)
        or _row_name(row)
        or _first_text(row, "ap_uuid", "uuid", "AP_UUID")
        or normalize_station_value(row)
        or str(row.get("站点") or "").strip()
    )


def _plan_site(row: dict[str, object | None]) -> str:
    return normalize_station_value(row) or str(row.get("站点") or "").strip() or UNASSIGNED_SITE_LABEL


def _resource_site(row: dict[str, object | None]) -> str:
    return normalize_station_value(row)


def _row_mac(row: dict[str, object | None]) -> str:
    for field in ("ap_mac", "AP_MAC", "mac", "MAC", "MAC地址"):
        mac = normalize_mac_key(row.get(field)) or ""
        if mac:
            return mac
    return ""


def _row_name(row: dict[str, object | None], *, compact: bool = False) -> str:
    for field in ("ap_name", "AP_NAME", "name", "AP名称"):
        name = _normalize_name(row.get(field), compact=compact)
        if name:
            return name
    return ""


def _normalize_match_text(value: object) -> str:
    return unicodedata.normalize("NFKC", str(value or "")).strip().casefold()


def _normalize_name(value: object, *, compact: bool = False) -> str:
    text = _normalize_match_text(value)
    text = re.sub(r"\s+", " ", text)
    text = "".join(char for char in text if unicodedata.category(char) not in {"Cf", "Cc"})
    if compact:
        return re.sub(r"[\s_\-:./\\|,;，。；、]+", "", text)
    return text


def _plan_unique_key(row: dict[str, object | None]) -> str:
    for prefix, field in (
        ("mac", "_match_mac"),
        ("uuid", "_match_uuid"),
        ("serial", "_match_serial"),
        ("name", "_match_name_compact"),
    ):
        value = str(row.get(field) or "")
        if value:
            return f"{prefix}:{value}"
    return ""


def _resource_unique_key(row: dict[str, object | None]) -> str:
    for prefix, field in (
        ("mac", "_match_mac"),
        ("uuid", "_match_uuid"),
        ("serial", "_match_serial"),
        ("name", "_match_name_compact"),
    ):
        value = str(row.get(field) or "")
        if value:
            return f"{prefix}:{value}"
    return f"row:{id(row)}"


def _append_unmatched_sample(stats: dict[str, object], resource: dict[str, object | None]) -> None:
    sample = stats.get("unmatched_sample")
    if not isinstance(sample, list) or len(sample) >= 20:
        return
    sample.append(
        {
            "ap_uuid": resource.get("ap_uuid"),
            "ap_name": resource.get("ap_name"),
            "ap_mac": resource.get("ap_mac"),
            "raw_site": resource.get("_resource_site"),
            "state": resource.get("state") or resource.get("state_raw") or resource.get("state_display"),
        }
    )


def _capacity_total_remark(value: object) -> tuple[int, str]:
    if isinstance(value, dict):
        return int(value.get("ap_total") or value.get("total") or 0), str(value.get("remark") or "")
    if value in (None, ""):
        return 0, ""
    return int(value), ""


def _with_online_rate(row: dict[str, object | None]) -> dict[str, object | None]:
    total = int(row.get("total") or 0)
    online = int(row.get("online") or 0)
    row["online_rate"] = (
        f"{online / total:.1%}"
        if total > 0 and online <= total
        else "—"
    )
    reonline = int(row.get("reonline_count") or 0)
    row["reonline_rate"] = (
        f"{reonline / total:.1%}"
        if total > 0
        else "—"
    )
    return row


def _display_value(value: object) -> str:
    return str(value) if value not in (None, "") else "-"


def _site_sort_key(site: str) -> tuple[int, str]:
    if site == UNASSIGNED_SITE_LABEL:
        return (1, site)
    return (0, site)


def overview_row_fill(row: dict[str, object | None]):
    from openpyxl.styles import PatternFill

    if str(row.get("site") or "") == TOTAL_SITE_LABEL:
        return PatternFill(fill_type="solid", fgColor="FFDBEAFE")
    total = int(row.get("total") or 0)
    online = int(row.get("online") or 0)
    if total and online == total:
        return PatternFill(fill_type="solid", fgColor="FFDCFCE7")
    if total and online / total < 0.8:
        return PatternFill(fill_type="solid", fgColor="FFFEF9C3")
    return None


def _format_export_sheet(sheet) -> None:
    from openpyxl.styles import Alignment, Border, Font, Side

    from netconsole.services.excel_autosize import apply_worksheet_autofit

    alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    header_font = Font(bold=True)
    # This standalone AC overview is not the trackside history sheet; keep
    # its existing header freeze behavior.
    sheet.freeze_panes = "A2"
    for row in sheet.iter_rows():
        sheet.row_dimensions[row[0].row].height = 24
        for cell in row:
            cell.alignment = alignment
            cell.border = border
            if cell.row == 1:
                cell.font = header_font
    apply_worksheet_autofit(sheet, maximum=60)
