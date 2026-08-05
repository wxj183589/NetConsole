from __future__ import annotations

from pathlib import Path
import re

from netconsole.core.ap_optical_capability import (
    OPTICAL_NOT_APPLICABLE_STATUS,
    is_ap_optical_applicable,
)
from netconsole.core.optical_severity_engine import compute_optical_severity, worse_optical_severity
from netconsole.core.sources.ap_source import compute_ap_status
from netconsole.core.state_engine import compute_state, display_optical_status
from netconsole.services.ap_online_overview import write_ap_online_overview_sheet
from netconsole.services.excel_autosize import apply_worksheet_column_widths
from netconsole.services.export.xlsx_style import apply_basic_sheet_style
from netconsole.services.fit_ap_link_info import lldp_display_status, lldp_source_label
from netconsole.services.offline_ap_ledger import OFFLINE_AP_STATUS_TEXT


OPTICAL_EXPORT_COLOR_RGB = {
    "normal": "DCFCE7",
    "notice": "FEF9C3",
    "warning": "FEF9C3",
    "alarm": "FEE2E2",
    "link_abnormal": "FFE4E6",
    "link_down": "FFE4E6",
    "no_light": "E5E7EB",
    "skipped": "F3F4F6",
    "not_applicable": "F3F4F6",
}


def evaluate_fit_ap_row_status(row: dict[str, object | None], neighbor_optical: dict[str, object | None] | None = None) -> str:
    if not is_ap_optical_applicable(row.get("model")):
        return OPTICAL_NOT_APPLICABLE_STATUS
    if neighbor_optical:
        ap_status = compute_ap_status(row)
        switch_status = _evaluate_neighbor_status(row.get("neighbor_rx_power"), neighbor_optical)
        return worse_optical_severity(switch_status, ap_status)
    result = compute_state(
        {
            "switch_device_name": row.get("neighbor_device_name"),
            "switch_interface_name": row.get("neighbor_interface"),
            "fit_ap_row": row,
        }
    )
    return result.optical_status


def evaluate_fit_ap_ap_status(row: dict[str, object | None]) -> str:
    if not is_ap_optical_applicable(row.get("model")):
        return OPTICAL_NOT_APPLICABLE_STATUS
    if bool(row.get("is_ap_offline")):
        return "offline"
    return compute_ap_status(row)


def evaluate_fit_ap_switch_status(row: dict[str, object | None], neighbor_optical: dict[str, object | None] | None = None) -> str:
    if neighbor_optical:
        return _evaluate_neighbor_status(row.get("neighbor_rx_power"), neighbor_optical)
    return compute_state(
        {
            "switch_device_name": row.get("neighbor_device_name"),
            "switch_interface_name": row.get("neighbor_interface"),
            "fit_ap_row": row,
        }
    ).switch_status


def fit_ap_optical_export_value(row: dict[str, object | None], field: str) -> str:
    if field == "switch_optical_status":
        return display_optical_status(evaluate_fit_ap_switch_status(row))
    if field == "optical_alarm_status":
        if bool(row.get("is_ap_offline")):
            return OFFLINE_AP_STATUS_TEXT
        return display_optical_status(evaluate_fit_ap_ap_status(row))
    return _display_link_value(row, field)


def export_fit_ap_optical_xlsx(
    path: Path,
    rows: list[dict[str, object | None]],
    columns: tuple[tuple[str, str], ...],
    headers: list[str],
    legend_text: str = "",
    overview_rows: list[dict[str, object | None]] | None = None,
    overview_headers: list[str] | None = None,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill

    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    overview_sheet = workbook.active
    overview_sheet.title = "AP上线情况概览"
    write_ap_online_overview_sheet(overview_sheet, list(overview_rows or []), list(overview_headers or []))
    overview_sheet.freeze_panes = "A2"

    optical_sheet = workbook.create_sheet("FIT-AP光衰")
    fields = [field for _key, field in columns]
    optical_sheet.append(headers)
    display_rows: list[dict[str, object | None]] = []
    for row in rows:
        display_row = {field: fit_ap_optical_export_value(row, field) for field in fields}
        display_rows.append(display_row)
        optical_sheet.append([display_row.get(field, "") for field in fields])
        status = evaluate_fit_ap_ap_status(row)
        color = OPTICAL_EXPORT_COLOR_RGB.get(status)
        if color:
            fill = PatternFill(fill_type="solid", fgColor=color)
            for cell in optical_sheet[optical_sheet.max_row]:
                cell.fill = fill
    if legend_text:
        optical_sheet.sheet_properties.pageSetUpPr.fitToPage = True
    optical_sheet.freeze_panes = "A2"
    apply_basic_sheet_style(optical_sheet, header_row=1, column_count=len(headers))
    apply_worksheet_column_widths(optical_sheet, headers, display_rows, fields, maximum=60)
    workbook.save(output)


def _evaluate_neighbor_status(rx_power: object, neighbor_optical: dict[str, object | None] | None) -> str:
    if neighbor_optical:
        return compute_optical_severity(
            {
                "switch_rx_power": neighbor_optical.get("rx_power"),
                "switch_port_status": neighbor_optical.get("port_status"),
                "alarm_low": neighbor_optical.get("rx_low_alarm"),
                "alarm_high": neighbor_optical.get("rx_high_alarm"),
                "warning_low": neighbor_optical.get("rx_low_warning"),
                "device_type": "switch",
            }
        ).severity
    value = _to_float(rx_power)
    return compute_optical_severity({"switch_rx_power": value, "device_type": "switch"}).severity


def _evaluate_ap_result(row: dict[str, object | None]):
    return compute_optical_severity(
        {
            "ap_rx_power": row.get("rx_power"),
            "ap_port_status": row.get("ap_port_status"),
            "alarm_low": row.get("rx_low_alarm"),
            "alarm_high": row.get("rx_high_alarm"),
            "warning_low": row.get("rx_low_warning"),
            "device_type": "ap",
        }
    )


def _to_float(value: object) -> float | None:
    match = re.search(r"[-+]?\d+(?:\.\d+)?", str(value or ""))
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def _display_value(value: object) -> str:
    return str(value) if value not in (None, "") else "-"


def _display_link_value(row: dict[str, object | None], field: str) -> str:
    value = row.get(field)
    if field == "lldp_source":
        return lldp_source_label(value)
    if field in {"lldp_match_status", "optical_match_status", "link_match_status"}:
        return lldp_display_status(value)
    return _display_value(value)
