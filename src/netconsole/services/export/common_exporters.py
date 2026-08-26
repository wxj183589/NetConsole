from __future__ import annotations

import csv
import json
import os
import shutil
import zipfile
from copy import copy
from collections.abc import Callable, Iterable, Mapping, Sequence
from pathlib import Path
from typing import Any

from netconsole.services.excel_autosize import apply_worksheet_column_widths

ProgressCallback = Callable[[str, int, int, str], None]
CancelCallback = Callable[[], bool]


class ExportCancelled(RuntimeError):
    pass


def _path_resolver_from_payload(payload: Mapping[str, Any]):
    from netconsole.core.paths import PathResolver

    app_root = str(payload.get("app_root") or "").strip() or None
    data_root = str(payload.get("data_root") or "").strip() or None
    return PathResolver(app_root=Path(app_root) if app_root else None, data_root=Path(data_root) if data_root else None)


def export_table_xlsx(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    sheets = [
        {
            "sheet_name": payload.get("sheet_name") or "Sheet1",
            "title": payload.get("title") or "",
            "columns": payload.get("columns") or [],
            "source": payload.get("source"),
            "rows": payload.get("rows") or [],
            "auto_width": payload.get("auto_width", True),
            "freeze_header": payload.get("freeze_header", True),
            "auto_filter": payload.get("auto_filter", True),
            "row_fill_field": payload.get("row_fill_field") or "",
            "row_fill_colors": payload.get("row_fill_colors") or {},
        }
    ]
    return export_multi_sheet_xlsx(path, {"sheets": sheets}, progress, should_cancel)


def export_multi_sheet_xlsx(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    from openpyxl import Workbook

    from netconsole.services.export.xlsx_style import apply_basic_sheet_style

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    sheets = list(payload.get("sheets") or [])
    if not sheets:
        sheets = [{"sheet_name": "Sheet1", "columns": [], "rows": []}]
    prepared_sheets = [(sheet, resolve_rows(sheet)) for sheet in sheets]
    total_rows = sum(len(rows) for _sheet, rows in prepared_sheets)
    written_rows = 0
    _emit(progress, "prepare", 0, total_rows, "正在准备工作簿")
    for sheet_index, (sheet_payload, rows) in enumerate(prepared_sheets, start=1):
        _check_cancel(should_cancel)
        columns = normalize_columns(sheet_payload.get("columns") or [])
        title = str(sheet_payload.get("title") or "")
        worksheet = workbook.create_sheet(_safe_sheet_title(str(sheet_payload.get("sheet_name") or f"Sheet{sheet_index}")))
        header_row = 2 if title else 1
        if title:
            worksheet.cell(row=1, column=1, value=title)
            if columns:
                worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        worksheet.append([column["title"] for column in columns])
        row_fill_field = str(sheet_payload.get("row_fill_field") or "")
        row_fill_colors = dict(sheet_payload.get("row_fill_colors") or {})
        for row_index, row in enumerate(rows, start=1):
            values = [_value_for_row(row, column["key"]) for column in columns]
            worksheet.append(values)
            _apply_row_fill(worksheet, row, row_fill_field, row_fill_colors)
            written_rows += 1
            if written_rows == total_rows or written_rows % 100 == 0:
                _emit(progress, "write_rows", written_rows, total_rows, f"正在写入数据 {written_rows}/{total_rows}")
                _check_cancel(should_cancel)
        if bool(sheet_payload.get("freeze_header", True)):
            worksheet.freeze_panes = f"A{header_row + 1}"
        if bool(sheet_payload.get("auto_filter", True)) and columns:
            worksheet.auto_filter.ref = worksheet.dimensions
        apply_basic_sheet_style(worksheet, header_row=header_row, column_count=len(columns))
        for row_number, row in enumerate(rows, start=header_row + 1):
            _apply_xlsx_column_styles(worksheet, row_number, columns)
            if _is_bold_export_row(row, sheet_payload):
                for cell in worksheet[row_number]:
                    font = copy(cell.font)
                    font.bold = True
                    cell.font = font
        if bool(sheet_payload.get("auto_width", True)) and columns:
            apply_worksheet_column_widths(
                worksheet,
                [column["title"] for column in columns],
                rows,
                [column["key"] for column in columns],
                maximum=60,
            )
            for index, column in enumerate(columns, start=1):
                if column.get("width"):
                    worksheet.column_dimensions[worksheet.cell(row=header_row, column=index).column_letter].width = float(column["width"])
    _emit(progress, "save_workbook", total_rows, total_rows, "正在保存 Excel 文件")
    _check_cancel(should_cancel)
    workbook.save(path)
    return total_rows


def export_table_csv(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    columns = normalize_columns(payload.get("columns") or [])
    rows = resolve_rows(payload)
    total = len(rows)
    _emit(progress, "write_csv", 0, total, "正在写入 CSV")
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow([column["title"] for column in columns])
        for index, row in enumerate(rows, start=1):
            writer.writerow([_value_for_row(row, column["key"]) for column in columns])
            if index == total or index % 500 == 0:
                _emit(progress, "write_csv", index, total, f"正在写入 CSV {index}/{total}")
                _check_cancel(should_cancel)
    return total


def export_markdown_text(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    text_file = str(payload.get("text_file") or "").strip()
    if text_file:
        text = Path(text_file).read_text(encoding="utf-8")
    else:
        text = str(payload.get("text") or "")
    _emit(progress, "write_text", 0, 1, "正在写入文本")
    _check_cancel(should_cancel)
    path.write_text(text, encoding="utf-8")
    _emit(progress, "write_text", 1, 1, "文本导出完成")
    return len(text)


def export_config_diff_text(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    from netconsole.core.database import Database
    from netconsole.repositories.config_snapshot_repository import ConfigSnapshotRepository
    from netconsole.services.config_lifecycle_service import ConfigLifecycleService

    path.parent.mkdir(parents=True, exist_ok=True)
    _emit(progress, "config_diff", 0, 1, "正在生成配置差异")
    _check_cancel(should_cancel)
    database = Database(Path(str(payload.get("db_path") or "")))
    repository = ConfigSnapshotRepository(database)
    service = ConfigLifecycleService(str(payload.get("site_name") or ""), database, _path_resolver_from_payload(payload), repository)
    left = repository.get(int(payload.get("left_snapshot_id") or 0))
    right = repository.get(int(payload.get("right_snapshot_id") or 0))
    diff = service.compare_snapshots(left, right)
    path.write_text(diff.raw_diff, encoding="utf-8")
    _emit(progress, "config_diff", 1, 1, "配置差异导出完成")
    return len(diff.raw_diff)


def export_zip_files(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    entries = _zip_entries(payload)
    total = len(entries)
    _emit(progress, "zip_files", 0, total, "正在打包文件")
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as archive:
        for index, (source, arcname) in enumerate(entries, start=1):
            archive.write(source, arcname)
            if index == total or index % 20 == 0:
                _emit(progress, "zip_files", index, total, f"正在打包文件 {index}/{total}")
                _check_cancel(should_cancel)
    return total


def copy_file_export(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    source = Path(str(payload.get("source") or ""))
    if not source.exists():
        raise FileNotFoundError(f"源文件不存在：{source}")
    path.parent.mkdir(parents=True, exist_ok=True)
    _emit(progress, "copy_file", 0, 1, "正在复制文件")
    _check_cancel(should_cancel)
    shutil.copyfile(source, path)
    _emit(progress, "copy_file", 1, 1, "文件复制完成")
    return 1


def export_app_logs_csv(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    from netconsole.services.log_display import display_log_row
    from netconsole.core.log_pagination import iter_logs_from_paths

    log_path = Path(str(payload.get("log_path") or ""))
    keyword = str(payload.get("keyword") or "").strip() or None
    level = str(payload.get("level") or "").strip() or None
    offset = max(0, int(payload.get("offset") or 0))
    limit = max(0, int(payload.get("limit") or 0))
    snapshot_value = payload.get("snapshot_size")
    snapshot_size = None if snapshot_value in (None, "") else max(0, int(snapshot_value))
    log_paths, snapshot_sizes = _resolve_app_log_snapshots(
        log_path,
        payload.get("log_files"),
        fallback_size=snapshot_size,
    )
    redact_web = bool(payload.get("redact_web"))
    path.parent.mkdir(parents=True, exist_ok=True)
    count = 0
    _emit(progress, "write_logs", 0, 0, "正在导出日志")
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["时间", "级别", "事件", "详情", "原始事件", "原始详情"])
        for matched_index, row in enumerate(
            iter_logs_from_paths(
                log_paths,
                keyword=keyword,
                level=level,
                parser=_parse_log_line,
                max_bytes_by_path=snapshot_sizes,
            )
        ):
            if matched_index < offset:
                continue
            if limit and count >= limit:
                break
            display = display_log_row(row)
            if redact_web:
                from netconsole.services.system_maintenance_redaction import redact_system_maintenance_text

                display = {key: redact_system_maintenance_text(value) for key, value in display.items()}
            writer.writerow(
                [
                    display.get("time", ""),
                    display.get("display_level", display.get("level", "")),
                    display.get("display_event", display.get("event", "")),
                    display.get("display_detail", display.get("detail", "")),
                    display.get("raw_event", display.get("event", "")),
                    display.get("raw_detail", display.get("detail", "")),
                ]
            )
            count += 1
            if count % 500 == 0:
                _emit(progress, "write_logs", count, 0, f"正在导出日志 {count} 条")
                _check_cancel(should_cancel)
    return count


def _resolve_app_log_snapshots(
    active_path: Path,
    value: object,
    *,
    fallback_size: int | None,
) -> tuple[list[Path], dict[Path, int | None]]:
    try:
        root = active_path.resolve().parent
    except OSError:
        root = active_path.parent
    rows = value if isinstance(value, list) else []
    paths: list[Path] = []
    sizes: dict[Path, int | None] = {}
    for row in rows:
        if not isinstance(row, Mapping):
            continue
        try:
            candidate = Path(str(row.get("path") or "")).resolve()
        except OSError:
            continue
        if candidate.parent != root or not (
            candidate.name == "app.log"
            or (candidate.name.startswith("app-") and candidate.name.endswith(".log"))
        ):
            continue
        if candidate in sizes:
            continue
        paths.append(candidate)
        sizes[candidate] = max(0, int(row.get("size") or 0))
    if not paths:
        candidate = active_path.resolve()
        paths = [candidate]
        sizes[candidate] = fallback_size
    return paths, sizes


def export_open_source_notices(
    path: Path,
    payload: Mapping[str, Any],
    progress: ProgressCallback | None = None,
    should_cancel: CancelCallback | None = None,
) -> int:
    from netconsole.services.open_source_notice_service import OpenSourceNoticeService

    service = OpenSourceNoticeService(Path(str(payload.get("base_dir") or "")))
    _emit(progress, "scan_dependencies", 0, 1, "正在扫描运行依赖")
    components = service.list_components()
    _check_cancel(should_cancel)
    rows = [component.__dict__ for component in components]
    if str(payload.get("format") or "") == "xlsx":
        return export_table_xlsx(
            path,
            {
                "columns": [
                    {"key": key, "title": title, "text": True}
                    for key, title in zip(
                        ("name", "version", "license", "purpose", "homepage", "note"),
                        ("组件名称", "版本", "许可证", "用途", "项目地址", "备注"),
                        strict=True,
                    )
                ],
                "source": {"type": "inline_rows", "rows": rows, "inline_reason": "small_static_notice"},
                "sheet_name": "开源许可",
                "freeze_header": True,
                "auto_filter": True,
                "auto_width": True,
            },
            progress,
            should_cancel,
        )
    text = ["NetConsole 开源许可说明", ""]
    for component in components:
        text.extend(
            (
                f"组件名称：{component.name}",
                f"版本：{component.version or '-'}",
                f"许可证：{component.license or '-'}",
                f"用途：{component.purpose or '-'}",
                f"项目地址：{component.homepage or '-'}",
                f"备注：{component.note or '-'}",
                "",
            )
        )
    return export_markdown_text(path, {"text": "\n".join(text)}, progress, should_cancel)


def export_command_reference_markdown(
    path: Path,
    payload: Mapping[str, Any],
    progress: ProgressCallback | None = None,
    should_cancel: CancelCallback | None = None,
) -> int:
    import json

    from netconsole.services.command_reference_service import CommandReference, export_command_references_markdown
    from netconsole.utils.text_encoding import read_text_with_fallback

    source = Path(str(payload.get("resource_path") or ""))
    if not source.is_file():
        raise FileNotFoundError(f"命令说明资源不存在：{source}")
    _emit(progress, "read_command_reference", 0, 1, "正在读取命令说明")
    _check_cancel(should_cancel)
    raw = json.loads(read_text_with_fallback(source))
    items = raw.get("items", raw) if isinstance(raw, Mapping) else raw
    references = [CommandReference.from_dict(dict(item)) for item in items or [] if isinstance(item, Mapping)]
    selected_ids = {str(value) for value in payload.get("selected_ids") or [] if str(value)}
    if "selected_ids" in payload:
        references = [item for item in references if item.id in selected_ids]
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(export_command_references_markdown(references), encoding="utf-8")
    _emit(progress, "write_command_reference", len(references), len(references), "命令说明导出完成")
    return len(references)


def export_vehicle_mr_history_xlsx(
    path: Path,
    payload: Mapping[str, Any],
    progress: ProgressCallback | None = None,
    should_cancel: CancelCallback | None = None,
) -> int:
    from openpyxl import Workbook

    from netconsole.services.vehicle_mr_online import VehicleMrOnlineStore

    filters = dict(payload.get("filters") or {})
    store = VehicleMrOnlineStore(_path_resolver_from_payload(payload), str(payload.get("site_name") or ""))
    _emit(progress, "query_vehicle_mr_history", 0, 1, "正在查询车载 MR 历史")
    _check_cancel(should_cancel)
    rows = store.query_events(
        str(payload.get("train_id") or ""),
        str(filters.get("start_time") or ""),
        str(filters.get("end_time") or ""),
        car_end_label=str(filters.get("car_end_label") or ""),
        status=str(filters.get("status") or ""),
        station=str(filters.get("station") or ""),
        ap_name=str(filters.get("ap_name") or ""),
        limit=1_000_000,
    )
    headers = ["时间", "端别", "状态", "车站", "轨旁AP", "RSSI", "事件类型", "判断说明"]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "历史记录"
    sheet.append(headers)
    for index, row in enumerate(rows, start=1):
        sheet.append(
            [
                row.get("event_time") or "",
                row.get("car_end_label") or "",
                row.get("status") or "",
                row.get("station") or "",
                row.get("ap_name") or "",
                row.get("rssi") if row.get("rssi") is not None else "-",
                row.get("event_type") or "",
                _vehicle_mr_status_reason_label(str(row.get("status_reason") or "")),
            ]
        )
        if index % 500 == 0:
            _emit(progress, "write_vehicle_mr_history", index, len(rows), f"正在写入历史记录 {index}/{len(rows)}")
            _check_cancel(should_cancel)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column, width in enumerate((22, 10, 12, 20, 24, 10, 16, 28), start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=column).column_letter].width = width
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return len(rows)


def _vehicle_mr_status_reason_label(reason: str) -> str:
    labels = {
        "both_offline": "双端均离线",
        "dual_active_ok": "双端在线",
        "tc1_missing": "双活缺TC1",
        "tc2_missing": "双活缺TC2",
        "both_ends_online": "双端在线",
        "expected_tc1_online": "TC1符合预期在线",
        "expected_tc2_online": "TC2符合预期在线",
        "unexpected_tc1_online": "非预期TC1在线",
        "unexpected_tc2_online": "非预期TC2在线",
        "expected_tail_online": "尾端在线",
        "unexpected_end_online": "非预期端在线",
        "direction_unknown_any_end_online": "方向未知，任意一端在线视为在线",
        "policy_unknown_any_end_online": "自动/未知策略，任意一端在线视为在线",
    }
    return labels.get(reason, reason or "-")


def export_device_csv(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    from netconsole.core.database import Database
    from netconsole.models.device import Device
    from netconsole.repositories.device_group_repository import DeviceGroupRepository
    from netconsole.repositories.device_repository import DeviceRepository
    from netconsole.services.device_import_export import DeviceImportExportService

    _emit(progress, "prepare_device_csv", 10, 100, "正在读取设备导出参数")
    _check_cancel(should_cancel)
    path.parent.mkdir(parents=True, exist_ok=True)
    devices_payload = list(payload.get("devices") or [])
    site_name = str(payload.get("site_name") or "")
    if devices_payload:
        devices = [Device.from_mapping(dict(row)) for row in devices_payload if isinstance(row, Mapping)]
        database = Database(Path(str(payload.get("db_path") or ":memory:")))
        group_repository = DeviceGroupRepository(database, site_name) if site_name else None
        service = DeviceImportExportService(DeviceRepository(database), group_repository)
    else:
        _emit(progress, "query_device_csv", 30, 100, "正在查询设备清单")
        _check_cancel(should_cancel)
        source = payload.get("source") if isinstance(payload.get("source"), Mapping) else {}
        db_path = str(payload.get("db_path") or source.get("db_path") or "")
        database = Database(Path(db_path))
        repository = DeviceRepository(database)
        group_repository = DeviceGroupRepository(database, site_name) if site_name else None
        filters = dict(payload.get("filters") or source.get("filters") or {})
        devices = repository.list(
            search=filters.get("search") or None,
            vendor=filters.get("vendor") or None,
            device_type=filters.get("device_type") or None,
            group_filter=filters.get("group_filter"),
            project_phase=filters.get("project_phase"),
            work_scope_status=filters.get("work_scope_status"),
        )
        service = DeviceImportExportService(repository, group_repository)
    selected_uuids = {
        str(value).strip()
        for value in payload.get("selected_device_uuids") or []
        if str(value).strip()
    }
    if selected_uuids:
        devices = [
            device
            for device in devices
            if str(device.device_uuid or "").strip() in selected_uuids
        ]
    _emit(progress, "generate_device_csv", 55, 100, "正在生成设备 CSV")
    _check_cancel(should_cancel)
    service.export_csv(
        path, devices, include_sensitive=not bool(payload.get("omit_credentials"))
    )
    _emit(progress, "verify_device_csv", 75, 100, "正在校验设备 CSV")
    if not path.is_file() or path.stat().st_size <= 0:
        raise RuntimeError("设备 CSV 写入后为空")
    _emit(progress, "register_device_csv", 90, 100, "正在注册设备 CSV Artifact")
    return len(devices)


def export_device_template_csv(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    from netconsole.services.device_import_export import (
        DEVICE_CSV_COLUMNS,
        TEMPLATE_EXAMPLE_ROWS,
    )

    _emit(progress, "prepare_device_template", 10, 100, "正在读取设备模板定义")
    _check_cancel(should_cancel)
    path.parent.mkdir(parents=True, exist_ok=True)
    _emit(progress, "generate_device_template", 35, 100, "正在生成设备导入模板")
    _check_cancel(should_cancel)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(DEVICE_CSV_COLUMNS)
        writer.writerows(TEMPLATE_EXAMPLE_ROWS)
    _emit(progress, "verify_device_template", 65, 100, "正在校验设备导入模板")
    _check_cancel(should_cancel)
    if not path.read_bytes().startswith(b"\xef\xbb\xbf"):
        raise RuntimeError("设备导入模板缺少 UTF-8 BOM")
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.reader(handle))
    if rows != [DEVICE_CSV_COLUMNS, *TEMPLATE_EXAMPLE_ROWS]:
        raise RuntimeError("设备导入模板字段校验失败")
    _emit(progress, "register_device_template", 85, 100, "正在注册模板 Artifact")
    return 0


def export_securecrt_sessions_task(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> dict[str, Any]:
    from netconsole.core.database import Database
    from netconsole.repositories.device_group_repository import DeviceGroupRepository
    from netconsole.repositories.device_repository import DeviceRepository
    from netconsole.services.securecrt_session_export import export_securecrt_sessions

    output_parent = Path(str(payload.get("output_dir") or path))
    database = Database(Path(str(payload.get("db_path") or "")))
    repository = DeviceRepository(database)
    filters = dict(payload.get("filters") or {})
    selected_uuids = {str(value).strip() for value in payload.get("selected_device_uuids") or [] if str(value).strip()}
    devices = repository.list(
        search=filters.get("search") or None,
        vendor=filters.get("vendor") or None,
        device_type=filters.get("device_type") or None,
        group_filter=filters.get("group_filter"),
        project_phase=filters.get("project_phase"),
        work_scope_status=filters.get("work_scope_status"),
    )
    if selected_uuids:
        devices = [device for device in devices if str(device.device_uuid or "").strip() in selected_uuids]
    group_names = {
        int(group.id): str(group.name)
        for group in DeviceGroupRepository(database, str(payload.get("site_name") or "")).list()
        if group.id is not None
    }
    template_value = str(payload.get("template_ini") or "")
    template_ini = Path(template_value) if template_value else None
    _emit(progress, "write_securecrt_sessions", 0, len(devices), "正在生成 SecureCRT 会话")
    _check_cancel(should_cancel)
    result = export_securecrt_sessions(
        devices,
        str(payload.get("site_name") or ""),
        output_parent,
        group_names=group_names,
        template_ini=template_ini if template_ini and template_ini.is_file() else None,
    )
    marker = Path(path)
    marker.parent.mkdir(parents=True, exist_ok=True)
    marker.write_text(str(result.output_dir), encoding="utf-8")
    _emit(progress, "write_securecrt_sessions", result.generated, len(devices), "SecureCRT 会话生成完成")
    return {"path": str(result.output_dir), "row_count": result.generated, "skipped": result.skipped}


def export_config_snapshots_zip(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    from netconsole.core import app_logger
    from netconsole.core.database import Database
    from netconsole.core.paths import PathResolver
    from netconsole.repositories.config_snapshot_repository import ConfigSnapshotRepository
    from netconsole.services.config_lifecycle_service import ConfigLifecycleService

    database = Database(Path(str(payload.get("db_path") or "")))
    site_name = str(payload.get("site_name") or "")
    paths = PathResolver()
    repository = ConfigSnapshotRepository(database)
    service = ConfigLifecycleService(site_name, database, paths, repository)
    snapshot_entries = list(payload.get("snapshot_entries") or [])
    file_entries = list(payload.get("file_entries") or [])
    failures_text = str(payload.get("failures_text") or "")
    total = len(snapshot_entries) + len(file_entries) + (1 if failures_text else 0)
    written = 0
    path.parent.mkdir(parents=True, exist_ok=True)
    _emit(progress, "zip_config_snapshots", 0, total, "正在打包配置快照")
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as archive:
        for entry in snapshot_entries:
            _check_cancel(should_cancel)
            if not isinstance(entry, Mapping):
                continue
            snapshot_id = entry.get("snapshot_id")
            archive_name = str(entry.get("archive_name") or "")
            if snapshot_id is None or not archive_name:
                continue
            snapshot = repository.get(int(snapshot_id))
            if snapshot.type in {"running", "saved"}:
                archive.writestr(archive_name, service.snapshot_text(snapshot))
            else:
                source = paths.site_dir(site_name) / snapshot.file_path
                if source.exists():
                    archive.write(source, archive_name)
                else:
                    app_logger.log_warning("CONFIG_SNAPSHOT_EXPORT_MISSING", f"snapshot_id={snapshot.id} path={source}")
            written += 1
            if written == total or written % 20 == 0:
                _emit(progress, "zip_config_snapshots", written, total, f"正在打包配置快照 {written}/{total}")
        for entry in file_entries:
            _check_cancel(should_cancel)
            if not isinstance(entry, Mapping):
                continue
            source = Path(str(entry.get("path") or ""))
            archive_name = str(entry.get("archive_name") or source.name)
            if source.exists() and source.is_file():
                archive.write(source, archive_name)
            else:
                app_logger.log_warning("CONFIG_SNAPSHOT_EXPORT_MISSING", f"path={source}")
            written += 1
            if written == total or written % 20 == 0:
                _emit(progress, "zip_config_snapshots", written, total, f"正在打包配置快照 {written}/{total}")
        if failures_text:
            archive.writestr("failed_devices.txt", failures_text.rstrip("\n") + "\n")
            written += 1
    return written


def _fit_ap_import_export_service(payload: Mapping[str, Any]):
    from netconsole.core.database import Database
    from netconsole.repositories.ac_repository import AcRepository
    from netconsole.services.fit_ap_import_export import FitApImportExportService

    return FitApImportExportService(AcRepository(Database(Path(str(payload.get("db_path") or "")))))


def export_fit_ap_csv_task(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    ac_uuid = str(payload.get("ac_uuid") or "").strip()
    rows = _ac_repository(payload).list_fit_ap_resources_with_metadata(ac_uuid) if ac_uuid else []
    selected_keys = {str(key).strip() for key in payload.get("selected_ap_keys") or [] if str(key).strip()}
    if selected_keys:
        rows = [row for row in rows if _fit_ap_selection_key(row) in selected_keys]
    else:
        rows = _filter_fit_ap_rows(rows, dict(payload.get("filters") or {}))
    path.parent.mkdir(parents=True, exist_ok=True)
    _emit(progress, "write_fit_ap_csv", 0, len(rows), "正在导出 FIT-AP CSV")
    _check_cancel(should_cancel)
    _fit_ap_import_export_service(payload).export_ap_csv(path, rows)
    _emit(progress, "write_fit_ap_csv", len(rows), len(rows), "FIT-AP CSV 导出完成")
    return len(rows)


def export_fit_ap_extension_xlsx_task(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    filters = dict(payload.get("filters") or {})
    rows = _ac_repository(payload).list_ap_extension_points(
        search=str(payload.get("search") or filters.get("search") or ""),
        station_name=str(filters.get("station_name") or ""),
        line_side=str(filters.get("line_side") or ""),
        direction=str(filters.get("direction") or ""),
        match_status=str(filters.get("match_status") or ""),
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    _emit(progress, "write_fit_ap_extension", 0, len(rows), "正在导出 FIT-AP 扩展信息")
    _check_cancel(should_cancel)
    _fit_ap_import_export_service(payload).export_standard_ap_extension_xlsx(path, rows)
    _emit(progress, "write_fit_ap_extension", len(rows), len(rows), "FIT-AP 扩展信息导出完成")
    return len(rows)


def export_fit_ap_extension_template_xlsx_task(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    ac_uuid = str(payload.get("ac_uuid") or "").strip()
    rows = _ac_repository(payload).list_fit_ap_resources_with_metadata(ac_uuid) if ac_uuid else []
    ap_entities = _ac_repository(payload).list_ap_entities(ac_uuid) if ac_uuid else []
    path.parent.mkdir(parents=True, exist_ok=True)
    _emit(progress, "write_fit_ap_extension_template", 0, len(rows), "正在导出 AP 扩展模板")
    _check_cancel(should_cancel)
    _fit_ap_import_export_service(payload).export_ap_extension_template_xlsx(path, rows, ap_entities)
    _emit(progress, "write_fit_ap_extension_template", len(rows), len(rows), "AP 扩展模板导出完成")
    return len(rows)


def export_ap_online_overview_xlsx_task(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    from netconsole.services.ap_online_overview import export_ap_online_overview_xlsx

    _emit(progress, "prepare_ap_online_overview", 0, 1, "正在查询 AP 在线概览数据")
    _check_cancel(should_cancel)
    overview = _build_ap_online_overview_payload(payload)
    rows = [dict(row) for row in overview.get("overview_rows") or [] if isinstance(row, Mapping)]
    path.parent.mkdir(parents=True, exist_ok=True)
    _emit(progress, "write_ap_online_overview", 0, len(rows), "正在导出 AP 在线概览")
    _check_cancel(should_cancel)
    export_ap_online_overview_xlsx(
        path,
        rows,
        [str(value) for value in payload.get("headers") or []],
        dict(overview.get("offline_ap_stats") or {}),
        [dict(row) for row in overview.get("offline_ap_ledger_rows") or [] if isinstance(row, Mapping)],
        [str(value) for value in payload.get("offline_ap_stats_headers") or []],
        [str(value) for value in payload.get("offline_ap_ledger_headers") or []],
    )
    _emit(progress, "write_ap_online_overview", len(rows), len(rows), "AP 在线概览导出完成")
    return len(rows)


def export_fit_ap_optical_xlsx_task(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    from netconsole.services.fit_ap_optical_export import export_fit_ap_optical_xlsx

    _emit(progress, "prepare_fit_ap_optical", 0, 1, "正在查询 FIT-AP 光衰数据")
    _check_cancel(should_cancel)
    overview = _build_ap_online_overview_payload(payload)
    optical_rows = _build_fit_ap_optical_rows(payload)
    columns_payload = [dict(column) for column in payload.get("columns") or [] if isinstance(column, Mapping)]
    columns = tuple((str(column.get("title") or column.get("key") or ""), str(column.get("key") or "")) for column in columns_payload)
    headers = [str(value) for value in payload.get("headers") or []]
    overview_headers = [str(value) for value in payload.get("overview_headers") or []]
    path.parent.mkdir(parents=True, exist_ok=True)
    _emit(progress, "write_fit_ap_optical", 0, len(optical_rows), "正在导出 FIT-AP 光衰")
    _check_cancel(should_cancel)
    export_fit_ap_optical_xlsx(
        path,
        optical_rows,
        columns,
        headers,
        overview_rows=[dict(row) for row in overview.get("overview_rows") or [] if isinstance(row, Mapping)],
        overview_headers=overview_headers,
    )
    _emit(progress, "write_fit_ap_optical", len(optical_rows), len(optical_rows), "FIT-AP 光衰导出完成")
    return len(optical_rows)


def _fit_ap_selection_key(row: Mapping[str, Any]) -> str:
    return str(row.get("ap_uuid") or row.get("ap_name") or "").strip()


def _filter_fit_ap_rows(rows: list[dict[str, Any]], filters: Mapping[str, Any]) -> list[dict[str, Any]]:
    search = "".join(char for char in str(filters.get("search") or "").casefold() if char.isalnum())
    if search:
        fields = (
            "ap_name",
            "ap_mac",
            "ap_ip",
            "serial_number",
            "rid1_bbssid",
            "rid2_bbssid",
            "rid3_bbssid",
            "lldp_neighbor_mac",
            "lldp_neighbor_interface",
            "mileage",
        )
        rows = [
            row
            for row in rows
            if any(search in "".join(char for char in str(row.get(field) or "").casefold() if char.isalnum()) for field in fields)
        ]
    group = str(filters.get("group") or "").strip()
    if group:
        rows = [row for row in rows if str(row.get("group_name") or "").strip() == group]
    state = str(filters.get("state") or "").strip()
    if state:
        rows = [row for row in rows if _fit_ap_state_matches(row, state)]
    return rows


def _fit_ap_state_matches(row: Mapping[str, Any], selected: str) -> bool:
    from netconsole.services.offline_ap_ledger import fit_ap_online_status

    if selected == "__offline__":
        return fit_ap_online_status(dict(row)) == "offline"
    if selected == "__online__":
        return fit_ap_online_status(dict(row)) == "online"
    values = {
        str(row.get(field) or "").strip().casefold()
        for field in ("state", "state_raw", "state_display")
        if str(row.get(field) or "").strip()
    }
    return selected.casefold() in values


def _build_fit_ap_optical_rows(payload: Mapping[str, Any]) -> list[dict[str, object | None]]:
    from netconsole.core.database import Database
    from netconsole.core.sources.switch_source import build_switch_data_lookup, compute_switch_status
    from netconsole.repositories.ac_repository import AcRepository
    from netconsole.repositories.device_fact_repository import DeviceFactRepository
    from netconsole.repositories.device_repository import DeviceRepository
    from netconsole.services.fit_ap_optical_export import evaluate_fit_ap_ap_status
    from netconsole.services.offline_ap_ledger import OFFLINE_AP_STATUS_TEXT, is_fit_ap_offline
    from netconsole.services.ap_identity.normalizers import normalize_mac
    from netconsole.utils.interface_sort import interface_sort_key

    ac_uuid = str(payload.get("ac_uuid") or "").strip()
    database = Database(Path(str(payload.get("db_path") or "")))
    ac_repository = AcRepository(database)
    device_repository = DeviceRepository(database)
    fact_repository = DeviceFactRepository(database)
    resources = ac_repository.list_fit_ap_resources_with_metadata(ac_uuid)
    optical_rows = ac_repository.list_fit_ap_optical(ac_uuid)
    devices = device_repository.list()
    optical_by_device = {str(device.device_uuid or ""): fact_repository.list_optical_modules(str(device.device_uuid or "")) for device in devices}
    switch_lookup = build_switch_data_lookup(devices, optical_by_device)
    resources_by_uuid = {str(row.get("ap_uuid") or ""): row for row in resources if row.get("ap_uuid")}
    resources_by_mac = {normalize_mac(row.get("ap_mac")): row for row in resources if normalize_mac(row.get("ap_mac"))}
    enriched: list[dict[str, object | None]] = []
    for row in optical_rows:
        resource = resources_by_uuid.get(str(row.get("ap_uuid") or "")) or resources_by_mac.get(normalize_mac(row.get("ap_mac")), {})
        neighbor_name = None if _invalid_fit_ap_neighbor_text(row.get("neighbor_device_name")) else row.get("neighbor_device_name")
        is_offline = is_fit_ap_offline(dict(resource)) or bool(row.get("is_ap_offline"))
        enriched.append(
            {
                **row,
                "ap_mac": row.get("ap_mac") or resource.get("ap_mac"),
                "site": row.get("site") or resource.get("site_name") or resource.get("site") or "未归属",
                "neighbor_device_name": neighbor_name,
                "switch_optical_status": compute_switch_status(
                    device_name=neighbor_name,
                    interface_name=row.get("neighbor_interface"),
                    lookup=switch_lookup,
                ),
                "is_ap_offline": is_offline,
                "optical_alarm_status": OFFLINE_AP_STATUS_TEXT if is_offline else row.get("optical_alarm_status"),
                "ap_optical_status": "offline" if is_offline else row.get("ap_optical_status"),
                "data_source": "historical" if is_offline else row.get("data_source"),
            }
        )
    filters = dict(payload.get("filters") or {})
    rows = _filter_fit_ap_optical_export_rows(enriched, filters, evaluate_fit_ap_ap_status)
    return sorted(
        rows,
        key=lambda row: (
            1 if str(row.get("neighbor_device_name") or "").strip() in {"", "-"} else 0,
            str(row.get("neighbor_device_name") or "").strip().casefold(),
            interface_sort_key(row.get("neighbor_interface")),
            str(row.get("ap_name") or ""),
        ),
    )


def _filter_fit_ap_optical_export_rows(
    rows: list[dict[str, object | None]],
    filters: Mapping[str, Any],
    status_func: Callable[[dict[str, object | None]], str],
) -> list[dict[str, object | None]]:
    result = rows
    for field in ("ap_name", "site"):
        needle = str(filters.get(field) or "").strip().casefold()
        if needle:
            result = [row for row in result if needle in str(row.get(field) or "").casefold()]
    status = str(filters.get("optical_alarm_status") or "").strip()
    if status:
        result = [row for row in result if status_func(row) == status]
    return result


def _invalid_fit_ap_neighbor_text(value: object) -> bool:
    lowered = str(value or "").casefold()
    return any(token.casefold() in lowered for token in ("Nearest", "Chassis ID", "Default", "customer bridge", "nontpmr"))


def _build_ap_online_overview_payload(payload: Mapping[str, Any]) -> dict[str, Any]:
    from netconsole.core.database import Database
    from netconsole.repositories.ac_repository import AcRepository
    from netconsole.repositories.device_repository import DeviceRepository
    from netconsole.services.ap_online_overview import ApOnlineOverviewService
    from netconsole.services.offline_ap_ledger import build_current_ap_history_indexes, build_device_lookup_by_name, build_offline_ap_ledger

    ac_uuid = str(payload.get("ac_uuid") or "").strip()
    database = Database(Path(str(payload.get("db_path") or "")))
    repository = AcRepository(database)
    resources = repository.list_fit_ap_resources_with_metadata(ac_uuid)
    optical_rows = repository.list_fit_ap_optical(ac_uuid)
    capacity_details = repository.list_active_trackside_plan_capacity_details()
    if not capacity_details:
        capacity_details = repository.list_station_ap_capacity_details()
    overview_rows = ApOnlineOverviewService.build_rows(
        metadata_rows=repository.list_fit_ap_metadata(),
        fit_ap_resources=resources,
        optical_rows=optical_rows,
        capacity_details=capacity_details,
    )
    latest_lldp, _latest_optical = build_current_ap_history_indexes(
        repository.list_current_ap_lldp_states(), resources
    )
    stats, ledger = build_offline_ap_ledger(
        fit_ap_resources=resources,
        latest_lldp_by_ap=latest_lldp,
        device_lookup_by_name=build_device_lookup_by_name(DeviceRepository(database).list()),
    )
    return {"overview_rows": overview_rows, "offline_ap_stats": stats, "offline_ap_ledger_rows": ledger}


def export_omnipeek_name_table_task(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> dict[str, Any]:
    from netconsole.core.database import Database
    from netconsole.models.omnipeek_name_table import SOURCE_DEVICE_MANAGEMENT, OmniPeekExportConfig
    from netconsole.repositories.ac_repository import AcRepository
    from netconsole.repositories.device_group_repository import DeviceGroupRepository
    from netconsole.repositories.device_repository import DeviceRepository
    from netconsole.services.omnipeek_name_table_service import OmniPeekNameTableService, export_items_to_omnipeek_nam

    source = dict(payload.get("source") or {})
    database = Database(Path(str(payload.get("db_path") or "")))
    site_name = str(payload.get("site_name") or "")
    device_repository = DeviceRepository(database)
    filters = dict(source.get("device_filters") or {})
    allowed_filters = {key: filters.get(key) for key in ("search", "vendor", "device_type", "group_filter") if filters.get(key) is not None}
    devices = device_repository.list(**allowed_filters)
    selected_device_uuids = {str(value) for value in source.get("selected_device_uuids") or [] if str(value)}
    if selected_device_uuids:
        devices = [device for device in devices if str(device.device_uuid or "") in selected_device_uuids]
    groups = DeviceGroupRepository(database, site_name).list() if site_name else []
    group_names = {int(group.id): group.name for group in groups if group.id is not None}
    config_data = dict(payload.get("config") or {})
    config_data["output_path"] = path
    config = OmniPeekExportConfig(**config_data)
    service = OmniPeekNameTableService(AcRepository(database), device_repository)
    selected_fit_ap_ids = [str(value) for value in source.get("selected_fit_ap_ids") or [] if str(value)]
    items = service.collect_items(
        include_ac_fit_ap=config.include_ac_fit_ap,
        include_ap_extensions=config.include_ap_extensions,
        include_device_mr=config.include_device_mr,
        ac_device_uuid=str(source.get("ac_uuid") or "") or None,
        devices=devices,
        group_names=group_names,
        selected_fit_ap_ids=selected_fit_ap_ids,
        scope_extensions_to_fit_ap=bool(source.get("scope_extensions_to_fit_ap", False)),
    )
    selected_keys = {str(value) for value in payload.get("selected_item_keys") or [] if str(value)}
    excluded_keys = {str(value) for value in payload.get("excluded_item_keys") or [] if str(value)}
    force_keys = {str(value) for value in payload.get("force_export_keys") or [] if str(value)}
    for item in items:
        if item.key in excluded_keys:
            item.selected = False
        elif item.key in selected_keys:
            item.selected = True
        if item.key in force_keys:
            item.selected = True
            item.force_export = True
    source_counts = service.source_counts(
        ac_device_uuid=str(source.get("ac_uuid") or "") or None,
        devices=devices,
        selected_fit_ap_ids=selected_fit_ap_ids,
        scope_extensions_to_fit_ap=bool(source.get("scope_extensions_to_fit_ap", False)),
    )
    source_counts[SOURCE_DEVICE_MANAGEMENT] = sum(1 for item in items if SOURCE_DEVICE_MANAGEMENT in (item.sources or [item.source]))
    _emit(progress, "write_omnipeek_name_table", 0, len(items), "正在导出 OmniPeek 名称表")
    _check_cancel(should_cancel)
    result = export_items_to_omnipeek_nam(items, config, source_counts=source_counts)
    row_count = result.total_entries
    _emit(progress, "write_omnipeek_name_table", row_count, row_count, "OmniPeek 名称表导出完成")
    return {
        "path": str(result.output_path),
        "row_count": row_count,
        "log_path": str(result.log_path),
        "skipped_count": result.skipped_count,
        "error_count": result.error_count,
    }


def export_online_mr_report_xlsx(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    from netconsole.services.vehicle_mr_offline_excel_report import VehicleMrOfflineExcelReportExporter

    session_dir = Path(str(payload.get("session_dir") or ""))
    _emit(progress, "online_mr_report_prepare", 1, 3, "正在读取 parsed 数据")
    _check_cancel(should_cancel)
    output_path = VehicleMrOfflineExcelReportExporter().export(session_dir, path)
    _emit(progress, "online_mr_report_save", 2, 3, "正在保存 Excel 报告")
    _check_cancel(should_cancel)
    _emit(progress, "online_mr_report_done", 3, 3, "离线分析报告导出完成")
    return 1 if output_path.exists() else 0


def export_car_network_point_table(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    from netconsole.core.paths import PathResolver
    from netconsole.services.rail_transit.car_network_diagnostic import CarNetworkPointTableStore

    site_name = str(payload.get("site_name") or "")
    store = CarNetworkPointTableStore(PathResolver(), site_name)
    nodes = store.load()
    _emit(progress, "write_car_network_point_table", 0, len(nodes), "正在导出车内通信点表")
    _check_cancel(should_cancel)
    path.parent.mkdir(parents=True, exist_ok=True)
    suffix = _effective_suffix(path)
    target = path if path.suffix.casefold() == suffix else path.with_name(f"{path.stem}.netconsole{suffix or '.xlsx'}")
    try:
        store.export_file(target, nodes)
        if target != path:
            os.replace(target, path)
    finally:
        if target != path and target.exists():
            try:
                target.unlink()
            except OSError:
                pass
    _emit(progress, "write_car_network_point_table", len(nodes), len(nodes), "车内通信点表导出完成")
    return len(nodes)


def normalize_columns(columns: Iterable[Any]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for column in columns:
        if isinstance(column, Mapping):
            key = str(column.get("key") or column.get("field") or "")
            title = str(column.get("title") or column.get("header") or key)
            result.append(
                {
                    "key": key,
                    "title": title,
                    "width": column.get("width"),
                    "text": bool(column.get("text", True)),
                    "number_format": str(column.get("number_format") or ""),
                    "wrap": bool(column.get("wrap", False)),
                    "horizontal": str(column.get("horizontal") or ""),
                }
            )
            continue
        if isinstance(column, Sequence) and not isinstance(column, (str, bytes)) and len(column) >= 2:
            title, key = column[0], column[1]
            result.append({"key": str(key), "title": str(title), "width": None, "text": True})
            continue
        key = str(column)
        result.append({"key": key, "title": key, "width": None, "text": True})
    return result


def resolve_rows(payload: Mapping[str, Any]) -> list[Any]:
    source = payload.get("source")
    if not source:
        return list(payload.get("rows") or [])
    if not isinstance(source, Mapping):
        raise ValueError("导出 source 必须是 dict")
    source_type = str(source.get("type") or "").strip()
    if source_type == "inline_rows":
        return list(source.get("rows") or [])
    if source_type == "jsonl_rows":
        return _resolve_jsonl_rows(source)
    if source_type == "repository_query":
        return _resolve_repository_rows(source)
    raise ValueError(f"不支持的导出数据源：{source_type or '<empty>'}")


def _resolve_jsonl_rows(source: Mapping[str, Any]) -> list[dict[str, Any]]:
    result_file = Path(str(source.get("result_file") or ""))
    if not result_file.is_file():
        raise FileNotFoundError(f"运行结果文件不存在：{result_file}")
    rows: list[dict[str, Any]] = []
    with result_file.open("r", encoding="utf-8") as handle:
        for line_no, line in enumerate(handle, start=1):
            text = line.strip()
            if not text:
                continue
            value = json.loads(text)
            if not isinstance(value, Mapping):
                raise ValueError(f"运行结果文件第 {line_no} 行不是对象")
            rows.append(dict(value))
    return rows


def _resolve_repository_rows(source: Mapping[str, Any]) -> list[dict[str, Any]]:
    repository = str(source.get("repository") or "").strip()
    method = str(source.get("method") or "").strip()
    filters = dict(source.get("filters") or {})
    db_path = Path(str(source.get("db_path") or ""))
    if repository == "device_repository" and method == "list":
        from netconsole.core.database import Database
        from netconsole.repositories.device_repository import DeviceRepository

        devices = DeviceRepository(Database(db_path)).list(
            search=filters.get("search") or None,
            vendor=filters.get("vendor") or None,
            device_type=filters.get("device_type") or None,
            group_filter=filters.get("group_filter"),
        )
        return [device.to_record() for device in devices]
    if repository == "wireless_scan_repository" and method == "list_results":
        from netconsole.repositories.wireless_scan_repository import WirelessScanRepository
        from netconsole.services.network_tools.wireless_scan_service import repository_row_to_display_row

        scan_id = str(filters.get("scan_id") or "").strip()
        if not scan_id:
            raise ValueError("wireless_scan_repository.list_results 缺少 scan_id")
        return [repository_row_to_display_row(row) for row in WirelessScanRepository(db_path).list_results(scan_id)]
    if repository == "ac_repository" and method == "list_trackside_ap_plan":
        from netconsole.core.database import Database
        from netconsole.repositories.ac_repository import AcRepository, TRACKSIDE_AP_PLAN_MODE

        mode = str(filters.get("mode") or TRACKSIDE_AP_PLAN_MODE)
        return AcRepository(Database(db_path)).list_trackside_ap_plan(mode)
    if repository == "ap_management_vlan_repository" and method == "list_export_rows":
        from netconsole.core.database import Database
        from netconsole.repositories.ap_management_vlan_repository import (
            ApManagementVlanRepository,
        )

        return ApManagementVlanRepository(Database(db_path)).list_export_rows()
    if repository == "ac_repository" and method == "list_station_online_summary_history":
        from netconsole.core.database import Database
        from netconsole.repositories.ac_repository import AcRepository

        site_name = str(filters.get("site_name") or "").strip() or None
        return AcRepository(Database(db_path)).list_station_online_summary_history(site_name, int(filters.get("limit") or 1_000_000))
    if repository == "ac_repository" and method == "list_fit_ap_history":
        from netconsole.core.database import Database
        from netconsole.repositories.ac_repository import AcRepository
        from netconsole.services.history_export_service import OPTICAL_HISTORY_COLORS, history_display_value

        history_kind = str(filters.get("history_kind") or "")
        ap_uuid = str(filters.get("ap_uuid") or "")
        color_field = str(filters.get("color_field") or "")
        language = str(filters.get("language") or "zh")
        rows = AcRepository(Database(db_path)).list_fit_ap_history_page(history_kind, ap_uuid, limit=1_000_000)
        result = []
        for row in rows:
            display_row = {key: history_display_value(row, key, color_field or None, language) for key in row}
            if color_field:
                display_row["__row_fill"] = OPTICAL_HISTORY_COLORS.get(str(row.get(color_field) or ""), "")
            result.append(display_row)
        return result
    if repository == "device_fact_repository" and method == "list_trackside_interface_history":
        from netconsole.core.database import Database
        from netconsole.repositories.device_fact_repository import DeviceFactRepository
        from netconsole.repositories.device_repository import DeviceRepository

        database = Database(db_path)
        device_uuid = str(filters.get("device_uuid") or "")
        interface_name = str(filters.get("interface_name") or "")
        device = next((item for item in DeviceRepository(database).list() if str(item.device_uuid or "") == device_uuid), None)
        rows = DeviceFactRepository(database).list_object_history_page("optical", device_uuid, interface_name, limit=1_000_000)
        return [
            {
                **row,
                "source_device_name": device.name if device is not None else row.get("device_uuid"),
                "source_device_id": device_uuid,
                "host": device.ip_address if device is not None else "",
                "optical_status": row.get("status"),
                "session_id": row.get("collect_run_uuid"),
            }
            for row in rows
        ]
    raise ValueError(f"不支持的 repository_query：{repository}.{method}")


def _ac_repository(payload: Mapping[str, Any]):
    from netconsole.core.database import Database
    from netconsole.repositories.ac_repository import AcRepository

    return AcRepository(Database(Path(str(payload.get("db_path") or ""))))


def _value_for_row(row: Any, key: str) -> object:
    if isinstance(row, Mapping):
        value = row.get(key)
    elif isinstance(row, Sequence) and not isinstance(row, (str, bytes)):
        try:
            value = row[int(key)]
        except (ValueError, IndexError):
            value = ""
    else:
        value = ""
    return "" if value is None else value


def _apply_row_fill(worksheet, row: Any, row_fill_field: str, row_fill_colors: Mapping[str, str]) -> None:
    if not row_fill_field or not isinstance(row, Mapping):
        return
    raw = str(row.get(row_fill_field) or "").strip()
    color = row_fill_colors.get(raw, raw)
    if not color:
        return
    normalized = color.lstrip("#").upper()
    if len(normalized) not in {6, 8}:
        return
    from openpyxl.styles import PatternFill

    fill = PatternFill(fill_type="solid", fgColor=normalized[-6:])
    for cell in worksheet[worksheet.max_row]:
        cell.fill = fill


def _apply_xlsx_column_styles(
    worksheet,
    row_number: int,
    columns: list[dict[str, Any]],
) -> None:
    from openpyxl.styles import Alignment

    for column_number, column in enumerate(columns, start=1):
        cell = worksheet.cell(row=row_number, column=column_number)
        number_format = str(column.get("number_format") or "")
        if number_format:
            cell.number_format = number_format
        if column.get("wrap") or column.get("horizontal"):
            cell.alignment = Alignment(
                horizontal=str(column.get("horizontal") or "left"),
                vertical="center",
                wrap_text=bool(column.get("wrap")),
            )


def _is_bold_export_row(row: Any, sheet_payload: Mapping[str, Any]) -> bool:
    if not isinstance(row, Mapping):
        return False
    field = str(sheet_payload.get("bold_row_field") or "")
    if not field:
        return False
    values = {
        str(value)
        for value in sheet_payload.get("bold_row_values") or []
    }
    return str(row.get(field) or "") in values


def _zip_entries(payload: Mapping[str, Any]) -> list[tuple[Path, str]]:
    source = payload.get("source") if isinstance(payload.get("source"), Mapping) else {}
    if str(source.get("type") or "") == "file_or_directory":
        items = source.get("paths") or []
        base_dir_value = source.get("base_dir") or payload.get("base_dir")
        exclude_names = {str(name) for name in source.get("exclude_names") or []}
    else:
        items = payload.get("items") or []
        base_dir_value = payload.get("base_dir")
        exclude_names = {str(name) for name in payload.get("exclude_names") or []}
    base_dir = Path(str(base_dir_value)).resolve() if base_dir_value else None
    entries: list[tuple[Path, str]] = []
    for item in items:
        source = Path(str(item)).resolve()
        if not source.exists():
            continue
        if source.is_dir():
            for child in source.rglob("*"):
                if child.is_file() and child.name not in exclude_names:
                    entries.append((child, _archive_name(child, base_dir or source.parent)))
        elif source.name not in exclude_names:
            entries.append((source, _archive_name(source, base_dir or source.parent)))
    return entries


def _archive_name(path: Path, base_dir: Path) -> str:
    try:
        return path.relative_to(base_dir).as_posix()
    except ValueError:
        return path.name


def _safe_sheet_title(title: str) -> str:
    value = "".join("_" if char in "[]:*?/\\\"" else char for char in title).strip() or "Sheet1"
    return value[:31]


def _effective_suffix(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".tmp" and path.stem:
        return Path(path.stem).suffix.lower()
    return suffix


def _parse_log_line(line: str) -> dict[str, str] | None:
    parts = line.split(" | ", 3)
    if len(parts) != 4:
        return None
    time, level, event, detail = parts
    return {"time": time, "level": level, "event": event, "detail": detail}


def _emit(progress: ProgressCallback | None, stage: str, current: int, total: int, message: str) -> None:
    if progress:
        progress(stage, current, total, message)


def _check_cancel(should_cancel: CancelCallback | None) -> None:
    if should_cancel and should_cancel():
        raise ExportCancelled("导出已取消")


def replace_output(tmp_path: Path, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    os.replace(tmp_path, output_path)
