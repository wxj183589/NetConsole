from __future__ import annotations

import warnings
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def load_workbook_without_unsupported_image_warning(path: str | Path, **kwargs: Any):
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message=r".*wmf image format is not supported so the image is being dropped.*",
            category=UserWarning,
        )
        return load_workbook(path, **kwargs)
