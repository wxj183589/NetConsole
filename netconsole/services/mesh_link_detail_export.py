from __future__ import annotations

import json
from pathlib import Path
from typing import Callable, Iterable

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from netconsole.models.mesh_log_models import format_mac_h3c
from netconsole.services.excel_report_utils import format_link_state
from netconsole.ui.table.table_autosize_engine import excel_column_width

MESH_LINK_EXPORT_ACTIVE_FONT_COLOR = "15803D"
MESH_LINK_EXPORT_GROUP_FILL_1 = "FFFFFF"
MESH_LINK_EXPORT_GROUP_FILL_2 = "F3F4F6"

ProgressCallback = Callable[[int, int, str], None]
CancelCallback = Callable[[], bool]


class MeshLinkDetailExportCancelled(Exception):
    pass


LINK_DETAIL_COLUMNS: tuple[tuple[str, str], ...] = (
    ("序号", "record_seq"),
    ("采样时间", "sample_time"),
    ("Radio", "radio"),
    ("链路状态", "link_state"),
    ("Peer MAC", "peer_mac"),
    ("对端AP名称", "peer_ap_name"),
    ("对端AP MAC", "peer_ap_mac"),
    ("归属站点", "peer_site"),
    ("归属区间", "belong_section"),
    ("归属类型", "belong_type"),
    ("对端射频口", "peer_radio"),
    ("建链时间", "establish_time"),
    ("链路时长", "duration_text"),
    ("链路数", "link_count"),
    ("MR侧RSSI", "local_rssi_db"),
    ("对端RSSI", "peer_rssi_db"),
    ("MR侧CPU", "local_cpu_percent"),
    ("对端CPU", "peer_cpu_percent"),
    ("MR侧内存", "local_mem_percent"),
    ("对端内存", "peer_mem_percent"),
    ("MR侧发送繁忙度", "local_tx_busy"),
    ("MR侧接收繁忙度", "local_rx_busy"),
    ("对端发送繁忙度", "peer_tx_busy"),
    ("对端接收繁忙度", "peer_rx_busy"),
    ("源文件", "archived_filename"),
    ("源行号", "source_line_number"),
)


ACTIVE_BUILD_ORDER_COLUMNS: tuple[tuple[str, str], ...] = (
    ("序号", "sequence"),
    ("Radio", "radio"),
    ("主链路 Peer MAC", "active_peer_mac"),
    ("对端AP名称", "peer_ap_name"),
    ("归属站点", "peer_site"),
    ("对端射频口", "peer_radio"),
    ("建链开始时间", "build_start_time"),
    ("建链结束时间", "build_end_time"),
    ("主链路持续时长(秒)", "main_link_duration_seconds"),
    ("日志上报时长(秒)", "reported_duration_seconds"),
    ("采样点数", "sample_count"),
    ("MR侧平均RSSI", "avg_mr_rssi"),
    ("MR侧最低RSSI", "min_mr_rssi"),
    ("MR侧最高RSSI", "max_mr_rssi"),
    ("发送繁忙度", "avg_tx_busy"),
    ("接收繁忙度", "avg_rx_busy"),
    ("配置切换时间(ms)", "main_link_switch_time_ms"),
    ("短时判定容差(ms)", "short_link_tolerance_ms"),
    ("是否同AP射频切换", "is_same_physical_ap_radio_switch"),
    ("建链结果", "build_result"),
    ("判定原因", "judge_reason"),
    ("是否AP回切", "is_ap_return_event"),
    ("是否乒乓异常", "is_pingpong_abnormal"),
    ("乒乓类型", "pingpong_type"),
    ("乒乓组ID", "pingpong_group_id"),
    ("乒乓返回耗时(ms)", "pingpong_return_duration_ms"),
    ("中间AP驻留时长(ms)", "middle_ap_dwell_ms"),
    ("前一AP", "previous_ap"),
    ("中间AP", "middle_ap"),
    ("返回AP", "return_ap"),
    ("乒乓次数", "pingpong_count"),
    ("乒乓判定原因", "pingpong_judgment_reason"),
    ("源文件", "source_file"),
)


def link_detail_row_values(row: dict[str, object]) -> list[object]:
    metrics = _json_dict(row.get("metrics_json"))
    peer = format_mac_h3c(row.get("peer_mac_normalized")) if row.get("peer_mac_normalized") else row.get("peer_mac_raw")
    return [
        row.get("record_seq") or row.get("source_line_number"),
        row.get("sample_time"),
        row.get("radio"),
        format_link_state(row.get("link_state")),
        row.get("peer_mac_raw") or peer,
        row.get("peer_ap_name") or "-",
        format_mac_h3c(row.get("peer_ap_mac")) if row.get("peer_ap_mac") else "-",
        row.get("peer_site") or "-",
        row.get("belong_section") or row.get("peer_section") or "-",
        _belong_type_text(row.get("belong_type")),
        row.get("peer_radio") or row.get("peer_radio_label") or "-",
        row.get("establish_time"),
        row.get("duration_text"),
        row.get("link_count"),
        metrics.get("local_rssi_db"),
        metrics.get("peer_rssi_db"),
        metrics.get("local_cpu_percent"),
        metrics.get("peer_cpu_percent"),
        metrics.get("local_mem_percent"),
        metrics.get("peer_mem_percent"),
        metrics.get("local_tx_busy"),
        metrics.get("local_rx_busy"),
        metrics.get("peer_tx_busy"),
        metrics.get("peer_rx_busy"),
        row.get("archived_filename"),
        row.get("source_line_number"),
    ]


def active_build_order_row_values(row: dict[str, object]) -> list[object]:
    result = str(row.get("build_result") or "")
    return [
        row.get("sequence"),
        row.get("radio"),
        format_mac_h3c(row.get("active_peer_mac")) if row.get("active_peer_mac") else "",
        row.get("peer_ap_name") or "",
        row.get("peer_site") or "",
        row.get("peer_radio") or "",
        row.get("build_start_time") or "",
        row.get("build_end_time") or "",
        row.get("main_link_duration_seconds") or "",
        row.get("reported_duration_seconds") or "",
        row.get("sample_count") or "",
        _rssi_stat_value(row.get("avg_mr_rssi")),
        _rssi_stat_value(row.get("min_mr_rssi")),
        _rssi_stat_value(row.get("max_mr_rssi")),
        row.get("avg_tx_busy") or "",
        row.get("avg_rx_busy") or "",
        row.get("main_link_switch_time_ms") or "",
        row.get("short_link_tolerance_ms") or "",
        "是" if row.get("is_same_physical_ap_radio_switch") else "否",
        _build_result_text(result),
        row.get("judge_reason") or "",
        "是" if row.get("is_ap_return_event") else "否",
        "是" if row.get("is_pingpong_abnormal") else "否",
        row.get("pingpong_type") or "",
        row.get("pingpong_group_id") or "",
        row.get("pingpong_return_duration_ms") or "",
        row.get("middle_ap_dwell_ms") or "",
        row.get("previous_ap") or "",
        row.get("middle_ap") or "",
        row.get("return_ap") or "",
        row.get("pingpong_count") or "",
        row.get("pingpong_judgment_reason") or "",
        row.get("source_file") or "",
    ]


def export_mesh_link_details_xlsx(
    path: Path,
    rows: Iterable[dict[str, object]],
    active_build_order_rows: Iterable[dict[str, object]] | None = None,
    *,
    total_rows: int | None = None,
    progress_callback: ProgressCallback | None = None,
    should_cancel: CancelCallback | None = None,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook(write_only=True)
    _write_link_detail_sheet(
        workbook.create_sheet("链路明细"),
        LINK_DETAIL_COLUMNS,
        rows,
        total_rows=max(int(total_rows or 0), 0),
        progress_callback=progress_callback,
        should_cancel=should_cancel,
    )
    _write_basic_sheet(
        workbook.create_sheet("主链路建链顺序"),
        ACTIVE_BUILD_ORDER_COLUMNS,
        active_build_order_rows or [],
        active_build_order_row_values,
        should_cancel=should_cancel,
    )
    _raise_if_cancelled(should_cancel)
    workbook.save(path)


def _write_link_detail_sheet(
    worksheet,
    columns: tuple[tuple[str, str], ...],
    rows: Iterable[dict[str, object]],
    *,
    total_rows: int,
    progress_callback: ProgressCallback | None,
    should_cancel: CancelCallback | None,
) -> int:
    headers = [header for header, _field in columns]
    alignment = Alignment(horizontal="center", vertical="center")
    header_font = Font(bold=True)
    active_font = Font(bold=True, color=MESH_LINK_EXPORT_ACTIVE_FONT_COLOR)
    standby_font = Font(bold=False)
    group_fill_1 = PatternFill(fill_type="solid", fgColor=MESH_LINK_EXPORT_GROUP_FILL_1)
    group_fill_2 = PatternFill(fill_type="solid", fgColor=MESH_LINK_EXPORT_GROUP_FILL_2)
    widths = _initial_widths(columns)
    worksheet.freeze_panes = "A2"
    worksheet.append(_styled_cells(worksheet, headers, alignment, header_font))
    group_index = -1
    previous_group_key: tuple[object, object, object] | None = None
    written = 0
    for row in rows:
        _raise_if_cancelled(should_cancel)
        values = link_detail_row_values(row)
        _update_widths(widths, columns, values)
        group_key = (row.get("source_file_id"), row.get("sample_time"), row.get("radio"))
        if group_key != previous_group_key:
            group_index += 1
            previous_group_key = group_key
        fill = group_fill_1 if group_index % 2 == 0 else group_fill_2
        font = active_font if str(row.get("link_state") or "").upper() == "ACTIVE" else standby_font
        worksheet.append(_styled_cells(worksheet, values, alignment, font, fill))
        written += 1
        if progress_callback is not None and (written % 500 == 0 or (total_rows and written >= total_rows)):
            progress_callback(written, total_rows, "mesh_analysis.export_progress_write_links")
    data_row_count = written
    if written == 0:
        values = ["未找到可导出的链路明细数据；请检查筛选条件、源文件和解析结果。"] + ["" for _ in headers[1:]]
        _update_widths(widths, columns, values)
        worksheet.append(_styled_cells(worksheet, values, alignment, standby_font))
        data_row_count = 1
    worksheet.auto_filter.ref = _sheet_range(len(columns), data_row_count + 1)
    _apply_widths(worksheet, widths)
    if progress_callback is not None:
        progress_callback(written, total_rows, "mesh_analysis.export_progress_write_links")
    return written


def _write_basic_sheet(worksheet, columns: tuple[tuple[str, str], ...], rows: Iterable[dict[str, object]], row_factory, should_cancel: CancelCallback | None = None) -> None:
    headers = [header for header, _field in columns]
    alignment = Alignment(horizontal="center", vertical="center")
    header_font = Font(bold=True)
    widths = _initial_widths(columns)
    worksheet.freeze_panes = "A2"
    worksheet.append(_styled_cells(worksheet, headers, alignment, header_font))
    written = 0
    for row in rows:
        _raise_if_cancelled(should_cancel)
        values = row_factory(row)
        _update_widths(widths, columns, values)
        worksheet.append(_styled_cells(worksheet, values, alignment))
        written += 1
    data_row_count = written
    if written == 0:
        values = ["未生成主链路建链顺序；请确认当前日志存在 ACTIVE 主链路采样。"] + ["" for _ in headers[1:]]
        _update_widths(widths, columns, values)
        worksheet.append(_styled_cells(worksheet, values, alignment))
        data_row_count = 1
    worksheet.auto_filter.ref = _sheet_range(len(columns), data_row_count + 1)
    _apply_widths(worksheet, widths)


def _styled_cells(worksheet, values: list[object] | tuple[object, ...], alignment: Alignment, font: Font | None = None, fill: PatternFill | None = None) -> list[WriteOnlyCell]:
    cells: list[WriteOnlyCell] = []
    for value in values:
        cell = WriteOnlyCell(worksheet, value="" if value is None else value)
        cell.alignment = alignment
        if font is not None:
            cell.font = font
        if fill is not None:
            cell.fill = fill
        cells.append(cell)
    return cells


def _initial_widths(columns: tuple[tuple[str, str], ...]) -> list[float]:
    return [excel_column_width(header, maximum=80.0, field=field, header=header) for header, field in columns]


def _update_widths(widths: list[float], columns: tuple[tuple[str, str], ...], values: list[object]) -> None:
    for index, value in enumerate(values):
        if index < len(widths):
            header, field = columns[index]
            widths[index] = max(widths[index], excel_column_width(value, maximum=80.0, field=field, header=header))


def _apply_widths(worksheet, widths: list[float]) -> None:
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = min(max(width, 8.0), 80.0)


def _sheet_range(column_count: int, row_count: int) -> str:
    return f"A1:{get_column_letter(max(column_count, 1))}{max(row_count, 1)}"


def _raise_if_cancelled(should_cancel: CancelCallback | None) -> None:
    if should_cancel is not None and should_cancel():
        raise MeshLinkDetailExportCancelled("导出已取消")


def _build_result_text(value: str) -> str:
    if value == "normal":
        return "正常"
    if value == "short":
        return "短时建链"
    if value == "same_ap_radio_switch":
        return "同AP射频切换"
    return value


def _belong_type_text(value: object) -> str:
    text = str(value or "").strip().lower()
    mapping = {
        "station": "车站",
        "section": "区间",
        "yard": "场段",
        "area": "区域",
        "empty": "空链路",
        "unknown": "未知",
    }
    if not text:
        return "-"
    return mapping.get(text, str(value))


def _rssi_stat_value(value: object) -> object:
    return "N/A" if value is None or value == "" else value


def _sample_group_indexes(rows: list[dict[str, object]]) -> list[int]:
    result: list[int] = []
    group_indexes: dict[tuple[object, object], int] = {}
    for row in rows:
        raw_group = row.get("sample_group_index")
        if raw_group not in (None, ""):
            try:
                result.append(int(raw_group))
                continue
            except (TypeError, ValueError):
                pass
        key = (row.get("sample_time"), row.get("radio"))
        if key not in group_indexes:
            group_indexes[key] = len(group_indexes)
        result.append(group_indexes[key])
    return result


def _json_dict(value: object) -> dict[str, object]:
    if isinstance(value, dict):
        return value
    if not value:
        return {}
    try:
        parsed = json.loads(str(value))
    except json.JSONDecodeError:
        return {}
    return parsed if isinstance(parsed, dict) else {}
