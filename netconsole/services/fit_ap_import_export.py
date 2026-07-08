from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from netconsole.repositories.ac_repository import AcRepository
from netconsole.services.ap_extension_import import (
    ApExtensionImportService,
    ImportPreview,
    standard_export_row,
    standard_template_headers,
)
from netconsole.utils.excel_workbook import load_workbook_without_unsupported_image_warning
from netconsole.utils.mileage import format_track_mileage, mileage_storage_text, parse_track_mileage
from netconsole.utils.station_normalize import normalize_station_value


def normalize_ap_direction(value: str) -> str:
    text = str(value or "").strip()
    if text.upper() == "CW":
        return "上行"
    if text.upper() == "CT":
        return "下行"
    if text.upper() in {"ZDK", "YDK", "CDK", "RDK"}:
        return text.upper()
    return text


AP_METADATA_LEGACY_IMPORT_FIELDS = ["AP名称", "AP_MAC", "归属站点", "里程", "点位说明", "上下行"]
AP_METADATA_IMPORT_FIELDS = [
    "AP名称",
    "AP_MAC",
    "归属类型",
    "归属站点",
    "归属区间",
    "区间起点站",
    "区间终点站",
    "场段",
    "区域",
    "网络",
    "线别",
    "里程",
    "点位说明",
    "方向",
    "备注",
]
AP_EXPORT_FIELDS = [
    "AP名称",
    "APID",
    "AP_IP",
    "AP_MAC",
    "型号",
    "SN",
    "状态",
    "AP组",
    "在线时长",
    "RID1信道",
    "RID1频宽",
    "RID1功率",
    "RID2信道",
    "RID2频宽",
    "RID2功率",
    "归属站点",
    "归属区间",
    "归属类型",
    "里程",
    "点位说明",
    "上下行",
    "更新时间",
]

AP_EXTENSION_TEMPLATE_FIELDS = [
    "AP名称",
    "AP_MAC",
    "归属类型",
    "归属站点",
    "归属区间",
    "区间起点站",
    "区间终点站",
    "场段",
    "区域",
    "网络",
    "线别",
    "里程",
    "点位说明",
    "方向",
    "备注",
]
AP_EXTENSION_TEMPLATE_EDITABLE_FIELDS = {"归属类型", "归属站点", "归属区间", "区间起点站", "区间终点站", "场段", "区域", "网络", "线别", "里程", "点位说明", "方向", "上下行", "备注"}


@dataclass(frozen=True)
class ApMetadataImportResult:
    updated: int
    skipped: int
    errors: list[str]


def make_fit_ap_export_filename(site_name: str, now: datetime | None = None) -> str:
    timestamp = (now or datetime.now()).strftime("%Y-%m-%d-%H%M")
    safe_site_name = "".join("_" if char in '<>:"/\\|?*' or ord(char) < 32 else char for char in site_name).strip().strip(".") or "site"
    return f"{safe_site_name}_fit_ap_{timestamp}.csv"


def make_ap_extension_template_filename(ac_name: str | None, now: datetime | None = None) -> str:
    timestamp = (now or datetime.now()).strftime("%Y%m%d_%H%M%S")
    safe_ac = "".join("_" if char in '<>:"/\\|?*' or ord(char) < 32 else char for char in str(ac_name or "AC")).strip().strip(".") or "AC"
    return f"AP扩展信息模板_{safe_ac}_{timestamp}.xlsx"


class FitApImportExportService:
    def __init__(self, repository: AcRepository) -> None:
        self.repository = repository
        self.ap_extension_import_service = ApExtensionImportService()

    def preview_ap_extension_import(self, path: Path, import_mode: str) -> ImportPreview:
        return self.ap_extension_import_service.preview_file(path, import_mode=import_mode)

    def commit_ap_extension_import(
        self,
        preview: ImportPreview,
        duplicate_strategy: str = "update_by_priority",
    ) -> dict[str, int | str]:
        return self.repository.import_ap_extension_points(
            preview.standard_rows,
            source_file=preview.file_name,
            template_type=preview.template_type,
            duplicate_strategy=duplicate_strategy,
        )

    def import_metadata_csv(self, path: Path) -> ApMetadataImportResult:
        with Path(path).open("r", newline="", encoding="utf-8-sig") as file:
            rows = list(csv.reader(file))
        if not rows:
            return ApMetadataImportResult(0, 0, [])
        return self.import_metadata_rows([header.strip() for header in rows[0]], rows[1:])

    def import_metadata_file(self, path: Path) -> ApMetadataImportResult:
        suffix = Path(path).suffix.casefold()
        if suffix == ".xlsx":
            sheet = load_workbook_without_unsupported_image_warning(path, data_only=True).active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                return ApMetadataImportResult(0, 0, [])
            headers = [str(value or "").strip() for value in rows[0]]
            values = [[str(value or "").strip() for value in row] for row in rows[1:]]
            return self.import_metadata_rows(headers, values)
        return self.import_metadata_csv(path)

    def import_metadata_rows(self, headers: list[str], rows: list[list[object]]) -> ApMetadataImportResult:
        if not _is_supported_metadata_header(headers):
            raise ValueError("Unsupported AP metadata template header")
        updated = 0
        skipped = 0
        errors: list[str] = []
        entity_rows = self.repository.list_ap_entities()
        for line_number, values in enumerate(rows, start=2):
            payload = {field: (_text(values[index]) if index < len(values) else "") for index, field in enumerate(headers)}
            ap_mac = normalize_ap_mac(payload["AP_MAC"])
            if not ap_mac:
                skipped += 1
                errors.append(f"Row {line_number}: AP_MAC is empty or invalid")
                continue
            mileage = parse_track_mileage(payload["里程"])
            direction = normalize_ap_direction(payload.get("方向") or payload.get("上下行")) or mileage.prefix or ""
            matched = self.repository.update_ap_entity_extension_by_mac(
                ap_mac,
                {
                    "station": payload["归属站点"],
                    "milestone": mileage_storage_text(payload["里程"]),
                    "location_note": payload["点位说明"],
                    "direction": direction,
                },
            )
            if not matched:
                skipped += 1
                errors.append(f"Row {line_number}: AP_MAC {payload['AP_MAC']} not matched")
                continue
            entity = _lookup_ap_entity_by_mac(entity_rows, ap_mac)
            metadata_payload = {
                "ap_uuid": (entity or {}).get("ap_uuid"),
                "ap_name": payload.get("AP名称") or (entity or {}).get("ap_name"),
                "site_name": payload["归属站点"],
                "belong_type": _normalize_belong_type(payload.get("归属类型")) or _infer_belong_type(payload.get("归属站点"), payload.get("归属区间"), payload.get("场段"), payload.get("区域")),
                "belong_section": payload.get("归属区间"),
                "section_start_station": payload.get("区间起点站"),
                "section_end_station": payload.get("区间终点站"),
                "yard_name": payload.get("场段"),
                "area_name": payload.get("区域"),
                "network_domain": payload.get("网络"),
                "line_side": payload.get("线别"),
                "mileage": mileage_storage_text(payload["里程"]),
                "location_note": payload["点位说明"],
                "direction": direction,
                "remark": payload.get("备注"),
            }
            if metadata_payload.get("ap_uuid") or metadata_payload.get("ap_name"):
                self.repository.upsert_fit_ap_metadata(metadata_payload)
            updated += 1
        return ApMetadataImportResult(updated, skipped, errors)

    def export_ap_csv(self, path: Path, rows: list[dict[str, object | None]]) -> None:
        with Path(path).open("w", newline="", encoding="utf-8-sig") as file:
            writer = csv.writer(file)
            writer.writerow(AP_EXPORT_FIELDS)
            for row in rows:
                writer.writerow(
                    [
                        row.get("ap_name") or "",
                        row.get("apid") or "",
                        row.get("ap_ip") or "",
                        row.get("ap_mac") or "",
                        row.get("model") or "",
                        row.get("serial_number") or "",
                        row.get("state_display") or row.get("state") or "",
                        row.get("group_name") or "",
                        row.get("online_time") or "",
                        row.get("rid1_channel") or "",
                        row.get("rid1_bandwidth") or "",
                        row.get("rid1_tx_power") or "",
                        row.get("rid2_channel") or "",
                        row.get("rid2_bandwidth") or "",
                        row.get("rid2_tx_power") or "",
                        row.get("site") or "",
                        row.get("section_name") or row.get("belong_section") or row.get("extension_section_name") or "",
                        _belong_type_label(row.get("belong_type") or row.get("extension_belong_type")),
                        _format_ap_mileage(row.get("mileage"), row.get("direction"), row.get("extension_line_side") or row.get("line_side")),
                        row.get("location_note") or "",
                        row.get("direction") or "",
                        row.get("updated_at") or "",
                    ]
                )

    def export_ap_extension_template_xlsx(
        self,
        path: Path,
        rows: list[dict[str, object | None]],
        ap_entities: list[dict[str, object | None]] | None = None,
    ) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "AP扩展信息模板"
        sheet.append(AP_EXTENSION_TEMPLATE_FIELDS)
        editable_fill = PatternFill(fill_type="solid", fgColor="FFF7D6")
        entity_lookup = _build_ap_entity_lookup(ap_entities or [])
        for row in rows:
            entity = _lookup_ap_entity(row, entity_lookup)
            sheet.append(_ap_extension_template_row(row, entity))
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            if cell.value in AP_EXTENSION_TEMPLATE_EDITABLE_FIELDS:
                cell.fill = editable_fill
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if sheet.cell(row=1, column=cell.column).value in AP_EXTENSION_TEMPLATE_EDITABLE_FIELDS:
                    cell.fill = editable_fill
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        _auto_width(sheet)
        workbook.save(path)

    def export_standard_ap_extension_xlsx(
        self,
        path: Path,
        rows: list[dict[str, object | None]],
    ) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "FIT-AP扩展信息"
        headers = list(standard_template_headers())
        sheet.append(headers)
        editable_fill = PatternFill(fill_type="solid", fgColor="FFF7D6")
        for row in rows:
            sheet.append(standard_export_row(row))
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.fill = editable_fill
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        _auto_width(sheet)
        workbook.save(path)


def _ap_extension_template_row(row: dict[str, object | None], entity: dict[str, object | None] | None = None) -> list[str]:
    station = normalize_station_value(entity) or normalize_station_value(row)
    direction = _text(row.get("direction") or (entity or {}).get("direction"))
    line_side = _text(row.get("extension_line_side") or row.get("line_side"))
    section = _text(row.get("section_name") or row.get("belong_section") or row.get("extension_section_name"))
    belong_type = _text(row.get("belong_type") or row.get("extension_belong_type"))
    return [
        _text(row.get("ap_name") or (entity or {}).get("ap_name")),
        normalize_ap_mac(row.get("ap_mac") or (entity or {}).get("ap_mac")),
        _belong_type_label(belong_type),
        station,
        section,
        _text(row.get("section_start_station") or row.get("extension_section_start_station")),
        _text(row.get("section_end_station") or row.get("extension_section_end_station")),
        _text(row.get("yard_name") or row.get("extension_yard_name")),
        _text(row.get("area_name") or row.get("extension_area_name")),
        _text(row.get("network_domain") or row.get("extension_network_domain")),
        line_side,
        _format_ap_mileage(_first_non_empty(row.get("mileage"), row.get("milestone"), (entity or {}).get("milestone")), direction, line_side),
        _text(row.get("location_note") or (entity or {}).get("location_note")),
        direction,
        _text(row.get("remark") or row.get("extension_remark")),
    ]


def _is_supported_metadata_header(headers: list[str]) -> bool:
    if headers == AP_METADATA_LEGACY_IMPORT_FIELDS:
        return True
    required = {"AP名称", "AP_MAC", "归属站点", "里程", "点位说明"}
    direction_fields = {"方向", "上下行"}
    known = set(AP_METADATA_IMPORT_FIELDS) | set(AP_EXTENSION_TEMPLATE_FIELDS) | direction_fields
    header_set = set(headers)
    return required <= header_set and bool(header_set & direction_fields) and header_set <= known


def _build_ap_entity_lookup(rows: list[dict[str, object | None]]) -> dict[tuple[str, str], dict[str, object | None]]:
    lookup: dict[tuple[str, str], dict[str, object | None]] = {}
    for row in rows:
        ap_mac = normalize_ap_mac(row.get("ap_mac"))
        if ap_mac:
            lookup[("ap_mac", ap_mac)] = row
    return lookup


def _lookup_ap_entity(row: dict[str, object | None], lookup: dict[tuple[str, str], dict[str, object | None]]) -> dict[str, object | None] | None:
    return lookup.get(("ap_mac", normalize_ap_mac(row.get("ap_mac"))))


def _auto_width(sheet) -> None:
    for column_cells in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells) + 2
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width, 10), 36)


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _format_ap_mileage(value: object, direction: object = None, line_side: object = None) -> str:
    display = format_track_mileage(value, direction=str(direction or ""), line_side=str(line_side or ""))
    return "" if display == "-" else display


def _first_non_empty(*values: object) -> object:
    for value in values:
        if value not in (None, ""):
            return value
    return ""


def _lookup_ap_entity_by_mac(rows: list[dict[str, object | None]], ap_mac: str) -> dict[str, object | None] | None:
    normalized = normalize_ap_mac(ap_mac)
    if not normalized:
        return None
    return next((row for row in rows if normalize_ap_mac(row.get("ap_mac")) == normalized), None)


def _normalize_belong_type(value: object) -> str:
    text = str(value or "").strip().casefold()
    if not text:
        return ""
    if text in {"station", "站点"}:
        return "station"
    if text in {"section", "interval", "区间"}:
        return "section"
    if text in {"yard", "场段", "场段/库内", "车辆段", "停车场", "库内"}:
        return "yard"
    if text in {"unknown", "未知"}:
        return "unknown"
    return ""


def _infer_belong_type(station: object, section: object, yard_name: object, area_name: object) -> str:
    if _text(yard_name) or _text(area_name):
        return "yard"
    if _text(section) and not _text(station):
        return "section"
    if _text(station):
        return "station"
    return "unknown"


def _belong_type_label(value: object) -> str:
    normalized = _normalize_belong_type(value)
    return {
        "station": "站点",
        "section": "区间",
        "yard": "场段/库内",
        "unknown": "未知",
    }.get(normalized, _text(value))


def normalize_ap_mac(value: object) -> str:
    import re

    hex_text = re.sub(r"[^0-9a-fA-F]", "", str(value or ""))
    if len(hex_text) != 12:
        return ""
    hex_text = hex_text.casefold()
    return f"{hex_text[0:4]}-{hex_text[4:8]}-{hex_text[8:12]}"
