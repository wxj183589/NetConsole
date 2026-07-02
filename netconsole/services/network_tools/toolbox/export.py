from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def export_rows_csv(path: Path, headers: list[str], rows: list[dict[str, object]]) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        for row in rows:
            writer.writerow([row.get(header, "") for header in headers])
    return path


def export_rows_xlsx(path: Path, headers: list[str], rows: list[dict[str, object]], *, sheet_name: str = "结果") -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name[:31] or "结果"
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        values = [str(cell.value or "") for cell in column]
        width = min(max((len(value) for value in values), default=8) + 2, 42)
        sheet.column_dimensions[column[0].column_letter].width = width
    workbook.save(path)
    return path
