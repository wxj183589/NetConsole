from __future__ import annotations


def apply_basic_sheet_style(sheet, *, header_row: int = 1, column_count: int | None = None) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    max_column = column_count or sheet.max_column
    if header_row > 1:
        for cell in sheet[1]:
            cell.alignment = alignment
            cell.font = Font(bold=True, size=13)
    for row in sheet.iter_rows():
        sheet.row_dimensions[row[0].row].height = 24 if row[0].row <= header_row else 22
        for cell in row:
            cell.alignment = alignment
            cell.border = border
            if cell.row == header_row and cell.column <= max_column:
                cell.font = header_font
                cell.fill = header_fill
