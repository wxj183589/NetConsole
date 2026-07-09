from __future__ import annotations

import csv
import json
import os
import shutil
import zipfile
from collections.abc import Callable, Iterable, Mapping, Sequence
from pathlib import Path
from typing import Any

from netconsole.services.excel_autosize import apply_worksheet_column_widths

ProgressCallback = Callable[[str, int, int, str], None]
CancelCallback = Callable[[], bool]


class ExportCancelled(RuntimeError):
    pass


def _path_resolver_from_payload(payload: Mapping[str, Any]):
    from netconsole.core.paths import PathResolver

    app_root = str(payload.get("app_root") or "").strip() or None
    data_root = str(payload.get("data_root") or "").strip() or None
    return PathResolver(app_root=Path(app_root) if app_root else None, data_root=Path(data_root) if data_root else None)


def export_table_xlsx(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    sheets = [
        {
            "sheet_name": payload.get("sheet_name") or "Sheet1",
            "title": payload.get("title") or "",
            "columns": payload.get("columns") or [],
            "source": payload.get("source"),
            "rows": payload.get("rows") or [],
            "auto_width": payload.get("auto_width", True),
            "freeze_header": payload.get("freeze_header", True),
            "auto_filter": payload.get("auto_filter", True),
            "row_fill_field": payload.get("row_fill_field") or "",
            "row_fill_colors": payload.get("row_fill_colors") or {},
        }
    ]
    return export_multi_sheet_xlsx(path, {"sheets": sheets}, progress, should_cancel)


def export_multi_sheet_xlsx(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    from openpyxl import Workbook

    from netconsole.services.export.xlsx_style import apply_basic_sheet_style

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    sheets = list(payload.get("sheets") or [])
    if not sheets:
        sheets = [{"sheet_name": "Sheet1", "columns": [], "rows": []}]
    prepared_sheets = [(sheet, resolve_rows(sheet)) for sheet in sheets]
    total_rows = sum(len(rows) for _sheet, rows in prepared_sheets)
    written_rows = 0
    _emit(progress, "prepare", 0, total_rows, "正在准备工作簿")
    for sheet_index, (sheet_payload, rows) in enumerate(prepared_sheets, start=1):
        _check_cancel(should_cancel)
        columns = normalize_columns(sheet_payload.get("columns") or [])
        title = str(sheet_payload.get("title") or "")
        worksheet = workbook.create_sheet(_safe_sheet_title(str(sheet_payload.get("sheet_name") or f"Sheet{sheet_index}")))
        header_row = 2 if title else 1
        if title:
            worksheet.cell(row=1, column=1, value=title)
            if columns:
                worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        worksheet.append([column["title"] for column in columns])
        row_fill_field = str(sheet_payload.get("row_fill_field") or "")
        row_fill_colors = dict(sheet_payload.get("row_fill_colors") or {})
        for row_index, row in enumerate(rows, start=1):
            values = [_value_for_row(row, column["key"]) for column in columns]
            worksheet.append(values)
            _apply_row_fill(worksheet, row, row_fill_field, row_fill_colors)
            written_rows += 1
            if written_rows == total_rows or written_rows % 100 == 0:
                _emit(progress, "write_rows", written_rows, total_rows, f"正在写入数据 {written_rows}/{total_rows}")
                _check_cancel(should_cancel)
        if bool(sheet_payload.get("freeze_header", True)):
            worksheet.freeze_panes = f"A{header_row + 1}"
        if bool(sheet_payload.get("auto_filter", True)) and columns:
            worksheet.auto_filter.ref = worksheet.dimensions
        apply_basic_sheet_style(worksheet, header_row=header_row, column_count=len(columns))
        if bool(sheet_payload.get("auto_width", True)) and columns:
            apply_worksheet_column_widths(
                worksheet,
                [column["title"] for column in columns],
                rows,
                [column["key"] for column in columns],
                maximum=60,
            )
            for index, column in enumerate(columns, start=1):
                if column.get("width"):
                    worksheet.column_dimensions[worksheet.cell(row=header_row, column=index).column_letter].width = float(column["width"])
    _emit(progress, "save_workbook", total_rows, total_rows, "正在保存 Excel 文件")
    _check_cancel(should_cancel)
    workbook.save(path)
    return total_rows


def export_table_csv(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    columns = normalize_columns(payload.get("columns") or [])
    rows = resolve_rows(payload)
    total = len(rows)
    _emit(progress, "write_csv", 0, total, "正在写入 CSV")
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow([column["title"] for column in columns])
        for index, row in enumerate(rows, start=1):
            writer.writerow([_value_for_row(row, column["key"]) for column in columns])
            if index == total or index % 500 == 0:
                _emit(progress, "write_csv", index, total, f"正在写入 CSV {index}/{total}")
                _check_cancel(should_cancel)
    return total


def export_markdown_text(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    text = str(payload.get("text") or "")
    _emit(progress, "write_text", 0, 1, "正在写入文本")
    _check_cancel(should_cancel)
    path.write_text(text, encoding="utf-8")
    _emit(progress, "write_text", 1, 1, "文本导出完成")
    return len(text)


def export_config_diff_text(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    from netconsole.core.database import Database
    from netconsole.repositories.config_snapshot_repository import ConfigSnapshotRepository
    from netconsole.services.config_lifecycle_service import ConfigLifecycleService

    path.parent.mkdir(parents=True, exist_ok=True)
    _emit(progress, "config_diff", 0, 1, "正在生成配置差异")
    _check_cancel(should_cancel)
    database = Database(Path(str(payload.get("db_path") or "")))
    repository = ConfigSnapshotRepository(database)
    service = ConfigLifecycleService(str(payload.get("site_name") or ""), database, _path_resolver_from_payload(payload), repository)
    left = repository.get(int(payload.get("left_snapshot_id") or 0))
    right = repository.get(int(payload.get("right_snapshot_id") or 0))
    diff = service.compare_snapshots(left, right)
    path.write_text(diff.raw_diff, encoding="utf-8")
    _emit(progress, "config_diff", 1, 1, "配置差异导出完成")
    return len(diff.raw_diff)


def export_zip_files(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    entries = _zip_entries(payload)
    total = len(entries)
    _emit(progress, "zip_files", 0, total, "正在打包文件")
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as archive:
        for index, (source, arcname) in enumerate(entries, start=1):
            archive.write(source, arcname)
            if index == total or index % 20 == 0:
                _emit(progress, "zip_files", index, total, f"正在打包文件 {index}/{total}")
                _check_cancel(should_cancel)
    return total


def copy_file_export(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    source = Path(str(payload.get("source") or ""))
    if not source.exists():
        raise FileNotFoundError(f"源文件不存在：{source}")
    path.parent.mkdir(parents=True, exist_ok=True)
    _emit(progress, "copy_file", 0, 1, "正在复制文件")
    _check_cancel(should_cancel)
    shutil.copyfile(source, path)
    _emit(progress, "copy_file", 1, 1, "文件复制完成")
    return 1


def export_app_logs_csv(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    from netconsole.ui.logs.log_display import display_log_row
    from netconsole.ui.logs.log_pagination_engine import iter_logs

    log_path = Path(str(payload.get("log_path") or ""))
    keyword = str(payload.get("keyword") or "").strip() or None
    level = str(payload.get("level") or "").strip() or None
    path.parent.mkdir(parents=True, exist_ok=True)
    count = 0
    _emit(progress, "write_logs", 0, 0, "正在导出日志")
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["时间", "级别", "事件", "详情", "原始事件", "原始详情"])
        for row in iter_logs(log_path, keyword=keyword, level=level, parser=_parse_log_line):
            display = display_log_row(row)
            writer.writerow(
                [
                    display.get("time", ""),
                    display.get("display_level", display.get("level", "")),
                    display.get("display_event", display.get("event", "")),
                    display.get("display_detail", display.get("detail", "")),
                    display.get("raw_event", display.get("event", "")),
                    display.get("raw_detail", display.get("detail", "")),
                ]
            )
            count += 1
            if count % 500 == 0:
                _emit(progress, "write_logs", count, 0, f"正在导出日志 {count} 条")
                _check_cancel(should_cancel)
    return count


def export_device_csv(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    from netconsole.core.database import Database
    from netconsole.models.device import Device
    from netconsole.repositories.device_group_repository import DeviceGroupRepository
    from netconsole.repositories.device_repository import DeviceRepository
    from netconsole.services.device_import_export import DeviceImportExportService

    path.parent.mkdir(parents=True, exist_ok=True)
    devices_payload = list(payload.get("devices") or [])
    site_name = str(payload.get("site_name") or "")
    if devices_payload:
        devices = [Device.from_mapping(dict(row)) for row in devices_payload if isinstance(row, Mapping)]
        database = Database(Path(str(payload.get("db_path") or ":memory:")))
        group_repository = DeviceGroupRepository(database, site_name) if site_name else None
        service = DeviceImportExportService(DeviceRepository(database), group_repository)
    else:
        source = payload.get("source") if isinstance(payload.get("source"), Mapping) else {}
        db_path = str(payload.get("db_path") or source.get("db_path") or "")
        database = Database(Path(db_path))
        repository = DeviceRepository(database)
        group_repository = DeviceGroupRepository(database, site_name) if site_name else None
        filters = dict(payload.get("filters") or source.get("filters") or {})
        devices = repository.list(
            search=filters.get("search") or None,
            vendor=filters.get("vendor") or None,
            device_type=filters.get("device_type") or None,
            group_filter=filters.get("group_filter"),
        )
        service = DeviceImportExportService(repository, group_repository)
    _emit(progress, "write_device_csv", 0, len(devices), "正在导出设备 CSV")
    _check_cancel(should_cancel)
    service.export_csv(path, devices)
    _emit(progress, "write_device_csv", len(devices), len(devices), "设备 CSV 导出完成")
    return len(devices)


def export_device_template_csv(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    from netconsole.services.device_import_export import TEMPLATE_EXAMPLE_ROWS, TEMPLATE_FIELDS

    path.parent.mkdir(parents=True, exist_ok=True)
    _emit(progress, "write_device_template", 0, len(TEMPLATE_EXAMPLE_ROWS), "正在导出设备模板")
    _check_cancel(should_cancel)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(TEMPLATE_FIELDS)
        writer.writerows(TEMPLATE_EXAMPLE_ROWS)
    _emit(progress, "write_device_template", len(TEMPLATE_EXAMPLE_ROWS), len(TEMPLATE_EXAMPLE_ROWS), "设备模板导出完成")
    return len(TEMPLATE_EXAMPLE_ROWS)


def export_securecrt_sessions_task(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> dict[str, Any]:
    from netconsole.models.device import Device
    from netconsole.services.securecrt_session_export import export_securecrt_sessions

    output_parent = Path(str(payload.get("output_dir") or path))
    devices = [Device.from_mapping(dict(row)) for row in payload.get("devices") or [] if isinstance(row, Mapping)]
    group_names = {int(key): str(value) for key, value in dict(payload.get("group_names") or {}).items()}
    template_value = str(payload.get("template_ini") or "")
    template_ini = Path(template_value) if template_value else None
    _emit(progress, "write_securecrt_sessions", 0, len(devices), "正在生成 SecureCRT 会话")
    _check_cancel(should_cancel)
    result = export_securecrt_sessions(
        devices,
        str(payload.get("site_name") or ""),
        output_parent,
        group_names=group_names,
        template_ini=template_ini if template_ini and template_ini.is_file() else None,
    )
    marker = Path(path)
    marker.parent.mkdir(parents=True, exist_ok=True)
    marker.write_text(str(result.output_dir), encoding="utf-8")
    _emit(progress, "write_securecrt_sessions", result.generated, len(devices), "SecureCRT 会话生成完成")
    return {"path": str(result.output_dir), "row_count": result.generated, "skipped": result.skipped}


def export_config_snapshots_zip(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    from netconsole.core import app_logger
    from netconsole.core.database import Database
    from netconsole.core.paths import PathResolver
    from netconsole.repositories.config_snapshot_repository import ConfigSnapshotRepository
    from netconsole.services.config_lifecycle_service import ConfigLifecycleService

    database = Database(Path(str(payload.get("db_path") or "")))
    site_name = str(payload.get("site_name") or "")
    paths = PathResolver()
    repository = ConfigSnapshotRepository(database)
    service = ConfigLifecycleService(site_name, database, paths, repository)
    snapshot_entries = list(payload.get("snapshot_entries") or [])
    file_entries = list(payload.get("file_entries") or [])
    failures_text = str(payload.get("failures_text") or "")
    total = len(snapshot_entries) + len(file_entries) + (1 if failures_text else 0)
    written = 0
    path.parent.mkdir(parents=True, exist_ok=True)
    _emit(progress, "zip_config_snapshots", 0, total, "正在打包配置快照")
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as archive:
        for entry in snapshot_entries:
            _check_cancel(should_cancel)
            if not isinstance(entry, Mapping):
                continue
            snapshot_id = entry.get("snapshot_id")
            archive_name = str(entry.get("archive_name") or "")
            if snapshot_id is None or not archive_name:
                continue
            snapshot = repository.get(int(snapshot_id))
            if snapshot.type in {"running", "saved"}:
                archive.writestr(archive_name, service.snapshot_text(snapshot))
            else:
                source = paths.site_dir(site_name) / snapshot.file_path
                if source.exists():
                    archive.write(source, archive_name)
                else:
                    app_logger.log_warning("CONFIG_SNAPSHOT_EXPORT_MISSING", f"snapshot_id={snapshot.id} path={source}")
            written += 1
            if written == total or written % 20 == 0:
                _emit(progress, "zip_config_snapshots", written, total, f"正在打包配置快照 {written}/{total}")
        for entry in file_entries:
            _check_cancel(should_cancel)
            if not isinstance(entry, Mapping):
                continue
            source = Path(str(entry.get("path") or ""))
            archive_name = str(entry.get("archive_name") or source.name)
            if source.exists() and source.is_file():
                archive.write(source, archive_name)
            else:
                app_logger.log_warning("CONFIG_SNAPSHOT_EXPORT_MISSING", f"path={source}")
            written += 1
            if written == total or written % 20 == 0:
                _emit(progress, "zip_config_snapshots", written, total, f"正在打包配置快照 {written}/{total}")
        if failures_text:
            archive.writestr("failed_devices.txt", failures_text.rstrip("\n") + "\n")
            written += 1
    return written


def export_wifi_survey_csv(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    from netconsole.core.database import Database
    from netconsole.repositories.wifi_survey_repository import WifiSurveyRepository

    db_path = Path(str(payload.get("db_path") or ""))
    session_id = int(payload.get("session_id") or 0)
    session_name = str(payload.get("session_name") or "")
    fields = [str(field) for field in payload.get("fields") or []]
    if not fields:
        raise ValueError("无线勘测 CSV 缺少字段列表")
    repository = WifiSurveyRepository(Database(db_path))
    rows = repository.list_observations_by_session(session_id)
    path.parent.mkdir(parents=True, exist_ok=True)
    _emit(progress, "write_wifi_survey_csv", 0, len(rows), "正在导出无线勘测 CSV")
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for index, row in enumerate(rows, start=1):
            _check_cancel(should_cancel)
            writer.writerow({field: (session_name if field == "session_name" else row.get(field)) for field in fields})
            if index == len(rows) or index % 500 == 0:
                _emit(progress, "write_wifi_survey_csv", index, len(rows), f"正在导出无线勘测 CSV {index}/{len(rows)}")
    return len(rows)


def export_wifi_survey_heatmap_png(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    from PySide6.QtGui import QGuiApplication, QPixmap

    from netconsole.core.database import Database
    from netconsole.repositories.wifi_survey_repository import WifiSurveyRepository
    from netconsole.services.wifi_survey.heatmap import build_heatmap_samples, generate_idw_heatmap, render_heatmap_png

    app = QGuiApplication.instance()
    if app is None:
        app = QGuiApplication(["netconsole-export-worker"])
    db_path = Path(str(payload.get("db_path") or ""))
    floor_plan_id = int(payload.get("floor_plan_id") or 0)
    session_id = int(payload.get("session_id") or 0)
    mode = str(payload.get("mode") or "strongest")
    selected_ssids = {str(value) for value in payload.get("selected_ssids") or []}
    selected_bssids = {str(value) for value in payload.get("selected_bssids") or []}
    repository = WifiSurveyRepository(Database(db_path))
    floor_plan = repository.get_floor_plan(floor_plan_id)
    base = QPixmap(str(floor_plan.get("image_path") or ""))
    if base.isNull():
        raise ValueError("无线勘测图纸文件无法读取")
    _emit(progress, "wifi_survey_heatmap_prepare", 0, 1, "正在生成无线热力图")
    _check_cancel(should_cancel)
    points = repository.list_points(session_id)
    observations = repository.list_observations_by_session(session_id)
    samples = build_heatmap_samples(points, observations, mode, selected_ssids, selected_bssids)
    overlay = None
    if len(samples) >= 3:
        overlay = generate_idw_heatmap(
            base.width(),
            base.height(),
            [(sample.x_px, sample.y_px, sample.rssi_dbm) for sample in samples],
        )
    rendered = render_heatmap_png(base, overlay)
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rendered.save(str(path), "PNG"):
        raise OSError(f"保存无线热力图失败：{path}")
    _emit(progress, "wifi_survey_heatmap_save", 1, 1, "无线热力图导出完成")
    return len(samples)


def export_snmp_query_result(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    result = _snmp_result_payload(payload)
    request = dict(result.get("request") or {})
    rows = [
        {
            "时间": request.get("started_at") or "",
            "设备": request.get("device_name") or "",
            "OID": row.get("oid") or "",
            "名称": row.get("name") or "",
            "实例": row.get("instance") or "",
            "类型": row.get("value_type") or "",
            "原始值": "" if row.get("value") is None else str(row.get("value")),
            "解码值": row.get("decoded_value") or "",
            "延迟": row.get("latency_ms") or 0,
            "状态": row.get("status") or "",
            "错误信息": row.get("error_message") or "",
        }
        for row in result.get("rows") or []
        if isinstance(row, Mapping)
    ]
    headers = list(rows[0].keys()) if rows else ["时间", "设备", "OID", "名称", "实例", "类型", "原始值", "解码值", "延迟", "状态", "错误信息"]
    path.parent.mkdir(parents=True, exist_ok=True)
    _emit(progress, "write_snmp_query_result", 0, len(rows), "正在导出 SNMP 查询结果")
    _check_cancel(should_cancel)
    suffix = _effective_suffix(path)
    if suffix == ".json":
        path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    elif suffix == ".xlsx":
        export_table_xlsx(
            path,
            {
                "sheet_name": "SNMP查询结果",
                "columns": [{"key": header, "title": header, "text": True} for header in headers],
                "source": {"type": "inline_rows", "rows": rows},
                "freeze_header": True,
                "auto_filter": True,
            },
            progress,
            should_cancel,
        )
    else:
        export_table_csv(
            path,
            {
                "columns": [{"key": header, "title": header, "text": True} for header in headers],
                "source": {"type": "inline_rows", "rows": rows},
            },
            progress,
            should_cancel,
        )
    _emit(progress, "write_snmp_query_result", len(rows), len(rows), "SNMP 查询结果导出完成")
    return len(rows)


def export_mib_product_compare(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    from netconsole.repositories.global_mib_repository import GlobalMibRepository
    from netconsole.services.mib_product_reference_compare_service import COMPARE_HEADERS, MibProductReferenceCompareService, _row_values

    db_path = Path(str(payload.get("db_path") or ""))
    left_reference_id = int(payload.get("left_reference_id") or 0)
    right_reference_id = int(payload.get("right_reference_id") or 0)
    _emit(progress, "write_mib_product_compare", 0, 0, "正在导出产品参考对比结果")
    _check_cancel(should_cancel)
    service = MibProductReferenceCompareService(GlobalMibRepository(db_path))
    rows = service.list_results(left_reference_id, right_reference_id, limit=200000)
    path.parent.mkdir(parents=True, exist_ok=True)
    if _effective_suffix(path) == ".xlsx":
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "产品参考对比"
        sheet.append(COMPARE_HEADERS)
        for row in rows:
            sheet.append(_row_values(row))
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in sheet.columns:
            letter = column[0].column_letter
            sheet.column_dimensions[letter].width = min(60, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
        workbook.save(path)
    else:
        with path.open("w", encoding="utf-8-sig", newline="") as file:
            writer = csv.writer(file)
            writer.writerow(COMPARE_HEADERS)
            for row in rows:
                writer.writerow(_row_values(row))
    _emit(progress, "write_mib_product_compare", len(rows), len(rows), "产品参考对比结果导出完成")
    return len(rows)


def _fit_ap_import_export_service(payload: Mapping[str, Any]):
    from netconsole.core.database import Database
    from netconsole.repositories.ac_repository import AcRepository
    from netconsole.services.fit_ap_import_export import FitApImportExportService

    return FitApImportExportService(AcRepository(Database(Path(str(payload.get("db_path") or "")))))


def export_fit_ap_csv_task(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    rows = [dict(row) for row in payload.get("rows") or [] if isinstance(row, Mapping)]
    if not rows:
        ac_uuid = str(payload.get("ac_uuid") or "").strip()
        if ac_uuid:
            rows = _ac_repository(payload).list_fit_ap_resources_with_metadata(ac_uuid)
    path.parent.mkdir(parents=True, exist_ok=True)
    _emit(progress, "write_fit_ap_csv", 0, len(rows), "正在导出 FIT-AP CSV")
    _check_cancel(should_cancel)
    _fit_ap_import_export_service(payload).export_ap_csv(path, rows)
    _emit(progress, "write_fit_ap_csv", len(rows), len(rows), "FIT-AP CSV 导出完成")
    return len(rows)


def export_fit_ap_extension_xlsx_task(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    rows = [dict(row) for row in payload.get("rows") or [] if isinstance(row, Mapping)]
    if not rows:
        filters = dict(payload.get("filters") or {})
        rows = _ac_repository(payload).list_ap_extension_points(
            search=str(payload.get("search") or filters.get("search") or ""),
            station_name=str(filters.get("station_name") or ""),
            line_side=str(filters.get("line_side") or ""),
            direction=str(filters.get("direction") or ""),
            match_status=str(filters.get("match_status") or ""),
        )
    path.parent.mkdir(parents=True, exist_ok=True)
    _emit(progress, "write_fit_ap_extension", 0, len(rows), "正在导出 FIT-AP 扩展信息")
    _check_cancel(should_cancel)
    _fit_ap_import_export_service(payload).export_standard_ap_extension_xlsx(path, rows)
    _emit(progress, "write_fit_ap_extension", len(rows), len(rows), "FIT-AP 扩展信息导出完成")
    return len(rows)


def export_fit_ap_extension_template_xlsx_task(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    rows = [dict(row) for row in payload.get("rows") or [] if isinstance(row, Mapping)]
    ap_entities = [dict(row) for row in payload.get("ap_entities") or [] if isinstance(row, Mapping)]
    ac_uuid = str(payload.get("ac_uuid") or "").strip()
    if ac_uuid and not rows:
        rows = _ac_repository(payload).list_fit_ap_resources_with_metadata(ac_uuid)
    if ac_uuid and not ap_entities:
        ap_entities = _ac_repository(payload).list_ap_entities(ac_uuid)
    path.parent.mkdir(parents=True, exist_ok=True)
    _emit(progress, "write_fit_ap_extension_template", 0, len(rows), "正在导出 AP 扩展模板")
    _check_cancel(should_cancel)
    _fit_ap_import_export_service(payload).export_ap_extension_template_xlsx(path, rows, ap_entities)
    _emit(progress, "write_fit_ap_extension_template", len(rows), len(rows), "AP 扩展模板导出完成")
    return len(rows)


def export_ap_online_overview_xlsx_task(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    from netconsole.services.ap_online_overview import export_ap_online_overview_xlsx

    rows = [dict(row) for row in payload.get("rows") or [] if isinstance(row, Mapping)]
    path.parent.mkdir(parents=True, exist_ok=True)
    _emit(progress, "write_ap_online_overview", 0, len(rows), "正在导出 AP 在线概览")
    _check_cancel(should_cancel)
    export_ap_online_overview_xlsx(
        path,
        rows,
        [str(value) for value in payload.get("headers") or []],
        dict(payload.get("offline_ap_stats") or {}),
        [dict(row) for row in payload.get("offline_ap_ledger_rows") or [] if isinstance(row, Mapping)],
        [str(value) for value in payload.get("offline_ap_stats_headers") or []],
        [str(value) for value in payload.get("offline_ap_ledger_headers") or []],
    )
    _emit(progress, "write_ap_online_overview", len(rows), len(rows), "AP 在线概览导出完成")
    return len(rows)


def export_omnipeek_name_table_task(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> dict[str, Any]:
    from netconsole.models.omnipeek_name_table import OmniPeekDeviceItem, OmniPeekExportConfig
    from netconsole.services.omnipeek_name_table_service import export_items_to_omnipeek_nam

    items = [OmniPeekDeviceItem(**dict(row)) for row in payload.get("items") or [] if isinstance(row, Mapping)]
    config_data = dict(payload.get("config") or {})
    config_data["output_path"] = path
    config = OmniPeekExportConfig(**config_data)
    _emit(progress, "write_omnipeek_name_table", 0, len(items), "正在导出 OmniPeek 名称表")
    _check_cancel(should_cancel)
    result = export_items_to_omnipeek_nam(items, config, source_counts=dict(payload.get("source_counts") or {}))
    row_count = result.total_entries
    _emit(progress, "write_omnipeek_name_table", row_count, row_count, "OmniPeek 名称表导出完成")
    return {
        "path": str(result.output_path),
        "row_count": row_count,
        "log_path": str(result.log_path),
        "skipped_count": result.skipped_count,
        "error_count": result.error_count,
    }


def export_online_mr_report_xlsx(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    from netconsole.services.vehicle_mr_offline_excel_report import VehicleMrOfflineExcelReportExporter

    session_dir = Path(str(payload.get("session_dir") or ""))
    _emit(progress, "online_mr_report_prepare", 1, 3, "正在读取 parsed 数据")
    _check_cancel(should_cancel)
    output_path = VehicleMrOfflineExcelReportExporter().export(session_dir, path)
    _emit(progress, "online_mr_report_save", 2, 3, "正在保存 Excel 报告")
    _check_cancel(should_cancel)
    _emit(progress, "online_mr_report_done", 3, 3, "离线分析报告导出完成")
    return 1 if output_path.exists() else 0


def export_car_network_point_table(path: Path, payload: Mapping[str, Any], progress: ProgressCallback | None = None, should_cancel: CancelCallback | None = None) -> int:
    from netconsole.core.paths import PathResolver
    from netconsole.services.rail_transit.car_network_diagnostic import CarNetworkPointTableStore

    site_name = str(payload.get("site_name") or "")
    store = CarNetworkPointTableStore(PathResolver(), site_name)
    nodes = store.load()
    _emit(progress, "write_car_network_point_table", 0, len(nodes), "正在导出车内通信点表")
    _check_cancel(should_cancel)
    path.parent.mkdir(parents=True, exist_ok=True)
    suffix = _effective_suffix(path)
    target = path if path.suffix.casefold() == suffix else path.with_name(f"{path.stem}.netconsole{suffix or '.xlsx'}")
    try:
        store.export_file(target, nodes)
        if target != path:
            os.replace(target, path)
    finally:
        if target != path and target.exists():
            try:
                target.unlink()
            except OSError:
                pass
    _emit(progress, "write_car_network_point_table", len(nodes), len(nodes), "车内通信点表导出完成")
    return len(nodes)


def normalize_columns(columns: Iterable[Any]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for column in columns:
        if isinstance(column, Mapping):
            key = str(column.get("key") or column.get("field") or "")
            title = str(column.get("title") or column.get("header") or key)
            result.append({"key": key, "title": title, "width": column.get("width"), "text": bool(column.get("text", True))})
            continue
        if isinstance(column, Sequence) and not isinstance(column, (str, bytes)) and len(column) >= 2:
            title, key = column[0], column[1]
            result.append({"key": str(key), "title": str(title), "width": None, "text": True})
            continue
        key = str(column)
        result.append({"key": key, "title": key, "width": None, "text": True})
    return result


def resolve_rows(payload: Mapping[str, Any]) -> list[Any]:
    source = payload.get("source")
    if not source:
        return list(payload.get("rows") or [])
    if not isinstance(source, Mapping):
        raise ValueError("导出 source 必须是 dict")
    source_type = str(source.get("type") or "").strip()
    if source_type == "inline_rows":
        return list(source.get("rows") or [])
    if source_type == "repository_query":
        return _resolve_repository_rows(source)
    raise ValueError(f"不支持的导出数据源：{source_type or '<empty>'}")


def _resolve_repository_rows(source: Mapping[str, Any]) -> list[dict[str, Any]]:
    repository = str(source.get("repository") or "").strip()
    method = str(source.get("method") or "").strip()
    filters = dict(source.get("filters") or {})
    db_path = Path(str(source.get("db_path") or ""))
    if repository == "device_repository" and method == "list":
        from netconsole.core.database import Database
        from netconsole.repositories.device_repository import DeviceRepository

        devices = DeviceRepository(Database(db_path)).list(
            search=filters.get("search") or None,
            vendor=filters.get("vendor") or None,
            device_type=filters.get("device_type") or None,
            group_filter=filters.get("group_filter"),
        )
        return [device.to_record() for device in devices]
    if repository == "wireless_scan_repository" and method == "list_results":
        from netconsole.repositories.wireless_scan_repository import WirelessScanRepository
        from netconsole.services.network_tools.wireless_scan_service import repository_row_to_display_row

        scan_id = str(filters.get("scan_id") or "").strip()
        if not scan_id:
            raise ValueError("wireless_scan_repository.list_results 缺少 scan_id")
        return [repository_row_to_display_row(row) for row in WirelessScanRepository(db_path).list_results(scan_id)]
    if repository == "ac_repository" and method == "list_trackside_ap_plan":
        from netconsole.core.database import Database
        from netconsole.repositories.ac_repository import AcRepository, TRACKSIDE_AP_PLAN_MODE

        mode = str(filters.get("mode") or TRACKSIDE_AP_PLAN_MODE)
        return AcRepository(Database(db_path)).list_trackside_ap_plan(mode)
    raise ValueError(f"不支持的 repository_query：{repository}.{method}")


def _ac_repository(payload: Mapping[str, Any]):
    from netconsole.core.database import Database
    from netconsole.repositories.ac_repository import AcRepository

    return AcRepository(Database(Path(str(payload.get("db_path") or ""))))


def _snmp_result_payload(payload: Mapping[str, Any]) -> dict[str, Any]:
    result_file = str(payload.get("result_file") or "").strip()
    if result_file:
        with Path(result_file).open("r", encoding="utf-8") as handle:
            data = json.load(handle)
        return dict(data or {})
    return dict(payload.get("result") or {})


def _value_for_row(row: Any, key: str) -> object:
    if isinstance(row, Mapping):
        value = row.get(key)
    elif isinstance(row, Sequence) and not isinstance(row, (str, bytes)):
        try:
            value = row[int(key)]
        except (ValueError, IndexError):
            value = ""
    else:
        value = ""
    return "" if value is None else value


def _apply_row_fill(worksheet, row: Any, row_fill_field: str, row_fill_colors: Mapping[str, str]) -> None:
    if not row_fill_field or not isinstance(row, Mapping):
        return
    raw = str(row.get(row_fill_field) or "").strip()
    color = row_fill_colors.get(raw, raw)
    if not color:
        return
    normalized = color.lstrip("#").upper()
    if len(normalized) not in {6, 8}:
        return
    from openpyxl.styles import PatternFill

    fill = PatternFill(fill_type="solid", fgColor=normalized[-6:])
    for cell in worksheet[worksheet.max_row]:
        cell.fill = fill


def _zip_entries(payload: Mapping[str, Any]) -> list[tuple[Path, str]]:
    source = payload.get("source") if isinstance(payload.get("source"), Mapping) else {}
    if str(source.get("type") or "") == "file_or_directory":
        items = source.get("paths") or []
        base_dir_value = source.get("base_dir") or payload.get("base_dir")
        exclude_names = {str(name) for name in source.get("exclude_names") or []}
    else:
        items = payload.get("items") or []
        base_dir_value = payload.get("base_dir")
        exclude_names = {str(name) for name in payload.get("exclude_names") or []}
    base_dir = Path(str(base_dir_value)).resolve() if base_dir_value else None
    entries: list[tuple[Path, str]] = []
    for item in items:
        source = Path(str(item)).resolve()
        if not source.exists():
            continue
        if source.is_dir():
            for child in source.rglob("*"):
                if child.is_file() and child.name not in exclude_names:
                    entries.append((child, _archive_name(child, base_dir or source.parent)))
        elif source.name not in exclude_names:
            entries.append((source, _archive_name(source, base_dir or source.parent)))
    return entries


def _archive_name(path: Path, base_dir: Path) -> str:
    try:
        return path.relative_to(base_dir).as_posix()
    except ValueError:
        return path.name


def _safe_sheet_title(title: str) -> str:
    value = "".join("_" if char in "[]:*?/\\\"" else char for char in title).strip() or "Sheet1"
    return value[:31]


def _effective_suffix(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".tmp" and path.stem:
        return Path(path.stem).suffix.lower()
    return suffix


def _parse_log_line(line: str) -> dict[str, str] | None:
    parts = line.split(" | ", 3)
    if len(parts) != 4:
        return None
    time, level, event, detail = parts
    return {"time": time, "level": level, "event": event, "detail": detail}


def _emit(progress: ProgressCallback | None, stage: str, current: int, total: int, message: str) -> None:
    if progress:
        progress(stage, current, total, message)


def _check_cancel(should_cancel: CancelCallback | None) -> None:
    if should_cancel and should_cancel():
        raise ExportCancelled("导出已取消")


def replace_output(tmp_path: Path, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    os.replace(tmp_path, output_path)
