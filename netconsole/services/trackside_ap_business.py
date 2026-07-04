from __future__ import annotations

from pathlib import Path
import re
from time import perf_counter

from netconsole.core import app_logger
from netconsole.core.optical_severity_engine import compute_optical_severity, display_optical_status, worse_optical_severity
from netconsole.core.sources.switch_source import build_switch_data_lookup
from netconsole.models.device import Device
from netconsole.repositories.device_group_repository import DeviceGroupRepository
from netconsole.services.ap_online_overview import AP_ONLINE_OVERVIEW_COLUMNS, write_ap_online_overview_sheet
from netconsole.services.offline_ap_ledger import (
    OFFLINE_AP_LEDGER_COLUMNS,
    OFFLINE_AP_STATS_COLUMNS,
    OFFLINE_AP_STATUS_TEXT,
    write_offline_ap_ledger_sheet,
    write_offline_ap_stats_sheet,
)
from netconsole.utils.interface_normalize import normalize_interface_name
from netconsole.utils.interface_sort import interface_sort_key
from netconsole.utils.station_normalize import normalize_station_value


TRACKSIDE_AP_BUSINESS_INTERNAL_FIELDS = {
    "host",
    "host_address",
    "management_ip",
    "ap_ip",
    "source_device",
    "collection_status",
    "offline_reason",
    "status_reason",
    "data_source",
    "switch_collection_status",
    "ap_rx_low_alarm",
    "ap_rx_low_warning",
}

TRACKSIDE_AP_BUSINESS_VISIBLE_COLUMNS = (
    ("ac.station", "site"),
    ("ac.indoor_switch", "device_name"),
    ("details.interface_name", "interface_name"),
    ("details.link", "link_status"),
    ("details.port_type", "port_type"),
    ("details.port_description", "description"),
    ("details.pvid", "pvid"),
    ("details.vlan", "vlan"),
    ("ac.indoor_switch_rx_power", "switch_rx_power"),
    ("trackside.switch_optical_status", "switch_optical_status"),
    ("ac.ap_mac", "ap_mac"),
    ("ac.ap_name", "ap_name"),
    ("ac.ap_side_rx_power", "ap_rx_power"),
    ("trackside.ap_optical_status", "ap_optical_status"),
    ("trackside_ap.last_collected_at", "updated_at"),
)

TRACKSIDE_AP_BUSINESS_COLUMNS = TRACKSIDE_AP_BUSINESS_VISIBLE_COLUMNS

TRACKSIDE_AP_BUSINESS_EXPORT_COLUMNS = (
    ("ac.station", "site"),
    ("ac.indoor_switch", "device_name"),
    ("details.interface_name", "interface_name"),
    ("details.link", "link_status"),
    ("ac.indoor_switch_rx_power", "switch_rx_power"),
    ("trackside.switch_optical_status", "switch_optical_status"),
    ("ac.ap_mac", "ap_mac"),
    ("ac.ap_name", "ap_name"),
    ("ac.ap_side_rx_power", "ap_rx_power"),
    ("trackside.ap_optical_status", "ap_optical_status"),
    ("trackside_ap.last_collected_at", "updated_at"),
    ("trackside.export.switch_optical_change", "switch_optical_change"),
    ("trackside.export.ap_optical_change", "ap_optical_change"),
    ("trackside.export.ap_port_change", "ap_port_change"),
    ("trackside.export.previous_switch", "previous_switch"),
    ("trackside.export.previous_interface", "previous_interface"),
    ("trackside.export.current_switch", "current_switch"),
    ("trackside.export.current_interface", "current_interface"),
    ("trackside.export.history_compared_at", "history_compared_at"),
)

NEW_ONLINE_AP_OVERVIEW_COLUMNS = (
    ("trackside.export.station", "site"),
    ("trackside.export.switch", "device_name"),
    ("details.interface_name", "interface_name"),
    ("details.link", "link_status"),
    ("details.port_type", "port_type"),
    ("details.port_description", "description"),
    ("details.pvid", "pvid"),
    ("details.vlan", "vlan"),
    ("ac.indoor_switch_rx_power", "switch_rx_power"),
    ("trackside.switch_optical_status", "switch_optical_status"),
    ("ac.ap_mac", "ap_mac"),
    ("ac.ap_name", "ap_name"),
    ("ac.ap_side_rx_power", "ap_rx_power"),
    ("trackside.ap_optical_status", "ap_optical_status"),
    ("trackside_ap.last_collected_at", "updated_at"),
    ("ac.register_status", "register_status"),
    ("ac.new_online_status", "new_online_status"),
    ("trackside.export.identity_source", "identity_source"),
    ("ac.current_unauthenticated", "current_unauthenticated"),
    ("ac.current_resource_exists", "current_resource_exists"),
    ("ac.last_unauthenticated_at", "last_unauthenticated_at"),
    ("trackside.export.first_seen_at", "first_seen_at"),
    ("trackside.export.ac_device", "ac_device_name"),
    ("APID", "apid"),
    ("ap.ip_address", "ap_ip"),
    ("ap.model", "model"),
    ("ap.serial_number", "serial_number"),
    ("ac.group_name", "group_name"),
    ("ac.state", "state_display"),
    ("trackside.export.suggestion", "suggestion"),
)

AP_OPTICAL_TREATMENT_RECORD_COLUMNS = (
    ("ac.station", "site"),
    ("ac.ap_name", "ap_name"),
    ("ac.ap_mac", "ap_mac"),
    ("ap.serial_number", "serial_number"),
    ("trackside.export.side", "side"),
    ("ac.indoor_switch", "device_name"),
    ("details.interface_name", "interface_name"),
    ("trackside.export.issue_type", "issue_type"),
    ("trackside.export.first_found_at", "first_found_at"),
    ("trackside.export.first_rx_power", "first_rx_power"),
    ("trackside.export.fixed_rx_power", "fixed_rx_power"),
    ("trackside.export.current_rx_power", "current_rx_power"),
    ("trackside.export.current_status", "current_status"),
    ("trackside.export.treatment_status", "treatment_status"),
    ("trackside.export.remark", "remark"),
    ("trackside.export.completed_at", "completed_at"),
)

OPTICAL_TREATMENT_ISSUE_STATUSES = {"notice", "warning", "alarm", "link_abnormal", "link_down", "no_light"}
OPTICAL_TREATMENT_IGNORED_STATUSES = {"normal", "unknown", "not_collected", "skipped", "failed", "timeout", "offline", "no_module", ""}
AP_SIDE_LABEL = "AP\u4fa7"
SWITCH_SIDE_LABEL = "\u4ea4\u6362\u673a\u4fa7"
TREATMENT_OPEN_LABEL = "\u672a\u5904\u7406"
TREATMENT_CLOSED_LABEL = "\u5df2\u5904\u7406"
ISSUE_TYPE_NOTICE_LABEL = "\u5149\u8870\u9884\u8b66"
ISSUE_TYPE_ALARM_LABEL = "\u5149\u8870\u544a\u8b66"
ISSUE_TYPE_LINK_ABNORMAL_LABEL = "\u94fe\u8def\u5f02\u5e38"
ISSUE_TYPE_NO_LIGHT_LABEL = "\u65e0\u5149"
ISSUE_TYPE_OPTICAL_ABNORMAL_LABEL = "\u5149\u8870\u5f02\u5e38"

TRACKSIDE_AP_BUSINESS_HEADER_TOOLTIPS = {
    "site": "trackside.tooltip.station",
    "link_status": "trackside.tooltip.link",
    "port_type": "trackside.tooltip.port_type",
    "switch_rx_power": "trackside.tooltip.switch_rx_power",
    "ap_optical_status": "trackside.tooltip.ap_optical_status",
}

TRACKSIDE_AP_DEVICE_COLUMNS = (
    ("details.interface_name", "interface_name"),
    ("details.link", "link_status"),
    ("details.port_type", "port_type"),
    ("details.port_description", "description"),
    ("details.pvid", "pvid"),
    ("details.vlan", "vlan"),
    ("ac.indoor_switch_rx_power", "switch_rx_power"),
    ("trackside.switch_optical_status", "switch_optical_status"),
    ("ac.ap_mac", "ap_mac"),
    ("ac.ap_name", "ap_name"),
    ("ac.ap_side_rx_power", "ap_rx_power"),
    ("trackside.ap_optical_status", "ap_optical_status"),
    ("field.updated_at", "updated_at"),
)

TRACKSIDE_OPTICAL_COLOR_RGB = {
    "normal": "DCFCE7",
    "notice": "FEF9C3",
    "warning": "FEF9C3",
    "alarm": "FEE2E2",
    "link_abnormal": "FFE4E6",
    "link_down": "FFE4E6",
    "no_light": "E5E7EB",
    "no_module": "F3F4F6",
    "skipped": "F3F4F6",
    "offline": "E5E7EB",
}
TRACKSIDE_EXPORT_HEADER_FILL = "DBEAFE"
TRACKSIDE_EXPORT_NORMAL_FILL = TRACKSIDE_OPTICAL_COLOR_RGB["normal"]
TRACKSIDE_EXPORT_WARNING_FILL = TRACKSIDE_OPTICAL_COLOR_RGB["warning"]
TRACKSIDE_EXPORT_ALARM_FILL = TRACKSIDE_OPTICAL_COLOR_RGB["alarm"]
CURRENT_OPTICAL_ABNORMAL_SHEET_TITLE = "当前异常光衰"
CURRENT_OPTICAL_ABNORMAL_EMPTY_TEXT = "当前无异常光衰（已排除无光端口）"

AP_SIDE_DISPLAY_FIELDS = {"ap_rx_power", "ap_tx_power"}
AP_SIDE_MISSING_DISPLAY = "-"
MATCH_SOURCE_LABELS = {
    "description": "\u63cf\u8ff0\u5339\u914d",
    "pvid": "PVID\u5339\u914d",
    "description+pvid": "\u63cf\u8ff0+PVID",
    "none": "-",
}


def description_contains_ap(description: object) -> bool:
    return "ap" in str(description or "").casefold()


def parse_vlan_set(value: object) -> set[int]:
    if value in (None, ""):
        return set()
    tokens = [item for item in re.split(r"[,，;；\s]+", str(value).strip()) if item]
    vlans: set[int] = set()
    for token in tokens:
        if "-" in token:
            parts = token.split("-", 1)
            if len(parts) != 2 or not parts[0].isdigit() or not parts[1].isdigit():
                raise ValueError(f"invalid VLAN range: {token}")
            start = int(parts[0])
            end = int(parts[1])
            if start > end:
                raise ValueError(f"invalid VLAN range: {token}")
            for vlan in range(start, end + 1):
                _validate_vlan(vlan)
                vlans.add(vlan)
            continue
        if not token.isdigit():
            raise ValueError(f"invalid VLAN: {token}")
        vlan = int(token)
        _validate_vlan(vlan)
        vlans.add(vlan)
    return vlans


def normalize_vlan_text(value: object) -> str:
    return ",".join(str(vlan) for vlan in sorted(parse_vlan_set(value)))


def pvid_matches_trackside_plan(device_station: object, pvid: object, active_plan: dict | None) -> bool:
    if not active_plan:
        return False
    try:
        vlan = int(str(pvid or "").strip())
    except ValueError:
        return False
    if vlan <= 0:
        return False
    station_name = str(device_station or "").strip()
    station_vlans = active_plan.get("station_vlans") if isinstance(active_plan, dict) else {}
    if isinstance(station_vlans, dict) and station_name and station_name in station_vlans:
        return vlan in set(station_vlans.get(station_name) or set())
    if isinstance(station_vlans, dict) and station_name and station_vlans and station_name not in station_vlans:
        try:
            from netconsole.core import app_logger

            app_logger.log_warning("TRACKSIDE_AP_PLAN_STATION_FALLBACK", f"station={station_name}, pvid={vlan}")
        except Exception:
            pass
    return vlan in set(active_plan.get("all_vlans") or set()) if isinstance(active_plan, dict) else False


def is_switch_device_type(device_type: object) -> bool:
    return str(device_type or "").strip().casefold() in {"sw", "switch", "交换机"}


def filter_station_switch_devices(devices: list[Device], database, site_name: str) -> list[Device]:
    groups = {group.id: group.name for group in DeviceGroupRepository(database, site_name).list()}
    return [
        device
        for device in devices
        if groups.get(device.group_id or -1, "") == "车站" and is_switch_device_type(device.device_type)
    ]


def is_trackside_layer2_interface(interface: dict[str, object | None]) -> bool:
    name = str(interface.get("interface_name") or "").strip()
    name_key = name.casefold()
    if not name_key:
        return False
    if name_key.startswith(("vlan-interface", "loopback", "inloopback", "null")):
        return False
    if str(interface.get("interface_type") or "").strip().casefold() in {"三层", "l3", "layer3", "route", "routed"}:
        return False
    if str(interface.get("port_status") or "").strip().casefold() == "route":
        return False
    if str(interface.get("ip_address") or "").strip():
        return False
    layer2_markers = (
        interface.get("pvid"),
        interface.get("vlan"),
        interface.get("port_status"),
        interface.get("link_type"),
        interface.get("port_link_type"),
    )
    if any(str(value or "").strip() for value in layer2_markers):
        return True
    return bool(re.match(r"^(?:GigabitEthernet|Ten-GigabitEthernet|XGE|Bridge-Aggregation)", name, re.IGNORECASE))


def is_trackside_ap_interface(device: Device, interface: dict[str, object | None], active_plan: dict | None = None) -> tuple[bool, str]:
    if not is_trackside_layer2_interface(interface):
        return False, "none"
    by_description = description_contains_ap(interface.get("description"))
    by_pvid = pvid_matches_trackside_plan(getattr(device, "station", ""), interface.get("pvid"), active_plan)
    if by_description and by_pvid:
        return True, "description+pvid"
    if by_description:
        return True, "description"
    if by_pvid:
        return True, "pvid"
    return False, "none"


def _validate_vlan(vlan: int) -> None:
    if vlan < 1 or vlan > 4094:
        raise ValueError(f"VLAN out of range: {vlan}")


def build_device_optical_status_lookup(
    devices: list[Device],
    optical_by_device: dict[str, list[dict[str, object | None]]],
) -> dict:
    """Backward-compatible alias — delegates to ``switch_source.build_switch_data_lookup``."""
    return build_switch_data_lookup(devices, optical_by_device)


def build_trackside_ap_business_rows(
    devices: list[Device],
    interfaces_by_device: dict[str, list[dict[str, object | None]]],
    optical_by_device: dict[str, list[dict[str, object | None]]],
    fit_ap_optical_rows: list[dict[str, object | None]],
    lldp_by_device: dict[str, list[dict[str, object | None]]] | None = None,
    fit_ap_resource_rows: list[dict[str, object | None]] | None = None,
    device_optical_status_lookup: dict[tuple[str, str], str] | None = None,
    trackside_ap_plan: dict | None = None,
    offline_ap_ledger_rows: list[dict[str, object | None]] | None = None,
    historical_lldp_rows: list[dict[str, object | None]] | None = None,
) -> list[dict[str, object | None]]:
    optical_indexes = {device_uuid: _latest_rows_by_normalized_interface(rows, "interface_name") for device_uuid, rows in optical_by_device.items()}
    lldp_indexes = {device_uuid: _latest_rows_by_normalized_interface(rows, "local_interface") for device_uuid, rows in (lldp_by_device or {}).items()}
    fit_ap_optical_rows = merge_fit_ap_rows_by_identity(fit_ap_optical_rows)
    fit_ap_resource_rows = merge_fit_ap_rows_by_identity(fit_ap_resource_rows or [])
    fit_ap_index: dict[tuple[str, str], dict[str, object | None]] = {}
    fit_ap_optical_by_mac: dict[str, dict[str, object | None]] = {}
    fit_ap_optical_by_identity: dict[tuple[str, str], dict[str, object | None]] = {}
    fit_ap_optical_by_name_mac: dict[str, dict[str, object | None]] = {}
    fit_ap_resource_by_mac: dict[str, dict[str, object | None]] = {}
    fit_ap_resource_by_identity: dict[tuple[str, str], dict[str, object | None]] = {}
    for row in fit_ap_optical_rows:
        key = (_normalize_name(row.get("neighbor_device_name")), normalize_interface_name(row.get("neighbor_interface")).casefold())
        if key[0] and key[1]:
            fit_ap_index[key] = row
        mac = normalize_mac(row.get("ap_mac"))
        if mac:
            fit_ap_optical_by_mac[mac] = row
        identity = ap_identity_key(row)
        if identity:
            fit_ap_optical_by_identity[identity] = row
        name_as_mac = normalize_mac(row.get("ap_name"))
        if name_as_mac:
            fit_ap_optical_by_name_mac[name_as_mac] = row
    for row in fit_ap_resource_rows or []:
        mac = normalize_mac(row.get("ap_mac"))
        if mac:
            fit_ap_resource_by_mac[mac] = row
        identity = ap_identity_key(row)
        if identity:
            fit_ap_resource_by_identity[identity] = row
    historical_lldp_index = _build_historical_lldp_index(historical_lldp_rows or [])

    result: list[dict[str, object | None]] = []
    for device in devices:
        device_uuid = str(device.device_uuid or "")
        device_names = {_normalize_name(device.name), _normalize_name(device.system_name)}
        device_names.discard("")
        optical_index = optical_indexes.get(device_uuid, {})
        lldp_index = lldp_indexes.get(device_uuid, {})
        for interface in interfaces_by_device.get(device_uuid, []):
            matched, match_source = is_trackside_ap_interface(device, interface, trackside_ap_plan)
            if not matched:
                continue
            interface_name = str(interface.get("interface_name") or "")
            normalized_interface = normalize_interface_name(interface_name).casefold()
            link_state = normalize_link_state(interface.get("link_status") or interface.get("link"))
            optical = optical_index.get(normalized_interface, {})
            lldp = lldp_index.get(normalized_interface, {})
            historical_lldp = _find_historical_lldp_row(historical_lldp_index, device_names, interface_name)
            neighbor_mac = normalize_mac(lldp.get("neighbor_mac"))
            historical_lldp_used = False
            if not neighbor_mac and historical_lldp:
                neighbor_mac = normalize_mac(historical_lldp.get("ap_mac") or historical_lldp.get("neighbor_mac"))
                historical_lldp_used = bool(neighbor_mac or historical_lldp.get("ap_name") or historical_lldp.get("ap_uuid"))
            resource_from_neighbor = fit_ap_resource_by_mac.get(neighbor_mac)
            identity_from_neighbor = ap_identity_key(resource_from_neighbor or historical_lldp or {})
            fit_ap_from_identity = fit_ap_optical_by_identity.get(identity_from_neighbor) if identity_from_neighbor else None
            resource_from_identity = fit_ap_resource_by_identity.get(identity_from_neighbor) if identity_from_neighbor else None
            fit_ap = (
                fit_ap_optical_by_mac.get(neighbor_mac)
                or fit_ap_from_identity
            ) or (
                _merge_resource_with_optical(resource_from_neighbor, fit_ap_optical_by_mac)
                or resource_from_identity
                or fit_ap_optical_by_name_mac.get(neighbor_mac)
                or _find_fit_ap_row(fit_ap_index, device_names, interface_name)
            )
            switch_collection_status = _switch_collection_status(device, interface, optical)
            switch_result = compute_optical_severity(
                {
                    "module_present": bool(_has_optical_module_data(optical)),
                    "no_module": _explicit_no_module(optical),
                    "switch_rx_power": optical.get("rx_power"),
                    "switch_port_status": optical.get("port_status"),
                    "alarm_low": optical.get("rx_low_alarm"),
                    "alarm_high": optical.get("rx_high_alarm"),
                    "warning_low": optical.get("rx_low_warning"),
                    "device_type": "switch",
                }
            )
            switch_status = switch_result.severity
            switch_offline = _is_switch_collection_offline(switch_collection_status)
            if _should_mark_switch_link_abnormal(link_state, switch_result, optical, switch_collection_status):
                switch_status = "link_abnormal"
            ap_candidate = {
                "ap_mac": normalize_mac(fit_ap.get("ap_mac")) or neighbor_mac,
                "ap_name": fit_ap.get("ap_name") or (historical_lldp or {}).get("ap_name"),
                "ap_rx_power": fit_ap.get("rx_power"),
                "ap_tx_power": fit_ap.get("tx_power"),
            }
            ac_idle = _is_ac_idle(fit_ap)
            ap_identity_known = any(ap_candidate.get(field) for field in ("ap_mac", "ap_name"))
            ap_side_has_data = _has_ap_side_optical_data(fit_ap, ap_candidate) or ac_idle or (switch_offline and ap_identity_known)
            ap_status = ""
            offline_reason = ""
            status_reason = ""
            data_source = "current"
            if historical_lldp_used and not lldp:
                status_reason = "沿用历史LLDP映射，本轮未采到当前LLDP"
                data_source = "historical_lldp"
            if switch_offline:
                switch_status = "offline"
                ap_status = "offline"
                offline_reason = "switch_offline"
                status_reason = "室内交换机离线，轨旁AP跟随离线"
                data_source = "mixed" if ap_identity_known else "stale"
            elif ac_idle:
                ap_status = "offline"
                offline_reason = "ac_idle"
                status_reason = "AC FIT-AP状态为Idle，轨旁AP离线"
            elif ap_side_has_data:
                ap_result = compute_optical_severity(
                    {
                        "module_present": bool(_has_optical_module_data(fit_ap)) or _explicit_no_module(fit_ap),
                        "no_module": _explicit_no_module(fit_ap),
                        "ap_rx_power": fit_ap.get("rx_power"),
                        "ap_port_status": fit_ap.get("ap_port_status"),
                        "alarm_low": fit_ap.get("rx_low_alarm"),
                        "alarm_high": fit_ap.get("rx_high_alarm"),
                        "warning_low": fit_ap.get("rx_low_warning"),
                        "device_type": "ap",
                    }
                )
                ap_status = ap_result.severity
            row = {
                    "site": device.station or normalize_station_value(fit_ap) or "",
                    "ac_device_uuid": fit_ap.get("ac_device_uuid"),
                    "ap_uuid": fit_ap.get("ap_uuid") or (historical_lldp or {}).get("ap_uuid"),
                    "serial_number": fit_ap.get("serial_number") or fit_ap_resource_by_identity.get(ap_identity_key(fit_ap) or ("", ""), {}).get("serial_number"),
                    "device_uuid": device_uuid,
                    "device_name": device.name,
                    "interface_name": interface_name,
                    "link_status": "DOWN" if switch_offline else link_state,
                    "protocol_status": interface.get("protocol_status") or interface.get("protocol"),
                    "description": interface.get("description"),
                    "port_type": _port_type(interface.get("port_status")),
                    "port_status": interface.get("port_status"),
                    "pvid": interface.get("pvid"),
                    "match_source": match_source,
                    "vlan": interface.get("vlan"),
                    "switch_rx_power": optical.get("rx_power"),
                    "switch_tx_power": optical.get("tx_power"),
                    "switch_optical_status": switch_status,
                    "ap_mac": ap_candidate["ap_mac"],
                    "ap_name": ap_candidate["ap_name"],
                    "ap_ip": fit_ap.get("ap_ip"),
                    "ap_state": fit_ap.get("state"),
                    "ap_state_display": fit_ap.get("state_display") or fit_ap.get("state_raw"),
                    "ap_rx_power": ap_candidate["ap_rx_power"] if ap_side_has_data and not switch_offline else None,
                    "ap_tx_power": ap_candidate["ap_tx_power"] if ap_side_has_data and not switch_offline else None,
                    "ap_rx_low_alarm": fit_ap.get("rx_low_alarm"),
                    "ap_rx_low_warning": fit_ap.get("rx_low_warning"),
                    "ap_optical_status": ap_status,
                    "ap_side_has_data": ap_side_has_data,
                    "updated_at": fit_ap.get("updated_at") or optical.get("updated_at") or interface.get("updated_at") or interface.get("collected_at"),
                    "source_device": fit_ap.get("device_name") or fit_ap.get("neighbor_device_name") or device.name,
                    "collection_status": fit_ap.get("status") or ("success" if optical else "not_collected"),
                    "switch_collection_status": switch_collection_status,
                    "offline_reason": offline_reason,
                    "status_reason": status_reason,
                    "data_source": data_source,
                    "has_current_lldp": bool(lldp),
                    "has_historical_lldp": bool(historical_lldp),
                    "has_fit_ap_resource": bool(resource_from_neighbor or resource_from_identity),
                    "is_ap_offline": bool(offline_reason),
                }
            _ensure_ap_optical_status(row)
            app_logger.log_info(
                "TRACKSIDE_AP_ROW_SOURCE",
                (
                    f"site={row.get('site')}, switch={row.get('device_name')}, interface={row.get('interface_name')}, "
                    f"switch_status={row.get('switch_optical_status')}, ap={row.get('ap_name') or row.get('ap_mac')}, "
                    f"ap_status={row.get('ap_optical_status')}, updated_at={row.get('updated_at')}"
                ),
            )
            result.append(row)
    result.extend(_offline_ledger_to_trackside_rows(offline_ap_ledger_rows or [], interfaces_by_device, optical_by_device))
    result = _merge_duplicate_ap_rows(result)
    result = _merge_duplicate_trackside_rows(result)
    _log_trackside_identity_coverage(
        devices,
        result,
        interfaces_by_device,
        lldp_by_device or {},
        fit_ap_resource_rows,
        fit_ap_optical_rows,
    )
    return sorted(result, key=lambda row: (str(row.get("site") or ""), str(row.get("device_name") or ""), interface_sort_key(row.get("interface_name"))))


def merge_fit_ap_rows_by_identity(rows: list[dict[str, object | None]]) -> list[dict[str, object | None]]:
    merged: dict[tuple[str, str], dict[str, object | None]] = {}
    passthrough: list[dict[str, object | None]] = []
    for row in rows or []:
        key = ap_identity_key(row)
        if not key:
            passthrough.append(dict(row))
            continue
        existing = merged.get(key)
        if existing is None or _fit_ap_prefer_score(row) >= _fit_ap_prefer_score(existing):
            merged[key] = {**existing, **row} if existing else dict(row)
        else:
            for field, value in row.items():
                if _is_missing_display(existing.get(field)) and not _is_missing_display(value):
                    existing[field] = value
    return [*merged.values(), *passthrough]


def _latest_rows_by_normalized_interface(rows: list[dict[str, object | None]], field: str) -> dict[str, dict[str, object | None]]:
    latest: dict[str, dict[str, object | None]] = {}
    for row in rows or []:
        key = normalize_interface_name(row.get(field)).casefold()
        if not key:
            continue
        existing = latest.get(key)
        if existing is None or _fact_prefer_score(row) >= _fact_prefer_score(existing):
            latest[key] = row
    return latest


def _build_historical_lldp_index(rows: list[dict[str, object | None]]) -> dict[tuple[str, str], dict[str, object | None]]:
    latest: dict[tuple[str, str], dict[str, object | None]] = {}
    for row in rows or []:
        interface = normalize_interface_name(row.get("neighbor_interface")).casefold()
        if not interface:
            continue
        for name_field in ("neighbor_device_name", "neighbor_switch_name", "neighbor_switch_sysname", "lldp_neighbor"):
            switch_name = _normalize_name(row.get(name_field))
            if not switch_name:
                continue
            key = (switch_name, interface)
            current = latest.get(key)
            if current is None or _fact_prefer_score(row) >= _fact_prefer_score(current):
                latest[key] = row
    return latest


def _find_historical_lldp_row(
    index: dict[tuple[str, str], dict[str, object | None]],
    device_names: set[str],
    interface_name: object,
) -> dict[str, object | None]:
    interface = normalize_interface_name(interface_name).casefold()
    if not interface:
        return {}
    for name in device_names:
        row = index.get((name, interface))
        if row:
            return row
    return {}


def _log_trackside_identity_coverage(
    devices: list[Device],
    rows: list[dict[str, object | None]],
    interfaces_by_device: dict[str, list[dict[str, object | None]]],
    lldp_by_device: dict[str, list[dict[str, object | None]]],
    fit_ap_resource_rows: list[dict[str, object | None]],
    fit_ap_optical_rows: list[dict[str, object | None]],
) -> None:
    candidate_ap_interface_count = sum(
        1
        for device in devices
        for interface in interfaces_by_device.get(str(device.device_uuid or ""), [])
        if is_trackside_ap_interface(device, interface)[0]
    )
    rows_with_ap_identity = sum(1 for row in rows if normalize_mac(row.get("ap_mac")) or str(row.get("ap_name") or "").strip())
    rows_without_ap_identity = max(len(rows) - rows_with_ap_identity, 0)
    current_lldp_port_count = sum(len(items) for items in lldp_by_device.values())
    preserved_lldp_port_count = sum(1 for row in rows if row.get("data_source") == "historical_lldp")
    fit_ap_optical_success_count = sum(1 for row in fit_ap_optical_rows if str(row.get("status") or "").casefold() == "success")
    fit_ap_optical_failed_count = sum(1 for row in fit_ap_optical_rows if str(row.get("status") or "").casefold() not in {"", "success"})
    app_logger.log_info(
        "TRACKSIDE_AP_IDENTITY_COVERAGE",
        (
            f"station_switch_total={len(devices)}, candidate_ap_interface_count={candidate_ap_interface_count}, "
            f"current_lldp_port_count={current_lldp_port_count}, preserved_lldp_port_count={preserved_lldp_port_count}, "
            f"fit_ap_resource_count={len(fit_ap_resource_rows)}, fit_ap_optical_success_count={fit_ap_optical_success_count}, "
            f"fit_ap_optical_failed_count={fit_ap_optical_failed_count}, trackside_rows_total={len(rows)}, "
            f"rows_with_ap_identity={rows_with_ap_identity}, rows_without_ap_identity={rows_without_ap_identity}"
        ),
    )
    for row in rows:
        if normalize_mac(row.get("ap_mac")) or str(row.get("ap_name") or "").strip():
            continue
        app_logger.log_warning(
            "TRACKSIDE_AP_IDENTITY_MISSING",
            (
                f"site={row.get('site')}, switch={row.get('device_name')}, interface={row.get('interface_name')}, "
                f"pvid={row.get('pvid')}, description={row.get('description')}, "
                f"has_current_lldp={bool(row.get('has_current_lldp'))}, "
                f"has_historical_lldp={bool(row.get('has_historical_lldp'))}, "
                f"has_fit_ap_resource={bool(row.get('has_fit_ap_resource'))}"
            ),
        )


def ap_identity_key(row: dict[str, object | None] | None) -> tuple[str, str] | None:
    if not row:
        return None
    serial = str(row.get("serial_number") or row.get("serial") or "").strip()
    if serial and serial not in {"-", "N/A", "n/a"}:
        return ("serial", serial.casefold())
    mac = normalize_mac(row.get("ap_mac") or row.get("mac"))
    if mac:
        return ("mac", mac.casefold())
    name = str(row.get("ap_name") or "").strip()
    if name and name not in {"-", "N/A", "n/a"}:
        return ("name", name.casefold())
    return None


def _fit_ap_prefer_score(row: dict[str, object | None]) -> tuple[int, int, int, str]:
    return (
        1 if _ap_state(row) == "online" else 0,
        1 if str(row.get("ap_ip") or "").strip() else 0,
        1 if _has_optical_module_data(row) or _has_ap_lldp_data(row) else 0,
        f"{row.get('collected_at') or ''}|{row.get('updated_at') or ''}|{_int_value(row.get('id')):020d}",
    )


def _fact_prefer_score(row: dict[str, object | None]) -> tuple[str, str, int]:
    return (
        str(row.get("collected_at") or ""),
        str(row.get("updated_at") or ""),
        _int_value(row.get("id")),
    )


def _int_value(value: object) -> int:
    try:
        return int(str(value or "0"))
    except ValueError:
        return 0


def _has_ap_lldp_data(row: dict[str, object | None]) -> bool:
    return any(str(row.get(field) or "").strip() for field in ("neighbor_device_name", "neighbor_interface", "lldp_neighbor"))


def _merge_duplicate_ap_rows(rows: list[dict[str, object | None]]) -> list[dict[str, object | None]]:
    merged: dict[tuple[str, str], dict[str, object | None]] = {}
    passthrough: list[dict[str, object | None]] = []
    for row in rows:
        key = ap_identity_key(row)
        if not key:
            passthrough.append(row)
            continue
        existing = merged.get(key)
        if existing is None:
            merged[key] = dict(row)
            continue
        if _trackside_ap_row_prefer_score(row) >= _trackside_ap_row_prefer_score(existing):
            merged[key] = _merged_trackside_ap_row(row, existing)
        else:
            merged[key] = _merged_trackside_ap_row(existing, row)
    return [*merged.values(), *passthrough]


def _trackside_ap_row_prefer_score(row: dict[str, object | None]) -> tuple[int, int, int, str]:
    return (
        0 if bool(row.get("is_ap_offline")) or row.get("offline_reason") in {"ac_idle", "switch_offline"} else 1,
        1 if _ap_state(row) == "online" else 0,
        1 if str(row.get("ap_ip") or "").strip() else 0,
        str(row.get("updated_at") or ""),
    )


def _merged_trackside_ap_row(primary: dict[str, object | None], secondary: dict[str, object | None]) -> dict[str, object | None]:
    result = dict(primary)
    for field, value in secondary.items():
        if _is_missing_display(result.get(field)) and not _is_missing_display(value):
            result[field] = value
    return result


def build_new_online_ap_overview_rows(
    current_resource_rows: list[dict[str, object | None]],
    resource_history_rows: list[dict[str, object | None]],
    trackside_rows: list[dict[str, object | None]],
    unauthenticated_rows: list[dict[str, object | None]] | None = None,
    unauthenticated_history_rows: list[dict[str, object | None]] | None = None,
) -> list[dict[str, object | None]]:
    if unauthenticated_rows is not None:
        return _build_unauthenticated_new_online_rows(
            current_resource_rows,
            trackside_rows,
            unauthenticated_rows or [],
        )
    history_by_identity: dict[tuple[str, str], list[dict[str, object | None]]] = {}
    for row in resource_history_rows or []:
        key = ap_identity_key(row)
        if key:
            history_by_identity.setdefault(key, []).append(row)
    trackside_by_identity = _trackside_rows_by_ap_identity(trackside_rows)
    rows: list[dict[str, object | None]] = []
    for resource in current_resource_rows or []:
        if _ap_state(resource) != "online":
            continue
        key = ap_identity_key(resource)
        if not key:
            continue
        current_collect_run = str(resource.get("collect_run_uuid") or "")
        current_collected_at = str(resource.get("collected_at") or resource.get("updated_at") or "")
        had_previous_history = False
        for history in history_by_identity.get(key, []):
            if current_collect_run and str(history.get("collect_run_uuid") or "") == current_collect_run:
                continue
            history_collected_at = str(history.get("collected_at") or history.get("created_at") or "")
            if current_collected_at and history_collected_at >= current_collected_at:
                continue
            had_previous_history = True
            break
        if had_previous_history:
            continue
        trackside = trackside_by_identity.get(key, {})
        rows.append(
            {
                "site": resource.get("site") or resource.get("site_name") or trackside.get("site"),
                "device_name": trackside.get("device_name") or "-",
                "interface_name": trackside.get("interface_name") or "-",
                "link_status": trackside.get("link_status") or "-",
                "port_type": trackside.get("port_type") or "-",
                "description": trackside.get("description") or "-",
                "pvid": trackside.get("pvid") or "-",
                "vlan": trackside.get("vlan") or "-",
                "switch_rx_power": trackside.get("switch_rx_power") or "-",
                "switch_optical_status": trackside.get("switch_optical_status") or "-",
                "ap_mac": normalize_mac(resource.get("ap_mac")) or "-",
                "ap_name": resource.get("ap_name") or "-",
                "ap_rx_power": trackside.get("ap_rx_power") or "-",
                "ap_optical_status": trackside.get("ap_optical_status") or "-",
                "updated_at": resource.get("updated_at") or resource.get("collected_at"),
                "register_status": resource.get("register_status") or "-",
                "new_online_status": resource.get("new_online_status") or "当前新上线Auto AP",
                "identity_source": resource.get("new_online_source") or "-",
                "current_unauthenticated": resource.get("current_unauthenticated") or "-",
                "current_resource_exists": "是",
                "last_unauthenticated_at": resource.get("last_unauthenticated_at") or "",
                "first_seen_at": resource.get("collected_at") or resource.get("updated_at"),
                "ac_device_name": resource.get("ac_device_name") or resource.get("device_name") or resource.get("ac_device_uuid"),
                "apid": resource.get("apid"),
                "ap_ip": resource.get("ap_ip"),
                "model": resource.get("model"),
                "serial_number": resource.get("serial_number"),
                "group_name": resource.get("group_name"),
                "state_display": resource.get("state_display") or resource.get("state_raw") or resource.get("state"),
                "suggestion": "新上线Auto AP，确认点位后在AC手动固化AP",
            }
        )
    return sorted(rows, key=lambda row: (str(row.get("site") or ""), str(row.get("ap_name") or ""), str(row.get("ap_mac") or "")))


def _build_unauthenticated_new_online_rows(
    current_resource_rows: list[dict[str, object | None]],
    trackside_rows: list[dict[str, object | None]],
    unauthenticated_rows: list[dict[str, object | None]],
) -> list[dict[str, object | None]]:
    resources_by_key = _new_online_identity_index(current_resource_rows)
    trackside_by_key = _new_online_identity_index(trackside_rows)
    rows: list[dict[str, object | None]] = []
    emitted: set[tuple[str, str] | tuple[str, int]] = set()
    for index, source in enumerate(unauthenticated_rows):
        keys = _new_online_identity_keys(source)
        primary_key: tuple[str, str] | tuple[str, int] = keys[0] if keys else ("row", index)
        if primary_key in emitted:
            continue
        emitted.add(primary_key)
        resource = _first_index_match(keys, resources_by_key)
        trackside = _first_index_match(keys, trackside_by_key)
        rows.append(
            {
                "identity_source": "AC未固化Auto AP",
                "register_status": "未固化",
                "new_online_status": "当前新上线Auto AP",
                "site": (resource or {}).get("site") or (resource or {}).get("site_name") or trackside.get("site"),
                "device_name": trackside.get("device_name") or "-",
                "interface_name": trackside.get("interface_name") or "-",
                "link_status": trackside.get("link_status") or "-",
                "port_type": trackside.get("port_type") or "-",
                "description": trackside.get("description") or "-",
                "pvid": trackside.get("pvid") or "-",
                "vlan": trackside.get("vlan") or "-",
                "switch_rx_power": trackside.get("switch_rx_power") or "-",
                "switch_optical_status": trackside.get("switch_optical_status") or "-",
                "ap_mac": normalize_mac((resource or {}).get("ap_mac")) or normalize_mac(source.get("inferred_ap_mac") or source.get("ap_name")) or "-",
                "ap_name": source.get("ap_name") or (resource or {}).get("ap_name") or "-",
                "ap_rx_power": trackside.get("ap_rx_power") or "-",
                "ap_optical_status": trackside.get("ap_optical_status") or "-",
                "updated_at": (resource or {}).get("updated_at") or source.get("updated_at") or source.get("collected_at"),
                "apid": source.get("apid") or (resource or {}).get("apid"),
                "current_unauthenticated": "是",
                "current_resource_exists": "是" if bool(resource) else "否",
                "serial_number": source.get("serial_number") or (resource or {}).get("serial_number"),
                "model": source.get("model") or (resource or {}).get("model"),
                "ap_ip": (resource or {}).get("ap_ip"),
                "group_name": (resource or {}).get("group_name"),
                "state_display": source.get("state_display") or source.get("state_raw") or source.get("state") or (resource or {}).get("state_display"),
                "ac_device_name": (resource or {}).get("ac_device_name") or (resource or {}).get("device_name") or source.get("ac_device_uuid"),
                "last_unauthenticated_at": source.get("collected_at") or source.get("created_at"),
                "suggestion": "新上线Auto AP，确认点位后在AC手动固化AP",
                "first_seen_at": source.get("collected_at") or source.get("created_at"),
            }
        )
    return sorted(rows, key=lambda row: (str(row.get("site") or ""), str(row.get("ap_name") or "")))


def _new_online_identity_index(rows: list[dict[str, object | None]]) -> dict[tuple[str, str], dict[str, object | None]]:
    index: dict[tuple[str, str], dict[str, object | None]] = {}
    for row in rows or []:
        for key in _new_online_identity_keys(row):
            index.setdefault(key, row)
    return index


def _new_online_identity_keys(row: dict[str, object | None]) -> list[tuple[str, str]]:
    keys: list[tuple[str, str]] = []
    serial = str(row.get("serial_number") or row.get("serial") or "").strip()
    if serial and serial not in {"-", "N/A", "n/a"}:
        keys.append(("serial", serial.casefold()))
    mac = normalize_mac(row.get("inferred_ap_mac") or row.get("ap_mac") or row.get("mac") or row.get("ap_name"))
    if mac:
        keys.append(("mac", mac.casefold()))
    ap_name = str(row.get("ap_name") or "").strip()
    if ap_name and ap_name not in {"-", "N/A", "n/a"}:
        keys.append(("name", ap_name.casefold()))
    ac_uuid = str(row.get("ac_device_uuid") or "").strip()
    apid = str(row.get("apid") or row.get("ap_id") or "").strip()
    if ac_uuid and apid:
        keys.append(("apid", f"{ac_uuid.casefold()}:{apid.casefold()}"))
    return keys


def _first_index_match(keys: list[tuple[str, str]], index: dict[tuple[str, str], dict[str, object | None]]) -> dict[str, object | None]:
    for key in keys:
        row = index.get(key)
        if row:
            return row
    return {}


def build_ap_optical_treatment_records(
    trackside_rows: list[dict[str, object | None]],
    ap_optical_history_rows: list[dict[str, object | None]],
    switch_optical_history_rows: list[dict[str, object | None]],
    fit_ap_resource_rows: list[dict[str, object | None]] | None = None,
    resource_history_rows: list[dict[str, object | None]] | None = None,
    offline_ledger_rows: list[dict[str, object | None]] | None = None,
) -> list[dict[str, object | None]]:
    records: list[dict[str, object | None]] = []
    identity_lookup = _ap_identity_lookup_with_sources(
        (
            ("trackside_row", trackside_rows or []),
            ("ap_optical_history", ap_optical_history_rows or []),
            ("fit_ap_resource", fit_ap_resource_rows or []),
            ("resource_history", resource_history_rows or []),
            ("offline_ledger", offline_ledger_rows or []),
        )
    )
    trackside_by_identity = _trackside_rows_by_ap_identity(trackside_rows)
    ap_history_by_identity = _ap_optical_history_by_identity(ap_optical_history_rows)
    switch_history_by_interface = _switch_optical_history_by_interface(switch_optical_history_rows)
    trackside_interface_lookup = _trackside_rows_by_switch_name_interface(trackside_rows)
    offline_ledger_interface_lookup = _offline_ledger_rows_by_switch_interface(offline_ledger_rows or [])
    for _key, current in trackside_by_identity.items():
        history = _ap_history_for_trackside(current, ap_history_by_identity)
        current_item = {
            **current,
            "side": "ap",
            "rx_power": current.get("ap_rx_power"),
            "optical_alarm_status": current.get("ap_optical_status"),
            "collected_at": current.get("updated_at"),
        }
        for record in _build_treatment_records_for_series(AP_SIDE_LABEL, current, history, current_item):
            enrich_treatment_record_ap_identity(
                record,
                identity_lookup=identity_lookup,
                trackside_interface_lookup=trackside_interface_lookup,
                offline_ledger_interface_lookup=offline_ledger_interface_lookup,
            )
            records.append(record)

    switch_rows = _trackside_rows_by_switch_interface(trackside_rows)
    for key, current in switch_rows.items():
        history = switch_history_by_interface.get(key, [])
        current_item = {
            **current,
            "side": "switch",
            "rx_power": current.get("switch_rx_power"),
            "optical_alarm_status": current.get("switch_optical_status"),
            "collected_at": current.get("updated_at"),
        }
        for record in _build_treatment_records_for_series(SWITCH_SIDE_LABEL, current, history, current_item):
            enrich_treatment_record_ap_identity(
                record,
                identity_lookup=identity_lookup,
                trackside_interface_lookup=trackside_interface_lookup,
                offline_ledger_interface_lookup=offline_ledger_interface_lookup,
            )
            records.append(record)
    for record in records:
        enrich_treatment_record_ap_identity(
            record,
            identity_lookup=identity_lookup,
            trackside_interface_lookup=trackside_interface_lookup,
            offline_ledger_interface_lookup=offline_ledger_interface_lookup,
        )
    return sorted(records, key=lambda row: (str(row.get("first_found_at") or ""), str(row.get("site") or ""), str(row.get("ap_name") or "")))


def _build_treatment_records_for_series(
    side_label: str,
    trackside: dict[str, object | None],
    history_rows: list[dict[str, object | None]],
    current_item: dict[str, object | None],
) -> list[dict[str, object | None]]:
    timeline = sorted([dict(row) for row in history_rows or []], key=lambda row: (str(row.get("collected_at") or ""), _int_value(row.get("id"))))
    if current_item:
        timeline.append(dict(current_item))
    records: list[dict[str, object | None]] = []
    open_record: dict[str, object | None] | None = None
    latest_status = ""
    latest_rx_power = None
    for item in timeline:
        status = _optical_treatment_status(item)
        if status in OPTICAL_TREATMENT_IGNORED_STATUSES and status != "normal":
            continue
        collected_at = item.get("collected_at") or item.get("updated_at") or item.get("created_at")
        rx_power = item.get("rx_power")
        latest_status = status or latest_status
        latest_rx_power = rx_power if rx_power not in (None, "") else latest_rx_power
        if status in OPTICAL_TREATMENT_ISSUE_STATUSES:
            if open_record is None:
                open_record = _new_treatment_record(side_label, trackside, item, status, collected_at, rx_power)
            continue
        if status == "normal" and open_record is not None:
            open_record["completed_at"] = collected_at
            open_record["fixed_rx_power"] = rx_power
            open_record["current_rx_power"] = rx_power
            open_record["current_status"] = display_optical_status("normal")
            open_record["treatment_status"] = TREATMENT_CLOSED_LABEL
            records.append(open_record)
            open_record = None
    if open_record is not None:
        open_record["current_rx_power"] = latest_rx_power
        open_record["current_status"] = display_optical_status(latest_status)
        open_record["treatment_status"] = TREATMENT_OPEN_LABEL
        records.append(open_record)
    return records


def _new_treatment_record(
    side_label: str,
    trackside: dict[str, object | None],
    item: dict[str, object | None],
    status: str,
    collected_at: object,
    rx_power: object,
) -> dict[str, object | None]:
    return {
        "site": trackside.get("site"),
        "ap_name": trackside.get("ap_name"),
        "ap_mac": trackside.get("ap_mac"),
        "serial_number": trackside.get("serial_number"),
        "side": side_label,
        "device_name": trackside.get("device_name"),
        "interface_name": trackside.get("interface_name"),
        "issue_type": _optical_issue_type(status),
        "first_found_at": collected_at,
        "completed_at": "",
        "first_rx_power": rx_power,
        "fixed_rx_power": "",
        "current_rx_power": rx_power,
        "current_status": display_optical_status(status),
        "treatment_status": TREATMENT_OPEN_LABEL,
        "remark": "",
    }


def _optical_treatment_status(row: dict[str, object | None]) -> str:
    status = row.get("optical_alarm_status") or row.get("alarm_status") or row.get("status")
    text = str(status or "").strip().casefold()
    if text in OPTICAL_TREATMENT_ISSUE_STATUSES or text in OPTICAL_TREATMENT_IGNORED_STATUSES:
        return text
    if not row:
        return ""
    return _optical_status_from_history(row, "ap").casefold()


def _optical_issue_type(status: str) -> str:
    return {
        "notice": ISSUE_TYPE_NOTICE_LABEL,
        "warning": ISSUE_TYPE_NOTICE_LABEL,
        "alarm": ISSUE_TYPE_ALARM_LABEL,
        "link_abnormal": ISSUE_TYPE_LINK_ABNORMAL_LABEL,
        "link_down": ISSUE_TYPE_LINK_ABNORMAL_LABEL,
        "no_light": ISSUE_TYPE_NO_LIGHT_LABEL,
    }.get(status, ISSUE_TYPE_OPTICAL_ABNORMAL_LABEL)


def _ap_optical_history_matches_trackside(row: dict[str, object | None], trackside: dict[str, object | None]) -> bool:
    row_serial = str(row.get("serial_number") or "").strip()
    current_serial = str(trackside.get("serial_number") or "").strip()
    if row_serial and current_serial:
        return row_serial == current_serial
    row_mac = normalize_mac(row.get("ap_mac"))
    current_mac = normalize_mac(trackside.get("ap_mac"))
    if row_mac and current_mac:
        return row_mac == current_mac
    row_name = str(row.get("ap_name") or "").strip().casefold()
    current_name = str(trackside.get("ap_name") or "").strip().casefold()
    if row_name and current_name:
        return row_name == current_name
    row_uuid = str(row.get("ap_uuid") or "").strip()
    current_uuid = str(trackside.get("ap_uuid") or "").strip()
    return bool(row_uuid and current_uuid and row_uuid == current_uuid)


def _ap_identity_lookup(rows: list[dict[str, object | None]]) -> dict[tuple[str, str], dict[str, object | None]]:
    return _ap_identity_lookup_with_sources((("unknown", rows),))


def _ap_identity_lookup_with_sources(
    sources: tuple[tuple[str, list[dict[str, object | None]]], ...],
) -> dict[tuple[str, str], dict[str, object | None]]:
    lookup: dict[tuple[str, str], dict[str, object | None]] = {}
    for source, rows in sources:
        for row in rows or []:
            payload = _ap_identity_payload(row, source)
            for key in _ap_identity_keys(payload):
                existing = lookup.get(key, {})
                merged = _merge_ap_identity_payload(existing, payload)
                if not merged.get("_source"):
                    merged["_source"] = source
                lookup[key] = merged
    return lookup


def _ap_identity_payload(row: dict[str, object | None], source: str = "") -> dict[str, object | None]:
    ap_name = row.get("ap_name")
    ap_mac = normalize_mac(row.get("ap_mac"))
    if _is_missing_display(ap_mac):
        name_as_mac = normalize_mac(ap_name)
        if _is_mac_like(name_as_mac):
            ap_mac = name_as_mac
    return {
        "ap_uuid": row.get("ap_uuid"),
        "ap_name": ap_name,
        "ap_mac": ap_mac,
        "serial_number": row.get("serial_number"),
        "site": row.get("site") or row.get("site_name") or row.get("station"),
        "device_name": row.get("device_name") or row.get("historical_switch_name"),
        "interface_name": row.get("interface_name") or row.get("historical_switch_interface"),
        "_source": source,
    }


def _ap_optical_history_by_identity(rows: list[dict[str, object | None]]) -> dict[tuple[str, str], list[dict[str, object | None]]]:
    grouped: dict[tuple[str, str], list[dict[str, object | None]]] = {}
    for row in rows or []:
        for key in _ap_identity_keys(row):
            grouped.setdefault(key, []).append(row)
    return grouped


def _ap_history_for_trackside(
    trackside: dict[str, object | None],
    history_by_identity: dict[tuple[str, str], list[dict[str, object | None]]],
) -> list[dict[str, object | None]]:
    seen: set[int] = set()
    history: list[dict[str, object | None]] = []
    for key in _ap_identity_keys(trackside):
        for row in history_by_identity.get(key, []):
            marker = id(row)
            if marker in seen:
                continue
            seen.add(marker)
            history.append(row)
    return history


def _switch_optical_history_by_interface(rows: list[dict[str, object | None]]) -> dict[tuple[str, str], list[dict[str, object | None]]]:
    grouped: dict[tuple[str, str], list[dict[str, object | None]]] = {}
    for row in rows or []:
        device_uuid = str(row.get("device_uuid") or "")
        interface_key = normalize_interface_name(row.get("interface_name")).casefold()
        if device_uuid and interface_key:
            grouped.setdefault((device_uuid, interface_key), []).append(row)
    return grouped


def _trackside_rows_by_switch_name_interface(rows: list[dict[str, object | None]]) -> dict[tuple[str, str], dict[str, object | None]]:
    lookup: dict[tuple[str, str], dict[str, object | None]] = {}
    for row in rows or []:
        key = (_normalize_name(row.get("device_name")), normalize_interface_name(row.get("interface_name")).casefold())
        if all(key):
            lookup[key] = row
    return lookup


def _offline_ledger_rows_by_switch_interface(rows: list[dict[str, object | None]]) -> dict[tuple[str, str], dict[str, object | None]]:
    lookup: dict[tuple[str, str], dict[str, object | None]] = {}
    for row in rows or []:
        key = (
            _normalize_name(row.get("historical_switch_name") or row.get("device_name")),
            normalize_interface_name(row.get("historical_switch_interface") or row.get("interface_name")).casefold(),
        )
        if all(key):
            lookup[key] = row
    return lookup


def _ap_identity_keys(row: dict[str, object | None]) -> list[tuple[str, str]]:
    keys: list[tuple[str, str]] = []
    serial = str(row.get("serial_number") or row.get("serial") or "").strip()
    if serial and serial not in {"-", "N/A", "n/a"}:
        keys.append(("serial", serial.casefold()))
    mac = normalize_mac(row.get("ap_mac") or row.get("mac"))
    if mac:
        keys.append(("mac", mac.casefold()))
    name = str(row.get("ap_name") or "").strip()
    if name and name not in {"-", "N/A", "n/a"}:
        keys.append(("name", name.casefold()))
        name_as_mac = normalize_mac(name)
        if not mac and _is_mac_like(name_as_mac):
            keys.append(("mac", name_as_mac.casefold()))
    ap_uuid = str(row.get("ap_uuid") or "").strip()
    if ap_uuid:
        keys.append(("uuid", ap_uuid.casefold()))
    return keys


def _is_mac_like(value: object) -> bool:
    return bool(re.fullmatch(r"[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}", str(value or "").strip().casefold()))


def _merge_ap_identity_payload(
    primary: dict[str, object | None],
    secondary: dict[str, object | None],
) -> dict[str, object | None]:
    result = dict(primary)
    for field in ("ap_uuid", "ap_name", "ap_mac", "serial_number", "site", "device_name", "interface_name"):
        if _is_missing_display(result.get(field)) and not _is_missing_display(secondary.get(field)):
            result[field] = secondary.get(field)
    if _is_missing_display(result.get("_source")) and not _is_missing_display(secondary.get("_source")):
        result["_source"] = secondary.get("_source")
    return result


def _complete_treatment_record_ap_identity(
    record: dict[str, object | None],
    identity_lookup: dict[tuple[str, str], dict[str, object | None]],
) -> None:
    matched: dict[str, object | None] = {}
    for key in _ap_identity_keys(record):
        matched = identity_lookup.get(key, {})
        if matched:
            break
    if not matched:
        return
    for field in ("ap_name", "ap_mac", "serial_number"):
        if _is_missing_display(record.get(field)) and not _is_missing_display(matched.get(field)):
            record[field] = matched.get(field)
    if _is_missing_display(record.get("ap_mac")):
        name_as_mac = normalize_mac(record.get("ap_name"))
        if _is_mac_like(name_as_mac):
            record["ap_mac"] = name_as_mac


def enrich_treatment_record_ap_identity(
    record: dict[str, object | None],
    *,
    identity_lookup: dict[tuple[str, str], dict[str, object | None]],
    trackside_interface_lookup: dict[tuple[str, str], dict[str, object | None]],
    offline_ledger_interface_lookup: dict[tuple[str, str], dict[str, object | None]],
) -> tuple[dict[str, object | None], str]:
    before_ap_name = record.get("ap_name")
    before_ap_mac = record.get("ap_mac")
    source = "not_found"
    for key in _ap_identity_keys(record):
        matched = identity_lookup.get(key)
        if matched:
            _fill_treatment_record_identity(record, matched)
            source = f"{matched.get('_source') or 'identity'}_{key[0]}"
            break
    if _record_identity_missing(record):
        interface_key = (_normalize_name(record.get("device_name")), normalize_interface_name(record.get("interface_name")).casefold())
        matched = trackside_interface_lookup.get(interface_key)
        if matched:
            _fill_treatment_record_identity(record, _ap_identity_payload(matched, "trackside_row"))
            source = "trackside_row"
        if _record_identity_missing(record):
            matched = offline_ledger_interface_lookup.get(interface_key)
            if matched:
                _fill_treatment_record_identity(record, _ap_identity_payload(matched, "offline_ledger"))
                source = "offline_ledger_interface"
    if _is_missing_display(record.get("ap_mac")):
        name_as_mac = normalize_mac(record.get("ap_name"))
        if _is_mac_like(name_as_mac):
            record["ap_mac"] = name_as_mac
            if source == "not_found":
                source = "ap_name_mac_fallback"
    changed = before_ap_name != record.get("ap_name") or before_ap_mac != record.get("ap_mac")
    if changed:
        app_logger.log_info(
            "TRACKSIDE_AP_TREATMENT_IDENTITY_ENRICH",
            " ".join(
                [
                    f"side={record.get('side') or ''}",
                    f"station={record.get('site') or ''}",
                    f"switch={record.get('device_name') or ''}",
                    f"interface={record.get('interface_name') or ''}",
                    f"before_ap_name={before_ap_name or '-'}",
                    f"before_ap_mac={before_ap_mac or '-'}",
                    f"serial={record.get('serial_number') or ''}",
                    f"after_ap_name={record.get('ap_name') or '-'}",
                    f"after_ap_mac={record.get('ap_mac') or '-'}",
                    f"source={source}",
                ]
            ),
        )
    elif _record_identity_missing(record):
        app_logger.log_info(
            "TRACKSIDE_AP_TREATMENT_IDENTITY_MISSING",
            " ".join(
                [
                    f"side={record.get('side') or ''}",
                    f"station={record.get('site') or ''}",
                    f"switch={record.get('device_name') or ''}",
                    f"interface={record.get('interface_name') or ''}",
                    f"serial={record.get('serial_number') or ''}",
                    "reason=no_trackside_or_offline_ledger_match",
                ]
            ),
        )
    return record, source


def _fill_treatment_record_identity(record: dict[str, object | None], source: dict[str, object | None]) -> None:
    for field in ("ap_name", "ap_mac", "serial_number", "site"):
        if _is_missing_display(record.get(field)) and not _is_missing_display(source.get(field)):
            record[field] = source.get(field)


def _record_identity_missing(record: dict[str, object | None]) -> bool:
    return _is_missing_display(record.get("ap_name")) or _is_missing_display(record.get("ap_mac")) or _is_missing_display(record.get("serial_number"))


def _trackside_rows_by_ap_identity(rows: list[dict[str, object | None]]) -> dict[tuple[str, str], dict[str, object | None]]:
    result: dict[tuple[str, str], dict[str, object | None]] = {}
    for row in rows or []:
        key = ap_identity_key(row)
        if key and (key not in result or _trackside_ap_row_prefer_score(row) >= _trackside_ap_row_prefer_score(result[key])):
            result[key] = row
    return result


def _trackside_rows_by_switch_interface(rows: list[dict[str, object | None]]) -> dict[tuple[str, str], dict[str, object | None]]:
    result: dict[tuple[str, str], dict[str, object | None]] = {}
    for row in rows or []:
        device_uuid = str(row.get("device_uuid") or "")
        interface_key = normalize_interface_name(row.get("interface_name")).casefold()
        if device_uuid and interface_key:
            result[(device_uuid, interface_key)] = row
    return result


def trackside_row_status(row: dict[str, object | None]) -> str:
    if bool(row.get("is_ap_offline")) or row.get("offline_reason") in {"switch_offline", "ac_idle"}:
        return "offline"
    switch_status = str(row.get("switch_optical_status") or "")
    ap_status = str(row.get("ap_optical_status") or "") if has_ap_side_optical_data(row) else ""
    return worse_optical_severity(switch_status, ap_status)


def has_ap_side_optical_data(row: dict[str, object | None]) -> bool:
    if not row:
        return False
    if bool(row.get("is_ap_offline")):
        return True
    if "ap_side_has_data" in row:
        return bool(row.get("ap_side_has_data"))
    if _explicit_no_module(row):
        return True
    if _is_missing_display(row.get("ap_mac")) or _is_missing_display(row.get("ap_name")) or _is_missing_display(row.get("ap_rx_power")):
        return False
    return _has_optical_module_data({"rx_power": row.get("ap_rx_power"), "tx_power": row.get("ap_tx_power")})


def format_ap_side_alarm(row: dict[str, object | None], language: str = "zh") -> str:
    status = _ap_optical_status_for_display(row)
    if not has_ap_side_optical_data(row) and not _has_valid_rx_power(row.get("ap_rx_power")):
        return AP_SIDE_MISSING_DISPLAY
    if status == "unknown":
        return AP_SIDE_MISSING_DISPLAY
    return display_optical_status(status, language) if status else AP_SIDE_MISSING_DISPLAY


def _ensure_ap_optical_status(row: dict[str, object | None]) -> None:
    if bool(row.get("is_ap_offline")) or row.get("offline_reason") in {"switch_offline", "ac_idle"}:
        return
    status = str(row.get("ap_optical_status") or "").strip().casefold()
    if status not in {"", "unknown"} or not _has_valid_rx_power(row.get("ap_rx_power")):
        return
    row["ap_optical_status"] = _compute_ap_optical_status_from_row(row)
    row["ap_side_has_data"] = True


def _ap_optical_status_for_display(row: dict[str, object | None]) -> str:
    if bool(row.get("is_ap_offline")) or row.get("offline_reason") in {"switch_offline", "ac_idle"}:
        return "offline"
    status = str(row.get("ap_optical_status") or "").strip().casefold()
    if status in {"", "unknown"} and _has_valid_rx_power(row.get("ap_rx_power")):
        return _compute_ap_optical_status_from_row(row)
    return status


def _compute_ap_optical_status_from_row(row: dict[str, object | None]) -> str:
    return compute_optical_severity(
        {
            "module_present": True,
            "no_module": _explicit_no_module(row),
            "ap_rx_power": row.get("ap_rx_power") or row.get("rx_power"),
            "ap_port_status": row.get("ap_port_status") or row.get("port_status"),
            "alarm_low": row.get("ap_rx_low_alarm") or row.get("rx_low_alarm"),
            "warning_low": row.get("ap_rx_low_warning") or row.get("rx_low_warning"),
            "device_type": "ap",
        }
    ).severity


def _switch_has_valid_light(optical: dict[str, object | None]) -> bool:
    rx_power = _float_value((optical or {}).get("rx_power"))
    return rx_power is not None and rx_power > -35


def _should_mark_switch_link_abnormal(
    link_state: object,
    switch_result: object,
    optical: dict[str, object | None],
    switch_collection_status: object,
) -> bool:
    if normalize_link_state(link_state) != "DOWN":
        return False
    if _is_switch_collection_offline(switch_collection_status):
        return False
    severity = str(getattr(switch_result, "severity", "") or "").strip().casefold()
    if severity in {"no_module", "no_light"}:
        return False
    if _explicit_no_module(optical):
        return False
    return _switch_has_valid_light(optical or {})


def _has_valid_rx_power(value: object) -> bool:
    return _float_value(value) is not None


def _float_value(value: object) -> float | None:
    if value is None:
        return None
    match = re.search(r"[-+]?\d+(?:\.\d+)?", str(value))
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def format_trackside_display_value(field: str, row: dict[str, object | None], language: str = "zh") -> str:
    if field == "match_source":
        return _match_source_label(row.get(field), language)
    if field == "port_type":
        return _port_type(row.get("port_type") or row.get("port_status"))
    if field == "link_status":
        return normalize_link_state(row.get("link_status") or row.get("link") or row.get("status"))
    if field == "protocol_status":
        return normalize_link_state(row.get("protocol_status") or row.get("protocol"))
    if field == "ap_optical_status":
        if bool(row.get("is_ap_offline")) or row.get("offline_reason") in {"switch_offline", "ac_idle"}:
            return OFFLINE_AP_STATUS_TEXT
        return format_ap_side_alarm(row, language)
    if field in AP_SIDE_DISPLAY_FIELDS and not has_ap_side_optical_data(row):
        return AP_SIDE_MISSING_DISPLAY
    value = row.get(field)
    if field == "switch_optical_status" and value:
        if row.get("offline_reason") == "switch_offline" or _is_switch_collection_offline(row.get("switch_collection_status")):
            return "交换机离线" if not language.startswith("en") else "Switch Offline"
        return display_optical_status(str(value), language)
    if field == "ap_optical_status" and value:
        return display_optical_status(str(value), language)
    return str(value) if value not in (None, "") else AP_SIDE_MISSING_DISPLAY


def _match_source_label(value: object, language: str = "zh") -> str:
    source = str(value or "none")
    if language.startswith("en"):
        return {
            "description": "Description",
            "pvid": "PVID",
            "description+pvid": "Description+PVID",
            "none": "-",
        }.get(source, source or "-")
    return MATCH_SOURCE_LABELS.get(source, source or "-")


def filter_trackside_ap_business_rows(rows: list[dict[str, object | None]], site: object = "", search: object = "") -> list[dict[str, object | None]]:
    site_text = str(site or "").strip()
    search_text = str(search or "").strip().casefold()
    result = rows
    if site_text:
        result = [row for row in result if str(row.get("site") or "") == site_text]
    if search_text:
        fields = ("ap_name", "ap_mac", "device_name", "interface_name", "site", "match_source", "pvid", "vlan")
        result = [row for row in result if any(search_text in str(row.get(field) or "").casefold() for field in fields)]
    return result


def build_trackside_site_filter_items(rows: list[dict[str, object | None]], all_label: str) -> list[tuple[str, str]]:
    sites = sorted({str(row.get("site") or "").strip() for row in rows if str(row.get("site") or "").strip()})
    return [(all_label, ""), *[(site, site) for site in sites]]


def export_trackside_ap_business_xlsx(
    path: Path,
    rows: list[dict[str, object | None]],
    columns: tuple[tuple[str, str], ...],
    headers: list[str],
    ap_online_overview_rows: list[dict[str, object | None]] | None = None,
    ap_online_overview_columns: tuple[tuple[str, str], ...] | None = None,
    ap_online_overview_headers: list[str] | None = None,
    new_online_ap_rows: list[dict[str, object | None]] | None = None,
    new_online_ap_columns: tuple[tuple[str, str], ...] | None = None,
    new_online_ap_headers: list[str] | None = None,
    new_online_ap_sheet_title: str = "\u65b0\u589e\u4e0a\u7ebfAP\u6982\u89c8",
    ap_optical_treatment_rows: list[dict[str, object | None]] | None = None,
    ap_optical_treatment_columns: tuple[tuple[str, str], ...] | None = None,
    ap_optical_treatment_headers: list[str] | None = None,
    ap_optical_treatment_sheet_title: str = "AP\u5149\u8870\u5904\u7406\u8bb0\u5f55",
    offline_ap_stats: dict[str, object | None] | None = None,
    offline_ap_ledger_rows: list[dict[str, object | None]] | None = None,
    offline_ap_stats_headers: list[str] | None = None,
    offline_ap_ledger_headers: list[str] | None = None,
    progress_callback=None,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    from netconsole.ui.table.table_autosize_engine import apply_worksheet_autofit

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "\u8f68\u65c1AP\u4e1a\u52a1"
    def log_write_phase(phase: str, start: float, **values: object) -> None:
        elapsed_ms = int((perf_counter() - start) * 1000)
        details = " ".join(f"{key}={value}" for key, value in values.items())
        app_logger.log_info("TRACKSIDE_EXPORT_PROFILE", f"phase={phase} elapsed_ms={elapsed_ms}" + (f" {details}" if details else ""))

    if progress_callback:
        progress_callback("trackside.export.progress_rows")
    phase_start = perf_counter()
    sheet.append(headers)
    fills = {
        status: PatternFill(fill_type="solid", fgColor=color)
        for status, color in TRACKSIDE_OPTICAL_COLOR_RGB.items()
        if color
    }
    header_fill = PatternFill(fill_type="solid", fgColor=TRACKSIDE_EXPORT_HEADER_FILL)
    for row in rows:
        sheet.append([_export_value(field, row) for _key, field in columns])
        fill = fills.get(trackside_row_status(row))
        for cell in sheet[sheet.max_row]:
            if fill:
                cell.fill = fill
    alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    header_font = Font(bold=True)
    _format_export_sheet(sheet, alignment, border, header_font, header_fill)
    sheet.auto_filter.ref = sheet.dimensions
    log_write_phase("write_trackside_sheet", phase_start, rows=len(rows))
    phase_start = perf_counter()
    build_current_optical_abnormal_sheet(workbook, sheet, rows)
    log_write_phase("write_current_optical_abnormal_sheet", phase_start, rows=sum(1 for row in rows if is_current_optical_abnormal_export_row(row)))
    if progress_callback:
        progress_callback("trackside.export.progress_overview")
    phase_start = perf_counter()
    _append_ap_overview_sheet(
        workbook,
        rows,
        alignment,
        border,
        header_font,
        ap_online_overview_rows,
        ap_online_overview_columns,
        ap_online_overview_headers,
        header_fill,
    )
    log_write_phase("write_ap_online_overview_sheet", phase_start, rows=len(ap_online_overview_rows or []))
    if progress_callback:
        progress_callback("trackside.export.progress_new_online")
    phase_start = perf_counter()
    _append_export_rows_sheet(
        workbook,
        new_online_ap_sheet_title,
        new_online_ap_rows or [],
        new_online_ap_columns or NEW_ONLINE_AP_OVERVIEW_COLUMNS,
        new_online_ap_headers or [key for key, _field in NEW_ONLINE_AP_OVERVIEW_COLUMNS],
        alignment,
        border,
        header_font,
        header_fill,
    )
    log_write_phase("write_new_online_ap_sheet", phase_start, rows=len(new_online_ap_rows or []))
    if progress_callback:
        progress_callback("trackside.export.progress_optical_treatment")
    phase_start = perf_counter()
    sorted_treatment_rows = _sort_ap_optical_treatment_rows(ap_optical_treatment_rows or [])
    _append_export_rows_sheet(
        workbook,
        ap_optical_treatment_sheet_title,
        sorted_treatment_rows,
        ap_optical_treatment_columns or AP_OPTICAL_TREATMENT_RECORD_COLUMNS,
        ap_optical_treatment_headers or [key for key, _field in AP_OPTICAL_TREATMENT_RECORD_COLUMNS],
        alignment,
        border,
        header_font,
        header_fill,
        fills,
        _ap_optical_treatment_row_fill_status,
        preserve_ap_identity=True,
    )
    log_write_phase("write_optical_treatment_sheet", phase_start, rows=len(sorted_treatment_rows))
    if offline_ap_stats is not None and offline_ap_ledger_rows is not None:
        if progress_callback:
            progress_callback("trackside.export.progress_offline")
        phase_start = perf_counter()
        stats_sheet = workbook.create_sheet("AP\u79bb\u7ebf\u60c5\u51b5")
        write_offline_ap_stats_sheet(stats_sheet, offline_ap_stats, offline_ap_stats_headers or [key for key, _field in OFFLINE_AP_STATS_COLUMNS])
        ledger_sheet = workbook.create_sheet("\u79bb\u7ebfAP\u53f0\u8d26")
        write_offline_ap_ledger_sheet(ledger_sheet, offline_ap_ledger_rows, offline_ap_ledger_headers or [key for key, _field in OFFLINE_AP_LEDGER_COLUMNS])
        log_write_phase("write_offline_ap_sheets", phase_start, rows=len(offline_ap_ledger_rows or []))
    if progress_callback:
        progress_callback("trackside.export.progress_summary")
    phase_start = perf_counter()
    _append_switch_optical_summary_sheet(workbook, rows, alignment, border, header_font, header_fill)
    log_write_phase("write_switch_optical_summary_sheet", phase_start, rows=len(rows))
    phase_start = perf_counter()
    for worksheet in workbook.worksheets:
        apply_worksheet_autofit(worksheet, maximum=60)
    _set_switch_optical_summary_widths(workbook)
    log_write_phase("autofit_sheets", phase_start, sheets=len(workbook.worksheets))
    if progress_callback:
        progress_callback("trackside.export.progress_save")
    phase_start = perf_counter()
    workbook.save(path)
    log_write_phase("save_workbook", phase_start, path=Path(path).name)


# Legacy aliases removed — status is now computed real-time from raw data.


def _find_fit_ap_row(fit_ap_index: dict[tuple[str, str], dict[str, object | None]], device_names: set[str], interface_name: object) -> dict[str, object | None]:
    interface_key = normalize_interface_name(interface_name).casefold()
    for device_name in device_names:
        row = fit_ap_index.get((device_name, interface_key))
        if row:
            return row
    return {}


def _merge_resource_with_optical(resource: dict[str, object | None] | None, optical_by_mac: dict[str, dict[str, object | None]]) -> dict[str, object | None]:
    if not resource:
        return {}
    optical = optical_by_mac.get(normalize_mac(resource.get("ap_mac")), {})
    return {**resource, **optical}


def _offline_ledger_to_trackside_rows(
    rows: list[dict[str, object | None]],
    interfaces_by_device: dict[str, list[dict[str, object | None]]] | None = None,
    optical_by_device: dict[str, list[dict[str, object | None]]] | None = None,
) -> list[dict[str, object | None]]:
    result: list[dict[str, object | None]] = []
    interface_indexes = {device_uuid: _latest_rows_by_normalized_interface(items, "interface_name") for device_uuid, items in (interfaces_by_device or {}).items()}
    optical_indexes = {device_uuid: _latest_rows_by_normalized_interface(items, "interface_name") for device_uuid, items in (optical_by_device or {}).items()}
    for row in rows:
        device_uuid = str(row.get("device_uuid") or "")
        interface_key = normalize_interface_name(row.get("historical_switch_interface")).casefold()
        interface = interface_indexes.get(device_uuid, {}).get(interface_key, {})
        optical = optical_indexes.get(device_uuid, {}).get(interface_key, {})
        link_state = normalize_link_state(interface.get("link_status") or interface.get("link"))
        switch_collection_status = row.get("switch_collection_status") or interface.get("switch_collection_status") or interface.get("collection_status")
        switch_result = compute_optical_severity(
            {
                "module_present": bool(_has_optical_module_data(optical)),
                "no_module": _explicit_no_module(optical),
                "switch_rx_power": optical.get("rx_power"),
                "switch_port_status": optical.get("port_status"),
                "alarm_low": optical.get("rx_low_alarm"),
                "alarm_high": optical.get("rx_high_alarm"),
                "warning_low": optical.get("rx_low_warning"),
                "device_type": "switch",
            }
        )
        switch_status = switch_result.severity
        if _should_mark_switch_link_abnormal(link_state, switch_result, optical, switch_collection_status):
            switch_status = "link_abnormal"
        result.append(
            {
                "site": row.get("site"),
                "ac_device_uuid": row.get("ac_device_uuid"),
                "ap_uuid": row.get("ap_uuid"),
                "device_uuid": device_uuid or row.get("device_uuid"),
                "device_name": row.get("historical_switch_name"),
                "interface_name": row.get("historical_switch_interface"),
                "link_status": link_state,
                "protocol_status": interface.get("protocol_status"),
                "description": interface.get("description") or row.get("offline_remark"),
                "port_type": _port_type(interface.get("port_status")),
                "port_status": interface.get("port_status"),
                "pvid": interface.get("pvid"),
                "match_source": "historical",
                "vlan": interface.get("vlan"),
                "switch_rx_power": optical.get("rx_power"),
                "switch_tx_power": optical.get("tx_power"),
                "switch_optical_status": switch_status,
                "ap_mac": row.get("ap_mac"),
                "ap_name": row.get("ap_name"),
                "ap_ip": row.get("ap_ip"),
                "ap_state": "Idle",
                "ap_state_display": row.get("ap_status"),
                "ap_rx_power": None,
                "ap_tx_power": None,
                "ap_optical_status": "offline",
                "ap_side_has_data": True,
                "updated_at": row.get("last_lldp_at"),
                "source_device": row.get("historical_switch_name"),
                "collection_status": "historical",
                "is_ap_offline": True,
                "offline_reason": row.get("offline_reason") or "ac_idle",
                "status_reason": row.get("status_reason") or "AC FIT-AP状态为Idle，轨旁AP离线",
                "data_source": row.get("data_source") or "historical",
                "switch_collection_status": switch_collection_status,
                "offline_remark": row.get("offline_remark"),
            }
        )
    return result


def summarize_trackside_ap_online_counts(rows: list[dict[str, object | None]]) -> tuple[int, int]:
    online = 0
    offline = 0
    seen: set[str] = set()
    for row in rows:
        key = str(row.get("ap_uuid") or row.get("ap_mac") or row.get("ap_name") or "").strip()
        if key and key in seen:
            continue
        if key:
            seen.add(key)
        state = _ap_state(row)
        if state == "online":
            online += 1
        elif state == "offline":
            offline += 1
    return online, offline


def _normalize_name(value: object) -> str:
    return str(value or "").strip().casefold()


def _port_type(value: object) -> str:
    text = str(value or "").strip().casefold()
    return text if text in {"access", "trunk", "hybrid"} else "unknown"


def normalize_link_state(value: object) -> str:
    text = str(value or "").strip().upper()
    if not text:
        return "-"
    if text in {"UP", "DOWN"}:
        return text
    if "DOWN" in text:
        return "DOWN"
    if "UP" in text:
        return "UP"
    return "-"


def normalize_mac(value: object) -> str:
    import re

    hex_text = re.sub(r"[^0-9a-fA-F]", "", str(value or ""))
    if len(hex_text) != 12:
        return str(value or "").strip().casefold()
    hex_text = hex_text.casefold()
    return f"{hex_text[0:4]}-{hex_text[4:8]}-{hex_text[8:12]}"


def _switch_collection_status(device: Device, interface: dict[str, object | None], optical: dict[str, object | None]) -> str:
    for source in (interface, optical):
        for field in ("switch_collection_status", "collection_status", "collect_status"):
            value = str(source.get(field) or "").strip()
            if value:
                return value
    for field in ("switch_collection_status", "collection_status", "collect_status"):
        value = str(getattr(device, field, "") or "").strip()
        if value:
            return value
    return "success" if optical else "not_collected"


def _is_switch_collection_offline(value: object) -> bool:
    text = str(value or "").strip().casefold()
    if not text:
        return False
    return text in {"offline", "switch_offline", "device_offline", "link_down", "离线", "交换机离线"} or "真实离线" in text


def enrich_trackside_export_rows(
    rows: list[dict[str, object | None]],
    fact_repository=None,
    ac_repository=None,
    switch_optical_history_rows: list[dict[str, object | None]] | None = None,
    ap_optical_history_rows: list[dict[str, object | None]] | None = None,
    ap_lldp_history_rows: list[dict[str, object | None]] | None = None,
) -> list[dict[str, object | None]]:
    switch_history_by_interface = _switch_optical_history_by_interface(switch_optical_history_rows or [])
    ap_history_by_identity = _ap_optical_history_by_identity(ap_optical_history_rows or [])
    ap_lldp_history_by_identity = _ap_optical_history_by_identity(ap_lldp_history_rows or [])
    enriched: list[dict[str, object | None]] = []
    for row in rows:
        item = dict(row)
        _apply_switch_optical_change(item, fact_repository, switch_history_by_interface)
        _apply_ap_optical_change(item, ac_repository, ap_history_by_identity)
        _apply_ap_port_change(item, ac_repository, ap_lldp_history_by_identity)
        enriched.append(item)
    return enriched


def optical_change_text(previous_status: object, current_status: object) -> str:
    previous = _normal_abnormal_state(previous_status)
    current = _normal_abnormal_state(current_status)
    if previous and current and previous != current:
        return f"{previous} → {current}"
    return "-"


def ap_port_change_text(previous_switch: object, previous_interface: object, current_switch: object, current_interface: object) -> str:
    previous_switch_text = str(previous_switch or "").strip()
    previous_interface_text = str(previous_interface or "").strip()
    current_switch_text = str(current_switch or "").strip()
    current_interface_text = str(current_interface or "").strip()
    if not all((previous_switch_text, previous_interface_text, current_switch_text, current_interface_text)):
        return "-"
    if previous_switch_text == current_switch_text and normalize_interface_name(previous_interface_text).casefold() == normalize_interface_name(current_interface_text).casefold():
        return "-"
    return f"AP端口变化: {previous_switch_text} {previous_interface_text} → {current_switch_text} {current_interface_text}"


def _apply_switch_optical_change(
    row: dict[str, object | None],
    fact_repository,
    history_by_interface: dict[tuple[str, str], list[dict[str, object | None]]] | None = None,
) -> None:
    row.setdefault("switch_optical_change", "-")
    device_uuid = str(row.get("device_uuid") or "")
    interface_key = normalize_interface_name(row.get("interface_name")).casefold()
    baseline = _optical_transition_baseline_before(
        (history_by_interface or {}).get((device_uuid, interface_key), []),
        row.get("updated_at"),
        row.get("switch_optical_status"),
        "switch",
    )
    if baseline:
        baseline_status = _optical_status_from_history(baseline, "switch")
        row["switch_optical_change"] = optical_change_text(baseline_status, row.get("switch_optical_status"))
        row["history_compared_at"] = baseline.get("collected_at") or baseline.get("created_at") or row.get("history_compared_at")
        return
    if fact_repository is None:
        return
    getter = getattr(fact_repository, "get_previous_optical_history", None)
    if not callable(getter):
        return
    previous = getter(str(row.get("device_uuid") or ""), str(row.get("interface_name") or ""), str(row.get("updated_at") or ""))
    if not previous:
        return
    previous_status = _optical_status_from_history(previous, "switch")
    row["switch_optical_change"] = optical_change_text(previous_status, row.get("switch_optical_status"))
    if row["switch_optical_change"] != "-":
        row["history_compared_at"] = previous.get("collected_at") or previous.get("created_at") or row.get("history_compared_at")


def _apply_ap_optical_change(
    row: dict[str, object | None],
    ac_repository,
    history_by_identity: dict[tuple[str, str], list[dict[str, object | None]]] | None = None,
) -> None:
    row.setdefault("ap_optical_change", "-")
    history_rows = _ap_history_for_trackside(row, history_by_identity or {})
    latest_valid = _latest_valid_ap_optical_history(history_rows)
    if _is_missing_display(row.get("ap_rx_power")) and latest_valid and not _is_missing_display(latest_valid.get("rx_power")):
        row["ap_rx_power"] = latest_valid.get("rx_power")
        row["ap_rx_low_alarm"] = latest_valid.get("rx_low_alarm") or row.get("ap_rx_low_alarm")
        row["ap_rx_low_warning"] = latest_valid.get("rx_low_warning") or row.get("ap_rx_low_warning")
        row["ap_last_valid_rx_power"] = latest_valid.get("rx_power")
        row["ap_last_valid_collected_at"] = latest_valid.get("collected_at") or latest_valid.get("created_at")
        row["ap_optical_data_source"] = "沿用历史"
        row["ap_optical_missing_reason"] = "not_collected" if str(row.get("collection_status") or "").casefold() in {"", "not_collected"} else "overwritten_by_failed_row"
        _ensure_ap_optical_status(row)
    baseline = _optical_transition_baseline_before(history_rows, row.get("updated_at"), row.get("ap_optical_status"), "ap")
    if baseline:
        baseline_status = _optical_status_from_history(baseline, "ap")
        row["ap_optical_change"] = optical_change_text(baseline_status, row.get("ap_optical_status"))
        row["history_compared_at"] = baseline.get("collected_at") or baseline.get("created_at") or row.get("history_compared_at")
        return
    if ac_repository is None:
        return
    getter = getattr(ac_repository, "get_previous_ap_optical_history", None)
    if not callable(getter):
        return
    previous = getter(ap_identity_filter(row), str(row.get("updated_at") or ""))
    if not previous:
        return
    previous_status = _optical_status_from_history(previous, "ap")
    row["ap_optical_change"] = optical_change_text(previous_status, row.get("ap_optical_status"))
    if row["ap_optical_change"] != "-":
        row["history_compared_at"] = previous.get("collected_at") or previous.get("created_at") or row.get("history_compared_at")


def _previous_row_before(rows: list[dict[str, object | None]], before: object) -> dict[str, object | None] | None:
    candidates = _history_rows_before(rows, before)
    if not candidates:
        return None
    return candidates[0]


def _history_rows_before(rows: list[dict[str, object | None]], before: object) -> list[dict[str, object | None]]:
    before_text = str(before or "")
    candidates = [
        row
        for row in rows or []
        if not before_text or str(row.get("collected_at") or row.get("created_at") or "") < before_text
    ]
    return sorted(
        candidates,
        key=lambda row: (str(row.get("collected_at") or row.get("created_at") or ""), _int_value(row.get("id"))),
        reverse=True,
    )


def _optical_transition_baseline_before(
    rows: list[dict[str, object | None]],
    before: object,
    current_status: object,
    side: str,
) -> dict[str, object | None] | None:
    current_state = _normal_abnormal_key(current_status)
    if not current_state:
        return None
    target_state = "abnormal" if current_state == "normal" else "normal"
    for row in _history_rows_before(rows, before):
        status = _normal_abnormal_key(_optical_status_from_history(row, side))
        if status == target_state:
            return row
    return None


def _latest_valid_ap_optical_history(rows: list[dict[str, object | None]]) -> dict[str, object | None] | None:
    candidates = [row for row in rows or [] if not _is_missing_display(row.get("rx_power"))]
    if not candidates:
        return None
    return max(candidates, key=lambda row: (str(row.get("collected_at") or row.get("created_at") or ""), _int_value(row.get("id"))))


def _apply_ap_port_change(
    row: dict[str, object | None],
    ac_repository,
    history_by_identity: dict[tuple[str, str], list[dict[str, object | None]]] | None = None,
) -> None:
    row.setdefault("ap_port_change", "-")
    row["current_switch"] = row.get("device_name") or "-"
    row["current_interface"] = row.get("interface_name") or "-"
    previous = _ap_port_transition_baseline_before(
        _ap_history_for_trackside(row, history_by_identity or {}),
        row.get("updated_at"),
        row.get("device_name"),
        row.get("interface_name"),
    )
    if previous:
        previous_switch = _previous_lldp_switch(previous)
        previous_interface = _previous_lldp_interface(previous)
        row["previous_switch"] = previous_switch or "-"
        row["previous_interface"] = previous_interface or "-"
        row["ap_port_change"] = ap_port_change_text(previous_switch, previous_interface, row.get("device_name"), row.get("interface_name"))
        row["history_compared_at"] = previous.get("collected_at") or previous.get("created_at") or row.get("history_compared_at") or "-"
        return
    if ac_repository is None:
        row.setdefault("previous_switch", "-")
        row.setdefault("previous_interface", "-")
        row.setdefault("history_compared_at", "-")
        return
    getter = getattr(ac_repository, "get_previous_ap_lldp_history", None)
    previous = getter(ap_identity_filter(row), str(row.get("updated_at") or "")) if callable(getter) else None
    previous_switch = _previous_lldp_switch(previous or {})
    previous_interface = _previous_lldp_interface(previous or {})
    row["previous_switch"] = previous_switch or "-"
    row["previous_interface"] = previous_interface or "-"
    row["ap_port_change"] = ap_port_change_text(previous_switch, previous_interface, row.get("device_name"), row.get("interface_name"))
    row["history_compared_at"] = (previous or {}).get("collected_at") or (previous or {}).get("created_at") or row.get("history_compared_at") or "-"


def _ap_port_transition_baseline_before(
    rows: list[dict[str, object | None]],
    before: object,
    current_switch: object,
    current_interface: object,
) -> dict[str, object | None] | None:
    current_switch_text = str(current_switch or "").strip()
    current_interface_key = normalize_interface_name(current_interface).casefold()
    if not current_switch_text or not current_interface_key:
        return None
    for row in _history_rows_before(rows, before):
        previous_switch = str(_previous_lldp_switch(row) or "").strip()
        previous_interface_key = normalize_interface_name(_previous_lldp_interface(row)).casefold()
        if not previous_switch or not previous_interface_key:
            continue
        if previous_switch.casefold() == current_switch_text.casefold() and previous_interface_key == current_interface_key:
            continue
        return row
    return None


def ap_identity_filter(row: dict[str, object | None]) -> dict[str, str]:
    return {
        "ap_uuid": str(row.get("ap_uuid") or "").strip(),
        "serial_number": str(row.get("serial_number") or "").strip(),
        "ap_mac": normalize_mac(row.get("ap_mac")),
        "ap_name": str(row.get("ap_name") or "").strip(),
    }


def _normal_abnormal_key(status: object) -> str:
    text = str(status or "").strip().casefold()
    if not text or text in {"unknown", "not_collected", "skipped", "offline", "failed", "timeout", "no_module", "-"}:
        return ""
    if text == "normal":
        return "normal"
    return "abnormal"


def _normal_abnormal_state(status: object) -> str:
    text = str(status or "").strip().casefold()
    if not text or text in {"unknown", "not_collected", "skipped", "offline", "failed", "timeout", "no_module", "-"}:
        return ""
    if text == "normal":
        return "正常"
    return "不正常"


def _optical_status_from_history(row: dict[str, object | None], device_type: str) -> str:
    status = row.get("alarm_status") or row.get("optical_alarm_status") or row.get("status")
    normalized_status = str(status or "").strip().casefold()
    if normalized_status in {"normal", "notice", "warning", "alarm", "link_abnormal", "link_down", "no_light", "no_module", "not_collected", "skipped", "offline"}:
        return str(status or "")
    if normalized_status == "unknown" and not _has_valid_rx_power(row.get("rx_power")):
        return str(status or "")
    result = compute_optical_severity(
        {
            "module_present": bool(_has_optical_module_data(row)),
            "rx_power": row.get("rx_power"),
            "alarm_low": row.get("rx_low_alarm"),
            "warning_low": row.get("rx_low_warning"),
            "port_status": row.get("port_status"),
            "device_type": device_type,
        }
    )
    return result.severity


def _previous_lldp_switch(row: dict[str, object | None]) -> object:
    return row.get("neighbor_switch_name") or row.get("neighbor_switch_sysname") or row.get("neighbor_device_name") or row.get("lldp_neighbor")


def _previous_lldp_interface(row: dict[str, object | None]) -> object:
    return row.get("neighbor_interface")


def _is_ac_idle(row: dict[str, object | None]) -> bool:
    state = " ".join(str(row.get(field) or "") for field in ("state", "state_raw", "state_display", "ap_state", "ap_state_display")).strip().casefold()
    if not state:
        return False
    token = state.split("=", 1)[0].strip()
    return token in {"i", "idle"} or "idle" in state


def _trackside_merge_key(row: dict[str, object | None]) -> tuple[str, str, str]:
    site = str(row.get("site") or "").strip().casefold()
    switch = str(row.get("device_uuid") or row.get("switch_uuid") or row.get("device_name") or row.get("switch_name") or "").strip().casefold()
    interface = normalize_interface_name(row.get("interface_name")).casefold()
    return site, switch, interface


def _merge_duplicate_trackside_rows(rows: list[dict[str, object | None]]) -> list[dict[str, object | None]]:
    merged: dict[tuple[str, str, str], dict[str, object | None]] = {}
    passthrough: list[dict[str, object | None]] = []
    for row in rows:
        key = _trackside_merge_key(row)
        if not all(key):
            passthrough.append(row)
            continue
        existing = merged.get(key)
        if existing is None:
            merged[key] = dict(row)
            continue
        _merge_trackside_row(existing, row)
    result = [*merged.values(), *passthrough]
    for row in result:
        _apply_trackside_offline_priority(row)
    return result


def _merge_trackside_row(target: dict[str, object | None], source: dict[str, object | None]) -> None:
    prefer_source_fields = (
        "ap_uuid",
        "ap_mac",
        "ap_name",
        "ap_ip",
        "ap_state",
        "ap_state_display",
        "ap_rx_power",
        "ap_tx_power",
        "ap_rx_low_alarm",
        "ap_rx_low_warning",
        "offline_remark",
    )
    for field in prefer_source_fields:
        if _is_missing_display(target.get(field)) and not _is_missing_display(source.get(field)):
            target[field] = source.get(field)
    for field, value in source.items():
        if field not in target or _is_missing_display(target.get(field)):
            target[field] = value
    if bool(source.get("ap_side_has_data")):
        target["ap_side_has_data"] = True
    if bool(source.get("is_ap_offline")):
        target["is_ap_offline"] = True
    if source.get("offline_reason") == "switch_offline" or target.get("offline_reason") != "switch_offline":
        for field in ("offline_reason", "status_reason", "data_source", "switch_collection_status"):
            if source.get(field):
                target[field] = source.get(field)
    if normalize_link_state(source.get("link_status")) == "DOWN":
        target["link_status"] = "DOWN"
    target["port_type"] = _port_type(target.get("port_type") or source.get("port_type") or target.get("port_status") or source.get("port_status"))


def _apply_trackside_offline_priority(row: dict[str, object | None]) -> None:
    switch_offline = row.get("offline_reason") == "switch_offline" or _is_switch_collection_offline(row.get("switch_collection_status"))
    ac_idle = _is_ac_idle(row) or row.get("offline_reason") == "ac_idle"
    if switch_offline:
        row["link_status"] = "DOWN"
        row["switch_optical_status"] = "offline"
        row["ap_optical_status"] = "offline"
        row["ap_side_has_data"] = True
        row["is_ap_offline"] = True
        row["offline_reason"] = "switch_offline"
        row["status_reason"] = "室内交换机离线，轨旁AP跟随离线"
        row["data_source"] = row.get("data_source") or "mixed"
    elif ac_idle:
        row["ap_optical_status"] = "offline"
        row["ap_side_has_data"] = True
        row["is_ap_offline"] = True
        row["offline_reason"] = "ac_idle"
        row["status_reason"] = row.get("status_reason") or "AC FIT-AP状态为Idle，轨旁AP离线"
    else:
        _ensure_ap_optical_status(row)
    row["port_type"] = _port_type(row.get("port_type") or row.get("port_status"))


def _export_value(field: str, row: dict[str, object | None], *, preserve_ap_identity: bool = False) -> str:
    if field == "completed_at" and row.get("treatment_status") == TREATMENT_OPEN_LABEL:
        return ""
    if preserve_ap_identity and field in {"ap_name", "ap_mac", "serial_number"}:
        value = row.get(field)
        return str(value) if value not in (None, "") else AP_SIDE_MISSING_DISPLAY
    return format_trackside_display_value(field, row)


def _has_optical_module_data(row: dict[str, object | None]) -> bool:
    if not row:
        return False
    module_fields = (
        "rx_power",
        "tx_power",
        "module_model",
        "module_serial_number",
        "module_vendor",
        "wavelength",
        "transmission_distance",
        "connector_type",
        "rx_low_alarm",
        "rx_low_warning",
    )
    return any(row.get(field) not in (None, "") for field in module_fields)


def _has_ap_side_optical_data(fit_ap: dict[str, object | None], candidate: dict[str, object | None]) -> bool:
    if not fit_ap:
        return False
    if _explicit_no_module(fit_ap):
        return True
    if _is_missing_display(candidate.get("ap_mac")) or _is_missing_display(candidate.get("ap_name")) or _is_missing_display(candidate.get("ap_rx_power")):
        return False
    return _has_optical_module_data(fit_ap)


def _explicit_no_module(row: dict[str, object | None]) -> bool:
    text = " ".join(
        str(row.get(field) or "")
        for field in (
            "optical_alarm_status",
            "status",
            "raw_status",
            "ap_raw_status",
            "error_message",
            "message",
        )
    ).strip().casefold()
    if not text:
        return False
    return any(token in text for token in ("no_module", "no module", "no transceiver", "no-transceiver", "\u65e0\u5149\u6a21\u5757"))


def _is_missing_display(value: object) -> bool:
    return str(value or "").strip() in {"", "-"}


def _format_export_sheet(sheet, alignment, border, header_font, header_fill=None) -> None:
    sheet.freeze_panes = "A2"
    for row in sheet.iter_rows():
        sheet.row_dimensions[row[0].row].height = 24 if row[0].row == 1 else 22
        for cell in row:
            cell.alignment = alignment
            cell.border = border
            if cell.row == 1:
                cell.font = header_font
                if header_fill is not None:
                    cell.fill = header_fill


def is_no_light_optical_row(row: dict[str, object | None]) -> bool:
    status_values = [
        row.get("switch_optical_status"),
        row.get("ap_optical_status"),
        row.get("optical_alarm_status"),
        row.get("alarm_status"),
        row.get("current_status"),
        row.get("status"),
        row.get("raw_status"),
        row.get("ap_raw_status"),
        row.get("module_status"),
        row.get("transceiver_status"),
    ]
    status_text = " ".join(str(value or "") for value in status_values).strip().casefold()
    return any(
        token in status_text
        for token in (
            "no_light",
            "no light",
            "no-light",
            "no_module",
            "no module",
            "no-module",
            "no transceiver",
            "no-transceiver",
            "无光",
            "无光模块",
            "未插光模块",
            "光模块不存在",
        )
    )


def is_current_optical_abnormal_export_row(row: dict[str, object | None]) -> bool:
    return trackside_row_status(row) in OPTICAL_TREATMENT_ISSUE_STATUSES and not is_no_light_optical_row(row)


def is_current_optical_abnormal_row(row: dict[str, object | None]) -> bool:
    return is_current_optical_abnormal_export_row(row)


def build_current_optical_abnormal_sheet(workbook, source_sheet, rows: list[dict[str, object | None]]) -> None:
    from copy import copy

    if CURRENT_OPTICAL_ABNORMAL_SHEET_TITLE in workbook.sheetnames:
        del workbook[CURRENT_OPTICAL_ABNORMAL_SHEET_TITLE]
    source_index = workbook.worksheets.index(source_sheet)
    sheet = workbook.create_sheet(CURRENT_OPTICAL_ABNORMAL_SHEET_TITLE, source_index + 1)
    _copy_worksheet_columns(source_sheet, sheet)
    _copy_worksheet_row(source_sheet, sheet, 1, 1)
    target_row = 2
    for source_row, data in enumerate(rows, start=2):
        if not is_current_optical_abnormal_export_row(data):
            continue
        _copy_worksheet_row(source_sheet, sheet, source_row, target_row)
        target_row += 1
    if target_row == 2:
        cell = sheet.cell(row=2, column=1, value=CURRENT_OPTICAL_ABNORMAL_EMPTY_TEXT)
        header_cell = source_sheet.cell(row=1, column=1)
        cell.font = copy(header_cell.font)
        cell.fill = copy(header_cell.fill)
        cell.border = copy(header_cell.border)
        cell.alignment = copy(header_cell.alignment)
        cell.number_format = header_cell.number_format
        cell.protection = copy(header_cell.protection)
        sheet.row_dimensions[2].height = 22
    sheet.freeze_panes = source_sheet.freeze_panes
    sheet.auto_filter.ref = sheet.dimensions


def _copy_worksheet_columns(source_sheet, target_sheet) -> None:
    from copy import copy

    for key, dimension in source_sheet.column_dimensions.items():
        target = target_sheet.column_dimensions[key]
        target.width = dimension.width
        target.hidden = dimension.hidden
        target.outlineLevel = dimension.outlineLevel
        target.collapsed = dimension.collapsed
        if dimension.style:
            target.style = copy(dimension.style)


def _copy_worksheet_row(source_sheet, target_sheet, source_row: int, target_row: int) -> None:
    from copy import copy

    target_sheet.row_dimensions[target_row].height = source_sheet.row_dimensions[source_row].height
    for source_cell in source_sheet[source_row]:
        target_cell = target_sheet.cell(row=target_row, column=source_cell.column, value=source_cell.value)
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)
        if source_cell.hyperlink:
            target_cell._hyperlink = copy(source_cell.hyperlink)
        if source_cell.comment:
            target_cell.comment = copy(source_cell.comment)


def _append_export_rows_sheet(
    workbook,
    title: str,
    rows: list[dict[str, object | None]],
    columns: tuple[tuple[str, str], ...],
    headers: list[str],
    alignment,
    border,
    header_font,
    header_fill=None,
    fills: dict[str, object] | None = None,
    row_fill_status_getter=None,
    preserve_ap_identity: bool = False,
) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    text_fields = {"ap_mac", "ap_name", "serial_number", "apid", "ap_ip"}
    for row in rows:
        sheet.append([_export_value(field, row, preserve_ap_identity=preserve_ap_identity) for _key, field in columns])
        for index, (_key, field) in enumerate(columns, start=1):
            if field in text_fields:
                sheet.cell(row=sheet.max_row, column=index).number_format = "@"
        status = row_fill_status_getter(row) if callable(row_fill_status_getter) else None
        fill = (fills or {}).get(status)
        if fill is not None:
            for cell in sheet[sheet.max_row]:
                cell.fill = fill
    _format_export_sheet(sheet, alignment, border, header_font, header_fill)
    sheet.auto_filter.ref = sheet.dimensions


def _sort_ap_optical_treatment_rows(rows: list[dict[str, object | None]]) -> list[dict[str, object | None]]:
    return sorted(
        [dict(row) for row in rows],
        key=lambda row: (
            _blank_last_text_sort_key(row.get("site") or row.get("station") or row.get("归属站点")),
            _blank_last_text_sort_key(row.get("device_name") or row.get("switch_name") or row.get("indoor_switch") or row.get("室内交换机")),
            interface_sort_key(row.get("interface_name") or row.get("接口名称")),
        ),
    )


def _blank_last_text_sort_key(value: object) -> tuple[int, str]:
    text = str(value or "").strip()
    if not text or text == "-":
        return (1, "")
    return (0, text.casefold())


def _ap_optical_treatment_row_fill_status(row: dict[str, object | None]) -> str:
    if row.get("treatment_status") == TREATMENT_CLOSED_LABEL:
        return "normal"
    status_text = " ".join(str(row.get(field) or "") for field in ("current_status", "issue_type")).casefold()
    if "预警" in status_text or "notice" in status_text or "warning" in status_text:
        return "warning"
    if any(token in status_text for token in ("无光", "告警", "异常", "alarm", "abnormal", "no_light")):
        return "alarm"
    return "alarm" if row.get("treatment_status") == TREATMENT_OPEN_LABEL else "normal"


def _append_ap_overview_sheet(
    workbook,
    rows: list[dict[str, object | None]],
    alignment,
    border,
    header_font,
    overview_rows: list[dict[str, object | None]] | None = None,
    overview_columns: tuple[tuple[str, str], ...] | None = None,
    overview_headers: list[str] | None = None,
    header_fill=None,
) -> None:
    sheet = workbook.create_sheet("AP上线情况概览")
    sheet.title = "AP\u4e0a\u7ebf\u60c5\u51b5\u6982\u89c8"
    if overview_rows is not None and overview_columns is not None and overview_headers is not None:
        write_ap_online_overview_sheet(sheet, overview_rows, overview_headers)
        return
    write_ap_online_overview_sheet(sheet, [], [key for key, _field in AP_ONLINE_OVERVIEW_COLUMNS])
    return


def _append_switch_optical_summary_sheet(workbook, rows: list[dict[str, object | None]], alignment, border, header_font, header_fill=None) -> None:
    sheet = workbook.create_sheet("\u4ea4\u6362\u673a\u5149\u6a21\u5757\u7edf\u8ba1")
    sheet.append(["\u4ea4\u6362\u673a", "\u5149\u6a21\u5757\u6570\u91cf", "\u672a\u63d2\u5149\u6a21\u5757\u7aef\u53e3\u6570\u91cf", "\u672a\u63d2\u5149\u6a21\u5757\u7aef\u53e3"])
    grouped: dict[str, dict[str, object]] = {}
    for row in rows:
        switch_name = str(row.get("device_name") or "-")
        item = grouped.setdefault(switch_name, {"module_count": 0, "missing_ports": []})
        missing_ports = item["missing_ports"]
        if isinstance(missing_ports, list):
            missing_ports.extend(_normalize_missing_module_ports(row.get("missing_module_ports")))
        if row.get("switch_optical_status") == "no_module" and isinstance(missing_ports, list):
            missing_ports.extend(_normalize_missing_module_ports(row.get("interface_name") or "-"))
        else:
            item["module_count"] = int(item["module_count"]) + 1
    for switch_name in sorted(grouped):
        item = grouped[switch_name]
        missing_ports = _normalize_missing_module_ports(item["missing_ports"])
        sheet.append(
            [
                switch_name,
                item["module_count"],
                len(missing_ports),
                ", ".join(_short_interface_name(port) for port in missing_ports) if missing_ports else "-",
            ]
        )
    _format_export_sheet(sheet, alignment, border, header_font, header_fill)
    sheet.auto_filter.ref = sheet.dimensions


def _normalize_missing_module_ports(value: object) -> list[str]:
    if value in (None, "", "-"):
        return []
    if isinstance(value, (list, tuple, set)):
        return [str(item).strip() for item in value if str(item or "").strip() and str(item or "").strip() != "-"]
    return [part.strip() for part in re.split(r"[,，;；]", str(value)) if part.strip() and part.strip() != "-"]


def _set_switch_optical_summary_widths(workbook) -> None:
    if "\u4ea4\u6362\u673a\u5149\u6a21\u5757\u7edf\u8ba1" not in workbook.sheetnames:
        return
    sheet = workbook["\u4ea4\u6362\u673a\u5149\u6a21\u5757\u7edf\u8ba1"]
    for column, width in {"A": 22, "B": 14, "C": 20, "D": 80}.items():
        sheet.column_dimensions[column].width = width


def _ap_state(row: dict[str, object | None]) -> str:
    text = " ".join(str(row.get(field) or "") for field in ("ap_state", "ap_state_display", "state", "state_display")).strip().casefold()
    if not text:
        return ""
    token = text.split("=", 1)[0].strip()
    if "idle" in text or "offline" in text or "离线" in text:
        return "offline"
    if token in {"r", "r/m", "run"} or any(value in text for value in ("online", "run", "up", "normal", "在线")):
        return "online"
    if any(token in text for token in ("offline", "down", "fault", "离线")):
        return "offline"
    return ""


def _short_interface_name(value: object) -> str:
    return str(value or "-").replace("GigabitEthernet", "GE")
