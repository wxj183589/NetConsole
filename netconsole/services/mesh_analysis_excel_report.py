from __future__ import annotations

from pathlib import Path
from typing import Callable, Iterable

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from netconsole.services.mesh_analysis_report import MeshAnalysisReportModel


EMPTY_PARSE_ISSUES_TEXT = "未发现解析问题"
CancelCallback = Callable[[], bool]
ProgressCallback = Callable[[int, str], None]

EMPTY_PARSE_ISSUES_TEXT = "未发现解析问题"


REPORT_FIELD_LABELS: dict[str, str] = {
    "key": "项目",
    "value": "值",
    "report_name": "报告名称",
    "mr_name": "MR名称",
    "data_source_type": "数据来源",
    "generated_at": "生成时间",
    "sample_time": "采样时间",
    "radio": "射频口",
    "total_peer_count": "Peer总数",
    "active_peer_count": "Active数量",
    "active_peer_mac": "主链路Peer MAC",
    "active_peer": "主链路Peer",
    "active_mr_rssi": "MR侧RSSI",
    "active_peer_rssi": "对端RSSI",
    "standby_peer_count": "备链数量",
    "available_backup_count": "可用备链数量",
    "strong_backup_count": "强备链数量",
    "best_backup_peer_mac": "最佳备链Peer MAC",
    "best_backup_rssi": "最佳备链RSSI",
    "active_tx_busy": "主链路TxBusy",
    "active_rx_busy": "主链路RxBusy",
    "max_tx_busy": "最大TxBusy",
    "max_rx_busy": "最大RxBusy",
    "link_count": "链路计数",
    "active_establish_time": "主链路建立时间",
    "active_duration_time": "主链路持续时间",
    "source_file": "源文件",
    "source_line_number": "源行号",
    "quality_level": "质量等级",
    "quality_score": "质量分",
    "quality_reasons": "质量原因",
    "fping_loss_rate": "fping丢包率",
    "sequence": "序号",
    "peer_mac": "Peer MAC",
    "peer_mac_display": "Peer MAC",
    "peer_ap_name": "对端AP名称",
    "peer_ap_mac": "对端AP MAC",
    "peer_site": "对端站点",
    "peer_radio": "对端射频口",
    "start_time": "开始时间",
    "end_time": "结束时间",
    "duration_seconds": "持续秒数",
    "sample_count": "采样点数",
    "first_mr_rssi": "首个MR侧RSSI",
    "last_mr_rssi": "最后MR侧RSSI",
    "avg_mr_rssi": "平均MR侧RSSI",
    "min_mr_rssi": "最低MR侧RSSI",
    "p10_mr_rssi": "P10 MR侧RSSI",
    "max_mr_rssi": "最高MR侧RSSI",
    "rssi_jitter": "RSSI抖动",
    "avg_peer_rssi": "平均对端RSSI",
    "min_peer_rssi": "最低对端RSSI",
    "avg_tx_busy": "平均TxBusy",
    "avg_rx_busy": "平均RxBusy",
    "available_backup_ratio": "可用备链占比",
    "strong_backup_ratio": "强备链占比",
    "no_backup_seconds": "无可用备链秒数",
    "weak_rssi_seconds": "弱RSSI秒数",
    "busy_seconds": "空口繁忙秒数",
    "link_count_delta_count": "链路计数变化次数",
    "duration_reset_count": "持续时间回退次数",
    "establish_reset_count": "建立时间重置次数",
    "segment_quality_score": "区段质量分",
    "segment_level": "区段等级",
    "segment_problem_tags": "区段问题标签",
    "source_files": "源文件列表",
    "first_seen_time": "首次出现时间",
    "last_seen_time": "最后出现时间",
    "seen_sample_count": "出现采样数",
    "active_sample_count": "Active采样数",
    "standby_sample_count": "Standby采样数",
    "active_segment_count": "Active区段数",
    "switch_in_count": "切入次数",
    "switch_out_count": "切出次数",
    "active_total_seconds": "Active总秒数",
    "active_total_ratio": "Active占比",
    "avg_active_rssi": "Active平均RSSI",
    "min_active_rssi": "Active最低RSSI",
    "p10_active_rssi": "Active P10 RSSI",
    "max_active_rssi": "Active最高RSSI",
    "weak_active_seconds": "弱Active秒数",
    "no_backup_when_active_seconds": "Active期间无备链秒数",
    "link_rebuild_count": "链路重建次数",
    "short_segment_count": "短区段次数",
    "flap_related_count": "乒乓相关次数",
    "peer_quality_score": "Peer质量分",
    "problem_tags": "问题标签",
    "suggestion": "建议",
    "switch_time": "切换时间",
    "from_peer": "原Peer",
    "from_peer_ap_name": "原AP名称",
    "to_peer": "新Peer",
    "to_peer_ap_name": "新AP名称",
    "previous_segment_duration": "原区段持续秒数",
    "new_segment_duration": "新区段持续秒数",
    "from_last_rssi": "切出前RSSI",
    "from_avg_rssi_before_switch": "切换前平均RSSI",
    "to_first_rssi": "切入首个RSSI",
    "to_avg_rssi_after_switch": "切入后平均RSSI",
    "best_backup_peer_before_switch": "切换前最佳备链Peer",
    "best_backup_rssi_before_switch": "切换前最佳备链RSSI",
    "tx_busy_before_switch": "切换前TxBusy",
    "rx_busy_before_switch": "切换前RxBusy",
    "tx_busy_after_switch": "切换后TxBusy",
    "rx_busy_after_switch": "切换后RxBusy",
    "switch_type": "切换类型",
    "severity": "严重级别",
    "diagnosis": "诊断",
    "evidence_id": "证据ID",
    "event_sequence": "事件序号",
    "event_time_start": "事件开始时间",
    "event_time_end": "事件结束时间",
    "event_type": "事件类型",
    "active_rssi_min": "主链路最低RSSI",
    "active_rssi_avg": "主链路平均RSSI",
    "backup_count_min": "最小备链数量",
    "tx_busy_max": "最大TxBusy",
    "rx_busy_max": "最大RxBusy",
    "source_line_number_start": "起始源行号",
    "source_line_number_end": "结束源行号",
    "busy_warning_seconds": "繁忙关注秒数",
    "busy_bad_seconds": "繁忙严重秒数",
    "busy_ratio": "繁忙占比",
    "busy_level": "繁忙等级",
    "event_time": "事件时间",
    "previous_link_cnt": "上一链路计数",
    "current_link_cnt": "当前链路计数",
    "previous_duration_time": "上一持续时间",
    "current_duration_time": "当前持续时间",
    "previous_establish_time": "上一建立时间",
    "current_establish_time": "当前建立时间",
    "rebuild_type": "重建类型",
    "related_sheet": "关联Sheet",
    "related_sequence": "关联序号",
    "related_event_type": "关联事件类型",
    "link_state": "链路状态",
    "mr_rssi": "MR侧RSSI",
    "peer_rssi": "对端RSSI",
    "tx_busy": "TxBusy",
    "rx_busy": "RxBusy",
    "link_cnt": "链路计数",
    "establish_time": "建立时间",
    "duration_time": "持续时间",
    "raw_line": "原始日志行",
    "original_filename": "原始文件名",
    "archived_filename": "归档文件名",
    "file_size": "文件大小",
    "sha256": "SHA256",
    "encoding": "编码",
    "is_gzip": "是否Gzip",
    "first_sample_time": "首个采样时间",
    "last_sample_time": "最后采样时间",
    "lines_read": "读取行数",
    "records_parsed": "解析记录数",
    "records_skipped": "跳过记录数",
    "duplicate_records": "重复记录数",
    "issue_count": "问题数量",
    "parse_status": "解析状态",
    "error_message": "错误信息",
    "issue_sequence": "问题序号",
    "issue_type": "问题类型",
    "message": "问题描述",
    "dimension": "评分维度",
    "weight": "权重",
    "score": "得分",
}

REPORT_FIELD_LABELS.update(
    {
        "key": "项目",
        "value": "值",
        "report_name": "报告名称",
        "mr_name": "MR名称",
        "data_source_type": "数据来源类型",
        "generated_at": "生成时间",
        "sample_time": "采样时间",
        "radio": "射频口",
        "total_peer_count": "Peer总数",
        "active_peer_count": "Active数量",
        "active_peer_mac": "主链路Peer MAC",
        "active_peer": "主链路Peer",
        "active_mr_rssi": "MR侧RSSI",
        "active_peer_rssi": "对端RSSI",
        "standby_peer_count": "备用Peer数量",
        "available_backup_count": "可用备份数量",
        "strong_backup_count": "强备份数量",
        "best_backup_peer_mac": "最佳备份Peer MAC",
        "best_backup_rssi": "最佳备份RSSI",
        "active_tx_busy": "主链路TxBusy",
        "active_rx_busy": "主链路RxBusy",
        "max_tx_busy": "最大TxBusy",
        "max_rx_busy": "最大RxBusy",
        "link_count": "链路记录数",
        "active_establish_time": "Active建立时间",
        "active_duration_time": "Active持续时间",
        "source_file": "源文件",
        "source_line_number": "源行号",
        "quality_level": "质量等级",
        "quality_score": "质量评分",
        "quality_reasons": "质量原因",
        "fping_loss_rate": "fping丢包率",
        "sequence": "序号",
        "peer_mac": "Peer MAC",
        "peer_mac_display": "Peer MAC",
        "peer_ap_name": "对端AP名称",
        "peer_ap_mac": "对端AP MAC",
        "peer_site": "对端站点",
        "peer_radio": "对端Radio",
        "start_time": "开始时间",
        "end_time": "结束时间",
        "duration_seconds": "持续秒数",
        "sample_count": "采样点数",
        "first_mr_rssi": "首个MR RSSI",
        "last_mr_rssi": "最后MR RSSI",
        "avg_mr_rssi": "平均MR RSSI",
        "min_mr_rssi": "最小MR RSSI",
        "p10_mr_rssi": "P10 MR RSSI",
        "max_mr_rssi": "最大MR RSSI",
        "rssi_jitter": "RSSI抖动",
        "avg_peer_rssi": "平均对端RSSI",
        "min_peer_rssi": "最小对端RSSI",
        "avg_tx_busy": "平均TxBusy",
        "avg_rx_busy": "平均RxBusy",
        "p90_tx_busy": "P90 TxBusy",
        "p90_rx_busy": "P90 RxBusy",
        "available_backup_ratio": "可用备份比例",
        "strong_backup_ratio": "强备份比例",
        "no_backup_seconds": "无备份秒数",
        "weak_rssi_seconds": "弱RSSI秒数",
        "busy_seconds": "繁忙秒数",
        "link_count_delta_count": "LinkCnt增量次数",
        "duration_reset_count": "DurationTime回退次数",
        "establish_reset_count": "EstablishTime重置次数",
        "segment_quality_score": "区段质量评分",
        "segment_level": "区段等级",
        "segment_problem_tags": "区段问题标签",
        "source_files": "源文件",
        "first_seen_time": "首次出现时间",
        "last_seen_time": "最后出现时间",
        "seen_sample_count": "出现采样数",
        "active_sample_count": "Active采样数",
        "standby_sample_count": "备用采样数",
        "active_segment_count": "Active区段数",
        "switch_in_count": "切入次数",
        "switch_out_count": "切出次数",
        "active_total_seconds": "Active总秒数",
        "active_total_ratio": "Active占比",
        "avg_active_rssi": "Active平均RSSI",
        "min_active_rssi": "Active最小RSSI",
        "p10_active_rssi": "Active P10 RSSI",
        "max_active_rssi": "Active最大RSSI",
        "weak_active_seconds": "弱Active秒数",
        "no_backup_when_active_seconds": "Active时无备份秒数",
        "link_rebuild_count": "链路重建次数",
        "short_segment_count": "短区段次数",
        "flap_related_count": "乒乓相关次数",
        "peer_quality_score": "Peer质量评分",
        "problem_tags": "问题标签",
        "suggestion": "建议",
        "switch_time": "切换时间",
        "from_peer": "切出Peer",
        "from_peer_ap_name": "切出AP名称",
        "to_peer": "切入Peer",
        "to_peer_ap_name": "切入AP名称",
        "previous_segment_duration": "上一段持续秒数",
        "new_segment_duration": "新段持续秒数",
        "from_last_rssi": "切出前最后RSSI",
        "from_avg_rssi_before_switch": "切出前平均RSSI",
        "to_first_rssi": "切入后首个RSSI",
        "to_avg_rssi_after_switch": "切入后平均RSSI",
        "best_backup_peer_before_switch": "切换前最佳备份Peer",
        "best_backup_rssi_before_switch": "切换前最佳备份RSSI",
        "tx_busy_before_switch": "切换前TxBusy",
        "rx_busy_before_switch": "切换前RxBusy",
        "tx_busy_after_switch": "切换后TxBusy",
        "rx_busy_after_switch": "切换后RxBusy",
        "switch_type": "切换类型",
        "severity": "严重程度",
        "diagnosis": "诊断",
        "evidence_id": "证据ID",
        "event_sequence": "事件序号",
        "event_time_start": "事件开始时间",
        "event_time_end": "事件结束时间",
        "event_type": "事件类型",
        "active_rssi_min": "Active最小RSSI",
        "active_rssi_avg": "Active平均RSSI",
        "backup_count_min": "最小备份数量",
        "tx_busy_max": "最大TxBusy",
        "rx_busy_max": "最大RxBusy",
        "source_line_number_start": "起始源行号",
        "source_line_number_end": "结束源行号",
        "busy_warning_seconds": "繁忙关注秒数",
        "busy_bad_seconds": "繁忙严重秒数",
        "busy_ratio": "繁忙占比",
        "busy_level": "繁忙等级",
        "event_time": "事件时间",
        "previous_link_cnt": "前一LinkCnt",
        "current_link_cnt": "当前LinkCnt",
        "previous_duration_time": "前一DurationTime",
        "current_duration_time": "当前DurationTime",
        "previous_establish_time": "前一EstablishTime",
        "current_establish_time": "当前EstablishTime",
        "rebuild_type": "重建类型",
        "related_sheet": "关联Sheet",
        "related_sequence": "关联序号",
        "related_event_type": "关联事件类型",
        "link_state": "链路状态",
        "mr_rssi": "MR RSSI",
        "peer_rssi": "对端RSSI",
        "tx_busy": "TxBusy",
        "rx_busy": "RxBusy",
        "link_cnt": "LinkCnt",
        "establish_time": "EstablishTime",
        "duration_time": "DurationTime",
        "raw_line": "原始日志行",
        "original_filename": "原始文件名",
        "archived_filename": "归档文件名",
        "file_size": "文件大小",
        "sha256": "SHA256",
        "encoding": "编码",
        "is_gzip": "是否Gzip",
        "first_sample_time": "首个采样时间",
        "last_sample_time": "最后采样时间",
        "lines_read": "读取行数",
        "records_parsed": "解析记录数",
        "records_skipped": "跳过记录数",
        "duplicate_records": "重复记录数",
        "issue_count": "问题数量",
        "parse_status": "解析状态",
        "error_message": "错误信息",
        "issue_sequence": "问题序号",
        "issue_type": "问题类型",
        "message": "消息",
        "dimension": "评分维度",
        "weight": "权重",
        "score": "得分",
    }
)

VALUE_TRANSLATIONS: dict[str, dict[str, str]] = {
    "quality_level": {"EXCELLENT": "优秀", "GOOD": "良好", "WARNING": "关注", "BAD": "异常", "CRITICAL": "严重"},
    "segment_level": {"优秀": "优秀", "良好": "良好", "关注": "关注", "异常": "异常", "严重": "严重"},
    "busy_level": {"EXCELLENT": "优秀", "GOOD": "良好", "WARNING": "关注", "BAD": "异常", "CRITICAL": "严重"},
    "severity": {"INFO": "信息", "WARNING": "关注", "BAD": "异常", "CRITICAL": "严重", "GOOD": "良好"},
    "switch_type": {
        "NORMAL_SWITCH": "正常切换",
        "LATE_SWITCH": "切换滞后",
        "WEAK_TARGET_SWITCH": "切入质量差",
        "FLAP_SWITCH": "乒乓切换",
        "SHORT_SEGMENT_SWITCH": "短时切换",
        "NO_ACTIVE_GAP_SWITCH": "切换时Active空洞",
        "UNKNOWN_SWITCH": "未知切换",
    },
    "event_type": {
        "NO_ACTIVE": "无Active",
        "MULTI_ACTIVE": "多Active",
        "WEAK_ACTIVE": "弱主链路",
        "BAD_ACTIVE": "严重弱主链路",
        "NO_BACKUP": "无可用备链",
        "HIGH_BUSY": "空口严重繁忙",
        "BUSY_WARNING": "空口繁忙关注",
        "LINK_REBUILD": "链路重建",
        "COUNTER_RESET": "计数器回退",
        "PARSE_GAP": "解析/时间线间隔",
    },
    "rebuild_type": {
        "LINKCNT_INCREASE": "链路计数增加",
        "DURATION_RESET": "持续时间回退",
        "ESTABLISH_RESET": "建立时间重置",
        "STATE_FLAP": "链路状态抖动",
        "UNKNOWN_REBUILD": "未知重建",
    },
    "link_state": {"ACTIVE": "主链路", "Active": "主链路", "STANDBY": "备链", "Standby": "备链", "Standy": "备链", "UNKNOWN": "未知"},
    "data_source_type": {"MR_RAW_MESH_LOG": "MR原始MESH日志", "VEHICLE_MR_REALTIME_OFFLINE": "车载MR实时收集离线数据"},
}

VALUE_TRANSLATIONS.update(
    {
        "quality_level": {
            "EXCELLENT": "优秀",
            "GOOD": "良好",
            "WARNING": "关注",
            "BAD": "异常",
            "CRITICAL": "严重",
        },
        "segment_level": {
            "EXCELLENT": "优秀",
            "GOOD": "良好",
            "WARNING": "关注",
            "BAD": "异常",
            "CRITICAL": "严重",
        },
        "busy_level": {
            "GOOD": "正常",
            "WARNING": "关注",
            "BAD": "严重",
            "CRITICAL": "严重",
        },
        "severity": {
            "INFO": "提示",
            "WARNING": "关注",
            "BAD": "异常",
            "CRITICAL": "严重",
        },
        "switch_type": {
            "NORMAL_SWITCH": "正常切换",
            "LATE_SWITCH": "切换滞后",
            "WEAK_TARGET_SWITCH": "切入质量差",
            "FLAP_SWITCH": "乒乓切换",
            "SHORT_SEGMENT_SWITCH": "短时切换",
            "NO_ACTIVE_GAP_SWITCH": "切换无Active空洞",
            "UNKNOWN_SWITCH": "数据不足",
        },
        "event_type": {
            "NO_ACTIVE": "无Active",
            "MULTI_ACTIVE": "多Active",
            "WEAK_ACTIVE": "弱Active",
            "BAD_ACTIVE": "严重弱Active",
            "NO_BACKUP": "无可用备份",
            "HIGH_BUSY": "高繁忙",
            "BUSY_WARNING": "繁忙关注",
            "LINK_REBUILD": "链路重建",
            "COUNTER_RESET": "计数器回退",
            "PARSE_GAP": "解析跳点",
        },
        "rebuild_type": {
            "LINKCNT_INCREASE": "LinkCnt增加",
            "DURATION_RESET": "持续时间回退",
            "ESTABLISH_RESET": "EstablishTime重置",
            "STATE_FLAP": "链路状态抖动",
            "UNKNOWN_REBUILD": "未知重建",
        },
        "link_state": {
            "ACTIVE": "主链路",
            "BACKUP": "备份链路",
            "STANDBY": "备用链路",
            "DOWN": "Down",
        },
        "data_source_type": {
            "MR_RAW_MESH_LOG": "MR原始MESH日志",
            "VEHICLE_MR_REALTIME_OFFLINE": "车载MR实时收集离线数据",
        },
    }
)

SHEET_DEFINITIONS: tuple[tuple[str, tuple[str, ...], str], ...] = (
    ("报告总览", ("key", "value"), "overview"),
    ("质量评分", ("dimension", "weight", "score", "diagnosis"), "score_rows"),
    ("原始文件清单", ("original_filename", "archived_filename", "file_size", "sha256", "encoding", "is_gzip", "first_sample_time", "last_sample_time", "lines_read", "records_parsed", "records_skipped", "duplicate_records", "issue_count", "parse_status", "error_message"), "source_files"),
    ("采样点质量统计", ("sample_time", "radio", "total_peer_count", "active_peer_count", "active_peer_mac", "active_mr_rssi", "active_peer_rssi", "standby_peer_count", "available_backup_count", "strong_backup_count", "best_backup_peer_mac", "best_backup_rssi", "active_tx_busy", "active_rx_busy", "max_tx_busy", "max_rx_busy", "link_count", "active_establish_time", "active_duration_time", "source_file", "source_line_number", "quality_level", "quality_score", "quality_reasons", "fping_loss_rate"), "sample_quality"),
    ("Active主链路区段", ("sequence", "radio", "active_peer_mac", "peer_ap_name", "peer_ap_mac", "peer_site", "peer_radio", "start_time", "end_time", "duration_seconds", "sample_count", "first_mr_rssi", "last_mr_rssi", "avg_mr_rssi", "min_mr_rssi", "p10_mr_rssi", "max_mr_rssi", "rssi_jitter", "avg_peer_rssi", "min_peer_rssi", "avg_tx_busy", "max_tx_busy", "avg_rx_busy", "max_rx_busy", "available_backup_ratio", "strong_backup_ratio", "no_backup_seconds", "weak_rssi_seconds", "busy_seconds", "link_count_delta_count", "duration_reset_count", "establish_reset_count", "segment_quality_score", "segment_level", "segment_problem_tags", "source_files"), "active_segments"),
    ("Peer质量排名", ("radio", "peer_mac", "peer_ap_name", "peer_ap_mac", "peer_site", "peer_radio", "first_seen_time", "last_seen_time", "seen_sample_count", "active_sample_count", "standby_sample_count", "active_segment_count", "switch_in_count", "switch_out_count", "active_total_seconds", "active_total_ratio", "avg_active_rssi", "min_active_rssi", "p10_active_rssi", "max_active_rssi", "rssi_jitter", "weak_active_seconds", "no_backup_when_active_seconds", "avg_tx_busy", "max_tx_busy", "avg_rx_busy", "max_rx_busy", "link_rebuild_count", "short_segment_count", "flap_related_count", "peer_quality_score", "problem_tags", "suggestion"), "peer_ranking"),
    ("切换事件分析", ("sequence", "radio", "switch_time", "from_peer", "from_peer_ap_name", "to_peer", "to_peer_ap_name", "previous_segment_duration", "new_segment_duration", "from_last_rssi", "from_avg_rssi_before_switch", "to_first_rssi", "to_avg_rssi_after_switch", "best_backup_peer_before_switch", "best_backup_rssi_before_switch", "tx_busy_before_switch", "rx_busy_before_switch", "tx_busy_after_switch", "rx_busy_after_switch", "switch_type", "severity", "diagnosis", "suggestion", "evidence_id"), "switch_events"),
    ("异常事件分析", ("event_sequence", "event_time_start", "event_time_end", "duration_seconds", "radio", "event_type", "severity", "active_peer", "peer_ap_name", "active_rssi_min", "active_rssi_avg", "backup_count_min", "tx_busy_max", "rx_busy_max", "source_file", "source_line_number_start", "source_line_number_end", "diagnosis", "suggestion", "evidence_id"), "anomaly_events"),
    ("无备份链路风险", ("event_sequence", "event_time_start", "event_time_end", "duration_seconds", "radio", "active_peer", "backup_count_min", "diagnosis", "suggestion", "evidence_id"), "no_backup_risks"),
    ("空口繁忙度分析", ("radio", "peer_mac", "peer_ap_name", "sample_count", "avg_tx_busy", "max_tx_busy", "p90_tx_busy", "avg_rx_busy", "max_rx_busy", "p90_rx_busy", "busy_warning_seconds", "busy_bad_seconds", "busy_ratio", "busy_level", "diagnosis"), "busy_analysis"),
    ("链路重建计数异常", ("sequence", "event_time", "radio", "peer_mac", "peer_ap_name", "previous_link_cnt", "current_link_cnt", "previous_duration_time", "current_duration_time", "previous_establish_time", "current_establish_time", "rebuild_type", "severity", "diagnosis", "source_file", "source_line_number", "raw_line"), "link_rebuild_events"),
    ("原始证据片段", ("evidence_id", "related_sheet", "related_sequence", "related_event_type", "radio", "sample_time", "source_file", "source_line_number", "link_state", "peer_mac", "peer_ap_name", "mr_rssi", "peer_rssi", "tx_busy", "rx_busy", "link_cnt", "establish_time", "duration_time", "raw_line"), "raw_evidence"),
    ("解析问题", ("issue_sequence", "source_file", "source_line_number", "issue_type", "severity", "message", "raw_line"), "parse_issues"),
)

ALL_LINK_COLUMNS: tuple[str, ...] = (
    "sample_time",
    "radio",
    "link_state",
    "peer_mac_display",
    "peer_ap_name",
    "peer_ap_mac",
    "peer_site",
    "peer_radio",
    "establish_time",
    "duration_time",
    "link_cnt",
    "mr_rssi",
    "peer_rssi",
    "tx_busy",
    "rx_busy",
    "source_file",
    "source_line_number",
    "raw_line",
)

FIXED_WIDTH_FIELDS = {"raw_line", "diagnosis", "suggestion", "quality_reasons", "segment_problem_tags", "problem_tags", "source_files", "source_file", "error_message", "message", "sha256"}
FIXED_WIDTH_BY_FIELD = {
    "raw_line": 60,
    "diagnosis": 50,
    "suggestion": 50,
    "message": 50,
    "source_file": 50,
    "source_files": 50,
    "sha256": 66,
    "sample_time": 24,
    "switch_time": 24,
    "event_time": 24,
    "event_time_start": 24,
    "event_time_end": 24,
    "start_time": 24,
    "end_time": 24,
    "first_seen_time": 24,
    "last_seen_time": 24,
    "first_sample_time": 24,
    "last_sample_time": 24,
    "active_establish_time": 24,
    "establish_time": 24,
    "previous_establish_time": 24,
    "current_establish_time": 24,
    "radio": 10,
    "sequence": 10,
    "event_sequence": 10,
    "mr_rssi": 12,
    "peer_rssi": 12,
    "active_mr_rssi": 14,
    "active_peer_rssi": 14,
    "tx_busy": 12,
    "rx_busy": 12,
    "active_tx_busy": 14,
    "active_rx_busy": 14,
    "avg_tx_busy": 14,
    "avg_rx_busy": 14,
}
MAC_FIELDS = {"active_peer_mac", "peer_mac", "peer_mac_display", "peer_ap_mac", "from_peer", "to_peer", "best_backup_peer_mac", "best_backup_peer_before_switch", "active_peer"}
LARGE_SHEET_ATTRS = {"sample_quality", "raw_evidence", "all_link_details", "active_segments", "peer_ranking", "anomaly_events", "switch_events"}
SHEET_DEFINITIONS = tuple(
    (sheet_name, fields, attr_name)
    for sheet_name, fields, attr_name in (
        ("报告总览", SHEET_DEFINITIONS[0][1], "overview"),
        ("质量评分", SHEET_DEFINITIONS[1][1], "score_rows"),
        ("原始文件清单", SHEET_DEFINITIONS[2][1], "source_files"),
        ("采样点质量统计", SHEET_DEFINITIONS[3][1], "sample_quality"),
        ("Active 主链路区段", SHEET_DEFINITIONS[4][1], "active_segments"),
        ("Peer质量排名", SHEET_DEFINITIONS[5][1], "peer_ranking"),
        ("切换事件分析", SHEET_DEFINITIONS[6][1], "switch_events"),
        ("异常事件分析", SHEET_DEFINITIONS[7][1], "anomaly_events"),
        ("无备份链路风险", SHEET_DEFINITIONS[8][1], "no_backup_risks"),
        ("空口繁忙度分析", SHEET_DEFINITIONS[9][1], "busy_analysis"),
        ("链路重建计数异常", SHEET_DEFINITIONS[10][1], "link_rebuild_events"),
        ("原始证据片段", SHEET_DEFINITIONS[11][1], "raw_evidence"),
        ("解析问题", SHEET_DEFINITIONS[12][1], "parse_issues"),
    )
)

STAGE_BY_ATTR = {
    "overview": "excel_overview",
    "sample_quality": "excel_sample_quality",
    "active_segments": "excel_active_segments",
    "peer_ranking": "excel_peer_ranking",
    "raw_evidence": "excel_raw_evidence",
}


class MeshAnalysisExcelReportExporter:
    def export(
        self,
        model: MeshAnalysisReportModel,
        path: Path,
        progress: ProgressCallback | None = None,
        should_cancel: CancelCallback | None = None,
    ) -> Path:
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        progress = progress or (lambda _value, _stage: None)
        should_cancel = should_cancel or (lambda: False)
        workbook = Workbook(write_only=True)
        workbook._mesh_report_options = model.options
        total_sheets = len(SHEET_DEFINITIONS) + (1 if getattr(model.options, "include_all_link_details", False) else 0)
        written = 0
        for sheet_name, fields, attr_name in SHEET_DEFINITIONS:
            _raise_if_cancelled(should_cancel)
            progress(90 + int(written / max(total_sheets, 1) * 8), STAGE_BY_ATTR.get(attr_name, f"excel_{attr_name}"))
            rows = self._rows_for(model, attr_name)
            self._write_sheet(workbook, sheet_name, fields, rows, attr_name, should_cancel, progress)
            written += 1
        if getattr(model.options, "include_all_link_details", False):
            _raise_if_cancelled(should_cancel)
            progress(98, "excel_all_link_details")
            self._write_sheet(workbook, "全量链路明细", ALL_LINK_COLUMNS, getattr(model, "all_link_details", []), "all_link_details", should_cancel, progress)
        progress(99, "excel_save")
        workbook.save(path)
        return path

    def _write_sheet(
        self,
        workbook: Workbook,
        sheet_name: str,
        fields: Iterable[str],
        rows: list[dict[str, object]] | list[tuple[object, ...]],
        attr_name: str,
        should_cancel: CancelCallback,
        progress: ProgressCallback,
    ) -> None:
        field_list = list(fields)
        sheet = workbook.create_sheet(sheet_name)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(field_list))}{max(len(rows) + 1, 1)}"
        scan_limit = int(getattr(getattr(workbook, "_mesh_report_options", None), "autofit_scan_limit", 2000) or 2000)
        width_tracker = _WidthTracker(field_list, scan_limit=scan_limit if attr_name in LARGE_SHEET_ATTRS else 100000)
        width_tracker.feed([REPORT_FIELD_LABELS.get(field, field) for field in field_list])
        if attr_name == "parse_issues" and not rows:
            rows = [{"issue_sequence": 1, "issue_type": "N/A", "severity": "INFO", "message": EMPTY_PARSE_ISSUES_TEXT}]
        for index, row in enumerate(rows[: width_tracker.scan_limit], 1):
            width_tracker.feed(self._row_values(field_list, row, attr_name, index))
        width_tracker.apply(sheet)
        sheet.append([_header_cell(sheet, REPORT_FIELD_LABELS.get(field, field)) for field in field_list])
        total_rows = len(rows)
        for index, row in enumerate(rows, 1):
            if index % 1000 == 0:
                _raise_if_cancelled(should_cancel)
                progress(95, f"excel_sheet_rows:{sheet_name}:{index}:{total_rows}")
            values = self._row_values(field_list, row, attr_name, index)
            sheet.append(values)

    def _row_values(self, fields: list[str], row: dict[str, object] | tuple[object, ...], attr_name: str, index: int) -> list[object]:
        if isinstance(row, tuple):
            return [translate_report_value(field, value) for field, value in zip(fields, row)]
        materialized = dict(row)
        if attr_name == "parse_issues":
            materialized.setdefault("issue_sequence", index)
            if "source_line_number" not in materialized and "line_number" in materialized:
                materialized["source_line_number"] = materialized.get("line_number")
        if attr_name == "all_link_details":
            materialized.setdefault("duration_time", materialized.get("duration_seconds"))
            materialized.setdefault("link_cnt", materialized.get("link_count"))
            materialized.setdefault("source_file", materialized.get("archived_filename"))
        return [translate_report_value(field, materialized.get(field)) for field in fields]

    def _rows_for(self, model: MeshAnalysisReportModel, attr_name: str) -> list[dict[str, object]] | list[tuple[object, ...]]:
        if attr_name == "overview":
            return [{"key": REPORT_FIELD_LABELS.get(str(key), str(key)), "value": translate_report_value(str(key), value)} for key, value in model.overview.items()]
        return list(getattr(model, attr_name, []))


def translate_report_value(field_name: str, value: object) -> object:
    if value is None or value == "":
        return "N/A" if field_name == "fping_loss_rate" else ""
    if field_name == "fping_loss_rate":
        return value if value not in (None, "") else "N/A"
    text = str(value)
    if field_name == "related_event_type":
        for mapping_name in ("switch_type", "event_type", "rebuild_type"):
            translated = VALUE_TRANSLATIONS.get(mapping_name, {}).get(text)
            if translated:
                return translated
    return VALUE_TRANSLATIONS.get(field_name, {}).get(text, value)


def apply_fast_report_autofit(sheet, headers: list[str], sample_rows: list[list[object]], max_scan_rows: int = 1000) -> None:
    tracker = _WidthTracker(headers, scan_limit=max_scan_rows)
    tracker.feed(headers)
    for row in sample_rows[:max_scan_rows]:
        tracker.feed(row)
    tracker.apply(sheet)


def _header_cell(sheet, value: str) -> WriteOnlyCell:
    cell = WriteOnlyCell(sheet, value=value)
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    return cell


class _WidthTracker:
    def __init__(self, fields: list[str], scan_limit: int) -> None:
        self.fields = fields
        self.scan_limit = scan_limit
        self.scanned = 0
        self.widths = [10 for _field in fields]

    def feed(self, values: list[object]) -> None:
        if self.scanned >= self.scan_limit:
            return
        self.scanned += 1
        for index, value in enumerate(values[: len(self.fields)]):
            field = self.fields[index]
            if field in FIXED_WIDTH_BY_FIELD or field in MAC_FIELDS:
                self.widths[index] = max(self.widths[index], FIXED_WIDTH_BY_FIELD.get(field, 20))
                continue
            if field in FIXED_WIDTH_FIELDS:
                self.widths[index] = max(self.widths[index], 40)
                continue
            self.widths[index] = max(self.widths[index], min(_display_width(value) + 2, 80))

    def apply(self, sheet) -> None:
        for index, width in enumerate(self.widths, 1):
            field = self.fields[index - 1]
            if field in FIXED_WIDTH_BY_FIELD or field in MAC_FIELDS:
                width = FIXED_WIDTH_BY_FIELD.get(field, 20)
            elif field in FIXED_WIDTH_FIELDS:
                width = 42
            sheet.column_dimensions[get_column_letter(index)].width = min(max(width, 10), 80)


def _display_width(value: object) -> int:
    text = str(value or "")
    width = 0
    for char in text[:120]:
        width += 2 if ord(char) > 127 else 1
    return width


def _raise_if_cancelled(should_cancel: CancelCallback) -> None:
    if should_cancel():
        raise RuntimeError("cancelled")
