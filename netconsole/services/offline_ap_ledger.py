from __future__ import annotations

from datetime import datetime
import json
from pathlib import Path
import re

from netconsole.services.ap_online_overview import is_fit_ap_online


OFFLINE_AP_STATUS_TEXT = "离线"
HISTORICAL_DATA_TEXT = "历史数据"
NO_HISTORY_DATA_TEXT = "none"
UNASSIGNED_SITE_TEXT = "未归属"
NO_HISTORY_LLDP_TEXT = "离线AP无历史LLDP记录"

OFFLINE_AP_LEDGER_COLUMNS = (
    ("ac.ap_name", "ap_name"),
    ("ac.ap_mac", "ap_mac"),
    ("field.status", "ap_status"),
    ("ac.station", "site"),
    ("ac.historical_switch", "historical_switch_name"),
    ("ac.historical_interface", "historical_switch_interface"),
)

OFFLINE_AP_STATS_COLUMNS = (
    ("ac.ap_total", "total_aps"),
    ("ac.online", "online_aps"),
    ("ac.offline", "offline_aps"),
    ("ac.offline_rate", "offline_rate"),
    ("ac.offline_with_lldp", "offline_with_lldp"),
    ("ac.offline_without_lldp", "offline_without_lldp"),
    ("ac.offline_locatable", "offline_locatable"),
    ("ac.offline_unlocatable", "offline_unlocatable"),
)

OFFLINE_AP_HEADER_LABELS = {
    "ac.ap_name": "AP名称",
    "ac.ap_mac": "AP_MAC",
    "field.status": "AP状态",
    "ac.station": "归属站点",
    "ac.historical_switch": "历史邻居交换机",
    "ac.historical_interface": "历史邻居接口",
    "ac.ap_total": "AP总数",
    "ac.online": "在线AP数",
    "ac.offline": "离线AP数",
    "ac.offline_rate": "离线率",
    "ac.offline_with_lldp": "有历史LLDP记录",
    "ac.offline_without_lldp": "无历史LLDP记录",
    "ac.offline_locatable": "可定位交换机接口",
    "ac.offline_unlocatable": "无法定位交换机接口",
}


def offline_ap_headers(columns: tuple[tuple[str, str], ...]) -> list[str]:
    return [OFFLINE_AP_HEADER_LABELS.get(key, key) for key, _field in columns]


def build_device_lookup_by_name(devices: list[object]) -> dict[str, dict[str, object | None]]:
    lookup: dict[str, dict[str, object | None]] = {}
    for device in devices or []:
        payload = {
            "device_uuid": getattr(device, "device_uuid", None),
            "name": getattr(device, "name", None),
            "sysname": getattr(device, "sysname", None),
            "station": getattr(device, "station", None),
        }
        for value in (payload["name"], payload["sysname"]):
            key = _normalize_name(value)
            if key:
                lookup[key] = payload
    return lookup


def build_offline_ap_ledger(
    *,
    fit_ap_resources: list[dict[str, object | None]],
    latest_lldp_by_ap: dict[str, dict[str, object | None]] | None = None,
    device_lookup_by_name: dict[str, dict[str, object | None]] | None = None,
    **_ignored: object,
) -> tuple[dict[str, object | None], list[dict[str, object | None]]]:
    resources = list(fit_ap_resources or [])
    lldp_index = latest_lldp_by_ap or {}
    device_lookup = device_lookup_by_name or {}

    ledger: list[dict[str, object | None]] = []
    for resource in resources:
        if not is_fit_ap_offline(resource):
            continue
        lldp = lldp_index.get(_ap_key(resource), {})
        switch_name = _historical_switch_name(lldp)
        switch_device = device_lookup.get(_normalize_name(switch_name), {}) if switch_name else {}
        switch_interface = _historical_switch_interface(lldp)
        has_lldp = bool(lldp)
        ledger.append(
            {
                "ap_uuid": resource.get("ap_uuid"),
                "ac_device_uuid": resource.get("ac_device_uuid"),
                "ap_name": resource.get("ap_name"),
                "ap_mac": _ap_mac(resource, lldp),
                "ap_ip": resource.get("ap_ip"),
                "ap_status": "Idle",
                "site": switch_device.get("station") or UNASSIGNED_SITE_TEXT,
                "device_uuid": switch_device.get("device_uuid"),
                "historical_switch_name": switch_name,
                "historical_switch_interface": switch_interface,
                "last_lldp_at": lldp.get("collected_at"),
                "is_ap_offline": True,
                "data_source": HISTORICAL_DATA_TEXT if has_lldp else NO_HISTORY_DATA_TEXT,
                "lldp_data_source": HISTORICAL_DATA_TEXT if has_lldp else "none",
                "optical_data_source": "none",
                "historical_interface_occupied": False,
                "offline_reason": "ap_idle" if has_lldp else "ap_idle,no_history_lldp",
                "offline_remark": "" if has_lldp else NO_HISTORY_LLDP_TEXT,
                "_history_lldp": lldp,
                "_device_match": switch_device,
            }
        )
    return build_offline_ap_stats(resources, ledger), ledger


def build_latest_ap_history_indexes(
    repository,
    resources: list[dict[str, object | None]],
) -> tuple[dict[str, dict[str, object | None]], dict[str, dict[str, object | None]]]:
    lldp: dict[str, dict[str, object | None]] = {}
    for resource in resources:
        if not is_fit_ap_offline(resource):
            continue
        key = _ap_key(resource)
        ap_uuid = str(resource.get("ap_uuid") or "")
        if not key or not ap_uuid:
            continue
        rows = repository.list_fit_ap_lldp_history_by_ap(ap_uuid, limit=1)
        if rows:
            lldp[key] = rows[0]
    return lldp, {}


def build_offline_ap_stats(
    fit_ap_resources: list[dict[str, object | None]],
    ledger_rows: list[dict[str, object | None]],
) -> dict[str, object | None]:
    total = len(fit_ap_resources or [])
    offline = len(ledger_rows or [])
    online = max(total - offline, 0)
    with_lldp = sum(1 for row in ledger_rows if row.get("last_lldp_at"))
    locatable = sum(1 for row in ledger_rows if row.get("historical_switch_name") and row.get("historical_switch_interface"))
    return {
        "total_aps": total,
        "online_aps": online,
        "offline_aps": offline,
        "offline_rate": f"{offline / total:.1%}" if total else "0.0%",
        "offline_with_lldp": with_lldp,
        "offline_without_lldp": offline - with_lldp,
        "offline_locatable": locatable,
        "offline_unlocatable": offline - locatable,
    }


def append_offline_fit_ap_optical_rows(
    current_rows: list[dict[str, object | None]],
    ledger_rows: list[dict[str, object | None]],
) -> list[dict[str, object | None]]:
    by_key = {_ap_key(row): dict(row) for row in current_rows}
    for offline in ledger_rows:
        key = _ap_key(offline)
        base = by_key.get(key, {})
        by_key[key] = {
            **base,
            "ac_device_uuid": offline.get("ac_device_uuid") or base.get("ac_device_uuid"),
            "ap_uuid": offline.get("ap_uuid") or base.get("ap_uuid"),
            "ap_name": offline.get("ap_name") or base.get("ap_name"),
            "ap_mac": offline.get("ap_mac") or base.get("ap_mac"),
            "ap_ip": offline.get("ap_ip") or base.get("ap_ip"),
            "site": offline.get("site") or base.get("site"),
            "neighbor_device_name": offline.get("historical_switch_name") or base.get("neighbor_device_name"),
            "neighbor_interface": offline.get("historical_switch_interface") or base.get("neighbor_interface"),
            "optical_alarm_status": OFFLINE_AP_STATUS_TEXT,
            "ap_optical_status": "offline",
            "is_ap_offline": True,
            "data_source": offline.get("data_source"),
            "optical_data_source": offline.get("optical_data_source"),
            "lldp_data_source": offline.get("lldp_data_source"),
            "offline_remark": offline.get("offline_remark"),
            "updated_at": offline.get("last_lldp_at"),
        }
    return list(by_key.values())


def write_offline_ap_stats_sheet(sheet, stats: dict[str, object | None], headers: list[str]) -> None:
    sheet.append(headers)
    sheet.append([_display_value(stats.get(field)) for _key, field in OFFLINE_AP_STATS_COLUMNS])
    _format_sheet(sheet)


def write_offline_ap_ledger_sheet(sheet, rows: list[dict[str, object | None]], headers: list[str]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append([_display_value(row.get(field)) for _key, field in OFFLINE_AP_LEDGER_COLUMNS])
    _format_sheet(sheet)


def export_offline_ap_ledger_xlsx(
    path: Path,
    stats: dict[str, object | None],
    rows: list[dict[str, object | None]],
    stats_headers: list[str],
    ledger_headers: list[str],
) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    stats_sheet = workbook.active
    stats_sheet.title = "AP离线情况"
    write_offline_ap_stats_sheet(stats_sheet, stats, stats_headers)
    ledger_sheet = workbook.create_sheet("离线AP台账")
    write_offline_ap_ledger_sheet(ledger_sheet, rows, ledger_headers)
    workbook.save(path)


def load_offline_ap_cache(path: Path, ac_device_uuid: str | None = None) -> dict[str, object | None] | None:
    if not path.exists():
        return None
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    if ac_device_uuid and str(payload.get("ac_device_uuid") or "") != str(ac_device_uuid):
        return None
    stats = payload.get("stats")
    rows = payload.get("ledger_rows")
    if not isinstance(stats, dict) or not isinstance(rows, list):
        return None
    return payload


def save_offline_ap_cache(
    path: Path,
    *,
    ac_device_uuid: str,
    stats: dict[str, object | None],
    ledger_rows: list[dict[str, object | None]],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "schema": 2,
        "ac_device_uuid": ac_device_uuid,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "stats": stats,
        "ledger_rows": [_cache_safe_row(row) for row in ledger_rows],
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def is_fit_ap_offline(row: dict[str, object | None]) -> bool:
    for field in ("state", "state_raw", "state_display"):
        token = _state_token(row.get(field))
        if token in {"I", "IDLE"}:
            return True
    return False


def _cache_safe_row(row: dict[str, object | None]) -> dict[str, object | None]:
    return {key: value for key, value in row.items() if not str(key).startswith("_")}


def _state_token(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    text = text.split("=", 1)[0].strip()
    return text.upper()


def _ap_key(row: dict[str, object | None]) -> str:
    for prefix, field in (("uuid", "ap_uuid"), ("mac", "ap_mac"), ("sn", "serial_number"), ("name", "ap_name"), ("ip", "ap_ip")):
        value = str(row.get(field) or "").strip().casefold()
        if value:
            return f"{prefix}:{value}"
    return ""


def _normalize_name(value: object) -> str:
    return str(value or "").strip().casefold()


def _historical_switch_name(lldp: dict[str, object | None]) -> object:
    return lldp.get("neighbor_device_name") or lldp.get("lldp_neighbor") or lldp.get("neighbor_sysname")


def _historical_switch_interface(lldp: dict[str, object | None]) -> object:
    return lldp.get("neighbor_interface")


def _ap_mac(resource: dict[str, object | None], lldp: dict[str, object | None]) -> str:
    for value in (
        resource.get("ap_mac"),
        resource.get("mac"),
        lldp.get("ap_mac"),
        lldp.get("neighbor_mac"),
        resource.get("ap_name"),
    ):
        mac = _mac_from_text(value)
        if mac:
            return mac
    return ""


def _mac_from_text(value: object) -> str:
    hex_text = re.sub(r"[^0-9a-fA-F]", "", str(value or ""))
    if len(hex_text) != 12:
        return ""
    hex_text = hex_text.casefold()
    return f"{hex_text[0:4]}-{hex_text[4:8]}-{hex_text[8:12]}"


def _display_value(value: object) -> str:
    return str(value) if value not in (None, "") else "-"


def _format_sheet(sheet) -> None:
    from openpyxl.styles import Alignment, Border, Font, Side
    from netconsole.ui.table.table_autosize_engine import apply_worksheet_autofit

    alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    sheet.freeze_panes = "A2"
    for row in sheet.iter_rows():
        sheet.row_dimensions[row[0].row].height = 24 if row[0].row == 1 else 22
        for cell in row:
            cell.alignment = alignment
            cell.border = border
            if cell.row == 1:
                cell.font = Font(bold=True)
    apply_worksheet_autofit(sheet, maximum=60)
