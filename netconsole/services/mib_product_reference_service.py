from __future__ import annotations

import hashlib
import re
import shutil
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from netconsole.core.paths import PathResolver
from netconsole.repositories.global_mib_repository import GlobalMibRepository


@dataclass(frozen=True)
class ProductReferenceImportReport:
    source_path: str
    stored_path: str
    file_hash: str
    status: str
    reference_id: int | None = None
    reference_name: str = ""
    object_count: int = 0
    trap_count: int = 0
    sheet_names: list[str] = field(default_factory=list)
    error_message: str = ""
    duplicate_reference_id: int | None = None


@dataclass(frozen=True)
class ProductReferenceMeta:
    vendor: str = "H3C"
    product_line: str = ""
    product_name: str = ""
    device_type: str = ""
    os_family: str = "Comware"
    os_major: str = ""
    release_series: str = ""
    doc_version: str = ""
    reference_name: str = ""


OBJECT_FIELDS = {
    "module_name": ("模块", "模块名", "MIB模块", "MIB Module", "Module"),
    "mib_file_name": ("MIB文件", "MIB文件名", "文件名", "MIB File", "File"),
    "object_name": ("节点名称", "对象名称", "对象名", "节点名", "MIB节点", "全局节点名称及OID", "子节点名称及OID", "Object Name", "Node Name", "Name"),
    "numeric_oid": ("OID", "节点OID", "数字OID", "Object OID", "Node OID"),
    "object_scope": ("范围", "对象类型", "节点类型", "Scope", "Object Type"),
    "access_from_reference": ("访问权限", "最大访问权限", "操作权限", "MAX-ACCESS", "Access"),
    "data_type_from_reference": ("数据类型", "类型", "Syntax", "Data Type"),
    "value_range": ("取值范围", "有效范围", "值域", "Value Range", "Range"),
    "chinese_description": ("中文含义", "中文描述", "描述", "含义", "Description"),
    "implementation_spec": ("实现规格", "实现说明", "功能描述", "规格", "Implementation", "Spec"),
    "operation_support": ("支持情况", "操作支持情况", "支持操作", "读写支持", "Operation Support", "Support"),
    "table_parent_name": ("所属表", "父表", "父节点名称", "表名", "Table"),
    "table_index_info": ("索引", "INDEX", "索引信息", "Index"),
}

TRAP_FIELDS = {
    "module_name": ("模块", "模块名", "MIB模块", "MIB Module", "Module"),
    "mib_file_name": ("MIB文件", "MIB文件名", "文件名", "MIB File", "File"),
    "trap_name": ("Trap名称", "告警名称", "告警节点名称", "通知名称", "Trap Name", "Notification Name", "Name"),
    "trap_oid": ("Trap OID", "告警OID", "通知OID", "OID"),
    "trap_title": ("告警标题", "标题", "Alarm Title", "Title"),
    "trap_type": ("告警类型", "Trap类型", "类型", "Alarm Type", "Trap Type"),
    "trap_level": ("告警级别", "级别", "Severity", "Level"),
    "clear_trap_oid": ("恢复Trap OID", "清除Trap OID", "Clear Trap OID"),
    "clear_trap_name": ("恢复Trap名称", "清除Trap名称", "Clear Trap Name"),
    "default_status": ("缺省状态", "默认状态", "Default Status"),
    "trigger_reason": ("触发原因", "产生原因", "原因", "Trigger Reason", "Reason"),
    "system_impact": ("系统影响", "影响", "Impact", "System Impact"),
    "status_control": ("状态控制", "Status Control"),
    "varbind_oids": ("绑定变量OID", "变量OID", "Varbind OID", "Binding OID"),
    "varbind_names": ("绑定变量名称", "变量名称", "Varbind Name", "Binding Name"),
    "varbind_descriptions": ("绑定变量说明", "变量说明", "Varbind Description", "Binding Description"),
    "varbind_index_nodes": ("绑定变量索引节点", "索引节点", "Varbind Index"),
    "varbind_types": ("绑定变量类型", "变量类型", "Varbind Type"),
    "varbind_value_ranges": ("绑定变量取值范围", "变量取值范围", "Varbind Range"),
    "suggestion": ("处理建议", "建议", "Suggestion", "Recommended Action"),
}

OBJECT_FIELDS.update(
    {
        "category_name": ("分册名", "分类名", "功能分类", "Category"),
        "module_name": (*OBJECT_FIELDS["module_name"], "模块名"),
        "mib_file_name": (*OBJECT_FIELDS["mib_file_name"], "MIB文件名"),
        "root_node_name": ("根节点", "Root Node"),
        "parent_node_name": ("父节点名称", "父节点", "表节点信息", "Parent Node"),
        "object_name": (*OBJECT_FIELDS["object_name"], "全局节点名称及OID", "子节点名称及OID", "节点名", "节点名称"),
        "access_from_reference": (*OBJECT_FIELDS["access_from_reference"], "最大访问权限"),
        "data_type_from_reference": (*OBJECT_FIELDS["data_type_from_reference"], "数据类型"),
        "value_range": (*OBJECT_FIELDS["value_range"], "有效范围"),
        "chinese_description": (*OBJECT_FIELDS["chinese_description"], "含义", "中文含义"),
        "function_description": ("功能描述", "功能介绍", "Function Description"),
        "implementation_spec": (*OBJECT_FIELDS["implementation_spec"], "实现规格"),
        "operation_support": (*OBJECT_FIELDS["operation_support"], "操作支持情况"),
    }
)
TRAP_FIELDS.update(
    {
        "category_name": ("分册名", "分类名", "功能分类", "Category"),
        "module_name": (*TRAP_FIELDS["module_name"], "模块名"),
        "mib_file_name": (*TRAP_FIELDS["mib_file_name"], "MIB文件名"),
    }
)


OBJECT_FIELDS.update(
    {
        "category_name": (*OBJECT_FIELDS.get("category_name", ()), "分册名", "分类名", "功能分类", "Category"),
        "module_name": (*OBJECT_FIELDS["module_name"], "模块名", "MIB模块", "MIB Module", "Module"),
        "mib_file_name": (*OBJECT_FIELDS["mib_file_name"], "MIB文件名", "MIB文件", "文件名", "MIB File", "File"),
        "root_node_name": (*OBJECT_FIELDS.get("root_node_name", ()), "根节点", "Root Node"),
        "parent_node_name": (*OBJECT_FIELDS.get("parent_node_name", ()), "父节点名称", "父节点", "表节点信息", "Parent Node"),
        "object_name": (*OBJECT_FIELDS["object_name"], "全局节点名称及OID", "子节点名称及OID", "节点名", "节点名称", "对象名", "对象名称"),
        "numeric_oid": (*OBJECT_FIELDS["numeric_oid"], "OID", "节点OID", "数字OID"),
        "access_from_reference": (*OBJECT_FIELDS["access_from_reference"], "最大访问权限", "访问权限", "MAX-ACCESS", "Access"),
        "data_type_from_reference": (*OBJECT_FIELDS["data_type_from_reference"], "数据类型", "类型", "Syntax"),
        "value_range": (*OBJECT_FIELDS["value_range"], "有效范围", "取值范围", "值域", "Range"),
        "chinese_description": (*OBJECT_FIELDS["chinese_description"], "含义", "中文含义", "中文描述", "描述"),
        "function_description": (*OBJECT_FIELDS.get("function_description", ()), "功能描述", "功能介绍", "Function Description"),
        "implementation_spec": (*OBJECT_FIELDS["implementation_spec"], "实现规格", "实现说明"),
        "operation_support": (*OBJECT_FIELDS["operation_support"], "操作支持情况", "支持情况", "支持操作"),
        "table_parent_name": (*OBJECT_FIELDS["table_parent_name"], "所属表", "父表", "表名"),
        "table_index_info": (*OBJECT_FIELDS["table_index_info"], "索引", "INDEX", "索引信息"),
    }
)
TRAP_FIELDS.update(
    {
        "category_name": (*TRAP_FIELDS.get("category_name", ()), "分册名", "分类名", "功能分类", "Category"),
        "module_name": (*TRAP_FIELDS["module_name"], "模块名", "MIB模块", "MIB Module", "Module"),
        "mib_file_name": (*TRAP_FIELDS["mib_file_name"], "MIB文件名", "MIB文件", "文件名", "MIB File", "File"),
        "trap_name": (*TRAP_FIELDS["trap_name"], "告警节点名称", "告警名称", "Trap名称", "通知名称"),
        "trap_oid": (*TRAP_FIELDS["trap_oid"], "告警OID", "Trap OID", "通知OID", "OID"),
        "trap_title": (*TRAP_FIELDS["trap_title"], "告警标题", "标题"),
        "trap_type": (*TRAP_FIELDS["trap_type"], "告警类型", "Trap类型", "类型"),
        "trap_level": (*TRAP_FIELDS["trap_level"], "告警级别", "级别", "Severity", "Level"),
        "clear_trap_oid": (*TRAP_FIELDS["clear_trap_oid"], "清除告警OID", "恢复Trap OID", "Clear Trap OID"),
        "clear_trap_name": (*TRAP_FIELDS["clear_trap_name"], "清除告警名称", "恢复Trap名称", "Clear Trap Name"),
        "default_status": (*TRAP_FIELDS["default_status"], "缺省状态", "默认状态"),
        "trigger_reason": (*TRAP_FIELDS["trigger_reason"], "触发原因", "产生原因", "原因"),
        "system_impact": (*TRAP_FIELDS["system_impact"], "系统影响", "影响"),
        "status_control": (*TRAP_FIELDS["status_control"], "状态控制", "Status Control"),
        "varbind_oids": (*TRAP_FIELDS["varbind_oids"], "绑定变量OID", "变量OID", "Varbind OID"),
        "varbind_names": (*TRAP_FIELDS["varbind_names"], "绑定变量名称", "变量名称", "Varbind Name"),
        "varbind_descriptions": (*TRAP_FIELDS["varbind_descriptions"], "绑定变量说明", "变量说明", "Varbind Description"),
        "varbind_index_nodes": (*TRAP_FIELDS["varbind_index_nodes"], "绑定变量索引节点", "索引节点", "Varbind Index"),
        "varbind_types": (*TRAP_FIELDS["varbind_types"], "绑定变量类型", "变量类型", "Varbind Type"),
        "varbind_value_ranges": (*TRAP_FIELDS["varbind_value_ranges"], "绑定变量取值范围", "变量取值范围", "Varbind Range"),
        "suggestion": (*TRAP_FIELDS["suggestion"], "处理建议", "建议", "Suggestion"),
    }
)


def _extend_fields(target: dict[str, tuple[str, ...]], extra: dict[str, tuple[str, ...]]) -> None:
    for field, synonyms in extra.items():
        target[field] = (*target.get(field, ()), *synonyms)


_extend_fields(
    OBJECT_FIELDS,
    {
        "category_name": ("分册名", "分类名", "功能分类", "类别"),
        "module_name": ("模块名", "MIB模块", "MIB模块名"),
        "mib_file_name": ("MIB文件名", "MIB文件", "文件名"),
        "root_node_name": ("根节点", "根节点名称"),
        "parent_node_name": ("父节点名称", "父节点", "表节点名称"),
        "object_name": ("全局节点名称及OID", "子节点名称及OID", "节点名", "节点名称", "对象名", "对象名称"),
        "numeric_oid": ("OID", "节点OID", "数字OID"),
        "access_from_reference": ("最大访问权限", "访问权限"),
        "data_type_from_reference": ("数据类型", "类型"),
        "value_range": ("有效范围", "取值范围", "值域"),
        "chinese_description": ("含义", "中文含义", "中文描述", "描述"),
        "function_description": ("功能描述", "功能介绍"),
        "implementation_spec": ("实现规格", "实现说明"),
        "operation_support": ("操作支持情况", "支持情况", "支持操作"),
        "table_parent_name": ("所属表", "父表", "表名"),
        "table_index_info": ("表节点信息", "索引", "索引信息"),
    },
)
_extend_fields(
    TRAP_FIELDS,
    {
        "category_name": ("分册名", "分类名", "功能分类", "类别"),
        "module_name": ("模块名", "MIB模块", "MIB模块名"),
        "mib_file_name": ("MIB文件名", "MIB文件", "文件名"),
        "trap_name": ("告警节点名称", "告警名称", "Trap名称", "通知名称"),
        "trap_oid": ("告警OID", "Trap OID", "通知OID", "OID"),
        "trap_title": ("告警标题", "标题"),
        "trap_type": ("告警类型", "Trap类型", "类型"),
        "trap_level": ("告警级别", "级别"),
        "clear_trap_oid": ("清除告警OID", "恢复Trap OID", "清除Trap OID"),
        "clear_trap_name": ("清除告警名称", "恢复Trap名称", "清除Trap名称"),
        "default_status": ("缺省状态", "默认状态"),
        "trigger_reason": ("触发原因", "产生原因", "原因"),
        "system_impact": ("系统影响", "影响"),
        "status_control": ("状态控制",),
        "varbind_oids": ("绑定变量OID", "变量OID"),
        "varbind_names": ("绑定变量节点名称", "绑定变量名称", "变量名称"),
        "varbind_descriptions": ("绑定变量含义", "绑定变量说明", "变量说明"),
        "varbind_index_nodes": ("绑定变量索引节点", "索引节点"),
        "varbind_types": ("绑定变量类型", "变量类型"),
        "varbind_value_ranges": ("绑定变量取值范围", "变量取值范围"),
        "suggestion": ("处理建议", "建议"),
    },
)


class MibProductReferenceService:
    def __init__(self, paths: PathResolver, repository: GlobalMibRepository | None = None) -> None:
        self.paths = paths
        self.repository = repository or GlobalMibRepository(paths.global_mib_db_path())

    def import_file(
        self,
        source: str | Path,
        *,
        vendor: str = "",
        product_line: str = "",
        product_name: str = "",
        software_version: str = "",
        reference_name: str = "",
    ) -> ProductReferenceImportReport:
        self.paths.ensure_global_mib_dirs()
        self.repository.initialize()
        source_path = Path(source)
        file_hash = _sha256_file(source_path)
        duplicate = self.repository.get_product_reference_by_hash(file_hash)
        if duplicate is not None:
            workbook = load_workbook(source_path, read_only=True, data_only=True)
            sheet_names: list[str] = []
            try:
                sheet_names = list(workbook.sheetnames)
                object_rows: list[dict[str, str]] = []
                trap_rows: list[dict[str, str]] = []
                for sheet in workbook.worksheets:
                    parsed_objects, parsed_traps = _parse_sheet(sheet.title, sheet.iter_rows(values_only=True))
                    object_rows.extend(parsed_objects)
                    trap_rows.extend(parsed_traps)
                duplicate_id = int(duplicate["id"])
                if not object_rows and not trap_rows:
                    return ProductReferenceImportReport(
                        source_path=str(source_path),
                        stored_path=str(duplicate.get("source_file") or ""),
                        file_hash=file_hash,
                        status="failed",
                        reference_id=duplicate_id,
                        reference_name=str(duplicate.get("reference_name") or ""),
                        sheet_names=sheet_names,
                        error_message="产品 MIB 参考表未解析到对象或告警，请检查 Excel 表头并重新导入。",
                        duplicate_reference_id=duplicate_id,
                    )
                self.repository.replace_product_reference_content(
                    duplicate_id,
                    sheet_names=list(workbook.sheetnames),
                    object_overrides=object_rows,
                    trap_overrides=trap_rows,
                )
            finally:
                workbook.close()
            return ProductReferenceImportReport(
                source_path=str(source_path),
                stored_path=str(duplicate.get("source_file") or ""),
                file_hash=file_hash,
                status="duplicate_reindexed",
                reference_id=duplicate_id,
                reference_name=str(duplicate.get("reference_name") or ""),
                object_count=len(object_rows),
                trap_count=len(trap_rows),
                sheet_names=sheet_names,
                duplicate_reference_id=int(duplicate["id"]),
            )
        stored_dir = self.paths.global_mib_references_dir()
        stored_dir.mkdir(parents=True, exist_ok=True)
        stored_path = _unique_target(stored_dir / source_path.name)
        if source_path.resolve() != stored_path.resolve():
            shutil.copy2(source_path, stored_path)

        workbook = load_workbook(stored_path, read_only=True, data_only=True)
        try:
            meta = _parse_reference_meta(source_path)
            object_rows: list[dict[str, str]] = []
            trap_rows: list[dict[str, str]] = []
            for sheet in workbook.worksheets:
                parsed_objects, parsed_traps = _parse_sheet(sheet.title, sheet.iter_rows(values_only=True))
                object_rows.extend(parsed_objects)
                trap_rows.extend(parsed_traps)
            name = reference_name or meta.reference_name or _reference_name_from_file(source_path)
            if not object_rows and not trap_rows:
                return ProductReferenceImportReport(
                    source_path=str(source_path),
                    stored_path=str(stored_path),
                    file_hash=file_hash,
                    status="failed",
                    reference_name=name,
                    sheet_names=list(workbook.sheetnames),
                    error_message="产品 MIB 参考表未解析到对象或告警，请检查 Excel 表头并重新导入。",
                )
            reference_id = self.repository.insert_product_reference(
                vendor=vendor or meta.vendor,
                product_line=product_line or meta.product_line,
                product_name=product_name,
                device_type=meta.device_type,
                os_family=meta.os_family,
                os_major=meta.os_major,
                release_series=meta.release_series,
                doc_version=meta.doc_version,
                software_version=software_version or meta.release_series or _version_hint(source_path.stem),
                reference_name=name,
                source_file=str(stored_path),
                file_hash=file_hash,
                source_path=str(source_path),
                stored_path=str(stored_path),
                sheet_names=workbook.sheetnames,
                object_overrides=object_rows,
                trap_overrides=trap_rows,
            )
            return ProductReferenceImportReport(
                source_path=str(source_path),
                stored_path=str(stored_path),
                file_hash=file_hash,
                status="reference_imported",
                reference_id=reference_id,
                reference_name=name,
                object_count=len(object_rows),
                trap_count=len(trap_rows),
                sheet_names=list(workbook.sheetnames),
            )
        finally:
            workbook.close()


def _parse_sheet(sheet_name: str, rows: Iterable[tuple[object, ...]]) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    buffered = list(rows)
    header_index, header = _find_header(buffered)
    if header_index < 0:
        return [], []
    object_map = _field_map(header, OBJECT_FIELDS)
    trap_map = _field_map(header, TRAP_FIELDS)
    object_rows: list[dict[str, str]] = []
    trap_rows: list[dict[str, str]] = []
    sheet_is_trap = any(keyword in sheet_name.lower() for keyword in ("trap", "notification", "告警", "通知"))
    for raw_row in buffered[header_index + 1 :]:
        if not raw_row or not any(_text(value) for value in raw_row):
            continue
        object_record = _extract_record(raw_row, object_map)
        trap_record = _extract_record(raw_row, trap_map)
        object_record["sheet_name"] = sheet_name
        trap_record["sheet_name"] = sheet_name
        object_record["module_name"] = _normalize_module_name(object_record.get("module_name", ""))
        trap_record["module_name"] = _normalize_module_name(trap_record.get("module_name", ""))
        _split_name_oid(object_record, "object_name", "numeric_oid")
        _split_name_oid(trap_record, "trap_name", "trap_oid")
        if _is_trap_row(sheet_is_trap, trap_record):
            if trap_record.get("trap_name") or trap_record.get("trap_oid"):
                trap_rows.append(trap_record)
            continue
        if object_record.get("object_name") or object_record.get("numeric_oid"):
            object_rows.append(object_record)
    return object_rows, trap_rows


def _find_header(rows: list[tuple[object, ...]]) -> tuple[int, list[str]]:
    for index, row in enumerate(rows[:30]):
        headers = [_text(value) for value in row]
        normalized = {_normalize_header(value) for value in headers if value}
        matches = 0
        for synonyms in (*OBJECT_FIELDS.values(), *TRAP_FIELDS.values()):
            if any(_normalize_header(item) in normalized for item in synonyms):
                matches += 1
        if matches >= 2:
            return index, headers
    return -1, []


def _field_map(headers: list[str], fields: dict[str, tuple[str, ...]]) -> dict[str, int]:
    normalized_headers = [_normalize_header(header) for header in headers]
    result: dict[str, int] = {}
    for field, synonyms in fields.items():
        for synonym in synonyms:
            needle = _normalize_header(synonym)
            if needle in normalized_headers:
                result[field] = normalized_headers.index(needle)
                break
    return result


def _extract_record(row: tuple[object, ...], field_map: dict[str, int]) -> dict[str, str]:
    return {field: _text(row[index]) if index < len(row) else "" for field, index in field_map.items()}


def _split_name_oid(record: dict[str, str], name_field: str, oid_field: str) -> None:
    value = record.get(name_field, "")
    if not value:
        return
    match = re.search(r"([A-Za-z][A-Za-z0-9-]*)\s*[（(]\s*([0-9]+(?:\.[0-9]+)+)\s*[）)]", value)
    if not match:
        return
    record[name_field] = match.group(1)
    if not record.get(oid_field):
        record[oid_field] = match.group(2)


def _is_trap_row(sheet_is_trap: bool, record: dict[str, str]) -> bool:
    if sheet_is_trap and (record.get("trap_name") or record.get("trap_oid") or record.get("trap_title")):
        return True
    return bool(record.get("trap_title") or record.get("trap_level") or record.get("trigger_reason") or record.get("varbind_names"))


def _normalize_header(value: object) -> str:
    text = _text(value).lower()
    for char in (" ", "\n", "\t", "\r", "_", "-", "：", ":", "（", "）", "(", ")", "/", "\\"):
        text = text.replace(char, "")
    return text


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _split_name_oid(record: dict[str, str], name_field: str, oid_field: str) -> None:
    value = record.get(name_field, "")
    if not value:
        return
    match = re.search(r"([A-Za-z][A-Za-z0-9-]*)\s*[\(（]\s*([0-9]+(?:\.[0-9]+)+)\s*[\)）]", value)
    if match:
        record[name_field] = match.group(1)
        if not record.get(oid_field):
            record[oid_field] = match.group(2)
        return
    oid_match = re.search(r"([0-9]+(?:\.[0-9]+)+)", value)
    if oid_match and not record.get(oid_field):
        record[oid_field] = oid_match.group(1)


def _normalize_module_name(value: str) -> str:
    text = _text(value)
    return re.sub(r"^\d+\s*[-_]\s*", "", text).strip()


def _normalize_header(value: object) -> str:
    text = _text(value).lower()
    for char in (" ", "\n", "\t", "\r", "_", "-", "：", ":", "，", ",", "、", "(", ")", "（", "）", "/", "\\"):
        text = text.replace(char, "")
    return text


def _reference_name_from_file(path: Path) -> str:
    return path.stem


def _parse_reference_meta(path: Path) -> ProductReferenceMeta:
    stem = path.stem.replace("_", " ")
    release_tokens = re.findall(r"\b[RE]?\d{2}xx\b", stem, flags=re.IGNORECASE)
    release_series = ",".join(_normalize_release_series(token) for token in release_tokens)
    doc_match = re.search(r"\b(\d+W\d+)\b", stem, flags=re.IGNORECASE)
    doc_version = doc_match.group(1).upper() if doc_match else ""
    product_line = "无线控制器" if "无线控制器" in stem or "WX" in stem.upper() else ""
    device_type = "wireless_ac" if product_line == "无线控制器" else ""
    os_major = "V9" if "r16xx" in {item.lower() for item in release_series.split(",")} else ""
    return ProductReferenceMeta(
        product_line=product_line,
        device_type=device_type,
        os_major=os_major,
        release_series=release_series,
        doc_version=doc_version,
        reference_name=stem,
    )


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("_x000D_", " ").replace("\r", " ").replace("\n", " ").strip()


def _normalize_release_series(value: str) -> str:
    value = value.strip()
    match = re.fullmatch(r"([A-Za-z]?)(\d{2})xx", value, flags=re.IGNORECASE)
    if not match:
        return value
    prefix, digits = match.groups()
    return f"{prefix.upper() if prefix else 'R'}{digits}xx"


def _version_hint(name: str) -> str:
    match = re.search(r"R\d+[A-Za-z0-9-]*", name)
    return match.group(0) if match else ""


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _unique_target(path: Path) -> Path:
    if not path.exists():
        return path
    index = 2
    while True:
        candidate = path.with_name(f"{path.stem}_{index}{path.suffix}")
        if not candidate.exists():
            return candidate
        index += 1


def _split_name_oid(record: dict[str, str], name_field: str, oid_field: str) -> None:
    value = record.get(name_field, "")
    if not value:
        return
    match = re.search(r"([A-Za-z][A-Za-z0-9-]*)\s*[\(（]\s*([0-9]+(?:\.[0-9]+)+)\s*[\)）]", value)
    if match:
        record[name_field] = match.group(1)
        if not record.get(oid_field):
            record[oid_field] = match.group(2)
        return
    oid_match = re.search(r"([0-9]+(?:\.[0-9]+)+)", value)
    if oid_match and not record.get(oid_field):
        record[oid_field] = oid_match.group(1)


def _normalize_header(value: object) -> str:
    text = _text(value).lower()
    for char in (" ", "\n", "\t", "\r", "_", "-", "－", "—", ":", "：", ",", "，", "。", "(", ")", "（", "）", "/", "\\"):
        text = text.replace(char, "")
    return text


def _parse_reference_meta(path: Path) -> ProductReferenceMeta:
    stem = path.stem.replace("_", " ")
    release_tokens = re.findall(r"\b[RE]?\d{2}xx\b", stem, flags=re.IGNORECASE)
    release_series = ",".join(_normalize_release_series(token) for token in release_tokens)
    doc_match = re.search(r"\b(\d+W\d+)\b", stem, flags=re.IGNORECASE)
    doc_version = doc_match.group(1).upper() if doc_match else ""
    product_line = "无线控制器" if "无线控制器" in stem or "WX" in stem.upper() else ""
    device_type = "wireless_ac" if product_line == "无线控制器" else ""
    os_major = "V9" if "r16xx" in {item.lower() for item in release_series.split(",")} else ""
    return ProductReferenceMeta(
        product_line=product_line,
        device_type=device_type,
        os_major=os_major,
        release_series=release_series,
        doc_version=doc_version,
        reference_name=stem,
    )
