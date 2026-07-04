from __future__ import annotations

import csv
import json
import logging
import os
import platform
import re
import subprocess
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import asdict, dataclass, field, replace
from datetime import datetime, timedelta
from pathlib import Path
from typing import Callable, Iterable

from netconsole.core.paths import PathResolver
from netconsole.models.device import Device
from netconsole.services.netmiko_connection import (
    H3C_NETMIKO_DEVICE_TYPE,
    build_netmiko_params,
    connection_targets,
    encoding_for_vendor,
    normalize_command_output,
    safe_send_command,
)
from netconsole.services.vehicle_mr_online import (
    H3CComwareV9VehicleMrMeshLinkParser,
    TRAIN_STATUS_ABNORMAL_SINGLE,
    TRAIN_STATUS_DUAL_ONLINE,
    TRAIN_STATUS_ONLINE,
    TRAIN_STATUS_UNEXPECTED_END,
    VehicleMrOnlineStore,
    VehicleMrTrainState,
    load_group_names,
    normalize_train_no,
    parse_train_identity,
    canonical_peer_name,
)
from netconsole.utils.excel_workbook import load_workbook_without_unsupported_image_warning
from netconsole.utils.natural_sort import train_natural_sort_key

try:
    from netmiko import ConnectHandler
except ImportError:  # pragma: no cover
    ConnectHandler = None  # type: ignore[assignment]


NODE_ORDER = ("TC1-MR", "TC1-SW", "TC1-SRV", "TC2-MR", "TC2-SW", "TC2-SRV")
POINT_TABLE_FIELDS = (
    "train_id",
    "train_no",
    "display_name",
    "tc",
    "end",
    "node_name",
    "node_type",
    "device_id",
    "device_name",
    "device_group",
    "station",
    "primary_address",
    "backup_address",
    "ip_vehicle",
    "ip_uplink",
    "ssh_host",
    "vrrp_ip",
    "address_mapping_mode",
    "primary_address_role",
    "backup_address_role",
    "remark",
)
AC_COMMANDS = ("display wlan mesh-link ap", "display wlan ap all", "display wlan ap all radio")
LOGGER = logging.getLogger(__name__)
ONLINE_MESH_STATUSES = {"forwarding", "active", "up"}
MR_REMOTE_PING_CONCURRENCY_PER_MR = 2
MR_REMOTE_PING_CONCURRENCY_TOTAL = 4
CAR_NETWORK_QUICK_DETECT_SECONDS = 4
CAR_NETWORK_QUICK_PING_COUNT = 4
CAR_NETWORK_QUICK_PING_TIMEOUT = 4
CAR_NETWORK_QUICK_CLI_READ_TIMEOUT = 8
CAR_NETWORK_CROSS_TC_PING_COUNT = 50
CAR_NETWORK_CROSS_TC_PING_TIMEOUT = 60
DEFAULT_GLOBAL_CONFIG = {
    "address_mapping": {
        "MR": {"primary_address_role": "vehicle_ip", "backup_address_role": "uplink_ip", "ssh_source": "primary_address"},
        "3SW": {"primary_address_role": "vehicle_ip", "backup_address_role": "uplink_ip", "ssh_source": "primary_address"},
        "SRV": {"primary_address_role": "vehicle_ip", "backup_address_role": "ignore", "ssh_source": "empty"},
    },
    "srv_generation": {"enabled": True, "tc1_host": 1, "tc2_host": 2, "vrrp_host": 254, "mode": "same_vehicle_subnet"},
    "point_table_locked": False,
}


@dataclass(frozen=True)
class CarNetworkNode:
    train_id: str
    node_name: str
    node_type: str
    display_name: str = ""
    ip_vehicle: str = ""
    ip_uplink: str = ""
    remark: str = ""
    train_no: str = ""
    tc: str = ""
    end: str = ""
    device_id: str = ""
    device_name: str = ""
    device_group: str = ""
    station: str = ""
    primary_address: str = ""
    backup_address: str = ""
    ssh_host: str = ""
    vrrp_ip: str = ""
    address_mapping_mode: str = "global"
    primary_address_role: str = ""
    backup_address_role: str = ""

    @property
    def normalized_name(self) -> str:
        return normalize_node_name(self.node_name)

    @property
    def is_mr(self) -> bool:
        return self.normalized_name.endswith("-MR") or self.node_type.upper() in {"MR", "AP", "FAT-AP"}

    @property
    def vehicle_ping_required(self) -> bool:
        return bool(self.ip_vehicle)

    @property
    def ping_ips(self) -> tuple[str, ...]:
        values = [self.ip_uplink, self.ip_vehicle]
        if self.ssh_host and self.ssh_host not in values:
            values.append(self.ssh_host)
        return tuple(dict.fromkeys(ip for ip in values if ip))

    def ssh_address(self, device: Device | None = None) -> str:
        if device is not None and device.primary_address:
            return device.primary_address
        return self.ssh_host or self.ip_uplink or self.ip_vehicle


@dataclass(frozen=True)
class CarNetworkTrain:
    train_id: str
    train_no: str
    display_name: str
    tc1_device: Device | None = None
    tc2_device: Device | None = None

    @property
    def mr_devices(self) -> dict[str, Device]:
        result: dict[str, Device] = {}
        if self.tc1_device is not None:
            result["TC1-MR"] = self.tc1_device
        if self.tc2_device is not None:
            result["TC2-MR"] = self.tc2_device
        return result


def get_train_sort_key(value: CarNetworkTrain | CarNetworkNode | tuple[object, ...] | object) -> tuple[object, ...]:
    if isinstance(value, CarNetworkTrain):
        return train_natural_sort_key(value.train_no, value.display_name, value.train_id)
    if isinstance(value, CarNetworkNode):
        return train_natural_sort_key(value.train_no, value.display_name, value.train_id)
    if isinstance(value, tuple):
        return train_natural_sort_key(*value)
    return train_natural_sort_key(value)


def sort_car_network_trains(trains: Iterable[CarNetworkTrain]) -> list[CarNetworkTrain]:
    return sorted(trains, key=get_train_sort_key)


@dataclass(frozen=True)
class PingResult:
    ip: str
    ok: bool
    loss_percent: float = 100.0
    avg_rtt_ms: float | None = None
    raw: str = ""
    error: str = ""
    transmitted: int | None = None
    received: int | None = None
    min_rtt_ms: float | None = None
    max_rtt_ms: float | None = None


@dataclass(frozen=True)
class SshResult:
    host: str
    ok: bool
    node_name: str = ""
    command_results: dict[str, PingResult] = field(default_factory=dict)
    task_results: dict[str, PingResult] = field(default_factory=dict)
    task_metadata: dict[str, dict[str, object]] = field(default_factory=dict)
    error: str = ""


@dataclass(frozen=True)
class MrRemotePingTask:
    task_id: str
    source_node: str
    source_host: str
    target_node: str
    target_ip: str
    direction: str
    command: str
    packet_count: int | None
    layer: str
    timeout: int


@dataclass(frozen=True)
class CoreRemotePingTask:
    task_id: str
    core_device_id: str
    core_device_name: str
    core_host: str
    target_node: str
    target_ip: str
    command: str
    layer: str
    timeout: int


@dataclass(frozen=True)
class CoreDiscoveryCandidate:
    device_name: str
    system_name: str
    group: str
    host: str
    selected: bool
    reason: str


@dataclass(frozen=True)
class AcApStatus:
    mesh_link: bool = False
    ap_all: bool = False
    radio_ok: bool | None = None
    selected: bool = False
    raw: dict[str, str] = field(default_factory=dict)
    error: str = ""

    @property
    def online(self) -> bool:
        return self.mesh_link or self.ap_all

    def as_json(self) -> dict[str, object]:
        return {
            "selected": self.selected,
            "mesh_link": "ok" if self.mesh_link else "fail" if self.selected else "unknown",
            "ap_all": "ok" if self.ap_all else "fail" if self.selected else "unknown",
            "radio": "ok" if self.radio_ok is True else "fail" if self.radio_ok is False else "unknown",
        }


@dataclass(frozen=True)
class AcControllerProbe:
    device_id: str
    device_name: str
    host: str
    success: bool
    output: str = ""
    error: str = ""
    matched_mrs: list[dict[str, object]] = field(default_factory=list)
    fallback_matched_mrs: list[dict[str, object]] = field(default_factory=list)
    parsed_peer_count: int = 0


@dataclass(frozen=True)
class MeshLinkPeer:
    local_ap_name: str
    peer_name: str
    peer_mac: str
    local_mac: str
    status: str
    rssi: int | None = None
    raw_line: str = ""


@dataclass(frozen=True)
class AcProbeResult:
    enabled: bool
    query_success: bool
    controllers: list[AcControllerProbe] = field(default_factory=list)
    raw_outputs: dict[str, str] = field(default_factory=dict)
    error: str = ""

    def as_json(self, train_status: TrainAcStatus | None = None) -> dict[str, object]:
        payload: dict[str, object] = {
            "enabled": self.enabled,
            "auto_discovered": True,
            "query_success": self.query_success,
            "controllers": [
                {
                    "device_id": item.device_id,
                    "device_name": item.device_name,
                    "host": item.host,
                    "success": item.success,
                    "output_length": len(item.output or ""),
                    "output_excerpt": (item.output or "")[:2000],
                    "parsed_peer_count": item.parsed_peer_count,
                    "matched_mrs": item.matched_mrs,
                    "fallback_matched_mrs": item.fallback_matched_mrs,
                    "error": item.error,
                }
                for item in self.controllers
            ],
        }
        if self.error:
            payload["error"] = self.error
        if train_status is not None:
            payload.update(
                {
                    "tc1_mr_online": train_status.tc1_mr_online,
                    "tc2_mr_online": train_status.tc2_mr_online,
                    "both_mr_offline": train_status.both_mr_offline,
                    "matched_by": train_status.matched_by,
                    "name_match_mode": "semantic_train_no_end",
                    "matched_mrs": [item for values in train_status.matched_details.values() for item in values],
                    "parse_success": train_status.parse_success,
                    "parse_warning": train_status.parse_warning,
                    "suspected_current_train_lines": train_status.suspected_current_train_lines,
                    "ac_output_nonempty": train_status.ac_output_nonempty,
                    "online_source": train_status.online_source,
                    "parser": train_status.parser,
                }
            )
        return payload


@dataclass(frozen=True)
class TrainAcStatus:
    tc1_mr_online: bool = False
    tc2_mr_online: bool = False
    both_mr_offline: bool = False
    any_query_success: bool = False
    matched_by: dict[str, list[str]] = field(default_factory=dict)
    matched_details: dict[str, list[dict[str, object]]] = field(default_factory=dict)
    parse_success: bool = False
    parse_warning: bool = False
    suspected_current_train_lines: list[str] = field(default_factory=list)
    ac_output_nonempty: bool = False
    online_source: str = ""
    parser: str = "H3CComwareV9VehicleMrMeshLinkParser"


@dataclass
class CarNetworkDiagnosticResult:
    train_id: str
    status: str
    nodes: dict[str, str]
    cross_train: dict[str, str]
    ac_status: str
    ssh_status: str
    conclusion: str
    train_no: str = ""
    display_name: str = ""
    ends: dict[str, dict[str, object]] = field(default_factory=dict)
    vrrp: dict[str, object] = field(default_factory=dict)
    ping_results: dict[str, PingResult] = field(default_factory=dict)
    core_results: dict[str, PingResult] = field(default_factory=dict)
    ssh_results: dict[str, SshResult] = field(default_factory=dict)
    ac_detail: AcApStatus = field(default_factory=AcApStatus)
    ac_probe: AcProbeResult | None = None
    train_ac_status: TrainAcStatus | None = None
    tables: dict[str, list[dict[str, object]]] = field(default_factory=dict)
    core_discovery: dict[str, object] = field(default_factory=dict)
    cross_tc_ping: dict[str, object] = field(default_factory=dict)

    def to_json_dict(self) -> dict[str, object]:
        return {
            "train_id": self.train_id,
            "train_no": self.train_no,
            "display_name": self.display_name,
            "status": self.status,
            "conclusion": self.conclusion,
            "nodes": self.nodes,
            "ends": self.ends,
            "vrrp": self.vrrp,
            "cross_train": self.cross_train,
            "primary_diagnosis_scope": "vehicle_internal_network",
            "ac_status": self.ac_detail.as_json(),
            "ac_probe": self.ac_probe.as_json(self.train_ac_status) if self.ac_probe is not None else {},
            "ssh_status": _ssh_status_json(self.ssh_results) or self.ssh_status,
            "core_results": {key: "ok" if result.ok else "fail" for key, result in self.core_results.items()},
            "core_discovery": self.core_discovery,
            "access_entry": _access_entry_json(self.ssh_results),
            "ground_access_status": _ground_access_status_json(self.ac_detail, self.ssh_results, self.core_results),
            "vehicle_internal_status": _vehicle_internal_status_json(self.ssh_results),
            "cross_tc_ping": self.cross_tc_ping,
            "diagnosis_items": _remote_diagnosis_items(self.ssh_results) + _diagnosis_items_from_tables(self.tables),
            "tables": self.tables,
        }


def normalize_node_name(value: object) -> str:
    text = str(value or "").strip()
    return {"TC1-AP": "TC1-MR", "TC2-AP": "TC2-MR"}.get(text, text)


def _is_vehicle_mr_group(group_name: str, *, allow_legacy: bool = False) -> bool:
    text = str(group_name or "")
    if "车载-3SW" in text:
        return False
    if "车载-MR" in text:
        return True
    return allow_legacy and text == "车载"


def is_vehicle_mr_device(device: Device, group_name: str = "", *, allow_legacy_group: bool = False) -> bool:
    name = str(device.name or "")
    type_text = str(device.device_type or "").upper()
    if not _is_vehicle_mr_group(group_name, allow_legacy=allow_legacy_group):
        return False
    return type_text in {"FAT-AP", "CLOUD-AP", "AP", "MR"} or "-MR-CT" in name.upper() or "-MR-CW" in name.upper()


def build_car_network_trains(repository, site_name: str) -> list[CarNetworkTrain]:
    group_names = load_group_names(repository, site_name)
    by_no: dict[str, dict[str, object]] = {}
    for device in repository.list():
        group_name = group_names.get(int(device.group_id or 0), "")
        if not _is_vehicle_group(group_name):
            continue
        classified = _classify_vehicle_device(device, group_name)
        train_no = _device_train_no_any(device)
        tc, _end = _device_tc_end(device)
        if classified is None or not train_no:
            continue
        train_id = _device_train_id(device, train_no)
        entry = by_no.setdefault(
            train_no,
            {
                "train_id": train_id,
                "train_no": train_no,
                "tc1_device": None,
                "tc2_device": None,
            },
        )
        if classified[0] == "MR":
            if tc == "TC1":
                entry["tc1_device"] = device
            elif tc == "TC2":
                entry["tc2_device"] = device
    trains = [
        CarNetworkTrain(
            str(item["train_id"]),
            str(item["train_no"]),
            f"{item['train_no']}车" if item["train_no"] else str(item["train_id"]),
            item["tc1_device"] if isinstance(item.get("tc1_device"), Device) else None,
            item["tc2_device"] if isinstance(item.get("tc2_device"), Device) else None,
        )
        for item in by_no.values()
    ]
    return sort_car_network_trains(trains)


def build_train_3sw_bindings(repository, site_name: str, trains: list[CarNetworkTrain]) -> dict[str, dict[str, Device]]:
    group_names = load_group_names(repository, site_name)
    train_nos = {train.train_no for train in trains if train.train_no}
    result: dict[str, dict[str, Device]] = {}
    for device in repository.list():
        group_name = group_names.get(int(device.group_id or 0), "")
        if "车载-3SW" not in group_name:
            continue
        train_no = _device_train_no(device, train_nos)
        node_name = _sw_node_name(device)
        if train_no and node_name:
            result.setdefault(train_no, {})[node_name] = device
    return result


def discover_ac_devices(repository) -> list[Device]:
    devices: list[Device] = []
    for device in repository.list():
        text = " ".join(str(value or "") for value in (device.device_type, device.name, device.system_name)).upper()
        cn_text = " ".join(str(value or "") for value in (device.device_type, device.name, device.system_name))
        if "AC" in text or "WX" in text or "WIRELESS CONTROLLER" in text or "无线控制器" in cn_text:
            devices.append(device)
    return devices


def discover_core_switches(repository, site_name: str = "") -> list[Device]:
    return [device for device, candidate in discover_core_switch_candidates(repository, site_name) if candidate.selected]
    group_names = load_group_names(repository, site_name) if site_name else {}
    result: list[Device] = []
    for device in repository.list():
        group_name = group_names.get(int(device.group_id or 0), "")
        if site_name and group_name != "COCC":
            continue
        text = " ".join(str(value or "") for value in (device.name, device.system_name, device.device_type, device.remark))
        upper = text.upper()
        type_upper = str(device.device_type or "").upper()
        if (
            type_upper in {"SW", "SWITCH", "L3 SWITCH", "三层交换机", "交换机"}
            or "CORE" in upper
            or "SWITCH" in upper
            or "SW" in upper
            or "核心" in text
            or "交换机" in text
        ):
            result.append(device)
    return result


def discover_core_switch_candidates(repository, site_name: str = "") -> list[tuple[Device, CoreDiscoveryCandidate]]:
    group_names = load_group_names(repository, site_name) if site_name else {}
    candidates: list[tuple[Device, str, str]] = []
    for device in repository.list():
        group_name = group_names.get(int(device.group_id or 0), "")
        excluded, exclude_reason = _is_excluded_core_ac(device)
        if excluded:
            if "COCC" in group_name.upper():
                candidates.append((device, group_name, f"excluded: {exclude_reason}"))
            continue
        selected, reason = _is_core_switch_candidate(device)
        device_text = " ".join(str(value or "") for value in (device.name, device.system_name, device.device_type, device.remark)).upper()
        group_is_cocc = "COCC" in group_name.upper()
        device_mentions_cocc = "COCC" in device_text
        if group_is_cocc:
            candidates.append((device, group_name, reason if selected else "group=COCC non-AC fallback candidate"))
        elif selected and device_mentions_cocc:
            candidates.append((device, group_name, reason or "device identity contains COCC core switch"))
    selectable = [(device, group, reason) for device, group, reason in candidates if not reason.startswith("excluded:")]
    preferred = [(device, group, reason) for device, group, reason in selectable if not reason.endswith("fallback candidate")]
    selected = preferred or (selectable if len(selectable) == 1 else [])
    selected_ids = {_device_id(device) or device.name for device, _group, _reason in selected}
    result: list[tuple[Device, CoreDiscoveryCandidate]] = []
    for device, group_name, reason in candidates:
        device_key = _device_id(device) or device.name
        is_selected = device_key in selected_ids
        if is_selected and reason.endswith("fallback candidate"):
            reason = "group=COCC and only one non-AC device"
        result.append(
            (
                device,
                CoreDiscoveryCandidate(
                    device_name=device.name or "",
                    system_name=device.system_name or "",
                    group=group_name,
                    host=device.primary_address or "",
                    selected=is_selected,
                    reason=reason,
                ),
            )
        )
    return result


def _is_excluded_core_ac(device: Device) -> tuple[bool, str]:
    text = " ".join(str(value or "") for value in (device.name, device.system_name, device.device_type, device.remark))
    upper = text.upper()
    type_upper = str(device.device_type or "").upper()
    if type_upper in {"AC", "WIRELESS CONTROLLER", "无线控制器"}:
        return True, "device_type is AC/Wireless Controller"
    if "无线控制器" in text or "WIRELESS CONTROLLER" in upper or "WX" in upper:
        return True, "name/system contains wireless controller/WX"
    if "AC" in upper and "CORE" not in upper and "SWITCH" not in upper and "交换机" not in text and "核心" not in text:
        return True, "name/system contains AC"
    return False, ""


def _is_core_switch_candidate(device: Device) -> tuple[bool, str]:
    text = " ".join(str(value or "") for value in (device.name, device.system_name, device.device_type, device.remark))
    upper = text.upper()
    type_upper = str(device.device_type or "").upper()
    reasons: list[str] = []
    if "核心" in text:
        reasons.append("name contains 核心")
    if "CORE" in upper:
        reasons.append("system_name/name contains CORE")
    if "交换机" in text:
        reasons.append("name contains 交换机")
    if type_upper in {"SW", "SWITCH", "L3 SWITCH", "三层交换机", "交换机"}:
        reasons.append("device_type is Switch")
    if "SWITCH" in upper:
        reasons.append("name/system contains SWITCH")
    if reasons:
        return True, "group=COCC and " + " / ".join(reasons)
    return False, ""


def default_point_table(
    train_id: str = "",
    train_no: str | None = None,
    mr_devices: dict[str, Device] | None = None,
    global_config: dict[str, object] | None = None,
) -> list[CarNetworkNode]:
    train_no = train_no or normalize_train_no(train_id)
    config = merge_global_config(global_config)
    mr_devices = mr_devices or {}
    tc1_device = mr_devices.get("TC1-MR")
    tc2_device = mr_devices.get("TC2-MR")
    base = _vehicle_prefix_from_addresses(
        [
            tc1_device.primary_address if tc1_device else "",
            tc1_device.backup_address if tc1_device else "",
            tc2_device.primary_address if tc2_device else "",
            tc2_device.backup_address if tc2_device else "",
        ],
        train_no or "",
    )
    vrrp = _build_host_ip(base, _srv_generation(config).get("vrrp_host", 254)) if base else ""
    display_name = f"{train_no}车" if train_no else train_id
    nodes = [
        CarNetworkNode(
            train_id=train_id,
            train_no=train_no or "",
            display_name=display_name,
            tc="TC1",
            end="CT",
            node_name="TC1-MR",
            node_type="MR",
            device_id=_device_id(tc1_device),
            device_name=tc1_device.name if tc1_device else "",
            primary_address=tc1_device.primary_address if tc1_device else "",
            backup_address=tc1_device.backup_address if tc1_device else "",
            vrrp_ip=vrrp,
            remark="CT MR",
        ),
        CarNetworkNode(train_id=train_id, train_no=train_no or "", display_name=display_name, tc="TC1", end="CT", node_name="TC1-SW", node_type="SW", ip_vehicle=_build_host_ip(base, 251) if base else "", vrrp_ip=vrrp, remark="三层交换机"),
        CarNetworkNode(train_id=train_id, train_no=train_no or "", display_name=display_name, tc="TC1", end="CT", node_name="TC1-SRV", node_type="Server", ip_vehicle=_build_host_ip(base, _srv_generation(config).get("tc1_host", 1)) if base else "", vrrp_ip=vrrp, remark="车载服务器"),
        CarNetworkNode(
            train_id=train_id,
            train_no=train_no or "",
            display_name=display_name,
            tc="TC2",
            end="CW",
            node_name="TC2-MR",
            node_type="MR",
            device_id=_device_id(tc2_device),
            device_name=tc2_device.name if tc2_device else "",
            primary_address=tc2_device.primary_address if tc2_device else "",
            backup_address=tc2_device.backup_address if tc2_device else "",
            vrrp_ip=vrrp,
            remark="CW MR",
        ),
        CarNetworkNode(train_id=train_id, train_no=train_no or "", display_name=display_name, tc="TC2", end="CW", node_name="TC2-SW", node_type="SW", ip_vehicle=_build_host_ip(base, 252) if base else "", vrrp_ip=vrrp, remark="三层交换机"),
        CarNetworkNode(train_id=train_id, train_no=train_no or "", display_name=display_name, tc="TC2", end="CW", node_name="TC2-SRV", node_type="Server", ip_vehicle=_build_host_ip(base, _srv_generation(config).get("tc2_host", 2)) if base else "", vrrp_ip=vrrp, remark="车载服务器"),
    ]
    return [apply_address_mapping(node, config) for node in nodes]


def merge_train_nodes(existing: list[CarNetworkNode], train: CarNetworkTrain, sw_devices: dict[str, Device] | None = None) -> list[CarNetworkNode]:
    by_name = {node.normalized_name: node for node in existing if node.train_id == train.train_id or node.train_no == train.train_no}
    defaults = default_point_table(train.train_id, train.train_no, train.mr_devices)
    merged: list[CarNetworkNode] = []
    for node in defaults:
        old = by_name.get(node.node_name)
        sw_device = (sw_devices or {}).get(node.node_name)
        if old is None:
            merged.append(_bind_sw_node(node, sw_device))
            continue
        merged.append(_preserve_manual_mapping(old, _bind_sw_node(node, sw_device), train))
    return merged


def _bind_sw_node(node: CarNetworkNode, device: Device | None) -> CarNetworkNode:
    if device is None or node.is_mr or "-SW" not in node.node_name:
        return node
    return apply_address_mapping(
        replace(
            node,
            device_id=_device_id(device),
            device_name=device.name,
            station=device.station or node.station,
            primary_address=device.primary_address or node.primary_address,
            backup_address=device.backup_address or node.backup_address,
        )
    )


def generate_point_table_from_devices(
    repository,
    site_name: str,
    existing_nodes: list[CarNetworkNode] | None = None,
    global_config: dict[str, object] | None = None,
) -> list[CarNetworkNode]:
    existing_nodes = list(existing_nodes or [])
    config = merge_global_config(global_config)
    group_names = load_group_names(repository, site_name)
    generated: list[CarNetworkNode] = []
    discovered_keys: set[tuple[str, str]] = set()
    train_nos: set[str] = set()
    for device in repository.list():
        group_name = group_names.get(int(device.group_id or 0), "")
        if not _is_vehicle_group(group_name):
            continue
        classified = _classify_vehicle_device(device, group_name)
        if classified is None:
            continue
        node_type, node_suffix = classified
        train_no = _device_train_no_any(device)
        if not train_no:
            continue
        tc, end = _device_tc_end(device)
        if not tc:
            continue
        train_id = _device_train_id(device, train_no)
        display_name = f"{train_no}车"
        node_name = f"{tc}-{node_suffix}"
        base_node = CarNetworkNode(
            train_id=train_id,
            train_no=train_no,
            display_name=display_name,
            tc=tc,
            end=end,
            node_name=node_name,
            node_type=node_type,
            device_id=_device_id(device),
            device_name=device.name,
            device_group=group_name,
            station=device.station or "",
            primary_address=device.primary_address or "",
            backup_address=device.backup_address or "",
            remark=_normalized_node_remark(node_name, node_type, end),
            address_mapping_mode="global",
        )
        old = _find_existing_node(existing_nodes, base_node)
        mapped = apply_address_mapping(base_node, config, overwrite=True)
        node = _preserve_manual_mapping(old, mapped, None) if old is not None else mapped
        if node.address_mapping_mode == "global":
            node = apply_address_mapping(node, config, overwrite=True)
        generated.append(node)
        discovered_keys.add((_device_id(device), node_name))
        train_nos.add(train_no)

    for train_no in sorted(train_nos, key=get_train_sort_key):
        train_id = _train_id_for_no(train_no, [*existing_nodes, *generated])
        existing_for_train = [node for node in [*existing_nodes, *generated] if node.train_no == train_no or node.train_id == train_id]
        present = {node.node_name for node in existing_for_train}
        prefix = infer_vehicle_prefix(existing_for_train, train_no)
        for node in default_point_table(train_id, train_no, global_config=config):
            if node.node_name not in present:
                generated.append(_apply_generated_network_defaults(node, prefix, config))
        generated = [_apply_generated_network_defaults(node, infer_vehicle_prefix([item for item in generated if item.train_no == node.train_no], node.train_no), config) for node in generated]

    current_keys = {(node.device_id, node.node_name) for node in generated if node.device_id}
    for old in existing_nodes:
        if old.device_id and (old.device_id, old.node_name) not in current_keys:
            generated.append(replace(old, remark=_normalized_node_remark(old.node_name, old.node_type, old.end)))

    nodes = _sort_nodes(_dedupe_nodes(generated))
    nodes = normalize_train_network_defaults(nodes, config, overwrite_custom=False)
    return _sort_nodes(_dedupe_nodes(nodes))


def apply_global_rules_to_nodes(nodes: list[CarNetworkNode], global_config: dict[str, object] | None = None, *, overwrite_custom: bool = False) -> list[CarNetworkNode]:
    config = merge_global_config(global_config)
    result: list[CarNetworkNode] = []
    for node in nodes:
        if node.address_mapping_mode == "custom" and not overwrite_custom:
            result.append(node)
            continue
        mapped = apply_address_mapping(replace(node, address_mapping_mode="global"), config, overwrite=True)
        same_train_nodes = [
            item
            for item in [*nodes, mapped]
            if (item.train_no and item.train_no == mapped.train_no) or (item.train_id and item.train_id == mapped.train_id)
        ]
        prefix = infer_vehicle_prefix(same_train_nodes, mapped.train_no)
        normalized = replace(mapped, remark=_normalized_node_remark(mapped.node_name, mapped.node_type, mapped.end))
        result.append(_apply_generated_network_defaults(normalized, prefix, config, overwrite_vrrp=overwrite_custom or node.address_mapping_mode != "custom"))
    result = normalize_train_network_defaults(result, config, overwrite_custom=overwrite_custom)
    return _sort_nodes(result)


def normalize_train_network_defaults(
    nodes: list[CarNetworkNode],
    global_config: dict[str, object] | None = None,
    *,
    overwrite_custom: bool = False,
) -> list[CarNetworkNode]:
    config = merge_global_config(global_config)
    grouped: dict[str, list[CarNetworkNode]] = {}
    for node in nodes:
        key = node.train_no or node.train_id
        grouped.setdefault(key, []).append(node)

    normalized: list[CarNetworkNode] = []
    for key, group in grouped.items():
        train_no = next((node.train_no for node in group if node.train_no), key)
        prefix = infer_vehicle_prefix(group, train_no)
        srv = _srv_generation(config)
        vrrp_ip = _build_host_ip(prefix, srv.get("vrrp_host", 254)) if prefix else ""
        for node in group:
            mode = _normalize_mapping_mode(node.address_mapping_mode)
            can_overwrite = overwrite_custom or mode == "global"
            next_node = node
            if can_overwrite and vrrp_ip:
                next_node = replace(next_node, vrrp_ip=vrrp_ip)
            if can_overwrite and srv.get("enabled", True) and prefix and _node_config_key(next_node) == "SRV":
                host = srv.get("tc1_host", 1) if next_node.tc == "TC1" else srv.get("tc2_host", 2)
                next_node = replace(next_node, ip_vehicle=_build_host_ip(prefix, host))
            if can_overwrite or _legacy_remark(next_node.remark):
                next_node = replace(next_node, remark=_normalized_node_remark(next_node.node_name, next_node.node_type, next_node.end))
            normalized.append(next_node)
    return _sort_nodes(normalized)


def apply_address_mapping(
    node: CarNetworkNode,
    global_config: dict[str, object] | None = None,
    *,
    overwrite: bool = False,
) -> CarNetworkNode:
    config = merge_global_config(global_config)
    mode = _normalize_mapping_mode(node.address_mapping_mode)
    global_rule = _global_mapping_rule(config, node)
    if mode == "global":
        primary_role = _normalize_address_role(str(global_rule.get("primary_address_role") or _default_primary_role(node)))
        backup_role = _normalize_address_role(str(global_rule.get("backup_address_role") or _default_backup_role(node)))
        ssh_source = str(global_rule.get("ssh_source") or "primary_address")
    else:
        primary_role = _normalize_address_role(node.primary_address_role or _default_primary_role(node))
        backup_role = _normalize_address_role(node.backup_address_role or _default_backup_role(node))
        ssh_source = "primary_address"
    has_address_source = bool(node.primary_address or node.backup_address)
    ip_vehicle = "" if overwrite and has_address_source else node.ip_vehicle
    ip_uplink = "" if overwrite and has_address_source else node.ip_uplink
    ssh_host = "" if overwrite and has_address_source else node.ssh_host
    ip_vehicle, ip_uplink, ssh_host = _apply_address_role(node.primary_address, primary_role, ip_vehicle, ip_uplink, ssh_host)
    ip_vehicle, ip_uplink, ssh_host = _apply_address_role(
        node.backup_address,
        backup_role,
        ip_vehicle,
        ip_uplink,
        ssh_host,
        overwrite_existing=primary_role == "all",
    )
    if not ssh_host:
        ssh_host = _ssh_from_source(ssh_source, node, ip_vehicle, ip_uplink)
    if not ssh_host and _node_config_key(node) in {"MR", "3SW"}:
        ssh_host = node.primary_address or ip_vehicle or ip_uplink
    return replace(
        node,
        ip_vehicle=ip_vehicle,
        ip_uplink=ip_uplink,
        ssh_host=ssh_host,
        primary_address_role=primary_role,
        backup_address_role=backup_role,
        address_mapping_mode=mode,
    )


def _preserve_manual_mapping(old: CarNetworkNode, new: CarNetworkNode, train: CarNetworkTrain | None) -> CarNetworkNode:
    mode = _normalize_mapping_mode(old.address_mapping_mode or new.address_mapping_mode)
    vrrp_ip = (old.vrrp_ip or new.vrrp_ip) if mode == "custom" else new.vrrp_ip
    remark = old.remark if mode == "custom" and not _legacy_remark(old.remark) else _normalized_node_remark(new.node_name, new.node_type, new.end)
    return replace(
        new,
        train_id=train.train_id if train is not None else new.train_id or old.train_id,
        train_no=train.train_no if train is not None else new.train_no or old.train_no,
        display_name=train.display_name if train is not None else new.display_name or old.display_name,
        ip_vehicle=old.ip_vehicle or new.ip_vehicle,
        ip_uplink=old.ip_uplink or new.ip_uplink,
        ssh_host=old.ssh_host or new.ssh_host,
        vrrp_ip=vrrp_ip,
        address_mapping_mode=mode,
        primary_address_role=old.primary_address_role or new.primary_address_role,
        backup_address_role=old.backup_address_role or new.backup_address_role,
        remark=remark,
    )


def _normalize_mapping_mode(value: str) -> str:
    if value == "manual":
        return "custom"
    if value == "auto" or not value:
        return "global"
    return "custom" if value == "custom" else "global"


def _normalize_address_role(value: str) -> str:
    text = str(value or "").strip()
    normalized = text.casefold()
    if text == "全部" or normalized in {"all", "all_addresses"}:
        return "all"
    if text == "车内IP":
        return "vehicle_ip"
    if text == "落地IP":
        return "uplink_ip"
    if text == "SSH地址":
        return "ssh_host"
    if text == "忽略":
        return "ignore"
    if normalized in {"vehicle_ip", "uplink_ip", "ssh_host", "ignore"}:
        return normalized
    return "ignore"


def _apply_address_role(
    address: str,
    role: str,
    ip_vehicle: str,
    ip_uplink: str,
    ssh_host: str,
    *,
    overwrite_existing: bool = False,
) -> tuple[str, str, str]:
    if not address:
        return ip_vehicle, ip_uplink, ssh_host
    role = _normalize_address_role(role)
    if role == "all":
        return address, address, address
    if role == "vehicle_ip" and (overwrite_existing or not ip_vehicle):
        ip_vehicle = address
    elif role == "uplink_ip" and (overwrite_existing or not ip_uplink):
        ip_uplink = address
    elif role == "ssh_host" and (overwrite_existing or not ssh_host):
        ssh_host = address
    return ip_vehicle, ip_uplink, ssh_host


def _normalized_node_remark(node_name: str, node_type: str, end: str = "") -> str:
    normalized = normalize_node_name(node_name)
    end_label = "CT" if normalized.startswith("TC1-") else "CW" if normalized.startswith("TC2-") else end
    if normalized.endswith("-MR") or str(node_type).upper() in {"MR", "AP", "FAT-AP", "CLOUD-AP"}:
        return f"{end_label} MR" if end_label else "MR"
    if normalized.endswith("-SW") or str(node_type).upper() in {"SW", "3SW", "L3 SWITCH"}:
        return "三层交换机"
    if normalized.endswith("-SRV") or str(node_type).upper() in {"SERVER", "SRV"}:
        return "车载服务器"
    return str(node_type or normalized)


def _legacy_remark(value: str) -> bool:
    text = str(value or "")
    return text in {"车载AP（主）", "车载AP（备）", "车载AP(主)", "车载AP(备)", "AP（主）", "AP（备）", "AP(主)", "AP(备)"}


def _node_config_key(node: CarNetworkNode) -> str:
    value = node.node_type.upper()
    if value in {"SW", "3SW", "L3 SWITCH"} or node.node_name.endswith("-SW"):
        return "3SW"
    if value in {"SERVER", "SRV"} or node.node_name.endswith("-SRV"):
        return "SRV"
    return "MR" if value in {"MR", "AP", "FAT-AP", "CLOUD-AP"} or node.node_name.endswith("-MR") else value


def _global_mapping_rule(config: dict[str, object], node: CarNetworkNode) -> dict[str, object]:
    mapping = config.get("address_mapping")
    if isinstance(mapping, dict):
        rule = mapping.get(_node_config_key(node))
        if isinstance(rule, dict):
            return rule
    return {}


def _ssh_from_source(source: str, node: CarNetworkNode, ip_vehicle: str, ip_uplink: str) -> str:
    return {
        "primary_address": node.primary_address,
        "backup_address": node.backup_address,
        "ip_vehicle": ip_vehicle,
        "ip_uplink": ip_uplink,
        "empty": "",
    }.get(source, "")


def _srv_generation(config: dict[str, object]) -> dict[str, object]:
    value = config.get("srv_generation")
    return value if isinstance(value, dict) else {}


def _default_primary_role(node: CarNetworkNode) -> str:
    if node.node_type.upper() in {"MR", "SW", "3SW", "SERVER", "SRV"}:
        return "vehicle_ip"
    return "ignore"


def _default_backup_role(node: CarNetworkNode) -> str:
    if node.node_type.upper() in {"MR", "SW", "3SW"} and node.backup_address:
        return "uplink_ip"
    return "ignore"


def _is_vehicle_group(group_name: str) -> bool:
    return str(group_name or "").startswith("车载")


def _classify_vehicle_device(device: Device, group_name: str) -> tuple[str, str] | None:
    group_text = str(group_name or "")
    text = " ".join(str(value or "") for value in (device.name, device.system_name, device.station, device.device_type, device.remark))
    upper = text.upper()
    type_upper = str(device.device_type or "").upper()
    if "车载-MR" in group_text or type_upper in {"CLOUD-AP", "FAT-AP", "AP", "MR"} or "MR" in upper or ("AP" in upper and "车载-MR" in group_text):
        return "MR", "MR"
    if "车载-3SW" in group_text or "车载-交换机" in group_text or any(token in upper for token in ("3SW", "TC1-SW", "TC2-SW", "L3 SWITCH", "SWITCH")) or "三层交换机" in text or "交换机" in text:
        return "SW", "SW"
    if "车载-SRV" in group_text or "车载-服务器" in group_text or any(token in upper for token in ("SRV", "SERVER")) or "服务器" in text:
        return "Server", "SRV"
    if group_text == "车载":
        return "MR", "MR"
    return None


def _device_train_no_any(device: Device) -> str:
    return normalize_train_no(" ".join(str(value or "") for value in (device.name, device.system_name, device.station, device.remark)))


def _device_train_id(device: Device, train_no: str) -> str:
    text = canonical_peer_name(" ".join(str(value or "") for value in (device.name, device.system_name, device.station, device.remark)))
    match = re.search(r"(?i)(?P<prefix>[^\s,;]*?LC\s*0*(?P<train_no>\d{1,3}))", text)
    if match and normalize_train_no(match.group("train_no")) == train_no:
        return re.sub(r"\s+", "", match.group("prefix"))
    return _train_id_for_no(train_no, [])


def _train_id_for_no(train_no: str, nodes: list[CarNetworkNode]) -> str:
    for node in nodes:
        if node.train_no == train_no and node.train_id:
            return node.train_id
    return train_no


def _vehicle_prefix_from_addresses(addresses: Iterable[str], train_no: str = "") -> str:
    prefixes = []
    for address in addresses:
        prefix = _ip_prefix(str(address or ""))
        if prefix:
            prefixes.append(prefix)
    if not prefixes:
        return ""
    counts: dict[str, int] = {}
    for prefix in prefixes:
        counts[prefix] = counts.get(prefix, 0) + 1
    return sorted(counts, key=lambda value: (-counts[value], value))[0]


def _device_tc_end(device: Device) -> tuple[str, str]:
    text = " ".join(str(value or "") for value in (device.name, device.system_name, device.station, device.remark))
    upper = text.upper()
    if "TC1" in upper or " CT" in f" {upper} " or "-CT" in upper or "车头" in text:
        return "TC1", "CT"
    if "TC2" in upper or " CW" in f" {upper} " or "-CW" in upper or "车尾" in text:
        return "TC2", "CW"
    return "", ""


def infer_vehicle_prefix(nodes: list[CarNetworkNode], train_no: str = "") -> str:
    candidates: list[str] = []
    for node_name in ("TC1-SW", "TC2-SW", "TC1-MR", "TC2-MR"):
        candidates.extend(node.ip_vehicle for node in nodes if node.node_name == node_name and node.ip_vehicle)
    candidates.extend(node.ip_vehicle for node in nodes if node.ip_vehicle)
    candidates.extend(node.primary_address for node in nodes if node.primary_address_role == "vehicle_ip" and node.primary_address)
    for ip in candidates:
        prefix = _ip_prefix(ip)
        if prefix:
            return prefix
    return _fallback_vehicle_prefix(train_no)


def _apply_generated_network_defaults(
    node: CarNetworkNode,
    prefix: str,
    config: dict[str, object],
    *,
    overwrite_vrrp: bool = True,
) -> CarNetworkNode:
    srv = _srv_generation(config)
    vrrp_ip = node.vrrp_ip
    if prefix and (overwrite_vrrp or not vrrp_ip):
        vrrp_ip = _build_host_ip(prefix, srv.get("vrrp_host", 254))
    ip_vehicle = node.ip_vehicle
    if prefix and _node_config_key(node) == "SRV" and srv.get("enabled", True) and not ip_vehicle:
        host = srv.get("tc1_host", 1) if node.tc == "TC1" else srv.get("tc2_host", 2)
        ip_vehicle = _build_host_ip(prefix, host)
    return replace(node, ip_vehicle=ip_vehicle, vrrp_ip=vrrp_ip)


def _ip_prefix(ip: str) -> str:
    match = re.match(r"^(\d{1,3}\.\d{1,3}\.\d{1,3})\.\d{1,3}$", str(ip or "").strip())
    return match.group(1) if match else ""


def _fallback_vehicle_prefix(train_no: str | None) -> str:
    return ""


def _build_host_ip(prefix: str, host: object) -> str:
    try:
        host_int = int(host)
    except (TypeError, ValueError):
        return ""
    return f"{prefix}.{host_int}" if prefix else ""


def _find_existing_node(nodes: list[CarNetworkNode], node: CarNetworkNode) -> CarNetworkNode | None:
    if node.device_id:
        found = next((item for item in nodes if item.device_id == node.device_id and item.node_name == node.node_name), None)
        if found is not None:
            return found
    return next((item for item in nodes if item.train_no == node.train_no and item.node_name == node.node_name), None)


def _dedupe_nodes(nodes: list[CarNetworkNode]) -> list[CarNetworkNode]:
    result: dict[tuple[str, str], CarNetworkNode] = {}
    for node in nodes:
        result[(node.train_no or node.train_id, node.node_name)] = node
    return list(result.values())


def _sort_nodes(nodes: list[CarNetworkNode]) -> list[CarNetworkNode]:
    order = {name: index for index, name in enumerate(NODE_ORDER)}
    return sorted(nodes, key=lambda node: (get_train_sort_key(node), order.get(node.node_name, 99), node.node_name))


def car_network_root(paths: PathResolver, site_name: str) -> Path:
    root = paths.car_network_diagnostic_parsed_dir(site_name)
    root.mkdir(parents=True, exist_ok=True)
    return root


class CarNetworkPointTableStore:
    def __init__(self, paths: PathResolver, site_name: str) -> None:
        self.root = car_network_root(paths, site_name)
        self.path = self.root / "point_table.json"

    def load(self) -> list[CarNetworkNode]:
        if not self.path.exists():
            return []
        data = json.loads(self.path.read_text(encoding="utf-8"))
        return _sort_nodes([node_from_mapping(row) for row in data])

    def save(self, nodes: Iterable[CarNetworkNode]) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.path.write_text(json.dumps([asdict(node) for node in _sort_nodes(list(nodes))], ensure_ascii=False, indent=2), encoding="utf-8")

    def import_file(self, path: Path) -> int:
        nodes = [node_from_mapping(row) for row in read_point_table_file(path)]
        self.save(nodes)
        return len(nodes)

    def export_file(self, path: Path, nodes: Iterable[CarNetworkNode] | None = None) -> None:
        write_point_table_file(path, _sort_nodes(list(nodes or self.load())))


class CarNetworkGlobalConfigStore:
    def __init__(self, paths: PathResolver, site_name: str) -> None:
        self.root = car_network_root(paths, site_name)
        self.path = self.root / "global_config.json"

    def load(self) -> dict[str, object]:
        if not self.path.exists():
            return _copy_global_config(DEFAULT_GLOBAL_CONFIG)
        try:
            data = json.loads(self.path.read_text(encoding="utf-8"))
        except Exception:
            return _copy_global_config(DEFAULT_GLOBAL_CONFIG)
        return merge_global_config(data)

    def save(self, config: dict[str, object]) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.path.write_text(json.dumps(merge_global_config(config), ensure_ascii=False, indent=2), encoding="utf-8")


def merge_global_config(config: dict[str, object] | None) -> dict[str, object]:
    merged = _copy_global_config(DEFAULT_GLOBAL_CONFIG)
    if not isinstance(config, dict):
        return merged
    mapping = config.get("address_mapping")
    if isinstance(mapping, dict):
        target = merged["address_mapping"]
        assert isinstance(target, dict)
        for node_type, values in mapping.items():
            if isinstance(values, dict):
                current = target.setdefault(str(node_type), {})
                if isinstance(current, dict):
                    for key, value in values.items():
                        text = str(value)
                        if str(key) in {"primary_address_role", "backup_address_role"}:
                            text = _normalize_address_role(text)
                        current[str(key)] = text
    srv = config.get("srv_generation")
    if isinstance(srv, dict):
        target_srv = merged["srv_generation"]
        assert isinstance(target_srv, dict)
        target_srv.update(srv)
    if "point_table_locked" in config:
        merged["point_table_locked"] = bool(config.get("point_table_locked"))
    return merged


def _copy_global_config(config: dict[str, object]) -> dict[str, object]:
    return json.loads(json.dumps(config, ensure_ascii=False))


def node_from_mapping(row: dict[str, object]) -> CarNetworkNode:
    source = dict(row)
    if "role" in source and "tc" not in source:
        source["tc"] = source.get("role")
    if "ip" in source and "ip_vehicle" not in source:
        source["ip_vehicle"] = source.get("ip")
    if "backup_ip" in source and "ip_uplink" not in source:
        source["ip_uplink"] = source.get("backup_ip")
    data = {field: str(source.get(field, "") or "") for field in POINT_TABLE_FIELDS}
    data["node_name"] = normalize_node_name(data["node_name"])
    if data["node_type"].upper() == "AP":
        data["node_type"] = "MR"
    if not data["train_no"]:
        data["train_no"] = normalize_train_no(data["train_id"])
    if not data["display_name"]:
        data["display_name"] = f"{data['train_no']}车" if data["train_no"] else data["train_id"]
    if not data["tc"]:
        data["tc"] = data["node_name"].split("-", 1)[0]
    if not data["end"]:
        data["end"] = "CT" if data["tc"] == "TC1" else "CW" if data["tc"] == "TC2" else ""
    if data["address_mapping_mode"] == "auto":
        data["address_mapping_mode"] = "global"
    elif data["address_mapping_mode"] == "manual":
        data["address_mapping_mode"] = "custom"
    data["primary_address_role"] = _normalize_address_role(data["primary_address_role"])
    data["backup_address_role"] = _normalize_address_role(data["backup_address_role"])
    node = CarNetworkNode(**data)
    return apply_address_mapping(node) if node.address_mapping_mode != "custom" else node


def read_point_table_file(path: Path) -> list[dict[str, object]]:
    if path.suffix.casefold() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as file:
            return [dict(row) for row in csv.DictReader(file)]
    workbook = load_workbook_without_unsupported_image_warning(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    return [{headers[index]: value for index, value in enumerate(values) if index < len(headers)} for values in rows[1:]]


def write_point_table_file(path: Path, nodes: list[CarNetworkNode]) -> None:
    if path.suffix.casefold() == ".csv":
        with path.open("w", encoding="utf-8-sig", newline="") as file:
            writer = csv.DictWriter(file, fieldnames=list(POINT_TABLE_FIELDS))
            writer.writeheader()
            writer.writerows(asdict(node) for node in nodes)
        return
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "car_network_point_table"
    sheet.append(list(POINT_TABLE_FIELDS))
    for node in nodes:
        sheet.append([getattr(node, field) for field in POINT_TABLE_FIELDS])
    widths = (14, 10, 12, 8, 8, 14, 12, 10, 24, 14, 16, 16, 16, 16, 16, 16, 16, 18, 18, 18, 32)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    workbook.save(path)


class CarNetworkDiagnosticService:
    def __init__(
        self,
        nodes: list[CarNetworkNode],
        *,
        train: CarNetworkTrain | None = None,
        mr_devices: dict[str, Device] | None = None,
        ac_device: Device | None = None,
        ac_devices: list[Device] | None = None,
        core_devices: list[Device] | None = None,
        paths: PathResolver | None = None,
        site_name: str = "",
        ping_func: Callable[[str], PingResult] | None = None,
        ac_command_func: Callable[[str], str] | None = None,
        ssh_command_func: Callable[[str, str], str] | None = None,
        core_command_func: Callable[[Device, str], str] | None = None,
        core_discovery: dict[str, object] | None = None,
        cancel_checker: Callable[[], bool] | None = None,
    ) -> None:
        self.nodes = nodes
        self.train = train
        self.mr_devices = mr_devices or (train.mr_devices if train is not None else {})
        self.ac_device = ac_device
        self.ac_devices = list(ac_devices or ([ac_device] if ac_device is not None else []))
        self.core_devices = list(core_devices or [])
        self.paths = paths
        self.site_name = site_name
        self.ping_func = ping_func or _run_local_aux_ping
        self.ac_command_func = ac_command_func
        self.ssh_command_func = ssh_command_func
        self.core_command_func = core_command_func
        self.core_discovery = core_discovery or _core_discovery_from_selected(self.core_devices)
        self.cancel_checker = cancel_checker

    def run(self, progress: Callable[[str, object], None] | None = None) -> CarNetworkDiagnosticResult:
        mr_nodes = [node for node in self.nodes if node.is_mr]
        remote_ping_count = sum(len(_mr_remote_ping_tasks(node, self.nodes, self._mr_ssh_host(node))) for node in mr_nodes if self._mr_ssh_host(node))
        mr_ssh_count = len([node for node in mr_nodes if self._mr_ssh_host(node)])
        core_ping_count = len(_core_remote_ping_tasks(self.core_devices, self.nodes))
        total_tasks = 1 + mr_ssh_count + remote_ping_count + 1 + core_ping_count + 1
        completed_tasks = 0

        def emit(stage: str, payload: object) -> None:
            if progress:
                progress(stage, payload)

        def finish_task(payload: dict[str, object]) -> None:
            nonlocal completed_tasks
            completed_tasks += 1
            payload["completed"] = completed_tasks
            payload["total"] = total_tasks
            emit("task_finished", payload)
            emit("progress_meta", {"completed": completed_tasks, "total": total_tasks, "percent": int(completed_tasks * 100 / max(1, total_tasks)), "message": payload.get("message", "")})

        emit("progress_meta", {"completed": 0, "total": total_tasks, "percent": 0, "message": "准备检测"})
        emit("stage", "获取列车在线情况 / AC mesh-link")
        self._raise_if_cancelled()
        emit("task_started", {"task_id": "ac_status", "layer": "AC层", "status": "running", "message": "正在获取列车在线情况或查询 AC mesh-link"})
        train_ac_status = get_vehicle_mr_online_status_from_store(self.paths, self.site_name, _current_train_no(self.train, self.nodes))
        if train_ac_status is not None:
            ac_probe = AcProbeResult(enabled=True, query_success=True)
        else:
            ac_probe = self._probe_ac_mesh_links()
            train_ac_status = match_train_mr_in_mesh_links(self.train, self.nodes, ac_probe)
        finish_task({"task_id": "ac_status", "layer": "AC层", "status": "ok" if train_ac_status.tc1_mr_online or train_ac_status.tc2_mr_online else "fail" if train_ac_status.both_mr_offline else "unknown", "message": _ac_task_message(train_ac_status)})
        if progress:
            progress("ac_probe", (ac_probe, train_ac_status))
        if should_skip_deep_probe_by_ac(train_ac_status):
            emit("stage", "检测完成")
            result = build_ac_offline_result(self.nodes, ac_probe, train_ac_status, self.train)
            result.core_discovery = self.core_discovery
            return result

        ping_results: dict[str, PingResult] = {}
        ac_status = ac_status_from_probe(ac_probe, train_ac_status)
        if progress:
            progress("ac", ac_status)

        ssh_results: dict[str, SshResult] = {}
        emit("stage", "MR SSH 登录与4秒快速检测")
        with ThreadPoolExecutor(max_workers=max(1, min(MR_REMOTE_PING_CONCURRENCY_TOTAL, len(mr_nodes)))) as executor:
            future_map = {executor.submit(self._check_ssh_from_mr, node, emit, finish_task): node for node in mr_nodes if self._mr_ssh_host(node)}
            for future in as_completed(future_map):
                self._raise_if_cancelled()
                result = future.result()
                ssh_results[result.node_name or result.host] = result
                if progress:
                    progress("ssh", result)

        emit("stage", "跨TC通信丢包检测")
        cross_tc_ping = self._check_cross_tc_ping(ssh_results, emit, finish_task)
        if progress:
            progress("cross_tc_ping", cross_tc_ping)

        emit("stage", "核心侧辅助 ping")
        core_results = self._check_from_core(progress, emit, finish_task)
        emit("stage", "生成诊断结论")
        result = evaluate_diagnostic(self.nodes, ping_results, ssh_results, ac_status, self.train, ac_probe, train_ac_status, core_results, self.core_devices, cross_tc_ping)
        result.core_discovery = self.core_discovery
        finish_task({"task_id": "summary", "layer": "诊断汇总", "status": result.status, "message": result.conclusion})
        emit("stage", "检测完成")
        return result


    def _raise_if_cancelled(self) -> None:
        if self.cancel_checker is not None and self.cancel_checker():
            raise RuntimeError("检测已取消")

    def _check_from_core(
        self,
        progress: Callable[[str, object], None] | None = None,
        emit: Callable[[str, object], None] | None = None,
        finish_task: Callable[[dict[str, object]], None] | None = None,
    ) -> dict[str, PingResult]:
        tasks = _core_remote_ping_tasks(self.core_devices, self.nodes)
        results: dict[str, PingResult] = {}
        devices_by_task_id = {_core_remote_ping_task_id(device, node.ip_uplink): device for device in self.core_devices for node in self.nodes if node.ip_uplink and node.node_type.upper() in {"MR", "SW", "3SW"}}
        for task in tasks:
            self._raise_if_cancelled()
            if emit:
                emit("task_started", _core_ping_task_payload(task, None, "running"))
            device = devices_by_task_id.get(task.task_id)
            if device is None:
                continue
            key = f"{task.core_device_id}->{task.target_ip}"
            try:
                output = self._send_core_ping(device, task.target_ip)
                result = parse_ping_output(task.target_ip, output)
            except Exception as exc:
                result = PingResult(task.target_ip, False, error=_clean_ping_error(str(exc)))
            results[key] = result
            if finish_task:
                finish_task(_core_ping_task_payload(task, result, _ping_task_status(result)))
            if progress:
                progress("core", (device, result))
        return results

    def _check_cross_tc_ping(
        self,
        ssh_results: dict[str, SshResult],
        emit: Callable[[str, object], None] | None = None,
        finish_task: Callable[[dict[str, object]], None] | None = None,
    ) -> dict[str, object]:
        task = _cross_tc_ping_task(self.nodes, ssh_results)
        if task is None:
            payload = {
                "status": "skipped",
                "source": "",
                "target": "",
                "target_ip": "",
                "loss_percent": None,
                "avg_rtt_ms": None,
                "command": "",
                "note": "无可登录 MR，跨TC通信未检测",
            }
            if finish_task:
                finish_task({"task_id": "cross_tc_ping", "layer": "跨TC通信", "status": "skipped", "message": payload["note"]})
            return payload
        if emit:
            emit("task_started", {"task_id": "cross_tc_ping", "source": task.source_node, "target": task.target_node, "target_ip": task.target_ip, "layer": "跨TC通信", "status": "running", "message": f"{task.source_node} 正在执行 {task.command}，目标 {task.target_node} / {task.target_ip}"})
        self._raise_if_cancelled()
        source_node = next((node for node in self.nodes if node.node_name == task.source_node), None)
        ssh = ssh_results.get(task.source_node)
        if source_node is None or ssh is None:
            ping = PingResult(task.target_ip, False, 100.0, error="跨TC ping 入口不存在")
        else:
            username, password = self._mr_credentials(source_node)
            try:
                ping = self._run_mr_remote_ping_task(task, username, password, source_node)
            except Exception as exc:
                ping = PingResult(task.target_ip, False, 100.0, error=str(exc))
            task_results = dict(ssh.task_results)
            task_metadata = dict(ssh.task_metadata)
            command_results = dict(ssh.command_results)
            task_results[task.task_id] = ping
            task_metadata[task.task_id] = asdict(task)
            command_results[task.target_ip] = ping
            ssh_results[task.source_node] = replace(ssh, command_results=command_results, task_results=task_results, task_metadata=task_metadata)
        status = "ok" if _ping_ok_no_loss(ping) else "loss" if ping.ok and ping.loss_percent > 0 else "fail"
        note = _ping_result_message(task.source_node, task.target_node, task.target_ip, ping, via_ssh=True, command=task.command)
        if finish_task:
            finish_task({"task_id": "cross_tc_ping", "source": task.source_node, "target": task.target_node, "target_ip": task.target_ip, "layer": "跨TC通信", "status": "unstable" if status == "loss" else status, "loss": ping.loss_percent, "avg_rtt": ping.avg_rtt_ms, "message": note})
        return {
            "status": status,
            "source": task.source_node,
            "target": task.target_node,
            "target_ip": task.target_ip,
            "loss_percent": ping.loss_percent,
            "avg_rtt_ms": ping.avg_rtt_ms,
            "command": task.command,
            "note": note,
        }

    def _send_core_ping(self, device: Device, ip: str) -> str:
        if self.core_command_func is not None:
            return self.core_command_func(device, ip)
        if ConnectHandler is None:
            raise RuntimeError("netmiko is not installed")
        targets = connection_targets(device)
        if targets:
            target = build_netmiko_params(targets[0])
        else:
            username = device.ssh_username or device.username or ""
            password = device.ssh_password or device.password or ""
            if not device.primary_address or not username or not password:
                raise RuntimeError("核心交换机连接信息不完整")
            from netconsole.services.netmiko_connection import ConnectionTarget

            target = build_netmiko_params(
                ConnectionTarget("SSH", H3C_NETMIKO_DEVICE_TYPE, device.primary_address, int(device.ssh_port or 22), username, password, encoding_for_vendor(device.device_vendor))
            )
        conn = ConnectHandler(**target)
        try:
            return safe_send_command(
                conn,
                build_h3c_ping_command(ip, packet_count=CAR_NETWORK_QUICK_PING_COUNT),
                read_timeout=CAR_NETWORK_QUICK_PING_TIMEOUT,
                strip_prompt=False,
                strip_command=False,
                use_timing=True,
                encoding=str(target.get("encoding") or "gb2312"),
            )
        finally:
            conn.disconnect()

    def _check_ac(self) -> AcApStatus:
        if self.ac_command_func is None and self.ac_device is None:
            return AcApStatus(selected=False, error="未选择 AC")
        raw: dict[str, str] = {}
        try:
            if self.ac_command_func is not None:
                for command in AC_COMMANDS:
                    raw[command] = self.ac_command_func(command)
            else:
                assert self.ac_device is not None
                raw = run_ac_commands(self.ac_device, AC_COMMANDS)
        except Exception as exc:
            return AcApStatus(selected=True, error=str(exc), raw=raw)
        train_no = _current_train_no(self.train, self.nodes)
        mesh_status = match_train_mr_in_mesh_links(
            self.train,
            self.nodes,
            AcProbeResult(True, True, [AcControllerProbe("", "manual", "", True, raw.get("display wlan mesh-link ap", ""))]),
        )
        mesh = mesh_status.tc1_mr_online or mesh_status.tc2_mr_online
        ap_all = _output_contains_train_mr_peer(raw.get("display wlan ap all", ""), train_no)
        radio_text = raw.get("display wlan ap all radio", "")
        radio_ok = None if not radio_text.strip() else not re.search(r"\b(down|fault|disable)\b", radio_text, re.IGNORECASE)
        return AcApStatus(mesh, ap_all, radio_ok, True, raw)

    def _probe_ac_mesh_links(self) -> AcProbeResult:
        if self.ac_command_func is not None:
            return probe_ac_mesh_links(self.ac_devices, command_func=self.ac_command_func)
        if not self.ac_devices:
            return AcProbeResult(enabled=True, query_success=False, error="未发现无线控制器")
        return probe_ac_mesh_links(self.ac_devices, command_func=self.ac_command_func)

    def _check_ssh_from_mr(
        self,
        node: CarNetworkNode,
        progress: Callable[[str, object], None] | None = None,
        finish_task: Callable[[dict[str, object]], None] | None = None,
    ) -> SshResult:
        def emit(stage: str, payload: object) -> None:
            if progress:
                progress(stage, payload)

        host = self._mr_ssh_host(node)
        if not host:
            return SshResult("", False, node.node_name, error="无 SSH 主机")
        username, password = self._mr_credentials(node)
        if not username or not password:
            return SshResult(host, False, node.node_name, error="跳过：未配置MR SSH账号密码")
        emit("task_started", {"task_id": f"{node.node_name}_ssh", "source": node.node_name, "target_ip": host, "layer": "SSH层", "status": "running", "message": f"正在登录 {node.node_name} / {host}"})
        ssh_finished = False
        try:
            command_results: dict[str, PingResult] = {}
            task_results: dict[str, PingResult] = {}
            task_metadata: dict[str, dict[str, object]] = {}
            tasks = _mr_remote_ping_tasks(node, self.nodes, host)
            if not tasks and finish_task:
                finish_task({"task_id": f"{node.node_name}_ssh", "source": node.node_name, "target_ip": host, "layer": "SSH层", "status": "ok", "message": f"{node.node_name} SSH登录成功"})
                ssh_finished = True
            with ThreadPoolExecutor(max_workers=max(1, min(MR_REMOTE_PING_CONCURRENCY_PER_MR, len(tasks)))) as executor:
                future_map = {}
                for task in tasks:
                    self._raise_if_cancelled()
                    emit("task_started", {"task_id": task.task_id, "source": task.source_node, "target": task.target_node, "target_ip": task.target_ip, "layer": task.layer, "status": "running", "message": f"{task.source_node} 正在执行 {task.command}，目标 {task.target_node} / {task.target_ip}"})
                    future_map[executor.submit(self._run_mr_remote_ping_task, task, username, password, node)] = task
                first_error: Exception | None = None
                for future in as_completed(future_map):
                    self._raise_if_cancelled()
                    task = future_map[future]
                    try:
                        ping = future.result()
                    except Exception as exc:
                        if first_error is None:
                            first_error = exc
                        ping = PingResult(task.target_ip, False, 100.0, error=str(exc))
                    else:
                        if finish_task and not ssh_finished:
                            finish_task({"task_id": f"{node.node_name}_ssh", "source": node.node_name, "target_ip": host, "layer": "SSH层", "status": "ok", "message": f"{node.node_name} SSH登录成功"})
                            ssh_finished = True
                    command_results[task.target_ip] = ping
                    task_results[task.task_id] = ping
                    task_metadata[task.task_id] = asdict(task)
                    if finish_task:
                        finish_task(_remote_ping_task_payload(task, ping))
                if not ssh_finished and first_error is not None:
                    raise first_error
            return SshResult(host, True, node.node_name, command_results, task_results, task_metadata)
        except Exception as exc:
            if finish_task and not ssh_finished:
                finish_task({"task_id": f"{node.node_name}_ssh", "source": node.node_name, "target_ip": host, "layer": "SSH层", "status": "fail", "message": f"{node.node_name} SSH登录失败：{exc}"})
            return SshResult(host, False, node.node_name, error=str(exc))

    def _run_mr_remote_ping_task(self, task: MrRemotePingTask, username: str, password: str, node: CarNetworkNode) -> PingResult:
        self._raise_if_cancelled()
        output = self._send_ssh_command(task.source_host, task.command, username, password, node, task.timeout)
        return parse_ping_output(task.target_ip, output)

    def _send_ssh_command(self, host: str, command: str, username: str, password: str, node: CarNetworkNode, read_timeout: int = 30) -> str:
        if self.ssh_command_func is not None:
            return self.ssh_command_func(host, command)
        if ConnectHandler is None:
            raise RuntimeError("netmiko is not installed")
        device = self.mr_devices.get(node.node_name)
        target = {
            "device_type": H3C_NETMIKO_DEVICE_TYPE,
            "host": host,
            "username": username,
            "password": password,
            "port": int(device.ssh_port if device is not None and device.ssh_port else 22),
            "timeout": 8,
            "conn_timeout": 8,
            "auth_timeout": 8,
            "banner_timeout": 8,
            "encoding": encoding_for_vendor(device.device_vendor if device else "H3C"),
            "session_log": None,
            "global_delay_factor": 1,
        }
        conn = ConnectHandler(**target)
        try:
            if _is_h3c_ping_command(command):
                return _send_h3c_ping_command(conn, command, str(target["encoding"]), _h3c_ping_read_timeout(command, read_timeout))
            return safe_send_command(
                conn,
                command,
                read_timeout=read_timeout,
                strip_prompt=False,
                strip_command=False,
                use_timing=True,
                encoding=str(target["encoding"]),
            )
        finally:
            conn.disconnect()

    def _mr_ssh_host(self, node: CarNetworkNode) -> str:
        return node.ssh_address(self.mr_devices.get(node.node_name))

    def _mr_credentials(self, node: CarNetworkNode) -> tuple[str, str]:
        device = self.mr_devices.get(node.node_name)
        username = ""
        password = ""
        if device is not None:
            username = device.ssh_username or device.username or ""
            password = device.ssh_password or device.password or ""
        return username, password


def evaluate_diagnostic(
    nodes: list[CarNetworkNode],
    ping_results: dict[str, PingResult],
    ssh_results: dict[str, SshResult],
    ac_status: AcApStatus,
    train: CarNetworkTrain | None = None,
    ac_probe: AcProbeResult | None = None,
    train_ac_status: TrainAcStatus | None = None,
    core_results: dict[str, PingResult] | None = None,
    core_devices: list[Device] | None = None,
    cross_tc_ping: dict[str, object] | None = None,
) -> CarNetworkDiagnosticResult:
    train_id = train.train_id if train is not None else (nodes[0].train_id if nodes else "")
    train_no = train.train_no if train is not None else (nodes[0].train_no if nodes else normalize_train_no(train_id))
    display_name = train.display_name if train is not None else (f"{train_no}车" if train_no else train_id)
    core_results = core_results or {}
    cross_tc_ping = cross_tc_ping or _cross_tc_ping_from_results(nodes, ssh_results)
    node_states = {node.normalized_name: _node_state(node, ping_results, ssh_results, ac_status, train_ac_status, nodes) for node in nodes}
    cross = _cross_status(nodes, ssh_results)
    ssh_ok_count = sum(1 for item in ssh_results.values() if item.ok)
    ssh_attempted = [item for item in ssh_results.values() if not _ssh_skipped(item)]
    ssh_status = "ok" if ssh_ok_count >= len([n for n in nodes if n.is_mr and n.ssh_address()]) and ssh_ok_count > 0 else "partial" if ssh_ok_count else "skipped" if ssh_results and not ssh_attempted else "fail"
    ac_text = "ok" if ac_status.online else "partial" if ac_status.radio_ok else "unknown" if not ac_status.selected else "fail"

    mr_nodes = [node for node in nodes if node.is_mr]
    remote_ping_values = [ping for ssh in ssh_results.values() for ping in (ssh.task_results or ssh.command_results).values()]
    core_ping_values = list(core_results.values())
    probe_values = remote_ping_values + core_ping_values + list(ping_results.values())
    all_ping_failed = not probe_values or all(not ping.ok for ping in probe_values)
    all_ssh_failed = not ssh_results or (bool(ssh_attempted) and ssh_ok_count == 0)
    no_ac_mr_online = not ac_status.online
    vehicle_loss = any(_vehicle_loss(ping) for ping in remote_ping_values) or any(_vehicle_loss(ping) for ping in ping_results.values())
    wired_bad = any(value in {"fail", "unstable"} for key, value in node_states.items() if not key.endswith("-MR"))
    mr_online_evidence = ac_status.online or ssh_ok_count > 0
    ssh_failed_but_mr_reachable = _ssh_failed_but_mr_reachable(mr_nodes, core_results, ssh_results)
    vehicle_internal_known = bool(remote_ping_values)
    vehicle_internal_ok = vehicle_internal_known and all(_ping_ok_no_loss(ping) for ping in remote_ping_values)
    cannot_execute_vehicle_probe = ssh_ok_count == 0 and not vehicle_internal_known
    peer_mr_issue = _peer_mr_issue(ssh_results)
    any_remote_fail = any(not _ping_ok_no_loss(ping) for ping in remote_ping_values)
    cross_tc_status = str(cross_tc_ping.get("status") or "")
    single_entry_verified = _single_entry_verified(nodes, ssh_results)
    dual_entry_verified = ssh_ok_count >= 2 and cross_tc_status == "ok"
    ground_aux_abnormal = (ac_status.selected and not ac_status.online) or any(not ping.ok for ping in core_ping_values)
    cross_vehicle_ok = cross_tc_status == "ok" and any(value == "ok" for value in cross.values())

    if cannot_execute_vehicle_probe:
        status = "partial_fail"
        conclusion = "无法从地面接入 MR 执行车内链路检测。可能原因：列车下电、MR射频关闭、地面到车载链路不可达或账号配置异常；不能直接判定车内网络故障。"
    elif vehicle_loss or cross_tc_status == "loss":
        status = "partial_fail"
        conclusion = "车内通信存在丢包"
    elif cross_tc_status == "fail":
        status = "partial_fail"
        conclusion = "车内通信异常"
    elif vehicle_internal_ok or single_entry_verified or cross_vehicle_ok:
        status = "ok"
        if ground_aux_abnormal:
            conclusion = "车内通信正常，地面/AC辅助状态异常"
        elif dual_entry_verified:
            conclusion = "车内通信正常（双端验证）"
        elif single_entry_verified or ssh_ok_count == 1:
            conclusion = "车内通信正常（单端激活）"
        else:
            conclusion = "车内通信正常（双端验证）"
    elif peer_mr_issue:
        status = "partial_fail"
        conclusion = peer_mr_issue
    elif any(value == "fail" for value in cross.values()):
        status = "partial_fail"
        conclusion = "跨TC链路 / VRRP / 中间骨干链路异常"
    elif any_remote_fail:
        status = "partial_fail"
        conclusion = "车内通信异常：MR CLI ping 车内设备存在不通路径，请查看分层检测结果"
    elif mr_online_evidence and wired_bad:
        status = "partial_fail"
        conclusion = "车内有线网络异常，重点检查 MR 到三层交换机、交换机到服务器链路"
    elif ssh_failed_but_mr_reachable:
        status = "partial_fail"
        conclusion = "MR管理SSH不可达或认证失败，但MR IP可达"
    elif all(state == "ok" for state in node_states.values()) and ssh_status in {"ok", "partial"} and (ac_status.online or not ac_status.selected):
        status = "ok"
        conclusion = "FULL_OK"
    else:
        status = "partial_fail" if any(value in {"fail", "unstable"} for value in node_states.values()) else "ok"
        conclusion = "存在异常，请查看分层检测结果" if status != "ok" else "FULL_OK"

    return CarNetworkDiagnosticResult(
        train_id=train_id,
        status=status,
        nodes=node_states,
        cross_train=cross,
        ac_status=ac_text,
        ssh_status=ssh_status,
        conclusion=conclusion,
        train_no=train_no,
        display_name=display_name,
        ends=_ends_json(nodes, node_states),
        vrrp=_vrrp_json(nodes, cross),
        ping_results=ping_results,
        core_results=core_results,
        ssh_results=ssh_results,
        ac_detail=ac_status,
        ac_probe=ac_probe,
        train_ac_status=train_ac_status,
        tables=build_result_tables(nodes, ping_results, ssh_results, ac_status, core_results, core_devices or [], train_ac_status),
        cross_tc_ping=cross_tc_ping,
    )


def build_ping_targets(nodes: list[CarNetworkNode]) -> list[str]:
    targets: list[str] = []
    for node in nodes:
        targets.extend(node.ping_ips)
    return list(dict.fromkeys(targets))


def _node_by_any_ip(nodes: list[CarNetworkNode], ip: str) -> CarNetworkNode | None:
    return next((node for node in nodes if ip in {node.primary_address, node.backup_address, node.ip_vehicle, node.ip_uplink, node.ssh_host}), None)


def _node_by_vehicle_ip(nodes: list[CarNetworkNode], ip: str) -> CarNetworkNode | None:
    return next((node for node in nodes if node.ip_vehicle == ip), None)


def _ping_task_status(result: PingResult) -> str:
    if result.ok and result.loss_percent == 0:
        return "ok"
    if result.ok and result.loss_percent > 0:
        return "unstable"
    return "fail"


def _ping_task_payload(node: CarNetworkNode | None, ip: str, layer: str, status: str, result: PingResult | None = None) -> dict[str, object]:
    name = node.node_name if node is not None else ip
    message = f"正在 ping {name} / {ip}"
    if result is not None:
        message = _ping_result_message("本机", name, ip, result)
    return {
        "task_id": f"local_ping_{ip}",
        "source": "本机",
        "target": name,
        "target_ip": ip,
        "layer": layer,
        "status": status,
        "loss": result.loss_percent if result is not None else None,
        "avg_rtt": result.avg_rtt_ms if result is not None else None,
        "message": message,
    }


def _remote_ping_task_payload(task: MrRemotePingTask, result: PingResult) -> dict[str, object]:
    return {
        "task_id": task.task_id,
        "source": task.source_node,
        "target": task.target_node,
        "target_ip": task.target_ip,
        "layer": task.layer,
        "status": _ping_task_status(result),
        "loss": result.loss_percent,
        "avg_rtt": result.avg_rtt_ms,
        "message": _ping_result_message(task.source_node, task.target_node, task.target_ip, result, via_ssh=True, command=task.command),
    }


def _core_ping_task_payload(task: CoreRemotePingTask, result: PingResult | None, status: str) -> dict[str, object]:
    core_name = _core_display_name(task.core_device_name)
    return {
        "task_id": task.task_id,
        "source": core_name,
        "target": task.target_node,
        "target_ip": task.target_ip,
        "layer": task.layer,
        "status": status,
        "loss": result.loss_percent if result is not None else None,
        "avg_rtt": result.avg_rtt_ms if result is not None else None,
        "message": f"{core_name} 正在执行 {task.command}，目标 {task.target_node} / {task.target_ip}" if result is None else _core_ping_note(task, result),
    }


def _core_display_name(name: str) -> str:
    text = str(name or "").strip()
    if not text:
        return "核心交换机"
    if "COCC核心交换机" in text:
        return text.replace("COCC核心交换机", "核心交换机").strip() or "核心交换机"
    return text


def _ping_result_message(source: str, target: str, target_ip: str, result: PingResult, via_ssh: bool = False, command: str = "") -> str:
    if via_ssh:
        return _h3c_ping_note(f"{source} CLI", command or build_h3c_ping_command(target_ip), result)
    return _h3c_ping_summary(result)


def _is_h3c_ping_command(command: str) -> bool:
    return bool(re.match(r"^\s*ping(?:\s|$)", str(command or ""), re.IGNORECASE))


def _h3c_ping_read_timeout(command: str, requested_timeout: int) -> int:
    match = re.search(r"(?:^|\s)-c\s+(\d+)(?:\s|$)", str(command or ""), re.IGNORECASE)
    packet_count = int(match.group(1)) if match else CAR_NETWORK_QUICK_PING_COUNT
    if packet_count <= CAR_NETWORK_QUICK_PING_COUNT:
        return max(requested_timeout, CAR_NETWORK_QUICK_CLI_READ_TIMEOUT)
    return max(requested_timeout, packet_count + 10)


def _send_h3c_ping_command(conn: object, command: str, encoding: str, read_timeout: int) -> str:
    if not hasattr(conn, "write_channel") or not hasattr(conn, "read_channel"):
        return safe_send_command(
            conn,
            command,
            read_timeout=read_timeout,
            strip_prompt=False,
            strip_command=False,
            use_timing=True,
            encoding=encoding,
        )
    try:
        if hasattr(conn, "clear_buffer"):
            conn.clear_buffer()
    except Exception:
        pass
    conn.write_channel(command.rstrip() + "\n")
    chunks: list[str] = []
    deadline = time.monotonic() + max(1, read_timeout)
    last_data_at = time.monotonic()
    while time.monotonic() < deadline:
        raw = conn.read_channel()
        text = normalize_command_output(raw, encoding)
        if text:
            chunks.append(text)
            last_data_at = time.monotonic()
            joined = "".join(chunks)
            if _h3c_ping_output_complete(joined):
                return joined
        elif chunks and time.monotonic() - last_data_at >= 1.0:
            joined = "".join(chunks)
            if _h3c_ping_output_complete(joined):
                return joined
        time.sleep(0.1)
    return "".join(chunks)


def _h3c_ping_output_complete(output: str) -> bool:
    text = str(output or "")
    return bool(re.search(r"\d+(?:\.\d+)?%\s+packet\s+loss", text, re.IGNORECASE) or re.search(r"\d+(?:\.\d+)?%\s*丢包", text))


def _h3c_ping_note(location: str, command: str, result: PingResult) -> str:
    return f"执行位置：{location}；命令：{command}；结果：{_h3c_ping_summary(result)}"


def _h3c_ping_summary(result: PingResult) -> str:
    packet_text = "-"
    if result.transmitted is not None and result.received is not None:
        packet_text = f"{result.received}/{result.transmitted}"
    avg = "-" if result.avg_rtt_ms is None else f"{result.avg_rtt_ms} ms"
    error = f"；错误：{_clean_ping_error(result.error)}" if result.error else ""
    return f"{packet_text}，丢包 {result.loss_percent}%，平均 {avg}{error}"


def _core_ping_note(task: CoreRemotePingTask, result: PingResult) -> str:
    note = _h3c_ping_note(f"{_core_display_name(task.core_device_name)} CLI", task.command, result)
    if _ping_ok_no_loss(result):
        return f"{note}；核心侧可达落地IP，本机不可达可能是本机路由或网关限制，不影响车内链路判断。"
    return f"{note}；核心侧也不可达落地IP，可能是列车离线、MR射频关闭或落地链路不可用。该结果仅表示地面接入异常，不等同车内通信异常。"


def _clean_ping_error(error: str) -> str:
    text = str(error or "")
    if "Command ['ping'" in text or 'Command ["ping"' in text:
        return "本机辅助 ping 超时，已排除出车内通信主诊断"
    return text


def _ac_task_message(status: TrainAcStatus) -> str:
    if status.tc1_mr_online and status.tc2_mr_online:
        return "AC/列车在线情况发现双端 MR 在线，继续检测车内链路"
    if status.tc1_mr_online or status.tc2_mr_online:
        return "AC/列车在线情况发现单端 MR 在线，继续检测车内链路"
    if status.both_mr_offline:
        return "AC mesh-link 未发现 TC1-MR 和 TC2-MR"
    return "AC mesh-link 查询失败或状态未知，继续通过 IP/SSH 辅助判断"


def _ssh_status_json(ssh_results: dict[str, SshResult]) -> dict[str, dict[str, object]]:
    payload: dict[str, dict[str, object]] = {}
    for name, result in ssh_results.items():
        remote_pings = list(result.task_results.values()) or list(result.command_results.values())
        payload[name] = {
            "connected": result.ok,
            "host": result.host,
            "error": result.error,
            "remote_ping_count": len(remote_pings),
            "remote_ping_ok_count": sum(1 for ping in remote_pings if _ping_ok_no_loss(ping)),
        }
    return payload


def _ground_access_status_json(ac_status: AcApStatus, ssh_results: dict[str, SshResult], core_results: dict[str, PingResult]) -> dict[str, object]:
    attempted_ssh = [item for item in ssh_results.values() if not _ssh_skipped(item)]
    if any(item.ok for item in attempted_ssh):
        mr_management = "ok"
    elif attempted_ssh:
        mr_management = "fail"
    else:
        mr_management = "unknown"
    if core_results:
        core_status = "ok" if any(_ping_ok_no_loss(item) for item in core_results.values()) else "fail"
    else:
        core_status = "skipped"
    return {
        "ac_mesh_link": "ok" if ac_status.online else "unknown" if not ac_status.selected else "fail",
        "mr_management_reachable": mr_management,
        "uplink_ip_reachable_from_core": core_status,
        "used_for_primary_diagnosis": False,
        "note": "地面侧接入仅作辅助，不直接判定车内网络故障。",
    }


def _access_entry_json(ssh_results: dict[str, SshResult]) -> dict[str, object]:
    entry_mrs = [name for name, result in ssh_results.items() if result.ok]
    return {
        "any_mr_ssh_ok": bool(entry_mrs),
        "entry_mrs": entry_mrs,
        "note": "本地可登录MR，具备车内链路检测入口" if entry_mrs else "本地无法登录任意一台MR，不能执行车内链路检测",
    }


def _vehicle_internal_status_json(ssh_results: dict[str, SshResult]) -> dict[str, object]:
    def status_for(source: str, target_prefix: str) -> str:
        ssh = ssh_results.get(source)
        if ssh is None or not ssh.ok:
            return "unknown"
        values = [
            ping
            for task_id, ping in ssh.task_results.items()
            if str(ssh.task_metadata.get(task_id, {}).get("target_node", "")).startswith(f"{target_prefix}-")
        ]
        if not values:
            values = [ping for ip, ping in ssh.command_results.items() if _vehicle_ip_tc(ip) == target_prefix]
        if not values:
            return "unknown"
        return "ok" if all(_ping_ok_no_loss(ping) for ping in values) else "fail"

    entry_count = sum(1 for result in ssh_results.values() if result.ok)
    all_remote = [ping for result in ssh_results.values() if result.ok for ping in (result.task_results or result.command_results).values()]
    single_entry_verified = entry_count == 1 and any(_ping_ok_no_loss(ping) for ping in all_remote)
    return {
        "validated": bool(all_remote),
        "validation_mode": "dual_entry" if entry_count >= 2 else "single_entry" if entry_count == 1 else "none",
        "single_entry_verified": single_entry_verified,
        "tc1_local": status_for("TC1-MR", "TC1"),
        "tc2_local": status_for("TC2-MR", "TC2"),
        "tc1_to_tc2": status_for("TC1-MR", "TC2"),
        "tc2_to_tc1": status_for("TC2-MR", "TC1"),
        "packet_loss_found": any(ping.ok and ping.loss_percent > 0 for ping in all_remote),
        "mr_peer_reachability": {
            "TC1-MR->TC2-MR": _task_status_for_target(ssh_results.get("TC1-MR"), "TC2-MR"),
            "TC2-MR->TC1-MR": _task_status_for_target(ssh_results.get("TC2-MR"), "TC1-MR"),
        },
    }


def _task_status_for_target(result: SshResult | None, target_node: str) -> str:
    if result is None or not result.ok:
        return "unknown"
    matches = [ping for task_id, ping in result.task_results.items() if f"_ping_{target_node}_" in task_id]
    if not matches:
        return "unknown"
    if any(ping.ok and ping.loss_percent > 0 for ping in matches):
        return "warning"
    if any(_ping_ok_no_loss(ping) for ping in matches):
        return "ok"
    return "fail"


def _vehicle_ip_tc(ip: str) -> str:
    if ip.endswith(".251") or ip.endswith(".1"):
        return "TC1"
    if ip.endswith(".252") or ip.endswith(".2"):
        return "TC2"
    return ""


def _peer_mr_issue(ssh_results: dict[str, SshResult]) -> str:
    for source, peer, peer_tc in (("TC1-MR", "TC2-MR", "TC2"), ("TC2-MR", "TC1-MR", "TC1")):
        result = ssh_results.get(source)
        if result is None or not result.ok:
            continue
        peer_values = [ping for task_id, ping in result.task_results.items() if f"_ping_{peer}_" in task_id]
        if not peer_values or any(_ping_ok_no_loss(ping) for ping in peer_values):
            continue
        peer_side_values = [
            ping
            for task_id, ping in result.task_results.items()
            if f"_ping_{peer_tc}-SW_" in task_id or f"_ping_{peer_tc}-SRV_" in task_id
        ]
        if peer_side_values and all(_ping_ok_no_loss(ping) for ping in peer_side_values):
            return f"{source} 到 {peer} 车内通信异常；跨TC主链路到对端SW/SRV基本可达，优先检查 {peer} 车内接口、车内IP、VLAN、MR状态或路由配置。"
    return ""


def _single_entry_verified(nodes: list[CarNetworkNode], ssh_results: dict[str, SshResult]) -> bool:
    ok_entries = [name for name, result in ssh_results.items() if result.ok]
    if len(ok_entries) != 1:
        return False
    missing_mrs = [node for node in nodes if node.is_mr and node.node_name not in ok_entries]
    return any(_mr_vehicle_reachable_by_any_entry(node, ssh_results, nodes)[0] for node in missing_mrs)


def _remote_diagnosis_items(ssh_results: dict[str, SshResult]) -> list[dict[str, object]]:
    items: list[dict[str, object]] = []
    for result in ssh_results.values():
        for task_id, ping in result.task_results.items():
            meta = result.task_metadata.get(task_id, {})
            items.append(
                {
                    "source": meta.get("source_node", result.node_name),
                    "target": meta.get("target_node", ""),
                    "target_ip": meta.get("target_ip", ping.ip),
                    "layer": meta.get("layer", ""),
                    "command": meta.get("command", ""),
                    "loss": ping.loss_percent,
                    "avg_rtt": ping.avg_rtt_ms,
                    "result": "ok" if _ping_ok_no_loss(ping) else "warning" if ping.ok and ping.loss_percent > 0 else "fail",
                }
            )
    return items


def _diagnosis_items_from_tables(tables: dict[str, list[dict[str, object]]]) -> list[dict[str, object]]:
    items: list[dict[str, object]] = []
    for rows in tables.values():
        for row in rows:
            status = str(row.get("status") or "")
            layer = str(row.get("layer") or "")
            node_text = str(row.get("node") or "")
            severity = "ok" if status == "OK" else "warning" if status in {"丢包异常", "AC未发现", "跳过"} else "fail" if status in {"不通", "SSH失败", "AC离线"} else "unknown"
            if severity == "ok" and layer not in {"MR远程本端检测", "MR远程跨TC检测", "MR远程跨TC长Ping", "SSH层", "AC层"}:
                continue
            source, target, target_ip = _parse_diagnosis_node_text(node_text)
            result = f"{status}"
            if row.get("loss") not in {"", "-", None}:
                result = f"{result}，丢包 {row.get('loss')}"
            reason = str(row.get("note") or "") or f"{node_text} {status}"
            items.append(
                {
                    "severity": severity,
                    "source": source,
                    "target": target,
                    "target_ip": target_ip,
                    "layer": layer,
                    "result": result,
                    "reason": reason,
                    "suggestion": _diagnosis_suggestion(layer, severity),
                }
            )
    return items


def _parse_diagnosis_node_text(text: str) -> tuple[str, str, str]:
    source = ""
    target = text
    target_ip = ""
    if "->" in text:
        source, target = [part.strip() for part in text.split("->", 1)]
    if "/" in target:
        target, target_ip = [part.strip() for part in target.rsplit("/", 1)]
    return source, target, target_ip


def _diagnosis_suggestion(layer: str, severity: str) -> str:
    if severity == "ok":
        return ""
    if "跨TC" in layer:
        return "检查跨TC链路、VRRP 状态、中间骨干链路和两端三层交换机配置"
    if layer == "MR远程本端检测":
        return "检查本端三层交换机端口、VLAN、服务器网卡和网关配置"
    if layer == "SSH层":
        return "检查 MR 管理地址、SSH账号密码、认证方式和访问控制"
    if layer == "AC层":
        return "检查 AC mesh-link 输出、MR 注册状态和列车号/端别命名映射"
    return "结合分层检测结果检查对应链路和设备配置"


def probe_ac_mesh_links(
    ac_devices: list[Device],
    command_func: Callable[[str], str] | None = None,
) -> AcProbeResult:
    controllers: list[AcControllerProbe] = []
    raw_outputs: dict[str, str] = {}
    if command_func is not None:
        try:
            output = command_func("display wlan mesh-link ap")
        except Exception as exc:
            return AcProbeResult(True, False, [AcControllerProbe("", "manual", "", False, error=str(exc))], error=str(exc))
        controller = AcControllerProbe("", "manual", "", True, output)
        return AcProbeResult(True, True, [controller], {"manual": output})
    with ThreadPoolExecutor(max_workers=max(1, min(6, len(ac_devices)))) as executor:
        futures = {executor.submit(_probe_single_ac_mesh_link, device): device for device in ac_devices}
        for future in as_completed(futures):
            controller = future.result()
            controllers.append(controller)
            if controller.output:
                raw_outputs[controller.device_id or controller.host or controller.device_name] = controller.output
    query_success = any(item.success for item in controllers)
    error = "" if query_success else "; ".join(item.error for item in controllers if item.error)
    return AcProbeResult(True, query_success, controllers, raw_outputs, error)


def _probe_single_ac_mesh_link(device: Device) -> AcControllerProbe:
    device_id = str(device.id or device.device_uuid or "")
    host = device.primary_address
    try:
        output = "\n".join(run_ac_commands(device, ("display wlan mesh-link ap",)).values())
        return AcControllerProbe(device_id, device.name, host, True, output)
    except Exception as exc:
        return AcControllerProbe(device_id, device.name, host, False, error=str(exc))


def match_train_mr_in_mesh_links(train: CarNetworkTrain | None, nodes: list[CarNetworkNode], ac_probe: AcProbeResult) -> TrainAcStatus:
    matched_by: dict[str, list[str]] = {"TC1-MR": [], "TC2-MR": []}
    matched_details: dict[str, list[dict[str, object]]] = {"TC1-MR": [], "TC2-MR": []}
    train_no = _current_train_no(train, nodes)
    parse_success = False
    parse_warning = False
    suspected_lines: list[str] = []
    output_nonempty = False
    if not ac_probe.query_success:
        return TrainAcStatus(
            tc1_mr_online=False,
            tc2_mr_online=False,
            both_mr_offline=False,
            any_query_success=False,
            matched_by=matched_by,
            matched_details=matched_details,
            parse_success=False,
            parse_warning=True,
            ac_output_nonempty=False,
            online_source="ac_query_failed",
        )
    for controller in ac_probe.controllers:
        if not controller.success:
            continue
        output_nonempty = output_nonempty or bool(str(controller.output or "").strip())
        controller_matched: list[dict[str, object]] = []
        parse_result = H3CComwareV9VehicleMrMeshLinkParser().parse(controller.output)
        links = parse_result.links
        object.__setattr__(controller, "parsed_peer_count", len(links))
        object.__setattr__(controller, "fallback_matched_mrs", [])
        parse_success = parse_success or parse_result.parse_status == "OK"
        suspected_lines.extend(_suspected_train_lines(controller.output, train_no, links))
        if parse_result.parse_status != "OK" or (suspected_lines and not links):
            parse_warning = True
        LOGGER.info("[车内通信检测] AC %s(%s) mesh-link 查询成功，列车在线情况 parser 解析 peer %s 条", controller.device_name, controller.host, len(links))
        for link in links:
            if not _mesh_status_online(link.status):
                continue
            identity = parse_train_identity(link.peer_name)
            if identity is None or identity.train_no != train_no:
                continue
            node_name = "TC1-MR" if identity.car_end == "CT" else "TC2-MR"
            detail = _mesh_match_detail(node_name, link.peer_name, identity.train_no, identity.car_end, link.status, source="ac_realtime_parser")
            if link.peer_name not in matched_by[node_name]:
                matched_by[node_name].append(link.peer_name)
                matched_details[node_name].append(detail)
                controller_matched.append(detail)
                LOGGER.info("[车内通信检测] 当前列车 %s，匹配到 %s: %s status=%s", train_no, node_name, link.peer_name, link.status)
        object.__setattr__(controller, "matched_mrs", controller_matched)
    tc1 = bool(matched_by["TC1-MR"])
    tc2 = bool(matched_by["TC2-MR"])
    both_offline = bool(ac_probe.query_success and output_nonempty and parse_success and not tc1 and not tc2)
    if both_offline and suspected_lines:
        parse_warning = True
    if not tc1:
        LOGGER.info("[车内通信检测] 当前列车 %s，未匹配到 TC1-MR，继续 ping/SSH", train_no)
    if not tc2:
        LOGGER.info("[车内通信检测] 当前列车 %s，未匹配到 TC2-MR，继续 ping/SSH", train_no)
    return TrainAcStatus(tc1, tc2, both_offline, ac_probe.query_success, matched_by, matched_details, parse_success, parse_warning, suspected_lines, output_nonempty, "ac_realtime_parser")


def get_vehicle_mr_online_status_from_store(paths: PathResolver | None, site_name: str, train_no: str) -> TrainAcStatus | None:
    train_no = normalize_train_no(train_no)
    if paths is None or not site_name or not train_no:
        return None
    LOGGER.info("[车内通信检测] 尝试读取列车在线情况 site=%s train_no=%s", site_name, train_no)
    try:
        store = VehicleMrOnlineStore(paths, site_name)
        states = store.list_current_states()
    except Exception as exc:
        LOGGER.info("[车内通信检测] 读取列车在线情况当前状态失败：%s", exc)
        return None
    LOGGER.info("[车内通信检测] 列车在线情况当前状态数量=%s", len(states))
    state = next((item for item in states if _vehicle_mr_state_matches_train_no(item, train_no)), None)
    if state is None or _vehicle_mr_state_is_expired(state):
        if state is not None:
            LOGGER.info("[车内通信检测] 列车在线情况状态已过期: train_no=%s tc1=%s tc2=%s status=%s", state.train_no, state.tc1.seen, state.tc2.seen, state.status)
        return None
    tc1 = bool(state.tc1.seen)
    tc2 = bool(state.tc2.seen)
    LOGGER.info("[车内通信检测] 匹配到列车在线情况状态: train_no=%s tc1=%s tc2=%s status=%s", state.train_no, tc1, tc2, state.status)
    online_statuses = {TRAIN_STATUS_ONLINE, TRAIN_STATUS_DUAL_ONLINE, TRAIN_STATUS_ABNORMAL_SINGLE, TRAIN_STATUS_UNEXPECTED_END}
    if not (tc1 or tc2 or state.status in online_statuses):
        return None
    matched_by: dict[str, list[str]] = {
        "TC1-MR": [_store_peer_label(state, "TC1")] if tc1 else [],
        "TC2-MR": [_store_peer_label(state, "TC2")] if tc2 else [],
    }
    matched_details: dict[str, list[dict[str, object]]] = {
        "TC1-MR": [_store_match_detail("TC1-MR", state, "CT")] if tc1 else [],
        "TC2-MR": [_store_match_detail("TC2-MR", state, "CW")] if tc2 else [],
    }
    return TrainAcStatus(
        tc1_mr_online=tc1,
        tc2_mr_online=tc2,
        both_mr_offline=not tc1 and not tc2,
        any_query_success=True,
        matched_by=matched_by,
        matched_details=matched_details,
        parse_success=True,
        parse_warning=False,
        ac_output_nonempty=True,
        online_source="vehicle_mr_online_current_state",
    )


def _vehicle_mr_state_is_expired(state: VehicleMrTrainState, max_age: timedelta = timedelta(seconds=600)) -> bool:
    timestamps = [
        state.last_ac_time,
        state.last_seen_at,
        state.tc1.last_seen_at,
        state.tc2.last_seen_at,
    ]
    parsed = [_parse_state_time(value) for value in timestamps if value]
    parsed = [value for value in parsed if value is not None]
    if not parsed:
        return False
    return datetime.now() - max(parsed) > max_age


def _vehicle_mr_state_matches_train_no(state: VehicleMrTrainState, train_no: str) -> bool:
    target = normalize_train_no(train_no)
    if not target:
        return False
    values = (
        state.train_no,
        state.train_id,
        state.display_name,
        state.tc1.ap_name,
        state.tc2.ap_name,
    )
    return any(normalize_train_no(value) == target for value in values if value)


def _parse_state_time(value: str) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    for candidate in (text, text.replace("T", " ")):
        try:
            return datetime.fromisoformat(candidate)
        except ValueError:
            pass
    return None


def _store_peer_label(state: VehicleMrTrainState, tc: str) -> str:
    end = state.tc1 if tc == "TC1" else state.tc2
    return end.ap_name or "online_mr_current_state"


def _store_match_detail(node: str, state: VehicleMrTrainState, end: str) -> dict[str, object]:
    end_state = state.tc1 if end == "CT" else state.tc2
    return {
        "node": node,
        "peer_name": end_state.ap_name or "online_mr_current_state",
        "train_no": state.train_no,
        "end": end,
        "status": state.status,
        "source": "vehicle_mr_online_current_state",
        "match_mode": "vehicle_mr_online_current_state",
    }


def should_skip_deep_probe_by_ac(train_ac_status: TrainAcStatus) -> bool:
    return (
        train_ac_status.any_query_success
        and train_ac_status.ac_output_nonempty
        and train_ac_status.parse_success
        and train_ac_status.both_mr_offline
        and not train_ac_status.parse_warning
        and not train_ac_status.suspected_current_train_lines
    )


def ac_status_from_probe(ac_probe: AcProbeResult, train_ac_status: TrainAcStatus) -> AcApStatus:
    return AcApStatus(
        mesh_link=train_ac_status.tc1_mr_online or train_ac_status.tc2_mr_online,
        ap_all=False,
        radio_ok=None,
        selected=ac_probe.enabled,
        raw=ac_probe.raw_outputs,
        error=ac_probe.error,
    )


def build_ac_offline_result(
    nodes: list[CarNetworkNode],
    ac_probe: AcProbeResult,
    train_ac_status: TrainAcStatus,
    train: CarNetworkTrain | None,
) -> CarNetworkDiagnosticResult:
    node_states = {node.node_name: ("fail" if node.is_mr else "skipped") for node in nodes}
    train_id = train.train_id if train is not None else (nodes[0].train_id if nodes else "")
    train_no = train.train_no if train is not None else (nodes[0].train_no if nodes else normalize_train_no(train_id))
    display_name = train.display_name if train is not None else (f"{train_no}车" if train_no else train_id)
    ac_status = ac_status_from_probe(ac_probe, train_ac_status)
    conclusion = "AC mesh-link 未发现 TC1-MR 和 TC2-MR，无法从地面接入 MR 执行车内链路检测；可能是列车下电、MR射频关闭或地面接入不可达，不能直接判定车内网络故障。"
    tables = build_ac_offline_tables(nodes)
    return CarNetworkDiagnosticResult(
        train_id=train_id,
        status="offline",
        nodes=node_states,
        cross_train={"TC1->TC2": "skipped", "TC2->TC1": "skipped"},
        ac_status="fail",
        ssh_status="skipped",
        conclusion=conclusion,
        train_no=train_no,
        display_name=display_name,
        ends=_ends_json(nodes, node_states),
        vrrp={"ip": next((node.vrrp_ip for node in nodes if node.vrrp_ip), ""), "status": "skipped"},
        ac_detail=ac_status,
        ac_probe=ac_probe,
        train_ac_status=train_ac_status,
        tables=tables,
    )


def build_ac_offline_tables(nodes: list[CarNetworkNode]) -> dict[str, list[dict[str, object]]]:
    tables = {"TC1": [], "TC2": []}
    for tc in ("TC1", "TC2"):
        mr = next((node for node in nodes if node.tc == tc and node.is_mr), None)
        if mr is not None:
            tables[tc].append(_table_row(mr.node_name, "AC层", "AC离线", "-", "-", "AC mesh-link 离线"))
        for node in [item for item in nodes if item.tc == tc and not item.is_mr]:
            tables[tc].append(_table_row(node.node_name, "车内有线层", "跳过", "-", "-", "列车双端MR均离线"))
    return tables


def build_result_tables(
    nodes: list[CarNetworkNode],
    ping_results: dict[str, PingResult],
    ssh_results: dict[str, SshResult],
    ac_status: AcApStatus,
    core_results: dict[str, PingResult] | None = None,
    core_devices: list[Device] | None = None,
    train_ac_status: TrainAcStatus | None = None,
) -> dict[str, list[dict[str, object]]]:
    tables = {"TC1": [], "TC2": []}
    core_results = core_results or {}
    core_devices = core_devices or []
    nodes_by_vehicle_ip = {ip: node for node in nodes for ip in _vehicle_candidate_ips(node, nodes)}
    for tc in ("TC1", "TC2"):
        for node in [item for item in nodes if item.tc == tc]:
            if node.is_mr:
                if train_ac_status is not None:
                    online = train_ac_status.tc1_mr_online if node.tc == "TC1" else train_ac_status.tc2_mr_online
                    details = train_ac_status.matched_details.get(node.node_name, [])
                    if online and details:
                        first = details[0]
                        note = f"AC mesh-link 在线：{first.get('peer_name', '-')}"
                        status = first.get("status")
                        if status:
                            note = f"{note} / {status}"
                    elif online:
                        note = "AC mesh-link 在线"
                    else:
                        note = f"AC mesh-link 未发现 {node.node_name}，继续后续检测"
                    tables[tc].append(_table_row(node.node_name, "AC层", "OK" if online else "AC未发现", "-", "-", note))
                if not node.ip_vehicle:
                    tables[tc].append(_table_row(node.node_name, "车内IP", "不适用", "-", "-", "MR 未配置车内IP"))
                ssh = ssh_results.get(node.node_name)
                ssh_label = "可管理" if ssh and ssh.ok else "跳过" if _ssh_skipped(ssh) else "SSH失败" if ssh else "未检测"
                ssh_note = f"SSH 登录 {node.node_name} 成功" if ssh and ssh.ok else "" if ssh is None else _clean_ping_error(ssh.error)
                ssh_ip = node.ssh_host or node.ip_vehicle or node.ip_uplink or (ssh.host if ssh else "")
                tables[tc].append(_table_row(f"{node.node_name} / {ssh_ip}" if ssh_ip else f"{node.node_name} SSH", "SSH地址 / MR管理", ssh_label, "-", "-", ssh_note))
                reachable, remote_ping, remote_note = _mr_vehicle_reachable_by_any_entry(node, ssh_results, nodes)
                if ssh is not None and not ssh.ok and remote_ping is not None:
                    status = "OK" if reachable else "丢包异常" if remote_ping.ok and remote_ping.loss_percent > 0 else "不通"
                    note = "本端SSH不可达，但对端MR可ping通本端MR车内地址，判定车内通信正常（单端激活）" if reachable else remote_note
                    tables[tc].append(_table_row(node.node_name, "单端激活验证", status, "-" if remote_ping.avg_rtt_ms is None else f"{remote_ping.avg_rtt_ms} ms", f"{remote_ping.loss_percent}%", note))
                if ssh is not None:
                    for target, ping in ssh.command_results.items():
                        target_node = nodes_by_vehicle_ip.get(target)
                        layer = "MR远程跨TC快速检测" if target_node is not None and target_node.tc != node.tc else "MR远程本端检测"
                        target_label = target_node.node_name if target_node is not None else target
                        tables[tc].append(_table_row(f"{node.node_name} -> {target_label} / {target}", layer, _ping_label(ping), "-" if ping.avg_rtt_ms is None else f"{ping.avg_rtt_ms} ms", f"{ping.loss_percent}%", _remote_ping_note(node, target_node, target, layer, ping)))
            if node.ip_uplink and node.node_type.upper() in {"MR", "SW", "3SW"}:
                matches = [(key, ping) for key, ping in core_results.items() if key.endswith(f"->{node.ip_uplink}")]
                if matches:
                    for key, ping in matches:
                        task = _core_task_for_result_key(key, nodes, core_devices)
                        note = _core_ping_note(task, ping) if task is not None else _h3c_ping_note("核心交换机", build_h3c_ping_command(node.ip_uplink, packet_count=CAR_NETWORK_QUICK_PING_COUNT), ping)
                        tables[tc].append(_table_row(f"{node.node_name} / {node.ip_uplink}", "核心侧落地IP检测", _core_ping_label(ping), "-" if ping.avg_rtt_ms is None else f"{ping.avg_rtt_ms} ms", f"{ping.loss_percent}%", note))
    return tables


def _table_row(name: str, layer: str, status: str, rtt: str, loss: str, note: str) -> dict[str, object]:
    return {"node": name, "layer": layer, "status": status, "rtt": rtt, "loss": loss, "note": note}


def _core_task_for_result_key(key: str, nodes: list[CarNetworkNode], core_devices: list[Device]) -> CoreRemotePingTask | None:
    for task in _core_remote_ping_tasks(core_devices, nodes):
        if key == f"{task.core_device_id}->{task.target_ip}":
            return task
    return None


def _ping_label(ping: PingResult) -> str:
    if ping.ok and ping.loss_percent == 0:
        return "OK"
    if ping.ok and ping.loss_percent > 0:
        return "丢包异常"
    return "不通"


def _core_ping_label(ping: PingResult) -> str:
    if ping.ok and ping.loss_percent == 0:
        return "OK"
    if ping.ok and ping.loss_percent > 0:
        return "辅助不稳定"
    return "辅助不可达"


def _remote_ping_note(source: CarNetworkNode, target_node: CarNetworkNode | None, target_ip: str, layer: str, ping: PingResult) -> str:
    target = target_node.node_name if target_node is not None else target_ip
    if layer == "跨TC通信":
        packet_count = CAR_NETWORK_CROSS_TC_PING_COUNT
    else:
        packet_count = CAR_NETWORK_QUICK_PING_COUNT
    command = build_h3c_ping_command(target_ip, packet_count=packet_count)
    base = _ping_result_message(source.node_name, target, target_ip, ping, via_ssh=True, command=command)
    if ping.ok and ping.loss_percent == 0:
        return f"{base}。{source.node_name} 到 {target} 车内通信正常。"
    if target_node is not None and target_node.is_mr:
        return f"{base}。{source.node_name} 到 {target} 不通，可能是 {target} 车内接口、车内IP、VLAN、三层路径或对端 MR 状态异常。"
    if "跨TC" in layer:
        return f"{base}。推断：跨TC链路、VRRP 或中间骨干链路存在异常，建议检查两端三层交换机、VRRP 状态和跨车骨干链路。"
    return f"{base}。推断：{source.node_name} 到 {target} 的本端车内链路异常，建议检查本端三层交换机端口、VLAN、服务器网卡和网关配置。"


def parse_h3c_mesh_link_ap_output(output: str) -> list[MeshLinkPeer]:
    text = _strip_ansi(str(output or "")).replace("\r\n", "\n").replace("\r", "\n")
    peers: list[MeshLinkPeer] = []
    local_ap_name = ""
    mac_pattern = r"(?:[0-9a-fA-F]{4}[-.][0-9a-fA-F]{4}[-.][0-9a-fA-F]{4}|[0-9a-fA-F]{2}(?::[0-9a-fA-F]{2}){5})"
    row_pattern = re.compile(rf"^\s*(?P<peer>\S+)\s+(?P<peer_mac>{mac_pattern})\s+(?P<local_mac>{mac_pattern})\s+(?P<status>\S+)(?:\s+(?P<rssi>-?\d+))?", re.IGNORECASE)
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        ap_match = re.match(r"^AP\s+name\s*:\s*(.+)$", line, re.IGNORECASE)
        if ap_match:
            local_ap_name = ap_match.group(1).strip()
            continue
        if line.casefold().startswith("peer name"):
            continue
        match = row_pattern.match(line)
        if not match:
            continue
        rssi_text = match.group("rssi")
        peers.append(
            MeshLinkPeer(
                local_ap_name=local_ap_name,
                peer_name=match.group("peer"),
                peer_mac=match.group("peer_mac"),
                local_mac=match.group("local_mac"),
                status=match.group("status"),
                rssi=int(rssi_text) if rssi_text is not None else None,
                raw_line=raw_line,
            )
        )
    return peers


def match_train_mr_in_raw_mesh_output(output: str, train_no: str) -> dict[str, list[dict[str, object]]]:
    result: dict[str, list[dict[str, object]]] = {"TC1-MR": [], "TC2-MR": []}
    if not train_no:
        return result
    for raw_line in _strip_ansi(str(output or "")).splitlines():
        line = raw_line.strip()
        if not line:
            continue
        identity = parse_train_mr_identity(line)
        if identity is None or identity["train_no"] != train_no:
            continue
        status = _online_status_in_line(line)
        if not status:
            continue
        peer_name = str(identity["source_name"])
        node_name = f"{identity['tc']}-MR"
        detail = _mesh_match_detail(node_name, peer_name, str(identity["train_no"]), str(identity["end"]), status)
        if peer_name not in {str(item.get("peer_name")) for item in result[node_name]}:
            result[node_name].append(detail)
    return result


def parse_train_mr_identity(name: str) -> dict[str, str] | None:
    train_no, end = extract_train_identity_from_name(name)
    if not train_no or end not in {"CT", "CW"}:
        return None
    text = str(name or "")
    if not re.search(r"(?i)(^|[-_\s])MR($|[-_\s])|MR[-_\s]?(?:CT|CW)|Cloud-AP|车载AP|AP", text):
        return None
    return {
        "train_no": train_no,
        "end": end,
        "tc": "TC1" if end == "CT" else "TC2",
        "role": "MR",
        "source_name": _identity_source_name(text),
    }


def extract_train_identity_from_name(name: str) -> tuple[str | None, str | None]:
    text = str(name or "")
    train_no = ""
    patterns = (
        r"(?i)(?<![A-Z0-9])LC\s*0*(\d{1,3})(?!\d)",
        r"列车\s*0*(\d{1,3})(?!\d)",
        r"(?<!\d)0*(\d{1,3})\s*车",
    )
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            train_no = f"{int(match.group(1)):02d}"
            break
    upper = text.upper()
    end = None
    if re.search(r"(?i)(?:MR-CT|-CT\b|\bCT\b|TC1)", upper) or "车头" in text:
        end = "CT"
    elif re.search(r"(?i)(?:MR-CW|-CW\b|\bCW\b|TC2)", upper) or "车尾" in text:
        end = "CW"
    return train_no or None, end


def _identity_source_name(text: str) -> str:
    for token in re.split(r"\s+", str(text or "").strip()):
        if extract_train_identity_from_name(token)[0]:
            return token
    return str(text or "").strip()


def _mesh_match_detail(node: str, peer_name: str, train_no: str, end: str, status: str, source: str = "ac_realtime_parser") -> dict[str, object]:
    return {
        "node": node,
        "peer_name": peer_name,
        "train_no": train_no,
        "end": end,
        "status": status,
        "source": source,
        "match_mode": "train_no_end",
    }


def _strip_ansi(text: str) -> str:
    return re.sub(r"\x1b\[[0-?]*[ -/]*[@-~]", "", text)


def _mesh_status_online(status: str) -> bool:
    return str(status or "").strip().casefold() in ONLINE_MESH_STATUSES


def _online_status_in_line(line: str) -> str:
    for token in re.split(r"\s+", str(line or "")):
        if _mesh_status_online(token):
            return token
    return ""


def _peer_label_name(value: str) -> str:
    return str(value or "").split(" status=", 1)[0]


def _current_train_no(train: CarNetworkTrain | None, nodes: list[CarNetworkNode]) -> str:
    value = train.train_no if train is not None else next((node.train_no for node in nodes if node.train_no), "")
    normalized = normalize_train_no(value)
    return f"{int(normalized):02d}" if normalized.isdigit() else normalized


def _suspected_train_lines(output: str, train_no: str, parsed_peers: list[object]) -> list[str]:
    if not train_no:
        return []
    parsed_peer_names = {str(getattr(peer, "peer_name", "")).strip() for peer in parsed_peers}
    suspected: list[str] = []
    for raw_line in _strip_ansi(str(output or "")).splitlines():
        line = raw_line.strip()
        if not line:
            continue
        parsed_no, _end = extract_train_identity_from_name(line)
        if parsed_no == train_no and not any(peer_name and peer_name in line for peer_name in parsed_peer_names):
            suspected.append(line)
    return suspected


def _run_local_aux_ping(ip: str, *, count: int = CAR_NETWORK_QUICK_PING_COUNT, timeout_ms: int = CAR_NETWORK_QUICK_DETECT_SECONDS * 1000) -> PingResult:
    args = _ping_args(ip, count, timeout_ms)
    started = time.monotonic()
    creationflags = subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0
    try:
        completed = subprocess.run(args, capture_output=True, text=True, encoding="gbk", errors="replace", timeout=CAR_NETWORK_QUICK_DETECT_SECONDS, creationflags=creationflags)
    except Exception as exc:
        return PingResult(ip, False, error=str(exc))
    output = f"{completed.stdout or ''}\n{completed.stderr or ''}"
    parsed = parse_ping_output(ip, output)
    if parsed.avg_rtt_ms is None and completed.returncode == 0:
        parsed = PingResult(ip, parsed.ok, parsed.loss_percent, round((time.monotonic() - started) * 1000 / max(1, count), 1), output)
    return parsed


def parse_ping_output(ip: str, output: str) -> PingResult:
    text = str(output or "")
    if not text.strip():
        return PingResult(ip, False, 100.0, error="ping output empty")
    transmitted: int | None = None
    received: int | None = None
    packet_match = re.search(r"(\d+)\s+packet\(s\)\s+transmitted,\s*(\d+)\s+packet\(s\)\s+received", text, re.IGNORECASE)
    if not packet_match:
        packet_match = re.search(r"(\d+)\s+packets?\s+transmitted,\s*(\d+)\s+(?:packets?\s+)?received", text, re.IGNORECASE)
    if packet_match:
        transmitted = int(packet_match.group(1))
        received = int(packet_match.group(2))
    else:
        transmitted_match = re.search(r"(\d+)\s+packet\(s\)\s+transmitted", text, re.IGNORECASE)
        received_match = re.search(r"(\d+)\s+packet\(s\)\s+received", text, re.IGNORECASE)
        if not transmitted_match:
            transmitted_match = re.search(r"(\d+)\s+packets?\s+transmitted", text, re.IGNORECASE)
        if not received_match:
            received_match = re.search(r"(\d+)\s+(?:packets?\s+)?received", text, re.IGNORECASE)
        if transmitted_match:
            transmitted = int(transmitted_match.group(1))
        if received_match:
            received = int(received_match.group(1))
    loss_match = re.search(r"(\d+(?:\.\d+)?)\s*%\s*(?:packet\s+)?loss", text, re.IGNORECASE)
    if not loss_match:
        loss_match = re.search(r"(\d+(?:\.\d+)?)\s*%\s*丢失", text, re.IGNORECASE)
    if loss_match:
        loss = float(loss_match.group(1))
    else:
        cn_loss = re.search(r"丢失\s*=\s*\d+\s*\((\d+(?:\.\d+)?)%\s*丢失\)", text)
        success_evidence = re.search(r"\bttl[=\s]|bytes\s+from|reply\s+from|bytes=|time[=<]\s*\d+|min/avg/max", text, re.IGNORECASE)
        loss = float(cn_loss.group(1)) if cn_loss else (0.0 if success_evidence else 100.0)
    min_rtt: float | None = None
    max_rtt: float | None = None
    min_avg_max = re.search(r"min/avg/max(?:/[a-z-]+)?\s*=\s*(\d+(?:\.\d+)?)/(\d+(?:\.\d+)?)/(\d+(?:\.\d+)?)(?:/\d+(?:\.\d+)?)?\s*ms", text, re.IGNORECASE)
    avg_match = re.search(r"(?:Average|平均)\s*=\s*(\d+(?:\.\d+)?)\s*ms", text, re.IGNORECASE)
    if not avg_match:
        avg_match = min_avg_max
    if not avg_match:
        avg_match = re.search(r"(?:avg|平均)[^=/]*[/=]\s*(\d+(?:\.\d+)?)", text, re.IGNORECASE)
    if not avg_match:
        avg_match = re.search(r"time[=<]\s*(\d+(?:\.\d+)?)\s*ms", text, re.IGNORECASE)
    if min_avg_max:
        min_rtt = float(min_avg_max.group(1))
        avg = float(min_avg_max.group(2))
        max_rtt = float(min_avg_max.group(3))
    else:
        avg = float(avg_match.group(1)) if avg_match else None
    return PingResult(ip, loss < 100.0, loss, avg, text, "", transmitted, received, min_rtt, max_rtt)


def run_ac_commands(ac: Device, commands: Iterable[str]) -> dict[str, str]:
    if ConnectHandler is None:
        raise RuntimeError("netmiko is not installed")
    targets = connection_targets(ac)
    if targets:
        target = build_netmiko_params(targets[0])
    else:
        username = ac.ssh_username or ac.username or ""
        password = ac.ssh_password or ac.password or ""
        if not ac.primary_address or not username or not password:
            raise RuntimeError("AC 连接信息不完整")
        from netconsole.services.netmiko_connection import ConnectionTarget

        target = build_netmiko_params(
            ConnectionTarget("SSH", H3C_NETMIKO_DEVICE_TYPE, ac.primary_address, int(ac.ssh_port or 22), username, password, encoding_for_vendor(ac.device_vendor))
        )
    conn = ConnectHandler(**target)
    encoding = str(target.get("encoding") or "gb2312")
    try:
        try:
            safe_send_command(
                conn,
                "screen-length disable",
                read_timeout=5,
                strip_prompt=False,
                strip_command=False,
                use_timing=True,
                encoding=encoding,
            )
        except Exception as exc:
            LOGGER.warning("AC分页关闭失败，继续执行采集命令: %s", exc)
        result: dict[str, str] = {}
        errors: list[str] = []
        for command in commands:
            try:
                result[command] = _send_ac_collect_command(conn, command, encoding)
            except Exception as exc:
                errors.append(f"{command}: {exc}")
                result[command] = ""
        if not any(output.strip() for output in result.values()) and errors:
            raise RuntimeError("AC采集命令失败：" + "; ".join(errors))
        return result
    finally:
        conn.disconnect()


def _send_ac_collect_command(conn: object, command: str, encoding: str) -> str:
    output = safe_send_command(
        conn,
        command,
        read_timeout=30,
        strip_prompt=False,
        strip_command=False,
        use_timing=True,
        encoding=encoding,
    )
    if output.strip() or command.strip().casefold() != "display wlan mesh-link ap":
        return output
    fallback = safe_send_command(
        conn,
        "dis wlan mesh-link ap",
        read_timeout=30,
        strip_prompt=False,
        strip_command=False,
        use_timing=True,
        encoding=encoding,
    )
    return fallback


def _ping_args(ip: str, count: int, timeout_ms: int) -> list[str]:
    if platform.system().casefold() == "windows":
        return ["ping", "-n", str(count), "-w", str(timeout_ms), ip]
    return ["ping", "-c", str(count), "-W", str(max(1, int(timeout_ms / 1000))), ip]


def build_h3c_ping_command(target_ip: str, *, packet_count: int | None = None) -> str:
    if packet_count is None:
        return f"ping {target_ip}"
    return f"ping -c {packet_count} {target_ip}"


def _core_remote_ping_tasks(core_devices: list[Device], nodes: list[CarNetworkNode]) -> tuple[CoreRemotePingTask, ...]:
    tasks: list[CoreRemotePingTask] = []
    targets = [node for node in nodes if node.ip_uplink and node.node_type.upper() in {"MR", "SW", "3SW"}]
    for device in core_devices:
        core_id = _device_id(device) or device.name
        core_name = device.name or core_id or "核心交换机"
        core_host = device.primary_address or ""
        for node in targets:
            tasks.append(
                CoreRemotePingTask(
                    task_id=_core_remote_ping_task_id(device, node.ip_uplink),
                    core_device_id=core_id,
                    core_device_name=core_name,
                    core_host=core_host,
                    target_node=node.node_name,
                    target_ip=node.ip_uplink,
                    command=build_h3c_ping_command(node.ip_uplink, packet_count=CAR_NETWORK_QUICK_PING_COUNT),
                    layer="核心侧落地IP检测",
                    timeout=CAR_NETWORK_QUICK_PING_TIMEOUT,
                )
            )
    return tuple(tasks)


def _core_remote_ping_task_id(device: Device, target_ip: str) -> str:
    return f"core_ping_{_device_id(device) or device.name}_{target_ip}"


def _core_discovery_from_selected(core_devices: list[Device]) -> dict[str, object]:
    candidates = [
        {
            "device_name": device.name or "",
            "system_name": device.system_name or "",
            "group": "",
            "host": device.primary_address or "",
            "selected": True,
            "reason": "provided as selected core switch",
        }
        for device in core_devices
    ]
    return {"candidates": candidates, "selected_count": len(candidates)}


def _mr_remote_ping_tasks(source: CarNetworkNode, nodes: list[CarNetworkNode], source_host: str) -> tuple[MrRemotePingTask, ...]:
    own = source.tc or source.node_name.split("-", 1)[0]
    local_first = [node for node in nodes if node.tc == own and not node.is_mr]
    remote = [node for node in nodes if node.tc != own and node.node_name != source.node_name]
    tasks: list[MrRemotePingTask] = []
    for node in local_first:
        for target_ip in _vehicle_candidate_ips(node, nodes):
            tasks.append(
                MrRemotePingTask(
                    task_id=f"{source.node_name}_ping_{node.node_name}_{target_ip}",
                    source_node=source.node_name,
                    source_host=source_host,
                    target_node=node.node_name,
                    target_ip=target_ip,
                    direction="local",
                    command=build_h3c_ping_command(target_ip, packet_count=CAR_NETWORK_QUICK_PING_COUNT),
                    packet_count=CAR_NETWORK_QUICK_PING_COUNT,
                    layer="MR远程本端检测",
                    timeout=CAR_NETWORK_QUICK_PING_TIMEOUT,
                )
            )
    for node in remote:
        for target_ip in _vehicle_candidate_ips(node, nodes):
            tasks.append(
                MrRemotePingTask(
                    task_id=f"{source.node_name}_ping_{node.node_name}_{target_ip}",
                    source_node=source.node_name,
                    source_host=source_host,
                    target_node=node.node_name,
                    target_ip=target_ip,
                    direction="cross_tc",
                    command=build_h3c_ping_command(target_ip, packet_count=CAR_NETWORK_QUICK_PING_COUNT),
                    packet_count=CAR_NETWORK_QUICK_PING_COUNT,
                    layer="MR远程跨TC快速检测",
                    timeout=CAR_NETWORK_QUICK_PING_TIMEOUT,
                )
            )
    return tuple(tasks)


def _cross_tc_ping_task(nodes: list[CarNetworkNode], ssh_results: dict[str, SshResult]) -> MrRemotePingTask | None:
    candidates: list[tuple[int, MrRemotePingTask]] = []
    for source_name, target_name in (("TC1-MR", "TC2-MR"), ("TC2-MR", "TC1-MR")):
        ssh = ssh_results.get(source_name)
        if ssh is None or not ssh.ok or _ssh_skipped(ssh):
            continue
        source = next((node for node in nodes if node.node_name == source_name), None)
        target = next((node for node in nodes if node.node_name == target_name), None)
        if source is None or target is None:
            continue
        target_ip = next(iter(_vehicle_candidate_ips(target, nodes)), "")
        if not target_ip:
            continue
        task = MrRemotePingTask(
            task_id=f"{source.node_name}_cross_tc_ping_{target.node_name}_{target_ip}",
            source_node=source.node_name,
            source_host=ssh.host,
            target_node=target.node_name,
            target_ip=target_ip,
            direction="cross_tc_loss",
            command=build_h3c_ping_command(target_ip, packet_count=CAR_NETWORK_CROSS_TC_PING_COUNT),
            packet_count=CAR_NETWORK_CROSS_TC_PING_COUNT,
            layer="跨TC通信",
            timeout=CAR_NETWORK_CROSS_TC_PING_TIMEOUT,
        )
        candidates.append((_cross_tc_direction_score(source, target, ssh, nodes), task))
    if not candidates:
        return None
    return min(candidates, key=lambda item: item[0])[1]


def _cross_tc_direction_score(source: CarNetworkNode, target: CarNetworkNode, ssh: SshResult, nodes: list[CarNetworkNode]) -> int:
    peer_results: list[PingResult] = []
    cross_results: list[PingResult] = []
    all_results = {**ssh.command_results}
    all_results.update(ssh.task_results)
    for key, ping in all_results.items():
        meta = ssh.task_metadata.get(key, {})
        target_node = str(meta.get("target_node") or "")
        if not target_node and f"_ping_{target.node_name}_" in str(key):
            target_node = target.node_name
        if not target_node:
            matched = _node_for_vehicle_ping_ip(nodes, ping.ip)
            target_node = matched.node_name if matched is not None else ""
        if target_node == target.node_name:
            peer_results.append(ping)
        elif target_node.startswith(f"{target.tc}-"):
            cross_results.append(ping)
    if any(_ping_ok_no_loss(ping) for ping in peer_results):
        return 0
    if any(_ping_ok_no_loss(ping) for ping in cross_results):
        return 1
    if any(ping.ok and ping.loss_percent > 0 for ping in peer_results):
        return 2
    if any(ping.ok and ping.loss_percent > 0 for ping in cross_results):
        return 3
    if not peer_results and not cross_results:
        return 4
    return 5


def _node_for_vehicle_ping_ip(nodes: list[CarNetworkNode], ip: str) -> CarNetworkNode | None:
    return next((node for node in nodes if ip in _vehicle_candidate_ips(node, nodes)), None)


def _cross_tc_ping_from_results(nodes: list[CarNetworkNode], ssh_results: dict[str, SshResult]) -> dict[str, object]:
    task = _cross_tc_ping_task(nodes, ssh_results)
    if task is None:
        return {"status": "skipped", "source": "", "target": "", "target_ip": "", "loss_percent": None, "avg_rtt_ms": None, "command": "", "note": "无可登录 MR，跨TC通信未检测"}
    ssh = ssh_results.get(task.source_node)
    ping = ssh.task_results.get(task.task_id) if ssh is not None else None
    if ping is None and ssh is not None:
        matches = [value for task_id, value in ssh.task_results.items() if f"_ping_{task.target_node}_" in task_id]
        ping = matches[0] if matches else ssh.command_results.get(task.target_ip)
    if ping is None:
        return {"status": "skipped", "source": task.source_node, "target": task.target_node, "target_ip": task.target_ip, "loss_percent": None, "avg_rtt_ms": None, "command": task.command, "note": "跨TC通信未检测"}
    status = "ok" if _ping_ok_no_loss(ping) else "loss" if ping.ok and ping.loss_percent > 0 else "fail"
    return {
        "status": status,
        "source": task.source_node,
        "target": task.target_node,
        "target_ip": task.target_ip,
        "loss_percent": ping.loss_percent,
        "avg_rtt_ms": ping.avg_rtt_ms,
        "command": task.command,
        "note": _ping_result_message(task.source_node, task.target_node, task.target_ip, ping, via_ssh=True, command=task.command),
    }


def _vehicle_candidate_ips(node: CarNetworkNode, nodes: list[CarNetworkNode]) -> tuple[str, ...]:
    candidates: list[str] = []
    if node.ip_vehicle:
        candidates.append(node.ip_vehicle)
    if _normalize_address_role(node.primary_address_role) in {"vehicle_ip", "all"} and node.primary_address:
        candidates.append(node.primary_address)
    if _normalize_address_role(node.backup_address_role) in {"vehicle_ip", "all"} and node.backup_address:
        candidates.append(node.backup_address)
    prefix = infer_vehicle_prefix(nodes, node.train_no)
    for address in (node.primary_address, node.backup_address, node.ssh_host, node.ip_uplink):
        if prefix and address and _ip_prefix(address) == prefix:
            candidates.append(address)
    if not candidates and prefix and node.is_mr:
        host = 249 if node.tc == "TC1" else 250 if node.tc == "TC2" else None
        if host is not None:
            candidates.append(_build_host_ip(prefix, host))
    return tuple(dict.fromkeys(candidates))


def _output_contains_train_mr_peer(output: str, train_no: str) -> bool:
    train_no = normalize_train_no(train_no)
    if not train_no:
        return False
    for token in re.split(r"\s+", _strip_ansi(str(output or ""))):
        identity = parse_train_identity(token)
        if identity is not None and identity.train_no == train_no:
            return True
    return False


def _node_state(
    node: CarNetworkNode,
    ping_results: dict[str, PingResult],
    ssh_results: dict[str, SshResult],
    ac_status: AcApStatus,
    train_ac_status: TrainAcStatus | None = None,
    all_nodes: list[CarNetworkNode] | None = None,
) -> str:
    if node.is_mr:
        ping_values = [ping_results[ip] for ip in node.ping_ips if ip in ping_results]
        ssh = ssh_results.get(node.node_name)
        remote_reachable, remote_ping, _remote_note = _mr_vehicle_reachable_by_any_entry(node, ssh_results, all_nodes or [node])
        if remote_ping is not None and remote_ping.ok and remote_ping.loss_percent > 0:
            return "unstable"
        if remote_reachable:
            return "ok"
        ac_end_online = False
        if train_ac_status is not None:
            ac_end_online = train_ac_status.tc1_mr_online if node.tc == "TC1" else train_ac_status.tc2_mr_online
        online_evidence = ac_end_online or (ac_status.online and train_ac_status is None) or bool(ssh and ssh.ok) or any(_ping_ok_no_loss(result) for result in ping_values)
        if any(result.ok and result.loss_percent > 0 for result in ping_values):
            return "unstable"
        if online_evidence:
            if any(not result.ok for result in ping_values):
                return "unstable"
            return "ok"
        if ping_values or ssh is not None or ac_status.selected:
            return "fail"
        return "pending"
    if node.ip_vehicle:
        results = _remote_results_for_node(node, ssh_results)
        if node.ip_vehicle in ping_results and not _ping_ok_no_loss(ping_results[node.ip_vehicle]):
            results = [ping_results[node.ip_vehicle]]
        elif not results and node.ip_vehicle in ping_results:
            results = [ping_results[node.ip_vehicle]]
        if not results:
            return "pending"
        if any(result.ok and result.loss_percent > 0 for result in results):
            return "unstable"
        if any(result.ok and result.loss_percent == 0 for result in results):
            return "ok"
        return "fail"
    return "not_applicable"


def _remote_results_for_node(node: CarNetworkNode, ssh_results: dict[str, SshResult]) -> list[PingResult]:
    own_mr = f"{node.tc}-MR" if node.tc else ""
    own_result = ssh_results.get(own_mr)
    if own_result is not None:
        matches = [ping for task_id, ping in own_result.task_results.items() if f"_ping_{node.node_name}_" in task_id]
        if matches:
            return matches
    return [ping for ssh in ssh_results.values() for task_id, ping in ssh.task_results.items() if f"_ping_{node.node_name}_" in task_id]


def _mr_vehicle_reachable_by_any_entry(
    node: CarNetworkNode,
    ssh_results: dict[str, SshResult],
    nodes: list[CarNetworkNode],
) -> tuple[bool, PingResult | None, str]:
    candidates = set(_vehicle_candidate_ips(node, nodes))
    if not candidates:
        return False, None, ""
    first_failure: tuple[PingResult, str] | None = None
    first_loss: tuple[PingResult, str] | None = None
    for ssh in ssh_results.values():
        if not ssh.ok:
            continue
        all_results = {**ssh.command_results}
        all_results.update(ssh.task_results)
        for key, ping in all_results.items():
            meta = ssh.task_metadata.get(key, {})
            target_ip = str(meta.get("target_ip") or ping.ip or key)
            if target_ip not in candidates:
                continue
            source = str(meta.get("source_node") or ssh.node_name or ssh.host)
            note = f"{source} CLI ping {node.node_name} 车内IP"
            if _ping_ok_no_loss(ping):
                return True, ping, f"{note}成功"
            if ping.ok and ping.loss_percent > 0 and first_loss is None:
                first_loss = (ping, f"{note}存在丢包")
            elif not ping.ok and first_failure is None:
                first_failure = (ping, f"{note}失败")
    if first_loss is not None:
        return False, first_loss[0], first_loss[1]
    if first_failure is not None:
        return False, first_failure[0], first_failure[1]
    return False, None, ""


def _cross_status(nodes: list[CarNetworkNode], ssh_results: dict[str, SshResult]) -> dict[str, str]:
    tc1_targets = {ip for node in nodes if node.tc == "TC2" and not node.is_mr for ip in _vehicle_candidate_ips(node, nodes)}
    tc2_targets = {ip for node in nodes if node.tc == "TC1" and not node.is_mr for ip in _vehicle_candidate_ips(node, nodes)}
    tc1 = ssh_results.get("TC1-MR")
    tc2 = ssh_results.get("TC2-MR")
    return {
        "TC1->TC2": _remote_ping_status(tc1, tc1_targets),
        "TC2->TC1": _remote_ping_status(tc2, tc2_targets),
    }


def _remote_ping_status(result: SshResult | None, targets: set[str]) -> str:
    if not targets:
        return "not_applicable"
    if result is None:
        return "pending"
    if _ssh_skipped(result):
        return "skipped"
    if not result.ok:
        return "ssh_failed"
    relevant = [ping for ip, ping in result.command_results.items() if ip in targets]
    if not relevant:
        return "pending"
    if any(_ping_ok_no_loss(ping) for ping in relevant):
        return "ok"
    if any(ping.ok and ping.loss_percent > 0 for ping in relevant):
        return "unstable"
    return "fail"


def _ping_ok_no_loss(result: PingResult | None) -> bool:
    return bool(result and result.ok and result.loss_percent == 0)


def _ssh_skipped(result: SshResult | None) -> bool:
    return bool(result and str(result.error).startswith("跳过："))


def _ssh_failed_but_mr_reachable(
    mr_nodes: list[CarNetworkNode],
    core_results: dict[str, PingResult],
    ssh_results: dict[str, SshResult],
) -> bool:
    for node in mr_nodes:
        ssh = ssh_results.get(node.node_name)
        if ssh is None or ssh.ok or _ssh_skipped(ssh):
            continue
        if node.ip_uplink and any(key.endswith(f"->{node.ip_uplink}") and _ping_ok_no_loss(ping) for key, ping in core_results.items()):
            return True
    return False


def _vehicle_loss(result: PingResult | None) -> bool:
    return bool(result and result.ok and result.loss_percent > 0)


def _vehicle_ping_bad(node: CarNetworkNode, ping_results: dict[str, PingResult]) -> bool:
    if not node.ip_vehicle:
        return False
    result = ping_results.get(node.ip_vehicle)
    return bool(result and (not result.ok or result.loss_percent > 0))


def _ends_json(nodes: list[CarNetworkNode], node_states: dict[str, str]) -> dict[str, dict[str, object]]:
    result: dict[str, dict[str, object]] = {}
    for tc, end, name in (("TC1", "CT", "TC1-MR"), ("TC2", "CW", "TC2-MR")):
        node = next((item for item in nodes if item.node_name == name), None)
        result[tc] = {
            "end": end,
            "mr_device_name": node.device_name if node else "",
            "mr_host": node.ssh_host or node.ip_uplink or node.ip_vehicle if node else "",
            "status": node_states.get(name, "pending"),
        }
    return result


def _vrrp_json(nodes: list[CarNetworkNode], cross: dict[str, str]) -> dict[str, object]:
    vrrp_ip = next((node.vrrp_ip for node in nodes if node.vrrp_ip), "")
    status = "ok" if all(value in {"ok", "not_applicable"} for value in cross.values()) else "fail"
    return {"ip": vrrp_ip, "status": status}


def _device_id(device: Device | None) -> str:
    return "" if device is None or device.id is None else str(device.id)


def _device_train_no(device: Device, known_train_nos: set[str]) -> str:
    text = " ".join(str(value or "") for value in (device.name, device.system_name, device.station, device.remark))
    parsed = normalize_train_no(text)
    if parsed in known_train_nos:
        return parsed
    if len(known_train_nos) == 1:
        return next(iter(known_train_nos))
    return parsed if parsed in known_train_nos else ""


def _sw_node_name(device: Device) -> str:
    text = " ".join(str(value or "") for value in (device.name, device.system_name, device.station, device.remark)).upper()
    if "TC1-SW" in text or "3SW-CT" in text or "CT" in text or "车头" in text:
        return "TC1-SW"
    if "TC2-SW" in text or "3SW-CW" in text or "CW" in text or "车尾" in text:
        return "TC2-SW"
    return ""
