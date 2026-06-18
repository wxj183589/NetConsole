from __future__ import annotations

from pathlib import Path

from PySide6.QtCore import Qt
from PySide6.QtWidgets import QComboBox, QFileDialog, QHBoxLayout, QPushButton, QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget

from netconsole.core.i18n import I18n
from netconsole.ui.pagination import DEFAULT_PAGE_SIZE, paginate_rows
from netconsole.ui.theme.table_style_engine import set_table_column_fields
from netconsole.ui.table_utils import auto_resize_table_columns, configure_readonly_table, create_table_context_menu
from netconsole.ui.widgets.pagination_widget import PaginationWidget


STATION_ONLINE_HISTORY_COLUMNS = (
    ("history.collected_at", "collected_at"),
    ("ac.station", "site_name"),
    ("ac.ap_total", "ap_total"),
    ("ac.online", "online_count"),
    ("ac.offline", "offline_count"),
    ("ac.online_rate", "online_rate"),
    ("field.remark", "remark"),
)


def export_station_online_history_xlsx(path: Path, rows: list[dict[str, object | None]], headers: list[str]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "AP Online History"
    alignment = Alignment(horizontal="center", vertical="center")
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = alignment
    sheet.freeze_panes = "A2"
    for row in rows:
        sheet.append([str(row.get(field) or "") for _key, field in STATION_ONLINE_HISTORY_COLUMNS])
        for cell in sheet[sheet.max_row]:
            cell.alignment = alignment
    for column_index in range(1, sheet.max_column + 1):
        letter = get_column_letter(column_index)
        width = max(len(str(cell.value or "")) for cell in sheet[letter]) + 2
        sheet.column_dimensions[letter].width = min(width, 48)
    workbook.save(path)


class StationOnlineHistoryDialog(QWidget):
    def __init__(self, i18n: I18n, rows: list[dict[str, object | None]], site_name: str | None = None) -> None:
        super().__init__()
        self.i18n = i18n
        self.rows = rows
        self.site_name = site_name
        self.page = 1
        self.page_size = DEFAULT_PAGE_SIZE

        self.station_filter = QComboBox()
        self.export_button = QPushButton()
        self.table = QTableWidget()
        self.pagination = PaginationWidget(self.i18n)
        set_table_column_fields(self.table, [field for _key, field in STATION_ONLINE_HISTORY_COLUMNS])
        configure_readonly_table(self.table)
        self.table.setColumnCount(len(STATION_ONLINE_HISTORY_COLUMNS))
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)

        actions = QHBoxLayout()
        actions.addWidget(self.station_filter)
        actions.addWidget(self.export_button)
        actions.addStretch(1)
        layout = QVBoxLayout()
        layout.addLayout(actions)
        layout.addWidget(self.table, 1)
        layout.addWidget(self.pagination)
        self.setLayout(layout)

        self.station_filter.currentIndexChanged.connect(self.filter_changed)
        self.export_button.clicked.connect(self.export_history)
        self.pagination.pageChanged.connect(self.set_page)
        self.pagination.pageSizeChanged.connect(self.set_page_size)
        self.retranslate()
        self.populate_station_filter()
        self.refresh_table()
        self.resize(920, 520)

    def retranslate(self) -> None:
        self.setWindowTitle(self.i18n.t("ac.online_history"))
        self.export_button.setText(self.i18n.t("ac.export_table"))
        self.table.setHorizontalHeaderLabels([self.i18n.t(key) for key, _field in STATION_ONLINE_HISTORY_COLUMNS])

    def populate_station_filter(self) -> None:
        self.station_filter.blockSignals(True)
        self.station_filter.clear()
        self.station_filter.addItem(self.i18n.t("field.all"), "")
        for site in sorted({str(row.get("site_name") or "") for row in self.rows if row.get("site_name")}):
            self.station_filter.addItem(site, site)
        if self.site_name:
            index = self.station_filter.findData(self.site_name)
            if index >= 0:
                self.station_filter.setCurrentIndex(index)
        self.station_filter.blockSignals(False)

    def filtered_rows(self) -> list[dict[str, object | None]]:
        site = str(self.station_filter.currentData() or "")
        if not site:
            return self.rows
        return [row for row in self.rows if str(row.get("site_name") or "") == site]

    def refresh_table(self) -> None:
        rows, state = paginate_rows(self.filtered_rows(), self.page_size, self.page)
        self.page = state.current_page
        self.pagination.set_state(state)
        self.table.setUpdatesEnabled(False)
        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(rows))
        for row_index, row in enumerate(rows):
            for column_index, (_key, field) in enumerate(STATION_ONLINE_HISTORY_COLUMNS):
                item = QTableWidgetItem(str(row.get(field) or ""))
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.table.setItem(row_index, column_index, item)
        self.table.setSortingEnabled(False)
        self.table.setUpdatesEnabled(True)
        auto_resize_table_columns(self.table)

    def filter_changed(self) -> None:
        self.page = 1
        self.refresh_table()

    def set_page(self, page: int) -> None:
        self.page = page
        self.refresh_table()

    def set_page_size(self, page_size: int) -> None:
        self.page_size = page_size
        self.page = 1
        self.refresh_table()

    def show_context_menu(self, position) -> None:
        index = self.table.indexAt(position)
        menu = create_table_context_menu(self.table, index.row(), index.column(), self.i18n.language, include_history=False)
        menu.exec(self.table.viewport().mapToGlobal(position))

    def export_history(self) -> None:
        path, _ = QFileDialog.getSaveFileName(self, self.i18n.t("ac.export_table"), "AP上线历史.xlsx", "Excel Files (*.xlsx)")
        if not path:
            return
        export_station_online_history_xlsx(Path(path), self.filtered_rows(), [self.i18n.t(key) for key, _field in STATION_ONLINE_HISTORY_COLUMNS])
