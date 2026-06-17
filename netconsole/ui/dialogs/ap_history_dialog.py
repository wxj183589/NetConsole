from __future__ import annotations

from pathlib import Path

from PySide6.QtCore import Qt
from PySide6.QtGui import QColor
from PySide6.QtWidgets import QFileDialog, QHBoxLayout, QPushButton, QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget

from netconsole.core.i18n import I18n
from netconsole.ui.pagination import DEFAULT_PAGE_SIZE, paginate_rows
from netconsole.ui.table_utils import auto_resize_table_columns, configure_readonly_table, create_table_context_menu
from netconsole.ui.widgets.pagination_widget import PaginationWidget
from netconsole.utils.optical_status import display_optical_status


AP_RADIO_HISTORY_COLUMNS = (
    ("history.collected_at", "collected_at"),
    ("ac.ap_name", "ap_name"),
    ("RID", "rid"),
    ("ap.channel", "channel"),
    ("ap.bandwidth", "bandwidth"),
    ("ap.tx_power", "tx_power"),
    ("history.raw_log_path", "raw_log_path"),
)
AP_LLDP_HISTORY_COLUMNS = (
    ("history.collected_at", "collected_at"),
    ("details.local_interface", "local_interface"),
    ("ac.lldp_neighbor", "lldp_neighbor"),
    ("ap.neighbor_interface", "neighbor_interface"),
    ("ap.neighbor_mac", "neighbor_mac"),
    ("ap.neighbor_device_name", "neighbor_device_name"),
    ("history.raw_log_path", "raw_log_path"),
)
AP_OPTICAL_HISTORY_COLUMNS = (
    ("history.collected_at", "collected_at"),
    ("ap.interface", "interface_name"),
    ("ap.optical_alarm_status", "optical_alarm_status"),
    ("ap.temperature", "temperature"),
    ("details.voltage", "voltage"),
    ("details.bias_current", "bias_current"),
    ("ap.tx_power", "tx_power"),
    ("ap.rx_power", "rx_power"),
    ("details.rx_low_alarm", "rx_low_alarm"),
    ("details.rx_high_alarm", "rx_high_alarm"),
    ("details.tx_low_alarm", "tx_low_alarm"),
    ("details.tx_high_alarm", "tx_high_alarm"),
    ("details.rx_low_warning", "rx_low_warning"),
    ("details.rx_high_warning", "rx_high_warning"),
    ("details.tx_low_warning", "tx_low_warning"),
    ("details.tx_high_warning", "tx_high_warning"),
    ("details.module_model", "module_model"),
    ("details.module_serial_number", "module_serial_number"),
    ("details.vendor", "module_vendor"),
    ("details.wavelength", "wavelength"),
    ("details.transmission_distance", "transmission_distance"),
    ("details.connector_type", "connector_type"),
    ("history.raw_log_path", "raw_log_path"),
)
OPTICAL_HISTORY_COLORS = {
    "normal": "#dcfce7",
    "warning": "#fef9c3",
    "alarm": "#fee2e2",
    "link_abnormal": "#ffe4e6",
    "no_light": "#e5e7eb",
    "skipped": "#f3f4f6",
}


def export_ap_history_xlsx(path: Path, rows: list[dict[str, object | None]], columns: tuple[tuple[str, str], ...], headers: list[str], color_field: str | None = None) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "AP History"
    alignment = Alignment(horizontal="center", vertical="center")
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = alignment
    sheet.freeze_panes = "A2"
    for row in rows:
        sheet.append([_history_display_value(row, field, color_field) for _key, field in columns])
        fill = None
        if color_field:
            color = OPTICAL_HISTORY_COLORS.get(str(row.get(color_field) or ""))
            fill = PatternFill(fill_type="solid", fgColor=color.lstrip("#").upper()) if color else None
        for cell in sheet[sheet.max_row]:
            cell.alignment = alignment
            if fill:
                cell.fill = fill
    for column_index in range(1, sheet.max_column + 1):
        letter = get_column_letter(column_index)
        width = max(len(str(cell.value or "")) for cell in sheet[letter]) + 2
        sheet.column_dimensions[letter].width = min(width, 56)
    workbook.save(path)


class ApHistoryDialog(QWidget):
    def __init__(
        self,
        i18n: I18n,
        ap_name: str,
        history_type: str,
        rows: list[dict[str, object | None]],
        columns: tuple[tuple[str, str], ...],
        color_field: str | None = None,
        parent=None,
    ) -> None:
        super().__init__(parent)
        self.i18n = i18n
        self.ap_name = ap_name
        self.history_type = history_type
        self.rows = rows
        self.columns = columns
        self.color_field = color_field
        self.page = 1
        self.page_size = DEFAULT_PAGE_SIZE
        self.setAttribute(Qt.WA_DeleteOnClose, True)
        self.setWindowTitle(self.i18n.t("history.title_with_object", device=ap_name, object=history_type))
        self.resize(900, 560)

        self.back_button = QPushButton()
        self.close_button = QPushButton()
        self.always_on_top_button = QPushButton()
        self.always_on_top_button.setCheckable(True)
        self.export_button = QPushButton()
        self.table = QTableWidget()
        self.pagination = PaginationWidget(self.i18n)
        configure_readonly_table(self.table)
        self.table.setColumnCount(len(columns))
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)

        actions = QHBoxLayout()
        actions.addWidget(self.back_button)
        actions.addWidget(self.close_button)
        actions.addStretch(1)
        actions.addWidget(self.export_button)
        actions.addWidget(self.always_on_top_button)
        layout = QVBoxLayout()
        layout.addLayout(actions)
        layout.addWidget(self.table)
        layout.addWidget(self.pagination)
        self.setLayout(layout)

        self.back_button.clicked.connect(self.return_to_parent)
        self.close_button.clicked.connect(self.close)
        self.always_on_top_button.toggled.connect(self.set_always_on_top)
        self.export_button.clicked.connect(self.export_history)
        self.pagination.pageChanged.connect(self.set_page)
        self.pagination.pageSizeChanged.connect(self.set_page_size)
        self.retranslate()
        self.refresh_table()

    def retranslate(self) -> None:
        self.back_button.setText(self.i18n.t("history.back"))
        self.close_button.setText(self.i18n.t("dialog.close"))
        self.always_on_top_button.setText(self.i18n.t("window.always_on_top"))
        self.export_button.setText(self.i18n.t("ac.export_table"))
        self.table.setHorizontalHeaderLabels([self.i18n.t(key) for key, _field in self.columns])

    def refresh_table(self) -> None:
        rows, state = paginate_rows(self.rows, self.page_size, self.page)
        self.page = state.current_page
        self.pagination.set_state(state)
        self.table.setUpdatesEnabled(False)
        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(rows))
        for row_index, row in enumerate(rows):
            color = OPTICAL_HISTORY_COLORS.get(str(row.get(self.color_field or "") or ""))
            for column_index, (_key, field) in enumerate(self.columns):
                item = QTableWidgetItem(_history_display_value(row, field, self.color_field, self.i18n.language))
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                if color:
                    item.setBackground(QColor(color))
                self.table.setItem(row_index, column_index, item)
        self.table.setSortingEnabled(False)
        self.table.setUpdatesEnabled(True)
        auto_resize_table_columns(self.table, column_min_widths={0: 170}, max_width=560)

    def set_page(self, page: int) -> None:
        self.page = page
        self.refresh_table()

    def set_page_size(self, page_size: int) -> None:
        self.page_size = page_size
        self.page = 1
        self.refresh_table()

    def show_context_menu(self, position) -> None:
        index = self.table.indexAt(position)
        menu = create_table_context_menu(self.table, index.row(), index.column(), self.i18n.language, include_history=False)
        menu.exec(self.table.viewport().mapToGlobal(position))

    def set_always_on_top(self, enabled: bool) -> None:
        self.setWindowFlag(Qt.WindowStaysOnTopHint, enabled)
        self.always_on_top_button.setText(self.i18n.t("window.cancel_always_on_top" if enabled else "window.always_on_top"))
        self.show()

    def return_to_parent(self) -> None:
        parent = self.parentWidget()
        self.close()
        if parent is not None:
            parent.show()
            parent.raise_()
            parent.activateWindow()

    def export_history(self) -> None:
        path, _ = QFileDialog.getSaveFileName(self, self.i18n.t("ac.export_table"), f"{self.ap_name}_{self.history_type}_history.xlsx", "Excel Files (*.xlsx)")
        if not path:
            return
        export_ap_history_xlsx(Path(path), self.rows, self.columns, [self.i18n.t(key) for key, _field in self.columns], self.color_field)


def _history_display_value(row: dict[str, object | None], field: str, color_field: str | None = None, language: str = "zh") -> str:
    if color_field and field == color_field:
        return display_optical_status(row.get(field), language)
    return str(row.get(field) or "")
