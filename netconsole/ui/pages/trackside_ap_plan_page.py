from __future__ import annotations

from netconsole.ui.dialogs.message_service import MessageBox
import csv
from datetime import datetime
from pathlib import Path
import re

from PySide6.QtCore import Qt, QTimer, Signal
from PySide6.QtWidgets import (
    QFileDialog,
    QHBoxLayout,
    QHeaderView,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from netconsole.core import app_logger
from netconsole.core.i18n import I18n
from netconsole.repositories.ac_repository import AcRepository, TRACKSIDE_AP_PLAN_MODE
from netconsole.utils.excel_workbook import load_workbook_without_unsupported_image_warning
from netconsole.services.trackside_ap_business import parse_vlan_set
from netconsole.ui.render.table_render_engine import apply_table_style, set_table_column_fields
from netconsole.ui.shell.fluent_bridge import FIF
from netconsole.ui.table.table_autosize_engine import apply_worksheet_autofit


TRACKSIDE_PLAN_COLUMNS = (
    ("ac.trackside_plan.station_name", "station_name"),
    ("ac.trackside_plan.ap_count", "ap_count"),
    ("ac.trackside_plan.ap_start_address", "ap_start_address"),
    ("ac.trackside_plan.mask", "mask_length"),
    ("ac.trackside_plan.ap_gateway", "ap_gateway"),
    ("ac.trackside_plan.ap_management_vlan", "ap_management_vlans"),
    ("field.remark", "remark"),
)
TRACKSIDE_PLAN_HEADERS = ["车站名称", "AP数量", "AP起始地址", "掩码", "AP网关", "AP管理VLAN", "备注"]
TRACKSIDE_PLAN_COLUMN_WIDTHS = {
    "station_name": 260,
    "ap_count": 90,
    "ap_start_address": 170,
    "mask_length": 140,
    "ap_gateway": 170,
    "ap_management_vlans": 170,
    "remark": 220,
}
MASK_ERROR_TEXT = "必须是0-32或合法连续IPv4掩码"


def _set_button_icon(button: QPushButton, icon: object | None) -> None:
    if icon is None:
        return
    icon_factory = getattr(icon, "icon", None)
    resolved_icon = icon_factory() if callable(icon_factory) else icon
    try:
        button.setIcon(resolved_icon)
    except TypeError:
        pass


class TracksideApPlanPage(QWidget):
    plan_saved = Signal()

    def __init__(self, repository: AcRepository, i18n: I18n, site_name: str) -> None:
        super().__init__()
        self.repository = repository
        self.i18n = i18n
        self.site_name = site_name
        self.table = QTableWidget()
        self.add_button = QPushButton()
        self.delete_button = QPushButton()
        self.save_button = QPushButton()
        self.import_button = QPushButton()
        self.export_button = QPushButton()
        self.template_button = QPushButton()
        self.refresh_button = QPushButton()
        self._dirty = False
        self._build_ui()
        self.retranslate()
        self.refresh()

    def _build_ui(self) -> None:
        actions = QHBoxLayout()
        for button in (self.add_button, self.delete_button, self.save_button, self.import_button, self.export_button, self.template_button, self.refresh_button):
            actions.addWidget(button)
        actions.addStretch(1)

        self.table.setColumnCount(len(TRACKSIDE_PLAN_COLUMNS))
        set_table_column_fields(self.table, [field for _key, field in TRACKSIDE_PLAN_COLUMNS])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.horizontalHeader().setStretchLastSection(False)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        apply_table_style(self.table)
        self._apply_column_layout()

        layout = QVBoxLayout(self)
        layout.addLayout(actions)
        layout.addWidget(self.table, 1)

        self.add_button.clicked.connect(self.add_row)
        self.delete_button.clicked.connect(self.delete_selected)
        self.save_button.clicked.connect(self.save_plan)
        self.import_button.clicked.connect(self.import_plan)
        self.export_button.clicked.connect(self.export_plan)
        self.template_button.clicked.connect(self.download_template)
        self.refresh_button.clicked.connect(self.refresh)
        self.table.itemChanged.connect(self._mark_dirty)

    def retranslate(self) -> None:
        self.add_button.setText(self.i18n.t("ac.trackside_plan.add_row"))
        self.delete_button.setText(self.i18n.t("ac.trackside_plan.delete_selected"))
        self.save_button.setText(self.i18n.t("ac.trackside_plan.save"))
        self.import_button.setText(self.i18n.t("ac.trackside_plan.import"))
        self.export_button.setText(self.i18n.t("ac.trackside_plan.export"))
        self.template_button.setText(self.i18n.t("ac.trackside_plan.template"))
        self.refresh_button.setText(self.i18n.t("details.refresh"))
        self._apply_button_icons()
        self.table.setHorizontalHeaderLabels([self.i18n.t(key) for key, _field in TRACKSIDE_PLAN_COLUMNS])

    def _apply_button_icons(self) -> None:
        button_icons = (
            (self.add_button, FIF.ADD),
            (self.delete_button, FIF.DELETE),
            (self.save_button, FIF.SAVE),
            (self.import_button, FIF.DOWNLOAD),
            (self.export_button, FIF.SHARE),
            (self.template_button, FIF.CLOUD_DOWNLOAD),
            (self.refresh_button, FIF.SYNC),
        )
        for button, icon in button_icons:
            _set_button_icon(button, icon)

    def refresh(self) -> None:
        if not self._confirm_discard_or_save_changes():
            return
        self.reload_plan_table()

    def add_row(self) -> None:
        row = self.table.rowCount()
        self.table.insertRow(row)
        for column in range(len(TRACKSIDE_PLAN_COLUMNS)):
            self.table.setItem(row, column, self._make_item(""))
        self._apply_column_layout()
        self._dirty = True

    def delete_selected(self) -> None:
        for row in sorted({index.row() for index in self.table.selectedIndexes()}, reverse=True):
            self.table.removeRow(row)
            self._dirty = True

    def save_plan(self) -> bool:
        try:
            rows = self._read_table_rows()
        except ValueError as exc:
            MessageBox.warning(self, self.i18n.t("ac.trackside_plan.validation_failed"), str(exc))
            return False
        self.repository.replace_trackside_ap_plan_rows(TRACKSIDE_AP_PLAN_MODE, rows)
        for row in rows:
            app_logger.log_info(
                "TRACKSIDE_AP_PLAN_SAVED",
                f"保存轨旁AP规划：station={row.get('station_name')}, ap_count={row.get('ap_count')}, vlan={row.get('ap_management_vlans')}",
            )
        self._dirty = False
        self.reload_plan_table()
        self.plan_saved.emit()
        return True

    def import_plan(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, self.i18n.t("ac.trackside_plan.import"), "", "Excel/CSV (*.xlsx *.csv)")
        if not path:
            return
        try:
            rows = _dedupe_station_rows(read_trackside_plan_file(Path(path)))
            self._validate_rows(rows)
        except ValueError as exc:
            MessageBox.warning(self, self.i18n.t("ac.trackside_plan.validation_failed"), str(exc))
            return
        self.repository.replace_trackside_ap_plan_rows(TRACKSIDE_AP_PLAN_MODE, rows)
        for row in rows:
            app_logger.log_info(
                "TRACKSIDE_AP_PLAN_SAVED",
                f"保存轨旁AP规划：station={row.get('station_name')}, ap_count={row.get('ap_count')}, vlan={row.get('ap_management_vlans')}",
            )
        self._dirty = False
        self.reload_plan_table()
        self.plan_saved.emit()

    def export_plan(self) -> None:
        path, _ = QFileDialog.getSaveFileName(self, self.i18n.t("ac.trackside_plan.export"), self._default_export_name(), "Excel (*.xlsx)")
        if not path:
            return
        export_trackside_plan_xlsx(Path(path), self.repository.list_trackside_ap_plan(TRACKSIDE_AP_PLAN_MODE))

    def download_template(self) -> None:
        path, _ = QFileDialog.getSaveFileName(self, self.i18n.t("ac.trackside_plan.template"), self._default_export_name(template=True), "Excel (*.xlsx)")
        if not path:
            return
        export_trackside_plan_xlsx(Path(path), [])

    def reload_plan_table(self) -> None:
        self._set_rows(self.repository.list_trackside_ap_plan(TRACKSIDE_AP_PLAN_MODE))
        self._dirty = False

    def _set_rows(self, rows: list[dict[str, object | None]]) -> None:
        self.table.blockSignals(True)
        self.table.setRowCount(len(rows))
        for row_index, row in enumerate(rows):
            for column_index, (_key, field) in enumerate(TRACKSIDE_PLAN_COLUMNS):
                self.table.setItem(row_index, column_index, self._make_item(str(row.get(field) or "")))
            app_logger.log_info(
                "TRACKSIDE_AP_PLAN_LOADED",
                f"加载轨旁AP规划：station={row.get('station_name')}, ap_count={row.get('ap_count')}, vlan={row.get('ap_management_vlans')}",
            )
        self.table.blockSignals(False)
        self._apply_column_layout()
        QTimer.singleShot(0, self._apply_column_layout)

    def _apply_column_layout(self) -> None:
        header = self.table.horizontalHeader()
        header.setStretchLastSection(False)
        header.setSectionResizeMode(QHeaderView.Interactive)
        fields = [field for _key, field in TRACKSIDE_PLAN_COLUMNS]
        fixed_width = sum(TRACKSIDE_PLAN_COLUMN_WIDTHS[field] for field in fields if field != "station_name")
        available = max(self.table.viewport().width(), 0)
        station_width = TRACKSIDE_PLAN_COLUMN_WIDTHS["station_name"]
        if available > fixed_width + station_width:
            station_width = available - fixed_width
        for column, field in enumerate(fields):
            width = station_width if field == "station_name" else TRACKSIDE_PLAN_COLUMN_WIDTHS[field]
            header.setSectionResizeMode(column, QHeaderView.Interactive)
            self.table.setColumnWidth(column, width)

    def _read_table_rows(self) -> list[dict[str, object | None]]:
        rows = []
        for row_index in range(self.table.rowCount()):
            row = {field: (self.table.item(row_index, column_index).text() if self.table.item(row_index, column_index) else "") for column_index, (_key, field) in enumerate(TRACKSIDE_PLAN_COLUMNS)}
            if not any(str(value or "").strip() for value in row.values()):
                continue
            row["sort_order"] = len(rows)
            rows.append(row)
        self._validate_rows(rows)
        return _dedupe_station_rows(rows)

    def _validate_rows(self, rows: list[dict[str, object | None]]) -> None:
        seen: set[str] = set()
        for index, row in enumerate(rows, start=2):
            station = str(row.get("station_name") or "").strip()
            if not station:
                raise ValueError(f"第{index}行 车站名称：必填")
            if station.casefold() in seen:
                continue
            seen.add(station.casefold())
            try:
                row["ap_count"] = int(str(row.get("ap_count") or "0").strip())
            except ValueError:
                raise ValueError(f"第{index}行 AP数量：必须是整数") from None
            if int(row["ap_count"]) < 0:
                raise ValueError(f"第{index}行 AP数量：必须是非负整数")
            mask_length = _parse_mask_length(row.get("mask_length"))
            if mask_length is None and str(row.get("mask_length") or "").strip():
                raise ValueError(f"第{index}行 掩码：{MASK_ERROR_TEXT}")
            row["mask_length"] = mask_length
            vlans = parse_vlan_set(row.get("ap_management_vlans"))
            if not vlans:
                raise ValueError(f"第{index}行 AP管理VLAN：必填")
            row["station_name"] = station
            row["ap_management_vlans"] = ",".join(str(vlan) for vlan in sorted(vlans))
            start = str(row.get("ap_start_address") or "").strip()
            gateway = str(row.get("ap_gateway") or "").strip()
            if start and not _valid_ipv4_or_placeholder(start):
                raise ValueError(f"第{index}行 AP起始地址：格式无效")
            if gateway and not _valid_ipv4(gateway):
                raise ValueError(f"第{index}行 AP网关：必须是IPv4")

    def _default_export_name(self, template: bool = False) -> str:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        suffix = "轨旁AP规划模板" if template else "轨旁AP规划"
        safe_site = re.sub(r'[\\/:*?"<>|]+', "_", str(self.site_name or "").strip()) or "site"
        return f"{safe_site}_{suffix}_{stamp}.xlsx"

    def _make_item(self, value: str) -> QTableWidgetItem:
        item = QTableWidgetItem(value)
        item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
        return item

    def _mark_dirty(self, _item: QTableWidgetItem) -> None:
        self._dirty = True

    def _confirm_discard_or_save_changes(self) -> bool:
        if not self._dirty:
            return True
        result = MessageBox.question(
            self,
            "轨旁AP规划",
            "当前轨旁AP规划有未保存修改，是否先保存？",
            MessageBox.StandardButton.Save | MessageBox.StandardButton.Discard | MessageBox.StandardButton.Cancel,
        )
        if result == MessageBox.StandardButton.Save:
            return self.save_plan()
        if result == MessageBox.StandardButton.Cancel:
            return False
        return result == MessageBox.StandardButton.Discard


def read_trackside_plan_file(path: Path) -> list[dict[str, object | None]]:
    if path.suffix.casefold() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return [_row_from_named(row) for row in csv.DictReader(handle)]
    workbook = load_workbook_without_unsupported_image_warning(path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    rows = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        raw = {headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))}
        if any(value not in (None, "") for value in raw.values()):
            rows.append(_row_from_named(raw))
    return rows


def export_trackside_plan_xlsx(path: Path, rows: list[dict[str, object | None]]) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "轨旁AP规划"
    sheet.append(TRACKSIDE_PLAN_HEADERS)
    for row in rows:
        sheet.append([row.get(field) or "" for _key, field in TRACKSIDE_PLAN_COLUMNS])
    apply_worksheet_autofit(sheet, maximum=50)
    workbook.save(path)


def _row_from_named(row: dict[object, object]) -> dict[str, object | None]:
    mapping = dict(zip(TRACKSIDE_PLAN_HEADERS, [field for _key, field in TRACKSIDE_PLAN_COLUMNS]))
    return {field: row.get(header, "") for header, field in mapping.items()}


def _dedupe_station_rows(rows: list[dict[str, object | None]]) -> list[dict[str, object | None]]:
    by_station: dict[str, dict[str, object | None]] = {}
    order: list[str] = []
    for row in rows:
        station = str(row.get("station_name") or "").strip()
        key = station.casefold()
        if not key:
            order.append(f"__blank_{len(order)}")
            by_station[order[-1]] = row
            continue
        if key not in by_station:
            order.append(key)
        by_station[key] = row
    result = [by_station[key] for key in order if key in by_station]
    for index, row in enumerate(result):
        row["sort_order"] = index
    return result


def _valid_ipv4(value: str) -> bool:
    parts = value.split(".")
    if len(parts) != 4:
        return False
    try:
        return all(0 <= int(part) <= 255 for part in parts)
    except ValueError:
        return False


def _valid_ipv4_or_placeholder(value: str) -> bool:
    parts = value.split(".")
    if len(parts) != 4:
        return False
    for part in parts:
        if part.upper() == "X":
            continue
        try:
            if int(part) < 0 or int(part) > 255:
                return False
        except ValueError:
            return False
    return True


def _parse_mask_length(value: object) -> int | None:
    text = "" if value is None else str(value).strip()
    if not text:
        return None
    if text.isdigit():
        prefix = int(text)
        return prefix if 0 <= prefix <= 32 else None
    if "." in text:
        return _dotted_netmask_to_prefix(text)
    return None


def _dotted_netmask_to_prefix(mask: str) -> int | None:
    parts = mask.split(".")
    if len(parts) != 4:
        return None
    octets: list[int] = []
    for part in parts:
        if not part.isdigit():
            return None
        value = int(part)
        if value < 0 or value > 255:
            return None
        octets.append(value)
    bits = "".join(f"{octet:08b}" for octet in octets)
    if re.fullmatch(r"1*0*", bits) is None:
        return None
    return bits.count("1")
