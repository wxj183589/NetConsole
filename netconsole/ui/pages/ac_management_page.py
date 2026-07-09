from __future__ import annotations

from netconsole.ui.dialogs.message_service import MessageBox
from netconsole.ui.dialogs.input_dialog_service import InputDialog
from datetime import datetime
from pathlib import Path
import uuid

from PySide6.QtCore import QSignalBlocker, QStandardPaths, Qt, QThread, QUrl, Signal
from PySide6.QtGui import QAction, QDesktopServices
from PySide6.QtWidgets import (
    QAbstractItemView,
    QComboBox,
    QDialog,
    QFileDialog,
    QFormLayout,
    QGridLayout,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMenu,
    QProgressBar,
    QPushButton,
    QScrollArea,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

from netconsole.core import app_logger
from netconsole.core.feature_flags import FeatureGate, apply_feature_to_widget, default_feature_gate
from netconsole.core.i18n import I18n
from netconsole.core.optical_severity_engine import display_optical_status
from netconsole.core.paths import PathResolver
from netconsole.core.settings import SettingsStore
from netconsole.core.sources.switch_source import (
    build_switch_data_lookup,
    compute_switch_status,
)
from netconsole.core.state_engine import STATUS_COLORS
from netconsole.models.device import Device
from netconsole.repositories.ac_repository import AcRepository
from netconsole.repositories.device_group_repository import DeviceGroupRepository
from netconsole.repositories.device_fact_repository import DeviceFactRepository
from netconsole.repositories.device_repository import DeviceRepository
from netconsole.adapters.h3c.h3c_command_profile import H3cAcCommandProfile
from netconsole.services.trackside_ap_business import (
    TRACKSIDE_AP_BUSINESS_COLUMNS,
    TRACKSIDE_AP_BUSINESS_HEADER_TOOLTIPS,
    build_trackside_ap_business_rows,
    build_trackside_site_filter_items,
    filter_station_switch_devices,
    filter_trackside_ap_business_rows,
    trackside_row_status,
)
from netconsole.services.offline_ap_ledger import (
    OFFLINE_AP_LEDGER_COLUMNS,
    OFFLINE_AP_STATS_COLUMNS,
    OFFLINE_AP_STATUS_TEXT,
    build_device_lookup_by_name,
    build_latest_ap_history_indexes,
    build_offline_ap_ledger,
    is_fit_ap_offline,
    load_offline_ap_cache,
    offline_ap_headers,
    save_offline_ap_cache,
)
from netconsole.services.ap_online_overview import (
    AP_ONLINE_OVERVIEW_COLUMNS,
    ApOnlineOverviewService,
    build_ap_online_overview_rows,
    export_ap_online_overview_xlsx,
)
from netconsole.services.background_job import BackgroundJob
from netconsole.services.background_process_manager import BackgroundProcessManager
from netconsole.services.fit_ap_import_export import FitApImportExportService, make_ap_extension_template_filename, make_fit_ap_export_filename
from netconsole.services.fit_ap_optical_export import (
    OPTICAL_EXPORT_COLOR_RGB,
    evaluate_fit_ap_ap_status,
    evaluate_fit_ap_row_status,
    evaluate_fit_ap_switch_status,
    export_fit_ap_optical_xlsx,
    fit_ap_optical_export_value as _fit_ap_optical_export_value,
)
from netconsole.services.fit_ap_link_info import lldp_display_status, lldp_source_label
from netconsole.services.omnipeek_name_table_service import OmniPeekNameTableService, default_omnipeek_line_name
from netconsole.models.omnipeek_name_table import SOURCE_DEVICE_MANAGEMENT
from netconsole.services.export import ExportJob, ExportTaskSpec
from netconsole.services.export.export_task_builders import (
    ap_online_overview_xlsx_spec,
    fit_ap_csv_spec,
    fit_ap_extension_template_xlsx_spec,
    fit_ap_extension_xlsx_spec,
)
from netconsole.services.export.export_process_manager import ExportProcessManager
from netconsole.services.external_terminal import TERMINAL_LABELS, available_external_terminal_configs, launch_external_terminal
from netconsole.services.device_web_service import DEFAULT_HTTPS_PORT, build_https_url, effective_https_port, open_https_url
from netconsole.ui.ac_collect_worker import AcCommandActionThread, AcInfoCollectThread, AcResourceCollectThread, FitApOpticalCollectThread
from netconsole.ui.app_events import app_events
from netconsole.ui.dialogs.device_detail_dialog import DeviceDetailDialog
from netconsole.ui.dialogs.fit_ap_detail_dialog import FitApDetailDialog
from netconsole.ui.dialogs.omnipeek_export_dialog import OmniPeekExportDialog
from netconsole.ui.export_action_helper import submit_export_task
from netconsole.ui.dialogs.station_online_history_dialog import StationOnlineHistoryDialog
from netconsole.ui.pages.trackside_ap_plan_page import TracksideApPlanPage
from netconsole.ui.pagination import DEFAULT_PAGE_SIZE, PaginationState, paginate_rows
from netconsole.ui.render.table_render_engine import STATUS_COLOR_MAP, apply_table_style, set_table_column_fields
from netconsole.ui.shell.fluent_bridge import FIF
from netconsole.ui.theme.contrast_engine import apply_item_contrast, apply_status_item_contrast
from netconsole.ui.export_path import CSV_FILTER, EXCEL_FILTER, remember_export_path, select_export_path
from netconsole.ui.table_utils import auto_fit_table_columns, auto_resize_table_columns, create_table_context_menu, configure_readonly_table, make_text_selectable
from netconsole.ui.widgets.table_check_delegate import create_checkable_table_item, install_checkbox_only_delegate, invert_table_rows_checked, is_checked_value, set_all_table_rows_checked
from netconsole.ui.widgets.pagination_widget import PaginationWidget
from netconsole.ui.window_popup_service import show_non_focus_window
from netconsole.utils.interface_sort import interface_sort_key
from netconsole.utils.mileage import format_track_mileage, mileage_search_tokens, parse_mileage_to_meters



CHECK_COLUMN = 0
SUMMARY_FIELDS = (
    ("details.model", "model"),
    ("details.serial_number", "serial_number"),
    ("details.software_version", "software_version"),
    ("field.https_port", "https_port"),
    ("ac.total_aps", "total_aps"),
    ("ac.online_aps", "online_aps"),
    ("ac.offline_aps", "offline_aps"),
    ("ac.ap_licenses", "total_ap_licenses"),
    ("ac.local_ap_licenses", "local_ap_licenses"),
    ("ac.remaining_local_ap_licenses", "remaining_local_ap_licenses"),
    ("ac.cpu_usage", "cpu_usage"),
    ("ac.memory_usage", "memory_usage"),
    ("field.updated_at", "updated_at"),
)

FIT_AP_RESOURCE_COLUMNS = (
    ("", "select"),
    ("ac.ap_name", "ap_name"),
    ("APID", "apid"),
    ("AP_IP", "ap_ip"),
    ("AP_MAC", "ap_mac"),
    ("details.model", "model"),
    ("field.status", "state_display"),
    ("AP组", "group_name"),
    ("ac.online_time", "online_time"),
    ("RID1信道", "rid1_channel"),
    ("RID1频宽", "rid1_bandwidth"),
    ("RID1功率", "rid1_tx_power"),
    ("RID2信道", "rid2_channel"),
    ("RID2频宽", "rid2_bandwidth"),
    ("RID2功率", "rid2_tx_power"),
    ("ac.site", "site"),
    ("归属区间", "section_name"),
    ("归属类型", "belong_type"),
    ("ac.mileage", "mileage"),
    ("ac.location_note", "location_note"),
    ("ac.direction", "direction"),
    ("field.updated_at", "updated_at"),
)

FIT_AP_RESOURCE_COLUMN_MIN_WIDTHS = {
    "select": 48,
    "ap_name": 150,
    "apid": 70,
    "ap_ip": 120,
    "ap_mac": 140,
    "model": 100,
    "state_display": 90,
    "group_name": 110,
    "online_time": 100,
    "rid1_channel": 80,
    "rid1_bandwidth": 80,
    "rid1_tx_power": 80,
    "rid2_channel": 80,
    "rid2_bandwidth": 80,
    "rid2_tx_power": 80,
    "site": 130,
    "section_name": 160,
    "belong_type": 90,
    "mileage": 90,
    "location_note": 180,
    "direction": 70,
    "updated_at": 170,
}

FIT_AP_RESOURCE_COLUMN_MAX_WIDTHS = {
    "select": 54,
    "ap_name": 320,
    "ap_mac": 170,
    "section_name": 260,
    "location_note": 360,
    "updated_at": 240,
}

AP_EXTENSION_COLUMNS = (
    ("ID", "id"),
    ("归属类型", "belong_type"),
    ("归属站点", "station_name"),
    ("归属区间", "section_name"),
    ("区间起点站", "section_start_station"),
    ("区间终点站", "section_end_station"),
    ("场段", "yard_name"),
    ("区域", "area_name"),
    ("网络", "network_domain"),
    ("线别", "line_side"),
    ("方向", "direction"),
    ("里程", "mileage_text"),
    ("点位说明", "location_desc"),
    ("AP编号", "ap_point_code"),
    ("AP名称", "ap_name"),
    ("AP MAC", "ap_mac_display"),
    ("距上一个AP", "distance_to_prev_m"),
    ("曲线半径", "curve_radius_m"),
    ("供电站", "power_station"),
    ("电源分配", "power_distribution"),
    ("光缆接入站", "fiber_access_station"),
    ("光缆分配", "fiber_distribution"),
    ("匹配状态", "match_status"),
    ("备注", "remark"),
)

AP_EXTENSION_SUMMARY_COLUMNS = (
    ("车站", "station_name"),
    ("AP点位数", "total"),
    ("已绑定MAC", "bound"),
    ("未绑定MAC", "unbound"),
    ("左线数量", "left"),
    ("右线数量", "right"),
    ("曲线段AP数", "curve"),
    ("间隔异常数", "risk"),
)

AP_EXTENSION_ISSUE_COLUMNS = (
    ("类型", "type"),
    ("级别", "severity"),
    ("来源行", "source_row"),
    ("AP MAC", "ap_mac_norm"),
    ("说明", "message"),
)

FIT_AP_OPTICAL_COLUMNS = (
    ("ac.ap_name", "ap_name"),
    ("ac.ap_mac", "ap_mac"),
    ("ac.station", "site"),
    ("ac.indoor_switch", "neighbor_device_name"),
    ("ac.indoor_port", "neighbor_interface"),
    ("ac.indoor_switch_rx_power", "neighbor_rx_power"),
    ("fit_ap.switch_optical_status", "switch_optical_status"),
    ("ac.ap_side_rx_power", "rx_power"),
    ("ap.optical_alarm_status", "optical_alarm_status"),
    ("field.updated_at", "updated_at"),
)

FIT_AP_OPTICAL_DETAIL_COLUMNS = (
    ("ac.ap_name", "ap_name"),
    ("field.ip_address", "ap_ip"),
    ("ac.site", "site"),
    ("ac.lldp_neighbor", "lldp_neighbor"),
    ("ap.neighbor_interface", "neighbor_interface"),
    ("ap.neighbor_mac", "neighbor_mac"),
    ("ap.neighbor_device_name", "neighbor_device_name"),
    ("ap.neighbor_rx_power", "neighbor_rx_power"),
    ("ap.interface", "interface_name"),
    ("ap.temperature", "temperature"),
    ("ap.tx_power", "tx_power"),
    ("ap.rx_power", "rx_power"),
    ("ap.optical_alarm_status", "optical_alarm_status"),
    ("field.updated_at", "updated_at"),
    ("field.status", "status"),
    ("ac.error_message", "error_message"),
)

OPTICAL_STATUS_COLORS = STATUS_COLOR_MAP
OPTICAL_STATUS_SEVERITY = {"unknown": 0, "not_collected": 0, "skipped": 1, "normal": 2, "notice": 3, "warning": 4, "alarm": 5, "link_abnormal": 6, "link_down": 6, "no_light": 7}
AC_FEATURE_ORDER = (
    "ac.trackside_ap_plan",
    "ac.ap_online_overview",
    "ac.fit_ap_resources",
    "ac.fit_ap_optical",
    "ac.fit_ap_extensions",
)


def _set_button_icon(button: QPushButton, icon: object | None) -> None:
    if icon is None:
        return
    icon_factory = getattr(icon, "icon", None)
    resolved_icon = icon_factory() if callable(icon_factory) else icon
    try:
        button.setIcon(resolved_icon)
    except TypeError:
        pass
AC_TAB_ACTION_LABELS = {
    "ac.trackside_ap_plan": ("新增行", "删除选中", "保存", "导入", "导出", "下载模板", "更新"),
    "ac.ap_online_overview": ("获取 AP 列表", "刷新", "导出", "清空结果"),
    "ac.fit_ap_resources": ("获取 AP 列表", "获取 AP 地址", "获取 AP 射频", "更新", "批量删除", "导出 AP 信息", "导出 OmniPeek 名称表", "清空选择", "反选"),
    "ac.fit_ap_optical": ("更新光衰", "并发数", "导出表格", "清除筛选"),
    "ac.fit_ap_extensions": (
        "标准模板导入",
        "原始设计/布点表智能识别导入",
        "下载标准模板",
        "导出",
        "新增",
        "编辑",
        "批量删除",
        "清空当前局点扩展信息",
        "刷新",
    ),
}


class OfflineApLedgerLoadThread(QThread):
    load_finished = Signal(object)
    load_failed = Signal(str)

    def __init__(self, device_repository: DeviceRepository, ac_device_uuid: str, cache_path: Path, parent=None) -> None:
        super().__init__(parent)
        self.device_repository = device_repository
        self.ac_device_uuid = ac_device_uuid
        self.cache_path = cache_path

    def run(self) -> None:
        try:
            ac_repository = AcRepository(self.device_repository.database)
            resources = ac_repository.list_ap_entities(self.ac_device_uuid)
            if not resources:
                resources = ac_repository.list_fit_ap_resources_with_metadata(self.ac_device_uuid)
            devices = self.device_repository.list()
            latest_lldp, _latest_optical = build_latest_ap_history_indexes(ac_repository, resources)
            stats, ledger = build_offline_ap_ledger(
                fit_ap_resources=resources,
                latest_lldp_by_ap=latest_lldp,
                device_lookup_by_name=build_device_lookup_by_name(devices),
            )
            save_offline_ap_cache(self.cache_path, ac_device_uuid=self.ac_device_uuid, stats=stats, ledger_rows=ledger)
            self.load_finished.emit({"stats": stats, "ledger_rows": ledger})
        except Exception as exc:
            self.load_failed.emit(str(exc))


def sort_fit_ap_optical_rows(rows: list[dict[str, object | None]]) -> list[dict[str, object | None]]:
    def key(row: dict[str, object | None]) -> tuple[int, str, tuple[object, ...], str]:
        name = str(row.get("neighbor_device_name") or "").strip()
        missing_name = 1 if name in {"", "-"} else 0
        return (missing_name, name.casefold(), interface_sort_key(row.get("neighbor_interface")), str(row.get("ap_name") or ""))

    return sorted(rows, key=key)


def enrich_fit_ap_optical_rows(
    rows: list[dict[str, object | None]],
    resources: list[dict[str, object | None]],
    device_optical_status_lookup: dict[tuple[str, str], dict[str, object | None]] | None = None,
) -> list[dict[str, object | None]]:
    resources_by_uuid = {str(row.get("ap_uuid") or ""): row for row in resources if row.get("ap_uuid")}
    resources_by_name = {str(row.get("ap_name") or ""): row for row in resources if row.get("ap_name")}
    lookup = device_optical_status_lookup or {}
    result: list[dict[str, object | None]] = []
    for row in rows:
        resource = resources_by_uuid.get(str(row.get("ap_uuid") or "")) or resources_by_name.get(str(row.get("ap_name") or ""), {})
        neighbor_name = None if _is_invalid_neighbor_text(row.get("neighbor_device_name")) else row.get("neighbor_device_name")
        switch_status = _lookup_switch_status(neighbor_name, row.get("neighbor_interface"), lookup)
        is_offline = is_fit_ap_offline(resource) or bool(row.get("is_ap_offline"))
        result.append(
            {
                **row,
                "ap_mac": row.get("ap_mac") or resource.get("ap_mac"),
                "site": row.get("site") or resource.get("site_name") or resource.get("site") or "未归属",
                "neighbor_device_name": neighbor_name,
                "switch_optical_status": switch_status,
                "is_ap_offline": is_offline,
                "optical_alarm_status": OFFLINE_AP_STATUS_TEXT if is_offline else row.get("optical_alarm_status"),
                "ap_optical_status": "offline" if is_offline else row.get("ap_optical_status"),
                "data_source": "historical" if is_offline else row.get("data_source"),
            }
        )
    return result


def _lookup_switch_status(
    neighbor_device_name: object,
    neighbor_interface: object,
    lookup: dict[tuple[str, str], dict[str, object | None]],
) -> str:
    """Compute switch-side status real-time from raw optical module data."""
    return compute_switch_status(
        device_name=neighbor_device_name,
        interface_name=neighbor_interface,
        lookup=lookup,
    )


def filter_fit_ap_optical_rows(rows: list[dict[str, object | None]], filters: dict[str, object | None]) -> list[dict[str, object | None]]:
    text_fields = ("ap_name", "site")
    result = rows
    for field in text_fields:
        needle = str(filters.get(field) or "").strip().casefold()
        if needle:
            result = [row for row in result if needle in str(row.get(field) or "").casefold()]
    status = str(filters.get("optical_alarm_status") or "").strip()
    if status:
        result = [row for row in result if evaluate_fit_ap_ap_status(row) == status]
    return result


def build_site_filter_items(rows: list[dict[str, object | None]], all_label: str) -> list[tuple[str, str]]:
    sites = sorted({str(row.get("site") or "").strip() for row in rows if str(row.get("site") or "").strip()})
    return [(all_label, ""), *[(site, site) for site in sites]]


def _normalize_resource_search(value: object) -> str:
    return "".join(char for char in str(value or "").casefold() if char.isalnum())


def _resource_state_token(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    return text.split("=", 1)[0].strip().casefold()


def _resource_state_values(row: dict[str, object | None]) -> set[str]:
    values: set[str] = set()
    for field in ("state", "state_raw", "state_display"):
        value = row.get(field)
        text = str(value or "").strip()
        token = _resource_state_token(value)
        if text:
            values.add(text.casefold())
        if token:
            values.add(token)
    return values


def _resource_state_matches(row: dict[str, object | None], selected: str) -> bool:
    if selected == "__offline__":
        return is_fit_ap_offline(row)
    if selected == "__online__":
        return not is_fit_ap_offline(row)
    return selected.casefold() in _resource_state_values(row)


def _resource_state_label(row: dict[str, object | None]) -> str:
    return str(row.get("state_display") or row.get("state_raw") or row.get("state") or "").strip()


def make_fit_ap_optical_export_filename(site_name: str, now: datetime | None = None) -> str:
    stamp = (now or datetime.now()).strftime("%Y-%m-%d-%H%M")
    safe_site = "".join(char if char not in '<>:"/\\|?*' else "_" for char in site_name or "site")
    return f"{safe_site}_FIT-AP光衰_{stamp}.xlsx"


def _normalize_mac_text(value: object) -> str:
    import re

    text = re.sub(r"[^0-9a-fA-F]", "", str(value or ""))
    return text.casefold() if len(text) == 12 else ""


def _display_value(value: object) -> str:
    return str(value) if value not in (None, "") else "-"


def _display_link_value(row: dict[str, object | None], field: str) -> str:
    value = row.get(field)
    if field == "lldp_source":
        return lldp_source_label(value)
    if field in {"lldp_match_status", "optical_match_status", "link_match_status"}:
        return lldp_display_status(value)
    return _display_value(value)


def _display_mileage_for_row(row: dict[str, object | None], field: str) -> str:
    return format_track_mileage(
        _mileage_value_for_row(row, field),
        direction=str(row.get("direction") or row.get("metadata_direction") or ""),
        line_side=str(row.get("extension_line_side") or row.get("line_side") or ""),
        mileage_type=str(row.get("mileage_type") or row.get("line_type") or ""),
    )


def _display_belong_type(value: object) -> str:
    text = str(value or "").strip().casefold()
    return {
        "station": "站点",
        "section": "区间",
        "yard": "场段/库内",
        "unknown": "未知",
    }.get(text, str(value or "").strip())


def _mileage_value_for_row(row: dict[str, object | None], field: str) -> object:
    if field == "mileage_text":
        return row.get("mileage_m") if row.get("mileage_m") is not None else row.get("mileage_text")
    value = row.get(field)
    if value not in (None, ""):
        return value
    if row.get("extension_mileage_m") is not None:
        return row.get("extension_mileage_m")
    return row.get("extension_mileage_text")


def _mileage_meters_for_row(row: dict[str, object | None], field: str) -> int | float | None:
    return parse_mileage_to_meters(_mileage_value_for_row(row, field))


def build_fit_ap_resource_table_row(row: dict[str, object | None]) -> dict[str, object | None]:
    fields = {field for _key, field in FIT_AP_RESOURCE_COLUMNS if field != "select"}
    keep_fields = fields | {"ap_uuid", "state_raw", "state", "ap_name", "ap_mac", "serial_number"}
    result = {field: row.get(field) for field in keep_fields}
    result["mileage"] = _display_mileage_for_row(row, "mileage")
    result["_mileage_meters"] = _mileage_meters_for_row(row, "mileage")
    return result


def _resource_mileage_search_text(row: dict[str, object | None]) -> str:
    tokens = mileage_search_tokens(
        _mileage_value_for_row(row, "mileage"),
        direction=str(row.get("direction") or row.get("metadata_direction") or ""),
        line_side=str(row.get("extension_line_side") or row.get("line_side") or ""),
    )
    return " ".join(sorted(tokens))


def _friendly_external_terminal_error(message: object) -> str:
    text = str(message or "").strip()
    if "未配置" in text or "No external terminal path" in text:
        return "未配置外部终端路径"
    if "未找到" in text or "not found" in text.casefold():
        return "外部终端路径不存在"
    return text or "启动外部终端失败"


def _ap_unique_key(row: dict[str, object | None]) -> str:
    for field in ("ap_uuid", "serial_number", "ap_mac"):
        value = str(row.get(field) or "").strip()
        if value:
            return f"{field}:{value.casefold()}"
    return f"row:{id(row)}"


def _capacity_total_remark(value: object, default_total: int) -> tuple[int, str]:
    if isinstance(value, dict):
        return int(value.get("ap_total") or value.get("total") or default_total), str(value.get("remark") or "")
    if value is None:
        return default_total, ""
    return int(value), ""


def _is_invalid_neighbor_text(value: object) -> bool:
    text = str(value or "")
    lowered = text.casefold()
    return any(token.casefold() in lowered for token in ("Nearest", "Chassis ID", "Default", "customer bridge", "nontpmr"))


def _with_online_rate(row: dict[str, object | None]) -> dict[str, object | None]:
    total = int(row.get("total") or 0)
    online = int(row.get("online") or 0)
    row["online_rate"] = f"{online / total:.1%}" if total else "0.0%"
    return row


def _overview_row_fill(row: dict[str, object | None]):
    from openpyxl.styles import PatternFill

    if str(row.get("site") or "") == "合计":
        return PatternFill(fill_type="solid", fgColor="DBEAFE")
    total = int(row.get("total") or 0)
    online = int(row.get("online") or 0)
    if total and online == total:
        return PatternFill(fill_type="solid", fgColor="DCFCE7")
    if total and online / total < 0.8:
        return PatternFill(fill_type="solid", fgColor="FEF9C3")
    return None


class AcManagementPage(QWidget):
    def __init__(
        self,
        device_repository: DeviceRepository,
        i18n: I18n,
        site_name: str = "demo",
        feature_gate: FeatureGate | None = None,
        eager_load: bool = True,
    ) -> None:
        super().__init__()
        self.device_repository = device_repository
        self.repository = AcRepository(device_repository.database)
        self.fact_repository = DeviceFactRepository(device_repository.database)
        self.import_export_service = FitApImportExportService(self.repository)
        self.i18n = i18n
        self.site_name = site_name
        self.feature_gate = feature_gate or default_feature_gate()
        self.settings = SettingsStore(PathResolver())
        self.background_manager = BackgroundProcessManager(self, paths=self.settings.paths)
        self.background_manager.progress.connect(self._background_progress)
        self.background_manager.finished.connect(self._background_finished)
        self.background_manager.failed.connect(self._background_failed)
        self.background_manager.cancelled.connect(self._background_failed)
        self._background_jobs: dict[str, dict[str, object]] = {}
        self.ac_devices: list[Device] = []
        self.resource_thread: AcResourceCollectThread | None = None
        self.ac_info_thread: AcInfoCollectThread | None = None
        self.action_thread: AcCommandActionThread | None = None
        self.optical_thread: FitApOpticalCollectThread | None = None
        self.trackside_export_manager = ExportProcessManager(self, paths=PathResolver())
        self.trackside_export_job_id: str | None = None
        self.trackside_export_output_path: Path | None = None
        self.trackside_export_manager.progress.connect(self._handle_trackside_export_progress)
        self.trackside_export_manager.finished.connect(self._finish_trackside_export)
        self.trackside_export_manager.failed.connect(self._fail_trackside_export)
        app_events.ac_summary_changed.connect(self._handle_ac_summary_changed)
        self.detail_windows: list[FitApDetailDialog] = []
        self.optical_rows: list[dict[str, object | None]] = []
        self.trackside_rows: list[dict[str, object | None]] = []
        self.offline_ap_stats: dict[str, object | None] = {}
        self.offline_ap_ledger_rows: list[dict[str, object | None]] = []
        self._offline_ap_context_loaded = False
        self._device_list_loaded = False
        self._loaded_feature_ids: set[str] = set()
        self.tab_by_feature_id: dict[str, QWidget] = {}
        self._ui_ready = False
        self._refreshing_current = False
        self._loaded_once = False
        self._last_loaded_site_name = ""
        self.offline_load_thread: OfflineApLedgerLoadThread | None = None
        self.offline_cache_path = PathResolver().offline_ap_cache_path
        self._overview_uses_trackside_plan = False
        self.resource_rows: list[dict[str, object | None]] = []
        self.selected_ap_keys: set[str] = set()
        self._device_optical_status_lookup: dict[tuple[str, str], dict[str, object | None]] = {}
        self.resource_page = 1
        self.resource_page_size = DEFAULT_PAGE_SIZE
        self.optical_page = 1
        self.optical_page_size = DEFAULT_PAGE_SIZE
        self.trackside_page = 1
        self.trackside_page_size = DEFAULT_PAGE_SIZE
        self._updating_online_summary = False

        self.device_combo = QComboBox(self)
        self.open_web_button = QPushButton(self)
        self.update_ac_info_button = QPushButton(self)
        self.persist_auto_ap_button = QPushButton(self)
        self.enable_ap_remote_login_button = QPushButton(self)
        for moved_button in (
            self.open_web_button,
            self.update_ac_info_button,
            self.persist_auto_ap_button,
            self.enable_ap_remote_login_button,
        ):
            moved_button.hide()
            moved_button.setVisible(False)
        self.refresh_button = QPushButton()
        self.cancel_update_button = QPushButton()
        self.cancel_update_button.setVisible(False)
        self.update_progress = QProgressBar()
        self.update_progress.setRange(0, 0)
        self.update_progress.setTextVisible(False)
        self.update_progress.setFixedHeight(6)
        self.update_progress.setVisible(False)
        self.status_label = make_text_selectable(QLabel())
        self.summary_labels: dict[str, QLabel] = {field: make_text_selectable(QLabel("-")) for _key, field in SUMMARY_FIELDS}
        self.tabs = QTabWidget()
        self.resources_table = QTableWidget()
        self.resources_pagination = PaginationWidget(self.i18n)
        self.batch_delete_button = QPushButton()
        self.import_button = QPushButton()
        self.export_extension_template_button = QPushButton()
        self.export_extension_template_button.setObjectName("exportApExtensionTemplateButton")
        self.export_button = QPushButton()
        self.omnipeek_export_button = QPushButton()
        self.clear_selection_button = QPushButton()
        self.invert_selection_button = QPushButton()
        self.selection_label = make_text_selectable(QLabel())
        self.resource_search_input = QLineEdit()
        self.resource_group_filter = QComboBox()
        self.resource_state_filter = QComboBox()
        self.extension_rows: list[dict[str, object | None]] = []
        self.extension_table = QTableWidget()
        self.extension_summary_table = QTableWidget()
        self.extension_issue_table = QTableWidget()
        self.extension_inner_tabs = QTabWidget()
        self.extension_search_input = QLineEdit()
        self.extension_import_standard_button = QPushButton()
        self.extension_import_smart_button = QPushButton()
        self.extension_export_button = QPushButton()
        self.extension_template_button = QPushButton()
        self.extension_add_button = QPushButton()
        self.extension_edit_button = QPushButton()
        self.extension_delete_button = QPushButton()
        self.extension_clear_button = QPushButton()
        self.extension_refresh_button = QPushButton()
        self.optical_table = QTableWidget()
        self.optical_pagination = PaginationWidget(self.i18n)
        self.refresh_optical_button = QPushButton()
        self.optical_concurrency_combo = QComboBox()
        self.optical_export_button = QPushButton()
        self.clear_optical_filters_button = QPushButton()
        self.optical_ap_filter = QLineEdit()
        self.optical_site_filter = QComboBox()
        self.overview_table = QTableWidget()
        self.offline_stats_table = QTableWidget()
        self.offline_ledger_table = QTableWidget()
        self.overview_inner_tabs = QTabWidget()
        self.offline_loading_spinner = QProgressBar()
        self.offline_loading_spinner.setRange(0, 0)
        self.offline_loading_spinner.setTextVisible(False)
        self.offline_loading_spinner.setFixedHeight(6)
        self.offline_loading_spinner.setVisible(False)
        self.offline_loading_label = make_text_selectable(QLabel("正在加载离线AP数据..."))
        self.offline_loading_label.setVisible(False)
        self.export_overview_button = QPushButton()
        self.save_overview_history_button = QPushButton()
        self.view_overview_history_button = QPushButton()
        self.trackside_table = QTableWidget()
        self.trackside_pagination = PaginationWidget(self.i18n)
        self.trackside_site_filter = QComboBox()
        self.trackside_search_input = QLineEdit()
        self.trackside_export_button = QPushButton()
        self.optical_legend_label = make_text_selectable(QLabel())
        self.trackside_plan_page = TracksideApPlanPage(self.repository, self.i18n, self.site_name)

        configure_readonly_table(self.resources_table)
        install_checkbox_only_delegate(self.resources_table, CHECK_COLUMN)
        configure_readonly_table(self.extension_table)
        configure_readonly_table(self.extension_summary_table)
        configure_readonly_table(self.extension_issue_table)
        configure_readonly_table(self.optical_table)
        configure_readonly_table(self.overview_table)
        configure_readonly_table(self.offline_stats_table)
        configure_readonly_table(self.offline_ledger_table)
        configure_readonly_table(self.trackside_table)
        self.overview_table.setEditTriggers(
            QAbstractItemView.DoubleClicked
            | QAbstractItemView.EditKeyPressed
            | QAbstractItemView.SelectedClicked
        )
        self.resources_table.setColumnCount(len(FIT_AP_RESOURCE_COLUMNS))
        self.extension_table.setColumnCount(len(AP_EXTENSION_COLUMNS))
        self.extension_summary_table.setColumnCount(len(AP_EXTENSION_SUMMARY_COLUMNS))
        self.extension_issue_table.setColumnCount(len(AP_EXTENSION_ISSUE_COLUMNS))
        self.optical_table.setColumnCount(len(FIT_AP_OPTICAL_COLUMNS))
        self.overview_table.setColumnCount(len(AP_ONLINE_OVERVIEW_COLUMNS))
        self.offline_stats_table.setColumnCount(len(OFFLINE_AP_STATS_COLUMNS))
        self.offline_ledger_table.setColumnCount(len(OFFLINE_AP_LEDGER_COLUMNS))
        self.trackside_table.setColumnCount(len(TRACKSIDE_AP_BUSINESS_COLUMNS))
        set_table_column_fields(self.resources_table, [field for _key, field in FIT_AP_RESOURCE_COLUMNS])
        set_table_column_fields(self.extension_table, [field for _key, field in AP_EXTENSION_COLUMNS])
        set_table_column_fields(self.extension_summary_table, [field for _key, field in AP_EXTENSION_SUMMARY_COLUMNS])
        set_table_column_fields(self.extension_issue_table, [field for _key, field in AP_EXTENSION_ISSUE_COLUMNS])
        set_table_column_fields(self.optical_table, [field for _key, field in FIT_AP_OPTICAL_COLUMNS])
        set_table_column_fields(self.overview_table, [field for _key, field in AP_ONLINE_OVERVIEW_COLUMNS])
        set_table_column_fields(self.offline_stats_table, [field for _key, field in OFFLINE_AP_STATS_COLUMNS])
        set_table_column_fields(self.offline_ledger_table, [field for _key, field in OFFLINE_AP_LEDGER_COLUMNS])
        set_table_column_fields(self.trackside_table, [field for _key, field in TRACKSIDE_AP_BUSINESS_COLUMNS])
        self.overview_table.itemChanged.connect(self.save_overview_total)
        self.overview_inner_tabs.currentChanged.connect(self.handle_overview_inner_tab_changed)
        self.resources_table.horizontalHeader().sectionClicked.connect(self._resource_header_clicked)
        self.resources_table.itemChanged.connect(self.update_selection_state)
        self.resources_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.resources_table.customContextMenuRequested.connect(self.show_resource_context_menu)
        self.extension_table.doubleClicked.connect(lambda index: self.edit_ap_extension_point(index.row()))
        self.optical_table.doubleClicked.connect(lambda index: self.open_ap_detail_from_optical(index.row()))
        self.offline_ledger_table.doubleClicked.connect(lambda index: self.open_ap_detail_from_offline_ledger(index.row()))
        self.optical_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.optical_table.customContextMenuRequested.connect(self.show_optical_context_menu)
        self.trackside_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.trackside_table.customContextMenuRequested.connect(self.show_trackside_context_menu)

        top = QHBoxLayout()
        top.setContentsMargins(0, 0, 0, 0)
        top.setSpacing(8)
        self.device_combo.setMinimumWidth(360)
        top.addWidget(self.device_combo, 1)
        top.addWidget(self.status_label)
        top.addWidget(self.cancel_update_button)

        summary = QGridLayout()
        for index, (key, field) in enumerate(SUMMARY_FIELDS):
            label = QLabel()
            label.setObjectName(f"summary_label_{field}")
            label.setProperty("translation_key", key)
            make_text_selectable(label)
            summary.addWidget(label, index // 4 * 2, index % 4)
            summary.addWidget(self.summary_labels[field], index // 4 * 2 + 1, index % 4)

        resources_tab = QWidget()
        resources_layout = QVBoxLayout()
        resource_actions = QHBoxLayout()
        for button in (
            self.refresh_button,
            self.batch_delete_button,
            self.export_button,
            self.omnipeek_export_button,
            self.clear_selection_button,
            self.invert_selection_button,
        ):
            resource_actions.addWidget(button)
        resource_actions.addWidget(self.selection_label)
        resource_actions.addStretch(1)
        resource_filters = QHBoxLayout()
        resource_filters.addWidget(self.resource_search_input, 2)
        resource_filters.addWidget(self.resource_group_filter)
        resource_filters.addWidget(self.resource_state_filter)
        resources_layout.addLayout(resource_actions)
        resources_layout.addLayout(resource_filters)
        resources_layout.addWidget(self.resources_table, 1)
        resources_layout.addWidget(self.resources_pagination)
        resources_tab.setLayout(resources_layout)

        extension_tab = QWidget()
        extension_layout = QVBoxLayout()
        extension_actions = QHBoxLayout()
        for button in (
            self.extension_import_standard_button,
            self.extension_import_smart_button,
            self.extension_template_button,
            self.extension_export_button,
            self.extension_add_button,
            self.extension_edit_button,
            self.extension_delete_button,
            self.extension_clear_button,
            self.extension_refresh_button,
        ):
            extension_actions.addWidget(button)
        extension_actions.addStretch(1)
        extension_layout.addLayout(extension_actions)
        extension_layout.addWidget(self.extension_search_input)
        self.extension_inner_tabs.addTab(self.extension_table, "AP点位扩展表")
        self.extension_inner_tabs.addTab(self.extension_summary_table, "车站汇总")
        self.extension_inner_tabs.addTab(self.extension_issue_table, "异常检查")
        extension_layout.addWidget(self.extension_inner_tabs, 1)
        extension_tab.setLayout(extension_layout)

        optical_tab = QWidget()
        optical_layout = QVBoxLayout()
        optical_actions = QHBoxLayout()
        optical_actions.addWidget(self.refresh_optical_button)
        optical_actions.addWidget(self.optical_concurrency_combo)
        optical_actions.addWidget(self.optical_export_button)
        optical_actions.addWidget(self.clear_optical_filters_button)
        optical_actions.addStretch(1)
        optical_filters = QHBoxLayout()
        for widget in (
            self.optical_ap_filter,
            self.optical_site_filter,
        ):
            optical_filters.addWidget(widget)
        self.optical_legend_label.setWordWrap(True)
        optical_layout.addLayout(optical_actions)
        optical_layout.addLayout(optical_filters)
        optical_layout.addWidget(self.optical_legend_label)
        optical_layout.addWidget(self.optical_table, 1)
        optical_layout.addWidget(self.optical_pagination)
        optical_tab.setLayout(optical_layout)

        overview_tab = QWidget()
        overview_layout = QVBoxLayout()
        self.overview_inner_tabs.addTab(self.overview_table, "")
        self.overview_inner_tabs.addTab(self.offline_stats_table, "")
        self.overview_inner_tabs.addTab(self.offline_ledger_table, "")
        overview_layout.addWidget(self.offline_loading_spinner)
        overview_layout.addWidget(self.offline_loading_label)
        overview_layout.addWidget(self.overview_inner_tabs, 1)
        overview_tab.setLayout(overview_layout)

        self.tab_by_feature_id = {
            "ac.trackside_ap_plan": self.trackside_plan_page,
            "ac.ap_online_overview": overview_tab,
            "ac.fit_ap_resources": resources_tab,
            "ac.fit_ap_extensions": extension_tab,
            "ac.fit_ap_optical": optical_tab,
        }
        with QSignalBlocker(self.tabs):
            self.tabs.addTab(self.trackside_plan_page, "")
            self.tabs.addTab(overview_tab, "")
            self.tabs.addTab(resources_tab, "")
            self.tabs.addTab(optical_tab, "")
            self.tabs.addTab(extension_tab, "")
            self._apply_feature_gate()
        # Trackside AP Service is mounted under Rail Transit.

        layout = QVBoxLayout()
        layout.addLayout(top)
        layout.addWidget(self.update_progress)
        layout.addLayout(summary)
        layout.addWidget(self.tabs, 1)
        self.setLayout(layout)

        self.device_combo.currentIndexChanged.connect(self.refresh_data)
        self.open_web_button.clicked.connect(self.open_web)
        self.update_ac_info_button.clicked.connect(self.refresh_ac_info)
        self.persist_auto_ap_button.clicked.connect(lambda: self.run_ac_action("persist_auto_ap", "一键固化新上线AP"))
        self.enable_ap_remote_login_button.clicked.connect(lambda: self.run_ac_action("enable_ap_remote_login", "一键开启AP远程登入"))
        self.refresh_button.clicked.connect(self.refresh_ac_resources)
        self.cancel_update_button.clicked.connect(self.cancel_current_update)
        self.batch_delete_button.clicked.connect(self.batch_delete_aps)
        self.export_extension_template_button.clicked.connect(self.export_ap_extension_template)
        self.export_button.clicked.connect(self.export_aps)
        self.omnipeek_export_button.clicked.connect(self.export_omnipeek_name_table)
        self.extension_import_standard_button.clicked.connect(lambda: self.import_ap_extensions("standard_template"))
        self.extension_import_smart_button.clicked.connect(lambda: self.import_ap_extensions("smart_design"))
        self.extension_template_button.clicked.connect(self.export_ap_extension_template)
        self.extension_export_button.clicked.connect(self.export_ap_extensions)
        self.extension_add_button.clicked.connect(self.add_ap_extension_point)
        self.extension_edit_button.clicked.connect(lambda: self.edit_ap_extension_point(self.extension_table.currentRow()))
        self.extension_delete_button.clicked.connect(self.delete_selected_ap_extension_points)
        self.extension_clear_button.clicked.connect(self.clear_ap_extension_points)
        self.extension_refresh_button.clicked.connect(self.refresh_ap_extensions)
        self.clear_selection_button.clicked.connect(self.clear_selection)
        self.invert_selection_button.clicked.connect(self.invert_selection)
        self.refresh_optical_button.clicked.connect(self.refresh_fit_ap_optical)
        self.optical_export_button.clicked.connect(self.export_optical_table)
        self.export_overview_button.clicked.connect(self.export_overview_table)
        self.save_overview_history_button.clicked.connect(self.save_overview_history_snapshot)
        self.view_overview_history_button.clicked.connect(self.open_overview_history)
        self.trackside_export_button.clicked.connect(self.export_trackside_table)
        self.clear_optical_filters_button.clicked.connect(self.clear_optical_filters)
        self.optical_ap_filter.textChanged.connect(self.apply_optical_filters)
        self.optical_site_filter.currentIndexChanged.connect(self.apply_optical_filters)
        self.trackside_site_filter.currentIndexChanged.connect(self.apply_trackside_filters)
        self.trackside_search_input.textChanged.connect(self.apply_trackside_filters)
        self.resource_search_input.textChanged.connect(self.apply_resource_filters)
        self.extension_search_input.textChanged.connect(self.refresh_ap_extensions)
        self.resource_group_filter.currentIndexChanged.connect(self.apply_resource_filters)
        self.resource_state_filter.currentIndexChanged.connect(self.apply_resource_filters)
        self.resources_pagination.pageChanged.connect(self.set_resource_page)
        self.resources_pagination.pageSizeChanged.connect(self.set_resource_page_size)
        self.optical_pagination.pageChanged.connect(self.set_optical_page)
        self.optical_pagination.pageSizeChanged.connect(self.set_optical_page_size)
        self.trackside_pagination.pageChanged.connect(self.set_trackside_page)
        self.trackside_pagination.pageSizeChanged.connect(self.set_trackside_page_size)
        self.trackside_plan_page.plan_saved.connect(self._handle_trackside_plan_saved)
        self.retranslate()
        self._parent_hidden_legacy_buttons()
        self.tabs.currentChanged.connect(self._on_current_tab_changed)
        self._ui_ready = True
        if eager_load:
            self.refresh_devices()

    def _apply_feature_gate(self) -> None:
        self._reconcile_feature_tabs()
        apply_feature_to_widget(self.feature_gate, "ac.fit_ap_resources", self.refresh_button)
        apply_feature_to_widget(self.feature_gate, "ac.fit_ap_resources", self.export_button)
        apply_feature_to_widget(self.feature_gate, "ac.omnipeek_name_table_export", self.omnipeek_export_button)
        apply_feature_to_widget(self.feature_gate, "ac.ac_info_update", self.update_ac_info_button)
        apply_feature_to_widget(self.feature_gate, "ac.ac_actions", self.persist_auto_ap_button)
        apply_feature_to_widget(self.feature_gate, "ac.ac_actions", self.enable_ap_remote_login_button)
        for widget in (
            self.extension_import_standard_button,
            self.extension_import_smart_button,
            self.extension_export_button,
            self.extension_template_button,
            self.extension_add_button,
            self.extension_edit_button,
            self.extension_delete_button,
            self.extension_clear_button,
            self.extension_refresh_button,
        ):
            apply_feature_to_widget(self.feature_gate, "ac.fit_ap_extensions", widget)
        apply_feature_to_widget(self.feature_gate, "ac.fit_ap_optical", self.refresh_optical_button)
        apply_feature_to_widget(self.feature_gate, "ac.fit_ap_optical", self.optical_export_button)

    def _reconcile_feature_tabs(self) -> None:
        current_feature = self._current_feature_id()
        blocked = self.tabs.blockSignals(True)
        try:
            while self.tabs.count():
                self.tabs.removeTab(0)
            for feature_id in AC_FEATURE_ORDER:
                if not self.feature_gate.is_visible(feature_id):
                    continue
                widget = self.tab_by_feature_id[feature_id]
                self.tabs.addTab(widget, self.i18n.t(feature_id))
                self.tabs.setTabEnabled(self.tabs.count() - 1, self.feature_gate.is_enabled(feature_id))
            target = self._tab_index_for_feature(current_feature) if current_feature else -1
            self.tabs.setCurrentIndex(target if target >= 0 else (0 if self.tabs.count() else -1))
        finally:
            self.tabs.blockSignals(blocked)
        self._set_ac_tab_titles()
        app_logger.log_info(
            "AC_FEATURE_TABS_RECONCILED",
            (
                f"session_override_active={self.feature_gate.is_session_override_active()} profile={self.feature_gate.profile} "
                f"visible_features={','.join(self._visible_feature_ids())} current_feature={self._current_feature_id() or ''} "
                f"tab_count={self.tabs.count()}"
            ),
        )

    def _set_ac_tab_titles(self) -> None:
        for feature_id, widget in self.tab_by_feature_id.items():
            index = self.tabs.indexOf(widget)
            if index >= 0:
                self.tabs.setTabText(index, self.i18n.t(feature_id))

    def _current_feature_id(self) -> str | None:
        current = self.tabs.currentWidget()
        for feature_id, widget in self.tab_by_feature_id.items():
            if widget is current:
                return feature_id
        return None

    def current_tab_action_labels(self) -> list[str]:
        feature_id = self._current_feature_id()
        if feature_id is None:
            return []
        return list(AC_TAB_ACTION_LABELS.get(feature_id, ()))

    def _tab_index_for_feature(self, feature_id: str | None) -> int:
        if not feature_id:
            return -1
        widget = self.tab_by_feature_id.get(feature_id)
        return self.tabs.indexOf(widget) if widget is not None else -1

    def _visible_feature_ids(self) -> list[str]:
        visible: list[str] = []
        for index in range(self.tabs.count()):
            widget = self.tabs.widget(index)
            feature_id = next((key for key, value in self.tab_by_feature_id.items() if value is widget), "")
            if feature_id:
                visible.append(feature_id)
        return visible

    def set_repository(self, device_repository: DeviceRepository, site_name: str) -> None:
        self.device_repository = device_repository
        self.repository = AcRepository(device_repository.database)
        self.fact_repository = DeviceFactRepository(device_repository.database)
        self.import_export_service = FitApImportExportService(self.repository)
        self.site_name = site_name
        self._device_list_loaded = False
        self._loaded_feature_ids.clear()
        self._loaded_once = False
        self._last_loaded_site_name = ""
        self.trackside_plan_page.repository = self.repository
        self.trackside_plan_page.site_name = site_name
        self.trackside_plan_page.refresh()
        self.refresh_ap_extensions()
        self.refresh_devices()

    def on_enter(self) -> None:
        if not self._loaded_once or self._last_loaded_site_name != self.site_name:
            self.load_initial_data()
            return
        self.refresh_current_async_or_lazy()

    def load_initial_data(self) -> None:
        if not self._ui_ready:
            return
        self._loaded_once = True
        self._last_loaded_site_name = self.site_name
        self.refresh_devices(load_current_only=False)
        self.trackside_plan_page.refresh()
        if not self.ac_devices:
            self.status_label.setText("当前局点未配置 AC 设备")
            app_logger.log_info("AC_INITIAL_LOAD_EMPTY", f"site={self.site_name}")
        else:
            app_logger.log_info("AC_INITIAL_LOAD_DONE", f"site={self.site_name} ac_count={len(self.ac_devices)}")

    def retranslate(self) -> None:
        self.open_web_button.setText(self.i18n.t("ac.open_web"))
        self.update_ac_info_button.setText("更新AC信息")
        self.persist_auto_ap_button.setText("一键固化新上线AP")
        self.enable_ap_remote_login_button.setText("一键开启AP远程登入")
        self.refresh_button.setText(self.i18n.t("details.refresh"))
        self.cancel_update_button.setText(self.i18n.t("ac.cancel_update"))
        self.batch_delete_button.setText(self.i18n.t("devices.batch_delete"))
        self.import_button.setText(self.i18n.t("ap.import_metadata"))
        self.export_extension_template_button.setText(self.i18n.t("ap.export_extension_template"))
        self.export_button.setText(self.i18n.t("ap.export_info"))
        self.omnipeek_export_button.setText("导出 OmniPeek 名称表")
        self.extension_import_standard_button.setText("标准模板导入")
        self.extension_import_smart_button.setText("原始设计/布点表智能识别导入")
        self.extension_template_button.setText("下载标准模板")
        self.extension_export_button.setText("导出")
        self.extension_add_button.setText("新增")
        self.extension_edit_button.setText("编辑")
        self.extension_delete_button.setText("批量删除")
        self.extension_clear_button.setText("清空当前局点扩展信息")
        self.extension_refresh_button.setText("刷新")
        self.extension_search_input.setPlaceholderText("搜索 AP MAC、AP名称、AP编号、归属站点、归属区间、区间起止、方向、里程、点位说明、场段、区域、备注")
        self.clear_selection_button.setText(self.i18n.t("devices.clear_selection"))
        self.invert_selection_button.setText(self.i18n.t("devices.invert_selection"))
        self.refresh_optical_button.setText(self.i18n.t("ac.refresh_optical"))
        self.optical_export_button.setText(self.i18n.t("ac.export_table"))
        self.export_overview_button.setText(self.i18n.t("ac.export_overview"))
        self.save_overview_history_button.setText(self.i18n.t("ac.save_history_snapshot"))
        self.view_overview_history_button.setText(self.i18n.t("ac.view_history"))
        self.trackside_export_button.setText(self.i18n.t("trackside.export"))
        self.trackside_search_input.setPlaceholderText(self.i18n.t("trackside.search"))
        self.resource_search_input.setPlaceholderText(self.i18n.t("ap.resource_search_placeholder"))
        self._set_resource_filter_items(self.resource_rows)
        self.resources_pagination.retranslate()
        self.optical_pagination.retranslate()
        self.trackside_pagination.retranslate()
        self.clear_optical_filters_button.setText(self.i18n.t("ac.clear_filters"))
        self.optical_ap_filter.setPlaceholderText(self.i18n.t("ac.ap_name"))
        self._apply_button_icons()
        self._refresh_translated_runtime_state()
        self._hide_global_proxy_buttons()

    def _apply_button_icons(self) -> None:
        button_icons = (
            (self.refresh_button, FIF.SYNC),
            (self.cancel_update_button, FIF.CANCEL),
            (self.batch_delete_button, FIF.DELETE),
            (self.export_button, FIF.SHARE),
            (self.omnipeek_export_button, FIF.SHARE),
            (self.clear_selection_button, FIF.CLEAR_SELECTION),
            (self.invert_selection_button, FIF.CHECKBOX),
            (self.extension_import_standard_button, FIF.DOWNLOAD),
            (self.extension_import_smart_button, FIF.DOWNLOAD),
            (self.extension_template_button, FIF.CLOUD_DOWNLOAD),
            (self.extension_export_button, FIF.SHARE),
            (self.extension_add_button, FIF.ADD),
            (self.extension_edit_button, FIF.EDIT),
            (self.extension_delete_button, FIF.DELETE),
            (self.extension_clear_button, FIF.BROOM),
            (self.extension_refresh_button, FIF.SYNC),
            (self.refresh_optical_button, FIF.UPDATE),
            (self.optical_export_button, FIF.SHARE),
            (self.clear_optical_filters_button, FIF.FILTER),
            (self.import_button, FIF.DOWNLOAD),
            (self.export_extension_template_button, FIF.CLOUD_DOWNLOAD),
            (self.export_overview_button, FIF.SHARE),
            (self.save_overview_history_button, FIF.SAVE),
            (self.view_overview_history_button, FIF.HISTORY),
            (self.trackside_export_button, FIF.SHARE),
        )
        for button, icon in button_icons:
            _set_button_icon(button, icon)

    def _hide_global_proxy_buttons(self) -> None:
        for button in (
            self.open_web_button,
            self.update_ac_info_button,
            self.persist_auto_ap_button,
            self.enable_ap_remote_login_button,
        ):
            button.setParent(self)
            button.hide()
            button.setVisible(False)

    def _parent_hidden_legacy_buttons(self) -> None:
        for button in (
            self.import_button,
            self.export_extension_template_button,
            self.export_overview_button,
            self.save_overview_history_button,
            self.view_overview_history_button,
            self.trackside_export_button,
        ):
            button.setParent(self)
            button.hide()
            button.setVisible(False)

    def _refresh_translated_runtime_state(self) -> None:
        current_concurrency = self.optical_concurrency_combo.currentData()
        self.optical_concurrency_combo.clear()
        for value in (50, 100, 200, 500, 1000):
            self.optical_concurrency_combo.addItem(f"{self.i18n.t('batch_collect.concurrency')}: {value}", value)
        target_concurrency = current_concurrency if current_concurrency is not None else 1000
        index = self.optical_concurrency_combo.findData(target_concurrency)
        self.optical_concurrency_combo.setCurrentIndex(index if index >= 0 else self.optical_concurrency_combo.findData(1000))
        self.optical_legend_label.setText(self.i18n.t("details.optical_color_legend"))
        self.status_label.setText(self.i18n.t("ac.status.not_collected"))
        for index, (key, _field) in enumerate(SUMMARY_FIELDS):
            label = self.findChild(QLabel, f"summary_label_{SUMMARY_FIELDS[index][1]}")
            if label is not None:
                label.setText(self.i18n.t(key))
        self._set_ac_tab_titles()
        self.overview_inner_tabs.setTabText(0, self.i18n.t("ac.ap_online_overview"))
        self.overview_inner_tabs.setTabText(1, "AP离线情况")
        self.overview_inner_tabs.setTabText(2, "离线AP台账")
        self.trackside_plan_page.retranslate()
        self.resources_table.setHorizontalHeaderLabels([self.i18n.t(key) for key, _field in FIT_AP_RESOURCE_COLUMNS])
        self.resources_table.horizontalHeaderItem(CHECK_COLUMN).setText(self.i18n.t("ap.select_all"))
        self.extension_table.setHorizontalHeaderLabels([key for key, _field in AP_EXTENSION_COLUMNS])
        self.extension_summary_table.setHorizontalHeaderLabels([key for key, _field in AP_EXTENSION_SUMMARY_COLUMNS])
        self.extension_issue_table.setHorizontalHeaderLabels([key for key, _field in AP_EXTENSION_ISSUE_COLUMNS])
        self.optical_table.setHorizontalHeaderLabels([self.i18n.t(key) for key, _field in FIT_AP_OPTICAL_COLUMNS])
        self.overview_table.setHorizontalHeaderLabels([self.i18n.t(key) for key, _field in AP_ONLINE_OVERVIEW_COLUMNS])
        self.offline_stats_table.setHorizontalHeaderLabels(offline_ap_headers(OFFLINE_AP_STATS_COLUMNS))
        self.offline_ledger_table.setHorizontalHeaderLabels(offline_ap_headers(OFFLINE_AP_LEDGER_COLUMNS))
        self.trackside_table.setHorizontalHeaderLabels([self.i18n.t(key) for key, _field in TRACKSIDE_AP_BUSINESS_COLUMNS])
        self._apply_trackside_header_tooltips()
        apply_table_style(self.resources_table)
        apply_table_style(self.extension_table)
        apply_table_style(self.extension_summary_table)
        apply_table_style(self.extension_issue_table)
        apply_table_style(self.optical_table)
        apply_table_style(self.overview_table)
        apply_table_style(self.offline_stats_table)
        apply_table_style(self.offline_ledger_table)
        apply_table_style(self.trackside_table)
        self.update_selection_state()
        self.update_open_web_button()

    def _apply_trackside_header_tooltips(self) -> None:
        for index, (label_key, field) in enumerate(TRACKSIDE_AP_BUSINESS_COLUMNS):
            item = self.trackside_table.horizontalHeaderItem(index)
            if item is None:
                continue
            tooltip_key = TRACKSIDE_AP_BUSINESS_HEADER_TOOLTIPS.get(field)
            label = self.i18n.t(label_key)
            item.setToolTip(self.i18n.t(tooltip_key) if tooltip_key else self.i18n.t("trackside.tooltip.default", label=label))

    def refresh_devices(self, *, load_current_only: bool = False) -> None:
        current_uuid = self.current_device_uuid()
        self.ac_devices = self.device_repository.list(vendor="H3C", device_type="AC")
        self._device_list_loaded = True
        self.device_combo.blockSignals(True)
        self.device_combo.clear()
        for device in self.ac_devices:
            self.device_combo.addItem(f"{device.name} ({device.ip_address})", device.device_uuid)
        index = self.device_combo.findData(current_uuid)
        self.device_combo.setCurrentIndex(index if index >= 0 else (0 if self.ac_devices else -1))
        self.device_combo.blockSignals(False)
        if load_current_only:
            self.refresh_current_tab_data()
        else:
            self.refresh_data()

    def clear_results(self) -> None:
        self._set_summary(None)
        self.resource_rows = []
        self.selected_ap_keys.clear()
        self._set_resource_filter_items([])
        self.apply_resource_pagination()
        self.optical_rows = []
        self._set_site_filter_items([])
        self.apply_optical_filters()
        self.offline_ap_stats = {}
        self.offline_ap_ledger_rows = []
        self._offline_ap_context_loaded = False
        self._set_rows(self.overview_table, AP_ONLINE_OVERVIEW_COLUMNS, [])
        self._set_rows(self.offline_stats_table, OFFLINE_AP_STATS_COLUMNS, [])
        self._set_rows(self.offline_ledger_table, OFFLINE_AP_LEDGER_COLUMNS, [])
        self.status_label.setText("结果已清空")

    def refresh_current_async_or_lazy(self, force_if_empty: bool = False) -> None:
        if not self._ui_ready or self._refreshing_current:
            return
        feature_id = self._current_feature_id()
        if feature_id is None:
            return
        if not self.feature_gate.is_enabled(feature_id):
            return
        if not force_if_empty and feature_id in self._loaded_feature_ids:
            return
        self._refreshing_current = True
        try:
            if not self._device_list_loaded:
                self.refresh_devices(load_current_only=True)
            else:
                self.refresh_current_tab_data(feature_id)
        finally:
            self._refreshing_current = False

    def _on_current_tab_changed(self, index: int) -> None:
        if index < 0:
            return
        feature_id = self._current_feature_id() or ""
        tab_text = self.tabs.tabText(index)
        actions = ", ".join(self.current_tab_action_labels())
        app_logger.log_info("AC_TAB_ACTIONS", f"tab={tab_text} feature={feature_id} actions={actions}")
        self.refresh_current_async_or_lazy()

    def refresh_current_tab_data(self, feature_id: str | None = None) -> None:
        if not self._ui_ready:
            return
        if not self._device_list_loaded:
            self.refresh_devices(load_current_only=True)
            return
        feature_id = feature_id or self._current_feature_id()
        if feature_id is None:
            return
        if not self.feature_gate.is_enabled(feature_id):
            return
        ac_uuid = self.current_device_uuid()
        if not ac_uuid:
            self.refresh_data()
            return
        self._set_summary(self.repository.get_ac_ap_summary(ac_uuid))
        if feature_id == "ac.trackside_ap_plan":
            self.trackside_plan_page.refresh()
        elif feature_id == "ac.ap_online_overview":
            resources = self._load_resource_rows(ac_uuid)
            self.refresh_overview_table(resources)
        elif feature_id == "ac.fit_ap_resources":
            self._load_resource_rows(ac_uuid)
        elif feature_id == "ac.fit_ap_optical":
            resources = self._load_resource_rows(ac_uuid)
            self._load_optical_rows(ac_uuid, resources)
        elif feature_id == "ac.fit_ap_extensions":
            self.refresh_ap_extensions()
        if feature_id:
            self._loaded_feature_ids.add(feature_id)

    def _load_resource_rows(self, ac_uuid: str) -> list[dict[str, object | None]]:
        resources = self.repository.list_fit_ap_resources_with_metadata(ac_uuid)
        self.resource_rows = resources
        valid_keys = {self._ap_selection_key(row) for row in resources}
        self.selected_ap_keys.intersection_update({key for key in valid_keys if key})
        self._set_resource_filter_items(resources)
        self.apply_resource_pagination()
        self.update_selection_state()
        self.update_open_web_button()
        return resources

    def _load_optical_rows(self, ac_uuid: str, resources: list[dict[str, object | None]]) -> None:
        self._rebuild_device_optical_status_lookup()
        current_optical_rows = self.repository.list_fit_ap_optical(ac_uuid)
        self.optical_rows = sort_fit_ap_optical_rows(enrich_fit_ap_optical_rows(current_optical_rows, resources, self._device_optical_status_lookup))
        self._set_site_filter_items(self.optical_rows)
        self.apply_optical_filters()

    def refresh_data(self) -> None:
        ac_uuid = self.current_device_uuid()
        if not ac_uuid:
            self._set_summary(None)
            self.resource_rows = []
            self.selected_ap_keys.clear()
            self._set_resource_filter_items([])
            self.apply_resource_pagination()
            self.optical_rows = []
            self.offline_ap_stats = {}
            self.offline_ap_ledger_rows = []
            self._offline_ap_context_loaded = False
            self._set_site_filter_items([])
            self.apply_optical_filters()
            self._set_rows(self.overview_table, AP_ONLINE_OVERVIEW_COLUMNS, [])
            self._set_rows(self.offline_stats_table, OFFLINE_AP_STATS_COLUMNS, [])
            self._set_rows(self.offline_ledger_table, OFFLINE_AP_LEDGER_COLUMNS, [])
            self.refresh_ap_extensions()
            self.update_open_web_button()
            return
        self._set_summary(self.repository.get_ac_ap_summary(ac_uuid))
        resources = self.repository.list_fit_ap_resources_with_metadata(ac_uuid)
        self.resource_rows = resources
        valid_keys = {self._ap_selection_key(row) for row in resources}
        self.selected_ap_keys.intersection_update({key for key in valid_keys if key})
        self._set_resource_filter_items(resources)
        self.apply_resource_pagination()
        self._rebuild_device_optical_status_lookup()
        current_optical_rows = self.repository.list_fit_ap_optical(ac_uuid)
        self.offline_ap_stats = {}
        self.offline_ap_ledger_rows = []
        self._offline_ap_context_loaded = False
        self._set_rows(self.offline_stats_table, OFFLINE_AP_STATS_COLUMNS, [])
        self._set_rows(self.offline_ledger_table, OFFLINE_AP_LEDGER_COLUMNS, [])
        self.optical_rows = sort_fit_ap_optical_rows(enrich_fit_ap_optical_rows(current_optical_rows, resources, self._device_optical_status_lookup))
        self._set_site_filter_items(self.optical_rows)
        self.apply_optical_filters()
        self.refresh_overview_table(resources)
        self.refresh_ap_extensions()
        self.update_selection_state()
        self.update_open_web_button()

    def handle_overview_inner_tab_changed(self, index: int) -> None:
        if index in {1, 2}:
            self.ensure_offline_ap_context_loaded()

    def ensure_offline_ap_context_loaded(self, force: bool = False) -> None:
        ac_uuid = self.current_device_uuid()
        if not ac_uuid:
            return
        if not force and self._offline_ap_context_loaded:
            return
        self._load_offline_ap_cache(ac_uuid)
        self._start_offline_ap_async_load(ac_uuid, force=force)

    def _load_offline_ap_cache(self, ac_uuid: str) -> None:
        cached = load_offline_ap_cache(self.offline_cache_path, ac_uuid)
        if not cached:
            return
        stats = cached.get("stats")
        ledger = cached.get("ledger_rows")
        if not isinstance(stats, dict) or not isinstance(ledger, list):
            return
        self.offline_ap_stats = stats
        self.offline_ap_ledger_rows = ledger
        self._offline_ap_context_loaded = True
        self._set_rows(self.offline_stats_table, OFFLINE_AP_STATS_COLUMNS, [stats])
        self._set_rows(self.offline_ledger_table, OFFLINE_AP_LEDGER_COLUMNS, ledger)

    def _start_offline_ap_async_load(self, ac_uuid: str, force: bool = False) -> None:
        if self.offline_load_thread is not None and self.offline_load_thread.isRunning():
            self.offline_loading_spinner.setVisible(True)
            self.offline_loading_label.setVisible(True)
            return
        if self._offline_ap_context_loaded and not force:
            return
        self.offline_loading_label.setText("正在加载离线AP数据...")
        self.offline_loading_spinner.setVisible(True)
        self.offline_loading_label.setVisible(True)
        thread = OfflineApLedgerLoadThread(self.device_repository, ac_uuid, self.offline_cache_path, self)
        self.offline_load_thread = thread
        thread.load_finished.connect(self._finish_offline_ap_load)
        thread.load_failed.connect(self._fail_offline_ap_load)
        thread.finished.connect(thread.deleteLater)
        thread.finished.connect(lambda: self._clear_offline_load_thread(thread))
        thread.start()

    def _finish_offline_ap_load(self, result: object) -> None:
        payload = result if isinstance(result, dict) else {}
        stats = payload.get("stats") if isinstance(payload.get("stats"), dict) else {}
        ledger = payload.get("ledger_rows") if isinstance(payload.get("ledger_rows"), list) else []
        self.offline_ap_stats = stats
        self.offline_ap_ledger_rows = ledger
        self._offline_ap_context_loaded = True
        self._set_rows(self.offline_stats_table, OFFLINE_AP_STATS_COLUMNS, [stats] if stats else [])
        self._set_rows(self.offline_ledger_table, OFFLINE_AP_LEDGER_COLUMNS, ledger)
        self.offline_loading_spinner.setVisible(False)
        self.offline_loading_label.setText(f"离线AP数据已更新，共 {len(ledger)} 条")
        self.offline_loading_label.setVisible(True)
        app_logger.log_info("OFFLINE_AP_LEDGER_ASYNC_LOADED", f"count={len(ledger)}, cache={self.offline_cache_path}")

    def _fail_offline_ap_load(self, message: str) -> None:
        self.offline_loading_spinner.setVisible(False)
        self.offline_loading_label.setText(f"离线AP数据计算失败：{message}")
        self.offline_loading_label.setVisible(True)
        app_logger.log_error("OFFLINE_AP_LEDGER_ASYNC_FAILED", message)

    def _clear_offline_load_thread(self, thread: OfflineApLedgerLoadThread) -> None:
        if self.offline_load_thread is thread:
            self.offline_load_thread = None

    def open_web(self) -> None:
        device = self.current_device()
        if device is None:
            return
        port, _source = effective_https_port(device.https_port)
        if not build_https_url(device.ip_address, port):
            MessageBox.information(self, self.i18n.t("ac.open_web"), self.i18n.t("ac.https_port_not_collected"))
            return
        if not open_https_url(device.ip_address, port):
            MessageBox.warning(self, self.i18n.t("ac.open_web"), self.i18n.t("ac.open_web_failed"))

    def update_open_web_button(self) -> None:
        device = self.current_device()
        port, _source = effective_https_port(device.https_port if device is not None else None)
        enabled = device is not None and build_https_url(device.ip_address, port) is not None
        self.open_web_button.setEnabled(enabled)
        self.open_web_button.setToolTip("" if enabled else self.i18n.t("ac.https_port_not_collected"))
        app_logger.log_info("AC_HTTPS_PORT_BUTTON_STATE", f"device={device.name if device else ''}, effective_port={port}, button_enabled={enabled}")

    def _set_update_running(self, running: bool, message: str = "") -> None:
        selected_count = len(self.selected_ap_names())
        self.update_progress.setVisible(running)
        self.cancel_update_button.setVisible(running)
        self.cancel_update_button.setEnabled(running)
        self.device_combo.setEnabled(not running)
        self.update_ac_info_button.setEnabled(not running)
        self.persist_auto_ap_button.setEnabled(not running)
        self.enable_ap_remote_login_button.setEnabled(not running)
        self._hide_global_proxy_buttons()
        self.refresh_button.setEnabled(not running)
        self.refresh_optical_button.setEnabled(not running)
        self.batch_delete_button.setEnabled(False if running else selected_count > 0)
        self.import_button.setEnabled(not running)
        self.export_extension_template_button.setEnabled(not running)
        self.clear_selection_button.setEnabled(False if running else selected_count > 0)
        self.invert_selection_button.setEnabled(False if running else self.resources_table.rowCount() > 0)
        self.optical_concurrency_combo.setEnabled(not running)
        self.clear_optical_filters_button.setEnabled(not running)
        if message:
            self.status_label.setText(message)

    def _set_update_progress(self, message: str) -> None:
        self.update_progress.setVisible(True)
        self.status_label.setText(message)

    def cancel_current_update(self) -> None:
        if self.resource_thread is not None and self.resource_thread.isRunning():
            self.resource_thread.cancel()
        if self.ac_info_thread is not None and self.ac_info_thread.isRunning():
            self.ac_info_thread.cancel()
        if self.action_thread is not None and self.action_thread.isRunning():
            self.action_thread.cancel()
        if self.optical_thread is not None and self.optical_thread.isRunning():
            self.optical_thread.cancel()
        self.cancel_update_button.setEnabled(False)
        self.status_label.setText(self.i18n.t("ac.update_cancelled"))

    def shutdown(self) -> None:
        self.cancel_current_update()
        if self.trackside_export_job_id is not None:
            self.trackside_export_manager.cancel_export(self.trackside_export_job_id)
        for thread_name in ("offline_load_thread",):
            thread = getattr(self, thread_name, None)
            if thread is None:
                continue
            try:
                if hasattr(thread, "requestInterruption"):
                    thread.requestInterruption()
                if hasattr(thread, "quit"):
                    thread.quit()
                if hasattr(thread, "wait") and thread.isRunning():
                    thread.wait(500)
            except Exception as exc:
                app_logger.log_warning("AC_THREAD_SHUTDOWN_FAILED", f"{thread_name}: {exc}")
        for window in list(self.detail_windows):
            window.close()
        self.detail_windows.clear()

    def refresh_ac_resources(self) -> None:
        self.feature_gate.assert_enabled("ac.fit_ap_resources")
        device = self.current_device()
        if device is None:
            MessageBox.information(self, self.i18n.t("ac.title"), self.i18n.t("devices.select_first"))
            return
        self._set_update_running(True, self.i18n.t("ac.updating_resources"))
        self.resource_thread = AcResourceCollectThread(device, self.site_name, parent=self)
        self.resource_thread.progress.connect(self._set_update_progress)
        self.resource_thread.collect_finished.connect(self._finish_resource_collect)
        self.resource_thread.collect_failed.connect(self._fail_resource_collect)
        self.resource_thread.finished.connect(self.resource_thread.deleteLater)
        self.resource_thread.finished.connect(lambda: setattr(self, "resource_thread", None))
        self.resource_thread.start()

    def refresh_ac_info(self) -> None:
        self.feature_gate.assert_enabled("ac.ac_info_update")
        device = self.current_device()
        if not self._ensure_h3c_ac_selected(device):
            return
        self._set_update_running(True, "正在更新AC信息...")
        self.ac_info_thread = AcInfoCollectThread(device, self.site_name, parent=self)
        self.ac_info_thread.progress.connect(self._set_update_progress)
        self.ac_info_thread.collect_finished.connect(self._finish_ac_info_collect)
        self.ac_info_thread.collect_failed.connect(self._fail_resource_collect)
        self.ac_info_thread.finished.connect(self.ac_info_thread.deleteLater)
        self.ac_info_thread.finished.connect(lambda: setattr(self, "ac_info_thread", None))
        self.ac_info_thread.start()

    def run_ac_action(self, action: str, title: str) -> None:
        self.feature_gate.assert_enabled("ac.ac_actions")
        device = self.current_device()
        if not self._ensure_h3c_ac_selected(device):
            return
        profile = H3cAcCommandProfile(device)
        commands = getattr(profile, f"{action}_commands")
        command_text = "\n".join(commands)
        confirm_text = f"确认对当前AC执行以下命令？\n\n{command_text}"
        if action == "persist_auto_ap":
            confirm_text = f"该操作会将新上线 AP 固化，并执行 save force 保存配置。\n\n{confirm_text}"
        if MessageBox.question(self, title, confirm_text) != MessageBox.Yes:
            return
        self._set_update_running(True, f"正在执行：{title}...")
        self.action_thread = AcCommandActionThread(device, self.site_name, action, parent=self)
        self.action_thread.progress.connect(self._set_update_progress)
        self.action_thread.action_finished.connect(lambda result, action_title=title: self._finish_ac_action(result, action_title))
        self.action_thread.action_failed.connect(self._fail_resource_collect)
        self.action_thread.finished.connect(self.action_thread.deleteLater)
        self.action_thread.finished.connect(lambda: setattr(self, "action_thread", None))
        self.action_thread.start()

    def refresh_fit_ap_optical(self) -> None:
        self.feature_gate.assert_enabled("ac.fit_ap_optical")
        device = self.current_device()
        if device is None:
            return
        self._set_update_running(True, self.i18n.t("ac.updating_optical"))
        self.optical_thread = FitApOpticalCollectThread(device, self.site_name, int(self.optical_concurrency_combo.currentData() or 200), self)
        self.optical_thread.progress.connect(self._set_update_progress)
        self.optical_thread.collect_finished.connect(self._finish_optical_collect)
        self.optical_thread.collect_failed.connect(self._fail_optical_collect)
        self.optical_thread.finished.connect(self.optical_thread.deleteLater)
        self.optical_thread.finished.connect(lambda: setattr(self, "optical_thread", None))
        self.optical_thread.start()

    def _finish_resource_collect(self, result) -> None:
        self._set_update_progress(self.i18n.t("ac.refreshing_page"))
        self._set_update_running(False)
        if not result.success and result.error_message:
            self.status_label.setText(self.i18n.t("ac.status.failed"))
            if result.error_message != "用户已取消更新":
                MessageBox.warning(self, self.i18n.t("ac.title"), result.error_message)
            else:
                self.status_label.setText(self.i18n.t("ac.update_cancelled"))
            self.refresh_devices()
            return
        self.refresh_devices()
        if getattr(result, "https_port_collected", False) and result.https_port is not None:
            if not getattr(result, "https_port_persisted", False):
                self._apply_transient_https_port(result.https_port)
                self.status_label.setText(self.i18n.t("ac.update_done_https_save_failed", port=result.https_port))
            else:
                self.status_label.setText(self.i18n.t("ac.update_done_with_https", port=result.https_port))
        else:
            device = self.current_device()
            port, source = effective_https_port(device.https_port if device is not None else None)
            self.status_label.setText(
                self.i18n.t("ac.update_done_https_history", port=port)
                if source == "device"
                else self.i18n.t("ac.update_done_https_default", port=DEFAULT_HTTPS_PORT)
            )
        if hasattr(result, "fit_ap_resources_updated") or hasattr(result, "unauthenticated_rows_updated"):
            resource_count = int(getattr(result, "fit_ap_resources_updated", 0) or 0)
            unauth_count = int(getattr(result, "unauthenticated_rows_updated", 0) or 0)
            unauth_error = str(getattr(result, "unauthenticated_error", "") or "").strip()
            if unauth_error:
                extra = f"FIT-AP资源 {resource_count} 条；新上线AP识别失败：{unauth_error}"
            else:
                extra = f"FIT-AP资源 {resource_count} 条；新上线AP {unauth_count} 条"
            self.status_label.setText(f"{self.status_label.text()}；{extra}")
        device = self.current_device()
        app_logger.log_info("AC_HTTPS_PORT_UI_REFRESHED", f"device={device.name if device else ''}, ui_port={device.https_port if device else None}")

    def _finish_ac_info_collect(self, result) -> None:
        self._set_update_progress(self.i18n.t("ac.refreshing_page"))
        self._set_update_running(False)
        if not result.success and result.error_message:
            self.status_label.setText(self.i18n.t("ac.status.failed"))
            if result.error_message != "用户已取消更新":
                MessageBox.warning(self, self.i18n.t("ac.title"), result.error_message)
            else:
                self.status_label.setText(self.i18n.t("ac.update_cancelled"))
            self.refresh_devices()
            return
        self.refresh_devices()
        if getattr(result, "https_port_collected", False) and result.https_port is not None:
            if not getattr(result, "https_port_persisted", False):
                self._apply_transient_https_port(result.https_port)
                self.status_label.setText(f"AC信息更新完成；HTTPS端口 {result.https_port} 已解析但保存失败")
            else:
                self.status_label.setText(f"AC信息更新完成；HTTPS端口 {result.https_port}")
        else:
            self.status_label.setText("AC信息更新完成；HTTPS端口未解析，打开网页时将使用默认443")

    def _finish_ac_action(self, result, title: str) -> None:
        self._set_update_running(False)
        if result.success:
            if getattr(result, "action", "") == "persist_auto_ap":
                self.status_label.setText("一键固化新上线AP完成，已执行 save force")
                return
            if getattr(result, "action", "") == "enable_ap_remote_login" and any(
                item.success and str(item.error_message or "").startswith("warning: read timeout")
                for item in getattr(result, "command_results", [])
            ):
                self.status_label.setText("一键开启AP远程登入完成；设备尾部回显读取超时，已按命令成功处理")
                return
            self.status_label.setText(f"{title}执行成功")
            return
        self.status_label.setText(f"{title}执行失败")
        MessageBox.warning(self, title, result.error_message or "命令执行失败")

    def _ensure_h3c_ac_selected(self, device: Device | None) -> bool:
        if device is None:
            MessageBox.information(self, self.i18n.t("ac.title"), "请先选择 AC")
            return False
        if str(device.device_vendor or "").upper() != "H3C" or str(device.device_type or "").upper() != "AC":
            MessageBox.warning(self, self.i18n.t("ac.title"), "该功能只支持 H3C AC 设备")
            return False
        return True

    def _apply_transient_https_port(self, port: int) -> None:
        device = self.current_device()
        if device is None:
            return
        device.https_port = port
        self._set_summary(self.repository.get_ac_ap_summary(str(device.device_uuid or "")))
        self.update_open_web_button()

    def _fail_resource_collect(self, message: str) -> None:
        self._set_update_running(False)
        self.status_label.setText(self.i18n.t("ac.status.failed"))
        MessageBox.warning(self, self.i18n.t("ac.title"), message)

    def _finish_optical_collect(self, result) -> None:
        self._set_update_progress(self.i18n.t("ac.refreshing_page"))
        self._set_update_running(False)
        self.status_label.setText(self.i18n.t("ac.status.done" if result.success else "ac.status.failed"))
        if getattr(result, "error_message", None) == "用户已取消更新":
            self.status_label.setText(self.i18n.t("ac.update_cancelled"))
        self.refresh_data()

    def _fail_optical_collect(self, message: str) -> None:
        self._set_update_running(False)
        self.status_label.setText(self.i18n.t("ac.status.failed"))
        MessageBox.warning(self, self.i18n.t("ac.title"), message)

    def current_device_uuid(self) -> str | None:
        value = self.device_combo.currentData()
        return str(value) if value else None

    def current_device(self) -> Device | None:
        current_uuid = self.current_device_uuid()
        for device in self.ac_devices:
            if device.device_uuid == current_uuid:
                return device
        return None

    def _ap_selection_key(self, row: dict[str, object | None]) -> str:
        return str(row.get("ap_uuid") or row.get("ap_name") or "").strip()

    def _sync_visible_resource_selection(self) -> None:
        for row in range(self.resources_table.rowCount()):
            item = self.resources_table.item(row, CHECK_COLUMN)
            if not item:
                continue
            key = str(item.data(Qt.UserRole) or "").strip()
            if not key:
                continue
            if is_checked_value(item.checkState()):
                self.selected_ap_keys.add(key)
            else:
                self.selected_ap_keys.discard(key)

    def selected_ap_names(self) -> list[str]:
        self._sync_visible_resource_selection()
        return sorted(self.selected_ap_keys)

    def checked_or_all_ap_rows(self) -> list[dict[str, object | None]]:
        ac_uuid = self.current_device_uuid()
        if not ac_uuid:
            return []
        rows = self.repository.list_fit_ap_resources_with_metadata(ac_uuid)
        selected = set(self.selected_ap_names())
        return [row for row in rows if not selected or self._ap_selection_key(row) in selected]

    def update_selection_state(self, *_args) -> None:
        self._sync_visible_resource_selection()
        count = len(self.selected_ap_names())
        self.selection_label.setText(
            self.i18n.t(
                "ap.resource_filter_stats",
                total=len(self.resource_rows),
                visible=len(self.filtered_resource_rows()),
                selected=count,
            )
        )
        self.batch_delete_button.setEnabled(count > 0)
        self.clear_selection_button.setEnabled(count > 0)
        self.invert_selection_button.setEnabled(self.resources_table.rowCount() > 0)

    def _resource_header_clicked(self, column: int) -> None:
        if column == CHECK_COLUMN:
            self._set_all_checked(len(self.selected_ap_names()) != self.resources_table.rowCount())

    def _set_all_checked(self, checked: bool) -> None:
        self.resources_table.blockSignals(True)
        set_all_table_rows_checked(self.resources_table, checked, CHECK_COLUMN)
        for row in range(self.resources_table.rowCount()):
            item = self.resources_table.item(row, CHECK_COLUMN)
            if item:
                key = str(item.data(Qt.UserRole) or "").strip()
                if key:
                    if checked:
                        self.selected_ap_keys.add(key)
                    else:
                        self.selected_ap_keys.discard(key)
        self.resources_table.blockSignals(False)
        self.update_selection_state()

    def clear_selection(self) -> None:
        self.selected_ap_keys.clear()
        self._set_all_checked(False)

    def invert_selection(self) -> None:
        self.resources_table.blockSignals(True)
        invert_table_rows_checked(self.resources_table, CHECK_COLUMN)
        for row in range(self.resources_table.rowCount()):
            item = self.resources_table.item(row, CHECK_COLUMN)
            if item:
                key = str(item.data(Qt.UserRole) or "").strip()
                checked = is_checked_value(item.checkState())
                if key:
                    if checked:
                        self.selected_ap_keys.add(key)
                    else:
                        self.selected_ap_keys.discard(key)
        self.resources_table.blockSignals(False)
        self.update_selection_state()

    def batch_delete_aps(self) -> None:
        ac_uuid = self.current_device_uuid()
        names = self.selected_ap_names()
        if not ac_uuid or not names:
            return
        answer = MessageBox.question(self, self.i18n.t("ac.title"), self.i18n.t("ap.batch_delete_confirm"))
        if answer != MessageBox.Yes:
            return
        count = self.repository.delete_fit_aps(ac_uuid, names)
        app_logger.log_info("FIT_AP_BATCH_DELETE", f"ac={ac_uuid}, count={count}")
        self.refresh_data()

    def import_metadata(self) -> None:
        self.feature_gate.assert_enabled("ac.fit_ap_resources")
        path, _ = QFileDialog.getOpenFileName(self, self.i18n.t("ap.import_metadata"), "", "Excel Files (*.xlsx);;CSV Files (*.csv)")
        if not path:
            return
        self._start_background_job(
            "fit_ap_metadata_import",
            {"path": path, "db_path": str(self.repository.database.path)},
            title=self.i18n.t("ap.import_metadata"),
        )

    def refresh_ap_extensions(self) -> None:
        search = self.extension_search_input.text() if hasattr(self, "extension_search_input") else ""
        self.extension_rows = self.repository.list_ap_extension_points(search=search)
        self._set_rows(self.extension_table, AP_EXTENSION_COLUMNS, self.extension_rows)
        self._set_rows(self.extension_summary_table, AP_EXTENSION_SUMMARY_COLUMNS, self._extension_summary_rows(self.extension_rows))
        self._set_rows(self.extension_issue_table, AP_EXTENSION_ISSUE_COLUMNS, self._extension_issue_rows(self.extension_rows))

    def import_ap_extensions(self, import_mode: str) -> None:
        self.feature_gate.assert_enabled("ac.fit_ap_extensions")
        title = "标准模板导入" if import_mode == "standard_template" else "原始设计/布点表智能识别导入"
        path, _ = QFileDialog.getOpenFileName(self, title, "", "Excel/CSV Files (*.xlsx *.csv)")
        if not path:
            return
        self._start_background_job(
            "fit_ap_extension_preview",
            {"path": path, "db_path": str(self.repository.database.path), "import_mode": import_mode},
            title=title,
        )

    def _start_background_job(self, task_type: str, params: dict[str, object], *, title: str) -> str:
        job_id = uuid.uuid4().hex
        self._background_jobs[job_id] = {"task_type": task_type, "title": title, **params}
        self.status_label.setText(f"{title}：后台处理中...")
        self.background_manager.start_job(BackgroundJob(job_id=job_id, task_type=task_type, params=params))
        return job_id

    def _background_progress(self, event: dict) -> None:
        message = str(event.get("message") or "")
        if message:
            self.status_label.setText(message)

    def _background_finished(self, event: dict) -> None:
        job_id = str(event.get("job_id") or "")
        context = self._background_jobs.pop(job_id, {})
        task_type = str(context.get("task_type") or "")
        title = str(context.get("title") or "后台任务")
        result = dict(event.get("result") or {})
        if task_type == "fit_ap_metadata_import":
            app_logger.log_info("FIT_AP_IMPORT", f"updated={result.get('updated', 0)}, skipped={result.get('skipped', 0)}")
            self.status_label.setText(f"{title}：完成，更新 {result.get('updated', 0)} 条，跳过 {result.get('skipped', 0)} 条")
            self.refresh_data()
            return
        if task_type == "fit_ap_extension_preview":
            self._confirm_ap_extension_import(title, context, result)
            return
        if task_type == "fit_ap_extension_commit":
            MessageBox.information(
                self,
                title,
                f"新增：{result.get('success_rows', 0)}\n更新：{result.get('updated_rows', 0)}\n跳过：{result.get('skipped_rows', 0)}\n错误：{result.get('error_rows', 0)}",
            )
            self.status_label.setText(f"{title}：完成")
            self.refresh_ap_extensions()
            self.refresh_data()

    def _background_failed(self, event: dict) -> None:
        job_id = str(event.get("job_id") or "")
        context = self._background_jobs.pop(job_id, {})
        title = str(context.get("title") or "后台任务")
        message = str(event.get("message") or event.get("error") or "后台任务失败")
        if "Unsupported AP metadata template header" in message:
            message = self.i18n.t("ap.metadata_template_unsupported")
        self.status_label.setText(f"{title}：失败")
        MessageBox.warning(self, title, message)

    def _confirm_ap_extension_import(self, title: str, context: dict[str, object], preview: dict[str, object]) -> None:
        summary = dict(preview.get("summary") or {})
        lines = [
            f"文件名：{preview.get('file_name') or '-'}",
            f"模板类型：{preview.get('template_type') or '-'}",
            f"识别置信度：{preview.get('confidence_score') or 0}",
            f"工作表数：{preview.get('sheet_count') or 0}",
            f"预览数据行：{summary.get('total_rows', 0)}",
            f"未绑定 AP MAC 数量：{summary.get('unbound_rows', 0)}",
            f"无效 MAC 数量：{summary.get('invalid_mac_rows', 0)}",
        ]
        if bool(preview.get("low_confidence")):
            lines.append("识别置信度较低，请手动映射字段。")
            MessageBox.warning(self, title, "\n".join(lines))
            return
        if MessageBox.question(self, title, "\n".join(lines) + "\n\n确认导入以上预览数据？") != MessageBox.Yes:
            return
        self._start_background_job(
            "fit_ap_extension_commit",
            {
                "path": str(context.get("path") or ""),
                "db_path": str(self.repository.database.path),
                "import_mode": str(context.get("import_mode") or "standard_template"),
            },
            title=title,
        )

    def export_ap_extensions(self) -> None:
        self.feature_gate.assert_enabled("ac.fit_ap_extensions")
        path = select_export_path(self, "导出FIT-AP扩展信息", f"{self.site_name}_FIT-AP扩展信息.xlsx", EXCEL_FILTER)
        if not path:
            return
        search = self.extension_search_input.text().strip() if hasattr(self, "extension_search_input") else ""
        submit_export_task(
            self,
            fit_ap_extension_xlsx_spec(
                path,
                db_path=self.repository.database.path,
                ac_uuid=self.current_device_uuid(),
                search=search,
                title="导出FIT-AP扩展信息",
                open_dir_on_success=True,
            ),
            success_title="导出FIT-AP扩展信息",
            paths=self.settings.paths,
        )
        remember_export_path(Path(path))

    def add_ap_extension_point(self) -> None:
        self._edit_ap_extension_point({})

    def edit_ap_extension_point(self, row: int) -> None:
        if row < 0 or row >= len(self.extension_rows):
            return
        self._edit_ap_extension_point(self.extension_rows[row])

    def _edit_ap_extension_point(self, row: dict[str, object | None]) -> None:
        dialog = QDialog(self)
        dialog.setWindowTitle("FIT-AP扩展信息")
        dialog.resize(720, 620)
        dialog.setMinimumSize(640, 520)
        form_widget = QWidget()
        form = QFormLayout(form_widget)
        fields = (
            ("belong_type", "归属类型"),
            ("station_name", "归属站点"),
            ("section_name", "归属区间"),
            ("section_start_station", "区间起点站"),
            ("section_end_station", "区间终点站"),
            ("yard_name", "场段"),
            ("area_name", "区域"),
            ("network_domain", "网络"),
            ("line_side", "线别"),
            ("direction", "方向"),
            ("mileage_text", "里程"),
            ("location_desc", "点位说明"),
            ("ap_point_code", "AP编号"),
            ("ap_name", "AP名称"),
            ("ap_mac_display", "AP MAC"),
            ("power_station", "供电站"),
            ("power_distribution", "电源分配"),
            ("fiber_access_station", "光缆接入站"),
            ("fiber_distribution", "光缆分配"),
            ("remark", "备注"),
        )
        editors: dict[str, QLineEdit] = {}
        for field, label in fields:
            value = _display_mileage_for_row(row, field) if field == "mileage_text" and row else row.get(field)
            editor = QLineEdit("" if value is None else str(value))
            editors[field] = editor
            form.addRow(label, editor)
        buttons = QHBoxLayout()
        save_button = QPushButton("保存")
        cancel_button = QPushButton("取消")
        buttons.addStretch(1)
        buttons.addWidget(save_button)
        buttons.addWidget(cancel_button)
        layout = QVBoxLayout()
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setWidget(form_widget)
        layout.addWidget(scroll, 1)
        layout.addLayout(buttons)
        dialog.setLayout(layout)
        save_button.clicked.connect(dialog.accept)
        cancel_button.clicked.connect(dialog.reject)
        if dialog.exec() != QDialog.Accepted:
            return
        payload = {field: editor.text().strip() for field, editor in editors.items()}
        if row.get("id"):
            payload["id"] = row.get("id")
        self.repository.upsert_ap_extension_point(payload)
        self.refresh_ap_extensions()
        self.refresh_data()

    def delete_selected_ap_extension_points(self) -> None:
        selected_rows = sorted({index.row() for index in self.extension_table.selectionModel().selectedRows()})
        ids = [int(self.extension_rows[row].get("id") or 0) for row in selected_rows if 0 <= row < len(self.extension_rows)]
        if not ids:
            return
        if MessageBox.question(self, "批量删除", f"确认删除 {len(ids)} 条 FIT-AP扩展信息？") != MessageBox.Yes:
            return
        self.repository.delete_ap_extension_points(ids)
        self.refresh_ap_extensions()
        self.refresh_data()

    def clear_ap_extension_points(self) -> None:
        if MessageBox.question(self, "清空当前局点扩展信息", "该操作会删除当前局点全部 FIT-AP扩展信息，确认继续？") != MessageBox.Yes:
            return
        count = self.repository.clear_ap_extension_points()
        MessageBox.information(self, "清空当前局点扩展信息", f"已删除 {count} 条。")
        self.refresh_ap_extensions()
        self.refresh_data()

    @staticmethod
    def _extension_summary_rows(rows: list[dict[str, object | None]]) -> list[dict[str, object | None]]:
        summary: dict[str, dict[str, object | None]] = {}
        for row in rows:
            station = str(row.get("station_name") or row.get("section_name") or row.get("yard_name") or "未填写").strip()
            item = summary.setdefault(station, {"station_name": station, "total": 0, "bound": 0, "unbound": 0, "left": 0, "right": 0, "curve": 0, "risk": 0})
            item["total"] = int(item["total"] or 0) + 1
            item["bound" if row.get("ap_mac_norm") else "unbound"] = int(item["bound" if row.get("ap_mac_norm") else "unbound"] or 0) + 1
            if row.get("line_side") == "左线":
                item["left"] = int(item["left"] or 0) + 1
            if row.get("line_side") == "右线":
                item["right"] = int(item["right"] or 0) + 1
            if row.get("curve_flag"):
                item["curve"] = int(item["curve"] or 0) + 1
            if row.get("interval_risk_level") and row.get("interval_risk_level") != "正常":
                item["risk"] = int(item["risk"] or 0) + 1
        return list(summary.values())

    @staticmethod
    def _extension_issue_rows(rows: list[dict[str, object | None]]) -> list[dict[str, object | None]]:
        issues: list[dict[str, object | None]] = []
        seen: dict[str, int] = {}
        for row in rows:
            row_id = int(row.get("id") or 0)
            mac = str(row.get("ap_mac_norm") or "")
            if not mac:
                issues.append({"type": "缺 AP MAC", "severity": "提示", "source_row": row.get("source_row"), "message": "记录将作为未绑定点位保存"})
            elif mac in seen:
                issues.append({"type": "AP MAC 重复", "severity": "警告", "source_row": row.get("source_row"), "ap_mac_norm": mac, "message": f"与扩展记录 {seen[mac]} 重复"})
            else:
                seen[mac] = row_id
            if row.get("mileage_text") and row.get("mileage_m") is None:
                issues.append({"type": "里程无法解析", "severity": "提示", "source_row": row.get("source_row"), "message": row.get("mileage_text")})
            if row.get("interval_risk_level") and row.get("interval_risk_level") != "正常":
                issues.append({"type": "AP间隔风险", "severity": "提示", "source_row": row.get("source_row"), "message": row.get("interval_risk_reason")})
            if row.get("match_status") == "extension_not_online":
                issues.append({"type": "扩展信息未采集到", "severity": "提示", "source_row": row.get("source_row"), "ap_mac_norm": mac, "message": "扩展信息中存在，但 AC 当前没有采集到"})
        return issues

    def export_ap_extension_template(self) -> None:
        ac_uuid = self.current_device_uuid()
        filename = make_ap_extension_template_filename(self._current_ac_export_name())
        path = self._select_ap_extension_template_path(filename)
        if not path:
            return
        submit_export_task(
            self,
            fit_ap_extension_template_xlsx_spec(
                path,
                db_path=self.repository.database.path,
                ac_uuid=ac_uuid or "",
                title=self.i18n.t("ap.export_extension_template"),
                open_dir_on_success=True,
            ),
            success_title=self.i18n.t("ap.export_extension_template"),
            paths=self.settings.paths,
        )
        remember_export_path(path)
        app_logger.log_info("FIT_AP_EXTENSION_TEMPLATE_EXPORT", f"ac={ac_uuid or '-'}, file={path.name}")

    def _current_ac_export_name(self) -> str:
        device = self.current_device()
        if device is None:
            return self.site_name
        return str(device.name or device.ip_address or self.site_name)

    def _select_ap_extension_template_path(self, filename: str) -> Path | None:
        desktop = QStandardPaths.writableLocation(QStandardPaths.DesktopLocation)
        base_dir = Path(desktop) if desktop else Path.home()
        selected, _ = QFileDialog.getSaveFileName(
            self,
            self.i18n.t("ap.export_extension_template"),
            str(base_dir / filename),
            EXCEL_FILTER,
        )
        return Path(selected) if selected else None

    def export_aps(self) -> None:
        self.feature_gate.assert_enabled("ac.fit_ap_resources")
        path = select_export_path(self, self.i18n.t("ap.export_info"), make_fit_ap_export_filename(self.site_name), CSV_FILTER)
        if not path:
            return
        rows = self.checked_or_all_ap_rows()
        submit_export_task(
            self,
            fit_ap_csv_spec(
                path,
                db_path=self.repository.database.path,
                rows=rows,
                title=self.i18n.t("ap.export_info"),
                open_dir_on_success=True,
            ),
            success_title=self.i18n.t("ap.export_info"),
            paths=self.settings.paths,
        )
        remember_export_path(path)
        app_logger.log_info("FIT_AP_EXPORT", f"count={len(rows)}, file={path.name}")

    def export_omnipeek_name_table(self) -> None:
        self.feature_gate.assert_enabled("ac.omnipeek_name_table_export")
        service = OmniPeekNameTableService(self.repository, self.device_repository)
        ac_uuid = self.current_device_uuid()
        items = service.collect_items(
            include_ac_fit_ap=True,
            include_ap_extensions=True,
            include_device_mr=True,
            ac_device_uuid=ac_uuid,
            group_names=self._device_group_names(),
        )
        if not items:
            MessageBox.warning(self, "导出 OmniPeek 名称表", "当前没有可导出的轨旁 AP 或车载 MR 数据。")
            return
        source_counts = service.source_counts(ac_device_uuid=ac_uuid)
        source_counts[SOURCE_DEVICE_MANAGEMENT] = sum(1 for item in items if SOURCE_DEVICE_MANAGEMENT in (item.sources or [item.source]))
        dialog = OmniPeekExportDialog(
            items,
            source_counts,
            default_line_name=default_omnipeek_line_name(self.site_name, self.settings.paths),
            settings=self.settings,
            parent=self,
        )
        dialog.exec()

    def _device_group_names(self) -> dict[int, str]:
        try:
            groups = DeviceGroupRepository(self.device_repository.database, self.site_name).list()
        except Exception as exc:
            app_logger.log_error("OMNIPEEK_GROUP_LOAD_FAILED", str(exc))
            return {}
        return {int(group.id): group.name for group in groups if group.id is not None}

    def current_optical_filters(self) -> dict[str, object | None]:
        return {
            "ap_name": self.optical_ap_filter.text(),
            "site": self.optical_site_filter.currentData(),
        }

    def filtered_optical_rows(self) -> list[dict[str, object | None]]:
        return filter_fit_ap_optical_rows(sort_fit_ap_optical_rows(self.optical_rows), self.current_optical_filters())

    def current_optical_page_rows(self) -> list[dict[str, object | None]]:
        rows, _state = paginate_rows(self.filtered_optical_rows(), self.optical_page_size, self.optical_page)
        return rows

    def apply_optical_filters(self) -> None:
        self.optical_page = 1
        self.apply_optical_pagination()

    def current_resource_filters(self) -> dict[str, object | None]:
        return {
            "search": self.resource_search_input.text(),
            "group": self.resource_group_filter.currentData(),
            "state": self.resource_state_filter.currentData(),
        }

    def filtered_resource_rows(self) -> list[dict[str, object | None]]:
        rows = self.resource_rows
        filters = self.current_resource_filters()
        search = _normalize_resource_search(filters.get("search"))
        if search:
            rows = [
                row
                for row in rows
                if search in _normalize_resource_search(row.get("ap_name"))
                or search in _normalize_resource_search(row.get("ap_mac"))
                or search in _normalize_resource_search(row.get("ap_ip"))
                or search in _normalize_resource_search(row.get("serial_number"))
                or search in _normalize_resource_search(row.get("rid1_bbssid"))
                or search in _normalize_resource_search(row.get("rid2_bbssid"))
                or search in _normalize_resource_search(row.get("rid3_bbssid"))
                or search in _normalize_resource_search(row.get("lldp_neighbor_mac"))
                or search in _normalize_resource_search(row.get("lldp_neighbor_interface"))
                or search in _normalize_resource_search(row.get("mileage"))
                or search in _normalize_resource_search(_resource_mileage_search_text(row))
            ]
        group = str(filters.get("group") or "").strip()
        if group:
            rows = [row for row in rows if str(row.get("group_name") or "").strip() == group]
        state = str(filters.get("state") or "").strip()
        if state:
            rows = [row for row in rows if _resource_state_matches(row, state)]
        return rows

    def apply_resource_filters(self) -> None:
        self.resource_page = 1
        self.apply_resource_pagination()

    def apply_resource_pagination(self) -> None:
        rows, state = paginate_rows(self.filtered_resource_rows(), self.resource_page_size, self.resource_page)
        self.resource_page = state.current_page
        self.resources_pagination.set_state(state)
        self._set_rows(self.resources_table, FIT_AP_RESOURCE_COLUMNS, [build_fit_ap_resource_table_row(row) for row in rows])
        self.update_selection_state()

    def current_resource_page_rows(self) -> list[dict[str, object | None]]:
        rows, _state = paginate_rows(self.filtered_resource_rows(), self.resource_page_size, self.resource_page)
        return rows

    def set_resource_page(self, page: int) -> None:
        self.resource_page = page
        self.apply_resource_pagination()

    def set_resource_page_size(self, page_size: int) -> None:
        self.resource_page_size = page_size
        self.resource_page = 1
        self.apply_resource_pagination()

    def apply_optical_pagination(self) -> None:
        rows, state = paginate_rows(self.filtered_optical_rows(), self.optical_page_size, self.optical_page)
        self.optical_page = state.current_page
        self.optical_pagination.set_state(state)
        self._set_rows(self.optical_table, FIT_AP_OPTICAL_COLUMNS, rows)

    def set_optical_page(self, page: int) -> None:
        self.optical_page = page
        self.apply_optical_pagination()

    def set_optical_page_size(self, page_size: int) -> None:
        self.optical_page_size = page_size
        self.optical_page = 1
        self.apply_optical_pagination()

    def current_trackside_filters(self) -> dict[str, object | None]:
        return {
            "site": self.trackside_site_filter.currentData(),
            "search": self.trackside_search_input.text(),
        }

    def filtered_trackside_rows(self) -> list[dict[str, object | None]]:
        filters = self.current_trackside_filters()
        return filter_trackside_ap_business_rows(self.trackside_rows, filters.get("site"), filters.get("search"))

    def current_trackside_page_rows(self) -> list[dict[str, object | None]]:
        rows, _state = paginate_rows(self.filtered_trackside_rows(), self.trackside_page_size, self.trackside_page)
        return rows

    def apply_trackside_filters(self) -> None:
        self.trackside_page = 1
        self.apply_trackside_pagination()

    def apply_trackside_pagination(self) -> None:
        rows, state = paginate_rows(self.filtered_trackside_rows(), self.trackside_page_size, self.trackside_page)
        self.trackside_page = state.current_page
        self.trackside_pagination.set_state(state)
        self._set_rows(self.trackside_table, TRACKSIDE_AP_BUSINESS_COLUMNS, rows)

    def set_trackside_page(self, page: int) -> None:
        self.trackside_page = page
        self.apply_trackside_pagination()

    def set_trackside_page_size(self, page_size: int) -> None:
        self.trackside_page_size = page_size
        self.trackside_page = 1
        self.apply_trackside_pagination()

    def clear_optical_filters(self) -> None:
        self.optical_ap_filter.clear()
        self.optical_site_filter.setCurrentIndex(0)
        self.clear_selection()
        self.apply_optical_filters()

    def export_optical_table(self) -> None:
        self.feature_gate.assert_enabled("ac.fit_ap_optical")
        path = select_export_path(self, self.i18n.t("ac.export_table"), make_fit_ap_optical_export_filename(self.site_name), EXCEL_FILTER)
        if not path:
            return
        self.refresh_overview_table()
        rows = self.filtered_optical_rows()
        optical_rows = []
        for row in rows:
            export_row = dict(row)
            for _key, field in FIT_AP_OPTICAL_COLUMNS:
                export_row[field] = _fit_ap_optical_export_value(row, field)
            export_row["_row_status"] = evaluate_fit_ap_row_status(row)
            optical_rows.append(export_row)
        submit_export_task(
            self,
            ExportTaskSpec(
                task_type="multi_sheet_xlsx",
                output_path=str(path),
                title=self.i18n.t("ac.export_table"),
                open_dir_on_success=True,
                payload={
                    "sheets": [
                        {
                            "sheet_name": "AP上线情况概览",
                            "columns": [
                                {"key": field, "title": self.i18n.t(key), "text": True}
                                for key, field in AP_ONLINE_OVERVIEW_COLUMNS
                            ],
                            "rows": self.current_overview_rows(),
                            "freeze_header": True,
                            "auto_filter": True,
                        },
                        {
                            "sheet_name": "FIT-AP光衰",
                            "columns": [
                                {"key": field, "title": self.i18n.t(key), "text": True}
                                for key, field in FIT_AP_OPTICAL_COLUMNS
                            ],
                            "rows": optical_rows,
                            "freeze_header": True,
                            "auto_filter": True,
                            "row_fill_field": "_row_status",
                            "row_fill_colors": OPTICAL_EXPORT_COLOR_RGB,
                        },
                    ]
                },
            ),
            success_title=self.i18n.t("ac.export_table"),
            paths=self.settings.paths,
        )
        remember_export_path(path)
        app_logger.log_info("FIT_AP_OPTICAL_EXPORT", f"count={len(rows)}, file={path.name}")

    def export_overview_table(self) -> None:
        path = select_export_path(self, self.i18n.t("ac.export_overview"), make_fit_ap_optical_export_filename(self.site_name), EXCEL_FILTER)
        if not path:
            return
        self.refresh_overview_table()
        self.ensure_offline_ap_context_loaded(force=True)
        rows = self.current_overview_rows()
        submit_export_task(
            self,
            ap_online_overview_xlsx_spec(
                path,
                rows=rows,
                headers=[self.i18n.t(key) for key, _field in AP_ONLINE_OVERVIEW_COLUMNS],
                offline_ap_stats=self.offline_ap_stats,
                offline_ap_ledger_rows=self.offline_ap_ledger_rows,
                offline_ap_stats_headers=offline_ap_headers(OFFLINE_AP_STATS_COLUMNS),
                offline_ap_ledger_headers=offline_ap_headers(OFFLINE_AP_LEDGER_COLUMNS),
                title=self.i18n.t("ac.export_overview"),
                open_dir_on_success=True,
            ),
            success_title=self.i18n.t("ac.export_overview"),
            paths=self.settings.paths,
        )
        remember_export_path(path)
        app_logger.log_info("AP_ONLINE_OVERVIEW_EXPORT", f"count={len(rows)}, file={path.name}")

    def export_trackside_table(self) -> None:
        if self.trackside_export_job_id is not None:
            self.status_label.setText(self.i18n.t("trackside.export.progress_write"))
            return
        path = select_export_path(self, self.i18n.t("trackside.export"), f"{self.site_name}_trackside_ap_business_{datetime.now().strftime('%Y-%m-%d-%H%M')}.xlsx", EXCEL_FILTER)
        if not path:
            return
        job_id = uuid.uuid4().hex
        self.trackside_export_job_id = job_id
        self.trackside_export_output_path = Path(path)
        self._set_trackside_export_running(True, self.i18n.t("trackside.export.progress_load"))
        job = ExportJob(
            job_id=job_id,
            job_type="trackside_ap_business",
            site_name=self.site_name,
            db_path=str(self.device_repository.database.path),
            output_path=str(path),
            params={"language": self.i18n.language},
        )
        try:
            self.trackside_export_manager.start_export(job)
        except Exception as exc:
            self.trackside_export_job_id = None
            self.trackside_export_output_path = None
            self._set_trackside_export_running(False)
            MessageBox.warning(self, self.i18n.t("trackside.export"), f"导出失败：{exc}")

    def _set_trackside_export_running(self, running: bool, message: str = "") -> None:
        self.update_progress.setVisible(running or self.resource_thread is not None or self.optical_thread is not None)
        if running:
            self.update_progress.setRange(0, 0)
            self.update_progress.setValue(0)
        self.trackside_export_button.setEnabled(not running)
        self.refresh_button.setEnabled(not running and self.resource_thread is None and self.optical_thread is None)
        if message:
            self.status_label.setText(message)

    def _finish_trackside_export(self, result: dict[str, object]) -> None:
        if str(result.get("job_id") or "") != str(self.trackside_export_job_id or ""):
            return
        path = Path(result.get("output_path") or self.trackside_export_output_path or "")
        self.trackside_export_job_id = None
        self.trackside_export_output_path = None
        self._set_trackside_export_running(False)
        self.update_progress.setVisible(True)
        self.update_progress.setRange(0, 1)
        self.update_progress.setValue(1)
        if path:
            remember_export_path(path)
        count = int(result.get("row_count") or 0)
        self.status_label.setText(self.i18n.t("trackside.loaded_count", count=count))
        app_logger.log_info("TRACKSIDE_AP_BUSINESS_EXPORT", f"count={count}, file={path.name if path else ''}")
        self._show_trackside_export_finished_dialog(path)

    def _fail_trackside_export(self, payload: dict[str, object]) -> None:
        if str(payload.get("job_id") or "") != str(self.trackside_export_job_id or ""):
            return
        cancelled = bool(payload.get("cancelled"))
        message = str(payload.get("message") or payload.get("error") or ("已取消导出" if cancelled else "导出失败"))
        self.trackside_export_job_id = None
        self.trackside_export_output_path = None
        self._set_trackside_export_running(False)
        self.update_progress.setVisible(False)
        if cancelled:
            self.status_label.setText("已取消导出")
            app_logger.log_info("TRACKSIDE_AP_BUSINESS_EXPORT_CANCELLED", message)
            return
        self.status_label.setText(self.i18n.t("trackside.export"))
        app_logger.log_error("TRACKSIDE_AP_BUSINESS_EXPORT_FAILED", message)
        MessageBox.warning(self, self.i18n.t("trackside.export"), message)

    def _handle_trackside_export_progress(self, event: dict[str, object]) -> None:
        if str(event.get("job_id") or "") != str(self.trackside_export_job_id or ""):
            return
        current = int(event.get("current", event.get("done", 0)) or 0)
        total = int(event.get("total") or 0)
        if total > 0:
            self.update_progress.setRange(0, total)
            self.update_progress.setValue(min(current, total))
        else:
            self.update_progress.setRange(0, 0)
        message = str(event.get("message") or event.get("stage") or "")
        if message:
            self.status_label.setText(f"正在导出轨旁AP业务：{message}")

    def _show_trackside_export_finished_dialog(self, path: Path) -> None:
        if not path:
            return
        box = MessageBox(self)
        box.setIcon(MessageBox.Information)
        box.setWindowTitle(self.i18n.t("trackside.export"))
        box.setText(f"轨旁AP业务导出完成：\n{path}")
        open_file_button = box.addButton("打开文件", MessageBox.ActionRole)
        open_dir_button = box.addButton("打开目录", MessageBox.ActionRole)
        box.addButton("关闭", MessageBox.RejectRole)
        box.exec()
        clicked = box.clickedButton()
        if clicked is open_file_button:
            QDesktopServices.openUrl(QUrl.fromLocalFile(str(path)))
        elif clicked is open_dir_button:
            QDesktopServices.openUrl(QUrl.fromLocalFile(str(path.parent)))

    def _handle_ac_summary_changed(self, site_name: str) -> None:
        if str(site_name or "") != str(self.site_name or ""):
            return
        self.refresh_devices()

    def save_overview_history_snapshot(self) -> None:
        count = self.repository.save_station_online_summary_history(self.current_overview_rows())
        MessageBox.information(self, self.i18n.t("ac.ap_online_overview"), self.i18n.t("ac.history_snapshot_saved"))
        app_logger.log_info("AP_ONLINE_OVERVIEW_HISTORY_SAVE", f"count={count}")

    def open_overview_history(self) -> None:
        site_name = self.selected_overview_site()
        rows = self.repository.list_station_online_summary_history(site_name=site_name)
        dialog = StationOnlineHistoryDialog(self.i18n, rows, site_name)
        self.detail_windows.append(dialog)
        dialog.destroyed.connect(lambda _=None, window=dialog: self._forget_detail_window(window))
        show_non_focus_window(self, dialog, key=f"station_online_history:{site_name}", activate=False, raise_window=False)

    def selected_overview_site(self) -> str | None:
        selected = self.overview_table.selectionModel().selectedRows() if self.overview_table.selectionModel() else []
        if not selected:
            return None
        item = self.overview_table.item(selected[0].row(), 0)
        if not item or item.text() == "合计":
            return None
        return item.text()

    def current_overview_rows(self) -> list[dict[str, object | None]]:
        return [self._table_row_to_dict(self.overview_table, AP_ONLINE_OVERVIEW_COLUMNS, row) for row in range(self.overview_table.rowCount())]

    def refresh_overview_table(self, resources: list[dict[str, object | None]] | None = None) -> None:
        ac_uuid = self.current_device_uuid()
        if not ac_uuid:
            self._set_rows(self.overview_table, AP_ONLINE_OVERVIEW_COLUMNS, [])
            self._set_rows(self.offline_stats_table, OFFLINE_AP_STATS_COLUMNS, [])
            self._set_rows(self.offline_ledger_table, OFFLINE_AP_LEDGER_COLUMNS, [])
            return
        source_rows = resources if resources is not None else self.repository.list_fit_ap_resources_with_metadata(ac_uuid)
        current_optical_rows = self.repository.list_fit_ap_optical(ac_uuid)
        capacity_details = self.repository.list_active_trackside_plan_capacity_details()
        self._overview_uses_trackside_plan = bool(capacity_details)
        if not capacity_details:
            capacity_details = self.repository.list_station_ap_capacity_details()
        self._set_rows(
            self.overview_table,
            AP_ONLINE_OVERVIEW_COLUMNS,
            ApOnlineOverviewService.build_rows(
                metadata_rows=self.repository.list_fit_ap_metadata(),
                fit_ap_resources=source_rows,
                optical_rows=current_optical_rows,
                capacity_details=capacity_details,
            ),
        )

    def _handle_trackside_plan_saved(self) -> None:
        self.refresh_overview_table()
        self.refresh_trackside_table()

    def _rebuild_device_optical_status_lookup(self) -> None:
        """Rebuild the device optical status lookup from all devices and their optical modules.

        This lookup maps (device_name, interface_name) -> optical module status,
        which is the single source of truth from device detail.
        """
        devices = self.device_repository.list()
        optical_by_device = {str(device.device_uuid or ""): self.fact_repository.list_optical_modules(str(device.device_uuid or "")) for device in devices}
        self._device_optical_status_lookup = build_switch_data_lookup(devices, optical_by_device)

    def refresh_trackside_table(self) -> None:
        devices = filter_station_switch_devices(self.device_repository.list(), self.device_repository.database, self.site_name)
        interfaces_by_device = {str(device.device_uuid or ""): self.fact_repository.list_device_interfaces(str(device.device_uuid or "")) for device in devices}
        optical_by_device = {str(device.device_uuid or ""): self.fact_repository.list_optical_modules(str(device.device_uuid or "")) for device in devices}
        lldp_by_device = {str(device.device_uuid or ""): self.fact_repository.list_lldp_neighbors(str(device.device_uuid or "")) for device in devices}
        lookup = self._device_optical_status_lookup or build_switch_data_lookup(devices, optical_by_device)
        self.trackside_rows = build_trackside_ap_business_rows(
            devices,
            interfaces_by_device,
            optical_by_device,
            self.repository.list_all_fit_ap_optical(),
            lldp_by_device,
            self.repository.list_all_fit_ap_resources_with_metadata(),
            lookup,
            self.repository.get_active_trackside_pvid_plan(),
            self.offline_ap_ledger_rows,
        )
        self._set_trackside_site_filter_items(self.trackside_rows)
        self.apply_trackside_filters()

    def _set_site_filter_items(self, rows: list[dict[str, object | None]]) -> None:
        current = self.optical_site_filter.currentData()
        self.optical_site_filter.blockSignals(True)
        self.optical_site_filter.clear()
        for label, value in build_site_filter_items(rows, self.i18n.t("field.all")):
            self.optical_site_filter.addItem(label, value)
        index = self.optical_site_filter.findData(current)
        self.optical_site_filter.setCurrentIndex(index if index >= 0 else 0)
        self.optical_site_filter.blockSignals(False)

    def _set_resource_filter_items(self, rows: list[dict[str, object | None]]) -> None:
        current_group = self.resource_group_filter.currentData()
        current_state = self.resource_state_filter.currentData()
        self.resource_group_filter.blockSignals(True)
        self.resource_state_filter.blockSignals(True)
        self.resource_group_filter.clear()
        self.resource_state_filter.clear()
        self.resource_group_filter.addItem(self.i18n.t("ap.group_filter_all"), "")
        for group in sorted({str(row.get("group_name") or "").strip() for row in rows if str(row.get("group_name") or "").strip()}):
            self.resource_group_filter.addItem(group, group)
        self.resource_state_filter.addItem(self.i18n.t("ap.state_filter_all"), "")
        self.resource_state_filter.addItem(self.i18n.t("ap.state_filter_online"), "__online__")
        self.resource_state_filter.addItem(self.i18n.t("ap.state_filter_offline"), "__offline__")
        added: set[str] = {"", "__online__", "__offline__"}
        for label in sorted({_resource_state_label(row) for row in rows if _resource_state_label(row)}):
            key = label.casefold()
            if key in added:
                continue
            self.resource_state_filter.addItem(label, key)
            added.add(key)
        group_index = self.resource_group_filter.findData(current_group)
        state_index = self.resource_state_filter.findData(current_state)
        self.resource_group_filter.setCurrentIndex(group_index if group_index >= 0 else 0)
        self.resource_state_filter.setCurrentIndex(state_index if state_index >= 0 else 0)
        self.resource_group_filter.blockSignals(False)
        self.resource_state_filter.blockSignals(False)

    def _set_trackside_site_filter_items(self, rows: list[dict[str, object | None]]) -> None:
        current = self.trackside_site_filter.currentData()
        self.trackside_site_filter.blockSignals(True)
        self.trackside_site_filter.clear()
        for label, value in build_trackside_site_filter_items(rows, self.i18n.t("field.all")):
            self.trackside_site_filter.addItem(label, value)
        index = self.trackside_site_filter.findData(current)
        self.trackside_site_filter.setCurrentIndex(index if index >= 0 else 0)
        self.trackside_site_filter.blockSignals(False)

    def show_resource_context_menu(self, position) -> None:
        index = self.resources_table.indexAt(position)
        menu = self.build_resource_context_menu(index.row(), index.column())
        menu.exec(self.resources_table.viewport().mapToGlobal(position))

    def build_resource_context_menu(self, row: int, column: int) -> QMenu:
        menu = create_table_context_menu(self.resources_table, row, column, self.i18n.language, include_history=False)
        menu.insertSeparator(menu.actions()[0] if menu.actions() else None)
        refresh_ap = QAction(self.i18n.t("ap.refresh_optical_current"), menu)
        open_terminal = QAction("打开外部终端", menu)
        detail = QAction(self.i18n.t("ap.view_details"), menu)
        if menu.actions():
            menu.insertAction(menu.actions()[0], refresh_ap)
            menu.insertAction(menu.actions()[0], open_terminal)
            menu.insertAction(menu.actions()[0], detail)
        else:
            menu.addAction(refresh_ap)
            menu.addAction(open_terminal)
            menu.addAction(detail)
        refresh_ap.setEnabled(row >= 0 and self.optical_thread is None and self.resource_thread is None)
        open_terminal.setEnabled(row >= 0)
        detail.setEnabled(row >= 0)
        refresh_ap.triggered.connect(lambda: self.refresh_resource_ap_optical(row))
        open_terminal.triggered.connect(lambda: self.open_resource_ap_external_terminal(row))
        detail.triggered.connect(lambda: self.open_ap_detail(row))
        return menu

    def open_resource_ap_external_terminal(self, row: int) -> None:
        rows = self.current_resource_page_rows()
        if row < 0 or row >= len(rows):
            return
        current_row = rows[row]
        ap_name = str(current_row.get("ap_name") or "").strip() or "FIT-AP"
        ap_ip = str(current_row.get("ap_ip") or "").strip()
        if not ap_ip:
            message = "当前 AP 没有 IP，无法打开外部终端"
            self.status_label.setText(message)
            MessageBox.information(self, "打开外部终端", message)
            return
        config = self._select_external_terminal_config()
        if config is None:
            return
        self.status_label.setText(f"正在打开外部终端：{ap_name} {ap_ip}")
        device = Device(
            name=ap_name,
            ip_address=ap_ip,
            device_vendor="H3C",
            device_type="FIT-AP",
            ssh_enabled=0,
            telnet_enabled=1,
            telnet_port=23,
            telnet_username="",
            telnet_password="h3capadmin",
            username="",
            password="h3capadmin",
            protocol="Telnet",
            port=23,
        )
        result = launch_external_terminal(device, config)
        if result.success:
            self.status_label.setText(f"已打开外部终端：{ap_ip}")
            return
        reason = _friendly_external_terminal_error(result.message)
        self.status_label.setText(f"打开外部终端失败：{reason}")
        MessageBox.warning(self, "打开外部终端", f"打开外部终端失败：{reason}")

    def _select_external_terminal_config(self):
        configs = available_external_terminal_configs(self.settings)
        if not configs:
            MessageBox.information(self, "打开外部终端", self.i18n.t("external_terminal.not_configured"))
            self.status_label.setText("打开外部终端失败：未配置外部终端路径")
            return None
        if len(configs) == 1:
            return configs[0]
        labels = [TERMINAL_LABELS.get(config.terminal_type, config.terminal_type) for config in configs]
        label, accepted = InputDialog.getItem(self, self.i18n.t("external_terminal.select_terminal"), self.i18n.t("external_terminal.select_terminal"), labels, 0, False)
        if not accepted:
            return None
        return configs[labels.index(label)]

    def refresh_resource_ap_optical(self, row: int) -> None:
        device = self.current_device()
        rows = self.current_resource_page_rows()
        if device is None or row < 0 or row >= len(rows):
            return
        current_row = rows[row]
        ap_uuid = str(current_row.get("ap_uuid") or "").strip()
        ap_mac = str(current_row.get("ap_mac") or "").strip()
        ap_name = str(current_row.get("ap_name") or "").strip()
        if not any((ap_uuid, ap_mac, ap_name)):
            return
        label = ap_name or ap_mac or ap_uuid
        self._set_update_running(True, self.i18n.t("ap.refreshing_current_optical", ap=label))
        self.optical_thread = FitApOpticalCollectThread(
            device,
            self.site_name,
            int(self.optical_concurrency_combo.currentData() or 200),
            self,
            target_ap_uuids=[ap_uuid] if ap_uuid else None,
            target_ap_macs=[ap_mac] if ap_mac else None,
            target_ap_names=[ap_name] if ap_name else None,
        )
        self.optical_thread.progress.connect(self._set_update_progress)
        self.optical_thread.collect_finished.connect(self._finish_optical_collect)
        self.optical_thread.collect_failed.connect(self._fail_optical_collect)
        self.optical_thread.finished.connect(self.optical_thread.deleteLater)
        self.optical_thread.finished.connect(lambda: setattr(self, "optical_thread", None))
        self.optical_thread.start()

    def show_optical_context_menu(self, position) -> None:
        index = self.optical_table.indexAt(position)
        menu = self.build_optical_context_menu(index.row(), index.column())
        menu.exec(self.optical_table.viewport().mapToGlobal(position))

    def build_optical_context_menu(self, row: int, column: int) -> QMenu:
        menu = create_table_context_menu(self.optical_table, row, column, self.i18n.language, include_history=False)
        menu.insertSeparator(menu.actions()[0] if menu.actions() else None)
        detail = QAction(self.i18n.t("ap.view_details"), menu)
        if menu.actions():
            menu.insertAction(menu.actions()[0], detail)
        else:
            menu.addAction(detail)
        detail.setEnabled(row >= 0)
        detail.triggered.connect(lambda: self.open_ap_detail_from_optical(row))
        return menu

    def show_trackside_context_menu(self, position) -> None:
        index = self.trackside_table.indexAt(position)
        menu = self.build_trackside_context_menu(index.row(), index.column())
        menu.exec(self.trackside_table.viewport().mapToGlobal(position))

    def build_trackside_context_menu(self, row: int, column: int) -> QMenu:
        menu = create_table_context_menu(self.trackside_table, row, column, self.i18n.language, include_history=False)
        first_action = menu.actions()[0] if menu.actions() else None
        device_detail = QAction(self.i18n.t("trackside.view_device_detail"), menu)
        ap_detail = QAction(self.i18n.t("trackside.view_ap_detail"), menu)
        device_detail.setEnabled(row >= 0)
        ap_detail.setEnabled(row >= 0)
        device_detail.triggered.connect(lambda: self.open_device_detail_from_trackside(row))
        ap_detail.triggered.connect(lambda: self.open_ap_detail_from_trackside(row))
        if first_action:
            menu.insertAction(first_action, device_detail)
            menu.insertAction(first_action, ap_detail)
            menu.insertSeparator(first_action)
        else:
            menu.addAction(device_detail)
            menu.addAction(ap_detail)
        return menu

    def open_ap_detail(self, row: int) -> None:
        ac_uuid = self.current_device_uuid()
        item = self.resources_table.item(row, CHECK_COLUMN)
        ap_uuid = str(item.data(Qt.UserRole)) if item else ""
        if not ac_uuid or not ap_uuid:
            return
        dialog = FitApDetailDialog(self.i18n, self.repository, ac_uuid, ap_uuid)
        self.detail_windows.append(dialog)
        dialog.destroyed.connect(lambda _=None, window=dialog: self._forget_detail_window(window))
        show_non_focus_window(self, dialog, key=f"fit_ap_detail:{ap_uuid}", activate=False, raise_window=False)

    def open_ap_detail_from_optical(self, row: int) -> None:
        ac_uuid = self.current_device_uuid()
        item = self.optical_table.item(row, 0)
        if not ac_uuid or not item:
            return
        ap_name = item.text()
        page_rows = self.current_optical_page_rows()
        current_row = page_rows[row] if 0 <= row < len(page_rows) else {}
        ap_uuid = str(current_row.get("ap_uuid") or "")
        resource_rows = self.repository.list_fit_ap_resources_with_metadata(ac_uuid)
        resource_index = next((index for index, resource in enumerate(resource_rows) if resource.get("ap_uuid") == ap_uuid), -1)
        if resource_index >= 0:
            self.open_ap_detail(resource_index)
        else:
            dialog = FitApDetailDialog(self.i18n, self.repository, ac_uuid, ap_uuid or ap_name)
            self.detail_windows.append(dialog)
            dialog.destroyed.connect(lambda _=None, window=dialog: self._forget_detail_window(window))
            show_non_focus_window(self, dialog, key=f"fit_ap_detail:{ap_uuid or ap_name}", activate=False, raise_window=False)

    def open_device_detail_from_trackside(self, row: int) -> None:
        rows = self.current_trackside_page_rows()
        if row < 0 or row >= len(rows):
            return
        device_uuid = str(rows[row].get("device_uuid") or "")
        device = next((item for item in self.device_repository.list() if item.device_uuid == device_uuid), None)
        if device is None:
            return
        dialog = DeviceDetailDialog(self.i18n, self.fact_repository, device, self, self.site_name)
        self.detail_windows.append(dialog)
        dialog.destroyed.connect(lambda _=None, window=dialog: self._forget_detail_window(window))
        show_non_focus_window(self, dialog, key=f"device_detail:{device.device_uuid or device.id}", activate=False, raise_window=False)

    def open_ap_detail_from_offline_ledger(self, row: int) -> None:
        ac_uuid = self.current_device_uuid()
        if not ac_uuid or row < 0 or row >= len(self.offline_ap_ledger_rows):
            return
        offline = self.offline_ap_ledger_rows[row]
        ap_uuid = self._resolve_fit_ap_uuid(
            ac_uuid,
            ap_uuid=offline.get("ap_uuid"),
            ap_mac=offline.get("ap_mac"),
            ap_name=offline.get("ap_name"),
        )
        target = ap_uuid or str(offline.get("ap_name") or offline.get("ap_mac") or "")
        if not target:
            return
        dialog = FitApDetailDialog(self.i18n, self.repository, ac_uuid, target)
        self.detail_windows.append(dialog)
        dialog.destroyed.connect(lambda _=None, window=dialog: self._forget_detail_window(window))
        show_non_focus_window(self, dialog, key=f"fit_ap_detail:{target}", activate=False, raise_window=False)

    def _resolve_fit_ap_uuid(self, ac_uuid: str, *, ap_uuid: object = None, ap_mac: object = None, ap_name: object = None) -> str:
        uuid_text = str(ap_uuid or "").strip()
        mac_text = _normalize_mac_text(ap_mac)
        name_text = str(ap_name or "").strip()
        for resource in self.repository.list_fit_ap_resources_with_metadata(ac_uuid):
            if uuid_text and str(resource.get("ap_uuid") or "") == uuid_text:
                return uuid_text
            if mac_text and _normalize_mac_text(resource.get("ap_mac")) == mac_text:
                return str(resource.get("ap_uuid") or "")
            if name_text and str(resource.get("ap_name") or "") == name_text:
                return str(resource.get("ap_uuid") or "")
        return uuid_text

    def open_ap_detail_from_trackside(self, row: int) -> None:
        rows = self.current_trackside_page_rows()
        if row < 0 or row >= len(rows):
            return
        current_row = rows[row]
        ac_uuid = str(current_row.get("ac_device_uuid") or self.current_device_uuid() or "")
        ap_uuid = str(current_row.get("ap_uuid") or "")
        if not ac_uuid or not ap_uuid:
            MessageBox.information(self, self.i18n.t("trackside.view_ap_detail"), self.i18n.t("trackside.ap_not_found"))
            return
        dialog = FitApDetailDialog(self.i18n, self.repository, ac_uuid, ap_uuid)
        self.detail_windows.append(dialog)
        dialog.destroyed.connect(lambda _=None, window=dialog: self._forget_detail_window(window))
        show_non_focus_window(self, dialog, key=f"fit_ap_detail:{ap_uuid}", activate=False, raise_window=False)

    def _forget_detail_window(self, window: FitApDetailDialog) -> None:
        self.detail_windows = [item for item in self.detail_windows if item is not window]

    def _set_summary(self, row: dict[str, object | None] | None) -> None:
        for _key, field in SUMMARY_FIELDS:
            if field == "https_port":
                device = self.current_device()
                port, source = effective_https_port(device.https_port if device is not None else None)
                self.summary_labels[field].setText(str(port) if source == "device" else self.i18n.t("ac.https_port_default", port=port))
                continue
            value = row.get(field) if row else None
            self.summary_labels[field].setText(str(value) if value not in (None, "") else "-")
        if row and self.summary_labels.get("cpu_usage"):
            self.summary_labels["cpu_usage"].setToolTip(
                f"5秒：{row.get('cpu_5s') or '-'}%\n1分钟：{row.get('cpu_1m') or '-'}%\n5分钟：{row.get('cpu_5m') or '-'}%"
            )

    def _set_rows(self, table: QTableWidget, columns: tuple[tuple[str, str], ...], rows: list[dict[str, object | None]]) -> None:
        previous_updating_online_summary = self._updating_online_summary
        if table is self.overview_table:
            self._updating_online_summary = True
        table.blockSignals(True)
        table.setUpdatesEnabled(False)
        table.setSortingEnabled(False)
        try:
            table.setRowCount(len(rows))
            for row_index, row in enumerate(rows):
                for column_index, (_key, field) in enumerate(columns):
                    if field == "select":
                        key = self._ap_selection_key(row)
                        item = create_checkable_table_item(key in self.selected_ap_keys, user_data=key)
                    else:
                        value = row.get(field)
                        if table is self.optical_table and field == "switch_optical_status":
                            value = display_optical_status(evaluate_fit_ap_switch_status(row), self.i18n.language)
                        elif table is self.optical_table and field == "optical_alarm_status":
                            value = OFFLINE_AP_STATUS_TEXT if bool(row.get("is_ap_offline")) else display_optical_status(evaluate_fit_ap_ap_status(row), self.i18n.language)
                        elif field in {"lldp_source", "lldp_match_status", "optical_match_status", "link_match_status"}:
                            value = _display_link_value(row, field)
                        elif field in {"switch_optical_status", "ap_optical_status"}:
                            value = display_optical_status(str(value or ""), self.i18n.language) if value else value
                        elif field in {"mileage", "mileage_text"}:
                            value = _display_mileage_for_row(row, field)
                        elif field == "belong_type":
                            value = _display_belong_type(value)
                        item = QTableWidgetItem(str(value) if value not in (None, "") else "-")
                        if field in {"mileage", "mileage_text"}:
                            meters = row.get("_mileage_meters") if field == "mileage" else _mileage_meters_for_row(row, field)
                            if meters is not None:
                                item.setData(Qt.ItemDataRole.UserRole, float(meters))
                        plan_locked = table is self.overview_table and bool(row.get("source") == "trackside_plan" or row.get("remark") == "\u8f68\u65c1AP\u89c4\u5212")
                        if field == "remark":
                            plan_locked = False
                        if table is self.overview_table and field in {"total", "remark"} and row.get("site") != "合计" and not plan_locked:
                            item.setFlags(item.flags() | Qt.ItemIsEditable | Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                        else:
                            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                        if plan_locked and field == "total":
                            item.setToolTip(self.i18n.t("ac.trackside_plan_total_locked"))
                        if field == "state_display":
                            item.setToolTip(f"{self.i18n.t('ap.state_raw')}: {row.get('state_raw') or row.get('state') or '-'}")
                        if table is self.optical_table:
                            apply_status_item_contrast(item, evaluate_fit_ap_row_status(row))
                        if table is self.trackside_table:
                            apply_status_item_contrast(item, trackside_row_status(row))
                        if table is self.overview_table:
                            self._apply_overview_color(item, row, field)
                    item.setTextAlignment(Qt.AlignCenter)
                    table.setItem(row_index, column_index, item)
        finally:
            table.blockSignals(False)
            table.setSortingEnabled(False)
            table.setUpdatesEnabled(True)
            if table is self.overview_table:
                self._updating_online_summary = previous_updating_online_summary
            if table is self.resources_table:
                self._resize_resource_columns()
            elif table is self.optical_table:
                self._resize_optical_columns()
            elif table is self.overview_table:
                auto_resize_table_columns(table)
            elif table is self.trackside_table:
                auto_resize_table_columns(table)
            else:
                auto_resize_table_columns(table)

    def _resize_optical_columns(self) -> None:
        auto_resize_table_columns(self.optical_table)

    def _resize_resource_columns(self) -> None:
        fields = [field for _key, field in FIT_AP_RESOURCE_COLUMNS]
        min_widths = {index: FIT_AP_RESOURCE_COLUMN_MIN_WIDTHS.get(field, 90) for index, field in enumerate(fields)}
        max_widths = {index: FIT_AP_RESOURCE_COLUMN_MAX_WIDTHS.get(field, 260) for index, field in enumerate(fields)}
        auto_fit_table_columns(self.resources_table, min_widths=min_widths, max_widths=max_widths, padding=28)
        header = self.resources_table.horizontalHeader()
        header.setStretchLastSection(False)
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.resources_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)

    def _apply_overview_color(self, item: QTableWidgetItem, row: dict[str, object | None], field: str) -> None:
        total = int(row.get("total") or 0)
        online = int(row.get("online") or 0)
        offline = int(row.get("offline") or 0)
        if total and online == total:
            apply_item_contrast(item, STATUS_COLOR_MAP["normal"])
        elif total and online / total < 0.8:
            apply_item_contrast(item, STATUS_COLOR_MAP["warning"])
        if field == "offline" and offline > 0:
            apply_item_contrast(item, STATUS_COLOR_MAP["alarm"])

    @staticmethod
    def _table_row_to_dict(table: QTableWidget, columns: tuple[tuple[str, str], ...], row: int) -> dict[str, object | None]:
        return {field: table.item(row, column).text() if table.item(row, column) else None for column, (_key, field) in enumerate(columns)}

    def save_overview_total(self, item: QTableWidgetItem) -> None:
        if self._updating_online_summary:
            return
        if item.column() not in {1, 5}:
            return
        site_item = self.overview_table.item(item.row(), 0)
        if not site_item or site_item.text() == "合计":
            return
        if item.column() == 5:
            remark = item.text()
            if remark == "-":
                remark = ""
            if len(remark) > 500:
                MessageBox.warning(self, self.i18n.t("ac.ap_online_overview"), self.i18n.t("ac.remark_too_long"))
                self.refresh_overview_table()
                return
            self.repository.upsert_station_ap_remark(site_item.text(), remark)
            self.refresh_overview_table()
            return
        if self._overview_uses_trackside_plan:
            MessageBox.information(self, self.i18n.t("ac.ap_online_overview"), self.i18n.t("ac.trackside_plan_total_locked"))
            self.refresh_overview_table()
            return
        try:
            total = int(item.text())
            if total < 0:
                raise ValueError
        except ValueError:
            MessageBox.warning(self, self.i18n.t("ac.ap_online_overview"), self.i18n.t("ac.ap_total_invalid"))
            self.refresh_overview_table()
            return
        self.repository.upsert_station_ap_capacity(site_item.text(), total)
        self.refresh_overview_table()
