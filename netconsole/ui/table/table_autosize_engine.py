from __future__ import annotations

from dataclasses import dataclass
from typing import Iterable, Mapping, Sequence

from PySide6.QtCore import Qt
from PySide6.QtGui import QFontMetrics
from PySide6.QtWidgets import QHeaderView, QTableWidget


DEFAULT_PADDING = 28
DEFAULT_CHAR_PIXEL = 8
DEFAULT_MAX_SAMPLE_ROWS = 500
MIN_COLUMN_WIDTH = 72
MAX_COLUMN_WIDTH = 520
HIGH_PRIORITY_MIN_WIDTH = 180
MEDIUM_PRIORITY_MIN_WIDTH = 112
LOW_PRIORITY_MIN_WIDTH = 82
MEDIUM_PRIORITY_MAX_WIDTH = 190
LOW_PRIORITY_MAX_WIDTH = 150
ACTION_COLUMN_WIDTH = 320
CHECK_COLUMN_WIDTH = 48

HIGH_PRIORITY_FIELDS = {
    "ap_name",
    "name",
    "device_name",
    "neighbor_device_name",
    "interface_name",
    "local_interface",
    "neighbor_interface",
    "description",
    "site",
    "peer_site",
    "station_name",
    "source_file",
    "source_file_name",
    "file_name",
    "archived_filename",
    "sample_time",
    "start_time",
    "end_time",
    "first_sample_time",
    "last_sample_time",
    "event_time",
    "build_start_time",
    "build_end_time",
    "imported_at",
}
MEDIUM_PRIORITY_FIELDS = {
    "ap_mac",
    "mac_address",
    "ip_address",
    "ap_ip",
    "vlan",
    "pvid",
    "status",
    "state_display",
    "port_status",
    "link_status",
    "protocol_status",
    "switch_optical_status",
    "ap_optical_status",
    "optical_alarm_status",
}
LOW_PRIORITY_FIELDS = {
    "updated_at",
    "collected_at",
    "online_time",
    "elapsed",
    "rx_power",
    "tx_power",
    "neighbor_rx_power",
    "temperature",
    "voltage",
    "bias_current",
    "apid",
    "id",
    "total",
    "online",
    "offline",
}


@dataclass(frozen=True)
class AutosizeColumn:
    width: int
    priority: str


def weighted_text_length(value: object) -> int:
    length = 0
    for char in str(value or ""):
        length += 2 if _is_wide_char(char) else 1
    return length


def weighted_text_pixel_width(value: object, char_pixel: int = DEFAULT_CHAR_PIXEL) -> int:
    return weighted_text_length(value) * char_pixel


def calculate_column_widths(table: QTableWidget, screen_width: int | None = None, max_rows: int = DEFAULT_MAX_SAMPLE_ROWS) -> dict[int, int]:
    _ = screen_width
    if table.columnCount() <= 0:
        return {}

    fields = _column_fields(table)
    row_indexes = _sample_rows(table.rowCount(), max_rows)
    base: dict[int, AutosizeColumn] = {}
    for column in range(table.columnCount()):
        field = fields[column] if column < len(fields) else ""
        priority = _column_priority(field, _header_text(table, column))
        width = _base_column_width(table, column, priority, row_indexes)
        base[column] = AutosizeColumn(width=width, priority=priority)
    target = {column: sizing.width for column, sizing in base.items()}
    return target


def apply_table_autosize(table: QTableWidget, screen_width: int | None = None, max_rows: int = DEFAULT_MAX_SAMPLE_ROWS) -> dict[int, int]:
    widths = calculate_column_widths(table, screen_width, max_rows=max_rows)
    if not widths:
        return {}

    header = table.horizontalHeader()
    header.setSectionResizeMode(QHeaderView.Interactive)
    header.setStretchLastSection(False)
    header.setSectionsMovable(False)
    table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
    table.setWordWrap(False)
    table.setTextElideMode(Qt.TextElideMode.ElideRight)
    for column, width in widths.items():
        field = _field_for(table, column)
        if field == "actions":
            header.setSectionResizeMode(column, QHeaderView.Fixed)
            table.setColumnWidth(column, ACTION_COLUMN_WIDTH)
            widths[column] = ACTION_COLUMN_WIDTH
        elif field == "select":
            header.setSectionResizeMode(column, QHeaderView.Fixed)
            table.setColumnWidth(column, CHECK_COLUMN_WIDTH)
            widths[column] = CHECK_COLUMN_WIDTH
        else:
            table.setColumnWidth(column, width)
            header.setSectionResizeMode(column, QHeaderView.Interactive)
    table.setProperty("netconsole_auto_layout_widths", widths)
    return widths


def excel_column_width(value: object, minimum: float = 8.0, maximum: float = 60.0, field: str | None = None, header: str | None = None) -> float:
    field_min, field_max = excel_column_width_bounds(field, header)
    effective_min = max(minimum, field_min)
    effective_max = min(maximum, field_max) if maximum else field_max
    width = weighted_text_length(value) * 1.15 + 3
    return max(effective_min, min(width, effective_max))


def excel_column_width_bounds(field: str | None = None, header: str | None = None) -> tuple[float, float]:
    text = f"{field or ''} {header or ''}".casefold()
    if any(token in text for token in ("序号", "source_line_number", "line_number", "record_seq")):
        return 8.0, 12.0
    if "radio" in text:
        return 8.0, 12.0
    if any(token in text for token in ("状态", "state", "result", "结果")):
        return 10.0, 18.0
    if any(token in text for token in ("rssi", "busy", "采样点数", "sample_count", "链路数", "link_count")):
        return 10.0, 14.0
    if any(token in text for token in ("mac", "peermac")):
        return 18.0, 22.0
    if any(token in text for token in ("time", "时间", "duration", "时长")):
        return 18.0, 26.0
    if any(token in text for token in ("ap名称", "ap name", "peer_ap_name", "站点", "station", "peer_site")):
        return 18.0, 32.0
    if any(token in text for token in ("source_file", "archived_filename", "file", "文件", "路径", "path")):
        return 30.0, 80.0
    return 8.0, 40.0


def calculate_excel_column_widths(
    headers: Sequence[object],
    rows: Iterable[Mapping[str, object | None] | Sequence[object]],
    fields: Sequence[str] | None = None,
) -> list[float]:
    widths = [excel_column_width(header, field=fields[index] if fields and index < len(fields) else None, header=str(header)) for index, header in enumerate(headers)]
    for row in rows:
        for index in range(len(widths)):
            value: object | None
            field = fields[index] if fields and index < len(fields) else str(headers[index])
            if isinstance(row, Mapping):
                value = row.get(field)
            else:
                value = row[index] if index < len(row) else None
            widths[index] = max(widths[index], excel_column_width(value, field=field, header=str(headers[index])))
    return widths


def apply_worksheet_column_widths(
    sheet,
    headers: Sequence[object],
    rows: Iterable[Mapping[str, object | None] | Sequence[object]],
    fields: Sequence[str] | None = None,
    maximum: float = 60.0,
) -> None:
    from openpyxl.utils import get_column_letter

    for index, width in enumerate(calculate_excel_column_widths(headers, rows, fields), start=1):
        sheet.column_dimensions[get_column_letter(index)].width = min(width, maximum)


def apply_worksheet_autofit(sheet, maximum: float = 60.0) -> None:
    from openpyxl.utils import get_column_letter

    for column_index in range(1, sheet.max_column + 1):
        header_value = sheet.cell(row=1, column=column_index).value
        width = excel_column_width(header_value, header=str(header_value or ""), maximum=maximum)
        for cell in sheet[get_column_letter(column_index)]:
            width = max(width, excel_column_width(cell.value, header=str(header_value or ""), maximum=maximum))
        sheet.column_dimensions[get_column_letter(column_index)].width = min(width, maximum)


def _install_resize_hook(table: QTableWidget) -> None:
    if table.property("netconsole_table_autosize_hooked"):
        return
    original_resize_event = table.resizeEvent

    def resize_event(event):
        original_resize_event(event)
        apply_table_autosize(table)

    table.resizeEvent = resize_event
    table.setProperty("netconsole_table_autosize_hooked", True)


def _base_column_width(table: QTableWidget, column: int, priority: str, row_indexes: list[int]) -> int:
    header = table.horizontalHeader()
    header_text = _header_text(table, column)
    header_metrics = QFontMetrics(header.font())
    cell_metrics = QFontMetrics(table.font())
    header_width = max(header_metrics.horizontalAdvance(header_text), weighted_text_pixel_width(header_text)) + DEFAULT_PADDING + 12
    content_width = 0
    for row in row_indexes:
        item = table.item(row, column)
        text = item.text() if item else ""
        content_width = max(content_width, cell_metrics.horizontalAdvance(text), weighted_text_pixel_width(text))
        if item is not None and text and not item.toolTip():
            item.setToolTip(text)
    width = max(header_width, content_width + DEFAULT_PADDING)
    if priority == "high":
        return max(HIGH_PRIORITY_MIN_WIDTH, int(header_width), min(width, MAX_COLUMN_WIDTH))
    if priority == "medium":
        return max(MEDIUM_PRIORITY_MIN_WIDTH, int(header_width), min(width, MEDIUM_PRIORITY_MAX_WIDTH))
    if priority == "low":
        return max(LOW_PRIORITY_MIN_WIDTH, int(header_width), min(width, LOW_PRIORITY_MAX_WIDTH))
    return max(MIN_COLUMN_WIDTH, int(header_width), min(width, MAX_COLUMN_WIDTH))


def _compress_columns(target: dict[int, int], base: dict[int, AutosizeColumn], overflow: int) -> dict[int, int]:
    result = dict(target)
    for priority, minimum in (("low", LOW_PRIORITY_MIN_WIDTH), ("medium", MEDIUM_PRIORITY_MIN_WIDTH), ("normal", MIN_COLUMN_WIDTH), ("high", HIGH_PRIORITY_MIN_WIDTH)):
        columns = [column for column, sizing in base.items() if sizing.priority == priority]
        for column in columns:
            if overflow <= 0:
                return result
            reducible = max(result[column] - minimum, 0)
            reduction = min(reducible, overflow)
            result[column] -= reduction
            overflow -= reduction
    return result


def _available_width(table: QTableWidget, screen_width: int | None) -> int:
    if screen_width and screen_width > 0:
        return screen_width
    width = table.viewport().width() or table.width()
    return width if width > 0 else 1200


def _sample_rows(row_count: int, max_rows: int) -> list[int]:
    if row_count <= 0:
        return []
    max_rows = max(1, int(max_rows or DEFAULT_MAX_SAMPLE_ROWS))
    if row_count <= max_rows:
        return list(range(row_count))
    tail_count = min(50, max_rows // 5)
    head_count = max_rows - tail_count
    return list(range(head_count)) + list(range(row_count - tail_count, row_count))


def _column_priority(field: str, header: str) -> str:
    normalized_field = field.casefold()
    normalized_header = header.casefold()
    if normalized_field in {"actions", "select"}:
        return "utility"
    if normalized_field in HIGH_PRIORITY_FIELDS or any(token in normalized_header for token in ("ap名称", "ap name", "设备名称", "device name", "接口名称", "interface", "描述", "description", "车站", "站点", "文件", "station")):
        return "high"
    if normalized_field in MEDIUM_PRIORITY_FIELDS or any(token in normalized_header for token in ("mac", "ip", "vlan", "pvid", "状态", "status")):
        return "medium"
    if normalized_field in LOW_PRIORITY_FIELDS or any(token in normalized_header for token in ("time", "时间", "id", "rx", "tx", "power", "功率")):
        return "low"
    return "normal"


def _header_text(table: QTableWidget, column: int) -> str:
    item = table.horizontalHeaderItem(column)
    return item.text() if item else ""


def _field_for(table: QTableWidget, column: int) -> str:
    fields = _column_fields(table)
    return fields[column] if column < len(fields) else ""


def _column_fields(table: QTableWidget) -> list[str]:
    fields = table.property("netconsole_column_fields")
    if isinstance(fields, list):
        return [str(field) for field in fields]
    if isinstance(fields, tuple):
        return [str(field) for field in fields]
    return []


def _is_wide_char(char: str) -> bool:
    return ord(char) > 127
